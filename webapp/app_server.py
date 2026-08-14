# -*- coding: utf-8 -*-
"""
app_server.py — Coupang Service Operations System 앱 서버 (반응형 웹앱 백엔드)
============================================================
PC에서 실행하면 같은 와이파이의 휴대폰·다른 PC가 브라우저로 접속하는 ERP형 앱.
표준 라이브러리만 사용(설치 0). 데이터는 전부 사내 PC에 남는다(클라우드 전송 없음).

  실행:  python webapp/app_server.py            # 실서비스 (첫 실행 시 PIN 자동 생성)
         python webapp/app_server.py --demo     # 합성데이터 데모 (PIN 0000)
  접속:  PC      → http://localhost:8899
         휴대폰  → http://<PC IP>:8899   (같은 와이파이, 방화벽 허용 필요)

보안: 4자리 PIN(첫 요청 시 입력, 기기에 저장). 사내 LAN 전용 설계 — 외부 인터넷 개방 금지.
"""
import sys, os, re, json, glob, time, threading, random, hashlib, io, shutil, secrets, copy
import base64, hmac
import ipaddress
import socket
from collections import deque
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from email import policy as email_policy
from email.parser import BytesParser
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.error import HTTPError
from urllib.parse import parse_qs, urljoin, urlsplit
from urllib.request import HTTPRedirectHandler, Request, build_opener

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)
sys.path.insert(0, ROOT)
# 관리대장을 Z: 에서 매번 끌어오지 않고 메모리 사본에서 연다(속도 개선 2026-07-31).
# sys.path 를 세운 **뒤에** 임포트해야 한다 — 위로 올리면 ecount 모듈을 못 찾는다.
from ecount_reconcile import master_stream, master_book
from pct_fmt import pct, pct_text          # 비율 표기 단일 규칙 (2026-08-05 지시)
PY = sys.executable
ENV = {**os.environ, "PYTHONIOENCODING": "utf-8"}

DEMO = "--demo" in sys.argv
PORT = int(sys.argv[sys.argv.index("--port") + 1]) if "--port" in sys.argv else 8899
WEBCFG = os.path.join(ROOT, "config", "webapp.json")

# 폰 홈 화면 아이콘이 가리켜야 할 **바뀌지 않는 주소**.
# 터널 주소(trycloudflare)는 띄울 때마다 새로 받으므로 아이콘에 박으면 안 된다.
FIXED_ENTRY = "https://mulder4780.github.io/coupang-ecount-reconcile/"
FIXED_LIVE_ENTRY = "https://mulder.tailf14aae.ts.net"

# 담당자별 경로는 공개 이후 바꾸지 않는다. 표시명·체크리스트는 바꿀 수 있어도
# slug는 설치된 PWA의 id이자 담당자가 저장한 바로가기이므로 영구 식별자다.
STAFF_CENTERS = {
    "ryu-jiyeong": {
        "name": "류지영", "title": "류지영 쿠팡 AS 및 정기점검 업무센터",
        "checklist": [
            "신규 돌발AS 접수와 처리상태 확인",
            "정기점검 예정·실행·미실시 사유 입력",
            "카카오톡 정기점검방·돌발점검방 원본 업로드",
            "택배 발송·현장 조치 완료일과 근거 첨부",
            "거래명세서·세금계산서 발행 확인",
            "입금일·입금액·잔여 미수금 확인",
            "ERP 금액 불일치와 청구 미등록 보완",
            "리모컨 불출·납품 기록(증평 담당·AS 담당자당 3개)",
        ],
    },
    "oh-jonghyeon": {
        "name": "오종현", "title": "오종현 업무센터",
        "checklist": [
            "PO 원본·견적서 수신 여부 확인",
            "구매·입금 원천자료 누락 확인",
            "프로젝트번호·캠프·금액 불일치 보완",
            "리모컨 불출·납품 기록(부산 담당·AS 담당자당 3개)",
        ],
    },
    # 김미영 — 매출·계산서·입금 담당 (2026-08-12 지시로 신설).
    # ★ 이 센터의 기준은 **세금계산서 발행월**이다. 다른 센터는 완료월로 일한다 —
    #   같은 회사 숫자인데 기준이 달라서 어긋나 보이는 자리가 여기다.
    "kim-miyeong": {
        "name": "김미영", "title": "김미영 쿠팡 매출·정산 업무센터",
        "checklist": [
            "쿠팡 매출 실적(전체) 월별 확인 — 세금계산서 발행월 기준",
            "정기점검·유료AS 실적표와 앱 집계 대조",
            "세금계산서 미발행 건 확인과 발행 요청",
            "입금내역 대사 — 입금일·입금액·미수금",
            "보험사 입금 확인(자기부담금·감가로 청구액과 다를 수 있음)",
            "차이 나는 달의 원인 확인 — 미발행분인지 자료 밀림인지",
        ],
    },
    "yoo-hyeonmin": {
        "name": "유현민", "title": "유현민 업무센터",
        "checklist": [
            "돌발AS·정기점검 전체 진행상태와 예외 확인",
            "담당자·일정·비용구분 정정 승인",
            "거래서류·세금계산서·입금 근거 확인",
            "앱 DB 저장과 Excel 자동 보관본 상태 확인",
            "확인 필요 항목의 객관 근거와 완료 판정 검토",
        ],
    },
}

# 변재선 업무는 2026-07-30 류지영에게 이관했다. 공개된 기존 주소/PWA 바로가기는
# 깨뜨리지 않고 새 담당자의 고정 주소로 영구 연결한다.
STAFF_CENTER_ALIASES = {"byeon-jaeseon": "ryu-jiyeong"}


# ── AS 담당기사 전용 화면 (2026-08-08 지시) ─────────────────────────────────
# "업무센터에 각 AS 담당자 4명도 넣어서 **별도의 비밀번호 없는** 화면으로 딱 AS
#  담당자가 할 수 있는 업무만 넣어서 만들어줘 (링크 타고 열면 크롬으로 강제로
#  열어서 앱을 모바일에 설치할 수 있는 구조로 알고리즘 구성해)"
#
# ★ **'비밀번호 없음'은 '누구나'가 아니다.** 기사에게 PIN 을 외우게 하지 않는 대신,
#   추측할 수 없는 열쇠가 든 링크(`/t/<slug>?k=…`)로 연다. 그 링크로 얻는 권한은
#   `role="tech"` 이며 **관리자·업무센터 API 를 하나도 못 쓴다**(`_auth()` 가 막는다).
#   기사 화면이 쓰는 길은 `/api/tech/*` 뿐이고, 그 길은 **그 사람 것만** 돌려준다.
# ★ 금액·통화 메모는 이 화면에 올리지 않는다. 기사가 할 일에 필요 없고,
#   링크는 카톡으로 돌아다닌다 — 새어도 될 것만 싣는다.
AS_TECH_CENTERS = {
    "cha-dongho":    {"name": "차동호", "직함": "팀장"},
    "kim-junhyeong": {"name": "김준형", "직함": ""},
    "kwon-ocheol":   {"name": "권오철", "직함": ""},
    "kim-pilwoo":    {"name": "김필우", "직함": ""},
}
TECH_KEYS_PATH = os.path.join(ROOT, "config", "tech_keys.local.json")
_TECH_KEY_LOCK = threading.Lock()


def tech_keys():
    """기사별 링크 열쇠. 없으면 만들어 저장한다(파일은 git 밖이다).

    ★ 열쇠를 코드나 로그에 적지 않는다. 만들어 준 링크는 사람이 카톡으로 전한다."""
    with _TECH_KEY_LOCK:
        try:
            d = json.load(open(TECH_KEYS_PATH, encoding="utf-8"))
        except Exception:
            d = {}
        changed = False
        for slug in AS_TECH_CENTERS:
            if not str(d.get(slug) or "").strip():
                d[slug] = secrets.token_urlsafe(24)
                changed = True
        if changed:
            os.makedirs(os.path.dirname(TECH_KEYS_PATH), exist_ok=True)
            tmp = TECH_KEYS_PATH + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(d, f, ensure_ascii=False, indent=2)
            os.replace(tmp, TECH_KEYS_PATH)
        return d


def tech_check_key(slug, key):
    if slug not in AS_TECH_CENTERS:
        return False
    want = str(tech_keys().get(slug) or "")
    return bool(want) and hmac.compare_digest(want, str(key or ""))


def tech_board(slug, limit=60):
    """기사 한 사람의 일감. **그 사람 것만**, 금액 없이."""
    cfg = AS_TECH_CENTERS.get(slug)
    if not cfg:
        return {"ok": False, "error": "등록되지 않은 기사입니다"}
    name = cfg["name"]
    today = datetime.now().strftime("%Y-%m-%d")

    def mine(r):
        # 담당기사 칸에 여러 명이 들어가는 행이 있다("김준형, 김필우") — 이름 포함으로 본다.
        return name and name in str(r.get("담당기사") or "")

    try:
        works = get_works() or {}
    except Exception as exc:
        return {"ok": False, "error": "원장을 읽지 못했습니다: %s" % exc}

    밀린것, 예정, 최근완료 = [], [], []
    for r in works.get("as") or []:
        if not mine(r):
            continue
        done, got = norm_date(r.get("작업완료일")), norm_date(r.get("접수일자"))
        vis = norm_date(r.get("방문예정일"))
        base = {"종류": "돌발AS", "ID": r.get("접수ID") or "", "캠프명": r.get("캠프명") or "",
                "프로젝트NO": str(r.get("프로젝트NO") or "").split(" · ")[0],
                "신청내용": r.get("신청내용") or "", "긴급도": r.get("긴급도") or "",
                "진행상태": r.get("진행상태") or "", "밴드": r.get("밴드 바로가기") or ""}
        if done:
            최근완료.append(dict(base, 날짜=done))
        else:
            밀린것.append(dict(base, 날짜=got or vis or "", 접수일자=got,
                               방문예정일=vis, 경과일=_daydiff(got, today)))
        if not done and vis and vis >= today:
            예정.append(dict(base, 날짜=vis))
    for r in works.get("pm") or []:
        if not mine(r):
            continue
        plan, real = norm_date(r.get("점검예정일")), norm_date(r.get("실제점검일"))
        base = {"종류": "정기점검", "ID": r.get("점검ID") or "", "캠프명": r.get("캠프명") or "",
                "프로젝트NO": str(r.get("프로젝트NO") or "").split(" · ")[0],
                "점검상태": r.get("점검상태") or "", "이상발견": r.get("이상발견여부") or "",
                "밴드": ""}
        if real:
            최근완료.append(dict(base, 날짜=real))
        elif plan and plan <= today:
            밀린것.append(dict(base, 날짜=plan, 경과일=_daydiff(plan, today)))
        elif plan:
            예정.append(dict(base, 날짜=plan))

    밀린것.sort(key=lambda e: e.get("날짜") or "9999")
    예정.sort(key=lambda e: e.get("날짜") or "9999")
    최근완료.sort(key=lambda e: e.get("날짜") or "", reverse=True)
    return {"ok": True, "slug": slug, "이름": name, "직함": cfg.get("직함") or "",
            "기준": today,
            "밀린것": 밀린것[:limit], "예정": 예정[:limit], "최근완료": 최근완료[:20],
            "요약": {"밀린것": len(밀린것), "예정": len(예정), "이번달완료":
                     sum(1 for e in 최근완료 if str(e.get("날짜") or "")[:7] == today[:7])},
            "안내": "금액·정산 정보는 이 화면에 나오지 않습니다."}


def tech_report(slug, wid, when, note):
    """기사가 올리는 완료 보고. 앱 DB에는 즉시, Excel은 보관본으로만 남긴다."""
    cfg = AS_TECH_CENTERS.get(slug)
    if not cfg:
        return {"ok": False, "error": "등록되지 않은 기사입니다"}
    wid, when = str(wid or "").strip(), norm_date(when)
    note = str(note or "").strip()[:200]
    if not wid or not when:
        return {"ok": False, "error": "건과 완료일이 있어야 합니다"}
    board = tech_board(slug, limit=999)
    ids = {e["ID"]: e for e in (board.get("밀린것") or []) + (board.get("예정") or [])}
    if wid not in ids:
        # ★ 남의 건에 완료를 찍지 못하게 한다. 목록에 없는 ID 는 그 사람 것이 아니다.
        return {"ok": False, "error": "내 일감 목록에 없는 건입니다"}
    kind = ids[wid]["종류"]
    sheet = "02_돌발AS접수" if kind == "돌발AS" else "04_정기점검"
    keycol = "접수ID" if kind == "돌발AS" else "점검ID"
    col = "작업완료일" if kind == "돌발AS" else "실제점검일"
    # ★ 조치 내용은 **원장 칸에 쓰지 않고 근거로 남긴다.** 02시트에는 '조치 내용' 칸이
    #   없다. 가까워 보인다고 `신청내용`(고객이 무엇을 요청했나)에 적으면 그 칸의 뜻이
    #   망가지고, 나중에 아무도 그게 요청인지 조치인지 구별하지 못한다.
    ev = f"기사 앱 완료보고 · {cfg['name']}" + (f" · 조치: {note}" if note else "")
    saved = enqueue_for_scheduled_apply(
        [{"sheet": sheet, "key_col": keycol, "key": wid, "col": col,
          "value": when, "vtype": "date", "only_if_empty": True, "evidence": ev}],
        source="tech-app", actor=f"tech:{slug}",
        idempotency_key=f"tech:{slug}:{wid}:{when}")
    return {**saved, "건": wid, "완료일": when,
            "메모기록": bool(note),
            "안내": ("앱 DB 저장 완료 · Excel 보관본은 11:00·15:00 회차에 생성됩니다."
                   if saved.get("ok") else saved.get("msg"))}


def staff_centers_payload():
    return [{
        "slug": slug, **cfg,
        "path": f"/staff/{slug}",
        "url": f"{FIXED_LIVE_ENTRY}/staff/{slug}",
    } for slug, cfg in STAFF_CENTERS.items()]


def staff_completions_payload(limit=100):
    """세 담당자의 객관완료 정본을 앱·외부 자동화에 제공한다."""
    import ledger_db
    return {
        "summary": ledger_db.staff_resolution_summary(),
        "rows": ledger_db.staff_resolutions(limit=limit),
        "basis": "완료일·원천 증빙·ERP·PO 대조로 입증된 건만 표시",
    }


PIN_HASH_ITERS = 310_000
PIN_STATE_LOCK = threading.RLock()


def _pin_record(pin, *, salt=None):
    """4자리 PIN은 평문으로 남기지 않고 PBKDF2 해시만 로컬 설정에 보관한다."""
    pin = str(pin or "").strip()
    if not re.fullmatch(r"\d{4}", pin):
        raise ValueError("PIN은 숫자 4자리여야 합니다")
    salt_bytes = bytes.fromhex(salt) if salt else secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", pin.encode("utf-8"), salt_bytes, PIN_HASH_ITERS)
    return {
        "salt": salt_bytes.hex(),
        "hash": digest.hex(),
        "iterations": PIN_HASH_ITERS,
    }


def _pin_matches(pin, record):
    try:
        pin = str(pin or "").strip()
        if not re.fullmatch(r"\d{4}", pin):
            return False
        salt = bytes.fromhex(str(record["salt"]))
        iterations = int(record.get("iterations") or PIN_HASH_ITERS)
        expected = bytes.fromhex(str(record["hash"]))
        actual = hashlib.pbkdf2_hmac(
            "sha256", pin.encode("utf-8"), salt, iterations)
        return secrets.compare_digest(actual, expected)
    except Exception:
        return False


def _atomic_write_webcfg(data):
    os.makedirs(os.path.dirname(WEBCFG), exist_ok=True)
    tmp = WEBCFG + f".tmp.{os.getpid()}"
    with open(tmp, "w", encoding="utf-8") as out:
        json.dump(data, out, ensure_ascii=False, indent=2)
        out.write("\n")
    os.replace(tmp, WEBCFG)


def load_pin_state():
    """기존 단일 평문 PIN 설정을 역할별 해시 설정으로 안전하게 이관한다."""
    if DEMO:
        demo = _pin_record("0000")
        return {
            "admin": demo,
            "staff": {slug: _pin_record("0000") for slug in STAFF_CENTERS},
            "token_secret": hashlib.sha256(b"csos-synthetic-device-session").hexdigest(),
            "versions": {"admin": 1, "staff": {slug: 1 for slug in STAFF_CENTERS}},
            "port": PORT,
        }
    try:
        with open(WEBCFG, encoding="utf-8") as src:
            cfg = json.load(src)
    except Exception:
        cfg = {}
    changed = False
    legacy_pin = str(cfg.pop("pin", "") or "").strip()
    auth = cfg.get("auth")
    if not isinstance(auth, dict):
        auth = {}
        cfg["auth"] = auth
        changed = True
    if not isinstance(auth.get("admin"), dict):
        bootstrap = legacy_pin if re.fullmatch(r"\d{4}", legacy_pin) else \
            str(random.SystemRandom().randint(1000, 9999))
        auth["admin"] = _pin_record(bootstrap)
        changed = True
    staff = auth.get("staff")
    if not isinstance(staff, dict):
        staff = {}
        auth["staff"] = staff
        changed = True
    for slug in STAFF_CENTERS:
        if not isinstance(staff.get(slug), dict):
            # 새 담당자 업무센터는 관리자 화면에서 PIN을 정하기 전까지 예측할 수 없는
            # 임시 PIN을 사용한다. 실제 운영 기본값은 --configure-pins로 한 번만 설정한다.
            staff[slug] = _pin_record(str(random.SystemRandom().randint(1000, 9999)))
            changed = True
    token_secret = str(auth.get("token_secret") or "")
    if not re.fullmatch(r"[0-9a-f]{64}", token_secret):
        token_secret = secrets.token_hex(32)
        auth["token_secret"] = token_secret
        changed = True
    versions = auth.get("versions")
    if not isinstance(versions, dict):
        versions = {}
        auth["versions"] = versions
        changed = True
    try:
        versions["admin"] = max(1, int(versions.get("admin") or 1))
    except Exception:
        versions["admin"] = 1
        changed = True
    staff_versions = versions.get("staff")
    if not isinstance(staff_versions, dict):
        staff_versions = {}
        versions["staff"] = staff_versions
        changed = True
    for slug in STAFF_CENTERS:
        try:
            staff_versions[slug] = max(1, int(staff_versions.get(slug) or 1))
        except Exception:
            staff_versions[slug] = 1
            changed = True
    if cfg.get("port") != PORT:
        cfg["port"] = PORT
        changed = True
    if changed:
        _atomic_write_webcfg(cfg)
    return {
        "admin": auth["admin"], "staff": staff,
        "token_secret": token_secret, "versions": versions, "port": PORT,
    }


PIN_STATE = load_pin_state()


def verify_pin(pin, staff_slug=""):
    with PIN_STATE_LOCK:
        record = (PIN_STATE.get("staff") or {}).get(staff_slug) if staff_slug \
            else PIN_STATE.get("admin")
        return _pin_matches(pin, record or {})


def set_role_pin(new_pin, staff_slug=""):
    """현재 역할 PIN을 변경하고 해당 역할의 저장된 기기 인증을 모두 무효화한다."""
    record = _pin_record(new_pin)
    with PIN_STATE_LOCK:
        if staff_slug:
            if staff_slug not in STAFF_CENTERS:
                raise ValueError("등록되지 않은 업무센터입니다")
            PIN_STATE.setdefault("staff", {})[staff_slug] = record
            versions = PIN_STATE.setdefault("versions", {}).setdefault("staff", {})
            versions[staff_slug] = int(versions.get(staff_slug) or 1) + 1
        else:
            PIN_STATE["admin"] = record
            versions = PIN_STATE.setdefault("versions", {})
            versions["admin"] = int(versions.get("admin") or 1) + 1
        if not DEMO:
            try:
                with open(WEBCFG, encoding="utf-8") as src:
                    cfg = json.load(src)
            except Exception:
                cfg = {}
            cfg.pop("pin", None)
            cfg["port"] = PORT
            cfg["auth"] = {
                "admin": PIN_STATE["admin"],
                "staff": PIN_STATE["staff"],
                "token_secret": PIN_STATE["token_secret"],
                "versions": PIN_STATE["versions"],
            }
            _atomic_write_webcfg(cfg)

# PIN은 로그인 수단일 뿐 권한이 아니다. 담당자 업무센터에서 로그인한 브라우저에는
# 서버가 서명한 HttpOnly 기기 토큰으로 역할을 고정한다. 토큰 서명키와 역할별 버전을
# 로컬 설정에 보존하므로 서버를 재시작해도 PC·모바일은 PIN을 다시 묻지 않는다.
# 앱을 열 때마다 만료일을 다시 늘리는 rolling 인증이며, PIN 변경 시에만 기존 기기를
# 무효화한다. 10년 TTL은 장기간 접속하지 않은 설치 앱까지 현실적으로 유지하기 위한
# 안전망이고, 정상적으로 앱을 여는 기기는 /api/auth/session 호출 때 계속 연장된다.
AUTH_SESSION_TTL = 10 * 365 * 24 * 60 * 60


def _b64url(data):
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64url_decode(text):
    raw = str(text or "")
    return base64.urlsafe_b64decode(raw + "=" * (-len(raw) % 4))


def _role_auth_version(staff_slug=""):
    versions = PIN_STATE.get("versions") or {}
    if staff_slug:
        return int((versions.get("staff") or {}).get(staff_slug) or 1)
    return int(versions.get("admin") or 1)


def create_auth_session(staff_slug="", tech_slug=""):
    # ★ tech 는 **세 번째 역할**이다. staff 로 만들면 그 순간 기사 링크 하나로
    #   원장 전체가 열린다(staff 는 `_auth()` 를 통과한다). 반드시 갈라 둔다.
    if tech_slug:
        if tech_slug not in AS_TECH_CENTERS:
            raise ValueError("등록되지 않은 기사입니다")
        role, staff_slug = "tech", tech_slug
    else:
        role = "staff" if staff_slug else "admin"
    if role == "staff" and staff_slug not in STAFF_CENTERS:
        raise ValueError("등록되지 않은 업무센터입니다")
    now = time.time()
    session = {
        "role": role,
        "staff_slug": staff_slug,
        "expires_at": int(now + AUTH_SESSION_TTL),
        "version": _role_auth_version(staff_slug),
    }
    payload = _b64url(json.dumps(
        session, ensure_ascii=True, separators=(",", ":"), sort_keys=True
    ).encode("utf-8"))
    secret = bytes.fromhex(PIN_STATE["token_secret"])
    signature = _b64url(hmac.new(secret, payload.encode("ascii"), hashlib.sha256).digest())
    return payload + "." + signature, session


def auth_cookie(token):
    return (f"csos_session={token}; Path=/; Max-Age={AUTH_SESSION_TTL}; "
            "HttpOnly; SameSite=Strict")


def auth_session_from_cookie(cookie_header):
    token = ""
    for part in str(cookie_header or "").split(";"):
        key, sep, value = part.strip().partition("=")
        if sep and key == "csos_session":
            token = value.strip()
            break
    if not token:
        return {}
    try:
        payload, signature = token.split(".", 1)
        secret = bytes.fromhex(PIN_STATE["token_secret"])
        expected = _b64url(hmac.new(
            secret, payload.encode("ascii"), hashlib.sha256
        ).digest())
        if not hmac.compare_digest(signature, expected):
            return {}
        session = json.loads(_b64url_decode(payload).decode("utf-8"))
        staff_slug = str(session.get("staff_slug") or "")
        role = str(session.get("role") or "")
        if role not in ("admin", "staff", "tech"):
            return {}
        if role == "staff" and staff_slug not in STAFF_CENTERS:
            return {}
        if role == "tech" and staff_slug not in AS_TECH_CENTERS:
            return {}
        if role == "admin" and staff_slug:
            return {}
        if int(session.get("expires_at") or 0) <= int(time.time()):
            return {}
        if int(session.get("version") or 0) != _role_auth_version(staff_slug):
            return {}
        return session
    except Exception:
        return {}


# ── 담당자 실시간 접속(조직도 온/오프라인) — 2026-08-12 지시 ──────────────────
#   "각 담당자 온라인인지 오프라인인지 표시." 앱은 사람이 지금 접속했는지를
#   **지어내지 않는다** — real signal 은 '인증된 브라우저가 서버에 마지막으로 말한
#   시각' 하나뿐이다. auth 쿠키는 매 요청에 실려 오므로 _auth()·기사 경로에서 그
#   시각을 찍고, get_orgchart 가 읽어 온/오프라인을 정한다. 기록이 아예 없으면
#   '오프라인'이 아니라 '상태 모름'이다([169]). 클라이언트(index.html)는 안 고친다.
_PRESENCE_PATH = os.path.join(ROOT, "reports", ".담당자_접속.json")
_PRESENCE_LOCK = threading.Lock()
_PRESENCE = None          # {slug: last_ts} — 프로세스 메모리(늘 최신)
_PRESENCE_FLUSH = [0.0]   # 디스크 쓰기 throttle(폴링 thrash 방지)

# ★ 관리자 로그인은 **센터장(유현민) 자리**다 (2026-08-12 지시: "유현민은 여기에
#   접속되어있으면 온라인으로 항상 표기해야지"). 전에는 role=="staff" 만 찍어서,
#   관리자로 앱을 열면 **아무 자리도 안 찍혔다** — 본인이 화면을 보고 있는데 제
#   자리만 '앱 접속 기록 없음'이었다. 근거는 로스터다: '센터장' 딱지가 붙는 자리가
#   여기 하나뿐이고, 관리자 PIN 을 쥔 사람이 그 사람이다.
#   ⚠ **사람이 로그인한 쿠키 세션만** 찍는다. localhost 스크립트가 X-Pin 헤더로
#   부르는 길(_auth 아래쪽)은 사람이 아니다 — 거기서 찍으면 워치독이 도는 내내
#   유현민이 '온라인'으로 보인다. 그건 잰 것이 아니라 지어낸 신호다([169]).
ADMIN_PRESENCE_SLUG = "yoo-hyeonmin"


def _presence_map():
    global _PRESENCE
    if _PRESENCE is None:
        try:
            raw = json.load(open(_PRESENCE_PATH, encoding="utf-8"))
            _PRESENCE = {str(k): float(v) for k, v in raw.items()}
        except Exception:
            _PRESENCE = {}
    return _PRESENCE


def presence_touch(slug):
    """담당자(staff_slug/기사 slug)의 인증 요청 시각을 찍는다 — 지어내지 않는다."""
    slug = str(slug or "").strip()
    if not slug:
        return
    now = time.time()
    with _PRESENCE_LOCK:
        p = _presence_map()
        p[slug] = now
        if now - _PRESENCE_FLUSH[0] < 30:   # 메모리는 늘 최신, 디스크는 30초마다
            return
        _PRESENCE_FLUSH[0] = now
        try:
            tmp = _PRESENCE_PATH + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(p, f, ensure_ascii=False)
            os.replace(tmp, _PRESENCE_PATH)
        except Exception:
            pass


def presence_snapshot():
    """{slug: last_ts} 사본. get_orgchart 가 읽는다."""
    with _PRESENCE_LOCK:
        return dict(_presence_map())


# 앱에서 확정한 운영기준은 관리대장 수식과 섞지 않고 작은 런타임 DB로 보관한다.
# reports/는 git 제외 대상이며, 저장 성공 직후 대표보고·확인필요 화면이 같은 값을 읽는다.
POLICY_FILE = os.path.join(ROOT, "reports", "operating_policies.json")

# ── 크롬 전용 수집 되보고 (2026-08-13 지시 '앞으로 크롬에서만 긁어오는 알고리즘') ──
#   `band/band_auto_collect.user.js` 가 밴드 탭에서 `POST /api/collect_report` 로 보낸다.
#   읽는 자리는 `band/userscript_watch.py` 하나다 — 칸 이름을 여기서 지어내면
#   감시자가 오류 없이 한 건도 못 읽는다([165]).  그래서 그쪽이 읽는 모양 그대로 쌓는다:
#       {"밴드": {"<밴드번호>": {state, at, 요청, 수확, why, …}}, "최근": [ …최대 100건… ]}
COLLECT_REPORT_PATH = os.path.join(ROOT, "reports", "크롬수집_보고.json")
#: 되보고는 밴드마다 따로 온다 — 읽고-고쳐-쓰기가 겹치면 한쪽 밴드가 통째로 사라진다.
_COLLECT_REPORT_LOCK = threading.Lock()
#: 남기는 최근 이력 건수.  갈래 판정은 '밴드'만 보므로 이것은 사람이 뒤져 볼 몫이다.
COLLECT_REPORT_KEEP = 100


def collect_report_save(rec):
    """되보고 한 건을 정본 파일에 합친다.  실패하면 예외를 올린다 — 조용히 넘기지 않는다.

    ⚠ 밴드번호가 밴드번호처럼 안 생겼으면 **거부**한다.  `260807`(날짜 꼬리표)·
      `202608082047`(시각 도장) 같은 값이 한 번 키로 들어가면 유령 밴드가 생기고,
      그것이 '아는 번호'가 되어 스스로를 되살린다(2026-08-12 실사고).
    """
    band = re.sub(r"\D", "", str((rec or {}).get("band") or ""))
    if not (7 <= len(band) <= 10):
        raise ValueError("밴드번호가 아니다: %r" % ((rec or {}).get("band"),))
    row = {k: v for k, v in dict(rec).items() if k != "band"}
    row.setdefault("at", datetime.now().isoformat(timespec="seconds"))
    row["받은시각"] = datetime.now().isoformat(timespec="seconds")
    with _COLLECT_REPORT_LOCK:
        doc = {}
        try:
            with open(COLLECT_REPORT_PATH, encoding="utf-8") as fh:
                doc = json.load(fh) or {}
        except FileNotFoundError:
            doc = {}
        except Exception:
            # 깨진 파일을 근거로 삼지 않는다.  옆에 치워 두고 새로 시작한다 —
            # 지우면 "그때 무엇이 와 있었나"를 잃는다([186] 과 같은 자리).
            try:
                os.replace(COLLECT_REPORT_PATH,
                           COLLECT_REPORT_PATH + ".broken-%s" % datetime.now().strftime("%Y%m%d%H%M%S"))
            except OSError:
                pass
            doc = {}
        bands = doc.get("밴드")
        if not isinstance(bands, dict):
            bands = {}
        bands[band] = row
        recent = doc.get("최근")
        if not isinstance(recent, list):
            recent = []
        recent.append(dict(row, band=band))
        doc["밴드"] = bands
        doc["최근"] = recent[-COLLECT_REPORT_KEEP:]
        os.makedirs(os.path.dirname(COLLECT_REPORT_PATH), exist_ok=True)
        tmp = COLLECT_REPORT_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(doc, fh, ensure_ascii=False, indent=1)
        os.replace(tmp, COLLECT_REPORT_PATH)
    return {"band": band, "밴드수": len(bands), "최근": len(doc["최근"])}


def load_policy_state():
    try:
        data = json.load(open(POLICY_FILE, encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_policy_state(key, value):
    os.makedirs(os.path.dirname(POLICY_FILE), exist_ok=True)
    data = load_policy_state()
    data[str(key)] = value
    tmp = POLICY_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, POLICY_FILE)
    return data


def multipart_parts(content_type, raw):
    """표준 라이브러리만으로 multipart/form-data를 안전하게 푼다."""
    if "multipart/form-data" not in str(content_type or ""):
        raise ValueError("multipart/form-data 형식이 아닙니다")
    head = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("ascii", "ignore")
    msg = BytesParser(policy=email_policy.default).parsebytes(head + raw)
    fields, files = {}, {}
    for part in msg.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        data = part.get_payload(decode=True) or b""
        filename = part.get_filename()
        if filename:
            files[name] = {"filename": os.path.basename(str(filename)), "data": data,
                           "content_type": part.get_content_type()}
        else:
            charset = part.get_content_charset() or "utf-8"
            try:
                fields[name] = data.decode(charset, errors="replace").strip()
            except LookupError:
                fields[name] = data.decode("utf-8", errors="replace").strip()
    return fields, files


def _safe_upload_name(name):
    name = os.path.basename(str(name or "")).strip()
    name = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', "_", name)
    return name[:160] or "kakao.txt"


PO_SOURCE_EXTS = {
    ".xlsx", ".xls", ".csv", ".pdf", ".png", ".jpg", ".jpeg", ".webp",
    ".zip", ".txt", ".eml",
}
WORK_LOG_SOURCE_EXTS = {".xlsx"}


def _path_is_under(path, roots):
    """Resolve a local path and allow it only below explicitly approved source roots."""
    try:
        candidate = os.path.normcase(os.path.realpath(os.path.abspath(path)))
    except (OSError, TypeError, ValueError):
        return False
    for root in roots:
        try:
            root_real = os.path.normcase(os.path.realpath(os.path.abspath(root)))
            if os.path.commonpath([candidate, root_real]) == root_real:
                return True
        except (OSError, TypeError, ValueError):
            continue
    return False


def _approved_source_roots():
    from source_dirs import ORIGIN_ROOT, PO_DIRS
    home = os.path.expanduser("~")
    roots = [ORIGIN_ROOT, os.path.join(home, "Desktop"), os.path.join(home, "Downloads")]
    roots.extend(PO_DIRS)
    return [root for root in roots if root]


def _validate_remote_url(url):
    parsed = urlsplit(str(url or "").strip())
    if parsed.scheme not in ("http", "https") or not parsed.hostname:
        raise ValueError("URL은 http:// 또는 https:// 주소여야 합니다")
    if parsed.username or parsed.password:
        raise ValueError("계정정보가 포함된 URL은 사용할 수 없습니다")
    try:
        parsed_port = parsed.port
    except ValueError as exc:
        raise ValueError("URL 포트 형식이 올바르지 않습니다") from exc
    if parsed_port not in (None, 80, 443):
        raise ValueError("URL은 표준 웹 포트만 사용할 수 있습니다")
    try:
        infos = socket.getaddrinfo(parsed.hostname, parsed_port or (443 if parsed.scheme == "https" else 80))
    except socket.gaierror as exc:
        raise ValueError(f"URL 서버 주소를 확인할 수 없습니다: {exc}") from exc
    for info in infos:
        ip = ipaddress.ip_address(info[4][0].split("%", 1)[0])
        if not ip.is_global:
            raise ValueError("내부망·로컬 주소 URL은 보안상 다운로드할 수 없습니다")
    return parsed


class _NoRedirect(HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):
        return None


def _download_direct_file(url, allowed_exts, max_bytes=55_000_000):
    """Download a public direct-file URL, validating every redirect against SSRF."""
    opener = build_opener(_NoRedirect)
    current = str(url or "").strip()
    for _hop in range(4):
        parsed = _validate_remote_url(current)
        request = Request(
            current,
            headers={
                "User-Agent": "CSOS/2026 PO-source-import",
                "Accept": "application/octet-stream,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.5",
            },
        )
        try:
            response = opener.open(request, timeout=25)
        except HTTPError as exc:
            if exc.code in (301, 302, 303, 307, 308) and exc.headers.get("Location"):
                current = urljoin(current, exc.headers["Location"])
                continue
            raise ValueError(f"URL 다운로드 실패(HTTP {exc.code})") from exc
        final_url = response.geturl()
        _validate_remote_url(final_url)
        disposition = response.headers.get("Content-Disposition") or ""
        match = re.search(r"filename\*?=(?:UTF-8''|[\"']?)([^\"';]+)", disposition, re.I)
        name = _safe_upload_name(match.group(1) if match else os.path.basename(urlsplit(final_url).path))
        ext = os.path.splitext(name)[1].lower()
        if ext not in allowed_exts:
            response.close()
            raise ValueError("URL이 허용된 원본 파일을 직접 가리키지 않습니다")
        data = response.read(max_bytes + 1)
        response.close()
        if len(data) > max_bytes:
            raise ValueError(f"URL 파일은 {max_bytes // 1_000_000}MB 이하여야 합니다")
        if not data:
            raise ValueError("URL에서 받은 파일이 비어 있습니다")
        return name, data, final_url
    raise ValueError("URL 리디렉션이 너무 많습니다")


def _unique_path(folder, name):
    name = _safe_upload_name(name)
    base, ext = os.path.splitext(name)
    candidate = os.path.join(folder, name)
    index = 2
    while os.path.exists(candidate):
        candidate = os.path.join(folder, f"{base}_{index}{ext}")
        index += 1
    return candidate


def _copy_local_reference(source_ref, destination, allowed_exts, max_files=250,
                          max_total=250_000_000):
    source = os.path.expandvars(str(source_ref or "").strip().strip("\"'"))
    if not source:
        return []
    source = os.path.abspath(source)
    if not os.path.exists(source):
        raise ValueError("붙여넣은 파일·폴더 경로를 이 PC에서 찾을 수 없습니다")
    if not _path_is_under(source, _approved_source_roots()):
        raise ValueError("원본자료·바탕화면·다운로드 또는 승인된 PO 폴더의 경로만 가져올 수 있습니다")
    candidates = [source] if os.path.isfile(source) else [
        os.path.join(parent, name)
        for parent, _dirs, names in os.walk(source)
        for name in names
    ]
    saved, total = [], 0
    for path in candidates:
        if len(saved) >= max_files:
            raise ValueError(f"한 번에 가져올 수 있는 파일은 최대 {max_files}개입니다")
        if not _path_is_under(path, _approved_source_roots()):
            continue
        ext = os.path.splitext(path)[1].lower()
        if ext not in allowed_exts or os.path.basename(path).startswith("~$"):
            continue
        try:
            size = os.path.getsize(path)
        except OSError:
            continue
        total += size
        if total > max_total:
            raise ValueError(f"한 번에 가져올 수 있는 총 용량은 {max_total // 1_000_000}MB입니다")
        dest = _unique_path(destination, os.path.basename(path))
        shutil.copy2(path, dest)
        saved.append(dest)
    if not saved:
        raise ValueError("경로 안에서 지원되는 원본 파일을 찾지 못했습니다")
    return saved


def _save_source_submission(fields, files, *, kind, allowed_exts, destination_root,
                            upload_field, staff_slug):
    """Store file/path/URL submissions in an immutable dated source folder."""
    if str(fields.get("staff_slug") or "").strip() != staff_slug:
        raise ValueError("이 자료 등록은 지정된 담당자 업무센터에서만 사용할 수 있습니다")
    now = datetime.now()
    project = re.sub(r"[^0-9A-Za-z가-힣_-]", "_",
                     str(fields.get("project_no") or fields.get("po_no") or kind).strip())[:48]
    folder = os.path.join(
        destination_root, f"{now:%Y}", f"{now:%m}", f"{now:%Y-%m-%d}",
        f"{now:%H%M%S}_{project or kind}",
    )
    if DEMO:
        folder = os.path.join(ROOT, "tmp", "demo-source-submission", kind, f"{now:%H%M%S}")
    os.makedirs(folder, exist_ok=True)
    saved = []
    upload = files.get(upload_field)
    if upload and upload.get("data"):
        if len(upload["data"]) > 55_000_000:
            raise ValueError("첨부파일은 55MB 이하만 가능합니다")
        ext = os.path.splitext(upload.get("filename") or "")[1].lower()
        if ext not in allowed_exts:
            raise ValueError("지원되지 않는 원본 파일 형식입니다")
        dest = _unique_path(folder, upload.get("filename"))
        with open(dest, "wb") as out:
            out.write(upload["data"])
        saved.append(dest)
    source_ref = str(fields.get("source_ref") or "").strip()
    final_url = ""
    if source_ref:
        if re.match(r"^https?://", source_ref, re.I):
            name, data, final_url = _download_direct_file(source_ref, allowed_exts)
            dest = _unique_path(folder, name)
            with open(dest, "wb") as out:
                out.write(data)
            saved.append(dest)
        else:
            saved.extend(_copy_local_reference(source_ref, folder, allowed_exts))
    if not saved:
        raise ValueError("파일을 선택하거나 URL·파일·폴더 경로를 붙여넣어 주세요")
    manifest = {
        "kind": kind,
        "registered_at": now.isoformat(timespec="seconds"),
        "staff_slug": staff_slug,
        "staff": STAFF_CENTERS[staff_slug]["name"],
        "source_ref": source_ref,
        "resolved_url": final_url,
        "project_no": str(fields.get("project_no") or "").strip(),
        "po_no": str(fields.get("po_no") or "").strip(),
        "memo": str(fields.get("memo") or "").strip(),
        "files": [os.path.basename(path) for path in saved],
    }
    with open(os.path.join(folder, "submission.json"), "w", encoding="utf-8") as out:
        json.dump(manifest, out, ensure_ascii=False, indent=2)
    return folder, saved, manifest


def save_staff_po_submission(fields, files, source_ip=""):
    from source_dirs import PO_DIR
    folder, saved, manifest = _save_source_submission(
        fields, files, kind="po", allowed_exts=PO_SOURCE_EXTS,
        destination_root=PO_DIR, upload_field="po_file", staff_slug="oh-jonghyeon",
    )
    inbox = os.path.join(ROOT, "inbox")
    os.makedirs(inbox, exist_ok=True)
    queued_files = []
    for path in saved:
        if os.path.splitext(path)[1].lower() != ".xlsx":
            continue
        dest = _unique_path(inbox, "PO_오종현_" + os.path.basename(path))
        shutil.copy2(path, dest)
        queued_files.append(os.path.basename(dest))
    started = queued = False
    msg = "PO 원본을 보관했습니다"
    if queued_files:
        started, msg = start_task("po")
        queued = False if started else defer_task_until_free("po")
    manifest["source_ip"] = source_ip
    manifest["po_compare_files"] = queued_files
    return {
        "folder": folder, "files": [os.path.basename(x) for x in saved],
        "po_compare_files": queued_files, "auto_check_started": started,
        "auto_check_queued": queued, "msg": msg,
    }


RECEIPT_SOURCE_EXTS = {".xlsx", ".xls", ".csv", ".png", ".jpg", ".jpeg", ".pdf"}


def save_staff_receipt_submission(fields, files, source_ip=""):
    """입금내역 파일·사진·URL을 '7. 입금내역'(Z:)에 보관하고 즉시 대조를 돌린다.

    사용자 지시(2026-07-31): 오종현 업무센터에서 드래그앤드롭·첨부·URL 등록 →
    지정 저장소(Z: 원본 자료)와 DB에 반영, 바로 앱에 보이게.
    저장은 _save_source_submission(불변 날짜 폴더 + URL 은 SSRF 검증 다운로드) 그대로.
    xlsx 는 inbox 로도 복사해 receipt_fill 이 내용 판별로 집어가게 한다 —
    대조 결과는 DB 큐(11·15시 반영)와 입금현황 리포트로 즉시 나타난다."""
    fields = dict(fields or {})
    fields["staff_slug"] = "oh-jonghyeon"          # 관리자 대리 등록도 오종현 자료로 보관
    folder, saved, manifest = _save_source_submission(
        fields, files, kind="receipt", allowed_exts=RECEIPT_SOURCE_EXTS,
        destination_root=__import__("source_dirs").RECEIPT_DIR,
        upload_field="receipt_file", staff_slug="oh-jonghyeon",
    )
    # csv 는 xlsx 변환본을 같은 폴더에 함께 둔다 — receipt_fill.load_deposits 는
    # *.xlsx 만 읽으므로, 변환 없이 두면 '저장은 되는데 영영 안 읽히는' 자료가 된다
    # (읽지 않는 파일은 빈칸과 같다 — receipt_apply 와 같은 변환기를 그대로 쓴다).
    # 원본 csv 는 지우지 않는다. xls 는 변환기가 없다 — 못 읽는다고 정직하게 말한다.
    warn = ""
    converted = []
    for path in list(saved):
        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            from receipt_apply import _csv_to_xlsx
            tmp_x = _csv_to_xlsx(path)
            dest = _unique_path(folder, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
            shutil.move(tmp_x, dest)
            converted.append(dest)
        elif ext == ".xls":
            warn = " · ⚠ xls 는 자동으로 읽히지 않습니다 — xlsx 로 다시 저장해 올려 주세요"
    if converted:
        saved = saved + converted
        manifest["files"] = [os.path.basename(p) for p in saved]
        manifest["converted"] = [os.path.basename(p) for p in converted]
        with open(os.path.join(folder, "submission.json"), "w", encoding="utf-8") as out:
            json.dump(manifest, out, ensure_ascii=False, indent=2)
    inbox = os.path.join(ROOT, "inbox")
    os.makedirs(inbox, exist_ok=True)
    queued_files = []
    for path in saved:
        if os.path.splitext(path)[1].lower() not in (".xlsx", ".xls", ".csv"):
            continue
        dest = _unique_path(inbox, "입금_오종현_" + os.path.basename(path))
        shutil.copy2(path, dest)
        queued_files.append(os.path.basename(dest))
    started = queued = False
    msg = "입금 자료를 보관했습니다" + (
        " · csv→xlsx 변환 %d건" % len(converted) if converted else "") + warn
    if queued_files:
        started, msg = start_task("receipt")
        queued = False if started else defer_task_until_free("receipt")
    manifest["source_ip"] = source_ip
    return {"folder": folder, "files": [os.path.basename(x) for x in saved],
            "auto_check_started": started, "auto_check_queued": queued, "msg": msg}


def save_staff_work_log_submission(fields, files, source_ip=""):
    from source_dirs import WORK_LOG_DIR
    if str(fields.get("staff_slug") or "").strip() != "ryu-jiyeong":
        raise ValueError("대표보고 일지는 류지영 업무센터에서만 등록할 수 있습니다")
    if str(fields.get("use_current") or "").strip() == "1":
        from work_log_sync import find_latest_source
        current = find_latest_source()
        folder, saved = os.path.dirname(current), [current]
    else:
        folder, saved, _manifest = _save_source_submission(
            fields, files, kind="work-log", allowed_exts=WORK_LOG_SOURCE_EXTS,
            destination_root=WORK_LOG_DIR, upload_field="work_log_file", staff_slug="ryu-jiyeong",
        )
    started, msg = start_task("work_log")
    queued = False if started else defer_task_until_free("work_log")
    return {
        "folder": folder, "files": [os.path.basename(x) for x in saved],
        "auto_check_started": started, "auto_check_queued": queued, "msg": msg,
    }


def _kakao_text_kind(data):
    """두 카톡방 파일이 서로 바뀌어 올라와도 본문 제목으로 자동 분류한다."""
    text = ""
    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise ValueError("카카오톡 텍스트 인코딩을 읽을 수 없습니다")
    head = text[:1500]
    if "쿠팡정기점검" in head:
        return "정기점검", text
    if "쿠팡돌발점검" in head:
        return "돌발점검", text
    raise ValueError("파일 첫 부분에서 ‘쿠팡정기점검’ 또는 ‘쿠팡돌발점검’ 대화방을 확인하지 못했습니다")


def save_ryu_upload(fields, files):
    """류지영 업무센터 업로드를 원본 자료에 보존하고 카톡 자동대조 inbox로 넘긴다."""
    from source_dirs import KAKAO_DIR
    needed = [files.get("kakao_regular"), files.get("kakao_emergency")]
    if any(not x for x in needed):
        raise ValueError("정기점검방과 돌발점검방 텍스트 파일을 각각 첨부해 주세요")
    found = {}
    parsed = []
    for f in needed:
        if not f["filename"].lower().endswith(".txt"):
            raise ValueError("카카오톡 대화내역은 .txt 파일만 첨부할 수 있습니다")
        if len(f["data"]) > 20_000_000:
            raise ValueError("카카오톡 텍스트 파일은 각 20MB 이하만 가능합니다")
        kind, _text = _kakao_text_kind(f["data"])
        if kind in found:
            raise ValueError(f"{kind} 대화방 파일이 두 번 첨부되었습니다")
        found[kind] = f
    if set(found) != {"정기점검", "돌발점검"}:
        raise ValueError("정기점검방·돌발점검방 두 종류가 모두 필요합니다")
    evidence = files.get("evidence_file")
    if evidence and evidence.get("data"):
        if len(evidence["data"]) > 25_000_000:
            raise ValueError("추가 근거 파일은 25MB 이하만 가능합니다")
        evidence_ext = os.path.splitext(evidence["filename"])[1].lower()
        if evidence_ext not in (".png", ".jpg", ".jpeg", ".webp", ".pdf", ".xlsx", ".docx", ".txt"):
            raise ValueError("추가 근거는 이미지·PDF·Excel·Word·텍스트 파일만 가능합니다")

    now = datetime.now()
    day_dir = os.path.join(KAKAO_DIR, f"{now:%Y}", f"{now:%m}", f"{now:%Y-%m-%d}")
    inbox = os.path.join(ROOT, "kakao", "inbox")
    os.makedirs(day_dir, exist_ok=True)
    os.makedirs(inbox, exist_ok=True)
    stamp = now.strftime("%Y%m%d_%H%M%S")
    saved = []
    for kind in ("정기점검", "돌발점검"):
        f = found[kind]
        original = _safe_upload_name(f["filename"])
        name = f"(류지영)_{kind}_{stamp}_{original}"
        dest = os.path.join(day_dir, name)
        with open(dest, "wb") as out:
            out.write(f["data"])
        # 자동대조 도구는 로컬 inbox의 txt를 즉시 읽는다. 원본은 Z:에 보존하고,
        # 작은 텍스트 사본만 로컬에 두어 PC 용량과 대조 속도를 모두 지킨다.
        inbox_path = os.path.join(inbox, name)
        shutil.copy2(dest, inbox_path)
        try:
            from kakao.kakao_reconcile import parse_export
            count = len(parse_export(dest))
        except Exception:
            count = 0
        saved.append({"방": kind, "파일": name, "메시지": count})

    evidence_saved = ""
    if evidence and evidence.get("data"):
        evidence_saved = f"(류지영)_추가근거_{stamp}_{_safe_upload_name(evidence['filename'])}"
        with open(os.path.join(day_dir, evidence_saved), "wb") as out:
            out.write(evidence["data"])

    manifest = {
        "등록일시": now.isoformat(timespec="seconds"),
        "등록자": fields.get("submitter") or "류지영",
        "조사기준일": fields.get("survey_date") or f"{now:%Y-%m-%d}",
        "조사메모": fields.get("survey_note") or "",
        "업무구분": fields.get("work_kind") or "",
        "프로젝트NO": fields.get("project_no") or "",
        "캠프명": fields.get("camp_name") or "",
        "담당자": fields.get("assignee") or "",
        "처리상태": fields.get("work_status") or "",
        "완료일": fields.get("completed_date") or "",
        "조치내용": fields.get("action_note") or "",
        "추가근거": evidence_saved,
        "파일": saved,
    }
    with open(os.path.join(day_dir, f"(류지영)_업로드기록_{stamp}.json"), "w", encoding="utf-8") as out:
        json.dump(manifest, out, ensure_ascii=False, indent=2)
    os.makedirs(os.path.join(ROOT, "reports"), exist_ok=True)
    with open(os.path.join(ROOT, "reports", "ryu_submissions.jsonl"), "a", encoding="utf-8") as out:
        out.write(json.dumps(manifest, ensure_ascii=False) + "\n")
    return manifest


# 류지영 업무센터의 입력 항목은 원장의 수식·검증 열을 직접 건드리지 않는다.
# 아래 화이트리스트에 있는 "사람이 확인해서 보충하는 원천 열"만 빈 칸에 한해 기록한다.
RYU_ENTRY_CONFIG = {
    "as": {
        "label": "돌발AS", "sheet": "02_돌발AS접수", "key_col": "접수ID",
        "date_col": "접수일자", "kind": "as",
        "fields": [
            {"name": "담당기사", "label": "담당기사", "type": "text"},
            {"name": "진행상태", "label": "진행 상태", "type": "text"},
            {"name": "방문예정일", "label": "방문예정일", "type": "date"},
            {"name": "방문예정시간", "label": "방문예정시간", "type": "text"},
            {"name": "작업완료일", "label": "작업완료일", "type": "date"},
            {"name": "재방문여부", "label": "재방문 여부", "type": "select",
             "options": ["예", "아니오"]},
            {"name": "재방문예정일", "label": "재방문 예정일", "type": "date"},
            {"name": "유상·무상·보험", "label": "비용 구분", "type": "select",
             "options": ["유상", "무상", "보험"]},
            {"name": "사진등록", "label": "사진 등록", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "동영상등록", "label": "동영상 등록", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "완료보고서등록", "label": "완료보고서 등록", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "문제내용", "label": "문제 내용", "type": "textarea"},
            {"name": "조치내용", "label": "조치 내용", "type": "textarea"},
            {"name": "완료예정일", "label": "완료 예정일", "type": "date"},
            {"name": "관리자검증상태", "label": "관리자 검증 상태", "type": "text"},
            {"name": "최종확인일", "label": "최종 확인일", "type": "date"},
            # ★ 청구 제외 — 삭제와 다른 손잡이다 (2026-08-13 류지영 요청).
            #   "정기점검이랑 동시진행으로 돌발AS로는 청구를 안할꺼라". 현장은 실제로
            #   다녀왔으므로 삭제도 취소도 사실이 아니다 — 청구에서만 뺀다.
            {"name": "청구제외", "label": "청구 제외(다녀옴·이 건으로는 청구 안 함)",
             "type": "select", "options": ["예", "아니오"]},
            {"name": "청구제외사유", "label": "청구 제외 사유", "type": "textarea"},
            {"name": "비고", "label": "비고", "type": "textarea"},
        ],
    },
    "pm": {
        "label": "정기점검", "sheet": "04_정기점검", "key_col": "점검ID",
        "date_col": "점검예정일", "kind": "pm",
        "fields": [
            {"name": "점검상태", "label": "점검 상태", "type": "text"},
            {"name": "점검예정일", "label": "점검 예정일", "type": "date"},
            {"name": "점검예정시간", "label": "점검 예정시간", "type": "text"},
            {"name": "담당기사", "label": "담당기사", "type": "text"},
            {"name": "실제점검일", "label": "실제 점검일", "type": "date"},
            {"name": "점검내용", "label": "점검 내용", "type": "textarea"},
            # ★ 원장이 실제로 쓰는 값은 '있음/없음'이다(daily_brief 도 '있음'만 센다).
            #   선택지를 '예/아니오'로 두면 화면에서 고쳐도 집계가 안 바뀐다 — 2026-08-04 발견.
            {"name": "이상발견여부", "label": "이상 발견 여부", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "이상내용", "label": "이상 내용", "type": "textarea"},
            {"name": "돌발AS전환여부", "label": "돌발AS 전환 여부", "type": "select",
             "options": ["예", "아니오"]},
            {"name": "유상추가작업발생", "label": "유상 추가작업 발생", "type": "select",
             "options": ["예", "아니오"]},
            {"name": "추가작업내용", "label": "추가작업 내용", "type": "textarea"},
            {"name": "유상·무상·보험", "label": "비용 구분", "type": "select",
             "options": ["유상", "무상", "보험"]},
            {"name": "점검사진", "label": "점검 사진", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "점검동영상", "label": "점검 동영상", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "점검보고서", "label": "점검 보고서", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "문제내용", "label": "문제 내용", "type": "textarea"},
            {"name": "조치내용", "label": "조치 내용", "type": "textarea"},
            {"name": "최종확인일(유현민 체크)", "label": "최종 확인일", "type": "date"},
            # 돌발AS 와 같은 손잡이를 정기점검에도 둔다 — 반대 방향(돌발AS 로 청구하고
            # 정기점검은 제외)도 실재한다. 지금 안 쓰여도 생기는 날 저절로 걸린다.
            {"name": "청구제외", "label": "청구 제외(다녀옴·이 건으로는 청구 안 함)",
             "type": "select", "options": ["예", "아니오"]},
            {"name": "청구제외사유", "label": "청구 제외 사유", "type": "textarea"},
            {"name": "비고", "label": "비고", "type": "textarea"},
        ],
    },
    "field": {
        "label": "현장작업", "sheet": "03_현장작업실적", "key_col": "작업ID",
        "date_col": "작업일자", "kind": "field",
        "fields": [
            {"name": "작업일자", "label": "작업일자", "type": "date"},
            {"name": "작업시작시간", "label": "작업 시작시간", "type": "text"},
            {"name": "작업종료시간", "label": "작업 종료시간", "type": "text"},
            {"name": "담당기사", "label": "담당기사", "type": "text"},
            {"name": "실제작업항목", "label": "실제 작업항목", "type": "textarea"},
            {"name": "실제작업상세", "label": "실제 작업상세", "type": "textarea"},
            {"name": "접수외추가작업여부", "label": "접수 외 추가작업", "type": "select",
             "options": ["예", "아니오"]},
            {"name": "추가작업내용", "label": "추가작업 내용", "type": "textarea"},
            {"name": "사용부품", "label": "사용 부품", "type": "text"},
            {"name": "수량", "label": "수량", "type": "number"},
            {"name": "비용구분", "label": "비용 구분", "type": "select",
             "options": ["유상", "무상", "보험"]},
            {"name": "완료여부", "label": "완료 여부", "type": "select",
             "options": ["완료", "미완료"]},
            {"name": "재방문필요", "label": "재방문 필요", "type": "select",
             "options": ["예", "아니오"]},
            {"name": "재방문사유", "label": "재방문 사유", "type": "textarea"},
            {"name": "작업사진", "label": "작업 사진", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "작업동영상", "label": "작업 동영상", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "완료보고서", "label": "완료보고서", "type": "select",
             "options": ["있음", "없음"]},
            {"name": "기사보고내용", "label": "기사 보고내용", "type": "textarea"},
            {"name": "문제내용", "label": "문제 내용", "type": "textarea"},
            {"name": "조치내용", "label": "조치 내용", "type": "textarea"},
            {"name": "완료예정일", "label": "완료 예정일", "type": "date"},
            {"name": "비고", "label": "비고", "type": "textarea"},
        ],
    },
    "settle": {
        "label": "거래서류·청구", "sheet": "06_거래서류청구수금", "key_col": "정산ID",
        "date_col": "작업완료일", "kind": "settle",
        "fields": [
            {"name": "거래명세서번호", "label": "거래명세서 번호", "type": "text"},
            {"name": "거래명세서발행일", "label": "거래명세서 발행일", "type": "date"},
            {"name": "PO필요여부", "label": "PO 필요 여부", "type": "select",
             "options": ["예", "아니오"]},
            {"name": "PO번호", "label": "PO 번호", "type": "text"},
            {"name": "PO발행일", "label": "PO 발행일", "type": "date"},
            {"name": "세금계산서발행일", "label": "세금계산서 발행일", "type": "date"},
            {"name": "청구일", "label": "청구일", "type": "date"},
            {"name": "지급예정일", "label": "지급 예정일", "type": "date"},
            {"name": "입금일", "label": "입금일", "type": "date"},
            {"name": "입금액", "label": "입금액", "type": "number"},
            {"name": "담당자", "label": "담당자", "type": "text"},
            {"name": "문제내용", "label": "문제 내용", "type": "textarea"},
            {"name": "조치내용", "label": "조치 내용", "type": "textarea"},
            {"name": "완료예정일", "label": "완료 예정일", "type": "date"},
            {"name": "비고", "label": "비고", "type": "textarea"},
        ],
    },
}


# 직원 화면·API·보관본이 같은 열 권한을 본다. 화면에서 입력칸을 숨기는 것은 보안이
# 아니므로 서버가 이 표 밖의 열을 항상 거부한다. 자동 판정·수식 열은 애초에 넣지 않는다.
_ALL_STAFF_ENTRY_FIELDS = {
    category: {field["name"] for field in cfg["fields"]}
    for category, cfg in RYU_ENTRY_CONFIG.items()
}
STAFF_ENTRY_PERMISSIONS = {
    "ryu-jiyeong": {key: set(value) for key, value in _ALL_STAFF_ENTRY_FIELDS.items()},
    "yoo-hyeonmin": {key: set(value) for key, value in _ALL_STAFF_ENTRY_FIELDS.items()},
    "oh-jonghyeon": {
        # 오종현 통화 건처럼 원천 AS/점검의 비용 사실을 바로잡을 수 있다. 06시트
        # 비용구분은 원천업무를 참조하는 수식이므로 직접 쓰지 않는다.
        "as": {"유상·무상·보험", "문제내용", "조치내용", "비고"},
        "pm": {"유상·무상·보험", "문제내용", "조치내용", "비고"},
        "settle": set(_ALL_STAFF_ENTRY_FIELDS["settle"]),
    },
    # 김미영은 **청구·수금 쪽만** 손댄다. 현장 사실(as·pm)은 다녀온 사람이 적는 것이라
    # 여기서 열어 주면 서류를 보고 현장 기록을 고치게 된다 — 그 순간 근거가 뒤집힌다.
    "kim-miyeong": {
        "settle": {"거래명세서번호", "거래명세서발행일", "PO필요여부", "PO번호", "PO발행일",
                   "세금계산서발행일", "청구일", "지급예정일", "입금일", "입금액",
                   "문제내용", "조치내용", "비고"},
    },
}
STAFF_REASON_REQUIRED_FIELDS = {
    "유상·무상·보험", "비용구분", "거래명세서번호", "PO번호",
    "세금계산서발행일", "입금액",
}
STAFF_DERIVED_FIELDS = {
    "계산서", "발행상태", "발행상태(자동)", "청구상태", "수금상태",
}


def _ryu_display_value(value):
    if value in (None, ""):
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _ryu_field_records():
    """03_현장작업실적의 실제 행을 읽기 전용으로 반환한다."""
    if DEMO:
        return [
            {
                "작업ID": f"FW-2607-{i:03d}", "접수ID": f"AS-2607-{i:03d}",
                "프로젝트NO": f"UJ26{2000+i:05d}", "캠프명": f"데모{i}캠프",
                "작업일자": f"2026-07-{i+1:02d}", "담당기사": "김준형",
                "실제작업항목": "현장 조치", "완료여부": "완료" if i < 4 else "미완료",
                "검증결과": "정상" if i < 4 else "확인필요",
            }
            for i in range(1, 7)
        ]
    cached = _fresh("ryu_field")
    if cached is not None:
        try:
            from app_store import default_store
            return default_store().overlay_rows("03_현장작업실적", cached, "작업ID")
        except Exception:
            return cached
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wanted = [
        "작업ID", "접수ID", "프로젝트NO", "캠프명", "작업일자", "작업시작시간",
        "작업종료시간", "담당기사", "최초접수내용", "실제작업항목", "실제작업상세",
        "접수외추가작업여부", "추가작업내용", "사용부품", "수량", "비용구분",
        "완료여부", "재방문필요", "재방문사유", "작업사진", "작업동영상",
        "완료보고서", "기사보고내용", "관리자검증", "거래명세서반영", "ERP반영",
        "검증자", "검증일", "문제내용", "조치내용", "완료예정일", "비고", "검증결과",
    ]
    out = []
    wb = master_book(master)
    try:
        if "03_현장작업실적" not in wb.sheetnames:
            return _store_cache("ryu_field", out)
        ws = wb["03_현장작업실적"]
        header = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(v).strip(): i for i, v in enumerate(header) if v not in (None, "")}
        for row in ws.iter_rows(min_row=5, values_only=True):
            key = row[idx["작업ID"]] if "작업ID" in idx and idx["작업ID"] < len(row) else None
            project = row[idx["프로젝트NO"]] if "프로젝트NO" in idx and idx["프로젝트NO"] < len(row) else None
            work_date = row[idx["작업일자"]] if "작업일자" in idx and idx["작업일자"] < len(row) else None
            camp = row[idx["캠프명"]] if "캠프명" in idx and idx["캠프명"] < len(row) else None
            if not key or (not work_date and not camp):
                continue
            iso = norm_date(work_date)
            id_blob = f"{key} {project}"
            if iso:
                if not iso.startswith(APP_YEAR + "-"):
                    continue
            elif not (re.search(r"(?:FW|AS|PM|JS)-?26", id_blob, re.I)
                      or re.search(r"\bUJ26\d{5}\b", id_blob, re.I)):
                continue
            rec = {}
            for name in wanted:
                pos = idx.get(name)
                rec[name] = _ryu_display_value(row[pos]) if pos is not None and pos < len(row) else ""
            out.append(rec)
    finally:
        wb.close()
    out.sort(key=lambda r: (norm_date(r.get("작업일자")) == "",
                            norm_date(r.get("작업일자")), str(r.get("작업ID") or "")))
    cached = _store_cache("ryu_field", out)
    try:
        from app_store import default_store
        return default_store().overlay_rows("03_현장작업실적", cached, "작업ID")
    except Exception:
        return cached


def _ryu_upload_history():
    path = os.path.join(ROOT, "reports", "ryu_submissions.jsonl")
    rows = []
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                try:
                    item = json.loads(line)
                except Exception:
                    continue
                day = norm_date(item.get("등록일시") or item.get("조사기준일"))
                if day and not day.startswith(APP_YEAR + "-"):
                    continue
                files = item.get("파일") if isinstance(item.get("파일"), list) else []
                rows.append({
                    "key": str(item.get("등록일시") or ""),
                    "project_no": str(item.get("프로젝트NO") or ""),
                    "camp": str(item.get("캠프명") or ""),
                    "date": day,
                    "status": "원본 저장",
                    "assignee": str(item.get("담당자") or item.get("등록자") or "류지영"),
                    "summary": str(item.get("조치내용") or item.get("조사메모") or
                                   f"카카오톡 원본 {len(files)}개"),
                    "detail": item,
                    "editable": False,
                })
    except Exception:
        pass
    rows.sort(key=lambda r: (r["date"] == "", r["date"], r["key"]))
    return rows


def _ryu_issue_target(row):
    key = str(row.get("업무ID") or row.get("ID") or row.get("원천업무ID") or "").strip()
    upper = key.upper()
    if upper.startswith("AS-"):
        return "as", key
    if upper.startswith("PM-"):
        return "pm", key
    if upper.startswith("FW-"):
        return "field", key
    if upper.startswith("JS-"):
        return "settle", key
    return "", ""


def _ryu_row(rec, key_name, date_names, status_names, summary_names,
             assignee_names=(), editable=True):
    def first(names):
        for name in names:
            value = rec.get(name)
            if value not in (None, ""):
                return str(value)
        return ""
    detail = {str(k): _ryu_display_value(v) for k, v in rec.items() if v not in (None, "")}
    return {
        "key": first((key_name,)),
        "project_no": first(("프로젝트NO",)),
        "camp": first(("캠프명",)),
        "date": norm_date(first(date_names)),
        "status": first(status_names),
        "assignee": first(assignee_names),
        "summary": first(summary_names),
        "detail": detail,
        "editable": bool(editable),
        "store_id": str(rec.get("_store_id") or rec.get("앱DB_ID") or ""),
        "record_version": int(rec.get("_record_version") or rec.get("DB버전") or 0),
        "source": str(rec.get("_source") or rec.get("출처") or ""),
    }


def warm_caches():
    """서버가 뜨자마자 무거운 집계를 **미리 계산해 둔다**(2026-07-31).

    ★ 왜 — 담당자 업무센터가 "업무 현황을 불러오는 중입니다" 에서 멈춰 있었다.
      죽은 게 아니라 **첫 계산이 111초** 걸렸다(실측). 사람은 그때까지 안 기다린다.
      두 번째부터는 677ms 다 — 즉 문제는 계산이 아니라 **누가 그 111초를 맞느냐**였다.
      서버가 대신 맞게 한다. 사람이 열 때는 이미 데워져 있다.

    ★ 순서가 중요하다 — works → settlements → issues → ryu.
      뒤의 것이 앞의 것을 재사용하므로, 거꾸로 하면 같은 걸 두 번 계산한다.
    ★ 실패해도 서버는 그대로 간다. 데우기는 편의지 필수가 아니다.
    ★ 한 번만 데우면 안 된다 — 항목별 TTL 이 issues 300초 · works/settle 600초라
      만료되는 순간 **다음 사람이 다시 111초를 맞는다.** 게다가 원장 mtime 이 바뀌면
      (11:00·15:00 보관본 생성 직후) 캐시가 통째로 비워진다. 그래서 주기적으로 다시 데운다.
      간격은 가장 짧은 TTL(300초)보다 짧아야 의미가 있다 → 240초.
    """
    import time as _t
    first = True
    while True:
        t0 = _t.time()
        # ★ 데우기는 **진짜로 다시 계산해야** 한다 (2026-08-08). 2026-08-08 부터
        #   조회 함수들이 만료 시 옛 값을 즉시 돌려주므로(stale-while-revalidate),
        #   여기서 그냥 부르면 옛 값만 받고 아무것도 안 데운 채 240초를 또 잔다.
        #   그러면 정작 사람이 열 때 뒤늦은 재계산을 맞는다 — 데우기의 뜻이 사라진다.
        for name, fn in (("works", lambda: refresh_now("works", real_works)),
                         ("settlements", lambda: refresh_now("settle", _build_settlements)),
                         ("issues", lambda: refresh_now("issues", _build_issues)),
                         ("erpdocs", lambda: refresh_now("erpdocs", _build_erpdocs)),
                         # 대표보고는 최근 24시간 사용 기록에서 **느린 화면 1등**이었다
                         # (648회·평균 110초). 데우는 목록에 없던 것이 이유의 절반이다.
                         ("exec", lambda: get_exec_report(None, _force=True)),
                         ("ryu", get_ryu_records)):
            try:
                fn()
            except Exception as e:
                if first:
                    print(f"  [예열] {name} 실패: {type(e).__name__} — 서버는 계속 갑니다")
        if first:
            print(f"  [예열] 완료 ({_t.time()-t0:.0f}s) — 업무센터가 바로 열립니다")
            first = False
        _t.sleep(240)


def get_ryu_records():
    """류지영 업무센터: 2026년 업무를 카테고리별 과거→최근 목록으로 제공한다."""
    works = get_works() or {"as": [], "pm": []}
    settlements = get_settlements() or []
    issues = get_issues() or {"rows": []}
    as_rows = [
        _ryu_row(r, "접수ID", ("접수일자", "작업완료일"),
                 ("진행상태", "검증결과"), ("신청내용", "문제내용"),
                 ("담당기사", "담당관리자"),
                 editable=str(r.get("출처") or "") != "ERP")
        for r in drop_side_work(works.get("as") or [])
    ]
    pm_rows = [
        _ryu_row(r, "점검ID", ("점검예정일", "실제점검일"),
                 ("점검상태", "검증결과"), ("점검내용", "이상내용"),
                 ("담당기사", "담당관리자"),
                 editable=str(r.get("출처") or "") not in ("ERP", "정기점검 스케줄 원본"))
        for r in drop_side_work(works.get("pm") or [])
    ]
    field_rows = [
        _ryu_row(r, "작업ID", ("작업일자",), ("완료여부", "검증결과"),
                 ("실제작업항목", "기사보고내용"), ("담당기사",))
        for r in drop_side_work(_ryu_field_records())
    ]
    settle_rows = [
        _ryu_row(r, "정산ID", ("완료일", "명세서발행일", "계산서발행일"),
                 ("상태",), ("업무구분", "적요"), ("담당자",),
                 editable=str(r.get("출처") or "") != "ERP")
        for r in drop_side_work(settlements)
    ]
    target_details = {
        "as": {r["key"]: r.get("detail") or {} for r in as_rows if r.get("key")},
        "pm": {r["key"]: r.get("detail") or {} for r in pm_rows if r.get("key")},
        "field": {r["key"]: r.get("detail") or {} for r in field_rows if r.get("key")},
        "settle": {r["key"]: r.get("detail") or {} for r in settle_rows if r.get("key")},
    }
    issue_rows = []
    for rec in drop_side_work((issues or {}).get("rows") or []):
        row = _ryu_row(rec, "업무ID", ("기준일", "일자", "접수일자", "점검예정일", "완료일"),
                       ("상태", "심각도"), ("문제내용", "내용·근거", "문제유형"),
                       ("담당자",))
        target_category, target_key = _ryu_issue_target(rec)
        if not row["key"]:
            row["key"] = target_key or str(rec.get("ID") or rec.get("원천업무ID") or "")
        if target_category and target_key:
            row["detail"] = {**row["detail"],
                             **target_details.get(target_category, {}).get(target_key, {})}
        row["target_category"] = target_category
        row["target_key"] = target_key
        row["editable"] = bool(target_category and target_key)
        issue_rows.append(row)
    rows = {
        "as": as_rows, "pm": pm_rows, "field": field_rows, "settle": settle_rows,
        "issue": issue_rows, "upload": _ryu_upload_history(),
    }
    for items in rows.values():
        items.sort(key=lambda r: (r.get("date", "") == "", r.get("date", ""), r.get("key", "")))
    def needs_attention(row, completed):
        verify = str((row.get("detail") or {}).get("검증결과") or "").strip()
        if verify and verify != "정상":
            return True
        return str(row.get("status") or "").strip() not in completed
    categories = [
        {"key": "as", "label": "돌발AS", "count": len(as_rows),
         "attention": sum(1 for r in as_rows
                          if needs_attention(r, ("작업완료", "완료", "정상")))},
        {"key": "pm", "label": "정기점검", "count": len(pm_rows),
         "attention": sum(1 for r in pm_rows if needs_attention(r, ("완료", "정상")))},
        {"key": "field", "label": "현장작업", "count": len(field_rows),
         "attention": sum(1 for r in field_rows if needs_attention(r, ("완료", "정상")))},
        {"key": "settle", "label": "거래서류·청구", "count": len(settle_rows),
         "attention": sum(1 for r in settle_rows if needs_attention(
             # 무상·보험은 청구 대상이 아니라 '조치 필요'가 아니다.
             # 미입력은 여기 안 넣는다 — 아직 모르는 것이라 사람이 봐야 한다(2026-08-09).
             r, ("정상", "무상", "보험", "무상/보험", "ERP 계산서(묶음)",
                 "완료(ERP 수금확인)", "완료(ERP 발행확인)")))},
        {"key": "issue", "label": "확인 필요", "count": len(issue_rows),
         "attention": len(issue_rows)},
        {"key": "upload", "label": "자료 등록", "count": len(rows["upload"]), "attention": 0},
    ]
    schema = {
        key: {"label": cfg["label"], "fields": cfg["fields"]}
        for key, cfg in RYU_ENTRY_CONFIG.items()
    }
    schema["issue"] = {
        "label": "확인 필요",
        "fields": [
            {"name": "조치내용", "label": "확인·조치 내용", "type": "textarea"},
            {"name": "완료예정일", "label": "완료 예정일", "type": "date"},
            {"name": "비고", "label": "비고", "type": "textarea"},
        ],
    }
    if DEMO:
        updated = datetime.now().isoformat(timespec="minutes")
    else:
        try:
            updated = datetime.fromtimestamp(_master_mtime()).isoformat(timespec="minutes")
        except Exception:
            updated = datetime.now().isoformat(timespec="minutes")
    return {"updated_at": updated, "year": APP_YEAR, "categories": categories,
            "rows": rows, "schema": schema}


def _staff_allowed_fields(staff_slug, category):
    if staff_slug == "admin":
        return set(_ALL_STAFF_ENTRY_FIELDS.get(category) or set())
    return set((STAFF_ENTRY_PERMISSIONS.get(staff_slug) or {}).get(category) or set())


def _staff_store_row(store, category, record_key):
    cfg = RYU_ENTRY_CONFIG[category]
    key = str(record_key or "").strip()
    matches = [row for row in store.list_sheet_rows(cfg["sheet"])
               if str(row.get(cfg["key_col"]) or row.get("_business_key") or "").strip() == key]
    if not matches:
        raise ValueError("선택한 업무를 앱 DB에서 찾지 못했습니다. 먼저 원본을 수집해 주세요")
    if len(matches) != 1:
        raise ValueError("같은 업무 ID가 여러 건이라 자동 수정하지 않았습니다")
    return matches[0], store.get_work(matches[0]["_store_id"])


def _staff_normalize_value(spec, value, *, clear=False):
    if clear:
        return ""
    if value is None:
        raise ValueError(f"{spec['label']} 값을 입력해 주세요")
    raw = str(value).strip()
    if not raw:
        raise ValueError(f"{spec['label']} 값을 입력하거나 지우기를 명시해 주세요")
    kind = spec.get("type") or "text"
    if kind == "date":
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
            raise ValueError(f"{spec['label']}은 YYYY-MM-DD 형식이어야 합니다")
        try:
            date.fromisoformat(raw)
        except ValueError:
            raise ValueError(f"{spec['label']} 날짜가 올바르지 않습니다")
        return raw
    if kind == "number":
        cleaned = raw.replace(",", "")
        try:
            number = float(cleaned)
        except ValueError:
            raise ValueError(f"{spec['label']}은 숫자로 입력해 주세요")
        return int(number) if number.is_integer() else number
    options = [str(v) for v in (spec.get("options") or [])]
    if options and raw not in options:
        raise ValueError(f"{spec['label']}은 허용된 선택지만 저장할 수 있습니다")
    return raw[:5000]


def _staff_public_record(category, row):
    """앱 저장 응답과 재조회가 같은 별칭·파생 상태를 쓰게 한다."""
    out = {str(k): _ryu_display_value(v) for k, v in dict(row or {}).items()}
    if category == "settle":
        issued = str(out.get("세금계산서발행일") or out.get("실제발행일") or "")[:10]
        out["계산서발행일"] = issued
        if issued:
            out["계산서"] = "발행"
        elif not out.get("계산서"):
            out["계산서"] = "미발행"
    out["record_version"] = int(out.get("_record_version") or out.get("DB버전") or 0)
    out["store_id"] = str(out.get("_store_id") or out.get("앱DB_ID") or "")
    return out


def save_staff_entry(staff_slug, body, *, store=None, actor=None):
    """직원 업무센터의 여러 필드를 앱 DB 한 트랜잭션으로 저장한다.

    Excel 행·셀은 수정하지 않는다. 버전이 낡으면 자동 재시도하지 않고 충돌을 반환하며,
    수식·자동상태 열과 역할 허용표 밖의 열은 서버가 거부한다.
    """
    body = dict(body or {})
    requested = str(body.get("category") or "").strip()
    category = requested
    record_key = str(body.get("key") or body.get("record_key") or "").strip()
    if requested == "issue":
        category = str(body.get("target_category") or "").strip()
        record_key = str(body.get("target_key") or "").strip()
    if category not in RYU_ENTRY_CONFIG or not record_key:
        raise ValueError("카테고리와 업무를 먼저 선택해 주세요")
    if staff_slug not in STAFF_CENTERS and staff_slug != "admin":
        raise PermissionError("등록되지 않은 업무센터입니다")
    allowed = _staff_allowed_fields(staff_slug, category)
    if requested == "issue":
        allowed &= {"조치내용", "완료예정일", "비고"}
    if not allowed:
        raise PermissionError("이 업무센터에서 수정할 수 없는 카테고리입니다")

    values = body.get("values")
    if isinstance(values, str):
        try:
            values = json.loads(values)
        except Exception:
            values = None
    if not isinstance(values, dict):
        values = {name: body.get(name) for name in allowed if name in body}
    clear_fields = body.get("clear_fields") or []
    if isinstance(clear_fields, str):
        try:
            clear_fields = json.loads(clear_fields)
        except Exception:
            clear_fields = [x.strip() for x in clear_fields.split(",") if x.strip()]
    clear_fields = {str(x) for x in clear_fields if str(x)}
    supplied = set(values) | clear_fields
    forbidden = sorted(supplied - allowed)
    if forbidden:
        raise PermissionError("수정 권한이 없는 항목: " + ", ".join(forbidden))
    derived = set(STAFF_DERIVED_FIELDS)
    if category == "field":
        derived.discard("비용구분")
    forbidden_derived = sorted(supplied & derived)
    if forbidden_derived:
        raise ValueError("수식·자동 상태는 직접 수정할 수 없습니다: " + ", ".join(forbidden_derived))
    if not supplied:
        raise ValueError("저장할 값을 입력해 주세요")

    cfg = RYU_ENTRY_CONFIG[category]
    field_specs = {item["name"]: item for item in cfg["fields"]}
    normalized = {}
    for name in sorted(supplied):
        normalized[name] = _staff_normalize_value(
            field_specs[name], values.get(name), clear=name in clear_fields,
        )

    # --demo 서버는 실제 DB를 만지지 않는다. 직접 합성검증은 명시한 임시 store를 쓴다.
    if DEMO and store is None:
        return {"ok": True, "committed": False, "demo": True, "action": "updated",
                "changed": sorted(normalized), "record_version": int(body.get("record_version") or 1),
                "record": {cfg["key_col"]: record_key, **normalized},
                "msg": "데모 입력 확인"}

    if store is None:
        from app_store import default_store
        store = default_store()
    try:
        expected_version = int(body.get("record_version"))
    except (TypeError, ValueError):
        raise ValueError("화면 데이터 버전이 없습니다. 목록을 새로고침해 다시 저장해 주세요")
    if expected_version < 1:
        raise ValueError("화면 데이터 버전이 없습니다. 목록을 새로고침해 다시 저장해 주세요")
    reason = str(body.get("reason") or body.get("survey_note") or "").strip()
    # ★ 사유 없는 청구 제외는 받지 않는다 (2026-08-13). 몇 달 뒤 "이건 왜 청구가
    #   안 나갔나"를 물을 사람이 반드시 있고, 그때 답할 수 있는 것은 이 한 줄뿐이다.
    #   STAFF_REASON_REQUIRED_FIELDS 는 **기존값을 고칠 때만** 사유를 받으므로
    #   빈 칸 → '예' 인 첫 지정이 그냥 통과한다 — 그 구멍을 여기서 막는다.
    if str(normalized.get("청구제외") or "") == "예":
        note = str(normalized.get("청구제외사유") or "").strip() or reason
        if len(note) < 2:
            raise ValueError("청구 제외는 사유를 함께 적어야 합니다 "
                             "(예: 정기점검과 동시 진행 — 돌발AS로 청구 안 함)")
        normalized["청구제외사유"] = note
    who = str(actor or ("staff:" + staff_slug))[:200]
    request_id = str(body.get("idempotency_key") or body.get("request_id") or "").strip()
    request_payload = {
        "staff_slug": staff_slug, "actor": who, "category": category,
        "key": record_key, "record_version": expected_version,
        "values": normalized, "clear_fields": sorted(clear_fields), "reason": reason,
    }
    if not request_id:
        request_id = _stable_request_key(
            "staff-entry", who, [], record=request_payload,
        )
    from app_store import SHEET_SPECS, VersionConflict
    sheet_spec = SHEET_SPECS.get(cfg["sheet"], {})
    core_by_column = {column: core for core, column in sheet_spec.items()
                      if core in ("public_id", "project_no", "camp_name", "status") and column}
    # 멱등 확인 → 현재값/버전 확인 → work_item·감사이벤트·outbox 변경 → 최종 직원 응답
    # 저장까지 한 BEGIN IMMEDIATE에 묶는다. 그래야 같은 키의 동시 요청 둘이 모두
    # "아직 응답 없음"을 본 뒤 하나는 업무를 바꾸고 바깥 멱등 INSERT에서 409가 나는
    # 반쪽 성공이 생기지 않는다.
    with store.transaction() as conn:
        replay, request_hash = store._idempotency_replay(
            conn, "staff:entry", request_id, request_payload,
        )
        if replay is not None:
            result = replay
        else:
            # BEGIN IMMEDIATE 뒤에 읽는다. 같은 키가 아닌 동시 수정도 이 시점의 한
            # 버전을 기준으로 판정되며, 이 트랜잭션이 끝날 때까지 다른 쓰기는 못 낀다.
            sheet_row, current = _staff_store_row(store, category, record_key)
            changed = {}
            before_values = {}
            for name, value in normalized.items():
                core = core_by_column.get(name)
                old = current.get(core) if core else (current.get("fields") or {}).get(name)
                if json.dumps(old, ensure_ascii=False, sort_keys=True, default=str) == json.dumps(
                        value, ensure_ascii=False, sort_keys=True, default=str):
                    continue
                before_values[name] = old
                changed[name] = value

            if not changed:
                latest = _staff_public_record(category, sheet_row)
                result = {
                    "ok": True, "committed": True, "action": "no_change", "changed": [],
                    "skipped": sorted(normalized), "record_version": latest["record_version"],
                    "record": latest, "msg": "이미 같은 값입니다",
                }
            else:
                correction_fields = [
                    name for name in changed
                    if (name in clear_fields or
                        (name in STAFF_REASON_REQUIRED_FIELDS and
                         before_values.get(name) not in (None, "")))
                ]
                if correction_fields and len(reason) < 2:
                    raise ValueError(
                        "기존 값 정정 사유를 입력해 주세요: " + ", ".join(correction_fields)
                    )
                if int(current.get("record_version") or 0) != expected_version:
                    raise VersionConflict(
                        current["id"], expected_version,
                        int(current.get("record_version") or 0),
                    )

                patch = {"fields": dict(changed)}
                for name, value in changed.items():
                    core = core_by_column.get(name)
                    if core:
                        patch[core] = value
                evidence = (
                    f"{STAFF_CENTERS.get(staff_slug, {'name': '관리자'})['name']} 업무센터"
                    f" · {reason[:300] if reason else '확인 입력'}"
                )
                response = store.update_work(
                    current["id"], expected_version=expected_version, patch=patch,
                    actor=who, source="app-staff-entry", evidence=evidence,
                    source_ref=f"staff:{staff_slug}:{category}:{record_key}",
                    idempotency_key=(request_id[:220] + ":work"), _conn=conn,
                )

                # 별도 재조회 연결은 아직 커밋되지 않은 변경을 볼 수 없다. update_work가
                # 같은 연결에서 돌려준 권위 레코드를 기존 Excel 모양에 덮어 응답을 만든다.
                after = dict(response["work"])
                saved_row = dict(sheet_row)
                saved_row.update(after.get("fields") or {})
                for core, column in sheet_spec.items():
                    if core in ("public_id", "project_no", "camp_name", "status") and column:
                        saved_row[column] = after.get(core)
                saved_row.update({
                    "_store_id": after["id"],
                    "_record_version": after["record_version"],
                    "_business_key": after["business_key"],
                    "_source": after.get("source") or "",
                    "_evidence": after.get("evidence") or "",
                })
                authoritative = _staff_public_record(category, saved_row)
                outbox_pending = int(conn.execute(
                    "SELECT COUNT(*) FROM outbox "
                    "WHERE status IN ('pending','failed','leased')"
                ).fetchone()[0])
                result = {
                    "ok": True, "committed": True, "action": "updated",
                    "changed": sorted(changed),
                    "skipped": sorted(set(normalized) - set(changed)),
                    "before": before_values,
                    "record_version": authoritative["record_version"],
                    "record": authoritative,
                    "event_id": response.get("event_id"),
                    "event_seq": response.get("event_seq"),
                    "app_outbox_pending": outbox_pending,
                    "msg": "앱 DB 저장 완료 · Excel 자동 보관본 대기",
                }
            store._save_idempotency(
                conn, "staff:entry", request_id, request_hash, result,
            )

    if result.get("action") == "updated" and not result.get("idempotent_replay"):
        _invalidate_app_data_caches()
    return result


# ═══ 삭제 · 되살리기 (2026-08-13 지시: "위건 삭제 수정 가능하게 코딩") ═══════════
#
# 류지영 요청: "삭제할수있게 해주세요!! / 정기점검이랑 동시진행으로 돌발AS로는
# 청구를 안할꺼라 삭제가 필요합니당"
#
# ★ 손잡이는 **둘**이고 뜻이 다르다. 화면이 그 차이를 한 줄로 말한다:
#   · 삭제 = 잘못 등록된 건. 목록·집계·청구에서 통째로 뺀다.
#   · 제외 = 다녀왔지만 이 건으로는 청구 안 함(`청구제외` 칸). 기록은 그대로 남는다.
#   위 요청의 실제 사정은 **제외**다 — 실측 UJ2601032 는 작업완료·완료일 2026-06-12·
#   담당기사 김필우·사진·완료보고서·ERP등록이 다 있는 건이라 '안 갔다'가 아니다.
#   그래도 삭제를 빼지 않는다: 잘못 등록된 건은 실재하고, 무엇이 맞는지는 사람만 안다.
#
# ★ 물리 DELETE 를 쓰지 않는다. `soft_delete_work` 는 `deleted_at` 만 세우고
#   `app_store.overlay_rows` 가 그 행을 legacy Excel 행까지 통째로 감춘다(:2532).
#   즉 사람에게는 지워진 것과 똑같이 보이면서 **되살릴 수 있다.**
_WORK_DELETE_CATEGORIES = ("as", "pm")


def _staff_work_target(store, category, record_key, *, include_deleted=False):
    """삭제·되살리기가 가리키는 정본 행 하나. 여럿이면 손대지 않는다."""
    cfg = RYU_ENTRY_CONFIG[category]
    key = str(record_key or "").strip()
    matches = [row for row in store.list_sheet_rows(cfg["sheet"], include_deleted=True)
               if str(row.get(cfg["key_col"]) or row.get("_business_key") or "").strip() == key]
    if include_deleted:
        matches = [row for row in matches if row.get("_deleted")]
    else:
        matches = [row for row in matches if not row.get("_deleted")]
    if not matches:
        raise ValueError("선택한 업무를 앱 DB에서 찾지 못했습니다")
    if len(matches) != 1:
        # 짐작으로 하나를 고르면 엉뚱한 현장이 목록에서 사라진다([172] 의 문).
        raise ValueError("같은 업무 ID가 여러 건이라 자동으로 처리하지 않았습니다")
    return cfg, matches[0]


def delete_staff_work(staff_slug, body, *, store=None, actor=None):
    """업무 한 건을 삭제한다(soft). 사유 없이는 지나가지 못한다."""
    category = str(body.get("category") or "").strip()
    record_key = str(body.get("key") or body.get("target_key") or "").strip()
    if category not in _WORK_DELETE_CATEGORIES or not record_key:
        raise ValueError("삭제할 업무를 먼저 선택해 주세요")
    if staff_slug not in STAFF_CENTERS and staff_slug != "admin":
        raise PermissionError("등록되지 않은 업무센터입니다")
    if not _staff_allowed_fields(staff_slug, category):
        raise PermissionError("이 업무센터에서는 이 업무를 삭제할 수 없습니다")
    reason = str(body.get("reason") or "").strip()
    if len(reason) < 2:
        # 사유 없는 삭제는 몇 달 뒤 아무도 설명하지 못한다.
        raise ValueError("삭제 사유를 적어 주세요 (예: 잘못 등록 — 같은 건이 두 번 접수됨)")
    if store is None:
        from app_store import default_store
        store = default_store()
    cfg, row = _staff_work_target(store, category, record_key)
    try:
        expected_version = int(body.get("record_version"))
    except (TypeError, ValueError):
        raise ValueError("화면 데이터 버전이 없습니다. 목록을 새로고침해 다시 시도해 주세요")
    who = str(actor or ("staff:" + staff_slug))[:200]
    # ⚠ 청구자료는 지우지 않는다 — 있다는 사실만 응답에 실어 화면이 한 번 더 묻는다.
    docs = _settlement_documents_for_source(store, record_key)
    if docs and not body.get("confirm_documents"):
        return {"ok": False, "needs_confirm": True, "documents": docs,
                "error": "이 건에는 청구자료가 있습니다 — 그래도 삭제할까요?",
                "record_version": expected_version}
    response = store.soft_delete_work(
        row["_store_id"], expected_version=expected_version, actor=who,
        reason=reason, source="app-staff-delete",
        evidence=(f"{STAFF_CENTERS.get(staff_slug, {'name': '관리자'})['name']} 업무센터"
                  f" 삭제 · {reason[:300]}"),
        idempotency_key=str(body.get("idempotency_key") or "").strip() or None,
    )
    _invalidate_app_data_caches()
    return {"ok": True, "committed": True, "action": "deleted",
            "key": record_key, "category": category,
            "record_version": int(response["work"]["record_version"]),
            "documents": docs,
            "event_id": response.get("event_id"),
            "msg": "삭제했습니다 — [삭제된 건 보기]에서 되살릴 수 있습니다"}


def restore_staff_work(staff_slug, body, *, store=None, actor=None):
    """삭제한 업무를 되살린다. 삭제와 같은 문을 거꾸로 지난다."""
    category = str(body.get("category") or "").strip()
    record_key = str(body.get("key") or body.get("target_key") or "").strip()
    if category not in _WORK_DELETE_CATEGORIES or not record_key:
        raise ValueError("되살릴 업무를 먼저 선택해 주세요")
    if staff_slug not in STAFF_CENTERS and staff_slug != "admin":
        raise PermissionError("등록되지 않은 업무센터입니다")
    if not _staff_allowed_fields(staff_slug, category):
        raise PermissionError("이 업무센터에서는 이 업무를 되살릴 수 없습니다")
    if store is None:
        from app_store import default_store
        store = default_store()
    cfg, row = _staff_work_target(store, category, record_key, include_deleted=True)
    try:
        expected_version = int(body.get("record_version"))
    except (TypeError, ValueError):
        raise ValueError("화면 데이터 버전이 없습니다. 목록을 새로고침해 다시 시도해 주세요")
    who = str(actor or ("staff:" + staff_slug))[:200]
    reason = str(body.get("reason") or "삭제 취소").strip()
    response = store.restore_work(
        row["_store_id"], expected_version=expected_version, actor=who,
        reason=reason, source="app-staff-restore",
        evidence=(f"{STAFF_CENTERS.get(staff_slug, {'name': '관리자'})['name']} 업무센터"
                  f" 되살림 · {reason[:300]}"),
        idempotency_key=str(body.get("idempotency_key") or "").strip() or None,
    )
    _invalidate_app_data_caches()
    return {"ok": True, "committed": True, "action": "restored",
            "key": record_key, "category": category,
            "record_version": int(response["work"]["record_version"]),
            "event_id": response.get("event_id"),
            "msg": "되살렸습니다 — 목록에 다시 보입니다"}


def _settlement_documents_for_source(store, source_id):
    """이 업무에 붙은 실제 청구자료 요약. 삭제해도 **지우지 않는다**([208])."""
    out = []
    key = str(source_id or "").strip()
    if not key:
        return out
    try:
        for row in store.list_sheet_rows("06_거래서류청구수금"):
            if str(row.get("원천업무ID") or "").strip() != key:
                continue
            if not _settlement_has_documents(row):
                continue
            out.append({
                "정산ID": str(row.get("정산ID") or ""),
                "PO번호": str(row.get("PO번호") or ""),
                "거래명세서번호": str(row.get("거래명세서번호") or ""),
                "세금계산서발행일": str(row.get("세금계산서발행일") or "")[:10],
                "입금일": str(row.get("입금일") or "")[:10],
            })
    except Exception:
        # 못 읽었으면 '없다'고 하지 않는다 — 못 읽었다고 말한다([169]).
        return [{"정산ID": "", "확인못함": "청구자료를 확인하지 못했습니다"}]
    return out


def deleted_staff_works(staff_slug, category=None, *, store=None):
    """삭제된 건 목록 — 사람이 볼 수 있는 자리에 둔다.

    코드로만 되살릴 수 있으면 없는 것과 같다. 화면이 이 목록을 보여 주고
    거기서 [되살리기] 를 누른다.
    """
    if staff_slug not in STAFF_CENTERS and staff_slug != "admin":
        raise PermissionError("등록되지 않은 업무센터입니다")
    if store is None:
        from app_store import default_store
        store = default_store()
    cats = [category] if category in _WORK_DELETE_CATEGORIES else list(_WORK_DELETE_CATEGORIES)
    out = []
    for cat in cats:
        if not _staff_allowed_fields(staff_slug, cat):
            continue
        cfg = RYU_ENTRY_CONFIG[cat]
        spec_status = {"as": "진행상태", "pm": "점검상태"}[cat]
        for row in store.list_sheet_rows(cfg["sheet"], include_deleted=True):
            if not row.get("_deleted"):
                continue
            out.append({
                "category": cat, "label": cfg["label"],
                "key": str(row.get(cfg["key_col"]) or row.get("_business_key") or ""),
                "프로젝트NO": str(row.get("프로젝트NO") or ""),
                "캠프명": str(row.get("캠프명") or ""),
                "상태": str(row.get(spec_status) or ""),
                "담당기사": str(row.get("담당기사") or ""),
                "삭제근거": str(row.get("_evidence") or ""),
                "record_version": int(row.get("_record_version") or 0),
            })
    out.sort(key=lambda r: (r["category"], r["key"]))
    return {"ok": True, "rows": out, "count": len(out)}


def get_staff_records(staff_slug, *, store=None):
    """로그인한 직원의 서버 허용표와 권위 DB 버전을 함께 반환한다."""
    if staff_slug not in STAFF_CENTERS and staff_slug != "admin":
        raise PermissionError("등록되지 않은 업무센터입니다")
    if store is None:
        payload = get_ryu_records()
    else:
        # 합성검증·DB 장애 복구 화면은 Excel 없이도 정본 행만 볼 수 있다.
        rows = {}
        from app_store import SHEET_SPECS
        for category, cfg in RYU_ENTRY_CONFIG.items():
            rendered = []
            spec = SHEET_SPECS.get(cfg["sheet"], {})
            for source in store.list_sheet_rows(cfg["sheet"]):
                view = _staff_public_record(category, source)
                rendered.append(_ryu_row(
                    view, cfg["key_col"], (cfg["date_col"],),
                    (spec.get("status") or "상태",),
                    ("문제내용", "조치내용", "비고"), ("담당자", "담당기사"),
                ))
            rows[category] = rendered
        payload = {"updated_at": datetime.now().isoformat(timespec="minutes"),
                   "year": APP_YEAR, "rows": rows,
                   "categories": [{"key": key, "label": cfg["label"],
                                    "count": len(rows.get(key) or []), "attention": 0}
                                   for key, cfg in RYU_ENTRY_CONFIG.items()],
                   "schema": {key: {"label": cfg["label"], "fields": cfg["fields"]}
                              for key, cfg in RYU_ENTRY_CONFIG.items()}}
    permissions = {}
    for category, cfg in RYU_ENTRY_CONFIG.items():
        allowed = _staff_allowed_fields(staff_slug, category)
        permissions[category] = sorted(allowed)
        if category in (payload.get("schema") or {}):
            payload["schema"][category]["fields"] = [
                field for field in cfg["fields"] if field["name"] in allowed
            ]
        for row in (payload.get("rows") or {}).get(category, []):
            row["editable"] = bool(row.get("editable", True) and allowed)
            row["allowed_fields"] = sorted(allowed)
    payload["staff"] = STAFF_CENTERS.get(staff_slug, {"name": "관리자"})["name"]
    payload["staff_slug"] = staff_slug
    payload["permissions"] = permissions
    try:
        db_updated = datetime.fromtimestamp(os.path.getmtime(str(
            (store or __import__("app_store").default_store()).db_path
        ))).isoformat(timespec="minutes")
        if db_updated > str(payload.get("updated_at") or ""):
            payload["updated_at"] = db_updated
    except Exception:
        pass
    return payload


def _ryu_find_master_record(category, record_key):
    cfg = RYU_ENTRY_CONFIG[category]
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = master_book(master)
    try:
        if cfg["sheet"] not in wb.sheetnames:
            raise ValueError("대상 시트를 찾지 못했습니다")
        ws = wb[cfg["sheet"]]
        header = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(v).strip(): i for i, v in enumerate(header) if v not in (None, "")}
        if cfg["key_col"] not in idx:
            raise ValueError("업무 ID 열을 찾지 못했습니다")
        for row in ws.iter_rows(min_row=5, values_only=True):
            pos = idx[cfg["key_col"]]
            if pos >= len(row) or str(row[pos] or "").strip() != str(record_key).strip():
                continue
            rec = {name: _ryu_display_value(row[i]) for name, i in idx.items() if i < len(row)}
            day = norm_date(rec.get(cfg["date_col"]))
            blob = f"{rec.get(cfg['key_col'], '')} {rec.get('프로젝트NO', '')}"
            if not (day.startswith(APP_YEAR + "-")
                    or re.search(r"(?:AS|PM|FW|JS)-?26", blob, re.I)
                    or re.search(r"\bUJ26\d{5}\b", blob, re.I)):
                raise ValueError(f"{APP_YEAR}년 업무만 입력할 수 있습니다")
            return rec
    finally:
        wb.close()
    raise ValueError("선택한 업무를 최신 관리대장에서 찾지 못했습니다")


def _save_ryu_evidence(file_info, category, record_key, submitter="류지영"):
    if not file_info or not file_info.get("data"):
        return ""
    data = file_info["data"]
    if len(data) > 25_000_000:
        raise ValueError("근거 파일은 25MB 이하여야 합니다")
    ext = os.path.splitext(file_info.get("filename") or "")[1].lower()
    allowed = (".png", ".jpg", ".jpeg", ".webp", ".pdf", ".xlsx", ".docx", ".txt")
    if ext not in allowed:
        raise ValueError("근거는 이미지·PDF·Excel·Word·텍스트 파일만 가능합니다")
    from source_dirs import KAKAO_DIR
    now = datetime.now()
    folder = os.path.join(KAKAO_DIR, f"{now:%Y}", f"{now:%m}", f"{now:%Y-%m-%d}")
    os.makedirs(folder, exist_ok=True)
    safe_key = re.sub(r"[^0-9A-Za-z가-힣_-]", "_", str(record_key))[:50]
    safe_submitter = re.sub(r"[^0-9A-Za-z가-힣_-]", "_", str(submitter or "직원"))[:30]
    safe_category = re.sub(r"[^0-9A-Za-z가-힣_-]", "_", str(category or "업무"))[:30]
    name = (f"({safe_submitter})_업무근거_{safe_category}_{safe_key}_{now:%Y%m%d_%H%M%S}_"
            f"{_safe_upload_name(file_info.get('filename'))}")
    with open(os.path.join(folder, name), "wb") as out:
        out.write(data)
    return name


def _stable_request_key(source, actor, items, record=None, supplied=None):
    """헤더가 없어도 동일 요청 재전송은 같은 DB 명령이 되게 한다."""
    raw = str(supplied or "").strip()
    payload = {"source": source, "actor": actor, "items": items, "record": record or {}}
    digest = hashlib.sha256(json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str,
    ).encode("utf-8")).hexdigest()
    if not raw:
        return f"app:{digest}"
    # 같은 클라이언트 키를 다른 본문에 재사용하면 저장소가 request hash 충돌로 거부해야
    # 한다. 본문 digest를 키에 섞으면 서로 다른 키가 되어 그 관문을 우회하므로 넣지 않는다.
    clean = re.sub(r"[^0-9A-Za-z._:-]", "_", raw)[:160]
    return f"client:{clean}:{hashlib.sha256(raw.encode('utf-8')).hexdigest()[:24]}"


def _app_db_stamp():
    try:
        from app_store import default_store
        path = str(default_store().db_path)
        return tuple(os.path.getmtime(p) if os.path.exists(p) else 0
                     for p in (path, path + "-wal"))
    except Exception:
        return (0, 0)


def _invalidate_app_data_caches():
    """현재 캐시는 무효화하되 DB overlay를 입힌 last-good은 즉시 응답용으로 남긴다."""
    cache = globals().get("_cache")
    if not isinstance(cache, dict):
        return
    # Z: Excel 재독기는 오래 걸릴 수 있다. 이미 읽은 Excel base가 있으면 그것을 복사해
    # 방금 커밋한 DB만 즉시 overlay하고, 뒤쪽 refresh가 새 base를 다시 계산하게 한다.
    for key in ("works", "settle"):
        previous = cache.get(key) or cache.get(key + "_stale")
        if previous is not None:
            try:
                patched = copy.deepcopy(previous)
                patched = (_overlay_app_store_works(patched) if key == "works"
                           else _overlay_app_store_settlements(patched))
                cache[key + "_stale"] = patched
            except Exception:
                cache.pop(key + "_stale", None)
        cache.pop(key, None)
        cache.pop(key + "_ts", None)
        cache[key + "_app_db_stamp"] = _app_db_stamp()
    for key in ("issues", "exec", "status"):
        for suffix in ("", "_ts", "_stale"):
            cache.pop(key + suffix, None)


def _safety_spool(items, error):
    """DB 실패 때만 쓰는 복구용 JSON. 이것은 저장 성공이 아니다."""
    added = 0
    spool_error = ""
    try:
        if items:
            from ledger_writer import queue_add
            added = queue_add(items)
    except Exception as exc:
        spool_error = str(exc)[:200]
    return {
        "spooled": bool(items) and not spool_error,
        "spool_added": added,
        "spool_error": spool_error,
        "error": str(error)[:300],
    }


def _prepare_archive_export(actor="app"):
    """외부 쓰기 없이 로컬 DB 스냅샷·템플릿 사본·명령계획만 검증해 만든다."""
    from archive_export import ArchiveExporter
    from app_store import default_store
    from ecount_reconcile import load_config, resolve_master

    template = resolve_master(load_config()["reconcile"]["master_xlsx"])
    prepared = ArchiveExporter(default_store()).run_local(
        template_path=template, adapter=None, dry_run=True,
    )
    rendered = bool(prepared.get("archive"))
    return {
        "ok": True,
        "prepared": True,
        "rendered": rendered,
        "actor": actor,
        "label": "보관본 지금 생성",
        "msg": ("검증된 Excel 보관본 확인 완료" if rendered else
                "보관본 생성 계획·DB 스냅샷 준비 완료 · Excel 렌더 검증 대기"),
        **prepared,
    }


def enqueue_for_scheduled_apply(items, source="app", actor="app",
                                idempotency_key=None, record=None):
    """앱 DB를 먼저 커밋하고, 성공한 값만 Excel 보관본 outbox에 남긴다.

    JSON은 정본이 아니다. 앱 DB 커밋이 실패했을 때만 복구용 안전 스풀로 쓰며,
    그런 경우 호출자에게 반드시 ``ok=False``를 돌려 거짓 저장 성공을 막는다.
    ``record``는 Excel 행 번호 없이 만드는 신규 업무용 명세다.
    """
    items = [dict(item or {}) for item in (items or [])]
    actor = str(actor or "app")[:200]
    request_key = _stable_request_key(
        source, actor, items, record=record, supplied=idempotency_key,
    )
    canonical_committed = False
    records = []
    try:
        from app_store import NotFoundError, default_store
        store = default_store()
        records = []
        if record:
            spec = dict(record)
            kind = str(spec.get("kind") or "").strip()
            business_key = str(spec.get("business_key") or "").strip()
            if not kind or not business_key:
                raise ValueError("신규 업무의 kind/business_key가 없습니다")
            try:
                current = store.get_work(kind=kind, business_key=business_key)
            except NotFoundError:
                current = None
            public_id = str((current or {}).get("public_id") or "").strip()
            if not public_id:
                reserved = store.reserve_public_id(
                    kind, spec.get("work_date"), str(spec.get("id_prefix") or kind),
                )
                public_id = reserved["public_id"]
            fields = dict(spec.get("fields") or {})
            id_col = str(spec.get("id_col") or "").strip()
            if id_col:
                fields[id_col] = public_id
            desired_core = {
                "public_id": public_id,
                "project_no": spec.get("project_no"),
                "camp_name": spec.get("camp_name"),
                "status": str(spec.get("status") or ""),
            }
            if current:
                same = all(
                    json.dumps(current.get(name), ensure_ascii=False, sort_keys=True, default=str)
                    == json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
                    for name, value in desired_core.items()
                ) and all(
                    json.dumps((current.get("fields") or {}).get(name), ensure_ascii=False,
                               sort_keys=True, default=str)
                    == json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
                    for name, value in fields.items()
                )
                if same:
                    response = {"work": current, "event_id": None, "event_seq": None,
                                "idempotent_replay": True, "no_change": True}
                else:
                    response = store.update_work(
                        current["id"], expected_version=int(current["record_version"]),
                        patch={**desired_core, "fields": fields},
                        actor=actor, source=source,
                        evidence=str(spec.get("evidence") or ""),
                        source_ref=str(spec.get("source_ref") or ""),
                        idempotency_key=f"{request_key}:update",
                    )
            else:
                response = store.create_work(
                    kind=kind,
                    business_key=business_key,
                    fields=fields,
                    actor=actor,
                    source=source,
                    evidence=str(spec.get("evidence") or ""),
                    source_ref=str(spec.get("source_ref") or ""),
                    idempotency_key=request_key,
                    **desired_core,
                )
            records.append(response)
            canonical_committed = True
            archive_queued = 0  # DB-only 신규 행은 snapshot exporter가 행을 생성한다.
        else:
            applied = store.apply_legacy_items(
                items, source=source, actor=actor, idempotency_key=request_key,
            )
            canonical_committed = bool(
                int(applied.get("created") or 0)
                + int(applied.get("updated") or 0)
                + int(applied.get("settings") or 0)
                + int(applied.get("skipped") or 0)
            ) or not items
            if not applied.get("ok"):
                raise RuntimeError("; ".join(
                    str(e.get("message") or e) for e in (applied.get("errors") or [])
                ) or "앱 DB 입력 일부가 커밋되지 않았습니다")
            records = list(applied.get("items") or [])
            # 기존 셀 명령은 11·15시 보관본 회차와 호환되도록 DB outbox에도 넣는다.
            # ledger_db가 같은 값을 앱 DB에 한 번 더 제시하지만 already_equal/idempotency로
            # 감사 이벤트를 중복 생성하지 않는다.
            import ledger_db
            archive_queued = ledger_db.enqueue(
                items, source=source, ingest_prefix=f"archive:{request_key}",
            ) if items else 0

        state = store.status()
        try:
            import ledger_db
            archive_state = ledger_db.status()
        except Exception:
            archive_state = {}
        _invalidate_app_data_caches()
        pending = int(archive_state.get("대기") or state.get("outbox_pending") or 0)
        return {
            "ok": True,
            "committed": True,
            "queued": archive_queued,
            "ingested": len(records),
            "records": records,
            "pending": pending,
            "archive_outbox_pending": pending,
            "app_outbox_pending": int(state.get("outbox_pending") or 0),
            "applying": False,
            "next_archive": archive_state.get("다음반영"),
            "next_apply": archive_state.get("다음반영"),
            "msg": "앱 DB 저장·보관본 생성 예정",
        }
    except Exception as exc:
        if canonical_committed:
            _invalidate_app_data_caches()
        retryable = type(exc).__name__ not in {
            "ValidationError", "IdempotencyConflict", "VersionConflict", "ValueError",
        }
        recovery = (_safety_spool(items, exc) if retryable else {
            "spooled": False, "spool_added": 0, "spool_error": "",
            "error": str(exc)[:300],
        })
        return {
            "ok": False,
            "committed": canonical_committed,
            "partial_commit": canonical_committed,
            "retryable": retryable,
            "queued": 0,
            "ingested": 0,
            "pending": recovery.get("spool_added", 0),
            "applying": False,
            "msg": (("앱 DB 저장 완료 · 보관본 outbox 실패 · 안전 임시 큐 보관"
                     if recovery.get("spooled") else "앱 DB 저장 완료 · 보관본 outbox 실패")
                    if canonical_committed else
                    ("앱 DB 저장 실패 · 안전 임시 큐 보관(미반영)"
                     if recovery.get("spooled") else "앱 DB 저장 실패 · 미반영")),
            **recovery,
        }


def save_ryu_entry(fields, files, source_ip="", actor="app", idempotency_key=None):
    """선택한 기존 업무의 빈 원천 칸만 큐에 넣고, 첨부 근거는 원본 폴더에 보존한다."""
    requested = str(fields.get("category") or "").strip()
    category = requested
    record_key = str(fields.get("record_key") or "").strip()
    if requested == "issue":
        category = str(fields.get("target_category") or "").strip()
        record_key = str(fields.get("target_key") or "").strip()
    if category not in RYU_ENTRY_CONFIG or not record_key:
        raise ValueError("카테고리와 업무를 먼저 선택해 주세요")
    if DEMO:
        current = {}
        for row in (get_ryu_records().get("rows") or {}).get(category, []):
            if str(row.get("key") or "") == record_key:
                current = row.get("detail") or {}
                break
        if not current:
            raise ValueError("선택한 데모 업무를 찾지 못했습니다")
    else:
        current = _ryu_find_master_record(category, record_key)
    cfg = RYU_ENTRY_CONFIG[category]
    allowed = {item["name"]: item for item in cfg["fields"]}
    if requested == "issue":
        allowed = {k: v for k, v in allowed.items() if k in ("조치내용", "완료예정일", "비고")}
    items = []
    for name, spec in allowed.items():
        raw = str(fields.get(name) or "").strip()
        if raw == "":
            continue
        vtype = spec["type"] if spec["type"] in ("date", "number") else "text"
        if vtype == "date" and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
            raise ValueError(f"{spec['label']}은 YYYY-MM-DD 형식이어야 합니다")
        value = raw
        if vtype == "number":
            cleaned = raw.replace(",", "")
            try:
                number = float(cleaned)
            except ValueError:
                raise ValueError(f"{spec['label']}은 숫자로 입력해 주세요")
            value = int(number) if number.is_integer() else number
        items.append({
            "sheet": cfg["sheet"], "key_col": cfg["key_col"], "key": record_key,
            "col": name, "value": value, "vtype": vtype, "only_if_empty": True,
        })
    evidence_name = _save_ryu_evidence(files.get("evidence_file"), category, record_key)
    note = str(fields.get("survey_note") or "").strip()
    evidence = (f"류지영 업무센터 입력({source_ip or '앱'})"
                f"{' · ' + note[:160] if note else ''}"
                f"{' · 근거 ' + evidence_name if evidence_name else ''}")
    for item in items:
        item["evidence"] = evidence
    manifest = {
        "등록일시": datetime.now().isoformat(timespec="seconds"),
        "등록자": "류지영", "입력유형": "업무 보충입력",
        "카테고리": requested, "반영카테고리": category,
        "업무ID": record_key, "프로젝트NO": str(current.get("프로젝트NO") or ""),
        "캠프명": str(current.get("캠프명") or ""), "조사메모": note,
        "추가근거": evidence_name, "입력항목": [item["col"] for item in items],
    }
    if DEMO:
        return {"ok": True, "committed": False, "queued": len(items),
                "pending": len(items), "manifest": manifest,
                "applying": False, "msg": "데모 입력"}
    if not items and not evidence_name:
        raise ValueError("보충할 항목 또는 근거 파일을 입력해 주세요")
    queued = enqueue_for_scheduled_apply(
        items, source="app-ryu", actor=actor, idempotency_key=idempotency_key,
    )
    if queued.get("ok"):
        os.makedirs(os.path.join(ROOT, "reports"), exist_ok=True)
        with open(os.path.join(ROOT, "reports", "ryu_submissions.jsonl"),
                  "a", encoding="utf-8") as out:
            out.write(json.dumps(manifest, ensure_ascii=False) + "\n")
    if not items:
        queued["msg"] = "근거 파일만 저장했습니다"
    return {**queued, "manifest": manifest}


def save_new_workcenter_job(fields, files, source_ip="", actor="app",
                            idempotency_key=None):
    """신규 AS·정기점검을 Excel 행 번호 없이 앱 DB에 즉시 등록한다."""
    category = str(fields.get("category") or "").strip()
    if category not in ("as", "pm"):
        raise ValueError("신규 업무는 돌발AS 또는 정기점검만 등록할 수 있습니다")
    from project_resolve import norm
    project_no = norm(fields.get("project_no"))
    if not re.fullmatch(r"UJ26\d{5}", project_no or "", re.I):
        raise ValueError("2026년 프로젝트번호(UJ + 숫자 7자리)를 입력해 주세요")
    work_date = str(fields.get("work_date") or "").strip()
    try:
        parsed_work_date = date.fromisoformat(work_date)
    except ValueError:
        parsed_work_date = None
    if not parsed_work_date or parsed_work_date.year != 2026:
        raise ValueError("업무일은 2026년 YYYY-MM-DD 형식으로 입력해 주세요")
    camp = str(fields.get("camp_name") or "").strip()
    if not camp:
        raise ValueError("캠프명을 입력해 주세요")
    sheet = "02_돌발AS접수" if category == "as" else "04_정기점검"
    prefix = "AS" if category == "as" else "PM"
    id_col = "접수ID" if category == "as" else "점검ID"
    # ★ 단계 낱말을 여기 적지 않는다 (2026-08-10). 예전 기본값 '접수'·'예정' 중
    #   '접수' 는 관리대장 드롭다운에 아예 없는 낱말이었다 — 앱이 목록 밖 값을
    #   새로 만들어 넣고 있었다는 뜻이다. 정의는 `work_flow` 한 곳이 정한다.
    import work_flow
    status = str(fields.get("status") or "").strip()
    if not status:
        status = work_flow.default_stage(category)
    allowed_stages = work_flow.stage_words(category)
    if allowed_stages and status and status not in allowed_stages:
        raise ValueError("현재 상태는 관리대장 목록에 있는 값이어야 합니다: "
                         + " · ".join(allowed_stages))
    kind = "돌발AS" if category == "as" else "정기점검"
    tech = str(fields.get("assignee") or "").strip()
    cost = str(fields.get("cost_type") or "").strip()
    note = str(fields.get("description") or "").strip()
    data = {
        "프로젝트NO": project_no,
        "캠프명": camp,
        ("접수일자" if category == "as" else "점검예정일"): work_date,
        "담당기사": tech,
        ("진행상태" if category == "as" else "점검상태"): status,
        "유상·무상·보험": cost,
    }
    if note:
        data["신청내용" if category == "as" else "점검내용"] = note
    evidence_name = _save_ryu_evidence(files.get("evidence_file"), category, project_no)
    evidence_text = (f"업무센터 신규등록({source_ip or '앱'})"
                     f"{' · 근거 ' + evidence_name if evidence_name else ''}")
    # DB 장애 때 원문을 잃지 않기 위한 안전 스풀 모양이다. 정상 경로에서는 JSON에 쓰지 않는다.
    items = [{
        "sheet": sheet, "key_col": "프로젝트NO", "key": project_no,
        "col": col, "value": value, "vtype": "date" if "일" in col else "text",
        "evidence": evidence_text, "only_if_empty": False,
    } for col, value in data.items() if value not in (None, "")]
    record = {
        "kind": kind,
        "business_key": project_no,
        "project_no": project_no,
        "camp_name": camp,
        "status": status,
        "fields": data,
        "work_date": work_date,
        "id_prefix": prefix,
        "id_col": id_col,
        "evidence": evidence_text,
        "source_ref": evidence_name,
    }
    if DEMO:
        manifest = {
            "등록일시": datetime.now().isoformat(timespec="seconds"),
            "등록자": str(fields.get("submitter") or STAFF_CENTERS["ryu-jiyeong"]["name"]),
            "입력유형": "신규 업무", "업무구분": kind,
            "업무ID": f"{prefix}-DEMO", "프로젝트NO": project_no, "캠프명": camp,
            "업무일": work_date, "담당자": tech, "상태": status,
            "내용": note, "근거": evidence_name,
        }
        return {"ok": True, "committed": False, "queued": 0, "pending": 0,
                "manifest": manifest, "applying": False, "msg": "데모 신규등록"}
    queued = enqueue_for_scheduled_apply(
        items, source="app-new-job", actor=actor,
        idempotency_key=idempotency_key, record=record,
    )
    work = (((queued.get("records") or [{}])[0]).get("work") or {})
    work_id = str(work.get("public_id") or work.get("id") or "")
    manifest = {
        "등록일시": datetime.now().isoformat(timespec="seconds"),
        "등록자": str(fields.get("submitter") or STAFF_CENTERS["ryu-jiyeong"]["name"]),
        "입력유형": "신규 업무", "업무구분": kind,
        "업무ID": work_id, "프로젝트NO": project_no, "캠프명": camp,
        "업무일": work_date, "담당자": tech, "상태": status,
        "내용": note, "근거": evidence_name,
    }
    if queued.get("ok"):
        os.makedirs(os.path.join(ROOT, "reports"), exist_ok=True)
        with open(os.path.join(ROOT, "reports", "workcenter_new_jobs.jsonl"),
                  "a", encoding="utf-8") as out:
            out.write(json.dumps(manifest, ensure_ascii=False) + "\n")
    return {**queued, "manifest": manifest}


def rows_xlsx(payload):
    """담당자 회신용 독립 XLSX를 메모리에서 만든다(관리대장은 절대 열어 저장하지 않는다)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    title = str(payload.get("title") or "확인목록")[:80]
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    rows = [r for r in rows[:5000] if isinstance(r, dict)]
    columns = [
        ("프로젝트NO", "프로젝트NO"), ("업무ID", "업무ID"), ("캠프명", "캠프명"),
        ("구분", "구분"), ("확인사항", "확인사항"), ("현재상태", "현재상태"),
        ("기준일자", "기준일자"), ("담당자", "담당자"),
        ("담당자 입력", "담당자입력"), ("처리결과", "처리결과"),
        ("완료일", "완료일"), ("첨부파일·근거", "첨부파일근거"), ("회신메모", "회신메모"),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "담당자 회신"
    ws.append([x[0] for x in columns])
    head_fill = PatternFill("solid", fgColor="203A75")
    thin = Side(style="thin", color="D9E1EF")
    for c in ws[1]:
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=thin)

    def safe(v):
        if v is None:
            return ""
        s = str(v)
        return "'" + s if s.startswith(("=", "+", "-", "@")) else s

    for r in rows:
        ws.append([safe(r.get(key, "")) for _label, key in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{max(1, ws.max_row)}"
    widths = [18, 16, 28, 16, 42, 16, 14, 15, 22, 18, 14, 32, 38]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    guide = wb.create_sheet("작성안내")
    guide.append(["항목", "작성 방법"])
    guide.append(["담당자 입력", "확인한 사실이나 실제 조치내용을 적습니다."])
    guide.append(["처리결과", "완료 / 진행중 / 확인불가 중 하나를 적습니다."])
    guide.append(["완료일", "YYYY-MM-DD 형식으로 적습니다."])
    guide.append(["첨부파일·근거", "파일명, 밴드 글, 카카오톡 근거 또는 URL을 적습니다."])
    guide.append(["회신메모", "추가 확인이 필요한 내용을 자유롭게 적습니다."])
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 75
    for c in guide[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue(), title

# 브루트포스 차단: IP당 로그인 5회 실패 → 10분 잠금 (외부 터널 공개 대비)
_fails = {}
def _locked(ip):
    c, until = _fails.get(ip, (0, 0))
    return c >= 5 and time.time() < until
def _fail(ip):
    c, _ = _fails.get(ip, (0, 0))
    _fails[ip] = (c + 1, time.time() + 600)
def _ok_login(ip):
    _fails.pop(ip, None)

# ───────────────────────── 작업 러너 ─────────────────────────
TASKS = {
    "automation":    ("자료 자동수집 후 앱 DB·Excel 보관본 반영",
                      [os.path.join(ROOT, "automation_pipeline.py"), "--once",
                       "--trigger", "app"]),
    "daily":         ("전체 대조 실행", [os.path.join(ROOT, "daily_run.py")]),
    "synthetic":     ("합성검증", [os.path.join(ROOT, "tests", "synthetic_check.py")]),
    "writer_prev":   ("자동입력 미리보기", [os.path.join(ROOT, "ledger_writer.py")]),
    # 예전 키는 설치된 앱·브라우저 캐시와의 호환 때문에 유지한다. 동작은 DB 장애 때
    # 남긴 안전 스풀을 정본 DB로 회수하는 것이며 Excel을 정본으로 읽어들이지 않는다.
    "writer_apply":  ("입력 DB 적재", [os.path.join(ROOT, "ledger_db.py"), "--intake"]),
    # 앱 DB가 정본이다. 이 키는 더 이상 원본 Excel을 직접 고치지 않고, 로컬 스냅샷과
    # 명령계획을 만든 뒤 별도 렌더·검증 단계가 이어받을 보관본 작업만 준비한다.
    "ledger_now":    ("보관본 지금 생성",
                      [os.path.join(ROOT, "archive_worker.py"), "--run"]),
    "upload_dry":    ("전표 전송대기 확인", [os.path.join(ROOT, "ecount_upload.py")]),
    # "upload_post" 제거(2026-08-05 사용자 지시) — ERP 실전송은 앱·AI가 하지 않는다.
    # 옛 앱·브라우저 캐시가 이 키를 보내도 아래 가드가 거부한다. 되살리지 말 것.
    "kakao":         ("카톡 대조", [os.path.join(ROOT, "kakao", "kakao_reconcile.py")]),
    "erp_ledger":    ("ERP원장 대조", [os.path.join(ROOT, "erp_ledger_check.py")]),
    "po":            ("쿠팡 PO 대조", [os.path.join(ROOT, "po_reconcile.py")]),
    "work_log":      ("정기점검·돌발AS 대표보고 일지 대조",
                      [os.path.join(ROOT, "work_log_sync.py"), "--apply"]),
    "erp_docs":      ("ERP 매출서류 대조", [os.path.join(ROOT, "erp_docs_check.py")]),
    # 입금 자료 업로드 직후 즉시 대조·큐 적재(엑셀 쓰기는 11·15시 회차) — 2026-07-31
    "receipt":       ("입금 대조·자동입력", [os.path.join(ROOT, "receipt_fill.py"), "--queue"]),
    "band_ingest":   ("밴드 수집분 반영(24시트+백필)",
                      [os.path.join(ROOT, "band", "ingest.py"), "--sheet", "--backfill"]),
    "band_docs":     ("밴드 문서 이미지 대조", [os.path.join(ROOT, "band", "doc_ocr.py"), "--scan"]),
    "band_docs_apply": ("밴드 문서 → 대장 입력", [os.path.join(ROOT, "band", "doc_ocr.py"), "--scan", "--apply"]),
    # 확인 필요 화면의 안전 실행: 객관완료 DB·ERP 매칭키·실제 미발행만 다시 본다.
    # 계산서 발행·외부 전송·Excel 저장은 evidence_refresh.py 자체가 하지 않는다.
    "evidence_sync": ("확인 필요 근거 재대조", [os.path.join(ROOT, "evidence_refresh.py")]),
}
TASK_TIMEOUTS = {"daily": 9000, "automation": 1800, "ledger_now": 1800,
                 "synthetic": 1200, "evidence_sync": 4500,
                 "band_docs": 2400, "band_docs_apply": 2400, "work_log": 2400}
runner = {"busy": False, "task": "", "log": deque(maxlen=3000), "done_at": None,
          "agent_route": ""}
_rlock = threading.Lock()


_codes_cache = {"t": 0, "v": None}


def get_codes():
    """드롭다운 선택지를 **10_코드관리 시트에서** 읽어 온다.

    화면에 목록을 박아 두면 사람이 바뀔 때마다 코드를 고쳐야 하고, 결국 시트와 어긋난다.
    시트가 진실이므로 거기서 읽는다(류지영 매니저가 시트만 고치면 앱도 따라간다).
    관리자검증상태는 10시트에 없어 기본값을 함께 준다.
    """
    if _codes_cache["v"] and time.time() - _codes_cache["t"] < 300:
        return _codes_cache["v"]
    out = {"관리자검증상태": ["일치", "추가작업발생", "작업내용누락", "확인필요"]}
    try:
        import openpyxl
        from ecount_reconcile import load_config, resolve_master
        wb = openpyxl.load_workbook(
            master_stream(resolve_master(load_config()["reconcile"]["master_xlsx"])),
            read_only=True, data_only=True)
        ws = wb["10_코드관리"]
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        if rows:
            hdr = [str(h).strip() if h else "" for h in rows[0]]
            for i, name in enumerate(hdr):
                if not name:
                    continue
                vals = []
                for r in rows[1:]:
                    v = r[i] if i < len(r) else None
                    if v not in (None, "") and str(v).strip() not in vals:
                        vals.append(str(v).strip())
                if vals:
                    out[name] = vals
        wb.close()
    except Exception as e:
        out["_error"] = str(e)[:80]
    _codes_cache.update({"t": time.time(), "v": out})
    return out


def enqueue_codes(codes):
    """폰 예약을 앱 DB에 즉시 저장하고 다음 보관본 생성 outbox에 등록한다."""
    import project_resolve as P
    ev = P.evidence()
    items, done, skip = [], [], []
    for c in codes:
        r = P.resolve(c, ev)
        if not r.get("ok"):
            skip.append({"code": c, "why": r.get("reason", "형식 오류")})
        elif not app_project_result(c, r):
            skip.append({"code": c, "why": "2026년 업무로 확인되지 않아 제외"})
        elif r["state"] == "등록됨":
            skip.append({"code": c, "why": f"이미 {r['sheet']} {r.get('row')}행에 있습니다"})
        else:
            items += P.row_items(r, ev)
            done.append(c)
            # 같은 요청에 두 건이 오면 뒤엣것이 같은 행을 노린다 — 자리를 미리 물린다
            ev["tail"][r["sheet"]] = r["row"]
    if not items:
        return {"ok": True, "queued": 0, "applied": 0, "skipped": skip}
    queued = enqueue_for_scheduled_apply(items, source="phone-reservation")
    runner["log"].append(
        f"[폰 예약] {len(done)}건 DB 저장 — Excel은 다음 11:00·15:00 보관본 생성")
    return {"ok": True, "applied": 0, "codes": done, "skipped": skip, **queued}


def start_task(key):
    with _rlock:
        if runner["busy"]:
            return False, "다른 작업 실행 중"
        # ERP 실전송은 앱에서 하지 않는다(2026-08-05 사용자 지시). 옛 화면·캐시가
        # 이 키를 보내도 여기서 끊는다 — 되돌릴 수 없는 등록을 버튼 하나로 만들지 않는다.
        if key == "upload_post":
            return False, "전표 실전송은 제공하지 않습니다 — ERP에서 사람이 직접 등록하세요"
        if key not in TASKS:
            return False, "알 수 없는 작업"
        if DEMO:
            runner["log"].append(f"[데모] '{TASKS[key][0]}' — 합성 환경에서는 실행을 시뮬레이션합니다.")
            return True, "demo"
        # 성공하는 결정론적 작업을 매번 Claude/Codex에 넘기지 않는다. 같은 안전 작업이
        # 세 번 실패했을 때만 autopilot이 AI 티켓을 만든다(크레딧 절약 + 중복 작업 방지).
        runner["agent_route"] = "결정론적 로컬 실행 · 반복 실패만 AI 인계"
        runner["busy"], runner["task"] = True, TASKS[key][0]
        runner["log"].clear()

    def work():
        title, args = TASKS[key]
        local_returncode = 1
        timed_out = False
        timeout = TASK_TIMEOUTS.get(key, 1800)
        runner["log"].append(f"===== {title} 시작 {datetime.now():%H:%M:%S} =====")
        runner["log"].append(f"[자동화] {runner['agent_route']} · 제한시간 {timeout // 60}분")
        try:
            if key == "ledger_now" and not args:
                prepared = _prepare_archive_export(actor="admin-task")
                output = json.dumps(prepared, ensure_ascii=False, default=str)
                runner["log"].append(prepared["msg"])
                runner["log"].append(f"계획: {prepared.get('export_id', '')}")
                local_returncode = 0
                runner["log"].append("===== 보관본 준비 종료 (코드 0) =====")
            else:
                from proc_guard import run_tree
                result = run_tree([PY] + args, cwd=ROOT, env=ENV, timeout=timeout,
                                  drain_timeout=30, output_limit=240_000)
                output = (result.stdout or "") + (("\n" + result.stderr) if result.stderr else "")
                for ln in output.splitlines():
                    if "UserWarning" not in ln and "warn(msg)" not in ln:
                        runner["log"].append(ln.rstrip())
                local_returncode = result.returncode
                timed_out = bool(result.timed_out)
                if result.timed_out:
                    runner["log"].append(
                        f"===== 제한시간 종료 · 자식나무 정리 (pid {result.stuck_pid or '정리됨'}) =====")
                else:
                    runner["log"].append(f"===== 종료 (코드 {result.returncode}) =====")
            try:
                import autopilot
                if local_returncode == 0 and not timed_out:
                    autopilot.resolve(title, args)
                else:
                    item = autopilot.defer(title, args, timeout,
                                           output or f"returncode={local_returncode}")
                    if item:
                        runner["log"].append("[자동복구] 안전 재시도 대기열에 기록")
            except Exception as exc:
                runner["log"].append(f"[자동복구] 상태 기록 실패: {str(exc)[:160]}")
        except Exception as e:
            runner["log"].append(f"오류: {e}")
            try:
                import autopilot
                item = autopilot.defer(title, args, timeout, f"{type(e).__name__}: {e}")
                if item:
                    runner["log"].append("[자동복구] 안전 재시도 대기열에 기록")
            except Exception:
                pass
        finally:
            if key == "evidence_sync" and local_returncode == 0:
                # DB 완료가 늘어났으므로 현재값만 무효화한다. last-good(stale)는 남겨 화면을
                # 멈추지 않고, 다음 API 조회가 뒤에서 새 값을 한 번만 만든다.
                for cache_key in ("settle", "issues", "exec"):
                    _cache.pop(cache_key, None)
                    _cache.pop(cache_key + "_ts", None)
            runner["busy"], runner["done_at"] = False, datetime.now().isoformat()
            _note_last_run(key, title, local_returncode)
            # 업로드해 둔 것의 결과를 **여기서 바로** 확인한다. 화면을 열 때만
            # 확인하면 아무도 안 보는 동안 끝난 회차는 알림이 안 나간다.
            # 판정은 notify 가 회차 자국을 읽어서 한다 — 여기서 다시 세지 않는다.
            try:
                import notify
                notify.sweep_uploads()
            except Exception:
                pass
    threading.Thread(target=work, daemon=True).start()
    return True, "started"


# ── 마지막 실행 시각 (2026-08-08 지시: "이 화면 상단에 최근 마지막 실행 날짜 시간") ──
#   ★ 메모리에 두면 안 된다. 이 서버는 코드를 고칠 때마다 다시 뜬다(하루에도 여러 번).
#     그때마다 화면이 "실행 기록 없음"으로 돌아가면, 오늘 아침에 돌린 대조까지
#     안 돈 것처럼 보인다 — 없는 것보다 나쁜 표시다. 그래서 파일에 남긴다.
LAST_RUN_PATH = os.path.join(ROOT, "reports", "작업_마지막실행.json")


def _note_last_run(key, title, returncode):
    try:
        d = last_runs()
        d[key] = {"제목": title, "끝난시각": datetime.now().isoformat(timespec="seconds"),
                  "코드": int(returncode if returncode is not None else -1)}
        os.makedirs(os.path.dirname(LAST_RUN_PATH), exist_ok=True)
        tmp = LAST_RUN_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(d, fh, ensure_ascii=False, indent=1)
        os.replace(tmp, LAST_RUN_PATH)          # 반쯤 쓰인 파일을 남기지 않는다
    except Exception:
        pass                                     # 기록에 실패해도 작업 자체는 끝난 것이다


def last_runs(merge_auto=False):
    """{작업키: {제목, 끝난시각, 코드}} — 없으면 빈 딕셔너리.

    `merge_auto` 면 **앱 밖에서 돈 것**도 합친다. 전체 대조는 09:50 스케줄러가
    매일 돌리는데, 앱 단추 기록만 보여 주면 오늘 아침에 이미 돈 대조를 두고
    화면이 "3일 전"이라 말한다 — 사람이 그걸 보고 또 누른다.
    근거는 daily_run 이 끝나며 남기는 `reports/agent_status.json` 의 `time` 이다.
    """
    try:
        with open(LAST_RUN_PATH, encoding="utf-8") as fh:
            d = json.load(fh)
        d = d if isinstance(d, dict) else {}
    except Exception:
        d = {}
    if merge_auto:
        try:
            with open(os.path.join(ROOT, "reports", "agent_status.json"),
                      encoding="utf-8") as fh:
                st = json.load(fh)
            t = str(st.get("time") or "")[:19]
            if t and t > str((d.get("daily") or {}).get("끝난시각") or ""):
                d["daily"] = {"제목": TASKS["daily"][0], "끝난시각": t,
                              "코드": 1 if st.get("aborted") else 0, "자동": True}
        except Exception:
            pass
    return d


_deferred_tasks = set()


def defer_task_until_free(key, max_wait_seconds=1800):
    """다른 작업 중이면 사람이 다시 누르지 않아도 끝나는 즉시 한 번 실행한다."""
    with _rlock:
        if key not in TASKS or key in _deferred_tasks:
            return False
        _deferred_tasks.add(key)

    def wait_and_start():
        try:
            deadline = time.time() + max_wait_seconds
            while time.time() < deadline:
                with _rlock:
                    busy = bool(runner["busy"])
                if not busy:
                    ok, _ = start_task(key)
                    if ok:
                        return
                time.sleep(5)
            runner["log"].append(f"[자동 대기] {TASKS[key][0]} 실행 대기 시간이 초과되었습니다.")
        finally:
            with _rlock:
                _deferred_tasks.discard(key)

    threading.Thread(target=wait_and_start, daemon=True).start()
    return True


# ───────────────────────── 데이터 ─────────────────────────
_cache = {"t": 0, "settle": None, "status": None}
_readlock = threading.RLock()  # Z:드라이브 엑셀 동시 읽기 직렬화(스레드 충돌 방지)
# ★ RLock이어야 한다: 정산 조회가 락을 쥔 채 업무 조회(대표번호 색인용)를 부르므로
#   일반 Lock이면 같은 스레드에서 자기 자신을 기다리다 멈춘다(실제로 응답 없음 발생)


# ── 날짜 정렬 공통 규칙 ────────────────────────────────────────
# 앱·리포트 어디서나 **과거가 맨 위, 최근이 맨 아래**(오름차순)로 통일한다.
# 새로 추가되는 행도 반드시 이 함수를 거치므로 따로 정렬해 줄 필요가 없다.
DATE_KEYS = {
    "settle": ("완료일", "계산서발행일", "명세서발행일", "입금일"),
    "as":     ("접수일자", "작업완료일", "방문예정일"),
    "pm":     ("점검예정일", "실제점검일"),
}
_DATE_RE = re.compile(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})")
APP_YEAR = "2026"
APP_YEAR_SHORT = APP_YEAR[-2:]
_APP_PROJECT_RE = re.compile(r"(?<![A-Za-z0-9])UJ(?P<yy>\d{2})\d{5}(?!\d)", re.I)
_APP_ID_RE = re.compile(r"(?<![A-Za-z0-9])(?:AS|PM|JS)-(?P<yy>\d{2})\d{2}(?:-|$)", re.I)
_APP_SLIP_RE = re.compile(r"(?<!\d)(?P<yy>\d{2})/\d{2}/\d{2}\s*-\s*\d+")
_OLD_APP_REF_RE = re.compile(
    r"(?<![A-Za-z0-9])UJ25\d{5}(?!\d)|"
    r"(?<![A-Za-z0-9])(?:AS|PM|JS)-25\d{2}-\d{3}(?!\d)|"
    r"(?<!\d)2025[-./]\d{1,2}(?:[-./]\d{1,2})?|2025년|(?<!\d)25(?:년도|년)",
    re.I,
)


def norm_date(v):
    """'2026.6.3' · '2026-06-03 00:00' → '2026-06-03' (문자열 비교로 시간순이 되게)"""
    m = _DATE_RE.search(str(v or ""))
    return "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3))) if m else ""


def row_date(rec, keys=()):
    """행의 대표 날짜. 지정 키를 우선 보고, 없으면 아무 날짜 값이나 찾아 쓴다."""
    for k in keys:
        d = norm_date(rec.get(k))
        if d:
            return d
    for v in rec.values():
        d = norm_date(v)
        if d:
            return d
    return ""


def app_year_record(rec, kind=None):
    """앱에는 2026년 업무만 노출한다.

    원본 엑셀은 그대로 두고 표시 경계에서만 판정한다. 프로젝트NO/업무ID가 있으면
    그것을 날짜보다 우선하며, ERP 묶음처럼 번호가 없는 행은 월·전표·날짜로 판정한다.
    연도를 확인할 수 없는 행도 섞어 보여 주지 않고 제외한다.
    """
    if not isinstance(rec, dict):
        return False

    def years(pattern, values):
        return {m.group("yy") for v in values for m in pattern.finditer(str(v or ""))}

    def date_year(keys):
        vals = [rec.get(k) for k in keys]
        found = {m.group(1) for v in vals
                 for m in re.finditer(r"(?<!\d)(20\d{2})[-./]", str(v or ""))}
        if found:
            return APP_YEAR if found == {APP_YEAR} else "other"
        short = years(_APP_SLIP_RE, vals)
        if short:
            return APP_YEAR if short == {APP_YEAR_SHORT} else "other"
        return ""

    def id_year(keys):
        found = years(_APP_ID_RE, [rec.get(k) for k in keys])
        if found:
            return APP_YEAR if found == {APP_YEAR_SHORT} else "other"
        return ""

    def project_year():
        found = years(_APP_PROJECT_RE,
                      [rec.get(k) for k in ("프로젝트NO", "포함프로젝트", "프로젝트명")])
        if found:
            return APP_YEAR if found == {APP_YEAR_SHORT} else "other"
        return ""

    # 데이터 종류마다 '그 건의 연도'를 정하는 열이 다르다. 수정일·확인일에 2026이
    # 찍혔다고 2025 업무를 되살리지 않도록 업무 발생일을 가장 먼저 본다.
    rules = {
        "as": (("접수일자",), ("접수ID", "업무ID")),
        "pm": (("점검예정일",), ("점검ID", "업무ID")),
        "settle": (("완료일",), ("원천업무ID", "정산ID", "업무ID")),
        "erp": (("월", "전표"), ()),
        "visit": (("방문일",), ()),
        "visit_pending": (("예정일",), ()),
        "unbilled": (("발행일",), ()),
        "issue": (("기준일", "접수일자", "점검예정일", "완료일", "발행일", "일자"),
                  ("업무ID", "접수ID", "점검ID", "정산ID", "ID")),
    }
    if kind in rules:
        date_keys, id_keys = rules[kind]
        y = date_year(date_keys)
        if y:
            return y == APP_YEAR
        y = id_year(id_keys) if id_keys else ""
        if y:
            return y == APP_YEAR
        y = project_year()
        return y == APP_YEAR

    # 표준 종류를 모르는 행은 실제 업무 날짜를 먼저 찾고, 없을 때만 ID·프로젝트로
    # 보완한다. 여러 연도가 함께 든 혼합 행은 통째로 제외한다.
    for keys in (("접수일자",), ("점검예정일",), ("완료일",), ("월", "전표"),
                 ("방문일",), ("발행일",), ("기준일", "일자")):
        y = date_year(keys)
        if y:
            return y == APP_YEAR
    y = id_year(("업무ID", "접수ID", "점검ID", "정산ID", "원천업무ID", "ID"))
    if y:
        return y == APP_YEAR
    y = project_year()
    return y == APP_YEAR


def app_year_rows(rows, kind=None):
    """2026년으로 판정되는 행만 새 목록으로 돌려준다."""
    out = []
    for r in rows:
        if not app_year_record(r, kind):
            continue
        clean = {}
        for k, v in r.items():
            if isinstance(v, str):
                v = _OLD_APP_REF_RE.sub("", v)
                v = re.sub(r"\s*[,·/]\s*(?=([,·/]|$))", "", v).strip(" ,·/")
            clean[k] = v
        out.append(clean)
    return out


def app_project_result(code, result):
    """프로젝트 자동조회 결과도 코드뿐 아니라 내부 날짜·ID까지 2026년인지 검사한다.

    UJ26 코드에 과거 AS-25 작업이 잘못 연결된 실데이터가 있으므로 프로젝트 코드만
    보고 통과시키면 오프라인 앱에서 2025년 내용이 다시 노출된다.
    """
    if not re.fullmatch(r"UJ26\d{5}", str(code or ""), re.I):
        return False
    if not isinstance(result, dict):
        return False
    blob = json.dumps(result, ensure_ascii=False, default=str)
    date_years = set(re.findall(r"(?<!\d)(20\d{2})[-./]", blob))
    id_years = {m.group("yy") for m in _APP_ID_RE.finditer(blob)}
    project_years = {m.group("yy") for m in _APP_PROJECT_RE.finditer(blob)}
    if date_years and date_years != {APP_YEAR}:
        return False
    if id_years and id_years != {APP_YEAR_SHORT}:
        return False
    if project_years and project_years != {APP_YEAR_SHORT}:
        return False
    return True


def sort_by_date(rows, kind, idkey=None):
    """과거 → 최근. 날짜가 없는 행은 맨 뒤(=가장 최근으로 취급), 동률은 ID순."""
    keys = DATE_KEYS.get(kind, ())
    return sorted(rows, key=lambda r: ((d := row_date(r, keys)) == "", d,
                                       str(r.get(idkey) or "") if idkey else ""))


def demo_settlements():
    camps = ["송파5MB(감일동)", "울산2캠프", "인천7MB(마곡동)", "부천3(BUC3)", "대전1캠프",
             "구리1캠프", "제주1Sub-hub", "창원1MB(팔용동)", "군포1Sub-Hub", "광주2Sub-hub"]
    rows = []
    rnd = random.Random(42)
    for i in range(1, 16):
        amt = rnd.choice([380000, 418000, 470800, 760000, 1230000, 1472500])
        st = rnd.choice(["정상", "세금계산서 미발행", "ERP 미확인", "미청구", "입금 대기"])
        d = (date(2026, 7, 1) + timedelta(days=i)).isoformat()
        rows.append({"정산ID": f"JS-2607-{i:03d}", "업무구분": rnd.choice(["돌발AS", "정기점검"]),
                     "캠프명": camps[i % len(camps)], "프로젝트NO": f"UJ26{1000+i}",
                     "원천업무ID": f"AS-2607-{i:03d}",
                     "공급가액": amt, "부가세": int(amt * 0.1), "합계": int(amt * 1.1),
                     "명세서": "있음" if st != "미청구" else "없음",
                     "명세서번호": f"2026/07/{i:02d}-1" if st != "미청구" else "",
                     "명세서발행일": d if st != "미청구" else "",
                     "계산서": "발행" if st == "정상" else "미발행",
                     "계산서발행일": d if st == "정상" else "", "승인번호": "",
                     "입금일": d if st == "정상" else "", "입금액": int(amt * 1.1) if st == "정상" else 0,
                     "미수금": 0 if st == "정상" else int(amt * 1.1), "비용구분": "유상",
                     "상태": st, "완료일": d})
    return sort_by_date(app_year_rows(rows, "settle"), "settle", "정산ID")


def _mark_app_store_row(row):
    if row.get("_store_id"):
        row["앱DB_ID"] = row.get("_store_id")
        row["DB버전"] = row.get("_record_version") or 0
        row["출처"] = "앱 DB"
    return row


def _sidecar_day(value):
    """Normalize a sidecar date without inventing a day from non-date text."""
    raw = str(_ryu_display_value(value) or "").strip()[:10]
    return raw if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw) else ""


def _sidecar_amounts(rows):
    """Return the exact sum of nonblank receipt/invoice amounts, or None."""
    total = Decimal("0")
    seen = False
    for value in rows:
        if value in (None, "", "-"):
            continue
        cleaned = re.sub(r"[^0-9.\-]", "", str(value))
        if cleaned in ("", "-", ".", "-."):
            continue
        try:
            total += Decimal(cleaned)
            seen = True
        except InvalidOperation:
            continue
    if not seen:
        return None
    return int(total) if total == total.to_integral_value() else float(total)


def _sidecar_ids(rows, field):
    return sorted({str(row.get(field) or "").strip() for row in rows
                   if str(row.get(field) or "").strip()})


def _source_outcome_map(store=None):
    """원천업무 ID별 취소·원격해결 정본. 프로젝트 유사매칭은 하지 않는다."""
    try:
        from cancel_resolution import source_outcomes
        return source_outcomes(store=store)
    except Exception:
        return {}


def _settlement_has_documents(row):
    """취소 뒤에도 보존·확인해야 할 실제 청구자료가 하나라도 있는가."""
    keys = (
        "거래명세서번호", "명세서번호", "거래명세서발행일", "명세서발행일",
        "세금계산서발행일", "계산서발행일", "세금계산서승인번호", "승인번호",
        "PO번호", "PO발행일", "청구일", "입금일", "입금액",
    )
    return any(str(row.get(key) or "").strip() not in ("", "0", "0.0") for key in keys)


def _apply_cancelled_source_to_settlement(row, outcome):
    """취소·청구제외건은 미발행 경고에서 빼되 청구자료가 있으면 충돌로 보존한다.

    ★ 빠지는 결과는 같아도 **이유는 두 가지고 뜻이 다르다** (2026-08-13 류지영 요청).
      · 취소   = 현장에 아무도 안 갔다.
      · 제외   = 다녀왔는데 이 건으로는 청구를 안 한다(정기점검 동시 진행 등).
      뭉쳐서 '취소'라고 적으면 다녀온 사실이 기록에서 뒤집히고, 기사 실적이 줄고,
      나중에 "그때 정말 안 갔었나"를 아무도 답할 수 없다([243] 과 같은 자리).
      그래서 전파 경로는 하나로 두되(사본을 만들지 않는다, [162]) **딱지를 가른다.**
    """
    if not outcome:
        return row
    if not outcome.get("cancelled"):
        if outcome.get("billing_excluded"):
            return _apply_excluded_source_to_settlement(row, outcome)
        return row
    row["원천업무취소"] = True
    row["원천업무상태"] = outcome.get("status") or "취소"
    row["취소사유"] = outcome.get("reason") or "접수취소"
    row["처리구분"] = outcome.get("treatment") or "접수취소"
    row["취소근거"] = outcome.get("evidence") or ""
    row["청구대상"] = False
    row["청구제외사유"] = row["취소사유"]
    row["미발행사유"] = ""
    if _settlement_has_documents(row):
        row["취소건청구자료충돌"] = True
        row["상태"] = "취소건 청구자료 존재 — 교차입력 확인"
    else:
        row["취소건청구자료충돌"] = False
        row["상태"] = ("접수취소(유선해결)" if row["처리구분"] == "원격해결"
                       else "접수취소(청구대상 아님)")
    return row


def _apply_excluded_source_to_settlement(row, outcome):
    """'다녀왔지만 이 건으로는 청구 안 함' — 청구에서만 빼고 현장 기록은 그대로 둔다.

    ⚠ 청구자료(PO·명세서·계산서·입금)는 **한 글자도 지우지 않는다**([208] 그대로).
      이미 끊긴 서류가 있는데 제외로 접으면 그 돈이 어느 화면에도 안 뜬다 —
      못 받는 것보다 나쁜 것은 못 받는 줄도 모르는 것이다.
    """
    row["원천업무청구제외"] = True
    row["원천업무상태"] = outcome.get("status") or ""
    row["청구대상"] = False
    row["청구제외사유"] = outcome.get("excluded_reason") or "이 건으로는 청구하지 않음"
    row["미발행사유"] = ""
    if _settlement_has_documents(row):
        # 취소와 같은 칸을 쓴다 — 화면이 이미 이 칸을 보고 '교차입력 확인'을 띄운다.
        row["취소건청구자료충돌"] = True
        row["상태"] = "청구제외건 청구자료 존재 — 교차입력 확인"
    else:
        row["취소건청구자료충돌"] = False
        row["상태"] = "청구 제외(다녀옴 · 이 건으로는 청구 안 함)"
    return row


def _overlay_app_store_settlements(rows, store=None):
    """Excel 읽기모델 위에 정본 DB와 계산서·입금 sidecar를 덮는다."""
    try:
        if store is None:
            from app_store import default_store
            store = default_store()
        merged = store.overlay_rows("06_거래서류청구수금", rows, "정산ID")
        source_cost = {}
        source_outcome = _source_outcome_map(store)
        for sheet, key_col in (("02_돌발AS접수", "접수ID"), ("04_정기점검", "점검ID")):
            for source in store.list_sheet_rows(sheet):
                source_id = str(source.get(key_col) or source.get("_business_key") or "").strip()
                if source_id and "유상·무상·보험" in source:
                    source_cost[source_id] = source.get("유상·무상·보험")
        for row in merged:
            source_id = str(row.get("원천업무ID") or "").strip()
            if source_id in source_cost:
                row["비용구분"] = source_cost[source_id]
            if "거래명세서번호" in row:
                row["명세서번호"] = row.get("거래명세서번호") or ""
            if "거래명세서발행일" in row:
                row["명세서발행일"] = str(row.get("거래명세서발행일") or "")[:10]
            if "세금계산서발행일" in row:
                issued = str(row.get("세금계산서발행일") or "")[:10]
                row["계산서발행일"] = issued
                row["계산서"] = "발행" if issued else "미발행"
        by_id = {str(r.get("정산ID") or "").strip(): r for r in merged
                 if str(r.get("정산ID") or "").strip()}
        projects = {}
        for row in merged:
            pno = str(row.get("프로젝트NO") or "").strip()
            if pno:
                projects.setdefault(pno, []).append(row)
        grouped = {"15_세금계산서관리": {}, "16_입금수금관리": {}}
        target_refs = {}
        for sheet in grouped:
            id_field = "계산서관리ID" if sheet.startswith("15_") else "입금관리ID"
            for side in sorted(
                    store.list_sheet_rows(sheet),
                    key=lambda item: str(item.get(id_field) or item.get("_business_key") or "")):
                relation_key = str(side.get("정산ID") or "").strip()
                legacy_key = str(side.get("_business_key") or "").strip()
                pno = str(side.get("프로젝트NO") or "").strip()
                # 명시 정산ID가 있으면 그것만 신뢰한다. 못 찾았다고 프로젝트의 유일한
                # 다른 정산행에 붙이면 side의 정산ID가 그 행을 뒤집는 더 큰 오류가 된다.
                if relation_key:
                    target = by_id.get(relation_key)
                    if target is None:
                        target = {
                            "정산ID": relation_key,
                            "연결상태": "정산ID 미연결",
                            "프로젝트NO": pno,
                        }
                        merged.append(target)
                        by_id[relation_key] = target
                else:
                    # v586 이전 business_key=정산ID 레코드만 호환한다. 관리ID는
                    # by_id에 없으므로 프로젝트가 정확히 한 행일 때만 임시 연결한다.
                    management_id = str(side.get(id_field) or "").strip()
                    target = by_id.get(legacy_key)
                    if (target is None and not management_id and pno
                            and len(projects.get(pno) or []) == 1):
                        target = projects[pno][0]
                    if target is None:
                        target = {
                            "정산ID": "", "연결상태": "정산ID 미확인",
                            "프로젝트NO": pno,
                        }
                        merged.append(target)
                token = id(target)
                target_refs[token] = target
                grouped[sheet].setdefault(token, []).append(side)

        # 한 정산에 계산서가 여러 장일 수 있다. 빈 마지막 행이 확인된 발행일을
        # 지우지 않게 날짜는 전체를 집계하고, 일부만 발행됐으면 그대로 드러낸다.
        for token, side_rows in grouped["15_세금계산서관리"].items():
            target = target_refs[token]
            ids = _sidecar_ids(side_rows, "계산서관리ID")
            issued_row_days = [day for day in
                               (_sidecar_day(row.get("실제발행일")) for row in side_rows)
                               if day]
            issued_days = sorted(set(issued_row_days))
            prior_day = _sidecar_day(
                target.get("계산서발행일") or target.get("세금계산서발행일")
            )
            all_days = sorted(set(issued_days + ([prior_day] if prior_day else [])))
            if ids:
                target["계산서관리ID"] = " · ".join(ids)
            target["계산서건수"] = len(side_rows)
            target["계산서발행건수"] = len(issued_row_days)
            target["계산서발행일목록"] = " · ".join(issued_days)
            if all_days:
                target["계산서발행일"] = all_days[-1]
            if issued_row_days and len(issued_row_days) < len(side_rows):
                target["계산서"] = "일부 발행일 확인"
            elif issued_days or prior_day:
                target["계산서"] = "발행"
            else:
                target["계산서"] = "미발행"
            invoice_total = _sidecar_amounts(row.get("발행금액") for row in side_rows)
            if invoice_total is not None:
                target["계산서발행금액"] = invoice_total
            target["계산서앱DB_ID"] = " · ".join(
                sorted({str(row.get("_store_id") or "") for row in side_rows
                        if str(row.get("_store_id") or "")})
            )
            target["계산서DB버전"] = max(
                [int(row.get("_record_version") or 0) for row in side_rows] or [0]
            )
            target["출처"] = "앱 DB"

        # 부분입금은 각 입금 레코드의 합계와 마지막 입금일로 보여 준다. 마지막
        # side 행 하나를 대입하면 실제 수금액이 줄어들 수 있다.
        for token, side_rows in grouped["16_입금수금관리"].items():
            target = target_refs[token]
            ids = _sidecar_ids(side_rows, "입금관리ID")
            receipt_days = sorted({day for day in
                                   (_sidecar_day(row.get("입금일")) for row in side_rows)
                                   if day})
            receipt_total = _sidecar_amounts(row.get("입금액") for row in side_rows)
            if ids:
                target["입금관리ID"] = " · ".join(ids)
            target["입금건수"] = len(side_rows)
            target["입금일목록"] = " · ".join(receipt_days)
            if receipt_days:
                target["입금일"] = receipt_days[-1]
            if receipt_total is not None:
                target["입금액"] = receipt_total
            target["입금앱DB_ID"] = " · ".join(
                sorted({str(row.get("_store_id") or "") for row in side_rows
                        if str(row.get("_store_id") or "")})
            )
            target["입금DB버전"] = max(
                [int(row.get("_record_version") or 0) for row in side_rows] or [0]
            )
            target["출처"] = "앱 DB"
        for row in merged:
            source_id = str(row.get("원천업무ID") or "").strip()
            _apply_cancelled_source_to_settlement(row, source_outcome.get(source_id))
        return [_mark_app_store_row(dict(row)) for row in merged]
    except Exception:
        # 정본 DB 장애는 /api/app-store/status가 별도로 드러낸다. 기존 읽기 화면까지
        # 통째로 닫지는 않아 운영자가 상태 페이지에서 복구할 수 있게 한다.
        return [dict(row) for row in rows]


def _overlay_app_store_ledger_records(recs, store=None):
    """정산 판정 전에 앱 DB 사실을 read_ledger 모양으로 덮는다.

    예전에는 상태를 먼저 계산하고 마지막에 화면 열만 덮어, 저장된 발행일이 있어도
    `미발행` 딱지가 남았다. 명시적 빈 값도 정정 사실이므로 키 존재 여부로 덮는다.
    """
    out = {str(key): dict(value or {}) for key, value in dict(recs or {}).items()}
    try:
        if store is None:
            from app_store import default_store
            store = default_store()
        mapping = {
            "프로젝트NO": "프로젝트NO", "캠프명": "캠프명",
            "원천업무ID": "원천업무ID", "업무구분": "업무구분",
            "비용구분": "비용구분", "작업완료일": "작업완료일",
            "실제작업공급가액": "원장_공급가액", "공급가액": "원장_공급가액",
            "부가세": "원장_부가세", "합계(VAT포함)": "원장_합계", "합계": "원장_합계",
            "거래명세서번호": "원장_거래명세서번호",
            "거래명세서발행일": "원장_거래명세서발행일",
            "거래명세서합계": "원장_거래명세서합계",
            "세금계산서발행일": "원장_세금계산서발행일",
            "세금계산서승인번호": "원장_세금계산서승인번호",
            "청구일": "원장_청구일", "지급예정일": "원장_지급예정일",
            "입금일": "원장_입금일", "입금액": "원장_입금액",
            "미수금액": "원장_미수금액", "PO필요여부": "원장_PO필요여부",
            "PO번호": "원장_PO번호", "PO발행일": "원장_PO발행일",
            "청구상태": "원장_청구상태",
        }
        for row in store.list_sheet_rows("06_거래서류청구수금"):
            sid = str(row.get("정산ID") or row.get("_business_key") or "").strip()
            if not sid:
                continue
            record = out.setdefault(sid, {})
            record["정산ID"] = sid
            for source, target in mapping.items():
                if source in row:
                    record[target] = row.get(source)

        # 계산서/입금 sidecar는 정산ID가 정본 관계키다. 프로젝트NO는 같은 프로젝트에
        # 정산이 정확히 하나일 때만 과도기 호환키로 쓴다.
        by_project = {}
        for sid, record in out.items():
            pno = str(record.get("프로젝트NO") or "").strip()
            if pno:
                by_project.setdefault(pno, []).append(sid)
        invoice_groups = {}
        for side in store.list_sheet_rows("15_세금계산서관리"):
            relation = str(side.get("정산ID") or "").strip()
            if relation:
                key = relation if relation in out else ""
            else:
                management_id = str(side.get("계산서관리ID") or "").strip()
                legacy = str(side.get("_business_key") or "").strip()
                key = legacy if legacy in out else ""
                if not key and not management_id:
                    pno = str(side.get("프로젝트NO") or "").strip()
                    key = ((by_project.get(pno) or [""])[0]
                           if len(by_project.get(pno) or []) == 1 else "")
            if key:
                invoice_groups.setdefault(key, []).append(side)
        for key, side_rows in invoice_groups.items():
            row_days = [day for day in
                        (_sidecar_day(row.get("실제발행일")) for row in side_rows)
                        if day]
            days = sorted(set(row_days))
            if days:
                out[key]["원장_세금계산서실제발행일"] = days[-1]
            out[key]["원장_세금계산서건수"] = len(side_rows)
            out[key]["원장_세금계산서발행건수"] = len(row_days)
            out[key]["원장_세금계산서발행일부분확인"] = bool(
                row_days and len(row_days) < len(side_rows)
            )

        receipt_groups = {}
        for side in store.list_sheet_rows("16_입금수금관리"):
            relation = str(side.get("정산ID") or "").strip()
            if relation:
                key = relation if relation in out else ""
            else:
                management_id = str(side.get("입금관리ID") or "").strip()
                legacy = str(side.get("_business_key") or "").strip()
                key = legacy if legacy in out else ""
                if not key and not management_id:
                    pno = str(side.get("프로젝트NO") or "").strip()
                    key = ((by_project.get(pno) or [""])[0]
                           if len(by_project.get(pno) or []) == 1 else "")
            if key:
                receipt_groups.setdefault(key, []).append(side)
        for key, side_rows in receipt_groups.items():
            days = sorted({day for day in
                           (_sidecar_day(row.get("입금일")) for row in side_rows)
                           if day})
            amount = _sidecar_amounts(row.get("입금액") for row in side_rows)
            if days:
                out[key]["원장_입금일"] = days[-1]
            if amount is not None:
                out[key]["원장_입금액"] = amount
            out[key]["원장_입금건수"] = len(side_rows)

        # 06 비용구분 수식은 보관본 생성 전까지 옛 값을 보일 수 있다. 원천 AS/PM의
        # 사람이 확정한 비용 사실을 읽기모델에 즉시 연결하되 06 필드 자체는 쓰지 않는다.
        source_cost = {}
        source_outcome = _source_outcome_map(store)
        for sheet, key_col in (("02_돌발AS접수", "접수ID"), ("04_정기점검", "점검ID")):
            for row in store.list_sheet_rows(sheet):
                source_id = str(row.get(key_col) or row.get("_business_key") or "").strip()
                if source_id and "유상·무상·보험" in row:
                    source_cost[source_id] = row.get("유상·무상·보험")
        for record in out.values():
            source_id = str(record.get("원천업무ID") or "").strip()
            if source_id in source_cost:
                record["비용구분"] = source_cost[source_id]
            outcome = source_outcome.get(source_id) or {}
            if outcome.get("cancelled"):
                record["원천업무취소"] = True
                record["원천업무상태"] = outcome.get("status") or "취소"
                record["원천업무취소사유"] = outcome.get("reason") or "접수취소"
                record["원천업무처리구분"] = outcome.get("treatment") or "접수취소"
                record["원천업무취소근거"] = outcome.get("evidence") or ""
        return out
    except Exception:
        return out


def real_settlements():
    from ecount_reconcile import (read_ledger, load_config, settle_status, has_statement,
                                  supply_from_statement, erp_progress)
    try:
        _erp_progress_map = erp_progress() or {}
    except Exception:
        _erp_progress_map = {}
    try:
        from ledger_db import resolutions
        objective_done = resolutions()
    except Exception:
        objective_done = {}
    cfg = load_config()
    try:
        recs = read_ledger(cfg["reconcile"]["master_xlsx"])
    except Exception:
        # 앱 DB가 정본이므로 Excel 보관본을 못 읽어도 DB 신규행은 계속 조회된다.
        recs = {}
    recs = _overlay_app_store_ledger_records(recs)
    rows = []
    for sid, r in sorted(recs.items()):
        issued = r.get("원장_세금계산서실제발행일") or r.get("원장_세금계산서발행일")
        invoice_partial = bool(r.get("원장_세금계산서발행일부분확인"))
        invoice_count = int(r.get("원장_세금계산서건수") or 0)
        invoice_issued_count = int(r.get("원장_세금계산서발행건수") or 0)
        has_stmt = has_statement(r)
        # 판정은 ecount_reconcile.settle_status 한 곳에서만 한다 — 엑셀 산출물과 어긋나지 않게.
        resolved = objective_done.get(sid) or {}
        resolved_status = str(resolved.get("status") or "")
        st = resolved_status if resolved_status.startswith("완료(") else settle_status(r)
        issued_by_evidence = "계산서" in resolved_status and resolved_status.startswith("완료(")
        # ★ 0원 표시 해소(사용자 지시 2026-08-05) — 공급가액 출처 우선순위.
        #   ① 실제작업공급가액(06시트 수식) ② **ERP 판매조회 공급가액**(프로젝트NO 색인)
        #   ③ 거래명세서합계.
        #   ★★ ③ 은 **부가세 포함액**이다(UJ2600050: 476,300 = 433,000×1.1).
        #      첫 구현에서 이걸 "부가세 별도" 자리에 그대로 넣어 금액이 부풀었다 —
        #      그래서 ERP 공급가액을 ② 로 끼워 넣고, ③ 으로 떨어질 때는 화면이
        #      '부가세 포함'이라고 밝히도록 `금액출처`를 내려보낸다.
        _erp = _erp_sales_index().get(str(r.get("프로젝트NO") or "")) or {}
        # ★ 사용자 지시(2026-08-08): 계산서 발행율이 0.9% 로 나오던 것은 미발행이 아니라
        #   **06시트 '계산서' 칸이 사람 손 입력이라 비어 있어서**였다(유상 716건 중 6건).
        #   ERP 가 '6.세금계산서발행'·'7.수금완료' 라고 말하면 그건 이미 나간 것이다.
        #   판정은 erp_progress() 를 쓴다 — 한 프로젝트에 상태가 섞이면 '혼재(...)' 를
        #   돌려주므로 여기서도 발행으로 세지 않는다(settle_status 와 같은 근거).
        _erp_state = str(_erp_progress_map.get(str(r.get("프로젝트NO") or "")) or "")
        _erp_issued = _erp_state[:2] in ("6.", "7.")
        if invoice_partial and not resolved_status.startswith("완료(") and not _erp_issued:
            st = "세금계산서 발행일 일부 확인"
        # ★ '미발행' 한 덩어리 안에 **가야 할 사람이 다른 두 가지**가 섞여 있었다
        #   (2026-08-08 사용자 질문 "작업은 완료인데 왜 계산서 발행이 안된거지").
        #   · ERP 가 `4.세금계산서발행대기` 라고 적어 둔 것 — PO 도 왔고 금액도 맞고
        #     명세서도 나갔다. ERP 에서 발행으로 넘기는 **류지영 손**만 남았다(실측 122건).
        #   · ERP 색인에 **전표 자체가 없는** 것 — 발행 대기가 아니라 아직 안 올라간 것.
        #     이건 전표 등록부터다. (실측 UJ2600015)
        #   같은 빨간 딱지로 묶어 두면 화면이 "누구에게 넘길지"를 말해 주지 못한다.
        if _erp_issued:
            _why = ""
        elif _erp_state[:2] == "4.":
            _why = "발행 대기(ERP 4단계)"
        elif not _erp_state:
            _why = "ERP 전표 없음"
        else:
            _why = "ERP " + _erp_state
        if r.get("원장_공급가액"):
            amt, src = int(r["원장_공급가액"]), "실제작업"
        elif _erp.get("supply"):
            amt, src = int(_erp["supply"]), "ERP"
        elif supply_from_statement(r.get("원장_거래명세서합계")) is not None:
            # ★ 2026-08-08 지시로 ÷1.1 환산이 확정 금액이 됐다. 예전처럼 부가세 포함액을
            #   그대로 싣지 않는다 — 그 값이 '부가세 별도' 칸에 앉아 금액을 10% 부풀렸다.
            amt, src = supply_from_statement(r["원장_거래명세서합계"]), "명세서(부가세 환산)"
        elif r.get("원장_거래명세서합계"):
            amt, src = int(r["원장_거래명세서합계"]), "명세서(부가세포함)"
        else:
            amt, src = 0, ""
        rows.append({"정산ID": sid, "업무구분": r.get("업무구분"), "캠프명": r.get("캠프명"),
                     "프로젝트NO": r.get("프로젝트NO"), "원천업무ID": r.get("원천업무ID"),
                     "공급가액": amt, "합계": r.get("원장_합계") or 0,
                     "금액출처": src, "ERP진행상태": _erp.get("state") or "",
                     "거래처코드": (_customer_index().get(str(r.get("캠프명") or "")) or {}).get("code", ""),
                     "부가세": r.get("원장_부가세"),
                     "명세서": "있음" if has_stmt else "없음",
                     "명세서번호": r.get("원장_거래명세서번호") or "",
                     "명세서발행일": str(r.get("원장_거래명세서발행일") or "")[:10],
                     "계산서": ("발행(근거확인)" if issued_by_evidence else
                              "발행(ERP확인)" if _erp_issued else
                              "일부 발행일 확인" if invoice_partial else
                              "발행" if issued else "미발행"),
                     "계산서발행일": str(issued or "")[:10],
                     "계산서건수": invoice_count,
                     "계산서발행건수": invoice_issued_count,
                     "미발행사유": (
                         f"발행일 {invoice_issued_count}/{invoice_count}건 확인"
                         if invoice_partial else
                         "" if (issued or issued_by_evidence or _erp_issued) else _why
                     ),
                     "승인번호": r.get("원장_세금계산서승인번호") or "",
                     "청구일": str(r.get("원장_청구일") or "")[:10],
                     "지급예정일": str(r.get("원장_지급예정일") or "")[:10],
                     "입금일": str(r.get("원장_입금일") or "")[:10],
                     "입금액": r.get("원장_입금액") or 0,
                     "미수금": r.get("원장_미수금액") if r.get("원장_미수금액") is not None else "",
                     "비용구분": r.get("비용구분"),
                     "PO필요": r.get("원장_PO필요여부") or "",
                     "PO번호": r.get("원장_PO번호") or "",
                     "PO발행일": str(r.get("원장_PO발행일") or "")[:10],
                     "상태": st, "완료일": str(r.get("작업완료일") or "")[:10],
                     "완료근거": resolved.get("basis") or "",
                     "완료확인일": str(resolved.get("first_seen") or "")[:10]})
    rows = _overlay_app_store_settlements(rows)
    return sort_by_date(app_year_rows(rows, "settle"), "settle", "정산ID")


def _overlay_app_store_works(out):
    try:
        from app_store import default_store
        store = default_store()
        out["as"] = store.overlay_rows("02_돌발AS접수", out.get("as") or [], "접수ID")
        out["pm"] = store.overlay_rows("04_정기점검", out.get("pm") or [], "점검ID")
        for kind in ("as", "pm"):
            # 컷오버 전 Excel 수식 ID 캐시가 비었어도 같은 프로젝트를 두 카드로 만들지
            # 않는다. 프로젝트가 양쪽에서 각각 정확히 한 건일 때만 보수적으로 합친다.
            db_by_project = {}
            legacy_by_project = {}
            for row in out[kind]:
                pno = str(row.get("프로젝트NO") or "").strip().upper()
                if not pno:
                    continue
                target = db_by_project if row.get("_store_id") else legacy_by_project
                target.setdefault(pno, []).append(row)
            remove_ids = set()
            for pno, db_rows in db_by_project.items():
                legacy_rows = legacy_by_project.get(pno) or []
                if len(db_rows) == 1 and len(legacy_rows) == 1:
                    legacy_rows[0].update(db_rows[0])
                    remove_ids.add(id(db_rows[0]))
            if remove_ids:
                out[kind] = [row for row in out[kind] if id(row) not in remove_ids]
            for row in out[kind]:
                _mark_app_store_row(row)
                derive_status(row, kind)
                derive_effective_verification(row, kind)
    except Exception:
        pass
    return out


def real_works():
    """02 돌발AS·04 정기점검 + 27 원본일정 현황 (앱 '업무' 데이터)"""
    import openpyxl
    from verification_sync import derived_field_status_map
    from ecount_reconcile import load_config, resolve_master
    try:
        from ledger_db import work_resolutions
        objective_done = work_resolutions()
    except Exception:
        # DB가 잠깐 잠겨도 원장 자체 화면은 계속 열려야 한다.
        objective_done = {}
    try:
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
        wb = master_book(master)
    except Exception:
        out = _overlay_app_store_works({"as": [], "pm": []})
        out["as"] = sort_by_date(app_year_rows(out["as"], "as"), "as", "접수ID")
        out["pm"] = sort_by_date(app_year_rows(out["pm"], "pm"), "pm", "점검ID")
        return out
    # 03시트는 접수ID·프로젝트NO가 02 완료행을 순서대로 끌어오는 배열수식이라
    # 캐시가 비어도 같은 순서를 재현해 돌발AS 카드에 현장 검증 상태를 붙인다.
    field_status = derived_field_status_map(wb)
    out = {"as": [], "pm": []}
    spec = {
        # 뒤쪽 3개는 '확인 완료' 표시 — 관리자가 검증한 건인지 카드에서 바로 보이게 한다.
        "02_돌발AS접수": ("as", ["접수ID", "프로젝트NO", "캠프명", "접수일자", "담당기사", "진행상태",
                                "작업완료일", "유상·무상·보험", "신청내용", "긴급도", "방문예정일",
                                "관리자검증상태", "최종확인일", "사진등록", "동영상등록",
                                "완료보고서등록", "ERP등록", "재방문여부",
                                "최초접수외추가작업", "추가작업확인상태",
                                # 밴드 원문 바로가기 — 목록에서 근거를 바로 열어 볼 수 있게 한다
                                "밴드 바로가기",
                                "검증결과", "검증문제코드"]),
        "04_정기점검": ("pm", ["점검ID", "프로젝트NO", "캠프명", "점검예정일", "실제점검일", "점검상태",
                              "담당기사", "이상발견여부", "돌발AS전환여부",
                              "유상추가작업발생", "유상·무상·보험", "비용구분", "추가작업내용",
                              "최종확인일(유현민 체크)", "점검사진", "점검보고서",
                              "ERP판매전표", "거래명세서", "담당관리자",
                              "검증결과", "검증문제코드"]),
    }
    for sheet, (key, cols) in spec.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        for row in ws.iter_rows(min_row=5, values_only=True):
            # ID 열은 수식 — 새로 추가된 행은 엑셀을 열기 전까지 캐시값이 없다.
            # 프로젝트NO를 대체 키로 써서 백필 행도 앱에 표시한다.
            rid = row[idx[cols[0]]] if cols[0] in idx else None
            if not rid and "프로젝트NO" in idx and idx["프로젝트NO"] < len(row):
                rid = row[idx["프로젝트NO"]]
            if not rid:
                continue
            # 행 확장(expand_rows)으로 만들어 둔 **빈 예비행**은 ID 수식이 값을 내므로
            # rid만으로는 걸러지지 않는다. 날짜도 캠프명도 없으면 실제 업무가 아니다.
            _d = row[idx[cols[3]]] if cols[3] in idx and idx[cols[3]] < len(row) else None
            _c = row[idx["캠프명"]] if "캠프명" in idx and idx["캠프명"] < len(row) else None
            if not _d and not _c:
                continue
            rec = {}
            for c in cols:
                v = row[idx[c]] if c in idx and idx[c] < len(row) else None
                rec[c] = str(v)[:10] if hasattr(v, "year") else ("" if v is None else str(v))
            if key == "as":
                rec.update(field_status.get(str(rec.get("프로젝트NO") or "").upper(), {}))
            else:
                rec["검증자"] = rec.get("담당관리자") or ""
                rec["검증일"] = rec.get("최종확인일(유현민 체크)") or ""
            derive_status(rec, key)
            # 객관 완료 판정은 Excel 상태 수식보다 DB가 정본이다. 완료일이 원장에 아직
            # 일괄반영 전이어도 앱에서는 즉시 완료로 보이고, 빠진 서류 경고는 그대로 남는다.
            resolved = (objective_done.get((key, str(rec.get(cols[0]) or "")))
                        or objective_done.get((key, str(rec.get("프로젝트NO") or ""))))
            if resolved:
                status_col = "진행상태" if key == "as" else "점검상태"
                done_col = "작업완료일" if key == "as" else "실제점검일"
                rec[status_col] = resolved["status"]
                rec[done_col] = rec.get(done_col) or resolved["completed_on"]
                rec["객관완료근거"] = resolved["basis"]
                rec["객관완료최초확인"] = resolved["first_seen"]
            derive_effective_verification(rec, key)
            out[key].append(rec)
    # 류지영 원본 일정은 UJ번호가 없는 캠프·장비 일정이다. 04시트와 같은 캠프·같은 달이면
    # 이미 프로젝트 행으로 표시되므로 중복하지 않고, 아직 04에 없는 미래 월만 읽기 전용으로 보탠다.
    source_schedule = _sheet_records(wb, "27_정기점검원본일정")
    wb.close()
    _overlay_app_store_works(out)
    try:
        pm_report = json.load(open(os.path.join(ROOT, "reports", "pm_schedule_sync.json"),
                                   encoding="utf-8"))
        predicted = {str(r.get("일정ID") or ""): r for r in (pm_report.get("schedule") or [])}
    except Exception:
        predicted = {}
    try:
        out["as"] += erp_work_rows(out["as"], "as")
        out["pm"] += erp_work_rows(out["pm"], "pm")
        idx = build_prj_index(out)
        apply_rep_no(out["as"], idx, "접수ID")
        apply_rep_no(out["pm"], idx, "점검ID")
    except Exception:
        pass
    def pm_camp_key(v):
        return re.sub(r"[^0-9A-Za-z가-힣]", "", re.split(r"[（(]", str(v or ""))[0]).lower()

    represented = set()
    for r in out["pm"]:
        d = norm_date(r.get("점검예정일") or r.get("실제점검일"))
        if d and r.get("캠프명"):
            represented.add((pm_camp_key(r.get("캠프명")), d[:7]))
    for s in source_schedule:
        month = str(s.get("예정월") or "")[:7]
        key = (pm_camp_key(s.get("캠프명")), month)
        if not month.startswith(APP_YEAR + "-") or not key[0] or key in represented:
            continue
        projects = sorted(set(re.findall(r"\bUJ26\d{4,}\b", str(s.get("연결프로젝트NO") or ""),
                                         flags=re.I)))
        prediction = predicted.get(str(s.get("일정ID") or "")) or {}
        out["pm"].append({
            "점검ID": str(s.get("일정ID") or ""),
            "프로젝트NO": projects[0].upper() if len(projects) == 1 else "",
            "캠프명": str(s.get("캠프명") or ""),
            # 일자가 미확정이면 월까지만 보인다. 1일로 만들면 허위 지연 경고가 생긴다.
            "점검예정일": str(s.get("점검예정일") or month),
            "예측점검일": str(prediction.get("예측점검일") or ""),
            "예측근거": str(prediction.get("예측근거") or ""),
            "예측신뢰도": str(prediction.get("예측신뢰도") or ""),
            "실제점검일": "",
            "점검상태": "예정" if s.get("점검예정일") else "예정월",
            "담당기사": str(s.get("담당기사") or ""),
            "장비수": s.get("장비수") or 0,
            "장비내역": str(s.get("장비내역") or ""),
            "반영상태": str(s.get("반영상태") or ""),
            "원본행": str(s.get("원본행") or ""),
            "원본파일": str(s.get("원본파일") or ""),
            "출처": "정기점검 스케줄 원본",
        })
        # 같은 캠프·같은 달이라도 담당기사/점검일이 다른 원본 그룹은 모두 보여 준다.
        # represented 는 04·ERP의 기존 프로젝트 중복을 막는 용도이고 원본끼리는 합치지 않는다.
    out["as"] = sort_by_date(app_year_rows(out["as"], "as"), "as", "접수ID")
    out["pm"] = sort_by_date(app_year_rows(out["pm"], "pm"), "pm", "점검ID")
    return out


def derive_status(rec, kind):
    """상태 열은 **수식**이라 새로 넣은 행은 엑셀을 한 번 열기 전까지 캐시값이 없다(None).
    그대로 두면 완료된 점검 90여 건이 전부 '미점검'으로 보인다 → 원본 열로 직접 판정한다.
    (판정 규칙은 시트 수식과 동일: 완료일 있으면 완료, 예정일이 지났으면 미점검)"""
    today = date.today().isoformat()
    if kind == "pm":
        state = str(rec.get("점검상태") or "").strip()
        if str(rec.get("실제점검일") or "").strip():
            if state not in ("AS전환", "점검불가", "취소", "철회"):
                rec["점검상태"] = "완료"
            return
        if state:
            return
        if str(rec.get("돌발AS전환여부") or "").strip():
            rec["점검상태"] = "AS전환"
        elif not str(rec.get("점검예정일") or "").strip():
            rec["점검상태"] = ""
        else:
            rec["점검상태"] = "미점검" if str(rec["점검예정일"])[:10] < today else "예정"
    else:
        state = str(rec.get("진행상태") or "").strip()
        if str(rec.get("작업완료일") or "").strip():
            if state not in ("취소", "철회"):
                rec["진행상태"] = "작업완료"
            return
        if state:
            return
        rec["진행상태"] = "접수"


def derive_effective_verification(rec, kind):
    """ZIP 패치 뒤 Excel을 열기 전에도 확정 상태가 앱에서 즉시 정상으로 보이게 한다.

    수식 캐시는 Excel 재계산 전까지 이전 결과를 유지한다. 여기서는 모든 필수 원인
    값이 명확히 충족된 경우에만 ``정상``으로 승격하고, 하나라도 불명확하면 기존
    검증결과를 그대로 둔다. 따라서 확인되지 않은 건을 정상으로 오판하지 않는다.
    """
    def text(name):
        return str(rec.get(name) or "").strip()

    if kind == "as" and text("진행상태") == "작업완료":
        required = [
            bool(text("담당기사")),
            bool(text("방문예정일")),
            bool(text("작업완료일")),
            text("사진등록") == "등록",
            text("동영상등록") != "누락",
            text("완료보고서등록") == "등록",
            text("유상·무상·보험") not in ("", "미확정"),
            text("ERP등록") in ("완료", "등록완료"),
            bool(text("재방문여부")),
            text("관리자검증상태") in ("일치", "추가작업발생"),
        ]
        if text("최초접수외추가작업") == "있음":
            required.append(text("추가작업확인상태") == "반영완료")
        # 03 현장행이 연결된 건은 문서·ERP·검증자·검증일까지 전부 확인돼야 한다.
        if rec.get("현장작업행"):
            required.extend([
                text("현장관리자검증") in ("일치", "추가작업발생"),
                text("거래명세서반영") == "반영완료",
                text("ERP반영") == "반영완료",
                text("검증자") == "유현민",
                bool(text("검증일")),
            ])
        if all(required):
            rec["검증결과"] = "정상"
            rec["검증문제코드"] = ""
    elif kind == "pm" and text("점검상태") == "완료":
        required = [
            bool(text("실제점검일")),
            text("점검사진") == "등록",
            text("ERP판매전표") in ("완료", "등록완료"),
            text("거래명세서") == "발행완료",
            text("검증자") == "유현민",
            bool(text("검증일")),
        ]
        # 점검보고서는 기존 앱 열 목록에 없던 열이라 새 파일에서는 반드시 등록돼야 한다.
        if "점검보고서" in rec:
            required.append(text("점검보고서") == "등록")
        if all(required):
            rec["검증결과"] = "정상"
            rec["검증문제코드"] = ""


def demo_works():
    rnd = random.Random(7)
    camps = ["송파5MB(감일동)", "울산2캠프", "인천7MB(마곡동)", "대전1캠프", "구리1캠프"]
    techs = ["김준형", "권오철", "김필우", "차동호"]
    a = [{"접수ID": f"AS-2607-{i:03d}", "캠프명": rnd.choice(camps),
          "접수일자": (date(2026, 7, 1) + timedelta(days=i)).isoformat(), "담당기사": rnd.choice(techs),
          "진행상태": rnd.choice(["접수", "방문예정", "작업중", "작업완료"]),
          "작업완료일": "", "유상·무상·보험": rnd.choice(["유상", "무상"]),
          "신청내용": "도어 센서 교체 외", "긴급도": rnd.choice(["보통", "긴급"])} for i in range(1, 11)]
    p = [{"점검ID": f"PM-2607-{i:03d}", "캠프명": rnd.choice(camps),
          "점검예정일": (date(2026, 7, 5) + timedelta(days=i * 2)).isoformat(),
          "실제점검일": "" if i % 3 == 0 else (date(2026, 7, 5) + timedelta(days=i * 2)).isoformat(),
          "점검상태": "예정" if i % 3 == 0 else "완료", "담당기사": rnd.choice(techs),
          "이상발견여부": rnd.choice(["없음", "있음"]), "돌발AS전환여부": "미전환"} for i in range(1, 8)]
    return {"as": a, "pm": p}


_MT = {"at": 0.0, "v": 0}
_MT_TTL = 30.0
_MT_LOCK = threading.Lock()


def _master_mtime():
    """관리대장이 마지막으로 저장된 시각. **30초 단일 캐시**를 둔다.

    ★ `resolve_master` 는 Z: 폴더를 훑어 최신 vN 을 고른다(실측 1.24초). 그런데
      `_fresh()` 가 **모든 캐시 조회마다** 이걸 불렀다 — 화면 하나가 API 를 예닐곱 개
      부르므로 아무것도 안 바뀐 상태에서도 Z: 를 예닐곱 번 훑었다. 갱신이 오래
      걸리던 가장 큰 몫이 여기였다(사용자 지시 2026-08-08: "갱신 빨리빨리하게").
    ★ 30초는 안전하다. 원장 저장은 11:00·15:00 회차나 사람 손으로 일어나며,
      화면에는 원장 실제 mtime을 기준시각으로 표시한다. 매 요청이 Z:를 훑는 것보다
      최대 30초 뒤 다음 조회에서 바뀐 원장을 잡는 편이 빠르고 정직하다.
    """
    now = time.time()
    if now - _MT["at"] < _MT_TTL:
        return _MT["v"]
    # 첫 화면 API 7개가 동시에 들어와도 Z: 최신본 탐색은 한 스레드만 한다. 락을
    # 잡은 뒤 다시 확인하지 않으면 대기하던 여섯 요청이 차례로 같은 탐색을 반복한다.
    with _MT_LOCK:
        now = time.time()
        if now - _MT["at"] < _MT_TTL:
            return _MT["v"]
        try:
            from ecount_reconcile import load_config, resolve_master
            v = os.path.getmtime(resolve_master(load_config()["reconcile"]["master_xlsx"]))
        except Exception:
            v = 0
        _MT["at"], _MT["v"] = now, v
        return v


def _data_asof_iso():
    """이 숫자들이 **언제의 원본인가**. `datetime.now()` 를 쓰면 안 된다.

    응답은 캐시에서 나올 수 있고(TTL 300~600초) 그동안 원장은 그대로다. 그런데
    '데이터최종갱신일'에 지금 시각을 찍으면 화면과 캡처가 낡은 숫자를 **'방금'이라고
    말한다** — 이 프로젝트가 1순위로 막는 '조용한 사고'다(CLAUDE.md '수집 밀림').
    숫자의 나이는 그것을 만든 관리대장이 마지막으로 저장된 시각이다.
    잴 수 없으면 **빈 문자열** — 모르는 것을 아는 척하지 않는다.
    """
    mt = _master_mtime()
    return datetime.fromtimestamp(mt).isoformat(timespec="seconds") if mt else ""


_brief_cache = {"key": None, "value": None}
_brief_lock = threading.Lock()
_work_log_view_cache = {"key": None, "value": None}
_work_log_view_lock = threading.Lock()


def _brief_source_key(day):
    """Cheap invalidation key for every source used by daily_brief."""
    try:
        from source_dirs import WORK_LOG_DIR
        work_logs = glob.glob(os.path.join(WORK_LOG_DIR, "**", "*.xlsx"), recursive=True)
        work_log_mt = max((os.path.getmtime(p) for p in work_logs), default=0)
    except Exception:
        work_log_mt = 0
    try:
        event_mt = os.path.getmtime(os.path.join(ROOT, "reports", "manual_daily_events.json"))
    except Exception:
        event_mt = 0
    try:
        schedule_mt = os.path.getmtime(os.path.join(ROOT, "reports", "pm_schedule_sync.json"))
    except Exception:
        schedule_mt = 0
    return day, _master_mtime(), work_log_mt, event_mt, schedule_mt


def get_daily_brief(day=None):
    """Return the representative brief without re-reading the large workbook per request.

    The first brief read includes the master workbook and the field work log.  When that
    overlaps the other first-page API reads through Cloudflare, the browser can time out
    and the saved report loses Yoo Subi's daily activity.  Cache the canonical result for
    each exact source revision.  Do not take the global workbook read lock here: a slow
    work-log read must never block settlements/status/works and freeze the whole app.
    """
    day = day or (date.today() - timedelta(days=1)).isoformat()
    key = _brief_source_key(day)
    with _brief_lock:
        if _brief_cache["value"] is not None and _brief_cache["key"] == key:
            return _brief_cache["value"]
        import daily_brief as DB
        result = DB.brief(day, DB.load()[0])
        source_mtime = max((float(v or 0) for v in key[1:]), default=0)
        result["데이터업데이트일시"] = (
            datetime.fromtimestamp(source_mtime).isoformat(timespec="minutes")
            if source_mtime else datetime.now().isoformat(timespec="minutes")
        )
        _brief_cache.update({"key": key, "value": result})
        return result


def get_work_log_view(day="", category="", state="", query=""):
    """대표보고 일지 원본과 최신 관리대장을 같은 시점에 대조한 상세 목록.

    원본 또는 관리대장이 바뀌면 캐시 키가 즉시 바뀌므로 다른 화면·Excel 입력이
    정식 원장에 반영되는 즉시 이 목록도 자동으로 다시 계산된다.
    """
    if DEMO:
        now = datetime.now().isoformat(timespec="seconds")
        result = {
            "ok": True, "generated_at": now,
            "source": "정기점검, 돌발AS 일지 (합성데이터).xlsx",
            "records": [
                {"구분": "돌발AS", "종류": "as_done", "프로젝트NO": "UJ2600975",
                 "캠프명": "송파5MB(감일동)", "일자": "2026-07-28", "상태": "완료",
                 "담당자": "류지영", "요청내용": "리모컨 작동 불가",
                 "실제조치": "리모컨 택배 발송 완료", "대조결과": "완료일 일치"},
                {"구분": "돌발AS", "종류": "as_open", "프로젝트NO": "UJ2601191",
                 "캠프명": "구리3MB(배양리)", "일자": "2026-07-29", "상태": "미실시",
                 "담당자": "", "요청내용": "간헐적 작동 불가",
                 "미처리사유": "방문 일정 조율", "대조결과": "미실시 사유 반영"},
                {"구분": "정기점검", "종류": "pm", "프로젝트NO": "UJ2601141",
                 "캠프명": "김해2캠프", "일자": "2026-07-28", "상태": "실행",
                 "담당자": "권오철", "실제조치": "3개월 점검 완료",
                 "대조결과": "완료일 일치"},
                {"구분": "정기점검", "종류": "pm", "프로젝트NO": "UJ2601144",
                 "캠프명": "인천4MB(청천동)", "일자": "2026-07-30", "상태": "실행",
                 "담당자": "김준형", "실제조치": "윤활·도어락 점검 완료",
                 "대조결과": "원장 미매칭"},
            ],
        }
    else:
        from ecount_reconcile import load_config, resolve_master
        from work_log_sync import analyze, find_latest_source
        source = find_latest_source()
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
        key = (
            os.path.abspath(source), os.path.getmtime(source), os.path.getsize(source),
            os.path.abspath(master), os.path.getmtime(master), os.path.getsize(master),
        )
        with _work_log_view_lock:
            if _work_log_view_cache["key"] != key:
                _work_log_view_cache.update({"key": key, "value": analyze(master, source)})
            result = _work_log_view_cache["value"]

    all_records = [
        dict(row) for row in (result.get("records") or [])
        if str(row.get("일자") or "").startswith(APP_YEAR + "-")
    ]
    text_query = re.sub(r"\s+", "", str(query or "")).lower()

    def keep(row):
        if day and str(row.get("일자") or "") != day:
            return False
        if category and str(row.get("구분") or "") != category:
            return False
        if state and str(row.get("상태") or "") != state:
            return False
        if text_query:
            hay = re.sub(r"\s+", "", " ".join(str(row.get(k) or "") for k in (
                "프로젝트NO", "캠프명", "담당자", "요청내용", "실제조치",
                "미처리사유", "대조결과",
            ))).lower()
            if text_query not in hay:
                return False
        return True

    records = [row for row in all_records if keep(row)]
    records.sort(key=lambda row: (
        str(row.get("일자") or ""), str(row.get("구분") or ""),
        str(row.get("프로젝트NO") or ""),
    ), reverse=True)
    visible_fields = (
        "구분", "종류", "원본시트", "원본행", "프로젝트NO", "캠프명", "일자",
        "원본상태", "상태", "미처리사유", "사유분류", "담당자", "요청내용",
        "실제조치", "비고", "원장매칭", "원장상태", "대조결과",
    )
    records = [{key: row.get(key, "") for key in visible_fields} for row in records]
    issue_words = ("미매칭", "상이", "중복", "확인 필요", "보완 대기")
    view_summary = {
        "전체": len(records),
        "돌발AS": sum(row.get("구분") == "돌발AS" for row in records),
        "정기점검": sum(row.get("구분") == "정기점검" for row in records),
        "확인필요": sum(any(word in str(row.get("대조결과") or "")
                         for word in issue_words) for row in records),
    }
    source_path = str(result.get("source") or "")
    return {
        "ok": True,
        "source_name": os.path.basename(source_path),
        "source_updated_at": (
            datetime.fromtimestamp(os.path.getmtime(source_path)).isoformat(timespec="seconds")
            if not DEMO and os.path.isfile(source_path) else result.get("generated_at") or ""
        ),
        "verified_at": result.get("generated_at") or "",
        "summary": result.get("요약") or {},
        "view_summary": view_summary,
        "records": records,
    }


def _fresh(key):
    """원장이 바뀌면 전체 무효화하고, 그 외에는 항목별 TTL만 적용한다.

    예전에는 120초마다 모든 대형 엑셀 캐시를 한꺼번에 지워 앱이 주기적으로
    20~50초 멈췄다. 원장 변경은 mtime으로 즉시 잡고, 외부 리포트 의존 항목만
    짧게 갱신한다.
    """
    mt = _master_mtime()
    if _cache.get("mt") != mt:
        _cache.clear()
        _cache["mt"] = mt
    if key in ("works", "settle"):
        # WAL 커밋은 본 DB 파일 mtime을 바로 바꾸지 않을 수 있으므로 WAL도 본다.
        app_stamp = _app_db_stamp()
        stamp_key = key + "_app_db_stamp"
        if _cache.get(stamp_key) != app_stamp:
            for suffix in ("", "_ts", "_stale"):
                _cache.pop(key + suffix, None)
            _cache[stamp_key] = app_stamp
    if key == "works":
        try:
            from ledger_db import DB_PATH
            db_mt = os.path.getmtime(DB_PATH)
        except Exception:
            db_mt = 0
        if _cache.get("works_db_mt") != db_mt:
            _cache.pop("works", None)
            _cache.pop("works_ts", None)
            _cache["works_db_mt"] = db_mt
    ttl = {"status": 300, "exec": 300, "issues": 300, "erpdocs": 300,
           "works": 600, "settle": 600}.get(key, 600)
    if key in _cache and time.time() - _cache.get(key + "_ts", 0) > ttl:
        _cache.pop(key, None)
        _cache.pop(key + "_ts", None)
    return _cache.get(key)


def _store_cache(key, value):
    _cache[key] = value
    _cache[key + "_ts"] = time.time()
    return value


# ── TTL 이 끝났다고 사람을 기다리게 하지 않는다 (2026-08-08 지시: "갱신 빨리빨리하게") ──
#
# `status` 에만 있던 stale-while-revalidate 를 무거운 자료 전부로 넓힌다. 예전에는
# TTL(600초)이 끝나는 **그 순간에 들어온 요청**이 Z: 콜드 재계산을 통째로 뒤집어썼다
# (실측 get_works 첫 계산 111초). 화면은 그동안 '갱신 중'에 멈춰 있었다.
#
# ★ 원장이 바뀌면 `_fresh` 가 `_cache` 를 통째로 비우므로 `_stale` 도 함께 사라진다.
#   **일부러 그렇게 둔다** — 바뀐 뒤의 옛 숫자를 '지금 값'처럼 보여 주는 것이 이
#   프로젝트가 1순위로 막는 조용한 사고다. 그때 한 번은 정직하게 기다린다.
_REFRESHING = {}
_REFRESH_LOCK = threading.Lock()


def _compute_locked(key, build):
    """Z: 를 실제로 읽는 자리. 락은 여기서만 잡는다(캐시 조회는 락 없이)."""
    with _readlock:
        c = _fresh(key)
        if c is not None:
            return c
        v = build()
        _store_cache(key, v)
        _cache[key + "_stale"] = v
        return v


def _spawn_refresh(key, build):
    with _REFRESH_LOCK:
        if _REFRESHING.get(key):
            return
        _REFRESHING[key] = True

    def run():
        try:
            _compute_locked(key, build)
        except Exception:
            pass                     # 뒤에서 도는 갱신이 실패해도 화면은 옛 값으로 산다
        finally:
            with _REFRESH_LOCK:
                _REFRESHING[key] = False

    threading.Thread(target=run, name=f"refresh-{key}", daemon=True).start()


def refresh_now(key, build):
    """캐시를 **실제로** 다시 만든다(예열용). 만료돼 있으면 계산하고, 아니면 그대로."""
    return _compute_locked(key, build)


def cached_data(key, build):
    """캐시가 있으면 즉시 · 만료면 옛 값을 즉시 주고 뒤에서 한 번만 다시 만든다."""
    c = _fresh(key)
    if c is not None:
        return c
    stale = _cache.get(key + "_stale")
    if stale is not None:
        _spawn_refresh(key, build)
        return stale
    return _compute_locked(key, build)


def get_works():
    if DEMO:
        return demo_works()
    return cached_data("works", real_works)


def _fmtv(v):
    """01시트 값 표시용: 부동소수 오차 정리·천단위·날짜"""
    if v is None or v == "":
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = round(float(v), 2)
        n = int(n) if abs(n - round(n)) < 0.01 else n
        return f"{n:,}"
    return str(v).strip()


def read_exec_report(master):
    """01_대표보고 시트를 구조 그대로 읽는다(엑셀 수식이 곧 집계 로직 — 앱에서 재계산하지 않음)."""
    import openpyxl
    wb = master_book(master)
    if "01_대표보고" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["01_대표보고"]
    rows = [r for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True)]
    wb.close()
    out = {"meta": {}, "summary": [], "sections": []}
    cur = None
    for i, row in enumerate(rows, 1):
        g = lambda j: row[j] if j < len(row) else None
        a = _fmtv(g(0))
        if i == 4:                                     # 보고일·집계기준일·보고자
            for li, vi in ((0, 1), (3, 4), (6, 7)):
                lab, val = _fmtv(g(li)), _fmtv(g(vi))
                if lab:
                    out["meta"][lab] = val
            continue
        if a.startswith("■"):
            out["summary"].append(a)
            continue
        # 섹션 헤더는 "2.  당일 업무 실적"처럼 숫자 뒤에 한글이 온다.
        # "1. [돌발AS] …" 같은 TOP5 항목은 대괄호로 시작하므로 헤더가 아니다.
        if re.match(r"^\d+\.\s*[가-힣]", a):
            cur = {"title": re.sub(r"\s+", " ", a), "items": [], "lines": []}
            out["sections"].append(cur)
            continue
        if not cur or a.startswith("※"):
            continue
        # 블록 헤더 행: "[ 돌발 AS · 현장 ]  [ 정기점검 ]  [ 거래서류 · 청구 ]"
        # → 열 위치별 그룹을 만들어 이후 행의 항목을 각 그룹에 담는다(AS/점검 구분 유지)
        heads = {li: _fmtv(g(li)).strip("[] ") for li in (0, 3, 6) if _fmtv(g(li)).startswith("[")}
        if heads:
            cur["colgroups"] = {}
            for li, name in heads.items():
                grp = {"name": name, "items": []}
                cur.setdefault("groups", []).append(grp)
                cur["colgroups"][li] = grp
            continue
        for li, vi in ((0, 1), (3, 4), (6, 7)):        # 3열 그룹: 라벨|값
            lab, val = _fmtv(g(li)), _fmtv(g(vi))
            if not lab or lab.startswith("["):
                continue
            if val == "" and len(lab) > 20:            # 값 없는 긴 문장 = 서술형(TOP5 등)
                cur["lines"].append(lab)
            else:
                grp = (cur.get("colgroups") or {}).get(li)
                (grp["items"] if grp else cur["items"]).append([lab, val])
    for s in out["sections"]:
        s.pop("colgroups", None)                       # 내부 매핑은 응답에서 제외
    old = _OLD_APP_REF_RE
    if any(old.search(str(v or "")) for v in out["meta"].values()):
        return {}
    out["summary"] = [x for x in out["summary"] if not old.search(str(x))]
    for s in out["sections"]:
        s["items"] = [x for x in s.get("items", []) if not old.search(" ".join(map(str, x)))]
        s["lines"] = [x for x in s.get("lines", []) if not old.search(str(x))]
        for g in s.get("groups", []):
            g["items"] = [x for x in g.get("items", []) if not old.search(" ".join(map(str, x)))]
    return out


def _sheet_records(wb, sheet):
    """머리글 4행 기준으로 시트를 JSON 안전한 dict 목록으로 읽는다."""
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    heads = [(i, str(h).strip()) for i, h in enumerate(hdr) if h not in (None, "")]
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        rec = {}
        for i, name in heads:
            value = row[i] if i < len(row) else None
            if isinstance(value, (datetime, date)):
                value = value.strftime("%Y-%m-%d")
            rec[name] = "" if value is None else value
        if any(v not in ("", None) for v in rec.values()):
            out.append(rec)
    return out


def _metric_number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _report_date(value):
    """대표 예외보고 계산에 쓸 날짜. 모르는 값은 만들지 않고 None으로 둔다."""
    text = norm_date(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def _report_business_days(start, end):
    """start~end(양 끝 포함)의 평일 수. 휴일 마스터가 없으므로 '영업일 추정'으로 표시한다."""
    if not start or not end or start > end:
        return 0
    return sum(1 for n in range((end - start).days + 1)
               if (start + timedelta(days=n)).weekday() < 5)


def _report_complete(value, when=""):
    return bool(_report_date(when)) or str(value or "").strip() in {
        "작업완료", "완료", "정상", "종결", "취소", "철회", "AS전환"
    }


def _report_missing(value, good):
    text = str(value or "").strip()
    return not text or text in {"미등록", "미작성", "미발행", "누락", "미확인"}


def _report_project(rec):
    """화면의 대표번호. 내부 AS/PM/JS 번호를 프로젝트NO처럼 가장하지 않는다."""
    value = str(rec.get("프로젝트NO") or "").strip()
    if re.match(r"^UJ\d{6,}$", value, re.I) or re.match(r"^ERP(?:[-_\s]|$)", value, re.I):
        return value
    for item in rec.values():
        hit = _UJ_RE.search(str(item or ""))
        if hit:
            return hit.group()
    return ""


def representative_summary(works, settlements, base_date=""):
    """유수비 대표 통화 요구를 기존 원장 값으로 계산한 읽기 전용 보고 모델.

    새 업무상태를 추측해 원장에 쓰지 않는다. 현장완료 확정은 완료일/완료상태가 있는 경우만,
    서류 미정리는 그 완료행의 사진·완료보고서·ERP 값이 실제로 빠진 경우만 센다.
    """
    today = _report_date(base_date) or date.today()
    as_rows = list((works or {}).get("as") or [])
    pm_rows = list((works or {}).get("pm") or [])

    def as_item(r, *, issue, state):
        received = _report_date(r.get("접수일자"))
        age = max(0, (today - received).days) if received else -1
        return {
            "프로젝트NO": _report_project(r),
            "ID": str(r.get("접수ID") or ""),
            "레코드ID": str(r.get("접수ID") or ""),
            "종류": "as",
            "프로젝트명": "돌발AS",
            "캠프명": str(r.get("캠프명") or ""),
            "일자": norm_date(r.get("접수일자")),
            "담당자": str(r.get("담당기사") or ""),
            "문제": issue,
            "상태": state,
            "경과일": age,
            "접수내용": str(r.get("신청내용") or ""),
        }

    backlog, paperwork = [], []
    for r in as_rows:
        done = _report_complete(r.get("진행상태"), r.get("작업완료일"))
        received = _report_date(r.get("접수일자"))
        age = max(0, (today - received).days) if received else -1
        if not done:
            grade = ("장기" if age > 30 else "심각" if age > 7 else
                     "경고" if age > 2 else "관심" if age > 1 else "정상")
            backlog.append(as_item(
                r, issue=f"전산상 미완료 · {age if age >= 0 else '날짜 미상'}일 경과",
                state=grade))
        else:
            missing = []
            if _report_missing(r.get("사진등록"), {"등록"}):
                missing.append("사진")
            if _report_missing(r.get("완료보고서등록"), {"등록", "완료"}):
                missing.append("완료보고서")
            if _report_missing(r.get("ERP등록"), {"등록", "완료"}):
                missing.append("ERP")
            if missing:
                paperwork.append(as_item(
                    r, issue="현장완료 · " + "·".join(missing) + " 미정리",
                    state="전산·서류 미정리"))

    backlog.sort(key=lambda r: (r.get("일자") or "9999-99-99", r.get("ID") or ""))
    paperwork.sort(key=lambda r: (r.get("일자") or "9999-99-99", r.get("ID") or ""))
    d1 = [r for r in backlog if r.get("경과일", -1) > 1]
    d2 = [r for r in backlog if r.get("경과일", -1) > 2]
    d7 = [r for r in backlog if r.get("경과일", -1) > 7]
    d30 = [r for r in backlog if r.get("경과일", -1) > 30]

    qmonth = ((today.month - 1) // 3) * 3 + 1
    qstart = date(today.year, qmonth, 1)
    qend = (date(today.year + (1 if qmonth == 10 else 0),
                 1 if qmonth == 10 else qmonth + 3, 1) - timedelta(days=1))
    qrows = [r for r in pm_rows if
             (lambda d: bool(d and qstart <= d <= qend))(_report_date(r.get("점검예정일")))]
    qdone = [r for r in qrows if
             _report_complete(r.get("점검상태"), r.get("실제점검일"))]
    total_days = (qend - qstart).days + 1
    elapsed_days = min(total_days, max(0, (today - qstart).days + 1))
    target = len(qrows)
    # ★ 목표누계는 분기 균등 배분이 아니라 **점검예정일 기준**이다(2026-07-31 수정).
    #   균등 배분(대상×경과일/분기일수)은 "7월에 몰아서 도는" 실제 일정과 어긋나
    #   목표 17 대 실제 44, +27건 같은 무의미한 숫자를 만들었다. 계획은 사람이
    #   예정일로 이미 세워 뒀다 — 오늘까지 예정된 건수가 곧 목표 누계다.
    expected = sum(1 for r in qrows
                   if (lambda d: bool(d and d <= today))(_report_date(r.get("점검예정일"))))
    actual = len(qdone)
    gap = actual - expected
    shortage_ratio = ((expected - actual) / expected * 100) if expected and actual < expected else 0
    signal = "적색" if shortage_ratio > 10 else "황색" if shortage_ratio >= 5 else "녹색"
    remaining = max(0, target - actual)
    remaining_business = _report_business_days(max(today + timedelta(days=1), qstart), qend)
    required_daily = remaining / remaining_business if remaining_business else (float(remaining) if remaining else 0)
    required_weekly = required_daily * 5
    techs = set()
    for r in qrows:
        techs.update(x.strip() for x in re.split(r"[,·/]|\s{2,}", str(r.get("담당기사") or ""))
                     if x.strip())
    available = len(techs)
    per_tech = required_daily / available if available else 0
    elapsed_business = _report_business_days(qstart, min(today, qend))
    total_business = _report_business_days(qstart, qend)
    forecast = min(target, round(actual / elapsed_business * total_business)) if elapsed_business else 0
    forecast_shortfall = max(0, target - forecast)

    statement_groups = {
        "돌발 AS": [], "정기점검": [], "신규·납품·설치": [], "기타": []
    }

    def statement_kind(r):
        text = str(r.get("업무구분") or "")
        if "돌발" in text or text.upper() == "AS":
            return "돌발 AS"
        if "정기" in text or "점검" in text:
            return "정기점검"
        if any(x in text for x in ("신규", "납품", "설치")):
            return "신규·납품·설치"
        return "기타"

    for r in settlements or []:
        # 유상임이 확인된 정산만 발행대상으로 센다. 비용구분 미확정은 확인 필요로 남긴다.
        if "유상" not in str(r.get("비용구분") or ""):
            continue
        statement_groups[statement_kind(r)].append(r)

    statement_rows = []
    unissued_rows = []
    for kind, rows in statement_groups.items():
        # 거래명세서 번호를 별도로 적지 않고 발행일만 기록한 과거 원장이 많다.
        # 발행일이 있는데도 번호가 비었다는 이유만으로 미발행으로 세면 2026년
        # 완료분 대부분이 다시 경고로 살아난다. 번호·상태·발행일 중 하나라도
        # 확인되면 발행 완료 근거로 인정한다.
        issued = [
            r for r in rows
            if str(r.get("명세서번호") or "").strip()
            or str(r.get("명세서") or "").strip()
            in {"있음", "발행", "발행완료", "완료", "반영완료"}
            or norm_date(r.get("명세서발행일"))
        ]
        unissued = [r for r in rows if r not in issued]
        dates = sorted(d for d in (norm_date(r.get("완료일")) for r in rows) if d)
        item = {
            "업무유형": kind, "발행대상": len(rows), "발행완료": len(issued),
            "미발행": len(unissued), "발행률": pct(len(issued), len(rows)),
            "대상기간": f"{dates[0]} ~ {dates[-1]}" if dates else "대상 없음",
            "합계금액": sum(_metric_number(r.get("공급가액")) for r in rows),
        }
        statement_rows.append(item)
        for r in unissued:
            unissued_rows.append({
                "프로젝트NO": _report_project(r),
                "ID": str(r.get("정산ID") or ""),
                "레코드ID": str(r.get("정산ID") or ""),
                "종류": "settle",
                "프로젝트명": str(r.get("업무구분") or kind),
                "캠프명": str(r.get("캠프명") or ""),
                "일자": norm_date(r.get("완료일")),
                "담당자": str(r.get("담당자") or ""),
                "문제": "발행 대상이나 거래명세서 미발행",
                "상태": "미발행",
                "금액": _metric_number(r.get("공급가액")),
            })

    policy_names = [
        "돌발 AS·정기점검 거래명세서 묶음기간",
        "여러 거래명세서의 세금계산서 합산 기준",
        "세금계산서 건별·월합계 발행 기준",
        "PO 수신 전 세금계산서 발행 가능 여부",
        "다음 청구주기 이월 조건·승인자",
    ]
    saved_policy = load_policy_state()
    policies, confirmed_policies = [], []
    for name in policy_names:
        state = saved_policy.get(name) if isinstance(saved_policy.get(name), dict) else {}
        item = {
            "기준": name,
            "상태": str(state.get("상태") or "확인 필요"),
            "확정내용": str(state.get("확정내용") or ""),
            "저장일시": str(state.get("저장일시") or ""),
            "저장자": str(state.get("저장자") or ""),
        }
        (confirmed_policies if item["상태"] == "확정" and item["확정내용"] else policies).append(item)

    one_line = (
        f"전산상 미완료 돌발 AS {len(backlog)}건 중 D+2 초과 {len(d2)}건, "
        f"현장완료·서류미정리 {len(paperwork)}건입니다. "
        f"{today.month}월 기준 정기점검은 목표 누계 {expected}건 대비 {actual}건"
        f"({gap:+d}건), 거래명세서 미발행 대상은 {len(unissued_rows)}건입니다."
    )
    return {
        "meta": {
            "집계기준일": today.isoformat(), "적용마감시간": "관리대장 최신 저장 시점",
            "데이터최종갱신일": _data_asof_iso(),
            "원천업무건수": len(as_rows) + len(pm_rows) + len(settlements or []),
            "검증되지않은건수": len(backlog) + len(paperwork) + len(unissued_rows),
            "필터조건": f"{APP_YEAR}년·정상 상세 기본 접힘",
        },
        "한줄종합보고": one_line,
        "돌발AS": {
            "전산상미완료": len(backlog), "현장완료서류미정리": len(paperwork),
            "D+1초과": len(d1), "D+2초과": len(d2), "7일초과": len(d7),
            "30일초과": len(d30), "대표지속보고": len(d2),
            "미완료목록": backlog, "서류미정리목록": paperwork,
        },
        "정기점검": {
            "분기": f"{qstart.month}~{qend.month}월", "분기시작일": qstart.isoformat(),
            "분기종료일": qend.isoformat(), "전체대상": target,
            "경과율": pct(elapsed_days, total_days) or 0,
            "목표누계": expected, "실제완료": actual, "계획대비": gap,
            "잔여대상": remaining, "잔여평일추정": remaining_business,
            "필요일일처리량": round(required_daily, 2),
            "필요주간처리량": round(required_weekly, 2),
            "투입기사수": available, "기사1인당필요일일": round(per_tech, 2),
            "예상완료": forecast, "예상미달": forecast_shortfall, "신호": signal,
            "목록": [{
                "프로젝트NO": _report_project(r), "ID": str(r.get("점검ID") or ""),
                "레코드ID": str(r.get("점검ID") or ""), "종류": "pm",
                "프로젝트명": "정기점검", "캠프명": str(r.get("캠프명") or ""),
                "일자": norm_date(r.get("점검예정일")), "담당자": str(r.get("담당기사") or ""),
                "문제": "분기 점검 대상", "상태": str(r.get("점검상태") or ""),
            } for r in qrows],
        },
        "거래명세서": {"업무유형별": statement_rows, "미발행목록": unissued_rows},
        "업무기준확인필요": policies,
        "업무기준확정": confirmed_policies,
    }


# ── 철거·신규납품 숨김 (사용자 지시 2026-07-29) ──────────────────────
#   "철거 및 신규건은 DB만 보관하고 앱에 표시하지마 / 추후에 앱에 추가할 수도 있으니
#    감안해서 정리해줘" → 원장에서 지우지 않고 **화면에서만** 뺀다.
#   대표보고(보고 탭)의 숫자는 서버가 계산해서 내려주므로 앱쪽 필터가 닿지 않는다.
#   그래서 여기서도 같은 규칙을 한 번 더 적용한다. 켜려면 아래 한 줄만 True 로 바꾼다
#   (index.html 의 SHOW_SIDE_WORK 도 같이 켠다 — 검증 [76]이 둘의 짝을 지킨다).
SHOW_SIDE_WORK = False
SIDE_WORK_RE = re.compile(r"철거|이전|납품|설치|계단|안전바|경보장치|메자닌")


def is_side_work(r):
    if SHOW_SIDE_WORK or not isinstance(r, dict):
        return False
    return any(SIDE_WORK_RE.search(str(r.get(k) or ""))
               for k in ("업무구분", "업무유형", "구분", "종류", "품목"))


def drop_side_work(rows):
    return [r for r in (rows or []) if not is_side_work(r)]


def get_representative_report():
    works = get_works()
    works = {k: (drop_side_work(v) if isinstance(v, list) else v) for k, v in (works or {}).items()}
    return representative_summary(works, drop_side_work(get_settlements()))


def read_exec_details(master, base_date=""):
    """대표보고 3·4절의 숫자를 만든 **동일 원천 행**을 건별 목록으로 돌려준다.

    앱에서 숫자를 다시 추정하면 엑셀 카드와 목록 건수가 갈릴 수 있다. 따라서
    01_대표보고/00_대시보드 수식이 참조하는 열과 조건을 그대로 재현한다.
    """
    import openpyxl
    wb = master_book(master)
    s06 = _sheet_records(wb, "06_거래서류청구수금")
    s07 = _sheet_records(wb, "07_불일치누락현황")
    s15 = _sheet_records(wb, "15_세금계산서관리")
    s17 = _sheet_records(wb, "17_문서대조현황")
    s02 = _sheet_records(wb, "02_돌발AS접수")
    s04 = _sheet_records(wb, "04_정기점검")
    wb.close()

    # 프로젝트NO·업무ID·정산ID 어느 것을 눌러도 캠프와 대표 날짜를 찾을 수 있게 한다.
    lookup = {}

    def remember(rec, ids, camp, when, kind="", owner=""):
        info = {
            "프로젝트NO": str(rec.get("프로젝트NO") or ""),
            "캠프명": str(rec.get(camp) or ""),
            "일자": norm_date(rec.get(when)),
            "프로젝트명": str(rec.get(kind) or ""),
            "담당자": str(rec.get(owner) or ""),
        }
        for key in ids:
            value = str(rec.get(key) or "").strip()
            if value:
                lookup.setdefault(value, info)

    for r in s06:
        remember(r, ("정산ID", "원천업무ID", "프로젝트NO"), "캠프명",
                 "작업완료일(자동)", "업무구분", "담당자")
    for r in s02:
        remember(r, ("접수ID", "프로젝트NO"), "캠프명", "접수일자", "", "담당기사")
    for r in s04:
        remember(r, ("점검ID", "프로젝트NO"), "캠프명", "점검예정일", "", "담당기사")

    rep_idx = build_prj_index({"as": s02, "pm": s04})

    def joined(rec):
        for key in ("프로젝트NO", "정산ID", "원천업무ID", "업무ID", "접수ID", "점검ID"):
            value = str(rec.get(key) or "").strip()
            if value in lookup:
                return lookup[value]
        return {}

    def detail(rec, *, when="", amount=0, issue="", status="", source=""):
        base = joined(rec)
        project = str(rec.get("프로젝트NO") or base.get("프로젝트NO") or "")
        if not project:
            candidate = {**base, **rec}
            candidate.setdefault("완료일", rec.get(when) if when else base.get("일자"))
            project, _ = rep_no(
                candidate, rep_idx,
                str(rec.get("거래명세서번호") or rec.get("명세서번호") or ""))
        rid = str(rec.get("정산ID") or rec.get("원천업무ID") or rec.get("업무ID")
                  or rec.get("접수ID") or rec.get("점검ID") or project or "")
        record_kind = ("settle" if rid.startswith("JS-") else
                       "as" if rid.startswith("AS-") else
                       "pm" if rid.startswith("PM-") else "")
        return {
            "프로젝트NO": project,
            "ID": rid,
            "레코드ID": rid,
            "종류": record_kind,
            "프로젝트명": str(rec.get("업무구분") or base.get("프로젝트명") or source or ""),
            "캠프명": str(rec.get("캠프명") or base.get("캠프명") or ""),
            "일자": norm_date(rec.get(when)) if when else str(base.get("일자") or ""),
            "금액": _metric_number(amount),
            "문제": str(issue or ""),
            "상태": str(status or ""),
            "담당자": str(rec.get("담당자") or rec.get("담당기사") or base.get("담당자") or ""),
            "출처": source,
        }

    def is_2026_settlement(r):
        d = norm_date(r.get("작업완료일(자동)") or r.get("작업완료일"))
        return bool(r.get("정산ID")) and d.startswith(APP_YEAR + "-")

    def sorted_rows(rows):
        return sort_by_date(rows, "metric", "ID")

    details = {}

    def add(label, rows, basis, kind):
        # 임의 기준일 보고를 만들 때 미래 행이 과거 캡처에 섞이지 않게 한다.
        # 일자가 없는 현재 잔여·문서 경고는 원천상 시점을 판별할 수 없어 그대로 남긴다.
        if base_date:
            rows = [r for r in rows
                    if not norm_date(r.get("일자")) or norm_date(r.get("일자")) <= base_date]
        rows = sorted_rows(rows)
        details[label] = {
            "rows": rows,
            "basis": basis,
            "kind": kind,
            "count": len(rows),
            "amount": sum(_metric_number(r.get("금액")) for r in rows),
        }

    # 3. 당일 금액 · 잔여 현황 — 06시트의 대표보고 수식과 같은 조건.
    add("청구액 (당일)",
        [detail(r, when="거래명세서발행일", amount=r.get("거래명세서합계"),
                status=r.get("청구상태"), source="거래명세서")
         for r in s06 if is_2026_settlement(r) and norm_date(r.get("거래명세서발행일")) == base_date],
        f"06_거래서류청구수금 · 거래명세서발행일={base_date} · 거래명세서합계", "amount")
    add("세금계산서 발행액 (당일)",
        [detail(r, when="세금계산서발행일", amount=r.get("세금계산서합계"),
                status=r.get("청구상태"), source="세금계산서")
         for r in s06 if is_2026_settlement(r) and norm_date(r.get("세금계산서발행일")) == base_date],
        f"06_거래서류청구수금 · 세금계산서발행일={base_date} · 세금계산서합계", "amount")
    add("입금액 (당일)",
        [detail(r, when="입금일", amount=r.get("입금액"),
                status=r.get("청구상태"), source="입금")
         for r in s06 if is_2026_settlement(r) and norm_date(r.get("입금일")) == base_date],
        f"06_거래서류청구수금 · 입금일={base_date} · 입금액", "amount")
    add("잔여 미청구액",
        [detail(r, amount=r.get("미청구액"), issue=r.get("문제내용"),
                status=r.get("청구상태"), source="미청구")
         for r in s06 if is_2026_settlement(r) and _metric_number(r.get("미청구액")) > 0],
        "06_거래서류청구수금 · 2026년 정산ID 보유 · 미청구액>0", "amount")
    add("잔여 미수금액",
        [detail(r, amount=r.get("미수금액"), issue=r.get("문제내용"),
                status=r.get("청구상태"), source="미수")
         for r in s06 if is_2026_settlement(r) and _metric_number(r.get("미수금액")) > 0],
        "06_거래서류청구수금 · 2026년 정산ID 보유 · 미수금액>0", "amount")
    add("작업금액 불일치 (현재)",
        [detail(r, amount=r.get("작업대비거래명세서차액"), issue=r.get("문제내용"),
                status=r.get("검증결과"), source="금액 불일치")
         for r in s06 if is_2026_settlement(r)
         and _metric_number(r.get("작업대비거래명세서차액")) != 0
         and _metric_number(r.get("거래명세서합계")) != 0
         and str(r.get("업무구분") or "") != "신규·납품·설치"],
        "06_거래서류청구수금 · 작업/명세서 차액≠0 · 명세서합계≠0 · 신규납품 제외", "risk")

    # 4. 리스크 — 00_대시보드 수식의 실제 원천행.
    issue_2026 = [r for r in s07 if str(r.get("업무기준연도(자동·숨김)") or "") == APP_YEAR]
    unique_work = {}
    for r in issue_2026:
        key = str(r.get("최상위 업무키") or "").strip()
        if not key:
            continue
        if key not in unique_work:
            unique_work[key] = detail(
                {**r, "업무ID": r.get("원천업무ID") or key},
                issue=r.get("문제상세"), status=r.get("조치상태"), source="확인필요")
        elif r.get("문제상세"):
            old = unique_work[key]["문제"]
            new = str(r.get("문제상세"))
            if new not in old:
                unique_work[key]["문제"] = " · ".join(x for x in (old, new) if x)
    add("문제 업무 건수(중복 제거)", list(unique_work.values()),
        "07_불일치누락현황 · 업무기준연도=2026 · 최상위 업무키 중복 제거", "risk")

    add("문서 경고 총계",
        [detail(r, issue=r.get("경고내용"), status=r.get("우선순위"),
                source="문서 경고")
         for r in s17 if str(r.get("정산ID") or "").startswith("JS-26")
         and str(r.get("경고내용") or "").strip()],
        "17_문서대조현황 · 정산ID=JS-26* · 경고내용 있음", "risk")

    tax_rows = []
    for r in s15:
        if not str(r.get("정산ID") or "").startswith("JS-26"):
            continue
        if str(r.get("발행기한임박여부") or "") == "예":
            tax_rows.append(detail(r, when="법정발행기한", amount=r.get("발행금액"),
                                   issue="세금계산서 발행기한 임박", status=r.get("발행상태(자동)"),
                                   source="세금계산서 기한"))
        if str(r.get("기한초과여부") or "") == "예":
            tax_rows.append(detail(r, when="법정발행기한", amount=r.get("발행금액"),
                                   issue="세금계산서 발행기한 초과", status=r.get("발행상태(자동)"),
                                   source="세금계산서 기한"))
    add("세금계산서 기한 임박·초과", tax_rows,
        "15_세금계산서관리 · JS-26* · 발행기한임박=예 또는 기한초과=예", "risk")

    add("PO 미발행 · 확인필요",
        [detail(r, issue=r.get("경고내용") or r.get("PO상태"),
                status=r.get("PO상태"), source="PO")
         for r in s17 if str(r.get("정산ID") or "").startswith("JS-26")
         and str(r.get("PO상태") or "") in ("PO 발행대기", "PO관리행 없음")],
        "17_문서대조현황 · JS-26* · PO상태=PO 발행대기/PO관리행 없음", "risk")
    add("거래명세서 미작성",
        [detail(r, issue=r.get("경고내용") or "거래명세서 미작성",
                status=r.get("거래명세서상태"), source="거래명세서")
         for r in s17 if str(r.get("정산ID") or "").startswith("JS-26")
         and str(r.get("거래명세서상태") or "") == "미작성"],
        "17_문서대조현황 · JS-26* · 거래명세서상태=미작성", "risk")
    add("아리바 청구 미등록",
        [detail(r, when="법정발행기한", amount=r.get("발행금액"),
                issue="아리바 청구 등록대기", status=r.get("아리바청구상태"),
                source="아리바")
         for r in s15 if str(r.get("정산ID") or "").startswith("JS-26")
         and str(r.get("아리바청구상태") or "") == "등록대기"],
        "15_세금계산서관리 · JS-26* · 아리바청구상태=등록대기", "risk")

    problem_rows = [
        detail({**r, "업무ID": r.get("원천업무ID")},
               amount=r.get("미청구액") or r.get("미수금액"),
               issue=r.get("문제상세"), status=r.get("조치상태"), source="문제 행")
        for r in issue_2026 if str(r.get("원천업무ID") or "").strip()
    ]
    add("문제 프로젝트 / 문제 행", problem_rows,
        "07_불일치누락현황 · 업무기준연도=2026 · 원천업무ID 보유 행", "risk")
    details["문제 프로젝트 / 문제 행"]["project_count"] = len({
        str(r.get("프로젝트NO") or "").strip() for r in issue_2026
        if str(r.get("프로젝트NO") or "").strip()
    })
    return details


def _augment_exec_daily(report):
    """유수비 대표 지시(2026-08-04) 항목을 '당일 업무 실적' 카테고리 카드에 직접 얹는다.

    대표는 캡처 첫머리의 카드 3장(돌발 AS·현장/정기점검/거래서류·청구)을 본다 —
    브리핑 블록·아래쪽 별도 섹션에만 넣었더니 "반영 하나도 안 됐다"가 됐다(실측).
    값은 엑셀에 쓰지 않고(DB-only) 서버가 보고 시점에 daily_brief·정산 집계에서
    계산해 얹는다. 항목 3원소째는 표시 색 — 화면(execHtml)·캡처(groupCols) 공통.
    """
    if not isinstance(report, dict):
        return report
    sec = next((s for s in report.get("sections", [])
                if "당일 업무 실적" in str(s.get("title", ""))), None)
    if not sec or not sec.get("groups"):
        return report

    def grp(word):
        return next((g for g in sec["groups"] if word in str(g.get("name", ""))), None)

    def metric_key(value):
        # 원장 라벨의 '(금일)'·'(완료 기준)'·날짜가 달라도 같은 지표로 교체한다.
        text = re.sub(r"\([^)]*\)", "", str(value or ""))
        return re.sub(r"[^0-9A-Za-z가-힣]", "", text)

    def put(g, label, value, color=None):
        if not g:
            return
        items = g.setdefault("items", [])
        row = [label, value] + ([color] if color else [])
        for i, it in enumerate(items):
            if metric_key(it[0]) == metric_key(label):
                items[i] = row
                return
        items.append(row)

    base_date = norm_date((report.get("meta") or {}).get("집계기준일") or "")
    try:
        brief = get_daily_brief(base_date or None)
    except Exception:
        brief = {}
    as_day = (brief or {}).get("돌발AS") or {}
    pm = (brief or {}).get("정기점검") or {}
    wl = ((brief or {}).get("일지대조") or {}).get("돌발AS") or {}

    # ① 캡처의 원장 수식 캐시는 하루 이상 늦을 수 있으므로 카드 3장을 같은 기준일의
    # daily_brief·객관 정산 데이터로 전부 교체한다. 2026-08-04의 원장 카드 8/3은
    # 실제 원천 대조(중복 제거) 3/0과 달랐다.
    put(grp("돌발"), "신규 접수", f"{int(as_day.get('신규접수') or 0)}건")
    put(grp("돌발"), "신규 중 처리 완료", f"{int(as_day.get('신규처리완료') or 0)}건")
    put(grp("돌발"), "신규 처리율", "%.1f%%" % float(as_day.get('신규처리율') or 0))
    put(grp("돌발"), "작업 완료", f"{int(as_day.get('완료') or 0)}건")
    put(grp("돌발"), "현장 작업", f"{int(as_day.get('현장작업') or 0)}건")
    put(grp("돌발"), "유상 발생", f"{int(as_day.get('유상발생') or 0)}건")
    put(grp("돌발"), "재방문 예정", f"{int(as_day.get('재방문예정') or 0)}건")

    # 건바이건(장비 한 대=1건) — 그룹으로 묶지 않는다(2026-08-12 지시).
    # 원본은 호기 한 줄씩이라 장비 대수가 곧 건수다.
    put(grp("정기점검"), "점검 완료", f"{int(pm.get('완료장비') or 0)}건")
    put(grp("정기점검"), "점검 예정", f"{int(pm.get('예정장비') or 0)}건")
    put(grp("정기점검"), "이상 발견", f"{len((brief or {}).get('이상발견') or [])}건")
    put(grp("정기점검"), "돌발 AS 전환", f"{len((brief or {}).get('AS전환') or [])}건")
    put(grp("정기점검"), "유상 점검", f"{len((brief or {}).get('점검중유상') or [])}건")

    # ② 대표: "이번 분기에 일수를 따졌을 때 몇 %". 분모는 04에 들어온 58행이 아니라
    # 류지영 원본 전체 125그룹·302대다. 계획/완료 대수와 기준일까지 이행을 함께 보여 준다.
    if pm.get("분기예정"):
        pm_group = grp("정기점검")
        if pm_group:
            pm_group["items"] = [it for it in pm_group.get("items", [])
                                 if metric_key(it[0]) != metric_key("분기 진행률")]
        # 비율은 소수점 1자리로 그대로 보여 준다 — int()로 자르면 99.9%가 99%로,
        # 반올림하면 100%로 보여 "다 됐다"는 오해를 만든다(2026-08-05 지시).
        prog = float(pm.get("분기진행률") or 0)
        el = float(pm.get("분기경과율") or 0)
        gap = round(prog - el, 1)
        put(pm_group, "분기 예정·완료",
            f"{int(pm.get('분기완료') or 0)} / {int(pm.get('분기예정') or 0)}건")
        put(pm_group, "분기 진행률", "%.1f%%" % prog)
        due = int(pm.get("기준일까지예정") or 0)
        due_done = int(pm.get("기준일까지완료") or 0)
        if due:
            put(pm_group, "기준일까지 이행",
                f"{due_done} / {due}건 · " + pct_text(due_done, due, "대상 없음"))
        if pm.get("예측일정"):
            put(pm_group, "예측 점검(과거 이력)", f"{int(pm.get('예측장비') or 0)}건")
        put(pm_group, "기간 경과 %.1f%% 대비" % el,
            f"{gap:+.1f}%p {'앞섬' if gap >= 0 else '뒤짐'}",
            "#12813F" if gap >= 0 else "#B42318")

    # ③ 대표: "기사가 스케줄을 못 잡아 미루는 건 절대 안 된다" — 핫이슈는 0이어도
    #   자리를 지킨다. 숫자가 보여야 '없다'를 보고할 수 있다.
    if wl:
        hot = int(wl.get("핫이슈") or 0)
        put(grp("돌발"), "미실시 핫이슈(일정·사유)", f"{hot}건",
            "#B42318" if hot else "#12813F")

    # ④ 거래서류·청구는 ERP 판매·PO·거래명세서·세금계산서·밴드 객관완료가 합쳐진
    # get_settlements 한 원천으로 다시 센다. 날짜가 없는 상태값으로 임의 날짜를 만들지 않는다.
    try:
        st = drop_side_work(get_settlements())
        on = lambda row, field: norm_date(row.get(field)) == base_date
        statement_day = [r for r in st if on(r, "명세서발행일")]
        tax_day = [r for r in st if on(r, "계산서발행일")]
        erp_day = [r for r in st if on(r, "완료일") and (
            str(r.get("ERP진행상태") or "").strip() not in ("", "미등록", "확인필요")
            or str(r.get("금액출처") or "") == "ERP")]
        billing_day = [r for r in st if on(r, "청구일")]
        receipt_day = [r for r in st if on(r, "입금일")]
        put(grp("거래서류"), "거래명세서 발행", f"{len(statement_day)}건")
        put(grp("거래서류"), "세금계산서 발행", f"{len(tax_day)}건")
        put(grp("거래서류"), "ERP 등록(작업완료 기준)", f"{len(erp_day)}건")
        put(grp("거래서류"), "청구 진행", f"{len(billing_day)}건")
        put(grp("거래서류"), "입금 건수", f"{len(receipt_day)}건")

        # 모수 = 기준일까지 작업완료가 객관 확인된 유상 정산(사용자 확정 2026-08-04).
        target_rows = [r for r in st if "유상" in str(r.get("비용구분") or "")
                       and norm_date(r.get("완료일"))
                       and norm_date(r.get("완료일")) <= base_date]
        issued = [r for r in target_rows if str(r.get("명세서번호") or "").strip()
                  or str(r.get("명세서") or "").strip()
                  in {"있음", "발행", "발행완료", "완료", "반영완료"}
                  or norm_date(r.get("명세서발행일"))]
        target, wait = len(target_rows), len(target_rows) - len(issued)
        if target:
            put(grp("거래서류"), "명세서 발행 대기(완료 기준)",
                f"{wait}건 / {target}건", "#B42318" if wait else "#12813F")
    except Exception:
        pass
    return report


def get_exec_report(day=None, _force=False):
    """`_force` 는 **뒤에서 도는 갱신·예열**이 쓴다.

    ★ 이것이 없으면 갱신이 아무 일도 안 한다: 갱신 스레드가 그냥 이 함수를 부르면
      아래 stale 가지를 타고 **옛 값을 그대로 돌려주고**, 부른 쪽은 그걸 새 값으로
      저장한다 — 화면은 영원히 낡은 채로 '갱신했다'고 말한다. warm_caches 가
      2026-08-08 에 같은 함정을 밟았다(`refresh_now` 로 고쳤다).
    """
    requested = norm_date(day)
    # 사람이 날짜를 고른 것인지, 우리가 어제로 기본값을 채운 것인지 갈라 둔다.
    # ★ 이 구분이 없어서 **보고일까지 어제로 찍혔다**(2026-08-06 지시로 발견).
    #   아래 기본값 로직은 '집계기준일 = 어제'를 노린 것인데, 뒤에서 requested 를
    #   보고일에도 그대로 넣는 바람에 10:30 대표 보고 머리에 어제 날짜가 박혔다.
    picked_by_user = bool(requested)
    # ★ 캐시 확인은 Z:/Excel 접근보다 **반드시 먼저**다. 예전 코드는 기본 기준일을
    # 알아내려고 여기서 원장을 먼저 열고, 그 뒤에야 캐시를 봤다. 그래서 캐시가 살아
    # 있어도 /api/exec_report 요청마다 2~3초가 걸리고, SWR 중복 호출 때 같은 초에
    # 6~16번 원장을 다시 읽었다. 기본 조회는 마지막 정상값을 즉시 돌려준다.
    if not picked_by_user and not _force and not DEMO:
        cached = _fresh("exec")
        if cached:
            return cached
        stale = _cache.get("exec_stale")
        if stale is not None:
            _spawn_refresh("exec", lambda: get_exec_report(None, _force=True))
            return stale
    if not requested and not DEMO:
        # ★ 기본 집계기준일 = **어제**(사용자 지시 2026-08-04 "8월 3일 실적이 하나도
        #   업데이트 안 되었다"). 01_대표보고 시트의 집계기준일은 수동값이라 하루 이상
        #   뒤처지곤 한다(오늘 08-04 인데 시트는 08-02 로 굳어 있었다 — 08-02 는
        #   일요일이라 화면이 전부 0 이었다). 시트가 더 최신이면 시트를 따른다.
        try:
            from datetime import date, timedelta
            y = date.today() - timedelta(days=1)
            if y.year == int(APP_YEAR):
                requested = y.isoformat()
        except Exception:
            pass
    if DEMO:
        labels = [
            "청구액 (당일)", "세금계산서 발행액 (당일)", "입금액 (당일)",
            "잔여 미청구액", "잔여 미수금액", "작업금액 불일치 (현재)",
            "문제 업무 건수(중복 제거)", "문서 경고 총계", "세금계산서 기한 임박·초과",
            "PO 미발행 · 확인필요", "거래명세서 미작성", "아리바 청구 미등록",
            "문제 프로젝트 / 문제 행",
        ]
        details = {label: {"rows": [], "basis": "합성 데모 · 2026년 원천 행", "kind": "risk",
                           "count": 0, "amount": 0} for label in labels}
        money = {"프로젝트NO": "UJ261001", "ID": "JS-2607-001", "레코드ID": "JS-2607-001",
                 "종류": "settle", "프로젝트명": "돌발AS", "캠프명": "울산2캠프",
                 "일자": "2026-07-24", "금액": 22000, "문제": "미청구 합성 예시",
                 "상태": "미청구", "담당자": "김준형", "출처": "미청구"}
        details["잔여 미청구액"].update(rows=[money], count=1, amount=22000, kind="amount")
        risk_rows = []
        for i in range(75):
            n = i % 15 + 1
            risk_rows.append({
                "프로젝트NO": f"UJ26{1000+n}", "ID": f"JS-2607-{n:03d}",
                "레코드ID": f"JS-2607-{n:03d}", "종류": "settle",
                "프로젝트명": "돌발AS" if n % 2 else "정기점검",
                "캠프명": f"합성 캠프 {n}", "일자": f"2026-07-{n:02d}", "금액": 0,
                "문제": f"문서 확인 필요 합성 항목 {i+1}", "상태": "P1",
                "담당자": "김준형", "출처": "문서 경고",
            })
        details["문서 경고 총계"].update(rows=risk_rows, count=len(risk_rows))
        return {
            "meta": {"보고일": requested or "2026-07-25",
                     "집계기준일": requested or "2026-07-24", "보고자": "유현민"},
            "summary": ["■ 데모 요약"],
            "sections": [
                {"title": "1. 당일 업무 실적",
                 "items": [["신규 접수", "3"], ["작업 완료", "1"]], "lines": []},
                {"title": "2. 당일 금액 · 잔여 현황",
                 "items": [["청구액 (당일)", "0"], ["세금계산서 발행액 (당일)", "0"],
                           ["입금액 (당일)", "0"], ["잔여 미청구액", "22,000"],
                           ["잔여 미수금액", "0"], ["작업금액 불일치 (현재)", "0"]]},
                {"title": "3. 리스크 (현재 기준)",
                 "items": [["문제 업무 건수(중복 제거)", "0"], ["문서 경고 총계", "75"],
                           ["세금계산서 기한 임박·초과", "0"], ["PO 미발행 · 확인필요", "0"],
                           ["거래명세서 미작성", "0"], ["아리바 청구 미등록", "0"],
                           ["문제 프로젝트 / 문제 행", "0개 / 0건"]]},
            ],
            "details": details,
        }
    # ★ **캐시를 보는 데 락은 필요 없다** (2026-08-08). `/api/status` 가 2026-08-07 에
    #   같은 이유로 고쳐졌는데 여기만 그대로 남아 있었다. `_readlock` 은 다른 요청이
    #   Z: 를 콜드로 읽는 동안 계속 잡혀 있고, 그동안 여기서는 **이미 만들어 둔 값을
    #   돌려주는 것조차** 막혔다. 실측(최근 24시간 사용 기록): 대표보고가 느린 화면
    #   648회로 1등 — 평균 110초. 계산이 무거운 게 아니라 줄을 서 있었던 것이다.
    #   락이 지키는 것은 'Z: 를 동시에 읽지 않는 것'이지 캐시 딕셔너리가 아니다.
    with _readlock:
        r = _fresh("exec") if not picked_by_user and not _force else None
        if r:
            return r
        from ecount_reconcile import load_config, resolve_master
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
        r = read_exec_report(master)
        sheet_base = norm_date((r.get("meta") or {}).get("집계기준일"))
        if not picked_by_user and sheet_base and (not requested or sheet_base > requested):
            requested = sheet_base
        # ★ 보고일은 **오늘**이다 (2026-08-06 지시: "보고일 오늘 날짜로 자동 변경").
        #   그동안 이 값은 원장 B3 에서 왔고, B3 는 report_dates 가 큐에 넣어 **11:00
        #   회차**에 들어갔다. 대표 보고는 10:30 이다 — 매일 아침 보고 시각까지 B3 는
        #   어제 날짜였다. 어제 날짜로 오늘 보고가 나가는 구조였다.
        #   그래서 원장을 기다리지 않고 여기서 오늘로 맞춘다. 원장 값이 **미래**이거나
        #   사람이 date= 로 고른 경우는 그대로 존중한다(과거 보고를 다시 볼 때).
        if not picked_by_user:
            _today = datetime.now().strftime("%Y-%m-%d")
            _stored = norm_date((r.get("meta") or {}).get("보고일")) or ""
            if _stored < _today:
                _yday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
                r.setdefault("meta", {})["보고일"] = _today
                r["meta"]["집계기준일"] = _yday
                r["meta"]["보고일자동"] = True     # 화면이 '오늘로 맞춤'을 알 수 있게
        base = requested or norm_date((r.get("meta") or {}).get("집계기준일")
                                      or (r.get("meta") or {}).get("보고일"))
        r["details"] = read_exec_details(master, base)
        if picked_by_user:
            # 보고일과 집계기준일은 **다른 값**이다. 보고는 다음 날 아침에 하므로
            # 기본은 '보고일=오늘, 집계기준일=어제' 다. 사람이 과거 날짜를 직접 고른
            # 때만 보고일도 그 날짜로 맞춘다(그때는 그 날짜의 보고서를 다시 보는 것이다).
            r.setdefault("meta", {})["집계기준일"] = requested
            if picked_by_user:
                r["meta"]["보고일"] = requested
            else:
                r["meta"]["보고일"] = datetime.now().strftime("%Y-%m-%d")
                r["meta"]["보고일자동"] = True
            # 금액·리스크 타일은 선택일로 다시 계산한 상세 집계와 맞춘다.
            for sec in r.get("sections", []):
                for item in sec.get("items", []):
                    d = r["details"].get(str(item[0]))
                    if not d:
                        continue
                    if str(item[0]) == "문제 프로젝트 / 문제 행":
                        item[1] = f"{d.get('project_count', 0)}개 / {d.get('count', 0)}건"
                    elif d.get("kind") == "amount":
                        item[1] = f"{int(d.get('amount') or 0):,}"
                    else:
                        item[1] = f"{int(d.get('count') or 0):,}"
            _augment_exec_daily(r)
            _append_remote_section(r)
            _append_kakao_warning(r)
            return r
        if requested:
            r.setdefault("meta", {})["집계기준일"] = requested
        _augment_exec_daily(r)
        _append_remote_section(r)
        _append_kakao_warning(r)
        # 옛 값 자리도 같이 채운다 — 이게 없으면 stale 가지가 영영 안 열려
        # TTL 이 끝나는 순간마다 누군가 한 명은 콜드 재계산을 통째로 맞는다.
        _cache["exec_stale"] = r
        return _store_cache("exec", r)


_SRC_IDX = {"at": 0, "doc": None, "kinds": [], "months": []}


def _source_index():
    """원본 색인(8천 건+)을 5분 캐시로 읽는다.

    ★ 2026-08-05 실측: 캐시 없이 요청마다 json.load 하니 API 가 45초를 넘겼다
      (색인이 1,806 → 8,248건으로 커진 뒤). 종류·월 목록도 그때 함께 계산해 둔다.
    """
    now = time.time()
    if now - _SRC_IDX["at"] < 300 and _SRC_IDX["doc"] is not None:
        return _SRC_IDX["doc"]
    try:
        with open(os.path.join(ROOT, "reports", "원본색인.json"), encoding="utf-8") as fh:
            doc = json.load(fh)
    except Exception:
        doc = None
    if doc:
        rows = doc.get("rows") or []
        _SRC_IDX["kinds"] = sorted({r.get("kind", "") for r in rows})
        _SRC_IDX["months"] = sorted({((r.get("slip") or "")[:7] or (r.get("date") or "")[:7])
                                     for r in rows if (r.get("slip") or r.get("date"))},
                                    reverse=True)[:36]
    _SRC_IDX["doc"] = doc
    _SRC_IDX["at"] = now
    return doc


# ── 한 건의 원본을 한자리에 (2026-08-07 지시) ────────────────────────────────
#   사용자 지시: "버튼을 누르면 해당 원본이 바로 열리게 … 일일이 원본 데이터를
#   찾아다닐 필요 없이 하는 게 목적임".
#   색인(원본색인.json)은 파일마다 uj·po·slip·post 를 이미 들고 있다. 그래서
#   따로 모을 것이 없다 — **한 건의 번호들로 색인을 훑어 갈래별로 묶어** 준다.
_BAND_URL = {"at": 0.0, "map": {}}


def _band_urls():
    """프로젝트NO → 밴드 글 주소. 근거(원본 덤프)가 있을 때만 만든다.

    ★ 주소를 지어내지 않는다 — 틀린 링크는 빈칸보다 나쁘다(fill_links_status 와 같은 규칙).
      덤프에는 글 번호가 키로, 밴드 번호가 `band` 로 들어 있다. 그 둘이 다 있을
      때만 주소가 된다. 한 글이 여러 건을 묶은 목록형이면 **그 건만 다루는 글**을
      우선한다(그래야 눌렀을 때 찾던 글이 바로 보인다).
    """
    now = time.time()
    if now - _BAND_URL["at"] < 600 and _BAND_URL["map"]:
        return _BAND_URL["map"]
    import glob as _glob
    prj = re.compile(r"UJ\d{7}")
    cand = {}
    for f in sorted(_glob.glob(os.path.join(ROOT, "band", "cache", "raw*.json"))):
        try:
            with open(f, encoding="utf-8") as fh:
                d = json.load(fh)
        except Exception:
            continue
        band, posts = str(d.get("band") or "").strip(), d.get("posts")
        if not band or not isinstance(posts, dict):
            continue
        for no, p in posts.items():
            if not str(no).isdigit():
                continue
            c = str((p.get("content") if isinstance(p, dict) else p) or "")
            ks = set(prj.findall(c))
            for k in ks:
                score = (1 if len(ks) == 1 else 0, 1 if "완료" in c[:80] else 0, int(no))
                if k not in cand or score > cand[k][0]:
                    cand[k] = (score, band, str(no))
    _BAND_URL["map"] = {k: "https://www.band.us/band/%s/post/%s" % (b, n)
                        for k, (_sc, b, n) in cand.items()}
    _BAND_URL["at"] = now
    return _BAND_URL["map"]


# 갈래 이름은 **사람이 부르는 이름**이다(사용자가 말한 그대로: 거래명세표·세금계산서·
# 매출전표·PO). 색인의 kind 는 도구가 붙인 이름이라 여기서 한 번 옮겨 준다.
_ORIG_GROUPS = (
    ("밴드 글", lambda k: k.startswith("밴드") and "첨부" not in k),
    ("밴드 첨부", lambda k: k.startswith("밴드") and "첨부" in k),
    ("거래명세서", lambda k: "거래명세" in k),
    ("세금계산서", lambda k: "계산서" in k),
    ("쿠팡 PO", lambda k: "PO" in k),
    ("매출전표·ERP", lambda k: k.startswith("ERP")),
    ("사진", lambda k: k in ("사진", "이미지")),
)


def _originals_for(uj="", po="", slip="", limit=8):
    """한 건에 딸린 원본을 갈래별로 묶어 돌려준다."""
    doc = _source_index() or {}
    rows = doc.get("rows") or []
    uj, po, slip = (str(x or "").strip() for x in (uj, po, slip))
    # PO 칸에 'PO329774/PR463518' 처럼 둘이 붙어 오는 일이 있다 — 토막마다 맞춰 본다
    pos = [x for x in re.split(r"[\s,/·]+", po) if len(x) >= 5]
    hits = []
    for r in rows:
        ok = (uj and r.get("uj") == uj) or (slip and r.get("slip") == slip) \
            or (pos and r.get("po") in pos)
        if ok:
            hits.append(r)
    try:
        from source_index import is_private
        hits = [r for r in hits if not is_private(r.get("path") or "", r.get("name") or "")]
    except Exception:
        hits = [r for r in hits if not str(r.get("name") or "").startswith("통화_")]
    out = []
    used = set()
    for label, want in _ORIG_GROUPS:
        got = []
        for r in hits:
            p = r.get("path") or ""
            if p in used or not want(str(r.get("kind") or "")):
                continue
            used.add(p)
            got.append({"name": r.get("name"), "path": p, "ext": r.get("ext"),
                        "date": r.get("date"), "slip": r.get("slip")})
        if got:
            got.sort(key=lambda x: (x.get("date") or "", x.get("name") or ""), reverse=True)
            out.append({"label": label, "n": len(got), "files": got[:limit]})
    etc = [r for r in hits if (r.get("path") or "") not in used]
    if etc:
        out.append({"label": "그 밖의 원본", "n": len(etc),
                    "files": [{"name": r.get("name"), "path": r.get("path"),
                               "ext": r.get("ext"), "date": r.get("date")}
                              for r in etc[:limit]]})
    return {"groups": out, "band_url": _band_urls().get(uj, ""),
            "built": doc.get("built"), "총": len(hits)}


_CUST_IDX = {"at": 0, "data": {}}


def _customer_index():
    """캠프명 → ERP 거래처코드 색인(customer_index.py 산출물). 5분 캐시.
    사용자 지시 2026-08-05 "거래처 코드 앱과 엑셀에 표기"."""
    now = time.time()
    if now - _CUST_IDX["at"] < 300 and _CUST_IDX["data"]:
        return _CUST_IDX["data"]
    try:
        with open(os.path.join(ROOT, "reports", "거래처코드_색인.json"),
                  encoding="utf-8") as f:
            _CUST_IDX["data"] = json.load(f).get("linked") or {}
    except Exception:
        _CUST_IDX["data"] = {}
    _CUST_IDX["at"] = now
    return _CUST_IDX["data"]


_ERP_IDX = {"at": 0, "data": {}}


def _erp_sales_index():
    """ERP 판매조회 프로젝트 색인(erp_sales_index.py 산출물)을 5분 캐시로 읽는다.
    화면을 열 때마다 Z: 의 엑셀을 훑으면 앱이 느려지므로 JSON 만 본다."""
    now = time.time()
    if now - _ERP_IDX["at"] < 300 and _ERP_IDX["data"]:
        return _ERP_IDX["data"]
    try:
        with open(os.path.join(ROOT, "reports", "ERP판매_프로젝트색인.json"),
                  encoding="utf-8") as f:
            _ERP_IDX["data"] = json.load(f).get("index") or {}
    except Exception:
        _ERP_IDX["data"] = {}
    _ERP_IDX["at"] = now
    return _ERP_IDX["data"]


def _append_kakao_warning(report):
    """카톡 내보내기가 1일 넘게 오래되면 '3. 리스크'에 경고 한 줄을 얹는다
    (사용자 지시 2026-08-04). 판정은 kakao/export_watch.py 가 남긴 JSON 을 읽기만 한다 —
    여기서 다시 계산하면 보고 화면 열 때마다 Z: 를 훑게 된다."""
    if not isinstance(report, dict):
        return report
    try:
        p = os.path.join(ROOT, "reports", "카톡_내보내기_경과.json")
        with open(p, encoding="utf-8") as f:
            k = json.load(f)
    except Exception:
        return report
    if not k.get("stale"):
        return report
    age = k.get("age_hours")
    val = f"{age/24:.1f}일 경과 — 갱신 필요" if age else "내보내기 파일 없음 — 갱신 필요"
    for sec in report.get("sections", []):
        if str(sec.get("title", "")).startswith("3. 리스크"):
            items = sec.setdefault("items", [])
            if not any(i and i[0] == "카톡 내보내기 오래됨" for i in items):
                items.append(["카톡 내보내기 오래됨", val])
            break
    return report


def _append_remote_section(report):
    """대표보고 '4. 리모컨 현황' — 엑셀이 아니라 DB에서 붙인다(사용자 지시 2026-08-04).

    리모컨은 처음부터 DB 정본이라 01_대표보고 시트에 행이 없다. 엑셀에 칸을 새로 파는
    대신 서버가 보고 시점에 현재값을 얹는다(DB-only 원칙). 한도 초과는 대표가 바로
    봐야 하는 항목이라 숫자만이 아니라 사람 이름까지 적는다.
    """
    if not isinstance(report, dict):
        return report
    if any(str(s.get("title", "")).startswith("4. 리모컨")
           for s in report.get("sections", [])):
        return report
    try:
        import ledger_db
        s = ledger_db.remote_status()
    except Exception:
        return report
    t = s.get("totals") or {}
    bs = s.get("branch_stock") or {}
    over = s.get("over_limit") or {}
    holds = sorted((s.get("holdings") or {}).items(),
                   key=lambda kv: -kv[1].get("holding", 0))

    # ★ 타일 값은 **숫자 한 덩어리**로만 둔다(사용자 지시 2026-08-04 "배열 깔끔하게").
    #   괄호 설명을 값에 붙이면 좁은 타일에서 두세 줄로 접혀 표가 지저분해진다.
    #   버전·이름 같은 부연은 아래 lines 로 내린다.
    items = [
        ["전체 보유", f"{t.get('all', 0)}개"],
        ["개인 보유", f"{t.get('holding', 0)}개"],
        ["지점 재고", f"{t.get('stock', 0)}개"],
        ["한도 초과", f"{len(over)}명"],
        ["부산공장", f"{(bs.get('부산') or {}).get('stock', 0)}개"],
        ["시화공장", f"{(bs.get('시화') or {}).get('stock', 0)}개"],
        ["증평본사", f"{(bs.get('증평') or {}).get('stock', 0)}개"],
        ["AS 담당자", f"{len([1 for _, h in holds if h.get('holding')])}명"],
    ]
    jp = (bs.get("증평") or {}).get("versions") or {}
    jp_txt = " · ".join(f"{k} {v}개" for k, v in jp.items() if v)
    hold_txt = " · ".join(f"{n} {h.get('holding', 0)}개" for n, h in holds if h.get("holding"))
    over_txt = " · ".join(f"{n} {q}개" for n, q in sorted(over.items(), key=lambda kv: -kv[1]))
    lines = [f"담당자 보유 — {hold_txt}" if hold_txt else "담당자 보유 — 없음"]
    if jp_txt:
        lines.append(f"증평본사 버전별 — {jp_txt}")
    if over_txt:
        lines.append(f"⚠ 한도({s.get('limit', 3)}개) 초과 — {over_txt} · 추가 불출 전 "
                     f"기존 리모컨 납품·사용 내역 확인 필요")
    lines.append("불출 담당 — 부산: 오종현 · 시화: 안은숙 · 증평: 류지영 "
                 "(기준: 2026-07-29 재고표 + 이후 업무센터 입력)")
    sec = {"title": "4. 리모컨 현황 (현재 기준)", "items": items, "lines": lines}
    secs = report.setdefault("sections", [])
    # 사용자 지시: **리스크 바로 아래**에 놓는다. 맨 뒤에 붙이면 '오늘 우선 조치' 뒤로
    # 밀려 대표가 리스크와 이어서 못 본다.
    pos = next((i for i, s in enumerate(secs) if "리스크" in str(s.get("title", ""))), None)
    secs.insert(pos + 1, sec) if pos is not None else secs.append(sec)
    return report


def get_issues():
    """07_불일치누락현황 — 엑셀의 '검증 안 된·확인해야 할' 항목 그대로"""
    if DEMO:
        return {"rows": [{
            "구분": "정산", "ID": "JS-2607-002", "프로젝트NO": "UJ261002",
            "문제유형": "세금계산서 발행 승인 대기", "캠프명": "울산2캠프",
            "담당자": "류지영", "근거상태": "승인 필요",
            "내용·근거": "ERP 판매조회 4.세금계산서발행대기 — 실제 발행은 사람 승인",
        }], "cols": []}
    return cached_data("issues", _build_issues)


def _issue_truth_rows(rows):
    """23시트의 낡은 라벨을 현재 DB·ERP 근거로 재판정해 앱에만 즉시 반영한다.

    Excel은 11·15시 회차 전까지 일부러 건드리지 않는다. 대신 앱이 객관완료 정본을
    무시하거나 잘못된 ERP 키를 행동 목록으로 내보내지 않게 읽는 순간 진실층을 얹는다.
    정산 완료는 정산 행만 닫는다 — 같은 JS가 ERP/금액 문제에도 쓰인 경우까지 지우면
    서로 다른 사건이 같이 사라진다.
    """
    try:
        import ledger_db
        resolutions = ledger_db.resolutions()
    except Exception:
        resolutions = {}
    try:
        with open(os.path.join(ROOT, "reports", "ERP판매_프로젝트색인.json"),
                  encoding="utf-8") as f:
            sales = json.load(f).get("index") or {}
    except Exception:
        sales = {}
    try:
        with open(os.path.join(ROOT, "reports", "ERP원장대조_상태.json"),
                  encoding="utf-8") as f:
            key_state = json.load(f)
    except Exception:
        key_state = {}

    # 23시트의 낡은 정산 경고보다 exact 원천업무 관계를 우선한다. 취소건에 실제
    # 청구자료가 없으면 조치 목록에서 빼고, 있으면 삭제하지 않고 교차입력 충돌 한 건으로
    # 바꾼다. 프로젝트·캠프 유사매칭은 재무 판정에 쓰지 않는다.
    cancelled_settlements = {}
    try:
        from app_store import default_store
        store = default_store()
        outcomes = _source_outcome_map(store)
        from cancel_resolution import settlement_document_evidence
        side_documents = settlement_document_evidence(store)
        for settle in store.list_sheet_rows("06_거래서류청구수금"):
            settle_id = str(settle.get("정산ID") or settle.get("_business_key") or "").strip()
            source_id = str(settle.get("원천업무ID") or "").strip()
            outcome = outcomes.get(source_id) or {}
            if settle_id and outcome.get("cancelled"):
                cancelled_settlements[settle_id] = {
                    "outcome": outcome,
                    "has_documents": (
                        _settlement_has_documents(settle)
                        or bool(side_documents.get(settle_id))
                    ),
                }
    except Exception:
        cancelled_settlements = {}

    out, removed_done, removed_bad_key, removed_cancelled = [], 0, 0, 0
    cancel_conflict_added = set()
    for raw in rows or []:
        row = dict(raw)
        issue_id = str(row.get("ID") or row.get("정산ID") or row.get("업무ID") or "").strip()
        category = str(row.get("구분") or "").strip()
        issue = str(row.get("문제유형") or "").strip()
        cancelled = cancelled_settlements.get(issue_id) or {}
        if cancelled and (category == "정산" or issue_id.startswith("JS-")):
            if not cancelled.get("has_documents"):
                removed_cancelled += 1
                continue
            if issue_id in cancel_conflict_added:
                removed_cancelled += 1
                continue
            cancel_conflict_added.add(issue_id)
            outcome = cancelled.get("outcome") or {}
            row["문제유형"] = "취소건 청구자료 존재 — 교차입력 확인"
            row["근거상태"] = "교차입력 확인"
            row["내용·근거"] = (
                f"원천업무 {outcome.get('source_id') or '-'} {outcome.get('status') or '취소'}"
                f" · {outcome.get('reason') or '접수취소'}인데 PO·명세서·계산서·입금 중 "
                "하나 이상이 존재합니다. 서류는 삭제하지 않고 연결 건만 확인합니다."
            )
            out.append(row)
            continue
        resolved = str((resolutions.get(issue_id) or {}).get("status") or "")
        if category == "정산" and issue_id.startswith("JS-") and resolved.startswith("완료("):
            removed_done += 1
            continue
        if key_state.get("key_looks_wrong") and category == "ERP" and issue.startswith("ERP "):
            removed_bad_key += 1
            continue

        project = str(row.get("프로젝트NO") or "").strip().upper()
        sale = sales.get(project) or {}
        stage = str(sale.get("state") or sale.get("status") or "").strip()
        why, lane = "", "사람 확인"
        if issue in ("세금계산서 미발행", "입금 대기", "미청구(전표 없음)"):
            if stage.startswith("4."):
                row["문제유형"] = "세금계산서 발행 승인 대기"
                why, lane = f"ERP 판매조회 {stage} — 실제 발행은 사람 승인", "승인 필요"
            elif stage.startswith(("1.", "2.", "3.")) or stage in ("확인", "오더처리"):
                row["문제유형"] = "ERP 선행업무 진행 중"
                why, lane = f"ERP 판매조회 {stage} — 아직 계산서/입금 단계 전", "자동 추적"
            elif not stage:
                row["문제유형"] = ("세금계산서 ERP 연결자료 필요"
                                  if issue == "세금계산서 미발행" else "ERP 연결자료 필요")
                why, lane = "ERP 판매조회 프로젝트 색인에 없음 — 실제 미발행·미입금 확정 아님", "자료 필요"
            else:
                why = f"ERP 판매조회 {stage}"
        elif issue == "비용구분 미입력":
            why, lane = "유상·무상·보험을 입증할 PO/견적/보험 근거가 필요", "자료 필요"
        row["근거상태"] = lane
        if why:
            row["내용·근거"] = (why + (" · " + str(row.get("내용·근거") or "")
                                      if row.get("내용·근거") else ""))[:500]
        out.append(row)

    if key_state.get("key_looks_wrong") and removed_bad_key:
        counts = key_state.get("counts") or {}
        out.append({
            "구분": "시스템", "ID": "ERP-KEY", "문제유형": "ERP 매칭키 불일치",
            "담당자": "유현민", "근거상태": "시스템 진단",
            "내용·근거": (
                f"ERP 고유전표 {key_state.get('erp_unique_slips', 0)}건 중 전표번호 직접매칭 "
                f"{key_state.get('matched_by_slip', 0)}건. 원시 A {counts.get('A', 0)} / "
                f"B {counts.get('B', 0)} / C {counts.get('C', 0)} / D {counts.get('D', 0)}는 "
                "개별 미등록으로 단정하지 않고 프로젝트·PO 직접근거로 재대조합니다."
            ),
        })
    return out, {"objective_done_hidden": removed_done,
                 "unreliable_erp_hidden": removed_bad_key,
                 "cancelled_billing_hidden": removed_cancelled,
                 "cancelled_billing_conflicts": len(cancel_conflict_added),
                 "erp_key_healthy": not bool(key_state.get("key_looks_wrong"))}


def _build_issues():
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = master_book(master)
    rows = []
    # 1순위: 관리대장 통합 시트 23_확인필요현황 (에이전트가 매일 갱신 — 단일 엑셀 관리)
    if "23_확인필요현황" in wb.sheetnames:
        ws = wb["23_확인필요현황"]
        hdr = [str(h).strip() for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True)) if h]
        for row in ws.iter_rows(min_row=5, values_only=True):
            vals = {hdr[i]: ("" if i >= len(row) or row[i] is None else str(row[i]))
                    for i in range(len(hdr))}
            if any(v for v in vals.values()):
                rows.append(vals)
        wb.close()
        rows, truth = _issue_truth_rows(rows)
        from responsibility import assign_issue_row
        rows = [assign_issue_row(row) for row in rows]
        rows = app_year_rows(apply_rep_no(rows), "issue")
        cols = list(hdr)
        for name in ("근거상태", "내용·근거"):
            if name not in cols:
                cols.append(name)
        out = {"rows": sort_by_date(rows, "check"), "cols": cols,
               "source": "23_확인필요현황 + DB/ERP 현재근거", "truth": truth}
        return _store_cache("issues", out)
    if "07_불일치누락현황" in wb.sheetnames:
        ws = wb["07_불일치누락현황"]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        heads = [(i, str(h).strip()) for i, h in enumerate(hdr) if h is not None]
        for row in ws.iter_rows(min_row=5, values_only=True):
            vals = {h: ("" if i >= len(row) or row[i] is None else
                        (str(row[i])[:10] if hasattr(row[i], "year") else str(row[i])))
                    for i, h in heads}
            if any(v for v in vals.values()):
                rows.append(vals)
            if len(rows) >= 300:
                break
    wb.close()
    # 대조 결과(밴드·카톡·ERP원장·쿠팡PO) 통합 — 07시트 위에 얹어 한 화면에서 전부 확인
    merged = []
    try:
        from findings_export import latest_csv
        for src, pat, filt in (("밴드 게시 미확인", "밴드대조_*.csv", lambda r: r.get("밴드게시") == "미확인"),
                               ("카톡 보고 미확인", "카톡대조_*.csv", lambda r: r.get("카톡보고") == "미확인"),
                               ("ERP원장 문제", "ERP원장대조_*.csv", lambda r: True),
                               ("쿠팡PO 문제", "PO대조_*.csv", lambda r: True)):
            for r in latest_csv(pat):
                if filt(r):
                    merged.append({
                        "문제유형": src + (f"({r['유형']})" if r.get("유형") else ""),
                        "업무ID": r.get("ID") or r.get("정산ID") or r.get("전표") or r.get("PO번호") or "",
                        "프로젝트NO": r.get("프로젝트NO") or "",
                        "기준일": (r.get("접수일자") or r.get("점검예정일") or r.get("완료일") or
                                   r.get("발행일") or r.get("일자") or ""),
                        "캠프명": r.get("캠프명", ""), "담당자": r.get("담당기사", ""),
                        "문제내용": (r.get("판정") or r.get("내용") or
                                     f"완료 {r.get('완료일','')}" ) [:100]})
    except Exception:
        pass
    from responsibility import assign_issue_row
    rows = [assign_issue_row(row) for row in merged + rows]
    rows = sort_by_date(app_year_rows(apply_rep_no(rows), "issue"), "check")
    cols = []
    for r in rows[:50]:
        for k in r:
            if k not in cols:
                cols.append(k)
    return {"rows": rows, "cols": cols}


def build_id():
    """index.html이 바뀌면 값이 달라진다. 폰에 열려 있는 앱이 구버전인지 판별하는 기준."""
    try:
        st = os.stat(os.path.join(BASE, "index.html"))
        return hashlib.md5(f"{int(st.st_mtime)}-{st.st_size}".encode()).hexdigest()[:10]
    except Exception:
        return "0"


def icon_revision():
    """아이콘 원본이 바뀌면 설치 앱이 새 URL로 다시 받도록 내용 해시를 버전으로 쓴다."""
    try:
        path = os.path.join(BASE, "brand", "csos-app-icon-source.png")
        with open(path, "rb") as f:
            return "csos-" + hashlib.sha256(f.read()).hexdigest()[:12]
    except Exception:
        return "csos-default"


def sync_installed_app_icons():
    """Windows 앱 서버가 갱신될 때 설치 PWA·바로가기 아이콘도 같은 원본으로 맞춘다."""
    if os.name != "nt" or DEMO:
        return ""
    script = os.path.join(BASE, "sync_app_icons.ps1")
    if not os.path.isfile(script):
        return ""
    try:
        from proc_guard import run_tree
        result = run_tree(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", script],
            cwd=ROOT, env=ENV, timeout=60, drain_timeout=10, output_limit=20_000,
        )
        if result.timed_out:
            return ("아이콘 자동동기화 시간초과 · 자식나무 정리"
                    f" · 잔류 pid={result.stuck_pid or 0}")
        return (result.stdout or result.stderr or "").strip()[-300:]
    except Exception as exc:
        return f"아이콘 자동동기화 실패: {type(exc).__name__}"


def brand_logo():
    """webapp/brand/ 에 넣어 둔 고객사 로고 파일명. 없으면 빈 문자열(기본 CSOS 마크 사용).
    파일은 gitignore 대상 — 상표 자산을 공개 저장소에 올리지 않기 위해서다."""
    d = os.path.join(BASE, "brand")
    if not os.path.isdir(d):
        return ""
    # 보고서 캡처에는 사용자가 지정한 유니버셜리프트 가로 CI 한 장을 우선한다.
    # 쿠팡 CI는 앱바에서 별도 자산으로 표시하므로 여기서 먼저 고르면 보고서의
    # 유니버셜리프트 CI가 쿠팡 파일명 정렬순서에 밀려 사라진다.
    preferred = "universal-lift-horizontal.png"
    if os.path.isfile(os.path.join(d, preferred)):
        return preferred
    for f in sorted(os.listdir(d)):
        if os.path.splitext(f)[1].lower() in (".png", ".svg", ".jpg", ".jpeg", ".webp"):
            return f
    return ""


def get_erpdocs():
    """25_ERP매출서류 — 이카운트 매출(세금)계산서 원본(2026년 전체).
    ERP는 여러 작업을 한 장으로 묶어 발행하므로 1행 = 작업 1건이 아니다."""
    if DEMO:
        return {"rows": [], "months": {}, "total": 0}
    return cached_data("erpdocs", _build_erpdocs)


def _build_erpdocs():
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    out = {"rows": [], "months": {}, "total": 0, "kinds": {}}
    try:
        wb = master_book(master)
        if "25_ERP매출서류" in wb.sheetnames:
            for row in wb["25_ERP매출서류"].iter_rows(min_row=5, values_only=True):
                if not row or not row[0]:
                    continue
                slip, mo, kind, sup = row[0], row[1], row[2], int(row[3] or 0)
                rec = {"전표": str(slip), "월": str(mo), "유형": str(kind or ""),
                       "공급가액": sup, "거래처": str(row[6] or ""),
                       "프로젝트명": str(row[7] or "")}
                if not app_year_record(rec, "erp"):
                    continue
                out["rows"].append(rec)
                m = out["months"].setdefault(str(mo), {"합계": 0, "건수": 0})
                m["합계"] += sup
                m["건수"] += 1
                m[str(kind)] = m.get(str(kind), 0) + sup
                out["kinds"][str(kind)] = out["kinds"].get(str(kind), 0) + sup
                out["total"] += sup
        # 26_계산서구성 — 계산서 1장에 어떤 프로젝트가 묶였는지(추정). 전표번호로 붙인다.
        if "26_계산서구성" in wb.sheetnames:
            comp = {}
            for row in wb["26_계산서구성"].iter_rows(min_row=5, values_only=True):
                if not row or not row[0]:
                    continue
                comp[str(row[0])] = {"포함건수": int(row[5] or 0),
                                     "포함프로젝트": str(row[6] or ""),
                                     "후보합계": int(row[7] or 0),
                                     "판정": str(row[8] or "")}
            for r in out["rows"]:
                r.update(comp.get(r["전표"], {"포함건수": 0, "포함프로젝트": "",
                                             "후보합계": 0, "판정": "미상"}))
        wb.close()
    except Exception as e:
        out["error"] = str(e)
    out["rows"] = sort_by_date(app_year_rows(out["rows"], "erp"), "erpdocs")
    return out


# ── 매출 실적(전체) — 김미영 업무센터 (2026-08-12 지시) ──────────────────────
# 사용자 지시: **"김미영 업무센터 하나 추가해서 쿠팡 매출 실적(전체) 포함 관리할 수
#              있게 구성해서 만들어"**
#
# ★ **기준이 다르면 숫자는 안 맞는 것이 정상이다.** 김미영 표는 머리글에 적힌 그대로
#   **세금계산서 발행 기준달**로 묶여 있고, 우리 화면 대부분은 **완료월**로 묶는다.
#   그래서 이 화면만은 발행월로 센다 — 근거는 25_ERP매출서류(이카운트가 **실제로 끊은**
#   매출계산서)다. 완료월 집계를 그대로 대 놓으면 맞는 달까지 틀려 보인다.
# ★ **다를 때 원인을 단정하지 않는다.** 후보가 둘이고 둘 다 실재한다 —
#   ① 아직 안 끊긴 건(ERP 진행상태가 6·7 이 아님) ② 우리 자료가 아직 안 들어옴.
#   근거를 나란히 놓고 판단은 사람이 한다([172] — 잘못 지목하면 멀쩡한 값을 고치러 간다).
# ★ 여기서 Z: 를 훑지 않는다([168]). 읽는 것은 이미 만들어진 캐시·리포트 파일뿐이다.
KIM_TABLE_PATH = os.path.join(ROOT, "reports", "김미영_매출실적_2026.json")
ERP_PRJ_INDEX = os.path.join(ROOT, "reports", "ERP판매_프로젝트색인.json")
# 김미영 표 낱말 → 우리 유형. '유료AS' 는 **계산서가 끊긴 돌발AS** 다
# (무상·보험은 애초에 계산서가 없다 — 발행월 집계에서는 둘이 같은 것을 가리킨다).
KIM_KIND_MAP = {"정기점검": "정기점검", "유료AS": "돌발AS"}
REVENUE_KIND_ORDER = ["정기점검", "돌발AS", "신규납품", "철거", "계단", "기타"]
ISSUED_STATES = ("6.", "7.", "8.")      # 8.무상납품완료 는 청구가 없는 완료다
LEDGER_KINDS = ("돌발AS", "정기점검")   # 관리대장 원장이 담는 업무 둘([154])


def _erp_state_index():
    """ERP 판매 색인을 **파일에서만** 읽는다. 없으면 없다고 말한다.

    ★ 못 읽은 것을 0 으로 접으면 '미발행 0건'이 사실처럼 보인다([169])."""
    try:
        data = json.load(open(ERP_PRJ_INDEX, encoding="utf-8"))
        when = datetime.fromtimestamp(os.path.getmtime(ERP_PRJ_INDEX)).strftime("%Y-%m-%d %H:%M")
        return (data.get("index") or {}), when
    except Exception:
        return None, ""


def _revenue_compare(months):
    """김미영 표와 우리 발행월 집계를 나란히 놓는다. **고치지 않는다 — 보여 줄 뿐이다.**"""
    try:
        kim = json.load(open(KIM_TABLE_PATH, encoding="utf-8"))
    except Exception:
        return {"있음": False,
                "안내": "김미영 실적표가 아직 앱에 들어오지 않았습니다 — 자료가 없어 대조를 "
                        "못 한 것이지 '차이 없음'이 아닙니다"}
    ours = {m["월"]: (m.get("유형") or {}) for m in months}
    rows, same, diff = [], 0, 0
    for label, our_kind in KIM_KIND_MAP.items():
        table = kim.get(label + "_공급가액") or {}
        for mo in sorted(table):
            them = int(table[mo] or 0)
            us = int((ours.get(mo) or {}).get(our_kind) or 0)
            rows.append({"월": mo, "항목": label, "우리유형": our_kind,
                         "우리": us, "김미영": them, "차이": us - them,
                         "일치": us == them})
            if us == them:
                same += 1
            else:
                diff += 1
    rows.sort(key=lambda r: (r["월"], r["항목"]))
    총우리 = sum(r["우리"] for r in rows)
    총그쪽 = sum(r["김미영"] for r in rows)
    return {
        "있음": True, "행": rows,
        "출처": kim.get("출처") or "",
        "부가세": kim.get("부가세") or "",
        "환산근거": kim.get("공급가액_환산근거") or "",
        "요약": {"일치한 달": same, "다른 달": diff,
                 "우리 합": 총우리, "김미영 합": 총그쪽, "차이 합": 총우리 - 총그쪽},
        "주의": kim.get("주의") or [],
    }


def _revenue_unissued():
    """아직 계산서가 안 끊긴 건 — **ERP 진행상태**가 근거다(원장 빈 칸이 아니다).

    ★ 원장의 계산서 칸은 사람 손 입력이라 대부분 비어 있다. 그 빈 칸을 세면
      '미발행 190건' 같은 숫자가 나오는데 그건 미발행이 아니라 **빈 칸**이었다."""
    idx, when = _erp_state_index()
    if idx is None:
        return {"있음": False,
                "안내": "ERP 판매 색인(reports/ERP판매_프로젝트색인.json)이 없어 세지 "
                        "못했습니다 — **'0건'이 아니라 '못 셈'입니다**"}
    seen, out_rows, 근거없음 = {}, [], 0
    for r in get_settlements() or []:
        if str(r.get("출처") or "") == "ERP":
            continue                     # 묶음 계산서 행 — 이미 끊긴 것이다
        kind = str(r.get("업무구분") or "").strip()
        if kind not in LEDGER_KINDS:
            continue
        prj = str(r.get("프로젝트NO") or "").strip()
        rec = idx.get(prj) if prj else None
        if not rec:
            근거없음 += 1                # 번호가 없거나 ERP 에 아직 없다 — 단정하지 않는다
            continue
        state = str(rec.get("state") or "")
        if state.startswith(ISSUED_STATES):
            continue
        if prj in seen:                  # 같은 프로젝트가 여러 행이면 금액은 한 번만
            continue
        seen[prj] = True
        out_rows.append({
            "프로젝트NO": prj, "업무구분": kind,
            "완료월": str(r.get("완료일") or "")[:7].replace("/", "-"),
            "캠프명": str(r.get("캠프명") or rec.get("cust") or ""),
            "공급가액": int(rec.get("supply") or r.get("공급가액") or 0),
            "ERP진행상태": state,
        })
    by_kind = {}
    for row in out_rows:
        cur = by_kind.setdefault(row["업무구분"], {"건수": 0, "공급가액": 0})
        cur["건수"] += 1
        cur["공급가액"] += row["공급가액"]
    out_rows.sort(key=lambda x: -x["공급가액"])
    CAP = 60
    return {
        "있음": True, "근거": "ERP 진행상태가 6.세금계산서발행·7.수금완료가 아닌 건",
        "색인시각": when, "건수": len(out_rows),
        "공급가액": sum(r["공급가액"] for r in out_rows),
        "항목별": by_kind, "행": out_rows[:CAP],
        "더있음": max(0, len(out_rows) - CAP),
        "근거없음": 근거없음,
        "근거없음뜻": "프로젝트NO 가 없거나 ERP 판매전표에 아직 없는 행 — 미발행이라고 "
                      "세지 않았습니다(모르는 것을 아는 것처럼 세지 않습니다)",
    }


def _build_revenue():
    docs = get_erpdocs()
    months = []
    for mo in sorted((docs.get("months") or {})):
        v = docs["months"][mo] or {}
        months.append({
            "월": str(mo).replace("/", "-")[:7],
            "건수": int(v.get("건수") or 0), "합계": int(v.get("합계") or 0),
            "유형": {k: int(x) for k, x in v.items() if k not in ("합계", "건수")},
        })
    out = {
        "ok": True,
        "기준": "세금계산서 **발행월** 기준입니다 — 완료월이 아닙니다(김미영 표와 같은 기준)",
        "출처": "관리대장 25_ERP매출서류 — 이카운트가 실제로 끊은 매출계산서",
        "유형순서": REVENUE_KIND_ORDER,
        "months": months,
        "kinds": {k: int(v) for k, v in (docs.get("kinds") or {}).items()},
        "total": int(docs.get("total") or 0),
        "건수": len(docs.get("rows") or []),
        "만든때": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    if docs.get("error"):
        out["원본오류"] = str(docs["error"])
    if not months:
        out["안내"] = ("25_ERP매출서류에서 읽은 계산서가 한 건도 없습니다 — 매출이 "
                       "없는 것이 아니라 **자료가 아직 안 들어온 것**일 수 있습니다")
    out["대조"] = _revenue_compare(months)
    out["미발행"] = _revenue_unissued()
    return out


def get_revenue():
    if DEMO:
        return {"ok": True, "months": [], "kinds": {}, "total": 0, "대조": {"있음": False},
                "미발행": {"있음": False}, "안내": "데모 화면입니다"}
    return cached_data("revenue", _build_revenue)


def _ux_summary():
    """앱 사용 흔적 요약 — 다음 개선을 **추측이 아니라 기록**으로 정하기 위한 것."""
    try:
        import ledger_db
        return ledger_db.ux_summary()
    except Exception as e:
        return {"error": str(e)[:80]}


def get_apply_window():
    """다음 Excel 보관본 생성까지 얼마나 남았나 — 앱이 항상 보여 준다.

    앱 DB 입력은 즉시 확정된다. 11:00·15:00은 Excel 정본 반영 시간이 아니라
    단방향 보관본 생성 회차다. 기존 응답 키는 설치된 앱 호환을 위해 유지한다."""
    try:
        import ledger_db
        # DB 장애 때 남긴 안전 JSON 스풀도 여기서 정본 DB로 회수한다.
        # Excel은 정본으로 역수입하지 않으며 11시·15시는 보관본 생성 회차다.
        ledger_db.intake_json(source="json-queue")
        return ledger_db.status()
    except Exception:
        try:
            return json.load(open(
                os.path.join(ROOT, "reports", "반영대기.json"), encoding="utf-8"))
        except Exception:
            return {"대기": 0, "다음반영": "", "남은분": 0,
                    "반영시각": ["11:00", "15:00"]}


def get_recalc_pending():
    """원장엔 올라왔지만 엑셀이 아직 계산하지 않아 화면에 안 나오는 건수.

    이걸 안 알려주면 사용자는 '넣었다는데 왜 없지?' 로 읽는다 — 숫자가 틀린 게 아니라
    아직 안 나온 것이다. recalc_pending.py 가 만들어 둔 캐시만 읽는다(원장 재읽기는 느리다)."""
    try:
        return json.load(open(os.path.join(ROOT, "reports", "재계산대기.json"), encoding="utf-8"))
    except Exception:
        return {"대기합계": 0, "항목": [], "안내": ""}


# ── 캘린더 분류(2026-08-05 지시: "돌발 AS·정기점검 등 선택한 항목만 볼 수 있게") ──
#    화면 필터는 **한글 라벨이 아니라 이 키**로 건다. 라벨은 문장이 바뀌면 같이 바뀌고,
#    그때마다 필터가 조용히 아무것도 못 걸러내는 상태가 된다.
# ★ 색은 **서로 확실히 달라야 한다** (2026-08-06 지적: "색상 다르게 표시해 헷갈린다").
#   정기점검 완료(#30D158)와 돌발AS 완료(#34C759)가 거의 같은 초록이라, 칩·달력·캡처
#   어디서도 둘을 구별할 수 없었다. 완료끼리는 계열을 갈라 둔다 — 정기점검=초록,
#   돌발AS=청록. 빨강 계열은 쓰지 않는다(완료인데 문제로 읽힌다).
#   ※ 같은 표가 `webapp/index.html` 의 CAL_FALLBACK_KINDS 에도 있다. **둘을 같이 고칠 것** —
#     오늘 폴더 이름을 두 곳에 적었다가 어긋난 사고를 이미 겪었다.
CAL_KINDS = [
    ("pm_plan",  "정기점검 예정",   "#0A84FF"),   # 파랑
    ("pm_pred",  "정기점검 예측",   "#5E5CE6"),   # 보라
    ("pm_done",  "정기점검 완료",   "#30D158"),   # 초록
    ("as_visit", "돌발AS 예정일",   "#FF9F0A"),   # 주황
    ("as_done",  "돌발AS 완료",     "#00A6A0"),   # 청록 — 초록과 확실히 가른다
    # ★ 미처리(2026-08-06 지시) — "언제 들어왔는데 아직 안 끝났나"가 달력에 보여야
    #   밀린 것이 눈에 띈다. 완료·예정과 헷갈리지 않게 붉은 계열로 확실히 가른다.
    ("as_open",   "돌발AS 미처리",  "#FF453A"),   # 빨강
    ("pm_overdue", "정기점검 미처리", "#C2185B"),  # 자주 — 빨강과도 갈린다
    ("etc",      "기타 일정",       "#8E8E93"),   # 회색
]


_BAND_EV = {"at": 0.0, "d": None}
_BAND_EV_TTL = 300


def _band_completion_index():
    """밴드·카톡이 **스스로 완료라고 말한** 프로젝트 색인 (2026-08-13 지시).

    형님 통화: "완료가 됐는데 네가 지금 미처리로 남아 있는 거야. 이렇게 되면
    일 안 하는 줄 알잖아." 원장 완료일 칸은 사람 손 입력이라 비어 있을 뿐인데,
    캘린더가 **원장만 읽어서** 다녀온 현장을 밀린 것으로 그렸다.

    · **읽는 자리는 하나다** — `band_extract.load_records()`(밴드 글 + 카톡 내보내기를
      같은 양식으로 읽는다). 여기서 다시 파싱하면 판정이 갈린다([162]).
    · **원장 빈 칸을 근거로 삼지 않는다** — 근거는 밴드·카톡 쪽에 있다([166]).
    · 비싼 파싱(실측 8,049건)은 **캐시 검사 뒤**에 온다([168]).
    · `latest` 는 밴드가 **어디까지 수집됐나**다. 이것이 없으면 "완료 글이 없다"와
      "아직 안 긁었다"를 구별 못 한다 — 0건을 근거로 쓰는 순간 `[169]` 다.

    돌려주는 것: {"완료": {프로젝트NO: 기록}, "언급": set(프로젝트NO), "최신": "YYYY-MM-DD"}
    """
    now = time.time()
    cur = _BAND_EV.get("d")
    if cur is not None and now - _BAND_EV["at"] < _BAND_EV_TTL:
        return cur

    # ★ 실측 첫 계산 **23.5초**(밴드 8,049건 파싱). 웹 요청 하나가 그만큼 서면 폰은
    #   그냥 실패로 본다([197]). 그래서 결과를 디스크에도 둔다 — 지문이 같으면
    #   **서버를 다시 띄워도 다시 안 센다**. 지문은 밴드 캐시 파일의 (이름,크기,수정시각)
    #   뿐이라 만드는 값이 싸다(로컬 디스크 · Z: 를 안 훑는다, [168]·[198]).
    cpath = os.path.join(ROOT, "reports", "밴드_완료색인.json")
    try:
        fp = []
        with os.scandir(os.path.join(ROOT, "band", "cache")) as it:
            for e in it:
                if e.name.endswith(".json") and not e.name.startswith(("raw_", "dump_")):
                    st = e.stat()
                    fp.append("%s|%d|%d" % (e.name, st.st_size, int(st.st_mtime)))
        fp = hashlib.sha256("\n".join(sorted(fp)).encode()).hexdigest()[:16]
    except Exception:
        fp = ""
    if fp:
        try:
            d = json.load(open(cpath, encoding="utf-8"))
            if d.get("지문") == fp:
                d["언급"] = set(d.get("언급") or ())
                _BAND_EV["d"], _BAND_EV["at"] = d, now
                return d
        except Exception:
            pass

    out = {"완료": {}, "언급": set(), "최신": "", "읽음": False, "지문": fp}
    try:
        import band_extract
        for r in band_extract.load_records() or []:
            pj = str(r.get("프로젝트NO") or "").strip()
            if not pj:
                continue
            out["언급"].add(pj)
            when = norm_date(r.get("작업일")) or norm_date(r.get("게시일"))
            if when > out["최신"]:
                out["최신"] = when
            if r.get("진행상태") == "작업완료":
                # 같은 프로젝트에 완료 글이 여러 번이면 **가장 이른 것**을 쓴다 —
                # 다시 올라온 글로 완료일이 뒤로 밀리면 안 된다.
                old = out["완료"].get(pj)
                if not old or (when and when < (norm_date(old.get("작업일")) or "9999")):
                    out["완료"][pj] = r
        out["읽음"] = True
    except Exception:
        # 못 읽었으면 **거르지 않고 예전대로** 간다. '못 읽음'을 '근거 없음'으로
        # 치지 않는다([169]) — 부르는 쪽이 `읽음` 을 보고 말을 고른다.
        pass
    # **못 읽은 결과는 디스크에 남기지 않는다** — 남기면 다음 재시작이 그 빈 색인을
    # 지문째 믿어 "완료 근거 0건"을 조용히 확언한다([169]).
    if fp and out["읽음"]:
        try:
            os.makedirs(os.path.dirname(cpath), exist_ok=True)
            tmp = cpath + ".tmp"
            with open(tmp, "w", encoding="utf-8") as fh:
                json.dump(dict(out, 언급=sorted(out["언급"])), fh, ensure_ascii=False)
            os.replace(tmp, cpath)
        except Exception:
            pass
    _BAND_EV["d"], _BAND_EV["at"] = out, now
    return out


def _why_still_open(row, idx, got):
    """왜 아직 미처리인가 — **모르는 이유를 갈라 말한다**([169]).

    '미처리 88건'만 있고 이유가 없으면 사람은 다녀온 현장을 다시 찾아간다.
    단정할 수 없으면 단정하지 않는다."""
    if not idx.get("읽음"):
        return "밴드·카톡 기록을 못 읽었다 — 완료 여부를 확인 못 함"
    pj = str(row.get("프로젝트NO") or "").split(" · ")[0].strip()
    last = idx.get("최신") or ""
    if pj and pj in idx.get("완료", {}):
        return "밴드는 완료라 하는데 같은 프로젝트에 열린 건이 여럿 — 사람이 가른다"
    if last and got and got > last:
        return f"밴드 수집이 {last} 까지만 왔다 — 아직 못 본 것일 수 있다"
    if pj and pj not in idx.get("언급", ()):
        return "밴드·카톡에 이 프로젝트 글이 없다"
    return "밴드에 접수 글은 있으나 완료 글이 없다"


def _calendar_work_events():
    """돌발AS·정기점검 실적을 캘린더 일정으로 바꾼다.

    캘린더에 정기점검 예정만 있으면 "오늘 무슨 일이 있었나"를 볼 수 없다. 원장에
    이미 날짜가 있는 것(방문예정일·작업완료일·실제점검일)을 일정으로 세워 두면
    같은 화면에서 예정과 실적을 나란히 본다. **원장에 있는 날짜만** 쓴다 —
    없는 날짜를 추정해 캘린더에 그리지 않는다.
    """
    out = []
    # 취소·완료 낱말 판정은 `work_flow` 한 곳에서 온다([162]). 이 모듈은 표준 라이브러리만
    # 쓰므로 부르는 값이 싸다 — 못 불러오면 **거르지 않고 예전대로** 간다(거짓으로
    # 지우는 것보다 예전 그림이 낫다, [172]).
    import work_flow
    try:
        works = get_works() or {}
    except Exception:
        return out

    def add(when, kind, title, row, extra=None):
        d = norm_date(when)
        if not d:
            return
        e = {
            "날짜": d, "시간": "", "제목": title,
            "장소": row.get("캠프명") or "", "캠프명": row.get("캠프명") or "",
            "업무구분": dict((k, l) for k, l, _c in CAL_KINDS).get(kind, kind),
            "분류": kind,
            "프로젝트NO": str(row.get("프로젝트NO") or "").split(" · ")[0],
            "원천업무ID": row.get("접수ID") or row.get("점검ID") or "",
            "담당기사": row.get("담당기사") or "",
            "연결근거": "관리대장 원장에 기록된 날짜",
            "예측": False, "예측신뢰도": "", "장비수": 0,
        }
        e.update(extra or {})
        out.append(e)

    today = datetime.now().strftime("%Y-%m-%d")
    # ★ 밴드·카톡이 완료라 말하는 건을 미처리에서 내린다(2026-08-13 지시).
    #   ⚠ **문을 좁게 건다**([172]·[208]): 그 프로젝트에 열린 원장 행이 **하나뿐일 때만**.
    #   여럿이면 어느 건이 끝난 것인지 원본이 말해 주지 않는다 — 짐작으로 닫으면
    #   아무도 안 가는데 목록에도 없다. 그때는 닫지 말고 **이유를 적어 남긴다**.
    bidx = _band_completion_index()
    열린수 = {}
    for _r in works.get("as") or []:
        if not norm_date(_r.get("작업완료일")) and not work_flow.is_cancelled(_r, "as"):
            _p = str(_r.get("프로젝트NO") or "").split(" · ")[0].strip()
            if _p:
                열린수[("as", _p)] = 열린수.get(("as", _p), 0) + 1
    for _r in works.get("pm") or []:
        if not norm_date(_r.get("실제점검일")) and not work_flow.is_cancelled(_r, "pm"):
            _p = str(_r.get("프로젝트NO") or "").split(" · ")[0].strip()
            if _p:
                열린수[("pm", _p)] = 열린수.get(("pm", _p), 0) + 1

    def _band_done(row, kind):
        """이 행을 완료로 볼 밴드·카톡 근거 — 없으면 None."""
        pj = str(row.get("프로젝트NO") or "").split(" · ")[0].strip()
        if not pj or 열린수.get((kind, pj), 0) != 1:
            return None
        ev = bidx.get("완료", {}).get(pj)
        if not ev:
            return None
        when = norm_date(ev.get("작업일")) or norm_date(ev.get("게시일"))
        return (when, ev) if when else None

    for r in works.get("as") or []:
        camp = r.get("캠프명") or "캠프 미상"
        done = norm_date(r.get("작업완료일"))
        # ★ 접수취소는 미처리가 아니다 (2026-08-13 지시 — 형님 통화: "접수해서
        #   취소된 건 다 지워버려. 있으면 계속 접수돼 있는 상태가 되잖아").
        #   실측 2026-08-13: as_open 131건 중 **41건(31%)이 진행상태 '취소'** 였다.
        #   취소된 현장이 매일 빨간 미처리로 서 있으면 "일 안 하는 줄 안다" — 그리고
        #   경보의 3분의 1이 가짜면 나머지도 아무도 안 본다([170]).
        #   ★ 그렇다고 **조용히 지우지는 않는다**([169]) — 회색 `etc` 로 내려 기록은
        #     남긴다. 목록에서 통째로 없애면 "그때 정말 접수가 있었나"를 잃는다.
        if work_flow.is_cancelled(r, "as"):
            add(norm_date(r.get("접수일자")) or done or r.get("방문예정일"), "etc",
                f"돌발AS 접수취소 · {camp}", r,
                {"연결근거": "02_돌발AS접수 진행상태 = 취소",
                 "진행상태": r.get("진행상태") or "",
                 "신청내용": r.get("신청내용") or ""})
            continue
        if not done:                       # 아직 안 끝난 건만 '예정'으로 세운다
            add(r.get("방문예정일"), "as_visit", f"돌발AS 예정 · {camp}", r,
                {"연결근거": "02_돌발AS접수 방문예정일",
                 "긴급도": r.get("긴급도") or "", "진행상태": r.get("진행상태") or ""})
            # ★ 미처리는 **접수일** 자리에 세운다(2026-08-06 지시). 그래야 달력에서
            #   "언제 들어온 게 아직도 안 끝났나"가 바로 보인다. 방문예정일이 비어
            #   있어도(원장에 거의 안 채워진다) 이건 반드시 잡힌다.
            got = norm_date(r.get("접수일자"))
            if got:
                days = (datetime.strptime(today, "%Y-%m-%d")
                        - datetime.strptime(got, "%Y-%m-%d")).days
                # ★ 상태가 이미 '작업완료'라고 말하는데 **날짜 칸만 빈** 것을 '미처리'라
                #   부르지 않는다(실측 2026-08-13: 2건). 다녀온 현장이 안 간 것으로
                #   보이면 그게 형님이 지적한 "완료된 것도 다 들어가 있다"의 정체다.
                #   그렇다고 완료로 세지도 않는다 — 언제 갔는지를 모르기 때문이다.
                #   모르는 것은 모른다고 회색으로 둔다([169]).
                if work_flow.says_done(r, "as"):
                    add(got, "etc", f"돌발AS 완료일 미기입 · {camp}", r,
                        {"연결근거": "02_돌발AS접수 진행상태는 완료인데 작업완료일이 비어 있음",
                         "경과일": days, "진행상태": r.get("진행상태") or "",
                         "신청내용": r.get("신청내용") or ""})
                    continue
                # ★ 밴드·카톡이 완료라 말하면 미처리가 아니다 — 다녀온 현장이다.
                bd = _band_done(r, "as")
                if bd:
                    when, ev = bd
                    add(when, "as_done", f"돌발AS 완료 · {camp}", r,
                        {"연결근거": "밴드·카톡 완료 게시 — 원장 작업완료일은 아직 빈칸",
                         "근거출처": ev.get("밴드") or ev.get("게시글") or "밴드·카톡",
                         "원장미기입": True, "접수일자": got,
                         "진행상태": r.get("진행상태") or "",
                         "신청내용": r.get("신청내용") or ""})
                    continue
                add(got, "as_open", f"돌발AS 미처리 · {camp}", r,
                    {"연결근거": "02_돌발AS접수 — 접수 뒤 작업완료일이 비어 있음",
                     "경과일": days, "긴급도": r.get("긴급도") or "",
                     "미처리사유": _why_still_open(r, bidx, got),
                     "진행상태": r.get("진행상태") or "", "신청내용": r.get("신청내용") or ""})
        # ★ 완료 건에도 **무슨 일이었는지**를 실어 보낸다(2026-08-08 지시:
        #   "리스트 중간 빈 공간에 사유 진행내용 등 표기"). 캡처 목록의 가운데 칸이
        #   이 값을 쓴다 — 없으면 캠프명만 늘어서서 "무슨 건인지"를 다시 찾아야 한다.
        add(r.get("작업완료일"), "as_done", f"돌발AS 완료 · {camp}", r,
            {"연결근거": "02_돌발AS접수 작업완료일",
             "유무상": r.get("유상·무상·보험") or "",
             "진행상태": r.get("진행상태") or "",
             "신청내용": r.get("신청내용") or ""})
    # ★ 같은 점검이 **완료와 미처리에 동시에** 나오지 않게 한다 (2026-08-11 지시 —
    #   김미영 통화: "미처리에도 나와 있고 완료에도 나와 있다니까").
    #   실측 8/3 남양주4MB(UJ2601329)·부산1MB(UJ2601321)가 캘린더 양쪽에 다 있었다.
    #   원인은 04시트에 같은 프로젝트가 **여러 행**으로 있고 그중 한 행에만 실제점검일이
    #   채워진 것이다 — 행 하나만 보는 판정으로는 영영 안 걸린다(오류도 안 난다).
    #   ★ 근거는 좁게 잡는다: **예정일 이후에 실제로 다녀온 기록이 있을 때만** 그 예정을
    #     치른 것으로 본다. 완료일이 예정일보다 **앞서면** 다른 회차이므로 미처리로 남긴다.
    #     넓게 잡으면 진짜 안 간 현장이 목록에서 사라진다 — 겹쳐 보이는 것보다 나쁘다.
    pm_done_days = {}
    for r in works.get("pm") or []:
        key = str(r.get("프로젝트NO") or r.get("점검ID") or "").strip()
        d = norm_date(r.get("실제점검일"))
        if key and d:
            pm_done_days.setdefault(key, []).append(d)
    for r in works.get("pm") or []:
        camp = r.get("캠프명") or "캠프 미상"
        # 취소는 AS 와 같은 자리에서 같은 판단으로 거른다 — 판정이 갈리면 한쪽 화면만
        # 고쳐진다([162]). 실측 2026-08-13 기준 04시트 점검상태에는 취소 낱말이 아직
        # 없다(완료·미점검·예정·예정월) — **없다고 안 거는 것이 아니라**, 생기는 날
        # 저절로 걸리게 둔다.
        if work_flow.is_cancelled(r, "pm"):
            add(norm_date(r.get("점검예정일")) or r.get("실제점검일"), "etc",
                f"정기점검 취소 · {camp}", r,
                {"연결근거": "04_정기점검 점검상태 = 취소",
                 "점검상태": r.get("점검상태") or ""})
            continue
        add(r.get("실제점검일"), "pm_done", f"정기점검 완료 · {camp}", r,
            {"연결근거": "04_정기점검 실제점검일",
             "이상발견": r.get("이상발견여부") or ""})
        # 예정일이 지났는데 실제점검일이 없다 = 미처리. 앞날 예정은 pm_plan 이 맡는다.
        plan = norm_date(r.get("점검예정일"))
        key = str(r.get("프로젝트NO") or r.get("점검ID") or "").strip()
        served = any(d >= plan for d in pm_done_days.get(key, ())) if plan else False
        if plan and not norm_date(r.get("실제점검일")) and plan <= today and not served:
            days = (datetime.strptime(today, "%Y-%m-%d")
                    - datetime.strptime(plan, "%Y-%m-%d")).days
            bd = _band_done(r, "pm")
            if bd:
                when, ev = bd
                add(when, "pm_done", f"정기점검 완료 · {camp}", r,
                    {"연결근거": "밴드·카톡 완료 게시 — 원장 실제점검일은 아직 빈칸",
                     "근거출처": ev.get("밴드") or ev.get("게시글") or "밴드·카톡",
                     "원장미기입": True, "점검예정일": plan,
                     "점검상태": r.get("점검상태") or ""})
                continue
            add(plan, "pm_overdue", f"정기점검 미처리 · {camp}", r,
                {"연결근거": "04_정기점검 — 예정일이 지났는데 실제점검일이 비어 있음",
                 "경과일": days, "점검상태": r.get("점검상태") or "",
                 "미처리사유": _why_still_open(r, bidx, plan)})
    return out


def get_calendar():
    """구글 캘린더(COUPANG 설치+납품+AS) 대조 캐시.

    gcal_sync.py 가 매일 만들어 둔 파일만 읽는다 — 앱은 절대 네트워크를 타지 않는다.
    폰에서 열 때 구글을 기다리면 화면이 멈추고, 터널이 죽으면 통째로 안 뜬다."""
    p = os.path.join(ROOT, "reports", "gcal_events.json")
    try:
        d = json.load(open(p, encoding="utf-8"))
    except Exception:
        d = {"갱신": "", "일정": [], "원천": ["Google Calendar 아직 수집되지 않음"], "안내":
             "Google Calendar 원천은 아직 없으며, 정기점검 원본의 공식·예측 일정은 아래 앱 캘린더에 표시합니다."}

    # Google Calendar에 외부 쓰기를 하지 않고, 류지영 원본 스케줄의 공식 예정일과
    # 동일 장비 과거 점검일 기반 예측일을 앱 캐시에만 합친다. 예측은 제목·근거에서
    # 명확히 구분하며 사용자가 Google Calendar 초안 화면에서 저장하기 전에는 외부 일정이 아니다.
    try:
        pm = json.load(open(os.path.join(ROOT, "reports", "pm_schedule_sync.json"),
                            encoding="utf-8"))
        events = list(d.get("일정") or [])

        def event_key(event):
            camp = re.sub(r"[^0-9A-Za-z가-힣]", "", str(
                event.get("캠프명") or event.get("장소") or event.get("제목") or "")).lower()
            return str(event.get("날짜") or ""), camp

        existing = {event_key(e) for e in events}
        added = predicted_count = 0
        for row in pm.get("schedule") or []:
            official = norm_date(row.get("점검예정일"))
            predicted = norm_date(row.get("예측점검일"))
            when = official or predicted
            if not when:
                continue
            event = {
                "날짜": when,
                "시간": "",
                "제목": ("[예측] " if not official else "") +
                       f"정기점검 · {row.get('캠프명') or '캠프 미정'}",
                "장소": row.get("캠프명") or "",
                "캠프명": row.get("캠프명") or "",
                "업무구분": "정기점검 예정(예측)" if not official else "정기점검 예정",
                "분류": "pm_pred" if not official else "pm_plan",
                "프로젝트NO": str(row.get("연결프로젝트NO") or "").split(" · ")[0],
                "원천업무ID": row.get("일정ID") or "",
                "연결근거": (row.get("예측근거") if not official else
                           "류지영 정기점검 스케줄 원본의 공식 예정일"),
                "예측": not bool(official),
                "예측신뢰도": row.get("예측신뢰도") or "",
                "장비수": int(row.get("장비수") or 0),
            }
            key = event_key(event)
            if key in existing:
                continue
            existing.add(key)
            events.append(event)
            added += 1
            predicted_count += int(not bool(official))
        d["일정"] = events
        d["정기점검추가"] = added
        d["예측일정수"] = predicted_count
        d["원천"] = list(dict.fromkeys(list(d.get("원천") or []) +
                                  ["류지영 정기점검 스케줄 원본(공식·과거 이력 예측)"]))
    except Exception:
        pass

    # 돌발AS·정기점검 실적을 합친다(2026-08-05 지시). 같은 날·같은 캠프·같은 분류가
    # 두 번 들어가지 않게 키로 막는다 — 캘린더에 같은 일이 두 줄이면 세다가 틀린다.
    try:
        events = list(d.get("일정") or [])
        seen = {(str(e.get("날짜") or ""), str(e.get("분류") or ""),
                 str(e.get("원천업무ID") or e.get("캠프명") or "")) for e in events}
        for e in _calendar_work_events():
            key = (e["날짜"], e["분류"], str(e.get("원천업무ID") or e.get("캠프명") or ""))
            if key in seen:
                continue
            seen.add(key)
            events.append(e)
        d["일정"] = events
        d["원천"] = list(dict.fromkeys(list(d.get("원천") or []) +
                                       ["관리대장 02_돌발AS접수·04_정기점검 실적"]))
    except Exception:
        pass

    # 사람이 손으로 넣는 일정(회의·교육 등). 원장·구글·밴드 어디에도 없는 약속을
    # 캘린더에 올리는 유일한 자리다 — 없으면 "9월 말 증평 회의" 같은 것이 통화 기록에만
    # 남아 아무도 안 본다(2026-08-06 차동호 팀장 통화).
    # ★ 파일은 `config/manual_events.local.json` — **git 밖**이다. 이 저장소는 공개라
    #   회의 장소·참석자 같은 내부 내용을 평문으로 올리지 않는다. 공유 캘린더에는
    #   잠긴 꾸러미(cal.enc)를 통해 전달된다.
    try:
        mp = os.path.join(ROOT, "config", "manual_events.local.json")
        for e in (json.load(open(mp, encoding="utf-8")) or []):
            if not e.get("날짜"):
                continue
            e.setdefault("분류", "etc")
            e.setdefault("제목", e.get("캠프명") or "일정")
            e["수기"] = True
            d.setdefault("일정", []).append(e)
    except FileNotFoundError:
        pass
    except Exception as exc:
        print(f"  ! 수기 일정 건너뜀: {exc}")

    # 분류가 없는 옛 일정(Google 원천 등)도 필터에 걸리도록 자리를 준다.
    for e in d.get("일정") or []:
        e.setdefault("분류", "etc")
    d["일정"] = sorted(d.get("일정") or [], key=lambda e: (
        str(e.get("날짜") or "9999"), str(e.get("시간") or ""), str(e.get("제목") or "")))
    d["분류목록"] = [{"key": k, "label": l, "color": c} for k, l, c in CAL_KINDS]
    return d


# ── 프로젝트(현장) 한 곳의 내력 한 벌 ────────────────────────────────────────
# 2026-08-08 지시: "정기점검 예측에 프로젝트를 클릭하면 과거에 했던 내역들이 다 보이게 /
# 돌발AS 등 모든 프로젝트를 클릭하면 과거 돌발AS 또는 정기점검 리스트가 보이게 /
# 지금 현황 예측 현황도 다 같이"
#
# ★ 캘린더 일정을 다시 세지 않고 **원장 행에서 직접** 만든다. 캘린더는 '날짜가 있는 것'만
#   세우므로, 그것만 보면 날짜가 안 채워진 행이 조용히 사라진다. 이 화면은 "이 현장에서
#   무슨 일이 있었나"에 답해야 하므로 날짜 없는 행도 `날짜없음` 으로 따로 세어 보여 준다.
# ★ 맞추는 열쇠는 **캠프명**이다. 프로젝트NO 는 예측 일정에서 자주 비어 있다
#   (화면에 '프로젝트 미확정'으로 나오는 그것). 캠프명이 같으면 같은 현장으로 본다.
def _camp_key(s):
    return re.sub(r"[^0-9A-Za-z가-힣]", "", str(s or "")).lower()


def _pj_key(s):
    return str(s or "").split(" · ")[0].strip().upper()


def project_history(camp="", pj="", limit=400):
    """현장 한 곳의 과거 내력 + 지금 현황 + 앞으로 예정·예측."""
    ck, pk = _camp_key(camp), _pj_key(pj)
    if not ck and not pk:
        return {"ok": False, "error": "캠프명 또는 프로젝트NO가 필요합니다"}
    today = datetime.now().strftime("%Y-%m-%d")

    def mine(row):
        if ck and _camp_key(row.get("캠프명")) == ck:
            return True
        return bool(pk) and _pj_key(row.get("프로젝트NO")) == pk

    import work_flow          # 취소 판정은 캘린더와 같은 한 곳에서 온다([162])
    try:
        works = get_works() or {}
    except Exception as exc:
        return {"ok": False, "error": "원장을 읽지 못했습니다: %s" % exc}

    LAB = dict((k, l) for k, l, _c in CAL_KINDS)
    이력, 현황, 예정, 날짜없음 = [], [], [], []
    캠프이름 = str(camp or "").strip()
    프로젝트들 = []

    def item(kind, when, title, row, idkey, **extra):
        e = {"분류": kind, "라벨": LAB.get(kind, kind), "날짜": when or "",
             "제목": title, "캠프명": row.get("캠프명") or "",
             "프로젝트NO": _pj_key(row.get("프로젝트NO")),
             "원천업무ID": row.get(idkey) or "", "담당기사": row.get("담당기사") or ""}
        e.update({k: v for k, v in extra.items() if v not in (None, "")})
        return e

    for r in works.get("as") or []:
        if not mine(r):
            continue
        if not 캠프이름:
            캠프이름 = r.get("캠프명") or ""
        if _pj_key(r.get("프로젝트NO")):
            프로젝트들.append(_pj_key(r.get("프로젝트NO")))
        got, done = norm_date(r.get("접수일자")), norm_date(r.get("작업완료일"))
        vis = norm_date(r.get("방문예정일"))
        common = dict(긴급도=r.get("긴급도") or "", 진행상태=r.get("진행상태") or "",
                      신청내용=r.get("신청내용") or "", 유무상=r.get("유상·무상·보험") or "")
        # ★ 캘린더와 **같은 판단**을 쓴다(2026-08-13). 한 화면은 취소를 미처리라 하고
        #   다른 화면은 아니라고 하면, 같은 원장을 보면서 두 답이 나온다.
        if work_flow.is_cancelled(r, "as"):
            이력.append(item("etc", got or done, "돌발AS 접수취소", r, "접수ID",
                             라벨="접수취소", **common))
            continue
        if done:
            이력.append(item("as_done", done, "돌발AS 완료", r, "접수ID",
                             접수일자=got, 소요일=_daydiff(got, done), **common))
        else:
            # 아직 안 끝난 건. 접수일이 있으면 '미처리'로 현황에 세우고, 방문예정일이
            # 앞날이면 '예정'에도 함께 세운다 — 같은 건이 두 칸에 보이는 것이 맞다
            # (지금 밀려 있고, 언제 갈 예정인지는 다른 사실이다).
            if got:
                현황.append(item("as_open", got, "돌발AS 미처리", r, "접수ID",
                                 경과일=_daydiff(got, today), 방문예정일=vis, **common))
            elif not vis:
                # ★ 날짜를 모르는 행에 '완료'·'미처리' 색과 말을 붙이지 않는다 —
                #   모르는 것은 모른다고 회색(etc)으로 둔다.
                날짜없음.append(item("etc", "", "날짜 미기입", r, "접수ID",
                                     라벨="돌발AS", **common))
            if vis and vis >= today:
                예정.append(item("as_visit", vis, "돌발AS 방문예정", r, "접수ID", **common))

    for r in works.get("pm") or []:
        if not mine(r):
            continue
        if not 캠프이름:
            캠프이름 = r.get("캠프명") or ""
        if _pj_key(r.get("프로젝트NO")):
            프로젝트들.append(_pj_key(r.get("프로젝트NO")))
        plan, real = norm_date(r.get("점검예정일")), norm_date(r.get("실제점검일"))
        common = dict(점검상태=r.get("점검상태") or "", 이상발견=r.get("이상발견여부") or "",
                      돌발AS전환=r.get("돌발AS전환여부") or "")
        if work_flow.is_cancelled(r, "pm"):
            이력.append(item("etc", real or plan, "정기점검 취소", r, "점검ID",
                             라벨="취소", **common))
            continue
        if real:
            이력.append(item("pm_done", real, "정기점검 완료", r, "점검ID",
                             점검예정일=plan, **common))
        elif plan and plan <= today:
            현황.append(item("pm_overdue", plan, "정기점검 미처리", r, "점검ID",
                             경과일=_daydiff(plan, today), **common))
        elif plan:
            예정.append(item("pm_plan", plan, "정기점검 예정", r, "점검ID", **common))
        else:
            날짜없음.append(item("etc", "", "날짜 미기입", r, "점검ID",
                                 라벨="정기점검", **common))

    # 예측(류지영 스케줄 원본 기반) — 원장에 없는 '앞으로'는 여기서만 온다.
    try:
        pm = json.load(open(os.path.join(ROOT, "reports", "pm_schedule_sync.json"),
                            encoding="utf-8"))
        for row in pm.get("schedule") or []:
            if not mine(row) and not (ck and _camp_key(row.get("캠프명")) == ck):
                continue
            official, pred = norm_date(row.get("점검예정일")), norm_date(row.get("예측점검일"))
            when = official or pred
            if not when or when < today:
                continue
            예정.append({"분류": "pm_plan" if official else "pm_pred",
                         "라벨": LAB.get("pm_plan" if official else "pm_pred", ""),
                         "날짜": when, "제목": "정기점검 예정" if official else "정기점검 예측",
                         "캠프명": row.get("캠프명") or "",
                         "프로젝트NO": _pj_key(row.get("연결프로젝트NO")),
                         "원천업무ID": row.get("일정ID") or "", "담당기사": "",
                         "예측": not bool(official),
                         "예측신뢰도": row.get("예측신뢰도") or "",
                         "근거": row.get("예측근거") or "",
                         "장비수": int(row.get("장비수") or 0)})
    except Exception:
        pass

    이력.sort(key=lambda e: e["날짜"], reverse=True)
    현황.sort(key=lambda e: e["날짜"])
    예정.sort(key=lambda e: e["날짜"])

    pmdone = [e["날짜"] for e in 이력 if e["분류"] == "pm_done"]
    asdone = [e["날짜"] for e in 이력 if e["분류"] == "as_done"]
    # 평균 점검 주기 — 점검이 두 번 이상 있어야 말이 된다. 한 번뿐이면 셈하지 않는다.
    cyc = None
    if len(pmdone) >= 2:
        ds = sorted(pmdone)
        gaps = [_daydiff(a, b) for a, b in zip(ds, ds[1:])]
        gaps = [g for g in gaps if g and g > 0]
        if gaps:
            cyc = round(sum(gaps) / len(gaps))
    nxt = 예정[0] if 예정 else None
    return {
        "ok": True,
        "캠프명": 캠프이름 or (camp or "(장소 미입력)"),
        "프로젝트NO": " · ".join(sorted(set(프로젝트들))[:4]),
        "요약": {
            "돌발AS": len(asdone), "돌발AS미처리": sum(1 for e in 현황 if e["분류"] == "as_open"),
            "정기점검": len(pmdone), "정기점검미처리": sum(1 for e in 현황 if e["분류"] == "pm_overdue"),
            "마지막점검일": pmdone[0] if pmdone else "",
            "마지막점검경과": _daydiff(pmdone[0], today) if pmdone else None,
            "마지막AS일": asdone[0] if asdone else "",
            "평균점검주기": cyc,
            "다음일정": ({"날짜": nxt["날짜"], "라벨": nxt["라벨"],
                          "예측": bool(nxt.get("예측"))} if nxt else None),
        },
        "현황": 현황[:limit], "예정": 예정[:limit],
        "이력": 이력[:limit], "날짜없음": 날짜없음[:limit],
        "더있음": max(0, len(이력) - limit),
        "기준": today,
        "원천": "관리대장 02_돌발AS접수·04_정기점검 + 류지영 정기점검 스케줄(예측)",
    }


# ── 입력 도우미: 원장에 이미 있는 값을 추천한다 ──────────────────────────────
# 2026-08-08 지시: "입력란 입력할 때 DB 기반으로 자동 입력 추천 뜨게 전체 앱 시스템
# 코딩해 / 사용자가 매번 찾아 입력하기 불편해"
#
# ★ **없는 값을 만들지 않는다.** 추천은 오직 원장에 실제로 있는 값이다. 그럴듯한
#   후보를 지어내면 사람이 그걸 골라 새 오타가 원장에 들어간다 — 지금 있는 표기
#   흔들림("김준형 "·"김준형")을 줄이려고 만드는 기능이 반대로 늘리게 된다.
# ★ 순서는 **많이 쓰인 순**이다. 가나다순은 늘 같은 것을 맨 아래에 둔다.
_SUGGEST = {"at": 0.0, "idx": {}}
_SUGGEST_TTL = 180.0

# 어느 칸이 어느 열에서 오는가. (표, 열이름) 짝이며 여러 곳에서 모을 수 있다.
_SUGGEST_SRC = {
    "캠프명":     [("as", "캠프명"), ("pm", "캠프명")],
    "프로젝트NO": [("as", "프로젝트NO"), ("pm", "프로젝트NO")],
    "담당기사":   [("as", "담당기사"), ("pm", "담당기사")],
    "진행상태":   [("as", "진행상태")],
    "점검상태":   [("pm", "점검상태")],
    "긴급도":     [("as", "긴급도")],
    "유무상":     [("as", "유상·무상·보험"), ("pm", "유상·무상·보험")],
    "비용구분":   [("pm", "비용구분")],
    "담당관리자": [("pm", "담당관리자")],
    "이상발견":   [("pm", "이상발견여부")],
    "신청내용":   [("as", "신청내용")],
}


def _suggest_index():
    now = time.time()
    if _SUGGEST["idx"] and now - _SUGGEST["at"] < _SUGGEST_TTL:
        return _SUGGEST["idx"]
    idx = {}
    try:
        works = get_works() or {}
    except Exception:
        return _SUGGEST["idx"] or {}
    for field, srcs in _SUGGEST_SRC.items():
        cnt = {}
        for table, col in srcs:
            for r in works.get(table) or []:
                v = str(r.get(col) or "").strip()
                # 여러 명이 한 칸에 들어가는 열이 있다("김준형, 김필우") — 갈라서 센다.
                parts = re.split(r"[,/·]| 및 ", v) if field in ("담당기사", "담당관리자") else [v]
                for p in parts:
                    p = p.strip()
                    if p and len(p) <= 60:
                        cnt[p] = cnt.get(p, 0) + 1
        idx[field] = [v for v, _n in sorted(cnt.items(), key=lambda kv: (-kv[1], kv[0]))]
    # 사람 이름은 한 목록으로도 쓴다(제출자·작성자 칸처럼 역할이 섞이는 자리).
    try:
        from ledger_db import AS_TECHS
        fixed = list(AS_TECHS) + ["류지영", "오종현"]
    except Exception:
        fixed = []
    seen, people = set(), []
    for v in list(idx.get("담당기사") or []) + list(idx.get("담당관리자") or []) + fixed:
        if v not in seen:
            seen.add(v)
            people.append(v)
    idx["사람"] = people
    _SUGGEST["idx"], _SUGGEST["at"] = idx, now
    return idx


def suggest_values(field, q="", limit=40):
    """한 칸에 넣을 만한 값들. 원장에 있는 것만, 많이 쓰인 순."""
    idx = _suggest_index()
    vals = idx.get(str(field or "").strip()) or []
    q = str(q or "").strip().lower()
    if q:
        head = [v for v in vals if v.lower().startswith(q)]     # 앞에서 맞는 것이 먼저
        rest = [v for v in vals if q in v.lower() and v not in head]
        vals = head + rest
    return {"field": field, "총": len(idx.get(field) or []),
            "values": vals[:max(1, min(int(limit or 40), 200))],
            "원천": "관리대장 02_돌발AS접수·04_정기점검에 실제로 있는 값"}


def _daydiff(a, b):
    """a→b 일수. 둘 중 하나라도 날짜가 아니면 None — 0 으로 속이지 않는다."""
    try:
        return (datetime.strptime(str(b)[:10], "%Y-%m-%d")
                - datetime.strptime(str(a)[:10], "%Y-%m-%d")).days
    except Exception:
        return None


def _org_flow():
    """배치도를 잇는 **흐름 화살표**의 근거 — 2026-08-13 지시
    ("실제 배치도만 플로우 화살표 연결되게 … 어떤 식으로 일처리되는지 한눈에").

    ★ **단계 낱말도 순서도 담당도 여기서 지어내지 않는다** — `ledger_db.flow_steps()`
      한 곳이 정한다(`[196]`). 화면에 손으로 적으면 흐름도가 바뀐 날 기능만 옛 그림으로
      남고, 그게 바로 그 규칙이 막으려던 사고다.

    두 가지를 **가른다** — 그러지 않으면 그림이 거짓말을 한다:
      · **접수 경로**(`갈래 == "접수 경로"`)는 **나란한 갈래**다. 카톡·법인폰·기사폰·
        개인폰 어느 길로도 들어온다. 이것을 한 줄에 세우면 '카톡 다음에 법인폰'이라는
        **없는 순서**가 생긴다.
      · 나머지는 **차례**다. 담당이 같은 연속 단계는 한 마디로 접는다 —
        17개를 그대로 그리면 폰에서 아무도 못 읽는다(접는 것이지 빼는 것이 아니다:
        접힌 단계 이름은 `label` 에 그대로 남는다, `[236]`).

    못 읽으면 **빈 흐름이 아니라 `ok: False`** 를 준다 — 화면이 '흐름 없음'과
    '흐름을 못 읽음'을 갈라 말해야 한다(`[169]`).
    """
    try:
        import ledger_db as _ld
        steps = list(_ld.flow_steps() or [])
    except Exception:
        return {"ok": False, "entries": [], "chain": []}

    def _who(v):
        s = str(v or "")
        for sep in ("·", ",", "/"):
            s = s.replace(sep, "\n")
        return [w.strip() for w in s.split("\n") if w.strip()]

    entries, seq = [], []
    for s in steps:
        row = {"label": str(s.get("단계") or ""), "who": _who(s.get("담당")),
               "note": str(s.get("메모") or "")}
        (entries if str(s.get("갈래") or "") == "접수 경로" else seq).append(row)

    chain = []
    for r in seq:
        if chain and chain[-1]["who"] == r["who"]:
            chain[-1]["label"] += " · " + r["label"]
        else:
            chain.append(dict(r))
    for i, c in enumerate(chain, 1):
        c["n"] = i
    return {"ok": True, "entries": entries, "chain": chain}


def get_orgchart():
    """조직도 — 사무실 책상 배치도(2026-08-12 지시). 각 자리에 누가·역할·상태.

    ★ 자리와 역할은 로스터(STAFF_CENTERS·AS_TECH_CENTERS)가 정본이다.
      사람의 '지금 무엇을 하는가'는 앱이 진짜로는 모른다 — 그래서 지어내지 않고
      역할만 싣는다([169] '0을 안 본 것으로 만들지 마라'의 사람판). **실시간으로
      아는 것은 AI 세션뿐**이라, AI 구역만 ai_claim 점유로 살아 움직인다.
      _live 는 그 점유를 실제로 읽었는지를 말한다(못 읽으면 폰이 스냅샷으로 물러선다)."""
    import time as _t
    now = _t.time()
    PAL = ["#0062CC", "#12813F", "#B25000", "#7A3DB8",
           "#0E7490", "#B3261E", "#177245", "#8A5A00"]

    def _ago(mt):
        mins = int((now - mt) / 60) if mt else 0
        if mins < 1:
            return "방금"
        if mins < 60:
            return "%d분 전" % mins
        return "%d시간 %d분 전" % (mins // 60, mins % 60)

    def _ago_iso(v):
        """스케줄러가 주는 ISO 문자열을 같은 말투로 바꾼다 — **모르면 '기록 없음'**.
        빈 값을 '방금'이라 적으면 한 번도 안 돈 회차가 방금 돈 것으로 보인다([169])."""
        s = str(v or "").strip()
        if not s:
            return "기록 없음"
        for f in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M"):
            try:
                return _ago(datetime.strptime(s[:19], f).timestamp())
            except Exception:
                continue
        return s[:16]

    # ── 담당자 온/오프라인([169]): 인증 브라우저가 마지막으로 서버에 말한 시각만
    #    real signal 이다. 5분 내면 온라인. 기록이 아예 없으면 '오프라인'이 아니라
    #    '상태 모름'(dot 은 off 이되 문구로 못박지 않는다). ──
    ONLINE_SEC = 300
    try:
        seen = presence_snapshot()
    except Exception:
        seen = {}

    def _human_state(slug):
        ts = seen.get(str(slug), 0)
        if not ts:
            return ("off", "앱 접속 기록 없음 · 상태 모름", "")
        if now - ts <= ONLINE_SEC:
            return ("busy", "지금 접속 중 · 온라인", _ago(ts))
        return ("off", "오프라인 · 마지막 접속", _ago(ts))

    # ── 관리팀 — 로스터(STAFF_CENTERS)가 정본. duties = 그 센터의 체크리스트 전부.
    #    유현민은 '센터장' 딱지(예전 '대표'에서 바꿈, 2026-08-12 지시). ──
    mgmt = []
    for i, (slug, cfg) in enumerate(STAFF_CENTERS.items()):
        cl = list(cfg.get("checklist") or [])
        hst, hmsg, hago = _human_state(slug)
        mgmt.append({
            "name": cfg.get("name", slug),
            "role": cfg.get("title", ""),
            "color": PAL[i % len(PAL)],
            "state": hst,
            "msg": hmsg,
            "ago": hago,
            "badge": "센터장" if slug == "yoo-hyeonmin" else "",
            "duties": cl or ["업무센터 담당"],
        })

    # ── 현장 AS팀 — 로스터(AS_TECH_CENTERS)가 정본. 팀장은 배정·일정 한 줄을 앞에 둔다. ──
    FIELD_DUTIES = ["배정에 따라 현장 조치", "밴드에 조치·완료 보고", "리모컨 납품 기록"]
    field = []
    for i, (slug, cfg) in enumerate(AS_TECH_CENTERS.items()):
        title = cfg.get("직함") or ""
        duties = (["현장 AS 팀장 · 배정·일정 관리"] + FIELD_DUTIES
                  if title == "팀장" else list(FIELD_DUTIES))
        hst, hmsg, hago = _human_state(slug)
        field.append({
            "name": cfg.get("name", slug),
            "role": "현장 AS · 정기점검",
            "color": PAL[(i + 3) % len(PAL)],
            "state": hst,
            "msg": (title + " · " if title else "") + hmsg,
            "ago": hago,
            "badge": title,
            "duties": duties,
        })

    # ── AI 스테이션 — 열려 있는 **모든** 에이전트 세션을 한 자리씩 보여 준다
    #    (2026-08-12 지시: "N 세션 = N 데스크"). 예전엔 기본 자리 셋만 두고 ai_claim
    #    점유로만 덮어써서, 세션이 열려 있어도 자원을 안 잡았으면 전부 '오프라인'으로
    #    보였다. 이제 살아 있는 세션은 각자 하나의 live 데스크가 된다.
    #    · 살아 있는 세션 판정은 **새로 만들지 않고** session_wrapup 을 그대로 빌린다
    #      (대화기록 mtime + [225] .대화기록_폴더.txt). ai_claim 은 그 세션이 지금
    #      무엇을 잡고 있나(why·자원)만 덧댄다. 못 읽으면 옛 ai_claim 전용으로 물러난다.
    #    · 로스터 셋(Claude/CSOS/Codex)은 언제나 보이되, 살아 있는 세션이 채우지 못한
    #      역할만 'off' 로 남는다 — ai 구역은 절대 비지 않는다([169]).
    BASE = [
        {"name": "Claude 세션", "role": "앱 코딩·검증·인계", "color": "#0062CC",
         "duties": ["앱 코딩·리팩터링·합성검증", "세션 인계 문서 갱신",
                    "막힘 풀린 일 재개·보류 푸시"]},
        {"name": "CSOS 수집", "role": "밴드·ERP 자료 수집", "color": "#0E7490",
         "duties": ["밴드 글·댓글 수집(우선순위 계획대로)", "ERP 판매·수금 내려받기",
                    "원본 흡수·대조 회차 공급"]},
        {"name": "Codex", "role": "교대 코딩", "color": "#5A6785",
         "duties": ["Claude 크레딧 소진 시 코딩 교대", "같은 시작 체크리스트를 따름"]},
    ]
    WHO_LABEL = {"claude": "Claude 세션", "codex": "Codex"}
    WHO_COLOR = {"claude": "#0062CC", "codex": "#5A6785"}

    # 1) 점유판(ai_claim) — 세션별로 묶는다. why·자원·나이·생사.
    claims_ok = False
    by_sid = {}
    try:
        import ai_claim
        for res, c in (ai_claim.load() or {}).items():
            if not isinstance(c, dict):
                continue
            sid = c.get("sid") or c.get("who") or res
            g = by_sid.setdefault(sid, {"who": c.get("who") or "?", "res": [],
                                        "why": "", "at": 0.0, "dead": True})
            g["res"].append(ai_claim.LOCKS.get(res, (res, False))[0])
            if c.get("why"):
                g["why"] = c.get("why")
            g["at"] = max(g["at"], float(c.get("at") or 0))
            if not ai_claim._is_dead(c):
                g["dead"] = False
        claims_ok = True
    except Exception:
        claims_ok = False

    # 2) 살아 있는 세션(대화기록) — session_wrapup 을 그대로 빌린다(자기 자신 포함).
    #    live_transcripts() 가 sid 앞 8자리를 주고, transcript_dir() 로 마지막 활동
    #    시각(mtime)을 잰다. 못 읽으면 아래에서 ai_claim 전용으로 물러난다([169]).
    live_src_ok = False
    live_ids = []
    tmtime = {}
    try:
        import session_wrapup as _sw
        live_ids = list(dict.fromkeys(_sw.live_transcripts()))
        live_src_ok = True
        tdir = _sw.transcript_dir()
        if tdir:
            for nm in os.listdir(tdir):
                if nm.endswith(".jsonl") and nm[:-6][:8] in live_ids:
                    try:
                        tmtime[nm[:-6][:8]] = os.path.getmtime(
                            os.path.join(tdir, nm))
                    except OSError:
                        pass
    except Exception:
        live_src_ok = False
        live_ids = []

    # 3) 세션 제목(있으면) — ~/.claude/sessions/<pid>.json 의 사용자 지정 name.
    #    사람이 세션에 붙인 뜻있는 이름만 쓴다(자동 파생명 'derived' 는 뺀다).
    titles = {}
    try:
        import json as _json
        sdir = os.path.join(os.path.expanduser("~"), ".claude", "sessions")
        for nm in os.listdir(sdir):
            if not nm.endswith(".json"):
                continue
            try:
                with open(os.path.join(sdir, nm), encoding="utf-8") as fh:
                    m = _json.load(fh)
            except Exception:
                continue
            sid = str(m.get("sessionId") or "")
            nom = str(m.get("name") or "").strip()
            if sid and nom and m.get("nameSource") != "derived":
                titles[sid[:8]] = nom
    except Exception:
        titles = {}

    # 살아 있는 세션 원천(대화기록)이나 점유판 중 하나라도 읽혔나 — 그게 _live 다.
    live_ok = bool(live_src_ok or claims_ok)

    # ── 살아 있는 세션을 한 자리씩(sid 앞 8자리로 중복 제거). 대화기록에 안 잡혀도
    #    점유가 죽지 않았으면(ai_claim 전용 폴백) 함께 싣는다. ──
    ai = []
    roles_filled = set()
    order = list(live_ids)
    for sid, g in by_sid.items():
        if not g["dead"] and sid[:8] not in order:
            order.append(sid[:8])

    seen = set()
    for sid8 in order:
        if sid8 in seen:
            continue
        seen.add(sid8)
        g = next((gg for s, gg in by_sid.items() if s[:8] == sid8), None)
        who = str(g["who"]) if g and g["who"] in WHO_LABEL else "claude"
        who_label = WHO_LABEL.get(who, "CSOS 수집")
        roles_filled.add(who_label)
        name = titles.get(sid8) or ("%s ·%s" % (who_label, sid8[:4]))
        if g and g["res"]:
            role = " · ".join(dict.fromkeys(g["res"]))
            msg = g["why"] or "작업 중"
            duties = [g["why"] or "작업 중", "점유: " + role,
                      "세션 점유(ai_claim)에서 실시간으로 읽었습니다"]
        else:
            role = "작업 세션 · 대기"
            msg = "열려 있는 세션 · 지금 잡은 자원 없음"
            duties = ["열려 있는 에이전트 세션(대화기록에서 실시간)",
                      "지금 잡은 배타 자원 없음"]
        mt = tmtime.get(sid8) or (g["at"] if g else 0)
        ai.append({
            "name": name,
            "role": role,
            "color": WHO_COLOR.get(who, "#0E7490"),
            "state": "live",
            "msg": msg,
            "ago": _ago(mt),
            "badge": "세션",
            "duties": duties,
        })

    # 로스터 셋은 언제나 보인다 — 살아 있는 세션이 채우지 못한 역할만 'off'.
    for b in BASE:
        if b["name"] in roles_filled:
            continue
        ai.append(dict(b, state="off", msg="연결되면 실시간으로 바뀝니다",
                       ago="—", duties=list(b["duties"])))

    # ★ **회차도 조수다** (2026-08-13 지시: "조수 스테이션에 지금 돌아가는 에이전트나
    #   세션 모두 추가해서 구현해"). 전까지 이 자리는 **대화 세션만** 셌다. 그런데 이
    #   프로젝트에서 실제로 일을 하는 것은 대부분 무인 회차다 — 워치독 30분 ·
    #   증분 파이프라인 5분 · 09:50 일일대조 · 08:00 밴드재수집 · 12시 정오회차.
    #   그것들이 어느 화면에도 안 보이니 "지금 무엇이 돌고 있나"를 사람이 물어야 했다.
    #   ★ **판정을 새로 만들지 않는다**([162]) — `schedule_watch` 회차가 써 둔 갈래를
    #     읽기만 한다. 여기서 스케줄러를 다시 물으면 요청마다 비싼 조회가 된다([168]).
    try:
        sw = json.load(open(os.path.join(ROOT, "reports", "스케줄러_회차감시.json"),
                            encoding="utf-8"))
    except Exception:
        sw = None
    if sw is None:
        # ★ **못 읽은 것을 '도는 것이 없다'로 치지 않는다**([169]).
        ai.append({"name": "회차(스케줄러)", "role": "상태 못 읽음",
                   "color": "#8E8E93", "state": "off", "badge": "회차",
                   "msg": "스케줄러 감시 결과가 아직 없다 — 없는 것이 아니라 못 본 것이다",
                   "ago": "—",
                   "duties": ["`python schedule_watch.py` 가 30분마다 써 둔다",
                              "그 파일이 없으면 여기서 도는 회차를 셀 수 없다"]})
    else:
        # 갈래는 회차가 정한다 — 낱말을 여기 적으면 사본이 둘 된다.
        살아있음 = {"도는중", "성공"}
        나쁨 = {"죽음", "안돎", "확인못함", "밀림"}
        for t in (sw.get("작업") or []):
            갈래 = str(t.get("갈래") or "")
            if 갈래 == "꺼짐":
                continue           # 사람이 끈 것은 고장이 아니다 — 자리를 안 차지한다
            돎 = 갈래 == "도는중"
            ai.append({
                "name": str(t.get("작업") or "회차").replace("쿠팡업무_", ""),
                "role": "회차 · %s" % (갈래 or "모름"),
                "color": "#0E7490" if 돎 else ("#FF453A" if 갈래 in 나쁨 else "#5E5CE6"),
                "state": "live" if 갈래 in 살아있음 else "off",
                "badge": "회차",
                "msg": str(t.get("말") or ""),
                "ago": _ago_iso(t.get("마지막실행")),
                "duties": [str(t.get("말") or ""),
                           "마지막 실행: %s" % (t.get("마지막실행") or "기록 없음"),
                           "다음 예정: %s" % (t.get("다음예정") or "모름"),
                           "판정은 `schedule_watch` 가 30분마다 써 둔 것입니다"],
            })
    # ai 는 절대 비지 않는다 — 로스터 셋이 늘 뒤에 붙는다. 원천을 못 읽었으면
    # gen 에 그 사실을 적는다(0건을 '없다'로 못 박지 않는다, [169]).

    gen = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")
    return {
        "_live": live_ok,
        "gen": gen + (" · 실시간" if live_ok else " · 점유 못 읽음"),
        "flow": _org_flow(),
        "zones": [
            {"key": "mgmt", "title": "관리팀 · 업무센터",
             "tag": "쿠팡 AS · 정기점검 원장", "people": mgmt},
            {"key": "field", "title": "현장 AS팀 · 정기점검",
             "tag": "현장 조치 · 밴드 보고", "people": field},
            {"key": "ai", "title": "조수 스테이션",
             "tag": "자동화 · 수집 · 코딩", "people": ai},
        ],
    }


def get_checks():
    """최근 카톡·밴드·ERP원장·쿠팡PO 대조 CSV를 ID별로 조인 — 4원천 검증 배지"""
    import csv as _csv
    out = {}
    def latest(pat):
        fs = sorted(glob.glob(os.path.join(ROOT, "reports", pat)))
        return fs[-1] if fs else None
    f = latest("카톡대조_*.csv")
    if f:
        for r in _csv.DictReader(open(f, encoding="utf-8-sig")):
            out.setdefault(r.get("ID", ""), {})["kakao"] = r.get("카톡보고", "")
    f = latest("밴드대조_*.csv")
    if f:
        for r in _csv.DictReader(open(f, encoding="utf-8-sig")):
            out.setdefault(r.get("ID", ""), {})["band"] = r.get("밴드게시", "")
    f = latest("ERP원장대조_*.csv")
    if f:
        for r in _csv.DictReader(open(f, encoding="utf-8-sig")):
            sid = r.get("정산ID", "") or r.get("전표", "")
            for one in str(sid).split(","):
                if one.strip():
                    out.setdefault(one.strip(), {})["erp"] = r.get("유형", "") + " " + r.get("판정", "")
    f = latest("PO대조_*.csv")
    if f:
        for r in _csv.DictReader(open(f, encoding="utf-8-sig")):
            sid = r.get("정산ID", "") or r.get("ID", "")
            for one in str(sid).split(","):
                if one.strip():
                    out.setdefault(one.strip(), {})["po"] = (r.get("판정", "") or r.get("유형", "")).strip()
    if DEMO and not out:
        out = {"JS-2607-002": {"kakao": "확인", "band": "미확인", "erp": "D 금액불일치",
                               "po": "PO 미발행"}}
    return {
        k: {a: _OLD_APP_REF_RE.sub("", str(b or "")) for a, b in v.items()}
        for k, v in out.items() if app_year_record({"ID": k})
    }


_UJ_RE = re.compile(r"(?<![A-Za-z0-9])UJ\d{6,}(?![0-9])")
# 계산서 제목에서 캠프명만 뽑는다: '쿠팡신규_송파1MB(감일동)-이동식…' → '송파1MB(감일동)'
_CAMP_RE = re.compile(r"[가-힣A-Za-z]+\d*(?:BMB|MB|캠프|Sub-?FC|Sub-?hub|FC)(?:\([^)]*\))?",
                      re.I)   # sub-hub / Sub-Hub 표기가 섞여 있어 대소문자 무시


def camp_of(title):
    m = _CAMP_RE.search(str(title or ""))
    if m:
        return m.group()
    t = re.sub(r"^(쿠팡\S*|돌발AS|정기점검)[_\s-]*", "", str(title or "")).strip()
    return (t or str(title or ""))[:28]


def erp_settlement_rows(ledger_rows):
    """관리대장에 정산 행이 **아예 없는 달**만 ERP 계산서로 채워 넣는다.

    왜 이렇게 하나
      · 06시트는 '작업 1건 = 1행' 구조다(업무구분이 원천업무ID 기반 수식).
        ERP 계산서는 여러 작업을 묶은 것이라 그 시트에 그대로 넣으면 수식이 어긋난다.
      · 그렇다고 1~6월을 비워두면 앱에서 그 달 매출이 0으로 보인다(사실과 다름).
      → 대장에 자료가 있는 달은 대장 우선, 없는 달만 ERP로 보완하고 출처를 표시한다.
        (같은 달을 양쪽에서 세지 않으므로 이중 계상이 없다)
    """
    have = {str(r.get("완료일") or "")[:7].replace("-", "/") for r in ledger_rows
            if r.get("공급가액")}
    docs = get_erpdocs()
    out = []
    for d in docs.get("rows", []):
        mo = d.get("월") or ""
        if not mo or mo in have:
            continue
        slip = d.get("전표") or ""
        iso = slip[:10].replace("/", "-")
        title = d.get("프로젝트명") or ""
        prj = (_UJ_RE.search(title) or [""])
        prj = prj.group() if hasattr(prj, "group") else ""

        sup = int(d.get("공급가액") or 0)
        out.append({
            "정산ID": "ERP-" + slip.replace("/", "").replace("-", "-"),
            "업무구분": d.get("유형") or "", "캠프명": camp_of(title),
            "프로젝트NO": prj, "원천업무ID": "",
            "공급가액": sup, "합계": sup + round(sup * 0.1),
            "명세서": "있음", "명세서번호": slip, "명세서발행일": iso,
            "계산서": "발행", "계산서발행일": iso, "승인번호": "",
            "청구일": "", "지급예정일": "", "입금일": "", "입금액": 0, "미수금": "",
            "비용구분": "유상", "PO필요": "", "PO번호": "", "PO발행일": "",
            "상태": "ERP 계산서(묶음)", "완료일": iso,
            "출처": "ERP", "적요": title})
    return out


# ── 대표 프로젝트NO ────────────────────────────────────────────
# 모든 건이 번호로 식별되게 한다. 우선순위:
#   1) 행에 이미 있는 프로젝트NO
#   2) 행 안 어딘가(내용·근거 등)에 적힌 UJ 번호
#   3) 같은 캠프·같은 달의 실제 작업에서 찾은 대표 번호
#   4) 그래도 없으면 ERP 전표번호 기반 식별자(UJ처럼 보이지 않게 'ERP-' 접두)
#      — 없는 UJ 번호를 지어내면 실제 번호와 헷갈리므로 절대 만들지 않는다.
def _camp_key(v):
    return re.sub(r"[\s()·]", "", str(v or "")).lower()[:14]


def build_prj_index(works):
    idx = {}
    for kind, dk in (("as", "접수일자"), ("pm", "점검예정일")):
        for r in works.get(kind, []):
            if r.get("출처") == "ERP":
                continue
            prj = str(r.get("프로젝트NO") or "").strip()
            if not prj:
                continue
            mo = norm_date(r.get(dk) or r.get("작업완료일") or r.get("실제점검일"))[:7]
            idx.setdefault((_camp_key(r.get("캠프명")), mo), []).append(prj)
    return idx


def rep_no(rec, idx=None, slip=""):
    """대표 프로젝트NO를 정한다(순수 함수 — 합성검증 대상)"""
    cur = str(rec.get("프로젝트NO") or "").strip()
    if cur:
        return cur, ""
    for v in rec.values():                       # 내용·근거 등 본문에 적힌 UJ
        m = _UJ_RE.search(str(v or ""))
        if m:
            return m.group(), "본문"
    if idx:
        mo = ""
        for k in ("완료일", "접수일자", "점검예정일", "일자", "작업완료일"):
            mo = norm_date(rec.get(k))[:7]
            if mo:
                break
        hits = idx.get((_camp_key(rec.get("캠프명")), mo))
        if hits:
            return sorted(hits)[0], "동일캠프·동월"
    if slip:
        m = re.match(r"(\d{2})(\d{2})/(\d{2})/(\d{2})\s*-\s*(\d+)", str(slip))
        if m:
            return f"ERP-{m.group(2)}{m.group(3)}{m.group(4)}-{m.group(5)}", "전표"
        return "ERP-" + re.sub(r"[^0-9A-Za-z-]", "", str(slip))[-10:], "전표"
    # 최후: 그 행 자신의 ID를 대표번호로 쓴다 — 번호 없는 행이 하나도 남지 않게
    for k in ("정산ID", "접수ID", "점검ID", "업무ID", "ID"):
        v = str(rec.get(k) or "").strip()
        if v:
            return v, "자체ID"
    return "", ""


def apply_rep_no(rows, idx=None, slipkey=None):
    for r in rows:
        no, how = rep_no(r, idx, str(r.get(slipkey) or "") if slipkey else "")
        if no and not str(r.get("프로젝트NO") or "").strip():
            r["프로젝트NO"] = no
            r["대표번호출처"] = how
    return rows


def erp_work_rows(existing, kind):
    """02/04 시트에 자료가 **아예 없는 달**만 ERP 계산서로 보완한다(정산과 같은 규칙).
    ERP 계산서 1장 = 작업 여러 건 묶음이므로 '건수'가 아니라 '그 달에 이런 업무가 있었다'는
    사실을 보여주는 용도다. 자료가 있는 달은 건드리지 않아 이중 계상이 없다."""
    dk = {"as": ("접수일자", "작업완료일"), "pm": ("점검예정일", "실제점검일")}[kind]
    want = {"as": "돌발AS", "pm": "정기점검"}[kind]
    have = {norm_date(r.get(dk[0]) or r.get(dk[1]))[:7] for r in existing}
    have = {h.replace("-", "/") for h in have if h}
    out = []
    for d in get_erpdocs().get("rows", []):
        if (d.get("유형") or "") != want:
            continue
        mo = d.get("월") or ""
        if not mo or mo in have:
            continue
        slip = d.get("전표") or ""
        iso = slip[:10].replace("/", "-")
        title = d.get("프로젝트명") or ""
        prj = _UJ_RE.search(title)
        base = {"프로젝트NO": prj.group() if prj else "", "캠프명": camp_of(title),
                "담당기사": "", "유상·무상·보험": "유상", "비고": "ERP 계산서 기준(작업 묶음)",
                "출처": "ERP", "적요": title}
        if kind == "as":
            base.update({"접수ID": "ERP-" + slip.replace("/", ""), "접수일자": iso,
                         "작업완료일": iso, "진행상태": "작업완료", "신청내용": title,
                         "긴급도": "", "방문예정일": ""})
        else:
            base.update({"점검ID": "ERP-" + slip.replace("/", ""), "점검예정일": iso,
                         "실제점검일": iso, "점검상태": "완료",
                         "이상발견여부": "", "돌발AS전환여부": ""})
        out.append(base)
    return out


def _build_settlements():
    rows = real_settlements()
    try:
        rows = rows + erp_settlement_rows(rows)
        idx = build_prj_index(get_works())
        apply_rep_no(rows, idx, "명세서번호")
        rows = app_year_rows(rows, "settle")
        rows = sort_by_date(rows, "settle", "정산ID")
    except Exception:
        pass
    return rows


def get_settlements():
    if DEMO:
        return demo_settlements()
    return cached_data("settle", _build_settlements)


def get_live_state(*, store=None, state_path=None, lock_path=None):
    """Return a cheap, server-authoritative revision for every connected device.

    This endpoint deliberately does not read Excel, Z:, Band, ERP, or any of the
    expensive dashboard aggregates.  Phones and PCs poll it and only revalidate
    their current data when the stable revision changes.  ``generated_at`` is a
    clock value for display and is *not* part of the revision, so two devices do
    not refresh each other forever merely because their polling seconds differ.
    """

    now = datetime.now().astimezone().isoformat(timespec="seconds")
    try:
        from app_store import default_store

        app = (store or default_store()).status()
        state_file = os.fspath(
            state_path or os.path.join(ROOT, "reports", "automation_pipeline_state.json")
        )
        try:
            with open(state_file, encoding="utf-8") as handle:
                pipeline = json.load(handle)
            state_mtime_ns = int(os.stat(state_file).st_mtime_ns)
        except (OSError, ValueError, TypeError):
            pipeline, state_mtime_ns = {}, 0

        def latest_timestamp(*values):
            candidates = [str(value) for value in values if str(value or "").strip()]
            if not candidates:
                return ""

            def sort_key(value):
                try:
                    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
                    return (parsed.timestamp(), value)
                except (TypeError, ValueError, OSError):
                    return (float("-inf"), value)

            return max(candidates, key=sort_key)

        last_run = pipeline.get("last_run") or {}
        # state JSON은 프로세스가 비정상 종료되면 ``running:true``인 채 남는다.
        # 실제 잠금의 pid+생성시각 지문까지 맞을 때만 실행 중으로 인정한다. 임시
        # state_path를 쓰는 합성검증은 같은 폴더의 잠금만 보므로 운영 잠금을 읽지 않는다.
        try:
            from automation_pipeline import pipeline_lock_status
            resolved_lock_path = (
                os.fspath(lock_path)
                if lock_path is not None
                else os.path.join(os.path.dirname(state_file), ".automation_pipeline.lock")
            )
            lock_state = pipeline_lock_status(resolved_lock_path)
            lock_alive = lock_state.get("alive") is True
            lock_run_id = str(lock_state.get("run_id") or "")
            state_run_id = str(last_run.get("run_id") or pipeline.get("active_run_id") or "")
            pipeline_running = bool(
                pipeline.get("running") and lock_alive
                and (not state_run_id or not lock_run_id or state_run_id == lock_run_id)
            )
        except Exception:
            # 판정 자체를 못 하면 state를 거짓 정상으로 만들지 않고 기존 상태를 보수적으로
            # 유지한다. 정상 경로에서는 공용 판정기가 항상 사용된다.
            lock_state = {"exists": None, "alive": None}
            pipeline_running = bool(pipeline.get("running"))
        history = pipeline.get("history") or []
        completed_runs = [
            str(run.get("finished_at") or "")
            for run in ([last_run] + (history if isinstance(history, list) else []))
            if isinstance(run, dict) and run.get("finished_at")
        ]
        last_completed_at = latest_timestamp(*completed_runs)
        last_export = app.get("last_export") or {}
        last_completed_export = app.get("last_completed_export") or {}
        if not last_completed_export and str(last_export.get("status") or "") in {
            "verified",
            "partial",
        }:
            last_completed_export = last_export
        export_completed_at = str(last_completed_export.get("finished_at") or "")
        app_changed_at = str(app.get("last_change_at") or "")
        state_file_updated_at = (
            datetime.fromtimestamp(state_mtime_ns / 1_000_000_000)
            .astimezone()
            .isoformat(timespec="seconds")
            if state_mtime_ns
            else ""
        )
        stable = {
            "change_seq": int(app.get("change_seq") or 0),
            "outbox_pending": int(app.get("outbox_pending") or 0),
            # 실패/실행 중 export의 snapshot 번호를 최신 보관본처럼 내보내면
            # 모든 기기가 보관 완료로 오해한다. 검증이 끝난 마지막 export만 표시한다.
            "snapshot_seq": int(last_completed_export.get("snapshot_seq") or 0),
            "export_status": str(last_completed_export.get("status") or ""),
            "pipeline_running": pipeline_running,
            "pipeline_lock_alive": lock_state.get("alive"),
            "pipeline_lock_run_id": str(lock_state.get("run_id") or ""),
            "pipeline_updated_at": str(last_run.get("updated_at") or ""),
            "pipeline_finished_at": str(last_run.get("finished_at") or ""),
            "last_completed_at": last_completed_at,
            "pipeline_stage": str(last_run.get("current_stage") or ""),
            "pipeline_state_mtime_ns": state_mtime_ns,
        }
        # 업무 자료를 다시 받는 기준과 관제 상태만 다시 그리는 기준을 나눈다.
        # 파이프라인은 단계마다 state 파일을 갱신한다. 그것을 자료 revision으로 쓰면
        # 기기마다 7개 대형 API를 단계 수만큼 다시 불러, 바로 이 화면을 또 막는다.
        data_stable = {
            "change_seq": stable["change_seq"],
            "last_completed_at": stable["last_completed_at"],
        }
        revision = hashlib.sha256(
            json.dumps(data_stable, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            .encode("utf-8")
        ).hexdigest()[:20]
        state_revision = hashlib.sha256(
            json.dumps(stable, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            .encode("utf-8")
        ).hexdigest()[:20]
        running = bool(stable["pipeline_running"])
        pending = int(stable["outbox_pending"])
        if running:
            phase, label = "updating", "자료 갱신 중"
        elif pending:
            phase, label = "archive_pending", f"Excel 보관 대기 {pending}건"
        else:
            phase, label = "current", "최신 자료"
        data_updated_at = latest_timestamp(
            app_changed_at,
            export_completed_at,
            last_completed_at,
        )
        state_updated_at = latest_timestamp(
            data_updated_at,
            stable["pipeline_updated_at"],
            stable["pipeline_finished_at"],
            state_file_updated_at,
        )
        return {
            "ok": True,
            "revision": revision,
            "state_revision": state_revision,
            "change_seq": stable["change_seq"],
            "snapshot_seq": stable["snapshot_seq"],
            "phase": phase,
            "label": label,
            "running": running,
            "current_stage": stable["pipeline_stage"],
            "outbox_pending": pending,
            "data_updated_at": data_updated_at,
            "state_updated_at": state_updated_at,
            "updated_at": data_updated_at,
            "server_time": now,
        }
    except Exception as exc:
        return {
            "ok": False,
            "revision": "",
            "phase": "error",
            "label": "상태 확인 지연",
            "running": False,
            "data_updated_at": "",
            "state_updated_at": "",
            "updated_at": "",
            "server_time": now,
            "error": str(exc)[:240],
        }


def get_status():
    """★ 이 함수는 Z: 네트워크 드라이브를 여러 번 읽는다(원장·ERP 내보내기·대조 CSV).
    로컬 단독으로는 4초면 끝나지만, Codex·일일실행이 같은 드라이브를 쓰는 동안에는
    280초~600초까지 늘어난다. 대시보드는 이걸 주기적으로 폴링하므로 캐시가 없으면
    **앱이 열리지 않는다**(2026-07-29 실측). 다른 데이터 API와 같은 캐시를 쓴다:
    원장이 바뀌면 즉시, 아니면 120초 TTL.

    ★ **캐시를 보는 데는 락이 필요 없다** (2026-08-07). 예전에는 이 첫 조회에서도
      `_readlock` 을 잡았다. 그런데 그 락은 다른 요청이 Z: 를 콜드로 읽는 동안
      계속 잡혀 있고(실측 `get_works` 첫 계산 111초 · 대표 예외보고 187초),
      그동안 여기서 **옛 값을 돌려주는 것조차 막혔다** — 기다리지 않게 하려고
      만든 stale-while-revalidate 인데 정작 그 길이 잠겨 있었다.
      락이 지키는 것은 'Z: 를 동시에 읽지 않는 것'이지 캐시 딕셔너리가 아니다
      (issues·erpdocs 는 이미 락 없이 `_fresh` 를 쓴다). 실제로 읽을 때만 잡는다
      — 그 자리는 `_refresh_status_now()` 다."""
    c = _fresh("status")
    if c:
        return c
    # ★ TTL 만료 시 옛 값을 **즉시** 돌려주고 재계산은 뒤에서 한 번만 한다(2026-08-03 UX).
    #   실측: /api/status 1,550회 호출에 평균 5.2초 — 만료 순간마다 Z: 재계산이 요청을
    #   통째로 잡고 있었다. 원장이 바뀌면 _fresh 가 stale 까지 비우므로 낡은 값이 남지 않는다.
    stale = _cache.get("status_stale")
    if stale:
        _spawn_status_refresh()
        return stale
    return _refresh_status_now()


_STATUS_REFRESH = {"busy": False}


def _refresh_status_now():
    with _readlock:
        c = _fresh("status")
        if c:
            return c
        c = _compute_status()
        if "error" not in c:
            _store_cache("status", c)
            _cache["status_stale"] = c
        return c


_STATUS_REFRESH_LOCK = threading.Lock()


def _spawn_status_refresh():
    # ★ '보고 나서 세우기'는 둘이 동시에 통과할 수 있다 — 폴링이 30초마다 오고
    #   재계산이 100초 넘게 걸리므로 실제로 겹친다. 그러면 무거운 Z: 재계산이
    #   두 벌 떠서 서로 락을 기다린다. 세우는 순간을 잠가 한 벌만 뜨게 한다.
    with _STATUS_REFRESH_LOCK:
        if _STATUS_REFRESH["busy"]:
            return
        _STATUS_REFRESH["busy"] = True

    def _run():
        try:
            _refresh_status_now()
        finally:
            _STATUS_REFRESH["busy"] = False

    threading.Thread(target=_run, daemon=True).start()


def _compute_status():
    if DEMO:
        return {"master": "쿠팡_통합업무_일일보고_관리대장_v23.xlsx (데모)", "fork": [],
                "agent_last": "2026-07-24 09:50", "steps": [
                    {"n": "합성검증", "s": "ok"}, {"n": "판매·세금계산서 대조", "s": "ok"},
                    {"n": "ERP원장 4유형 대조", "s": "ok"}, {"n": "밴드 수집·대조", "s": "skip"},
                    {"n": "카톡 대조", "s": "ok"}, {"n": "관리대장 자동입력", "s": "ok"},
                    {"n": "전표 전송대기", "s": "ok"}],
                "pending_updates": 2, "inbox": 1, "kakao": 2, "band": False, "demo": True}
    try:
        from coupang_workbench import get_status as ws
        st = ws()
        # 동기화 백본: 에이전트가 쓴 agent_status.json 우선 (없으면 md 리포트 파싱)
        steps, rt = [], ""
        agent_aborted, agent_age_h, agent_failed = False, None, []
        try:
            aj = json.load(open(os.path.join(ROOT, "reports", "agent_status.json"), encoding="utf-8"))
            steps = aj.get("steps", [])
            rt = aj.get("time", "")[:16].replace("T", " ")
            # ★ 이 시각이 **어제 것이면 어제 것이라고 말해야 한다**(2026-08-07 사용자 지적:
            #   "에이전트 실행 시간이 맞지 않아 지금 날짜로 반영이 안되었어").
            #   08-06 21:01 에서 멈춘 회차가 초록 점과 함께 떠 있어 정상으로 보였다.
            #   스케줄러는 앞 회차가 도는 동안 다음 회차를 조용히 건너뛰므로, 시각만으로는
            #   사람이 알아챌 수 없다 — 중단 여부와 몇 시간째인지를 같이 내려보낸다.
            agent_aborted = bool(aj.get("aborted"))
            try:
                _mt = os.path.getmtime(os.path.join(ROOT, "reports", "agent_status.json"))
                agent_age_h = round((datetime.now().timestamp() - _mt) / 3600.0, 1)
            except OSError:
                agent_age_h = None
            agent_failed = [s.get("name") or s.get("n") for s in steps if not (s.get("ok") or s.get("s") == "ok")]
        except Exception:
            for s in st.get("report_summary", []):
                mark = "ok" if "✅" in s else ("skip" if "스킵" in s else "fail")
                steps.append({"n": re.sub(r"[✅❌⏭]|스킵|실패", "", s).strip(), "s": mark})
            rt = st.get("report_time", "")
            if rt:
                rt = f"{rt[:4]}-{rt[4:6]}-{rt[6:8]} {rt[9:11]}:{rt[11:13]}"
        tunnel = ""
        try:
            tunnel = open(os.path.join(ROOT, "reports", "tunnel_url.txt"), encoding="utf-8").read().strip()
        except Exception:
            pass
        # 3원천 검증 실집계 (최신 대조 CSV) — 보고서·대시보드가 하드코딩 없이 실데이터를 쓰도록
        srcs = {}
        try:
            from findings_export import REPORT_DIR as _report_dir
            from findings_export import latest_csv

            def _src_meta(pat):
                """대조 CSV 파일명(…_YYYYMMDD_HHMM)에서 기간·대조시각을 뽑는다(2026-08-03 지시).

                대상 기간은 앱 연도 1/1부터 대조 실행일까지 — 4원천 숫자가 '언제 자료 기준'
                인지 보고서만 봐도 알 수 있게 한다."""
                fs = sorted(glob.glob(os.path.join(_report_dir, pat)))
                m = re.search(r"(\d{8})_(\d{4})", os.path.basename(fs[-1])) if fs else None
                if not m:
                    return {}
                day, tm = m.groups()
                iso = f"{day[:4]}-{day[4:6]}-{day[6:]}"
                return {"period": f"{APP_YEAR}-01-01 ~ {iso}",
                        "asof": f"{iso} {tm[:2]}:{tm[2:]} 대조"}

            for key, pat, col, okv in (("band", "밴드대조_*.csv", "밴드게시", "확인"),
                                       ("kakao", "카톡대조_*.csv", "카톡보고", "확인")):
                rows = app_year_rows(latest_csv(pat), "issue")
                if rows:
                    ok = sum(1 for r in rows if r.get(col) == okv)
                    # ★ '자료없음'은 실패가 아니다 (2026-08-07 지시).
                    #   카톡 내보내기는 방에 들어간 뒤부터만 나온다. 그 전 작업은
                    #   아무리 잘해도 카톡에서 못 찾는다 — 그걸 '미확인'과 같이 세면
                    #   지워지지 않는 빨간 줄이 쌓이고, 진짜 누락(자료는 있는데 보고가
                    #   없는 건)이 그 속에 묻힌다. 분모에서 빼고 따로 보여 준다.
                    na = sum(1 for r in rows if r.get(col) == "자료없음")
                    miss = [r for r in rows if r.get(col) not in (okv, "자료없음")]
                    srcs[key] = {"total": len(rows) - na, "ok": ok, "miss": len(miss),
                                 "na": na,
                                 "miss_prj": [r.get("프로젝트NO") or r.get("ID") for r in miss[:8]],
                                 **_src_meta(pat)}
            erp = app_year_rows(latest_csv("ERP원장대조_*.csv"), "issue")
            if erp:
                srcs["erp"] = {"total": len(erp), "ok": 0, "miss": len(erp),
                               "miss_prj": [r.get("정산ID") or r.get("전표") for r in erp[:8]],
                               **_src_meta("ERP원장대조_*.csv")}
            po = app_year_rows(latest_csv("PO대조_*.csv"), "issue")
            if po:
                srcs["po"] = {"total": len(po), "ok": 0, "miss": len(po),
                              "miss_prj": [r.get("PO번호") or r.get("정산ID") for r in po[:8]],
                              **_src_meta("PO대조_*.csv")}
        except Exception:
            pass

        # ★ 대조 리포트만 보면 '자료가 있는데 왜 없다고 하냐'는 말이 나온다 — 리포트가
        #   없는 이유가 (1) 파일을 안 넣었다 (2) 넣었는데 **파일이 비어 있다**
        #   (3) 대조를 안 돌렸다 로 갈리기 때문이다. 2026-07-27에 실제로 ERP 파일 3개가
        #   회사명 한 줄만 있는 빈 파일이었고, 아무도 그걸 몰랐다. 그래서 여기서 같이 본다.
        try:
            from inbox_scan import pick
            import openpyxl as _ox
            for key, kinds in (("erp", ("ledger", "slips")), ("po", ("po",)),
                               ("tax", ("tax",)), ("stmt", ("stmt",))):
                files = []
                for kd in kinds:
                    files += pick(kd) or []
                files = [f for f in files
                         if not re.search(r"(?<!\d)2025(?!\d)|(?<!\d)25[/._-]\d{2}", os.path.basename(f))]
                info = {"files": len(files), "rows": 0, "empty": []}
                for f in files[:6]:
                    try:
                        w = _ox.load_workbook(f, read_only=True, data_only=True)
                        n = 0
                        for sn in w.sheetnames:
                            n += sum(1 for r in w[sn].iter_rows(values_only=True)
                                     if sum(1 for x in r if x not in (None, "")) >= 3)
                        w.close()
                        info["rows"] += n
                        if n < 2:                      # 회사명 한 줄만 있는 '빈 내보내기'
                            info["empty"].append(os.path.basename(f))
                    except Exception:
                        pass
                if files:
                    srcs.setdefault(key, {})["inbox"] = info
        except Exception:
            pass

        try:
            from agent_dispatch import status as agent_dispatch_status
            agent_route = agent_dispatch_status()
        except Exception:
            agent_route = {}
        try:
            import autopilot as _autopilot
            autopilot_status = _autopilot.status()
        except Exception:
            autopilot_status = {}
        return {"master": os.path.basename(st.get("master", "") or "") + "  " + st.get("master_label", ""),
                "fork": st.get("fork", []), "agent_last": rt or "기록 없음", "steps": steps,
                "agent_aborted": agent_aborted, "agent_age_h": agent_age_h,
                "agent_failed": [f for f in agent_failed if f][:4],
                # 20시간이면 오늘 09:50 회차가 통째로 빠졌다는 뜻이다(일일대조는 하루 한 번).
                "agent_stale": bool(agent_aborted or (agent_age_h is not None and agent_age_h >= 20)),
                "pending_updates": st["pending_updates"], "inbox": st["inbox"],
                "kakao": st["kakao"], "band": st["band_auth"], "demo": False, "tunnel": tunnel,
                "sources": srcs, "build": build_id(), "recalc": get_recalc_pending(),
                "applywin": get_apply_window(),
                "agent_dispatch": agent_route, "autopilot": autopilot_status}
    except Exception as e:
        return {"error": str(e)}


def latest_reports():
    out = []
    old_year = _OLD_APP_REF_RE
    # '자료현황' 을 맨 앞에 둔다 — "그거 지금 몇 건이지?" 를 매번 다시 세지 않으려고 만든 장이다
    # (사용자 지시 2026-07-29). data_status.py 가 만들고 daily_run 이 매일 갱신한다.
    # 정산분 보고자료를 맨 앞에 둔다 — 다음 날 아침 대표 보고에 그대로 쓰는 장이라
    # 앱을 열자마자 보여야 한다(2026-08-05 지시). settle_report.py 가 매일 만든다.
    for pat, name in [("보고자료_*정산분.md", "정산분 보고"),
                      ("자료현황.md", "자료현황"),
                      ("미비점_*.md", "미비점"),
                      ("견적명세_불일치.md", "견적·명세 불일치"),
                      ("세금계산서_미발행_경과.md", "계산서 미발행 경과"),
                      ("담당자_객관완료.md", "담당자 객관완료"),
                      ("종합리포트_*.md", "종합"), ("카톡대조_*.md", "카톡"), ("밴드대조_*.md", "밴드"),
                      ("ERP원장대조_*.md", "ERP원장"), ("이카운트대조_*.md", "판매·계산서")]:
        fs = sorted(glob.glob(os.path.join(ROOT, "reports", pat)))
        if fs:
            text = open(fs[-1], encoding="utf-8").read()[:20000]
            text = "\n".join(line for line in text.splitlines() if not old_year.search(line))
            out.append({"kind": name, "file": os.path.basename(fs[-1]),
                        "text": text})
    if DEMO and not out:
        out = [{"kind": "종합", "file": "demo.md",
                "text": "# 데모 리포트\n\n| 단계 | 결과 |\n|---|---|\n| 합성검증 | ✅ |\n| 카톡 대조 | ✅ |\n\n## 문제 예시\n- JS-2607-002 금액불일치(원장 620,000 / EC 500,000)"}]
    return out


WORKCENTER_ACTIVITY = os.path.join(ROOT, "reports", "workcenter_activity.json")
IMPROVEMENT_QUEUE = os.path.join(ROOT, "reports", "workcenter_improvements.jsonl")


def _atomic_json(path, value):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + f".{os.getpid()}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(value, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, path)


def record_workcenter_activity(slug, event="view"):
    """담당자 입력 중임을 AI 조정 파일에 남긴다.

    ai_claim.py는 최근 heartbeat가 있으면 새로운 code/ledger 점유를 잠시 미뤄,
    화면 입력·첨부가 진행되는 동안 배포 재시작이나 원장 교체가 끼어들지 않게 한다.
    """
    slug = str(slug or "").strip()
    if slug not in STAFF_CENTERS:
        raise ValueError("등록되지 않은 업무센터입니다")
    now = datetime.now()
    event = str(event or "view").strip().lower()[:40]
    editing_events = {"input", "change", "paste", "drop", "save", "upload", "submit"}
    state = {
        "slug": slug, "name": STAFF_CENTERS[slug]["name"],
        "event": event,
        "updated_at": now.isoformat(timespec="seconds"),
        "updated_ts": time.time(),
        "active_until_ts": time.time() + 120 if event in editing_events else 0,
        "editing": event in editing_events,
    }
    _atomic_json(WORKCENTER_ACTIVITY, state)
    return state


def install_staff_shortcut(slug):
    """PWA 프롬프트를 제공하지 않는 Windows 브라우저의 안전한 설치 대체 경로.

    담당자별 고정 HTTPS 주소를 Chrome/Edge 앱 창으로 여는 바탕화면 바로가기를 만든다.
    URL과 아이콘은 프로젝트에서 관리하는 고정값만 사용하며 요청값을 셸 문자열로
    조합하지 않는다.
    """
    slug = str(slug or "").strip()
    center = STAFF_CENTERS.get(slug)
    if not center:
        raise ValueError("등록되지 않은 업무센터입니다")
    if os.name != "nt":
        raise RuntimeError("PC 바로가기 설치는 Windows에서만 지원합니다")
    script = os.path.join(BASE, "install_staff_shortcut.ps1")
    icon = os.path.join(BASE, "csos-app.ico")
    if not os.path.isfile(script) or not os.path.isfile(icon):
        raise RuntimeError("설치 구성 파일이 준비되지 않았습니다")
    url = f"{FIXED_LIVE_ENTRY}/staff/{slug}"
    from proc_guard import run_tree
    result = run_tree(
        ["powershell", "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass",
         "-File", script, "-Url", url, "-Name", center["title"], "-Icon", icon],
        cwd=BASE, env=ENV, timeout=20, drain_timeout=10, output_limit=20_000,
    )
    if result.timed_out:
        raise RuntimeError("바로가기 생성 시간초과 · 자식나무 정리"
                           f" · 잔류 pid={result.stuck_pid or 0}")
    if result.returncode:
        raise RuntimeError((result.stderr or result.stdout or "바로가기 생성 실패").strip()[:240])
    return {"installed": True, "title": center["title"], "url": url}


# ── 업로드 알림 (2026-08-14 지시: "업로드 알림 만들어") ──────────────────────
#   ★ 알림을 **만드는 자리는 notify.push 하나다**([162]). 여기서는 누구에게 보낼지만
#     정하고 문구·채널·합치기·기록은 전부 그쪽이 한다.
#   ★ 알림 실패가 업로드를 죽이지 않는다 — notify.push 는 예외를 밖으로 내지 않지만,
#     import 자체가 실패할 수도 있으므로 여기서도 한 겹 감싼다.

def _upload_audience(actor):
    """형님(관리자) + **올린 사람**. 올린 사람이 모르면 또 올려 중복 원본이 된다."""
    try:
        import notify
    except Exception:
        return []
    who = [notify.ADMIN]
    mine = notify.actor_token(actor.get("role"), actor.get("staff_slug"))
    if mine and mine not in who:
        who.append(mine)
    return who


def _uploader_label(actor):
    """누가 올렸나 — 화면에 쓰는 이름. 코드에 사람 이름을 박지 않는다."""
    slug = str(actor.get("staff_slug") or "")
    if slug in STAFF_CENTERS:
        return STAFF_CENTERS[slug]["name"]
    return "관리자" if actor.get("role") == "admin" else "알 수 없는 사용자"


def _first_filename(files):
    """올린 파일 이름 한 개(여럿이면 '외 N건'). 없으면 빈 문자열 — 지어내지 않는다."""
    names = []
    for value in (files or {}).values():
        if isinstance(value, dict) and value.get("filename"):
            names.append(os.path.basename(str(value["filename"])))
        elif isinstance(value, list):
            for one in value:
                if isinstance(one, dict) and one.get("filename"):
                    names.append(os.path.basename(str(one["filename"])))
    if not names:
        return ""
    return names[0] if len(names) == 1 else f"{names[0]} 외 {len(names) - 1}건"


def _notify_upload_received(kind, filename, actor, started=False, queued=False,
                            duplicate=False, task_key="", extra=""):
    try:
        import notify
        who = _upload_audience(actor)
        if not who:
            return
        name = os.path.basename(str(filename or "")) or "(파일명 없음)"
        if duplicate:
            상태 = "이미 들어온 내용"
        elif started:
            상태 = "정리 시작"
        elif queued:
            상태 = "앞 작업이 끝나면 정리 시작"
        else:
            상태 = "다음 자동 회차가 정리"
        notify.push(
            f"{kind} 업로드",
            f"{_uploader_label(actor)} 님이 {kind}을 올렸습니다 — {name}",
            (extra or "") or f"{상태}.",
            evidence="/api/automation/kakao-upload" if task_key == "automation" else "",
            audience=who, severity="info", 상태=상태)
        # ★ '올렸다'만 알리면 반쪽이다 — 처리가 죽어도 된 줄 안다([169]).
        #   결과는 회차가 남긴 자국을 읽어 뒤따라 알린다.
        if task_key == "automation" and not duplicate:
            notify.expect_upload(kind, name, who, source="automation")
    except Exception:
        pass                                     # 알리려다 접수를 막지 않는다


def _notify_upload_rejected(kind, filename, actor, why):
    try:
        import notify
        who = _upload_audience(actor)
        if not who:
            return
        name = os.path.basename(str(filename or "")) or "(파일명 없음)"
        notify.push(
            f"{kind} 업로드 거절",
            f"{_uploader_label(actor)} 님의 {kind} 등록이 거절됐습니다 — {name}",
            str(why or "")[:300], evidence="", audience=who,
            severity="warning", 상태="거절")
    except Exception:
        pass


def get_notifications(actor=None):
    """사람이 조치해야 하는 실패·미반영·감시 이슈만 한곳에 모은다."""
    items = []
    # 업로드 알림 — 접수해 둔 것의 결과를 먼저 확인한다. 비싼 탐색이 아니라
    # 회차가 이미 써 둔 로컬 파일만 읽는다([168]). 바뀐 것이 없으면 파일도 안 쓴다.
    try:
        import notify
        notify.sweep_uploads()
        role = str((actor or {}).get("role") or "admin")
        slug = str((actor or {}).get("staff_slug") or "")
        for row in notify.feed(role, slug):
            n = int(row.get("건수") or 1)
            detail = str(row.get("본문") or "")
            if n > 1:                            # 합쳤으면 몇 건인지 말한다([169])
                detail = (detail + " · " if detail else "") + f"최근 5분 {n}건"
            items.append({
                "id": f"notify:{row.get('id')}",
                "severity": row.get("심각도") or "info",
                "title": row.get("제목") or "새 알림",
                "detail": detail[:220],
            })
    except Exception as exc:
        items.append({
            "id": "notify_read_error", "severity": "error",
            "title": "업로드 알림 확인 실패", "detail": str(exc)[:180],
        })
    try:
        from ledger_writer import load_queue
        pending = len(load_queue())
        if pending:
            items.append({
                "id": "ledger_pending", "severity": "warning",
                "title": f"Excel 보관본 생성 대기 {pending}개",
                "detail": "업무센터 입력은 앱 DB에 저장됐으며 Excel 보관본 생성을 기다립니다.",
            })
    except Exception as exc:
        items.append({
            "id": "ledger_queue_error", "severity": "error",
            "title": "Excel 보관본 생성 대기열 확인 실패", "detail": str(exc)[:180],
        })
    try:
        issue_path = os.path.join(ROOT, "reports", "realtime_issues.json")
        data = json.load(open(issue_path, encoding="utf-8"))
        for issue in data.get("issues") or []:
            if issue.get("status") not in ("new", "ongoing"):
                continue
            sev = str(issue.get("severity") or "P2")
            if issue.get("status") == "new" or sev in ("P0", "P1"):
                items.append({
                    "id": str(issue.get("id") or "monitor"),
                    "severity": "error" if sev in ("P0", "P1") else "warning",
                    "title": str(issue.get("title") or "확인 필요"),
                    "detail": str(issue.get("action") or issue.get("evidence") or "")[:220],
                })
    except FileNotFoundError:
        pass
    except Exception as exc:
        items.append({
            "id": "monitor_read_error", "severity": "error",
            "title": "감시 상태 확인 실패", "detail": str(exc)[:180],
        })
    if runner.get("busy"):
        items.append({
            "id": "runner_busy", "severity": "info",
            "title": f"{runner.get('task') or '자동 작업'} 실행 중",
            "detail": "완료 후 대기 중인 앱 DB 변경을 보관본 생성 순서로 넘깁니다.",
        })
    return {"count": sum(1 for x in items if x["severity"] != "info"),
            "items": items, "checked_at": datetime.now().isoformat(timespec="seconds")}


def save_workcenter_improvement(fields, files, source_ip=""):
    slug = str(fields.get("staff_slug") or "").strip()
    if slug not in STAFF_CENTERS:
        raise ValueError("등록되지 않은 업무센터입니다")
    title = str(fields.get("title") or "").strip()
    description = str(fields.get("description") or "").strip()
    if not title or not description:
        raise ValueError("불편한 점의 제목과 설명을 입력해 주세요")
    attachment = files.get("attachment")
    saved_name = ""
    if attachment and attachment.get("data"):
        if len(attachment["data"]) > 25_000_000:
            raise ValueError("첨부파일은 25MB 이하만 가능합니다")
        ext = os.path.splitext(attachment.get("filename") or "")[1].lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp", ".pdf", ".txt"):
            raise ValueError("이미지·PDF·텍스트 파일만 첨부할 수 있습니다")
        from source_dirs import ORIGIN_ROOT
        now = datetime.now()
        folder = os.path.join(ORIGIN_ROOT, "51. 업무센터 개선요청",
                              f"{now:%Y}", f"{now:%m}", f"{now:%Y-%m-%d}")
        os.makedirs(folder, exist_ok=True)
        saved_name = (f"{STAFF_CENTERS[slug]['name']}_{now:%Y%m%d_%H%M%S}_"
                      f"{_safe_upload_name(attachment.get('filename'))}")
        with open(os.path.join(folder, saved_name), "wb") as out:
            out.write(attachment["data"])
    ticket = {
        "id": f"WC-{datetime.now():%Y%m%d%H%M%S}-{random.randint(100,999)}",
        "registered_at": datetime.now().isoformat(timespec="seconds"),
        "staff_slug": slug, "staff": STAFF_CENTERS[slug]["name"],
        "title": title, "description": description,
        "attachment": saved_name, "source_ip": source_ip,
        "status": "new", "route": "Claude Code 우선 · 사용 불가 시 Codex",
    }
    os.makedirs(os.path.dirname(IMPROVEMENT_QUEUE), exist_ok=True)
    with open(IMPROVEMENT_QUEUE, "a", encoding="utf-8") as out:
        out.write(json.dumps(ticket, ensure_ascii=False) + "\n")
    return ticket


def datalake_records(qs):
    """자료창고(datalake record) 조회 — 읽기 전용 ([24] 앱 화면 노출).

    ERP Excel·밴드 글이 흡수된 `record` 표를 앱에서 **보고 캡처**할 수 있게 한다.
    · 정본 질의는 하나다 — CLI 도 앱도 `datalake.find(on='record')` 를 부른다
      (두 벌로 만들면 결과가 갈린다, 설계서).
    · **읽기 전용**: 큐에도 안 넣고 엑셀도 안 연다. 비싼 Z: glob 도 없다(datalake.db 는 로컬).
    · 자유문 검색은 FTS5 가 있을 때만. 없으면 조용히 목록만 준다(있는 척하지 않는다).
    """
    def g(k, d=""):
        return (qs.get(k, [d])[0] or d).strip()
    kind = g("kind")
    since = g("since")
    until = g("until")
    q = g("q")
    party = g("party")
    try:
        limit = max(1, min(500, int(g("limit", "100"))))
    except ValueError:
        limit = 100
    try:
        import datalake as D
    except Exception as e:  # pragma: no cover
        return {"error": f"datalake 로드 실패: {e}", "rows": [], "kinds": [], "total": 0}
    con = None
    try:
        con = D.connect()
        total = con.execute("SELECT COUNT(*) c FROM record").fetchone()["c"]
        # kind 요약 — 사람이 어떤 갈래가 얼마나 있는지 먼저 본다
        kinds = [{"kind": r["kind"], "count": r["c"], "last": r["last"]}
                 for r in con.execute(
                     "SELECT kind, COUNT(*) c, MAX(biz_date) last FROM record"
                     " GROUP BY kind ORDER BY c DESC")]
        fts = D.has_fts(con)
        note = ""
        rows = []
        try:
            recs = D.find(con, on="record", kind=(kind or None), since=(since or None),
                          until=(until or None), q=(q or None), party=(party or None),
                          limit=limit)
        except RuntimeError as e:
            # FTS 없는 파이썬에서 자유문 검색을 요청한 경우 — 목록은 주되 이유를 밝힌다
            note = str(e)
            recs = D.find(con, on="record", kind=(kind or None), since=(since or None),
                          until=(until or None), party=(party or None), limit=limit)
        for r in recs:
            rows.append({"id": r["id"], "kind": r["kind"], "key": r["natural_key"],
                         "date": r["biz_date"], "party": r["party"],
                         "amount": r["amount"], "status": r["status"]})
        return {"rows": rows, "kinds": kinds, "total": total, "shown": len(rows),
                "fts": fts, "note": note,
                "filter": {"kind": kind, "since": since, "until": until,
                           "q": q, "party": party, "limit": limit}}
    except Exception as e:  # pragma: no cover
        return {"error": f"{e}", "rows": [], "kinds": [], "total": 0}
    finally:
        if con is not None:
            try: con.close()
            except Exception: pass


# ───────────────────────── HTTP ─────────────────────────
class H(BaseHTTPRequestHandler):
    def handle_one_request(self):
        # 어떤 예외도 소켓을 조용히 죽이지 않게(ERR_EMPTY_RESPONSE 방지) 전역 가드
        try:
            super().handle_one_request()
        except (ConnectionError, BrokenPipeError):
            pass
        except Exception as e:
            try:
                self._send(500, {"error": str(e)[:300]})
            except Exception:
                pass

    def _send(self, code, body, ctype="application/json; charset=utf-8", headers=None):
        data = body if isinstance(body, bytes) else json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Private-Network", "true")
        for key, value in (headers or {}).items():
            self.send_header(str(key), str(value))
        self.end_headers()
        self.wfile.write(data)

    def _auth(self):
        # 잠금 카운트는 /api/login 에서만 증가 — 구 PIN이 저장된 브라우저의
        # 자동 폴링이 잠금을 유발하던 문제(자기 잠금) 방지
        session = auth_session_from_cookie(self.headers.get("Cookie", ""))
        if session:
            # 담당자 온/오프라인(조직도): 인증 세션이 서버에 말한 시각을 찍는다.
            # 관리자는 센터장(유현민) 자리다 — 근거는 ADMIN_PRESENCE_SLUG 주석.
            _role = str(session.get("role") or "")
            if _role == "staff":
                presence_touch(session.get("staff_slug"))
            elif _role == "admin" and not session.get("legacy_local"):
                presence_touch(ADMIN_PRESENCE_SLUG)
            # ★ 기사(tech)는 여기를 **통과하지 못한다.** 이 관문 뒤에는 원장 전체가
            #   있고, 기사 링크는 카톡으로 돌아다닌다. 기사 화면은 /api/tech/* 만 쓴다.
            return str(session.get("role") or "") != "tech"
        if _locked(self.client_address[0]):
            return False
        # 쿠키를 쓸 수 없는 기존 localhost 자동화만 관리자 PIN 헤더로 호환한다.
        return verify_pin(self.headers.get("X-Pin", ""))

    def _actor(self):
        session = auth_session_from_cookie(self.headers.get("Cookie", ""))
        if session:
            return session
        # 과거 로컬 스크립트 호환. 원격 요청은 로그인 세션 없이 관리자 권한을
        # 얻을 수 없다.
        if self.client_address[0] in ("127.0.0.1", "::1", "localhost"):
            return {"role": "admin", "staff_slug": "", "legacy_local": True}
        return {"role": "unknown", "staff_slug": ""}

    def _actor_name(self):
        actor = self._actor()
        role = str(actor.get("role") or "unknown")
        slug = str(actor.get("staff_slug") or "").strip()
        return f"{role}:{slug}" if slug else role

    def _idempotency_key(self, body=None):
        key = str(self.headers.get("Idempotency-Key") or "").strip()
        if not key and isinstance(body, dict):
            key = str(body.get("idempotency_key") or "").strip()
        return key[:240]

    def _require_admin(self):
        if self._actor().get("role") != "admin":
            self._send(403, {"ok": False, "error": "관리자 전용 기능입니다"})
            return False
        return True

    def _require_staff(self, *allowed_slugs):
        actor = self._actor()
        slug = str(actor.get("staff_slug") or "")
        if actor.get("role") != "staff" or (allowed_slugs and slug not in allowed_slugs):
            self._send(403, {"ok": False, "error": "이 업무센터에서 사용할 수 없는 기능입니다"})
            return ""
        return slug

    def do_OPTIONS(self):
        """브라우저 수집기(band.us 페이지)에서 보내는 사전 요청 허용 — 로컬에서만 쓴다"""
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers",
                         "Content-Type, X-Pin, X-Staff-Slug, Idempotency-Key")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        # Chrome의 Private Network Access: 공개 사이트(https)에서 로컬 주소로 보낼 때 필요
        self.send_header("Access-Control-Allow-Private-Network", "true")
        self.send_header("Access-Control-Max-Age", "600")
        self.end_headers()

    def do_GET(self):
        p = self.path.split("?")[0]
        staff_match = re.fullmatch(r"/staff/([a-z0-9-]+)", p)
        staff_slug = staff_match.group(1) if staff_match else ""
        if staff_slug in STAFF_CENTER_ALIASES:
            target = STAFF_CENTER_ALIASES[staff_slug]
            return self._send(
                308, b"", "text/plain; charset=utf-8",
                headers={"Location": f"/staff/{target}"},
            )
        if staff_slug and staff_slug not in STAFF_CENTERS:
            return self._send(404, {"error": "등록되지 않은 업무센터"})
        # ── AS 기사 전용 화면 (2026-08-08 지시) — 비밀번호 대신 링크 열쇠
        tech_match = re.fullmatch(r"/t/([a-z0-9-]+)", p)
        if tech_match:
            tslug = tech_match.group(1)
            if tslug not in AS_TECH_CENTERS:
                return self._send(404, {"error": "등록되지 않은 기사"})
            # ★ 여기서 `from urllib.parse import parse_qs` 를 하면 **안 된다.**
            #   함수 안 import 는 그 이름을 함수 전체의 지역변수로 만든다 — 이 가지를
            #   안 지나간 다른 가지(/manifest.json)에서 `parse_qs` 가 '아직 값이 없는
            #   지역변수'가 되어 500 이 났다. 2026-08-07~08 이틀간 매니페스트가 통째로
            #   500 이었고, 그래서 **설치 안내가 아예 안 뜬 것**이다(모듈 맨 위 25행에
            #   이미 import 돼 있다).
            key = (parse_qs(urlsplit(self.path).query).get("k", [""])[0] or "").strip()
            if key:
                # ★ 열쇠는 **주소창에서 지운다.** 카톡·밴드에 붙여 넣은 화면 갈무리로
                #   열쇠가 새는 것이 가장 흔한 사고다. 한 번 쓰고 쿠키로 바꾼 뒤 넘긴다.
                if not tech_check_key(tslug, key):
                    return self._send(403, {"error": "링크가 올바르지 않습니다. 담당자에게 새 링크를 요청하세요."})
                token, _s = create_auth_session(tech_slug=tslug)
                return self._send(302, b"", "text/plain; charset=utf-8", headers={
                    "Location": f"/t/{tslug}", "Set-Cookie": auth_cookie(token)})
            html = open(os.path.join(BASE, "tech.html"), encoding="utf-8").read()
            host = (self.headers.get("Host") or "").lower()
            if "trycloudflare.com" in host:      # 임시 터널 주소를 앱 아이콘에 박지 않는다
                html = html.replace('<link rel="manifest" href="/manifest.json">', "")
            else:
                html = html.replace('<link rel="manifest" href="/manifest.json">',
                                    f'<link rel="manifest" href="/manifest.json?tech={tslug}">')
            return self._send(200, html.encode("utf-8"), "text/html; charset=utf-8")
        if p == "/api/tech/links":
            # 기사에게 보낼 링크 — **관리자만** 본다(열쇠가 그대로 들어 있다).
            if not self._require_admin():
                return
            ks = tech_keys()
            return self._send(200, {"links": [
                {"slug": s, "이름": c["name"], "직함": c.get("직함") or "",
                 "링크": f"{FIXED_LIVE_ENTRY}/t/{s}?k={ks.get(s,'')}"}
                for s, c in AS_TECH_CENTERS.items()],
                "안내": "이 링크는 비밀번호와 같습니다. 본인에게만 보내세요."})
        if p.startswith("/api/tech/"):
            actor = self._actor()
            tslug = str(actor.get("staff_slug") or "")
            if str(actor.get("role") or "") != "tech" or tslug not in AS_TECH_CENTERS:
                return self._send(401, {"ok": False, "error": "기사 링크로 열어 주세요"})
            presence_touch(tslug)   # 담당자 온/오프라인(조직도) — 기사 화면 접속
            if p == "/api/tech/board":
                return self._send(200, tech_board(tslug))
            return self._send(404, {"error": "no route"})
        if p in ("/", "/index.html", "/ryu") or staff_slug:
            html = open(os.path.join(BASE, "index.html"), encoding="utf-8").read()
            if staff_slug:
                html = html.replace(
                    '<link rel="manifest" href="/manifest.json">',
                    f'<link rel="manifest" href="/manifest.json?staff={staff_slug}">'
                )
            # ★★ 터널 주소로 들어온 경우에는 **설치 가능하게 만들지 않는다**.
            #   터널 호스트는 띄울 때마다 바뀌는데, 여기서 [설치]를 하면 그 임시 호스트가
            #   앱 아이콘에 **영구히 박힌다**. 주소가 바뀌는 순간 그 아이콘은 영영
            #   'ERR_FAILED'만 띄운다 — PC·폰 둘 다 그렇게 죽었다(2026-07-28 실사고).
            #   설치는 오직 고정 주소에서만 되게 하고, 여기서는 매니페스트·서비스워커를 뺀다.
            host = (self.headers.get("Host") or "").lower()
            if "trycloudflare.com" in host:
                html = html.replace('<link rel="manifest" href="/manifest.json">', "")
                html = html.replace("navigator.serviceWorker.register('/sw.js')",
                                    "Promise.reject()")
            return self._send(200, html.encode("utf-8"), "text/html; charset=utf-8")
        if p.startswith("/brand/"):                    # 고객사 CI(쿠팡 로고) — 로컬 파일만 서빙
            fn = os.path.basename(p)
            fp = os.path.join(BASE, "brand", fn)
            ext = os.path.splitext(fn)[1].lower()
            ct = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                  ".svg": "image/svg+xml", ".webp": "image/webp"}.get(ext)
            if not ct or not os.path.exists(fp):
                return self._send(404, {"error": "no brand asset"})
            return self._send(200, open(fp, "rb").read(), ct)
        if p.startswith("/icons/"):
            fn = os.path.basename(p)
            fp = os.path.join(BASE, "icons", fn)
            if not fn.endswith(".svg") or not os.path.isfile(fp):
                return self._send(404, {"error": "no icon asset"})
            return self._send(200, open(fp, "rb").read(), "image/svg+xml")
        if p == "/api/brief":
            # 대표 보고용 '내용' 브리핑. 화면·PC 리포트·폰 사본이 **같은 문장**을 쓰도록
            # daily_brief 하나만 출처로 삼는다(따로 만들면 숫자가 갈린다).
            try:
                import daily_brief as DB
                day = None
                m = re.search(r"[?&]date=(\d{4}-\d{2}-\d{2})", self.path)
                if m:
                    day = m.group(1)
                    if not day.startswith(APP_YEAR + "-"):
                        return self._send(200, {"ok": False,
                                               "error": f"앱 브리핑은 {APP_YEAR}년만 표시합니다"})
                b = get_daily_brief(day)
                return self._send(200, {"ok": True, "text": DB.text(b), **b})
            except Exception as e:
                return self._send(200, {"ok": False, "error": str(e)[:200]})
        if p == "/api/codes":
            return self._send(200, get_codes())
        if p == "/api/brand":
            return self._send(200, {"logo": brand_logo()})
        if re.fullmatch(r"/icon(?:-\d+)?\.(svg|png)", p):      # 아이콘(벡터/래스터 공용)
            try:
                return self._send(200, open(os.path.join(BASE, p.lstrip("/")), "rb").read(),
                                  "image/svg+xml" if p.endswith(".svg") else "image/png")
            except Exception:
                return self._send(404, {"error": "no icon"})
        if p.startswith("/fonts/"):
            # 나눔고딕 웹폰트(사용자 지시 2026-07-31). 구글 폰트 주소를 그대로 쓰면
            # 인터넷이 없는 곳에서 글꼴이 안 뜨고, 외부로 요청이 나간다. 그래서 동봉해 여기서 낸다.
            # ★ 경로에 이름만 받는다 — '..' 를 그대로 이어 붙이면 서버 파일이 통째로 새 나간다.
            name = os.path.basename(p[len("/fonts/"):])
            ext = os.path.splitext(name)[1].lower()
            if ext not in (".woff2", ".woff", ".css"):
                return self._send(404, {"error": "no font"})
            try:
                raw = open(os.path.join(BASE, "fonts", name), "rb").read()
            except Exception:
                return self._send(404, {"error": "no font"})
            ct = {".woff2": "font/woff2", ".woff": "font/woff",
                  ".css": "text/css; charset=utf-8"}[ext]
            # ★ _send 를 쓰지 않는다. _send 는 Cache-Control: no-store 를 먼저 박아서
            #   폰이 화면을 열 때마다 4MB 를 다시 받는다. 폰트 조각은 파일명이 곧 판본이라
            #   (구글이 이름에 해시를 넣는다) 영구 캐시가 안전하고, 그래야 두 번째부터 즉시 뜬다.
            self.send_response(200)
            self.send_header("Content-Type", ct)
            self.send_header("Content-Length", str(len(raw)))
            self.send_header("Cache-Control", "public, max-age=31536000, immutable")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(raw)
            return
        if p == "/sw.js":
            # ★ 예전 주석: "캐시는 하지 않는다 — 옛 화면이 남으면 안 된다."
            #   그 걱정은 **데이터**에 대해서는 지금도 맞다. 그래서 /api/ 는 여전히 캐시하지 않는다.
            #   다만 아무것도 캐시하지 않으면 **PC 가 꺼졌을 때 앱이 아예 안 열린다.** 폰에서
            #   입력해 두려면 화면은 떠야 하므로(사용자 지시 2026-07-31), 껍데기만 캐시한다.
            #   · 화면(HTML)은 네트워크 먼저 — PC 가 켜져 있으면 언제나 최신을 받는다.
            #     실패할 때만 캐시본을 낸다. 그래서 '옛 화면이 남는' 일은 생기지 않는다.
            #   · 글꼴 조각은 파일명이 곧 판본이라 캐시 먼저(276조각을 매번 받을 이유가 없다).
            #   · /api/ 는 손대지 않는다. 쓰기 실패는 화면 쪽 오프라인 큐가 받는다.
            js = ("""
const V = 'csos-%s';
const SHELL = ['/', '/fonts/nanumgothic.css', '/icon-192.png', '/manifest.json'];
self.addEventListener('install', e => {
  e.waitUntil(caches.open(V).then(c => c.addAll(SHELL).catch(()=>{})).then(()=>self.skipWaiting()));
});
self.addEventListener('activate', e => {
  e.waitUntil(caches.keys()
    .then(ks => Promise.all(ks.filter(k => k !== V).map(k => caches.delete(k))))
    .then(() => self.clients.claim()));
});
self.addEventListener('fetch', e => {
  const req = e.request;
  if (req.method !== 'GET') return;                       // 쓰기는 큐가 맡는다
  const url = new URL(req.url);
  if (url.origin !== self.location.origin) return;
  if (url.pathname.startsWith('/api/')) return;           // 데이터는 캐시하지 않는다
  if (url.pathname.startsWith('/fonts/')) {               // 글꼴: 캐시 먼저
    e.respondWith(caches.match(req).then(hit => hit || fetch(req).then(r => {
      const copy = r.clone(); caches.open(V).then(c => c.put(req, copy)); return r;
    })));
    return;
  }
  // 화면: 네트워크 먼저, 끊겼을 때만 캐시본
  e.respondWith(fetch(req).then(r => {
    const copy = r.clone(); caches.open(V).then(c => c.put(req, copy)); return r;
  }).catch(() => caches.match(req).then(hit => hit || caches.match('/'))));
});
""" % build_id()).lstrip()
            # ★ _send 는 str을 받으면 JSON으로 감싼다 — 반드시 bytes로 넘겨야 스크립트가 된다
            return self._send(200, js.encode("utf-8"), "application/javascript; charset=utf-8")
        if p == "/manifest.json":                      # 홈 화면에 추가 시 앱처럼 보이게
            # ★ start_url 은 **이 페이지와 같은 출처**여야 한다. 다른 도메인을 넣으면
            #   크롬이 매니페스트를 통째로 무시해 [설치 및 바로가기 만들기]가 먹통이 된다
            #   (2026-07-27에 고정 주소를 넣었다가 실제로 설치가 안 됐다).
            #   그래서 여기는 "/" 로 두고, **오래 쓸 아이콘은 고정 주소에서 설치**한다.
            #   ★ 2026-08-13: 이 줄에는 원래 "터널 주소로 들어온 사람에게는 index.html이
            #     배너로 그 사실을 알린다"고 적혀 있었는데 **그 배너가 실재하지 않았다.**
            #     주석만 약속하고 화면은 아무 말도 안 하니, 터널로 연 사람은 설치가
            #     일부러 막힌 줄도 모르고 "앱 설치가 안 된다"만 겪었다. 지금은 실제로
            #     index.html 의 `installState()` · `noticeIfNotInstallable()` 이 이유와
            #     고정 주소를 말한다. 여기 판정(트라이클라우드플레어 차단)을 고치면
            #     그쪽 `installState()` 의 같은 조건도 함께 본다.
            icon_rev = icon_revision()
            query = parse_qs(urlsplit(self.path).query)
            staff_slug = str((query.get("staff") or [""])[0]).strip()
            center = STAFF_CENTERS.get(staff_slug)
            start_url = f"/staff/{staff_slug}" if center else "/"
            app_name = center["title"] if center else "Coupang Service Operations System"
            app_id = start_url if center else "/"
            # ★ 기사용(2026-08-08 지시 "링크를 열었을 때 크롬으로 자동 설치").
            #   `/t/<slug>` 는 매니페스트에 `?tech=` 를 붙여 왔는데 **여기서 그걸 안 읽었다.**
            #   그래서 기사가 설치해도 아이콘이 `start_url="/"` — 즉 **PIN 걸린 관리자
            #   화면**으로 갔다. 설치는 됐는데 못 들어가니 아무도 안 쓰게 된다.
            #   ★ start_url 에 **열쇠는 넣지 않는다.** 매니페스트는 캐시되고 기기에 남는다 —
            #     거기 열쇠가 박히면 비밀번호를 파일로 뿌리는 것과 같다. 세션 쿠키가
            #     `/t/<slug>` 를 열어 주므로 필요도 없다.
            tech_slug = str((query.get("tech") or [""])[0]).strip()
            tech = AS_TECH_CENTERS.get(tech_slug)
            if tech:
                start_url = app_id = f"/t/{tech_slug}"
                app_name = f"{tech['name']} · 쿠팡 AS"
            return self._send(200, {
                "name": app_name,
                "short_name": (tech["name"] if tech else
                               center["name"] if center else "CSOS"),
                "id": app_id, "start_url": start_url, "scope": "/", "display": "standalone",
                "background_color": "#060D2B", "theme_color": "#060D2B",
                "icons": [
                    {"src": f"/icon-192.png?v={icon_rev}", "sizes": "192x192",
                     "type": "image/png", "purpose": "any"},
                    {"src": f"/icon-512.png?v={icon_rev}", "sizes": "512x512",
                     "type": "image/png", "purpose": "any maskable"}]},
                "application/manifest+json")
        if p == "/api/ping":
            return self._send(200, {"app": "coupang-work", "demo": DEMO, "build": build_id()})
        # ── 밴드 자동수집 (2026-08-09 지시: Claude Code 없이 앱이 스스로 수집) ─────────
        #   로그인된 밴드 탭의 유저스크립트가 이 둘을 읽어 스스로 긁는다 — Claude Code 가
        #   수집 루프에서 빠져 크레딧을 안 쓴다. _send 가 이미 CORS(*)·사설망 허용을
        #   붙이므로 band.us 페이지에서 곧바로 fetch 된다. 계획은 회차가 미리 계산해 둔
        #   것을 그대로 내려 준다(원장 읽기가 비싸 요청마다 다시 계산하지 않는다).
        if p == "/grab_posts.js":
            try:
                js = open(os.path.join(ROOT, "band", "grab_posts.js"),
                          encoding="utf-8").read()
            except OSError:
                return self._send(404, {"error": "수집기 없음"})
            return self._send(200, js.encode("utf-8"),
                              "application/javascript; charset=utf-8")
        if p == "/band_auto_collect.user.js":
            try:
                js = open(os.path.join(ROOT, "band", "band_auto_collect.user.js"),
                          encoding="utf-8").read()
            except OSError:
                return self._send(404, {"error": "유저스크립트 없음"})
            return self._send(200, js.encode("utf-8"),
                              "application/javascript; charset=utf-8")
        # ERP 수집기 — 로그인된 ERP 탭이 그대로 fetch 해서 돌린다.
        #   const s = await (await fetch('http://127.0.0.1:<포트>/erp_grab.js?keys=hometax')).text();
        #   (0,eval)(s);   →  진행상황은 window.__ERPALL
        # ★ 이 자리가 있어야 수집기를 **대화 컨텍스트로 퍼 나르지 않는다**(8.8KB · 한 번에
        #   2~3천 토큰). 크롬은 localhost 를 신뢰 출처로 봐서 https 페이지에서도 http 로
        #   가져온다(밴드 쪽에서 실측된 것과 같은 길 — [217]).
        # ★ PIN 게이트 **앞**이다(위 /grab_posts.js 와 같은 이유): 보내는 쪽은 ecount.com
        #   탭이라 이 서버의 인증 쿠키를 못 싣는다. 뒤에 두면 전부 401 이 되고, 그러면
        #   "수집기가 안 돈다"가 아니라 "수집이 0건"으로 보인다.
        if p == "/erp_grab.js":
            q = parse_qs(urlsplit(self.path).query)
            keys = [k.strip() for k in (q.get("keys", [""])[0] or "").split(",") if k.strip()]
            preset = (q.get("preset", [""])[0] or "").strip() or None
            try:
                import erp_grab
                js, unknown = erp_grab.build_all(keys or None, preset)
            except Exception as e:                     # noqa: BLE001 — 서버를 죽이지 않는다
                return self._send(500, {"error": f"수집기를 못 만들었다: {e}"})
            # ★ 빈 스크립트를 200 으로 주지 않는다. 붙여 넣은 쪽은 '돌았는데 0건'으로 읽고
            #   등록 안 된 이름은 영영 안 드러난다([169] — 실패가 성공처럼 보이는 자리).
            if not js:
                return self._send(404, {"error": "등록부에 없는 화면이라 만들 것이 없다",
                                        "모르는이름": unknown,
                                        "할것": "python erp_grab.py --find <이름>"})
            return self._send(200, js.encode("utf-8"),
                              "application/javascript; charset=utf-8")
        if p == "/api/collect_plan":
            q = parse_qs(urlsplit(self.path).query)
            band = (q.get("band", [""])[0] or "").strip()
            try:
                sys.path.insert(0, os.path.join(ROOT, "band"))
                import comment_backfill as _cb
                return self._send(200, _cb.load_plan(band or None))
            except Exception as e:
                return self._send(200, {"band": band, "nos": [],
                                        "error": type(e).__name__})
        if p == "/api/auth/session":
            session = auth_session_from_cookie(self.headers.get("Cookie", ""))
            if not session:
                return self._send(401, {"ok": False, "error": "기기 인증 없음"})
            # 정상 기기는 앱을 열 때마다 만료일을 연장한다. PIN은 변경하거나 브라우저
            # 데이터를 직접 지우기 전까지 다시 물어보지 않는다.
            token, renewed = create_auth_session(
                session.get("staff_slug") if session.get("role") == "staff" else "")
            return self._send(200, {
                "ok": True,
                "authenticated": True,
                "role": renewed["role"],
                "staff_slug": renewed.get("staff_slug") or "",
                "expires_at": renewed["expires_at"],
            }, headers={"Set-Cookie": auth_cookie(token)})
        if p == "/api/flow":
            # AS 접수 → 수금 업무 흐름도(2026-08-07 지시). 정본은 DB 다.
            # ★ PIN 게이트 **앞**에 둔다(2026-08-12): 이것은 업무 절차 다이어그램이라
            #   원장 같은 민감 데이터가 아니고, 같은 내용이 이미 오프라인 꾸러미
            #   (cloud_publish d["flow"] → data.enc)로 공개돼 있다. 게이트 뒤에 있으면
            #   PIN·세션이 없는 폰이 loadFlow 로 받을 때 빈값을 받아 "흐름이 비어
            #   있습니다"가 뜬다(사용자 실측 화면 — 캡처도 그래서 실패했다).
            #   /grab_posts.js·/api/collect_plan 처럼 공개군에 속한다.
            # 차트는 여러 장(종전/개선, 추후 정기점검): ?chart=<key>
            import ledger_db
            fqs = parse_qs(urlsplit(self.path).query)   # 함수 안 import 금지(윗 주석)
            key = (fqs.get("chart", [""])[0] or "").strip() or "as_legacy"
            try:
                steps = ledger_db.flow_steps(key)
            except ValueError as e:
                return self._send(400, {"ok": False, "error": str(e)})
            return self._send(200, {"steps": steps, "chart": key,
                                    "charts": ledger_db.flow_charts(),
                                    "notes": ledger_db.flow_notes(key)})
        if not self._auth():
            return self._send(401, {"error": "PIN"})
        if p == "/api/live-state":
            return self._send(200, get_live_state())
        if p == "/api/status":
            return self._send(200, get_status())
        if p == "/api/app-store/status":
            try:
                from app_store import default_store
                state = default_store().status()
                db_path = str(state.pop("db_path", ""))
                return self._send(200, {
                    **state,
                    "canonical": True,
                    "db_file": os.path.basename(db_path),
                    "pending_meaning": "Excel 보관본 생성 outbox",
                })
            except Exception as exc:
                return self._send(503, {"ok": False, "canonical": True,
                                        "error": str(exc)[:240]})
        if p == "/api/automation/status":
            try:
                import automation_pipeline
                return self._send(200, automation_pipeline.status())
            except Exception as exc:
                # 관제 카드 한 장이 고장 나도 다른 업무 화면은 살아 있어야 한다.
                return self._send(200, {
                    "ok": False,
                    "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                    "running": False,
                    "last_run": {"status": "error", "summary": str(exc)[:240]},
                    "sources": {},
                    "app_db": {"status": "error", "detail": "자동화 상태를 읽지 못했습니다"},
                    "excel": {"status": "error", "error": str(exc)[:240]},
                    "human_gates": [],
                })
        if p == "/api/autopilot":
            # 읽기 전용 상태. 실제 재시도는 워치독 한 곳만 실행해 중복을 막는다.
            try:
                import autopilot as _autopilot
                return self._send(200, _autopilot.status())
            except Exception as exc:
                return self._send(200, {"active": 0, "error": str(exc)[:180]})
        if p == "/api/notifications":
            # 받는이에 따라 다른 목록이다 — 올린 사람은 제 업로드 결과를 본다.
            return self._send(200, get_notifications(self._actor()))
        if p == "/api/staff/centers":
            return self._send(200, {"centers": staff_centers_payload()})
        if p == "/api/staff/completions":
            return self._send(200, staff_completions_payload())
        if p == "/api/remote/status":
            # 리모컨 불출·납품 현황(2026-08-03 지시). 류지영·오종현 업무센터와 관리자 공용.
            import ledger_db
            return self._send(200, ledger_db.remote_status())
        if p == "/api/source-file":
            # 원본 파일을 **접속한 기기로** 내려보낸다 (2026-08-07 지시:
            # "원본 자료 클릭하면 접속한 디바이스에서 바로 열리게").
            # ★ 예전에는 클릭이 /api/open → os.startfile 이라 **서버 PC 에서** 열렸다.
            #   폰이나 다른 PC 로 접속하면 눌러도 아무 일이 없었고(원격은 아예 403),
            #   사무실 PC 에서는 아무도 안 보는 화면에 창만 떴다. 이제 파일을 그대로
            #   내려보내고, 무엇으로 열지는 **기기가** 정한다 — 폰이면 폰의 뷰어로 뜬다.
            if not self._require_admin():
                return
            from urllib.parse import quote          # parse_qs 는 모듈 맨 위에 있다(윗 주석)
            qs = parse_qs(urlsplit(self.path).query)
            want = (qs.get("path", [""])[0] or "").strip()
            doc = _source_index() or {}
            # ★ 화이트리스트는 **색인 하나**다. 임의 경로를 열어 주면 서버 파일이 통째로 샌다.
            allow = {r.get("path") for r in (doc.get("rows") or [])}
            if not want or want not in allow or not os.path.isfile(want):
                return self._send(404, {"ok": False, "error": "허용된 원본이 아닙니다"})
            # 민감 자료는 색인에서 이미 빠지지만, **색인은 회차로 다시 만들어진다.**
            # 옛 색인이 남아 있는 동안 여기로 직접 요청이 오면 그대로 나갈 수 있으므로
            # 파일을 내보내기 직전에 한 번 더 판정한다(2026-08-07 지시).
            try:
                from source_index import is_private
                if is_private(want):
                    return self._send(404, {"ok": False, "error": "허용된 원본이 아닙니다"})
            except ImportError:
                if os.path.basename(want).startswith("통화_"):
                    return self._send(404, {"ok": False, "error": "허용된 원본이 아닙니다"})
            name = os.path.basename(want)
            ext = os.path.splitext(name)[1].lower()
            inline_ct = {".pdf": "application/pdf", ".png": "image/png",
                         ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                         ".gif": "image/gif", ".webp": "image/webp",
                         ".txt": "text/plain; charset=utf-8",
                         ".md": "text/plain; charset=utf-8",
                         ".csv": "text/csv; charset=utf-8"}
            attach_ct = {".xlsx": "application/vnd.openxmlformats-officedocument"
                                  ".spreadsheetml.sheet",
                         ".xls": "application/vnd.ms-excel",
                         ".docx": "application/vnd.openxmlformats-officedocument"
                                  ".wordprocessingml.document",
                         ".zip": "application/zip"}
            ct = inline_ct.get(ext) or attach_ct.get(ext) or "application/octet-stream"
            # PDF·사진은 폰에서 바로 보여 주는 편이 낫고, 엑셀은 기기의 앱으로 넘겨야 한다.
            # `?dl=1` 이면 무엇이든 내려받기로 준다.
            how = ("inline" if ext in inline_ct and qs.get("dl", [""])[0] != "1"
                   else "attachment")
            try:
                size = os.path.getsize(want)
                fh = open(want, "rb")
            except OSError as exc:
                return self._send(500, {"ok": False, "error": str(exc)[:200]})
            self.send_response(200)
            self.send_header("Content-Type", ct)
            self.send_header("Content-Length", str(size))
            # 한글 파일 이름은 RFC 5987 로 싣는다 — 그냥 넣으면 헤더가 깨져 이름이 사라진다.
            self.send_header("Content-Disposition", f"{how}; filename*=UTF-8''{quote(name)}")
            self.send_header("Cache-Control", "private, max-age=60")
            self.end_headers()
            try:
                with fh:
                    while True:
                        chunk = fh.read(256 * 1024)
                        if not chunk:
                            break
                        self.wfile.write(chunk)
            except (BrokenPipeError, ConnectionResetError):
                pass       # 미리보기를 닫으면 흔히 난다 — 서버 오류가 아니다
            try:
                import ledger_db as _ldb      # 이 파일은 ledger_db 를 함수 안에서만 쓴다
                _ldb.ux_add([{"kind": "tap", "target": "원본열기", "detail": name}])
            except Exception:
                pass
            return
        if p == "/api/originals":
            # 한 건(정산/작업)에 딸린 원본을 갈래별로. 상세 화면의 '원본 바로 열기'가 쓴다.
            # ★ 인증은 /api/source-file 과 **같은 문턱**이어야 한다 — 여기서 경로를
            #   알려 주고 저기서 막으면, 어떤 파일이 있는지가 미인증에게 새어 나간다.
            if not self._require_admin():
                return
            qs = parse_qs(urlsplit(self.path).query)   # 함수 안 import 금지(윗 주석)
            g = lambda k: (qs.get(k, [""])[0] or "").strip()
            return self._send(200, _originals_for(g("uj"), g("po"), g("slip")))
        if p == "/api/sources":
            qs = parse_qs(urlsplit(self.path).query)   # 함수 안 import 금지(윗 주석)
            # 원본 자료 색인(source_index.py 산출물) — 앱에서 필터·검색해 클릭 한 번에 연다.
            # 파일이 수만 개가 될 수 있어 서버에서 먼저 거른다(q·kind·limit).
            doc = _source_index()
            if not doc:
                return self._send(200, {"count": 0, "rows": [], "kinds": [], "months": [],
                                        "note": "색인 없음 — source_index.py 실행 필요"})
            rows = doc.get("rows") or []
            # ★ 민감 자료(통화 메모)는 색인 단계에서 이미 빠진다. 그래도 응답에서 한 번 더
            #   막는다 — 색인은 하루 회차로 다시 만들어지므로 **옛 색인 파일이 남아 있는
            #   동안**에도 앱에 뜨면 안 된다(2026-08-07 지시). 화면 코드가 바뀌어도 서버가 먼저 막는다.
            try:
                from source_index import is_private
                rows = [r for r in rows
                        if not is_private(r.get("path") or "", r.get("name") or "")]
            except Exception:
                rows = [r for r in rows if not str(r.get("name") or "").startswith("통화_")]
            q = (qs.get("q", [""])[0] or "").strip().lower()
            kind = (qs.get("kind", [""])[0] or "").strip()
            month = (qs.get("month", [""])[0] or "").strip()   # 2026-01 형식
            year = (qs.get("year", [""])[0] or "").strip()     # 2026 형식(월 없이 연도만)
            # 갈래(세금계산서·거래명세서…)는 kind 여러 개 + 경로 패턴의 묶음이라
            # 화면이 kinds=A|B|C 와 path=문자열 로 넘긴다. 8천 건을 클라이언트로
            # 다 보내지 않고 **여기서** 거른다(2026-08-05: 400건만 보내 필터가 0이 됐다).
            kinds = [x for x in (qs.get("kinds", [""])[0] or "").split("|") if x]
            pathkey = (qs.get("path", [""])[0] or "").strip()
            if kind:
                rows = [r for r in rows if r.get("kind") == kind]
            if kinds or pathkey:
                rows = [r for r in rows
                        if (kinds and r.get("kind") in kinds)
                        or (pathkey and pathkey in (r.get("path") or ""))]
            if month:
                # 업무 발생 월 — 전표번호(건별 PDF)가 있으면 그것이 정본, 없으면 파일 날짜.
                rows = [r for r in rows
                        if (r.get("slip") or "")[:7].replace("-", "-") == month
                        or (not r.get("slip") and (r.get("date") or "")[:7] == month)]
            elif year:
                # 연도만 고른 경우(2026-08-06). 월과 같은 기준(전표일 우선)으로 본다 —
                # 기준이 다르면 '2026년 전체'와 월 12개의 합이 어긋나 사람이 숫자를 못 믿는다.
                rows = [r for r in rows
                        if (r.get("slip") or "")[:4] == year
                        or (not r.get("slip") and (r.get("date") or "")[:4] == year)]
            if q:
                rows = [r for r in rows
                        if q in (r.get("name", "") + r.get("uj", "") + r.get("slip", "")
                                 + r.get("po", "") + r.get("path", "")).lower()]
            kinds, months = _SRC_IDX["kinds"], _SRC_IDX["months"]
            try:
                limit = max(1, min(500, int(qs.get("limit", ["200"])[0])))
            except Exception:
                limit = 200
            return self._send(200, {"count": len(rows), "built": doc.get("built"),
                                    "kinds": kinds, "months": months, "rows": rows[:limit]})
        if p == "/api/tax-overdue":
            # 세금계산서 미발행 경과(통계 화면용) — tax_invoice_watch 가 daily_run 에서 갱신
            try:
                path = os.path.join(ROOT, "reports", "세금계산서_미발행_경과.json")
                return self._send(200, json.load(open(path, encoding="utf-8")))
            except Exception:
                return self._send(200, {"total": 0, "buckets": {}, "rows": []})
        if p == "/api/staff/work-log-status":
            actor = self._actor()
            if actor.get("role") == "staff" and actor.get("staff_slug") != "ryu-jiyeong":
                return self._send(403, {"ok": False, "error": "류지영 업무센터 전용 기능입니다"})
            if actor.get("role") not in ("admin", "staff"):
                return self._send(403, {"ok": False, "error": "업무센터 세션이 필요합니다"})
            try:
                params = parse_qs(urlsplit(self.path).query)
                return self._send(200, get_work_log_view(
                    day=str((params.get("date") or [""])[0]).strip(),
                    category=str((params.get("category") or [""])[0]).strip(),
                    state=str((params.get("state") or [""])[0]).strip(),
                    query=str((params.get("q") or [""])[0]).strip(),
                ))
            except Exception as exc:
                return self._send(200, {"ok": False, "error": str(exc)[:240]})
        # 철거·신규납품은 응답 단계에서 뺀다 — 앱이 아예 받지 않게 한다.
        # (앱에도 같은 필터가 있지만, 화면 코드가 바뀌어도 새어 나가지 않게 서버가 먼저 막는다)
        if p == "/api/settlements":
            return self._send(200, {"rows": drop_side_work(get_settlements())})
        if p == "/api/works":
            w = get_works()
            return self._send(200, {k: (drop_side_work(v) if isinstance(v, list) else v)
                                    for k, v in (w or {}).items()})
        if p == "/api/issues":
            iss = get_issues()
            if isinstance(iss, dict) and isinstance(iss.get("rows"), list):
                iss = {**iss, "rows": drop_side_work(iss["rows"])}
            return self._send(200, iss)
        if p in ("/api/staff/records", "/api/ryu/records"):
            actor = self._actor()
            if actor.get("role") == "staff":
                staff_slug = str(actor.get("staff_slug") or "")
            elif actor.get("role") == "admin":
                staff_slug = "admin"
            else:
                return self._send(403, {"ok": False, "error": "업무센터 권한이 필요합니다"})
            try:
                return self._send(200, get_staff_records(staff_slug))
            except PermissionError as exc:
                return self._send(403, {"ok": False, "error": str(exc)})
        if p == "/api/exec_report":
            m = re.search(r"[?&]date=(\d{4}-\d{2}-\d{2})", self.path)
            day = m.group(1) if m else None
            if day and not day.startswith(APP_YEAR + "-"):
                return self._send(400, {"error": f"{APP_YEAR}년 날짜만 선택할 수 있습니다"})
            return self._send(200, get_exec_report(day))
        if p in {
            "/api/v1/reports/daily/exceptions",
            "/api/v1/reports/daily/as-backlog",
            "/api/v1/reports/daily/inspection-progress",
            "/api/v1/reports/daily/statement-progress",
            "/api/v1/as-requests/backlog-summary",
            "/api/v1/as-requests/backlog-detail",
            "/api/v1/inspections/quarter-progress",
            "/api/v1/statements/eligibility-summary",
            "/api/v1/statements/unissued",
            "/api/v1/tax-invoices/composition-check",
        }:
            report = get_representative_report()
            if p.endswith(("as-backlog", "backlog-summary", "backlog-detail")):
                return self._send(200, {"meta": report["meta"], **report["돌발AS"]})
            if p.endswith(("inspection-progress", "quarter-progress")):
                return self._send(200, {"meta": report["meta"], **report["정기점검"]})
            if p.endswith(("statement-progress", "eligibility-summary", "unissued",
                           "composition-check")):
                return self._send(200, {"meta": report["meta"], **report["거래명세서"],
                                        "업무기준확인필요": report["업무기준확인필요"]})
            return self._send(200, report)
        if p == "/api/ux":
            return self._send(200, {"summary": _ux_summary()})
        if p == "/api/error_help":
            # 오류 하나를 **초등학생도 알아볼 말**로 풀어 준다. 읽기 전용이다.
            # ★ 모르는 오류에도 답한다 — 다만 '앎: false' 로 정직하게 말하고,
            #   지어내는 대신 '캡처해서 보내세요' 로 보낸다. 뭉쳐서 "알 수 없는
            #   오류입니다" 라고만 하면 사람이 엉뚱한 데를 고치러 간다.
            qs = parse_qs(urlsplit(self.path).query)
            tgt = (qs.get("target", [""])[0] or "")[:160]
            det = (qs.get("detail", [""])[0] or "")[:400]
            try:
                import error_book
                return self._send(200, {"ok": True, **error_book.help_for(tgt, det)})
            except Exception as exc:
                # 도움말을 못 만든 것을 '도움말 없음'으로 적지 않는다.
                return self._send(200, {"ok": False, "앎": False,
                                        "이름": "도움말을 불러오지 못함",
                                        "쉬운말": "오류 설명을 여는 것 자체가 실패했습니다.",
                                        "왜": str(exc)[:160],
                                        "하세요": ["이 화면을 캡처해서 관리자에게 보냅니다."],
                                        "신고문구": f"[앱 오류 신고]\n어디: {tgt}\n무엇: {det}\n"
                                                f"도움말 실패: {str(exc)[:160]}"})
        if p == "/api/ask":
            # ★ 앱이 스스로 답한다 — 클로드를 부르기 전에 여기서 먼저 묻는다.
            #   크레딧이 새던 자리는 계산이 아니라 **왕복**이었다: 사람이 이상한 숫자를
            #   보고 앱이 이유를 못 말해서 클로드에게 묻고, 클로드는 이미 디스크에 있는
            #   사실을 다시 조립했다. 이 길은 그 사실을 바로 돌려준다.
            #   읽기 전용이다 — 물어봤을 뿐인데 값이 바뀌는 일은 없다.
            q = (parse_qs(urlsplit(self.path).query).get("q", [""])[0] or "").strip()
            if not q:
                return self._send(400, {"ok": False, "error": "질문이 비었습니다"})
            try:
                import local_ai
                return self._send(200, {"ok": True, **local_ai.ask(q)})
            except Exception as exc:
                # 답변기가 깨져도 **화면은 살아 있어야 한다.** 여기서 500 을 주면
                # 사람은 앱을 못 믿고 결국 클로드에게 묻는다(없애려던 그 왕복이다).
                return self._send(200, {
                    "ok": False, "질문": q, "답함": False, "분류": None, "확신": "없음",
                    "답": "답변기가 답하지 못했습니다: %s" % str(exc)[:140],
                    "다음": "아래 문구를 클로드에게 붙여넣으십시오.",
                    "근거": "", "클로드문구": q})
        if p == "/api/calendar":
            return self._send(200, get_calendar())
        if p == "/api/suggest":
            # 입력 자동완성. 원장에 이미 있는 값만 돌려준다(문턱은 /api/works 와 같다).
            qs = parse_qs(urlsplit(self.path).query)   # 함수 안 import 금지(윗 주석)
            g = lambda k: (qs.get(k, [""])[0] or "").strip()
            return self._send(200, suggest_values(g("f"), g("q"), g("n") or 40))
        if p == "/api/project-history":
            # 현장 한 곳의 과거 내력 + 지금 현황 + 예측. 캘린더·업무 화면의 '이력' 창이 쓴다.
            # ★ 문턱은 `/api/calendar`·`/api/works` 와 **같다** — 같은 원장 행을 다시
            #   묶어 보여 줄 뿐이라 여기만 더 열거나 더 잠그면 화면마다 말이 달라진다.
            qs = parse_qs(urlsplit(self.path).query)   # 함수 안 import 금지(윗 주석)
            g = lambda k: (qs.get(k, [""])[0] or "").strip()
            return self._send(200, project_history(g("camp"), g("pj")))
        if p == "/api/erpdocs":
            return self._send(200, get_erpdocs())
        if p == "/api/revenue":
            return self._send(200, get_revenue())
        if p == "/api/checks":
            return self._send(200, get_checks())
        if p == "/api/orgchart":
            # 조직도 실시간 상태. 인증(쿠키/PIN)이 없으면 여기 못 오고, 그때
            # 화면은 오프라인 스냅샷으로 물러선다(index.html ORG_SNAP).
            try:
                return self._send(200, get_orgchart())
            except Exception as exc:
                return self._send(200, {"_live": False,
                                        "gen": "조직도 상태를 읽지 못했습니다",
                                        "error": str(exc)[:200], "zones": []})
        if p == "/api/flow-stages":
            # 돌발AS·정기점검 **단계 정의** (2026-08-10 지시). 화면이 단계 낱말을
            # 스스로 적지 않고 여기서 받아 간다 — 두 곳에 적으면 언젠가 갈린다([162]).
            import work_flow
            try:
                return self._send(200, {"ok": True,
                                        **work_flow.definition("refresh" in self.path)})
            except Exception as exc:
                return self._send(200, {"ok": False, "error": str(exc)[:200], "갈래": {}})
        if p == "/api/staff/deleted-works":
            # 삭제된 건 목록 — 되살리기는 **사람이 볼 수 있는 자리**에 있어야 한다.
            actor_session = self._actor()
            if actor_session.get("role") == "staff":
                actor_slug = str(actor_session.get("staff_slug") or "")
            elif actor_session.get("role") == "admin":
                actor_slug = "admin"
            else:
                return self._send(403, {"ok": False, "error": "업무센터 권한이 필요합니다"})
            rq = parse_qs(urlsplit(self.path).query)
            cat = (rq.get("category") or [""])[0].strip()
            try:
                return self._send(200, deleted_staff_works(actor_slug, cat or None))
            except PermissionError as e:
                return self._send(403, {"ok": False, "error": str(e)[:300]})
            except Exception as e:
                # 못 읽은 것을 '삭제된 건 없음'이라 하지 않는다([169]).
                return self._send(200, {"ok": False, "error": str(e)[:300],
                                        "rows": [], "count": None})
        if p == "/api/records":
            # 자료창고(datalake record) 조회 — 읽기 전용 ([24] 앱 화면 노출)
            #   qs 는 이 분기에서 직접 판다 — 위 분기들의 qs 는 그 분기 안에서만 산다
            rq = parse_qs(urlsplit(self.path).query)
            return self._send(200, datalake_records(rq))
        if p == "/api/reports":
            return self._send(200, {"reports": latest_reports()})
        if p == "/api/tasklog":
            return self._send(200, {"busy": runner["busy"], "task": runner["task"],
                                    "log": list(runner["log"])[-300:],
                                    "last": last_runs(merge_auto=True)})
        return self._send(404, {"error": "not found"})

    def do_POST(self):
        p = self.path.split("?")[0]
        ip = self.client_address[0]
        if p == "/api/tech/report":
            # 기사 완료 보고. 제 목록에 있는 건에만 찍을 수 있다(tech_report 가 확인).
            actor = self._actor()
            tslug = str(actor.get("staff_slug") or "")
            if str(actor.get("role") or "") != "tech" or tslug not in AS_TECH_CENTERS:
                return self._send(401, {"ok": False, "error": "기사 링크로 열어 주세요"})
            ln = int(self.headers.get("Content-Length", 0) or 0)
            if ln > 8192:
                return self._send(413, {"ok": False, "error": "내용이 너무 깁니다"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
            except Exception:
                return self._send(400, {"ok": False, "error": "형식이 올바르지 않습니다"})
            out = tech_report(tslug, body.get("id"), body.get("완료일"), body.get("메모"))
            return self._send(200 if out.get("ok") else 400, out)
        if p == "/api/login":
            if _locked(ip):
                return self._send(429, {"ok": False, "error": "시도 초과 — 10분 후 다시"})
            ln = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(ln) or b"{}")
            staff_slug = str(body.get("staff_slug") or "").strip()
            if staff_slug and staff_slug not in STAFF_CENTERS:
                return self._send(400, {"ok": False, "error": "등록되지 않은 업무센터입니다"})
            ok = verify_pin(body.get("pin", ""), staff_slug)
            (_ok_login if ok else _fail)(ip)
            if not ok:
                return self._send(401, {"ok": False})
            try:
                token, session = create_auth_session(staff_slug)
                cookie = auth_cookie(token)
                return self._send(200, {
                    "ok": True,
                    "role": session["role"],
                    "staff_slug": session["staff_slug"],
                }, headers={"Set-Cookie": cookie})
            except ValueError as exc:
                return self._send(400, {"ok": False, "error": str(exc)})
        if p == "/api/band_dump":
            return self._band_dump()
        if p == "/api/collect_report":
            # 크롬 전용 수집 되보고.  **PIN 게이트 앞이다** — 보내는 쪽은 band.us 탭이라
            # 이 서버의 인증 쿠키를 못 싣는다(다른 출처다).  뒤에 두면 401 만 쌓이고
            # 감시자는 '한 번도 안 옴'이라 말한다 — 스크립트는 멀쩡히 도는데도([169]).
            # 들어오는 것은 상태 한 줄뿐이고 밴드번호 모양까지 검사하므로 문을 연다.
            ln = int(self.headers.get("Content-Length", 0) or 0)
            if ln <= 0 or ln > 20_000:
                return self._send(400, {"ok": False, "error": "되보고 형식 오류"})
            try:
                rec = json.loads(self.rfile.read(ln) or b"{}")
            except Exception as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:160]})
            try:
                return self._send(200, {"ok": True, **collect_report_save(rec)})
            except ValueError as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:160]})
            except Exception as exc:
                # 못 적었으면 못 적었다고 말한다.  200 을 주면 스크립트는 보고가
                # 도착한 줄 알고, 감시자는 빈 파일을 보고 '안 옴'이라 한다.
                return self._send(500, {"ok": False, "error": str(exc)[:160]})
        if not self._auth():
            return self._send(401, {"error": "PIN"})
        if p == "/api/automation/kakao-upload":
            actor = self._actor()
            allowed = actor.get("role") == "admin" or (
                actor.get("role") == "staff"
                and str(actor.get("staff_slug") or "") == "ryu-jiyeong"
            )
            if not allowed:
                return self._send(403, {
                    "ok": False,
                    "error": "카카오톡 원본은 관리자 또는 류지영 업무센터에서 등록합니다",
                })
            ln = int(self.headers.get("Content-Length", 0) or 0)
            if ln <= 0 or ln > 31_000_000:
                # ★ 거절도 알린다 — 화면에만 뜨고 사라지면 형님은 그 파일이 올라온
                #   줄 알고, 올린 사람은 실패한 줄 모른 채 넘어간다(SPEC ③).
                _notify_upload_rejected("카톡 원본", "", self._actor(),
                                        "카카오톡 파일은 30MB 이하여야 합니다")
                return self._send(413, {"ok": False, "error": "카카오톡 파일은 30MB 이하여야 합니다"})
            filename = ""
            try:
                _fields, files = multipart_parts(
                    self.headers.get("Content-Type", ""), self.rfile.read(ln)
                )
                upload = files.get("kakao_file") or files.get("file")
                if not upload:
                    _notify_upload_rejected("카톡 원본", "", self._actor(),
                                            "카카오톡 .txt 파일이 선택되지 않았습니다")
                    return self._send(400, {"ok": False, "error": "카카오톡 .txt 파일을 선택해 주세요"})
                filename = str(upload.get("filename") or "KakaoTalk_upload.txt")
                import automation_pipeline
                saved = automation_pipeline.submit_kakao_file(
                    filename,
                    upload.get("data") or b"",
                )
                started, message = start_task("automation")
                queued = False if started else defer_task_until_free("automation")
                _notify_upload_received(
                    "카톡 원본", filename, self._actor(),
                    started=started, queued=queued,
                    duplicate=bool(saved.get("duplicate")),
                    task_key="automation")
                return self._send(200, {
                    "ok": True,
                    "saved": saved,
                    "auto_started": started,
                    "auto_queued": queued,
                    "msg": message,
                })
            except ValueError as exc:
                _notify_upload_rejected("카톡 원본", filename, self._actor(), str(exc)[:200])
                return self._send(400, {"ok": False, "error": str(exc)[:260]})
            except Exception as exc:
                _notify_upload_rejected("카톡 원본", filename, self._actor(), str(exc)[:200])
                return self._send(500, {"ok": False, "error": str(exc)[:300]})
        if p == "/api/notifications/ack":
            # 확인한 알림을 내린다. **내가 받는 알림만** 내려간다 — 남의 경보를
            # 대신 지우면 그 사람은 영영 그 사실을 모른다.
            ln = int(self.headers.get("Content-Length", 0) or 0)
            if ln > 20_000:
                return self._send(413, {"ok": False, "error": "알림 확인 요청이 너무 큽니다"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}") if ln else {}
                if not isinstance(body, dict):
                    raise ValueError("JSON 객체가 필요합니다")
                ids = body.get("ids")
                if ids is not None and not isinstance(ids, list):
                    raise ValueError("ids 는 목록이어야 합니다")
                import notify
                actor = self._actor()
                # 화면은 `/api/notifications` 가 준 `notify:<id>` 를 그대로 돌려준다.
                # 접두사를 안 떼면 한 건도 안 걸리면서 오류도 안 난다([165]).
                want = [str(i)[:64].split("notify:", 1)[-1] for i in (ids or [])][:200]
                n = notify.ack(want, actor.get("role"), actor.get("staff_slug"))
                return self._send(200, {"ok": True, "acked": n})
            except ValueError as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:160]})
            except Exception as exc:
                # 못 내렸으면 못 내렸다고 말한다 — 200 을 주면 화면은 사라진 줄 안다.
                return self._send(500, {"ok": False, "error": str(exc)[:160]})
        if p == "/api/archive/export":
            if not self._require_admin():
                return
            ln = int(self.headers.get("Content-Length", 0) or 0)
            if ln > 50_000:
                return self._send(413, {"ok": False, "error": "보관본 요청이 너무 큽니다"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}") if ln else {}
            except Exception:
                return self._send(400, {"ok": False, "error": "요청 형식이 올바르지 않습니다"})
            if not isinstance(body, dict):
                return self._send(400, {"ok": False, "error": "JSON 객체가 필요합니다"})
            if body.get("dry_run") is False or body.get("publish") or body.get("external"):
                return self._send(400, {"ok": False,
                                        "error": "앱에서는 검증된 로컬 보관본 준비만 허용합니다"})
            try:
                result = _prepare_archive_export(actor=self._actor_name())
                return self._send(200, result)
            except Exception as exc:
                return self._send(500, {"ok": False, "prepared": False,
                                        "error": str(exc)[:300]})
        if p == "/api/auth/change-pin":
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 20_000:
                return self._send(400, {"ok": False, "error": "PIN 변경 요청 형식 오류"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
                actor = self._actor()
                role = str(actor.get("role") or "")
                staff_slug = str(actor.get("staff_slug") or "") if role == "staff" else ""
                if role not in ("admin", "staff"):
                    return self._send(403, {"ok": False, "error": "로그인 역할을 확인할 수 없습니다"})
                current_pin = str(body.get("current_pin") or "")
                new_pin = str(body.get("new_pin") or "")
                if not verify_pin(current_pin, staff_slug):
                    return self._send(400, {"ok": False, "error": "현재 PIN이 올바르지 않습니다"})
                if new_pin == current_pin:
                    return self._send(400, {"ok": False, "error": "새 PIN은 현재 PIN과 달라야 합니다"})
                set_role_pin(new_pin, staff_slug)
                token, session = create_auth_session(staff_slug)
                cookie = auth_cookie(token)
                return self._send(200, {
                    "ok": True,
                    "role": role,
                    "staff_slug": staff_slug,
                    "expires_at": session["expires_at"],
                    "message": "PIN이 변경되었습니다. 현재 기기는 계속 로그인됩니다.",
                }, headers={"Set-Cookie": cookie})
            except ValueError as exc:
                return self._send(400, {"ok": False, "error": str(exc)})
            except Exception as exc:
                return self._send(500, {"ok": False, "error": str(exc)[:180]})
        if p == "/api/ux":
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 100_000:
                return self._send(400, {"ok": False, "error": "UX 기록 형식 오류"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
                events = (body.get("events") or [])[:50]
                import ledger_db
                saved = ledger_db.ux_add(events)
                # ux 표는 90일이 지나면 스스로 지운다(ux_add 의 DELETE). 그래서 오류만은
                # 따로 **덧붙이기만 하는** 보관본에 남긴다 — 지난달 일을 물을 수 있어야
                # '다시 안 나게' 를 말할 수 있다. 보관이 실패해도 UX 수집을 막지 않되,
                # 실패를 성공처럼 적지도 않는다(-1 을 그대로 내보낸다).
                archived = 0
                try:
                    import error_book
                    archived = error_book.archive(events)
                except Exception:
                    archived = -1
                return self._send(200, {"ok": True, "saved": saved, "archived": archived})
            except Exception as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:160]})
        if p == "/api/staff/activity":
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 20_000:
                return self._send(400, {"ok": False, "error": "업무센터 상태 형식 오류"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
                actor_slug = self._require_staff(*STAFF_CENTERS.keys())
                if not actor_slug:
                    return
                if str(body.get("staff_slug") or "") != actor_slug:
                    return self._send(403, {"ok": False, "error": "다른 담당자 상태는 변경할 수 없습니다"})
                state = record_workcenter_activity(body.get("staff_slug"), body.get("event"))
                return self._send(200, {"ok": True, **state})
            except Exception as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:180]})
        if p == "/api/staff/install-shortcut":
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 20_000:
                return self._send(400, {"ok": False, "error": "설치 요청 형식 오류"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
                actor_slug = self._require_staff(*STAFF_CENTERS.keys())
                if not actor_slug:
                    return
                if str(body.get("staff_slug") or "") != actor_slug:
                    return self._send(403, {"ok": False, "error": "다른 담당자 바로가기는 만들 수 없습니다"})
                result = install_staff_shortcut(body.get("staff_slug"))
                return self._send(200, {"ok": True, **result})
            except Exception as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:240]})
        if p == "/api/staff/improvement":
            actor_slug = self._require_staff(*STAFF_CENTERS.keys())
            if not actor_slug:
                return
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 30_000_000:
                return self._send(400, {"ok": False, "error": "개선요청·첨부 용량은 합계 30MB 이하여야 합니다"})
            try:
                fields, files = multipart_parts(self.headers.get("Content-Type", ""),
                                                self.rfile.read(ln))
                fields["staff_slug"] = actor_slug
                ticket = save_workcenter_improvement(fields, files, ip)
                _notify_upload_received(
                    "개선요청", _first_filename(files) or str(fields.get("title") or ""),
                    self._actor(), extra="담당자가 불편한 점을 올렸습니다.")
                return self._send(200, {"ok": True, "ticket": ticket})
            except Exception as exc:
                _notify_upload_rejected("개선요청", "", self._actor(), str(exc)[:200])
                return self._send(400, {"ok": False, "error": str(exc)[:260]})
        if p == "/api/staff/po-upload":
            actor_slug = self._require_staff("oh-jonghyeon")
            if not actor_slug:
                return
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 60_000_000:
                _notify_upload_rejected("PO 첨부", "", self._actor(),
                                        "PO 첨부는 합계 60MB 이하여야 합니다")
                return self._send(400, {"ok": False, "error": "PO 첨부는 합계 60MB 이하여야 합니다"})
            try:
                fields, files = multipart_parts(self.headers.get("Content-Type", ""),
                                                self.rfile.read(ln))
                fields["staff_slug"] = actor_slug
                result = save_staff_po_submission(fields, files, ip)
                _notify_upload_received("PO 첨부", _first_filename(files), self._actor(),
                                        extra="담당자가 PO 자료를 올렸습니다.")
                return self._send(200, {"ok": True, **result})
            except Exception as exc:
                _notify_upload_rejected("PO 첨부", "", self._actor(), str(exc)[:200])
                return self._send(400, {"ok": False, "error": str(exc)[:320]})
        if p in ("/api/remote/request", "/api/remote/deliver", "/api/remote/stock",
                 "/api/remote/edit", "/api/remote/delete", "/api/remote/restore"):
            # 리모컨 관리(2026-08-03 지시, 같은 날 개정): 승인 단계 없이 기록·관리·보고만.
            # 불출·납품은 류지영/오종현 업무센터와 관리자. 3개 한도는 ledger_db 가 강제한다.
            actor = self._actor()
            role = str(actor.get("role") or "")
            slug = str(actor.get("staff_slug") or "")
            allowed = role == "admin" or (
                role == "staff" and slug in ("ryu-jiyeong", "oh-jonghyeon"))
            if not allowed:
                return self._send(403, {"ok": False,
                                        "error": "리모컨 관리는 류지영·오종현 업무센터 또는 관리자만"})
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 20_000:
                return self._send(400, {"ok": False, "error": "리모컨 요청 형식 오류"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
                import ledger_db
                who = "관리자" if role == "admin" else STAFF_CENTERS[slug]["name"]
                if p == "/api/remote/stock":
                    # 지점 재고 등록(2026-08-03): 입고(add ±N) 또는 실사(set 절대값)
                    after = ledger_db.remote_stock_adjust(
                        body.get("branch"), body.get("qty"),
                        body.get("mode") or "add", body.get("note") or "", who,
                        version=body.get("version") or "")
                    return self._send(200, {"ok": True, "stock": after})
                if p == "/api/remote/request":
                    # 공지(2026-08-04): 불출 일자·투입 예정 캠프명까지 기록한다.
                    # 2026-08-06: 버전·지사 불출자 이름도 함께 남긴다.
                    rid = ledger_db.remote_request(
                        body.get("branch"), body.get("technician"), body.get("qty"),
                        who, body.get("note") or "",
                        body.get("issued_on") or "", body.get("camp") or "",
                        version=body.get("version") or "",
                        issuer=body.get("issuer") or "")
                    return self._send(200, {"ok": True, "id": rid, "status": "불출완료"})
                # 고치기·지우기·되돌리기(2026-08-06 지시). 수량이 서로 물려 있어
                # ledger_db 가 전후 상태를 비교하고, 지운 내용은 원장에 남아 복구된다.
                if p == "/api/remote/edit":
                    r = ledger_db.remote_edit(
                        body.get("kind"), body.get("id"), body.get("fields") or {},
                        edited_by=who, force=bool(body.get("force")),
                        reason=body.get("reason") or "")
                    return self._send(200, {"ok": True, **r})
                if p == "/api/remote/delete":
                    r = ledger_db.remote_delete(
                        body.get("kind"), body.get("id"), deleted_by=who,
                        force=bool(body.get("force")),
                        reason=body.get("reason") or "")
                    return self._send(200, {"ok": True, **r})
                if p == "/api/remote/restore":
                    row = ledger_db.remote_restore(body.get("audit_id"), actor=who)
                    return self._send(200, {"ok": True, "row": row})
                # 샘플발송·개인지급은 프로젝트가 없다(2026-08-11 지시) — 프로젝트NO·캠프명이
                # 둘 다 비면 지급 사유를 필수로 받아 감사 가능하게 남긴다(받는 사람 = technician).
                # 스키마는 그대로다: 사유는 납품처(camp) 표기와 note 꼬리표로 남는다.
                proj = str(body.get("project") or "").strip()
                camp = str(body.get("camp") or "").strip()
                note = str(body.get("note") or "")
                if not proj and not camp:
                    why = str(body.get("grant_reason") or "").strip()
                    if why not in ("샘플발송", "개인지급", "기타"):
                        return self._send(400, {"ok": False, "error":
                            "프로젝트NO·캠프명이 없으면 지급 사유(샘플발송·개인지급·기타)를 골라야 합니다"})
                    # 기타는 자유 사유를 함께 받는다(2026-08-11 류지영 요청) — 선택이지
                    # 필수는 아니다. 적으면 납품처 표기와 note 에 그대로 남는다.
                    why_text = str(body.get("grant_reason_text") or "").strip()
                    if why == "기타" and why_text:
                        camp = f"(기타: {why_text[:40]})"
                        note = (note + " " if note else "") + f"[지급사유: 기타 — {why_text[:120]}]"
                    else:
                        camp = f"({why})"
                        note = (note + " " if note else "") + f"[지급사유: {why}]"
                rid = ledger_db.remote_deliver(
                    body.get("technician"), proj,
                    camp, body.get("qty"),
                    body.get("delivered_on") or "", note, who,
                    kind=body.get("kind") or "납품",
                    version=body.get("version") or "")
                return self._send(200, {"ok": True, "id": rid})
            except ValueError as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:260]})
            except Exception as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:260]})
        if p == "/api/staff/receipt-upload":
            # 관리자(업무센터 탭) 또는 오종현 담당자 페이지에서만. 자료는 항상 오종현
            # 소유로 보관된다 — 입금 원천의 관리 책임이 그쪽이기 때문(2026-07-31).
            actor = self._actor()
            allowed = actor.get("role") == "admin" or (
                actor.get("role") == "staff"
                and str(actor.get("staff_slug") or "") == "oh-jonghyeon")
            if not allowed:
                return self._send(403, {"ok": False,
                                        "error": "입금 자료는 관리자 또는 오종현 업무센터에서만 등록합니다"})
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 60_000_000:
                _notify_upload_rejected("입금내역", "", self._actor(),
                                        "입금 자료는 합계 60MB 이하여야 합니다")
                return self._send(400, {"ok": False, "error": "입금 자료는 합계 60MB 이하여야 합니다"})
            try:
                fields, files = multipart_parts(self.headers.get("Content-Type", ""),
                                                self.rfile.read(ln))
                result = save_staff_receipt_submission(fields, files, ip)
                _notify_upload_received("입금내역", _first_filename(files), self._actor(),
                                        extra="입금 자료가 들어왔습니다 — 대조는 입금 대조 회차가 합니다.")
                return self._send(200, {"ok": True, **result})
            except Exception as exc:
                _notify_upload_rejected("입금내역", "", self._actor(), str(exc)[:200])
                return self._send(400, {"ok": False, "error": str(exc)[:320]})
        if p == "/api/staff/work-log-upload":
            actor_slug = self._require_staff("ryu-jiyeong")
            if not actor_slug:
                return
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 60_000_000:
                _notify_upload_rejected("대표보고 일지", "", self._actor(),
                                        "대표보고 일지는 60MB 이하여야 합니다")
                return self._send(400, {"ok": False, "error": "대표보고 일지는 60MB 이하여야 합니다"})
            try:
                fields, files = multipart_parts(self.headers.get("Content-Type", ""),
                                                self.rfile.read(ln))
                fields["staff_slug"] = actor_slug
                result = save_staff_work_log_submission(fields, files, ip)
                _notify_upload_received("대표보고 일지", _first_filename(files), self._actor())
                return self._send(200, {"ok": True, **result})
            except Exception as exc:
                _notify_upload_rejected("대표보고 일지", "", self._actor(), str(exc)[:200])
                return self._send(400, {"ok": False, "error": str(exc)[:320]})
        if p == "/api/staff/new-job":
            # ★ 관리자도 등록한다 (2026-08-10 지시 — 돌발AS·정기점검 화면에서 바로).
            #   전에는 류지영 업무센터로 로그인했을 때만 열려 있어, 관리자가 앱에서
            #   **새 건을 만들 길 자체가 없었다.** 기사(tech)는 위 `_auth()` 가 막는다.
            actor = self._actor()
            actor_slug = str(actor.get("staff_slug") or "")
            is_admin = actor.get("role") == "admin"
            if not (is_admin or (actor.get("role") == "staff"
                                 and actor_slug == "ryu-jiyeong")):
                return self._send(403, {
                    "ok": False,
                    "error": "신규 업무는 관리자 또는 류지영 업무센터에서 등록합니다"})
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 30_000_000:
                return self._send(400, {"ok": False, "error": "신규업무·첨부 용량은 합계 30MB 이하여야 합니다"})
            try:
                fields, files = multipart_parts(self.headers.get("Content-Type", ""),
                                                self.rfile.read(ln))
                # 등록자는 **서버가 정한다** — 화면이 보낸 이름을 믿으면 감사기록이
                # 누가 했는지를 못 말한다.
                fields["staff_slug"] = actor_slug
                fields["submitter"] = (STAFF_CENTERS[actor_slug]["name"]
                                       if actor_slug in STAFF_CENTERS else "관리자")
                result = save_new_workcenter_job(
                    fields, files, ip,
                    actor=self._actor_name(),
                    idempotency_key=self._idempotency_key(),
                )
                return self._send(200 if result.get("ok") else 503, result)
            except Exception as exc:
                return self._send(400, {"ok": False, "error": str(exc)[:300]})
        if p == "/api/export_xlsx":
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 2_000_000:
                return self._send(400, {"ok": False, "error": "내보낼 목록 크기가 올바르지 않습니다"})
            try:
                payload = json.loads(self.rfile.read(ln) or b"{}")
                data, _title = rows_xlsx(payload)
                return self._send(
                    200, data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                return self._send(500, {"ok": False, "error": f"엑셀 생성 실패: {str(e)[:160]}"})
        if p == "/api/flow":
            # 흐름 저장·되돌리기 (2026-08-07 지시). 통째로 받는다 — 순서 바꾸기·지우기가
            # 섞이면 부분 갱신은 어긋나고, 단계는 많아야 서른 개라 통째가 안전하다.
            if not self._require_admin():
                return
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 200_000:
                return self._send(400, {"ok": False, "error": "저장 내용 크기가 올바르지 않습니다"})
            b = json.loads(self.rfile.read(ln) or b"{}")
            import ledger_db
            who = str(b.get("저장자") or "앱 사용자")[:40]
            key = str(b.get("chart") or "").strip() or "as_legacy"
            try:
                n = (ledger_db.flow_restore(who, key) if b.get("되돌리기")
                     else ledger_db.flow_save(b.get("steps") or [], who, key))
                # 문제점·개선점(2026-08-11)도 같은 저장으로 온다 — 키가 있을 때만 바꾼다
                # (없는 요청이 빈 목록으로 지우면 화면 하나가 남의 메모를 지운다).
                if "notes" in b and not b.get("되돌리기"):
                    ledger_db.flow_notes_save(key, b.get("notes") or [], who)
            except ValueError as e:
                return self._send(400, {"ok": False, "error": str(e)})
            return self._send(200, {"ok": True, "단계수": n, "chart": key,
                                    "steps": ledger_db.flow_steps(key),
                                    "notes": ledger_db.flow_notes(key)})
        if p == "/api/policy":
            if not self._require_admin():
                return
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 100_000:
                return self._send(400, {"ok": False, "error": "저장 내용 크기가 올바르지 않습니다"})
            b = json.loads(self.rfile.read(ln) or b"{}")
            key = str(b.get("기준") or "").strip()
            value = str(b.get("확정내용") or "").strip()
            if not key or not value:
                return self._send(400, {"ok": False, "error": "기준과 확정 내용을 입력하세요"})
            state = {
                "상태": "확정", "확정내용": value,
                "저장자": str(b.get("저장자") or "앱 사용자")[:40],
                "저장일시": datetime.now().isoformat(timespec="seconds"),
            }
            save_policy_state(key, state)
            return self._send(200, {"ok": True, **state})
        if p == "/api/ryu/upload":
            actor_slug = self._require_staff("ryu-jiyeong")
            if not actor_slug:
                return
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 45_000_000:
                _notify_upload_rejected("업무센터 첨부", "", self._actor(),
                                        "첨부 용량은 합계 45MB 이하여야 합니다")
                return self._send(400, {"ok": False, "error": "첨부 용량은 합계 45MB 이하여야 합니다"})
            try:
                fields, files = multipart_parts(self.headers.get("Content-Type", ""),
                                                self.rfile.read(ln))
                fields["submitter"] = STAFF_CENTERS[actor_slug]["name"]
                result = save_ryu_upload(fields, files)
                started, msg = start_task("kakao")
                queued = False if started else defer_task_until_free("kakao")
                _notify_upload_received("업무센터 첨부", _first_filename(files), self._actor(),
                                        started=started, queued=queued)
                return self._send(200, {"ok": True, "saved": result,
                                        "auto_check_started": started,
                                        "auto_check_queued": queued, "msg": msg})
            except Exception as e:
                _notify_upload_rejected("업무센터 첨부", "", self._actor(), str(e)[:200])
                return self._send(400, {"ok": False, "error": str(e)[:260]})
        if p in ("/api/staff/entry", "/api/ryu/entry"):
            actor_session = self._actor()
            if actor_session.get("role") == "staff":
                actor_slug = str(actor_session.get("staff_slug") or "")
                if actor_slug not in STAFF_CENTERS:
                    return self._send(403, {"ok": False, "error": "등록되지 않은 업무센터입니다"})
            elif actor_session.get("role") == "admin":
                actor_slug = "admin"
            else:
                return self._send(403, {"ok": False, "error": "업무센터 권한이 필요합니다"})
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 30_000_000:
                return self._send(400, {"ok": False, "error": "입력·첨부 용량은 합계 30MB 이하여야 합니다"})
            try:
                content_type = self.headers.get("Content-Type", "")
                raw = self.rfile.read(ln)
                evidence_name = ""
                if "multipart/form-data" in content_type:
                    fields, files = multipart_parts(content_type, raw)
                    category = str(fields.get("category") or "").strip()
                    target_category = str(fields.get("target_category") or "").strip()
                    actual_category = target_category if category == "issue" else category
                    if actual_category not in RYU_ENTRY_CONFIG:
                        raise ValueError("수정할 업무 카테고리를 확인할 수 없습니다")
                    allowed = _staff_allowed_fields(actor_slug, actual_category)
                    values = {name: fields.get(name) for name in allowed
                              if name in fields and str(fields.get(name) or "").strip() != ""}
                    clear_fields = [name[len("clear__"):] for name, value in fields.items()
                                    if name.startswith("clear__") and str(value) == "1"]
                    body = {**fields, "key": fields.get("record_key"), "values": values,
                            "clear_fields": clear_fields}
                    upload = files.get("evidence_file")
                    if upload:
                        evidence_name = _save_ryu_evidence(
                            upload, actual_category, body.get("key") or body.get("target_key"),
                            submitter=STAFF_CENTERS.get(actor_slug, {"name": "관리자"})["name"],
                        )
                else:
                    body = json.loads(raw or b"{}")
                # 제출 본문의 직원명·slug는 신뢰하지 않는다. actor는 인증 쿠키에서만 온다.
                body.pop("staff_slug", None)
                body.pop("submitter", None)
                body["idempotency_key"] = self._idempotency_key(body)
                if evidence_name:
                    note = str(body.get("reason") or body.get("survey_note") or "").strip()
                    body["reason"] = (note + (" · " if note else "") + "근거 " + evidence_name)
                result = save_staff_entry(
                    actor_slug, body, actor=self._actor_name(),
                )
                return self._send(200, {**result, "evidence": evidence_name})
            except PermissionError as e:
                return self._send(403, {"ok": False, "error": str(e)[:300]})
            except Exception as e:
                if type(e).__name__ in ("VersionConflict", "IdempotencyConflict"):
                    return self._send(409, {"ok": False, "error": str(e)[:300],
                                            "conflict": True})
                return self._send(400, {"ok": False, "error": str(e)[:300]})
        if p in ("/api/staff/work-delete", "/api/staff/work-restore"):
            # 삭제·되살리기는 /api/staff/entry 와 **같은 문**을 지난다 — 권한·멱등키·
            # 낙관잠금·409 매핑이 갈리면 한쪽만 고쳐진다([162]).
            actor_session = self._actor()
            if actor_session.get("role") == "staff":
                actor_slug = str(actor_session.get("staff_slug") or "")
                if actor_slug not in STAFF_CENTERS:
                    return self._send(403, {"ok": False, "error": "등록되지 않은 업무센터입니다"})
            elif actor_session.get("role") == "admin":
                actor_slug = "admin"
            else:
                return self._send(403, {"ok": False, "error": "업무센터 권한이 필요합니다"})
            ln = int(self.headers.get("Content-Length", 0))
            if ln <= 0 or ln > 1_000_000:
                return self._send(400, {"ok": False, "error": "요청 본문을 확인해 주세요"})
            try:
                body = json.loads(self.rfile.read(ln) or b"{}")
                body.pop("staff_slug", None)
                body["idempotency_key"] = self._idempotency_key(body)
                handler = (delete_staff_work if p.endswith("work-delete")
                           else restore_staff_work)
                result = handler(actor_slug, body, actor=self._actor_name())
                # 청구자료 확인이 필요하면 200 으로 되묻는다(실패가 아니다).
                return self._send(200, result)
            except PermissionError as e:
                return self._send(403, {"ok": False, "error": str(e)[:300]})
            except Exception as e:
                if type(e).__name__ in ("VersionConflict", "IdempotencyConflict"):
                    return self._send(409, {"ok": False, "error": str(e)[:300],
                                            "conflict": True})
                return self._send(400, {"ok": False, "error": str(e)[:300]})
        if p == "/api/enqueue":
            if not self._require_admin():
                return
            # 폰이 **PC 꺼진 동안 예약해 둔** 프로젝트 코드를 받아 원장에 등록한다.
            # 오프라인 앱이 PC가 살아난 걸 확인하는 즉시 스스로 보낸다(사람 개입 없음).
            ln = int(self.headers.get("Content-Length", 0))
            codes = (json.loads(self.rfile.read(ln) or b"{}").get("codes") or [])[:50]
            if DEMO:
                return self._send(200, {"ok": True, "applied": 0, "msg": "데모"})
            result = enqueue_codes(codes)
            return self._send(200 if result.get("ok") else 503, result)
        m = re.match(r"^/api/run/(\w+)$", p)
        if m:
            if not self._require_admin():
                return
            ok, msg = start_task(m.group(1))
            return self._send(200 if ok else 409, {"ok": ok, "msg": msg})
        if p == "/api/open":
            if not self._require_admin():
                return
            # 워크벤치 대체: 관리대장·폴더 열기. 원격(터널)에서는 의미가 없고 위험하므로
            # 서버가 도는 PC에서 접속했을 때만 허용한다.
            if ip not in ("127.0.0.1", "::1", "localhost"):
                return self._send(403, {"ok": False, "error": "이 기능은 사무실 PC에서만 됩니다"})
            ln = int(self.headers.get("Content-Length", 0))
            what = json.loads(self.rfile.read(ln) or b"{}").get("what", "")
            try:
                from ecount_reconcile import load_config, resolve_master
                master = resolve_master(load_config()["reconcile"]["master_xlsx"])
            except Exception:
                master = ""
            targets = {"master": master, "master_dir": os.path.dirname(master),
                       "inbox": os.path.join(ROOT, "inbox"),
                       "kakao": os.path.join(ROOT, "kakao", "inbox"),
                       "band_docs": os.path.join(ROOT, "band", "docs_inbox"),
                       "band_cache": os.path.join(ROOT, "band", "cache"),
                       "reports": os.path.join(ROOT, "reports")}
            path = targets.get(what)
            # ★ 원본 색인에서 고른 파일도 연다(2026-08-05 "클릭 한번으로 찾고 열 수 있게").
            #   임의 경로 열기는 위험하므로 **색인에 있는 경로만** 허용한다(화이트리스트).
            if not path and what:
                try:
                    with open(os.path.join(ROOT, "reports", "원본색인.json"),
                              encoding="utf-8") as fh:
                        allow = {r.get("path") for r in (json.load(fh).get("rows") or [])}
                except Exception:
                    allow = set()
                if what in allow:
                    path = what
            if not path or not os.path.exists(path):
                return self._send(404, {"ok": False, "error": f"경로 없음: {what}"})
            try:
                os.startfile(path)
                return self._send(200, {"ok": True, "opened": os.path.basename(path) or path})
            except Exception as e:
                return self._send(500, {"ok": False, "error": str(e)[:200]})
        if p == "/api/set_dates":
            # 보고일·집계기준일은 앱 DB에 즉시 커밋한다. B3·B4는 보관본 렌더 대상일 뿐
            # Excel에서 다시 앱 DB로 역수입하지 않는다.
            ln = int(self.headers.get("Content-Length", 0))
            b = json.loads(self.rfile.read(ln) or b"{}")
            items = []
            for cell, key, label in (("B3", "보고일", "보고일"), ("B4", "집계기준일", "집계기준일")):
                v = str(b.get(key, "")).strip()
                if v:
                    if not re.match(r"^\d{4}-\d{2}-\d{2}$", v):
                        return self._send(400, {"ok": False, "error": f"{label} 형식 오류(YYYY-MM-DD)"})
                    items.append({"sheet": "00_대시보드", "cell": cell, "key": cell, "key_col": "-",
                                  "col": label, "value": v, "vtype": "date",
                                  "evidence": f"앱 기준일 설정({ip})", "only_if_empty": False})
            if not items:
                return self._send(400, {"ok": False, "error": "날짜 없음"})
            if DEMO:
                return self._send(200, {"ok": True, "demo": True})
            queued = enqueue_for_scheduled_apply(
                items, source="app-dates", actor=self._actor_name(),
                idempotency_key=self._idempotency_key(b),
            )
            if not queued.get("ok"):
                return self._send(503, queued)
            # 예전 '즉시 반영'은 Excel 정본 시절의 문맥이다. 지금은 앱 DB 커밋 뒤
            # 로컬 스냅샷·명령계획을 준비하고, 렌더·검증 전에는 생성 완료라 하지 않는다.
            if b.get("apply"):
                try:
                    archive = _prepare_archive_export(actor=self._actor_name())
                    return self._send(200, {**queued, "archive": archive})
                except Exception as exc:
                    return self._send(500, {**queued, "ok": False, "committed": True,
                                            "archive": {"ok": False,
                                                        "error": str(exc)[:240]}})
            return self._send(200, queued)
        if p == "/api/input":
            # 구버전 화면·오프라인 대기열 호환 경로. 실제 저장은 공용 직원 API와 같은
            # 역할×열 허용표·낙관잠금·감사로그를 거친다. 임의 열은 더 이상 만들 수 없다.
            ln = int(self.headers.get("Content-Length", 0))
            b = json.loads(self.rfile.read(ln) or b"{}")
            category = {"02_돌발AS접수": "as", "04_정기점검": "pm",
                        "06_거래서류청구수금": "settle"}.get(str(b.get("sheet") or ""))
            if not category or not b.get("key") or not b.get("col") or b.get("value") in (None, ""):
                return self._send(400, {"ok": False, "error": "sheet/key/col/value 필요"})
            actor_session = self._actor()
            actor_slug = (str(actor_session.get("staff_slug") or "")
                          if actor_session.get("role") == "staff" else
                          "admin" if actor_session.get("role") == "admin" else "")
            if not actor_slug:
                return self._send(403, {"ok": False, "error": "업무센터 권한이 필요합니다"})
            try:
                from app_store import default_store
                store = default_store()
                _row, current = _staff_store_row(store, category, str(b["key"]))
                cfg = RYU_ENTRY_CONFIG[category]
                from app_store import SHEET_SPECS
                sheet_spec = SHEET_SPECS.get(cfg["sheet"], {})
                core_by_column = {v: k for k, v in sheet_spec.items()
                                  if k in ("public_id", "project_no", "camp_name", "status")}
                core = core_by_column.get(str(b["col"]))
                old = current.get(core) if core else (current.get("fields") or {}).get(str(b["col"]))
                if old not in (None, "") and b.get("overwrite") is not True:
                    return self._send(200, {"ok": True, "committed": True,
                                            "action": "no_change", "changed": [],
                                            "skipped": [str(b["col"])],
                                            "record_version": int(current["record_version"]),
                                            "msg": "기존 값 보호 · 변경 없음"})
                body = {
                    "category": category, "key": str(b["key"]),
                    "record_version": int(b.get("record_version") or current["record_version"]),
                    "values": {str(b["col"]): b["value"]},
                    "reason": str(b.get("reason") or "").strip(),
                    "idempotency_key": self._idempotency_key(b),
                }
                result = save_staff_entry(
                    actor_slug, body, store=store, actor=self._actor_name(),
                )
                return self._send(200, result)
            except PermissionError as exc:
                return self._send(403, {"ok": False, "error": str(exc)[:300]})
            except Exception as exc:
                code = 409 if type(exc).__name__ in ("VersionConflict", "IdempotencyConflict") else 400
                return self._send(code, {"ok": False, "error": str(exc)[:300],
                                         "conflict": code == 409})
        return self._send(404, {"error": "not found"})

    def _band_dump(self):
        """브라우저 수집기가 밴드 게시글 원본을 직접 전송(no-cors POST) — PIN 쿼리로 보호"""
        from urllib.parse import parse_qs, urlparse
        q = parse_qs(urlparse(self.path).query)
        if not verify_pin((q.get("pin") or [""])[0]):
            return self._send(401, {"ok": False})
        ln = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(ln)
        try:
            d = json.loads(raw.decode("utf-8"))
            band = re.sub(r"\D", "", str(d.get("band", ""))) or "unknown"
            os.makedirs(os.path.join(ROOT, "band", "cache"), exist_ok=True)
            path = os.path.join(ROOT, "band", "cache", f"dump_{band}.json")
            open(path, "w", encoding="utf-8").write(json.dumps(d, ensure_ascii=False))
            started, _message = start_task("automation")
            if not started:
                defer_task_until_free("automation")
            return self._send(200, {
                "ok": True,
                "saved": len(d.get("posts", {})),
                "auto_started": started,
            })
        except Exception as e:
            return self._send(400, {"ok": False, "error": str(e)[:200]})

    def log_message(self, *a):
        pass


def lan_ip():
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


PUBLISH_EVERY = 3 * 3600      # 폰이 보는 사본을 몇 초마다 새로 올릴지


def publish_loop():
    """PC가 켜져 있는 동안 **주기적으로** 폰용 사본을 올린다.

    ★ PC가 꺼져도 폰이 쓰이려면 사본이 최신이어야 한다. 예전에는 daily_run 이 돌 때만
      올려서, 아침에 한 번 돌리고 저녁에 PC를 끄면 폰은 **아침 숫자**를 보게 됐다.
      그러면 '꺼져도 된다'는 말이 사실이 아니게 된다. 그래서 3시간마다 올린다.
      (사본은 잠겨 있고 60~80KB라 부담이 없다)
    """
    if DEMO:
        return
    time.sleep(120)                        # 기동 직후 혼잡할 때는 피한다
    while True:
        from operation_window import is_input_window
        if is_input_window():
            time.sleep(60)
            continue
        try:
            # 자동 게시도 사람·다른 AI의 수동 게시를 밟지 않게 publish 점유를 강제한다.
            publish_env = {**ENV, "CSOS_AI": "server"}
            from proc_guard import run_tree
            r = run_tree([PY, os.path.join(ROOT, "cloud_publish.py"), "--push"],
                         cwd=ROOT, env=publish_env, timeout=900,
                         drain_timeout=30, output_limit=80_000)
            if r.timed_out:
                runner["log"].append(
                    "[사본 자동 게시] 시간초과 · 자식나무 정리"
                    f" · 잔류 pid={r.stuck_pid or 0} · 다음 주기에 재시도")
            else:
                output = "\n".join(x for x in (r.stdout or "", r.stderr or "") if x)
                tail = [line for line in output.splitlines() if line.strip()][-1:] or [""]
                runner["log"].append(
                    f"[사본 자동 게시] 코드 {r.returncode} · {tail[0][:120]}")
        except Exception as e:
            runner["log"].append(f"[사본 자동 게시] 실패 {type(e).__name__}")
        time.sleep(PUBLISH_EVERY)


class _Server(ThreadingHTTPServer):
    """★ 사고 #16의 진짜 원인 — 윈도우에서 SO_REUSEADDR 이 켜져 있으면 **이미 남이 쓰는
    포트에도 바인드가 성공한다.** 그래서 새 서버가 '시작됨'을 찍고도 요청은 계속 옛
    프로세스가 받아 갔다. 코드를 고쳐도 화면이 안 바뀌는 증상의 정체가 이것이다.
    재사용을 끄면 두 번째 서버는 조용히 뜨는 대신 **에러로 죽는다** — 그게 옳다."""
    allow_reuse_address = False


def main():
    if "--configure-pins" in sys.argv:
        admin_pin = os.environ.get("CSOS_ADMIN_PIN", "").strip()
        staff_pin = os.environ.get("CSOS_STAFF_PIN", "").strip()
        if not admin_pin or not staff_pin:
            print("CSOS_ADMIN_PIN·CSOS_STAFF_PIN 환경변수가 모두 필요합니다.")
            return 2
        set_role_pin(admin_pin)
        for slug in STAFF_CENTERS:
            set_role_pin(staff_pin, slug)
        print("관리자·담당자 PIN 정책을 로컬 해시 설정으로 갱신했습니다.")
        return 0
    # ★ 새 업무센터 하나만 열어 줄 때 쓴다. `--configure-pins` 는 **모든 담당자 PIN을
    #   같은 값으로 덮으므로** 센터를 하나 새로 만들었다고 그것을 돌리면 안 된다.
    #   PIN 은 반드시 **환경변수**로 받는다 — 명령줄 인자는 작업 목록에 그대로 보인다.
    if "--set-staff-pin" in sys.argv:
        i = sys.argv.index("--set-staff-pin") + 1
        slug = sys.argv[i] if i < len(sys.argv) else ""
        pin = os.environ.get("CSOS_STAFF_PIN", "").strip()
        if slug not in STAFF_CENTERS:
            print("등록된 업무센터가 아닙니다:", slug or "(빈 값)")
            print("고를 수 있는 것:", ", ".join(STAFF_CENTERS))
            return 2
        if not pin.isdigit() or not (4 <= len(pin) <= 8):
            print("CSOS_STAFF_PIN 환경변수에 4~8자리 숫자를 넣어 주세요(명령줄에 적지 마세요).")
            return 2
        set_role_pin(pin, slug)
        print(f"{STAFF_CENTERS[slug]['name']} 업무센터 PIN을 새로 정했습니다 "
              f"— 기존에 로그인돼 있던 기기는 다시 로그인해야 합니다.")
        return 0
    icon_sync = sync_installed_app_icons()
    if icon_sync:
        print("  아이콘:", icon_sync)
    try:
        srv = _Server(("0.0.0.0", PORT), H)
    except OSError:
        print(f"★ {PORT} 포트를 이미 다른 앱 서버가 쓰고 있습니다. 새로 뜨지 않았습니다.")
        print("  옛 서버가 계속 응답하므로 **코드를 고쳐도 화면이 안 바뀝니다.**")
        print("  먼저 정리하세요 (PowerShell):")
        print("    Get-CimInstance Win32_Process -Filter \"Name='python.exe'\" |"
              " ? { $_.CommandLine -like '*app_server.py*' } | % { Stop-Process -Id $_.ProcessId -Force }")
        return 1
    threading.Thread(target=publish_loop, daemon=True).start()
    threading.Thread(target=warm_caches, daemon=True).start()
    mode = "데모(합성데이터)" if DEMO else "실서비스"
    print(f"Coupang Service Operations System 앱 서버 [{mode}] 시작")
    print(f"  PC:      http://localhost:{PORT}")
    print(f"  휴대폰:  http://{lan_ip()}:{PORT}   (같은 와이파이)")
    print("  PIN:     역할별 로컬 해시 설정 사용")
    srv.serve_forever()


if __name__ == "__main__":
    main()
