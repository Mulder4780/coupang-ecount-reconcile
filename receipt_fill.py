# -*- coding: utf-8 -*-
"""receipt_fill.py — 입금(수금) 자동입력.

왜: 폰 앱 상세의 '빈 항목 입력' 칸(PO번호·PO발행일·청구일·지급예정일·입금일·입금액) 중
PO 두 칸은 `po_reconcile` 이 이미 자동으로 채운다. **입금일·입금액만 사람이 손으로** 넣고
있었다. 근거 자료(거래처별계정별원장)가 들어오면 사람 손을 거치지 않게 한다
(사용자 지시 2026-07-28: "자료 확인되면 알아서 입력 자동화").

무엇을 채우나 — **입금일·입금액만** 채운다.
  · 근거: 거래처별계정별원장(외상매출금)의 **대변 = 입금**. 차변은 매출 발생이라 건드리지 않는다.
  · 청구일·지급예정일은 **채우지 않는다.** 사내에서 무엇을 '청구일'로 보는지(세금계산서 발행일인지
    아리바 업로드일인지), 지급예정일 산정기준이 무엇인지 확정된 규칙이 없다. 규칙 없이 날짜를
    만들어 넣으면 미수 집계가 통째로 틀어진다(AGENTS.md 절대규칙 10 — 원자료에 없는 값 임의 채움 금지).
    확정되면 여기에 한 줄 추가하면 된다.

어떻게 맞추나 — po_reconcile 과 같은 **양방향 유일 일치** 원칙.
  입금액이 06시트의 세금계산서합계(없으면 거래명세서합계)와 딱 맞고, 그런 후보가 **한 건뿐**이며,
  그 입금건과 금액이 같은 다른 입금도 없을 때만 자동입력한다. 쿠팡 입금은 여러 건을 묶어
  한 번에 넣는 일이 잦아 금액이 겹치면 엉뚱한 정산행에 붙는다 — 겹치면 사람에게 넘긴다.

실행
  python receipt_fill.py            # 미리보기(큐 적재 없음)
  python receipt_fill.py --queue    # ledger_writer 큐 적재(원장 반영은 --apply 가 한다)
"""
import os
import re
import sys
import hashlib
import collections
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")
sys.path.insert(0, BASE_DIR)

AMOUNT_TOL = 1          # 원 단위 반올림 차이만 허용


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(s) if s not in ("", "-", ".") else None
    except ValueError:
        return None


def _day(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.search(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", str(v or ""))
    if not m:
        return None
    try:
        return date(*(int(x) for x in m.groups()))
    except ValueError:
        return None


def parse_receipts(path):
    """거래처별계정별원장 → [{일자, 전표, 적요, 금액}] (대변 = 입금)

    합계/이월 행은 건너뛴다 — '월 계'·'누 계'·'전기이월' 이 금액을 갖고 있어
    그대로 두면 수천만 원짜리 가짜 입금이 하나 생긴다."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out, has_credit = [], False
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hdr_i, idx = None, {}
        for i, r in enumerate(rows[:25]):
            names = {str(c).strip(): j for j, c in enumerate(r) if c is not None}
            if any("적요" in n for n in names) and any("대변" in n for n in names):
                hdr_i = i
                for n, j in names.items():
                    if "일자" in n or "날짜" in n:
                        idx.setdefault("date", j)
                    if "No" in n or "번호" in n:
                        idx.setdefault("slip", j)
                    if "적요" in n:
                        idx["remark"] = j
                    if "대변" in n:
                        idx["credit"] = j
                break
        if hdr_i is None or "credit" not in idx:
            continue
        has_credit = True
        for r in rows[hdr_i + 1:]:
            if r is None or all(c is None for c in r):
                continue
            joined = " ".join(str(c) for c in r if c is not None)
            if re.search(r"(월|누|합)\s*계|이\s*월|소\s*계", joined):
                continue                                  # 합계·이월 행
            amt = _num(r[idx["credit"]])
            if not amt:
                continue
            d = _day(r[idx["date"]]) if idx.get("date") is not None else None
            if d is None:
                d = _day(joined)
            if d is None:
                continue                                  # 날짜 없는 입금은 쓸 수 없다
            out.append({
                "일자": d, "금액": amt,
                "전표": str(r[idx["slip"]] or "") if idx.get("slip") is not None else "",
                "적요": str(r[idx["remark"]] or "") if idx.get("remark") is not None else "",
                "출처": os.path.basename(path),
            })
    wb.close()
    return out, has_credit


# 같은 거래처가 표기만 다르게 적힌다 — '쿠팡로지스틱스'/'쿠팡로지스틱',
# '김진주(위더스)'/'김진주（위더스 )'(전각 괄호). 집계할 때 갈라지면 금액이 둘로 쪼개진다.
def norm_cust(name):
    s = str(name or "")
    # ★ 법인격을 **괄호를 지우기 전에** 뗀다. 괄호부터 지우면 '(주)모벤티스'가 '주모벤티스'가
    #   되어 '주식회사 모벤티스'와 영영 다른 거래처가 된다.
    s = re.sub(r"[（(]\s*[주유재사]\s*[）)]|㈜|㈕|주식회사|유한회사|합자회사", "", s)
    s = re.sub(r"[（）()\s\-·.]", "", s)
    if s.startswith("쿠팡로지스틱"):
        return "쿠팡로지스틱스"
    return s or "(미기재)"


def parse_deposit_list(path):
    """오종현 관리 '26년도 쿠팡 입금내역' → [{일자, 거래처, 금액}]

    머리글(날짜·거래처·입금액)이 몇 번째 행에 있는지는 사람이 제목·빈 줄을 넣는 만큼
    바뀐다(현재 5행). 행 번호를 박아 두면 다음 달에 한 줄 밀리는 순간 0건이 된다
    → **머리글을 찾아서** 시작 행을 정한다."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hdr_i, idx = None, {}
        for i, r in enumerate(rows[:30]):
            names = {str(c).strip(): j for j, c in enumerate(r) if c is not None}
            if any(n in ("날짜", "일자", "입금일") for n in names) and \
               any("입금" in n and "액" in n or n == "금액" for n in names):
                hdr_i = i
                for n, j in names.items():
                    if n in ("날짜", "일자", "입금일"):
                        idx["date"] = j
                    elif "거래처" in n or "업체" in n or "입금자" in n:
                        idx["cust"] = j
                    elif ("입금" in n and "액" in n) or n == "금액":
                        idx["amt"] = j
                break
        # ★ 은행에서 그대로 받은 **거래내역조회** 도 읽는다 (2026-08-07, 김미영 대리 파일).
        #   머리글이 `거래일시 · 출금 · 입금 · 거래후 잔액 · 거래내용 …` 이라 위 규칙에
        #   안 걸린다('입금'은 있는데 '입금액'이 아니고, '날짜'가 아니라 '거래일시'다).
        #   예전 모양(날짜·거래처·입금액)은 사람이 정리한 표였고, 이건 **원본 그대로**다 —
        #   손이 안 닿았으니 오히려 이쪽이 정본에 가깝다. 둘 다 받는다.
        #   · 출금 행은 버린다(우리가 세는 것은 받은 돈이다).
        #   · 거래처 이름은 '상대계좌예금주명'을 먼저 쓴다 — '거래내용'은 은행이 잘라
        #     `쿠팡로지스틱` 처럼 글자가 빠진 채 오는 일이 많다(이 파일에서도 28건).
        if hdr_i is None:
            for i, r in enumerate(rows[:30]):
                names = {str(c).strip(): j for j, c in enumerate(r) if c is not None}
                if "거래일시" in names and "입금" in names:
                    hdr_i = i
                    idx = {"date": names["거래일시"], "amt": names["입금"],
                           "__양식": "은행원본"}
                    for key in ("상대계좌예금주명", "거래내용"):
                        if key in names:
                            idx["cust"] = names[key]
                            break
                    if "출금" in names:
                        idx["out"] = names["출금"]
                    break
        if hdr_i is None or "date" not in idx or "amt" not in idx:
            continue
        for r in rows[hdr_i + 1:]:
            if r is None:
                continue
            d = _day(r[idx["date"]])
            a = _num(r[idx["amt"]])
            if d is None or not a:
                continue
            joined = " ".join(str(c) for c in r if c is not None)
            if re.search(r"(합|총|누)\s*계", joined):
                continue                      # 합계 행을 입금으로 세면 금액이 두 배가 된다
            out.append({
                "일자": d, "금액": a,
                "거래처": norm_cust(r[idx["cust"]] if idx.get("cust") is not None else ""),
                "전표": "", "적요": str(r[idx["cust"]] or "") if idx.get("cust") is not None else "",
                "양식": idx.get("__양식", "정리표"),
                "출처": os.path.basename(path),
            })
    wb.close()
    return out


def _unique_deposit_files(paths):
    """여러 정본 경로에 복제된 *같은* 입금 파일은 한 번만 읽는다.

    원본 자료 폴더에는 공유 폴더 정본을 날짜별로 보관한 복사본이 함께 있다.
    두 경로를 모두 읽되 내용이 완전히 같은 파일만 SHA-256으로 제거한다.
    파일명이 같아도 내용이 다르면 서로 다른 갱신본일 수 있으므로 보존한다.
    """
    unique, seen = [], set()
    for path in paths:
        digest = hashlib.sha256()
        with open(path, "rb") as fh:
            for block in iter(lambda: fh.read(1024 * 1024), b""):
                digest.update(block)
        key = digest.digest()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def dedupe_deposits(rows):
    """**같은 입금이 서로 다른 파일에 들어 있는 것**을 합친다 (2026-08-08 실사고).

    무엇이 문제였나: 입금 합계가 2,417,075,528원인데 청구 합계는 545,353,633원 —
    4.4배였다. 같은 입금이 오종현 정리본과 은행 원본 **양쪽에** 들어 있었고,
    46쌍 1,056,769,994원이 두 번 세어지고 있었다.
    위 `_unique_deposit_files` 는 **똑같은 파일**만 거른다(SHA-256). 이건 서로 다른
    파일이다 — 사람이 정리한 표와 은행 원본이라 **모양이 달라 안 걸렸다.** 같은 돈인데.
    입금액 칸이 비어 있어서 다행이었다. 자동입력이 돌았다면 10억이 두 번 들어갔다.

    ★ 어려운 것은 합치는 게 아니라 **멀쩡한 입금을 지우지 않는 것**이다.
      실측으로 갈라야 할 것이 둘 있었다:
        · 같은 날 같은 금액인데 **거래처가 다른** 입금 — 다른 돈이다(2026-05-06 387,200원
          세 건이 그랬다). 그래서 열쇠에 거래처가 들어간다.
        · **같은 파일 안에서** 같은 날·금액·거래처가 두 번 — 그건 실제로 두 번 들어온
          것이다(은행 원본에 두 줄이면 이체가 두 번이다). 그래서 **파일 경계를 넘을 때만**
          합친다. 파일 안의 개수는 그대로 둔다.
      지워서 나는 사고가 더 크다 — 두 번 세면 눈에 띄지만, 지운 입금은 영영 안 들어온
      것으로 남는다.

    어느 쪽을 남기나: **은행 원본**이다. 사람 손이 안 닿았으니 정본에 가깝다
    (parse_deposit_list 주석과 같은 판단). 은행 원본에 없는 기간은 정리표가 채운다.
    """
    groups = collections.OrderedDict()
    for r in rows:
        key = (r["일자"], int(round(r["금액"])), r.get("거래처") or "")
        groups.setdefault(key, []).append(r)
    out, dropped = [], []
    for key, g in groups.items():
        bank = [r for r in g if r.get("양식") == "은행원본"]
        keep = bank if bank else g
        # 파일 안의 개수는 보존한다 — 은행 원본에 두 줄이면 두 번 들어온 것이다.
        out.extend(keep)
        if bank and len(g) > len(bank):
            dropped.extend([r for r in g if r.get("양식") != "은행원본"])
    return out, dropped


def load_deposits():
    """입금내역 폴더의 엑셀을 전부 읽는다(사용자 지시 — 여기가 정본).

    같은 입금이 여러 파일에 겹쳐 있으면 **한 번만** 센다(dedupe_deposits).
    """
    import glob
    from source_dirs import receipt_dirs
    out, candidates = [], []
    for d in receipt_dirs():
        for f in sorted(glob.glob(os.path.join(d, "**", "*.xlsx"), recursive=True)):
            if os.path.basename(f).startswith("~$"):
                continue                      # 엑셀이 열려 있을 때 생기는 잠금 파일
            candidates.append(f)
    files = _unique_deposit_files(candidates)
    for f in files:
        out += parse_deposit_list(f)
    out, dropped = dedupe_deposits(out)
    if dropped:
        print("  ※ 파일이 겹쳐 두 번 세어진 입금 %d건 %s원을 뺐습니다"
              " (은행 원본을 남깁니다)"
              % (len(dropped), format(int(sum(r["금액"] for r in dropped)), ",")))
    return out, files


def open_settlements(master):
    """입금일이 비어 있는 06시트 정산행 — 청구액(세금계산서합계 우선)과 함께."""
    from ecount_reconcile import read_ledger
    out = []
    for sid, r in read_ledger(master).items():
        if r.get("원장_입금일"):
            continue
        billed = r.get("원장_세금계산서합계") or r.get("원장_거래명세서합계")
        if not billed:
            continue
        out.append({
            "정산ID": sid, "청구액": float(billed),
            "프로젝트NO": r.get("프로젝트NO", ""), "캠프명": r.get("캠프명", ""),
            "발행일": _day(r.get("원장_세금계산서발행일") or r.get("원장_거래명세서발행일")),
        })
    return out


def match(receipts, rows):
    """양방향 유일 일치만 자동입력. 겹치면 사람에게 넘긴다."""
    paired, spare = [], []
    for rc in receipts:
        cands = [s for s in rows
                 if abs(s["청구액"] - rc["금액"]) <= AMOUNT_TOL
                 and (s["발행일"] is None or s["발행일"] <= rc["일자"])]
        rivals = [q for q in receipts if abs(q["금액"] - rc["금액"]) <= AMOUNT_TOL]
        if len(cands) == 1 and len(rivals) == 1:
            paired.append((rc, cands[0]))
        else:
            spare.append((rc, "후보 %d건 · 같은 금액 입금 %d건" % (len(cands), len(rivals))))
    return paired, spare


# ── 묶음 입금 분해 ────────────────────────────────────────────────────────────
# ★ 왜 필요한가 (2026-08-08 지시).
#   위 match() 는 **한 입금 = 한 정산**일 때만 채운다. 그런데 쿠팡은 여러 건을 묶어
#   한 번에 넣는다. 그래서 이 도구가 매일 도는데도 결과가 **늘 0건**이었다 —
#   실측 입금 147건 2,417,075,528원이 전부 '보류'였고 06시트 입금액 칸은 724행이
#   통째로 비어 있었다. 코드도 자료도 멀쩡한데 규칙이 현실과 안 맞았던 것이다.
#   비어 보이지도 않는다: 매일 "보류 147건"을 성실히 찍으니 도는 줄로만 안다.
#
# ★ 어려운 것은 맞추는 게 아니라 **잘못 맞추지 않는 것**이다.
#   금액 하나에 우연히 합이 맞는 조합은 얼마든지 나온다. 엉뚱한 정산행에 입금을 붙이면
#   그 건은 받은 것으로 처리돼 미수 목록에서 사라지고, 정작 못 받은 건은 영영 안 보인다.
#   그래서 근거는 하나뿐이다 — **합이 딱 맞는 조합이 정확히 하나일 때만** 채운다.
#   둘 이상이면 어느 쪽이 맞는지 알 방법이 없으므로 사람에게 넘긴다(예전과 같다).
MAX_PARTS = 8          # 한 번에 묶이는 건수 상한 — 넘으면 우연한 합이 급격히 늘어난다
MAX_POOL = 40          # 후보 상한. 넘으면 경우의 수가 터져 계산이 끝나지 않는다
MAX_STATES = 400_000   # 부분합 표 상한 — 넘으면 '못 셌다'로 두고 사람에게 넘긴다


def unique_subset(pool, target):
    """합이 `target` 인 조합이 **정확히 하나**일 때만 그 조합을 준다.

    반환: (조합 or None, 왜)
      · 조합을 세다 2개째가 나오면 즉시 포기한다 — 전부 셀 이유가 없다.
      · 금액은 원 단위 정수로 다룬다(부동소수 오차로 합이 어긋나면 안 된다).
    """
    tgt = int(round(target))
    items = [(int(round(s["청구액"])), s) for s in pool if 0 < round(s["청구액"]) <= tgt]
    if not items or len(items) > MAX_POOL:
        return None, ("후보 없음" if not items else "후보 %d건 — 상한 초과" % len(items))
    # dp[합] = (조합수(2에서 멈춤), 조합 예시)
    dp = {0: (1, ())}
    for amt, row in items:
        nxt = dict(dp)
        for s, (n, ex) in dp.items():
            if len(ex) >= MAX_PARTS:
                continue
            t = s + amt
            if t > tgt:
                continue
            on, oex = nxt.get(t, (0, ()))
            nxt[t] = (min(2, on + n), oex or (ex + (row,)))
        dp = nxt
        if len(dp) > MAX_STATES:
            return None, "경우의 수가 너무 많다(%d) — 사람 확인" % len(dp)
    n, ex = dp.get(tgt, (0, ()))
    if n == 1 and len(ex) >= 2:
        return list(ex), "유일 조합 %d건" % len(ex)
    if n == 0:
        return None, "맞는 조합 없음"
    return None, "조합이 둘 이상 — 어느 쪽인지 알 수 없다"


def match_bundles(spare, rows, taken):
    """보류된 묶음 입금을 분해한다. 이미 쓴 정산행(taken)은 다시 쓰지 않는다."""
    made, still = [], []
    for rc, why in spare:
        pool = [s for s in rows
                if s["정산ID"] not in taken
                and (s["발행일"] is None or s["발행일"] <= rc["일자"])]
        combo, note = unique_subset(pool, rc["금액"])
        if combo:
            for s in combo:
                taken.add(s["정산ID"])
            made.append((rc, combo, note))
        else:
            still.append((rc, why + " · 묶음분해: " + note))
    return made, still


def billing_totals(master):
    """06시트 청구·입금 총액. 건별 귀속이 안 될 때 **총액으로는** 대사할 수 있다."""
    from ecount_reconcile import read_ledger
    billed = paid = 0.0
    n_billed = n_paid = 0
    biggest = 0.0
    for _sid, r in read_ledger(master).items():
        b = r.get("원장_세금계산서합계") or r.get("원장_거래명세서합계")
        if b:
            billed += float(b)
            n_billed += 1
            biggest = max(biggest, float(b))
        if r.get("원장_입금액"):
            paid += float(r["원장_입금액"])
            n_paid += 1
    return {"청구건": n_billed, "청구액": billed, "입금기록건": n_paid,
            "입금기록액": paid, "최대1건": biggest}


def summarize(receipts, files, master=None, save=True):
    """입금 현황 정리 — 콘솔엔 집계만, 상세는 reports/ 로."""
    import collections
    by = collections.defaultdict(lambda: [0, 0.0])
    for r in receipts:
        b = by[r["거래처"]]
        b[0] += 1
        b[1] += r["금액"]
    total = sum(r["금액"] for r in receipts)
    days = [r["일자"] for r in receipts]
    print("입금내역 %d건 · 합계 %s원 · %s ~ %s (파일 %d개)"
          % (len(receipts), format(int(total), ","), min(days), max(days), len(files)))
    for cust, (n, amt) in sorted(by.items(), key=lambda kv: -kv[1][1])[:8]:
        print("  %-18s %3d건 %16s원" % (cust[:18], n, format(int(amt), ",")))
    if len(by) > 8:
        print("  … 외 %d개 거래처" % (len(by) - 8))
    if not save:
        return
    os.makedirs(REPORT_DIR, exist_ok=True)
    out = os.path.join(REPORT_DIR, "입금현황.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("# 쿠팡 입금 현황\n\n")
        f.write("- 원천: %s\n" % " · ".join(os.path.basename(x) for x in files))
        f.write("- 입금 %d건 · 합계 %s원 · %s ~ %s\n\n" %
                (len(receipts), format(int(total), ","), min(days), max(days)))
        if master:
            t = billing_totals(master)
            gap = t["청구액"] - total
            f.write("## 총액 대사 (건별 귀속은 불가 — 아래 주의 참고)\n\n")
            f.write("| 항목 | 건 | 금액 |\n|---|---:|---:|\n")
            f.write("| 관리대장 청구액(세금계산서·명세서) | %d | %s |\n"
                    % (t["청구건"], format(int(t["청구액"]), ",")))
            f.write("| 관리대장에 기록된 입금 | %d | %s |\n"
                    % (t["입금기록건"], format(int(t["입금기록액"]), ",")))
            f.write("| **실제 입금(이 폴더 기준)** | %d | **%s** |\n"
                    % (len(receipts), format(int(total), ",")))
            f.write("| 청구 − 실제입금 | | %s |\n\n" % format(int(gap), ","))
            mid = sorted(r["금액"] for r in receipts)[len(receipts) // 2]
            f.write("> ★ **이 차액을 미수로 확정 보고하지 말 것.** 입금은 여러 건을 묶어 들어온다\n"
                    "> (입금 중앙값 %s원 vs 정산 1건 최대 %s원 — 한 번 입금에 여러 건이 섞여 있다).\n"
                    "> 관리대장에 아직 청구가 안 올라온 건도 있다. 건별로 어느 계산서에 대한 입금인지는\n"
                    "> **쿠팡 지급명세(remittance)가 있어야** 알 수 있다.\n\n"
                    % (format(int(mid), ","), format(int(t["최대1건"]), ",")))
        f.write("## 거래처별\n\n| 거래처 | 건수 | 금액 |\n|---|---:|---:|\n")
        for cust, (n, amt) in sorted(by.items(), key=lambda kv: -kv[1][1]):
            f.write("| %s | %d | %s |\n" % (cust, n, format(int(amt), ",")))
        f.write("\n## 전체 내역(날짜 오름차순)\n\n| 날짜 | 거래처 | 금액 |\n|---|---|---:|\n")
        for r in sorted(receipts, key=lambda x: (x["일자"], x["거래처"])):
            f.write("| %s | %s | %s |\n" % (r["일자"], r["거래처"], format(int(r["금액"]), ",")))
    print("  상세:", out)


def main():
    queue_mode = "--queue" in sys.argv
    from ecount_reconcile import load_config, resolve_master
    from inbox_scan import pick
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])

    # 1순위: 오종현 관리 입금내역 폴더(사용자 지시 — 여기가 정본)
    receipts, dep_files = load_deposits()
    if receipts:
        summarize(receipts, dep_files, master=master)

    # 2순위: 이카운트 거래처별계정별원장의 대변. 있으면 보태서 본다.
    for f in pick("ledger"):
        part, ok = parse_receipts(f)
        if ok:
            receipts += part

    if not receipts:
        print("입금 자료가 아직 없습니다 — 들어오면 자동으로 잡습니다.")
        for d in __import__("source_dirs").receipt_dirs() or ["(입금내역 폴더 없음)"]:
            print("  넣는 곳:", d)
        print("  또는 회계 I > 출력물 > 장부 > 거래처별계정별원장")
        print("      거래처=쿠팡로지스틱스 · 계정=1089(외상매출금) · ★개별거래처기준 · 기간지정 → Excel")
        return 0

    rows = open_settlements(master)
    paired, spare = match(receipts, rows)

    print("입금 대조 — 원장 입금행 %d건 / 입금일 빈 정산 %d건 → 유일매칭 %d건 · 보류 %d건"
          % (len(receipts), len(rows), len(paired), len(spare)))
    for rc, s in paired[:20]:
        print("  · %s %s %s %s원 → %s (%s)"
              % (s["정산ID"], s["프로젝트NO"], s["캠프명"][:14],
                 format(int(rc["금액"]), ","), rc["일자"], rc["적요"][:20]))
    # ★ 여기서 끝내면 매일 "보류 147건"만 찍는다 — 묶음 입금을 분해해 본다.
    taken = {s["정산ID"] for _, s in paired}
    bundles, spare = match_bundles(spare, rows, taken)
    if bundles:
        print("  [묶음분해] %d건 — 합이 딱 맞는 조합이 **하나뿐**일 때만 채운다"
              % len(bundles))
        for rc, combo, note in bundles[:10]:
            print("    · %s %s원 → %d건 (%s)"
                  % (rc["일자"], format(int(rc["금액"]), ","), len(combo), note))
            for s in combo[:6]:
                print("        %s %s %s %s원"
                      % (s["정산ID"], s["프로젝트NO"], s["캠프명"][:12],
                         format(int(s["청구액"]), ",")))
    if spare:
        print("  [보류] %d건 — 조합이 둘 이상이거나 맞는 조합이 없습니다(사람 확인)" % len(spare))
    if not receipts:
        print("  ※ 대변(입금) 행이 0건입니다 — '대표거래처로 합산' 이 켜져 있으면 캠프별 거래처의")
        print("     입금이 빠집니다. 검색 조건을 '개별거래처기준' 으로 다시 뽑아 주세요.")

    if not queue_mode:
        print("\n미리보기 — 실제 적재: python receipt_fill.py --queue")
        return 0

    items = []
    # 묶음으로 풀린 건도 같은 방식으로 채운다. 다만 **입금액은 그 정산행의 청구액**이다 —
    # 입금 총액을 각 행에 적으면 한 건이 여러 번 들어온 것처럼 보인다.
    pairs = [(rc, s, "거래처별계정별원장 대변 %s %d원 (금액 유일매칭)"
                     % (rc["일자"], int(rc["금액"])), int(rc["금액"]))
             for rc, s in paired]
    for rc, combo, note in bundles:
        ev = ("거래처별계정별원장 대변 %s %d원 묶음입금 %d건으로 분해 (%s)"
              % (rc["일자"], int(rc["금액"]), len(combo), note))
        for s in combo:
            pairs.append((rc, s, ev, int(round(s["청구액"]))))
    for rc, s, ev, amount in pairs:
        for sheet in ("06_거래서류청구수금", "16_입금수금관리"):
            items.append({"sheet": sheet, "key_col": "정산ID", "key": s["정산ID"],
                          "col": "입금일", "value": rc["일자"].isoformat(), "vtype": "date",
                          "evidence": ev, "only_if_empty": True})
            items.append({"sheet": sheet, "key_col": "정산ID", "key": s["정산ID"],
                          "col": "입금액", "value": str(amount), "vtype": "number",
                          "evidence": ev, "only_if_empty": True})
    if not items:
        print("적재할 항목 없음")
        return 0
    from ledger_writer import queue_add
    print("큐 적재:", queue_add(items), "개 셀 → ledger_db --intake 후 11:00·15:00 원장 반영")
    return 0


if __name__ == "__main__":
    sys.exit(main())
