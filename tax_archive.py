# -*- coding: utf-8 -*-
"""
tax_archive.py — 세금계산서를 **건별 PDF** 로 보관한다 (일자-No. 가 곧 파일명)

사용자 지시(2026-08-05): "1월부터 지금까지 세금계산서 pdf 건별 전체 다운로드 …
원본 저장". ERP `(세금)계산서진행단계` 엑셀은 246건이 한 파일에 들어 있어 사람이
특정 계산서를 찾을 수 없다. 한 건 = 한 PDF 로 굳혀 월별 폴더에 넣는다.

    세금계산서_건별/2026/01/[2026-01-15-3]_쿠팡로지스틱스_UJ2600123_1,650,000원_발행완료.pdf
      └ 대괄호=일자-No.(ERP 키) · 거래처 · 프로젝트NO · 금액 · 진행단계

프로젝트NO 는 프로젝트명에서 UJ 패턴이 보일 때만 넣는다(지어내지 않는다).
승인번호가 있으면 PDF 본문에 남긴다 — 홈택스 대조의 정본 키다.

  python tax_archive.py             # 새 건만
  python tax_archive.py --limit 300
  python tax_archive.py --force
"""
import argparse
import glob
import json
import os
import re
import sys
import time

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
import child_budget

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

UJ = re.compile(r"UJ\d{7}")
OUT_JSON = os.path.join(ROOT, "reports", "세금계산서_상세.json")
INCREMENTAL_RETURN_CODE = child_budget.INCREMENTAL_RETURN_CODE


def out_root():
    import source_dirs as S
    return os.path.join(S.ERP_DIR, "세금계산서_건별")


def safe(s, n=24):
    s = re.sub(r"[\\/:*?\"<>|\r\n\t]", " ", str(s or "")).strip()
    return re.sub(r"\s+", " ", s)[:n]


def path_key(path):
    return os.path.normcase(os.path.abspath(path))


def existing_pdf_paths(root):
    """기존 PDF를 디렉터리 열람 한 번으로 읽는다.

    Z:의 `os.path.exists(dst)`를 계산서마다 부르면 완료 회차도 원본 수만큼 SMB 왕복을
    한다. 스캔 오류는 누락으로 오인해 다시 렌더하지 않도록 실패로 올린다.
    """
    if not os.path.isdir(root):
        return set()

    def raise_walk_error(exc):
        raise exc

    found = set()
    for base, _dirs, files in os.walk(root, onerror=raise_walk_error):
        for name in files:
            if name.lower().endswith(".pdf"):
                found.add(path_key(os.path.join(base, name)))
    return found


def norm_slip(raw):
    """`2026/01/15 -3` → `2026-01-15-3`."""
    parts = re.findall(r"\d+", str(raw or ""))
    if len(parts) < 4:
        return ""
    return f"{parts[0]}-{parts[1]}-{parts[2]}-{int(parts[3])}"


def collect():
    """계산서진행단계 엑셀(내용 판별 taxinv)에서 건별 행을 모은다."""
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    import source_dirs as S
    # ★ **목록을 받을 때 딸려 온 stat 을 버리지 않는다**([198]·[409]).  예전에는
    #   `glob` 이 **이름만** 주고 `cands.sort(key=os.path.getmtime)` 가 파일마다
    #   Z: 를 **다시 찔렀다** — SMB 에서 그것은 파일 하나가 왕복 한 번이다.
    #   같은 폴더(`S.ERP_DIR`)에서 [409] 가 잰 값: **124.9초 -> 10.0초**
    #   (파일 170개 · 파일당 731.9ms 왕복이 사라졌다 · **목록·순서는 그대로**).
    #   여기는 계산서진행단계(taxinv)를 찾는 자리다.
    # ⚠ **`skip_dirs=()` 를 반드시 적는다**([198]).  공용 워커의 기본값은 *색인의*
    #   목록(`_보관`·`_바로가기`)이라, 말없이 물려받으면 거기 든 엑셀이 **조용히
    #   빠지면서 오류도 안 난다**([165]).  옛 `glob` 은 아무것도 안 걸렀으므로
    #   여기서도 안 거른다 — **결과가 한 톨도 바뀌면 안 된다**.
    try:
        from source_index import walk_stat
        pairs = [(st.st_mtime, os.path.join(dp, fn))
                 for dp, fn, st in walk_stat(S.ERP_DIR, skip_dirs=())
                 if fn.lower().endswith(".xlsx")
                 and not fn.startswith(("~$", "ESD007E"))]
        pairs.sort(key=lambda x: x[0], reverse=True)
        cands = [p for _, p in pairs]
    except Exception:
        # 공용 워커를 못 쓰면 예전 길로 간다 — 느릴 뿐 답은 같다([169]).
        cands = [p for p in glob.glob(os.path.join(S.ERP_DIR, "**", "*.xlsx"), recursive=True)
                 if not os.path.basename(p).startswith(("~$", "ESD007E"))]
        cands.sort(key=os.path.getmtime, reverse=True)
    rows, seen, src = [], set(), []
    for p in cands[:80]:
        try:
            wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
            ws = wb.active
            head = [str(c or "") for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
            j = "|".join(head)
            if "진행단계" not in j or "승인번호" not in j:
                wb.close()
                continue
            idx = {h: i for i, h in enumerate(head)}
            for r in ws.iter_rows(min_row=3, values_only=True):
                r = ["" if c is None else str(c).strip() for c in r]
                if not r or len(r) < 5:
                    continue
                slip = norm_slip(r[0])
                if not slip or slip in seen:
                    continue
                seen.add(slip)
                g = lambda k, d="": (r[idx[k]] if k in idx and len(r) > idx[k] else d)
                amt = (g("공급가액") or "").replace(",", "")
                rows.append({
                    "slip": slip, "project": g("프로젝트명"), "cust": g("거래처명"),
                    "issued": g("발행일자"), "supply": int(float(amt)) if re.fullmatch(r"-?\d+(\.\d+)?", amt) else 0,
                    "vat": g("부가세"), "total": g("합계금액"), "type": g("종류"),
                    "state": g("전자(세금)계산서 진행단계"), "approve": g("승인번호"),
                    "memo": g("적요명"), "src": os.path.basename(p),
                })
            src.append(os.path.basename(p))
            wb.close()
        except Exception:
            continue
    return rows, src


def render(row, path):
    from archive_render import html_to_pdf, esc
    uj = (UJ.search(row.get("project") or "") or [None])
    uj = uj.group(0) if hasattr(uj, "group") else ""
    html = f"""
<h1>세금계산서 {esc(row['slip'])}</h1>
<div class="meta">
 <b>거래처</b> {esc(row.get('cust'))} &nbsp;|&nbsp; <b>프로젝트</b> {esc(row.get('project'))}
 {f'({esc(uj)})' if uj else ''}<br>
 <b>발행일자</b> {esc(row.get('issued') or '(미발행)')} &nbsp;|&nbsp;
 <b>진행단계</b> {esc(row.get('state'))} &nbsp;|&nbsp; <b>종류</b> {esc(row.get('type'))}<br>
 <b>승인번호</b> {esc(row.get('approve') or '-')}
</div>
<table>
 <tr><th>공급가액</th><th>부가세</th><th>합계금액</th></tr>
 <tr><td class="num">{row.get('supply', 0):,}</td><td class="num">{esc(row.get('vat'))}</td>
     <td class="num">{esc(row.get('total'))}</td></tr>
</table>
<p><b>적요</b> {esc(row.get('memo'))}</p>
<div class="foot">원본: ERP (세금)계산서진행단계 {esc(row.get('src'))} ·
 보관 생성 {time.strftime('%Y-%m-%d %H:%M')} · 자동 생성본(원본 불변)</div>
"""
    return html_to_pdf(html, path)


def render_atomic(row, path):
    """시간 제한 중 끊긴 PDF를 정상 보관본으로 세지 않는다."""
    tmp = path + f".part-{os.getpid()}"
    try:
        made = render(row, tmp)
        if made and os.path.exists(tmp):
            os.replace(tmp, path)
            return path
        return None
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except OSError:
            pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=400)
    ap.add_argument("--force", action="store_true")
    a = ap.parse_args()

    rows, src = collect()
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    json.dump({"count": len(rows), "src": src, "rows": rows},
              open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False)
    root, made, skip, fail, attempted = out_root(), 0, 0, 0, 0
    existing = set() if a.force else existing_pdf_paths(root)
    cut, seen = False, 0
    for row in rows:
        if attempted >= a.limit:
            # 상한 뒤의 계산서를 아직 보지 않았다. 0으로 닫으면 잔량이 있어도 자율복구
            # 큐가 완료 처리하므로 다음 회차가 확인하도록 75를 돌린다.
            cut = True
            break
        seen += 1
        y, m = row["slip"][:4], row["slip"][5:7]
        uj = (UJ.search(row.get("project") or "") or [None])
        uj = uj.group(0) if hasattr(uj, "group") else ""
        state = "발행완료" if str(row.get("state", "")).startswith(("발행", "전송")) else \
                safe(row.get("state") or "미발행", 10)
        name = (f"[{row['slip']}]_{safe(row.get('cust'), 18)}"
                f"{'_' + uj if uj else ''}_{row.get('supply', 0):,}원_{state}.pdf")
        dst = os.path.join(root, y, m, name)
        if not a.force and path_key(dst) in existing:
            skip += 1
            continue
        attempted += 1
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        if render_atomic(row, dst):
            made += 1
            existing.add(path_key(dst))
        else:
            fail += 1
    print(f"세금계산서 건별 PDF: 새로 {made}건 · 건너뜀 {skip} · 실패 {fail} "
          f"(원본 {len(rows)}건) → {root}")
    if fail:
        print(f"  ★ PDF {fail}건을 만들지 못했다 — 완료로 닫지 않고 실패로 남긴다.")
        return 1
    if cut:
        print(f"  ★ 건수 상한({a.limit:,}건)에 닿아 여기까지 하고 돌아온다 — "
              f"원본 {len(rows):,}건 중 앞에서부터 {seen:,}건까지 봤다. "
              "잔량 확인은 다음 회차가 이어서 한다(PDF 는 한 건씩 저장돼 있다).")
        return INCREMENTAL_RETURN_CODE
    return 0


if __name__ == "__main__":
    sys.exit(main())
