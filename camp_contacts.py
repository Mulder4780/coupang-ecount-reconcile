# -*- coding: utf-8 -*-
"""
camp_contacts.py — **전국 쿠팡캠프 담당자 한 장**: 캠프명 · 담당자 이름 · 전화번호

유수비 대표 지시(2026-08-18 07:33 카톡): "쿠팡 전국정기점검 캠프와 담당자리스트
오늘중으로 주세요"

왜 새로 만드나 (실측 2026-08-18)
  · 캠프 목록은 이미 있다 — `reports/캠프마스터.json` 268개.
  · 그런데 그중 **연락처가 5개뿐**이었다. ERP 거래처등록(ESA001M)을 대 보니
    CU(캠프) 356개 중 전화번호가 **18개**다 — 열을 안 읽은 것이 아니라([165] 를
    먼저 의심했다) **ERP 에 안 적혀 있는 것**이다. 다른 거래처는 423개가 차 있다.
  · 진짜 원천은 **밴드 접수 글 본문**이다(9,557건 중 6,246건에 번호가 있다):
        ● 캠프이름 : 울산2캠프
        ● 캠프주소 : 울산광역시 …
        ● 현장책임 : 제석화
        ● 담당번호 : 010-7532-8543
        ● 안전관리 : 이상협
        ● 담당번호 : 010-2511-4947 / waynelee@coupang.com
    캠프마다 **현장책임·안전관리 두 사람**이 있고 셋(이름·전화·메일) 다 있다.

지키는 것
  · **읽는 자리는 하나다** — 뽑기는 `band_extract.parse_post` 가 한다([162]).
    여기서 본문을 다시 정규식으로 뜯지 않는다. 갈리면 한쪽만 고쳐진다.
  · **담당자는 바뀐다.** 캠프마다 **최신 게시일** 것을 채택하고 근거(게시일·밴드·
    글번호)를 같이 남긴다. 언제 자료인지 안 밝히면 낡은 번호를 확언하게 된다.
  · **못 뽑은 캠프를 '담당자 없음'이라 하지 않는다** — 밴드에 접수 글이 없으면
    그것은 **모르는 것**이다([169]). `근거` 칸이 비면 화면이 그렇게 적는다.
  · **짐작으로 채우지 않는다.** 캠프명이 비슷하다고 옆 캠프 번호를 붙이면 대표
    보고에 틀린 번호가 실린다 — 빈 칸보다 나쁘다([172]).
  · **Z: 를 안 훑는다** — 밴드 캐시만 읽는다([168]). 실측 3초.
  · 원문은 한 글자도 안 고친다.

산출: reports/캠프_담당자.json     (앱 `/api/camps` 가 이것만 읽는다)

  python camp_contacts.py            (사람이 볼 요약)
  python camp_contacts.py --write    (파일 생성 · 09:50 회차 단계)
"""
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
try:  # 무인 회차는 pythonw 라 sys.stdout 이 None 이다 — [235]
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import band_extract as BE  # noqa: E402

REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")
OUT = os.path.join(REPORT_DIR, "캠프_담당자.json")
MASTER = os.path.join(REPORT_DIR, "캠프마스터.json")

# 정기점검으로 세는 업무유형. `band_extract` 가 실제로 쓰는 낱말이며 여기서
# 새 낱말을 지어내지 않는다.
PM_KINDS = ("정기점검",)


# 캠프명 칸에 **사람이 메모를 적은** 글이 있다(실측: `서초1MB(양재동) = 서초1Sub-hub -
# 2RT 안전수칙은 …` 처럼 58자). 그런 이름은 **자동으로 자르지 않는다** — 잘라 붙이면
# 엉뚱한 캠프에 번호가 들어간다([172]). 대신 표시만 갈라 사람이 판단하게 둔다([169]).
NAME_MAX = 24


def _norm(name):
    """캠프명 맞추기 — **대소문자·공백·괄호·하이픈만** 없앤다.

    실측(2026-08-18): `울산1Sub-hub` / `울산1sub-hub`, `전주1 Sub Hub` /
    `전주1SUBHUB` / `전주1Sub-hub`, `M_익산1` / `M익산1` 처럼 **표기만 다른**
    묶음이 19개였다. 안 합치면 같은 캠프가 목록에 여러 줄로 뜬다.

    ★ 그래도 **숫자는 안 건드린다.** `송파1MB(감일동)` 과 `송파5MB(감일동)` 은
      실재하는 **다른 캠프**다([172] 의 실측 오탐) — 앞 숫자가 달라 안 합쳐진다.
    """
    import re as _re
    return _re.sub(r"[\s()\[\]_\-/·.]", "", str(name or "")).upper()


#: 캠프를 가리키는 낱말 — 이 중 하나도 없으면 캠프명 칸에 들어온 **메모**다.
#: 실측(2026-08-18 형님 지적): `안전수칙 부착 확인하기 - 부착안된 건들도 있음` 처럼
#: 기사가 캠프명 칸에 적어 둔 안내문이 목록에 캠프로 올라와 대표 보고에까지 갔다.
#: 캠프 이름에는 **절대 안 나오는** 서술어. 형님 지적(2026-08-18)의 그 줄
#: `안전수칙 부착 확인하기 - 부착안된 건들도 있음` 이 여기 걸린다.
MEMO_WORDS = ("하기", "있음", "없음", "필요", "요청", "바랍니다", "해주", "주세요",
              "확인바", "부착안", "미부착", "완료함", "예정임")

CAMP_WORDS = ("캠프", "캐프", "MB", "SUB-HUB", "SUBHUB", "SUB HUB", "SUB-FC", "SUBFC",
              "HUB", "FC", "센터", "물류", "터미널", "지점", "창고")


def looks_like_camp(name):
    """캠프 이름 모양인가 — **아닌 것만** 뺀다(`[172]`).

    ★ 넓게 거르면 실재하는 캠프가 목록에서 사라진다 — 그 캠프는 아무도 안 챙기면서
      목록에도 없다. 그래서 근거는 **캠프를 가리키는 낱말이 하나도 없다** 하나뿐이고,
      진짜 캠프명이 앞에 붙은 메모(`서초1모바일캠프(양재동/주유소) - 안전수칙은 …`)는
      **그대로 남긴다** — 그것은 `이름확인필요` 로 이미 표시돼 사람이 판단한다.
    ★ 뺀 것은 **숫자로 말한다**(`[169]`) — 조용히 빼면 '없는 캠프'가 된다.
    """
    nm = str(name or "").strip()
    if len(nm) < 3:
        return False                       # `)`·`000` 같은 파싱 부스러기
    up = nm.upper()
    if any(w in up for w in CAMP_WORDS):
        return True                        # 캠프 낱말이 있으면 무조건 남긴다
    # ★ **낱말이 없다고 바로 빼지 않는다.** 실측으로 그렇게 했더니 737→590,
    #   정기점검 402→**280** 이 됐다 — `=M_ASN1 (아산)`·`M_DAJ1 (당진)` 같은
    #   **실재하는 캠프 코드**가 통째로 사라졌다. 122개 캠프가 대표 보고에서
    #   빠지는 것은 메모 몇 줄이 섞이는 것보다 훨씬 나쁘다(`[172]`).
    #   그래서 남은 문은 **서술어**뿐이다 — 캠프 이름에는 동사가 없고, 메모에는 있다.
    #   길이로 자르면 `MC08(당진) M_DAJ1 (당진)` 처럼 **실재 캠프 둘이 한 칸에 적힌 것**
    #   까지 사라진다(실측 8개). 서술어는 캠프 이름에 절대 안 나온다.
    return not any(w in nm for w in MEMO_WORDS)


def _units(text):
    """본문 → 호기 번호들. 판정은 `pm_content` 것을 그대로 빌린다(`[162]`).

    ★ 못 읽으면 **빈 것으로 두고 넘어간다** — 호기 하나 때문에 캠프 목록
      전체가 안 만들어지면 연락처까지 같이 잃는다.
    """
    if not text:
        return ()
    try:
        import pm_content
        return tuple(pm_content.parse_units(text).get("units") or ())
    except Exception:                        # noqa: BLE001
        return ()


# ─────────────────────────────────────────────────────────────────────────────
# 모름을 채운다 — **원천 셋을 전수로 보되, 후보가 유일할 때만** (2026-08-19 지시)
#
# 사용자 지시: "모름 표시되어있는 부분 밴드 및 erp, 원본 데이터등 전수 조사해서
# 반영될 수 있도록 코딩해"
#
# 먼저 어디에 답이 있는지 쟀다(짐작으로 채우지 않는다). 실측 2026-08-19:
#   · 밴드·카톡 본문 — 이미 쓰지만 **최신 글이 사람 칸을 통째로 덮어** 25칸을 잃고 있었다
#   · 밴드·카톡 **사람 명부** — 안 쓰고 있었다. 같은 사람이 여러 캠프에 나오므로
#     한 곳에서 알아낸 메일을 다른 캠프에서 쓸 수 있다(실측 143칸)
#   · **ERP 거래처 마스터** — 안 쓰고 있었다. 캠프 249개에 담당자명 217 · 메일 198
#     이 있고 밴드와 같은 사람이다(권순환 → sukwon52@coupangls.com, 실측 일치)
#   · 관리대장 — **캠프 연락처 칸이 없다**(read_ledger 열 목록에 하나도 없다)
#
# ★ 어려운 것은 채우는 게 아니라 **잘못 채우지 않는 것**이다([172]). 메일을 틀리게
#   채우면 사람이 **엉뚱한 사람에게 캠프 업무를 보낸다** — 모름으로 두는 것보다 나쁘다.
#   그래서 문을 다섯 건다:
#     ① **후보가 유일할 때만** 채운다. 여럿이면 그대로 모름이고 숫자로 보고한다([169]).
#     ② **덮지 않는다.** 빈 칸만 채운다 — 원문에 적힌 값이 언제나 이긴다.
#     ③ **역할을 지어내지 않는다.** ERP 거래처 담당자는 현장책임인지 안전관리인지
#        말하지 않으므로 `담당자`(직책 미상) 칸에만 담는다.
#     ④ **번호가 이름보다 세다.** 이름은 양식에 박혀 돌아다니지만 번호는 그 사람이다
#        (`people_alias.py` 가 쓰는 것과 같은 근거). 전화로 먼저 묻고 없을 때만 이름.
#     ⑤ **채운 것을 원문처럼 보이게 하지 않는다.** 근거에 출처를 적어 사람이
#        "이건 어디서 온 값인가"를 물을 수 있게 한다.
# ─────────────────────────────────────────────────────────────────────────────
PERSON_FIELDS = ("이름", "전화", "메일")


def _same_person(a, b):
    """두 사람 칸이 같은 사람인가. **번호가 먼저**다([172] — 사람이 바뀌면 옛 메일을
    새 사람 이름 옆에 붙이면 안 된다. 그건 아무도 아닌 연락처가 된다)."""
    ta, tb = (a.get("전화") or "").strip(), (b.get("전화") or "").strip()
    if ta and tb:
        return ta == tb
    na, nb = (a.get("이름") or "").strip(), (b.get("이름") or "").strip()
    if na and nb:
        return na == nb
    # 한쪽이 통째로 비면 판단할 근거가 없다 — 이어 붙이지 않는다.
    return False


def _merge_person(prev, new):
    """새 글이 이기되 **같은 사람이면 빈 칸에 옛 값을 남긴다.**"""
    new = dict(new or {})
    if not prev:
        return new
    if not _same_person(prev, new):
        return new                      # 사람이 바뀌었다 — 섞지 않는다
    for f in PERSON_FIELDS:
        if not (new.get(f) or "").strip() and (prev.get(f) or "").strip():
            new[f] = prev[f]
    return new


def _first_name(text):
    """사람 이름 토막 하나를 고른다. 못 고르면 **빈 문자열**(=명부에 안 넣는다).

    ★ **한 글자는 이름이 아니다** (2026-08-19 실측). 첫 판이 앞 토막만 잘랐더니
      명부 열쇠에 `김`·`이`·`M`·`씬` 같은 한 글자와 직함 `SM` 이 들어왔고, 그것으로
      3칸이 채워질 참이었다. `김` 하나로 사람을 특정하면 **엉뚱한 사람의 메일**이
      캠프에 박힌다 — 모름으로 두는 것보다 나쁘다([172]).
    ★ 그래서 **한글이 든 토막을 먼저** 고른다: `SM 로건`→`로건` · `권순환 Cain`→`권순환`
      · `김대범 / Move CL`→`김대범`. 한글이 하나도 없으면 첫 토막을 쓰되 2글자 이상만.
    """
    t = str(text or "").strip()
    if not t:
        return ""
    toks = [x for x in t.replace("/", " ").split() if x]
    for x in toks:
        if any("가" <= ch <= "힣" for ch in x) and len(x) >= 2:
            return x
    for x in toks:
        if len(x) >= 2:
            return x
    return ""


def person_directory(recs, erp_rows=()):
    """사람 명부 — 전화·이름으로 이름/전화/메일을 되찾는다.

    ★ **세는 것은 사람이지 글이 아니다.** 같은 글이 여러 번 실려도 후보가 늘지
      않아야 '후보 여럿'(=안 채움) 판정이 흔들리지 않는다.
    """
    tel, name = {}, {}
    def put(box, key, field, val):
        if not key or not val:
            return
        box.setdefault(key, {}).setdefault(field, set()).add(val)
    for r in recs:
        for slot in ("현장책임", "안전관리", "담당자"):
            p = r.get(slot) or {}
            t = (p.get("전화") or "").strip()
            n = _first_name(p.get("이름"))
            m = (p.get("메일") or "").strip()
            if t:
                put(tel, t, "이름", n)
                put(tel, t, "메일", m)
            if n:
                put(name, n, "전화", t)
                put(name, n, "메일", m)
    # ERP 거래처등록의 담당자·메일도 같은 명부에 넣는다 — 같은 사람이다(실측).
    for r in erp_rows or ():
        n = _first_name(r.get("manager"))
        m = str(r.get("email") or "").strip()
        t = str(r.get("tel") or "").strip()
        if n:
            put(name, n, "메일", m)
            put(name, n, "전화", t)
        if t:
            put(tel, t, "이름", n)
            put(tel, t, "메일", m)
    return {"전화": tel, "이름": name}


def _lookup(directory, person, field):
    """이 칸을 채울 값 → (값, 근거) 또는 (None, 사유). 번호가 이름보다 세다."""
    t = (person.get("전화") or "").strip()
    n = _first_name(person.get("이름"))
    for 열쇠, 값, 라벨 in (("전화", t, "같은 번호"), ("이름", n, "같은 이름")):
        if not 값:
            continue
        cand = (directory.get(열쇠, {}).get(값) or {}).get(field)
        if not cand:
            continue
        if len(cand) == 1:
            return next(iter(cand)), 라벨
        return None, "후보 여럿"          # 여럿이면 이름으로도 다시 묻지 않는다
    return None, ""


def fill_gaps(rows, directory, erp_by_key=None):
    """빈 칸을 명부로 채운다. **덮지 않고 · 유일할 때만 · 출처를 적는다.**

    돌려주는 숫자가 곧 리포트가 말할 것이다 — 채운 칸, 후보가 여럿이라 **안 채운**
    칸, 그리고 끝내 모름인 칸. 조용히 채우면 사람이 원문에 적힌 값과 구별하지 못한다.
    """
    채움 = {}
    여럿 = {}
    erp_by_key = erp_by_key or {}
    for r in rows:
        # ① ERP 거래처등록의 담당자 — **직책 미상**이므로 `담당자` 칸에만 담는다.
        er = (erp_by_key.get((r.get("거래처코드") or "").strip())
              or erp_by_key.get(_norm(r.get("캠프명"))))
        if er:
            p = dict(r.get("담당자") or {})
            for 칸, 열 in (("이름", "manager"), ("메일", "email"), ("전화", "tel")):
                v = str(er.get(열) or "").strip()
                if 칸 == "이름":
                    v = _first_name(v)
                if v and not (p.get(칸) or "").strip():
                    p[칸] = v
                    채움["담당자 " + 칸] = 채움.get("담당자 " + 칸, 0) + 1
                    r.setdefault("보완", {})["담당자 " + 칸] = "ERP 거래처등록"
            if p:
                r["담당자"] = p
        # ② 사람 명부 — 같은 번호(우선) · 같은 이름
        for slot in ("현장책임", "안전관리", "담당자"):
            p = dict(r.get(slot) or {})
            if not p:
                continue
            # ★ **방금 채운 값을 다시 열쇠로 쓰지 않는다** (2026-08-19 실측).
            #   첫 판은 이름→전화→메일 순서로 채우면서 `p` 를 그대로 열쇠로 썼다.
            #   그러면 ERP 가 준 이름으로 전화를 찾고, **그 전화로** 메일을 찾는
            #   연쇄가 생긴다 — 한 칸이 틀리면 나머지도 같이 틀리면서 **똑같이
            #   확신에 차 보인다.** 조회는 언제나 **원문 그대로의 사람 칸**에 한다.
            원문 = dict(p)
            for f in PERSON_FIELDS:
                if (p.get(f) or "").strip():
                    continue
                v, 왜 = _lookup(directory, 원문, f)
                열쇠 = slot + " " + f
                if v:
                    p[f] = v
                    채움[열쇠] = 채움.get(열쇠, 0) + 1
                    r.setdefault("보완", {})[열쇠] = "사람 명부(%s)" % 왜
                elif 왜 == "후보 여럿":
                    여럿[열쇠] = 여럿.get(열쇠, 0) + 1
            r[slot] = p
    return {"보완채움": 채움, "보완후보여럿": 여럿,
            "보완합계": sum(채움.values()), "보완보류": sum(여럿.values())}


# -----------------------------------------------------------------------------
# ★ 원천 하나가 더 있었다 — **정기점검 스케줄 원본**이 정답지다 (2026-08-19 지시)
#
# 형님 캡처: `(류지영) ★01. 쿠팡 정기점검 스케줄표_원본.xlsx` 의 `2026년 N분기
# 정기점검` 시트. 열이 이렇다 — 기존 캠프명 · 변경 캠프명 · 캠프주소 · 호기 ·
# 종류 · 모델 · **현장책임자(CL)** · HP · **안전관리자** · HP · **E-mail**.
#
# ★ **왜 이 표가 이기나** — 밴드 접수 글은 그때 그 사람이 적어 넣은 것이고
#   이 표는 **목적 자체가 캠프 담당자 명부**다(쿠팡이 주고 류지영이 관리한다).
#   실측 2026-08-19: 이름이 겹치는 183캠프에서 **값이 서로 다른 칸 236개**
#   (안전관리 이름 49 · 메일 30 · 주소 46 · 호기 34) · 빈 칸을 채울 것 93개.
#   즉 화면이 지금까지 **낡은 값을 그럴듯하게** 보여 주고 있었다([165] 모양).
#
# ★ **덮지만 버리지 않는다**([169]) — 밀린 값은 `이전값` 에 남기고 **몇 칸을
#   고쳤는지 숫자로 말한다**. 조용히 덮으면 "그때 밴드에는 뭐라고 적혀 있었나"를
#   잃고, 잘못 덮었을 때 되돌릴 근거도 없어진다.
# ★ **빈 값으로 덮지 않는다** — 엑셀 칸이 비었으면 밴드 값이 그대로 남는다.
# ★ **호기는 뜻이 달라진다** — 밴드 것은 '관측된 호기'인데 이 표는 **설치 대수**다.
#   그래서 `호기출처` 를 같이 실어 화면이 갈라 말하게 한다([169]).
# ★ **비싼 읽기는 캐시 검사 뒤에**([168]) — Z: 의 4MB 워크북이다. 지문은
#   (파일명·크기·수정시각·규칙판)이고, 규칙을 고치면 `PM_SCHED_VER` 를 손으로
#   올린다 — 원본이 안 바뀌면 지문이 안 움직여 옛 답이 영원히 이긴다(이 프로젝트가
#   네 번 겪은 모양이다).
# ★ **못 읽으면 '없다'가 아니다**([169]) — 리포트가 그 사실을 그대로 적는다.
# -----------------------------------------------------------------------------
def _sched_files():
    """폴더의 스케줄 워크북 **전부**를 새로 손댄 것부터. 없으면 빈 목록.

    ★ **하나만 고르면 정본이 통째로 사라진다** (2026-08-19 실사고).
      그날 폴더에 둘이 있었다 — `(류지영) ★01. 쿠팡 정기점검 스케줄표_원본.xlsx`
      (16:43 · 분기 시트 **4개** · 193캠프)와 `정기점검 리스트(안전관리자).xlsx`
      (16:44 · 3분기 **한 장** · 152캠프). 예전 규칙은 mtime 최신 하나였으므로
      **1분 차이로 정본이 통째로 밀렸고**, 4분기가 안 읽혀 236칸이 3분기 값으로
      되돌아갔다(M광주2 안전관리 김장혁(Dino) -> 정지수). 어느 쪽이 이길지가
      업무 의미가 아니라 **누가 나중에 저장했나**로 정해지고 있었다.
    ★ 그래서 순위는 파일이 아니라 **분기**가 정한다([326] 의 '최신 시트가 이긴다').
      같은 분기가 두 파일에 다 있으면 그때만 mtime 최신이 이긴다.
    """
    try:
        import source_dirs as SD
        folder = SD.PM_SCHEDULE_DIR
    except Exception:
        return []
    out = []
    try:
        for n in os.listdir(folder):
            if n.startswith("~$") or not n.lower().endswith((".xlsx", ".xlsm")):
                continue
            p = os.path.join(folder, n)
            try:
                st = os.stat(p)
            except Exception:
                continue
            out.append((p, st.st_mtime, st.st_size))
    except Exception:
        return []
    out.sort(key=lambda x: -x[1])
    return out


def _quarter_of(sheet):
    """'2026년 4분기 정기점검' -> (2026, 4). 못 읽으면 (0, 0) 이라 뒤로 간다."""
    m = re.search(r"(\d{4})\s*년", sheet or "")
    q = re.search(r"(\d)\s*분기", sheet or "")
    return (int(m.group(1)) if m else 0, int(q.group(1)) if q else 0)


PM_SCHED_CACHE = os.path.join(REPORT_DIR, "캠프_정기점검원본.json")
PM_SCHED_VER = 3          # <- 파싱 규칙을 고치면 이 숫자를 올린다


def _sched_file():
    """가장 최근에 손댄 스케줄 워크북 하나. 없으면 None."""
    try:
        import source_dirs as SD
        folder = SD.PM_SCHEDULE_DIR
    except Exception:
        return None
    best = None
    try:
        for n in os.listdir(folder):
            if n.startswith("~$") or not n.lower().endswith((".xlsx", ".xlsm")):
                continue
            p = os.path.join(folder, n)
            st = os.stat(p)
            if best is None or st.st_mtime > best[1]:
                best = (p, st.st_mtime, st.st_size)
    except Exception:
        return None
    return best


def _cell(row, i):
    if i is None or i < 0 or i >= len(row):
        return ""
    v = row[i]
    if v is None:
        return ""
    return str(v).replace(chr(10), " ").strip()


def _tel(text):
    """전화 모양만 남긴다 — 엑셀 칸에 공백·메모가 붙어 온다."""
    t = "".join(ch for ch in str(text or "") if ch.isdigit() or ch == "-").strip("-")
    return t if sum(ch.isdigit() for ch in t) >= 9 else ""


def _mail(text):
    t = str(text or "").strip()
    return t if ("@" in t and " " not in t and "." in t.split("@")[-1]) else ""


# 그 워크북은 셀이 병합돼 있어 **캠프명 칸에 주소가 그대로 들어오는 행**이 있다
# (실측: `경상남도 통영시 광도면 덕포로 202` · `88`). `looks_like_camp` 는 밴드
# 본문용이라 이런 것을 통과시키는데, **여기서 그 함수를 넓히면 안 된다** — 밴드
# 쪽 판정까지 같이 바뀐다([172]). 그러니 문은 **이 표 전용**으로 단다.
_ADDR_HEAD = ("서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
              "경기", "강원", "충청", "충북", "충남", "전라", "전북", "전남",
              "경상", "경북", "경남", "제주")


def _looks_like_address(name):
    """행정 지명으로 시작하고 띄어쓴 덩어리가 셋 이상이면 주소다."""
    t = str(name or "").strip()
    if not t.startswith(_ADDR_HEAD):
        return False
    return len(t.split()) >= 3 or any(t.endswith(x) for x in ("로", "길")) \
        or any(w in t for w in ("읍 ", "면 ", "동 ", "리 ", "번길"))


def _sched_camp_ok(name):
    """이 표의 캠프명 칸이 정말 캠프인가 — 주소·숫자·메모는 아니다([169])."""
    t = str(name or "").strip()
    if not t or not looks_like_camp(t):
        return False
    return not _looks_like_address(t)


def pm_schedule_camps(force=False):
    """정기점검 스케줄 원본 -> ({정규화이름: {…}}, 왜). 실패도 말로 돌려준다."""
    files = _sched_files()
    if not files:
        return {}, {"길": "못 읽음: 정기점검 스케줄 원본 폴더에 워크북이 없다"}
    # ★ 지문에 **파일 전부**를 담는다 — 하나만 담으면 옆 파일이 바뀌어도 옛 답이 이긴다.
    sig = "|".join("%s:%s:%s" % (os.path.basename(p_), int(m_), z_)
                   for p_, m_, z_ in files) + ("|v%d" % PM_SCHED_VER)
    if not force:
        try:
            with open(PM_SCHED_CACHE, encoding="utf-8") as fh:
                old = json.load(fh)
            if old.get("지문") == sig:
                return old.get("camps") or {}, old.get("왜") or {}
        except Exception:
            pass
    try:
        import openpyxl
    except Exception as e:                                   # noqa: BLE001
        return {}, {"길": "못 읽음: %s" % str(e)[:80]}
    # ★ 파일이 여럿이면 **분기가 순위를 정한다**([326] '최신 시트가 이긴다').
    #   같은 분기가 두 파일에 다 있을 때만 mtime 최신이 이긴다 — 그때는 어느 쪽이
    #   새 값인지 파일 시각 말고는 근거가 없다.
    books, 시트목록, 못읽음 = [], [], []
    for path_, mtime_, _sz in files:
        try:
            wb_ = openpyxl.load_workbook(path_, read_only=True, data_only=True)
        except Exception as e:                               # noqa: BLE001
            # ★ 한 파일이 깨져도 나머지는 읽는다 — 다만 **못 읽었다고 말한다**([169]).
            못읽음.append("%s: %s" % (os.path.basename(path_), str(e)[:60]))
            continue
        books.append(wb_)
        for i_, t_ in enumerate(wb_.sheetnames):
            if "분기" in t_ and "정기점검" in t_:
                y_, q_ = _quarter_of(t_)
                시트목록.append((-y_, -q_, -mtime_, i_, wb_, t_,
                              os.path.basename(path_)))
    시트목록.sort(key=lambda x: x[:4])
    if not 시트목록 and 못읽음:
        return {}, {"길": "못 읽음: " + " / ".join(못읽음)[:120]}

    want = {"변경캠프명": "새이름", "기존캠프명": "옛이름", "캠프주소": "주소",
            "호기": "호기", "종류": "종류", "모델": "모델",
            "현장책임자(CL)": "cl", "현장책임자(CL)HP": "clhp",
            "안전관리자": "safe", "안전관리자HP": "safehp",
            "E-mail": "mail", "안전관리자e-mail": "mail", "설치일자": "설치일"}
    camps, 시트수, 행수 = {}, 0, 0
    쓴파일 = []
    for order, (_y, _q, _m, _i, wb, t, _fn) in enumerate(시트목록):
        if _fn not in 쓴파일:
            쓴파일.append(_fn)
        ws = wb[t]
        hdr, hr = None, 0
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
            vals = [_cell(row, i) for i in range(len(row))]
            if "변경 캠프명" in vals or "No." in vals:
                hdr = {}
                for i, v in enumerate(vals):
                    k = v.replace(" ", "")
                    if k in want and want[k] not in hdr:
                        hdr[want[k]] = i
                hr = r
                break
        if not hdr or "새이름" not in hdr:
            continue
        시트수 += 1
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            name = _cell(row, hdr.get("새이름")) or _cell(row, hdr.get("옛이름"))
            if not _sched_camp_ok(name):
                continue          # 숫자·주소·메모가 캠프명 칸에 섞여 온다([172])
            행수 += 1
            key = _norm(name)
            c = camps.setdefault(key, {"캠프명": name, "시트": t, "순": order,
                                       "호기": [], "옛이름": []})
            if order < c["순"]:                       # 더 최신 시트가 이긴다
                c["캠프명"], c["시트"], c["순"] = name, t, order
            for x in _units(_cell(row, hdr.get("호기"))):
                if x not in c["호기"]:
                    c["호기"].append(x)
            옛 = _cell(row, hdr.get("옛이름"))
            if (옛 and _norm(옛) != key and _sched_camp_ok(옛)
                    and len(옛) <= NAME_MAX and 옛 not in c["옛이름"]):
                c["옛이름"].append(옛)
            for k2 in ("주소", "종류", "모델", "설치일"):
                v = _cell(row, hdr.get(k2))
                if v and not c.get(k2):
                    c[k2] = v[:60]
            cl = {"이름": _cell(row, hdr.get("cl"))[:40],
                  "전화": _tel(_cell(row, hdr.get("clhp"))), "메일": ""}
            sf = {"이름": _cell(row, hdr.get("safe"))[:40],
                  "전화": _tel(_cell(row, hdr.get("safehp"))),
                  "메일": _mail(_cell(row, hdr.get("mail")))}
            for slot, p in (("현장책임", cl), ("안전관리", sf)):
                # 사람 칸도 **각각 따로** 최신을 고른다 — 통째로 덮으면 메일이 사라진다.
                merged = dict(c.get(slot) or {})
                for fld in ("이름", "전화", "메일"):
                    if p.get(fld) and (not merged.get(fld) or order <= c["순"]):
                        merged[fld] = p[fld]
                if any(merged.get(x) for x in ("이름", "전화", "메일")):
                    c[slot] = merged
    for wb_ in books:
        try:
            wb_.close()
        except Exception:
            pass
    for c in camps.values():
        c["호기"] = sorted(c["호기"])
    왜 = {"길": "읽음" if not 못읽음 else "읽음(일부 못 읽음)",
          "파일": " + ".join(쓴파일) if 쓴파일 else "",
          "파일수": len(files), "못읽은파일": 못읽음,
          "시트차례": [x[5] for x in 시트목록],
          "시트": 시트수,
          "행": 행수, "캠프": len(camps),
          "받은때": BE.datetime.now().isoformat(timespec="seconds")
          if hasattr(BE, "datetime") else ""}
    try:
        with open(PM_SCHED_CACHE, "w", encoding="utf-8") as fh:
            json.dump({"지문": sig, "왜": 왜, "camps": camps}, fh, ensure_ascii=False)
    except Exception:
        pass                       # 캐시를 못 써도 답은 답이다
    return camps, 왜


def _same(a, b):
    def flat(x):
        return str(x or "").replace(" ", "").replace("-", "")
    return flat(a) == flat(b)


def apply_pm_schedule(rows, camps):
    """정본을 얹는다 — **덮되 버리지 않고, 몇 칸인지 말한다**([169])."""
    stat = {"채운칸": 0, "고친칸": 0, "고친캠프": 0, "새캠프": 0, "칸별": {}, "예": []}
    by = {}
    for r in rows:
        by.setdefault(_norm(r.get("캠프명")), r)
    for key, c in camps.items():
        r = by.get(key)
        if r is None:
            # ★ 밴드에도 마스터에도 없는 캠프 — **빼면 '없는 캠프'가 된다**([169]).
            r = {"캠프명": c["캠프명"], "캠프주소": "", "거래처코드": "",
                 "이름확인필요": len(c["캠프명"]) > NAME_MAX, "다른표기": [],
                 "호기": [], "정기점검": True, "정기점검건수": 0, "돌발AS건수": 0,
                 "총건수": 0, "최근작업일": "", "현장책임": {}, "안전관리": {},
                 "담당자": {}, "근거": {}}
            rows.append(r)
            by[key] = r
            stat["새캠프"] += 1
        before = (stat["채운칸"], stat["고친칸"])
        keep = r.setdefault("이전값", {})

        def put(새, 자리, 이전):
            if not 새:
                return ""                   # 빈 값으로 덮지 않는다
            if 이전 and not _same(이전, 새):
                keep.setdefault(자리, {})["밴드"] = 이전
                stat["고친칸"] += 1
                stat["칸별"][자리] = stat["칸별"].get(자리, 0) + 1
                if len(stat["예"]) < 8:
                    stat["예"].append("%s %s: %s -> %s" % (r["캠프명"], 자리,
                                                           str(이전)[:22], str(새)[:22]))
            elif not 이전:
                stat["채운칸"] += 1
            return 새

        for slot in ("현장책임", "안전관리"):
            src = c.get(slot) or {}
            cur = dict(r.get(slot) or {})
            for fld in ("이름", "전화", "메일"):
                v = put(src.get(fld), "%s.%s" % (slot, fld), cur.get(fld))
                if v:
                    cur[fld] = v
            if any(cur.get(x) for x in ("이름", "전화", "메일")):
                r[slot] = cur
                r.setdefault("근거", {})[slot] = {"출처": "정기점검 스케줄 원본",
                                                  "시트": c.get("시트") or ""}
        v = put(c.get("주소"), "캠프주소", r.get("캠프주소"))
        if v:
            r["캠프주소"] = v
        if c.get("호기"):
            옛 = list(r.get("호기") or [])
            if 옛 and sorted(옛) != sorted(c["호기"]):
                keep.setdefault("호기", {})["밴드"] = 옛
                stat["고친칸"] += 1
                stat["칸별"]["호기"] = stat["칸별"].get("호기", 0) + 1
            elif not 옛:
                stat["채운칸"] += 1
            r["호기"] = list(c["호기"])
            # ★ 뜻이 다르다 — 밴드는 '관측', 이 표는 '설치 대수'다. 화면이 갈라 말한다.
            r["호기출처"] = "정기점검 스케줄 원본"
        for 옛이름 in c.get("옛이름") or []:
            alt = r.setdefault("다른표기", [])
            if 옛이름 != r.get("캠프명") and 옛이름 not in alt:
                alt.append(옛이름)
        for f2 in ("종류", "모델", "설치일"):
            if c.get(f2) and not r.get(f2):
                r[f2] = c[f2]
        # 이 표에 있다는 것 자체가 정기점검 대상이라는 뜻이다.
        r["정기점검"] = True
        if (stat["채운칸"], stat["고친칸"]) != before:
            stat["고친캠프"] += 1
        if not keep:
            r.pop("이전값", None)
    return stat


def build():
    recs = BE.load_records()
    camps = {}
    버린메모 = []                                # 캠프 이름 모양이 아니라 뺀 것(`[169]`)
    for r in recs:
        camp = (r.get("캠프명") or "").strip()
        if not camp:
            continue
        if not looks_like_camp(camp):
            버린메모.append(camp[:60])          # 캠프가 아니라 캠프명 칸에 적힌 메모다
            continue
        key = _norm(camp)
        posted = str(r.get("게시일") or "")
        c = camps.setdefault(key, {
            "캠프명": camp, "캠프주소": "", "현장책임": None, "안전관리": None,
            # 옛 양식(2023~2024)은 직책 없이 `● 담당자명` 한 사람이다 — 실측 2,714글.
            # 현장책임 칸에 합치지 않는다(원문이 직책을 말한 적 없다 · [172]).
            "담당자": None,
            "근거": {}, "정기점검건수": 0, "돌발AS건수": 0,
            "최근작업일": "", "총건수": 0, "호기": set(),
        })
        # ★ **호기를 새로 읽지 않는다** — `pm_content.parse_units` 한 곳이 판정한다
        #   (`[162]`). 거기는 '1,2호기'·'1~3호기'·'#4'·'3호' 를 다 받고 1~99 만 인정해
        #   날짜·금액 조각을 호기라고 우기지 않는다. 여기서 정규식을 또 쓰면 정기점검
        #   화면과 캠프 화면이 **같은 글을 놓고 다른 호기를 말하게 된다.**
        c["호기"].update(_units(r.get("본문")))
        # 표시 이름은 **가장 최근 글에 쓰인 그대로**를 쓴다(양식이 바뀌면 따라간다).
        c["총건수"] += 1
        if r.get("업무유형") in PM_KINDS:
            c["정기점검건수"] += 1
        elif r.get("업무유형") == "돌발AS":
            c["돌발AS건수"] += 1
        wd = str(r.get("작업일") or "")
        if wd > c["최근작업일"]:
            c["최근작업일"] = wd

        # ★ 사람 칸은 **각각 따로** 최신을 고른다. 한 글에 현장책임만 적히고
        #   안전관리는 빈 글이 흔하다 — 글 하나를 통째로 이기게 하면 애써 있는
        #   다른 칸이 빈 값에 덮인다.
        for slot in ("현장책임", "안전관리", "담당자"):
            p = r.get(slot)
            if not p:
                continue
            prev = c["근거"].get(slot, {})
            if posted >= str(prev.get("게시일") or ""):
                # ★ 바로 위 주석과 **같은 함정이 한 층 아래에** 있었다 (2026-08-19).
                #   역할별로는 갈랐는데 그 안의 이름·전화·메일은 **통째로 덮었다** —
                #   최신 글에 메일 칸이 없으면 옛 글의 메일이 사라진다. 실측 25칸.
                c[slot] = _merge_person(c.get(slot), p)
                c["근거"][slot] = {"게시일": posted, "밴드": r.get("밴드"),
                                   "글번호": r.get("게시글")}
        if r.get("캠프주소") and posted >= str(c["근거"].get("주소", {}).get("게시일") or ""):
            c["캠프주소"] = r["캠프주소"]
            c["근거"]["주소"] = {"게시일": posted}
        # ★ 표시 이름은 최신이 아니라 **가장 자주 쓰인 표기**다. 최신 글 하나가
        #   메모 섞인 이름이면(실측 있음) 그 캠프가 목록에서 이상해진다.
        c.setdefault("표기", {})
        c["표기"][camp] = c["표기"].get(camp, 0) + 1

    # ERP 거래처코드를 붙인다 — **이미 만들어 둔 캠프마스터를 읽기만** 한다.
    #   여기서 ERP 를 다시 훑으면 웹 요청 뒤에서 Z: 재귀 탐색이 된다([168]).
    master = {}
    try:
        with open(MASTER, encoding="utf-8") as f:
            for row in (json.load(f).get("rows") or []):
                master[_norm(row.get("캠프명"))] = row
    except Exception:
        master = {}   # 못 읽었으면 '없다'가 아니라 **안 붙인다**

    rows = []
    for key, c in camps.items():
        m = master.get(key) or {}
        # 같은 캠프의 여러 표기 중 **가장 많이 쓰인 것**. 같은 횟수면 짧은 쪽 —
        # 긴 쪽은 메모가 섞여 있을 확률이 높다.
        tally = c.get("표기") or {c["캠프명"]: 1}
        name = sorted(tally.items(), key=lambda kv: (-kv[1], len(kv[0])))[0][0]
        rows.append({
            "캠프명": name,
            # ★ 자동으로 자르지 않는다 — 화면이 '이름 확인 필요'로 갈라 적는다.
            "이름확인필요": len(name) > NAME_MAX,
            "다른표기": sorted(k for k in tally if k != name)[:4],
            "캠프주소": c["캠프주소"] or (m.get("주소") or ""),
            "거래처코드": m.get("거래처코드") or "",
            # ★ **'관측된 호기'이지 '보유 대수'가 아니다**(`[169]`). 접수·점검 글이
            #   한 번도 없던 호기는 여기 안 보인다. 그래서 `1,2,5` 처럼 사이가 빌 수
            #   있고, 그것을 '3·4호기가 없다'로 읽으면 안 된다 — 화면이 그렇게 적는다.
            "호기": sorted(c["호기"]),
            "정기점검": c["정기점검건수"] > 0,
            "정기점검건수": c["정기점검건수"],
            "돌발AS건수": c["돌발AS건수"],
            "총건수": c["총건수"],
            "최근작업일": c["최근작업일"],
            "현장책임": c["현장책임"] or {},
            "안전관리": c["안전관리"] or {},
            "담당자": c["담당자"] or {},
            "근거": c["근거"],
        })

    # 밴드에 글이 **한 번도 없는** 캠프도 목록에 남긴다 — 빼면 '없는 캠프'가 된다.
    #   담당자 칸은 비고 `근거` 도 비므로 화면이 '모름'이라 적는다([169]).
    for key, m in master.items():
        if key in camps:
            continue
        rows.append({
            "캠프명": m.get("캠프명") or "", "캠프주소": m.get("주소") or "",
            "이름확인필요": len(m.get("캠프명") or "") > NAME_MAX, "다른표기": [],
            "거래처코드": m.get("거래처코드") or "",
            "정기점검": False, "정기점검건수": 0, "돌발AS건수": 0, "총건수": 0,
            "최근작업일": m.get("최근작업일") or "",
            # ERP 에 적힌 담당자가 있으면 그것만은 싣는다(출처를 밝힌다).
            # ★ 여기도 **직책 미상**이다 — ERP 거래처등록의 담당자 칸은 현장책임인지
            #   안전관리인지 말하지 않는다. 그래서 `담당자` 칸으로 담는다([172]).
            "현장책임": {}, "안전관리": {},
            "담당자": ({"이름": m.get("담당자") or "", "전화": m.get("연락처") or "",
                        "메일": ""} if (m.get("담당자") or m.get("연락처")) else {}),
            "근거": ({"담당자": {"출처": "ERP 거래처등록"}}
                     if (m.get("담당자") or m.get("연락처")) else {}),
        })

    # ── 모름 채우기 (2026-08-19 지시) ────────────────────────────────────
    # ★ 비싼 탐색은 **캐시 검사 뒤**에 온다([168]) — `camp_code_match._erp_customers`
    #   한 곳을 빌린다. 여기서 `load_customers()` 를 직접 부르면 Z: 를 또 훑는다.
    erp_rows, erp_왜 = [], {"길": "안 읽음"}
    try:
        from camp_code_match import _erp_customers
        erp_rows, erp_왜 = _erp_customers()
    except Exception as e:                       # noqa: BLE001
        # ★ **못 읽은 것을 '없다'로 치지 않는다**([169]). 리포트가 그대로 적는다.
        erp_왜 = {"길": "못 읽음: %s" % str(e)[:60], "건수": 0}
    erp_by_key = {}
    for er in erp_rows:
        코드 = str(er.get("code") or "").strip()
        if 코드:
            erp_by_key[코드] = er
        이름 = _norm(er.get("name") or "")
        if 이름:
            erp_by_key.setdefault(이름, er)
    # ★ **정본이 먼저 얹힌다** — 그다음에 남은 빈 칸만 명부·ERP 로 채운다.
    #   순서를 뒤집으면 약한 근거가 먼저 자리를 잡아 정본이 '고침'으로 세인다.
    sched, sched_왜 = pm_schedule_camps()
    정본 = apply_pm_schedule(rows, sched) if sched else {}
    보완 = fill_gaps(rows, person_directory(recs, erp_rows), erp_by_key)

    sort_rows(rows)
    return {
        "갱신": BE.datetime.now().isoformat(timespec="seconds")
        if hasattr(BE, "datetime") else "",
        "출처": "정기점검 스케줄 원본(류지영 정본·우선) + 밴드 접수 글 본문"
                "(band_extract) + reports/캠프마스터.json + ERP 거래처등록(모름 보완)",
        "보완출처": erp_왜,
        # ★ 조용히 덮지 않는다([169]) — 몇 칸을 고쳤는지·무엇을 못 읽었는지 말한다.
        "정기점검원본": sched_왜,
        "정본반영": 정본,
        "rows": rows,
        **보완,
        # ★ 뺀 것은 숫자로 말한다(`[169]`) — 조용히 빼면 '원래 없던 것'이 된다.
        "캠프아님": len(버린메모),
        "캠프아님예": sorted(set(버린메모))[:8],
        **summarize(rows),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 정렬·요약은 **한 곳**이다([162]).
#
# 사람이 앱에서 고친 값을 덮어 쓴 뒤(`camp_edit.overlay`)에도 머리글 숫자를 다시
# 세어야 하는데, 거기서 따로 세면 언젠가 갈린다 — 그러면 표에는 전화가 있는데
# 머리글은 '모름'이라 말하는 화면이 된다(오류는 안 난다 · [165] 모양).
# ─────────────────────────────────────────────────────────────────────────────
def has_tel(row):
    """이 캠프에 **누구든** 전화가 있나. 역할 이름은 여기 한 곳에만 적는다."""
    for slot in ("현장책임", "안전관리", "담당자"):
        if (row.get(slot) or {}).get("전화"):
            return True
    return False


def sort_rows(rows):
    """이름이 수상한 것은 **지우지 않고 맨 뒤로** 보낸다([169])."""
    rows.sort(key=lambda r: (bool(r.get("이름확인필요")), not r.get("정기점검"),
                             -int(r.get("정기점검건수") or 0), str(r.get("캠프명") or "")))
    return rows


def summarize(rows):
    tel = sum(1 for r in rows if has_tel(r))
    return {
        "캠프수": len(rows),
        "정기점검캠프수": sum(1 for r in rows if r.get("정기점검")),
        "전화있음": tel,
        "전화모름": len(rows) - tel,
        "이름확인필요": sum(1 for r in rows if r.get("이름확인필요")),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 담당자가 **바뀐 것**을 자국으로 남긴다 (2026-08-18 지시: "추가 및 변경시 자동으로
# 이 카테고리에 반영해서 업데이트").
#
#   · 이 회차는 매번 밴드에서 **다시 뽑아** OUT 을 통째로 덮는다. 그래서 어제 번호와
#     오늘 번호가 달라도 **어느 화면에도 티가 안 난다** — 값이 비는 것이 아니라
#     조용히 바뀌는 것이다([165] 와 같은 종류).
#   · 그러므로 덮기 **전에** 옛 파일과 대 보고 달라진 캠프만 적어 둔다.
#   · ★ **바뀐 것 자체는 경보가 아니다**([297]). 잘 따라간 변경은 정상이다 —
#     화면이 '최근 바뀜'으로 보여 줄 뿐 인계 '먼저 처리할 것'에는 안 올린다.
# ─────────────────────────────────────────────────────────────────────────────
CHANGES = os.path.join(REPORT_DIR, "캠프_변경.json")
CHANGE_KEEP = 400          # 이 이상은 오래된 것부터 버린다(파일이 무한히 자라지 않게)


def contact_sig(row):
    """한 캠프의 **연락처 지문**. 건수·최근작업일은 매일 움직이므로 넣지 않는다 —
    넣으면 매일 전부 '바뀜'이 되어 아무도 안 본다([170] 의 재수집 사고와 같은 모양)."""
    out = []
    for slot in ("현장책임", "안전관리", "담당자"):
        p = row.get(slot) or {}
        out.append("%s=%s/%s/%s" % (slot, p.get("이름") or "", p.get("전화") or "",
                                    p.get("메일") or ""))
    return " · ".join(out)


def diff_changes(old_rows, new_rows, when):
    """옛 목록 → 새 목록에서 **연락처가 달라진 캠프**만 돌려준다."""
    prev = {_norm(r.get("캠프명")): r for r in (old_rows or [])}
    out = []
    for r in new_rows:
        key = _norm(r.get("캠프명"))
        before = prev.get(key)
        sig = contact_sig(r)
        if before is None:
            # 새로 나타난 캠프 — 연락처가 있을 때만 적는다(빈 캠프는 소식이 아니다).
            if sig.replace("현장책임=//", "").replace("안전관리=//", "") \
                   .replace("담당자=//", "").replace(" · ", "").strip():
                out.append({"때": when, "캠프명": r.get("캠프명"), "갈래": "새 캠프",
                            "이전": "", "지금": sig})
            continue
        old_sig = contact_sig(before)
        if old_sig != sig:
            out.append({"때": when, "캠프명": r.get("캠프명"), "갈래": "연락처 바뀜",
                        "이전": old_sig, "지금": sig})
    return out


def record_changes(new_rows, when):
    """덮기 **전에** 부른다. 못 읽으면 조용히 넘어간다 — 첫 실행에는 옛 파일이 없다."""
    try:
        with open(OUT, encoding="utf-8") as f:
            old_rows = (json.load(f).get("rows") or [])
    except Exception:
        return []                      # 옛 파일이 없으면 '전부 새것'이 아니라 **모름**이다
    fresh = diff_changes(old_rows, new_rows, when)
    if not fresh:
        return []
    try:
        with open(CHANGES, encoding="utf-8") as f:
            log = json.load(f).get("rows") or []
    except Exception:
        log = []
    log = (fresh + log)[:CHANGE_KEEP]
    tmp = CHANGES + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump({"갱신": when, "rows": log}, f, ensure_ascii=False, indent=1)
    os.replace(tmp, CHANGES)
    return fresh


def load_changes(limit=30):
    try:
        with open(CHANGES, encoding="utf-8") as f:
            return (json.load(f).get("rows") or [])[:limit]
    except Exception:
        return []


# -----------------------------------------------------------------------------
# * 원본이 새로 와도 아무도 말해 주지 않는다 (2026-08-19 실사고 . 분담판 [151])
#
# 그날 형님이 캡처를 들고 "이게 맞다는데 검증해봐" 라고 물어서야 드러났다:
# 정기점검 스케줄 원본은 16:44 에 갱신됐는데 `reports/캠프_담당자.json` 은 16:16
# 것이라, 정본 152캠프 754칸 중 **236칸이 옛 담당자**였다. 그런데 그 236칸은
# **빈칸이 아니라 '틀린 사람'** 이라 화면은 멀쩡히 이름.전화를 보여 준다([165]).
# 형님이 안 물었으면 다음 09:50 회차까지 옛 담당자가 대표 보고에 그대로 실렸다.
#
# * **여기서 다시 만들지 않는다**([168]) - `build()` 는 밴드 전체를 파싱한다.
#   재는 것은 mtime 둘뿐이고, 고치는 것은 사람이 명령한다(`--write`).
# * **못 읽으면 '정상'이 아니라 '모름'이다**([169]) - Z: 가 끊긴 것과 원본이
#   그대로인 것은 다른 사실이다. 뭉치면 끊긴 날 화면이 "최신"이라고 확언한다.
# * **웹 요청에서 부르지 않는다**([168]) - `_sched_file()` 실측 2.78초(SMB).
#   회차가 재서 `STALE_MARK` 에 적고 인계.화면은 그것을 읽기만 한다.
# -----------------------------------------------------------------------------
STALE_MARK = os.path.join(REPORT_DIR, "캠프원본_밀림.json")
STALE_SLACK_S = 60        # 회차가 원본을 읽고 쓰는 사이에 나는 잔여 초
# ★ 같은 원본이 **두 파일**을 먹인다 (2026-08-20 · 분담판 `[168]`).
#     · `캠프_담당자.json`      → 전국쿠팡캠프 화면(담당자 이름.전화)
#     · `pm_schedule_sync.json` → **달력의 정기점검 예정**
#   그런데 감시자는 앞쪽만 봤다. 그래서 뒤쪽이 낡으면 달력이 **옛 예정을 조용히**
#   보여 주고 어느 화면도 그 말을 안 했다 — 실측 2026-08-19: 원본 09:11 인데
#   그 파일은 08-18 14:04 에 멈춰 있었고, 손으로 돌리자 확정 50 → 53건이 됐다.
#   형님은 "자료를 안올렸나?" 를 물으셨다 — **자료는 올라와 있었다.**
#   ⚠ 조치가 서로 다르므로 갈래를 합치지 않는다(`[289]`): 담당자는
#     `camp_contacts.py --write`, 달력은 `pm_schedule_sync.py --apply` 다.
CAL_OUT = os.path.join(REPORT_DIR, "pm_schedule_sync.json")
CAL_FIX = "python pm_schedule_sync.py --apply"


def _cal_stale(mt, name, when, _t):
    """달력이 읽는 `pm_schedule_sync.json` 이 원본을 따라왔나 (2026-08-20 · `[168]`).

    ★ **이 사고는 빈칸으로 안 나타난다.** 달력은 옛 예정을 멀쩡히 그리고, 오류도
      안 나고, 개수도 그럴듯하다 — 그래서 형님이 "자료를 안올렸나?" 를 물으실 때까지
      아무도 몰랐다(실측 2026-08-19: 원본 09:11 · 그 파일 08-18 14:04 · 손으로
      돌리자 확정 50 → 53건).
    ★ **담당자 갈래와 합치지 않는다**(`[289]`) — 고치는 명령이 다르다. 합치면 한쪽만
      낡은 날 사람이 **엉뚱한 명령**을 돌리고 원인은 그대로 남는다(`[172]`).
    ★ **못 읽은 것을 '정상'이라 하지 않는다**(`[169]`) — Z: 가 끊긴 것과 파일이
      최신인 것은 다른 사실이다."""
    fix = CAL_FIX
    try:
        rt = os.path.getmtime(CAL_OUT)
    except Exception:
        return {"갈래": "없음", "자료시각": "", "늦은분": 0, "조치": fix,
                "말": ("달력이 읽는 정기점검 예정 파일이 아직 없다 - "
                       "원본 %s 는 %s 것이다" % (name, when))}
    if mt <= rt + STALE_SLACK_S:
        return {"갈래": "정상", "자료시각": _t(rt), "늦은분": 0, "조치": fix,
                "말": "달력 정기점검 예정이 스케줄 원본보다 새롭다"}
    late = int((mt - rt) // 60)
    return {"갈래": "밀림", "자료시각": _t(rt), "늦은분": late, "조치": fix,
            "말": ("정기점검 스케줄 원본이 **달력 예정**보다 %d분 새롭다 "
                   "(원본 %s = %s . 달력 자료 %s) - 달력은 그동안 **옛 예정**을 "
                   "오류 없이 그대로 보여 준다" % (late, name, when, _t(rt)))}


def sched_stale():
    """정기점검 스케줄 원본이 앱 자료보다 새로운가. 갈래는 정상.밀림.모름 셋."""
    import datetime as _dt

    def _t(ts):
        try:
            return _dt.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return ""

    out = {"갈래": "모름", "말": "", "원본": "", "원본시각": "", "자료시각": "",
           "늦은분": 0, "조치": "python camp_contacts.py --write",
           "잰때": _dt.datetime.now().isoformat(timespec="seconds"),
           # ★ 달력 예정은 **다른 파일.다른 조치**다(`[289]`). 원본 시각은 한 번만
           #   재고 둘이 그것을 나눠 쓴다(`[162]`) — 여기서 또 재면 판정이 갈린다.
           "달력": {"갈래": "모름", "말": "정기점검 스케줄 원본을 못 읽어 달력 예정이 "
                                  "최신인지 확인 못 함",
                    "자료시각": "", "늦은분": 0, "조치": CAL_FIX}}
    try:
        f = _sched_file()
    except Exception as e:
        out["말"] = "정기점검 스케줄 원본 폴더를 못 읽었다(%s) - 최신인지 확인 못 함" % e
        return out
    if not f:
        # 폴더는 열렸는데 워크북이 없다. '없다'와 '못 읽었다'를 가른다([169]).
        out["말"] = "정기점검 스케줄 원본 폴더에 워크북이 없다 - 최신인지 확인 못 함"
        return out
    path, mt, _sz = f
    out["원본"] = os.path.basename(path)
    out["원본시각"] = _t(mt)
    out["달력"] = _cal_stale(mt, out["원본"], out["원본시각"], _t)
    try:
        rt = os.path.getmtime(OUT)
    except Exception:
        out["갈래"] = "없음"
        out["말"] = ("캠프 담당자 자료를 아직 한 번도 안 만들었다 - "
                     "원본 %s 는 %s 것이다" % (out["원본"], out["원본시각"]))
        return out
    out["자료시각"] = _t(rt)
    if mt <= rt + STALE_SLACK_S:
        out["갈래"] = "정상"
        out["말"] = "앱 담당자 자료가 정기점검 스케줄 원본보다 새롭다"
        return out
    out["갈래"] = "밀림"
    out["늦은분"] = int((mt - rt) // 60)
    # * 조용한 사고라는 것을 문장이 직접 말한다 - 빈칸이 아니라 '틀린 사람'이다.
    out["말"] = ("정기점검 스케줄 원본이 앱 담당자 자료보다 %d분 새롭다 "
                 "(원본 %s = %s . 앱 자료 %s) - 그 사이 바뀐 담당자는 "
                 "빈칸이 아니라 **옛 사람 이름.전화**로 화면에 그대로 보인다"
                 % (out["늦은분"], out["원본"], out["원본시각"], out["자료시각"]))
    return out


def stale_mark(rec=None):
    """회차가 판정을 적어 둔다. 인계.화면은 이 파일만 읽는다([168])."""
    rec = rec or sched_stale()
    try:
        os.makedirs(REPORT_DIR, exist_ok=True)
        tmp = STALE_MARK + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(rec, f, ensure_ascii=False, indent=1)
        os.replace(tmp, STALE_MARK)
    except Exception as e:
        # * 못 적었으면 못 적었다고 말한다([247]) - 200 을 주고 넘어가면
        #   감시자는 빈 파일을 보고 '이상 없음'이라 한다.
        rec = dict(rec, 적기실패=str(e))
    return rec


def stale_read():
    """회차가 써 둔 판정을 읽기만 한다. 없으면 None(= 아직 안 쟀다)."""
    try:
        with open(STALE_MARK, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def main():
    import datetime as _dt
    if "--stale" in sys.argv:
        # * 여기서는 build() 를 부르지 않는다([168]) - mtime 둘만 재고 끝낸다.
        rec = stale_mark() if "--write" in sys.argv else sched_stale()
        if hasattr(sys.stdout, "reconfigure"):
            try:
                sys.stdout.reconfigure(encoding="utf-8")   # [235] pythonw 대비
            except Exception:
                pass
        print("[%s] %s" % (rec["갈래"], rec["말"]))
        if rec["갈래"] == "밀림":
            print("  조치: " + rec["조치"])
        return
    d = build()
    d["갱신"] = _dt.datetime.now().isoformat(timespec="seconds")
    changed = []
    if "--write" in sys.argv:
        os.makedirs(REPORT_DIR, exist_ok=True)
        # ★ 덮기 **전에** 무엇이 달라졌는지 적는다 — 덮은 뒤에는 물어볼 곳이 없다.
        changed = record_changes(d["rows"], d["갱신"])
        tmp = OUT + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(d, f, ensure_ascii=False, indent=1)
        os.replace(tmp, OUT)
    print(f"캠프 {d['캠프수']}개 · 정기점검 {d['정기점검캠프수']}개 · "
          f"전화 있음 {d['전화있음']} · 모름 {d['전화모름']}"
          + (f" · 이번에 바뀐 캠프 {len(changed)}개" if changed else "")
          + (f" → {OUT}" if "--write" in sys.argv else ""))
    for r in d["rows"][:5]:
        # 화면과 같은 차례로 고른다 — 현장책임이 없으면 직책 미상 담당자가 ①이다.
        s = r["현장책임"] or r["담당자"] or {}
        print(f"  {r['캠프명']:<18} 정기{r['정기점검건수']:>3} "
              f"{s.get('이름','') or '-':<8} {s.get('전화','') or '-'}")


if __name__ == "__main__":
    main()
