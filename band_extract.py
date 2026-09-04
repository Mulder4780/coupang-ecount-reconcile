# -*- coding: utf-8 -*-
"""
band_extract.py — 밴드 게시글 → 구조화 업무 레코드 추출 (월별 백필 원천)
==========================================================================
밴드 게시글은 아래 규격으로 작성되어 있어 기계 파싱이 가능하다.

    ☑️판매전표 +거래명세서 +견적서 = 메일발송 完 ⭕
    ♣ ［ 2026년 02분기 3개월 유료 A/S 완료 ]
    ● A/S 일자 : 2026.06.01 (월요일)
    ● A/S 담당 : 김필우
    ● 프로젝트NO : UJ2600931
    ● 캠프이름 : 양주1캠프

이를 파싱해 [프로젝트NO·업무유형·유상무상·작업일·담당기사·캠프명·진행상태·문서상태]로 만든다.
관리대장에 없는 과거 월(2026-06, 05 …) 백필의 1차 원천이며,
이미 원장에 있는 건은 '원장등록됨'으로 표시해 중복 입력을 막는다.

실행:
  python band_extract.py --month 2026-06            # 6월 추출 → 리포트
  python band_extract.py --month 2026-06 --sheet    # + 관리대장 24_밴드업무추출 시트 반영(vN+1)
  python band_extract.py --all                      # 전체 기간
"""
import sys, os, re, csv, json, glob, hashlib
import time
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import people_alias as _ALIAS

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(BASE_DIR, "band", "cache")
# ★ 카톡 원본이 있는 자리도 **여기 한 곳**이 정한다 (2026-08-18).
#   읽는 쪽이 제 손으로 경로를 조립하면 사본이 둘 되고, 원본이 이사한 날
#   한쪽만 고쳐진다([162]). 지금 읽는 쪽: load_kakao_records ·
#   app_server._band_completion_index(카톡 근거 지문).
KAKAO_INBOX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kakao", "inbox")
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")

RE_PRJ = re.compile(r"프로젝트\s*NO\s*[:：]?\s*(UJ\d{6,})", re.I)
RE_DATE = re.compile(r"A/?S\s*일자\s*[:：]?\s*(\d{4})[.\-/](\d{1,2})[.,\-/](\d{1,2})")
RE_TECH = re.compile(r"A/?S\s*담당\s*[:：]?\s*([^\n●]*)")
RE_CAMP = re.compile(r"캠프\s*(?:이름|명)\s*[:：]?\s*([^\n●]*)")
RE_TITLE = re.compile(r"♣\s*[［\[]([^\]］]+)[\]］]")

# ★ 캠프 **담당자**(쿠팡 쪽 사람) — 우리 기사(`A/S 담당`)와 다른 사람이다.
#   접수 양식이 이렇게 생겼다(실측 2026-08-18 · 밴드 글 6,246건):
#       ● 캠프이름 : 울산2캠프
#       ● 캠프주소 : 울산광역시 …
#       ● 현장책임 : 제석화
#       ● 담당번호 : 010-7532-8543
#       ● 안전관리 : 이상협
#       ● 담당번호 : 010-2511-4947 / waynelee@coupang.com
#   ⚠ `담당번호` 가 **한 글에 두 번** 나온다. 순서로 고르면 언젠가 어긋나므로
#     **이름 줄 바로 다음 줄**의 번호만 그 사람 것으로 본다(짐작 금지 · [172]).
#   ⚠ 번호 칸에 이메일이 붙어 오므로 전화·메일을 갈라 담는다.
# ★ **메일이 오는 자리는 둘이다** (2026-08-18 실측 — 한쪽만 읽고 있었다).
#   ① `● 담당번호 : 010-2511-4947 / waynelee@coupang.com` — 번호 줄에 붙어 온다 (1,414줄)
#   ② `● E- MAIL : nominchul@coupangls.com` — **바로 다음 별도 줄**로 온다 (약 2,750줄)
#   ②를 안 읽어서 실측 현장책임 407명 중 메일이 **1명**뿐이었다 — 원본에는 있는데
#   화면이 비어 있었으니 '메일이 없는 캠프'로 보였다([165]: 안 읽은 칸은 빈칸과 같다).
#   라벨 표기가 제각각이라(`E- MAIL`·`E - MAIL`·`E-MAIL`·`E-mail`·`이메일`·`이- 메일`·
#   `email`) 공백·하이픈을 흘려 읽는다. **그 줄이 없으면 없는 대로 둔다** — 다음 사람의
#   메일을 끌어오면 대표 보고에 엉뚱한 주소가 박힌다(못 채우는 것보다 나쁘다 · [172]).
_MAIL_LINE = (r"(?:\n\s*●?\s*(?:E\s*[-–]?\s*MAIL|이\s*[-–]?\s*메일|이메일|메일\s*주소)"
              r"\s*[:：]?\s*([^\n●]*))?")
# ★ 라벨은 `현장책임`·`현장책임자`·`현장책임자명` 셋 다 온다 (2026-08-18 실측).
#   `자`·`자명` 을 안 먹으면 그것이 **이름 자리로 새어** `자 : One님`·`자명 : Turner님`
#   이 대표 보고에 그대로 실린다(실측 2건). 그렇다고 통째로 옵셔널하게 두면 이번엔
#   `현장책임 : 자명수` 같은 **멀쩡한 이름의 앞 글자를 라벨이 먹는다** — 못 읽는 것보다
#   나쁘다([172]). 그래서 뒤에 **콜론이 따라올 때만** 라벨로 친다(lookahead).
_MGR_SUFFIX = r"(?:\s*자\s*명?(?=\s*[:：]))?"
RE_SITE_MGR = re.compile(
    r"현장\s*책임" + _MGR_SUFFIX +
    r"\s*[:：]?\s*([^\n●]*)\n\s*●?\s*담당\s*(?:번호|자폰|자\s*폰)\s*[:：]?\s*([^\n●]*)"
    + _MAIL_LINE, re.IGNORECASE)
RE_SAFE_MGR = re.compile(
    r"안전\s*관리" + _MGR_SUFFIX +
    r"\s*[:：]?\s*([^\n●]*)\n\s*●?\s*담당\s*(?:번호|자폰|자\s*폰)\s*[:：]?\s*([^\n●]*)"
    + _MAIL_LINE, re.IGNORECASE)
# ★ **옛 양식은 한 사람이다** (2026-08-18 실측 — 안 읽고 있었다).
#   2023~2024 년 글은 현장책임·안전관리로 갈리기 전이라 이렇게 생겼다:
#       ● 캠프주소 : 경상남도 창원시 …
#       ● 담당자명 : 김재민
#       ● 담당번호 : 010-2003-3349
#   실측 캐시 7,813글 중 **옛 양식 2,714 · 새 양식 2,177**(둘 다 99). 즉 **옛 것이 더
#   많은데 위 두 정규식이 통째로 못 읽어** 그 캠프들이 전부 '전화 모름'으로 보였다 —
#   빈칸과 구별이 안 되는 종류의 잘못이다([165]).
# ⚠ **현장책임 칸에 넣지 않는다.** 원문은 이 사람이 현장책임인지 안전관리인지 **말한
#   적이 없다.** 넣으면 대표 보고에 근거 없는 직책이 박힌다 — 못 채우는 것보다 나쁘다
#   ([172]). 그래서 `담당자` 라는 제 칸으로 담고, 화면이 '직책 미상'으로 밝힌다.
# ⚠ `담당자명` 은 `담당번호`·`담당자폰` 과 겹치지 않는다(`명` 으로 끝난다).
#   `A/S 담당`(우리 기사)과도 다르다 — 그쪽은 RE_TECH 가 따로 읽는다.
RE_OLD_MGR = re.compile(
    r"담당\s*자\s*명\s*[:：]?\s*([^\n●]*)\n\s*●?\s*담당\s*(?:번호|자폰|자\s*폰)\s*[:：]?\s*([^\n●]*)"
    + _MAIL_LINE, re.IGNORECASE)
RE_CAMP_ADDR = re.compile(r"캠프\s*주소\s*[:：]?\s*([^\n●]*)")
# * **원본이 직접 말한 ERP 거래처코드** (2026-09-02 형님 지시 "거래처 코드는 최신
#   카톡 자료 및 최신 자료가 맞는거야 그걸로 업데이트해").
#
#   노승용 매니저 요청(2026-08-18)의 실체가 여기다 - *"거래처코드+거래처명만
#   업체정보에 매칭이 되어도 관리가 200,000% 쉬워지니 카톡이나 밴드 등에 넣으실
#   때 꼭 기입 요청드립니다."*  그 뒤로 접수 양식에 **[거래처코드 : CU021] 칸이
#   실제로 적혀 들어오기 시작했다**(실측 2026-09-02 - 35글 - 캠프 22곳).
#
#   * 그런데 **읽는 코드가 한 줄도 없었다** - 이 파일에 "거래처코드" 라는 낱말이
#     아예 없었다. 사람이 적어 보내는데 기계가 안 읽으면 **없는 것과 같다**([169]).
#     그래서 camp_contacts 는 여전히 이름/주소로 **추정**만 하고 있었다([426]).
#
#   * **모양을 CU 로 못 박지 않는다**([165]) - ERP 거래처코드 접두가 늘어나는 날
#     그 갈래만 조용히 안 걸리면서 오류도 안 난다. 대신 글자 1~3 + 숫자 2~6 으로
#     좁혀 문장 속 아무 낱말이 걸리지 않게 한다.
#   * 값이 비면([거래처코드 : ] 뒤가 빈 칸) **안 걸린다** - 빈 칸을 코드로 읽지 않는다.
RE_POST_CODE = re.compile(r"거래처\s*코드\s*[:：]?\s*([A-Za-z]{1,3}\d{2,6})")
RE_TEL = re.compile(r"01[016789][-. ]?\d{3,4}[-. ]?\d{4}")
RE_MAIL = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")


def _tel_fmt(raw):
    """번호 표기를 고른다. 자릿수가 아는 모양(10·11)일 때만 끊고, 아니면 원문 그대로
    둔다 — 모르는 모양을 억지로 끊으면 틀린 번호가 확정처럼 보인다."""
    d = re.sub(r"\D", "", raw)
    if len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10:
        return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return re.sub(r"[.\s]", "-", raw)


def _person(m):
    """(이름, 전화, 메일) — **셋 다 있는 것만 참으로 치지 않는다.**
    이름만 적힌 캠프도 있고 번호만 적힌 캠프도 있다. 없는 칸은 빈 문자열로 두고
    **지어내지 않는다** — 빈 칸은 사람이 채울 수 있지만 틀린 번호는 아무도 못 고친다."""
    if not m:
        return None
    name = (m.group(1) or "").strip(" :：-·").strip()
    rest = m.group(2) or ""
    tel = RE_TEL.search(rest)
    # 메일은 번호 줄(①)이 먼저고, 없으면 바로 다음 `E- MAIL` 줄(②)이다. 둘 다 없으면
    # **빈 칸으로 둔다** — 아래 사람의 메일을 끌어오지 않는다([172]).
    mail = RE_MAIL.search(rest)
    if not mail and m.lastindex and m.lastindex >= 3:
        mail = RE_MAIL.search(m.group(3) or "")
    # 표기만 고른다 — 숫자는 한 자도 안 건드린다. 원문에 `01038797505`·`010-71091794`
    # 처럼 하이픈이 없거나 어긋난 것이 실측 6건 있었고, 그대로 두면 대표 보고에서
    # 사람이 번호를 잘못 끊어 읽는다.
    tel = _tel_fmt(tel.group(0)) if tel else ""
    if not name and not tel:
        return None
    return {"이름": name, "전화": tel, "메일": mail.group(0) if mail else ""}
TECHS = ("김준형", "권오철", "김필우", "차동호", "김경원")
# 밴드 원문에서 실제 확인된 오탈자. 원문 캐시는 보존하되 구조화 결과에는
# 기준 기사명만 기록해 관리대장·기사별 집계로 오탈자가 전파되지 않게 한다.
TECH_ALIASES = {
    "권오절": "권오철",
    "권오처르": "권오철",
}


# 사람 이름처럼 생겼지만 사람이 아닌 것들. 알고리즘으로는 '하이테크'와 '엄진언'을
# 가를 수 없어 목록으로 둔다(2026-07-28 실데이터에서 확인).
VENDORS = ("하이테크", "대신택배")
# 이름에 붙여 쓴 직책 — '김승기기장' 처럼 띄어쓰기 없이 오는 경우가 있다
TITLES = ("기장", "기사", "과장", "차장", "부장", "대리", "팀장", "소장", "실장", "반장")
# 미배정을 뜻하는 자리표시자
PLACEHOLDER = re.compile(r"^(0{2,}|-+|자\)|미배정|미정|없음)$")


def _looks_like_name(t):
    return bool(re.fullmatch(r"[가-힣]{2,4}", t))


def normalize_tech(raw, when=""):
    """기사명만 남긴다 — 설명·직책·업체·자리표시자는 걸러낸다.

    ★ 2026-07-28 실사고: 이 함수가 '첫 조각'을 그대로 통과시켜
      `000 (캠프상태확인 및 스케쥴 세팅)`·`자) - 각캠프담당자 …` 같은 **작업 메모가
      담당기사 칸에 그대로 들어갔다.** 대표보고 'TOP 5' 에 `담당: 000 (…)` 로 노출됐고,
      기사별 집계도 오염됐다. 걸러 주는 clean_tech 는 리포트에만 쓰이고 있었다.
      버린 원문은 호출 쪽에서 비고에 남긴다(tech_note) — 정보를 없애지는 않는다."""
    cleaned = str(raw or "").strip()
    for wrong, right in TECH_ALIASES.items():
        cleaned = cleaned.replace(wrong, right)
    # ★ 그만둔 사람 이름이 양식 문구를 타고 담당 칸에 들어온다(2026-08-08 실측 35건).
    #   원문은 두고 **읽을 때만** 지금 담당자로 옮긴다 — people_alias 가 근거를 갖는다.
    for _h in _ALIAS.HANDOVERS:
        if _h["before"] in cleaned:
            cleaned = cleaned.replace(_h["before"],
                                      _ALIAS.resolve_person(_h["before"], when=when))
    tech = ", ".join(t for t in TECHS if t in cleaned)
    if tech:
        return tech
    # ★ 조각을 하나씩 보면 '스케쥴'·'체크' 같은 낱말도 이름처럼 생겨서 통과한다.
    #   그래서 **칸 전체를 먼저 판단한다** — 낱말이 넷 이상이면 이름이 아니라 문장이다.
    cleaned = re.sub(r"\(.*?\)", " ", cleaned)          # 괄호 설명은 통째로 뗀다
    cleaned = re.sub(r"\.{2,}.*$", " ", cleaned)
    tokens = [t for t in re.split(r"[,·+/\s]+", cleaned) if t]
    if not tokens or len(tokens) > 3:
        return ""
    names = []
    for part in tokens:
        for t in TITLES:                                # 김승기기장 → 김승기
            if part.endswith(t) and _looks_like_name(part[: -len(t)]):
                part = part[: -len(t)]
                break
        if PLACEHOLDER.match(part) or part in VENDORS:
            continue
        if _looks_like_name(part):
            names.append(part)
    return ", ".join(dict.fromkeys(names))


# ── 접수 취소 ────────────────────────────────────────────────────────────
# 사용자 지시(2026-08-08): **"접수 했다가 접수 취소하는 경우도 많은데 이것도
# 잡아내는 알고리즘 추가해"**
#
# 실제 사례(밴드 댓글): "통화 완료 했습니다 / 작동 원활함. 접수 취소 하세요"
#
# ★ 예전 규칙은 `"접수취소" in 본문` **한 줄**이었다. 두 군데서 새어 나갔다:
#   ① **띄어쓰기** — 사람은 '접수 취소'라고 쓴다. 붙여 쓴 것만 잡고 있었다.
#   ② **자리** — 취소는 본문이 아니라 **댓글**로 온다. 접수 글은 이미 올라간
#      뒤이므로 고칠 것이 댓글밖에 없다. 본문만 보면 영영 못 본다.
#   그래서 취소된 건이 '돌발AS 미처리'로 남아 AS 미실시 숫자를 계속 부풀렸다.
# ★★ '취소'라는 낱말만으로 판정하면 **멀쩡한 건이 죽는다.** 실측에서 그대로 나왔다:
#      "바디부분 아크릴판은 캠프담당 취소요청함" · "택배발송 취소요청하심"
#    둘 다 부품·택배 취소지 AS 접수 취소가 아니다. 취소로 처리하면 그 현장은
#    아무도 안 가는데 목록에서도 사라진다 — 미실시로 남는 것보다 나쁘다.
#    그래서 **'취소' 곁에 '접수'나 A/S 가 있을 때만** 접수 취소로 본다.
#    ★ '취소' 곁에 A/S 가 있으면 되게 했더니 **모든 글이 걸렸다** — 밴드 양식에
#      `● A/S 완료 :` 줄이 늘 따라붙기 때문이다(실측 2620: '택배발송 취소요청하심'
#      바로 뒤가 그 줄이었다). 그래서 근거는 **'접수'가 '취소'에 붙어 있는 것** 하나다.
_CANCEL = re.compile(
    r"(접\s*수\s*[^가-힣A-Za-z0-9\n]{0,3}(를|건|은|이)?\s*(요\s*청)?\s*취\s*소"  # 접수(를/건) 취소
    r"|취\s*소\s*[^가-힣A-Za-z0-9\n]{0,3}(된|할|하실)?\s*접\s*수"                # 취소 접수
    r"|오\s*접\s*수"                                   # 오접수
    r"|중\s*복\s*접\s*수"                              # 중복접수
    r"|접\s*수\s*(를\s*)?(철\s*회|반\s*려))")          # 접수 철회/반려
# 취소가 아닌데 '취소'가 들어간 말 — 걸러 내지 않으면 멀쩡한 건이 취소로 죽는다.
_NOT_CANCEL = re.compile(r"(접\s*수\s*취\s*소\s*(불\s*가|없|아\s*님|안\s*됨|보\s*류)"
                         r"|접\s*수(?:는|를)?\s*취\s*소\s*(하\s*지|안\s*함)"
                         r"|(?:보\s*험|택\s*배|부\s*품|예\s*약|주\s*문|발\s*송)\s*접\s*수\s*취\s*소"
                         r"|접\s*수\s*유\s*지|취\s*소\s*된\s*건\s*없)")

# 취소를 **되돌린** 말도 하나의 상태 사건이다. 예전처럼 댓글을 한 덩어리로
# 합친 뒤 `_NOT_CANCEL.search()`를 하면, 오래된 "접수 취소"와 나중의 "접수 유지"가
# 같은 문자열에 있다는 이유만으로 둘 다 사라졌다. 이제 문장별·댓글별 사건을 시간순으로
# 놓고 마지막 명시 상태가 이긴다(2026-08-11 엄격검토).
_ACTIVE = re.compile(
    r"(접\s*수\s*(?:는\s*)?(?:유\s*지|계\s*속|진\s*행|재\s*개)"
    r"|접\s*수(?:는|를)?\s*취\s*소\s*(?:하\s*지\s*않|안\s*함|안\s*하|아\s*님|불\s*가|안\s*됨|보\s*류)"
    r"|접\s*수\s*취\s*소\s*(?:철\s*회|해\s*제|번\s*복)"
    r"|취\s*소\s*(?:철\s*회|해\s*제|번\s*복)"
    r"|취\s*소\s*된\s*건\s*없)"
)
_NON_SERVICE_CANCEL = re.compile(
    r"(?:보\s*험|택\s*배|부\s*품|예\s*약|주\s*문|발\s*송)\s*접\s*수\s*취\s*소"
)


def cancel_state(text):
    """한 본문/댓글 안의 마지막 명시 상태: ``cancel``·``active``·빈 문자열.

    보험·택배·부품 접수 취소는 서비스 접수 상태가 아니므로 사건에서 제외한다.
    같은 댓글 안에 정정이 함께 적혀도 **뒤에 쓴 말**이 이긴다. 이 함수는 한 댓글만
    판단하고, 댓글 사이의 시간 순서는 :func:`latest_cancel_event`가 맡는다.
    """
    s = str(text or "")
    if not s:
        return ""
    events = []
    excluded = list(_NON_SERVICE_CANCEL.finditer(s))
    for match in _CANCEL.finditer(s):
        # "보험접수 취소"처럼 다른 종류의 접수 취소와 겹친 매치는 버린다.
        if any(not (match.end() <= bad.start() or match.start() >= bad.end())
               for bad in excluded):
            continue
        # 바로 뒤의 부정·보류는 취소가 아니라 active 사건으로 아래 _ACTIVE가 잡는다.
        tail = s[match.start():min(len(s), match.end() + 16)]
        if _NOT_CANCEL.search(tail):
            continue
        events.append((match.start(), match.end(), "cancel"))
    for match in _ACTIVE.finditer(s):
        events.append((match.start(), match.end(), "active"))
    if not events:
        return ""
    events.sort(key=lambda item: (item[0], item[1]))
    return events[-1][2]


def cancel_hit(text):
    """이 글/댓글이 **접수 취소**를 말하고 있나.

    부품·택배 취소와 갈라야 한다(위 주석). 근거는 '접수'가 '취소'에 붙어 있는 것,
    그리고 오접수·중복접수·접수철회다. 그냥 '취소'는 판정하지 않는다 —
    취소로 잘못 처리하면 그 현장은 아무도 안 가는데 목록에서도 사라진다.
    """
    return cancel_state(text) == "cancel"


def _event_epoch(value):
    """밴드 밀리초·초·ISO 시각을 정렬 가능한 epoch 초로 바꾼다."""
    if value in (None, ""):
        return None
    try:
        number = float(value)
        return number / 1000.0 if number > 10_000_000_000 else number
    except (TypeError, ValueError):
        pass
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).timestamp()
    except (TypeError, ValueError):
        return None


def latest_cancel_event(post):
    """본문과 댓글을 시간순으로 놓고 마지막 접수상태 사건을 돌려준다.

    ``{state, source, text, created_at, epoch, order}`` 모양이다. 시각 없는 댓글은
    수집 단계의 원칙과 같이 버린다. 본문은 게시시각, 댓글은 댓글시각을 사용하므로
    오래된 취소 댓글 뒤의 최신 "접수 유지"가 정상적으로 취소를 해제한다.
    """
    if not isinstance(post, dict):
        return None
    events = []
    body = str(post.get("content") or "")
    body_state = cancel_state(body)
    body_epoch = _event_epoch(post.get("created_at"))
    if body_state:
        events.append({"state": body_state, "source": "본문", "text": body,
                       "created_at": post.get("created_at"), "epoch": body_epoch,
                       "order": 0})
    for index, comment in enumerate(post.get("comments") or [], 1):
        if not isinstance(comment, dict):
            continue
        text = str(comment.get("content") or comment.get("body") or "")
        state = cancel_state(text)
        epoch = _event_epoch(comment.get("created_at"))
        if not state or epoch is None:  # 시각 없는 댓글로 순서를 지어내지 않는다.
            continue
        events.append({"state": state, "source": "댓글", "text": text,
                       "created_at": comment.get("created_at"), "epoch": epoch,
                       "order": index})
    if not events:
        return None
    # 본문 시각이 결측이면 알려진 댓글보다 앞선 것으로만 취급한다.
    events.sort(key=lambda event: (
        event.get("epoch") is not None,
        event.get("epoch") if event.get("epoch") is not None else float("-inf"),
        event.get("order", 0),
    ))
    return events[-1]


def comment_text(post):
    """글에 달린 댓글 본문을 한 덩어리로 — 캐시에 댓글이 없으면 빈 문자열.

    ★ 캐시 모양은 `comments: [{author, created_at, content}]` 하나다 (2026-08-08부터
      화면 긁기·API 양쪽이 같은 모양으로 담는다). 담는 쪽이 둘이라 **읽는 쪽은
      반드시 하나**여야 한다 — 갈리면 한쪽만 고쳐지고 다른 쪽은 조용히 옛것으로 남는다.
      수집 자체는 'CSOS 리서치 및 자료 수집' 세션 몫이다(CLAUDE.md).
      적힌 수만큼 못 읽은 글은 `cancel_blind_count()` 가 센다 —
      "댓글은 있는데 못 읽는다"를 조용히 넘기지 않기 위해서다.
    """
    if not isinstance(post, dict):
        return ""
    out = []
    for c in (post.get("comments") or []):
        if isinstance(c, dict):
            out.append(str(c.get("content") or c.get("body") or ""))
        else:
            out.append(str(c or ""))
    return "\n".join(out)


def cancel_blind_count(posts):
    """댓글이 달렸는데 **다 못 읽은** 글 수 — 취소를 놓칠 수 있는 사각지대의 크기.

    ★ '하나도 못 읽음'이 아니라 '적힌 수만큼 못 읽음'으로 센다 (2026-08-08).
      접힌 댓글을 한 개만 펴서 담은 글은 본문이 있으니 예전 기준으로는 안 걸렸는데,
      정작 취소 통보는 **못 편 그 댓글**일 수 있다. 반쯤 읽은 것을 다 읽은 것으로
      세면 사각지대가 0으로 보인다 — 제일 나쁜 종류의 안심이다.

    ★★ 그런데 그 세는 법마저 `comment_count` 에 기대고 있었다 (2026-08-08 저녁 실측).
      캐시 10,312글 중 `comment_count>0` 은 **6글**이고 댓글 본문은 **0글**이다.
      밴드에 댓글이 없어서가 아니라 **수집기가 그 숫자를 안 담아서**다. 그래서
      사각지대 계기가 "사각지대 0" 이라고 말했다 — **재는 도구가 같은 결측에 눈이
      멀어 있었다.** 없는 것과 안 본 것을 구별하지 못하면 계기는 늘 안심을 준다.
      이제 셋을 가른다:
        · `comments` 키가 **아예 없다**  → 한 번도 안 들여다봤다 → **사각지대**
        · `comments: []`                → 보긴 봤고 없었다 → 사각지대 아님
        · 적힌 수보다 적게 담겼다        → 반쯤 읽었다 → 사각지대(예전 기준)
    """
    n = 0
    for p in (posts or {}).values():
        if not isinstance(p, dict):
            continue
        if "comments" not in p:
            # 들여다본 적이 없다. comment_count 가 0 이어도 그 0 을 믿을 근거가 없다.
            n += 1
            continue
        try:                       # 캐시에 따라 숫자가 문자열로 들어 있다
            cnt = int(str(p.get("comment_count") or 0).strip() or 0)
        except ValueError:
            cnt = 0
        if cnt <= 0:
            continue
        got = len([c for c in (p.get("comments") or []) if isinstance(c, dict)])
        if got < cnt and not p.get("comments_full"):
            n += 1
    return n


def tech_note(raw, kept):
    """정규화하며 버린 부분 — 호출 쪽이 비고에 남겨 정보를 잃지 않게 한다."""
    raw = str(raw or "").strip()
    return "" if not raw or raw == kept else raw


#: 게시 양식의 머리 — `♣ ［ 돌발유료 A/S 완료 ]` 처럼 생겼다.
RE_FORM_HEAD = re.compile(r"♣\s*[［\[]")


def split_forms(c):
    """한 글에 붙어 있는 **게시 양식마다** 본문을 가른다.

    ★ 2026-08-21 형님 지시("완료건이 자꾸 캘린더 캡처에 표시된다 · AS기사들
      곤란해지고 있음")에서 나온 뿌리다.  화면 긁기가 이웃 글까지 한 덩어리로 담는
      일이 있는데(실측 밴드 글 2,214개 중 159개 · 7.2%), `parse_post` 는 프로젝트
      번호를 **맨 앞 하나만** 읽었다.  그래서 뒤에 붙은 408개 번호가 기계에 아예
      안 보였고 그중 108개는 그 글에 **'완료' 제목이 있었다** — 다녀와서 밴드에
      완료를 올렸는데도 미처리로 서고, 그것이 대표 보고에 그대로 실렸다.
    ★ **양식 머리(`♣ ［…]`)로만 가른다.**  프로젝트 번호로 가르면 '미실시 AS 공유'
      같은 **목록 글**의 번호들이 저마다 게시글이 되어, 안 한 일이 한 것으로도
      한 일이 안 한 것으로도 뒤집힌다([172]).  양식이 하나뿐이면 예전 그대로다.
    ⚠ 두 번째 양식부터는 그 앞에 붙는 문서 표시(`✅판매전표 …`)가 **앞 덩어리 끝에**
      남는다.  진행상태·프로젝트NO·작업일은 양식 **안**에 있어 정확하지만 `문서상태`
      는 첫 양식 쪽으로 기운다 — 아는 한계라 여기 적어 둔다([169]).
    """
    heads = [m.start() for m in RE_FORM_HEAD.finditer(c or "")]
    if len(heads) <= 1:
        return [c]
    # 첫 덩어리는 머리 앞의 문서 표시까지 안고 간다(0 부터).
    bounds = [0] + heads[1:] + [len(c)]
    return [c[a:b] for a, b in zip(bounds[:-1], bounds[1:])]


def parse_post_all(no, p, band):
    """한 글에서 **모든** 게시 양식을 읽는다. 하나뿐이면 `parse_post` 와 같다.

    ★ `parse_post` 의 반환 모양은 한 글자도 안 바꿨다 — 읽는 곳이 여럿이다
      (`cancel_watch`·`cross_signal`·`app_server` 완료색인 · `comment_backfill`).
    """
    c = p.get("content") or ""
    forms = split_forms(c)
    if len(forms) <= 1:
        r = parse_post(no, p, band)
        return [r] if r else []
    out, seen = [], set()
    for i, part in enumerate(forms):
        q = dict(p)
        q["content"] = part
        r = parse_post("%s#%d" % (no, i) if i else no, q, band)
        if not r:
            continue
        key = (r.get("프로젝트NO"), r.get("진행상태"), r.get("작업일"))
        if key in seen:            # 같은 양식이 두 번 담긴 글 — 한 번만 센다
            continue
        seen.add(key)
        out.append(r)
    return out


def form_field(rx, c):
    """양식 머리(`♣ ［…]`) **뒤**의 값이 그 양식의 주인공이다.

    ★ 2026-08-21 실사고(형님 지시 "처리됐는데 미처리로 뜼는 건 해결").
      완료 글 머리말에 **참조용 다른 프로젝트NO** 를 적어 두는 일이 있다:
        `✔️ 동일 캠프 돌발AS와 동시 진행 / ● 프로젝트NO : UJ2600756`
        `♣ ［ 돌발유료 A/S 완료 ] … ● 프로젝트NO : UJ2600793`
      그런데 `RE_PRJ.search(c)` 는 **맨 앞 = 참조 번호**를 읽어, 완료 글이
      **엉뚝한 프로젝트에 붙었다.** 실측 17건 · 그중 15건이 완료 글이다.
    ★ **두 방향으로 다 틀린다.** 다녀온 현장이 미처리로 서고(형님이 짚으신
      그 문제) **동시에 안 간 현장이 완료로 찍힌다** — 뒤쪽이 더 나쁘다([172]).
    ★ **머리 뒤에 값이 없거나 비면 예전 그대로**다 — 빈 값으로 덮으면 알던 것을
      잃는다([326]). 머리가 없는 글(댓글·메모)도 예전 그대로라 넓히는 것이 아니다.
    """
    m = RE_FORM_HEAD.search(c or "")
    if m:
        got = rx.search(c, m.end())
        if got and (got.lastindex is None or (got.group(1) or "").strip()):
            return got
    return rx.search(c or "")


def form_prj(c):
    """이 글의 **주인공** 프로젝트NO. 밴드 본문을 읽는 곳은 전부 이것을 쓴다([162]).

    대상이 댓글·카톡처럼 양식 머리가 없으면 `RE_PRJ.search` 와 같다.
    """
    return form_field(RE_PRJ, c)


def parse_post(no, p, band):
    c = p.get("content") or ""
    prj = form_prj(c)          # ★ 양식 머리 뒤가 주인공([172])
    title = (RE_TITLE.search(c).group(1).strip() if RE_TITLE.search(c) else "")
    if not prj and not title:
        return None                       # 업무 게시글이 아님(공지·자료 등)

    md = RE_DATE.search(c)
    work_date = ""
    if md:
        y, mo, d = int(md.group(1)), int(md.group(2)), int(md.group(3))
        work_date = f"{y:04d}-{mo:02d}-{d:02d}" if mo and d else ""   # 2026.00.00 = 미정

    prj_no = prj.group(1) if prj else ""
    if prj_no and set(prj_no[2:]) == {"0"}:      # UJ000000 = 양식 템플릿 게시글
        return None

    ts = p.get("created_at")
    posted = datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d") if ts else ""

    tech_raw = (RE_TECH.search(c).group(1).strip() if RE_TECH.search(c) else "")
    # 게시일을 함께 넘긴다 — 인계 **전** 글은 그때 담당자 그대로 둬야 한다.
    tech = normalize_tech(tech_raw, when=posted)

    _mc = form_field(RE_CAMP, c)   # 서두 메모가 캠프명을 가로챈다(실측 5건)
    camp = _mc.group(1).strip() if _mc else ""
    camp = re.sub(r"\s*\.{3}더보기.*$", "", camp).strip()

    # 업무유형·유상무상·상태
    if "정기점검" in title or "3개월" in title or "분기" in title:
        kind = "정기점검"
    elif "돌발" in title:
        kind = "돌발AS"
    elif "설치" in title or "납품" in title:
        kind = "신규납품설치"
    else:
        kind = "기타"
    if "동시" in title or "동시진행" in c:
        kind += "(동시진행)"
    cost = "유상" if "유료" in title else ("무상" if "무료" in title else "")
    # ★ 본문과 댓글을 한 문자열로 합치지 않는다. 각 댓글의 시각을 읽어 마지막 명시
    #   상태가 이긴다. 단 완료 제목과 같은 시각의 본문 취소는 기존 안전규칙대로 완료에
    #   양보한다(실측 4979: 일부 작업만 취소하고 실제 작업은 완료).
    state_event = latest_cancel_event(p)
    if state_event and state_event["state"] == "cancel" and state_event["source"] == "댓글":
        status = "취소"
    elif "완료" in title:
        status = "작업완료"
    elif state_event and state_event["state"] == "cancel":
        status = "취소"
    elif "안내" in title:
        status = "접수·예정"
    else:
        status = ""

    docs = [d for d, kw in (("판매전표", "판매전표"), ("거래명세서", "거래명세서"),
                            ("견적서", "견적서"), ("메일발송", "메일발송")) if kw in c]
    # ★ 캠프 담당자는 **기존 칸을 하나도 안 건드리고** 옆에 더한다 — 이 반환값을 읽는
    #   곳이 여럿이다(cancel_watch · cross_signal · app_server 완료색인).
    addr = RE_CAMP_ADDR.search(c)
    _pc = form_field(RE_POST_CODE, c)   # 양식 머리 뒤가 주인공([172]-[376])
    return {"프로젝트NO": prj_no, "업무유형": kind, "비용구분": cost,
            "작업일": work_date, "담당기사": tech, "캠프명": camp, "진행상태": status,
            "문서상태": "+".join(docs), "사진": p.get("photo_count", 0),
            "게시일": posted, "밴드": band, "게시글": no,
            # ★ **본문을 옆에 더한다** (2026-08-18 지시 "카톡에 메시지에 근거가 있으면
            #   그것도 표시해 캡처화면에"). 지금까지는 파싱한 칸만 돌려줘서, 카톡에
            #   근거가 적혀 있어도 **원문을 보여 줄 길이 없었다** — 화면은 '완료 글이
            #   없다'만 말하고 그 옆 카톡 한 줄은 아무 데도 안 떴다.
            #   길이를 묶는 이유: 이 반환값은 8천 건을 한꺼번에 만든다. 통째로 담으면
            #   색인 한 번에 메모리가 몇십 MB 늘고 디스크 캐시도 같이 부푼다.
            #   자르는 것은 **원문이 아니라 사본**이며, 잘렸다는 것은 읽는 쪽이
            #   `본문잘림` 으로 안다(조용히 자르지 않는다, [273]).
            "본문": c[:800],
            "본문잘림": len(c) > 800,
            "캠프주소": (addr.group(1).strip() if addr else ""),
            # * 원본이 적어 보낸 ERP 거래처코드 - **옆에 더하기만** 한다.
            #   이 반환값을 읽는 곳이 여럿이라 기존 칸의 모양을 안 바꾼다([314]).
            "거래처코드": (_pc.group(1).strip().upper() if _pc else ""),
            "현장책임": _person(RE_SITE_MGR.search(c)),
            "안전관리": _person(RE_SAFE_MGR.search(c)),
            # 옛 양식(직책 미상) — 위 둘과 **다른 칸**이다. 기존 칸은 안 건드린다.
            "담당자": _person(RE_OLD_MGR.search(c))}


BAND_READ_ENV = "COUPANG_BAND_READ"


def band_read_on():
    """밴드 글을 읽을 것인가 (형님 2026-09-04 지시: 카톡만 반영).

    ★ 끄는 것은 **읽기**뿐이다 — 밴드 캐시는 한 글자도 안 지운다([500] 과 같은
      자리: 거기서는 긁는 길만 껐고 여기서는 읽는 길이다). 되돌리려면 이 환경변수를
      지우거나 `include_band=True` 로 부른다.
    ★ 못 읽으면 **켜짐**이다([169] 를 이 자리에 맞게 정한 것). 잘못 끄면 캠프
      연락처를 잃는데(2026-09-04 실측 전화 695 → 308곳) 그것은 다시 만들 수 없고,
      잘못 켜는 값은 옛 근거를 한 번 더 읽는 것뿐이다.
    """
    v = str(os.environ.get(BAND_READ_ENV, "")).strip().lower()
    return v not in ("0", "off", "false", "no")


def load_records(include_band=None):
    """밴드 글 + 카톡 내보내기를 **같은 양식**으로 읽는 한 곳([162]).

    include_band=None 이면 환경변수(`COUPANG_BAND_READ`)가 정한다. 부르는 쪽이
    명시로 주면 그것이 이긴다 — 검증이 목 없이 갈래를 재려고 쓴다([295]).

    ⚠ 밴드를 빼면 무엇을 잃는지 잰 값(2026-09-04 · 관리대장 v630):
      캠프 751 → 360곳 · 캠프 전화 695 → 308곳 · 밴드 사진 4,986 → 0장 ·
      프로젝트NO 1,902 → 1,120개 · 완료 근거 1,636 → 911개.
      그런데 **캘린더 갈래는 한 건도 안 바뀐다**(as_open 68 · as_done 523 ·
      pm_overdue 17 · pm_done 389 그대로) — 밴드 근거는 원장이 빈 것을 닫을 때만
      쓰이는데([244]) 지금 열린 건은 밴드에도 완료 글이 없기 때문이다.
    """
    if include_band is None:
        include_band = band_read_on()
    out = []
    for f in (glob.glob(os.path.join(CACHE_DIR, "*.json")) if include_band else []):
        b = os.path.basename(f)
        if b.startswith(("raw_", "dump_")):
            continue
        d = json.load(open(f, encoding="utf-8"))
        # ★ 밴드 이름은 화면에서 긁힌 문자가 아니라 **파일의 밴드 ID**가 정본이다.
        # 2026-08-21 실측으로 캐시 이름이 깨진 문자 또는 단순히 `밴드 홈`으로
        # 저장돼, 쿠팡AS와 매출처업무의 수집 기준일이 서로 섞였다. 파일명은 두
        # 밴드의 영구 ID라 인코딩·화면 제목과 무관하게 안정적이다. 알려진 ID만
        # 사람 이름으로 정규화하고, 모르는 밴드는 원문 이름을 그대로 보존한다.
        band_id = os.path.splitext(b)[0]
        band = {
            "90610953": "(주)유니버셜리프트 쿠팡AS",
            "84789192": "(주)유니버셜리프트 매출처업무",
        }.get(band_id, d.get("band_name", b))
        for no, p in d.get("posts", {}).items():
            # ★ 한 글에 양식이 여럿이면 **전부** 읽는다(2026-08-21) — 맨 앞 하나만
            #   읽던 때는 뒤에 붙은 완료 글이 기계에 아예 안 보였다.
            out.extend(parse_post_all(no, p, band))
    out += load_kakao_records()
    out.sort(key=lambda r: (r["작업일"] or r["게시일"], r["프로젝트NO"]))
    return out


KAKAO_ROOM_MARKERS = ("쿠팡돌발점검", "쿠팡정기점검")
KAKAO_SPAN_TRUNCATED = False
KAKAO_SELECTION_STATUS = {}

def _span_cache_path():
    """카톡 원본 구간 캐시의 자리. **`REPORT_DIR` 에서 온다.**

    2026-08-25 실사고(분담판 [227]): 여기가 `BASE_DIR/reports` 를 직접 적어 두어
    `REPORT_DIR` 을 **안 봤다.** 그래서 한 모듈 안에서 리포트 자리가 둘로 갈렸다 —
    지문(`_kakao_selection_trigger`)과 빠른 길은 `REPORT_DIR` 을 읽는데 캐시만
    옛 자리를 읽고 썼다. 결과가 둘이다:
      · 운영: `COUPANG_REPORT_DIR` 로 자리를 옮기면 새 자리의 지문에 **옛 자리의
        결과**가 조용히 실린다 — 오류도 안 나고 목록도 그럴듯하다([165]).
      · 관문: `t201` 이 `REPORT_DIR` 을 격리해도 `_extend_early` 가 **실제** 구간
        캐시의 `__결과__` 에서 이전 경로를 이어 붙여 **실데이터가 섞였다**
        (2026-08-25 실측 합성 2 + 실제 6 = 8). 그래서 형님이 그날 카톡을 올리기만
        해도 관문이 빨개졌고, 그 관문은 `daily_run` 의 0단계라 **그날 아침 대조가
        통째로 안 돌았다**. 자료를 넣었다고 빨개지는 관문은 아무도 안 믿는다
        ([170]) — 그리고 검사가 실측 증거를 읽고 쓰면 안 된다([247]).
    ★ **상수가 아니라 함수여야 한다.** 모듈 상수는 import 때 굳어서, 부르는 쪽이
      `REPORT_DIR` 을 옮겨도 따라오지 못한다([371] — 모듈 전역은 프로세스의 것이다).
    ★ 기본값은 한 글자도 안 바뀐다 — `REPORT_DIR` 의 기본이 `BASE_DIR/reports` 다([172]).
    """
    return os.path.join(REPORT_DIR, ".카톡_원본구간.json")


def _selection_cache_path():
    """카톡 원본 선택 캐시의 자리. 근거는 `_span_cache_path` 와 같다([227])."""
    return os.path.join(REPORT_DIR, ".카톡_원본선택.json")


def _selection_status_path():
    """자국 파일을 실제로 찾았는지 남기는 자리 — 캐시와 같은 REPORT_DIR 을 쓴다."""
    return os.path.join(REPORT_DIR, "카톡_원본선택_상태.json")


def _kakao_selection_complete(recent_names, named_hits):
    """흡수 자국이 말한 파일을 모두 찾았을 때만 그 선택을 캐시해도 된다."""
    expected = {os.path.basename(str(name)) for name in (recent_names or []) if str(name)}
    found = {os.path.basename(str(name)) for name in (named_hits or []) if str(name)}
    return not expected or expected <= found


def _write_kakao_selection_status(trigger, recent_names, named_hits, cache_saved):
    """못 읽은 답이 그럴듯한 빈 목록으로 보이지 않게 원인을 원자 저장한다([169])."""
    expected = {os.path.basename(str(name)) for name in (recent_names or []) if str(name)}
    found = {os.path.basename(str(name)) for name in (named_hits or []) if str(name)}
    missing = sorted(expected - found)
    row = {
        "시각": datetime.now().isoformat(timespec="seconds"),
        "trigger": trigger,
        "정상": not missing,
        "자국파일수": len(expected),
        "찾은파일수": len(expected & found),
        "못찾은파일": missing,
        "캐시저장": bool(cache_saved),
        "왜": ("" if not missing else
               "카톡 흡수 자국이 말한 파일 %d개를 원본 선택기가 못 찾았습니다 — "
               "최신일은 사람 미보고가 아니라 원본 폴더를 다 못 읽은 값일 수 있습니다"
               % len(missing)),
    }
    global KAKAO_SELECTION_STATUS
    KAKAO_SELECTION_STATUS = row
    try:
        os.makedirs(os.path.dirname(_selection_status_path()), exist_ok=True)
        tmp = _selection_status_path() + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(row, fh, ensure_ascii=False)
        os.replace(tmp, _selection_status_path())
    except OSError:
        pass
    return row


def kakao_selection_status():
    """현재 선택 규칙·흡수 자국에 맞는 상태만 돌려준다(옛 경보 재사용 금지)."""
    trigger = _kakao_selection_trigger()
    if KAKAO_SELECTION_STATUS.get("trigger") == trigger:
        return dict(KAKAO_SELECTION_STATUS)
    try:
        with open(_selection_status_path(), encoding="utf-8") as fh:
            row = json.load(fh)
    except (OSError, ValueError, TypeError):
        return {}
    return dict(row) if isinstance(row, dict) and row.get("trigger") == trigger else {}


def _kakao_selection_trigger():
    """새 카톡 흡수·로컬 원본·선택 규칙이 바뀌었는지만 로컬에서 잰다.

    Z: 파일은 날짜·시각 이름으로 보관되어 제자리 수정하지 않는 원본이다. 새 원본은 반드시
    `download_intake` 또는 `카톡_반영회차` 자국을 남긴다. 그 두 자국과 로컬 inbox,
    선택 코드가 그대로면 이전 경로 목록도 그대로다. 이 검사를 Z: `stat`보다 먼저 해야
    화면 갱신 때 SMB 왕복 6회를 되풀이하지 않는다([168]).
    """
    base = os.path.dirname(os.path.abspath(__file__))
    h = hashlib.sha256()
    try:
        for path in (os.path.join(REPORT_DIR, "download_intake.json"),
                     os.path.join(REPORT_DIR, "카톡_반영회차.json"),
                     os.path.abspath(__file__)):
            h.update(os.path.basename(path).encode("utf-8"))
            try:
                with open(path, "rb") as fh:
                    h.update(fh.read())
            except OSError:
                h.update(b"<missing>")
        local_rows = []
        for path in glob.glob(os.path.join(KAKAO_INBOX, "**", "*.txt"), recursive=True):
            st = os.stat(path)
            local_rows.append((os.path.normcase(os.path.abspath(path)), int(st.st_size),
                               int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))))
        h.update(json.dumps(sorted(local_rows), ensure_ascii=False,
                            separators=(",", ":")).encode("utf-8"))
    except OSError:
        return ""
    return h.hexdigest()


def _load_kakao_selection(trigger, dedupe_content):
    if not trigger:
        return None
    try:
        with open(_selection_cache_path(), encoding="utf-8") as fh:
            raw = json.load(fh)
    except (OSError, ValueError, TypeError):
        return None
    key = "dedupe" if dedupe_content else "all"
    row = raw.get(key) if isinstance(raw, dict) and raw.get("trigger") == trigger else None
    paths = row.get("paths") if isinstance(row, dict) else None
    return list(paths) if isinstance(paths, list) and paths else None


def _save_kakao_selection(trigger, dedupe_content, paths):
    if not trigger or not paths:
        return
    try:
        with open(_selection_cache_path(), encoding="utf-8") as fh:
            raw = json.load(fh)
        if not isinstance(raw, dict) or raw.get("trigger") != trigger:
            raw = {"version": 1, "trigger": trigger}
    except (OSError, ValueError, TypeError):
        raw = {"version": 1, "trigger": trigger}
    key = "dedupe" if dedupe_content else "all"
    raw[key] = {"paths": list(paths), "saved_at": datetime.now().isoformat(timespec="seconds")}
    try:
        os.makedirs(os.path.dirname(_selection_cache_path()), exist_ok=True)
        tmp = _selection_cache_path() + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(raw, fh, ensure_ascii=False)
        os.replace(tmp, _selection_cache_path())
    except OSError:
        pass


def _kakao_span(path, cache):
    """그 원본이 **어느 방의 언제부터**를 담고 있나 — 앞부분만 읽는다.

    카톡 내보내기는 '방의 전체 대화'가 아니라 **내보낸 사람이 그 방에 들어온 뒤**를
    담는다. 실측 2026-08-24: 같은 `쿠팡돌발점검` 방인데 `쿠팡돌발점검_25.txt` 는
    2025-12-08 부터, 형님이 오늘 내보낸 것은 **2026-07-20**(그 방에 초대된 날)부터다.
    그래서 '방별 최신 하나'만 고르면 그 이전이 통째로 사라지는데 **오류도 안 나고
    건수도 나온다**([169]). 실측 피해: 카톡 완료 근거 130 -> 901 · 미처리 83건 중
    '카톡에 글이 있다'가 20 -> 79 였다(59건이 '글이 없다'로 잘못 분류돼 있었다).

    비싼 것은 파일 열기다(Z: 는 파일당 왕복 한 번) — 캐시 검사가 먼저다([168]).
    """
    try:
        st = os.stat(path)
    except OSError:
        return None
    key = os.path.normcase(os.path.abspath(path))
    sig = [st.st_size, int(st.st_mtime)]
    hit = cache.get(key)
    if isinstance(hit, dict) and hit.get("sig") == sig:
        return hit
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as fh:
            head = fh.read(4000)
    except OSError:
        return None
    room = next((w for w in KAKAO_ROOM_MARKERS if w in head), "")
    m = re.search(r"-{3,}\s*(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", head)
    rec = {"sig": sig, "방": room,
           "시작": ("%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3)))
                    if m else "")}
    cache[key] = rec
    return rec


def _extend_early(paths, folders, per_room=3):
    """방마다 **더 이른 구간**을 덮는 원본을 보탠다.

    최신본은 그대로 둔다(끝이 가장 늦다). 시작일이 더 이른 것을 최신순으로 보태되
    보탤 때마다 그 방의 '가장 이른 시작일'을 낮춘다 — 이미 덮인 구간은 안 보탠다.

    ★ 시작일을 **못 읽은 파일은 안 보탠다**([169]). 모르는 것을 '더 이르다'고
      우기면 엉뚱한 누적본이 딸려 들어와 대조가 남의 방 글을 읽는다([172]).
    ★ 방마다 `per_room` 개까지만 — 누적본이 쌓인 폴더에서 폭주하지 않게.
    """
    if not folders:
        return paths
    try:
        cache = json.load(open(_span_cache_path(), encoding="utf-8"))
        if not isinstance(cache, dict):
            cache = {}
    except (OSError, ValueError):
        cache = {}
    before = dict(cache)

    floor, out, seen = {}, list(paths), set()
    for p in paths:
        seen.add(os.path.normcase(os.path.abspath(p)))
        rec = _kakao_span(p, cache)
        if rec and rec.get("방") and rec.get("시작"):
            cur = floor.get(rec["방"])
            if cur is None or rec["시작"] < cur:
                floor[rec["방"]] = rec["시작"]

    # 비싼 것은 파일 열기가 아니라 **Z: 재귀 glob** 이다 — 실측 2026-08-24 로
    # 구간 캐시가 다 찬 뒤에도 162초였다. 그래서 **결과 자체**를 캐시한다.
    # 입력이 같으면 결과도 같고, 새 원본이 들어오면 회차 자국이 바뀌어 입력이
    # 바뀌므로 저절로 무효가 된다([168] — 비싼 탐색은 캐시 검사 뒤에).
    ckey = "|".join(sorted(os.path.normcase(os.path.abspath(p)) for p in paths))
    hit = cache.get("__결과__")
    if (isinstance(hit, dict) and hit.get("열쇠") == ckey
            and time.time() - float(hit.get("때") or 0) < 86400):
        keep = [q for q in (hit.get("경로") or []) if os.path.isfile(q)]
        # 하나라도 사라졌으면 캐시를 믿지 않는다 — 반쪽 목록은 조용히 자료를 줄인다.
        if len(keep) == len(hit.get("경로") or []) and keep:
            return keep

    # 최신 내보내기 두 파일이 바뀌어도 **과거 구간은 없어지지 않는다**. 이전 결과는
    # 아래 전수 탐색을 끝낸 때만 저장되므로, 거기에 들어 있던 더 이른 구간을 새 최신본에
    # 이어 붙이면 Z: 전체를 다시 훑을 이유가 없다. 2026-08-24 실측으로 새 카톡 두 개가
    # 들어올 때마다 이 재귀 탐색이 60~162초씩 되살아나 세 소비자가 차례로 기다렸다.
    # 다만 이전 경로가 하나라도 사라졌거나 구간을 못 읽으면 안전하게 기존 전수 탐색으로
    # 물러난다 — 반쪽 목록을 빠른 정답으로 만들지는 않는다([169]).
    previous = hit.get("경로") if isinstance(hit, dict) else None
    if isinstance(previous, list) and previous:
        candidate = list(out)
        candidate_seen = set(seen)
        complete = True
        for q in previous:
            qkey = os.path.normcase(os.path.abspath(str(q)))
            if qkey in candidate_seen:
                continue
            rec = _kakao_span(q, cache)
            if not rec or not rec.get("방") or not rec.get("시작"):
                complete = False
                break
            room, start = rec["방"], rec["시작"]
            cur = floor.get(room)
            if cur is None or start < cur:
                candidate.append(q)
                candidate_seen.add(qkey)
                floor[room] = start
        if complete and all(room in floor for room in KAKAO_ROOM_MARKERS):
            cache["__결과__"] = {"열쇠": ckey, "때": time.time(), "경로": candidate}
            try:
                os.makedirs(os.path.dirname(_span_cache_path()), exist_ok=True)
                tmp = _span_cache_path() + ".tmp"
                with open(tmp, "w", encoding="utf-8") as fh:
                    json.dump(cache, fh, ensure_ascii=False)
                os.replace(tmp, _span_cache_path())
            except OSError:
                pass
            return candidate

    # ★ 로컬 inbox 를 **먼저** 본다. 옛 누적본은 사람이 손으로 여기 넣고,
    #   Z: 정본은 회차가 옮긴 최신본이라 대개 시작일이 같다. 그리고 로컬은 공짜지만
    #   Z: 는 파일 하나가 SMB 왕복 한 번이다([198]).
    local = os.path.normcase(os.path.abspath(KAKAO_INBOX))
    order = sorted(folders,
                   key=lambda f: 0 if os.path.normcase(os.path.abspath(f)) == local else 1)

    # ★ **예산 안에서만 훑는다**([180]·[324]). 첫 실행은 캐시가 비어 파일을 다 열어야
    #   하는데 실측 2026-08-24 로 Z: 79개에 10분이 넘었다 — 그대로 두면 09:50 회차와
    #   앱 요청이 거기서 죽는다(사고 #29). 넘으면 그만 보되 **캐시는 남기므로**
    #   다음 회차가 이어받아 저절로 수렴한다([406] 과 같은 모양).
    budget = 400.0
    try:
        budget = float(os.environ.get("COUPANG_KAKAO_SPAN_BUDGET_S", "400"))
    except ValueError:
        pass
    t0 = time.time()
    global KAKAO_SPAN_TRUNCATED
    KAKAO_SPAN_TRUNCATED = False

    added = {}
    for folder in order:
        # ★ 검사는 glob **앞**에 온다. Z: 재귀 glob 은 한 번 시작하면 못 멈추고
        #   실측 60~140초다 — 파일 루프 안에서만 보면 예산이 아무것도 안 막는다.
        if time.time() - t0 > budget:
            KAKAO_SPAN_TRUNCATED = True
            break
        cand = []
        try:
            found = glob.glob(os.path.join(folder, "**", "*.txt"), recursive=True)
        except OSError:
            continue
        for p in found:
            key = os.path.normcase(os.path.abspath(p))
            if key in seen:
                continue
            seen.add(key)
            try:
                cand.append((os.path.getmtime(p), p))
            except OSError:
                continue
        for _, p in sorted(cand, reverse=True):
            if time.time() - t0 > budget:
                # ★ 조용히 멈추지 않는다([169]) — 다 못 봤다는 사실을 남긴다.
                KAKAO_SPAN_TRUNCATED = True
                break
            rec = _kakao_span(p, cache)
            if not rec:
                continue
            room, start = rec.get("방"), rec.get("시작")
            if not room or not start:
                continue
            if added.get(room, 0) >= per_room:
                continue
            cur = floor.get(room)
            if cur is not None and start >= cur:
                continue
            out.append(p)
            floor[room] = start
            added[room] = added.get(room, 0) + 1
        if KAKAO_SPAN_TRUNCATED:
            break

    # 다 못 본 회차의 결과는 캐시하지 않는다 — 반쪽을 하루 동안 정답으로 쓰게 된다.
    if not KAKAO_SPAN_TRUNCATED:
        cache["__결과__"] = {"열쇠": ckey, "때": time.time(), "경로": list(out)}

    if cache != before:
        try:
            os.makedirs(os.path.dirname(_span_cache_path()), exist_ok=True)
            tmp = _span_cache_path() + ".tmp"
            with open(tmp, "w", encoding="utf-8") as fh:
                json.dump(cache, fh, ensure_ascii=False)
            os.replace(tmp, _span_cache_path())
        except OSError:
            pass
    return out


def kakao_source_paths(dedupe_content=True):
    """공유 정본과 로컬 inbox에서 대화방별 최신 원본을 돌려준다.

    카톡 즉시반영은 원본을 Z: 정본으로 옮긴다. 여기서 로컬 ``kakao/inbox``만
    보면 신규 접수는 원장에 들어가도 완료 보고는 대표보고 색인에 영영 안 들어온다.
    소비자가 제각각 경로를 만들지 않고 ``source_dirs.kakao_dirs()``를 그대로 쓴다.

    카톡 내보내기는 방의 전체 대화를 매번 다시 담는다. 73개 누적본을 Z:에서 전부
    열면 첫 대표보고가 6분 넘게 멈춘다. 파일 메타데이터로 최신순을 만든 뒤, 알려진
    두 업무방의 최신본을 찾는 즉시 멈춘다. 보통 파일 두 개만 열고 끝난다.

    대표보고 캐시 지문은 파일 내용까지 읽을 필요가 없으므로
    ``dedupe_content=False``를 쓴다. 실제 파싱 때만 선택된 최신본끼리 내용 중복을
    제거한다.
    """
    try:
        from source_dirs import kakao_dirs
        folders = list(kakao_dirs())
    except Exception:
        folders = []
    if os.path.isdir(KAKAO_INBOX) and KAKAO_INBOX not in folders:
        folders.append(KAKAO_INBOX)

    selection_trigger = _kakao_selection_trigger()
    selected = _load_kakao_selection(selection_trigger, dedupe_content)
    if selected is not None:
        return selected

    # 방금 흡수한 원본의 **정확한 목적지**가 있으면 그것부터 쓴다. 예전에는 파일명만
    # 읽고 각 이름을 모든 공유폴더 후보에 대입해 `isfile`을 반복했다. SMB 한 번이 수초인
    # 날에는 최신 파일을 이미 알고도 30초 넘게 주소 찾기만 했다. 목적지는 흡수기가 실제
    # 저장한 자국이라 추측이 아니며, 두 업무방이 모두 확인되지 않으면 아래 복구 탐색으로
    # 그대로 물러난다.
    intake_candidates = []
    try:
        base = os.path.dirname(os.path.abspath(__file__))
        intake = json.load(open(os.path.join(REPORT_DIR, "download_intake.json"),
                                encoding="utf-8"))
        for row in intake.get("이동") if isinstance(intake, dict) else []:
            path = str(row.get("목적지") or "") if isinstance(row, dict) else ""
            if path.lower().endswith(".txt") and os.path.isfile(path):
                intake_candidates.append((os.path.getmtime(path), path))
    except (OSError, ValueError, TypeError):
        pass

    intake_rooms = set()
    for _, path in intake_candidates:
        try:
            with open(path, encoding="utf-8-sig", errors="replace") as fh:
                head = fh.read(400)
        except OSError:
            continue
        room = next((word for word in KAKAO_ROOM_MARKERS if word in head), "")
        if room:
            intake_rooms.add(room)
    intake_complete = all(room in intake_rooms for room in KAKAO_ROOM_MARKERS)

    # 정상 반영 회차가 남긴 최신 파일명을 먼저 쓴다. 공유폴더 전체를 훑는 것은
    # 복구용 차선이다. Z:에서 73개 누적본을 열면 6분, 최신 두 개만 열면 수초다.
    recent_names = []
    if not intake_complete:
        try:
            base = os.path.dirname(os.path.abspath(__file__))
            history = json.load(open(os.path.join(REPORT_DIR, "카톡_반영회차.json"),
                                     encoding="utf-8"))
            for run in history if isinstance(history, list) else []:
                names = [os.path.basename(str(name)) for name in (run.get("받은파일") or [])
                         if str(name).lower().endswith(".txt")]
                if names:
                    recent_names = names
                    break
        except (OSError, ValueError, TypeError):
            pass

    # ★ **쓰는 쪽이 실제로 넣는 자리를 빌린다**([162]).  흡수기는
    #   `<카톡폴더>/2026/<이름>` 에 넣는데 아래 `direct` 는 그 자리를 안 봤다 —
    #   실측 2026-08-27 자국이 말한 이름 **2개 중 0개**를 잡았다([277]).
    #   ⚠ 자리를 여기 손으로 적으면 흡수기가 옮긴 날 **그때부터 또 조용히 빠진다.**
    #   ⚠ `create=False` — 읽기가 폴더를 만들면 안 된다.
    canon = ""
    try:
        import kakao_apply
        canon = kakao_apply.canon_dir(create=False) or ""
    except Exception:
        canon = ""

    fast_candidates = list(intake_candidates)
    named_hits = set()
    for name in recent_names:
        for folder in folders:
            direct = [os.path.join(folder, name)]
            stamp = re.search(r"KakaoTalk_(\d{4})(\d{2})(\d{2})_", name)
            if stamp:
                y, m, d = stamp.groups()
                direct.append(os.path.join(folder, y, m, f"{y}-{m}-{d}", name))
            if canon:
                direct.append(os.path.join(canon, name))
            for path in direct:
                if not os.path.isfile(path):
                    continue
                try:
                    fast_candidates.append((os.path.getmtime(path), path))
                except OSError:
                    continue
                named_hits.add(name)

    def choose(candidates):
        markers = KAKAO_ROOM_MARKERS
        chosen, fallback = {}, []
        for _, path in sorted(candidates, reverse=True):
            try:
                with open(path, encoding="utf-8-sig", errors="replace") as fh:
                    head = fh.read(400)
            except OSError:
                continue
            marker = next((word for word in markers if word in head), "")
            if marker:
                chosen.setdefault(marker, path)
                if all(word in chosen for word in markers):
                    break
            elif not fallback:
                fallback.append(path)
        paths = [chosen[word] for word in markers if word in chosen]
        return paths if paths else fallback

    paths = choose(fast_candidates)
    # ★ **자국이 말한 이름을 다 못 찾았으면 그 답을 믿지 않는다**([169]).
    #   예전 문은 `len(paths) < 2` 하나였다 — 옛 파일들이 두 방을 다 채우면 새 원본이
    #   통째로 빠져도 이 문이 안 열리고, 그 덜 읽은 답이 캐시에 박힌다.  실측
    #   2026-08-27 이 그 모양이었다: 자국 2개 중 0개를 잡았는데 옛 파일 여섯이 두 방을
    #   채워 복구 탐색이 안 돌았다 — **오류도 안 나고 개수도 그럴듯했다**([165]).
    #   ⚠ 물러나는 값은 복구 탐색 한 번(실측 144초)이고, 부딪히는 값은 **오늘 자료가
    #     통째로 빠진 채 아무 화면에도 안 뜨는 것**이다.  위 (A)가 맞으면 거의 안 열린다.
    named_blind = bool(recent_names) and len(named_hits) < len(recent_names)
    if len(paths) < 2 or named_blind:
        # 보고 자국이 없거나 한 방만 받은 새 PC는 공유 정본에서 스스로 복구한다.
        candidates, seen_path = list(fast_candidates), set()
        for _, path in fast_candidates:
            seen_path.add(os.path.normcase(os.path.abspath(path)))
        for folder in folders:
            for path in glob.glob(os.path.join(folder, "**", "*.txt"), recursive=True):
                key = os.path.normcase(os.path.abspath(path))
                if key in seen_path:
                    continue
                seen_path.add(key)
                try:
                    candidates.append((os.path.getmtime(path), path))
                except OSError:
                    continue
                if os.path.basename(path) in recent_names:
                    named_hits.add(os.path.basename(path))
        paths = choose(candidates)

    paths = _extend_early(paths, folders)

    # 로컬 inbox 는 사람이 보존한 두 업무방의 **서로 다른 시점 사본**이다. 시작일이 더
    # 이른 파일 하나가 나중 구간을 전부 포함한다고 단정할 수 없다. 실측으로
    # `쿠팡돌발점검_25.txt` 에만 실제 업무 UJ2600067·0068·0189 세 건이 있었고,
    # 더 일찍 시작한 팀채팅 사본에는 없었다. 로컬 18개는 재귀 탐색 비용이 사실상 없으므로
    # 전부 포함하고, 아래 내용 해시로 완전 같은 사본만 한 번 센다. 공유 Z: 99개를 전부
    # 읽는 것보다 빠르면서 실제 업무 세 건을 잃지 않는 경계다([169]·[172]).
    known = {os.path.normcase(os.path.abspath(path)) for path in paths}
    for path in sorted(glob.glob(os.path.join(KAKAO_INBOX, "**", "*.txt"), recursive=True)):
        key = os.path.normcase(os.path.abspath(path))
        if key not in known:
            paths.append(path)
            known.add(key)

    # ★ 복구 탐색까지 했는데 흡수 자국의 파일이 하나라도 없으면 **그 답은 캐시하지
    #   않는다**([277]). Z: 가 잠깐 끊긴 2026-08-27 회차는 옛 로컬 18개로 두 방이
    #   그럴듯하게 채워졌고, 그 반쪽 답이 같은 trigger 에 박혀 다음 흡수 전까지
    #   `카톡 최신 2026-07-31` 로 남았다. 목록은 종전처럼 돌려주되 매 호출에서 다시
    #   확인하게 하고, 대표보고가 사람 탓을 하지 않도록 원인을 별도 자국으로 남긴다.
    cache_allowed = _kakao_selection_complete(recent_names, named_hits)
    if not dedupe_content:
        if cache_allowed:
            _save_kakao_selection(selection_trigger, False, paths)
        _write_kakao_selection_status(selection_trigger, recent_names, named_hits,
                                      cache_allowed)
        return paths
    import hashlib
    out, seen_hash = [], set()
    for path in paths:
        try:
            with open(path, "rb") as fh:
                digest = hashlib.sha256(fh.read()).digest()
        except OSError:
            continue
        if digest not in seen_hash:
            seen_hash.add(digest)
            out.append(path)
    if cache_allowed:
        _save_kakao_selection(selection_trigger, True, out)
    _write_kakao_selection_status(selection_trigger, recent_names, named_hits,
                                  cache_allowed)
    return out


def load_kakao_records():
    """카톡 내보내기(.txt)도 같은 양식(♣ ［…] ● 프로젝트NO / ● 캠프이름)을 쓴다.

    밴드에 안 올라오고 카톡에만 보고된 건이 있어(2026-07-27 기준 39건) 함께 읽는다.
    한 메시지에 여러 건이 담기므로 ♣ 로 덩어리를 나눠 게시글처럼 취급한다.
    """
    DAY = re.compile(r"-{3,}\s*(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일")
    out, seen = [], set()
    for f in kakao_source_paths():
        room = os.path.splitext(os.path.basename(f))[0]
        try:
            txt = open(f, encoding="utf-8", errors="replace").read()
        except OSError:
            continue
        day = ""
        for chunk in re.split(r"(?=-{3,}\s*\d{4}년)", txt):
            m = DAY.search(chunk)
            if m:
                day = "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3)))
            for i, blk in enumerate(re.split(r"♣", chunk)[1:]):
                if "프로젝트NO" not in blk:
                    continue
                key = (day, blk[:200])
                if key in seen:
                    continue
                seen.add(key)
                r = parse_post(f"kakao-{room}-{day}-{i}",
                               {"content": "♣" + blk[:2000], "author": room,
                                "created_at": None, "photo_count": 0, "comment_count": 0},
                               f"카톡 {room}")
                if r:
                    if not r.get("게시일"):
                        r["게시일"] = day
                    out.append(r)
    return out


def ledger_projects(master):
    """원장에 이미 있는 프로젝트NO 집합 (02·04·05·06 시트)"""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    seen = set()
    for sh in ("02_돌발AS접수", "04_정기점검", "05_신규납품설치", "06_거래서류청구수금"):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        try:
            j = [i for i, h in enumerate(hdr) if str(h).strip() == "프로젝트NO"][0]
        except IndexError:
            continue
        for row in ws.iter_rows(min_row=5, values_only=True):
            if j < len(row) and row[j]:
                seen.add(str(row[j]).strip())
    wb.close()
    return seen


HEADERS = ["프로젝트NO", "업무유형", "비용구분", "작업일", "담당기사", "캠프명",
           "진행상태", "문서상태", "사진", "원장등록", "게시일", "밴드"]
WIDTHS = [13, 17, 9, 12, 14, 22, 11, 26, 6, 10, 12, 24]


def main():
    args = sys.argv[1:]
    month = args[args.index("--month") + 1] if "--month" in args else None
    recs = load_records()
    if month:
        recs = [r for r in recs if (r["작업일"] or r["게시일"]).startswith(month)]

    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    known = ledger_projects(master)
    for r in recs:
        r["원장등록"] = "등록됨" if r["프로젝트NO"] in known else "미등록"

    new = [r for r in recs if r["원장등록"] == "미등록" and r["프로젝트NO"]]
    os.makedirs(REPORT_DIR, exist_ok=True)
    tag = month or "전체"
    base = os.path.join(REPORT_DIR, f"밴드업무추출_{tag}_{datetime.now():%Y%m%d_%H%M}")
    with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS); w.writeheader()
        w.writerows([{k: r.get(k, "") for k in HEADERS} for r in recs])

    from collections import Counter
    ck, cs = Counter(r["업무유형"] for r in recs), Counter(r["진행상태"] for r in recs)
    ct = Counter(r["담당기사"] for r in recs if r["담당기사"])
    with open(base + ".md", "w", encoding="utf-8") as f:
        f.write(f"# 밴드 업무 추출 — {tag}\n\n")
        f.write(f"- 생성 {datetime.now():%Y-%m-%d %H:%M} / 추출 {len(recs)}건 "
                f"(원장 등록됨 {len(recs)-len(new)} · **미등록 {len(new)}**)\n")
        f.write(f"- 업무유형: {dict(ck)}\n- 진행상태: {dict(cs)}\n- 담당기사: {dict(ct)}\n\n")
        f.write("## 원장 미등록 건 (백필 후보)\n\n")
        f.write("| 프로젝트NO | 유형 | 비용 | 작업일 | 기사 | 캠프 | 상태 | 문서 |\n|---|---|---|---|---|---|---|---|\n")
        for r in new:
            f.write(f"| {r['프로젝트NO']} | {r['업무유형']} | {r['비용구분']} | {r['작업일']} | "
                    f"{r['담당기사']} | {r['캠프명']} | {r['진행상태']} | {r['문서상태']} |\n")

    print(f"추출 {len(recs)}건 (등록됨 {len(recs)-len(new)} / 미등록 {len(new)})")
    print(f"유형 {dict(ck)}")
    print("리포트:", base + ".md")

    if "--sheet" in args:
        from findings_sheet import upsert, build_generic_sheet
        xml = build_generic_sheet(
            "24_밴드업무추출", HEADERS, WIDTHS,
            [[r.get(k, "") for k in HEADERS] for r in recs],
            f"[사용법] 밴드 게시글에서 자동 추출한 업무 원천({tag}). '원장등록=미등록' 행이 백필 후보입니다. "
            f"에이전트가 갱신하며 수기 입력은 하지 마세요.")
        dst, msg = upsert(master, xml, sheet_name="24_밴드업무추출", headers=HEADERS)
        print(f"24_밴드업무추출: {msg}")
        if dst:
            print("   ", dst)


if __name__ == "__main__":
    main()
