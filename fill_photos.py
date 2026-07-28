# -*- coding: utf-8 -*-
"""fill_photos.py — 밴드에 사진이 있는 건의 **사진 열**을 채운다.

사용자 지시(2026-07-28~29): 밴드 사진을 원장에 반영해 달라.

★ 사진을 **내려받지 않고** 채운다.
  원장의 사진 열이 묻는 것은 "사진이 있느냐"이지 "사진 파일을 갖고 있느냐"가 아니다.
  밴드 캐시의 `photo_count` 가 곧 그 답이다. 밴드 사진첩에는 13만 장이 있어서
  전량 내려받으면 12GB·수십 시간이 드는데, 이 열을 채우는 데는 한 장도 필요 없다.

무엇을 채우나
  02_돌발AS접수 · 사진등록     ← 밴드 글에 사진이 있으면 '등록'
  04_정기점검   · 점검사진     ← 〃

무엇을 **안** 채우나 (근거가 없다 — AGENTS.md 절대규칙 10)
  · 동영상등록 / 점검동영상 — 밴드 추출에 **동영상 항목이 아예 없다.**
    사진이 있다고 동영상도 있다고 적으면 그건 지어내는 것이다.
  · 사진 근거가 없는 건에 '누락' 을 적지 않는다. 아직 작업 전인 건은 사진이 없는 게 정상이고,
    '누락' 은 "있어야 하는데 없다" 는 뜻이라 미완료 건에 붙이면 거짓 지적이 된다.
    (작업완료 건의 '누락' 판정은 fill_report_photo 소관이다)
  · 이미 값이 있는 칸은 건드리지 않는다.

실행
  python fill_photos.py            # 미리보기
  python fill_photos.py --apply    # 큐 적재 + 원장 반영(vN+1)
"""
import os
import sys

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# (시트, 채울 열) — 동영상 열은 일부러 뺐다. 근거가 없다.
TARGET = [("02_돌발AS접수", "사진등록"), ("04_정기점검", "점검사진")]
VALUE = "등록"          # 두 열 모두 기존 값이 '등록' 이다(유효성 목록 안)


def band_photo_counts():
    """프로젝트NO → 밴드 글의 사진 장수(여러 글이면 가장 많은 쪽)."""
    from band_extract import load_records
    out = {}
    for r in load_records():
        k = str(r.get("프로젝트NO") or "").strip()
        if not k:
            continue
        try:
            n = int(str(r.get("사진") or 0) or 0)
        except ValueError:
            n = 0
        if n > out.get(k, 0):
            out[k] = n
    return out


def plan(master):
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    photo = band_photo_counts()
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    items, tally = [], {}
    for sheet, col in TARGET:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        if col not in hdr or "프로젝트NO" not in hdr:
            continue
        jc, jp = hdr.index(col), hdr.index("프로젝트NO")
        hit = blank = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            k = str(row[jp] or "").strip()
            if not k:
                continue
            if str(row[jc] or "").strip():
                continue                       # 이미 값이 있으면 그대로 둔다
            blank += 1
            if photo.get(k, 0) <= 0:
                continue                       # 근거 없음 — 비워 둔다('누락' 을 찍지 않는다)
            hit += 1
            items.append({
                "sheet": sheet, "key_col": "프로젝트NO", "key": k,
                "col": col, "value": VALUE, "vtype": "text",
                "evidence": "밴드 글 사진 %d장" % photo[k], "only_if_empty": True,
            })
        tally[(sheet, col)] = (hit, blank)
    wb.close()
    return items, tally, len(photo)


def main():
    do = "--apply" in sys.argv
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    items, tally, nph = plan(master)
    print("밴드 사진 근거 %d개 프로젝트" % nph)
    for (sheet, col), (hit, blank) in tally.items():
        print("  %-14s %-8s 빈칸 %3d → 채움 %3d (근거 없어 그대로 둠 %d)"
              % (sheet[:14], col, blank, hit, blank - hit))
    print("\n채울 칸 %d개" % len(items))
    print("  ※ 동영상 열은 채우지 않습니다 — 밴드에 동영상 정보가 없습니다.")
    if not do:
        print("\n실제로 채우려면:  python fill_photos.py --apply")
        return 0
    if not items:
        print("채울 것 없음")
        return 0
    from ledger_writer import queue_add
    print("큐 적재:", queue_add(items), "개 셀 → python ledger_writer.py --apply")
    return 0


if __name__ == "__main__":
    sys.exit(main())
