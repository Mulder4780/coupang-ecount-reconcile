# -*- coding: utf-8 -*-
"""
erp_bundle.py — ERP 계산서 1장에 **어떤 프로젝트가 묶여 있는지** 추정해 표기
================================================================================
ERP는 여러 작업을 한 장으로 묶어 발행한다(최대 15건). 그래서 25_ERP매출서류의
1행은 작업 1건이 아니다. 품목 단위 판매현황이 들어오기 전까지는 건별 배분을 못 하지만,
**캠프·유형·기간**으로 후보를 좁히고 **금액 합이 맞는지**로 확신도를 매길 수는 있다.

판정
  확정   후보 작업들의 공급가액 합 = 계산서 공급가액 (±1원)   → 구성이 확실하다
  유력   합이 ±3% 이내                                        → 거의 맞다
  추정   캠프·유형·기간은 맞지만 금액이 안 맞음                → 사람이 확인
  미상   후보를 못 찾음                                        → 사람이 확인

**추정을 확정처럼 쓰지 않는다** — 판정 열을 함께 적어 어디까지 믿을지 보이게 한다.

실행
  python erp_bundle.py             # 리포트만
  python erp_bundle.py --sheet     # + 26_계산서구성 시트 반영(vN+1)
"""
import sys, os, re, csv
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")

_CAMP_RE = re.compile(r"[가-힣A-Za-z]+\d*(?:BMB|MB|캠프|Sub-?FC|Sub-?hub|FC)(?:\([^)]*\))?", re.I)
_PAREN_RE = re.compile(r"\([^)]*\)")


def camp_of(title):
    """계산서 제목·캠프명에서 캠프를 뽑는다"""
    m = _CAMP_RE.search(str(title or ""))
    if m:
        return m.group()
    t = re.sub(r"^(쿠팡\S*|돌발AS|정기점검)[_\s-]*", "", str(title or "")).strip()
    return (t or str(title or ""))[:28]


def camp_key(name):
    """'송파1MB(감일동)' 과 '송파1MB' 를 같은 캠프로 본다 — 괄호 안 지명은 표기 흔들림이 크다"""
    s = _PAREN_RE.sub("", str(name or ""))
    return re.sub(r"[\s_·\-]", "", s).lower()


def kind_of(title):
    """계산서 제목 → 업무 유형 (erp_docs_check.work_kind 와 같은 규칙)"""
    s = str(title or "")
    if "돌발" in s or re.search(r"\bAS\b", s, re.I):
        return "돌발AS"
    if "정기점검" in s or "분기" in s:
        return "정기점검"
    if "철거" in s:
        return "철거"
    if "계단" in s:
        return "계단"
    if "신규" in s or "납품" in s or re.search(r"\d+\s*EA\b", s, re.I):
        return "신규납품"
    return "기타"


def load_erp():
    """25_ERP매출서류 → [{slip, month, kind, amt, title, camp}]"""
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    if "25_ERP매출서류" not in wb.sheetnames:
        wb.close()
        return master, []
    ws = wb["25_ERP매출서류"]
    hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
    out = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        g = lambda c: (r[idx[c]] if c in idx and idx[c] < len(r) else None)
        slip = str(g("일자-No.") or "").strip()
        if not slip:
            continue
        title = str(g("프로젝트명") or "")
        out.append({"slip": slip, "month": str(g("월") or "")[:7],
                    "kind": str(g("유형") or "") or kind_of(title),
                    "amt": int(g("공급가액") or 0), "title": title,
                    "camp": camp_of(title)})
    wb.close()
    return master, out


def load_works(master):
    """02·04·06 → 후보 작업 [{prj, camp, kind, date, amt, src}]"""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    out = []
    spec = [("02_돌발AS접수", "돌발AS", "접수일자", "작업완료일", None),
            ("04_정기점검", "정기점검", "점검예정일", "실제점검일", None),
            ("05_신규납품설치", None, "요청일", None, "견적금액"),
            ("06_거래서류청구수금", None, "작업완료일", None, "실제작업공급가액")]
    for sheet, kind, dcol, dcol2, amtcol in spec:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        for r in ws.iter_rows(min_row=5, values_only=True):
            g = lambda c: (r[idx[c]] if c in idx and idx[c] < len(r) else None)
            prj = str(g("프로젝트NO") or "").strip()
            camp = str(g("캠프명") or "").strip()
            if not prj or not camp:
                continue
            d = g(dcol2) or g(dcol)
            out.append({"prj": prj, "camp": camp,
                        "kind": kind or str(g("업무구분") or ""),
                        "date": str(d or "")[:10],
                        "amt": int(g(amtcol) or 0) if amtcol else 0,
                        "src": sheet})
    wb.close()
    return out


# ── 밴드 '세금계산서 발행 완료' 목록글 = 계산서 구성 원본 ────────────────────
# 매출처 밴드에는 계산서 한 장이 어떤 건들로 이뤄졌는지 사람이 직접 적어 둔 글이 있다.
#   1. 송파1MB(감일동)(1/10) : 2R/T Mobile-lift 2EA 19,780,000원 / PO326259
#      UJ2501950 - 발행완료
#   6. 김해2MB외 철거 및 이전 설치(1/25) : 4,558,500원 / PO330304
#      1. 김해2MB(장유동) … : 906,000원 UJ2600136      ← 묶인 하위 건들
# 추정보다 이게 훨씬 정확하다. 발행일+금액으로 ERP 전표와 맞춰 '확정(밴드)'으로 쓴다.
_ITEM_RE = re.compile(r"^\s*(\d{1,2})\.\s*(.+?)\((\d{1,2})/(\d{1,2})\)\s*[:：]\s*(.*)$")
_AMT_RE = re.compile(r"([\d,]{4,})\s*원")
_UJ_RE = re.compile(r"UJ\d{7}")


def band_invoice_index():
    """(발행일 'YYYY-MM-DD', 공급가액) → {프로젝트들, 캠프, PO}"""
    import json, glob
    cache = os.path.join(ROOT, "band", "cache")
    idx = {}
    for f in glob.glob(os.path.join(cache, "*.json")):
        b = os.path.basename(f)
        if b.startswith(("dump_", "raw_")):
            continue
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for p in (d.get("posts") or {}).values():
            ts = p.get("created_at")
            if not ts:
                continue
            year = datetime.fromtimestamp(ts / 1000).year
            pm = datetime.fromtimestamp(ts / 1000).month
            lines = (p.get("content") or "").splitlines()
            cur = None
            for ln in lines:
                m = _ITEM_RE.match(ln)
                if m:
                    mo, day = int(m.group(3)), int(m.group(4))
                    # 1월에 올린 글이 12월 건을 적기도 한다 — 그때는 작년으로 본다
                    y = year - 1 if (pm <= 2 and mo >= 11) else year
                    am = _AMT_RE.search(m.group(5))
                    if not am:
                        cur = None
                        continue
                    amt = int(am.group(1).replace(",", ""))
                    po = re.search(r"PO\d+", m.group(5))
                    cur = idx.setdefault(("%04d-%02d-%02d" % (y, mo, day), amt),
                                         {"프로젝트": [], "캠프": m.group(2).strip(),
                                          "PO": po.group() if po else ""})
                    for x in _UJ_RE.findall(m.group(5)):
                        if x not in cur["프로젝트"]:
                            cur["프로젝트"].append(x)
                elif cur is not None:
                    # 다음 항목이 나오기 전까지의 줄(하위 건 포함)에서 프로젝트NO를 줍는다
                    for x in _UJ_RE.findall(ln):
                        if x not in cur["프로젝트"]:
                            cur["프로젝트"].append(x)
    return idx


# 쿠팡 PO 발주글의 총금액 표기가 제각각이다:
#   '★ 총금액 : 25,223,400원'  '★ 총금액 :8,626,500원'  '★ 총금액 : 13,866,500 KRW'
# '원'만 찾으면 KRW로 적힌 글을 통째로 놓친다(2026-07-27에 추정 23건이 이 때문이었다).
_TOTAL_RE = re.compile(r"총\s*금\s*액\s*[:：]?\s*([\d,]{4,})\s*(?:원|KRW|₩)?", re.I)


def band_po_index():
    """쿠팡 PO 발주 글 → 금액별 프로젝트NO.

        ★ 총금액 : 50,400,000 원
        ★ 품  목 : M_무안1 4R/T Mobile-lift 3EA
        ★ 프로젝트 No. : UJ2600232 (리프트만)

    계산서 목록글에 프로젝트NO가 안 적힌 건들은 여기서 채운다(금액이 같으면 같은 건).
    """
    import json, glob
    cache = os.path.join(ROOT, "band", "cache")
    idx = {}
    for f in glob.glob(os.path.join(cache, "*.json")):
        b = os.path.basename(f)
        if b.startswith(("dump_", "raw_")):
            continue
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for p in (d.get("posts") or {}).values():
            body = p.get("content") or ""
            m = _TOTAL_RE.search(body)
            if not m:
                continue
            prjs = _UJ_RE.findall(body)
            if not prjs:
                continue
            amt = int(m.group(1).replace(",", ""))
            # 총금액은 부가세 포함일 수도, 공급가일 수도 있다 — 둘 다 열쇠로 넣는다
            for k in (amt, round(amt / 1.1)):
                cur = idx.setdefault(k, [])
                for x in prjs:
                    if x not in cur:
                        cur.append(x)
    return idx


_ITEM_LINE = re.compile(r"품\s*목\s*[:：]\s*([^\n★]{2,60})")
_PO_NO = re.compile(r"PO\d{5,}")
_CNT = re.compile(r"(\d{1,3})\s*건")


def band_po_meta():
    """PO 발주글에서 **프로젝트NO가 안 적힌** 건의 단서를 모은다.

    이런 글이 대부분이다:
        ★ 총금액 : 25,223,400원
        ★ 품  목 : 정기점검 29건 (5/21-30 6건, 6/1-10 23건)
        ★ 쿠팡오더 No. : PO364055/PR511170
    개별 프로젝트NO는 없지만 **PO번호·유형·건수**는 확실하다.
    이것만 알아도 아리바에서 품목을 조회할 수 있어, '추정'보다 훨씬 쓸모 있다.
    """
    import json, glob
    cache = os.path.join(ROOT, "band", "cache")
    kakao = os.path.join(ROOT, "kakao", "inbox")
    bodies = []
    for f in glob.glob(os.path.join(cache, "*.json")):
        b = os.path.basename(f)
        if b.startswith(("dump_", "raw_")):
            continue
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        bodies += [(p.get("content") or "") for p in (d.get("posts") or {}).values()]
    for f in glob.glob(os.path.join(kakao, "*.txt")):
        try:
            bodies.append(open(f, encoding="utf-8", errors="replace").read())
        except OSError:
            pass

    idx = {}
    for body in bodies:
        # 한 글에 여러 발주가 담기기도 한다 — '구매 오더' 단위로 쪼갠다
        for blk in re.split(r"(?=Coupang이\(가\) 새 구매 오더)", body):
            m = _TOTAL_RE.search(blk)
            if not m:
                continue
            amt = int(m.group(1).replace(",", ""))
            po = _PO_NO.search(blk)
            item = _ITEM_LINE.search(blk)
            item_s = item.group(1).strip() if item else ""
            cnt = _CNT.search(item_s)
            rec = {"PO": po.group() if po else "",
                   "품목": item_s,
                   "건수": int(cnt.group(1)) if cnt else 0,
                   "유형": kind_of(item_s)}
            for k in (amt, round(amt / 1.1)):
                idx.setdefault(k, rec)
    return idx


_QTR_RE = re.compile(r"(\d{2})년\s*(\d)\s*분기")


def window(doc):
    """후보로 볼 기간 (시작월, 끝월).

    ERP는 작업이 끝난 뒤 묶어서 끊으므로 발행월보다 앞선다 → 기본은 발행월 이전 3개월.
    다만 제목에 '25년 4분기'처럼 분기가 적혀 있으면 그 분기를 그대로 쓴다
    (분기 점검은 발행이 한참 늦어져 3개월 창을 벗어난다)."""
    mo = doc["month"].replace("/", "-")
    if len(mo) < 7:
        return "", ""
    q = _QTR_RE.search(doc["title"])
    if q:
        y, n = 2000 + int(q.group(1)), int(q.group(2))
        return "%04d-%02d" % (y, n * 3 - 2), "%04d-%02d" % (y, n * 3)
    y, m = int(mo[:4]), int(mo[5:7])
    lo = "%04d-%02d" % (y - 1, m + 9) if m <= 3 else "%04d-%02d" % (y, m - 3)
    return lo, mo


def bundle(doc, works, binx=None, poinx=None, pometa=None):
    """계산서 1장 → (후보 프로젝트 목록, 판정, 합계)

    후보를 못 찾으면 판정 자리에 **왜 못 찾았는지**를 적는다.
    '미상' 한 마디로는 사람이 무엇을 해야 할지 알 수 없다.
    """
    band_camp = ""
    # ① 밴드에 사람이 적어 둔 계산서 목록이 최우선이다 — 추정이 아니라 기록이다.
    if binx:
        got = binx.get((doc["slip"][:10].replace("/", "-"), doc["amt"]))
        if got:
            if got["프로젝트"]:
                return got["프로젝트"], "확정(밴드)", doc["amt"]
            # 번호가 없으면 캠프라도 쓴다. 다만 목록글의 이름은 'A/S_1'·'정기점검_2' 같은
            # **묶음 라벨**인 경우가 많아, 캠프 형태일 때만 갈아끼운다.
            band_camp = got["캠프"][:18]
            if _CAMP_RE.search(got["캠프"]):
                doc = {**doc, "camp": got["캠프"]}
    # ② 쿠팡 PO 발주 글은 계산서 목록에 없는 건도 커버한다 — 금액만 같으면 같은 발주다.
    #    (목록글이 있을 때만 보면 PO365213 처럼 목록에 안 실린 건을 놓친다)
    po = (poinx or {}).get(doc["amt"]) or []
    if po:
        return po, "확정(밴드PO)", doc["amt"]

    # ③ 프로젝트NO는 없어도 **쿠팡 PO 번호와 건수**는 밴드에 적혀 있다.
    #    추정으로 뭉개지 말고 그 사실을 그대로 알려 준다 — 아리바에서 품목을 볼 수 있다.
    meta = (pometa or {}).get(doc["amt"])
    # 계단·철거·신규납품은 AS/점검이 아니라 별도 공사다. PO 번호를 알아도 그 사실은 그대로다 —
    # 둘 다 적어 준다(어디서 확인할지 + 왜 작업 행이 없는지).
    if doc["kind"] in ("계단", "철거", "신규납품", "기타"):
        po = f" · {meta['PO']}" if (meta and meta.get("PO")) else ""
        return [], f"대상외({doc['kind']} — 별도 공사, 원장에 작업 행 없음{po})", 0
    if meta and meta.get("PO"):
        kd = meta.get("유형") or doc["kind"]
        n = meta.get("건수") or 0
        cand = [w["prj"] for w in works
                if w["date"] and w["kind"] and kd in w["kind"]
                and window(doc)[0] <= w["date"][:7] <= window(doc)[1]]
        tag = f"PO확인({meta['PO']} · {meta['품목'][:24]})"
        # 후보 건수가 PO 건수와 딱 맞으면 그 목록을 그대로 쓴다
        if n and len(set(cand)) == n:
            return sorted(set(cand)), f"유력({meta['PO']} · {n}건 일치)", doc["amt"]
        return [], tag, doc["amt"]

    ck, kind = camp_key(doc["camp"]), doc["kind"]
    # 계단·철거·신규납품·기타는 AS/점검이 아니라 **별도 공사**다. 02·04 시트에 작업 행이
    # 아예 없으므로 '미상'으로 두면 영원히 안 풀린다. 대상 아님으로 분명히 적는다.
    # (신규 납품건 제외는 사용자 지시 2026-07-26)
    if kind in ("계단", "철거", "신규납품", "기타"):
        return [], f"대상외({kind} — 별도 공사, 원장에 작업 행 없음)", 0
    lo, hi = window(doc)
    tail = f" · 밴드캠프 {band_camp}" if band_camp else ""
    if not ck or not lo:
        return [], "미상(캠프 못 읽음)" , 0
    same_camp = [w for w in works if camp_key(w["camp"]) == ck]
    if not same_camp:
        if band_camp:
            return [], f"밴드확인({band_camp})", 0
        return [], "미상(그 캠프 작업이 대장에 없음)", 0
    same_kind = [w for w in same_camp
                 if not kind or kind == "기타" or kind in w["kind"] or w["kind"] in kind]
    if not same_kind:
        return [], (f"밴드확인({band_camp})" if band_camp else f"미상({kind} 자료 없음)"), 0
    hits = [w for w in same_kind if w["date"] and lo <= w["date"][:7] <= hi]
    if not hits:
        gotm = sorted({w["date"][:7] for w in same_kind if w["date"]})
        return [], f"미상(기간 {lo}~{hi} 밖 · 보유 {','.join(gotm[:3])})", 0
    seen, prjs = set(), []
    for w in hits:
        if w["prj"] not in seen:
            seen.add(w["prj"])
            prjs.append(w["prj"])
    tot = sum(w["amt"] for w in hits)
    if doc["amt"] and tot:
        if abs(tot - doc["amt"]) <= 1:
            v = "확정"
        elif abs(tot - doc["amt"]) <= doc["amt"] * 0.03:
            v = "유력"
        else:
            v = "추정"
    else:
        v = "추정"
    return prjs, v, tot


def main():
    args = sys.argv[1:]
    master, docs = load_erp()
    if not docs:
        sys.exit("25_ERP매출서류 시트가 비어 있습니다 — 먼저 erp_docs_check.py --sheet 를 돌리세요")
    works = load_works(master)
    binx, poinx, pometa = band_invoice_index(), band_po_index(), band_po_meta()
    print(f"밴드 근거: 계산서 목록 {len(binx)}항목 · PO(프로젝트 포함) {len(poinx)}금액 · "
          f"PO(번호·건수만) {len(pometa)}금액")
    rows, stat, why = [], {}, {}
    for d in sorted(docs, key=lambda x: x["slip"]):
        prjs, verdict, tot = bundle(d, works, binx, poinx, pometa)
        stat[verdict.split("(")[0]] = stat.get(verdict.split("(")[0], 0) + 1
        why[verdict] = why.get(verdict, 0) + 1
        rows.append([d["slip"], d["month"], d["kind"], d["camp"], d["amt"],
                     len(prjs), ", ".join(prjs[:15]), tot, verdict, d["title"]])

    os.makedirs(REPORT_DIR, exist_ok=True)
    base = os.path.join(REPORT_DIR, f"계산서구성_{datetime.now():%Y%m%d_%H%M}")
    HDR = ["일자-No.", "월", "유형", "캠프명", "계산서공급가액", "포함건수",
           "포함프로젝트NO", "후보합계", "판정", "프로젝트명"]
    with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(HDR)
        w.writerows(rows)
    print(f"ERP 계산서 {len(rows)}장 — " + " · ".join(f"{k} {v}" for k, v in sorted(stat.items())))
    print("  미상 사유: " + " · ".join(f"{k.split('(',1)[-1].rstrip(')')} {v}"
                                       for k, v in sorted(why.items(), key=lambda x: -x[1])
                                       if k.startswith("미상"))[:300])
    print("리포트:", base + ".csv")

    if "--sheet" in args:
        from findings_sheet import upsert, build_generic_sheet
        W = [14, 9, 10, 22, 15, 9, 62, 14, 8, 46]
        xml = build_generic_sheet(
            "26_계산서구성", HDR, W, rows,
            "ERP 계산서 1장에 묶인 작업을 캠프·유형·기간으로 좁혀 추정한 것. "
            "판정이 '확정'(금액 합 일치)일 때만 그대로 믿고, '추정'·'미상'은 사람이 확인한다. "
            "품목 단위 판매현황이 들어오면 건별로 정확히 배분된다. erp_bundle.py 자동 갱신 — 수기 입력 금지.")
        dst, msg = upsert(master, xml, sheet_name="26_계산서구성", headers=HDR)
        print(f"26_계산서구성 시트: {msg}")


if __name__ == "__main__":
    main()
