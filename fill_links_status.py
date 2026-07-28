# -*- coding: utf-8 -*-
"""
fill_links_status.py — 밴드 바로가기 URL + 03_현장작업실적 상태 자동 표기
================================================================================
사용자 지시(2026-07-28)
  ① 02_돌발AS접수 '밴드 바로가기' 에 실제 URL 을 넣는다.
  ② 03_현장작업실적 작업사진·작업동영상·완료보고서를 확인되면 표기한다.
  ③ 관리자검증은 일치하면 '일치', 검증결과는 확인된 건은 '정상'.

■ 검증결과(AK)는 **수식이다** — 직접 '정상'이라 쓰지 않는다.
  원인 칸(작업내용·사진·완료보고서·관리자검증…)이 채워지면 저절로 '정상'이 된다.
  값으로 덮으면 다음부터 자동 판정이 죽는다.

■ 밴드 URL 은 어디서 오나
  가공된 캐시(band/cache/<밴드>.json)에는 **글 번호가 없다**(author·content·created_at·
  photo_count 뿐). 원본 덤프(raw_*.json)의 `posts` 가 **글 번호를 키로** 갖고 있어
  거기서 얻는다. 형식은 원장에 이미 있던 한 건과 같다:
      https://band.us/band/<밴드ID>/post/<글번호>
  한 글에 여러 프로젝트가 묶인 목록형 글이 있으므로 **그 건만 다루는 글**을 우선한다.
  근거가 없으면 **주소를 지어내지 않는다** — 틀린 링크는 빈칸보다 나쁘다.

■ 03시트 완료보고서의 근거
  02시트에서는 '밴드에 완료 글이 있음' 말고 근거가 없어 보류했지만, 03시트는 다르다.
  **기사보고내용·실제작업상세가 곧 기사가 올린 보고 내용**이다. 이걸 근거로 삼는다.
  ('(자동 초안)' 문구는 우리가 넣은 자리표시자이므로 근거로 치지 않는다.)

■ 작업동영상은 근거원이 없다 → '해당없음'. 없는 근거로 '등록'이라 적지 않는다.

사용
  python fill_links_status.py            # 집계만 (파일 안 건드림)
  python fill_links_status.py --apply    # vN+1 생성
"""
import os
import re
import sys
import glob
import json
import zipfile
from collections import Counter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path, esc  # noqa: E402

HDR = 4
PRJ = re.compile(r"\b(UJ\d{7})\b")
DRAFT = "(자동 초안)"

# 유효성 목록(10_코드관리) — 목록 밖 값은 쓰지 않는다
DOC = ("등록", "누락", "해당없음")
ADMIN = ("일치", "추가작업발생", "작업내용누락", "확인필요")


def _s(v):
    return "" if v is None else str(v).strip()


def band_links():
    """프로젝트NO → URL. 그 건만 다루는 글을 우선하고, 없으면 만들지 않는다."""
    cand = {}
    for f in sorted(glob.glob(os.path.join(BASE, "band", "cache", "raw*.json"))):
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        band, posts = _s(d.get("band")), d.get("posts")
        if not band or not isinstance(posts, dict):
            continue
        for no, p in posts.items():
            if not str(no).isdigit():
                continue
            c = str((p.get("content") if isinstance(p, dict) else p) or "")
            ks = set(PRJ.findall(c))
            if not ks:
                continue
            for k in ks:
                # 점수: ① 그 건만 다루는 글(프로젝트 1개) ② 완료 글 ③ 최신
                score = (1 if len(ks) == 1 else 0, 1 if "완료" in c[:80] else 0, int(no))
                if k not in cand or score > cand[k][0]:
                    cand[k] = (score, band, str(no))
    return {k: f"https://band.us/band/{b}/post/{n}" for k, (_s_, b, n) in cand.items()}


def load(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def rows(sn):
        ws = wb[sn]
        hdr = [_s(h) for h in next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        out = []
        for n, r in enumerate(ws.iter_rows(min_row=HDR + 1, values_only=True), start=HDR + 1):
            if r[0] in (None, ""):
                continue
            g = {h: (r[i] if i < len(r) else None) for h, i in ix.items()}
            g["_행"] = n
            out.append(g)
        return hdr, out

    h2, as_rows = rows("02_돌발AS접수")
    h3, fw_rows = rows("03_현장작업실적")
    _, band = rows("24_밴드업무추출")
    wb.close()

    photo = {}
    for b in band:
        k = _s(b.get("프로젝트NO"))
        if not k:
            continue
        try:
            photo[k] = max(photo.get(k, 0), int(_s(b.get("사진")) or 0))
        except ValueError:
            pass
    return (h2, as_rows), (h3, fw_rows), photo


def plan(path):
    from openpyxl.utils import get_column_letter as GL
    (h2, as_rows), (h3, fw_rows), photo = load(path)
    links = band_links()
    fills, tally = {}, {}

    def put(col_letter, row, val, key):
        fills[f"{col_letter}{row}"] = val
        tally.setdefault(key, Counter())[val if len(val) < 12 else "URL"] += 1

    # ── ① 02시트 밴드 바로가기 ──────────────────────────────────────────────
    cL = GL(h2.index("밴드 바로가기") + 1)
    nolink = 0
    for r in as_rows:
        if _s(r.get("밴드 바로가기")):
            continue                              # 사람이 넣은 건 건드리지 않는다
        u = links.get(_s(r.get("프로젝트NO")))
        if u:
            put(cL, r["_행"], u, "밴드 바로가기")
        else:
            nolink += 1

    # ── ② 03시트 상태 ──────────────────────────────────────────────────────
    C = {h: GL(h3.index(h) + 1) for h in
         ("작업사진", "작업동영상", "완료보고서", "관리자검증") if h in h3}
    for r in fw_rows:
        k = _s(r.get("프로젝트NO"))
        detail = _s(r.get("실제작업상세"))
        if DRAFT in detail:
            detail = ""
        report = _s(r.get("기사보고내용"))
        item = _s(r.get("실제작업항목"))
        extra = _s(r.get("추가작업내용")) or _s(r.get("접수외추가작업여부")) == "있음"
        written = bool(detail or report or item)     # 기사가 무엇이든 적어 냈는가

        if "작업사진" in C and not _s(r.get("작업사진")):
            put(C["작업사진"], r["_행"], "등록" if photo.get(k, 0) > 0 else "누락", "작업사진")
        if "작업동영상" in C and not _s(r.get("작업동영상")):
            put(C["작업동영상"], r["_행"], "해당없음", "작업동영상")   # 근거원 없음
        if "완료보고서" in C and not _s(r.get("완료보고서")):
            put(C["완료보고서"], r["_행"], "등록" if written else "누락", "완료보고서")
        if "관리자검증" in C and not _s(r.get("관리자검증")):
            v = "추가작업발생" if extra else ("일치" if written else "작업내용누락")
            put(C["관리자검증"], r["_행"], v, "관리자검증")

    for k, c in tally.items():
        for v in c:
            if k != "밴드 바로가기" and v not in DOC + ADMIN:
                raise RuntimeError(f"허용되지 않은 값 {k}={v!r}")
    return fills, tally, nolink, len(as_rows), len(fw_rows), len(links)


def apply_cells(xml, fills):
    """XML을 한 번만 훑는다 — 칸마다 찾으면 수천 번 훑게 된다."""
    seen = set()

    def repl(m):
        ref, attrs = m.group(1), m.group(2) or ""
        if ref not in fills or 't="shared"' in m.group(0) or 't="array"' in m.group(0):
            return m.group(0)
        seen.add(ref)
        a = re.sub(r'\s+t="[^"]*"', "", attrs)
        return f'<c r="{ref}"{a} t="inlineStr"><is><t>{esc(fills[ref])}</t></is></c>'

    out = re.sub(r'<c r="([A-Z]+\d+)"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)',
                 repl, xml, flags=re.S)
    return out, seen


def verify(src, out, fills, allowed):
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"
    wa = openpyxl.load_workbook(src, data_only=True)
    wb_ = openpyxl.load_workbook(out, data_only=True)
    for sn in ("02_돌발AS접수", "03_현장작업실적"):
        sa, sb = wa[sn], wb_[sn]
        mine = {r for r in fills if _sheet_of(r, fills) == sn}
        diff = []
        for rr in range(1, max(sa.max_row, sb.max_row) + 1):
            for cc in range(1, max(sa.max_column, sb.max_column) + 1):
                ca = sa.cell(rr, cc)
                if f"{sn}!{ca.coordinate}" in fills or ca.coordinate in mine:
                    continue
                if ca.value != sb.cell(rr, cc).value:
                    diff.append(f"{sn}!{ca.coordinate}")
                    if len(diff) > 5:
                        break
        assert not diff, f"의도치 않은 셀 변경: {diff}"
    # 수식이 살아 있어야 자동 판정이 계속 돈다
    wf = openpyxl.load_workbook(out, data_only=False)
    for sn, refs in (("02_돌발AS접수", ("AK5", "AL5", "AN5")),
                     ("03_현장작업실적", ("AK5", "AM5", "AN5"))):
        for ref in refs:
            assert str(wf[sn][ref].value or "").startswith("="), f"{sn} {ref} 수식이 사라졌다"
    wa.close(); wb_.close()
    zs = zipfile.ZipFile(src)
    other = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not other, f"의도치 않은 파트 변경: {other}"
    zs.close(); z.close()


def _sheet_of(ref, fills):
    return fills.get("_sheet_" + ref, "")


def main():
    do = "--apply" in sys.argv
    m = latest_master()
    src = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src)}\n")

    fills, tally, nolink, n2, n3, nlink = plan(src)
    print(f"밴드 원본에서 찾은 프로젝트 {nlink}개 · 02시트 {n2}행 · 03시트 {n3}행")
    print(f"채울 칸 {len(fills)}개\n")
    for k, c in tally.items():
        print(f"  {k:<12}{sum(c.values()):>5}  " + " · ".join(f"{v} {x}" for x, v in c.most_common()))
    if nolink:
        print(f"\n  밴드 글을 못 찾아 주소를 비워 두는 행 {nolink}개"
              " — 없는 주소를 지어내지 않습니다.")
    print("  ※ 검증결과는 **수식이라 건드리지 않습니다** — 원인 칸이 채워지면 저절로 '정상'이 됩니다.")
    print("  ※ 작업동영상은 근거원이 없어 '해당없음'으로 둡니다.")
    if not do:
        print("\n실제로 채우려면:  python fill_links_status.py --apply")
        return 0

    # 시트별로 갈라 넣는다
    by_sheet = {"02_돌발AS접수": {}, "03_현장작업실적": {}}
    from openpyxl.utils import column_index_from_string as CI  # noqa: F401
    (h2, as_rows), (h3, fw_rows), _ = load(src)
    rows2 = {r["_행"] for r in as_rows}
    for ref, v in fills.items():
        rn = int(re.search(r"\d+$", ref).group(0))
        col = re.match(r"[A-Z]+", ref).group(0)
        # 02시트에 넣는 것은 '밴드 바로가기' 열뿐이다
        tgt = "02_돌발AS접수" if (col == _colof(h2, "밴드 바로가기") and rn in rows2) else "03_현장작업실적"
        by_sheet[tgt][ref] = v

    zin = zipfile.ZipFile(src)
    changed = {}
    for sn, f in by_sheet.items():
        if not f:
            continue
        sp = sheet_xml_path(zin, sn)
        new, seen = apply_cells(zin.read(sp).decode("utf-8"), f)
        miss = [r for r in f if r not in seen]
        if miss:
            print(f"★ {sn}: XML에 칸이 없어 못 채운 곳 {len(miss)}개 {miss[:4]}")
        changed[sp] = new.encode("utf-8")
    from dashboard_clean import force_recalc
    changed["xl/workbook.xml"] = force_recalc(
        zin.read("xl/workbook.xml").decode("utf-8")).encode("utf-8")

    mm = re.search(r"_v(\d+)\.xlsx$", src)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm.group(1)) + 1}.xlsx", src)
    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close(); zin.close()

    try:
        _verify_simple(src, tmp, by_sheet, set(changed))
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise
    os.replace(tmp, dst)
    print(f"\n{len(fills)}칸 채움 → {os.path.basename(dst)}")
    print("엑셀에서 한 번 열어 주세요 — 검증결과가 다시 계산됩니다.")
    return 0


def _colof(hdr, name):
    from openpyxl.utils import get_column_letter as GL
    return GL(hdr.index(name) + 1)


def _verify_simple(src, out, by_sheet, allowed):
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"
    wa = openpyxl.load_workbook(src, data_only=True)
    wb_ = openpyxl.load_workbook(out, data_only=True)
    for sn, f in by_sheet.items():
        if not f:
            continue
        sa, sb = wa[sn], wb_[sn]
        for ref, v in f.items():
            assert _s(sb[ref].value) == v, f"{sn} {ref} 반영 실패 ({sb[ref].value!r})"
        diff = []
        for rr in range(1, max(sa.max_row, sb.max_row) + 1):
            for cc in range(1, max(sa.max_column, sb.max_column) + 1):
                ca = sa.cell(rr, cc)
                if ca.coordinate in f:
                    continue
                if ca.value != sb.cell(rr, cc).value:
                    diff.append(f"{sn}!{ca.coordinate}")
                    if len(diff) > 5:
                        break
        assert not diff, f"의도치 않은 셀 변경: {diff}"
    wf = openpyxl.load_workbook(out, data_only=False)
    for sn, refs in (("02_돌발AS접수", ("AK5", "AL5", "AN5")),
                     ("03_현장작업실적", ("AK5", "AM5", "AN5"))):
        for ref in refs:
            assert str(wf[sn][ref].value or "").startswith("="), f"{sn} {ref} 수식이 사라졌다"
    wa.close(); wb_.close()
    zs = zipfile.ZipFile(src)
    other = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not other, f"의도치 않은 파트 변경: {other}"
    zs.close(); z.close()
    print("  검증 통과 — 수식 보존 · 다른 셀·파트 변경 없음")


if __name__ == "__main__":
    sys.exit(main())
