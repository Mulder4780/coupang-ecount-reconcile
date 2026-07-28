# -*- coding: utf-8 -*-
"""
fix_tech_names.py — 담당기사 칸에 들어간 **사람 아닌 값**을 바로잡는다
================================================================================
사용자 확인(2026-07-28)으로 값이 확정된 10건을 고친다.

왜 생겼나
  밴드/카톡 글에서 기사명을 뽑는 `band_extract.normalize_tech` 가 '첫 조각'을
  그대로 통과시켰다. 그래서 작업 메모가 담당기사 칸에 들어갔고, 대표보고
  'TOP 5' 에 `담당: 000 (캠프상태확인 및 스케쥴 세팅)` 로 노출됐다.
  (걸러 주는 함수는 있었지만 리포트에서만 쓰이고 원장에 쓸 때는 안 거쳤다.)

원칙
  · **원문을 없애지 않는다.** 담당기사 칸에서 뺀 값은 비고(AJ)에 그대로 남긴다.
    나중에 "왜 이름이 바뀌었지?" 를 되짚을 수 있어야 한다.
  · 행 번호를 박지 않고 **값으로 찾는다** — 행이 밀려도 맞는 곳을 고친다.
  · 관리대장은 zip 패치로만 만진다(차트·도형 보호).

사용
  python fix_tech_names.py            # 무엇을 어떻게 고칠지만 보여 준다
  python fix_tech_names.py --apply    # vN+1 생성
"""
import os
import re
import sys
import zipfile

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path, replace_inline_cell, esc  # noqa: E402

# ★ 처음엔 02시트만 봤다가 03_현장작업실적에 같은 값이 남았다(2026-07-28).
#   같은 값이 여러 시트에 퍼지므로 **기사 열이 있는 시트를 전부** 훑는다.
SHEETS = ("02_돌발AS접수", "03_현장작업실적", "04_정기점검", "05_신규납품설치")
HDR_ROW = 4
TECH_COLS = ("담당기사", "담당자", "작업자")
NOTE_COL = "비고"

# 현재 값 → 고칠 값. 사용자 확인 결과(2026-07-28):
#   · 김혜진  = 유니버셜 퇴사자, 이전 담당  → 이름이 맞다
#   · 하이테크 = 협력업체, 엄진언 = 사람     → 담당기사는 엄진언
#   · 기장    = 직책                      → 김승기
#   · 000 / 자) = 미배정 자리표시자 + 작업 메모 → 담당기사는 비운다
FIX = {
    "000 (캠프상태확인 및 스케쥴 세팅)": "",
    "자) - 각캠프담당자 캠프 컨디션상태 체크": "",
    "하이테크 + 엄진언": "엄진언",
    "김혜진 대신택배": "김혜진",
    "김승기기장": "김승기",
}


def _cols(ws):
    """(기사열들, 비고열) — 없는 열은 건너뛴다. 시트마다 구성이 다르다."""
    hdr = [str(h or "").strip() for h in
           next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))]
    from openpyxl.utils import get_column_letter as GL
    ix = {h: GL(i + 1) for i, h in enumerate(hdr) if h}
    return [ix[c] for c in TECH_COLS if c in ix], ix.get(NOTE_COL)


def survey(path):
    """[(시트, 행, 프로젝트NO, 기사열, 현재기사, 새기사, 비고열, 새비고)]"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sn in SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        tcs, nc = _cols(ws)
        for tc in tcs:
            for r in range(HDR_ROW + 1, ws.max_row + 1):
                cur = str(ws[f"{tc}{r}"].value or "").strip()
                if cur not in FIX:
                    continue
                keep = f"담당기사 원본: {cur}"
                newnote = None
                if nc:
                    note = str(ws[f"{nc}{r}"].value or "").strip()
                    newnote = note if keep in note else (f"{note} | {keep}" if note else keep)
                out.append((sn, r, str(ws[f"B{r}"].value or ""), tc,
                            cur, FIX[cur], nc, newnote))
    wb.close()
    return out


def blank_cell(xml, ref):
    """스타일만 남기고 완전히 비운다(빈 문자열이 아니라 진짜 빈 칸)."""
    pat = re.compile(r'<c r="%s"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)' % ref, re.S)
    m = pat.search(xml)
    if not m:
        raise AssertionError(f"{ref} 셀 XML 없음")
    attrs = re.sub(r'\s+t="[^"]*"', "", m.group(1) or "")
    return xml[:m.start()] + f'<c r="{ref}"{attrs}/>' + xml[m.end():]


def apply(src, dst, rows):
    zin = zipfile.ZipFile(src)
    changed, paths = {}, {}
    for sn in sorted({r[0] for r in rows}):
        paths[sn] = sheet_xml_path(zin, sn)
        xml = zin.read(paths[sn]).decode("utf-8")

        # 02시트에는 공유수식이 5,321칸 있다(A·F·M·AK·AL·AN·AP·AR — 채번·자동판정 열).
        # 시트에 있다는 이유로 막으면 아무것도 못 고친다.
        # **내가 건드릴 칸이 그 그룹에 속하는지**만 본다 — 속한 칸을 갈아엎으면 나머지가 깨진다.
        for _sn, r, _prj, tc, _cur, _new, nc, nn in [x for x in rows if x[0] == sn]:
            for ref in ([f"{tc}{r}"] + ([f"{nc}{r}"] if nc and nn is not None else [])):
                m = re.search(r'<c r="%s"[^>]*>\s*<f([^>]*)>' % ref, xml)
                if m and ('t="shared"' in m.group(1) or 't="array"' in m.group(1)):
                    raise RuntimeError(f"{sn} {ref} 이 공유·배열 수식이라 손대면 위험 — 중단")

        for _sn, r, _prj, tc, _cur, new, nc, nn in [x for x in rows if x[0] == sn]:
            xml = (blank_cell(xml, f"{tc}{r}") if not new
                   else replace_inline_cell(xml, f"{tc}{r}", new))
            if nc and nn is not None:
                xml = replace_inline_cell(xml, f"{nc}{r}", nn)
        changed[paths[sn]] = xml.encode("utf-8")

    # 담당기사를 비우면 기사배정상태·대표보고 '담당자 미배정' 수가 따라 바뀌어야 한다.
    # 저장된 계산 결과는 옛 값 그대로이므로 열 때 다시 계산하도록 표시해 둔다.
    from dashboard_clean import force_recalc
    wbp = "xl/workbook.xml"
    changed[wbp] = force_recalc(zin.read(wbp).decode("utf-8")).encode("utf-8")

    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        raise FileExistsError(f"임시 결과가 이미 존재: {tmp}")
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close()
    zin.close()

    # 검증이 실패하면 **중간 결과를 남기지 않는다** — 남으면 다음 실행이 막히고,
    # 무엇보다 검증에 떨어진 파일이 폴더에 굴러다니면 안 된다.
    try:
        verify(src, tmp, rows, set(changed) | {wbp})
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise
    os.replace(tmp, dst)


def verify(src, out, rows, allowed):
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"
    wa = openpyxl.load_workbook(src, data_only=True)
    wb_ = openpyxl.load_workbook(out, data_only=True)

    touched = {}
    for sn, r, prj, tc, _cur, new, nc, nn in rows:
        sb = wb_[sn]
        got = str(sb[f"{tc}{r}"].value or "").strip()
        assert got == new, f"{sn} {r}행 담당기사 {got!r} ≠ {new!r}"
        assert str(sb[f"B{r}"].value or "") == prj, f"{sn} {r}행이 밀렸다"
        touched.setdefault(sn, set()).add(f"{tc}{r}")
        if nc and nn is not None:
            assert str(sb[f"{nc}{r}"].value or "").strip() == nn, f"{sn} {r}행 비고 반영 실패"
            touched[sn].add(f"{nc}{r}")

    # ★ 고친 칸 말고는 **한 칸도** 안 바뀌어야 한다 — 건드린 시트를 전수 대조한다
    for sn, refs in touched.items():
        sa, sb = wa[sn], wb_[sn]
        diff = []
        for r in range(1, max(sa.max_row, sb.max_row) + 1):
            for c in range(1, max(sa.max_column, sb.max_column) + 1):
                ca = sa.cell(r, c)
                if ca.coordinate in refs:
                    continue
                if ca.value != sb.cell(r, c).value:
                    diff.append(f"{sn}!{ca.coordinate}")
                    if len(diff) > 5:
                        break
        assert not diff, f"의도치 않은 셀 변경: {diff}"
    assert wa.sheetnames == wb_.sheetnames, "시트 구성이 바뀌었다"
    wa.close(); wb_.close()

    zs = zipfile.ZipFile(src)
    other = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not other, f"의도치 않은 파트 변경: {other}"
    zs.close(); z.close()
    print(f"  검증 통과 — {len(rows)}행 수정 · 시트 {len(touched)}개 전수 대조 · 다른 파트 변경 없음")


def main():
    do = "--apply" in sys.argv
    m = latest_master()
    src = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src)}\n")

    rows = survey(src)
    if not rows:
        print("고칠 행이 없습니다 — 이미 정리됐습니다.")
        return 0
    for sn, r, prj, tc, cur, new, nc, _nn in rows:
        print(f"  {sn} {r:>4}행 {prj:11s} {tc}열 {cur!r}")
        shown = repr(new) if new else "(비움 — 미배정)"
        note = f" · 비고에 '담당기사 원본: {cur}' 추가" if nc else " · (비고 열 없음)"
        print(f"        → {shown}{note}")
    print(f"\n합계 {len(rows)}행")
    if not do:
        print("\n실제로 고치려면:  python fix_tech_names.py --apply")
        return 0

    mm = re.search(r"_v(\d+)\.xlsx$", src)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm.group(1)) + 1}.xlsx", src)
    print(f"\n결과: {os.path.basename(dst)}")
    apply(src, dst, rows)
    return 0


if __name__ == "__main__":
    sys.exit(main())
