# -*- coding: utf-8 -*-
"""zscan.py — 쿠팡 업무 폴더(Z:) 전수 조사 · 관련 자료만 골라 원장과 대조.

사용자 지시(2026-07-28): "이 폴더에서 관련있는 자료는 전부 긁어와서 비교 검토해보고
관련 없는 자료는 db에 적용하지마."

핵심 판단 — **파일을 열지 않는다.**
  네트워크 드라이브에 2만 개가 넘고 PDF 1만·JPG 5천이다. 전부 열면 몇 시간이 걸리고
  대부분은 현장 사진·안전서류라 얻을 게 없다. 그런데 이 회사는 파일명에
  `… 클로징 리프트 철거  906,000원 UJ2600136.PDF` 처럼 **프로젝트NO·금액·PO번호를
  이미 적어 둔다.** 파일명만 읽어도 대조가 된다. 내용 확인이 필요한 건만 뒤에 따로 연다.

무엇을 '관련 있다'고 보나
  · 파일명에 프로젝트NO(UJ+7자리) 또는 PO번호(PO+6자리)가 있으면 무조건 관련.
  · 그게 없어도 서류 성격(거래명세서·세금계산서·견적서·발주·정산)이면 관련.
  · 안전보건대장·지게차·보험·교육·근태·도면·스티커 폴더는 **DB 무관**으로 분류한다.
    업무상 중요하지만 관리대장 02·04·05·06 어느 열에도 들어갈 자리가 없다.

★ **기본은 읽기 전용이다.** 조사·대조 결과만 리포트로 낸다.
  원장에 쓰는 경로는 `--queue-docs` 하나뿐이고, 그것도
  **캠프명이 같고 날짜가 ±7일인 1:1 확정 건**의 04시트 `거래명세서` 한 열만 건드린다.
  "관련 없는 자료는 db에 적용하지마"(사용자 지시)를 지키는 방법은 쓰기 경로를 좁게 만드는 것이다.
  후보가 여럿이거나 짝이 없는 건은 **쓰지 않고 리포트로만** 남긴다.

실행
  python zscan.py                 # 전수 조사 + 원장 대조 리포트
  python zscan.py --folder "♣ 10. 세금계산서 발행"
"""
import argparse
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict


def configure_stdout(stream):
    """워치독의 CP949 파이프에서도 유니코드 진행 문구가 실패하지 않게 한다."""
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8", errors="replace")
    return stream


configure_stdout(sys.stdout)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")
sys.path.insert(0, BASE_DIR)
import child_budget

ROOT = r"Z:\2. Cost\★★★쿠팡 업무 폴더★★★"
BUDGET_ENV = "ZSCAN_BUDGET_SEC"
INCREMENTAL_RETURN_CODE = child_budget.INCREMENTAL_RETURN_CODE
DOC_PROGRESS_SCHEMA = 1
DOC_PROGRESS_NAME = ".zscan_docs_progress.json"
DOC_CHECKPOINT_DIRS = 250

# 기본 갈래(원장 누락·금액 공백 스캔)의 진도. `--docs` 와 **파일을 나눠 쓴다** —
# 담는 것이 서로 다르고, 한 파일을 같이 쓰면 한쪽이 다른 쪽 진도를 지운다.
SCAN_PROGRESS_NAME = ".zscan_scan_progress.json"
SCAN_PROGRESS_SCHEMA = 1

RE_UJ = re.compile(r"UJ\s?(\d{7})", re.I)
RE_PO = re.compile(r"PO\s?(\d{6})", re.I)
RE_MONEY = re.compile(r"([0-9]{1,3}(?:,[0-9]{3})+)\s*원")
RE_DATE = re.compile(r"(20\d{2})[.\-_]?(\d{2})[.\-_]?(\d{2})")

SKIP_NAMES = {"thumbs.db", "desktop.ini"}

# DB 무관 — 업무상 필요하지만 관리대장 어느 열에도 들어갈 자리가 없다.
IRRELEVANT_FOLDERS = (
    "설치공사 안전보건대장", "지게차 서류", "쿠팡 보험", "안전수칙",
    "교육, 위임장", "인건비", "기사님들 근태", "기본 도면", "사다리_계단",
)
# 서류 성격 — 이름에 프로젝트NO가 없어도 대조 대상이다.
DOC_WORDS = ("거래명세서", "세금계산서", "계산서", "견적", "발주", "정산", "청구", "입금", "수금", "PO")


def classify(path_rel, name):
    low = name.lower()
    if any(k in path_rel for k in IRRELEVANT_FOLDERS):
        return "무관(안전·교육·근태 등)"
    if RE_UJ.search(name) or RE_PO.search(name):
        return "관련(프로젝트NO·PO 명시)"
    if any(w in name for w in DOC_WORDS):
        return "관련(서류)"
    if low.endswith((".jpg", ".jpeg", ".png", ".heic")):
        return "무관(사진)"
    if low.endswith((".dwg", ".pptx", ".docx", ".hwp", ".zip")):
        return "참고(도면·문서)"
    return "미분류"


def scan(root=ROOT, only=None):
    out = []
    base = os.path.join(root, only) if only else root
    for dp, _dn, fn in os.walk(base):
        rel = os.path.relpath(dp, root)
        for f in fn:
            if f.lower() in SKIP_NAMES or f.startswith("~$"):
                continue
            kind = classify(rel, f)
            rec = {"폴더": rel, "파일": f, "분류": kind}
            if kind.startswith("관련"):
                rec["UJ"] = ["UJ" + m for m in RE_UJ.findall(f)]
                rec["PO"] = ["PO" + m for m in RE_PO.findall(f)]
                money = [int(x.replace(",", "")) for x in RE_MONEY.findall(f)]
                rec["금액"] = max(money) if money else None
            out.append(rec)
    return out


def _scan_progress_path(path=None):
    return path or os.path.join(REPORT_DIR, SCAN_PROGRESS_NAME)


def _new_scan_progress(root, only):
    return {"schema": SCAN_PROGRESS_SCHEMA,
            "root": os.path.normcase(os.path.abspath(root)),
            "only": only or "",
            "pending_dirs": [""],
            "files": 0, "kinds": {}, "codes": {}, "po": [], "folders": {},
            "scanned_dirs": 0,
            "started_at": time.strftime("%Y-%m-%dT%H:%M:%S%z")}


def _load_scan_progress(root, only, path=None):
    """같은 루트·같은 --folder 의 완전한 체크포인트만 이어받는다.

    깨진 JSON 을 빈 진행으로 삼키면 이미 센 디렉터리를 처음부터 다시 훑고도
    이유를 숨긴다([169]). 그래서 깨진 파일은 실패로 올린다 — `--docs` 와 같다.
    """
    path = _scan_progress_path(path)
    if not os.path.exists(path):
        return _new_scan_progress(root, only)
    with open(path, encoding="utf-8") as f:
        value = json.load(f)
    if (value.get("schema") != SCAN_PROGRESS_SCHEMA
            or value.get("root") != os.path.normcase(os.path.abspath(root))
            or value.get("only") != (only or "")):
        return _new_scan_progress(root, only)
    if not isinstance(value.get("pending_dirs"), list) or not isinstance(value.get("codes"), dict):
        raise ValueError("Z폴더 스캔 진도 파일 형식이 올바르지 않습니다: %s" % path)
    return value


def _save_scan_progress(value, path=None):
    _atomic_json(_scan_progress_path(path), value)


def _clear_scan_progress(path=None):
    try:
        os.unlink(_scan_progress_path(path))
    except FileNotFoundError:
        pass


def _fold_into(value, rel, name):
    """파일 한 개를 집계에 접는다. 레코드는 담지 않는다."""
    kind = classify(rel, name)
    value["files"] += 1
    value["kinds"][kind] = value["kinds"].get(kind, 0) + 1
    for k in IRRELEVANT_FOLDERS:
        if k in rel:
            value["folders"][k] = value["folders"].get(k, 0) + 1
    if not kind.startswith("관련"):
        return
    money = [int(x.replace(",", "")) for x in RE_MONEY.findall(name)]
    first_money = max(money) if money else None
    for m in RE_PO.findall(name):
        po = "PO" + m
        if po not in value["po"]:
            value["po"].append(po)
    for m in RE_UJ.findall(name):
        code = "UJ" + m
        cur = value["codes"].get(code)
        if cur is None:
            # `f`·`fm` = 그 코드의 **첫 파일**(리포트 표가 쓴다),
            # `m` = 금액이 있는 것만 모은 목록(금액 불일치 판정이 쓴다).
            cur = {"f": name, "fm": first_money, "m": []}
            value["codes"][code] = cur
        if first_money is not None:
            cur["m"].append(first_money)


def scan_incremental(root=ROOT, only=None, state_path=None, stop=None):
    """디렉터리 단위로 이어서 세고 **집계만** 남긴다.

    반환은 ``(집계, complete, 집계)``. ``complete`` 가 거짓이면 호출자는 기존
    리포트를 절대 덮지 않고 75 로 돌아가야 한다([169] — 반쪽 숫자를 완전한
    리포트처럼 내면 '원장에 없는 프로젝트NO' 가 실제보다 적게 나온다).

    ★ 레코드를 통째로 담지 않는다. 실측 2026-09-03 파일 **232,384개**라 그것을
      JSON 으로 250 디렉터리마다 저장하면 그 저장이 다시 회차를 먹는다
      (`--docs` 는 서류만 걸러 담아 이 문제가 없었다).
    """
    if not os.path.isdir(root):
        raise FileNotFoundError("폴더에 닿지 못했습니다(네트워크 드라이브 확인): %s" % root)
    stop = stop or child_budget.over
    value = _load_scan_progress(root, only, state_path)
    pending = value["pending_dirs"]
    since_checkpoint = 0
    base_rel = only or ""

    while pending:
        if stop():
            _save_scan_progress(value, state_path)
            return value, False, value

        rel = pending.pop()
        folder = root if not rel else os.path.join(root, rel)
        try:
            with os.scandir(folder) as it:
                entries = list(it)
            child_dirs, found = [], []
            for entry in entries:
                if entry.is_dir(follow_symlinks=False):
                    child_dirs.append(os.path.relpath(entry.path, root))
                    continue
                if entry.name.lower() in SKIP_NAMES or entry.name.startswith("~$"):
                    continue
                found.append(entry.name)
        except OSError:
            # 이 디렉터리는 **완주하지 않았다**. 다시 넣고 자국을 남긴 뒤 실패로 올린다.
            pending.append(rel)
            _save_scan_progress(value, state_path)
            raise

        # 디렉터리 하나를 끝까지 읽은 뒤에만 집계에 접는다(재개해도 중복되지 않는다).
        for name in found:
            _fold_into(value, rel, name)
        value["scanned_dirs"] = int(value.get("scanned_dirs") or 0) + 1
        child_dirs.sort(reverse=True)
        pending.extend(child_dirs)
        since_checkpoint += 1
        if since_checkpoint >= DOC_CHECKPOINT_DIRS:
            _save_scan_progress(value, state_path)
            since_checkpoint = 0

    _save_scan_progress(value, state_path)
    return value, True, value

def ledger_index():
    """원장 프로젝트NO → (시트, 금액). 06시트 금액이 있으면 그걸 쓴다."""
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from workbook_patch import latest_master
    path, ver = latest_master()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    idx = {}
    for sn, amt_col in (("02_돌발AS접수", None), ("04_정기점검", None),
                        ("05_신규납품설치", None), ("06_거래서류청구수금", "거래명세서합계")):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdr = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        if "프로젝트NO" not in hdr:
            continue
        j = hdr.index("프로젝트NO")
        ja = hdr.index(amt_col) if amt_col and amt_col in hdr else None
        for row in ws.iter_rows(min_row=5, values_only=True):
            code = row[j]
            if not code:
                continue
            code = str(code).strip().upper()
            amt = row[ja] if ja is not None and isinstance(row[ja], (int, float)) else None
            prev = idx.get(code)
            idx[code] = (sn, amt if amt else (prev[1] if prev else None))
    wb.close()
    return idx, ver


RE_DOC_DATE = re.compile(r"(20\d{2})[.\-_](\d{1,2})[.\-_](\d{1,2})")
DOC_KINDS = (("세금계산서", "세금계산서"), ("거래명세서", "거래명세서"), ("계산서", "계산서"))


def camp_key(s):
    """캠프명 비교용 — 괄호·공백·기호를 털어낸 핵심부.

    파일명은 `2026-06-19 구로1MB(독산동B) 거래명세서.pdf`, 원장은 `구로1MB(독산동B)` 처럼
    같은 표기를 쓰지만 띄어쓰기·괄호 안 표기가 흔들린다. 괄호 앞부분이 가장 안정적이다."""
    s = re.sub(r"\s+", "", str(s or ""))
    s = re.split(r"[（(]", s)[0]
    return re.sub(r"[^0-9A-Za-z가-힣]", "", s).lower()


def _doc_record(root, folder, name):
    """PDF 파일명 하나를 서류 대조 레코드로 바꾼다. 대상이 아니면 ``None``."""
    if not name.lower().endswith(".pdf"):
        return None
    kind = next((v for k, v in DOC_KINDS if k in name), None)
    if not kind:
        return None
    m = RE_DOC_DATE.search(name)
    if not m:
        return None
    y, mo, d = (int(x) for x in m.groups())
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    camp = name[m.end():]
    camp = re.sub(r"\.(pdf|PDF)$", "", camp)
    for k, _v in DOC_KINDS:
        camp = camp.replace(k, "")
    camp = camp.replace("(수정)", "").strip(" _-")
    if not camp:
        return None
    return {"일자": "%04d-%02d-%02d" % (y, mo, d), "캠프키": camp_key(camp),
            "캠프원문": camp, "종류": kind, "파일": name,
            "폴더": os.path.relpath(folder, root)}


def doc_catalog(root=ROOT):
    """거래명세서·세금계산서 PDF → [{일자, 캠프키, 캠프원문, 종류, 파일}]

    ★ 파일을 열지 않는다. 이 회사는 `날짜 + 캠프명 + 서류종류` 로 파일명을 짓고 있어
      **발행 여부와 발행일은 파일명만으로 확정된다.** 금액은 파일명에 거의 없으므로
      (3,856개 중 17개) 금액이 필요하면 그때 해당 PDF만 연다."""
    out = []
    for dp, _dn, fn in os.walk(root):
        for f in fn:
            rec = _doc_record(root, dp, f)
            if rec:
                out.append(rec)
    return out


def _atomic_json(path, value):
    """체크포인트를 같은 폴더 임시파일 뒤 원자 교체한다."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    temp = "%s.%d.tmp" % (path, os.getpid())
    try:
        with open(temp, "w", encoding="utf-8", newline="") as f:
            json.dump(value, f, ensure_ascii=False, separators=(",", ":"))
            f.write("\n")
            f.flush()
            os.fsync(f.fileno())
        os.replace(temp, path)
    finally:
        if os.path.exists(temp):
            os.unlink(temp)


def _doc_progress_path(path=None):
    return path or os.path.join(REPORT_DIR, DOC_PROGRESS_NAME)


def _new_doc_progress(root):
    return {"schema": DOC_PROGRESS_SCHEMA,
            "root": os.path.normcase(os.path.abspath(root)),
            "pending_dirs": [""], "docs": [],
            "scanned_dirs": 0, "scanned_files": 0,
            "started_at": time.strftime("%Y-%m-%dT%H:%M:%S%z")}


def _load_doc_progress(root, path=None):
    """같은 루트의 완전한 체크포인트만 이어받는다.

    깨진 JSON을 빈 진행으로 삼키면 이미 센 디렉터리를 처음부터 다시 훑고도 이유를
    숨긴다. 따라서 깨진 파일은 실패로 올려 사람이 원인을 볼 수 있게 한다.
    """
    path = _doc_progress_path(path)
    if not os.path.exists(path):
        return _new_doc_progress(root)
    with open(path, encoding="utf-8") as f:
        value = json.load(f)
    expected = os.path.normcase(os.path.abspath(root))
    if value.get("schema") != DOC_PROGRESS_SCHEMA or value.get("root") != expected:
        return _new_doc_progress(root)
    if not isinstance(value.get("pending_dirs"), list) or not isinstance(value.get("docs"), list):
        raise ValueError("Z폴더 서류대조 진도 파일 형식이 올바르지 않습니다: %s" % path)
    return value


def _save_doc_progress(value, path=None):
    _atomic_json(_doc_progress_path(path), value)


def _clear_doc_progress(path=None):
    try:
        os.unlink(_doc_progress_path(path))
    except FileNotFoundError:
        pass


def doc_catalog_incremental(root=ROOT, state_path=None, stop=None):
    """서류 목록을 디렉터리 단위로 이어서 센다.

    반환은 ``(docs, complete, progress)``. ``complete``가 거짓이면 호출자는 기존
    리포트를 절대 덮지 않고 75로 돌아가야 한다. 체크포인트에는 아직 볼 디렉터리와
    지금까지 확인한 **완전한 디렉터리**의 결과만 들어가므로 재개해도 중복되지 않는다.
    """
    if not os.path.isdir(root):
        raise FileNotFoundError("폴더에 닿지 못했습니다(네트워크 드라이브 확인): %s" % root)
    stop = stop or child_budget.over
    value = _load_doc_progress(root, state_path)
    pending = value["pending_dirs"]
    since_checkpoint = 0

    while pending:
        if stop():
            _save_doc_progress(value, state_path)
            return value["docs"], False, value

        rel = pending.pop()
        folder = root if not rel else os.path.join(root, rel)
        try:
            with os.scandir(folder) as it:
                entries = list(it)
            child_dirs, found = [], []
            file_count = 0
            for entry in entries:
                if entry.is_dir(follow_symlinks=False):
                    child_dirs.append(os.path.relpath(entry.path, root))
                    continue
                file_count += 1
                rec = _doc_record(root, folder, entry.name)
                if rec:
                    found.append(rec)
        except OSError:
            # 이 디렉터리는 **완주하지 않았다**. 다시 넣고 자국을 남긴 뒤 실패로 올린다.
            pending.append(rel)
            _save_doc_progress(value, state_path)
            raise

        # 디렉터리 하나를 끝까지 읽은 뒤에만 결과와 자식 목록을 확정한다.
        value["docs"].extend(found)
        value["scanned_files"] = int(value.get("scanned_files") or 0) + file_count
        value["scanned_dirs"] = int(value.get("scanned_dirs") or 0) + 1
        child_dirs.sort(reverse=True)
        pending.extend(child_dirs)
        since_checkpoint += 1
        if since_checkpoint >= DOC_CHECKPOINT_DIRS:
            _save_doc_progress(value, state_path)
            since_checkpoint = 0

    # 리포트 원자 교체가 실패해도 이 완주 체크포인트로 다시 쓸 수 있게 먼저 저장한다.
    _save_doc_progress(value, state_path)
    return value["docs"], True, value


def _write_docs_report(path, docs, ver, paired, orphan, ambiguous):
    """완주 결과만 임시파일에 다 쓴 뒤 기존 리포트를 원자 교체한다."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    temp = "%s.%d.tmp" % (path, os.getpid())
    try:
        with open(temp, "w", encoding="utf-8", newline="") as f:
            f.write("# 거래명세서·세금계산서 ↔ 관리대장 대조 (파일명 근거)\n\n")
            f.write("- 서류 PDF %d개 · 관리대장 v%d\n" % (len(docs), ver))
            f.write("- 1:1 확정 **%d** · 후보 여럿 %d · 원장에 짝 없음 %d\n" %
                    (len(paired), len(ambiguous), len(orphan)))
            f.write("- 판정 기준: 캠프명이 같고 날짜가 ±7일. 둘 다 맞을 때만 인정한다.\n")
            f.write("- ★ 읽기 전용 — 원장에 쓰지 않는다.\n\n")
            f.write("## 1:1 확정 (서류 확보 완료)\n\n| 시트 | ID | 프로젝트NO | 캠프 | 원장일자 | 서류 | 파일 |\n")
            f.write("|---|---|---|---|---|---|---|\n")
            for d, r in paired[:600]:
                f.write("| %s | %s | %s | %s | %s | %s | %s |\n" %
                        (r["시트"][:2], r["ID"], r["프로젝트NO"], r["캠프명"][:16],
                         r["일자"], d["종류"], d["파일"][:52]))
            if len(paired) > 600:
                f.write("\n(상위 600건만 표기 — 전체 %d건)\n" % len(paired))
            f.write("\n## 원장에 짝이 없는 서류 (%d개) — 원장에 그 작업이 없다는 뜻\n\n" % len(orphan))
            f.write("| 일자 | 캠프 | 종류 | 폴더 |\n|---|---|---|---|\n")
            for d in orphan[:300]:
                f.write("| %s | %s | %s | %s |\n" %
                        (d["일자"], d["캠프원문"][:22], d["종류"], d["폴더"][:34]))
            if len(orphan) > 300:
                f.write("\n(상위 300건만 표기 — 전체 %d건)\n" % len(orphan))
            f.flush()
            os.fsync(f.fileno())
        os.replace(temp, path)
    finally:
        if os.path.exists(temp):
            os.unlink(temp)


def ledger_rows_for_docs():
    """02·04 시트의 (캠프키, 날짜, ID, 프로젝트NO) — 서류 대조용."""
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from workbook_patch import latest_master
    path, ver = latest_master()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []
    for sn, idcol, datecols in (("02_돌발AS접수", "접수ID", ("작업완료일", "접수일자")),
                                ("04_정기점검", "점검ID", ("실제점검일", "점검예정일"))):
        ws = wb[sn]
        hdr = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        gi = {c: hdr.index(c) for c in hdr if c}
        for r in ws.iter_rows(min_row=5, values_only=True):
            prj = r[gi["프로젝트NO"]] if "프로젝트NO" in gi else None
            camp = r[gi["캠프명"]] if "캠프명" in gi else None
            if not (prj or camp):
                continue
            day = ""
            for dc in datecols:
                v = r[gi[dc]] if dc in gi else None
                if hasattr(v, "date"):
                    day = v.date().isoformat()
                    break
                if isinstance(v, str) and RE_DOC_DATE.search(v):
                    day = v[:10]
                    break
            rows.append({"시트": sn, "ID": r[gi[idcol]] if idcol in gi else "",
                         "프로젝트NO": str(prj or ""), "캠프명": str(camp or ""),
                         "캠프키": camp_key(camp), "일자": day})
    wb.close()
    return rows, ver


def match_docs(docs, rows, tol_days=7):
    """캠프키가 같고 날짜가 ±7일이면 그 작업의 서류로 본다.

    날짜만·캠프만으로는 못 잇는다 — 같은 캠프에 한 달에 여러 건이 있고,
    같은 날 여러 캠프를 돈다. 둘 다 맞을 때만 인정하고, 후보가 여럿이면 사람에게 넘긴다."""
    from datetime import date as _d

    def dt(s):
        try:
            y, m, dd = (int(x) for x in s.split("-"))
            return _d(y, m, dd)
        except Exception:
            return None

    by_camp = defaultdict(list)
    for r in rows:
        if r["캠프키"] and r["일자"]:
            by_camp[r["캠프키"]].append(r)
    paired, orphan, ambiguous = [], [], []
    for d in docs:
        cands = [r for r in by_camp.get(d["캠프키"], [])
                 if dt(r["일자"]) and dt(d["일자"])
                 and abs((dt(r["일자"]) - dt(d["일자"])).days) <= tol_days]
        if len(cands) == 1:
            paired.append((d, cands[0]))
        elif not cands:
            orphan.append(d)
        else:
            # ★ 후보가 여럿이면 **날짜가 정확히 같은 것**을 먼저 본다. 그런 게 딱 하나면
            #   그게 그 작업이다(같은 캠프를 같은 날 두 번 가는 일은 없다).
            #   그래도 여럿이면 손대지 않는다 — 엉뚱한 행에 '발행완료'가 찍히면 되돌리기 어렵다.
            exact = [r for r in cands if r["일자"][:10] == d["일자"]]
            if len(exact) == 1:
                paired.append((d, exact[0]))
            else:
                ambiguous.append((d, cands))
    return paired, orphan, ambiguous


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--folder", help="특정 하위 폴더만")
    ap.add_argument("--docs", action="store_true", help="거래명세서·세금계산서 PDF ↔ 원장 대조")
    ap.add_argument("--queue-docs", action="store_true",
                    help="1:1 확정 건만 04시트 거래명세서='발행완료' 로 큐 적재")
    ap.add_argument("--root", default=ROOT)
    args = ap.parse_args()

    if args.queue_docs:
        docs = doc_catalog(args.root)
        rows, ver = ledger_rows_for_docs()
        paired, _orphan, ambiguous = match_docs(docs, rows)
        # 04시트만 — 02_돌발AS에는 거래명세서 열이 아예 없다(06시트 소관인데 그 행이 없다).
        # 드롭다운(10_코드관리!AB) 이 허용하는 값은 발행완료·미발행·해당없음 뿐이다.
        best = {}
        for d, r in paired:
            if r["시트"] != "04_정기점검" or not r["프로젝트NO"]:
                continue
            best.setdefault(r["프로젝트NO"], (d, r))
        items = []
        for prj, (d, r) in sorted(best.items()):
            items.append({"sheet": "04_정기점검", "key_col": "프로젝트NO", "key": prj,
                          "col": "거래명세서", "value": "발행완료", "vtype": "text",
                          "evidence": "쿠팡 업무 폴더 서류 %s (%s · %s)" % (d["파일"][:60], d["일자"], d["종류"]),
                          "only_if_empty": True})
        skipped02 = len({r["프로젝트NO"] for d, r in paired if r["시트"] == "02_돌발AS접수"})
        print("1:1 확정 중 04_정기점검 %d건 적재 · 02_돌발AS %d건은 **열이 없어 보류**(06시트 소관)"
              % (len(items), skipped02))
        print("후보 여럿 %d건은 쓰지 않는다 — 같은 캠프에 여러 건이라 사람이 골라야 한다" % len(ambiguous))
        if not items:
            return 0
        from ledger_writer import queue_add
        print("큐 적재:", queue_add(items), "개 셀 → ledger_db --intake 후 11:00·15:00 원장 반영")
        return 0

    if args.docs:
        budget = child_budget.start(BUDGET_ENV)  # 안 주면 0 = 직접 실행은 종전처럼 무제한
        docs, complete, progress = doc_catalog_incremental(args.root)
        if not complete:
            print("Z폴더 서류 확인 진행: 디렉터리 %d개 · 파일 %d개 · 서류 %d개 · 남은 디렉터리 %d개"
                  % (progress["scanned_dirs"], progress["scanned_files"], len(docs),
                     len(progress["pending_dirs"])), flush=True)
            print("★ 시간 예산(%d초)이 다 되어 기존 완전한 리포트는 고치지 않고 다음 회차로 넘깁니다."
                  % budget, flush=True)
            print("증분 수집 계속 필요 — 실패가 아니며 저장된 디렉터리부터 이어서 확인합니다.",
                  flush=True)
            return INCREMENTAL_RETURN_CODE
        rows, ver = ledger_rows_for_docs()
        paired, orphan, ambiguous = match_docs(docs, rows)
        kinds = Counter(d["종류"] for d in docs)
        print("서류 PDF %d개 (%s) · 원장(v%d) 대조행 %d"
              % (len(docs), " · ".join("%s %d" % kv for kv in kinds.most_common()), ver, len(rows)))
        print("  ★ 1:1 확정 %d · 후보 여럿 %d · 원장에 짝 없음 %d"
              % (len(paired), len(ambiguous), len(orphan)))
        have = {(p[1]["시트"], p[1]["ID"] or p[1]["프로젝트NO"]) for p in paired}
        print("  서류가 확인된 원장 행 %d개" % len(have))
        os.makedirs(REPORT_DIR, exist_ok=True)
        out = os.path.join(REPORT_DIR, "Z폴더_서류대조.md")
        _write_docs_report(out, docs, ver, paired, orphan, ambiguous)
        _clear_doc_progress()
        print("리포트:", out)
        return 0

    if not os.path.isdir(args.root):
        print("폴더에 닿지 못했습니다(네트워크 드라이브 확인):", args.root)
        return 1

    recs = scan(args.root, args.folder)
    kinds = Counter(r["분류"] for r in recs)
    rel = [r for r in recs if r["분류"].startswith("관련")]
    print("파일 %d개 — %s" % (len(recs), " · ".join("%s %d" % kv for kv in kinds.most_common())))

    idx, ver = ledger_index()
    by_code = defaultdict(list)
    for r in rel:
        for c in r.get("UJ", []):
            by_code[c].append(r)
    codes = set(by_code)
    known = codes & set(idx)
    unknown = codes - set(idx)
    print("파일명에서 얻은 프로젝트NO %d개 — 원장(v%d)에 있음 %d · **없음 %d**"
          % (len(codes), ver, len(known), len(unknown)))

    # 금액이 파일명과 원장 양쪽에 있는 것만 비교한다. 한쪽만 있으면 비교가 아니다.
    gaps = []
    for c in sorted(known):
        led = idx[c][1]
        fam = [r["금액"] for r in by_code[c] if r.get("금액")]
        if led and fam and all(abs(led - x) > 1 for x in fam):
            gaps.append((c, idx[c][0], led, fam[0]))

    po = {p for r in rel for p in r.get("PO", [])}
    print("PO번호 %d개 · 금액 비교 가능 %d건 중 **불일치 %d건**"
          % (len(po), sum(1 for c in known if idx[c][1] and any(r.get("금액") for r in by_code[c])), len(gaps)))

    os.makedirs(REPORT_DIR, exist_ok=True)
    out = os.path.join(REPORT_DIR, "Z폴더_스캔.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("# 쿠팡 업무 폴더 전수 조사\n\n")
        f.write("- 대상: `%s`%s\n" % (args.root, (" / " + args.folder) if args.folder else ""))
        f.write("- 파일 %d개 · 관리대장 v%d 기준\n" % (len(recs), ver))
        f.write("- ★ 이 도구는 원장에 아무것도 쓰지 않는다(읽기 전용). 넣을 것은 사람이 고른다.\n\n")
        f.write("## 분류\n\n| 분류 | 건수 |\n|---|---:|\n")
        for k, v in kinds.most_common():
            f.write("| %s | %d |\n" % (k, v))
        f.write("\n## 원장에 없는 프로젝트NO (%d개)\n\n" % len(unknown))
        f.write("| 프로젝트NO | 금액(파일명) | 파일 |\n|---|---:|---|\n")
        for c in sorted(unknown):
            r = by_code[c][0]
            f.write("| %s | %s | %s |\n" % (c, format(r["금액"], ",") if r.get("금액") else "-",
                                            r["파일"][:70]))
        f.write("\n## 금액 불일치 (%d건) — 파일명 vs 원장\n\n" % len(gaps))
        f.write("| 프로젝트NO | 시트 | 원장 | 파일명 | 차액 |\n|---|---|---:|---:|---:|\n")
        for c, sn, led, fam in gaps:
            f.write("| %s | %s | %s | %s | %s |\n" % (c, sn, format(int(led), ","),
                                                     format(int(fam), ","), format(int(fam - led), ",")))
        f.write("\n## 무관으로 분류해 **적용하지 않은** 폴더\n\n")
        f.write("업무상 필요한 자료지만 관리대장 02·04·05·06 어느 열에도 들어갈 자리가 없다.\n\n")
        for k in IRRELEVANT_FOLDERS:
            n = sum(1 for r in recs if k in r["폴더"])
            if n:
                f.write("- %s — %d개\n" % (k, n))
    print("리포트:", out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
