# -*- coding: utf-8 -*-
"""
side_excel_retire.py — 관리대장 시트로 옮겨 간 **별도 엑셀**을 접는다
================================================================================
사용자 지시(2026-08-07): **"관리대장 이 파일로만 관리해, 그리고 필요없는 엑셀은 지워."**

무엇을 하나
  『23_확인필요현황』·『29_거래처코드』 시트가 관리대장 안에 **실제로 채워진 것을 확인한
  뒤에만**, 같은 내용을 담던 옛 별도 엑셀을 `OLD/` 로 옮긴다.

★ 왜 사람이 손으로 지우지 않고 이 도구가 하나
  "시트를 만들었으니 이제 파일을 지워도 된다"를 **사람이 판단하면 순서를 놓친다.**
  시트가 비어 있는데 파일을 먼저 지우면 그 자료는 그 순간 아무 데도 없다. 그래서
  판단 기준을 기계가 가진다 — 시트가 있고, 보고 시트 표식을 달고 있고, 데이터가
  한 줄이라도 있어야 옮긴다. 하나라도 아니면 **아무것도 하지 않는다.**

★ 왜 지우지 않고 OLD 로 옮기나
  이 프로젝트의 '지운다'는 늘 `OLD/` 다(2026-07-27 지시 "이전 버전은 전부 OLD 폴더에
  저장", `ledger_versions.ARCHIVE`). 폴더 하나 차이일 뿐 눈앞에서는 사라지고,
  잘못됐을 때 되돌릴 수 있다. 지우기는 되돌릴 수 없다.

실행
  python side_excel_retire.py            # 무엇을 옮길지만 보여 준다
  python side_excel_retire.py --apply    # 실제로 옮긴다 (11:00·15:00 회차가 부른다)
"""
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# (관리대장 시트, 그 시트가 대신하게 된 옛 별도 엑셀)
PAIRS = [
    ("23_확인필요현황", "쿠팡_확인필요현황_최신.xlsx"),
    ("29_거래처코드", "쿠팡_거래처코드_최신.xlsx"),
]
ARCHIVE = "OLD"          # ledger_versions 와 같은 자리를 쓴다
MIN_ROWS = 5             # 4행이 머리글 — 5행이 있어야 데이터가 한 줄이라도 있는 것


def check(master, sheet):
    """이 시트가 '옛 파일을 접어도 될 만큼' 채워졌나. (된다, 이유) 를 돌려준다."""
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from findings_export import is_agent_sheet
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return False, "시트가 아직 없다"
        ws = wb[sheet]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not is_agent_sheet(row1):
            # 이름만 같고 사람이 만든 다른 시트일 수 있다 — 그러면 손대지 않는다.
            return False, "보고 시트 표식이 없다(사람이 만든 시트일 수 있다)"
        if (ws.max_row or 0) < MIN_ROWS:
            return False, f"시트가 비어 있다({ws.max_row}행)"
        return True, f"{ws.max_row - 4}행"
    finally:
        wb.close()


def free_name(folder, name):
    """OLD 에 같은 이름이 있으면 덮지 않고 옆에 둔다 — 옛것도 근거다."""
    dst = os.path.join(folder, name)
    if not os.path.exists(dst):
        return dst
    stem, ext = os.path.splitext(name)
    for i in range(2, 100):
        cand = os.path.join(folder, f"{stem} ({i}){ext}")
        if not os.path.exists(cand):
            return cand
    return None


def run(apply=False, master=None):
    from ecount_reconcile import load_config, resolve_master
    master = master or resolve_master(load_config()["reconcile"]["master_xlsx"])
    folder = os.path.dirname(master)
    out = []
    for sheet, side in PAIRS:
        path = os.path.join(folder, side)
        if not os.path.exists(path):
            out.append((sheet, side, None, "이미 정리됨"))
            continue
        ok, why = check(master, sheet)
        if not ok:
            out.append((sheet, side, False, f"보류 — {why}"))
            continue
        if not apply:
            out.append((sheet, side, True, f"옮길 수 있다({why}) — --apply 로 실행"))
            continue
        dest_dir = os.path.join(folder, ARCHIVE)
        os.makedirs(dest_dir, exist_ok=True)
        dst = free_name(dest_dir, side)
        try:
            os.replace(path, dst)          # 같은 드라이브라 이동은 원자적이다
            out.append((sheet, side, True, f"{ARCHIVE}/ 로 옮김({why})"))
        except OSError as exc:
            # 사람이 열어 두면 잠긴다. 실패해도 회차 전체를 멈추지 않는다 — 다음에 또 온다.
            out.append((sheet, side, False, f"옮기지 못함(열려 있나?) — {type(exc).__name__}"))
    return out


def main():
    apply = "--apply" in sys.argv
    rows = run(apply=apply)
    for sheet, side, ok, msg in rows:
        mark = "·" if ok is None else ("✅" if ok else "…")
        print(f"  {mark} {sheet:<16} {side:<28} {msg}")
    moved = sum(1 for _s, _f, ok, m in rows if ok and ARCHIVE in str(m))
    held = sum(1 for _s, _f, ok, _m in rows if ok is False)
    print(f"별도 엑셀 정리: 옮김 {moved} · 보류 {held} · 대상 {len(rows)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
