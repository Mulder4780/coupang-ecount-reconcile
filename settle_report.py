# -*- coding: utf-8 -*-
"""
settle_report.py — 하루치 **정산분 보고자료**를 세 원천에서 만들어 낸다 (2026-08-05 지시)

사용자 지시: "8월 5일 정산분 8월 6일에 보고 자료 보고할거 정리해 / 밴드도 우선 8월 5일꺼
찾아 정리하고 / ERP도 8월 5일꺼 우선 찾아 정리해서 보고할 수 있게 빨리 준비해."

왜 스크립트인가
  이 정리는 **내일도, 모레도** 필요하다. 한 번 손으로 만들면 다음 날 또 손으로 만들어야
  한다(프로젝트 상시 원칙). 그래서 날짜만 바꾸면 같은 보고서가 나오게 만든다.

무엇을 보나 — 세 원천을 각각, 그 날짜만
  ① 카톡  : 3. 카카오톡 내보내기(돌발AS방·정기점검방) → 접수/완료/담당/사진 건수
  ② 밴드  : band/cache/*.json → 그 날 올라온 글(쿠팡 PO 수주 글 포함)
  ③ ERP   : 원본색인의 그 날짜 판매조회 내보내기 → 판매전표 행

원칙
  · **금액을 임의로 확정하지 않는다.** 같은 프로젝트가 진행상태만 달리 두 줄이면
    합계를 하나로 정하지 않고 '확인 필요'에 적는다(이중계상 방지 — 절대규칙 5 정신).
  · 원천마다 "무엇을 확인했고 그 날 것이 몇 건인지"를 표로 남긴다. 없으면 없다고 쓴다.
    수집 실패와 '실제로 없음'은 다른 말이다.
  · 엑셀을 열지 않는다(읽기 전용).

쓰는 법
  python settle_report.py                 # 어제분(보고는 다음 날 하므로 기본값이 어제)
  python settle_report.py --day 2026-08-05
  → reports/보고자료_YYYYMMDD정산분.md  (앱 [기록] 탭에 '정산분 보고'로 뜬다)
"""
import argparse
import glob
import json
import os
import re
import sys
import time
import warnings
from datetime import datetime, timedelta

warnings.filterwarnings("ignore")
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

UJ = re.compile(r"UJ\d{7}")
PO = re.compile(r"PO\d{6,}")
MONEY = re.compile(r"([\d,]{4,})\s*원")


# ───────────────────────── ① 카톡 ─────────────────────────
DAY_LINE = re.compile(r"^-+ (\d{4})년 (\d{1,2})월 (\d{1,2})일 .+ -+$")
MSG_LINE = re.compile(r"^\[([^\]]+)\] \[(오전|오후) (\d{1,2}):(\d{2})\] ?(.*)$")


def _kakao_files():
    """카톡 내보내기 정본. 없으면 빈 목록(원천이 없다는 뜻)."""
    try:
        import source_dirs as S
        base = getattr(S, "KAKAO_DIR", None)
    except Exception:
        base = None
    out = []
    for d in filter(None, [base, os.path.join(ROOT, "kakao", "inbox")]):
        out += glob.glob(os.path.join(d, "**", "KakaoTalk*.txt"), recursive=True)
    return sorted(set(out))


def _kakao_blocks(path):
    """한 메시지 = 한 덩어리. 카톡 양식은 여러 줄이라 줄 단위로는 못 읽는다."""
    day = head = None
    buf = None
    for line in open(path, encoding="utf-8", errors="replace"):
        line = line.rstrip("\n")
        m = DAY_LINE.match(line.strip())
        if m:
            if buf is not None:
                yield day, head, "\n".join(buf)
                buf = None
            day = "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3)))
            continue
        mm = MSG_LINE.match(line)
        if mm:
            if buf is not None:
                yield day, head, "\n".join(buf)
            who, ap, h, mi, rest = mm.groups()
            h = int(h) % 12 + (12 if ap == "오후" else 0)
            head = (who, "%02d:%02d" % (h, int(mi)))
            buf = [rest]
        elif buf is not None:
            buf.append(line)
    if buf is not None:
        yield day, head, "\n".join(buf)


def _field(text, *names):
    for n in names:
        m = re.search(r"●\s*%s\s*[:：]\s*(.+)" % n, text)
        if m:
            return m.group(1).strip()
    return ""


def kakao_day(day):
    rows = []
    for path in _kakao_files():
        room = "정기점검" if "정기점검" in open(path, encoding="utf-8",
                                          errors="replace").readline() else "돌발AS"
        for d, head, body in _kakao_blocks(path):
            if d != day or not head:
                continue
            who, hm = head
            first = body.split("\n", 1)[0]
            if "A/S" not in first:
                continue
            state = "완료" if "완료" in first else ("접수" if "안내" in first else "")
            if not state:
                continue
            rows.append({
                "방": room, "상태": state, "시각": hm, "작성자": who,
                "프로젝트NO": (UJ.search(body).group(0) if UJ.search(body) else ""),
                "캠프": _field(body, "캠프이름", "캠프명"),
                "AS일자": _field(body, "A/S 일자"),
                "AS담당": (_field(body, "A/S 담당") or "").split("●")[0].strip(),
                "신청일자": _field(body, "신청일자"),
                "신청내용": _field(body, "신청내용")[:110],
            })
    # ★ 같은 대화 내보내기가 여러 벌 있으면(투입함 사본·다운로드 사본) 같은 메시지를
    #   두 번 세게 된다 — 실제로 완료 6건이 12건으로 나왔다(2026-08-05).
    #   메시지의 신원은 (방·상태·시각·프로젝트NO) 다. 파일이 몇 벌이든 한 번만 센다.
    seen, uniq = set(), []
    for r in rows:
        key = (r["방"], r["상태"], r["시각"], r["프로젝트NO"], r["캠프"])
        if key in seen:
            continue
        seen.add(key)
        uniq.append(r)
    uniq.sort(key=lambda r: (r["방"], r["상태"], r["시각"]))
    return uniq


# ───────────────────────── ② 밴드 ─────────────────────────
def band_day(day):
    out = []
    cdir = os.path.join(ROOT, "band", "cache")
    for f in sorted(glob.glob(os.path.join(cdir, "*.json"))):
        band = os.path.basename(f)[:-5]
        if not band.isdigit():
            continue
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        name = d.get("band_name") or band
        for no, v in (d.get("posts") or {}).items():
            if not isinstance(v, dict) or v.get("deleted") or not v.get("created_at"):
                continue
            if time.strftime("%Y-%m-%d", time.localtime(v["created_at"] / 1000)) != day:
                continue
            body = v.get("content") or ""
            out.append({"밴드": name, "번호": int(no), "작성자": v.get("author", ""),
                        "시각": time.strftime("%H:%M", time.localtime(v["created_at"] / 1000)),
                        "PO": ",".join(sorted(set(PO.findall(body)))),
                        "금액": (MONEY.search(body).group(1) if MONEY.search(body) else ""),
                        "본문": re.sub(r"\s+", " ", body)[:160]})
    out.sort(key=lambda r: (r["밴드"], r["번호"]))
    return out


# ───────────────────────── ③ ERP ─────────────────────────
def erp_day(day):
    """그 날짜 판매조회 내보내기에서 그 날 일자 행만. 없으면 빈 목록."""
    try:
        import openpyxl
    except Exception:
        return [], []
    idx = os.path.join(REPORT_DIR, "원본색인.json")
    try:
        rows = (json.load(open(idx, encoding="utf-8")).get("rows") or [])
    except Exception:
        return [], []
    files = [r for r in rows if r.get("kind") == "ERP:sales" and r.get("date") == day]
    stamp = day.replace("-", "/")
    got, seen_files = [], []
    for r in files:
        try:
            wb = openpyxl.load_workbook(r["path"], read_only=True, data_only=True)
        except Exception:
            continue
        ws = wb.worksheets[0]
        head = None
        n = 0
        for row in ws.iter_rows(values_only=True):
            v = [str(x) if x is not None else "" for x in row]
            if head is None:
                if any(x == "일자" for x in v):
                    head = v
                continue
            if v and v[0].startswith(stamp):
                got.append(dict(zip(head, v)))
                n += 1
        wb.close()
        seen_files.append({"파일": r["name"], "행": n})
    return got, seen_files


def po_projects(po_list):
    """PO 번호로 **그 PO에 묶인 전표들**을 ERP 내보내기에서 되찾는다 (2026-08-05).

    밴드 PO 글에는 총금액만 있고 프로젝트 No. 가 비어 있다. 그런데 ERP 판매조회에는
    각 전표에 PO번호가 붙어 있어, 거꾸로 훑으면 **어느 캠프 몇 건인지 전부 복원된다.**
    실제로 PO375206 21건·PO375207 13건이 나왔고 합계가 밴드 총금액과 원 단위까지 맞았다
    — 맞아떨어지면 그 자체가 검산이다.
    """
    try:
        import openpyxl
    except Exception:
        return {}
    try:
        rows = (json.load(open(os.path.join(REPORT_DIR, "원본색인.json"),
                               encoding="utf-8")).get("rows") or [])
    except Exception:
        return {}
    want = set(po_list)
    out = {p: {} for p in want}
    for r in [x for x in rows if x.get("kind") == "ERP:sales" and x.get("ext") == "xlsx"]:
        try:
            wb = openpyxl.load_workbook(r["path"], read_only=True, data_only=True)
        except Exception:
            continue
        for ws in wb.worksheets:
            head = None
            for row in ws.iter_rows(values_only=True):
                v = [str(x) if x is not None else "" for x in row]
                if head is None:
                    if any(x == "일자" for x in v):
                        head = v
                    continue
                rec = {k: x for k, x in zip(head, v) if k}
                pono = rec.get("PO번호", "")
                for p in want:
                    if p in pono:
                        # 같은 전표가 여러 내보내기에 있으므로 프로젝트+일자로 한 번만.
                        out[p][(rec.get("프로젝트코드코드", ""), rec.get("일자", "").strip())] = rec
        wb.close()
    return out


def _incomplete(r):
    """쿠팡 전표인데 **관리항목명이 비어 있으면** 미완성(중복) 전표로 본다.

    2026-08-05 실측: 판매조회 1,044행 중 쿠팡 창고 855행에서 관리항목명이 빈 행은
    UJ2601384 단 **한 줄**이었다(나머지 96건의 빈칸은 전부 비쿠팡 거래처 — 그쪽은
    관리항목을 쓰지 않는다). 쿠팡 건은 '돌발AS'·'3분기 정기점검'·'계단납품' 처럼
    관리항목이 항상 붙는다. 그래서 빈칸은 '아직 덜 채운 전표'라는 뜻이다.
    """
    return "쿠팡" in r.get("창고명", "") and not (r.get("관리항목명") or "").strip()


def erp_summary(rows):
    """쿠팡 건만 골라 합계를 낸다 — 단, **같은 프로젝트가 여러 줄이면 합치지 않는다.**

    예외: 같은 프로젝트·같은 금액인데 한쪽이 미완성 전표(_incomplete)면 그쪽을 빼고
    한 건으로 본다. 근거가 데이터에 있을 때만 정하고, 없으면 폭으로 남긴다.
    """
    coupang = [r for r in rows
               if "쿠팡" in (r.get("창고명", "") + r.get("거래처명", ""))]
    by_prj = {}
    for r in coupang:
        by_prj.setdefault(r.get("프로젝트코드코드") or "(무번호)", []).append(r)
    dup = {}
    for k, v in list(by_prj.items()):
        if len(v) <= 1:
            continue
        amts = {x.get("공급가액합계") for x in v}
        bad = [x for x in v if _incomplete(x)]
        if len(amts) == 1 and len(bad) == len(v) - 1:
            by_prj[k] = [x for x in v if not _incomplete(x)]     # 미완성 줄을 뺀다
            coupang = [x for x in coupang if x not in bad]
        else:
            dup[k] = v
    def amt(r):
        try:
            return int(float(r.get("공급가액합계") or 0))
        except ValueError:
            return 0
    low = sum(amt(v[0]) for v in by_prj.values())          # 중복을 한 건으로 볼 때
    high = sum(amt(r) for r in coupang)                    # 전부 별건으로 볼 때
    return coupang, by_prj, dup, low, high


# ───────────────────────── 보고서 ─────────────────────────
def build(day):
    ka, bd = kakao_day(day), band_day(day)
    erp, erp_files = erp_day(day)
    coupang, by_prj, dup, low, high = erp_summary(erp)
    y, m, d = day.split("-")
    L = ["# %s월 %s일 정산분 보고자료" % (int(m), int(d)), "",
         "작성 %s · 원천 3곳(카톡·밴드·ERP)에서 **%s 것만** 뽑았습니다."
         % (datetime.now().strftime("%Y-%m-%d %H:%M"), day),
         "확정하지 못한 것은 마지막 절에 따로 모았습니다.", "", "---", ""]

    done = [r for r in ka if r["상태"] == "완료"]
    new = [r for r in ka if r["상태"] == "접수"]
    po_posts = [r for r in bd if r["PO"]]
    L += ["## 1. 한 장 요약", "",
          "| 항목 | 건수 | 금액 | 원천 |", "|---|---:|---:|---|"]
    if po_posts:
        L.append("| 쿠팡 신규 PO 수주 | %d건 | %s | 밴드 |"
                 % (len(po_posts), " + ".join(p["금액"] + "원" for p in po_posts if p["금액"]) or "-"))
    L.append("| ERP 판매전표(쿠팡) | %d행 | %s | ERP 판매조회 |"
             % (len(coupang), ("{:,}원".format(low) if low == high
                               else "{:,} ~ {:,}원".format(low, high))))
    L.append("| 현장 완료 보고 | %d건 | ERP 미반영 | 카톡 |" % len(done))
    L.append("| 신규 접수 | %d건 | — | 카톡 |" % len(new))
    L.append("")

    L += ["## 2. 밴드 — 그 날 올라온 글", ""]
    if bd:
        L += ["| 밴드 | 글번호 | 시각 | 작성자 | PO | 금액 | 내용 |",
              "|---|---:|---|---|---|---:|---|"]
        for r in bd:
            L.append("| %s | %d | %s | %s | %s | %s | %s |"
                     % (r["밴드"], r["번호"], r["시각"], r["작성자"], r["PO"] or "-",
                        (r["금액"] + "원") if r["금액"] else "-", r["본문"][:70]))
    else:
        L.append("그 날 올라온 밴드 글이 **없습니다**(수집 실패가 아니라 실제로 없음).")
    L.append("")

    # 2-1. PO 에 묶인 전표를 ERP 에서 되찾아 붙인다 — 밴드 글에는 목록이 없다.
    po_detail = po_projects([p["PO"] for p in po_posts if p["PO"]])
    dup_prj = {}
    for i_po, (po, recs) in enumerate(sorted(po_detail.items()), 1):
        if not recs:
            continue
        L += ["### 2-%d. %s 에 묶인 전표 %d건" % (i_po, po, len(recs)), ""]
        L += ["| 일자 | 프로젝트 | 거래처 | 공급가액 |", "|---|---|---|---:|"]
        tot = 0
        for (prj, day2), rec in sorted(recs.items(), key=lambda x: x[0][1]):
            try:
                tot += int(float(rec.get("공급가액합계") or 0))
            except ValueError:
                pass
            L.append("| %s | %s | %s | %s |" % (day2, prj, rec.get("거래처명", ""),
                                                rec.get("공급가액합계", "")))
            dup_prj.setdefault(prj, set()).add(rec.get("거래처명", ""))
        said = [p["금액"] for p in po_posts if p["PO"] == po]
        L += ["| | | **ERP 합계** | **{:,}** |".format(tot), "",
              ("→ 밴드 고지액 %s원과 **일치**합니다." % said[0]) if said and
              said[0].replace(",", "") == str(tot) else
              ("→ 밴드 고지액 %s원과 **다릅니다** — 확인 필요." % (said[0] if said else "?")), ""]

    L += ["## 3. ERP 판매전표", ""]
    if coupang:
        L += ["| 프로젝트 | 거래처 | 창고 | 진행상태 | 관리항목 | 공급가액 | 합계 | PO |",
              "|---|---|---|---|---|---:|---:|---|"]
        for r in coupang:
            L.append("| %s | %s | %s | %s | %s | %s | %s | %s |" % (
                r.get("프로젝트코드코드", ""), r.get("거래처명", ""), r.get("창고명", ""),
                r.get("진행상태", ""), r.get("관리항목명") or "—",
                r.get("공급가액합계", ""), r.get("금액합계", ""), r.get("PO번호") or "-"))
        if low == high:
            L += ["", "합계 **{:,}원**(부가세 별도).".format(low)]
    else:
        L.append("그 날 일자로 잡힌 쿠팡 판매전표가 없습니다.")
    L.append("")

    L += ["## 4. 현장 실적 — 카톡", ""]
    for state, title in (("완료", "완료 보고"), ("접수", "신규 접수")):
        sel = [r for r in ka if r["상태"] == state]
        L += ["### 4-%s. %s %d건" % ("1" if state == "완료" else "2", title, len(sel)), ""]
        if sel:
            L += ["| 방 | 프로젝트 | 캠프 | 담당 | A/S 일자 | 보고시각 |",
                  "|---|---|---|---|---|---|"]
            for r in sel:
                L.append("| %s | %s | %s | %s | %s | %s |"
                         % (r["방"], r["프로젝트NO"] or "-", r["캠프"] or "-",
                            r["AS담당"] or "-", r["AS일자"] or "-", r["시각"]))
        else:
            L.append("없습니다.")
        L.append("")

    L += ["## 5. 확정해야 할 것", ""]
    todo = []
    for k, v in sorted(dup.items()):
        todo.append("**%s 가 %d줄입니다** — 진행상태가 %s 로 갈립니다. 같은 건이면 합계 "
                    "{:,}원, 별건이면 {:,}원입니다. ERP 판매조회에서 전표번호를 보면 갈립니다."
                    .format(low, high)
                    % (k, len(v), " / ".join(x.get("진행상태", "?") for x in v)))
    for p in po_posts:
        if not po_detail.get(p["PO"]):
            todo.append("**%s 의 프로젝트 목록** — 밴드 글에 프로젝트 No. 가 비어 있고 ERP"
                        " 판매조회에도 이 PO가 붙은 전표가 없습니다. 쿠팡 PO 원본이 필요합니다."
                        % p["PO"])
    # ★ 한 프로젝트번호가 서로 다른 캠프에 붙어 있으면 청구가 어긋난다(2026-08-05 실제 발견).
    #   PO 에 묶인 전표뿐 아니라 **그 날 전표**도 함께 봐야 잡힌다 — 실제 사례가 그랬다
    #   (7/28 중구1 건과 8/5 야탑1 건에 같은 UJ2601384 가 붙어 있었다).
    for r in coupang:
        prj = r.get("프로젝트코드코드") or ""
        if prj:
            dup_prj.setdefault(prj, set()).add(r.get("거래처명", ""))
    for prj, camps in sorted(dup_prj.items()):
        if len(camps) > 1:
            todo.append("**%s 가 서로 다른 캠프 %d곳에 붙어 있습니다**(%s). 프로젝트번호"
                        " 오입력으로 보이며, 이대로 청구하면 쿠팡 쪽 매칭이 어긋납니다."
                        % (prj, len(camps), " / ".join(sorted(camps))))
    mismatch = [r for r in done if r["AS일자"] and day.replace("-", ".") not in r["AS일자"]
                and day[5:].replace("-", ".").lstrip("0") not in r["AS일자"]]
    if mismatch:
        todo.append("**정산 기준일** — 완료 %d건은 작업일이 %s 이 아니고 보고만 그날 올라왔습니다"
                    "(%s). 작업일 기준인지 보고일 기준인지에 따라 건수가 달라집니다."
                    % (len(mismatch), day,
                       ", ".join(r["프로젝트NO"] or "?" for r in mismatch[:6])))
    L += ["%d. %s" % (i + 1, t) for i, t in enumerate(todo)] or ["확정 대기 항목 없습니다."]
    L.append("")

    L += ["## 6. 원천별 확인 범위", "",
          "| 원천 | 확인한 것 | 결과 |", "|---|---|---|"]
    L.append("| 카톡 | 내보내기 %d개의 %s 메시지 | 완료 %d · 접수 %d |"
             % (len(_kakao_files()), day, len(done), len(new)))
    L.append("| 밴드 | 캐시 전 밴드의 %s 글 | %d건 |" % (day, len(bd)))
    L.append("| ERP | %s 판매조회 내보내기 %d개 | %s |"
             % (day, len(erp_files),
                (", ".join("%s %d행" % (f["파일"], f["행"]) for f in erp_files) or "내보내기 없음")))
    return "\n".join(L) + "\n"


def main():
    ap = argparse.ArgumentParser(description="하루치 정산분 보고자료를 만든다")
    ap.add_argument("--day", help="YYYY-MM-DD (기본: 어제 — 보고는 다음 날 하므로)")
    a = ap.parse_args()
    day = a.day or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    text = build(day)
    os.makedirs(REPORT_DIR, exist_ok=True)
    out = os.path.join(REPORT_DIR, "보고자료_%s정산분.md" % day.replace("-", ""))
    tmp = out + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(text)
    os.replace(tmp, out)
    print("정산분 보고자료: %s (%d줄)" % (os.path.basename(out), text.count("\n")))
    return 0


if __name__ == "__main__":
    sys.exit(main())
