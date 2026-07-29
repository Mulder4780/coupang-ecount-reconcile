# -*- coding: utf-8 -*-
"""
billing_fill.py — 06_거래서류청구수금에 **청구(거래명세서) 근거를 올린다**
===============================================================================
왜 필요한가
  06시트는 "실제 작업금액 ↔ 거래명세서 ↔ 세금계산서 ↔ 입금"을 대사하는 정산 원장이다.
  그런데 데이터가 67행·4,100만원뿐인데 ERP 판매(쿠팡분)는 4억을 넘는다.
  **청구 자체가 원장에 안 올라와 있어서** 미수·수금률이 어떤 숫자든 의미가 없다.
  입금 대조보다 이게 먼저다.

무엇을 근거로 채우나
  ERP **판매조회** 내보내기(`ERP_DIR/판매조회_*.xlsx`). 판매조회 = 판매전표 = 거래명세서
  발행 근거다. 그래서 실제작업(I열)이 아니라 **거래명세서 자리(N·O)** 에 들어간다.
  · I 실제작업공급가액 ← 03_현장작업실적 (수식, 건드리지 않는다)
  · O 거래명세서공급가액 ← 판매조회 공급가액합계   → P·Q 는 수식이 자동
  · N 거래명세서발행일   ← 판매조회 일자
  · S PO번호            ← 판매조회 PO번호(있을 때만)
  두 값이 어긋나면 AE(작업대비거래명세서차액) 수식이 스스로 드러낸다 — 그게 이 시트의 목적이다.

무엇을 쓰지 않나 (★ 사용자 지시: 관련 없는 자료는 DB에 넣지 않는다)
  · 판매조회에는 쿠팡 외 거래처(엘에스일렉트릭·태림포장·시스콘 등)가 섞여 있다.
    **관리대장 02/04/05시트에 프로젝트NO가 등록된 건만** 통과시킨다. 그 교집합이 곧 쿠팡 범위다.
  · 원장 상태가 완료가 아닌 건은 넣지 않고 `reports/` 에 따로 보고한다
    (ERP엔 매출이 있는데 원장은 미완료 — 그 자체가 확인이 필요한 사실이다).
  · 이미 06시트에 있는 원천업무ID는 건드리지 않는다.

행을 어떻게 고르나 (★ v259 사고 재발 방지)
  06시트는 03시트와 **반대**다. 03시트는 B열(접수ID)이 수식이라 행 배정을 엑셀이 정하므로
  빈 행에 미리 쓰면 안 됐다. 06시트는 **C열(원천업무ID)이 입력열**이고 A열(정산ID)이 C에서
  파생된다 — 행을 정하는 건 우리다. 그래서 빈 행에 순서대로 쓰는 것이 설계대로다.
  작업완료일 오름차순으로 넣어 정산ID(JS-yymm-###)가 시간순이 되게 한다.

사용
  python billing_fill.py                 # 미리보기(무엇을 몇 건 쓸지)
  python billing_fill.py --queue         # 입력 큐에 적재 (ledger_writer --apply 로 반영)
  python billing_fill.py --self-test     # 합성 검증
"""
import sys, os, re, glob, json
from datetime import datetime, date

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

SHEET = "06_거래서류청구수금"
HDR = 4
FIRST = 5
PRJ = re.compile(r"^UJ\d{7}$")
DONE = {"작업완료", "완료"}
# (시트, 상태열) — 05는 프로젝트NO가 거의 없지만 있으면 받는다
SRC = (("02_돌발AS접수", "진행상태"), ("04_정기점검", "점검상태"), ("05_신규납품설치", "진행상태"))
COL = {"원천업무ID": "C", "거래명세서발행일": "N", "거래명세서공급가액": "O", "PO번호": "S"}


def _s(v):
    return "" if v is None else str(v).strip()


# ── 순수 판정 로직 (합성 검증 대상) ────────────────────────────────
def eligible(prj, sales_keys, ledger, in06):
    """이 프로젝트를 06시트에 올릴 수 있나 → (가능여부, 사유)

    순서가 곧 규칙이다:
      1) ERP 판매조회에 매출이 있어야 청구 근거가 있다
      2) 관리대장에 등록된 프로젝트여야 쿠팡 건이다(비쿠팡 거래처 차단)
      3) 원장이 완료라고 해야 청구 대상이다
      4) 이미 06시트에 있으면 다시 넣지 않는다
    """
    if prj not in sales_keys:
        return False, "ERP 매출 없음"
    if prj not in ledger:
        return False, "원장 미등록(비쿠팡 추정)"
    wid, _sheet, st = ledger[prj]
    if st not in DONE:
        return False, f"원장 미완료({st or '상태없음'})"
    if wid in in06:
        return False, "이미 06시트에 있음"
    return True, ""


def sort_key(item):
    """작업완료일 오름차순 — 정산ID(ROW 기반)가 시간순이 되게. 날짜 없으면 맨 뒤."""
    d = item.get("일자")
    return (0, d) if isinstance(d, date) else (1, date(9999, 1, 1))


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", _s(v))
    if m:
        try:
            return date(*(int(x) for x in m.groups()))
        except ValueError:
            return None
    return None


# ── 원천 읽기 ─────────────────────────────────────────────────
def sales_rows():
    """판매조회 파일들 → {프로젝트NO: {공급가액, 일자, PO번호, 거래처, 건수}}"""
    import openpyxl
    from source_dirs import ERP_DIR
    out, files = {}, []
    for p in sorted(glob.glob(os.path.join(ERP_DIR, "판매조회*.xls*"))):
        if os.path.basename(p).startswith("~$"):
            continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            hdr, hrow = None, 0
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
                cells = [_s(c) for c in row]
                if any("프로젝트코드" in c for c in cells):
                    hdr, hrow = cells, i
                    break
            if not hdr:
                continue
            ix = {}
            for i, h in enumerate(hdr):
                for k, want in (("prj", "프로젝트코드"), ("amt", "공급가액"), ("dt", "일자"),
                                ("po", "PO번호"), ("cust", "거래처명")):
                    if want in h and k not in ix:
                        ix[k] = i
            n = 0
            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                k = _s(row[ix["prj"]]) if ix.get("prj") is not None and ix["prj"] < len(row) else ""
                if not PRJ.match(k):
                    continue
                amt = row[ix["amt"]] if ix.get("amt") is not None and ix["amt"] < len(row) else 0
                amt = int(amt) if isinstance(amt, (int, float)) else 0
                d = to_date(row[ix["dt"]]) if ix.get("dt") is not None and ix["dt"] < len(row) else None
                po = _s(row[ix["po"]]) if ix.get("po") is not None and ix["po"] < len(row) else ""
                cu = _s(row[ix["cust"]]) if ix.get("cust") is not None and ix["cust"] < len(row) else ""
                cur = out.setdefault(k, {"공급가액": 0, "일자": None, "PO번호": "", "거래처": cu, "건수": 0})
                cur["공급가액"] += amt
                cur["건수"] += 1
                if d and (cur["일자"] is None or d > cur["일자"]):   # 마지막 발행일
                    cur["일자"] = d
                if po and not cur["PO번호"]:
                    cur["PO번호"] = po
                n += 1
            if n:
                files.append((os.path.basename(p), sn, n))
        wb.close()
    return out, files


def ledger_map(path):
    """관리대장 02/04/05 → {프로젝트NO: (원천업무ID, 시트, 상태)}"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sn, stn in SRC:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdr = next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))
        H = {_s(h): i for i, h in enumerate(hdr) if h is not None}
        pc, sc = H.get("프로젝트NO"), H.get(stn)
        if pc is None:
            continue
        for row in ws.iter_rows(min_row=FIRST, values_only=True):
            wid = _s(row[0]) if row else ""
            prj = _s(row[pc]) if pc < len(row) else ""
            st = _s(row[sc]) if sc is not None and sc < len(row) else ""
            if wid and PRJ.match(prj):
                out.setdefault(prj, (wid, sn, st))
    wb.close()
    return out


def sheet06_state(path):
    """(이미 있는 원천업무ID, 비어있는 입력행 목록)

    빈 행 = C열이 비어 있고 **A열에 정산ID 수식이 깔려 있는** 행.
    수식이 없는 행에 쓰면 정산ID·금액·차액이 따라오지 않아 반쪽짜리가 된다."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[SHEET]
    have, empty = set(), []
    for r in range(FIRST, ws.max_row + 1):
        c = ws.cell(row=r, column=3).value
        a = ws.cell(row=r, column=1).value
        if isinstance(c, str) and c.strip():
            have.add(c.strip())
        elif isinstance(a, str) and a.startswith("="):
            empty.append(r)
    wb.close()
    return have, empty


# ── 계획 ─────────────────────────────────────────────────────
def plan(path):
    sales, files = sales_rows()
    ledger = ledger_map(path)
    have, empty = sheet06_state(path)
    take, drop = [], {}
    for prj, s in sales.items():
        ok, why = eligible(prj, sales, ledger, have)
        if not ok:
            drop.setdefault(why, []).append(prj)
            continue
        wid, sn, st = ledger[prj]
        take.append({"프로젝트NO": prj, "원천업무ID": wid, "시트": sn,
                     "공급가액": s["공급가액"], "일자": s["일자"],
                     "PO번호": s["PO번호"], "거래처": s["거래처"]})
    take.sort(key=sort_key)
    return take, drop, empty, files


def build_queue(take, empty):
    """행 부족분은 잘라내고 **몇 건을 못 넣었는지 반드시 알린다**(조용한 절단 금지)."""
    n = min(len(take), len(empty))
    q = []
    for item, row in zip(take[:n], empty[:n]):
        base = {"sheet": SHEET, "evidence": f"ERP 판매조회 {item.get('프로젝트NO', '')}".strip(),
                "only_if_empty": True}
        q.append({**base, "cell": f"C{row}", "value": item["원천업무ID"], "vtype": "text"})
        if item["공급가액"]:
            q.append({**base, "cell": f"O{row}", "value": item["공급가액"], "vtype": "number"})
        if item["일자"]:
            q.append({**base, "cell": f"N{row}", "value": item["일자"], "vtype": "date"})
        if item["PO번호"]:
            q.append({**base, "cell": f"S{row}", "value": item["PO번호"], "vtype": "text"})
    return q, len(take) - n


# ── 합성 검증 ─────────────────────────────────────────────────
def self_test():
    sales = {"UJ2600001": 1, "UJ2600002": 1, "UJ2600003": 1, "UJ2600009": 1}
    ledger = {"UJ2600001": ("AS-2601-001", "02_돌발AS접수", "작업완료"),
              "UJ2600002": ("PM-2601-001", "04_정기점검", "완료"),
              "UJ2600003": ("AS-2601-003", "02_돌발AS접수", "접수"),
              "UJ2600004": ("AS-2601-004", "02_돌발AS접수", "작업완료")}
    in06 = {"AS-2601-001"}
    cases = [
        ("UJ2600001", False, "이미"),      # 중복 방지
        ("UJ2600002", True, ""),           # 정상
        ("UJ2600003", False, "미완료"),    # 원장 미완료
        ("UJ2600004", False, "ERP"),       # 판매조회에 없음
        ("UJ2600009", False, "미등록"),    # ★ 비쿠팡 — 원장에 없으면 절대 통과 못 한다
    ]
    bad = 0
    for prj, want, frag in cases:
        ok, why = eligible(prj, sales, ledger, in06)
        if ok != want or (not want and frag not in why):
            print(f"  [FAIL] {prj} → {ok} / {why}")
            bad += 1
    # 정렬: 날짜 없는 건은 맨 뒤로
    items = [{"일자": date(2026, 3, 1)}, {"일자": None}, {"일자": date(2026, 1, 5)}]
    got = [x["일자"] for x in sorted(items, key=sort_key)]
    if got != [date(2026, 1, 5), date(2026, 3, 1), None]:
        print(f"  [FAIL] 정렬 {got}"); bad += 1
    # 행 부족: 조용히 자르지 않고 남은 수를 돌려준다
    q, left = build_queue([{"원천업무ID": "AS-1", "공급가액": 100, "일자": date(2026, 1, 1), "PO번호": ""},
                           {"원천업무ID": "AS-2", "공급가액": 0, "일자": None, "PO번호": "P1"}], [72])
    if left != 1 or [x["cell"] for x in q] != ["C72", "O72", "N72"]:
        print(f"  [FAIL] 큐 {left} {[x['cell'] for x in q]}"); bad += 1
    # 실제작업(I)·정산ID(A)는 수식이므로 절대 쓰지 않는다
    if any(x["cell"][0] in ("A", "I", "P", "Q") for x in q):
        print("  [FAIL] 수식열에 쓰려 함"); bad += 1
    print("billing_fill self-test:", "OK" if not bad else f"{bad}건 실패")
    return bad == 0


def main():
    if "--self-test" in sys.argv:
        sys.exit(0 if self_test() else 1)
    from workbook_patch import latest_master
    path, ver = latest_master()
    take, drop, empty, files = plan(path)
    print(f"관리대장 v{ver} · 판매조회 파일 {len(files)}개")
    for f in files:
        print(f"  · {f[0]} [{f[1]}] {f[2]}행")
    print(f"\n06시트에 올릴 건 {len(take)}건 · 공급가액 {sum(x['공급가액'] for x in take):,}원")
    from collections import Counter
    print("  시트별:", dict(Counter(x["시트"] for x in take)))
    print(f"  빈 입력행 {len(empty)}개" + ("" if len(empty) >= len(take) else
          f"  ★ {len(take) - len(empty)}건이 남는다 — expand_rows.py 로 행을 늘려야 한다"))
    print("\n제외:")
    for why, lst in sorted(drop.items(), key=lambda x: -len(x[1])):
        print(f"  · {why}: {len(lst)}건")

    os.makedirs(os.path.join(ROOT, "reports"), exist_ok=True)
    rp = os.path.join(ROOT, "reports", "청구원장_채우기.md")
    with open(rp, "w", encoding="utf-8") as fh:
        fh.write(f"# 06_거래서류청구수금 채우기 (v{ver}, {datetime.now():%Y-%m-%d %H:%M})\n\n")
        fh.write("판매조회(ERP 판매전표) = 거래명세서 발행 근거. 실제작업(I열)은 03시트 수식이라 건드리지 않는다.\n\n")
        fh.write(f"- 올릴 건: **{len(take)}건 / {sum(x['공급가액'] for x in take):,}원**\n")
        for why, lst in sorted(drop.items(), key=lambda x: -len(x[1])):
            fh.write(f"- 제외 {why}: {len(lst)}건\n")
        mi = drop.get("원장 미등록(비쿠팡 추정)", [])
        if mi:
            fh.write(f"\n## 원장에 없는 프로젝트 {len(mi)}건 (쿠팡 건인지 확인 필요)\n\n")
            for prj in sorted(mi)[:60]:
                fh.write(f"- {prj}\n")
        nd = [w for w in drop if w.startswith("원장 미완료")]
        if nd:
            fh.write("\n## ERP엔 매출이 있는데 원장은 미완료 (상태가 밀린 것으로 보인다)\n\n")
            for w in nd:
                fh.write(f"- {w}: {', '.join(sorted(drop[w])[:30])}\n")
    print(f"\n리포트: {rp}")

    if "--queue" in sys.argv:
        import claim_guard
        claim_guard.require("ledger")
        q, left = build_queue(take, empty)
        qp = os.path.join(ROOT, "updates", "pending_updates.json")
        os.makedirs(os.path.dirname(qp), exist_ok=True)
        old = []
        if os.path.exists(qp):
            try:
                old = json.load(open(qp, encoding="utf-8"))
            except Exception:
                old = []
        for x in q:
            if isinstance(x.get("value"), date):
                x["value"] = x["value"].isoformat()
        json.dump(old + q, open(qp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"큐 적재 {len(q)}셀 ({len(q and take) and min(len(take), len(empty))}건)"
              + (f" · ★ 행이 모자라 {left}건 못 넣음" if left else ""))


if __name__ == "__main__":
    main()
