# -*- coding: utf-8 -*-
"""
recalc_pending.py — "원장엔 있는데 화면엔 아직 안 나오는 건"을 세어 앱이 말하게 한다
===============================================================================
왜 필요한가 (2026-07-29 실제로 겪은 오해)
  06_거래서류청구수금에 청구 636건(4.6억)을 넣었다. C열(원천업무ID)은 값이라 703행이 보이는데,
  정산ID·업무구분·금액·차액은 **수식**이라 엑셀이 한 번 계산해야 값이 생긴다.
  그때까지 앱은 옛 67건만 읽어 "정산 112건"이라고 말한다. 화면만 보면 **넣은 게 사라진 줄 안다.**
  숫자가 틀린 게 아니라 **아직 안 나온 것**인데, 그 사실을 아무도 말해 주지 않는 게 문제다.

  ★ 이건 도구로 못 고친다. 엑셀 수식을 계산하는 건 엑셀뿐이고(이 프로젝트는 pywin32를 쓰지
    않는다), 빈 행에 값을 미리 써넣는 건 v259에서 엉뚱한 건에 붙는 사고로 이미 겪었다.
    그래서 **고치는 대신 드러낸다** — 앱이 "N건이 대기 중, 엑셀을 한 번 열면 나옵니다"라고 말한다.

무엇을 세나
  입력열에는 값이 있는데 그 행의 ID열(수식) 캐시값이 비어 있는 행. 그게 곧 "대기 중"이다.

사용
  python recalc_pending.py            # 세어서 reports/재계산대기.json 갱신
  python recalc_pending.py --print    # 사람이 읽는 한 줄
  python recalc_pending.py --self-test
"""
import sys, os, json
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
CACHE = os.path.join(ROOT, "reports", "재계산대기.json")
FIRST = 5

# (시트, 입력열 번호, ID열 번호, 사람이 읽는 이름)
#   06: C 원천업무ID 를 넣으면 A 정산ID·금액·차액이 따라온다
#   03: 02시트가 늘면 B 접수ID 가 스스로 채워진다(입력열이 따로 없어 ID열만 본다)
#   ★ 2026-07-30: 카톡 신규 40건을 02시트에 넣었는데 여기 목록에 02가 없어서 앱이
#     아무 말도 하지 않았다. 값(B 프로젝트NO)은 들어갔지만 접수ID(A)가 수식이라
#     재계산 전에는 앱이 549행만 읽는다 — **넣은 40건이 없어 보인다.**
#     새 행을 만드는 시트는 전부 여기에 있어야 한다. 하나 빠지면 그 시트만 조용히 사라진다.
SHEETS = [("06_거래서류청구수금", 3, 1, "청구·정산"),
          ("02_돌발AS접수", 2, 1, "돌발AS 접수"),
          ("04_정기점검", 2, 1, "정기점검"),
          ("05_신규납품설치", 2, 1, "신규납품·철거"),
          ("03_현장작업실적", None, 2, "현장 작업실적")]


def _has(v):
    return v is not None and str(v).strip() != ""


def count_rows(input_vals, id_vals):
    """(입력 있음, ID 있음) 두 줄을 받아 (대기, 정상) 을 센다.

    입력열이 없는 시트(03)는 ID열만 본다 — 셀 수 있는 게 그것뿐이다."""
    wait = ready = 0
    n = max(len(input_vals or []), len(id_vals or []))
    for i in range(n):
        iv = input_vals[i] if input_vals and i < len(input_vals) else None
        dv = id_vals[i] if id_vals and i < len(id_vals) else None
        if input_vals is None:
            continue
        if _has(iv) and not _has(dv):
            wait += 1
        elif _has(iv) and _has(dv):
            ready += 1
    return wait, ready


def scan(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for name, icol, idcol, label in SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        ivals, dvals = [], []
        for r in ws.iter_rows(min_row=FIRST, values_only=True):
            ivals.append(r[icol - 1] if icol and icol - 1 < len(r) else None)
            dvals.append(r[idcol - 1] if idcol - 1 < len(r) else None)
        if icol is None:
            continue
        wait, ready = count_rows(ivals, dvals)
        if wait:
            out.append({"시트": name, "이름": label, "대기": wait, "표시중": ready})
    wb.close()
    return out


def save(items, ver):
    os.makedirs(os.path.dirname(CACHE), exist_ok=True)
    total = sum(x["대기"] for x in items)
    doc = {"갱신": datetime.now().isoformat(timespec="seconds"), "관리대장": ver,
           "대기합계": total, "항목": items,
           "안내": ("관리대장에 올라와 있지만 아직 화면에 나오지 않은 건입니다. "
                  "엑셀에서 관리대장을 한 번 열었다 닫으면 계산이 끝나고 바로 반영됩니다.")
           if total else ""}
    json.dump(doc, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    return doc


def self_test():
    bad = 0
    # 입력은 있는데 ID(수식)가 아직 없는 행 = 대기
    wait, ready = count_rows(["AS-1", "AS-2", "AS-3", None], ["JS-1", None, None, None])
    if (wait, ready) != (2, 1):
        print("  [FAIL] 기본 계수", wait, ready); bad += 1
    # 공백 문자열은 값이 아니다 — 엑셀 수식이 ""를 돌려주는 걸 '있음'으로 세면 대기가 0이 된다
    if _has("") or _has("   ") or _has(None):
        print("  [FAIL] 빈 값 판정"); bad += 1
    w2, _ = count_rows(["AS-1"], [""])
    if w2 != 1:
        print("  [FAIL] 수식이 돌려준 빈 문자열을 값으로 셌다"); bad += 1
    # 입력열이 없는 시트는 세지 않는다(추측으로 숫자를 만들지 않는다)
    if count_rows(None, ["a", "b"]) != (0, 0):
        print("  [FAIL] 입력열 없는 시트"); bad += 1
    print("recalc_pending self-test:", "OK" if not bad else f"{bad}건 실패")
    return bad == 0


def main():
    if "--self-test" in sys.argv:
        sys.exit(0 if self_test() else 1)
    from workbook_patch import latest_master
    path, ver = latest_master()
    doc = save(scan(path), ver)
    if doc["대기합계"]:
        detail = " · ".join(f"{x['이름']} {x['대기']}건" for x in doc["항목"])
        print(f"재계산 대기 {doc['대기합계']}건 ({detail}) — 엑셀을 한 번 열면 반영됩니다")
    else:
        print("재계산 대기 없음 — 원장과 화면이 일치합니다")


if __name__ == "__main__":
    main()
