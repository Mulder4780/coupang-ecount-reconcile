# -*- coding: utf-8 -*-
"""
fill_report_photo.py — 02시트 완료보고서·사진 칸을 채워 검증을 정상으로 넘긴다
================================================================================
왜 남아 있었나
  02시트에서 '완료보고서'의 근거가 '밴드에 완료 글이 있음'뿐이었는데, 그건 밴드수정 칸과
  같은 근거라 보류했다. 그 사이 03_현장작업실적에는 **기사보고내용·실제작업상세**라는
  훨씬 확실한 근거로 완료보고서를 채워 뒀다. 그걸 프로젝트NO로 물려받는다.

근거 순서 (지어내지 않는다)
  ① 03시트에 그 프로젝트의 완료보고서 판정이 있으면 → 그대로 따른다
  ② 03시트에 기사보고내용·실제작업상세가 있으면      → 등록
  ③ 밴드에 그 프로젝트의 '작업완료' 글이 있으면       → 등록 (이 회사는 밴드 게시가 보고 절차다)
  ④ 아무것도 없으면                                → 누락
  사진: 밴드 사진 장수 > 0 → 등록 / 아니면 누락

작업완료가 아닌 건은 건드리지 않는다(검증 수식도 작업완료일 때만 본다).
이미 값이 있는 칸은 덮지 않는다. 검증결과는 수식이라 손대지 않는다.

  python fill_report_photo.py            # 집계만
  python fill_report_photo.py --apply    # vN+1
"""
import os
import re
import sys
import glob
import json
import zipfile
from collections import Counter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path, esc  # noqa: E402

SHEET = "02_돌발AS접수"
HDR = 4
PRJ = re.compile(r"\b(UJ\d{7})\b")
DRAFT = "(자동 초안)"
ALLOWED = ("등록", "누락", "해당없음")


def _s(v):
    return "" if v is None else str(v).strip()


def evidence(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def rows(sn):
        ws = wb[sn]
        hdr = [_s(h) for h in next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        out = []
        for r in ws.iter_rows(min_row=HDR + 1, values_only=True):
            if r[0] in (None, ""):
                continue
            out.append({h: (r[i] if i < len(r) else None) for h, i in ix.items()})
        return out

    fw, band = rows("03_현장작업실적"), rows("24_밴드업무추출")
    wb.close()

    rep, photo = {}, {}
    for f in fw:
        k = _s(f.get("프로젝트NO"))
        if not k:
            continue
        v = _s(f.get("완료보고서"))
        if v in ALLOWED and k not in rep:
            rep[k] = v                                   # ① 03시트 판정
        if k not in rep:
            d = _s(f.get("실제작업상세"))
            if DRAFT in d:
                d = ""
            if d or _s(f.get("기사보고내용")) or _s(f.get("실제작업항목")):
                rep[k] = "등록"                          # ② 보고 내용이 있다
    doneposted = set()
    for b in band:
        k = _s(b.get("프로젝트NO"))
        if not k:
            continue
        try:
            photo[k] = max(photo.get(k, 0), int(_s(b.get("사진")) or 0))
        except ValueError:
            pass
        if "완료" in _s(b.get("진행상태")):
            doneposted.add(k)
    return rep, doneposted, photo


def plan(path):
    import openpyxl
    from openpyxl.utils import get_column_letter as GL
    rep, posted, photo = evidence(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    hdr = [_s(h) for h in next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))]
    ix = {h: i for i, h in enumerate(hdr) if h}
    C = {h: GL(ix[h] + 1) for h in ("완료보고서등록", "사진등록") if h in ix}
    fills, tally, src = {}, {}, Counter()
    for n, r in enumerate(ws.iter_rows(min_row=HDR + 1, values_only=True), start=HDR + 1):
        if r[0] in (None, ""):
            continue
        get = lambda h: _s(r[ix[h]]) if h in ix and ix[h] < len(r) else ""
        if get("진행상태") != "작업완료":
            continue
        k = get("프로젝트NO")
        if "완료보고서등록" in C and not get("완료보고서등록"):
            if k in rep:
                v, why = rep[k], "03시트 보고내용"
            elif k in posted:
                v, why = "등록", "밴드 완료 글"
            else:
                v, why = "누락", "근거 없음"
            fills[f'{C["완료보고서등록"]}{n}'] = v
            tally.setdefault("완료보고서등록", Counter())[v] += 1
            src[why] += 1
        if "사진등록" in C and not get("사진등록"):
            v = "등록" if photo.get(k, 0) > 0 else "누락"
            fills[f'{C["사진등록"]}{n}'] = v
            tally.setdefault("사진등록", Counter())[v] += 1
    wb.close()
    for c in tally.values():
        for v in c:
            assert v in ALLOWED, f"허용되지 않은 값 {v!r}"
    return fills, tally, src


def main():
    do = "--apply" in sys.argv
    m = latest_master()
    src_path = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src_path)}\n")
    fills, tally, why = plan(src_path)
    print(f"채울 칸 {len(fills)}개")
    for k, c in tally.items():
        print(f"  {k:<14}{sum(c.values()):>5}  " + " · ".join(f"{v} {x}" for x, v in c.most_common()))
    print("  근거:", dict(why))
    print("  ※ 검증결과는 수식이라 건드리지 않습니다.")
    if not do:
        print("\n실제로 채우려면:  python fill_report_photo.py --apply")
        return 0

    zin = zipfile.ZipFile(src_path)
    sp = sheet_xml_path(zin, SHEET)
    seen = set()

    def repl(mm):
        ref, attrs = mm.group(1), mm.group(2) or ""
        if ref not in fills or 't="shared"' in mm.group(0) or 't="array"' in mm.group(0):
            return mm.group(0)
        seen.add(ref)
        a = re.sub(r'\s+t="[^"]*"', "", attrs)
        return f'<c r="{ref}"{a} t="inlineStr"><is><t>{esc(fills[ref])}</t></is></c>'

    xml = re.sub(r'<c r="([A-Z]+\d+)"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)',
                 repl, zin.read(sp).decode("utf-8"), flags=re.S)
    miss = [r for r in fills if r not in seen]
    if miss:
        print(f"★ 칸이 없어 못 채운 곳 {len(miss)}개 {miss[:4]}")
    from dashboard_clean import force_recalc
    changed = {sp: xml.encode("utf-8"),
               "xl/workbook.xml": force_recalc(
                   zin.read("xl/workbook.xml").decode("utf-8")).encode("utf-8")}
    mm2 = re.search(r"_v(\d+)\.xlsx$", src_path)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm2.group(1)) + 1}.xlsx", src_path)
    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close(); zin.close()
    try:
        import openpyxl
        z = zipfile.ZipFile(tmp)
        assert z.testzip() is None
        wa = openpyxl.load_workbook(src_path, data_only=True)[SHEET]
        wb_ = openpyxl.load_workbook(tmp, data_only=True)[SHEET]
        for ref, v in fills.items():
            assert _s(wb_[ref].value) == v, f"{ref} 반영 실패"
        diff = []
        for r in range(1, max(wa.max_row, wb_.max_row) + 1):
            for c in range(1, max(wa.max_column, wb_.max_column) + 1):
                ca = wa.cell(r, c)
                if ca.coordinate in fills:
                    continue
                if ca.value != wb_.cell(r, c).value:
                    diff.append(ca.coordinate)
                    if len(diff) > 5:
                        break
        assert not diff, f"의도치 않은 셀 변경: {diff}"
        wf = openpyxl.load_workbook(tmp, data_only=False)[SHEET]
        for ref in ("AK5", "AN5"):
            assert str(wf[ref].value or "").startswith("="), f"{ref} 수식이 사라졌다"
        zs = zipfile.ZipFile(src_path)
        other = [n for n in zs.namelist() if n not in changed and zs.read(n) != z.read(n)]
        assert not other, f"의도치 않은 파트 변경: {other}"
        zs.close(); z.close()
        print("  검증 통과 — 수식 보존 · 다른 셀·파트 변경 없음")
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise
    os.replace(tmp, dst)
    print(f"\n{len(fills)}칸 채움 → {os.path.basename(dst)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
