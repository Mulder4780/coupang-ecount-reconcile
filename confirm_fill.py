# -*- coding: utf-8 -*-
"""
confirm_fill.py — 근거가 있는 건의 **날짜·확인 표시**를 한꺼번에 채운다
================================================================================
사용자 지시(2026-07-26): "확인된 접수일자 및 작업완료 처리된 건 날짜 모두 입력해서 보이게,
내가 확인 완료된 것과 같이 모두 체크해서 반영."

채우는 것 (전부 **빈칸일 때만**, 기존 값은 절대 덮지 않는다)
  · 접수일자 / 점검예정일   ← 밴드 작업일(없으면 게시일)
  · 작업완료일 / 실제점검일 ← 밴드 작업일. 상태가 완료인데 날짜만 빠진 건.
  · 사진등록                ← 밴드 글에 사진이 있으면 '등록'
  · 관리자검증상태·담당관리자·최종확인일
                            ← **작업이 완료됐고 밴드에 그 글이 실제로 있는 건만** '일치'로 체크.
                              근거 없는 건까지 체크하면 확인했다는 거짓 기록이 남는다.

밴드의 `작업완료` 글과 완료 날짜가 함께 있으면 완료 판정은 SQLite DB에 기록한다.
상태 수식 셀은 덮지 않으며 취소·철회·AS전환·점검불가와 날짜 모순은 충돌로 남긴다.

실행
  python confirm_fill.py            # 무엇을 채울지 미리보기
  python confirm_fill.py --queue    # DB 완료판정 즉시 + 근거 셀 일괄반영 대기
  python confirm_fill.py --apply    # 위 큐를 SQLite로 흡수(Excel은 다음 11:00·15:00)
"""
import sys, os
from datetime import date, datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

MANAGER = "유현민"
CONFLICT_STATES = {"취소", "철회", "AS전환", "점검불가"}


def _day(value):
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()[:10]
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        return ""


def objective_completion_day(current, band, base=None, existing_done=None, today=None):
    """모순 없는 밴드 완료 근거의 실제 완료일을 돌려준다. 아니면 빈 문자열."""
    state = str(current or "").strip()
    if (not band or str(band.get("status") or "").strip() != "작업완료"
            or state in CONFLICT_STATES):
        return ""
    # 원장에 이미 실제 완료일이 있으면 그것을 우선한다. 밴드 원문의 연도 생략이
    # 전년도처럼 파싱된 2건이 있어, 명시된 원장 날짜보다 추정 연도를 우선하면 안 된다.
    completed = _day(existing_done) or _day(band.get("date"))
    if not completed:
        return ""
    limit = _day(today) or date.today().isoformat()
    if completed > limit:
        return ""
    started = _day(base)
    if started and completed < started:
        return ""
    return completed


def objective_completion(current, done_value, band, base=None, existing_done=None, today=None):
    """합성검증·호출부 호환용 객관 완료 여부."""
    return bool(objective_completion_day(current, band, base, existing_done, today))


def band_map():
    """프로젝트NO → {날짜, 사진, 상태}

    ★ 한 건에 글이 여러 개 올라온다(접수 안내 → 작업완료). 예전에는 **먼저 나온 것 하나만**
      집었는데, 그러면 접수 글이 잡혀 완료 글을 통째로 놓친다.
      전체로 보면 '첫 기록이 완료' 382건 vs '완료 기록이 하나라도 있음' 859건 —
      477건을 미완료로 잘못 본다. 완료 확인 체크가 여기에 걸려 있어 그대로 두면
      **완료된 건을 확인 못 했다고 넘기거나, 반대로 근거를 잘못 읽는다**(2026-07-27).
      그래서 완료 글을 우선으로, 정보가 많은 쪽을 고른다(project_resolve와 같은 규칙).
    """
    import band_extract as B
    out = {}
    for r in B.load_records():
        p = (r.get("프로젝트NO") or "").strip()
        if not p:
            continue
        cur = {"date": (r.get("작업일") or r.get("게시일") or "")[:10],
               "photo": bool(r.get("사진수") or r.get("photo_count") or r.get("사진")),
               "kind": r.get("업무유형") or "",
               "status": str(r.get("진행상태") or "")}
        cur["_score"] = (3 if cur["status"] == "작업완료" else 0) + \
                        sum(1 for k in ("date", "photo", "kind") if cur[k])
        old = out.get(p)
        if not old or cur["_score"] > old["_score"]:
            out[p] = cur
    return out


def _plan():
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    bm = band_map()
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    items, stat, completions = [], {}, []

    def bump(k):
        stat[k] = stat.get(k, 0) + 1

    # 시트마다 확인 열 이름이 다르다(02: 관리자검증상태 / 04: 최종확인일(유현민 체크)).
    # 있는 열만 골라 쓰고, 없는 열은 조용히 건너뛴다.
    spec = [("02_돌발AS접수", "접수일자", "작업완료일", "진행상태", "작업완료",
             "사진등록", "관리자검증상태", "최종확인일"),
            ("04_정기점검", "점검예정일", "실제점검일", "점검상태", "완료",
             "점검사진", None, "최종확인일(유현민 체크)")]
    for sheet, dcol, donecol, statcol, donev, photocol, verifycol, datecol in spec:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        has = lambda c: c in idx
        for row_no, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            g = lambda c: (row[idx[c]] if c in idx and idx[c] < len(row) else None)
            prj = str(g("프로젝트NO") or "").strip()
            if not prj:
                continue
            b = bm.get(prj)
            # ★ ledger_writer가 보는 키는 "vtype" 이다. "type"으로 넘기면 조용히 text로 들어가
            #   날짜가 문자열이 되고 집계·정렬이 어긋난다(2026-07-26에 348개가 이렇게 들어갔다).
            add = lambda col, val, why: items.append(
                {"sheet": sheet, "key": prj, "key_col": "프로젝트NO", "col": col,
                 # ★ 계획 단계에서 빈칸만 고르지만, **쓰기 단계에서도** 한 번 더 막는다.
                 #   수식 셀은 캐시가 비어 보일 수 있어 계획만 믿으면 남의 확인 기록을 덮는다.
                 #   확인 기록을 덮는 것은 되돌릴 수 없다.
                 "value": val, "only_if_empty": True,
                 "vtype": "date" if ("일" in col and "등록" not in col) else "text",
                 "evidence": why})

            # 상태열은 수식이다. 새로 넣은 행은 엑셀을 한 번 열기 전까지 캐시값이 없어
            # 04시트는 197행이 빈칸으로 읽힌다 — 상태 문자열만 믿으면 전부 미완료로 샌다.
            # 완료 날짜가 있으면 그것으로 완료라고 본다(= 앱의 derive_status와 같은 판단).
            band_completed = b is not None and "완료" in b["status"]
            done = (str(g(statcol) or "") == donev or bool(g(donecol))
                    or band_completed)

            if b and b["date"]:
                if not g(dcol) or isinstance(g(dcol), str):
                    add(dcol, b["date"], "밴드 게시"); bump(f"{sheet} {dcol}")
                # 상태가 완료인데 날짜만 빠진 건 — 화면·집계에서 통째로 빠진다
                if has(donecol) and (not g(donecol) or isinstance(g(donecol), str)) and done:
                    add(donecol, b["date"], "밴드 게시(완료)"); bump(f"{sheet} {donecol}")
                    done = True

            # 완료 글+날짜는 현장 실시가 끝났다는 객관 근거다. 상태 수식 셀을 텍스트로
            # 덮지 않고 DB 정본에 기록한다. AS는 완료일이 접수일보다 앞서면 모순이므로 제외한다.
            completed_on = objective_completion_day(
                g(statcol), b, base=g(dcol) if sheet == "02_돌발AS접수" else None,
                existing_done=g(donecol))
            if completed_on:
                id_col = "접수ID" if sheet == "02_돌발AS접수" else "점검ID"
                kind = "as" if sheet == "02_돌발AS접수" else "pm"
                record_id = str(g(id_col) or prj).strip()
                completions.append({
                    "kind": kind, "record_id": record_id, "project": prj,
                    "status": donev, "completed_on": completed_on,
                    "basis": f"{sheet}!{row_no} 밴드 작업완료 게시+완료일",
                })
                bump(f"{sheet} DB 객관완료")

            # 확인 체크는 **완료 + 밴드 근거**가 둘 다 있을 때만
            confirmed = done and band_completed
            if not confirmed:
                continue
            if has("완료보고서등록") and not g("완료보고서등록"):
                add("완료보고서등록", "등록", "완료보고서(밴드·카톡)")
                bump(f"{sheet} 완료보고서등록")
            if photocol and has(photocol) and not g(photocol) and b["photo"]:
                add(photocol, "등록", "밴드 사진"); bump(f"{sheet} {photocol}")
            if verifycol and has(verifycol) and not g(verifycol):
                add(verifycol, "일치", "완료+밴드 근거"); bump(f"{sheet} {verifycol}")
            if has("담당관리자") and not g("담당관리자"):
                add("담당관리자", MANAGER, "확인자"); bump(f"{sheet} 담당관리자")
            if datecol and has(datecol) and (not g(datecol) or isinstance(g(datecol), str)):
                add(datecol, (g(donecol) and str(g(donecol))[:10]) or b["date"], "확인일")
                bump(f"{sheet} {datecol}")
    wb.close()
    # 동일 행이 밴드 중복 레코드로 두 번 보이더라도 DB upsert 전 목록부터 한 건으로 만든다.
    unique = {(r["kind"], r["record_id"]): r for r in completions}
    return master, items, stat, list(unique.values())


def plan():
    master, items, stat, _completions = _plan()
    return master, items, stat


def resolution_plan():
    master, _items, _stat, completions = _plan()
    return master, completions


def main():
    from operation_window import input_window_label, is_input_window
    if is_input_window():
        print(f"입력 보호시간({input_window_label()}) — 완료보고서 반영 생략")
        return
    do_apply = "--apply" in sys.argv
    do_queue = "--queue" in sys.argv
    if do_apply:
        # 실제 원장 쓰기만 배타 점유를 요구한다. --queue는 기존 원자적 대기열에만 추가한다.
        from claim_guard import require
        require("ledger", "confirm_fill")
    master, items, stat, completions = _plan()
    if not items and not completions:
        print("채울 항목·객관 완료 판정 없음")
        return
    print(f"채울 셀 {len(items)}개 · DB 객관 완료 {len(completions)}건")
    for k, v in sorted(stat.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}건")
    if not do_apply and not do_queue:
        print("\n큐에 넣기: python confirm_fill.py --queue")
        print("DB에 넣기: python confirm_fill.py --apply (Excel은 다음 11:00·15:00)")
        return
    import ledger_db
    resolved = ledger_db.work_resolution_sync(completions)
    print(f"DB 객관 완료 동기화 {resolved}건")
    import ledger_writer as L
    n = L.queue_add(items) if items else 0
    print(f"큐 추가 {n}건")
    if do_queue:
        return
    print("DB 흡수:", ledger_db.intake_json(source="confirm_fill"))
    st = ledger_db.status()
    print(f"Excel 반영 대기 {st['대기']}건 · 다음 {st['다음반영']}")


if __name__ == "__main__":
    main()
