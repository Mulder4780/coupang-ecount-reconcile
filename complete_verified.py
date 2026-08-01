# -*- coding: utf-8 -*-
"""검증·매칭과 실제 완료 근거가 모두 있는 2026년 행을 DB 완료 정본에 기록한다.

상태값만 보고 완료시키지 않는다. 업무별 필수 완료일과 검증결과가 함께 있을 때,
상태 수식 셀을 덮지 않고 ledger_db.work_resolution에 근거와 최초 확인시각을 남긴다.
"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def _s(value):
    return "" if value is None else str(value).strip()


def _year(value):
    if isinstance(value, (date, datetime)):
        return value.year
    text = _s(value)
    return int(text[:4]) if len(text) >= 4 and text[:4].isdigit() else None


def _as_completion_ready(get):
    """작업완료 입력 후에도 02시트 검증문제코드가 생기지 않는지 선검사한다."""
    if _year(get("작업완료일")) != 2026:
        return False
    if _s(get("관리자검증상태")) != "일치":
        return False
    if _s(get("완료보고서등록")) != "등록":
        return False
    if _s(get("사진등록")) in ("", "누락"):
        return False
    if _s(get("동영상등록")) == "누락":
        return False
    if _s(get("비용구분")) in ("", "미확정"):
        return False
    if _s(get("ERP등록")) in ("", "미등록"):
        return False
    if not _s(get("재방문여부")):
        return False
    if _s(get("최초접수외추가작업")) == "있음" and _s(get("추가작업확인상태")) == "미반영":
        return False
    return True


def plan(path: str):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items, evidence = [], []

    def rows(sheet):
        ws = wb[sheet]
        header = list(next(ws.iter_rows(min_row=4, max_row=4, values_only=True)))
        index = {str(value).strip(): i for i, value in enumerate(header) if value is not None}
        for row_no, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            get = lambda name: row[index[name]] if name in index and index[name] < len(row) else None
            yield row_no, get

    # 돌발AS: 완료일 + 관리자 실제작업 일치 + 검증결과 정상 + 완료보고서 등록.
    for row_no, get in rows("02_돌발AS접수"):
        project = _s(get("프로젝트NO"))
        if not project or _year(get("접수일자")) != 2026:
            continue
        if _s(get("진행상태")) in ("취소", "철회"):
            continue
        if _s(get("검증결과")) == "정상" and _as_completion_ready(get):
            items.append({
                "kind": "as", "record_id": _s(get("접수ID")) or project,
                "project": project, "status": "작업완료",
                "completed_on": _s(get("작업완료일"))[:10],
                "basis": (
                    f"02!R{row_no} 작업완료일 + AD{row_no} 관리자검증 일치 + "
                    f"W/Y/Z/U/S{row_no} 완료 필수근거 + AK{row_no} 검증 정상"
                ),
            })
            evidence.append(("02_돌발AS접수", row_no, project, "작업완료"))

    # 정기점검: 실제점검일 + 검증 정상. 취소·AS전환 같은 충돌 상태는 제외.
    for row_no, get in rows("04_정기점검"):
        project = _s(get("프로젝트NO"))
        if (not project or _year(get("점검예정일")) != 2026
                or _s(get("점검상태")) in ("AS전환", "점검불가", "취소", "철회")):
            continue
        if _year(get("실제점검일")) == 2026 and _s(get("검증결과")) == "정상":
            items.append({
                "kind": "pm", "record_id": _s(get("점검ID")) or project,
                "project": project, "status": "완료",
                "completed_on": _s(get("실제점검일"))[:10],
                "basis": f"04!H{row_no} 실제점검일 + AB{row_no} 검증 정상",
            })
            evidence.append(("04_정기점검", row_no, project, "완료"))

    # 신규·납품·설치: 업무구분에 맞는 실제 완료일이 있어야 한다.
    for row_no, get in rows("05_신규납품설치"):
        project = _s(get("프로젝트NO"))
        if (not project or _year(get("요청일")) != 2026
                or _s(get("진행상태")) in ("취소", "철회")):
            continue
        kind = _s(get("업무구분"))
        actual = (
            get("철거·이전완료일")
            if ("철거" in kind or "이전" in kind)
            else get("실제설치일")
            if "설치" in kind
            else get("실제납품일")
        )
        if _year(actual) == 2026 and _s(get("검증결과")) == "정상":
            items.append({
                "kind": "install", "record_id": _s(get("업무ID")) or project,
                "project": project, "status": "완료",
                "completed_on": _s(actual)[:10],
                "basis": f"05!실제완료일({row_no}) + AG{row_no} 검증 정상",
            })
            evidence.append(("05_신규납품설치", row_no, project, "완료"))

    wb.close()
    return items, evidence


def main():
    from workbook_patch import latest_master

    path = latest_master()[0]
    items, evidence = plan(path)
    print(f"원본: {os.path.basename(path)}")
    print(f"완료 보완 대상: {len(items)}건")
    for sheet, row, project, status in evidence:
        print(f"  {sheet}!{row} {project} → {status}")
    if not items:
        print("새로 DB 완료 판정할 행 없음")
        return
    do_queue = "--queue" in sys.argv
    do_apply = "--apply" in sys.argv
    if not do_queue and not do_apply:
        print("DB에 기록하려면: python complete_verified.py --queue")
        return
    if do_apply:
        from claim_guard import require
        require("ledger", "complete_verified")
    import ledger_db
    added = ledger_db.work_resolution_sync(items)
    print(f"DB 객관 완료 동기화: {added}건")


if __name__ == "__main__":
    main()
