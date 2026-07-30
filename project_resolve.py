# -*- coding: utf-8 -*-
"""
project_resolve.py — 프로젝트 코드 하나로 나머지를 전부 채운다
================================================================================
류지영 매니저가 새 건을 넣을 때 손으로 적는 건 **프로젝트NO 하나뿐**이어야 한다.
접수ID·점검ID·캠프명·업무유형·담당기사·일자는 이미 밴드/카톡/원장 어딘가에 있으므로
사람이 다시 칠 이유가 없다. 이 모듈이 그 조회와 채번을 담당한다.

  UJ2601138  →  02_돌발AS접수 531행
                접수ID   AS-2606-527   (채번: 접두어-작업월-행번호)
                캠프명   창원1MB(팔용동)  (밴드 본문)
                담당기사 김필우          (밴드 본문)
                접수일자 2026-06-03     (밴드 작업일)

증거 우선순위 — 위에 있을수록 강하다
  ① 관리대장 기존 행   이미 등록된 코드면 그 행이 진실이다. 새로 만들지 않고 그 행을 알려준다.
  ② 24_밴드업무추출    구조화된 밴드 추출분(원장 코드의 99%를 덮는다)
  ③ 밴드·카톡 원문     24시트에 아직 안 들어온 신규 건의 캠프명(camp_fill.camp_book 재사용)

채번 규칙은 **시트 수식과 똑같다**(접두어-yymm-행번호). 규칙이 갈리면 사람이 연 순간
엑셀이 다시 계산해 번호가 어긋나므로, 여기서 만든 값과 수식 결과는 반드시 일치해야 한다.

  python project_resolve.py UJ2601138            # 한 건 조회
  python project_resolve.py UJ2601138 --apply    # 원장에 새 행으로 등록(빈 칸만·vN+1)
  python project_resolve.py --export             # 폰·태블릿용 색인 JSON 게시
  python project_resolve.py --audit              # 전 코드 자동채움 가능률 점검
"""
import sys, os, re, json
from datetime import datetime, date

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

CODE_RE = re.compile(r"UJ\d{7}")
HDR_ROW = 4

# 업무유형 → 어느 시트로 갈 건인가. 02와 04에 같은 코드가 동시에 나온 적은 한 번도 없다
# (v171 기준 0건) — 코드 하나는 돌발AS이거나 정기점검이거나 둘 중 하나다.
KIND_SHEET = {
    "돌발AS": "02_돌발AS접수", "돌발AS(동시진행)": "02_돌발AS접수",
    "정기점검": "04_정기점검", "정기점검(동시진행)": "04_정기점검",
    "신규납품설치": "05_신규납품설치",
}

# 시트별: (ID열, 접두어, 채번에 쓰는 날짜열, 프로젝트NO열)
SHEET_ID = {
    "02_돌발AS접수":       ("접수ID", "AS", "접수일자",   "프로젝트NO"),
    "03_현장작업실적":     ("작업ID", "FW", "작업일자",   "프로젝트NO"),
    "04_정기점검":         ("점검ID", "PM", "점검예정일", "프로젝트NO"),
    "05_신규납품설치":     ("업무ID", "NS", "요청일",     "프로젝트NO"),
    "06_거래서류청구수금": ("정산ID", "JS", "작업완료일", "프로젝트NO"),
    "13_PO발주관리":       ("PO관리ID", "PO", "PO요청일", "프로젝트NO"),
}

# 새 행에 채워 넣을 열 — 시트마다 이름이 다르다. 값은 resolve()가 만든 키.
ROW_MAP = {
    "02_돌발AS접수": {"프로젝트NO": "code", "캠프명": "camp", "접수일자": "date",
                      "담당기사": "tech", "유상·무상·보험": "cost", "진행상태": "status",
                      "작업완료일": "done", "신청내용": "note", "비고": "src"},
    "04_정기점검":   {"프로젝트NO": "code", "캠프명": "camp", "점검예정일": "date",
                      "담당기사": "tech", "유상·무상·보험": "cost", "실제점검일": "done",
                      "비고": "src"},
    "05_신규납품설치": {"프로젝트NO": "code", "캠프명": "camp", "업무구분": "kind",
                        "요청일": "date", "비고": "src"},
}
DATE_COLS = {"접수일자", "작업완료일", "점검예정일", "실제점검일", "요청일", "작업일자"}


def norm(code):
    """'uj 2601138' · 'UJ2601138 ' · '2601138' → 'UJ2601138'. 아니면 None."""
    s = re.sub(r"\s|-", "", str(code or "")).upper()
    if re.fullmatch(r"\d{7}", s):
        s = "UJ" + s
    m = CODE_RE.fullmatch(s)
    return m.group(0) if m else None


def _s(v):
    return str(v).strip() if v not in (None, "") else ""


def clean_tech(v):
    """담당기사 칸에 이름이 아닌 게 들어오는 걸 막는다.
    밴드 본문 파싱이 가끔 '000 (캠프상태확인 및 스케쥴 세팅)'·'자)' 같은 조각을 물어 온다.
    '김준형, 김필우'처럼 **두 사람이 함께 간 건은 정상**이므로 쉼표는 살린다."""
    t = _s(v)
    if not t or len(t) > 14:
        return ""
    if re.search(r"[()\[\]:：/]|\d", t):        # 괄호·숫자가 섞이면 이름이 아니다
        return ""
    return t if re.fullmatch(r"[가-힣A-Za-z]{2,4}(\s*[,.]\s*[가-힣A-Za-z]{2,4})*", t) else ""


def _d(v):
    """엑셀 날짜·문자열 → 'YYYY-MM-DD'"""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    m = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", str(v or ""))
    return "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3))) if m else ""


# ───────────────────────── 증거 수집 ─────────────────────────
def evidence(master=None):
    """모든 증거를 한 번에 읽어 프로젝트 코드별로 접어 둔다(원장은 한 번만 연다)."""
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = master or resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)

    def sheet_rows(sh, hr=HDR_ROW):
        if sh not in wb.sheetnames:
            return
        ws = wb[sh]
        hdr = [_s(h) for h in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        for n, r in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), start=hr + 1):
            yield n, {h: (r[i] if i < len(r) else None) for h, i in ix.items()}, ix

    # ① 관리대장 기존 행 — 코드가 어느 시트 몇 행에 있고 어떤 ID를 받았는지
    ledger, cols, tail = {}, {}, {}
    for sh, (idc, _pfx, _dc, keyc) in SHEET_ID.items():
        last = HDR_ROW
        for n, r, ix in sheet_rows(sh):
            cols[sh] = ix
            code = norm(r.get(keyc))
            if not code:
                continue
            last = n
            e = ledger.setdefault(code, {"sheets": {}, "ids": {}})
            e["sheets"][sh] = n
            if _s(r.get(idc)):
                e["ids"][idc] = _s(r.get(idc))
            if sh in ("02_돌발AS접수", "04_정기점검", "05_신규납품설치"):
                e.setdefault("camp", _s(r.get("캠프명")))
                e.setdefault("tech", clean_tech(r.get("담당기사")))
                e.setdefault("date", _d(r.get(SHEET_ID[sh][2])))
        tail[sh] = last          # 데이터가 있는 마지막 행 → 다음 행이 새 행 자리

    # 시트 용량(수식이 깔린 마지막 행). 넘겨서 쓰면 채번 수식이 없는 맨땅에 값만 남는다.
    cap = {sh: wb[sh].max_row for sh in SHEET_ID if sh in wb.sheetnames}

    # ② 24_밴드업무추출 — 가장 넓은 증거. 같은 코드가 여러 번 올라오면 정보가 많은 쪽을 쓴다.
    band = {}
    for _n, r, _ix in sheet_rows("24_밴드업무추출"):
        code = norm(r.get("프로젝트NO"))
        if not code:
            continue
        cur = {"camp": _s(r.get("캠프명")), "kind": _s(r.get("업무유형")),
               "cost": _s(r.get("비용구분")), "tech": clean_tech(r.get("담당기사")),
               "date": _d(r.get("작업일")), "status": _s(r.get("진행상태")),
               "posted": _d(r.get("게시일"))}
        old = band.get(code)
        # '작업완료' 글이 '접수·예정' 글보다 정확하다(일자·기사가 확정된 뒤 올라온다)
        score = sum(1 for v in cur.values() if v) + (3 if cur["status"] == "작업완료" else 0)
        if not old or score > old["_score"]:
            cur["_score"] = score
            band[code] = cur

    wb.close()

    # ③ 밴드·카톡 원문 — 24시트에 아직 안 잡힌 신규 건의 캠프명
    try:
        from camp_fill import camp_book
        book = camp_book()
    except Exception:
        book = {}

    return {"master": master, "ledger": ledger, "band": band, "book": book,
            "tail": tail, "cap": cap, "cols": cols}


# ───────────────────────── 채번 ─────────────────────────
def mint(prefix, when, rown):
    """시트 수식과 **완전히 동일**: 접두어-yymm-행번호(3자리).
       수식: =IF($B{r}="","","AS-"&TEXT(IF($D{r}="",TODAY(),$D{r}),"yymm")&"-"&TEXT(ROW()-4,"000")
       → 날짜가 없으면 오늘로 계산되는 것까지 똑같이 맞춘다. 어긋나면 사람이 파일을 연
         순간 엑셀이 다시 계산해 ID가 바뀌고, 03·06이 값으로 들고 있는 참조가 끊긴다."""
    m = re.match(r"(\d{4})-(\d{2})-", when or "")
    yymm = (m.group(1)[2:] + m.group(2)) if m else datetime.now().strftime("%y%m")
    return f"{prefix}-{yymm}-{rown - HDR_ROW:03d}"


# ───────────────────────── 리졸브 ─────────────────────────
def resolve(code, ev):
    """프로젝트 코드 하나 → 시트·행·ID·필드 전부. 출처를 필드마다 같이 돌려준다."""
    code = norm(code)
    if not code:
        return {"ok": False, "reason": "프로젝트 코드 형식이 아닙니다 (UJ + 숫자 7자리)"}

    out = {"ok": True, "code": code, "src": {}, "ids": {}, "unknown": []}
    b = ev["band"].get(code, {})
    led = ev["ledger"].get(code)

    def put(key, value, source):
        if value and not out.get(key):
            out[key] = value
            out["src"][key] = source

    # ① 원장이 최우선 — 이미 등록된 건이면 그 값이 진실이다
    if led:
        for k in ("camp", "tech", "date"):
            put(k, led.get(k), "관리대장")
        out["ids"] = dict(led["ids"])
        out["sheets"] = led["sheets"]
    # ② 밴드 추출
    for k, lab in (("camp", "밴드 캠프이름"), ("kind", "밴드 업무유형"), ("cost", "밴드 비용구분"),
                   ("tech", "밴드 담당기사"), ("date", "밴드 작업일"), ("status", "밴드 진행상태")):
        put(k, b.get(k), lab)
    # ③ 밴드·카톡 원문(캠프명 보조)
    put("camp", ev["book"].get(code), "밴드·카톡 본문")
    # 작업일이 없으면 게시일이라도 — 채번의 연월은 있어야 한다
    if not out.get("date"):
        put("date", b.get("posted"), "밴드 게시일(작업일 미상)")

    out["state"] = "등록됨" if led else "신규"

    # 어느 시트로 가는 건인가
    kind = out.get("kind") or ""
    sheet = None
    if led:
        for sh in ("02_돌발AS접수", "04_정기점검", "05_신규납품설치"):
            if sh in led["sheets"]:
                sheet = sh
                break
    if not sheet:
        sheet = KIND_SHEET.get(kind)
    out["sheet"] = sheet

    if led:
        out["row"] = led["sheets"].get(sheet)
        return out

    # ── 신규 건: 들어갈 자리와 받을 ID를 미리 계산해 보여 준다 ──
    if not sheet:
        out["ok"] = False
        out["reason"] = (f"업무유형을 알 수 없어 어느 시트인지 정하지 못했습니다"
                         f"{' (밴드 업무유형: ' + kind + ')' if kind else ' — 밴드에 이 코드가 없습니다'}")
        return out
    row = ev["tail"].get(sheet, HDR_ROW) + 1
    if row > ev["cap"].get(sheet, 0):
        out["ok"] = False
        out["reason"] = (f"{sheet} 빈 행이 없습니다 — "
                         f"python expand_rows.py --sheet {sheet} --add 50 --apply 로 늘린 뒤 다시 실행")
        return out
    out["row"] = row
    idc, pfx, _dc, _kc = SHEET_ID[sheet]
    out["ids"][idc] = mint(pfx, out.get("date"), row)
    out["src"][idc] = f"채번(시트 수식과 동일: {pfx}-작업월-{row}행)"

    # 채울 수 없는 칸을 숨기지 않는다 — 사람이 무엇을 손으로 넣어야 하는지 알아야 한다
    for k, lab in (("camp", "캠프명"), ("date", "일자"), ("tech", "담당기사"), ("cost", "비용구분")):
        if not out.get(k):
            out["unknown"].append(lab)
    return out


def row_items(res, ev):
    """신규 행을 원장에 쓰기 위한 ledger_writer 큐 항목. 좌표 지정 모드(cell)를 쓴다 —
       아직 그 행에 프로젝트NO가 없어서 키 조회로는 행을 찾을 수 없다."""
    from backfill_rows import col_letter
    sheet, row = res["sheet"], res["row"]
    cols = ev["cols"].get(sheet, {})
    vals = dict(res)
    vals["src"] = "프로젝트코드 자동채움 " + datetime.now().strftime("%Y-%m-%d")
    vals["note"] = (res.get("kind") or "").strip()
    vals["done"] = res.get("date") if res.get("status") == "작업완료" else ""
    vals["status"] = {"작업완료": "작업완료", "취소": "취소"}.get(res.get("status"), "접수")

    items = []
    for led_col, key in ROW_MAP.get(sheet, {}).items():
        if led_col not in cols:
            continue
        v = vals.get(key)
        if v in (None, ""):
            continue
        items.append({"sheet": sheet, "cell": f"{col_letter(cols[led_col] + 1)}{row}",
                      "key": f"{col_letter(cols[led_col] + 1)}{row}", "key_col": "-",
                      "col": led_col, "value": v,
                      "vtype": "date" if led_col in DATE_COLS else "text",
                      "only_if_empty": True,
                      "evidence": f"프로젝트코드 {res['code']} 자동채움 · "
                                  f"{res['src'].get(key, '자동')}"})
    return items


# ───────────────────────── 색인 게시(폰·태블릿용) ─────────────────────────
def export(ev, path=None):
    """프로젝트 코드 색인을 정적 JSON으로 뽑는다.
       고정 주소(GitHub Pages)에 올려 두면 **사무실 PC가 꺼져 있어도** 폰·태블릿에서
       코드 조회와 자동채움 미리보기가 그대로 된다. 원장 기록만 PC가 필요하다."""
    path = path or os.path.join(ROOT, "docs", "resolve_index.json")
    codes = sorted(set(ev["band"]) | set(ev["ledger"]) | set(ev["book"]))
    out = {}
    for c in codes:
        r = resolve(c, ev)
        if not r.get("ok") and not r.get("camp"):
            continue
        rec = {k: r.get(k) for k in ("camp", "kind", "cost", "tech", "date", "status",
                                     "sheet", "row", "state") if r.get(k)}
        if r.get("ids"):
            rec["ids"] = r["ids"]
        if r.get("unknown"):
            rec["unknown"] = r["unknown"]
        if not r.get("ok"):
            rec["reason"] = r.get("reason", "")
        out[c] = rec
    doc = {"generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
           "count": len(out),
           "tail": ev["tail"], "cap": ev["cap"],
           "codes": out}
    os.makedirs(os.path.dirname(path), exist_ok=True)
    json.dump(doc, open(path, "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
    return path, len(out)


# ───────────────────────── CLI ─────────────────────────
def show(r):
    if not r.get("ok"):
        print(f"  ✗ {r.get('code', '')} — {r.get('reason')}")
        return
    tag = "이미 등록됨" if r["state"] == "등록됨" else "신규 등록 대상"
    print(f"\n  {r['code']}  [{tag}]  {r.get('sheet') or '시트 미정'} {r.get('row') or ''}행")
    for k, v in r.get("ids", {}).items():
        print(f"    {k:<8} {v:<16} {r['src'].get(k, '관리대장')}")
    for key, lab in (("camp", "캠프명"), ("kind", "업무유형"), ("date", "일자"),
                     ("tech", "담당기사"), ("cost", "비용구분"), ("status", "진행상태")):
        if r.get(key):
            print(f"    {lab:<8} {str(r[key]):<16} {r['src'].get(key, '')}")
    if r.get("unknown"):
        print(f"    ! 못 채운 칸: {', '.join(r['unknown'])} — 손으로 넣어야 합니다")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    ev = evidence()

    if "--export" in flags:
        p, n = export(ev)
        print(f"색인 게시: {n}개 코드 → {p}")
        print("  git push 하면 고정 주소에서 폰·태블릿이 바로 씁니다(PC 꺼져 있어도 조회 가능).")
        return

    if "--audit" in flags:
        codes = sorted(set(ev["band"]) | set(ev["ledger"]))
        res = [resolve(c, ev) for c in codes]
        new = [r for r in res if r.get("state") == "신규"]
        full = [r for r in new if r.get("ok") and not r.get("unknown")]
        part = [r for r in new if r.get("ok") and r.get("unknown")]
        bad = [r for r in new if not r.get("ok")]
        print(f"전체 코드 {len(res)} · 원장 등록됨 {len(res)-len(new)} · 미등록 {len(new)}")
        print(f"  미등록 중 → 전부 자동 {len(full)} · 일부 수동 {len(part)} · 시트 미정 {len(bad)}")
        from collections import Counter
        miss = Counter(u for r in part for u in r["unknown"])
        if miss:
            print("  손이 필요한 칸:", dict(miss))
        for r in bad[:3]:
            print("  [시트 미정]", r["code"], "-", r.get("reason", "")[:60])
        return

    if not args:
        print(__doc__.strip().split("\n\n")[0])
        print("\n  python project_resolve.py UJ2601138 [--apply]")
        print("  python project_resolve.py --export | --audit")
        return

    results = [resolve(c, ev) for c in args]
    for r in results:
        show(r)

    todo = [r for r in results if r.get("ok") and r["state"] == "신규"]
    if "--apply" not in flags:
        if todo:
            print(f"\n미리보기 — 실제 등록: python project_resolve.py {' '.join(args)} --apply")
        return
    if not todo:
        print("\n등록할 신규 건 없음")
        return

    items = []
    for r in todo:
        items += row_items(r, ev)
    import ledger_writer as L
    import ledger_db
    print(f"\n큐 추가 {L.queue_add(items)}개 셀 ({len(todo)}행)")
    print("DB 흡수:", ledger_db.intake_json(source="project_resolve"))
    st = ledger_db.status()
    print(f"Excel 반영 대기 {st['대기']}건 · 다음 {st['다음반영']}")


if __name__ == "__main__":
    main()
