# -*- coding: utf-8 -*-
"""
findings_export.py — 확인필요 사항 통합 엑셀 생성기
=====================================================
모든 대조 결과(정산 조치필요·밴드·카톡·ERP원장·쿠팡PO)를 한 권의 엑셀로 모아
관리대장 폴더에 『쿠팡_확인필요현황_최신.xlsx』로 저장한다(항상 덮어쓰기 = 항상 최신).
관리대장 본체는 건드리지 않는다. 매일 에이전트(daily_run)가 자동 실행.

시트: 요약 / 정산_조치필요 / 밴드_미확인 / 카톡_미확인 / ERP원장_문제 / 쿠팡PO_문제
"""
import sys, os, csv, glob
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from ecount_reconcile import (read_ledger, load_config, resolve_master, settle_status,
                              erp_progress_statuses)
from responsibility import confirmed_owner

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.path.join(BASE_DIR, "reports")
OUT_NAME = "쿠팡_확인필요현황_최신.xlsx"


def latest_csv(pat):
    fs = sorted(glob.glob(os.path.join(REPORT_DIR, pat)))
    if not fs:
        return []
    return list(csv.DictReader(open(fs[-1], encoding="utf-8-sig")))


def settle_issues(master):
    rows, resolved, retracted = [], [], []
    try:
        import ledger_db
        objective_done = ledger_db.resolutions()
    except Exception:
        objective_done = {}
    prog_basis = None
    raw_progress = erp_progress_statuses()
    for sid, r in sorted(read_ledger(master).items()):
        # 판정은 ecount_reconcile.settle_status 한 곳에서만 한다 — 화면과 엑셀이 어긋나지 않게.
        st = settle_status(r)
        if st.startswith("완료("):
            # 객관 입증 완료(사용자 지시 2026-07-31) — 조치 목록에서 빼고 **DB에만** 기록한다.
            # 엑셀 셀 백필은 하지 않는다(판매조회에 발행일이 없다 — 절대규칙 10).
            resolved.append({"settle_id": sid, "project": r.get("프로젝트NO"),
                             "status": st,
                             "basis": ("ERP 판매조회 프로젝트번호 직접 일치·동일 프로젝트 "
                                       "전체 전표 진행상태(" + st[3:-1] + ")")})
            continue
        db_done = objective_done.get(sid) or {}
        if str(db_done.get("status") or "").startswith("완료("):
            # Excel 수식·발행일은 그대로 두되 객관근거 완료 DB가 조치 목록의 정본이다.
            continue
        project = str(r.get("프로젝트NO") or "").strip()
        raw = set(raw_progress.get(project, ()))
        if "7.수금완료" in raw and raw != {"7.수금완료"}:
            # 원천이 사라졌다는 이유로 과거 완료를 지우지는 않는다. 현재 ERP에서 같은
            # 프로젝트의 완료·미완료 전표가 **동시에** 보이는 명시적 충돌만 철회한다.
            retracted.append(sid)
        if st in ("무상/보험", "정상"):
            continue
        rows.append({"정산ID": sid, "문제유형": st, "캠프명": r.get("캠프명"),
                     "프로젝트NO": r.get("프로젝트NO"), "공급가액": r.get("원장_공급가액") or 0,
                     "완료일": str(r.get("작업완료일") or "")[:10],
                     "명세서번호": r.get("원장_거래명세서번호") or "", "PO번호": r.get("원장_PO번호") or ""})
    if resolved:
        try:
            import ledger_db
            ledger_db.resolution_sync(resolved)
            print(f"  객관 입증 완료 {len(resolved)}건 → DB(resolution) 기록(엑셀 백필 없음)")
        except Exception as exc:                      # 기록 실패가 보고서 생성을 막지는 않는다
            print(f"  ! 완료 기록 실패: {exc}")
    if retracted:
        try:
            import ledger_db
            removed = ledger_db.resolution_retract(retracted)
            if removed:
                print(f"  명시적 ERP 상태 충돌 완료 {removed}건 → 정확한 정산ID로 철회")
        except Exception as exc:
            print(f"  ! 완료 철회 기록 실패: {exc}")
    return rows


def collect(master):
    data = {}
    data["정산_조치필요"] = settle_issues(master)
    data["밴드_미확인"] = [r for r in latest_csv("밴드대조_*.csv") if r.get("밴드게시") == "미확인"]
    data["카톡_미확인"] = [r for r in latest_csv("카톡대조_*.csv") if r.get("카톡보고") == "미확인"]
    data["ERP원장_문제"] = latest_csv("ERP원장대조_*.csv")
    data["쿠팡PO_문제"] = latest_csv("PO대조_*.csv")
    data["날짜_미상"] = dateless(master)
    data["문서_원장미등록"] = doc_unregistered(master)
    data["금액_불일치"] = amount_gap(master)
    apply_confirmed_responsibility(data)
    return data


def apply_confirmed_responsibility(data):
    """제안 범위 그대로 확정된 내부 확인 담당자를 모든 산출물에 공통 적용한다."""
    specs = {
        "정산_조치필요": ("정산", lambda r: r.get("문제유형")),
        "밴드_미확인": ("밴드", lambda _r: "밴드 게시 미확인"),
        "카톡_미확인": ("카톡", lambda _r: "카톡 보고 미확인"),
        "ERP원장_문제": ("ERP", lambda r: f"ERP {r.get('유형', '')}"),
        "쿠팡PO_문제": ("PO", lambda r: f"PO {r.get('유형', '')}"),
        "날짜_미상": ("빈칸", lambda r: f"{r.get('빈칸', '')} 비어 있음"),
        "문서_원장미등록": ("문서", lambda _r: "원장 미등록"),
        "금액_불일치": ("금액", lambda _r: "작업금액 불일치"),
    }
    for group, rows in data.items():
        category, issue_of = specs.get(group, ("", lambda _r: ""))
        for row in rows:
            row["담당자"] = confirmed_owner(
                issue_of(row), category, row.get("담당자") or row.get("담당기사"))
    return data


def amount_gap(master):
    """실제 작업금액과 거래명세서 금액이 다른 건.
    신규·납품·설치는 제외한다(사용자 지시 2026-07-26 — 별도 절차로 관리).
    명세서를 아직 안 끊은 건도 제외한다 — 그건 '미청구'로 이미 따로 잡힌다.
    ★ 작업금액이 **아직 계산되지 않은 건**도 제외한다(2026-07-31).
      실제작업합계(K열)는 `N(I)+N(J)` 수식이고 I열은 03시트를 합산하는 수식이다. 엑셀
      재계산 전에는 둘 다 비어 있어서 work=0 이 되는데, 이걸 명세서금액과 비교하면
      **명세서가 있는 전 건이 차액 -전액으로 '불일치'가 된다**(633건 거짓 경보의 정체).
      0원이 진짜 0원인지 미계산인지 수식으로는 구분할 수 없으므로, 계산되기 전에는
      비교하지 않는다 — 이 건들은 '금액 재계산 대기'로 이미 따로 세고 있다."""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    out = []
    if "06_거래서류청구수금" in wb.sheetnames:
        ws = wb["06_거래서류청구수금"]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        for row in ws.iter_rows(min_row=5, values_only=True):
            g = lambda c: (row[idx[c]] if c in idx and idx[c] < len(row) else None)
            if not g("정산ID") or "신규" in str(g("업무구분") or ""):
                continue
            work, inv = g("실제작업합계") or 0, g("거래명세서합계") or 0
            if not inv or not work or work == inv:
                continue
            out.append({"정산ID": g("정산ID"), "프로젝트NO": g("프로젝트NO"), "캠프명": g("캠프명"),
                        "업무구분": g("업무구분"), "작업금액": work, "명세서금액": inv,
                        "차액": work - inv, "명세서번호": g("거래명세서번호"),
                        "확인방법": "밴드 거래명세서 사진 또는 이카운트 전표에서 실제 청구금액을 확인해 "
                                    "작업금액·명세서금액 중 틀린 쪽을 고치세요"})
    wb.close()
    return out


def doc_unregistered(master):
    """밴드 문서 사진 OCR에서 읽힌 프로젝트NO 중 관리대장 어디에도 없는 건.
    견적·판매전표처럼 AS/점검이 아닌 업무라 02·04 백필 대상이 아니어서 자동 등록하지 않는다."""
    rows = latest_csv("밴드문서OCR_*.csv")
    if not rows:
        return []
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    known = set()
    for sh in wb.sheetnames:
        ws = wb[sh]
        try:
            hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        except StopIteration:
            continue
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        if "프로젝트NO" not in idx:
            continue
        i = idx["프로젝트NO"]
        for r in ws.iter_rows(min_row=5, values_only=True):
            v = r[i] if i < len(r) else None
            if v:
                known.add(str(v).strip())
    wb.close()
    out, seen = [], set()
    for r in rows:
        p = (r.get("프로젝트NO") or "").strip()
        if not p or p in known or p in seen:
            continue
        seen.add(p)
        out.append({"프로젝트NO": p, "발행일": r.get("발행일"), "유형": r.get("유형"),
                    "문서파일": r.get("파일"), "공급가액": r.get("공급가액"),
                    "확인방법": "밴드 해당 게시글을 열어 캠프·업무유형·금액을 확인한 뒤 "
                                "알맞은 시트(02·04·13_PO발주관리)에 등록"})
    return out


def dateless(master):
    """날짜가 비어 있는 업무 행 — 밴드·카톡 어디에도 근거가 없어 사람이 확인해야 한다.
    (밴드에 글이 있는 건은 ledger_writer 로 이미 자동으로 채워진다)"""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    out = []
    for sheet, dcol in (("02_돌발AS접수", "접수일자"), ("04_정기점검", "점검예정일")):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        for row in ws.iter_rows(min_row=5, values_only=True):
            g = lambda c: (row[idx[c]] if c in idx and idx[c] < len(row) else None)
            date, camp = g(dcol), g("캠프명")
            if not date and camp:
                out.append({"시트": sheet, "프로젝트NO": g("프로젝트NO"), "캠프명": camp,
                            "담당기사": g("담당기사"), "빈칸": dcol,
                            "확인방법": "밴드 게시글·카톡 보고에서 실제 작업일을 찾아 해당 셀에 입력"})
            # 캠프명이 없으면 어느 현장인지 알 수 없다 — 앱 카드에도 '캠프 미상'으로 뜬다.
            # 문서 발행 알림 글(판매전표·명세서)만 밴드에 있고 작업 글이 없는 건들이 여기 해당.
            elif date and not camp:
                out.append({"시트": sheet, "프로젝트NO": g("프로젝트NO"), "캠프명": "",
                            "담당기사": g("담당기사"), "빈칸": "캠프명",
                            "확인방법": "밴드에서 그 프로젝트NO로 검색해 어느 캠프 작업인지 확인 후 "
                                        "캠프명 칸에 입력(앱 카드에 '캠프 미상'으로 표시됩니다)"})
    wb.close()
    return out


def write_xlsx(data, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    NAVY, LINE = "0E1B3F", "E4E9F0"
    hd_font = Font(bold=True, color="FFFFFF", size=10)
    hd_fill = PatternFill("solid", fgColor=NAVY)

    # 요약 시트
    ws = wb.active; ws.title = "요약"
    ws["A1"] = "Coupang Service Operations System — 확인필요 현황"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"생성 {datetime.now():%Y-%m-%d %H:%M} · 에이전트 자동 생성(매일 갱신) · 관리대장 본체는 별도"
    ws["A2"].font = Font(color="777777", size=9)
    ws.append([]); ws.append(["구분", "건수", "설명"])
    for c in ws[4]:
        c.font, c.fill = hd_font, hd_fill
    desc = {"정산_조치필요": "유상 정산 중 미청구·계산서미발행·입금대기·금액미입력",
            "밴드_미확인": "작업완료인데 밴드 게시글을 찾지 못한 건",
            "카톡_미확인": "작업완료인데 카톡 보고를 찾지 못한 건",
            "ERP원장_문제": "ERP에만/원장에만/계산서X/금액차 (A~D)",
            "쿠팡PO_문제": "원장미등록 PO·오기입·금액차·연결제안 (A~D)",
            "날짜_미상": "작업일·점검일 또는 캠프명이 비어 있어 밴드·카톡에서 찾아 채워야 하는 건",
            "문서_원장미등록": "밴드 문서 사진에는 있는데 관리대장 어디에도 없는 프로젝트NO",
            "금액_불일치": "실제 작업금액과 거래명세서 금액이 다른 건(신규·납품 제외, 미청구 제외)"}
    for k, rows in data.items():
        ws.append([k.replace("_", " "), len(rows), desc.get(k, "")])
    for col, w in (("A", 22), ("B", 8), ("C", 62)):
        ws.column_dimensions[col].width = w

    # 상세 시트들
    for name, rows in data.items():
        s = wb.create_sheet(name)
        if not rows:
            s["A1"] = "해당 없음 ✅ (또는 대조 데이터 미투입 — inbox에 파일을 넣으면 채워짐)"
            continue
        cols = []
        for r in rows:
            for k in r:
                if k not in cols:
                    cols.append(k)
        s.append(cols)
        for c in s[1]:
            c.font, c.fill = hd_font, hd_fill
            c.alignment = Alignment(vertical="center")
        for r in rows:
            s.append([r.get(c, "") for c in cols])
        s.freeze_panes = "A2"
        s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
        for i, c in enumerate(cols, 1):
            width = max(10, min(46, max([len(str(c))] + [len(str(r.get(c, ""))) for r in rows[:50]]) + 2))
            s.column_dimensions[get_column_letter(i)].width = width
    wb.save(out_path)


def main():
    cfg = load_config()
    master = resolve_master(cfg["reconcile"]["master_xlsx"])
    data = collect(master)
    out = os.path.join(os.path.dirname(master), OUT_NAME)
    try:
        write_xlsx(data, out)
    except PermissionError:
        out = os.path.join(REPORT_DIR, OUT_NAME)   # 열려 있으면 reports/에 대체 저장
        write_xlsx(data, out)
    total = sum(len(v) for v in data.values())
    print("확인필요 통합:", " / ".join(f"{k} {len(v)}" for k, v in data.items()))
    print(f"총 {total}건 → {out}")


if __name__ == "__main__":
    main()
