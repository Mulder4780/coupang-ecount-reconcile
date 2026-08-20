# -*- coding: utf-8 -*-
"""
inbox_scan.py — inbox에 들어온 엑셀이 **무슨 자료인지 내용으로 판별**한다
==============================================================================
이카운트에서 내려받으면 파일명이 `8W1JR7MGB50PHOP.xlsx` 처럼 무작위다.
파일명으로 고르면(‘원장’ 포함 등) 이런 파일은 전부 무시돼 "파일이 없습니다"가 뜬다.
→ 파일을 열어 머리글을 보고 종류를 정한다. 사용자는 그냥 넣기만 하면 된다.

판별 종류
  ledger : 거래처별계정별원장  (적요 + 차변/대변)
  po     : 쿠팡 PO 목록        (PO번호 컬럼 또는 PO+숫자 값이 다수)
  sales  : 판매/세금계산서 내보내기 (공급가액 + 거래처/품목)
  unknown: 판별 실패

사용
  python inbox_scan.py                 # inbox 폴더 목록·판별 결과 출력
  from inbox_scan import pick          # pick("ledger") → 해당 종류 파일 경로 리스트
"""
import sys, os, re, glob, time

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INBOX_DIR = os.environ.get("COUPANG_INBOX_DIR") or os.path.join(BASE_DIR, "inbox")
PO_VAL = re.compile(r"\bPO[\s-]?\d{3,}\b", re.I)


def _cells(path, sheets=4, rows=30):
    """앞부분만 읽는다 — 판별에 필요한 건 머리글 근처뿐.
    read_only=True는 쓰지 않는다: 이카운트가 내보낸 파일은 <dimension>이 'A1:A1'로
    잘못 적혀 있어 read_only 모드가 1행만 읽고 멈춘다(판별 전량 실패의 원인이었다)."""
    import openpyxl
    out = []
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        for ws in wb.worksheets[:sheets]:
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i >= rows:
                    break
                out.append([("" if c is None else str(c)).strip() for c in r])
    finally:
        wb.close()
    return out


# ── 회계 원장류 네 화면 ────────────────────────────────────────
# ★ 2026-08-08 실측 사고 — **네 화면이 한 통에 들어가 있었다.**
#   예전 규칙은 "'적요' 와 차변/대변이 같은 표에 있으면 ledger" 한 줄이었다. 그런데
#   이카운트 회계 원장류는 넷 다 그 모양이다. 실측으로 `pick("ledger")` 12개 중
#   **거래처별계정별원장은 5개뿐**이고 계정별원장 5 · 분개장 1 · 현금출납장 1 이었다.
#
#   그래서 `erp_ledger_check` 가 **전 거래처 전표를 쿠팡 것으로 읽었다**:
#   전표 5,157건을 대조해 **원장 매칭 정상 0건** · 유형A(‘설치·작업 근거 확인 필요 ★’)
#   **1,856건**. 그 1,856건 대부분은 남의 회사 거래다 — 아무도 가서 확인할 수 없는
#   현장 확인 지시가 매일 09:50 회차마다 새로 찍혔다. 파일도 있고 숫자도 나오니
#   **실패한 티가 안 났다.** ‘매칭 0건’ 이 유일한 신호였는데 아무도 안 봤다.
#
#   가르는 근거는 **머리글**이다(제목줄은 사람이 바꿀 수 있고 잘리기도 한다):
#     거래처별계정별원장 E010809 : 일자-No. 적요 차변금액 대변금액 잔액
#                                  → 거래처가 **조회 조건**이라 표에는 거래처 열이 없다
#     계정별원장        E010807 : 위 + **거래처명** 열 (계정 하나, 전 거래처)
#     분개장                     : 전표번호 계정명 거래처 차변 대변 적요
#     현금출납장                 : 일자-No. **상대계정명 상대거래처명** 적요 차변 대변 잔액
#   회계거래조회(slips)는 금액이 `금액` 한 열이라 차변/대변 관문에서 이미 걸러진다(실측).
LEDGER_SCREENS = {
    "ledger":      "거래처별계정별원장",
    "ledger_acct": "계정별원장(전 거래처)",
    "journal":     "분개장",
    "cashbook":    "현금출납장",
}


def ledger_kind(rows):
    """회계 원장류 네 화면을 가른다. 원장류가 아니면 None. (순수 함수 — 합성검증 대상)"""
    for r in rows:
        names = [c for c in r if c]
        # 차변/대변이 없으면 원장류가 아니다 — 이 관문이 회계거래조회를 막는다
        if not any(("대변" in n or "차변" in n) for n in names):
            continue
        if any(("상대계정" in n or "상대거래처" in n) for n in names):
            return "cashbook"
        # 분개장 검사가 '적요' 검사보다 **먼저**여야 한다 — 분개장에도 적요·거래처가 있어서
        # 순서가 뒤집히면 분개장이 계정별원장으로 읽힌다
        if any("전표번호" in n for n in names) and any("계정명" in n for n in names):
            return "journal"
        if not any("적요" in n for n in names):
            continue
        if any("거래처" in n for n in names):
            return "ledger_acct"
        return "ledger"
    return None


# ★ 발행·청구 파이프라인 열 — 입금 원본에는 없고 대조·현황표에만 있다 (분담판 [91]).
#   여기 낱말을 늘릴 때는 **입금 정리표에 있을 수 있는 말인지** 먼저 본다 —
#   있을 수 있는 말을 넣으면 진짜 입금이 조용히 안 읽힌다([165]).
BILLING_PIPELINE_MARKS = ("명세서", "세금계산서", "청구일", "PO번호", "지급예정일")


def classify_rows(rows):
    """머리글·값 패턴으로 종류 결정 (순수 함수 — 합성검증 대상)"""
    flat = [c for r in rows for c in r if c]
    joined = " ".join(flat)
    has = lambda *ks: any(k in c for c in flat for k in ks)

    # 회계 원장류 — 네 화면이 서로 비슷하게 생겼다. ledger_kind() 가 가른다.
    k = ledger_kind(rows)
    if k:
        return k
    # 표를 못 읽었을 때의 차선: 제목줄 낱말.
    # ★ 순서가 뜻을 갖는다 — '거래처별계정별원장' 은 '계정별원장' 을 **부분문자열로 품는다**.
    #   계정별원장을 먼저 보면 거래처별이 통째로 그쪽으로 간다.
    if "거래처별계정별" in joined:
        return "ledger"
    if "현금출납" in joined:
        return "cashbook"
    if "분개장" in joined:
        return "journal"
    if "계정별원장" in joined:
        return "ledger_acct"

    # 홈택스 전자(세금)계산서 리스트 — '승인번호'가 있으면 이것 말고 없다.
    # (회계 I > 전자(세금)계산서 > 홈택스자료조회 에서 Excel 로 내려받은 것)
    if has("승인번호") and has("공급자사업자번호", "공급받는자사업자번호", "공급자상호"):
        return "hometax"

    # 회계 I 출력물 [매출(세금)계산서현황] — 일자-No. + 매출부가세/매출합계 조합이
    # 정확한 지문이다. ★ 2026-08-03 실측: 쿠팡매출청구서현황 양식에도 '내역보기' 열이
    # 있어서 아래 taxinv 휴리스틱이 먼저 삼켰다 — 이 검사가 반드시 taxinv 보다 앞서야
    # erp_docs_check 가 원본을 찾는다.
    for r in rows:
        n = [c for c in r if c]
        if (any("일자" in x and "No" in x.replace(" ", "") for x in n)
                and any("매출부가세" in x or "매출합계" in x for x in n)):
            return "tax"

    # ★ (세금)계산서진행단계 (E010849) — 사용자가 말한 **'잔량'** 이다 (2026-08-08).
    #   "아직 발행 안 한 것"의 목록이라 다른 매출 자료와 **묻으면 안 된다**. 그런데
    #   `내역보기`+`공급가액`+`부가세` 를 다 가지고 있어 바로 아래 taxinv 가 먼저
    #   삼켰다(실측 101행이 '매출세금계산서'로 들어갔다). 그러면 잔량이 밀려도
    #   erp_grab 의 밀림 보고에 **따로 안 잡힌다** — 조용한 사고다.
    #   `단계별기능` 열은 이 화면에만 있다.
    if has("전자(세금)계산서 진행단계", "전자(세금)계산서진행단계") and has("단계별기능"):
        return "taxstep"

    # 매출(세금)계산서조회(재고) — 재고 I > 영업관리 > 판매일괄회계반영.
    # 거래명세서 현황과 머리글이 겹쳐(공급가액+부가세) 예전엔 'stmt' 로 잘못 잡혔다.
    if "매출(세금)계산서조회" in joined:
        return "taxinv"
    if has("내역보기") and has("공급가액") and has("부가세"):
        return "taxinv"

    # ★ 견적서조회 — **거래명세서로 잘못 분류되고 있었다** (2026-08-08 발견).
    #   `일자-No.` + `공급가액` + `부가세` 가 있어 아래 stmt 규칙에 먼저 걸렸다.
    #   그래서 900행짜리 견적 4장이 '거래명세서' 통에 앉아 있었고, ERP 엑셀 흡수기가
    #   머리행을 못 찾아 **한 건도 못 읽었다**(그것을 신고해서 드러났다).
    #   견적만 가진 것: 진행상태(03.입찰참여·07.수주확정) · 견적○○합계 · 영업지원 담당자.
    #   이 판정은 stmt·tax 보다 **먼저** 와야 한다.
    for r in rows:
        n = [c for c in r if c]
        if any("진행상태" in x for x in n) and (
                any(x.startswith("견적") and "합계" in x for x in n)
                or any("영업지원" in x for x in n)):
            return "quote"
    if "견적서조회" in joined:
        return "quote"

    # 이카운트 매출 관련 현황 3종 (파일명이 무작위라 머리글로만 구분된다)
    for r in rows:
        n = [c for c in r if c]
        if any("일자" in x and "No" in x.replace(" ", "") for x in n) or any(x.replace(" ", "") == "일자-번호" for x in n):
            if any("매출부가세" in x or "매출합계" in x for x in n):
                return "tax"          # 매출(세금)계산서현황
            if any("부가세" in x for x in n) and any("공급가액" in x for x in n):
                return "stmt"         # 거래명세서 현황
        if any("전표번호" in x for x in n) and any("거래처명" in x for x in n):
            return "slips"            # 회계거래조회 / 회계거래현황
    if "매출(세금)계산서현황" in joined:
        return "tax"
    if "거래명세서" in joined and has("공급가액"):
        return "stmt"

    # ★ PO·청구 대조 현황표는 **입금 원본이 아니다** (2026-08-20, 분담판 [91]).
    #   실측: 오종현 'CSOS PO관련 누락 및 취소 건 현황' 이 아래 receipt 규칙에 걸려
    #   `7. 입금내역` 에 18개(사본 포함)나 앉아 있었다 — receipt 통 20개 중 18개다.
    #   ★ **오늘 뽑히는 가짜 입금은 0건이다** — 그 표의 입금일이 전부 '-' 라서일 뿐이고,
    #     한 칸만 채워지면 그 줄이 그대로 입금으로 세어진다(거래처는 'CU141' 같은 코드로).
    #     receipt_fill 은 입금일을 06시트에 **써 넣는** 길이라(쓰는 길, [170]) 돈이 안
    #     들어왔는데 들어온 것으로 보이게 된다. 0건은 '없는 것'이 아니라 '아직 안 채워진
    #     것'이다([169]).
    #   가르는 근거는 **발행·청구 파이프라인 열**이다. 입금 원본은 좁다 — 언제 얼마가
    #   누구에게서 들어왔나만 적는다(실측: 정리표 '날짜·거래처·입금액' 3열 · 은행 원본
    #   '거래일시·출금·입금·거래후 잔액'). 명세서·세금계산서·청구일을 같이 들고 있으면
    #   그것은 입금 원본이 아니라 대조표다.
    #   ★ 표시가 **둘 이상**일 때만 본다([172]). 하나로 잡으면 사람이 입금 정리표에
    #     '세금계산서 발행일' 한 열을 더하는 날 **진짜 입금이 통째로 안 읽힌다** —
    #     못 읽는 파일은 빈칸과 같다([165]). 실측 가름은 5개 대 0개로 경계가 없다.
    if (has("입금일", "수금일", "입금일자", "수금일자") and has("입금액", "수금액")
            and has("거래처", "프로젝트", "캠프")
            and sum(1 for m in BILLING_PIPELINE_MARKS if has(m)) >= 2):
        return "billing_status"

    # 입금/수금 원본 — 파일명이 무작위여도 날짜+금액 머리글 조합이면 독립 원천이다.
    if has("입금일", "수금일", "입금일자", "수금일자") and has("입금액", "수금액") \
            and has("거래처", "프로젝트", "캠프"):
        return "receipt"
    # ★ 은행에서 그대로 내려받은 **거래내역조회** (2026-08-07 김미영 대리 파일).
    #   사람이 정리한 표가 아니라 원본이라 머리글이 다르다 — '입금액'이 아니라 '입금',
    #   '입금일'이 아니라 '거래일시'. 이걸 못 알아보면 `9. 미분류` 로 가서 아무도 안 읽는다
    #   (실제로 unknown 이었다). 통장 사본이니 '거래후 잔액'까지 있어야 이것으로 본다 —
    #   '거래일시+입금'만 보면 다른 표를 잘못 물 수 있다.
    if has("거래일시") and has("입금") and has("거래후 잔액", "거래후잔액"):
        return "receipt"

    # ERP 판매조회에는 PO번호 열이 함께 있어 아래 PO 규칙이 먼저 잡으면 쿠팡 목록으로
    # (이 아래로는 내용 휴리스틱 — 파일명 지문은 classify() 맨 앞에서 이미 봤다)
    # 잘못 분류된다. 판매 진행상태+공급가액+거래처/품목 조합은 ERP 판매가 우선이다.
    if has("진행상태") and has("공급가액") and has("거래처", "거래처명", "품목", "품목명"):
        return "sales"

    # PO: 컬럼명에 PO가 있거나, PO12345 형태 값이 여러 개
    if has("PO번호", "PO No", "PO NO", "P/O", "발주번호", "Purchase Order"):
        return "po"
    if len(PO_VAL.findall(joined)) >= 3:
        return "po"

    # 판매/세금계산서 내보내기
    if has("공급가액") and has("거래처", "거래처명", "품목", "품목명"):
        return "sales"
    return "unknown"


# ★ ERP 내보내기 파일명은 **화면코드를 그대로 단다** — 내용보다 확실한 지문이다
#   (2026-08-08 실측). 매출(세금)계산서조회(E010727)·거래명세서현황·재고쪽 조회는
#   내려받은 엑셀이 **셋 다 똑같이 생겼다**: 시트 '거래명세서' · 머리글 '일자 - 번호'.
#   그래서 내용 휴리스틱으로는 못 가르고, E010727 이 통째로 `taxinv` 로 빨려 들어갔다.
#
#   조용한 사고인 이유: 화면을 **제대로 긁어 왔는데도** 신선도표의 '매출계산서조회'는
#   계속 "3일 밀림"으로 남는다. 종류가 다른 통에 들어갔으니 그 통은 영영 안 차고,
#   다음 세션이 같은 화면을 또 긁는다. 받은 사람만 받았다고 믿는다.
#
#   ★ **증거가 있는 것만 적는다.** 화면을 실제로 긁어 그 파일이 떨어지는 것을 본 것만
#     넣을 것 — 짐작으로 적으면 멀쩡한 자료를 엉뚱한 통으로 보낸다.
ERP_FILE_PREFIX = {
    "EBG006M": "sales",      # 매출(세금)계산서조회 E010727 (2026-08-08 실측으로 확인)
}


def classify(path):
    base = os.path.basename(path or "")
    for pre, kind in ERP_FILE_PREFIX.items():
        if base.startswith(pre):
            return kind
    try:
        return classify_rows(_cells(path))
    except Exception:
        return "unknown"


# ── 분류 결과 캐시 ────────────────────────────────────────────
#   ★ 2026-07-30 실측: classify() 는 파일을 **열어서** 내용으로 종류를 정한다.
#     scan() 이 '0. 원본 자료' 트리의 엑셀 전부를 열고, /api/status 가 pick() 을
#     5번(ledger·slips·po·tax·stmt) 부르므로 같은 파일을 5번 열었다. Z: 네트워크
#     드라이브라 대시보드 첫 로딩이 **280초** 걸렸다(측정값).
#     파일 내용이 바뀌지 않았으면 다시 열 이유가 없다 — (크기, 수정시각) 으로 판정한다.
_CLS_MEM = {}                                   # 프로세스 안: 경로 → (크기, mtime, 종류)
_CLS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "reports", "inbox_classify.json")
_CLS_DISK = None                                # 프로세스 간: 재시작해도 유지


def _cls_load():
    global _CLS_DISK
    if _CLS_DISK is None:
        try:
            import json as _json
            _CLS_DISK = _json.load(open(_CLS_FILE, encoding="utf-8"))
        except Exception:
            _CLS_DISK = {}
    return _CLS_DISK


def _cls_save():
    try:
        import json as _json
        os.makedirs(os.path.dirname(_CLS_FILE), exist_ok=True)
        _json.dump(_CLS_DISK, open(_CLS_FILE, "w", encoding="utf-8"), ensure_ascii=False)
    except Exception:
        pass


# ★ **규칙을 고쳐도 캐시가 옛 답을 붙들고 있었다** (2026-08-08).
#   캐시 열쇠가 (크기, 수정시각) 뿐이라, 파일이 안 바뀌면 영원히 옛 종류를 돌려준다.
#   원본 엑셀은 한 번 떨어지면 다시 안 바뀌므로 **분류 규칙을 고친 보람이 아무 데도
#   안 닿는다** — 고친 사람은 고쳤다고 믿고, 화면은 어제와 똑같다. 오류도 안 난다.
#   그래서 규칙을 손댈 때마다 이 숫자를 올린다. 열쇠에 들어가므로 전부 한 번 다시 읽는다.
RULES_VERSION = 3       # 2: 회계 원장류 네 화면 가름(ledger/ledger_acct/journal/cashbook)
                        # 3: PO·청구 대조 현황표를 입금 원본에서 가름([91])


def classify_cached(path):
    """내용이 그대로면 캐시값을 쓴다. 바뀌었으면(또는 규칙이 바뀌었으면) 다시 열어 본다."""
    try:
        st = os.stat(path)
        sig = [st.st_size, int(st.st_mtime), RULES_VERSION]
    except OSError:
        return classify(path)
    hit = _CLS_MEM.get(path)
    if hit and list(hit[:3]) == sig:
        return hit[3]
    disk = _cls_load().get(path)
    if disk and len(disk) == 4 and list(disk[:3]) == sig:
        _CLS_MEM[path] = tuple(sig) + (disk[3],)
        return disk[3]
    kind = classify(path)
    _CLS_MEM[path] = tuple(sig) + (kind,)
    _cls_load()[path] = sig + [kind]
    _cls_save()
    return kind


_SCAN_MEM = {}                                  # 폴더 → (만료시각, [(경로, 종류)])
SCAN_TTL = 60.0                                 # 초. 한 요청 안의 반복 호출을 합치는 용도


SKIP_DIRS = {"_보관", "_중복사본_보관"}      # 백업 사본 — 새 자료가 아니다


def scan(folder=None, ttl=None):
    """[(경로, 종류)] — 임시파일(~$)은 제외. 분류는 캐시, 폴더 훑기는 짧게 메모이즈한다.

    ★ 분류를 캐시한 뒤에도 대시보드가 느렸다. `pick()` 이 5종(ledger·slips·po·tax·stmt)
      마다 **네트워크 드라이브를 다시 훑었기** 때문이다(재귀 glob + 파일마다 os.stat).
      종류가 다르다고 폴더가 달라지는 게 아니므로 훑기 결과를 잠깐 재사용한다.
      TTL 을 짧게(60초) 두어 새 파일이 들어오면 다음 분에는 잡힌다."""
    folder = folder or INBOX_DIR
    ttl = SCAN_TTL if ttl is None else ttl
    now = time.monotonic()
    hit = _SCAN_MEM.get(folder)
    if hit and hit[0] > now:
        return hit[1]
    out = []
    for p in sorted(glob.glob(os.path.join(folder, "**", "*.xls*"), recursive=True)):
        if os.path.basename(p).startswith("~$"):
            continue
        # ★ 백업 폴더는 훑지 않는다 (2026-08-07 실측). `_보관` 에는 복구용 보관이 매일
        #   쌓아 둔 사본이 들어간다 — 9일치 1.3GB, 관리대장 사본만 21개다. 그것까지
        #   재귀로 훑느라 이 함수가 **222초**를 썼고, daily_run 의 '자료현황 갱신'이
        #   600초 한도를 넘겨 매 회차 FAIL 이었다. 사본은 새 자료가 아니므로 셀 이유가 없다.
        if any(seg in SKIP_DIRS for seg in p.replace("\\", "/").split("/")):
            continue
        out.append((p, classify_cached(p)))
    _SCAN_MEM[folder] = (now + ttl, out)
    return out


def pick(kind, folder=None):
    """해당 종류 파일 경로. 내용 판별이 우선이고, 판별 실패 시 파일명으로 한 번 더 본다.

    ★ 폴더를 지정하지 않으면 **원본 자료 폴더 + 로컬 inbox** 를 모두 본다.
      사용자 지시(2026-07-28)로 원본은 '0. 원본 자료' 폴더에 모으지만, 급할 때
      PC inbox 에 바로 떨어뜨리는 경우가 있어 둘 다 훑는다."""
    if folder is None:
        try:
            from source_dirs import excel_dirs
            got, seen = [], set()
            for d in excel_dirs():
                for p_, k in scan(d):
                    if k == kind and os.path.basename(p_) not in seen:
                        seen.add(os.path.basename(p_))
                        got.append(p_)
            if got:
                return got
        except Exception:
            pass
    got = [p for p, k in scan(folder) if k == kind]
    if got:
        return got
    hints = {"ledger": ("원장", "계정"), "po": ("PO",), "sales": ("판매", "매출"),
             "tax": ("계산서",), "stmt": ("명세서",), "slips": ("전표", "거래조회"),
             "receipt": ("입금", "수금", "송금")}
    return [p for p, k in scan(folder) if k == "unknown"
            and any(h.lower() in os.path.basename(p).lower() for h in hints.get(kind, ()))]


LABEL = {"ledger": "거래처별계정별원장",
         "ledger_acct": "계정별원장(전 거래처)", "journal": "분개장", "cashbook": "현금출납장",
         "taxstep": "전자(세금)계산서 진행단계", "quote": "견적서조회",
         "po": "쿠팡 PO 목록",
         "sales": "판매·세금계산서 내보내기", "tax": "매출(세금)계산서현황",
         "stmt": "거래명세서 현황", "slips": "회계거래(전표) 현황",
         "taxinv": "매출(세금)계산서조회(재고)", "hometax": "홈택스 전자(세금)계산서",
         "receipt": "입금·수금 내역",
         "billing_status": "PO·청구 대조 현황표(입금 원본 아님)",
         "unknown": "판별 실패"}

if __name__ == "__main__":
    rows = scan()
    if not rows:
        print(f"inbox 비어 있음 — {INBOX_DIR}")
    # ★ 파일이 있어도 **내용이 비어 있으면** 대조가 안 된다. 2026-07-27에 ERP 내보내기
    #   3개가 '회사명' 한 줄만 있는 빈 파일이었는데, 크기(13KB·20KB)만 보면 정상처럼 보여
    #   아무도 몰랐다. 그래서 행 수를 세어 같이 보여준다.
    def rowcount(path):
        try:
            import openpyxl
            w = openpyxl.load_workbook(path, read_only=True, data_only=True)
            n = 0
            for sn in w.sheetnames:
                n += sum(1 for r in w[sn].iter_rows(values_only=True)
                         if sum(1 for x in r if x not in (None, "")) >= 3)
            w.close()
            return n
        except Exception:
            return -1
    empty = []
    for p, k in rows:
        n = rowcount(p)
        mark = ""
        if n == 0:
            mark = "  ★ 비어 있음 — 다시 내보내세요"
            empty.append(os.path.basename(p))
        elif n > 0:
            mark = f"  {n}행"
        print(f"  [{LABEL[k]:14s}] {os.path.basename(p)}  ({os.path.getsize(p)//1024}KB){mark}")
    if empty:
        print(f"\n★ 내용이 없는 파일 {len(empty)}개: " + ", ".join(empty))
        print("  이카운트에서 조회 조건(기간·거래처)을 넣고 **화면에 행이 보이는 상태**에서 내보내세요.")
    if rows:
        from collections import Counter
        c = Counter(k for _, k in rows)
        print("집계:", ", ".join(f"{LABEL[k]} {n}건" for k, n in c.items()))
