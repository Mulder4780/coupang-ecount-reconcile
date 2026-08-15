# -*- coding: utf-8 -*-
"""
stmt_link.py — 거래명세서(인쇄본) ↔ 판매조회(프로젝트코드) ↔ 원장 교차 매칭 (읽기 전용)

왜(2026-08-05 야간): 명세서 인쇄본 790건에는 **프로젝트코드가 없다**(거래처가 캠프명뿐).
판매조회에는 UJ코드·진행상태·PO가 건별로 있다. 둘을 잇는 공식 키는 이카운트 화면의
'불러온 전표'(엑셀에는 안 나옴)라서, **금액합계+일자**로 잇는다:
  · 같은 일자에 같은 금액합계가 판매조회에 **유일하게** 있으면 → 그 UJ코드로 확정
  · 둘 이상이면 모호 — 추측하지 않고 '모호'로 남긴다(원장 반영 금지 원칙)
결과는 명세서마다 UJ코드·진행상태·PO를 붙인 색인이며, 후속 대조·완료 판정의 근거 자료다.

산출: reports/명세서_프로젝트_매칭.json / .md   (엑셀에 쓰지 않는다)
"""
import json
import os
import re
import sys
from collections import defaultdict

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

OUT_JSON = os.path.join(ROOT, "reports", "명세서_프로젝트_매칭.json")
OUT_MD = os.path.join(ROOT, "reports", "명세서_프로젝트_매칭.md")
UJ = re.compile(r"UJ\d{7}")
# 쿠팡 현장으로 보이는 거래처 이름 — 캠프·MB·허브·FC 계열.
# 여기 걸리지 않는 곳(물류사·제조사)은 UJ 프로젝트가 없어 짝이 없는 것이 정상이다.
CAMPLIKE = re.compile(r"캠프|MB|Sub-?hub|Sub-?FC|허브|HUB|FC|물류센터|센터|터미널", re.I)


def load_sales():
    """판매조회 엑셀 **전부**에서 [일자, UJ, 진행상태, PO, 거래처, 금액합계] 행을 모은다.

    ★ 2026-08-05 — 예전에는 조건에 맞는 **첫 파일 하나만** 읽고 멈췄다(`break`).
      2025 판매조회를 내려받자 그것이 가장 최근 파일이 되어, 2026 명세서 793건이
      **전부 짝 없음**으로 떨어졌다(실측 확정 593 → 0). 한 해를 받으면 다른 해가
      사라지는 구조라 고친다. 같은 행이 두 파일에 겹쳐도 UJ 는 집합으로 다루므로
      '모호'가 늘지 않는다.
    """
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from erp_sales_index import sales_candidate_paths
    # 최근 파일만 자르면 다른 ERP 내보내기가 쌓인 날 유효 판매조회가 순위 밖으로 밀린다.
    # ERP 판매 색인과 같은 후보 선택기를 써서, 최신 파일과 이미 sales 로 입증된 파일을
    # 모두 머리글로 다시 확인한다(캐시는 채택 근거가 아니라 후보 우선순위일 뿐이다).
    cands = sales_candidate_paths()
    rows, srcs = [], []
    for p in cands:
        try:
            wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
            ws = wb.active
            head = [str(c or "") for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
            joined = "|".join(head)
            if "프로젝트코드" in joined and "진행상태" in joined and "금액합계" in joined:
                idx = {h: i for i, h in enumerate(head)}
                for r in ws.iter_rows(min_row=3, values_only=True):
                    r = ["" if c is None else str(c).strip() for c in r]
                    if len(r) <= max(idx.values()):
                        continue
                    uj = (UJ.search(r[idx.get("프로젝트코드코드", 1)] or "") or UJ.search(" ".join(r)))
                    amt = (r[idx["금액합계"]] or "").replace(",", "")
                    rows.append({
                        "date": (r[idx["일자"]] or "").strip()[:10].replace("-", "/"),
                        "uj": uj.group(0) if uj else "",
                        "state": r[idx["진행상태"]],
                        "po": r[idx.get("PO번호", 4)] if "PO번호" in idx else "",
                        "cust": r[idx.get("거래처명", 7)] if "거래처명" in idx else "",
                        "total": int(float(amt)) if re.fullmatch(r"-?\d+(\.\d+)?", amt) else 0,
                    })
                srcs.append(os.path.basename(p))
            wb.close()
        except Exception:
            continue
    return rows, srcs


def main():
    docs = json.load(open(os.path.join(ROOT, "reports", "거래명세서_상세.json"),
                          encoding="utf-8"))["docs"]
    sales, srcs = load_sales()
    src = ", ".join(srcs[:4]) + (f" 외 {len(srcs) - 4}개" if len(srcs) > 4 else "")
    if not sales:
        print("판매조회 원본을 찾지 못함 — 중단")
        return 1

    # (일자, 금액합계) → 판매행들.  일자는 전표일과 같아야 정상이지만,
    # 하루이틀 밀리는 손입력이 있어 ±2일까지 본다(정확일 우선).
    by_amt = defaultdict(list)
    for s in sales:
        by_amt[s["total"]].append(s)

    def dnum(d):  # '2026/01/26' → 20260126
        p = re.findall(r"\d+", d)
        return int("".join(p)) if len(p) >= 3 else 0

    linked, ambiguous, unmatched = [], [], []
    for d in docs:
        slip_date = d["slip"].split("-")[0]
        cands = by_amt.get(d["amount"], [])
        exact = [s for s in cands if s["date"] == slip_date]
        near = exact or [s for s in cands if abs(dnum(s["date"]) - dnum(slip_date)) <= 2]
        ujs = {s["uj"] for s in near if s["uj"]}
        if len(ujs) == 1:
            s = near[0]
            linked.append({"slip": d["slip"], "cust": d["cust"], "amount": d["amount"],
                           "uj": ujs.pop(), "state": s["state"], "po": s["po"],
                           "how": "합계", "exact_date": bool(exact)})
            continue
        if len(ujs) > 1:
            # 같은 금액·일자가 여럿이면 **거래처(캠프명)** 로 가른다 — 명세서 수신 거래처와
            # 판매행 거래처명이 같은 후보만 남겨 유일해지면 확정.
            byc = [s for s in near if s["uj"] and d["cust"] and
                   (s["cust"] == d["cust"] or d["cust"] in s["cust"] or s["cust"] in d["cust"])]
            cujs = {s["uj"] for s in byc}
            if len(cujs) == 1:
                s = byc[0]
                linked.append({"slip": d["slip"], "cust": d["cust"], "amount": d["amount"],
                               "uj": cujs.pop(), "state": s["state"], "po": s["po"],
                               "how": "합계+거래처", "exact_date": s in exact})
            else:
                ambiguous.append({"slip": d["slip"], "amount": d["amount"],
                                  "ujs": sorted(ujs)[:5]})
            continue
        # 합계로 못 찾으면 **품목 단위**로 — 명세서 한 장이 판매행 여러 개를 묶은 경우다.
        # 품목 하나(공급+부가세)가 같은 일자(±2일) 판매행과 유일하게 맞으면 그 UJ 를 모은다.
        item_ujs, item_states, hit = set(), set(), 0
        for it in d.get("items", []):
            t = (it.get("supply") or 0) + (it.get("vat") or 0)
            if not t:
                continue
            ic = [s for s in by_amt.get(t, [])
                  if abs(dnum(s["date"]) - dnum(slip_date)) <= 2 and s["uj"]]
            iu = {s["uj"] for s in ic}
            if len(iu) == 1:
                item_ujs |= iu
                item_states |= {s["state"] for s in ic}
                hit += 1
        if item_ujs and hit >= max(1, len(d.get("items", [])) // 2):
            linked.append({"slip": d["slip"], "cust": d["cust"], "amount": d["amount"],
                           "uj": ";".join(sorted(item_ujs)),
                           "state": ";".join(sorted(item_states)), "po": "",
                           "how": f"품목 {hit}/{len(d.get('items', []))}",
                           "exact_date": False})
        else:
            unmatched.append({"slip": d["slip"], "cust": d["cust"], "amount": d["amount"]})

    # ── 마지막 수단: **여러 판매행을 묶은 한 장**(부분합) ──────────────────────
    # 원인을 세어 보니 짝 없는 쿠팡 현장 명세서의 92%가 이것이었다(2026-08-06).
    # 현장은 프로젝트 두셋을 한 장에 묶어 청구한다. 조합이 **유일할 때만** 확정한다.
    sub_amb = []
    try:
        import stmt_subset
        taken = {u for x in linked for u in str(x.get("uj") or "").split(";") if u}
        sub_linked, sub_amb, _rest = stmt_subset.find(
            docs, sales, taken, {x["slip"] for x in unmatched})
        if sub_linked:
            got = {s["slip"] for s in sub_linked}
            unmatched = [x for x in unmatched if x["slip"] not in got]
            for s in sub_linked:
                linked.append({"slip": s["slip"], "cust": s["cust"], "amount": s["amount"],
                               "uj": ";".join(s["ujs"]),
                               "state": ";".join(s["states"]), "po": "",
                               "how": s["how"], "exact_date": False,
                               "parts": s["parts"]})
        ambiguous.extend({"slip": a["slip"], "amount": a["amount"],
                          "ujs": [";".join(c) for c in a["candidates"]]} for a in sub_amb)
    except Exception as exc:                 # 부분합이 실패해도 앞의 결과는 그대로 낸다
        print(f"  ! 부분합 매칭 건너뜀: {exc}")

    state_cnt = defaultdict(int)
    for x in linked:
        state_cnt[x["state"] or "?"] += 1

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    json.dump({"sales_src": src, "linked": linked, "ambiguous": ambiguous,
               "unmatched": unmatched},
              open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False)

    L = ["# 거래명세서 ↔ 판매조회(프로젝트코드) 매칭", "",
         f"- 판매조회 원본: `{src}` ({len(sales)}행)",
         f"- 명세서 {len(docs)}건 중 **UJ 확정 {len(linked)}건** · 모호 {len(ambiguous)}건 · "
         f"판매조회에 없음 {len(unmatched)}건",
         "- 매칭 방식: ① 금액합계+일자(±2일) 유일 일치 ② 같은 금액이 여럿이면 거래처로 가름"
         " ③ 품목 단위 ④ **여러 판매행을 묶은 한 장(부분합)** — 어느 단계든 후보가"
         " 둘 이상이면 확정하지 않는다(추측 금지)", "",
         "  ※ 방식별 건수: "
         + " · ".join(f"{k} {v}건" for k, v in sorted(
             __import__("collections").Counter(x.get("how", "?") for x in linked).items(),
             key=lambda kv: -kv[1])), "",
         "## 확정 건의 진행상태 분포", "", "| 진행상태 | 건수 |", "|---|---:|"]
    for k in sorted(state_cnt, key=lambda x: -state_cnt[x]):
        L.append(f"| {k} | {state_cnt[k]} |")
    if unmatched:
        # ★ 짝 없음을 한 덩어리로 보여 주면 사람이 전부 문제라고 읽는다. 실제로는
        #   쿠팡 아닌 거래처(물류사·제조사)가 섞여 있고, 그쪽은 **UJ 프로젝트 자체가 없어**
        #   짝이 없는 게 정상이다. 확인해야 할 것만 앞에 세운다.
        camp = [x for x in unmatched if CAMPLIKE.search(x.get("cust") or "")]
        other = [x for x in unmatched if not CAMPLIKE.search(x.get("cust") or "")]
        amt = lambda rows: sum(r.get("amount") or 0 for r in rows)
        L += ["", f"## 판매조회에 짝이 없는 명세서 ({len(unmatched)}건)", "",
              f"- **확인 필요 — 쿠팡 현장 {len(camp)}건 / {amt(camp):,}원**",
              f"- 참고 — 그 밖의 거래처 {len(other)}건 / {amt(other):,}원 "
              f"(UJ 프로젝트가 없는 거래라 짝이 없는 것이 정상)", "",
              "### 확인 필요(쿠팡 현장) — 금액 큰 순 30건", "",
              "| 전표번호 | 거래처(캠프) | 금액 |", "|---|---|---:|"]
        for x in sorted(camp, key=lambda r: -(r.get("amount") or 0))[:30]:
            L.append(f"| {x['slip']} | {x['cust'][:26]} | {x['amount']:,} |")
        if other:
            L += ["", "### 참고 — 그 밖의 거래처 20건", "",
                  "| 전표번호 | 거래처 | 금액 |", "|---|---|---:|"]
            for x in sorted(other, key=lambda r: -(r.get("amount") or 0))[:20]:
                L.append(f"| {x['slip']} | {x['cust'][:26]} | {x['amount']:,} |")
    L += ["", "> 읽기 전용 색인이다. 원장·완료 처리에 쓰려면 후속 도구가 이 근거를 인용한다."]
    open(OUT_MD, "w", encoding="utf-8").write("\n".join(L))
    print(f"명세서 {len(docs)}건: UJ 확정 {len(linked)} · 모호 {len(ambiguous)} · "
          f"짝 없음 {len(unmatched)} → reports/명세서_프로젝트_매칭.md")
    return 0


if __name__ == "__main__":
    sys.exit(main())
