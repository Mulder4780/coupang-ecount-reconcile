# -*- coding: utf-8 -*-
"""
fix_ids.py — 비어 보이는 ID를 확정값으로 채워 전 행에 표기
==============================================================
02/04 시트의 ID 열은 수식이라, 새로 추가된 행은 엑셀을 한 번 열기 전까지
프로그램(openpyxl)에서 값이 비어 보인다. 이 도구가 **수식과 똑같은 규칙**으로
ID를 계산해 값으로 확정한다.

원본 수식:
    =IF($B{r}="","", "AS-" & TEXT(IF($D{r}="",TODAY(),$D{r}),"yymm") & "-" & TEXT(ROW()-4,"000"))
    → 접두어 + 접수일자(D열)의 yymm + "-" + (행번호-4, 3자리)

안전 원칙:
  1. **데이터가 있는 행(프로젝트NO 존재)만** 확정 — 빈 행의 수식은 그대로 두어
     류지영 매니저가 새 데이터를 넣을 때 자동 채번이 계속 작동한다
  2. 이미 값이 있는 ID는 건드리지 않음 (기존 채번 보존)
  3. 생성 전 **전 시트 중복 검사** — 하나라도 충돌하면 아무것도 쓰지 않고 중단
  4. 항상 vN+1 새 파일 생성

실행:
  python fix_ids.py            # 미리보기(무엇을 채울지)
  python fix_ids.py --apply    # 실제 확정(vN+1)
"""
import sys, os, re, csv, zipfile
from datetime import datetime, date
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")

# 시트: (ID열, 접두어, 판단기준 프로젝트열, 날짜열)
SHEETS = {
    "02_돌발AS접수": {"id": "접수ID", "prefix": "AS", "key": "프로젝트NO", "date": "접수일자"},
    "04_정기점검":   {"id": "점검ID", "prefix": "PM", "key": "프로젝트NO", "date": "점검예정일"},
}


def make_id(prefix, d, rown):
    """수식과 동일: 접두어-yymm-순번(행번호-4, 3자리). 날짜 없으면 오늘 기준(TODAY()와 동일)."""
    base = d if isinstance(d, (datetime, date)) else None
    yymm = (base or datetime.now()).strftime("%y%m")
    return f"{prefix}-{yymm}-{rown - 4:03d}"


def scan(master):
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    plans, existing = [], defaultdict(set)
    for sheet, spec in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [str(h).strip() if h else "" for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        idx = {h: i for i, h in enumerate(hdr) if h}
        for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
            rn = 5 + i
            g = lambda k: row[idx[k]] if k in idx and idx[k] < len(row) else None
            key, cur = g(spec["key"]), g(spec["id"])
            if cur not in (None, ""):
                existing[sheet].add(str(cur).strip())      # 기존 ID 보존·중복검사용
                continue
            if key in (None, ""):
                continue                                    # 빈 행 → 수식 그대로 유지
            plans.append({"sheet": sheet, "row": rn, "col": idx[spec["id"]] + 1,
                          "id": make_id(spec["prefix"], g(spec["date"]), rn),
                          "prj": str(key).strip(), "date": str(g(spec["date"]) or "")[:10]})
    wb.close()
    # 중복 검사 (기존 + 신규 사이, 신규끼리)
    dups = []
    seen = defaultdict(set)
    for p in plans:
        if p["id"] in existing[p["sheet"]] or p["id"] in seen[p["sheet"]]:
            dups.append(p)
        seen[p["sheet"]].add(p["id"])
    return plans, dups


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def apply_ids(master, plans):
    """수식 셀을 값 셀로 교체 (해당 행만). vN+1 생성 + 검증."""
    from ledger_writer import sheet_file_map, esc
    mv = re.search(r"_v(\d+)\.xlsx$", master)
    v = int(mv.group(1))
    dst = master.replace(f"_v{v}.xlsx", f"_v{v+1}.xlsx")
    if os.path.exists(dst):
        sys.exit(f"{os.path.basename(dst)} 이미 존재 — 중단")

    zin = zipfile.ZipFile(master)
    smap = sheet_file_map(zin)
    by_file = defaultdict(list)
    for p in plans:
        by_file[smap[p["sheet"]]].append(p)

    edits, done = {}, 0
    for fname, items in by_file.items():
        xml = zin.read(fname).decode("utf-8")
        for p in items:
            cl = col_letter(p["col"])
            ref = f"{cl}{p['row']}"
            mrow = re.search(r'(<row r="%d"[^>]*>)(.*?)(</row>)' % p["row"], xml, re.S)
            if not mrow:
                continue
            head, body, tail = mrow.groups()
            mopen = re.search(r'<c r="%s"[^>]*>' % ref, body)
            if not mopen:
                continue
            if mopen.group(0).endswith("/>"):
                span = mopen.span()
            else:
                span = (mopen.start(), body.find("</c>", mopen.end()) + 4)
            style = re.search(r' s="(\d+)"', mopen.group(0))
            s_attr = f' s="{style.group(1)}"' if style else ""
            new_cell = f'<c r="{ref}"{s_attr} t="inlineStr"><is><t>{esc(p["id"])}</t></is></c>'
            body = body[:span[0]] + new_cell + body[span[1]:]
            xml = xml[:mrow.start()] + head + body + tail + xml[mrow.end():]
            done += 1
        edits[fname] = xml.encode("utf-8")

    zout = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, edits.get(it.filename) or zin.read(it.filename))
    zout.close(); zin.close()

    z = zipfile.ZipFile(dst)
    assert z.testzip() is None, "zip 무결성 실패"
    import openpyxl
    w = openpyxl.load_workbook(dst, read_only=True, data_only=True)
    miss = [p for p in plans[:40]
            if str(w[p["sheet"]].cell(row=p["row"], column=p["col"]).value or "") != p["id"]]
    w.close()
    zsrc = zipfile.ZipFile(master)
    diff = [n for n in zsrc.namelist() if n not in edits and zsrc.read(n) != z.read(n)]
    zsrc.close(); z.close()
    assert not miss, f"기록 검증 실패: {miss[:3]}"
    assert not diff, f"의도치 않은 파트 변경: {diff}"
    return dst, done, f"v{v} → v{v+1}"


def main():
    apply_mode = "--apply" in sys.argv
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    plans, dups = scan(master)

    from collections import Counter
    per = Counter(p["sheet"] for p in plans)
    print(f"ID 미표기 행 {len(plans)}건 — " + " / ".join(f"{k} {v}건" for k, v in per.items()))
    for p in plans[:5]:
        print(f"  {p['sheet']} {p['row']}행: {p['prj']} ({p['date'] or '일자없음'}) → {p['id']}")
    if len(plans) > 5:
        print(f"  … 외 {len(plans)-5}건")
    if dups:
        print(f"\n★ 중복 충돌 {len(dups)}건 — 안전을 위해 중단합니다:")
        for d in dups[:5]:
            print("   ", d["sheet"], d["row"], d["id"])
        return

    os.makedirs(REPORT_DIR, exist_ok=True)
    if plans:
        base = os.path.join(REPORT_DIR, f"ID확정_{datetime.now():%Y%m%d_%H%M}.csv")
        with open(base, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["sheet", "row", "id", "prj", "date"], extrasaction="ignore")
            w.writeheader(); w.writerows(plans)
        print("리포트:", base)

    if not apply_mode:
        print("\n미리보기 — 실제 확정: python fix_ids.py --apply")
        return
    if not plans:
        print("확정할 ID 없음 (모두 표기됨)"); return
    dst, done, msg = apply_ids(master, plans)
    print(f"\nID 확정 완료: {msg} — {done}건 기록 (빈 행 수식은 보존)")
    print("   ", dst)


if __name__ == "__main__":
    main()
