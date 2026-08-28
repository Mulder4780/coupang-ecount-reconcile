# -*- coding: utf-8 -*-
"""업무 단계 정의 — **한 곳에서만** 정한다 (2026-08-10 지시).

사용자 지시: "돌발 AS 카테고리 및 정기점검 카테고리에 입력창을 만들어 … 플로우차트
참고해서 진행할 수 있게 하고 나중에 플로우차트가 변경되면 기능도 따라서 연동되게".

'연동'은 기능이 흐름도를 **읽는다**는 뜻이다. 단계 낱말을 화면·서버·검증에 각각 적어
두면 그 순간 사본이 셋이 되고, 흐름도가 바뀌어도 **아무 데도 안 따라간다**([162]).
그래서 낱말을 정하는 자리는 이 파일 하나이고, 이 파일은 스스로 낱말을 **짓지 않는다**.

## 근거는 셋, 층이 다르다
① **관리대장 `10_코드관리` 드롭다운** — 낱말의 정본이다. 02시트 `진행상태`(Q)·
   04시트 `점검상태`(G) 칸의 데이터유효성이 가리키는 그 열을 그대로 읽는다.
   어느 열을 가리키는지도 **워크북에서 읽는다** — 여기 'G열'이라 적어 두면 사람이
   드롭다운 출처를 옮긴 날 그 사실이 아무 화면에도 안 뜬다([165] 와 같은 종류).
② **앱 흐름도** `ledger_db.flow_steps()` — 낱말이 아니라 **일의 순서와 담당**이다.
   ①의 단계에 붙일 수 있으면 붙이고, 근거가 없으면 **붙이지 않는다**.
③ **원장에 실제로 쓰인 값** — ①에 없는 낱말을 찾기 위해서만 본다. 지우지 않는다.
   이미 그렇게 적힌 행이 있으니 화면에서 고를 수 있어야 하고, 다만 '목록 밖'이라 적는다.

## 두 가지를 지킨다
- **낱말을 지어내지 않는다**([166]). 새 낱말을 넣으면 사람이 쓰던 정렬·필터가 어긋난다.
- **바뀐 것은 반드시 티가 나게 한다.** 흐름도가 바뀌었는데 아무 화면에도 안 뜨는 것이
  제일 위험하다. 지문이 달라지면 `reports/업무흐름_변경.json` 에 남고 인계 문서 위로 오른다.

읽기 전용이다 — 큐에도 안 넣고 엑셀도 열어 고치지 않는다.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import time
import zipfile
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

REPORT_DIR = os.path.join(ROOT, "reports")
DEF_PATH = os.path.join(REPORT_DIR, "업무흐름_정의.json")
CHANGE_PATH = os.path.join(REPORT_DIR, "업무흐름_변경.json")

# 관리대장을 여는 것은 1.3초쯤 걸린다. 요청마다 열면 화면이 그만큼 늦는다 —
# **비싼 탐색은 언제나 캐시 검사 뒤에 온다**([168]).
TTL_SEC = int(os.environ.get("COUPANG_WORKFLOW_TTL") or 600)

# 어느 시트의 어느 칸이 '단계'인가. 갈래를 늘리려면 여기 한 줄이다.
KINDS = {
    "as": {"이름": "돌발AS", "시트": "02_돌발AS접수", "칸": "진행상태",
           "ID칸": "접수ID", "날짜칸": "방문예정일", "완료칸": "작업완료일"},
    "pm": {"이름": "정기점검", "시트": "04_정기점검", "칸": "점검상태",
           "ID칸": "점검ID", "날짜칸": "점검예정일", "완료칸": "실제점검일"},
}

# '완료'를 뜻하는 단계를 고르는 실마리. **후보가 유일할 때만** 고른다([172] 의 문).
DONE_HINT = "완료"

# ★ '이 건은 더 이상 갈 곳이 아니다' 를 뜻하는 낱말 (2026-08-13 지시).
#   낱말을 지어내지 않는다([166]) — 이미 코드 세 곳이 같은 짝을 쓰고 있었고
#   (`complete_verified`·`ledger_db` 의 conflicts·`backfill_rows`) 그 짝을 여기로 모았다.
#   **사본이 넷이면 한 곳만 고쳐진다**([162]) — 새로 판정하는 자리는 여기를 부른다.
#   ⚠ 낱말을 넓히지 말 것. 취소로 잘못 부르면 그 현장은 아무도 안 가는데 목록에서도
#   사라진다 — 미처리로 남는 것보다 나쁘다([172] 의 문). 근거는 **드롭다운 낱말과
#   정확히 같을 때**뿐이다(공백만 무시).
CANCELLED = ("취소", "철회")


def is_cancelled(row, kind):
    """이 행이 '접수취소'인가 — 상태 칸의 낱말이 CANCELLED 와 정확히 같을 때만 참.

    kind 는 'as'·'pm'. 어느 칸을 읽을지는 KINDS 가 정한다 — 칸 이름을 부르는 쪽에
    적으면 시트가 바뀐 날 **오류 없이 한 건도 안 걸린다**([165]).
    """
    col = (KINDS.get(kind) or {}).get("칸")
    if not col:
        return False
    return _norm((row or {}).get(col)) in tuple(_norm(w) for w in CANCELLED)


def says_done(row, kind):
    """상태 칸이 '완료'라고 말하는가 — 완료 **날짜**가 비어 있어도 참.

    날짜가 없다고 '미처리'라 부르면 다녀온 현장이 안 간 것으로 보인다.
    '완료라는데 날짜만 없다'와 '아직 안 갔다'는 다른 사실이다([169])."""
    col = (KINDS.get(kind) or {}).get("칸")
    return bool(col) and DONE_HINT in _norm((row or {}).get(col))

_MEM = {"at": 0.0, "def": None}


# ───────────────────────────────── 워크북에서 읽기 ─────────────────────────────

def _norm(s):
    """견주기용 정규화 — 공백만 없앤다. 낱말 자체는 절대 고치지 않는다."""
    return re.sub(r"\s+", "", str(s or ""))


def _sheet_xml(z, title):
    """시트 이름 → `xl/worksheets/sheetN.xml`. 자리(sheet3.xml)를 짐작하지 않는다."""
    wb = z.read("xl/workbook.xml").decode("utf-8", "ignore")
    rid = None
    for m in re.finditer(r"<sheet\b[^>]*>", wb):
        tag = m.group(0)
        name = re.search(r'name="([^"]*)"', tag)
        rel = re.search(r'r:id="([^"]*)"', tag)
        if name and rel and name.group(1) == title:
            rid = rel.group(1)
            break
    if not rid:
        return None
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
    for m in re.finditer(r"<Relationship\b[^>]*>", rels):
        tag = m.group(0)
        i = re.search(r'Id="([^"]*)"', tag)
        t = re.search(r'Target="([^"]*)"', tag)
        if i and t and i.group(1) == rid:
            target = t.group(1).lstrip("/")
            if target.startswith("xl/"):
                return target
            return "xl/" + target
    return None


def _dv_source(sheet_xml, col_letter):
    """`진행상태` 칸의 드롭다운이 **어느 열**을 보고 있나 → ("10_코드관리", "G").

    엑셀은 같은 목록을 두 곳에 적는다(구형 `dataValidation` · x14 확장). 둘 다 본다.
    """
    pat = re.compile(r"<(?:x14:)?dataValidation\b.*?</(?:x14:)?dataValidation>", re.S)
    want = re.compile(r"\b%s\d+" % re.escape(col_letter))
    for block in pat.findall(sheet_xml):
        sq = (re.search(r'sqref="([^"]*)"', block)
              or re.search(r"<xm:sqref>(.*?)</xm:sqref>", block, re.S))
        if not sq or not want.search(sq.group(1)):
            continue
        f1 = re.search(r"<(?:x14:)?formula1>(.*?)</(?:x14:)?formula1>", block, re.S)
        if not f1:
            continue
        ref = re.sub(r"</?xm:f>", "", f1.group(1)).strip()
        m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]{1,3})\$?\d+", ref)
        if m:
            return m.group(1), m.group(2)
    return None


def _col_letter(ws, want, max_row=6):
    """머리글 이름 → 열 문자. 자리를 짐작하지 않는다 — 사람이 열을 옮길 수 있다."""
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for c in row:
            if c.value is not None and _norm(c.value) == _norm(want):
                return c.column_letter
    return None


def _read_master():
    """관리대장에서 갈래별 단계 낱말을 읽어 온다. 실패하면 `None` — 지어내지 않는다."""
    import openpyxl
    import source_dirs as S
    import ecount_reconcile as E
    path = E.resolve_master(os.path.join(S.LEDGER_DIR, "x.xlsx"))
    if not path or not os.path.isfile(path):
        return None
    out = {"원본": os.path.basename(path), "갈래": {}}
    with zipfile.ZipFile(path) as z:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for key, cfg in KINDS.items():
                if cfg["시트"] not in wb.sheetnames:
                    continue
                ws = wb[cfg["시트"]]
                letter = _col_letter(ws, cfg["칸"])
                if not letter:
                    continue
                xml_name = _sheet_xml(z, cfg["시트"])
                src = None
                if xml_name and xml_name in z.namelist():
                    src = _dv_source(z.read(xml_name).decode("utf-8", "ignore"), letter)
                if not src or src[0] not in wb.sheetnames:
                    continue
                code_ws = wb[src[0]]
                words, seen = [], set()
                for r in range(1, 260):
                    v = code_ws[src[1] + str(r)].value
                    v = str(v).strip() if v is not None else ""
                    if not v:
                        continue
                    if _norm(v) == _norm(cfg["칸"]):
                        continue            # 목록 맨 위의 제목 칸이다 — 단계가 아니다
                    if v in seen:
                        continue
                    seen.add(v)
                    words.append(v)
                if words:
                    out["갈래"][key] = {"낱말": words,
                                        "출처": "%s!%s열" % (src[0], src[1]),
                                        "칸": cfg["칸"], "시트": cfg["시트"]}
        finally:
            wb.close()
    return out if out["갈래"] else None


EXTRA_KEY = "work_flow_extra:"      # app_setting 열쇠 접두 — 갈래마다 한 줄


def extra_stages(kind):
    """사람이 **앱에서 더한** 단계 낱말. `([낱말...], 못읽은이유)`.

    ★ **관리대장 시트를 안 건드린다** (2026-08-28 형님 지시 "현재단계 진행상태에
      '택배발송' 추가"). 넣을 자리는 `10_코드관리!G13`(비어 있다)이지만
      2026-08-10 정본 규칙이 **새 Excel 직접 mutator 를 금지**하고, 2026-08-24
      지시가 *"엑셀은 저장용으로만"* 이다. 그러므로 낱말을 더하는 길은 앱 DB 뿐이고
      그것이 `[196]` 이 이관 대기 목록에 적어 둔 방향 그대로다.

    ★ **`app_setting` 표를 빌린다**([162]) — 낙관잠금·감사로그가 공짜로 붙는다.

    ⚠ **엑셀을 열어 보는 사람에게는 안 보인다**([169]). 관리대장 드롭다운은
      `10_코드관리` 열을 그대로 읽으므로 여기 더한 낱말이 거기엔 없다. 엑셀은
      이제 읽기 전용 보관본이라([212]·[474]) 그 칸을 사람이 고르지 않지만,
      **모르는 채로 두지 않기 위해** 리포트가 그 사실을 적는다.

    ★ **못 읽으면 빈 목록이다** — 이것 하나 때문에 단계 정의가 통째로 안 만들어지면
      새 접수가 등록조차 안 된다([172]). 대신 **이유를 같이 돌려준다**([169]).
    """
    key = str(kind or "").strip()
    if not key:
        return [], ""
    try:
        from app_store import default_store
        rec = default_store().get_setting(EXTRA_KEY + key)
    except Exception as exc:
        return [], "%s: %s" % (type(exc).__name__, exc)
    val = rec.get("value")
    if not isinstance(val, dict):
        return [], ""
    out, seen = [], set()
    for item in (val.get("낱말") or []):
        w = str((item or {}).get("단계") or "").strip()
        if w and w not in seen:
            seen.add(w)
            out.append(w)
    return out, ""


def _extra_raw(kind):
    from app_store import default_store
    rec = default_store().get_setting(EXTRA_KEY + str(kind))
    val = rec.get("value")
    return ({} if not isinstance(val, dict) else val), int(rec.get("record_version") or 0)


def add_stage(kind, word, *, actor="app", why=""):
    """단계 낱말 하나를 더한다. 이미 있으면 아무것도 안 한다(멱등).

    ★ **관리대장에 있는 낱말은 안 받는다** — 같은 낱말이 두 출처로 들어오면
      화면이 그것을 두 번 보여 주고, 어느 쪽이 정본인지 아무도 모른다([162]).
    ★ **되돌릴 수 있다** — `remove_stage` 가 짝이다. 되살리기 없는 추가는
      잘못 넣은 순간 사람이 할 수 있는 일이 없어진다.
    """
    if kind not in KINDS:
        raise ValueError("모르는 갈래입니다: %s (%s 중 하나)"
                         % (kind, " · ".join(sorted(KINDS))))
    w = str(word or "").strip()
    if not (1 <= len(w) <= 30):
        raise ValueError("단계 낱말을 1~30자로 적어 주세요")
    d = definition(refresh=True)
    official = [s["단계"] for s in ((d.get("갈래") or {}).get(kind) or {}).get("단계") or []
                if s.get("출처") == "관리대장 목록"]
    if w in official:
        return {"ok": True, "더함": False,
                "msg": "'%s' 은(는) 이미 관리대장 목록에 있습니다" % w}
    cur, ver = _extra_raw(kind)
    words = list(cur.get("낱말") or [])
    if any(str((x or {}).get("단계") or "") == w for x in words):
        return {"ok": True, "더함": False, "msg": "'%s' 은(는) 이미 더해져 있습니다" % w}
    words.append({"단계": w, "때": datetime.now().isoformat(timespec="seconds"),
                  "누가": str(actor or "app")[:80], "왜": str(why or "")[:200]})
    from app_store import default_store
    default_store().set_setting(
        EXTRA_KEY + kind, {"낱말": words},
        expected_version=(None if ver == 0 else ver),
        actor=actor or "app", source="app-work-flow-extra",
        evidence="단계 낱말 추가: %s" % w)
    return {"ok": True, "더함": True, "낱말수": len(words),
            "msg": "'%s' 을(를) %s 단계에 더했습니다" % (w, KINDS[kind]["이름"])}


def remove_stage(kind, word, *, actor="app"):
    """더했던 낱말을 뺀다. **관리대장 낱말은 못 뺀다** — 그것은 원본의 몫이다."""
    if kind not in KINDS:
        raise ValueError("모르는 갈래입니다: %s" % kind)
    w = str(word or "").strip()
    cur, ver = _extra_raw(kind)
    words = [x for x in (cur.get("낱말") or [])
             if str((x or {}).get("단계") or "") != w]
    if len(words) == len(cur.get("낱말") or []):
        return {"ok": True, "뺌": False,
                "msg": "'%s' 은(는) 앱에서 더한 낱말이 아닙니다" % w}
    from app_store import default_store
    default_store().set_setting(
        EXTRA_KEY + kind, {"낱말": words},
        expected_version=(None if ver == 0 else ver),
        actor=actor or "app", source="app-work-flow-extra",
        evidence="단계 낱말 제거: %s" % w)
    return {"ok": True, "뺌": True, "낱말수": len(words),
            "msg": "'%s' 을(를) 뺐습니다" % w}


def _used_words():
    """원장에 **실제로 쓰인** 값과 건수. 목록 밖 낱말을 찾기 위해서만 본다."""
    used = {}
    try:
        sys.path.insert(0, os.path.join(ROOT, "webapp"))
        import app_server as A
        works = A.get_works() or {}
    except Exception:
        return used
    for key, cfg in KINDS.items():
        counts = {}
        for row in (works.get(key) or []):
            w = str(row.get(cfg["칸"]) or "").strip()
            if w:
                counts[w] = counts.get(w, 0) + 1
        if counts:
            used[key] = counts
    return used


def _flow():
    """앱 흐름도. 못 읽으면 빈 목록 — 흐름도가 없다고 단계 목록까지 잃지는 않는다."""
    try:
        import ledger_db
        return ledger_db.flow_steps() or []
    except Exception:
        return []


def _link(word, flow):
    """단계 낱말에 흐름도 단계를 붙인다. **후보가 유일할 때만** 붙인다.

    잘못 붙이면 화면이 엉뚱한 담당자를 가리킨다 — 안 붙은 것보다 나쁘다([172]).
    근거는 '한쪽이 다른 쪽을 통째로 담고 있다' 하나뿐이다. 낱말을 쪼개 견주지 않는다.
    """
    a = _norm(word)
    hits = [s for s in flow if a and (a in _norm(s.get("단계")) or _norm(s.get("단계")) in a)]
    if len(hits) != 1:
        return None
    s = hits[0]
    return {"흐름단계": s.get("단계"), "담당": s.get("담당") or "",
            "근거": s.get("근거") or "", "순서": s.get("순서")}


# ───────────────────────────────── 정의 만들기 ────────────────────────────────

def build():
    """지금의 단계 정의를 만든다(캐시를 보지 않는다)."""
    now = datetime.now().isoformat(timespec="seconds")
    flow = _flow()
    used = _used_words()
    try:
        master = _read_master()
    except Exception as e:                              # 관리대장이 없거나 못 열릴 때
        master = None
        master_err = str(e)[:200]
    else:
        master_err = ""
    out = {"만든시각": now, "원본": (master or {}).get("원본", ""),
           "흐름단계수": len(flow), "갈래": {}, "경고": []}
    if not master:
        out["경고"].append("관리대장 드롭다운을 읽지 못했습니다 — 단계 목록은 "
                           "**원장에 이미 쓰인 값**만으로 만들었습니다"
                           + (" (%s)" % master_err if master_err else ""))
    for key, cfg in KINDS.items():
        words_cfg = ((master or {}).get("갈래") or {}).get(key) or {}
        # ★ **관리대장 목록과 앱에서 더한 낱말을 갈라 둔다**([166]).
        #   기본단계·완료단계는 **관리대장에서만** 와야 한다 — 앱 추가분이 그
        #   자리를 차지하면 Z: 가 끊긴 날 새 접수가 엉뚱한 단계로 박힌다.
        master_words = list(words_cfg.get("낱말") or [])
        added, add_why = extra_stages(key)
        if add_why:
            out["경고"].append("%s — 앱에서 더한 단계 낱말을 못 읽었습니다: %s"
                               % (cfg["이름"], add_why))
        # 더한 낱말은 **맨 뒤**에 붙인다 — 중간에 끼우면 기본단계가 밀린다.
        official = master_words + [w for w in added if w not in master_words]
        counts = used.get(key) or {}
        stages = []
        for w in official:
            item = {"단계": w,
                    "출처": ("관리대장 목록" if w in master_words else "앱에서 더함"),
                    "쓰인건수": counts.get(w, 0)}
            link = _link(w, flow)
            if link:
                item.update(link)
            stages.append(item)
        # 목록에 없는데 원장에 쓰여 있는 낱말 — 지우지 않는다. 고를 수는 있어야 한다.
        extra = sorted([w for w in counts if w not in official],
                       key=lambda w: -counts[w])
        for w in extra:
            item = {"단계": w, "출처": "원장에 쓰인 값(목록 밖)", "쓰인건수": counts[w]}
            link = _link(w, flow)
            if link:
                item.update(link)
            stages.append(item)
        done = [s["단계"] for s in stages
                if DONE_HINT in s["단계"] and s["출처"] == "관리대장 목록"]
        out["갈래"][key] = {
            "이름": cfg["이름"], "시트": cfg["시트"], "칸": cfg["칸"],
            "ID칸": cfg["ID칸"], "날짜칸": cfg["날짜칸"], "완료칸": cfg["완료칸"],
            "목록출처": words_cfg.get("출처", ""),
            "단계": stages,
            # ★ 기본단계는 **관리대장 목록에서만** 온다. 목록을 못 읽었다고
            # '원장에 쓰인 목록 밖 낱말 중 제일 흔한 것'으로 대신하면 안 된다 —
            # 그것은 지나간 기록이지 새 건이 처음 가질 단계가 아니다.
            # 2026-08-18 실측: Z: 가 끊기자 그 자리가 **'취소'** 가 됐다. 그대로
            # 등록하면 새 접수가 취소로 박혀 미처리에서 사라진다([243]).
            # 못 읽으면 빈칸이다 — 부르는 쪽이 '모른다'를 보고 멈춘다([166]·[172]).
            # ★ 앱 추가분은 여기 못 온다 — `master_words` 만 본다([166]).
            "기본단계": (master_words[0] if master_words else ""),
            "완료단계": (done[0] if len(done) == 1 else ""),
            "목록밖": extra,
            "흐름연결": sum(1 for s in stages if s.get("흐름단계")),
        }
        if len(done) != 1:
            out["경고"].append(
                "%s — '완료'로 읽을 단계를 하나로 못 정했습니다(후보 %d개). "
                "원클릭 완료 단추를 감춥니다." % (cfg["이름"], len(done)))
        app_added = [w for w in official if w not in master_words]
        if app_added:
            out["경고"].append(
                "%s — 앱에서 더한 단계 %s. 앱 드롭다운에는 나오지만 "
                "**관리대장(엑셀)을 열어 보는 사람에게는 안 보입니다** — "
                "엑셀은 읽기 전용 보관본입니다."
                % (cfg["이름"], " · ".join(app_added)))
        if extra:
            out["경고"].append(
                "%s — 목록에 없는 낱말이 원장에 쓰여 있습니다: %s. "
                "고를 수는 있게 두었지만 '목록 밖'으로 표시합니다."
                % (cfg["이름"], " · ".join("%s(%d건)" % (w, counts[w]) for w in extra)))
    out["지문"] = fingerprint_of(out)
    return out


def fingerprint_of(d):
    """무엇이 바뀌면 '바뀌었다'고 할 것인가 — 낱말 목록과 흐름도 단계다.

    쓰인 건수는 매일 움직이므로 지문에 넣지 않는다. 넣으면 매일 '바뀜'이 떠서
    아무도 안 보게 된다(경보가 대부분이면 그 경보는 죽는다).
    """
    seed = {k: [s["단계"] for s in (v.get("단계") or [])]
            for k, v in sorted((d.get("갈래") or {}).items())}
    seed["흐름"] = [(s.get("단계"), s.get("담당")) for s in _flow()]
    blob = json.dumps(seed, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()[:16]


def definition(refresh=False):
    """화면·서버가 부르는 자리. **캐시 검사가 먼저**다([168])."""
    now = time.time()
    if not refresh and _MEM["def"] and (now - _MEM["at"]) < TTL_SEC:
        return _MEM["def"]
    if not refresh and os.path.isfile(DEF_PATH):
        try:
            age = now - os.path.getmtime(DEF_PATH)
            if age < TTL_SEC:
                with open(DEF_PATH, encoding="utf-8") as f:
                    d = json.load(f)
                _MEM.update(at=now, **{"def": d})
                return d
        except Exception:
            pass
    d = build()
    save(d)
    _MEM.update(at=now, **{"def": d})
    return d


def save(d):
    os.makedirs(REPORT_DIR, exist_ok=True)
    tmp = DEF_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        json.dump(d, f, ensure_ascii=False, indent=1)
    os.replace(tmp, DEF_PATH)


# ───────────────────────────────── 쓰는 쪽이 부르는 것 ─────────────────────────

def stages(kind, refresh=False):
    return ((definition(refresh).get("갈래") or {}).get(kind) or {}).get("단계") or []


def stage_words(kind, refresh=False):
    return [s["단계"] for s in stages(kind, refresh)]


def default_stage(kind, refresh=False):
    """새 건이 처음 갖는 단계. 화면·서버가 **같은 것**을 써야 한다."""
    return ((definition(refresh).get("갈래") or {}).get(kind) or {}).get("기본단계") or ""


def done_stage(kind, refresh=False):
    """'완료'를 뜻하는 단계. 하나로 못 정하면 빈 문자열 — 짐작해서 찍지 않는다."""
    return ((definition(refresh).get("갈래") or {}).get(kind) or {}).get("완료단계") or ""


# ───────────────────────────────── 바뀐 것 알리기 ──────────────────────────────

def check(write=True):
    """지문이 달라졌으면 **자국을 남긴다.** 조용히 따라가면 아무도 모른다."""
    prev = {}
    if os.path.isfile(CHANGE_PATH):
        try:
            with open(CHANGE_PATH, encoding="utf-8") as f:
                prev = json.load(f)
        except Exception:
            prev = {}
    old = {}
    if os.path.isfile(DEF_PATH):
        try:
            with open(DEF_PATH, encoding="utf-8") as f:
                old = json.load(f)
        except Exception:
            old = {}
    new = build()
    changed = bool(old) and old.get("지문") and old["지문"] != new["지문"]
    if write:
        save(new)
        _MEM.update(at=time.time(), **{"def": new})
    if not changed:
        return {"바뀜": False, "지문": new["지문"], "미확인": prev if prev.get("바뀜") else None}
    diff = []
    for key in sorted(new.get("갈래") or {}):
        a = [s["단계"] for s in ((old.get("갈래") or {}).get(key) or {}).get("단계") or []]
        b = [s["단계"] for s in (new["갈래"][key].get("단계") or [])]
        for w in b:
            if w not in a:
                diff.append("%s — 단계 추가: %s" % (new["갈래"][key]["이름"], w))
        for w in a:
            if w not in b:
                diff.append("%s — 단계 없어짐: %s" % (new["갈래"][key]["이름"], w))
    if old.get("흐름단계수") != new.get("흐름단계수"):
        diff.append("흐름도 단계 수 %s → %s"
                    % (old.get("흐름단계수"), new.get("흐름단계수")))
    if not diff:
        diff.append("흐름도 단계 이름·담당이 바뀌었습니다")
    rec = {"바뀜": True, "시각": new["만든시각"], "이전지문": old.get("지문"),
           "지문": new["지문"], "무엇이": diff}
    if write:
        os.makedirs(REPORT_DIR, exist_ok=True)
        with open(CHANGE_PATH, "w", encoding="utf-8", newline="") as f:
            json.dump(rec, f, ensure_ascii=False, indent=1)
    return rec


def banner():
    """인계 문서가 읽는 자리. 사람이 `--ack` 할 때까지 남는다."""
    if not os.path.isfile(CHANGE_PATH):
        return None
    try:
        with open(CHANGE_PATH, encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return None
    return d if d.get("바뀜") else None


def ack():
    if os.path.isfile(CHANGE_PATH):
        os.remove(CHANGE_PATH)
    return True


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    if "--ack" in argv:
        ack()
        print("업무 흐름 변경 알림을 내렸습니다")
        return 0
    # 단계 낱말 더하기·빼기 — **사람이 명령할 때만** 돈다(무인 회차는 안 부른다).
    for flag, fn, verb in (("--add", add_stage, "더함"), ("--remove", remove_stage, "뺌")):
        if flag in argv:
            i = argv.index(flag)
            rest = [a for a in argv[i + 1:] if not a.startswith("-")]
            if len(rest) < 2:
                print("쓰기: python work_flow.py %s <as|pm> <낱말> [사유]" % flag)
                return 2
            kw = {"actor": "사람"} if flag == "--remove" else {
                "actor": "사람", "why": " ".join(rest[2:])}
            try:
                r = fn(rest[0], rest[1], **kw)
            except Exception as exc:
                print("못 했습니다 — %s" % exc)
                return 2
            print(r.get("msg") or "")
            if r.get(verb):
                print("  → 앱 드롭다운에 곧바로 나옵니다. "
                      "관리대장(엑셀)에는 안 들어갑니다 — 엑셀은 보관본입니다.")
            return 0
    if "--check" in argv:
        r = check(write=True)
        if r.get("바뀜"):
            print("★ 업무 흐름이 바뀌었습니다 — " + " · ".join(r["무엇이"]))
        else:
            print("업무 흐름 그대로 (지문 %s)" % r.get("지문"))
        return 0
    d = definition(refresh="--refresh" in argv)
    print("정의 %s · 원본 %s · 흐름도 %d단계"
          % (d.get("만든시각"), d.get("원본") or "(못 읽음)", d.get("흐름단계수", 0)))
    for key, v in (d.get("갈래") or {}).items():
        print(" [%s] %s — %s" % (key, v["이름"], v.get("목록출처") or "목록 출처 없음"))
        for s in v.get("단계") or []:
            mark = "·" if s["출처"] == "관리대장 목록" else "!"
            link = ("  ← %s(%s)" % (s["흐름단계"], s["담당"])) if s.get("흐름단계") else ""
            print("   %s %-8s %5d건%s" % (mark, s["단계"], s.get("쓰인건수", 0), link))
        print("   기본 '%s' · 완료 '%s' · 흐름연결 %d"
              % (v.get("기본단계"), v.get("완료단계") or "(못 정함)", v.get("흐름연결", 0)))
    for w in d.get("경고") or []:
        print(" ! " + w)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
