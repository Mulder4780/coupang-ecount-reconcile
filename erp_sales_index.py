# -*- coding: utf-8 -*-
"""
erp_sales_index.py — ERP 판매조회를 **프로젝트NO 색인**으로 굳힌다 (읽기 전용)

왜(2026-08-05):
 ① 앱 정산 금액이 틀렸다. 실제작업공급가액(03시트 수식)이 비어 거래명세서합계로 대신
    보여 줬는데, **명세서합계는 부가세 포함액**이다(UJ2600050: 476,300 = 433,000×1.1).
    "부가세 별도"라고 적힌 자리에 포함액이 들어가 있었다. → ERP 공급가액을 정본으로 쓴다.
 ② "세금계산서 미발행을 찾아 완료 처리" — 발행 여부의 객관 근거는 ERP 진행상태다
    (6.세금계산서발행 / 7.수금완료). 그 상태를 프로젝트NO 로 바로 찾게 만든다.

산출: reports/ERP판매_프로젝트색인.json
  { "UJ2600050": {"supply":433000,"vat":43300,"total":476300,"state":"4.세금계산서발행대기",
                  "date":"2026/01/07","po":"","cust":"강서1MB(가양A)","rows":1}, ... }
같은 UJ 가 여러 행이면 금액은 합산하고 상태는 **가장 앞선 단계**를 남긴다
(한 건이라도 미발행이면 그 프로젝트는 아직 안 끝난 것으로 본다 — 낙관 금지).
"""
import glob
import hashlib
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

OUT = os.path.join(ROOT, "reports", "ERP판매_프로젝트색인.json")
UJ = re.compile(r"UJ\d{7}")
# 진행 단계 순서 — 작은 값이 덜 진행된 것. 합칠 때 **가장 덜 진행된 상태**를 남긴다.
ORDER = {"1.미확인": 1, "확인": 2, "2.메일발송": 3, "3.오더처리": 4,
         "4.세금계산서발행대기": 5, "5.": 6, "6.세금계산서발행": 7, "7.수금완료": 8,
         "8.무상납품완료": 9}
ISSUED = ("6.세금계산서발행", "7.수금완료")
INDEX_RULES_VERSION = 1


def rank(state):
    for k, v in ORDER.items():
        if str(state or "").startswith(k):
            return v
    return 0


def _path_key(path):
    return os.path.normcase(os.path.abspath(path or ""))


def prioritize_sales_candidates(cands, cached_sales=(), limit=60):
    """Keep recent unknown exports *and* every file already proven to be sales.

    The ERP folder contains many unrelated exports.  A plain ``cands[:limit]``
    silently loses an older valid sales export as soon as enough newer ledger or
    tax files arrive.  The classification cache is only a prioritization hint;
    callers still verify workbook headers before accepting a file.
    """
    ordered = list(cands)
    known = {_path_key(p) for p in cached_sales}
    selected = list(ordered[:max(0, int(limit))])
    selected.extend(p for p in ordered if _path_key(p) in known)
    out, seen = [], set()
    for path in selected:
        key = _path_key(path)
        if key not in seen:
            seen.add(key)
            out.append(path)
    return out


def sales_candidate_paths(limit=60):
    """Return recent exports plus cached sales exports, newest first.

    ★ **목록에 딸려 온 stat 을 버리지 않는다**([198]). 예전에는 `glob` 로 이름만 받아
      온 뒤 `cands.sort(key=os.path.getmtime)` 로 **파일마다 Z: 를 다시 찔렀다.**
      실측 2026-08-24: glob 14.2초 + **정렬 124.4초**(xlsx 170개 · 파일당 731.9ms).
      `os.scandir` 항목의 stat 은 목록을 받을 때 딸려 오므로 그 왕복이 통째로 없어진다.
      이것이 관문 `[7] 쿠팡 PO 대조` 가 3분 ↔ 121분을 오가던 큰 조각이다(분담판 [211]).
    ★ **`skip_dirs=()` 는 일부러 비운 것이다.** 공용 워커의 기본값은 *색인의* 목록
      (`_보관`·`_바로가기`)인데, 그것을 말없이 물려받으면 거기 든 판매조회가
      **조용히 빠지면서 오류도 안 난다**([198] 의 ⚠ · [165]). glob 은 아무것도 안
      걸렀으므로 여기서도 안 거른다 — **결과가 한 톨도 바뀌면 안 된다.**
    ★ 공용 워커가 없어도 죽지 않는다 — 예전 길로 간다(느릴 뿐 답은 같다).
    """
    import source_dirs as S
    inventory = None
    try:
        from inbox_scan import cached_inventory, scan
        inventory = cached_inventory([S.ERP_DIR], max_age_s=366 * 86400)
        if inventory is None:
            raise LookupError("최근 분류표 없음")
        pairs_by_path = {
            _path_key(row["path"]): (row["mtime"], row["path"])
            for row in inventory
            if row["path"].lower().endswith(".xlsx")
            and not os.path.basename(row["path"]).startswith(("~$", "ESD007E"))
        }
        # 오늘 막 들어와 아직 분류표 전체 목록에 없던 파일만 실제 stat 한다.
        for path, _kind in scan(S.ERP_DIR):
            key = _path_key(path)
            if (key in pairs_by_path or not path.lower().endswith(".xlsx")
                    or os.path.basename(path).startswith(("~$", "ESD007E"))):
                continue
            try:
                pairs_by_path[key] = (os.stat(path).st_mtime, path)
            except OSError:
                continue
        pairs = list(pairs_by_path.values())
        pairs.sort(key=lambda x: x[0], reverse=True)
        cands = [p for _, p in pairs]
    except Exception:
        try:
            from source_index import walk_stat
            pairs = [(st.st_mtime, os.path.join(dp, fn))
                     for dp, fn, st in walk_stat(S.ERP_DIR, skip_dirs=())
                     if fn.lower().endswith(".xlsx")
                     and not fn.startswith(("~$", "ESD007E"))]
            pairs.sort(key=lambda x: x[0], reverse=True)
            cands = [p for _, p in pairs]
        except Exception:
            cands = [p for p in glob.glob(os.path.join(S.ERP_DIR, "**", "*.xlsx"),
                                          recursive=True)
                     if not os.path.basename(p).startswith(("~$", "ESD007E"))]
            cands.sort(key=os.path.getmtime, reverse=True)
    cached_sales = [row["path"] for row in (inventory or []) if row["kind"] == "sales"]
    if not cached_sales:
        try:
            cache_path = os.path.join(ROOT, "reports", "inbox_classify.json")
            with open(cache_path, encoding="utf-8") as f:
                cache = json.load(f)
            cached_sales = [path for path, value in cache.items()
                            if isinstance(value, list) and len(value) >= 4
                            and value[3] == "sales"]
        except Exception:
            pass
    return prioritize_sales_candidates(cands, cached_sales, limit)


def sales_exports(limit=60, cands=None):
    """판매조회 엑셀을 **전부** 찾는다(새 것부터). 파일명이 무작위라 머리글로 판정한다.

    ★ 2026-08-05 — 예전에는 '가장 최근 것 하나'만 읽었다. 그런데 2025 자료를 받으려고
      2025 판매조회를 내려받는 순간 그것이 '가장 최근 파일'이 되어 색인이 통째로
      2025 로 바뀌었고, 2026 명세서 793건이 **전부 짝 없음**이 됐다(실측).
      한 해를 받으면 다른 해가 사라지는 구조 자체가 틀렸다 — 모두 읽어 합친다.
    """
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    cands = list(cands) if cands is not None else sales_candidate_paths(limit)
    found = []
    for p in cands:
        try:
            wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
            ws = wb.active
            head = [str(c or "") for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
            if "프로젝트코드코드" in "|".join(head) and "진행상태" in "|".join(head):
                found.append((p, ws, head, wb))
            else:
                wb.close()
        except Exception:
            continue
    return found


def build_one(ws, head):
    """엑셀 한 개 → {UJ: 집계}. 한 파일 안에서만 합산한다(파일 간 합산은 중복이다)."""
    idx = {h: i for i, h in enumerate(head)}

    def num(r, key):
        s = (r[idx[key]] if key in idx and len(r) > idx[key] else "").replace(",", "")
        return int(float(s)) if re.fullmatch(r"-?\d+(\.\d+)?", s) else 0

    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        r = ["" if c is None else str(c).strip() for c in r]
        m = UJ.search(" ".join(r))
        if not m:
            continue
        uj = m.group(0)
        cur = out.setdefault(uj, {"supply": 0, "vat": 0, "total": 0, "state": "",
                                  "date": "", "po": "", "cust": "", "rows": 0})
        cur["supply"] += num(r, "공급가액합계")
        cur["vat"] += num(r, "부가세합계")
        cur["total"] += num(r, "금액합계")
        cur["rows"] += 1
        st = r[idx["진행상태"]] if "진행상태" in idx else ""
        if not cur["state"] or rank(st) < rank(cur["state"]):
            cur["state"] = st
        if not cur["date"]:
            cur["date"] = (r[0] or "")[:10]
        for k, col in (("po", "PO번호"), ("cust", "거래처명")):
            if not cur[k] and col in idx and len(r) > idx[col]:
                cur[k] = r[idx[col]]
    return out


def _candidate_stamp(paths):
    """분류표에 적힌 불변 원본 서명으로 판매 색인의 정확한 입력판을 만든다."""
    try:
        from inbox_scan import _CLS_FILE, RULES_VERSION
        with open(_CLS_FILE, encoding="utf-8") as fh:
            cache = json.load(fh)
        rows = []
        for path in paths:
            row = cache.get(path)
            if not isinstance(row, list) or len(row) < 4:
                return ""
            rows.append([_path_key(path), row[0], row[1], row[2], row[3]])
        raw = json.dumps([INDEX_RULES_VERSION, RULES_VERSION, rows], ensure_ascii=False,
                         separators=(",", ":")).encode("utf-8")
        return hashlib.sha256(raw).hexdigest()
    except Exception:
        return ""


def _save_index(payload):
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    tmp = OUT + ".%s.tmp" % os.getpid()
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, OUT)


def build():
    """여러 판매조회를 합친다. **같은 UJ 는 새 파일 것이 이긴다** — 더하지 않는다.
    (더하면 두 내보내기가 겹치는 기간에서 금액이 두 배가 된다)"""
    paths = sales_candidate_paths()
    stamp = _candidate_stamp(paths)
    if stamp:
        try:
            with open(OUT, encoding="utf-8") as fh:
                old = json.load(fh)
            if (old.get("fingerprint") == stamp and isinstance(old.get("index"), dict)
                    and isinstance(old.get("src"), list)):
                return old["index"], old["src"]
        except Exception:
            pass

    merged, srcs = {}, []
    found = sales_exports(cands=paths)
    for path, ws, head, wb in found:                 # 새 것 → 옛 것 순서
        try:
            one = build_one(ws, head)
        finally:
            wb.close()
        if not one:
            continue
        srcs.append(os.path.basename(path))
        for uj, v in one.items():
            merged.setdefault(uj, v)                 # 이미 있으면(=더 새 파일) 그대로 둔다
    _save_index({"src": srcs, "count": len(merged), "index": merged,
                 "fingerprint": stamp, "rules": INDEX_RULES_VERSION})
    return merged, srcs


def main():
    idx, srcs = build()
    issued = sum(1 for v in idx.values() if str(v["state"]).startswith(ISSUED))
    years = {}
    for v in idx.values():
        years[str(v.get("date", ""))[:4]] = years.get(str(v.get("date", ""))[:4], 0) + 1
    print(f"ERP 판매 색인 {len(idx)}개 프로젝트 (원본 {len(srcs)}개) · 발행 이상 단계 {issued} · "
          + " ".join(f"{y}:{n}" for y, n in sorted(years.items()) if y)
          + " → reports/ERP판매_프로젝트색인.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
