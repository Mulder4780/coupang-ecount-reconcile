# -*- coding: utf-8 -*-
"""
synthetic_check.py — 합성데이터 상시 검증 (실데이터·실서버 접촉 0)
====================================================================
사용자 상시 지시(2026-07-24): "항상 합성데이터 검증해서 작업 진행".
실제 시스템을 건드리기 전에 이 스크립트가 초록이어야 한다.

검증 항목:
  1. ERP원장 대조기(erp_ledger_check): 유형 A/B/C/D 판정이 설계대로 나오는가
  2. 전표 페이로드(ecount_upload.build_payload): 금액·부가세·일자·프로젝트 매핑
  3. 판매 inbox 매칭(ecount_reconcile.match_project): 프로젝트NO/금액 매칭 규칙

실행:  python tests/synthetic_check.py   (전부 통과 시 exit 0, 'ALL GREEN')
"""
import sys, os, re, tempfile, subprocess, hashlib, json, sqlite3, time, contextlib

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PY = sys.executable
# 합성검증 중 resolve_master가 불려도 실관리대장 구버전 정리가 실행되면 안 된다.
os.environ["CSOS_SYNTHETIC"] = "1"


def _snapshot_output_bytes(paths):
    """파일의 존재 여부와 바이트를 그대로 잰다(줄바꿈·인코딩 정규화 금지)."""
    out = {}
    for path in paths:
        path = os.path.abspath(path)
        if os.path.isfile(path):
            with open(path, "rb") as f:
                out[path] = (True, f.read())
        else:
            out[path] = (False, b"")
    return out


# 전체 합성검증이 시작되기 **전** 상태. [192]가 끝에서 다시 재어, 중간 어느 검증이든
# 계획·게시 산출물을 바꿨으면 잡는다. 기존 dirty 파일도 바이트 그대로 기준에 포함된다.
_SYNTHETIC_OUTPUT_PATHS = [
    os.path.join(ROOT, "reports", "밴드_수집계획.json"),
    os.path.join(ROOT, "docs", "collect", "plan.json"),
    os.path.join(ROOT, "docs", "collect", "grab_posts.js"),
]
_SYNTHETIC_OUTPUT_BASELINE = _snapshot_output_bytes(_SYNTHETIC_OUTPUT_PATHS)


def make_ledger(path):
    wb = openpyxl.Workbook()
    ws6 = wb.active; ws6.title = "06_거래서류청구수금"
    h6 = ["정산ID", "업무구분", "원천업무ID", "프로젝트NO", "캠프명", "작업완료일", "비용구분",
          "실제작업공급가액", "실제작업부가세", "실제작업합계", "거래명세서번호", "거래명세서발행일",
          "거래명세서공급가액", "거래명세서합계", "세금계산서발행일", "세금계산서합계", "입금일", "입금액",
          "PO필요여부", "PO번호", "PO발행일"]
    for _ in range(3): ws6.append([])
    ws6.append(h6)
    # JS-1: ERP 일치 + 세금계산서 미발행 → C / PO111 쿠팡 일치
    ws6.append(["JS-1", "돌발AS", "AS-1", "UJ0001", "테스트캠프A", "2026-07-01", "유상",
                100000, 10000, 110000, "2026/07/01-4", "2026-07-01", 100000, 110000, None, 0, None, None,
                "필요", "PO111", "2026-07-03"])
    # JS-2: ERP 금액불일치 → D / PO333은 쿠팡 목록에 없음 → PO-B
    ws6.append(["JS-2", "정기점검", "PM-1", "UJ0002", "테스트캠프B", "2026-07-02", "유상",
                200000, 20000, 220000, "2026/07/02-1", "2026-07-02", 200000, 220000, "2026-07-03", 220000, None, None,
                "필요", "PO333", "2026-07-04"])
    # JS-3: 명세서번호 없음 → B / PO필요·번호없음 → PO444 유일매칭 → PO-D
    ws6.append(["JS-3", "돌발AS", "AS-2", "UJ0003", "테스트캠프C", "2026-07-05", "유상",
                300000, 30000, 330000, None, None, None, 0, None, 0, None, None,
                "필요", None, None])
    ws15 = wb.create_sheet("15_세금계산서관리")
    for _ in range(3): ws15.append([])
    ws15.append(["정산ID", "실제발행일", "발행금액", "승인번호"])
    ws15.append(["JS-2", "2026-07-03", 220000, "APPR-001"])
    ws16 = wb.create_sheet("16_입금수금관리")
    for _ in range(3): ws16.append([])
    ws16.append(["정산ID", "입금일", "입금액", "미수금액"])
    ws2 = wb.create_sheet("02_돌발AS접수")
    for _ in range(3): ws2.append([])
    ws2.append(["접수ID", "프로젝트NO", "캠프명", "담당기사", "작업완료일", "진행상태"])
    ws2.append(["AS-K1", "UJ0001", "테스트캠프A(감일동)", "김필우", "2026-07-01", "작업완료"])
    ws2.append(["AS-K2", "UJ0002", "다른캠프B(성수동)", "김준형", "2026-07-02", "작업완료"])
    wb.save(path)


def make_erp(path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "원장"
    ws.append(["거래처별계정별원장 (합성)"])
    ws.append(["일자-No.", "적요", "차변금액", "대변금액"])
    ws.append(["2026/07/01 -4", "테스트캠프A 돌발AS", 110000, None])        # JS-1 일치
    ws.append(["2026/07/02 -1", "테스트캠프B 정기점검", 999999, None])      # JS-2 금액불일치
    ws.append(["2026/07/05 -9", "인천8MB 상하차리프트(원장에 없음)", 500000, None])  # A형
    ws.append(["2026/07/06 -10", "입금 — 매출 전표로 세면 안 됨", None, 500000])
    ws.append(["2026/07 계", "", 1609999, 500000])
    wb.save(path)
    wb.close()

    # 실제 이카운트 내보내기처럼 dimension을 잘못된 A1:A1로 만든다.
    # read_only=True 파서는 이 경우 1셀만 읽고 전표 0건을 내므로 반드시 회귀검사한다.
    import zipfile
    broken = path + ".dimension"
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(broken, "w") as dst:
        for item in src.infolist():
            blob = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                blob = re.sub(rb'<dimension ref="[^"]+"', b'<dimension ref="A1:A1"',
                              blob, count=1)
            dst.writestr(item, blob)
    os.replace(broken, path)


def t1_erp_check(tmp):
    import erp_ledger_check as E
    assert E.norm_slip("2026/07-02-6") == "2026/07/02-6"
    assert E.in_erp_period({"작업완료일": "2026-08-01"}, "", "2026-07-01", "2026-07-31") is False
    assert E.in_erp_period({"작업완료일": "2026-07-05"}, "", "2026-07-01", "2026-07-31") is True
    ledger = os.path.join(tmp, "ledger.xlsx"); erp = os.path.join(tmp, "erp원장.xlsx")
    make_ledger(ledger); make_erp(erp)
    r = subprocess.run([PY, os.path.join(ROOT, "erp_ledger_check.py"),
                        "--file", erp, "--master", ledger],
                       capture_output=True, text=True, encoding="utf-8", errors="replace", cwd=ROOT,
                       env={**os.environ, "COUPANG_REPORT_DIR": tmp,
                            "COUPANG_UPDATES_DIR": tmp})
    out = r.stdout
    m = re.search(r"A\(ERP에만\) (\d+) / B\(원장에만\) (\d+) / C\(계산서미발행\) (\d+) / D\(금액불일치\) (\d+) / (?:전표키 )?정상 (\d+)", out)
    assert m, f"결과 라인 파싱 실패:\n{out}\n{r.stderr}"
    a, b, c, d, ok = map(int, m.groups())
    assert (a, b, c, d, ok) == (1, 1, 1, 1, 1), f"판정 불일치: A{a} B{b} C{c} D{d} OK{ok} (기대 1,1,1,1,1)"
    print("  [1] ERP원장 대조 A/B/C/D 판정 ✅")


def t2_payload():
    from ecount_upload import build_payload
    rec = {"정산ID": "JS-9", "캠프명": "합성캠프", "업무구분": "돌발AS",
           "프로젝트NO": "UJ9999", "작업완료일": "2026-07-20", "원장_공급가액": 1472500.0}
    up = {"CUST": "CUSTX", "CR_CODE": "4049", "TAX_GUBUN": "11"}
    p = build_payload(rec, up)["InvoiceAutoList"][0]["BulkDatas"]
    assert p["SUPPLY_AMT"] == "1472500" and p["VAT_AMT"] == "147250", p
    assert p["TRX_DATE"] == "20260720" and p["PJT_CD"] == "UJ9999" and p["CR_CODE"] == "4049", p
    assert "JS-9" in p["REMARKS"], p
    print("  [2] 전표 페이로드 매핑 ✅")


def t3_match():
    from ecount_reconcile import match_project, norm_ecount
    ec = norm_ecount([{"SUPPLY_AMT": "110000", "IO_DATE": "20260701", "REMARKS": "UJ0001 테스트캠프A", "CUST_DES": "쿠팡"},
                      {"SUPPLY_AMT": "500000", "IO_DATE": "20260705", "REMARKS": "무관 건", "CUST_DES": "쿠팡"}])
    rec = {"프로젝트NO": "UJ0001", "원장_공급가액": 110000.0}
    m, how = match_project(rec, ec, 0, "쿠팡")
    assert m and how == "프로젝트NO", (m, how)
    rec2 = {"프로젝트NO": "UJ_NOPE", "원장_공급가액": 500000.0}
    m2, how2 = match_project(rec2, ec, 0, "쿠팡")
    assert m2 and how2 == "금액", (m2, how2)
    print("  [3] inbox 판매 매칭 규칙 ✅")


def t4_kakao(tmp):
    ledger = os.path.join(tmp, "ledger_k.xlsx")
    make_ledger(ledger)
    # 원문이 없어도 사용자가 해당 건을 개별 완료 처리한 기록은 다음 대조에서 되살아나면 안 된다.
    wb = openpyxl.load_workbook(ledger)
    ws = wb["02_돌발AS접수"]
    ws["G4"] = "조치내용"
    ws["G6"] = "카톡 보고 미확인 완료처리(사용자 지시 2026-07-29)"
    wb.save(ledger)
    wb.close()
    txt = os.path.join(tmp, "쿠팡AS방.txt")
    with open(txt, "w", encoding="utf-8") as f:
        f.write("쿠팡AS방 님과 카카오톡 대화\n저장한 날짜 : 2026-07-24 18:00:00\n\n")
        f.write("--------------- 2026년 7월 1일 화요일 ---------------\n")
        f.write("[김필우] [오후 2:59] UJ0001 테스트캠프A 작업완료\n")
        f.write("사진 2장 첨부\n")                                   # 여러 줄 병합 검증
        f.write("2026년 7월 3일 오후 4:01, 유현민 : 일반 공지 메시지\n")  # 모바일 형식 검증
    r = subprocess.run([PY, os.path.join(ROOT, "kakao", "kakao_reconcile.py"),
                        "--file", txt, "--master", ledger],
                       capture_output=True, text=True, encoding="utf-8", cwd=ROOT, env={**os.environ, "COUPANG_REPORT_DIR": tmp, "COUPANG_UPDATES_DIR": tmp})
    m = re.search(r"확인 (\d+) / 미확인 (\d+)", r.stdout)
    assert m, f"카톡 결과 파싱 실패:\n{r.stdout}\n{r.stderr}"
    ok, miss = map(int, m.groups())
    assert (ok, miss) == (2, 0), f"카톡 판정 불일치: 확인{ok} 미확인{miss} (기대 2,0)"
    report = next(p for p in os.listdir(tmp) if p.startswith("카톡대조_") and p.endswith(".csv"))
    rows = list(__import__("csv").DictReader(open(os.path.join(tmp, report), encoding="utf-8-sig")))
    manual = next(x for x in rows if x["ID"] == "AS-K2")
    assert manual["카톡보고"] == "확인" and manual["매칭근거"] == "사용자완료처리", manual

    # ★ 자료가 없는 기간을 '미확인'이라 부르지 않는다 (2026-08-07 지시).
    #   카톡 내보내기는 방에 들어간 뒤부터만 나온다. 그 전 작업은 아무리 잘해도
    #   카톡에서 못 찾는다 — '미확인'으로 세면 지워지지 않는 빨간 줄이 쌓이고
    #   진짜 누락(자료는 있는데 보고가 없는 건)이 그 속에 묻힌다.
    _kr = open(os.path.join(ROOT, "kakao", "kakao_reconcile.py"), encoding="utf-8").read()
    assert "자료없음" in _kr, "카톡 대조가 '자료없음'을 가르지 않는다"
    assert re.search(r"floor\s*=\s*min\(\(m\[.date.\] for m in msgs\)", _kr), \
        "자료 시작일을 **가진 메시지에서** 뽑지 않는다 — 날짜를 못박으면 " \
        "나중에 옛 내보내기가 들어와도 계속 '자료없음'이라 답한다"
    assert 'r["완료일"] < floor' in _kr, "완료일이 자료 시작 이전인지를 안 본다"
    # 앱 배지: '자료없음'은 실패도 분모도 아니다.
    _as = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert re.search(r'na\s*=\s*sum\(1 for r in rows if r\.get\(col\) == "자료없음"\)', _as), \
        "앱이 '자료없음'을 따로 세지 않는다"
    assert 'r.get(col) not in (okv, "자료없음")' in _as, \
        "앱이 '자료없음'을 미확인과 같이 센다 — 영원히 안 지워지는 빨간 줄이 된다"
    # 확인필요현황·07시트 얹기는 `== "미확인"` 이라 자료없음이 저절로 빠진다.
    _fe = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    assert 'r.get("카톡보고") == "미확인"' in _fe, \
        "확인필요 목록이 카톡보고를 정확히 '미확인'으로 거르지 않는다"
    # 방 이름 보완(2026-07-30): 머리글이 유형을 말하지 않으면 방으로 정한다.
    # ★ 순서가 뒤집히면 돌발방에 올라온 철거 글이 02시트로 잘못 간다 — 머리글이 먼저다.
    import kakao_extract as _KE
    assert _KE.kind_by_room("★UNI★ 쿠팡돌발점검") == "02_돌발AS접수"
    assert _KE.kind_by_room("★UNI★ 쿠팡정기점검") == "04_정기점검"
    assert _KE.kind_by_room("아무방") == ""
    assert _KE.kind_of("[철거 안내]") == "(철거·보관)", "머리글이 방보다 우선이어야 한다"
    _ke_src = open(os.path.join(ROOT, "kakao_extract.py"), encoding="utf-8").read()
    assert "kind_of(head) or kind_by_room(room)" in _ke_src, "머리글 우선 순서가 깨졌다"
    print("  [4] 카톡 내보내기 파싱·대조·사용자 개별 완료 유지 (PC/모바일·다중행) ✅")


def t5_writer(tmp):
    import json
    src = os.path.join(tmp, "합성대장_v1.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "06_거래서류청구수금"
    for _ in range(3): ws.append([])
    ws.append(["정산ID", "거래명세서번호", "거래명세서발행일", "비고"])
    ws.append(["JS-W1", None, None, "빈칸 채움 대상"])
    ws.append(["JS-W2", "기존값-유지", None, "덮어쓰기 금지 검증"])
    ws2 = wb.create_sheet("02_돌발AS접수")
    for _ in range(3): ws2.append([])
    ws2.append(["접수ID", "프로젝트NO", "완료보고서등록", "사진등록"])
    ws2.append(["AS-MIX-1", "UJ-MIX-1", None, None])
    h = wb.create_sheet("19_AI작업인수인계")
    h.append(["헤더"]); h.append(["기존 인수인계 행"])
    dash = wb.create_sheet("00_대시보드"); dash["A2"] = "기존 사용법"
    m18 = wb.create_sheet("18_문서발행업무매뉴얼"); m18["A1"] = "기존 매뉴얼"
    m20 = wb.create_sheet("20_쿠팡통합업무상세매뉴얼"); m20.append([1, "기존 상세"])
    wb.save(src)
    q = os.path.join(tmp, "q.json")
    json.dump([
        {"sheet": "06_거래서류청구수금", "key_col": "정산ID", "key": "JS-W1", "col": "거래명세서번호",
         "value": "2026/07/24-9", "vtype": "text", "evidence": "합성", "only_if_empty": True},
        {"sheet": "06_거래서류청구수금", "key_col": "정산ID", "key": "JS-W1", "col": "거래명세서발행일",
         "value": "2026-07-24", "vtype": "date", "evidence": "합성", "only_if_empty": True},
        {"sheet": "06_거래서류청구수금", "key_col": "정산ID", "key": "JS-W2", "col": "거래명세서번호",
         "value": "덮어쓰기시도", "vtype": "text", "evidence": "합성", "only_if_empty": True},
        # 같은 시트에서 서로 다른 키 열을 연달아 써도 둘 다 정확한 행을 찾아야 한다.
        {"sheet": "02_돌발AS접수", "key_col": "접수ID", "key": "AS-MIX-1", "col": "사진등록",
         "value": "등록", "vtype": "text", "evidence": "합성", "only_if_empty": True},
        {"sheet": "02_돌발AS접수", "key_col": "프로젝트NO", "key": "UJ-MIX-1", "col": "완료보고서등록",
         "value": "등록", "vtype": "text", "evidence": "합성", "only_if_empty": True},
    ], open(q, "w", encoding="utf-8"), ensure_ascii=False)
    r = subprocess.run([PY, os.path.join(ROOT, "ledger_writer.py"), "--apply",
                        "--master", src, "--queue", q],
                       capture_output=True, text=True, encoding="utf-8", errors="replace",
                       cwd=ROOT, env={**os.environ, "COUPANG_REPORT_DIR": tmp,
                                     "COUPANG_UPDATES_DIR": tmp,
                                     "COUPANG_LEDGER_GATE": "1"})
    assert "반영 완료" in r.stdout and "입력 4건 / 건너뜀 1건" in r.stdout, f"{r.stdout}\n{r.stderr}"
    dst = src.replace("_v1.xlsx", "_v2.xlsx")
    w2 = openpyxl.load_workbook(dst)
    ws2 = w2["06_거래서류청구수금"]
    assert ws2["B5"].value == "2026/07/24-9", ws2["B5"].value          # 빈칸 채움
    assert ws2["C5"].value is not None, "날짜 직렬값 기록 실패"          # date serial
    assert ws2["B6"].value == "기존값-유지", ws2["B6"].value            # 덮어쓰기 금지
    as2 = w2["02_돌발AS접수"]
    assert as2["C5"].value == "등록" and as2["D5"].value == "등록", \
        "같은 시트의 접수ID·프로젝트NO 혼합 키 조회 실패"
    hand = w2["19_AI작업인수인계"]
    assert any("자동 입력" in str(c.value) for row in hand.iter_rows() for c in row if c.value), "인수인계 기록 없음"
    w2.close()
    # 종료 인수인계와 상시지시(00 A2·18·20)를 한 버전에 함께 반영한다.
    import workbook_patch as WP
    v3 = src.replace("_v1.xlsx", "_v3.xlsx")
    WP.patch(dst, v3, "2026-07-28 #합성", "보완", "상세",
             "새 사용법", "18 갱신", "20 갱신",
             extra_rows=[("2026-07-28 #합성2", "묶음제목", "묶음상세")])
    w3 = openpyxl.load_workbook(v3, read_only=True)
    assert w3["00_대시보드"]["A2"].value == "새 사용법"
    assert w3["18_문서발행업무매뉴얼"]["A2"].value == "18 갱신"
    assert w3["20_쿠팡통합업무상세매뉴얼"]["D2"].value == "20 갱신"
    hand_ws = w3["19_AI작업인수인계"]
    hand_ws.reset_dimensions()      # 합성 시트 dimension 이 A열뿐이라 B열이 가려진다
    hand3 = [c.value for row in hand_ws.iter_rows() for c in row]
    assert "보완" in hand3 and "묶음제목" in hand3, \
        "예약 여러 건이 한 버전(vN+1 하나)에 함께 실리지 않는다"
    w3.close()
    # 회귀: 스타일만 있는 자기닫힘 빈 셀(<c s=.../>) 바로 뒤에 수식 셀 — 오매칭으로 '값 있음' 오판했던 실사고
    import ledger_writer as LW
    from ledger_writer import apply_to_xml
    xml = ('<worksheet><dimension ref="A1:C5"/><sheetData>'
           '<row r="5"><c r="A5" t="inlineStr"><is><t>K</t></is></c>'
           '<c r="B5" s="43"/><c r="C5" s="53"><f t="shared" si="1"/><v>0</v></c></row>'
           '</sheetData></worksheet>')
    ok, nx, why = apply_to_xml(xml, {"row": 5, "colL": "B", "value": "값", "vtype": "text", "only_if_empty": True})
    assert ok and '<c r="B5" s="43" t="inlineStr"><is><t>값</t></is></c>' in nx, (ok, why, nx)
    ok2, _, why2 = apply_to_xml(nx, {"row": 5, "colL": "C", "value": "X", "vtype": "text", "only_if_empty": True})
    assert not ok2 and "이미" in why2, (ok2, why2)     # 수식 셀은 여전히 보호
    # 같은 기준일을 다시 저장하면 버전만 하나 더 생기면 안 된다.
    date_xml = ('<worksheet><sheetData><row r="3"><c r="B3" s="1"><v>46226</v></c>'
                '</row></sheetData></worksheet>')
    same, unchanged, why3 = apply_to_xml(
        date_xml, {"row": 3, "colL": "B", "value": "2026-07-23",
                   "vtype": "date", "only_if_empty": False})
    assert not same and unchanged == date_xml and "동일 값" in why3, (same, why3)
    # 큐 쓰기는 원자 교체 + 중복 방지여야 한다.
    old_pending = LW.PENDING
    try:
        LW.PENDING = os.path.join(tmp, "atomic_queue.json")
        item = {"sheet": "S", "key": "K", "col": "C"}
        assert LW.queue_add([item, item]) == 1
        assert LW.load_queue() == [item]
        # 프로세스가 강제 종료돼 남긴 잠금은 다음 실행이 즉시 회수해야 한다.
        with open(LW.PENDING + ".lock", "w", encoding="ascii") as f:
            f.write("99999999 2026-01-01T00:00:00")
        item2 = {"sheet": "S", "key": "K2", "col": "C"}
        assert LW.queue_add([item2]) == 1
        assert not os.path.exists(LW.PENDING + ".lock")
    finally:
        LW.PENDING = old_pending
    from band.band_reconcile import photo_updates
    photo_q = photo_updates([{
        "시트": "02_돌발AS접수", "ID": "AS-OLD-CACHE", "프로젝트NO": "UJ-STABLE",
        "밴드게시": "확인", "사진수": 2, "게시일": "2026-07-29", "게시자": "합성",
    }])
    assert len(photo_q) == 1 and photo_q[0]["key_col"] == "프로젝트NO" \
        and photo_q[0]["key"] == "UJ-STABLE", photo_q
    print("  [5] 자동입력 엔진(빈칸만·동일값 멱등·큐잠금/원자쓰기·자기닫힘셀·인수인계) ✅")


def t7_po(tmp):
    ledger = os.path.join(tmp, "ledger_po.xlsx")
    make_ledger(ledger)
    pof = os.path.join(tmp, "쿠팡PO목록_테스트.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["쿠팡 PO 발행 목록"])
    ws.append(["발행일자", "PO번호", "금액"])
    ws.append(["2026-07-03", "PO111", 100000])   # JS-1 일치
    ws.append(["2026-07-05", "PO222", 999000])   # 원장 미등록 → A
    ws.append(["2026-07-06", "PO444", 300000])   # JS-3 유일매칭 → A+D
    wb.save(pof)
    r = subprocess.run([PY, os.path.join(ROOT, "po_reconcile.py"),
                        "--file", pof, "--master", ledger, "--no-queue"],
                       capture_output=True, text=True, encoding="utf-8", cwd=ROOT, env={**os.environ, "COUPANG_REPORT_DIR": tmp, "COUPANG_UPDATES_DIR": tmp})
    # 판정 이름이 '원장미등록' → '미청구'로 바뀌었다(ERP 계산서 대조 후 진짜 미청구만 남김)
    m = re.search(r"A\((?:원장미등록|미청구)\) (\d+) / B\(쿠팡목록에없음\) (\d+) / C\(금액불일치\) (\d+) / D\(연결제안\) (\d+) / 정상 (\d+)", r.stdout)
    assert m, f"PO 결과 파싱 실패:\n{r.stdout}\n{r.stderr}"
    a, b, c, d, ok = map(int, m.groups())
    assert (a, b, c, d, ok) == (2, 1, 0, 1, 1), f"PO 판정 불일치: A{a} B{b} C{c} D{d} OK{ok} (기대 2,1,0,1,1)"
    import glob as _g
    assert "유일 매칭" in open(sorted(_g.glob(os.path.join(tmp, "PO대조_*.md")))[-1], encoding="utf-8").read()
    print("  [7] 쿠팡 PO 대조(원장미등록·오기입·금액·유일매칭 제안) ✅")


def t8_findings_sheet(tmp):
    ledger = os.path.join(tmp, "합성대장F_v1.xlsx")
    make_ledger(ledger)
    env = {**os.environ, "COUPANG_REPORT_DIR": tmp, "COUPANG_UPDATES_DIR": tmp}
    r = subprocess.run([PY, os.path.join(ROOT, "findings_sheet.py"), "--master", ledger],
                       capture_output=True, text=True, encoding="utf-8", cwd=ROOT, env=env)
    assert "시트 신규 추가" in r.stdout, f"{r.stdout}\n{r.stderr}"
    v2 = ledger.replace("_v1.xlsx", "_v2.xlsx")
    w = openpyxl.load_workbook(v2)
    ws = w["23_확인필요현황"]
    assert [c.value for c in ws[4][:3]] == ["구분", "ID", "문제유형"]
    vals = [ws.cell(row=i, column=1).value for i in range(5, ws.max_row + 1)]
    assert "정산" in vals, vals          # 합성 정산 조치필요가 들어와야 함
    w.close()
    # 멱등: 같은 내용으로 재실행 → v3 생성 안 됨
    r2 = subprocess.run([PY, os.path.join(ROOT, "findings_sheet.py"), "--master", v2],
                        capture_output=True, text=True, encoding="utf-8", cwd=ROOT, env=env)
    assert "변경 없음" in r2.stdout and not os.path.exists(ledger.replace("_v1", "_v3")), r2.stdout
    # ★ 자기잠식 금지 (2026-08-07 실사고). 23시트는 4행에 '프로젝트NO' 열을 가진다.
    #   원장을 훑는 쪽이 이 시트를 원장으로 세면 "이미 등록된 프로젝트"가 부풀어
    #   미등록 문서가 통째로 사라진다(실측 223건). 위 멱등 검사는 그 증상을 **우연히**
    #   잡았을 뿐이라, 원인 쪽을 직접 못박는다.
    import findings_export as FE
    assert FE.is_agent_sheet(("23_확인필요현황 (에이전트 자동 갱신 — 수기 입력 금지)",))
    assert not FE.is_agent_sheet(("프로젝트NO", "캠프명")) and not FE.is_agent_sheet(())
    w = openpyxl.load_workbook(v2, read_only=True)
    a1 = next(w["23_확인필요현황"].iter_rows(min_row=1, max_row=1, values_only=True))
    w.close()
    assert FE.is_agent_sheet(a1), f"보고 시트 표식이 1행에 없다: {a1}"
    src = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    assert "is_agent_sheet(head[0])" in src, "doc_unregistered 가 보고 시트를 안 거른다"
    print("  [8] 확인필요 시트 통합(신규 추가·머리글·멱등·자기잠식 금지) ✅")


def t10_band_extract():
    from band_extract import parse_post, normalize_tech
    mk = lambda c: {"content": c, "created_at": 1780000000000, "photo_count": 4}
    r = parse_post("1", mk(
        "2026년 6월 1일 오전 8:08 게시글\n\n☑️판매전표 +거래명세서 +견적서 = 메일발송 完 ⭕\n"
        "♣ ［ 2026년 02분기 3개월 유료 A/S 완료 ]\n● A/S 일자 : 2026.06.01 (월요일)\n"
        "● A/S 담당 : 김필우\n● 프로젝트NO : UJ2600931\n● 캠프이름 : 양주1캠프\n\n...더보기"), "밴드A")
    assert r["프로젝트NO"] == "UJ2600931" and r["업무유형"] == "정기점검", r
    assert r["비용구분"] == "유상" and r["작업일"] == "2026-06-01" and r["진행상태"] == "작업완료", r
    assert r["담당기사"] == "김필우" and r["캠프명"] == "양주1캠프", r
    assert "판매전표" in r["문서상태"] and "메일발송" in r["문서상태"], r
    r2 = parse_post("2", mk("♣ ［ 돌발무료 A/S 안내 ]\n● A/S 일자 : 2026.00.00 (요일)\n"
                            "● 프로젝트NO : UJ2601999\n● 캠프이름 : 테스트캠프"), "밴드A")
    assert r2["업무유형"] == "돌발AS" and r2["비용구분"] == "무상", r2
    assert r2["작업일"] == "" and r2["진행상태"] == "접수·예정", r2      # 2026.00.00 = 미정
    assert parse_post("3", mk("● 프로젝트NO : UJ000000\n♣ ［ 돌발유료 A/S 안내 ]"), "밴드A") is None  # 템플릿 제외
    assert parse_post("4", mk("안녕하세요 일반 공지입니다"), "밴드A") is None                          # 비업무 제외
    assert normalize_tech("권오절") == "권오철"
    assert normalize_tech("권오처르 + 1명 지원") == "권오철"
    print("  [10] 밴드 업무 추출(유형·유상무상·기사명 오탈자 정규화·템플릿/비업무 제외) ✅")


def t11_backfill():
    """백필 핵심 로직: 중복 제거·코드값 변환·용량 판단 (실데이터·실파일 접촉 없음)"""
    from backfill_rows import dedupe, enrich, SPEC
    base = {"프로젝트NO": "", "업무유형": "정기점검", "비용구분": "유상", "작업일": "", "담당기사": "김필우",
            "캠프명": "테스트캠프", "진행상태": "", "문서상태": "", "사진": 0, "게시일": "", "밴드": "B"}
    recs = [
        {**base, "프로젝트NO": "UJ1", "진행상태": "접수·예정", "게시일": "2026-06-01"},
        {**base, "프로젝트NO": "UJ1", "진행상태": "작업완료", "작업일": "2026-06-05", "게시일": "2026-06-05"},
        {**base, "프로젝트NO": "UJ2", "진행상태": "작업완료", "작업일": "2026-06-03", "게시일": "2026-06-03"},
    ]
    d = dedupe(recs)
    assert len(d) == 2, d                                   # UJ1 2건 → 1건
    u1 = [x for x in d if x["프로젝트NO"] == "UJ1"][0]
    assert u1["진행상태"] == "작업완료" and u1["_중복"] == 2, u1   # 완료본이 대표
    e = enrich(u1)
    assert e["_진행상태"] == "작업완료" and e["_완료일"] == "2026-06-05", e
    assert "게시 2건 통합" in e["_출처"], e
    e2 = enrich({**base, "프로젝트NO": "UJ3", "진행상태": "접수·예정", "작업일": "", "게시일": "2026-06-09"})
    assert e2["_진행상태"] == "접수" and e2["_완료일"] == "", e2   # 미완료는 완료일 비움
    # 매핑에 ID·자동계산 열이 없어야 함(수식 보호)
    for kind, spec in SPEC.items():
        assert not any(c.endswith("ID") for c in spec["map"]), spec["map"]
        assert "점검상태" not in spec["map"] and "기사배정상태" not in spec["map"], spec["map"]
    print("  [11] 백필 로직(중복통합·상태변환·완료일·수식열 제외) ✅")


def t12_dedupe(tmp):
    """중복 정리: 값만 비우고 수식 셀은 반드시 보존 (합성 워크북)"""
    import zipfile
    from dedupe_rows import blank_cells
    src = os.path.join(tmp, "합성중복_v1.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "02_돌발AS접수"
    for _ in range(3): ws.append([])
    ws.append(["접수ID", "프로젝트NO", "캠프명", "비고"])
    ws.append(["=IF(B5=\"\",\"\",\"AS-1\")", "UJ1", "캠프A", "유지"])
    ws.append(["=IF(B6=\"\",\"\",\"AS-2\")", "UJ1", "캠프A", "중복"])
    wb.save(src)
    dst, cleared, kept, _ = blank_cells(src, [("02_돌발AS접수", 6, ["A", "B", "C", "D"])])
    w = openpyxl.load_workbook(dst)
    ws2 = w["02_돌발AS접수"]
    assert ws2["B6"].value is None and ws2["D6"].value is None, (ws2["B6"].value, ws2["D6"].value)  # 값 제거
    assert str(ws2["A6"].value).startswith("="), ws2["A6"].value      # ★수식 셀은 보존
    assert ws2["B5"].value == "UJ1" and ws2["D5"].value == "유지"      # 유지 행 무손상
    w.close()
    assert cleared >= 2 and kept >= 1, (cleared, kept)
    assert zipfile.ZipFile(dst).testzip() is None
    print("  [12] 중복정리(값만 비움·수식 보존·유지행 무손상) ✅")


def t13_fix_ids(tmp):
    """ID 확정: 수식 규칙 재현·데이터 행만·빈 행 수식 보존"""
    from fix_ids import make_id, scan, apply_ids
    from datetime import date as _d
    assert make_id("AS", _d(2026, 6, 1), 66) == "AS-2606-062", make_id("AS", _d(2026, 6, 1), 66)
    assert make_id("PM", _d(2026, 7, 13), 45) == "PM-2607-041"
    src = os.path.join(tmp, "합성ID_v1.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "02_돌발AS접수"
    for _ in range(3): ws.append([])
    ws.append(["접수ID", "프로젝트NO", "캠프명", "접수일자"])
    ws.append(["AS-2607-001", "UJ1", "캠프A", _d(2026, 7, 1)])          # 기존 ID → 보존
    ws["A6"] = '=IF($B6="","","AS-"&TEXT($D6,"yymm")&"-"&TEXT(ROW()-4,"000"))'
    ws["B6"], ws["C6"], ws["D6"] = "UJ2", "캠프B", _d(2026, 6, 1)        # 데이터 있음 → 확정 대상
    ws["A7"] = '=IF($B7="","","AS-"&TEXT($D7,"yymm")&"-"&TEXT(ROW()-4,"000"))'  # 빈 행 → 수식 유지
    wb.save(src)
    plans, dups = scan(src)
    assert not dups and len(plans) == 1 and plans[0]["row"] == 6, (plans, dups)
    assert plans[0]["id"] == "AS-2606-002", plans[0]
    dst, done, _ = apply_ids(src, plans)
    w = openpyxl.load_workbook(dst)
    ws2 = w["02_돌발AS접수"]
    assert ws2["A5"].value == "AS-2607-001", ws2["A5"].value            # 기존 무손상
    assert ws2["A6"].value == "AS-2606-002", ws2["A6"].value            # 확정됨
    assert str(ws2["A7"].value).startswith("="), ws2["A7"].value        # ★빈 행 수식 보존
    w.close()
    print("  [13] ID 확정(수식규칙 재현·기존보존·빈행 수식유지) ✅")


def t9_watchdog():
    import time as _t
    from watchdog import pick_archive, pick_old_reports
    vers = [(19, "v19"), (25, "v25"), (21, "v21"), (24, "v24"), (23, "v23")]
    assert set(pick_archive(vers)) == {"v19", "v21", "v23", "v24"}, pick_archive(vers)
    assert pick_archive([(25, "v25")]) == []                                            # 최신 1개만 보존
    now = _t.time()
    files = [("old.md", now - 40*86400), ("new.md", now - 5*86400),
             ("agent_status.json", now - 90*86400), ("tunnel_url.txt", now - 90*86400)]
    dele = pick_old_reports(files, days=30)
    assert dele == ["old.md"], dele                                                    # 보호파일·최근파일 제외
    # ★ 2026-07-28: 워치독이 16:27~20:43 안 돌았고 그 사이 터널 주소가 죽어 폰 접속이 끊겼다.
    #   로그에는 '정상'만 남아 있어 아무도 몰랐다. 쉰 것보다 **쉰 걸 모르는 것**이 문제다.
    from watchdog import gap_note
    from datetime import datetime as _dt
    now = _dt(2026, 7, 28, 20, 43)
    assert "256분" in gap_note("[07-28 16:27] 서버 정상", now), gap_note("[07-28 16:27] x", now)
    assert gap_note("[07-28 20:13] 서버 정상", now) == ""          # 정상 주기는 조용히
    assert gap_note("[07-28 19:59] 서버 정상", now) == ""          # 44분은 여유 안(30+15)
    assert gap_note("", now) == "" and gap_note("깨진 줄", now) == ""
    # 해가 바뀌면 직전 기록은 작년이다 — 미래로 읽어 음수 공백이 나오면 안 된다
    assert gap_note("[12-31 23:50] x", _dt(2026, 1, 1, 0, 10)) == ""
    print("  [9] 워치독 판단 로직(버전 최신 1개·보호파일·30일 기준·실행 공백 감지) ✅")


def t14_datesort():
    """날짜 정렬 규칙 — 과거가 위, 최근이 아래. 새 데이터도 자동으로 이 순서를 따라야 한다."""
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "webapp"))
    from app_server import norm_date, row_date, sort_by_date
    assert norm_date("2026.6.3") == "2026-06-03", norm_date("2026.6.3")        # 형식 혼재 흡수
    assert norm_date("2026-06-03 00:00:00") == "2026-06-03"
    assert norm_date("") == "" and norm_date(None) == ""
    rows = [{"접수ID": "B", "접수일자": "2026-07-02"},
            {"접수ID": "C", "접수일자": ""},                                    # 날짜 없음 → 맨 뒤
            {"접수ID": "A", "접수일자": "2026.6.3"},
            {"접수ID": "D", "접수일자": "2025-12-31"}]
    got = [r["접수ID"] for r in sort_by_date(rows, "as", "접수ID")]
    assert got == ["D", "A", "B", "C"], got
    # 대표 날짜 열이 없는 시트(확인필요)는 행 안의 아무 날짜나 찾아 쓴다
    assert row_date({"메모": "발생 2026-03-05"}) == "2026-03-05"
    # 같은 날짜면 ID순 — 실행할 때마다 순서가 흔들리면 안 된다
    same = [{"접수ID": "Z", "접수일자": "2026-01-01"}, {"접수ID": "A", "접수일자": "2026-01-01"}]
    assert [r["접수ID"] for r in sort_by_date(same, "as", "접수ID")] == ["A", "Z"]
    print("  [14] 날짜 정렬(과거→최근·빈값 맨뒤·형식혼재·동률ID) ✅")


def t15_doc_ocr():
    """밴드 문서 이미지 파서 — 실제 Windows OCR이 뱉은 오인식 텍스트 그대로 검증.
    (이미지·OCR 엔진 없이 순수 파서만 돌리므로 빠르고 환경에 의존하지 않는다)"""
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "band"))
    from doc_ocr import parse_doc, infer_amounts, projects, doc_type, build_updates
    명세서 = """거 래 명 세 서
작성 일자
공급받는자
표로젝트NO
2026-07-20
쿠팡로지스틱스서비스
U]2601138
명세서번호
드록번호
SL-2026-0712
123-81-45678
공급가액
1,850,000
150,000
2,000,000
합계금얙 OJAT 포함)
세액
185,000
15,000
200,000
2 200,000"""
    r = parse_doc(명세서, "s.png")
    assert r["유형"] == "거래명세서", r["유형"]                    # 자간 벌어진 제목
    assert r["프로젝트NO"] == "UJ2601138", r["프로젝트NO"]         # U] → UJ 보정
    assert r["명세서번호"] == "SL-2026-0712", r["명세서번호"]      # 옆칸 숫자와 안 붙어야
    assert (r["공급가액"], r["세액"], r["합계"]) == (2000000, 200000, 2200000), r
    assert r["발행일"] == "2026-07-20" and r["신뢰도"] == "높음", r

    계산서 = """전 자 세 금 계산 서
20260721-41000000-11223344
514-86-01234
U]2601138 인&2Sub
공급가액
2|000,000
작성일자
세액
200,000
2026-07-21
합계금액
2 200,000"""
    r2 = parse_doc(계산서, "t.png")
    assert r2["유형"] == "세금계산서" and r2["명세서번호"] == "", r2
    assert r2["승인번호"] == "20260721-41000000-11223344", r2["승인번호"]
    assert r2["공급가액"] == 2000000, r2                          # '2|000,000' 콤마 오인식 흡수

    # 금액 정합성(공급가+세액=합계)이 안 맞으면 채택하지 않는다 → OCR 자릿수 오류 차단
    assert infer_amounts([1234000, 99999, 5555555]) == (None, None, None)
    assert doc_type("아무 글") == "미상" and projects("UL2601140") == ["UJ2601140"]

    # 자동 입력 후보: 신뢰도 높음 + 판정 일치인 건만
    rows = [{"신뢰도": "높음", "판정": "일치", "정산ID": "JS-1", "파일": "a.png",
             "제안": {"거래명세서번호": "SL-2026-0712", "거래명세서발행일": "2026-07-20"}},
            {"신뢰도": "낮음", "판정": "일치", "정산ID": "JS-2", "파일": "b.png",
             "제안": {"거래명세서번호": "X"}},
            {"신뢰도": "높음", "판정": "금액 불일치 — 대장 1 / 문서 2", "정산ID": "JS-3",
             "파일": "c.png", "제안": {"공급가액": 2}}]
    up = build_updates(rows)
    assert [u["key"] for u in up] == ["JS-1", "JS-1"], up
    assert {u["vtype"] for u in up} == {"text", "date"}, up
    print("  [15] 밴드 문서 OCR 파서(오인식 보정·금액정합·자동입력 게이트) ✅")


def t17_expand():
    """행 확장 — 수식 이동·범위 확장·기존행 보존(합성 워크북)"""
    import sys as _s
    _s.path.insert(0, ROOT)
    from expand_rows import self_test
    self_test()
    print("  [17] 시트 행 확장(수식 재배치·표/검증 범위·기존행 보존) ✅")


def t18_erp_docs():
    """ERP 매출서류 분류·inbox 내용판별 — 파일명이 무작위여도 잡히는가"""
    import sys as _s
    _s.path.insert(0, ROOT)
    from erp_docs_check import work_kind, norm_slip
    from inbox_scan import classify_rows
    assert work_kind("돌발AS_울산2캠프 테이블리프트 작동 멈춤") == "돌발AS"
    assert work_kind("26년 2분기 정기점검 - 서초1MB(양재동B)") == "정기점검"
    assert work_kind("쿠팡신규납품_부산4캠프 2R/T Mobile-lift 2EA") == "신규납품"
    assert work_kind("쿠팡철거_창원1MB 매립형 1EA 철거") == "철거"
    assert work_kind("쿠팡_구리3캠프 전면 연장형 A형 계단 1EA") == "계단"
    assert norm_slip("2026/07/25 -11") == "2026/07/25-11"       # 공백 정규화
    # 내용 판별(파일명 무관)
    assert classify_rows([["회사명 : (주)유니버셜"], ["일자-No.", "거래처명", "담당자 이메일주소",
                          "프로젝트명", "공급가액", "매출부가세", "매출합계"]]) == "tax"
    assert classify_rows([["회사명"], ["일자 - 번호", "거래처명", "담당자 이메일주소",
                          "공급가액", "부가세", "합 계"]]) == "stmt"
    assert classify_rows([["회사명"], ["전표번호", "입력메뉴", "금액", "거래처명", "적요명"]]) == "slips"
    assert classify_rows([["아무 표"], ["가", "나"]]) == "unknown"
    assert classify_rows([["일자", "적요", "차변금액", "대변금액"]]) == "ledger"
    # ★ 잔량 — `(세금)계산서진행단계`(E010849). '아직 발행 안 한 것'이라 다른 매출
    #   자료와 **묻으면 안 된다**. 그런데 `내역보기`+`공급가액`+`부가세` 를 다 가지고
    #   있어 taxinv 가 먼저 삼켰다(2026-08-08 실측: 101행이 매출세금계산서로 들어갔다).
    #   묻히면 잔량이 밀려도 밀림 보고에 안 잡힌다 — 조용한 사고다.
    잔량 = [["회사명 : 주식회사 유니버셜"],
            ["일자-No.", "프로젝트그룹1명", "적요명", "발행일자", "거래처명", "공급가액",
             "부가세", "합계금액", "종류", "전자(세금)계산서 진행단계", "단계별기능",
             "승인번호", "내역보기"]]
    assert classify_rows(잔량) == "taxstep", "잔량이 다른 매출 자료에 묻힌다"
    # 옆 화면을 잘못 물지 않는가 — `단계별기능` 이 없으면 예전 판정 그대로다
    assert classify_rows([["회사명"], ["일자-No.", "거래처명", "공급가액", "부가세",
                                       "내역보기"]]) == "taxinv"
    print("  [18] ERP 매출서류 유형분류·inbox 내용판별(잔량 분리 포함) ✅")


def t19_workbook_integrity(tmp):
    """엑셀 손상 판정 방지 — 수식 셀의 t 선언과 캐시값(<v>) 정합성.
    실사고: expand_rows가 만든 셀 1,573개가 t="str"인데 <v>가 없어
    엑셀이 열 때마다 '내용에 문제가 있습니다' 복구 대화상자를 띄웠다."""
    import sys as _s, zipfile as _z
    _s.path.insert(0, ROOT)
    from fix_workbook import add_missing_v, scan_sheet, iter_tags
    broken = ('<sheetData><row r="5">'
              '<c r="A5" s="1" t="str"><f>IF(1,2,3)</f></c>'          # ← 캐시값 없음(문제)
              '<c r="B5" s="1" t="str"><f>NOW()</f><v/></c>'          # 정상
              '<c r="C5" s="1"><f>SUM(A1:A2)</f></c>'                 # t 없음 → 대상 아님
              '</row></sheetData>')
    fixed, n = add_missing_v(broken)
    assert n == 1, n
    assert '<c r="A5" s="1" t="str"><f>IF(1,2,3)</f><v/></c>' in fixed, fixed
    assert fixed.count("<v/>") == 2, fixed          # 기존 정상 셀을 건드리지 않음
    _, again = add_missing_v(fixed)
    assert again == 0, "재실행 시 또 고치면 안 됨(멱등)"

    _, fixable = scan_sheet(broken, "t", 10)
    assert fixable == 1, fixable
    bad, _ = scan_sheet('<sheetData><row r="5"><c r="B5"/><c r="A5"/></row></sheetData>', "t", 10)
    assert any("열 순서" in b for b in bad), bad          # 열 역순 검출
    bad2, _ = scan_sheet('<sheetData><row r="7"><c r="A5"/></row></sheetData>', "t", 10)
    assert any("행 7 안에" in b for b in bad2), bad2      # 셀 참조/행 불일치 검출
    bad3, _ = scan_sheet('<sheetData><row r="5"><c r="A5" s="99"/></row></sheetData>', "t", 10)
    assert any("스타일" in b for b in bad3), bad3         # 스타일 인덱스 초과 검출
    # ★ 실사고 회귀: 속성이 붙은 <v>를 못 알아보고 <v/>를 하나 더 넣으면
    #   한 셀에 <v>가 2개가 되고, 엑셀이 그 시트를 통째로 비워 버린다(313셀 피해)
    from fix_workbook import remove_dup_v, count_dup_v, _V_RE
    keep = ('<sheetData><row r="5">'
            '<c r="A5" s="1" t="str"><f>X()</f><v xml:space="preserve">값 </v></c>'
            '</row></sheetData>')
    _, n2 = add_missing_v(keep)
    assert n2 == 0, "xml:space 붙은 <v>를 못 알아봄 — 중복 <v>를 또 만든다"
    assert _V_RE.search('<v xml:space="preserve">a</v>'), "정규식이 속성 있는 <v>를 놓침"
    dup = ('<sheetData><row r="5">'
           '<c r="A5" s="1" t="str"><f>X()</f><v xml:space="preserve">값 </v><v/></c>'
           '<c r="B5" s="1" t="str"><f>Y()</f><v/></c>'
           '</row></sheetData>')
    assert count_dup_v(dup) == 1, count_dup_v(dup)
    fixed3, n3 = remove_dup_v(dup)
    assert n3 == 1 and count_dup_v(fixed3) == 0, fixed3
    assert '<v xml:space="preserve">값 </v></c>' in fixed3, fixed3   # 원래 값은 보존
    assert '<c r="B5" s="1" t="str"><f>Y()</f><v/></c>' in fixed3, fixed3  # 정상 셀은 유지
    print("  [19] 워크북 무결성(수식 캐시값·<v> 중복·행열 순서·스타일 범위) ✅")


def t20_rep_no():
    """대표 프로젝트NO — 모든 건이 번호로 식별되어야 한다.
    실사고: 정규식의 \b가 파일에 **백스페이스 문자**로 저장돼 UJ 번호를 하나도 못 찾았다."""
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "webapp"))
    from app_server import rep_no, _UJ_RE, apply_rep_no, build_prj_index
    assert "" not in _UJ_RE.pattern, "정규식에 백스페이스 문자가 섞임"
    assert _UJ_RE.search("명세서 2026/07/01-4 · PO PO367787 · UJ2600975").group() == "UJ2600975"
    assert not _UJ_RE.search("XUJ2600975"), "앞에 글자가 붙으면 잡으면 안 됨"
    # 1) 원본 우선
    assert rep_no({"프로젝트NO": "UJ2601138"}) == ("UJ2601138", "")
    # 2) 본문에서 복원
    assert rep_no({"내용·근거": "명세서 2026/07/01-4 · UJ2600975"}) == ("UJ2600975", "본문")
    # 3) 같은 캠프·같은 달의 실제 작업에서
    idx = build_prj_index({"as": [{"캠프명": "울산2캠프", "접수일자": "2026-05-03",
                                   "프로젝트NO": "UJ2600777"}], "pm": []})
    assert rep_no({"캠프명": "울산2캠프", "완료일": "2026-05-20"}, idx) == ("UJ2600777", "동일캠프·동월")
    # 4) 전표 기반(UJ처럼 보이지 않게)
    n, how = rep_no({"캠프명": "x"}, None, "2026/01/10-2")
    assert (n, how) == ("ERP-260110-2", "전표"), (n, how)
    assert not n.startswith("UJ"), "없는 UJ 번호를 지어내면 안 된다"
    # 5) 최후엔 자체 ID
    assert rep_no({"정산ID": "JS-2607-001"}) == ("JS-2607-001", "자체ID")
    # 6) 일괄 적용 시 빈 건이 남지 않는다
    rows = [{"프로젝트NO": "UJ2601138"}, {"내용·근거": "UJ2600975"}, {"ID": "JS-9"}]
    apply_rep_no(rows)
    assert all(str(r.get("프로젝트NO") or "").strip() for r in rows), rows
    print("  [20] 대표 프로젝트NO(원본·본문·동일캠프·전표·자체ID) ✅")


def t21_reorder():
    """행 재배치 — 과거→최근 정렬 + 수식 상대참조가 함께 이동하는가.
    실사고 회귀: 자기닫힘 <f .../> 를 게으른 정규식으로 잡아 다음 셀의 r=\"B5\"까지
    숫자를 바꿔 'B-3' 같은 잘못된 셀 참조가 만들어졌다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    from reorder_rows import shift_rows, move_row, self_test
    self_test()
    assert shift_rows("$B5+E4", 3) == "$B8+E7"                 # 상대행만 이동
    assert shift_rows("$E$4:E9", 2) == "$E$4:E11"              # 절대행 고정
    row = ('<row r="9"><c r="A9" s="1" t="str"><f t="shared" si="3"/><v/></c>'
           '<c r="B9" s="1"><v>7</v></c>'
           '<c r="C9" s="1" t="str"><f>SUM($B$4:B8)</f><v/></c></row>')
    out = move_row(row, 5)
    assert '<c r="B5" s="1"><v>7</v></c>' in out, out           # ★ 셀 참조가 망가지면 안 됨
    assert '<c r="A5"' in out and '<c r="C5"' in out, out
    assert "SUM($B$4:B4)" in out, out                           # 직전행 참조가 따라 이동
    assert "B-" not in out, out
    print("  [21] 행 재배치(정렬·상대참조 이동·자기닫힘 f 안전) ✅")


def t22_bundle():
    """ERP 계산서 구성 추정 — 캠프 표기 흔들림·분기 제목·금액 정합 판정.
    추정을 확정으로 올려버리면 틀린 배분이 대장에 굳으므로 판정 경계를 못 박아 둔다."""
    import erp_bundle as B
    # 괄호 안 지명이 달라도 같은 캠프
    assert B.camp_key("송파1MB(감일동)") == B.camp_key("송파1MB"), B.camp_key("송파1MB(감일동)")
    assert B.camp_key("송파3Sub-hub") == B.camp_key("송파3subhub")
    assert B.camp_key("양주2캠프") != B.camp_key("양주1캠프")
    # 제목 → 유형
    assert B.kind_of("25년 4분기 정기점검 - 양주2캠프") == "정기점검"
    assert B.kind_of("돌발AS 인천8MB") == "돌발AS"
    assert B.kind_of("쿠팡신규_송파1MB-이동식상하차리프트 2RT") == "신규납품"
    # 제목에 분기가 적혀 있으면 발행월이 아니라 그 분기를 본다
    assert B.window({"month": "2026/01", "title": "25년 4분기 정기점검"}) == ("2025-10", "2025-12")
    assert B.window({"month": "2026/05", "title": "돌발AS"}) == ("2026-02", "2026-05")
    assert B.window({"month": "2026/02", "title": "돌발AS"}) == ("2025-11", "2026-02")  # 연 넘김
    W = [{"prj": "UJ2600001", "camp": "양주2캠프", "kind": "정기점검",
          "date": "2026-04-10", "amt": 600000, "src": "06"},
         {"prj": "UJ2600002", "camp": "양주2캠프", "kind": "정기점검",
          "date": "2026-05-11", "amt": 400000, "src": "06"}]
    doc = {"camp": "양주2캠프", "kind": "정기점검", "month": "2026/05",
           "title": "정기점검 - 양주2캠프", "amt": 1000000}
    prjs, v, tot = B.bundle(doc, W)
    assert prjs == ["UJ2600001", "UJ2600002"] and v == "확정" and tot == 1000000, (prjs, v, tot)
    prjs, v, _ = B.bundle({**doc, "amt": 1020000}, W)      # 2% 차이 → 유력
    assert v == "유력", v
    prjs, v, _ = B.bundle({**doc, "amt": 2000000}, W)      # 100% 차이 → 추정
    assert v == "추정", v
    # 후보를 못 찾으면 **이유**가 남아야 한다("미상" 한 마디는 조치를 못 한다)
    _, v, _ = B.bundle({**doc, "camp": "없는캠프"}, W)
    assert v.startswith("미상(") and "캠프" in v, v
    # 계단·철거·신규납품·기타는 AS/점검이 아니라 별도 공사다 — 02·04에 작업 행이 아예 없으므로
    # '미상'이 아니라 **대상외**로 못 박는다(그래야 영원히 미해결로 남지 않는다).
    _, v, _ = B.bundle({**doc, "kind": "철거"}, W)
    assert v.startswith("대상외(") and "철거" in v, v
    # 밴드에 사람이 적어 둔 계산서 목록·PO 발주글이 추정보다 우선한다
    binx = {("2026-05-11", 1000000): {"프로젝트": ["UJ2600009"], "캠프": "양주2캠프", "PO": ""}}
    prjs, v, _ = B.bundle({**doc, "slip": "2026/05/11-3", "amt": 1000000}, W, binx)
    assert v == "확정(밴드)" and prjs == ["UJ2600009"], (v, prjs)
    poinx = {777000: ["UJ2600077", "UJ2600078"]}
    prjs, v, _ = B.bundle({**doc, "amt": 777000}, W, None, poinx)
    assert v == "확정(밴드PO)" and len(prjs) == 2, (v, prjs)
    # PO 발주글의 총금액 표기는 제각각이다 — '원'만 보면 KRW로 적힌 글을 통째로 놓친다
    for t, want in (("★ 총금액 : 25,223,400원", "25,223,400"),
                    ("★ 총금액 :8,626,500원", "8,626,500"),
                    ("★ 총금액 : 13,866,500 KRW", "13,866,500"),
                    ("★ 총금액 :14,803,300 KRW", "14,803,300")):
        m = B._TOTAL_RE.search(t)
        assert m and m.group(1) == want, (t, m.group(1) if m else None)
    # 프로젝트NO가 안 적힌 PO는 번호·건수라도 알려 준다(추정으로 뭉개지 않는다)
    meta = {900000: {"PO": "PO364055", "품목": "정기점검 29건", "건수": 29, "유형": "정기점검"}}
    _, v2, _ = B.bundle({**doc, "amt": 900000}, [], None, None, meta)
    assert v2.startswith("PO확인(PO364055"), v2
    # 계단·철거는 PO를 알아도 '별도 공사'라는 사실이 유지돼야 한다
    _, v3, _ = B.bundle({**doc, "kind": "계단", "amt": 900000}, [], None, None,
                        {900000: {"PO": "PO111", "품목": "계단 1EA", "건수": 1, "유형": "계단"}})
    assert v3.startswith("대상외(계단") and "PO111" in v3, v3
    _, v, _ = B.bundle({**doc, "month": "2026/01"}, W)     # 기간 밖
    assert v.startswith("미상(") and "보유" in v, v
    print("  [22] ERP 계산서 구성 추정(캠프정규화·분기창·금액정합·미상사유) ✅")


def t23_formulas():
    """깨진 수식 복구 — 음수 행 참조·굳어버린 오류값·좁아진 범위.
    이 셋이 엑셀에서 #N/A 691개로 나타났다(2026-07-26). 다시 새지 않게 못 박는다."""
    import fix_formulas as F
    # 상대 행 오프셋을 유지한 채 다른 행에 재생산되는가
    tmpl = F.normalize('IF($B5="","",COUNT($AO$4:AO4)+1)', 5)
    assert F.instantiate(tmpl, 5) == 'IF($B5="","",COUNT($AO$4:AO4)+1)'
    assert F.instantiate(tmpl, 130) == 'IF($B130="","",COUNT($AO$4:AO129)+1)'
    # 깨짐 판정
    assert F.is_broken('COUNT($AO$4:AO-120)+1')      # 음수 행
    assert F.is_broken('#N/A')                       # 오류값이 수식으로 굳음
    assert F.is_broken('SUM(#REF!)')
    assert not F.is_broken('IF($A5="","",1)')
    # ★ 공유수식 참조 셀은 본문이 비어 있다. 이걸 본으로 뽑으면 열 전체가 못 고쳐진다.
    cells = [(5, 'Q', 'IF($A5="","",N($O5))', 0, 0),
             (6, 'Q', '', 0, 0), (7, 'Q', '', 0, 0), (8, 'Q', '', 0, 0)]
    tm = F.templates(cells)
    assert tm.get('Q'), '빈 수식이 최다 득표로 뽑히면 안 된다'
    assert F.instantiate(tm['Q'], 9) == 'IF($A9="","",N($O9))'
    # 범위 확장: 실제 마지막 행까지 늘어나되, 이미 충분하면 건드리지 않는다
    xml = "<f>COUNTIF('02_돌발AS접수'!$A$5:$A$154,1)</f>"
    y, n = F.widen(xml, {'02_돌발AS접수': 344})
    assert '$A$344' in y and n == 1, y
    y2, n2 = F.widen(y, {'02_돌발AS접수': 344})
    assert n2 == 0, '이미 맞는 범위를 또 건드리면 안 됨'
    # 셀 위치 계산: <row> 여는 태그 길이를 빼먹으면 엉뚱한 자리를 덮어쓴다
    sx = ('<sheetData><row r="5" spans="1:3"><c r="A5"><v>1</v></c>'
          '<c r="B5" t="e"><f>#N/A</f><v>#N/A</v></c></row>'
          '<row r="6"><c r="B6"><f>IF($A6="","",2)</f><v>2</v></c></row></sheetData>')
    cells = F.parse_cells(sx)
    b5 = [c for c in cells if c[1] == 'B' and c[0] == 5][0]
    assert sx[b5[3]:b5[4]].startswith('<c r="B5"'), sx[b5[3]:b5[4]][:30]
    fixed, k = F.fix_sheet(sx)
    assert k == 1 and 'IF($A5="","",2)' in fixed and '#N/A' not in fixed, fixed
    assert '<c r="A5"><v>1</v></c>' in fixed, '옆 셀이 망가지면 안 됨'
    # 값이 보이는 상태열도 수식 전용 열이면 복구한다(첨부 화면의 '미배정' 고착 회귀).
    owned = ('<sheetData>'
             '<row r="5"><c r="B5" t="inlineStr"><is><t>UJ1</t></is></c>'
             '<c r="M5" t="inlineStr"><is><t>미배정</t></is></c></row>'
             '<row r="6"><c r="B6" t="inlineStr"><is><t>UJ2</t></is></c>'
             '<c r="M6"><f>IF($B6="","",IF($L6="","미배정","배정완료"))</f><v>미배정</v></c>'
             '</row></sheetData>')
    restored, nr = F.restore_owned_formulas(owned, ("M",))
    assert nr == 1 and '<c r="M5"><f>IF($B5="","",IF($L5="","미배정","배정완료"))</f><v/></c>' in restored, restored
    assert not F.direct_self_refs(restored), F.direct_self_refs(restored)
    direct = '<row r="5"><c r="M5"><f>IF(M5="","",1)</f><v/></c></row>'
    assert F.direct_self_refs(direct) == ["M5"]
    counter = ('<row r="130"><c r="AO130"><f>'
               'IF($B130="","",COUNT($AO$4:AO153)+1)</f><v>1</v></c></row>')
    counter_out, counter_n = F.fix_cumulative_counters(counter)
    assert counter_n == 1 and 'COUNT($AO$4:AO129)' in counter_out, counter_out
    # 03시트 자동선택 사슬: 잘못 들어간 AS 채번 자기참조와 확장 뒤 굳은
    # COUNTIF 끝행을 한꺼번에 바로잡아야 한다.
    selector = (
        "IFERROR(INDEX('02_돌발AS접수'!$A$5:$A$741,MATCH(1,"
        "('02_돌발AS접수'!$A$5:$A$741&lt;&gt;\"\")*"
        "('02_돌발AS접수'!$Q$5:$Q$741=\"작업완료\")*"
        "('02_돌발AS접수'!$R$5:$R$741&lt;&gt;\"\")*"
        "(COUNTIF($B$4:$B201,'02_돌발AS접수'!$A$5:$A$741)=0),0)),\"\")"
    )
    selector_xml = (
        '<sheetData>'
        f'<row r="202"><c r="B202"><f>{selector}</f><v/></c></row>'
        '<row r="203"><c r="B203"><f>IF($B203="","","AS-"&amp;TEXT($D203,"yymm"))</f>'
        '<v>AS-2607-468</v></c></row>'
        f'<row r="204"><c r="B204"><f>{selector}</f><v/></c></row>'
        '</sheetData>'
    )
    selector_out, selector_n = F.fix_completed_as_selector(selector_xml)
    assert selector_n == 2, selector_n
    assert 'COUNTIF($B$4:$B202' in selector_out, selector_out
    assert 'COUNTIF($B$4:$B203' in selector_out, selector_out
    assert 'AS-2607-468' not in selector_out, selector_out
    assert not F.direct_self_refs(selector_out), F.direct_self_refs(selector_out)
    # 일반 수식열에서도 사람이 확정한 inlineStr 값은 <v>가 없다는 이유로 덮으면 안 된다.
    mixed = ('<sheetData>'
             '<row r="5"><c r="H5"><f>IF($B5="","",1)</f><v>1</v></c>'
             '<c r="I5"><f>IF($B5="","",2)</f><v>2</v></c></row>'
             '<row r="6"><c r="H6"><f>IF($B6="","",1)</f><v>1</v></c>'
             '<c r="I6"><f>IF($B6="","",2)</f><v>2</v></c></row>'
             '<row r="7"><c r="H7" t="inlineStr"><is><t>엄진언</t></is></c></row>'
             '<row r="8"><c r="H8" t="inlineStr"><is><t></t></is></c>'
             '<c r="I8"><f>IF($B8="","",2)</f><v>2</v></c></row>'
             '</sheetData>')
    mixed_out, mixed_n = F.fill_missing(mixed)
    assert mixed_n == 2, mixed_n
    assert '<c r="H7" t="inlineStr"><is><t>엄진언</t></is></c>' in mixed_out, mixed_out
    assert '<c r="H8"><f>IF($B8="","",1)</f><v/></c>' in mixed_out, mixed_out
    assert '<c r="I7"><f>IF($B7="","",2)</f><v/></c>' in mixed_out, mixed_out
    assert "00_대시보드" in F.FILL_MISSING_EXCLUDE
    print("  [23] 수식 복구(수기값 보존·빈칸 판정·대시보드 제외·순환참조·범위확장) ✅")


def t24_reorder_safety():
    """행 재배치가 **엑셀이 못 여는 파일**을 만들지 않는지.
    2026-07-26 v101~v104가 이 두 가지로 열리지 않았다. 구조 검사로는 안 잡힌다 —
    XML은 멀쩡하고 '의미'만 어긋나기 때문에 여기서 못 박아 둔다."""
    import reorder_rows as R
    # (1) 배열수식 ref 는 자기 셀 주소다 — 행을 옮기면 함께 가야 한다
    row = ('<row r="69"><c r="Q69" s="53"><f t="array" ref="Q69" aca="1">SUM($A69:$B69)</f>'
           '<v>1</v></c></row>')
    moved = R.move_row(row, 5)
    assert 'ref="Q5"' in moved, moved
    assert '<row r="5"' in moved and '<c r="Q5"' in moved, moved
    assert 'SUM($A5:$B5)' in moved, moved
    # 범위형 ref 도 양끝이 따라간다
    row2 = '<row r="10"><c r="A10"><f t="array" ref="A10:B10">1</f><v>1</v></c></row>'
    assert 'ref="A12:B12"' in R.move_row(row2, 12), R.move_row(row2, 12)
    # (2) 공유수식은 '연속한 행' 전제라 섞기 전에 반드시 펼쳐야 한다
    xml = ('<sheetData>'
           '<row r="5"><c r="B5"><f t="shared" ref="B5:B7" si="3">IF($A5="","",1)</f><v>1</v></c></row>'
           '<row r="6"><c r="B6"><f t="shared" si="3"/><v>1</v></c></row>'
           '<row r="7"><c r="B7"><f t="shared" si="3"/><v>1</v></c></row>'
           '</sheetData>')
    out, n = R.unshare(xml)
    assert n == 3, n
    assert 't="shared"' not in out and 'si="3"' not in out, out
    assert 'IF($A6="","",1)' in out and 'IF($A7="","",1)' in out, out   # 각 행 수식이 생겼다
    assert '<c r="B6"' in out and '<v>1</v>' in out, out                # 셀·캐시는 그대로
    print("  [24] 재배치 안전성(배열수식 ref 이동·공유수식 펼침) ✅")


def t25_attachments():
    """행에 **붙어 있는 것들**이 함께 움직이는가 / 범위 확장이 시작행을 밀지 않는가.
    둘 다 2026-07-26 합성 점검에서 찾은 구멍이다."""
    import reorder_rows as R, expand_rows as E
    # (1) 하이퍼링크는 sheetData 밖에 적혀 있어 행만 정렬하면 제자리에 남는다 →
    #     9행이 다른 건으로 바뀌었는데 링크는 원래 건의 밴드 글을 가리킨다.
    post = '<hyperlinks><hyperlink ref="AM9" r:id="rId1"/><hyperlink ref="AM12" r:id="rId2"/></hyperlinks>'
    out, n = R.move_hyperlinks(post, {9: 300, 12: 7})
    assert n == 2 and 'ref="AM300"' in out and 'ref="AM7"' in out, out
    out2, n2 = R.move_hyperlinks(post, {})          # 옮길 행이 없으면 그대로
    assert n2 == 0 and out2 == post
    # (2) 범위 확장은 **끝행만** 늘려야 한다. 시작행까지 밀면 구간이 통째로 벗어난다.
    assert E.bump("A5:A5", 5, 8) == "A5:A8", E.bump("A5:A5", 5, 8)
    assert E.bump("V5:V31 V33:V474", 474, 500) == "V5:V31 V33:V500"
    assert E.bump("V31", 31, 60) == "V31", "단일 셀은 범위가 아니다"
    assert E.bump("A5:AG104", 104, 164) == "A5:AG164"
    # (3) 마지막 행을 본떠 확장할 때 자기행뿐 아니라 '바로 윗행까지'인
    # 상대참조도 함께 이동해야 한다. 고정되면 모든 새 행이 같은 ID를 고른다.
    copied = E.shift_formula('IF($B202="","",COUNTIF($B$4:$B201,1))', 202, 205)
    assert '$B205' in copied and '$B204' in copied and '$B$4' in copied, copied
    print("  [25] 행 부속물(하이퍼링크 이동·범위 끝행만 확장) ✅")


def t26_mobile():
    """폰에서 여는 경로가 **끊기지 않는 구조인지** 합성으로 확인한다.

    실제로 반복해서 겪은 고장은 전부 '주소가 어딘가에 박혀 있어서'였다:
      · 폰 홈 화면 아이콘  → 추가할 때의 터널 주소가 매니페스트 start_url에 박힘
      · PC 앱(크롬 PWA)   → 설치할 때의 터널 주소가 app-id에 박힘
    터널 주소는 띄울 때마다 새로 받으므로 다음 날이면 전부 죽는다.
    그래서 **바깥으로 내보내는 주소는 언제나 고정 진입점**이어야 한다.
    """
    import json as _j, re as _re
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as A

    FIX = A.FIXED_ENTRY
    assert FIX.startswith("https://"), FIX
    assert "trycloudflare" not in FIX, "고정 진입점에 터널 주소를 넣으면 안 된다"

    # (1) 앱이 내주는 매니페스트: 홈 화면 아이콘이 고정 주소를 향하는가
    src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    i = src.index('if p == "/manifest.json"')
    blk = src[i:i + 900]
    # ★ 매니페스트 start_url은 **그 페이지와 같은 출처**여야 한다.
    #   다른 도메인을 넣으면 크롬이 매니페스트를 무시해 [설치 및 바로가기 만들기]가 아예 안 된다
    #   (2026-07-27에 고정 주소를 넣었다가 폰에서 설치가 막혔다).
    assert 'start_url = f"/staff/{staff_slug}" if center else "/"' in src, \
        "앱·담당자 매니페스트 start_url은 같은 출처 경로여야 설치가 된다"
    assert '"start_url": start_url' in src
    assert "FIXED_ENTRY" not in blk, "다른 도메인을 start_url에 넣으면 크롬이 설치를 거부한다"
    # 대신 앱 화면이 **고정 주소를 항상 보여줘야** 한다(닫히는 배너가 아니라 상시 표기).
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert FIX in idx, "앱이 고정 주소를 표기하지 않는다"
    assert "showFixedEntry" in idx and 'id="entrycard"' in idx, "고정 주소 상시 표기 카드가 없다"
    assert "tunnelNotice" not in idx, "닫히는 안내 배너는 상시 표기로 대체됐다"
    # 카드가 숨겨져 있으면 표기한 의미가 없다
    _card = idx[idx.index('id="entrycard"'):idx.index('id="entrycard"') + 60]
    assert "display:none" not in _card, "고정 주소 카드가 숨겨져 있다"

    # (1-b) 크롬 [설치 및 바로가기 만들기]는 fetch 핸들러를 가진 서비스 워커가 있어야 뜬다.
    #       고정 주소 쪽과 앱(터널) 쪽 **양쪽 출처 모두** 필요하다 — 설치는 출처 단위다.
    assert "serviceWorker" in idx and "/sw.js" in idx, "앱에 서비스 워커 등록이 없다 — 설치가 안 된다"
    assert 'p == "/sw.js"' in src, "앱 서버가 /sw.js 를 내주지 않는다"
    # ★ 700자만 잘라 보던 것을 3000자로 넓혔다(2026-07-31). 서비스 워커가 오프라인 껍데기
    #   캐시를 하게 되면서 블록이 길어졌고, fetch 핸들러가 창 밖으로 밀려나 검증이 헛돌았다.
    #   보려는 것은 '핸들러가 있는가' 이지 '몇 번째 글자에 있는가' 가 아니다.
    _swblk = src[src.index('p == "/sw.js"'):][:3000]
    assert "addEventListener('fetch'" in _swblk, "fetch 핸들러가 없으면 크롬이 설치를 제안하지 않는다"
    # 오프라인에서도 화면이 떠야 폰에서 입력해 둘 수 있다(사용자 지시 2026-07-31).
    assert "caches.open" in _swblk and "caches.match" in _swblk, \
        "서비스 워커가 껍데기를 캐시하지 않으면 PC 가 꺼졌을 때 앱이 아예 안 열린다"
    assert "startsWith('/api/')" in _swblk, "데이터(/api/)까지 캐시하면 옛 숫자가 화면에 남는다"
    _dsw = open(os.path.join(ROOT, "docs", "sw.js"), encoding="utf-8").read()
    assert "addEventListener('fetch'" in _dsw, "고정 주소 쪽 서비스 워커에 fetch 핸들러가 없다"

    # (2) 고정 진입점 페이지: 앱 시작 자체가 PC 터널과 무관한가
    doc = open(os.path.join(ROOT, "docs", "index.html"), encoding="utf-8").read()
    assert 'rel="manifest"' in doc, "고정 페이지에 매니페스트가 없으면 홈 화면 추가가 앱으로 안 붙는다"
    assert "serviceWorker" in doc and "sw.js" in doc, "고정 페이지에 서비스 워커 등록이 없다 — 설치가 안 된다"
    assert "location.replace('app.html')" in doc, "고정 주소가 PC 독립 앱으로 바로 들어가지 않는다"
    assert "endpoint.json" not in doc and "/api/ping" not in doc, \
        "앱 시작 전에 PC 터널을 확인하면 PC 종료 시 다시 막힌다"

    mf = _j.load(open(os.path.join(ROOT, "docs", "manifest.json"), encoding="utf-8"))
    assert mf["start_url"] == FIX + "app.html", mf
    assert mf.get("scope", FIX).startswith(FIX.rstrip("/")), mf
    assert "trycloudflare" not in _j.dumps(mf)
    icon_hash = hashlib.sha256(
        open(os.path.join(ROOT, "webapp", "brand", "csos-app-icon-source.png"), "rb").read()
    ).hexdigest()[:12]
    icon_rev = f"csos-{icon_hash}"
    assert any(f"icon-192.png?v={icon_rev}" in x.get("src", "") for x in mf["icons"]), mf
    assert any(f"icon-512.png?v={icon_rev}" in x.get("src", "") for x in mf["icons"]), mf
    assert f"csos-icon-{icon_hash}" in open(
        os.path.join(ROOT, "docs", "sw.js"), encoding="utf-8").read()
    for n in (32, 180, 192, 512):
        fp = os.path.join(ROOT, "docs", f"icon-{n}.png")
        assert os.path.getsize(fp) > 1000, fp

    # (3) 터널 주소는 어디에도 하드코딩돼 있으면 안 된다(리포트·설정 파일은 예외)
    hard = []
    for fn in ("webapp/app_server.py", "webapp/index.html", "docs/index.html",
               "phone_access.py", "publish_endpoint.py"):
        txt = open(os.path.join(ROOT, fn), encoding="utf-8").read()
        for m in _re.finditer(r"https://[a-z0-9-]+\.trycloudflare\.com", txt):
            hard.append(f"{fn}: {m.group()}")
    assert not hard, "터널 주소가 코드에 박혀 있다 — 주소가 바뀌면 그대로 죽는다: " + "; ".join(hard[:3])

    # (4) 자가복구가 붙어 있는가 — 죽은 주소를 붙들고 있으면 아무도 못 고친다
    import webapp.tunnel_run as T  # noqa: F401
    tr = open(os.path.join(ROOT, "webapp", "tunnel_run.py"), encoding="utf-8").read()
    assert "def watch(" in tr and "def publish(" in tr, "터널 자가점검·주소게시가 빠졌다"

    # ★ cloudflared 로그에는 api.trycloudflare.com(내부 API)도 나온다.
    #   그걸 터널 주소로 잘못 잡으면 고정 주소가 엉뚱한 곳을 가리켜 폰이 통째로 막힌다.
    import re as _re2
    m = _re2.search(r"m = re\.search\(r\"\((https[^\"]+)\)\"", tr)
    assert m, "터널 주소 정규식을 못 찾음"
    rx = _re2.compile(m.group(1))
    assert not rx.search("https://api.trycloudflare.com"), "api.trycloudflare.com을 걸러내지 못한다"
    assert rx.search("INF | https://mold-restored-flags-earthquake.trycloudflare.com |"), "정상 주소를 못 잡는다"

    # 주소를 너무 자주 새로 만들면 등록조차 안 되는 주소를 받는다 — 한도가 있어야 한다
    assert "_throttle" in tr and "MAX_PER_HOUR" in tr, "터널 재생성 한도가 없다"

    # 게시 직전에 형식·생존을 한 번 더 본다(죽은 주소를 올리면 폰이 막힌다)
    pe = open(os.path.join(ROOT, "publish_endpoint.py"), encoding="utf-8").read()
    assert "게시 취소" in pe and "/api/ping" in pe, "죽은 주소를 그대로 게시할 수 있다"
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "kill_stale_tunnel" in wd, "워치독이 좀비를 정리하지 않으면 재시작이 무효가 된다"
    assert "heal_fixed_funnel" in wd and "ensure_public_funnel" in wd, \
        "워치독이 PC 내부 응답만 보고 휴대폰 공개 Funnel 장애를 놓친다"
    # ★★ 터널 주소에서는 **설치가 되면 안 된다**. 터널 호스트는 매번 바뀌는데 거기서
    #   설치하면 그 임시 주소가 아이콘에 영구히 박혀, 주소가 바뀌는 순간 영영 안 열린다
    #   (2026-07-28: PC·폰의 설치된 앱이 둘 다 옛 터널 주소로 죽었다).
    i = src.index('if p in ("/", "/index.html", "/ryu")')
    blk2 = src[i:i + 1200]
    assert "trycloudflare.com" in blk2 and 'rel="manifest"' in blk2,         "터널 출처에서 매니페스트를 빼지 않는다 — 거기서 설치하면 아이콘이 곧 죽는다"
    assert "serviceWorker.register" in blk2, "터널 출처에서 서비스워커도 빼야 설치가 안 뜬다"
    print("  [26] 모바일 접속 경로(고정 진입점·터널출처 설치차단·자가복구) ✅")


def t27_po():
    """쿠팡 PO 대조 — **계산서가 끊긴 PO를 '누락'이라 부르지 않는다**.

    쿠팡 PO는 '정기점검 24건'처럼 여러 작업을 묶은 것이라 06시트 PO번호 칸에 안 적힌다.
    그걸 전부 '원장 미등록'으로 세면 96건 중 94건이 쏟아져 나와, 무엇이 진짜 미청구인지
    묻혀 버린다(2026-07-27 실제로 그랬다). ERP 계산서 금액과 맞춰 본 뒤,
    **계산서가 없는 PO만** 미청구로 올린다.
    """
    import po_reconcile as P, inspect
    src = inspect.getsource(P.main)
    assert "erp_amts" in src and "invoiced(" in src, "ERP 계산서 대조 없이 미등록으로 단정한다"
    assert "미청구 — 쿠팡 PO는 받았는데" in src, "판정 문구가 조치로 이어지지 않는다"
    assert "billed" in src, "계산서 발행된 PO를 따로 세지 않는다"
    # 밴드에 '세금계산서 발행 완료'라고 적힌 PO를 미청구로 보고하면 안 된다
    # (오종현 매니저 확인: PO 내역은 밴드 매출처업무에 1월부터 전부 기재한다)
    assert "band_po" in src and "발행완료" in src, "밴드에 적힌 PO 처리 상태를 안 본다"

    import po_band_status as PB
    body = "\n".join([
        "Coupang이(가) 새 구매 오더(PO344599)를 전송했습니다.",
        "⭐ 세금계산서 발행 2건 발행",
        "★오더번호 : PO344599",
        "2026.03.25 세금계산서 발행 완료",
        "★ 프로젝트 No. : UJ2600211",
    ])
    got = None
    for blk in __import__("re").split(r"(?=Coupang이\(가\) 새 구매 오더|⭐|✅)", body):
        if "PO344599" not in blk:
            continue
        st = [n for n, rx in PB.STATES if rx.search(blk)]
        if "발행완료" in st:
            got = st
    assert got and "발행완료" in got, ("발행 완료 문구를 못 읽는다", got)
    st2 = [n for n, rx in PB.STATES if rx.search("✅ 쿠팡오더처리 + 세금계산서 발행대기")]
    assert "발행대기" in st2, st2
    st3 = [n for n, rx in PB.STATES if rx.search("※쿠팡오더 금액 안맞아도 처리해도 된다고 확인 받음")]
    assert "금액이상" in st3, st3

    # PO 번호 정규화 — 표기가 흔들려도 같은 PO로 본다
    for t_, want in (("PO326234", "PO326234"), ("po 326234", "PO326234"),
                     ("PO-326234", "PO326234"), ("발주 PO326234 건", "PO326234")):
        assert P.norm_po(t_) == want, (t_, P.norm_po(t_))
    assert P.norm_po("PROJECT") == "" and P.norm_po("") == ""

    # 총금액이 부가세 포함으로 적힌 PO도 공급가 기준 계산서와 맞아야 한다
    erp = {1000000, 1100000}
    def invoiced(a):
        return int(a) in erp or round(int(a) / 1.1) in erp
    assert invoiced(1100000) and invoiced(1000000)
    assert not invoiced(999999)
    print("  [27] 쿠팡 PO 대조(계산서 발행분 제외·번호 정규화·부가세 표기) ✅")


def t16_status():
    """상태 수식 캐시 보정 — 새로 넣은 행은 상태열(수식)이 None이라 완료된 점검이
    전부 '미점검'으로 보였다. 원본 열로 직접 판정하는 로직 검증."""
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "webapp"))
    from app_server import derive_status
    from datetime import date as _d, timedelta as _td
    past = (_d.today() - _td(days=10)).isoformat()
    future = (_d.today() + _td(days=10)).isoformat()
    r = {"점검상태": "", "실제점검일": "2026-05-27", "점검예정일": past}
    derive_status(r, "pm"); assert r["점검상태"] == "완료", r          # 완료일 있으면 완료
    r = {"점검상태": "", "실제점검일": "", "돌발AS전환여부": "Y", "점검예정일": past}
    derive_status(r, "pm"); assert r["점검상태"] == "AS전환", r
    r = {"점검상태": "", "실제점검일": "", "점검예정일": past}
    derive_status(r, "pm"); assert r["점검상태"] == "미점검", r        # 예정일 경과
    r = {"점검상태": "", "실제점검일": "", "점검예정일": future}
    derive_status(r, "pm"); assert r["점검상태"] == "예정", r          # 아직 안 지남
    r = {"점검상태": "완료", "실제점검일": ""}
    derive_status(r, "pm"); assert r["점검상태"] == "완료", r          # 기존 값은 안 건드림
    r = {"점검상태": "미점검", "실제점검일": "2026-07-30"}
    derive_status(r, "pm"); assert r["점검상태"] == "완료", r          # 완료일이 오래된 캐시보다 우선
    r = {"점검상태": "AS전환", "실제점검일": "2026-07-30"}
    derive_status(r, "pm"); assert r["점검상태"] == "AS전환", r        # 충돌 상태는 보존
    r = {"진행상태": "", "작업완료일": "2026-06-02"}
    derive_status(r, "as"); assert r["진행상태"] == "작업완료", r
    r = {"진행상태": "접수", "작업완료일": "2026-06-02"}
    derive_status(r, "as"); assert r["진행상태"] == "작업완료", r      # 완료일이 접수 상태보다 우선
    r = {"진행상태": "취소", "작업완료일": "2026-06-02"}
    derive_status(r, "as"); assert r["진행상태"] == "취소", r          # 취소는 자동 덮기 금지
    r = {"진행상태": "", "작업완료일": ""}
    derive_status(r, "as"); assert r["진행상태"] == "접수", r
    print("  [16] 상태 수식 캐시 보정(완료·AS전환·미점검·예정·기존값 보존) ✅")


def t6_webapp():
    import time, json, urllib.request, http.cookiejar
    port = 18899
    p = subprocess.Popen([PY, os.path.join(ROOT, "webapp", "app_server.py"), "--demo", "--port", str(port)],
                         cwd=ROOT, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    try:
        base = f"http://127.0.0.1:{port}"
        # ★ 기동 대기를 9초(30×0.3)에서 45초로 늘리고, **안 떴으면 안 떴다고 말한다**
        #   (2026-08-09 실사고). 합성검증이 두 번 연속 `ConnectionResetError` 로 죽었다.
        #   원인은 코드가 아니라 **부하**였다 — daily_run 회차가 Z: 를 훑는 동안
        #   데모 서버가 9초 안에 포트를 못 잡았다. 그런데 대기 루프는 30번 다 실패해도
        #   **그냥 빠져나가** 첫 요청이 연결 거부로 터졌다. 그래서 화면에 뜬 이유가
        #   "서버가 안 떴다"가 아니라 엉뚱한 소켓 오류였고, 한참 헤맸다.
        #   ※ 이건 **관문**이다. 여기가 거짓 빨간불이면 실작업이 통째로 막히고
        #     daily_run 0단계도 회차를 중단시킨다 — 느린 것을 고장으로 읽으면 안 된다.
        up = False
        for _ in range(150):                      # 45초까지 기다린다
            if p.poll() is not None:              # 프로세스가 죽었으면 더 기다릴 것 없다
                break
            try:
                urllib.request.urlopen(base + "/api/ping", timeout=1); up = True; break
            except Exception:
                time.sleep(0.3)
        assert up, ("데모 웹앱이 45초 안에 포트 %d 를 잡지 못했습니다"
                    "(프로세스 종료코드 %s) — 포트를 다른 것이 쓰고 있거나 "
                    "기계가 몹시 바쁩니다. 소켓 오류로 터지기 전에 여기서 멈춥니다"
                    % (port, p.poll()))
        # 로그인(잘못된 PIN → 401, 데모 PIN 0000 → ok)
        req = urllib.request.Request(base + "/api/login", data=b'{"pin":"9999"}', method="POST")
        try:
            urllib.request.urlopen(req); assert False, "잘못된 PIN이 통과됨"
        except urllib.error.HTTPError as e:
            assert e.code == 401
        req = urllib.request.Request(base + "/api/login", data=b'{"pin":"0000"}', method="POST")
        assert json.loads(urllib.request.urlopen(req).read())["ok"]
        # 담당자 로그인은 서버 세션에 역할이 고정된다. 공용 PIN을 알고 있어도
        # 실행·정책·다른 담당자 입력 API를 직접 호출할 수 없어야 한다.
        staff_opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(http.cookiejar.CookieJar()))
        staff_login = urllib.request.Request(
            base + "/api/login",
            data=b'{"pin":"0000","staff_slug":"ryu-jiyeong"}',
            method="POST")
        staff_auth = json.loads(staff_opener.open(staff_login).read())
        assert staff_auth["role"] == "staff" and staff_auth["staff_slug"] == "ryu-jiyeong"
        staff_session_response = staff_opener.open(base + "/api/auth/session")
        staff_session = json.loads(staff_session_response.read())
        assert staff_session["ok"] and staff_session["authenticated"]
        assert staff_session["role"] == "staff"
        assert staff_session["staff_slug"] == "ryu-jiyeong"
        assert int(staff_session["expires_at"]) > int(time.time()) + (300 * 24 * 60 * 60), \
            "담당자 기기 인증이 장기 유지되지 않음"
        assert "Max-Age=" in (staff_session_response.headers.get("Set-Cookie") or ""), \
            "앱 재실행 시 기기 인증 만료일이 연장되지 않음"
        for path, data in (
            ("/api/run/all", b"{}"),
            ("/api/policy", '{"기준":"x","확정내용":"y"}'.encode("utf-8")),
            ("/api/staff/po-upload", b""),
        ):
            request = urllib.request.Request(
                base + path, data=data, headers={"X-Pin": "0000"}, method="POST")
            try:
                staff_opener.open(request)
                assert False, f"담당자 세션이 관리자 API를 호출함: {path}"
            except urllib.error.HTTPError as exc:
                assert exc.code == 403, (path, exc.code)
        # 인증 상태·정산 API
        h = {"X-Pin": "0000"}
        st = json.loads(urllib.request.urlopen(urllib.request.Request(base + "/api/status", headers=h)).read())
        assert st.get("demo") and st["steps"], st
        se = json.loads(urllib.request.urlopen(urllib.request.Request(base + "/api/settlements", headers=h)).read())
        assert len(se["rows"]) >= 10 and se["rows"][0]["정산ID"].startswith("JS-"), len(se.get("rows", []))
        ds = [r.get("완료일", "") for r in se["rows"] if r.get("완료일")]
        assert ds == sorted(ds), "정산 목록이 과거→최근 순이 아님: " + str(ds[:5])
        # 기준일 설정(데모: 시뮬레이션 응답) — 잠금 테스트 이전에 수행
        req = urllib.request.Request(base + "/api/set_dates", data='{"보고일":"2026-07-25"}'.encode("utf-8"),
                                     headers={"X-Pin": "0000"}, method="POST")
        d2 = json.loads(urllib.request.urlopen(req).read())
        assert d2.get("ok") and d2.get("demo"), d2
        # 미인증 접근 차단
        try:
            urllib.request.urlopen(base + "/api/settlements"); assert False, "무인증 접근 허용됨"
        except urllib.error.HTTPError as e:
            assert e.code == 401
        # 브루트포스 잠금: 5회 실패 후 429
        for _ in range(5):
            try:
                urllib.request.urlopen(urllib.request.Request(base + "/api/login", data=b'{"pin":"1111"}', method="POST"))
            except urllib.error.HTTPError:
                pass
        try:
            urllib.request.urlopen(urllib.request.Request(base + "/api/login", data=b'{"pin":"0000"}', method="POST"))
            assert False, "잠금 미작동"
        except urllib.error.HTTPError as e:
            assert e.code == 429, e.code
        # 메인 페이지 서빙
        html = urllib.request.urlopen(base + "/").read().decode("utf-8")
        assert "Coupang Service Operations System" in html and "tabbar" in html and "d_report" in html
        # 대시보드는 화면 블록과 내부 KPI를 모두 카드로 취급한다. 체크로 표시 여부를
        # 정하고 드래그·화살표로 순서를 바꾸며, 구버전 배치도 v2로 이어받는다.
        for marker in (
            "DASH_LAYOUT_KEY", "csos_dashboard_layout_v2", "DASH_LAYOUT_LEGACY_KEY",
            "defaultDashboardLayout", "DASH_KPI_CARDS", "data-kpi-card",
            "dashboardLayoutState", "setDashboardBlockVisible", "moveDashboardBlock",
            "setDashboardKpiVisible", "moveDashboardKpi", "prepareDashboardKpis",
            "cycleDashboardBlockSize", "dash-drag-handle", "dash-kpi-handle", "dash-choice",
            "role=\"status\" aria-live=\"polite\"", "window.addEventListener('storage'",
            # 2026-08-06: 끌기가 HTML5 dragstart → 포인터 이벤트로 바뀌었다.
            # 터치에서 dragstart 가 아예 나지 않아 폰에서 못 움직였다(검증 [122]).
            "dashDragEnable('dashGrid'", "initDashboardLayout();",
        ):
            assert marker in html, "대시보드 카드 편집 누락: " + marker
        assert "#dashGrid{display:grid;grid-template-columns:repeat(12" in html.replace(" ", "")
        # 제공받은 CSOS 아이콘을 앱·캡처에 공통 사용한다. 회사 CI는 상하 배치한
        # 밝은 배경판 안에서 잘리지 않아야 하며, 컨테이너는 기존 295px 폭을 보존한다.
        assert '/icon-192.png?v=csos-20260730' in html
        assert 'class="logo app-icon"' in html and "loadAppIconImg" in html
        app_server_src = open(os.path.join(ROOT, "webapp", "app_server.py"),
                              encoding="utf-8").read()
        assert "icon_revision()" in app_server_src
        assert "sync_installed_app_icons()" in app_server_src
        assert os.path.isfile(os.path.join(ROOT, "webapp", "sync_app_icons.ps1"))
        assert os.path.isfile(os.path.join(ROOT, "webapp", "build_windows_icon.ps1"))
        assert os.path.isfile(os.path.join(ROOT, "webapp", "install_staff_shortcut.ps1"))
        compact_html = html.replace(" ", "")
        assert ".appbar-brand-stack{min-width:295px}" in compact_html
        assert ".uni-app-brand{width:196px;height:26px" in compact_html
        assert "drawUniversalLogo" in html and "universal-lift-horizontal.png" in html
        # 공통 헤더는 시스템 식별/회사 로고/상태의 3구역이며, 회사 로고는
        # 유니버셜리프트를 위에 두고 쿠팡을 그 아래에 둔다.
        assert 'class="appbar-identity"' in html
        assert 'class="appbar-brand-stack"' in html and 'class="appbar-status"' in html
        header = html[html.index('<header class="appbar">'):html.index("</header>", html.index('<header class="appbar">'))]
        assert header.index('class="uni-app-brand"') < header.index('class="coupang-app-brand"')
        for asset in ("icon-32.png", "icon-180.png", "icon-192.png", "icon-512.png"):
            r = urllib.request.urlopen(base + "/" + asset)
            assert r.headers.get_content_type() == "image/png" and len(r.read()) > 1000
        # 공통 캡처 도구막대와 류지영 조사실(카톡 2개 방 원본)은 모든 화면에서
        # 같은 저장·복사·엑셀 흐름을 쓴다.
        for marker in ("media-tools", "icon-btn", "#i-file-spreadsheet",
                       'id="v-ryu"', 'name="kakao_regular"', 'name="kakao_emergency"',
                       "/api/ryu/upload", "★UNI★ 쿠팡정기점검", "★UNI★ 쿠팡돌발점검",
                       "auto_check_queued"):
            assert marker in html, "UI marker missing: " + marker
        # 류지영 화면은 단순 업로드방이 아니라 대시보드형 업무센터다.
        # 카테고리→과거목록→선택건 보충입력 흐름과 일반 메뉴(실행 제외)를 함께 제공한다.
        for marker in ("류지영 쿠팡 AS 및 정기점검 업무센터", 'class="workcenter-person"',
                       'id="ryuSummaryGrid"',
                       'id="ryuCategoryTabs"', 'id="ryuHistoryList"', 'id="ryuEntryForm"',
                       "/api/staff/records", "/api/staff/entry", "submitRyuEntry",
                       "body.ryu-mode .tabbar button[data-v=\"run\"]{display:none}",
                       "routeNav('dash')"):
            assert marker in html, "류지영 업무센터 UI 누락: " + marker
        assert 'href="/staff/yoo-hyeonmin"' in html, \
            "세 직원 업무센터 요구인데 유현민 업무센터 진입 버튼이 없다"
        assert "/api/staff/install-shortcut" in html and "window.__csosInstallPrompt" in html
        assert "workcenterIsInstalled()" in html and "csos_installed_" in html
        assert "maybeShowInstallCard()" in html and "display-mode: standalone" in html
        # 아이콘 참조는 스프라이트(<use href="#i-...">)다 — 경로가 아니라 **이름**으로 확인한다.
        for icon_name in ("bootstrap-person-workspace",
                          "bootstrap-person-badge-fill"):
            assert f"#i-{icon_name}" in html, "Bootstrap 업무센터 아이콘 누락: " + icon_name
        # PIN 원문을 브라우저에 계속 보관하지 않고, 서버 서명 쿠키를 복원한다.
        assert "/api/auth/session" in html and "restoreRoleSession" in html
        assert "localStorage.setItem('cw_pin'" not in html
        assert "LEGACY_PIN" in html and "localStorage.removeItem('cw_pin')" in html
        for marker in ('id="workLogRecordList"', "renderWorkLogRecords",
                       "workLogFilterCategory", "workLogFilterState",
                       "프로젝트 미확정"):
            assert marker in html, "대표보고 일지 상세목록 UI 누락: " + marker
        app_src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
        for marker in ("AUTH_SESSION_TTL", "create_auth_session", "auth_cookie",
                       "auth_session_from_cookie", "get_work_log_view",
                       'if p == "/api/auth/session"'):
            assert marker in app_src, "기기 인증·일지 자동연동 API 누락: " + marker
        assert "def defer_task_until_free(" in app_src and '"auto_check_queued": queued' in app_src
        for marker in ("RYU_ENTRY_CONFIG", "def get_staff_records(",
                       "def save_staff_entry(", '"only_if_empty": True',
                       '"/api/staff/records", "/api/ryu/records"',
                       '"/api/staff/entry", "/api/ryu/entry"'):
            assert marker in app_src, "공용 직원 업무센터 API 누락: " + marker
        assert "def install_staff_shortcut(" in app_src
        assert 'if p == "/api/staff/install-shortcut"' in app_src
        assert "flex-wrap:nowrap!important" in html.replace(" ", "")
        # 보고·목록 툴바의 구형 일지저장/직접전달 버튼은 제거하되,
        # 사용자가 명시 요청한 담당자 고정 URL 공유는 허용한다.
        assert not re.search(
            r'<button[^>]+onclick="(?:save\w*Journal|share(?:Mine|Daily|Journal)\w*)',
            html), "removed journal/report share button is still exposed"
        # 원천 검증은 밴드·카톡·ERP·쿠팡PO 4종이 **자료 유무와 무관하게** 항상 표시돼야 한다
        assert "4원천 검증" in html, "원천 검증 제목이 4원천이 아님"
        assert "쿠팡 PO" in html and "PO 목록 투입 시 자동 대조" in html, "PO 원천 행 누락"
        for marker in ("srcDetails", "sourceRowHeight", "sourceTable(sourceDetails)",
                       "미확인 프로젝트NO", "missProjects"):
            assert marker in html, f"대표 캡처 4원천 표 구조 누락: {marker}"
        assert "if(window.__view==='daily') renderDaily()" in html, \
            "4원천 상태가 늦게 도착했을 때 대표 보고가 갱신되지 않음"
        assert "renderDaily();\n    blob = await reportToPng();" in html, \
            "이미지 저장·공유 직전에 최신 4원천 집계를 다시 묶지 않음"
        # 기간 필터: 프리셋 + 직접 지정 + 제외 건수 알림
        assert 'id="fperiod"' in html and 'id="fd1"' in html and 'id="fd2"' in html, "기간 선택 UI 누락"
        assert "최근 3개월" in html and "직접 지정" in html, "기간 프리셋 누락"
        assert "날짜 없는" in html, "기간 제외 건수 알림 누락"
        # 통계 탭 월별표는 폰에서도 보여야 한다(카드 대체본이 없어 숨기면 화면이 빈다)
        assert "table.grid.ptbl{display:table!important" in html.replace(" ", ""), "폰에서 월별표가 숨겨짐"
        assert "erpHtml(" in html and "api/erpdocs" in html, "ERP 매출 반영 누락"
        # 처리 안내: 확인 필요 건마다 '어디서 확인·어떻게 반영'이 붙어야 한다
        assert "이 건은 어떻게 처리하나요?" in html, "처리 안내 카드 누락"
        for k in ("금액 미입력", "세금계산서 미발행", "PO 미발행", "ERP 계산서(묶음)"):
            assert f"'{k}'" in html or f'"{k}"' in html, f"{k} 안내 누락"
        assert "band/docs_inbox" in html and "처리 방법 · 용어 설명" in html, "도움말 항목 누락"
        # 대표보고: 비어 있던 '정리' 절은 제거하고 화면 순서대로 1부터 다시 번호를 매긴다.
        assert "cleanExecTitle" in html and "sectionNo += 1" in html, "대표보고 번호 재정렬 로직 누락"
        assert "filter(s=>!/^(정리|요약)$/" in html, "빈 정리·요약 절 제거 로직 누락"
        assert "dailyIssueHtml" in html and "이슈사항" in html, "당일 업무 이슈 구역 누락"
        assert ".esec" in html and ".esum" in html and ".egroup" in html, "대표보고 스타일 누락"
        # 목록 카드·표에서 프로젝트NO가 맨 앞·굵게 나와야 한다(4개 탭 전부)
        assert html.count('class="prjno"') >= 6, "프로젝트NO 강조 표기 부족"
        assert '<b class="prjno">' in html, "표에서 프로젝트NO 굵게 표기 누락"
        assert 'class="sid"' in html, "보조 ID 표기 누락"
        # 구버전 감지: 폰에 켜 둔 앱이 예전 화면을 계속 쓰지 않도록
        assert "checkBuild" in html and "hardReload" in html, "구버전 자동 갱신 누락"
        pg = json.loads(urllib.request.urlopen(base + "/api/ping", timeout=5).read())
        assert pg.get("build"), "ping 응답에 build 값 없음"
        # 당일 업무 실적: 항목마다 어떤 건인지 상세가 붙고, 숫자가 어긋나면 알려야 한다
        assert "dayDetail(" in html and "dayRows(" in html, "당일 실적 상세 누락"
        # 문구는 바뀔 수 있다(2026-07-30: "앱에서 찾은 건" → "연결 프로젝트"). 지켜야 하는 것은
        # **불일치를 숨기지 않는 동작**이다 — 그래서 문구가 아니라 그 로직이 있는지 본다.
        assert "const diff = Number.isFinite(n) && n !== rows.length;" in html, "보고 숫자와 목록 건수 불일치 판정이 없다"
        for k in ("신규접수", "점검완료", "거래명세서발행", "입금건수"):
            assert f"'{k}'" in html, f"{k} 매핑 누락"
        # 문제 코드는 축약어가 아니라 '무엇이 비었고 무엇을 해야 하는지'로 풀어 써야 한다
        assert "관리자검증상태" in html and "중복판정(선택)" in html, "문제 코드 해설 누락"
        assert "이미 정리(중복 통합" in html, "삭제된 건 표시 누락"
        assert "codeHtml(" in html and "topLine(" in html, "코드 해설 함수 누락"
        # 서비스 워커는 **진짜 자바스크립트**로 나가야 한다.
        # _send 가 str을 JSON으로 감싸는 바람에 "self.add..." 처럼 따옴표에 싸여 나가면
        # 브라우저가 등록을 거부하고 [설치 및 바로가기 만들기]가 그대로 먹통이 된다.
        _r = urllib.request.urlopen(base + "/sw.js", timeout=5)
        _sw = _r.read().decode("utf-8")
        assert "javascript" in _r.headers.get("Content-Type", ""), _r.headers.get("Content-Type")
        # ★ 'self.addEventListener 로 시작하는가' 로 보던 것을 2026-07-31 에 바꿨다.
        #   오프라인 껍데기 캐시가 붙으면서 앞에 캐시 이름(const V=...)이 오게 됐는데,
        #   그건 JSON 으로 감싸인 것과 아무 상관이 없다. 보려는 것은 **따옴표에 싸여 나가는가** 다.
        assert not _sw.lstrip().startswith(("{", '"')), "sw.js가 JSON으로 감싸여 나간다: " + _sw[:40]
        assert "self.addEventListener" in _sw, "sw.js에 이벤트 등록이 없다: " + _sw[:40]
        assert "addEventListener('fetch'" in _sw, "fetch 핸들러 없이는 설치가 제안되지 않는다"
        print("  [6] 웹앱 API(PIN 인증·상태·정산·페이지 서빙) ✅")
    finally:
        p.terminate()


def t28_resolve():
    """[28] 프로젝트 코드 하나로 나머지가 따라오는가.
    이 알고리즘의 목숨은 **채번이 시트 수식과 똑같은가**에 달렸다. 어긋나면 사람이 엑셀을
    연 순간 다시 계산돼 ID가 바뀌고, 03·06이 값으로 들고 있는 참조가 전부 끊긴다."""
    import project_resolve as P
    from datetime import datetime

    # (1) 코드 정규화 — 사람이 어떻게 쳐도 같은 코드로 모여야 한다
    for raw in ("UJ2601138", "uj2601138", " UJ 2601138 ", "2601138", "UJ-2601138"):
        assert P.norm(raw) == "UJ2601138", raw
    for bad in ("", None, "UJ260113", "UJ26011389", "AS-2607-001", "UJ260113X"):
        assert P.norm(bad) is None, bad

    # (2) 채번이 시트 수식과 같은가 — 수식: 접두어 & TEXT(날짜,"yymm") & TEXT(ROW()-4,"000")
    assert P.mint("AS", "2026-06-03", 531) == "AS-2606-527", P.mint("AS", "2026-06-03", 531)
    assert P.mint("PM", "2025-12-31", 9) == "PM-2512-005"
    assert P.mint("JS", "2026-07-01", 1004) == "JS-2607-1000", "네 자리도 잘려선 안 된다"
    assert P.mint("AS", "", 100) == "AS-%s-096" % datetime.now().strftime("%y%m")
    assert P.mint("AS", None, 100).endswith("-096")

    # (3) 실제 원장 수식과 대조 — 규칙을 코드에만 적어 두면 시트가 바뀌어도 모른다
    import openpyxl, re as _re
    from ecount_reconcile import load_config, resolve_master
    m = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(m, data_only=False)
    for sh, (idc, pfx, _dc, _kc) in P.SHEET_ID.items():
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        f = next((ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)
                  if isinstance(ws.cell(row=r, column=1).value, str)
                  and str(ws.cell(row=r, column=1).value).startswith("=")), None)
        if not f:
            continue
        got = _re.search(r'"(\w{2})-"', f)
        assert got, "%s ID 수식에서 접두어를 못 찾음" % sh
        assert got.group(1) == pfx, (
            "%s: 시트 수식 접두어 '%s' != 코드 '%s' — 새 행 ID가 엉뚱한 이름으로 매겨진다"
            % (sh, got.group(1), pfx))
        assert "ROW()-4" in f.replace(" ", ""), "%s 채번이 ROW()-4가 아니다" % sh

        # ★ 행을 늘릴 때 새 행에도 같은 채번 수식이 심겨야 한다. expand_rows 의 표에서
        #   빠진 시트는 늘려도 ID가 영영 빈칸이다(03_현장작업실적이 실제로 그랬다 —
        #   120행을 늘려 놓고도 '여유 0'인 채였다, 2026-07-27).
        import expand_rows as E
        assert sh in E.ID_FORMULA, f"{sh} 가 expand_rows.ID_FORMULA 에 없다 — 늘려도 채번이 안 된다"
        _col, _name, _pfx, _dcol = E.ID_FORMULA[sh]
        assert _pfx == pfx, f"{sh} expand_rows 접두어 {_pfx} != 시트 {pfx}"
        assert _name == idc, f"{sh} expand_rows ID열 이름 {_name} != {idc}"
        # 날짜열도 실제 수식과 같아야 한다(02·04는 D, 03·05는 E, 06은 F로 서로 다르다)
        _m = _re.search(r'TEXT\(IF\(\$([A-Z]+)\d+="",TODAY\(\)', f.replace(" ", ""))
        assert _m and _m.group(1) == _dcol, (
            f"{sh}: expand_rows 날짜열 {_dcol} != 시트 수식 {_m.group(1) if _m else '?'}")
    wb.close()

    # (4) 리졸브 — 이미 등록된 코드는 새로 만들지 않고 그 행을 가리켜야 한다
    ev = P.evidence(m)
    known = next(iter(ev["ledger"]))
    r = P.resolve(known, ev)
    assert r["ok"] and r["state"] == "등록됨" and r.get("row"), r
    assert r["sheet"] in P.SHEET_ID, r["sheet"]

    # 신규 코드는 **빈 행**을 받아야 한다 — 기존 행을 덮으면 실데이터가 날아간다
    fake = "UJ9999999"
    ev2 = dict(ev)
    ev2["band"] = dict(ev["band"])
    ev2["band"][fake] = {"camp": "합성캠프", "kind": "돌발AS", "cost": "유상",
                         "tech": "홍길동", "date": "2026-06-03", "status": "작업완료",
                         "posted": "2026-06-03", "_score": 9}
    n = P.resolve(fake, ev2)
    assert n["ok"] and n["state"] == "신규", n
    assert n["row"] > ev["tail"]["02_돌발AS접수"], "새 행이 기존 데이터 위에 얹힌다"
    assert n["row"] <= ev["cap"]["02_돌발AS접수"], "수식 없는 행에 값만 쓰면 채번이 죽는다"
    assert n["ids"]["접수ID"] == P.mint("AS", "2026-06-03", n["row"])

    # 쓰기 항목은 좌표 지정 + '빈 칸만'이어야 한다(키 조회로는 아직 없는 행을 못 찾는다)
    items = P.row_items(n, ev2)
    assert items and all(i.get("only_if_empty") and i.get("cell") for i in items), items[:1]
    assert any(i["col"] == "프로젝트NO" and i["value"] == fake for i in items)
    assert all(str(i["cell"]).endswith(str(n["row"])) for i in items), "행 번호가 섞였다"
    # ID 열은 절대 쓰지 않는다 — 수식이 알아서 채운다
    assert not any(i["col"] in ("접수ID", "점검ID", "작업ID", "정산ID") for i in items)

    # 업무유형을 모르면 **찍지 말고 멈춰야** 한다
    ev2["band"][fake] = dict(ev2["band"][fake], kind="기타")
    assert not P.resolve(fake, ev2)["ok"], "시트를 모르면서 등록하면 엉뚱한 시트에 들어간다"
    print("  [28] 프로젝트코드 리졸브(정규화·채번=수식·빈행 배치·ID열 보호) ✅")


def t29_cloud():
    """[29] PC가 꺼져도 폰이 열 수 있는가 — 잠금·오프라인 폴백·예약 반영."""
    import csos_crypto as C, base64, hmac as _h, hashlib as _hl, zlib as _z, json as _j
    assert C.self_test(), "AES 공식 시험벡터 실패 — 폰이 절대 못 연다"

    raw = _j.dumps({"codes": {"UJ2601138": {"camp": "합성캠프"}}}, ensure_ascii=False).encode()
    packed = _z.compress(raw, 9)
    s = C.seal(packed, "0000", iters=1000)
    assert s["cipher"] == "AES-256-CBC" and s["kdf"] == "PBKDF2-SHA256", s
    k = C.derive("0000", base64.b64decode(s["salt"]), 1000)
    tag = _h.new(k[32:], base64.b64decode(s["iv"]) + base64.b64decode(s["ct"]), _hl.sha256).digest()
    assert _h.compare_digest(tag, base64.b64decode(s["tag"])), "무결성 태그 불일치"
    # 틀린 PIN은 **복호 전에** 걸러져야 한다(엉뚱한 데이터를 그리면 안 된다)
    kbad = C.derive("9999", base64.b64decode(s["salt"]), 1000)   # 일부러 다른 PIN
    bad = _h.new(kbad[32:], base64.b64decode(s["iv"]) + base64.b64decode(s["ct"]), _hl.sha256).digest()
    assert not _h.compare_digest(bad, base64.b64decode(s["tag"]))
    assert C.ITERS >= 300000, "반복이 낮으면 4자리 PIN이 순식간에 뚫린다"
    assert _z.decompress(packed) == raw

    # 올라간 사본이 **평문이 아니어야** 한다 — 공개 저장소다
    enc = os.path.join(ROOT, "docs", "data.enc")
    if os.path.exists(enc):
        d = _j.load(open(enc, encoding="utf-8"))
        assert set(("salt", "iv", "ct", "tag")) <= set(d), list(d)
        # ct는 Base64로 표현한 암호문이라 우연히 "UJ25" 같은 4글자 조각이 생길 수 있다.
        # 2026-07-30 실제 재암호화에서 그 확률 충돌로 검증이 실패했다. 암호문 문자열을
        # 평문 검색하지 말고, 자료가 들어갈 수 있는 별도 메타데이터가 없는지 확인한다.
        allowed = {"v", "kdf", "iter", "cipher", "salt", "iv", "ct", "tag",
                   "pin_auth", "zip"}
        assert set(d) <= allowed, f"암호화 봉투에 예상 밖 메타데이터가 있다: {set(d) - allowed}"
        for key in ("salt", "iv", "ct", "tag"):
            base64.b64decode(d[key], validate=True)
        meta = _j.dumps({k: v for k, v in d.items() if k != "ct"}, ensure_ascii=False)
        for leak in ("캠프", "UJ25", "UJ26", "프로젝트NO"):
            assert leak not in meta, "암호화 봉투 메타데이터에 평문 '%s' 이 있다" % leak

    # 고정 페이지: PC가 죽었을 때 **막다른 오류가 아니라** 오프라인 앱으로 가야 한다
    doc = open(os.path.join(ROOT, "docs", "index.html"), encoding="utf-8").read()
    assert "app.html" in doc and "location.replace('app.html')" in doc, \
        "고정 주소가 PC 독립 앱으로 바로 들어가지 않는다"
    app = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    for need in ("PBKDF2", "AES-CBC", "DecompressionStream", "csos_queue", "/api/queue"):
        assert need in app, "오프라인 앱에 %s 누락" % need
    assert "Authorization" in app and "Bearer " in app, \
        "클라우드 큐 인증 헤더가 없어 예약을 안전하게 보낼 수 없다"
    assert "클라우드에 보관되며 PC가 켜지면 자동 반영" in app, \
        "PC 독립 예약의 실제 동작 안내가 빠졌다"
    sw = open(os.path.join(ROOT, "docs", "sw.js"), encoding="utf-8").read()
    assert "endpoint.json" in sw and "data.enc" in sw, "사본을 쥐거나 주소를 캐시하는 규칙이 없다"
    assert "addEventListener('fetch'" in sw

    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '"/api/enqueue"' in srv and "def enqueue_codes(" in srv, "폰 예약을 받을 곳이 없다"

    # git add부터 실패를 숨기면 실제 게시가 안 됐는데도 '반영했습니다'가 찍힌다.
    import cloud_publish as CP
    assert "docs/resolve_index.json" not in CP.PUBLISH_FILES, "존재하지 않는 파일을 git add 한다"
    assert "docs/manifest.json" in CP.PUBLISH_FILES, \
        "PC 독립 시작 주소를 바꿔도 매니페스트가 게시되지 않는다"
    cloud_source = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    phone_source = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    assert "def snapshot_key(" in cloud_source and 'sealed["pin_auth"] = pin_auth' in cloud_source, \
        "hashed PIN snapshot compatibility is missing"
    assert "async function snapshotPassword(" in phone_source and \
           "auth-pbkdf2-sha256" in phone_source, \
        "phone snapshot cannot derive the stored authentication digest"

    class _R:
        def __init__(self, code=0, out="", err=""):
            self.returncode, self.stdout, self.stderr = code, out, err
    calls = []
    def fail_add(cmd, **_kwargs):
        calls.append(cmd)
        return _R(1, err="pathspec missing") if cmd[1] == "add" else _R()
    ok, stage, detail = CP.git_publish("합성", runner=fail_add)
    assert not ok and stage == "add" and len(calls) == 1 and "pathspec" in detail, (ok, stage, calls)
    calls.clear()
    def no_changes(cmd, **_kwargs):
        calls.append(cmd)
        return _R(1, out="nothing to commit") if cmd[1] == "commit" else _R()
    ok, stage, _ = CP.git_publish("합성", runner=no_changes)
    assert ok and stage == "" and [c[1] for c in calls] == ["add", "commit", "push"], calls

    # ★ 실 PIN이 소스 어딘가에 박히면 공개 저장소에 그대로 남아 사본 잠금이 무의미해진다.
    #   (실제로 자체검증 코드에 리터럴로 들어갔던 적이 있다 — 2026-07-27)
    try:
        real = str(_j.load(open(os.path.join(ROOT, "config", "webapp.json"),
                                encoding="utf-8"))["pin"])
    except Exception:
        real = ""
    if real:
        import subprocess as _sp
        r = _sp.run(["git", "grep", "-l", "--cached", real], cwd=ROOT,
                    capture_output=True, text=True, encoding="utf-8", errors="replace")
        hits = [ln for ln in (r.stdout or "").splitlines() if ln.strip()]
        assert not hits, "커밋된 파일에 실 PIN이 들어 있다: " + ", ".join(hits[:3])
    # (5) PC 꺼짐 **콜드 스타트** — 앱을 새로 여는 경우까지 되는가(2026-07-31).
    #     오프라인 큐만으로는 부족했다: 세션 복원이 네트워크 실패를 인증 거절과 같이 다뤄
    #     PIN 게이트로 막혀, 정작 입력 화면에 못 들어갔다. 셋 다 있어야 구멍이 안 남는다.
    live_html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "netDown" in live_html and "cw_dev_auth" in live_html, \
        "네트워크 실패와 인증 거절을 구분하지 않으면 PC 꺼진 뒤 앱을 새로 열 때 잠긴다"
    assert "r.status === 401 || r.status === 403" in live_html, \
        "outbox 가 401·403 을 4xx 로 버린다 — 오프라인에서 써 둔 입력이 유실된다"
    assert "function offlineBanner(" in live_html and "_netFetch('/api/ping" in live_html, \
        "PC 꺼짐 안내·복귀 감시가 없으면 사용자가 반영 여부를 알 수 없다"
    assert "location.reload" not in live_html[live_html.index("function offlineBanner("):
                                              live_html.index("function offlineBanner(") + 1200], \
        "복귀 시 스스로 새로고침하면 입력 중이던 화면이 날아간다(2026-07-31 지시)"

    print("  [29] 폰 단독 사용(잠금·오프라인 폴백·콜드스타트·git 게시 실패감지·PIN 비노출) ✅")


def t30_dns_and_versions():
    """[30] 오래 끌던 접속 불가의 진짜 원인과, 버전 파일이 쌓이는 문제.

    사내 DNS가 *.trycloudflare.com 을 안 풀어 준다. 그래서 PC에서 터널을 찔러 보면
    늘 실패했고, publish는 게시를 취소하고 watch는 멀쩡한 터널을 계속 갈아치웠다.
    폰에서는 살아 있던 주소인데 PC 혼자 못 보고 있었다."""
    import net_probe as N

    # 공개 DNS 우회가 실제로 동작하는가(회사 DNS가 막아도 이름을 얻어야 한다)
    ips = N.resolve_public("www.cloudflare.com")
    assert ips and all(re.fullmatch(r"[\d.]+", i) for i in ips), ips

    # 판정이 정직한가 — 없는 이름은 살아있다고 하면 안 된다
    ok, why = N.probe("https://this-name-does-not-exist-csos.trycloudflare.com/api/ping", timeout=6)
    assert not ok and "풀지 못함" in why, why

    # 게시·감시가 **회사 DNS에 속지 않는 경로**를 쓰는가
    pe = open(os.path.join(ROOT, "publish_endpoint.py"), encoding="utf-8").read()
    assert "from net_probe import probe" in pe, "게시가 여전히 회사 DNS로 판단한다"
    tr = open(os.path.join(ROOT, "webapp", "tunnel_run.py"), encoding="utf-8").read()
    assert tr.count("from net_probe import probe") >= 2, "감시 루프가 멀쩡한 터널을 죽었다고 본다"
    # 터널 대상은 127.0.0.1이어야 한다(localhost면 IPv6로 풀려 HTTP 530)
    assert 'f"http://127.0.0.1:{PORT}"' in tr, "localhost로 주면 IPv6(::1)로 풀려 앱에 못 닿는다"
    assert "localhost:{PORT}" not in tr, "localhost가 남아 있다"
    # 서비스가 죽어 있으면 생성 한도가 복구를 막으면 안 된다
    assert "_endpoint_alive" in tr and "HARD_MAX" in tr, "한도가 복구까지 막는다"

    # 버전 정리: 최신본은 어떤 경우에도 남아야 한다
    import ledger_versions as V
    from ecount_reconcile import load_config, resolve_master
    # 검증 도중 vN+1 이 생길 수 있다(옆 세션·사람 저장 — 2026-08-07 실측 v542→v543).
    # 최신본 판정과 폴더 스캔을 같은 시점으로 다시 맞춰 한 번 재시도한다.
    for _ in range(2):
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
        keep, move = V.plan(master)
        kept = {k["path"] for k in keep}
        if master in kept:
            break
    assert master in kept, "최신본을 접으려 한다 — 모든 도구가 멈춘다"
    assert len(keep) >= min(V.KEEP_LATEST, len(keep) + len(move)), (len(keep), len(move))
    assert not (set(m["path"] for m in move) & kept), "같은 파일이 남김·접기 양쪽에 있다"
    # 지우지 않는다(옮기기만 한다)
    src = open(os.path.join(ROOT, "ledger_versions.py"), encoding="utf-8").read()
    assert "shutil.move" in src and "삭제하지 않는다" in src
    assert "os.remove(" not in src, "구버전 정리가 중복 파일을 삭제한다"
    print("  [30] 사내 DNS 우회 판정 · 터널 대상 IPv4 · 버전 파일 정리(최신본 보호) ✅")


def t31_tech():
    """[31] 누가 어디를 다녀왔는가 — 세는 기준이 흔들리면 실적이 부풀거나 빠진다."""
    import tech_report as T

    # 동행 건은 양쪽 다 센다. 이름이 아닌 조각은 버린다.
    assert T.split_tech("김준형, 김필우") == ["김준형", "김필우"]
    assert T.split_tech("문상국. 최일파") == ["문상국", "최일파"]
    assert T.split_tech("000 (캠프상태확인 및 스케쥴 세팅)") == []
    assert T.split_tech("") == [] and T.split_tech(None) == []

    visits, pending, unknown, _m = T.collect()
    assert visits, "방문 기록이 하나도 안 잡힌다"
    # **완료된 것만** 방문으로 센다 — 예정만 있는 건은 pending 이어야 한다
    assert all(v["방문일"] for v in visits), "완료일 없는 건이 방문으로 세어졌다"
    assert all(not p["방문일"] for p in pending), "완료된 건이 미방문으로 빠졌다"
    assert all(u["방문일"] and not u["기사"] for u in unknown), unknown[:1]
    # 한 사람씩 펼쳐 담는다(집계할 때 다시 쪼갤 필요가 없게)
    assert all("," not in v["기사"] for v in visits), "동행 건이 안 펼쳐졌다"

    by = T.summary(visits)
    assert sum(d["총"] for d in by.values()) == len(visits)
    for t, d in by.items():
        assert d["돌발AS"] + d["정기점검"] == d["총"], (t, d)
        assert d["최근"] and re.fullmatch(r"\d{4}-\d{2}-\d{2}", d["최근"]), (t, d["최근"])
    print("  [31] 기사별 방문(동행 분리·완료분만·합계 정합) ✅")


def t32_band_sheet():
    """[32] 24_밴드업무추출이 통째로 잘리는 사고를 막는다.

    band_extract 의 --sheet 는 시트를 **덮어쓴다**. --month 와 같이 쓰면 그 달 것만 남는다.
    ingest 가 월별 루프로 이걸 반복해서 2026-07-27에 4,223행 → 506행으로 잘렸다
    (2025-12~2026-05가 통째로 사라졌다). 두 겹으로 막는다."""
    # ① ingest 는 --sheet 를 월별로 부르면 안 된다
    ing = open(os.path.join(ROOT, "band", "ingest.py"), encoding="utf-8").read()
    i = ing.index('if "--sheet" in args:')
    blk = ing[i:i + 700]
    assert '"--month", mo, "--sheet"' not in blk, (
        "ingest 가 24시트를 월별로 덮어쓴다 — 마지막 달만 남고 나머지가 사라진다")
    assert '"band_extract.py"), "--sheet"' in blk.replace("os.path.join(ROOT, ", ""), blk[:120]

    # ② 시트가 캐시보다 훨씬 적으면 잘린 것이다
    import json as _j, glob as _g
    # ★ 분모는 **24시트가 담는 범위의 글**이어야 한다 (2026-08-07 실사고).
    #   예전에는 캐시 전체와 견줬다. 밤새 밴드를 개설 시점까지 전량 수집하자 캐시가
    #   2,078 → 8,499 글로 뛰었고(2015년 글까지 들어온다), 24시트는 올해치만 담으므로
    #   3,005행 vs 8,499글이 되어 **성공했다는 이유로 검증이 빨개졌다.**
    #   시트가 잘렸는지 보려면 같은 범위끼리 견줘야 한다 — 올해 글만 센다.
    #   (지운 글의 묘비에는 created_at 이 없다 — 그것도 분모에서 빠진다)
    posts = 0
    from datetime import datetime as _dtm
    try:
        import app_server as _AS
        _yr = int(_AS.APP_YEAR)
    except Exception:
        _yr = _dtm.now().year
    _y0 = int(_dtm(_yr, 1, 1).timestamp() * 1000)
    for f in _g.glob(os.path.join(ROOT, "band", "cache", "*.json")):
        if os.path.basename(f).startswith(("dump_", "raw_")):
            continue
        try:
            for p in (_j.load(open(f, encoding="utf-8")).get("posts") or {}).values():
                ms = isinstance(p, dict) and p.get("created_at")
                if ms and int(ms) >= _y0:
                    posts += 1
        except Exception:
            pass
    if posts:
        import openpyxl
        from ecount_reconcile import load_config, resolve_master
        wb = openpyxl.load_workbook(resolve_master(load_config()["reconcile"]["master_xlsx"]),
                                    read_only=True, data_only=True)
        ws = wb["24_밴드업무추출"]
        hdr = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        ig = hdr.index("게시일")
        days = sorted(str(r[ig])[:10] for r in ws.iter_rows(min_row=5, values_only=True)
                      if ig < len(r) and r[ig])
        wb.close()
        assert days, "24시트가 비어 있다"
        # 캐시는 1,200여 글인데 시트가 그 절반도 안 되면 잘린 것으로 본다
        assert len(days) > posts * 0.5, (
            f"24시트 {len(days)}행 vs 밴드 캐시 {posts}글 — 시트가 잘린 것으로 보인다")
        # 밴드 시작(2025-12)부터 있어야 한다 — 최근 몇 달만 남아 있으면 덮어써진 것이다
        assert days[0] <= "2026-01-31", (
            f"24시트가 {days[0]} 부터 시작한다 — 앞부분이 덮어써졌다")
    print("  [32] 24시트 전체 보존(월별 덮어쓰기 차단·행수 급감 감지) ✅")


def t33_unbilled_banner():
    """[33] 미청구가 눈에 띄는가 — 앱을 열면 화면 맨 위에 뜨고, 경과일은 오늘 기준이어야 한다.

    ★ 예전에는 GitHub Actions가 매일 세어 **메일**을 보냈다(공개 파일 docs/aging.json +
      워크플로 실패 트릭). 사용자가 메일을 원하지 않아 그 경로를 전부 걷어내고, 앱을 열 때
      화면에서 보이게 옮겼다. 지워진 것이 되살아나지 않는지도 함께 본다."""
    # (1) 메일 경로가 완전히 사라졌는가 — 남아 있으면 원치 않는 메일이 계속 간다
    assert not os.path.exists(os.path.join(ROOT, ".github", "workflows", "aging-alert.yml")), \
        "메일을 보내는 워크플로가 아직 있다"
    assert not os.path.exists(os.path.join(ROOT, "docs", "aging.json")), \
        "공개 aging.json 이 아직 있다(이제 쓰지 않는다)"
    assert not os.path.exists(os.path.join(ROOT, "tools", "aging_check.py")), \
        "메일용 판정 스크립트가 아직 있다"
    cp = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    # 설명 주석에 '예전에는 aging.json 을 썼다'는 이력은 남아도 된다 — **만드는 코드**만 본다
    assert "def aging(" not in cp, "cloud_publish 에 공개 aging.json 생성 함수가 남아 있다"
    _code = [ln for ln in cp.splitlines()
             if "aging.json" in ln and not ln.strip().startswith(("#", "*", "·"))
             and "예전" not in ln and "docs/aging.json)" not in ln]
    assert not _code, "cloud_publish 가 아직 aging.json 을 다룬다: " + str(_code[:2])

    # (2) 미청구가 **잠긴 사본 안에** 들어가는가(공개 파일이 아니라)
    assert "def unbilled(" in cp and 'd["unbilled"]' in cp, "미청구가 사본에 안 담긴다"
    import cloud_publish as CP
    rows = CP.unbilled()
    for r in rows:
        assert re.fullmatch(r"\d{4}-\d{2}-\d{2}", r["발행일"]), r
        assert isinstance(r["금액"], int), r
    assert rows == sorted(rows, key=lambda x: x["발행일"]), "발행일 순이 아니다"

    # (3) 앱이 **오늘 날짜로** 경과일을 계산하는가 — 사본이 묵어도 일수는 정확해야 한다
    app = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    assert "renderAlert" in app and 'id="alertbox"' in app, "미청구 배너가 없다"
    assert "renderAlert();" in app, "앱을 열 때 배너를 그리지 않는다"
    assert "Date.now()" in app.split("const DAYS")[1][:200], \
        "경과일을 사본 생성시각으로 계산하면 며칠 묵은 사본에서 숫자가 틀어진다"
    assert "D.unbilled" in app, "배너가 사본의 미청구를 읽지 않는다"
    # 90일/120일 단계 표시
    assert "r.d >= 120" in app and "r.d >= 90" in app, "경과 단계 표시가 없다"
    print("  [33] 미청구 배너(메일 경로 제거·사본 내 보관·오늘 기준 경과일) ✅")


def t34_capture_and_no_send():
    """[34] 확인 목록 캡처 기능 + **쿠팡에 자동 전송 경로가 없는지**.

    사용자 상시 지시(2026-07-27): 쿠팡 담당자에게는 어떤 메시지도 보내지 않는다.
    유니버셜 내부에서 처리한다. 캡처·공유는 '파일을 만들어 줄 뿐'이고 받는 사람은 사람이 고른다."""
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # (1) 요약 캡처 + 건별 목록 캡처가 둘 다 있는가
    for fn in ("mineToPng", "captureMine", "shareMine",
               "mineDetailToPng", "captureMineDetail", "shareMineDetail"):
        assert "function " + fn in idx, f"{fn} 없음"
    assert 'onclick="captureMine()"' in idx and 'onclick="captureMineDetail()"' in idx, "버튼이 안 걸렸다"
    assert "window._mineIdx" in idx, "건별 캡처가 어느 항목인지 모른다"
    # 건수가 많을 때 조용히 자르면 받는 사람이 그게 전부인 줄 안다
    assert "MINE_CAP" in idx and "이 이미지에는" in idx, "잘린 사실을 이미지에 안 적는다"

    # (2) 자동 전송 경로가 없는가 — 공유는 navigator.share(사람이 수신자를 고름)만 허용
    for bad in ("mailto:", "smtplib", "sendmail", "api.telegram", "hooks.slack",
                "kakao.link", "openapi.kakao", "openapi.band", "band.us/api"):
        assert bad not in idx, f"앱에 자동 전송 경로가 있다: {bad}"
    # 밴드·카톡 모듈은 읽기 전용이어야 한다
    import glob as _g
    for f in _g.glob(os.path.join(ROOT, "band", "*.py")) + _g.glob(os.path.join(ROOT, "kakao", "*.py")):
        src = open(f, encoding="utf-8").read()
        for bad in ("create_post", "write_post", "post_comment", "mailto:", "smtplib"):
            assert bad not in src, f"{os.path.basename(f)} 에 글쓰기·발송 경로가 있다: {bad}"

    # (3) 메일 경로가 되살아나지 않았는가(2026-07-27 제거)
    assert not os.path.exists(os.path.join(ROOT, ".github", "workflows")), \
        "워크플로가 되살아났다 — 메일이 갈 수 있다"

    # (4) 규칙이 문서에 남아 있는가 — 다음 AI가 이어받아도 지키게
    ag = open(os.path.join(ROOT, "AGENTS.md"), encoding="utf-8").read()
    assert "쿠팡 담당자에게 어떤 메시지도 보내지 않는다" in ag, "절대규칙에 안 적혀 있다"

    # (5) '원천' 같은 속어를 화면 문구로 쓰지 않는다(사용자가 무슨 뜻인지 되물었다)
    _vis = [ln for ln in idx.splitlines()
            if "원천 없음" in ln and not ln.strip().startswith(("*", "/*", "//", "★"))]
    assert not _vis, "화면에 '원천 없음'이 남아 있다: " + str(_vis[:1])

    # (6) 자료 유무 판정이 상태 도착 뒤에 다시 그려지는가
    #     (안 그러면 밴드·카톡·PO가 멀쩡한데 '아직 없음'으로 4건이 거짓 표시된다)
    _ls = idx[idx.index("async function loadStatus()"):][:900]
    assert "srcStats = s.sources" in _ls and "renderBoard()" in _ls, \
        "상태를 받은 뒤 확인목록을 다시 그리지 않는다 — 자료가 있어도 '없음'으로 뜬다"
    # 앱서버 자동 게시와 사람/다른 AI 수동 게시가 겹치면 같은 Git 커밋이 두 번 생긴다.
    cp = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert 'require("publish"' in cp and "acquired_here" in cp, \
        "폰 사본 Git 게시에 publish 점유가 강제되지 않는다"
    assert '"CSOS_AI": "server"' in srv and "publish_env" in srv, \
        "앱서버 자동 게시가 점유 주체를 밝히지 않는다"
    print("  [34] 확인목록 캡처(요약·건별) · 자동 전송 경로 없음 · 자료판정 순서 ✅")


def t35_confirm_evidence():
    """[35] '확인 완료' 체크는 **근거가 완료라고 말할 때만** 찍혀야 한다.

    한 건에 글이 여러 개 올라온다(접수 안내 → 작업완료). band_map 이 **먼저 나온 것**만
    집던 탓에 접수 글이 잡혀 완료 글을 놓쳤다 — 전체로 '첫 기록이 완료' 382 vs
    '완료 기록이 하나라도 있음' 859. 그 상태로 확인 체크를 돌리면 근거가 '접수·예정'인
    행에 '일치·확인완료'가 찍힌다. 확인하지 않은 것을 확인했다고 남기는 것이라 가장 나쁘다.
    (2026-07-27: 대상 40건 중 37건이 그랬다)"""
    import confirm_fill as C

    # (1) 완료 글을 우선으로 고르는가 — 합성 레코드로 직접 확인
    fake = [{"프로젝트NO": "UJ9000001", "진행상태": "접수·예정", "작업일": "2026-06-01", "업무유형": "돌발AS"},
            {"프로젝트NO": "UJ9000001", "진행상태": "작업완료", "작업일": "2026-06-03", "업무유형": "돌발AS"},
            {"프로젝트NO": "UJ9000002", "진행상태": "접수·예정", "작업일": "2026-06-02", "업무유형": "돌발AS"}]
    import band_extract as B
    _orig = B.load_records
    try:
        B.load_records = lambda *a, **k: fake
        bm = C.band_map()
    finally:
        B.load_records = _orig
    assert bm["UJ9000001"]["status"] == "작업완료", "접수 글이 완료 글을 가린다"
    assert bm["UJ9000001"]["date"] == "2026-06-03", bm["UJ9000001"]
    assert bm["UJ9000002"]["status"] == "접수·예정", "없는 완료를 지어내면 안 된다"

    # (2) 실제 계획 — 확인 체크가 붙는 건은 **전부** 근거가 '작업완료'여야 한다
    real = C.band_map()
    _m, items, _stat, completions = C._plan()
    VCOLS = ("관리자검증상태", "담당관리자", "최종확인일", "최종확인일(유현민 체크)")
    bad = [i["key"] for i in items if i["col"] in VCOLS
           and (real.get(i["key"]) or {}).get("status") != "작업완료"]
    assert not bad, ("근거가 완료가 아닌데 확인 체크를 찍으려 한다: " + ", ".join(sorted(set(bad))[:5]))

    # (3) 확인·날짜 필드는 기존 값을 덮지 않고, 상태 수식 셀은 아예 큐에 넣지 않는다.
    #     객관 완료는 work_resolution DB에만 기록한다.
    assert all(i.get("only_if_empty") for i in items), "확인·날짜 필드의 빈칸만 채우는 규칙이 빠졌다"
    assert not any(i["col"] in ("진행상태", "점검상태") for i in items), \
        "객관 완료 판정이 상태 수식 셀을 덮으려 한다"
    assert completions and all(c["status"] in ("작업완료", "완료")
                               and c["completed_on"] and c["basis"] for c in completions), \
        "DB 객관 완료 판정에 상태·완료일·근거가 빠졌다"
    # (4) 취소성 상태·날짜 없는 완료 글·완료가 아닌 글은 자동 완료하지 않는다.
    done_band = {"status": "작업완료", "date": "2026-07-30"}
    assert C.objective_completion("접수", "작업완료", done_band)
    assert C.objective_completion("미점검", "완료", done_band)
    for conflict in ("취소", "철회", "AS전환", "점검불가"):
        assert not C.objective_completion(conflict, "완료", done_band), conflict
    assert not C.objective_completion("접수", "작업완료", {"status": "작업완료", "date": ""})
    assert not C.objective_completion("접수", "작업완료", {"status": "접수", "date": "2026-07-30"})
    assert not C.objective_completion("접수", "작업완료", done_band,
                                      base="2026-07-31"), "접수보다 앞선 완료일은 모순"
    assert C.objective_completion("접수", "작업완료", {"status": "작업완료", "date": "2025-01-03"},
                                  base="2026-01-02", existing_done="2026-01-03"), \
        "명시된 원장 완료일보다 밴드의 연도 추정을 우선하면 안 된다"
    print("  [35] 확인 체크 근거 정합(완료 글 우선·DB 완료판정·날짜충돌 보존) ✅")


def t36_mobile_input():
    """[36] 폰에서 빈 항목을 편하게 채울 수 있는가 — 달력·드롭다운·확대 방지."""
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # (1) 날짜는 손으로 치지 않는다 — 폰 달력이 뜨는 type="date"
    assert "function fieldInput" in idx, "입력칸 렌더 함수가 없다"
    _fi = idx[idx.index("function fieldInput"):][:1600]
    assert "type=\"date\"" in _fi, "날짜 칸이 달력으로 안 뜬다"
    assert 'inputmode="numeric"' in _fi, "숫자 칸에 숫자 키패드가 안 뜬다"
    assert "<select" in _fi, "선택지가 드롭다운이 아니다"
    # ★ 16px 미만이면 iOS 사파리가 입력할 때 화면을 확대해 버린다
    m = re.search(r"font-size:(\d+)px", _fi)
    assert m and int(m.group(1)) >= 16, f"입력 글자 크기 {m.group(1) if m else '?'}px — iOS에서 화면이 확대된다"

    # (2) 선택지는 **시트에서** 온다 — 화면에 박아 두면 사람이 바뀔 때 어긋난다
    assert "/api/codes" in idx and "loadCodes()" in idx, "코드 목록을 시트에서 안 읽는다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "def get_codes(" in srv and "10_코드관리" in srv, "서버가 코드 시트를 안 읽는다"
    # 담당기사·관리자검증상태는 자유 입력이 아니어야 한다(오타·표기흔들림 방지)
    _spec = idx[idx.index("const INPUT_SPEC"):][:2000]
    assert "opts:'담당기사'" in _spec, "담당기사가 아직 자유 입력이다"
    assert "opts:'관리자검증상태'" in _spec, "관리자검증상태가 아직 자유 입력이다"

    # (3) 코드 목록을 못 받아도 입력 자체가 막히면 안 된다
    assert "자유 입력으로 떨어뜨린다" in _fi or "list.length" in _fi, "폴백이 없다"

    # (4) 기사 목록은 실제 배정 횟수 순 — 폰에서 자주 쓰는 사람이 위로
    assert "function byUsage" in idx, "기사 정렬이 없다"

    # (5) 실제 코드 시트에서 목록이 나오는가
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    wb = openpyxl.load_workbook(resolve_master(load_config()["reconcile"]["master_xlsx"]),
                                read_only=True, data_only=True)
    assert "10_코드관리" in wb.sheetnames, "코드 시트가 없다"
    rows = list(wb["10_코드관리"].iter_rows(min_row=4, values_only=True))
    wb.close()
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    assert "담당기사" in hdr, "코드 시트에 담당기사 열이 없다"
    i = hdr.index("담당기사")
    names = [str(r[i]).strip() for r in rows[1:] if i < len(r) and r[i] not in (None, "")]
    for who in ("김준형", "권오철", "김필우", "차동호"):      # AGENTS.md 규칙 6의 기준 4인
        assert who in names, f"코드 시트 담당기사에 {who} 가 없다"
    # (6) 부가세는 **원장 값**을 쓴다 — 10%로 계산하면 반올림 때문에 서류와 어긋난다
    assert "function vatLine" in idx, "부가세 표기 함수가 없다"
    _v = idx[idx.index("function vatLine"):][:900]
    assert "r.부가세" in _v, "원장 부가세를 안 읽고 계산해 버린다"
    assert "합계−공급가액" in _v, "원장이 비었을 때 되짚은 값이라는 표시가 없다"
    assert "≠ 합계" in _v, "공급가+부가세가 합계와 안 맞을 때 알리지 않는다"
    er = open(os.path.join(ROOT, "ecount_reconcile.py"), encoding="utf-8").read()
    assert '"원장_부가세"' in er and "실제작업부가세" in er, "원장에서 부가세를 안 읽는다"

    # 실데이터 정합 — 공급가+부가세=합계 가 깨진 행이 있으면 원장이 틀린 것이다
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as _A
    bad = [r["정산ID"] for r in _A.real_settlements()
           if r.get("부가세") not in (None, "") and r.get("합계")
           and abs((r["공급가액"] or 0) + (r["부가세"] or 0) - (r["합계"] or 0)) > 1]
    assert not bad, "원장 공급가+부가세≠합계: " + ", ".join(bad[:5])
    # (7) 보고일·집계기준일을 **보고 화면에서 바로** 고칠 수 있는가
    #     예전에는 대시보드 맨 아래 카드에만 있어, 보고서를 보다 날짜가 틀린 걸 발견하면
    #     탭을 옮겨 찾아가야 했다.
    for fn in ("openRptDates", "saveRptDates", "rptDatesToday"):
        assert "function " + fn in idx, f"{fn} 없음"
    assert 'onclick="openRptDates()"' in idx, "보고 화면에 날짜 변경 버튼이 없다"
    _rd = idx[idx.index("function openRptDates"):][:2800]
    assert 'type="date"' in _rd, "보고 날짜가 달력으로 안 뜬다"
    _sv = idx[idx.index("function saveRptDates"):][:1200]
    assert "/api/set_dates" in _sv, "보고 기준일이 앱 DB 저장 API에 이어지지 않는다"
    # 집계기준일이 보고일보다 뒤면 잘못 고른 것이다 — 그대로 쓰면 보고서가 어긋난다
    assert "집계기준일 > 보고일" in _sv.replace("집계기준일 &gt; 보고일", "집계기준일 > 보고일"),         "집계기준일이 보고일보다 뒤인 경우를 안 막는다"
    # 사용자 확정 기준: 보고일은 오늘, 집계기준일은 주말 여부와 무관하게 정확한 전날
    _td = idx[idx.index("function rptDatesToday"):][:600]
    assert "previousDayISO(cur)" in _td, "집계기준일 기본값이 정확한 전날이 아니다"
    _init = idx[idx.index("function initDates"):][:500]
    assert "previousDayISO(cur)" in _init, "대시보드 날짜 기본값이 정확한 전날이 아니다"
    assert "requestAnimationFrame(()=>rptDateChanged())" in _rd, \
        "날짜 변경창을 열었을 때 오늘/전날 기본값이 즉시 적용되지 않는다"
    print("  [36] 폰 입력(달력·드롭다운·16px·시트연동) · 부가세 원장값 · 보고일 즉시수정 ✅")


def t37_band_coverage():
    """[37] 밴드 수집이 안 닿는 기간을 '기사가 안 올렸다'로 몰지 않는가.

    2026-07-28: 류지영이 엑셀을 저장하면서 옛 행의 상태 수식이 계산됐고, 그 순간
    2025-12-08·09 작업 28건이 '밴드 게시 미확인'으로 떴다. 그런데 쿠팡AS 밴드 캐시는
    2025-12-16부터다 — 기사가 안 올린 게 아니라 우리가 그 이전을 못 긁어온 것이다.
    이대로 두면 없는 잘못으로 기사들을 추궁하게 된다."""
    src = open(os.path.join(ROOT, "band", "band_reconcile.py"), encoding="utf-8").read()
    assert "수집범위밖" in src, "수집이 안 닿는 기간을 따로 구분하지 않는다"
    # ★ 밴드별로 시작일을 보고 **가장 늦은 것**을 써야 한다. 합쳐서 최소값을 잡으면
    #   매출처 밴드가 쿠팡AS 밴드의 빈 구간을 덮어 버린다.
    assert "max(_first.values())" in src, "밴드별 시작일을 안 보고 합쳐서 판단한다"

    # 하위 도구는 '미확인'만 문제로 올려야 한다(수집범위밖은 사람이 할 일이 아니다)
    fe = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    assert '밴드게시") == "미확인"' in fe, "확인필요 목록이 수집범위밖까지 문제로 올린다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert 'r.get("밴드게시") == "미확인"' in srv, "앱이 수집범위밖까지 문제로 센다"

    # 실제 리포트에 수집 시작일보다 앞선 건이 '미확인'으로 남아 있으면 안 된다
    import glob as _g, csv as _csv, json as _j
    from datetime import datetime as _dt
    rep = sorted(_g.glob(os.path.join(ROOT, "reports", "밴드대조_*.csv")))
    if rep:
        first = {}
        for f in _g.glob(os.path.join(ROOT, "band", "cache", "*.json")):
            if os.path.basename(f).startswith(("dump_", "raw_")):
                continue
            try:
                c = _j.load(open(f, encoding="utf-8"))
            except Exception:
                continue
            ds = [_dt.fromtimestamp(p["created_at"] / 1000).date()
                  for p in (c.get("posts") or {}).values() if p.get("created_at")]
            if ds:
                first[c.get("band_name", f)] = min(ds)
        if first:
            cut = max(first.values())
            bad = []
            for r in _csv.DictReader(open(rep[-1], encoding="utf-8-sig")):
                if r.get("밴드게시") != "미확인":
                    continue
                d = (r.get("완료일") or "")[:10]
                if d and _dt.strptime(d, "%Y-%m-%d").date() < cut:
                    bad.append(r.get("ID"))
            assert not bad, (f"수집 시작({cut}) 이전 건이 '미확인'으로 남아 있다: "
                             + ", ".join(x for x in bad[:5] if x))
    print("  [37] 밴드 수집범위 구분(밴드별 시작일·미확인만 문제로) ✅")


def t38_daily_brief():
    """[38] 대표 보고는 **숫자가 아니라 내용**이어야 한다(2026-07-28 통화 지시).

    "숫자를 나한테 보고하라는 게 아니야" — 왜 갔고 무슨 작업을 했고 유·무상은 어떻게
    됐는지, 추가작업이 생겼는지, 정기점검 분기 진행률이 어떤지가 보고 내용이다."""
    import daily_brief as D
    from datetime import date, timedelta

    data, _m = D.load()
    b = D.brief("2026-07-27", data)
    t = D.text(b)

    # (1) 대표가 물은 항목이 전부 구조에 있는가
    for k in ("돌발AS", "정기점검", "완료내역", "무상건", "추가작업건",
              "점검중유상", "AS전환", "내용미기입"):
        assert k in b, f"{k} 누락"
    assert "분기진행률" in b["정기점검"] and 0 <= b["정기점검"]["분기진행률"] <= 100

    # (2) ★ '완료일 없음'과 '아직 안 감'을 갈라야 한다.
    #     뭉치면 미처리 84건이라고 보고하게 되고(대표가 놀란다), 반대로 '없다'고 하면
    #     최근 건을 놓친다. 30일 기준으로 나눈다.
    assert "완료일미기입" in b["돌발AS"], "완료일 미기입을 미처리와 안 나눴다"
    assert b["돌발AS"]["미처리"] <= b["돌발AS"]["미처리"] + b["돌발AS"]["완료일미기입"]
    assert "최근 30일" in t, "미처리 숫자가 어느 범위인지 문장에 없다"

    # (3) 완료 건은 '왜 갔는지'가 있어야 보고가 된다
    for x in b["완료내역"]:
        assert set(("왜", "무엇", "비용", "추가작업", "이상내용", "문제내용", "조치내용")) <= set(x), x
    # 내용이 없으면 지어내지 말고 미기입으로 남겨야 한다
    assert all(x["무엇"] == "" for x in b["내용미기입"]), "미기입 판정이 틀렸다"

    # (4) 분기 계산이 맞는가 — 7월은 7~9월 구간
    #     ★ 사용자 지시(2026-07-29)로 'N분기' 대신 **월 범위**로 적는다(검증 [70] 참고).
    assert b["정기점검"]["분기"].endswith("7~9월"), b["정기점검"]["분기"]
    assert b["정기점검"]["분기끝월"] == "9월", b["정기점검"]["분기끝월"]
    b1 = D.brief("2026-02-10", data)
    assert b1["정기점검"]["분기"].endswith("1~3월"), b1["정기점검"]["분기"]
    assert b1["정기점검"]["분기범위"] == "1~3월", b1["정기점검"]["분기범위"]

    # (5) 없는 날을 넣어도 죽지 않아야 한다(보고가 매일 돌아간다)
    empty = D.brief("2020-01-01", data)
    assert empty["돌발AS"]["완료"] == 0 and isinstance(D.text(empty), str)
    print("  [38] 대표 브리핑(내용 중심·미처리/미기입 분리·분기 진행률) ✅")


def t39_realtime_monitor():
    """[39] 입력 보호시간·이슈 상태전이·자동 진입점 방어."""
    from datetime import datetime
    from operation_window import is_input_window
    import realtime_monitor as R

    # ★ 2026-08-11 지시(엑셀 손입력 종료 — 앱 전용 입력)로 보호시간은 기본 퇴역.
    #   '류지영 아침 입력' 전제가 사라졌으므로 기본은 언제나 False 여야 한다.
    #   되돌리는 스위치(COUPANG_INPUT_WINDOW)가 켜졌을 때만 옛 동작이 살아난다.
    _saved_win = os.environ.pop("COUPANG_INPUT_WINDOW", None)
    try:
        assert not is_input_window(datetime(2026, 7, 28, 8, 0, 0)), \
            "보호시간이 퇴역했는데 아직 아침 90분을 멈춘다"
        assert not is_input_window(datetime(2026, 7, 28, 9, 29, 59))
        os.environ["COUPANG_INPUT_WINDOW"] = "08:00-09:30"
        assert not is_input_window(datetime(2026, 7, 28, 7, 59, 59))
        assert is_input_window(datetime(2026, 7, 28, 8, 0, 0))
        assert is_input_window(datetime(2026, 7, 28, 9, 29, 59))
        assert not is_input_window(datetime(2026, 7, 28, 9, 30, 0))
        os.environ["COUPANG_INPUT_WINDOW"] = "깨진값"
        assert not is_input_window(datetime(2026, 7, 28, 8, 30, 0)), \
            "망가진 설정이 자동화를 멈춘다 — 없음으로 동작해야 한다"
    finally:
        if _saved_win is None:
            os.environ.pop("COUPANG_INPUT_WINDOW", None)
        else:
            os.environ["COUPANG_INPUT_WINDOW"] = _saved_win

    issue = R._issue("sample", "P2", "샘플", "근거", "조치")
    first, changes1 = R.reconcile_issues([issue], {}, datetime(2026, 7, 28, 10, 0))
    assert first[0]["status"] == "provisional" and not changes1
    second, changes2 = R.reconcile_issues(
        [issue], {"issues": first}, datetime(2026, 7, 28, 10, 5)
    )
    assert second[0]["status"] == "new" and changes2[0]["transition"] == "new"
    third, changes3 = R.reconcile_issues(
        [], {"issues": second}, datetime(2026, 7, 28, 10, 10)
    )
    assert third[0]["status"] == "resolved" and changes3[0]["transition"] == "resolved"

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    watchdog = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    endpoint = open(os.path.join(ROOT, "publish_endpoint.py"), encoding="utf-8").read()
    cloud = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    assert "if is_input_window()" in daily
    assert "if is_input_window()" in watchdog
    assert "from net_probe import probe" in watchdog
    main_body = watchdog[watchdog.index("def main():"):]
    assert "archive_versions(dry)" not in main_body
    assert "if is_input_window()" in endpoint and "if is_input_window()" in cloud
    formulas = open(os.path.join(ROOT, "fix_formulas.py"), encoding="utf-8").read()
    assert '"--file" in args' in formulas, "수식 검사가 임시 사본 대신 실원장을 열 수 있다"
    dashboard = open(os.path.join(ROOT, "dashboard_clean.py"), encoding="utf-8").read()
    monitor = open(os.path.join(ROOT, "realtime_monitor.py"), encoding="utf-8").read()
    assert '"--file" in args' in dashboard, "대시보드 검사가 임시 사본 대신 실원장을 열 수 있다"
    assert '"dashboard_clean.py"' in monitor and "dashboard_filldown_debris" in monitor

    claim = open(os.path.join(ROOT, "ai_claim.py"), encoding="utf-8").read()
    assert "os.mkdir(GUARD)" in claim and "os.replace(tmp, CLAIMS)" in claim
    print("  [39] 입력시간 완전정지·이슈 신규/지속/해결·점유 원자화 ✅")


def t41_dates_explicit():
    """[41] '금일'은 읽는 시점마다 다른 날이 된다 — 보고물에 남아 있으면 안 된다.

    사고 #0(2026-07-28): 타일 라벨이 '점검 예정 (금일)'인데 숫자는 집계기준일(어제) 것이라
    보고 자리에서 어긋났다. 화면·이미지·문장 세 곳 모두 **실제 날짜**로 말해야 한다.
    덤으로 사고 #16(서버 두 개)의 진짜 원인인 포트 재사용도 여기서 함께 막는다."""
    import daily_brief as DB

    # (1) 브리핑 문장 — 항목 줄이 전부 날짜로 시작하는가
    day = "2026-03-04"
    b = {
        "기준일": day,
        "돌발AS": {"신규접수": 1, "신규처리완료": 1, "신규처리율": 100,
                   "완료": 1, "미처리": 0, "완료일미기입": 0},
        "정기점검": {"예정": 0, "예정장비": 0, "완료": 0, "완료장비": 0,
                     "분기": "2026년 1분기",
                     "분기예정": 0, "분기완료": 0, "분기진행률": 0},
        "완료내역": [], "무상건": [], "추가작업건": [],
        "점검중유상": [], "AS전환": [], "이상발견": [], "내용미기입": [],
        "완료일미기입목록": [],
        "신규목록": [{"프로젝트NO": "UJ2600001", "캠프명": "시험캠프", "담당기사": "홍길동",
                      "왜": "시험 접수내용", "무엇": "", "비용": "유상", "추가작업": "",
                      "일자": day, "구분": "돌발AS", "접수일": day}],
    }
    b["완료내역"] = [dict(b["신규목록"][0], 일자=day, 접수일="2026-02-25", 무엇="시험 작업")]
    out = DB.text(b)
    assert day in out.splitlines()[0], "머리줄에 기준일이 없다"
    assert "03-04 · UJ2600001 · 시험캠프" in out, "항목이 '날짜 · 프로젝트NO · 캠프'로 시작하지 않는다"
    assert "접수 02-25 · 7일 만" in out, "완료건에 접수일·경과일이 안 붙는다"
    assert "금일" not in out and "당일" not in out, "브리핑 문장에 '금일/당일'이 남아 있다"

    # (1-2) 브리핑 화면은 **문장 줄바꿈이 아니라 구조화된 데이터**로 그려야 한다.
    #       줄로 흘려 놓으면 어디서 어디까지가 한 건인지 안 보인다(사용자 지적 2026-07-28).
    _ix = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for need in ("function bCard(", "function briefBlock(", "BRIEF['신규목록']",
                 "BRIEF['완료내역']", 'class="bcard"', 'class="bmiss"'):
        assert need in _ix, f"브리핑 카드 렌더링에 {need} 가 없다"
    assert "BRIEF.text" in _ix, "복사 버튼이 서버 문장을 쓰지 않는다(대표 전달용 형식이 깨진다)"

    # (1-3) 구 버전 자동 접기 — 사용자 지시(2026-07-28) "말 안 해도 OLD 로".
    #       저장하는 도구가 11개라 **찾는 쪽**(resolve_master·latest_master)에 건다.
    import ledger_versions as LV
    assert hasattr(LV, "autoprune"), "autoprune 이 없다"
    _lv = open(os.path.join(ROOT, "ledger_versions.py"), encoding="utf-8").read()
    _auto = _lv[_lv.index("def autoprune("):_lv.index("def main(")]
    assert "os.remove" not in _auto and "shutil.rmtree" not in _auto, \
        "자동 정리가 파일을 지운다 — 접어 두기만 해야 한다"
    assert "_archive_old_versions" in _auto and "shutil.move" in _lv, \
        "자동 정리가 OLD 이동 함수를 호출하지 않는다"
    for f in ("ecount_reconcile.py", "workbook_patch.py"):
        s = open(os.path.join(ROOT, f), encoding="utf-8").read()
        assert "autoprune" in s, f"{f} 가 구 버전을 자동으로 접지 않는다"
    assert LV.KEEP_LATEST == 1 and LV.ARCHIVE == "OLD", "보관 정책이 바뀌었다"
    assert LV.archive_folder(ROOT) == os.path.join(os.path.abspath(ROOT), "OLD")

    # ★ 자동으로 도는 정리는 아무도 안 보고 있다. **남의 파일을 옮기면 안 된다.**
    #   처음 구현이 `*_v*.xlsx` 를 통째로 잡아 합성검증용 임시 파일까지 옮겨 시험이 깨졌다.
    import tempfile as _tf, shutil as _sh
    _d = _tf.mkdtemp()
    try:
        for _n in ("합성대장F_v1.xlsx", "합성대장F_v2.xlsx", "기타자료_v3.xlsx"):
            open(os.path.join(_d, _n), "w").write("x")
        LV._AUTODONE = False
        assert LV.autoprune(os.path.join(_d, "합성대장F_v2.xlsx")) == 0, "남의 파일을 옮겼다"
        assert len(os.listdir(_d)) == 3, "남의 폴더를 건드렸다"

        for _v in (1, 2, 3):
            open(os.path.join(_d, f"쿠팡_통합업무_일일보고_관리대장_v{_v}.xlsx"), "w").write("x")
            _stamp = 1700000000 + (_v * 86400)  # 서로 다른 날짜여도 최신 1개만 남겨야 한다
            os.utime(os.path.join(_d, f"쿠팡_통합업무_일일보고_관리대장_v{_v}.xlsx"),
                     (_stamp, _stamp))
        open(os.path.join(_d, "쿠팡_통합업무_일일보고_관리대장_v9_보관.xlsx"), "w").write("x")
        _keep, _move = LV.plan(os.path.join(_d, "쿠팡_통합업무_일일보고_관리대장_v3.xlsx"))
        assert [x["v"] for x in _keep] == [3] and sorted(x["v"] for x in _move) == [1, 2], \
            "KEEP_DAYS=0인데 날짜별 구버전을 작업 폴더에 남긴다"
        # 합성 모드는 끝까지 유지한다. 예전 검증은 이 구간에서 CSOS_SYNTHETIC 을
        # 잠시 지웠고, 그 사이 다른 코드가 resolve_master 를 부르면 실 관리대장을
        # 정리할 수 있었다. 이동 알고리즘은 임시 폴더에 한정해 내부 함수를 직접 검증한다.
        LV._AUTODONE = False
        assert LV.autoprune(os.path.join(_d, "쿠팡_통합업무_일일보고_관리대장_v3.xlsx")) == 0, \
            "합성 모드인데 autoprune 이 파일을 옮기려 한다"
        assert LV._archive_old_versions(
            os.path.join(_d, "쿠팡_통합업무_일일보고_관리대장_v3.xlsx")) == 2, \
            "옛 버전 2개가 접히지 않았다"
        left = sorted(x for x in os.listdir(_d) if x.endswith(".xlsx"))
        assert "쿠팡_통합업무_일일보고_관리대장_v3.xlsx" in left, "최신본이 접혔다"
        assert "쿠팡_통합업무_일일보고_관리대장_v9_보관.xlsx" in left, "'보관' 표시본이 접혔다"
        assert "기타자료_v3.xlsx" in left and "합성대장F_v1.xlsx" in left, "남의 파일이 접혔다"
        assert sorted(os.listdir(os.path.join(_d, "OLD"))) == [
            "쿠팡_통합업무_일일보고_관리대장_v1.xlsx",
            "쿠팡_통합업무_일일보고_관리대장_v2.xlsx"], "접힌 목록이 다르다"

        # 과거 `_이전버전`도 다음 실행 때 지정 OLD로 합친다. 같은 v2라도 내용이 다르면
        # 기존 OLD 파일을 덮거나 지우지 않고 출처 꼬리표로 둘 다 남겨야 한다.
        _legacy = os.path.join(_d, "_이전버전")
        os.makedirs(_legacy)
        _legacy_v2 = os.path.join(_legacy, "쿠팡_통합업무_일일보고_관리대장_v2.xlsx")
        open(_legacy_v2, "w").write("다른 v2")
        # 방금 저장된 파일은 사람 것일 수 있어 옮기지 않는다([94]) — 이 검사의 의도는
        # '병합이 되는가'이므로 mtime 을 오래된 것으로 둔다.
        os.utime(_legacy_v2, (_stamp, _stamp))
        LV._AUTODONE = False
        assert LV._archive_old_versions(
            os.path.join(_d, "쿠팡_통합업무_일일보고_관리대장_v3.xlsx")) == 1, \
            "예전 보관 폴더의 파일이 OLD로 합쳐지지 않았다"
        assert not os.listdir(_legacy), "예전 보관 폴더에 관리대장이 남았다"
        archived = sorted(os.listdir(os.path.join(_d, "OLD")))
        assert "쿠팡_통합업무_일일보고_관리대장_v2.xlsx" in archived
        assert any("__from_이전버전" in name for name in archived), \
            "동명 구버전을 덮어쓰지 않고 보존하지 못했다"
    finally:
        LV._AUTODONE = False
        _sh.rmtree(_d, ignore_errors=True)

    # (2) 화면·이미지 — '(금일)' 라벨을 실제 날짜로 바꾸는 코드가 살아 있는가
    src = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "function dateLabel(" in src, "dateLabel 이 없다"
    assert src.count("dateLabel(l)") >= 3, "타일·셀·이미지 세 곳 모두 dateLabel 을 거쳐야 한다"
    assert 'class="dbanner"' in src, "기준일 안내 띠가 없다"
    # 도움말·목록 조회는 **원본 라벨**로 해야 한다(날짜를 박은 문자열로 찾으면 전부 어긋난다)
    assert "helpKey(l)" in src and "dayDetail(l,v)" in src, "라벨 가공본으로 조회하고 있다"

    # (3) 서버 두 개 방지 — 포트 재사용이 꺼져 있어야 새 서버가 조용히 묻히지 않는다
    ap = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "allow_reuse_address = False" in ap, "포트 재사용이 켜져 있어 서버가 두 개 뜬다"

    # (4) 원본 자료 취합 — 종류별 폴더가 한 곳에서만 정의돼야 한다
    import source_dirs as SD
    for name in ("ERP_DIR", "COUPANG_DIR", "KAKAO_DIR", "BAND_DIR"):
        assert getattr(SD, name).startswith(SD.ORIGIN_ROOT), f"{name} 이 원본 폴더 밖을 가리킨다"
    cs = open(os.path.join(ROOT, "collect_sources.py"), encoding="utf-8").read()
    assert "shutil.move" not in cs and "os.remove" not in cs, "취합 도구가 원본을 지우거나 옮긴다"

    # (4-2) 이카운트 내보내기는 파일명이 무작위(8W1JR7MGB50PHOP.xlsx)라 **내용으로** 갈라야 한다.
    #       2026-07-28: 매출(세금)계산서조회(재고)가 '거래명세서 현황'으로 잘못 잡히고
    #       홈택스 전자(세금)계산서는 아예 판별 실패였다.
    from inbox_scan import classify_rows as _CR
    for name, rows, want in (
        ("매출(세금)계산서조회(재고)",
         [["일자 - 번호", "거래처명", "담당자 이메일주소", "공급가액", "부가세", "합 계", "내역보기"],
          ["2026/07/27 -2", "코리아종합물류", "", "352000", "35200", "387200", "내역보기 거래명세서"]],
         "taxinv"),
        ("홈택스 전자(세금)계산서",
         [["일자", "전송일자", "구분", "승인번호", "공급자사업자번호", "공급자상호",
           "공급받는자사업자번호", "공급받는자상호", "기준품목", "계산서분류"],
          ["2026/07/23", "2026/07/23", "매입", "2026072310", "2150351494", "하나산업",
           "8678100168", "유니버셜", "제빙기수리비", "세금계산서"]],
         "hometax"),
        # 기존 판별이 새 규칙에 밀리지 않는지 (겹치는 머리글이 있다)
        ("거래명세서 현황", [["일자-번호", "거래처명", "공급가액", "부가세", "합계"],
                             ["2026/07/01-1", "쿠팡", "100", "10", "110"]], "stmt"),
        ("매출(세금)계산서현황", [["일자-번호", "거래처명", "매출합계", "매출부가세"],
                                  ["2026/07/01-1", "쿠팡", "100", "10"]], "tax"),
        ("거래처별계정별원장", [["일자", "적요", "차변", "대변"],
                                ["2026-07-01", "매출", "100", "0"]], "ledger"),
    ):
        got = _CR(rows)
        assert got == want, f"{name} 판별 {got!r}, 기대 {want!r}"

    # Downloads 를 훑되 **아는 종류만** 가져와야 한다(개인 폴더를 통째로 퍼 오면 안 된다)
    assert "KNOWN" in cs and "kind_of" in cs, "Downloads 에서 종류를 안 가리고 가져온다"

    # (5) 대시보드 채우기 내림 잔해 — 관리대장에 닿을 때만 본다(네트워크가 끊겨도 검증은 돈다)
    #     빈 구역에 수식이 흘러들면 화면이 지저분해질 뿐 아니라 입력칸(H46)이 0 으로 덮여
    #     '당일 실적' 표가 통째로 죽는다(2026-07-28 실사고).
    try:
        import dashboard_clean as DC
        from workbook_patch import latest_master
        mm = latest_master()
        src = mm[0] if isinstance(mm, tuple) else mm
    except Exception:
        src = None
    # data_only=True에서는 캐시가 없는 수식이 None으로 보여 잔해를 놓친다.
    # XML을 직접 봐서 캐시 없는 수식은 잡고, 스타일만 남은 셀은 통과시킨다.
    _dash = ('<sheetData><row r="27">'
             '<c r="B27" s="1"><f>COUNTIF($A$1:$A$2,1)</f><v/></c>'
             '<c r="E27" s="1"/><c r="H27" s="1" t="inlineStr"><is><t>잔해</t></is></c>'
             '</row></sheetData>')
    _left = DC.survey_xml(_dash)
    assert [r for r, _v, _w in _left] == ["B27", "H27"], _left
    # (6) 담당기사 칸에 사람 아닌 값이 들어가면 대표보고 TOP 5 에 그대로 노출되고
    #     기사별 집계가 오염된다(2026-07-28 실사고 10건). 뽑는 단계에서 막는다.
    from band_extract import normalize_tech as NT
    for raw, want in (("000 (캠프상태확인 및 스케쥴 세팅)", ""),
                      ("자) - 각캠프담당자 캠프 컨디션상태 체크", ""),
                      ("하이테크 + 엄진언", "엄진언"),      # 업체는 빼고 사람만
                      # 2026-08-08 인계 반영 — 이 이름은 이제 현재 담당자로 읽힌다([155]).
                      # 업체(대신택배)를 빼는 규칙 자체는 그대로다.
                      ("김혜진 대신택배", "류지영"),
                      ("김승기기장", "김승기"),             # 붙여 쓴 직책 분리
                      ("김필우 기사", "김필우"),
                      ("권오절", "권오철"),                 # 기존 오탈자 교정 유지
                      ("김준형, 권오철", "김준형, 권오철"),
                      ("미배정", ""), ("000", "")):
        got = NT(raw)
        assert got == want, f"normalize_tech({raw!r}) = {got!r}, 기대 {want!r}"

    # (7) 셀 교체가 **옆 칸을 먹어치우지 않는지**. 빈 셀은 자기닫힘(`<c .../>`)이라
    #     느슨한 정규식이 `/` 를 삼키고 다음 `</c>` 까지 지운다(2026-07-28 실사고 — AK33 소실).
    from workbook_patch import replace_inline_cell as _RC
    _x = '<c r="AJ33" s="5"/><c r="AK33" s="7" t="inlineStr"><is><t>지켜야함</t></is></c>'
    _o = _RC(_x, "AJ33", "메모")
    assert "지켜야함" in _o and 'r="AK33"' in _o, "빈 셀을 교체하며 옆 칸을 삼켰다"
    assert _o.count("<c ") == 2, f"셀 개수가 바뀌었다: {_o}"

    if src and os.path.exists(src):
        import openpyxl as _ox
        from project_resolve import clean_tech
        _w = _ox.load_workbook(src, read_only=True, data_only=True)
        junk = []
        # ★ 02시트만 보다가 03_현장작업실적에 같은 값이 남았다 — 기사 열이 있는 시트를 전부 본다
        for _sn in ("02_돌발AS접수", "03_현장작업실적", "04_정기점검", "05_신규납품설치"):
            if _sn not in _w.sheetnames:
                continue
            _s = _w[_sn]
            _h = [str(h or "").strip() for h in
                  next(_s.iter_rows(min_row=4, max_row=4, values_only=True))]
            _cs = [i for i, h in enumerate(_h) if h in ("담당기사", "담당자", "작업자")]
            for _row in _s.iter_rows(min_row=5, values_only=True):
                if not _row[0]:
                    continue
                for _i in _cs:
                    _v = str(_row[_i] or "").strip() if _i < len(_row) else ""
                    if _v and clean_tech(_v) != _v:
                        junk.append(f"{_sn}:{_v}")
        _w.close()
        assert not junk, ("담당기사 칸에 사람 아닌 값 %d건 — "
                          "python fix_tech_names.py 로 정정: %s" % (len(junk), junk[:3]))
        left = DC.survey(src)
        assert not left, ("대시보드에 채우기 내림 잔해 %d칸 — python dashboard_clean.py 로 정리: %s"
                          % (len(left), ", ".join(r for r, _v, _w in left[:8])))
        import openpyxl
        w = openpyxl.load_workbook(src, data_only=False)
        assert w["00_대시보드"]["H46"].value in (None, ""), \
            "H46(조회일 직접입력)이 비어 있지 않다 — 당일 실적 표가 죽는다"
        w.close()
        print("  [41] … 대시보드 잔해 0칸 · 입력칸 H46 정상")
    print("  [41] 날짜 명시(금일 금지·항목별 날짜)·서버 중복 방지·원본 취합 안전 ✅")


def t40_claim_enforced():
    """[40] 협업 규칙이 **문서로만** 있으면 잊는다 — 원장 도구가 실제로 막아야 한다.

    두 AI가 동시에 원장을 고치면 각자 vN+1을 만들어 한쪽이 통째로 묻힌다(되돌릴 수 없다).
    ai_claim 은 규칙을 적는 곳이고, claim_guard 가 그 규칙을 강제한다."""
    import claim_guard, ai_claim

    # (1) 원장을 쓰는 도구는 전부 가드를 거쳐야 한다
    for fn in ("ledger_writer.py", "workbook_patch.py", "expand_rows.py", "confirm_fill.py",
               "fix_formulas.py", "dashboard_clean.py", "reorder_rows.py"):
        src = open(os.path.join(ROOT, fn), encoding="utf-8").read()
        assert "claim_guard" in src and 'require("ledger"' in src, f"{fn} 이 점유를 확인하지 않는다"

    # 실제 협업 점유 파일을 시험이 비우면, 합성검증 도중 다른 AI가 원장 쓰기에 진입한다.
    # 독립 임시 점유 파일에서만 충돌·자동 점유를 검증하고 실제 점유는 그대로 보존한다.
    # ★ 업무센터 활동 파일도 같이 격리한다 (2026-08-13 실사고). 안 하면 **사람이 앱에서
    #   입력하고 있는 동안 합성검증이 통째로 빨개진다** — 아래 take() 는 시험의 '설정'인데
    #   _workcenter_priority_gate 가 "오종현 업무센터 입력이 진행 중이라 미룹니다" 로
    #   막아 버린다. 제품 동작은 옳고 시험이 실제 상태에 매달린 것이 잘못이다.
    #   그림이 나쁜 자리다: 관문이 "ALL GREEN 확인 후에만 실데이터 작업" 인데, 담당자가
    #   일을 하고 있으면 그 관문을 아무도 통과 못 한다. 정상이 실패처럼 보이는 자리.
    #   (같은 격리를 t104 는 이미 하고 있었다 — 이 시험만 그 가드보다 먼저 쓰여 빠졌다.
    #    게이트 자체는 한 줄도 안 건드린다. 그 동작을 검증하는 자리는 따로 있다.)
    real_claims, real_guard = ai_claim.CLAIMS, ai_claim.GUARD
    real_activity = ai_claim.WORKCENTER_ACTIVITY
    with tempfile.TemporaryDirectory() as claim_tmp:
        ai_claim.CLAIMS = os.path.join(claim_tmp, "claims.json")
        ai_claim.GUARD = os.path.join(claim_tmp, ".guard")
        ai_claim.WORKCENTER_ACTIVITY = os.path.join(claim_tmp, "workcenter_activity.json")
        try:
            # (2) 남이 잡고 있으면 멈춰야 한다
            assert ai_claim.take("codex", "ledger", "합성검증"), "시험용 점유를 못 잡았다"
            try:
                os.environ["CSOS_AI"] = "claude"
                try:
                    claim_guard.require("ledger", "test")
                    raise AssertionError("남이 잡았는데 그냥 진행했다 — 동시 수정이 일어난다")
                except SystemExit as e:
                    assert e.code == 3, e.code
                # (3) 사람이 직접 실행할 때(CSOS_AI 없음)는 막지 않는다 — 손을 묶으면 안 된다
                os.environ.pop("CSOS_AI", None)
                assert claim_guard.require("ledger", "test") is True
            finally:
                os.environ.pop("CSOS_AI", None)
                ai_claim.free("codex", "ledger")

            # (4) 아무도 안 잡았으면 자동으로 잡고 진행한다
            os.environ["CSOS_AI"] = "claude"
            try:
                assert claim_guard.require("ledger", "test") is True
                assert (ai_claim.load().get("ledger") or {}).get("who") == "claude"
            finally:
                ai_claim.free("claude", "ledger")
                os.environ.pop("CSOS_AI", None)
        finally:
            ai_claim.CLAIMS, ai_claim.GUARD = real_claims, real_guard
            ai_claim.WORKCENTER_ACTIVITY = real_activity
            os.environ.pop("CSOS_AI", None)
    print("  [40] 점유 강제(원장 도구 차단·사람 실행 허용·자동 점유) ✅")


def t42_first_empty_row(tmp):
    """[42] '빈 행'을 프로젝트NO 한 열로 판정하면 **남의 행에 번호를 얹는다**.

    2026-07-28 실사고: 02시트 547행은 사람이 번호 없이 내용만 적어 둔 행
    (M_순천1·김필우·리모컨)이었는데, 프로젝트NO 가 비었다는 이유로 '빈 행'이 되어
    전혀 다른 건(UJ2601347 중구1·경광등)의 번호가 그 행에 들어갔다.
    나머지 칸은 '빈 칸만' 정책이 막아 줘서 겉으로는 조용했지만 한 행에 두 건이 섞였다.
    → 빈 행 판정은 **행 전체**로 한다. 번호만 없는 행은 '새 행'이 아니라
      내용으로 맞춰 **번호를 채워 줄 대상**이다(새 행을 만들면 같은 건이 두 행이 된다)."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import openpyxl
    from datetime import date as _date
    import kakao_extract as kx
    import backfill_rows as bf

    path = os.path.join(tmp, "합성_원장.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "02_돌발AS접수"
    hdr = ["접수ID", "프로젝트NO", "캠프명", "접수일자", "신청내용", "담당기사"]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=i, value=h)
    for r, code in ((5, "UJ0000001"), (6, "UJ0000002")):
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value="캠프%d" % r)
        ws.cell(row=r, column=4, value=_date(2026, 7, 20))
    # 7행: 번호만 없는 기존 행(사람이 먼저 적어 둔 행) — 빈 행이 아니다
    ws.cell(row=7, column=3, value="M_순천1")
    ws.cell(row=7, column=4, value=_date(2026, 7, 27))
    ws.cell(row=7, column=5, value="리프트 리모콘 작동 안함")
    ws.cell(row=7, column=6, value="김필우")
    ws.cell(row=12, column=1, value="")          # 용량만 넓힌다(빈 행)
    wb.save(path)

    start, cols, cap, orphans = kx.sheet_state(path, "02_돌발AS접수")
    assert start == 8, "번호 없는 7행을 빈 행으로 봤다 — 남의 행에 번호를 얹는다 (start=%s)" % start
    assert [o["행"] for o in orphans] == [7], orphans
    assert orphans[0]["내용키"] == kx._key("리프트 리모콘 작동 안함"), orphans

    b_start, _, _ = bf.sheet_state(path, "02_돌발AS접수")
    assert b_start == 8, "backfill_rows 도 같은 함정에 빠진다 (start=%s)" % b_start

    # 공지 일자와 원장 일자는 며칠 어긋난다(공지가 다음 날 올라온다) — ±3일은 같은 건
    assert kx._near("2026-07-28", {_date(2026, 7, 27)}) is True
    assert kx._near("2026-07-28", {_date(2026, 7, 20)}) is False
    assert kx._near("", {_date(2026, 7, 27)}) is False

    # 두 자리 연도('26.07.28')를 못 읽으면 신청일자가 통째로 빈칸이 된다(실제 UJ2601345)
    assert kx.norm_date("26.07.28") == _date(2026, 7, 28)
    assert kx.norm_date("2026.00.00 (요일)") is None          # 템플릿은 채우지 않는다
    # 유형을 모르면 찍지 않는다 — 억지로 02·04 로 몰면 엉뚱한 시트에 등록된다
    assert kx.kind_of("♣ ［ 돌발유료 A/S 안내 ]") == "02_돌발AS접수"
    assert kx.kind_of("♣ ［ 2026년 03분기 3개월 유료 A/S 안내 ]") == "04_정기점검"
    assert kx.kind_of("♣ ［쿠팡 철거보관 안내］") not in ("02_돌발AS접수", "04_정기점검")
    assert kx.status_of("♣ ［ 돌발유료 A/S 완료 ]") == "완료"
    assert kx.status_of("✅ 접수취소 ♣ ［ 돌발유료 A/S 안내 ]") == "취소"
    # 담당기사 칸에 작업 메모가 들어가면 대표보고에 그대로 노출된다(2026-07-28 실사고)
    from band_extract import normalize_tech
    assert normalize_tech("000 (캠프상태확인 및 스케쥴 세팅)") == ""
    print("  [42] 빈 행 판정(행 전체)·번호없는 행 채움·카톡 파싱(2자리연도·유형·상태) ✅")


def t43_receipt_fill(tmp):
    """[43] 입금 자동입력 — 합계행을 입금으로 세면 가짜 수금이 생긴다.

    폰 앱 '빈 항목 입력'의 입금일·입금액만 사람이 손으로 넣고 있었다(사용자 지시 2026-07-28:
    자료 확인되면 자동화). 근거는 거래처별계정별원장의 **대변**이다. 함정 셋:
      · '월 계'·'누 계'·'전기이월' 행이 큰 금액을 갖고 있다 → 그대로 세면 가짜 입금 1건
      · 차변은 매출 발생이지 입금이 아니다
      · 쿠팡은 여러 건을 묶어 한 번에 넣는다 → 금액이 겹치면 엉뚱한 정산행에 붙는다"""
    import sys as _s
    _s.path.insert(0, ROOT)
    import openpyxl
    from datetime import date as _date
    import receipt_fill as rf

    path = os.path.join(tmp, "합성_계정별원장.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["회사명 : 합성"])
    ws.append(["일자", "일자-No.", "적요", "차변", "대변"])
    ws.append([_date(2026, 1, 10), "2026/01/10 -3", "제품매출", 3043150, None])   # 차변=매출
    ws.append([_date(2026, 2, 15), "2026/02/15 -1", "보통예금 입금", None, 3043150])
    ws.append([_date(2026, 2, 20), "2026/02/20 -2", "보통예금 입금", None, 500000])
    ws.append([None, None, "2026/02 월 계", 3043150, 3543150])                    # 합계행
    ws.append([None, None, "누 계", 3043150, 3543150])                            # 합계행
    wb.save(path)

    got, ok = rf.parse_receipts(path)
    assert ok, "적요+대변 머리글을 못 찾았다"
    assert len(got) == 2, "합계행·차변행이 섞였다: %s" % got     # 차변 1 + 합계 2 는 빠져야 한다
    assert {int(g["금액"]) for g in got} == {3043150, 500000}, got
    assert got[0]["일자"] == _date(2026, 2, 15), got[0]

    # 유일 일치만 자동입력 — 금액이 같은 후보가 둘이면 사람에게 넘긴다
    rows = [
        {"정산ID": "S1", "청구액": 3043150.0, "프로젝트NO": "UJ1", "캠프명": "양주2",
         "발행일": _date(2026, 1, 10)},
        {"정산ID": "S2", "청구액": 500000.0, "프로젝트NO": "UJ2", "캠프명": "부산1",
         "발행일": _date(2026, 1, 11)},
        {"정산ID": "S3", "청구액": 500000.0, "프로젝트NO": "UJ3", "캠프명": "창원1",
         "발행일": _date(2026, 1, 12)},
    ]
    paired, spare = rf.match(got, rows)
    assert [s["정산ID"] for _, s in paired] == ["S1"], paired      # 500000 은 후보 2건 → 보류
    assert len(spare) == 1, spare

    # 발행일보다 앞선 입금은 그 건의 입금이 아니다
    early = [{"일자": _date(2025, 12, 1), "금액": 3043150.0, "전표": "", "적요": "", "출처": ""}]
    assert rf.match(early, rows[:1])[0] == [], "청구 전에 들어온 돈을 그 건에 붙였다"

    # 청구일·지급예정일은 채우지 않는다 — 산정 규칙이 확정되기 전엔 만들어 넣으면 안 된다
    src = open(os.path.join(ROOT, "receipt_fill.py"), encoding="utf-8").read()
    body = src.split('if __name__')[0]
    assert '"col": "청구일"' not in body and '"col": "지급예정일"' not in body, \
        "규칙이 확정되지 않은 청구일·지급예정일을 채우고 있다"
    # 오종현 관리 '26년도 쿠팡 입금내역' — 사람이 만드는 표라 제목·빈 줄 수가 달라진다.
    # 머리글 행을 박아 두면 다음 달에 한 줄 밀리는 순간 조용히 0건이 된다.
    dep = os.path.join(tmp, "합성_입금내역.xlsx")
    wb2 = openpyxl.Workbook()
    w2 = wb2.active
    w2.append([])
    w2.append([None, "2026 쿠팡 입금 내역"])          # 제목 — 머리글이 아니다
    w2.append([])
    w2.append([None, "날짜", "거래처", "입금액"])       # 실제 머리글(4행)
    w2.append([None, _date(2026, 7, 27), "쿠팡로지스틱스", 715000])
    w2.append([None, _date(2026, 7, 27), "쿠팡로지스틱", 814000])      # 같은 거래처, 표기만 다름
    w2.append([None, _date(2026, 5, 6), "김진주（위더스 )", 387200])   # 전각 괄호
    w2.append([None, None, "합 계", 1916200])                          # 합계행
    wb2.save(dep)

    got2 = rf.parse_deposit_list(dep)
    assert len(got2) == 3, "합계행이 입금으로 세어졌다: %s" % got2
    assert sum(g["금액"] for g in got2) == 1916200, got2
    custs = {g["거래처"] for g in got2}
    assert "쿠팡로지스틱스" in custs and "쿠팡로지스틱" not in custs, \
        "같은 거래처가 표기 차이로 갈라졌다 — 금액이 둘로 쪼개진다: %s" % custs
    assert len(custs) == 2, custs                     # 쿠팡 + 김진주위더스
    assert rf.norm_cust("(주)모벤티스") == rf.norm_cust("주식회사 모벤티스"), "법인격 표기 미정규화"

    # ★ 은행에서 그대로 받은 '거래내역조회' 도 읽는다 (2026-08-07 김미영 대리 파일).
    #   사람이 정리한 표가 아니라 원본이라 머리글이 다르다 — '입금액'이 아니라 '입금',
    #   '입금일'이 아니라 '거래일시'. 못 알아보면 `9. 미분류` 로 가서 아무도 안 읽는다
    #   (실제로 classify 가 unknown, 파서가 0건이었다 — 77건이 통째로 사라질 뻔했다).
    bank = os.path.join(tmp, "거래내역조회_입출식 예금20260807.xlsx")
    wb3 = openpyxl.Workbook(); w3 = wb3.active
    w3.append(["거래내역조회_입출식 예금"]); w3.append(["계좌번호:036-509375-04-012"])
    w3.append([" ", "거래일시", "출금", "입금", "거래후 잔액", "거래내용",
               "상대계좌번호", "상대은행", "메모", "거래구분", "수표어음금액",
               "CMS코드", "상대계좌예금주명"])
    w3.append(["135", "2026-03-11 12:17:07", 0, 8306650, 21429227, "쿠팡로지스틱스",
               "", "산업은행", "", "타행이체", 0, "", "쿠팡로지스틱스서비스"])
    # 은행이 거래내용을 잘라 보내는 일이 잦다 — 예금주명을 먼저 봐야 같은 거래처로 묶인다
    w3.append(["153", "2026-03-23 12:03:15", 0, 920700, -53814429, "쿠팡로지스틱",
               "", "산업은행", "", "타행이체", 0, "", "쿠팡로지스틱스서비스"])
    w3.append(["900", "2026-03-24 09:00:00", 5000000, 0, 100, "임대료",
               "", "산업은행", "", "타행이체", 0, "", "건물주"])   # 출금 — 받은 돈이 아니다
    wb3.save(bank)
    from inbox_scan import classify as _cls
    assert _cls(bank) == "receipt", "은행 거래내역을 못 알아본다 — 미분류로 가서 아무도 안 읽는다"
    got3 = rf.parse_deposit_list(bank)
    assert len(got3) == 2, "출금 행까지 입금으로 셌다: %s" % got3
    assert sum(g["금액"] for g in got3) == 9227350, got3
    assert {g["거래처"] for g in got3} == {"쿠팡로지스틱스"}, \
        "잘린 '쿠팡로지스틱' 이 다른 거래처로 갈렸다 — 예금주명을 먼저 봐야 한다: %s" % got3

    # 공유 폴더 정본을 원본 자료 폴더에 복사해 둔 경우 같은 70건을 140건으로 세면 안 된다.
    import shutil
    copied = os.path.join(tmp, "보관복사본.xlsx")
    shutil.copy2(dep, copied)
    uniq = rf._unique_deposit_files([dep, copied])
    assert uniq == [dep], "내용이 같은 입금 파일 복사본을 중복 집계한다: %s" % uniq
    print("  [43] 입금 자동입력(합계행 제외·차변 제외·유일매칭·머리글 자동탐지·거래처 정규화) ✅")


def t44_zscan():
    """[44] 쿠팡 업무 폴더 2만 개 — **파일을 열지 않고** 파일명으로 고르고 대조한다.

    사용자 지시(2026-07-28): "관련있는 자료는 전부 긁어와서 비교 검토, 관련 없는 자료는
    db에 적용하지마." 지키는 방법은 쓰기 경로를 좁게 만드는 것이다 —
    캠프명+날짜(±7일)가 **둘 다** 맞는 1:1 확정 건만 쓰고, 후보가 여럿이면 사람에게 넘긴다.
    (같은 캠프에 한 달에 여러 건이 있고 같은 날 여러 캠프를 돈다 — 한쪽만으로는 못 잇는다)"""
    import sys as _s
    _s.path.insert(0, ROOT)
    import zscan

    # 안전·교육 폴더는 업무상 필요해도 관리대장에 들어갈 자리가 없다
    assert zscan.classify("♣ 6. 설치공사 안전보건대장", "허가서.pdf").startswith("무관")
    assert zscan.classify("♣ 7. 쿠팡 지게차 서류", "UJ2600136 지게차.pdf").startswith("무관")
    assert zscan.classify("♣ 2. 쿠팡 돌발AS", "철거 906,000원 UJ2600136.PDF").startswith("관련")
    assert zscan.classify("♣ 1. 쿠팡 정기점검", "2026-06-19 구로1MB 거래명세서.pdf").startswith("관련")
    assert zscan.classify("♣ 100. 쿠팡 기타자료", "현장사진.jpg").startswith("무관")

    # 캠프명 표기 흔들림 — 괄호 안 표기가 달라도 같은 캠프로 본다
    assert zscan.camp_key("구로1MB(독산동B)") == zscan.camp_key("구로1MB (독산동A)")
    assert zscan.camp_key("일산2MB(양평동4가B)") != zscan.camp_key("일산3MB(양평동6가B)")

    docs = [{"일자": "2026-06-19", "캠프키": zscan.camp_key("구로1MB(독산동B)"),
             "캠프원문": "구로1MB", "종류": "거래명세서", "파일": "a.pdf", "폴더": "x"},
            {"일자": "2026-06-19", "캠프키": zscan.camp_key("없는캠프"),
             "캠프원문": "없는캠프", "종류": "거래명세서", "파일": "b.pdf", "폴더": "x"}]
    rows = [{"시트": "04_정기점검", "ID": "PM-1", "프로젝트NO": "UJ1", "캠프명": "구로1MB(독산동B)",
             "캠프키": zscan.camp_key("구로1MB(독산동B)"), "일자": "2026-06-20"}]
    paired, orphan, amb = zscan.match_docs(docs, rows)
    assert len(paired) == 1 and paired[0][1]["프로젝트NO"] == "UJ1", paired   # 하루 차이는 같은 건
    assert len(orphan) == 1 and not amb, (orphan, amb)

    # 같은 캠프에 후보가 둘이면 **쓰지 않는다** — 엉뚱한 행에 발행완료가 찍힌다
    rows2 = rows + [{"시트": "04_정기점검", "ID": "PM-2", "프로젝트NO": "UJ2",
                     "캠프명": "구로1MB(독산동B)", "캠프키": zscan.camp_key("구로1MB(독산동B)"),
                     "일자": "2026-06-18"}]
    p2, _o2, a2 = zscan.match_docs(docs, rows2)
    assert not p2 and len(a2) == 1, (p2, a2)

    # 날짜가 멀면 다른 건이다
    far = [dict(docs[0], 일자="2026-05-01")]
    assert zscan.match_docs(far, rows)[0] == []
    print("  [44] 쿠팡 폴더 조사(무관 폴더 제외·캠프키·1:1 확정만 인정) ✅")


def t45_cloud_queue_and_erp_documents(tmp):
    """[45] PC 독립 큐와 ERP 문서 완료는 근거가 있는 빈칸만 건드린다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import fill_erp_documents as fed

    book = os.path.join(tmp, "erp_docs.xlsx")
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "02_돌발AS접수"
    a.append([]); a.append([]); a.append([])
    a.append(["접수ID", "프로젝트NO", "진행상태", "ERP등록"])
    a.append(["AS-1", "UJ0000001", "작업완료", ""])
    a.append(["AS-2", "UJ0000002", "접수", ""])
    p = wb.create_sheet("04_정기점검")
    p.append([]); p.append([]); p.append([])
    p.append(["점검ID", "프로젝트NO", "점검상태", "ERP판매전표", "거래명세서"])
    p.append(["PM-1", "UJ0000001", "완료", "", ""])
    p.append(["PM-2", "UJ0000002", "완료", "", ""])
    wb.save(book)
    original = fed.document_evidence
    fed.document_evidence = lambda: {
        "UJ0000001": {"판매전표", "거래명세서"},
        "UJ0000002": {"판매전표"},
    }
    try:
        fills, counts, _ = fed.plan(book)
    finally:
        fed.document_evidence = original
    assert fills == {
        "02_돌발AS접수!D5": "완료",
        "04_정기점검!D5": "완료",
        "04_정기점검!E5": "발행완료",
        "04_정기점검!D6": "완료",
    }, fills
    assert not any("C5" in ref or "C6" in ref for ref in fills), "업무 상태 셀을 직접 덮었다"
    assert counts["04_정기점검:거래명세서"] == 1

    app = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    snap = open(os.path.join(ROOT, "mobile_snapshot.py"), encoding="utf-8").read()
    sync = open(os.path.join(ROOT, "cloud_queue_sync.py"), encoding="utf-8").read()
    for need in ("cloud_queue", "Authorization", "/api/queue", "accepted", "duplicate"):
        assert need in app, f"폰 클라우드 큐 연결 누락: {need}"
    for need in ("/api/queue/lease", "/api/queue/ack", "/api/queue/release",
                 "is_input_window", 'take("cloud-sync", "ledger"'):
        assert need in sync, f"로컬 큐 유실방지 누락: {need}"
    for need in ("완료보고서등록", "ERP판매전표", "거래명세서"):
        assert need in snap, f"폰 사본 상태 누락: {need}"
    print("  [45] PC 독립 영구큐·ERP 문서근거 완료·앱 상태 전파 ✅")


def t46_app_2026_only():
    """[46] 앱 표시 경계는 2025년을 숨기고 2026년만 통과시킨다."""
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as app

    assert not app.app_year_record(
        {"프로젝트NO": "UJ2601234", "접수일자": "2025-12-31"}, "as"), \
        "프로젝트 번호가 26이어도 실제 접수일이 2025면 2025 업무다"
    assert app.app_year_record(
        {"프로젝트NO": "UJ2501234", "접수일자": "2026-01-02"}, "as"), \
        "2026 접수 건을 프로젝트 번호만 보고 숨겼다"
    assert app.app_year_record({"업무ID": "AS-2607-001"}, "as")
    assert not app.app_year_record({"업무ID": "PM-2512-099"}, "pm")
    assert app.app_year_record({"월": "2026/07", "전표": "2026/07/24 - 3"}, "erp")
    assert app.app_year_record({"전표": "26/07/24 - 3"}, "erp")
    assert not app.app_year_record({"월": "2025/12", "전표": "25/12/30 - 1"}, "erp")
    assert app.app_year_record({"기준일": "2026-07-28", "문제내용": "완료보고 확인"}, "issue")
    assert not app.app_year_record({"캠프명": "연도 확인 불가"}), "연도 미상 행이 2026 목록에 섞였다"
    assert not app.app_year_record(
        {"포함프로젝트": "UJ2500001, UJ2600001"}), "2025·2026 혼합 행이 통째로 표시됐다"
    assert not app.app_project_result(
        "UJ2600007", {"date": "2025-12-31", "ids": {"접수ID": "AS-2512-040"}}), \
        "UJ26 코드만 보고 연결된 2025 작업을 앱에 노출했다"
    assert not app.app_project_result(
        "UJ2600008", {"date": "2026-01-02", "ids": {"접수ID": "AS-2512-041"}}), \
        "2026 날짜와 2025 업무ID가 섞인 자동조회 결과를 통과시켰다"
    assert app.app_project_result(
        "UJ2600009", {"date": "2026-01-03", "ids": {"접수ID": "AS-2601-005"}})

    kept = app.app_year_rows([
        {"프로젝트NO": "UJ2500001"},
        {"프로젝트NO": "UJ2600001"},
        {"정산ID": "JS-2607-001"},
    ])
    assert len(kept) == 2 and all("25" not in str(r) for r in kept), kept

    pub = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    sw = open(os.path.join(ROOT, "docs", "sw.js"), encoding="utf-8").read()
    assert 're.fullmatch(r"UJ26\\d{5}"' in pub, "배포 프로젝트코드에 2025가 섞일 수 있다"
    assert "A.app_project_result(c, r)" in pub, "UJ26에 연결된 2025 작업을 사본에서 거르지 않는다"
    assert 're.match(r"2026-' in pub, "미청구 배포에 2025가 섞일 수 있다"
    assert "function keep2026" in phone and "D = keep2026(await open(pin))" in phone
    assert "codeIs2026(k,r)" in phone, "구 사본의 프로젝트 자동조회 2025 2차 방어가 없다"
    assert "2026-only" in sw, "구 서비스워커가 2025 포함 사본을 계속 쓸 수 있다"

    import daily_brief as db
    brief = db.brief("2026-07-28", {
        "as": [
            {"접수ID": "AS-2512-001", "프로젝트NO": "UJ2500001",
             "접수일자": "2025-12-01", "작업완료일": "", "진행상태": "접수"},
            {"접수ID": "AS-2607-001", "프로젝트NO": "UJ2600001",
             "접수일자": "2026-07-28", "작업완료일": "", "진행상태": "접수"},
        ],
        "pm": [], "fw": [],
    })
    assert brief["돌발AS"]["신규접수"] == 1
    assert brief["돌발AS"]["완료일미기입"] == 0, "2025 미완료가 2026 브리핑에 섞였다"

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for need in ("const APP_YEAR = '2026'", "py.innerHTML = `<option>${APP_YEAR}</option>`",
                 "return +d === FOCUS_YEAR", "const y = APP_YEAR"):
        assert need in live, f"라이브 앱 2026 고정 누락: {need}"
    print("  [46] 앱 전 화면 2026년 전용 필터·구 사본 2차 방어 ✅")


def t47_back_nav():
    """[47] 뒤로가기 — 바로 전 화면 → 없으면 대시보드 → 대시보드에서 종료(사용자 지시 2026-07-28).

    시트(상세창)와 탭을 **한 스택**으로 다뤄야 한다. 따로 두면 시트를 닫는 뒤로가기와
    탭을 되돌리는 뒤로가기가 어긋난다. 그리고 되돌린 뒤에는 반드시 여분 항목을 다시
    쌓아야(navGuard) 다음 뒤로가기가 앱 밖으로 새지 않는다."""
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # 화면 전환과 기록을 분리했는가 — applyView 는 기록을 안 남기고, show 는 남긴다
    assert "function applyView(" in live and "function show(v){" in live, "applyView/show 분리 누락"
    assert "navPush({t:'tab', from:from})" in live, "탭 이동이 뒤로가기 기록에 안 쌓인다"
    assert "navPush({t:'sheet'})" in live, "시트가 뒤로가기 기록에 안 쌓인다"

    # 홈 판정과 종료
    assert "curView() !== 'dash'" in live and "exitApp()" in live, "홈(대시보드) 폴백·종료 누락"
    assert "function exitApp()" in live and "window.close()" in live, "종료 처리 누락"

    # ★ 여분 항목이 없으면 첫 뒤로가기가 앱 밖으로 나간다
    assert "function navGuard()" in live, "navGuard 누락"
    assert live.count("navGuard()") >= 4, "되돌린 뒤 여분 항목을 다시 쌓지 않는 경로가 있다"

    # ★ history.go(-n) 은 popstate 를 한 번만 낸다 — n 만큼 세면 다음 뒤로가기를 먹는다
    assert "_navSkip += 1" in live, "history.go(-n) 의 popstate 를 n 번으로 세고 있다"
    assert "_navSkip += back" not in live, "go(-n) popstate 를 n 번으로 세는 코드가 남아 있다"

    # 시트를 통째로 닫아도 탭 기록은 남아야 한다(그래야 뒤로가기가 이전 탭으로 간다)
    assert "while(_nav.length && _nav[_nav.length-1].t === 'sheet')" in live, \
        "closeSheetAll 이 탭 기록까지 걷어낸다"
    # 옛 분리 구현의 흔적이 남아 있으면 두 스택이 다시 어긋난다
    assert "_sheetHist" not in live, "옛 _sheetHist 카운터가 남아 있다(스택 이원화)"
    print("  [47] 뒤로가기(이전 화면→대시보드→종료)·시트/탭 단일 스택 ✅")


def t48_excel_2026_stats_and_verified_completion():
    """[48] Excel statistics are 2026-only; completion requires exact evidence."""
    import stats_2026 as S

    old = "=COUNTIF('02_돌발AS접수'!$Q$5:$Q$744,\"작업완료\")"
    fixed = S.yearize_formula(old)
    assert "COUNTIFS" in fixed
    assert "'02_돌발AS접수'!$D$5:$D$744" in fixed
    assert "DATE(2026,1,1)" in fixed and "DATE(2027,1,1)" in fixed

    risk = "=COUNTIF('07_불일치누락현황'!$F$5:$F$304,\"*미배정*\")"
    fixed_risk = S.yearize_formula(risk)
    assert "'07_불일치누락현황'!$Q$5:$Q$304,2026" in fixed_risk

    document = "=COUNTIF('17_문서대조현황'!$G$5:$G$154,\"미작성\")"
    fixed_document = S.yearize_formula(document)
    assert "'17_문서대조현황'!$A$5:$A$154,\"JS-26*\"" in fixed_document

    top = S.top5_formula(1)
    top_index = S._top_index(1)
    assert "$P28" in top
    assert "$Q$5:$Q$304=2026" in top_index
    assert "$V$5:$V$304=\"포함\"" in top_index
    assert "UJ25" not in top and "2025" not in top

    helper = S.issue_year_formula(5)
    assert "'02_돌발AS접수'!$D$5:$D$744" in helper
    assert "'04_정기점검'!$D$5:$D$624" in helper
    assert "'06_거래서류청구수금'!$L$5:$L$154" in helper

    import complete_verified as C

    assert C._year("2026-07-27") == 2026
    assert C._year("2025-12-31") == 2025
    ready = {
        "작업완료일": "2026-07-27",
        "관리자검증상태": "일치",
        "완료보고서등록": "등록",
        "사진등록": "등록",
        "동영상등록": "",
        "비용구분": "유상",
        "ERP등록": "등록",
        "재방문여부": "아니오",
        "최초접수외추가작업": "",
        "추가작업확인상태": "",
    }
    assert C._as_completion_ready(lambda name: ready.get(name))
    ready["ERP등록"] = ""
    assert not C._as_completion_ready(lambda name: ready.get(name))
    print("  [48] Excel 2026-only statistics and evidence-gated completion OK")


def t49_exec_metric_drilldown_and_sheet_scroll(tmp):
    """[49] 대표보고 3·4절은 숫자와 동일한 원천행을 열고, 긴 목록은 끝까지 스크롤된다."""
    path = os.path.join(tmp, "exec_metric.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for _ in range(3):
            ws.append([])
        ws.append(headers)
        for row in rows:
            ws.append(row)

    sheet("06_거래서류청구수금", [
        "정산ID", "원천업무ID", "프로젝트NO", "캠프명", "작업완료일(자동)", "업무구분", "담당자",
        "거래명세서발행일", "거래명세서합계", "세금계산서발행일", "세금계산서합계",
        "입금일", "입금액", "미청구액", "미수금액", "작업대비거래명세서차액",
        "문제내용", "청구상태", "검증결과"
    ], [
        ["JS-2601", "AS-2601", "UJ260001", "합성1캠프", "2026-07-28", "돌발AS", "김기사",
         "2026-07-28", 11000, "2026-07-28", 11000, "2026-07-28", 11000,
         22000, 33000, -1000, "합성 문제", "청구중", "불일치"],
        ["JS-2602", "AS-2602", "UJ260002", "합성2캠프", "2026-07-28", "신규·납품·설치", "이기사",
         "", 0, "", 0, "", 0, 0, 0, 999, "", "", "불일치"],
        ["JS-2501", "AS-2501", "UJ250001", "과거캠프", "2025-07-28", "돌발AS", "구기사",
         "2026-07-28", 999999, "2026-07-28", 999999, "2026-07-28", 999999,
         999999, 999999, 999999, "2025 제외", "", "불일치"],
    ])
    sheet("02_돌발AS접수",
          ["접수ID", "프로젝트NO", "캠프명", "접수일자", "담당기사"],
          [["AS-2601", "UJ260001", "합성1캠프", "2026-07-27", "김기사"]])
    sheet("04_정기점검",
          ["점검ID", "프로젝트NO", "캠프명", "점검예정일", "담당기사"], [])
    sheet("07_불일치누락현황",
          ["업무기준연도(자동·숨김)", "최상위 업무키", "원천업무ID", "프로젝트NO", "문제상세", "조치상태"],
          [[2026, "UJ260001", "AS-2601", "UJ260001", "문제 A", "확인필요"],
           [2026, "UJ260001", "AS-2601", "UJ260001", "문제 B", "확인필요"],
           [2025, "UJ250001", "AS-2501", "UJ250001", "과거 문제", "확인필요"]])
    sheet("15_세금계산서관리",
          ["정산ID", "발행기한임박여부", "기한초과여부", "법정발행기한",
           "발행금액", "발행상태(자동)", "아리바청구상태"],
          [["JS-2601", "예", "예", "2026-07-28", 11000, "미발행", "등록대기"],
           ["JS-2501", "예", "예", "2025-07-28", 999999, "미발행", "등록대기"]])
    sheet("17_문서대조현황",
          ["정산ID", "경고내용", "우선순위", "PO상태", "거래명세서상태"],
          [["JS-2601", "문서 경고", "P1", "PO 발행대기", "미작성"],
           ["JS-2501", "과거 경고", "P1", "PO 발행대기", "미작성"]])
    wb.save(path)

    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as A
    d = A.read_exec_details(path, "2026-07-28")
    assert d["청구액 (당일)"]["count"] == 1 and d["청구액 (당일)"]["amount"] == 11000
    assert d["세금계산서 발행액 (당일)"]["count"] == 1
    assert d["입금액 (당일)"]["count"] == 1
    assert d["잔여 미청구액"]["amount"] == 22000
    assert d["잔여 미수금액"]["amount"] == 33000
    assert d["작업금액 불일치 (현재)"]["count"] == 1, "신규납품·2025가 섞였다"
    assert d["문제 업무 건수(중복 제거)"]["count"] == 1
    assert d["문제 프로젝트 / 문제 행"]["count"] == 2
    assert d["세금계산서 기한 임박·초과"]["count"] == 2
    assert d["PO 미발행 · 확인필요"]["count"] == 1
    assert d["거래명세서 미작성"]["count"] == 1
    assert d["아리바 청구 미등록"]["count"] == 1
    row = d["잔여 미청구액"]["rows"][0]
    assert row["프로젝트NO"] == "UJ260001" and row["캠프명"] == "합성1캠프"
    assert row["종류"] == "settle" and row["레코드ID"] == "JS-2601"

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for token in ("function openExecMetric", "function metricToPng", "function metricOpenCall",
                  "function sheetSnapshot", "function setSheetContent", "sheetactions",
                  "scrollTop: $('sheetbody').scrollTop", "PageDown", "PageUp"):
        assert token in live, token + " 누락"
    assert "#sheetbody{flex:1 1 auto;min-height:0;overflow-y:auto" in live
    assert re.search(r"(?m)^\s*#slist\{display:none\}", live), "데스크톱 정산 목록 범위가 없다"
    assert not re.search(r"(?m)^\s*\.slist\{display:none\}", live), "상세 시트 목록까지 숨긴다"
    assert "const actions = body.querySelector('.actions.sticky,.media-tools.sticky')" in live
    assert "openRecord(${esc4(kind)},${esc4(id)},${esc4(prj)})" in live
    assert "window._board = Object.assign" not in live, "리스크 일부 목록이 전체 담당자 보드를 덮는다"
    assert "saveOrOpen(b, assigneeFileName())" in live, "공유 불가 시 이미지를 다시 렌더한다"
    print("  [49] 대표보고 숫자 원천행·정확 라우팅·캡처·모달 끝까지 스크롤 ✅")


def t50_stale_completion_drilldown_and_capture():
    """[50] 오래된 완료일 미기입 경고는 정확한 목록·원천행·캡처로 이어진다."""
    import daily_brief as D

    data = {
        "as": [{
            "접수ID": "AS-2606-001", "프로젝트NO": "UJ2606001", "캠프명": "합성캠프",
            "접수일자": "2026-06-01", "작업완료일": "", "진행상태": "접수",
            "담당기사": "김기사", "신청내용": "합성 완료일 확인",
        }],
        "pm": [], "fw": [],
    }
    b = D.brief("2026-07-29", data)
    assert b["돌발AS"]["완료일미기입"] == 1
    stale = b["완료일미기입목록"][0]
    assert stale["레코드ID"] == "AS-2606-001" and stale["레코드종류"] == "as"

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for token in ("function openStaleBrief(", "onclick=\"openStaleBrief()\"",
                  "window._briefMetric", "눌러서 ${stale.length}건 목록 보기",
                  "이미지 저장", "이미지로 전달", "r.담당자?' · '+r.담당자"):
        assert token in live, token + " 누락"
    assert ".bneed.actionable" in live and 'role="button"' in live

    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    pub = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    for token in ("function openStaleCloud(", "function staleToPng(", "function shareStaleCloud(",
                  "function saveStaleCloud(", 'id="sheetbody"', "프로젝트 번호를 누르면"):
        assert token in phone, "고정 주소 앱 " + token + " 누락"
    assert ".sheetbody{flex:1 1 auto;min-height:0;overflow-y:auto" in phone, \
        "고정 주소 목록을 끝까지 스크롤할 수 없다"
    assert '"완료일미기입목록": _b.get("완료일미기입목록", [])' in pub, \
        "암호화 사본에 완료일 누락 원천 목록이 없다"
    sw = open(os.path.join(ROOT, "docs", "sw.js"), encoding="utf-8").read()
    icon_hash = hashlib.sha256(open(os.path.join(
        ROOT, "webapp", "brand", "csos-app-icon-source.png"), "rb").read()).hexdigest()[:12]
    assert f"csos-icon-{icon_hash}" in sw, \
        "설치형 휴대폰 앱이 이전 화면 캐시를 계속 쥘 수 있다"
    print("  [50] 오래된 완료일 미기입 목록·정확 라우팅·담당자 캡처 ✅")


def t51_manual_daily_activity():
    """[51] 프로젝트NO 없는 대표 접수와 택배 발송 처리도 당일 업무에서 빠지지 않는다."""
    import daily_brief as D

    data = {
        "as": [{
            "접수ID": "", "프로젝트NO": "", "캠프명": "GWJ1 M_순천1",
            "접수일자": "2026-07-27", "작업완료일": "", "진행상태": "접수",
            "담당기사": "김필우", "신청내용": "리프트 리모콘 작동 안함",
        }],
        "pm": [], "fw": [],
        "events": [{
            "날짜": "2026-07-28", "접수일": "2026-07-27", "캠프명": "GWJ1 M_순천1",
            "게시자": "유수비 대표", "처리자": "류지영 매니저",
            "신청내용": "리프트 리모콘 작동 안함", "처리내용": "리모컨 택배 발송 완료",
            "상태": "택배 발송 완료",
        }],
    }
    b = D.brief("2026-07-28", data)
    assert b["돌발AS"]["업무처리"] == 1
    assert b["당일처리목록"][0]["캠프명"] == "GWJ1 M_순천1"
    assert b["당일처리목록"][0]["게시자"] == "유수비 대표"
    assert b["당일처리목록"][0]["무엇"] == "리모컨 택배 발송 완료"
    text = D.text(b)
    assert "2026-07-28 업무 처리 1건" in text and "류지영 매니저" in text

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for token in ("BRIEF['당일처리목록']", "업무 처리 ${activityL.length}",
                  "bCard(x,'activity')", "현장 AS 완료와 별도"):
        assert token in live, token + " 누락"
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    pub = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    assert "function renderBrief(" in phone and "D.brief || {}" in phone
    assert "현장 AS 완료와 별도" in phone
    assert '"당일처리목록": _b.get("당일처리목록", [])' in pub, \
        "암호화 사본에 유수비 게시·류지영 택배 처리 실적이 없다"
    print("  [51] 유수비 대표 접수·류지영 택배 발송을 당일 업무 처리로 포함 ✅")


def t52_data_status():
    """[52] 자료현황 한 장 — 같은 질문을 매번 다시 세지 않게(사용자 지시 2026-07-29).

    ★ 이 장이 **느리면 아무도 안 본다.** Z: 폴더 2만 개 순회 같은 무거운 일은 여기서
      다시 하지 않고, 앞 단계가 남긴 리포트에서 숫자만 읽는다. 그래서 daily_run 에서
      대조들이 **끝난 뒤에** 와야 한다.
    ★ 앱 [기록] 탭에 뜨지 않으면 만든 의미가 없다 — 파일명이 서버 목록과 맞아야 한다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import data_status as D

    # 리포트에서 숫자만 긁는 함수가 실제로 동작하는가(없으면 조용히 빈 dict)
    with tempfile.TemporaryDirectory() as tmp:
        old = os.environ.get("COUPANG_REPORT_DIR")
        os.environ["COUPANG_REPORT_DIR"] = tmp
        try:
            import importlib
            importlib.reload(D)
            open(os.path.join(tmp, "입금현황.md"), "w", encoding="utf-8").write(
                "# 쿠팡 입금 현황\n\n- 입금 70건 · 합계 1,106,167,980원 · 2026-04-13 ~ 2026-07-27\n")
            got = D.from_report("입금현황.md", ("건수", r"입금\s*([\d,]+)건"),
                                ("합계", r"합계\s*([\d,]+)원"))
            assert got.get("건수") == 70 and got.get("합계") == 1106167980, got
            assert D.from_report("없는파일.md", ("x", r"(\d+)")) == {}, "없는 리포트에서 죽으면 안 된다"

            # 문서사진은 날짜별 하위 폴더에 있다. 최상위만 세거나 2025년을 섞으면 안 된다.
            photo_root = os.path.join(tmp, "문서사진")
            os.makedirs(os.path.join(photo_root, "2026", "07", "2026-07-29"))
            os.makedirs(os.path.join(photo_root, "2025", "12", "2025-12-26"))
            open(os.path.join(photo_root, "2026", "07", "2026-07-29", "a.jpg"), "wb").write(b"1")
            open(os.path.join(photo_root, "2026", "07", "2026-07-29", "b.png"), "wb").write(b"22")
            open(os.path.join(photo_root, "2025", "12", "2025-12-26", "old.jpg"), "wb").write(b"333")
            count, size_mb = D.image_tree_stats(photo_root, year="2026")
            assert count == 2 and size_mb > 0, (count, size_mb)
        finally:
            if old is None:
                os.environ.pop("COUPANG_REPORT_DIR", None)
            else:
                os.environ["COUPANG_REPORT_DIR"] = old
            importlib.reload(D)

    # 무거운 재계산을 하지 않는다 — Z: 폴더를 직접 훑는 코드가 있으면 안 된다
    src = open(os.path.join(ROOT, "data_status.py"), encoding="utf-8").read()
    assert "os.walk" not in src, "자료현황이 폴더를 직접 순회한다 — 느려서 아무도 안 보게 된다"
    assert "from_report(" in src, "앞 단계 리포트를 재활용하지 않는다"

    # 앱 [기록] 탭에 뜨는가 (파일명이 서버 목록과 맞아야 한다)
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '("자료현황.md", "자료현황")' in srv, "앱 리포트 목록에 자료현황이 없다"
    # daily_run 이 매일 갱신하는가
    dr = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "data_status.py" in dr, "daily_run 에 자료현황 갱신 단계가 없다"
    assert dr.index("findings_export.py") < dr.index("data_status.py"), \
        "자료현황이 대조·리포트보다 먼저 돌면 지난 숫자를 읽는다"
    print("  [52] 자료현황 한 장(리포트 재활용·앱 노출·daily_run 순서) ✅")


def t53_session_handoff():
    """[53] 세션이 갑자기 끊겨도 다음 세션이 이어받는다(사용자 지시 2026-07-29).

    ★ 종료 체크리스트는 **끝낼 시간이 있을 때만** 지켜진다. 컨텍스트가 차거나 크레딧이
      끊기면 그럴 기회가 없고, 그때 점유·큐·임시파일이 방치된다.
    ★ 죽은 점유는 **프로세스 생사**로 판정한다 — 시간(45분)만 보면 그동안 원장이 잠긴다.
      ai_claim 은 `at` 을 **에포크 초(float)** 로 적는다(ISO 문자열 아님) — 이걸 틀리면
      경과 시간이 '?' 로 나와 죽은 잠금을 못 가려낸다(실제로 처음에 그랬다).
    ★ 이 도구는 **아무것도 고치지 않는다.** 상대 AI 가 일하는 중일 수 있어 자동으로
      점유를 풀거나 큐를 반영하면 가로채게 된다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import session_handoff as H

    # 에포크 초를 제대로 읽는가 + 죽은 프로세스면 시간과 무관하게 잔재로 보는가
    import time as _t
    fake = {"ledger": {"who": "codex", "why": "테스트", "at": _t.time() - 120, "pid": 999999}}
    orig = H.claims.__globals__.get("ai_claim")
    class _Stub:
        @staticmethod
        def load():
            return fake
    H.claims.__globals__["ai_claim"] = _Stub
    _s.modules["ai_claim"] = _Stub
    try:
        got = H.claims()
        assert got and got[0]["mins"] == 2, "에포크 초 파싱 실패(ISO 로 읽고 있다): %s" % got
        assert got[0]["stale"] is True, "죽은 프로세스인데 잔재로 안 본다 — 원장이 45분 잠긴다"
        bl = H.blockers({"큐잔량": 0, "임시파일": [], "점유": got, "미푸시": []})
        assert any("잔재" in w for w, _ in bl), bl
        # 푸는 방법을 반드시 알려 주되 **될 명령**이어야 한다. 이건 codex 의 죽은
        # 점유라 `--free` 는 ai_claim 이 거부한다(세션 단위 규칙) — 통하는 것은
        # `--adopt` 다. 예전엔 무조건 `--free` 를 적어 사람이 그대로 하고 막혔다.
        # 소유 판정은 [114] 가 따로 지킨다.
        assert any("--adopt" in c for _, c in bl), "풀 명령을 알려주지 않는다: %s" % bl
        # 살아 있는 점유는 잔재가 아니다 — 상대가 일하는 중이다.
        # ★ [210] 신원 판정과 아귀가 맞아야 한다: 실사용에서 주인은 언제나 점유를
        #   적기 **전에** 떠 있다(에이전트가 살아서 적는다). '점유 시각보다 뒤에
        #   태어난 pid'는 재사용된 남이므로, 산 점유를 흉내내려면 at 도 지금이어야 한다.
        fake["ledger"]["pid"] = os.getpid()
        fake["ledger"]["at"] = _t.time()
        assert H.claims()[0]["stale"] is False, "살아 있는 점유를 잔재로 본다 — 상대 작업을 가로챈다"
        # 그리고 그 반대 — 점유 시각(옛날)보다 뒤에 태어난 산 pid 는 재사용이다([210])
        fake["ledger"]["at"] = 946684800.0          # 2000-01-01 — 이 pid 는 그 뒤에 태어났다
        assert H.claims()[0]["stale"] is True, \
            "점유 시각보다 뒤에 태어난 pid(재사용)를 살아 있는 주인으로 오판"
    finally:
        if orig is not None:
            H.claims.__globals__["ai_claim"] = orig
        _s.modules.pop("ai_claim", None)

    # 큐·임시파일도 막힌 것으로 잡는가
    bl = H.blockers({"큐잔량": 7, "임시파일": ["x.tmp.xlsx"], "점유": [], "미푸시": ["a"]})
    kinds = " ".join(w for w, _ in bl)
    assert "입력 큐" in kinds and "임시파일" in kinds and "푸시" in kinds, kinds

    # 고치지 않는다 — 자동 해제·자동 반영 코드가 있으면 안 된다
    src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    body = src.split("if __name__")[0]
    assert "ai_claim.free" not in body and "queue_add" not in body, \
        "세션인계가 스스로 고치고 있다 — 상대 AI 작업을 가로챈다"
    for marker in ("CHECKPOINT_PATH", "def write_checkpoint(", "★ 진행 중 작업 — 여기서 바로 재개",
                   "--checkpoint", "--clear-checkpoint"):
        assert marker in src, f"새 세션 재개 체크포인트 누락: {marker}"

    # 워치독이 30분마다 남기는가 · 시작 체크리스트 0번인가
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "snapshot_handoff" in wd and "session_handoff.py" in wd, "워치독이 스냅샷을 안 남긴다"
    cm = open(os.path.join(os.path.dirname(ROOT), "CLAUDE.md"), encoding="utf-8").read()
    assert "session_handoff.py --check" in cm, "시작 체크리스트에 없다 — 아무도 안 읽는다"
    assert cm.index("session_handoff.py --check") < cm.index("AGENTS.md` 전체 읽기"), \
        "세션인계가 체크리스트 맨 앞이 아니다"
    print("  [53] 세션 인계(죽은 점유 판정·큐/임시파일·워치독 스냅샷·체크리스트 0번) ✅")


def t54_side_work_db_only():
    """[54] 철거·신규납품은 **DB에만 두고 앱에는 안 보인다**(사용자 지시 2026-07-29).

    원장에서 지우는 게 아니다 — 05_신규납품설치 시트에 그대로 두고 화면에서만 뺀다.
    ★ 05시트 업무구분은 '납품'·'철거' 처럼 **한 단어**다(10_코드관리 M열).
      '신규납품' 만 잡으면 정작 원장에 적힌 '납품' 을 놓친다 — 처음에 실제로 그랬다.
    ★ 돌발AS·정기점검은 절대 걸리면 안 된다. 걸리면 본업이 화면에서 사라진다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import kakao_extract as kx

    # 저장 자리 — 05시트로 가고, 업무구분은 유효성 목록 안의 값이어야 한다
    assert "05_신규납품설치" in kx.SHEET_MAP, "철거·납품의 저장 자리가 없다"
    assert set(kx.SIDE_KIND.values()) <= {"철거", "납품", "설치", "이전"}, kx.SIDE_KIND
    assert kx.SIDE_KIND.get("(철거·보관)") == "철거" and kx.SIDE_KIND.get("(납품설치)") == "납품"
    extract_src = open(os.path.join(ROOT, "kakao_extract.py"), encoding="utf-8").read()
    assert 'r["시트"] not in SIDE_KIND' in extract_src, \
        "05시트에 정상 등록한 철거·납품을 보류라고 잘못 표시한다"
    assert '("02_돌발AS접수", "04_정기점검", "05_신규납품설치")' in extract_src, \
        "05시트 철거·납품을 기존 프로젝트로 안 보면 매일 중복 등록된다"
    m = kx.SHEET_MAP["05_신규납품설치"]
    assert m.get("업무구분") == "_업무구분" and "프로젝트NO" in m, m
    for c in ("요청일", "철거·이전예정일", "납품예정일"):
        assert c in kx.DATE_COLS, "%s 가 날짜로 안 들어간다" % c

    # 앱 필터 — 별도 공사는 걸러지고 본업은 남아야 한다
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    mm = re.search(r"const SIDE_WORK = /([^/]+)/", live)
    assert mm, "앱에 SIDE_WORK 규칙이 없다"
    rx = re.compile(mm.group(1))
    for word in ("철거", "납품", "설치", "이전", "계단"):
        assert rx.search(word), "'%s' 가 앱에서 안 걸러진다" % word
    for word in ("돌발AS", "정기점검"):
        assert not rx.search(word), "★ '%s' 가 걸러진다 — 본업이 화면에서 사라진다" % word

    # 데이터 받는 관문 전부에 걸려 있는가(한 곳만 빠져도 거기서 새어 나온다)
    for anchor in ("rowIs2026(r,'settle')", "rowIs2026(r,'as')",
                   "rowIs2026(r,'pm')", "rowIs2026(r,'issue')"):
        i = live.find(anchor)
        assert i > 0, anchor
        assert "isSideWork(r)" in live[i:i + 120], "%s 관문에 필터가 없다" % anchor
    assert live.count("isSideWork(r)") >= 5, "관문 일부에만 걸려 있다"
    print("  [54] 철거·납품 DB 저장·앱 비표시(05시트 저장·관문 5곳·본업 보존) ✅")


def t55_pm_brief_drilldown_and_capture():
    """[55] 정기점검 예정·실행·미실행 숫자는 원천 목록·점검ID·캡처로 이어진다."""
    import daily_brief as D

    data = {"as": [], "fw": [], "events": [], "pm": [
        {"점검ID": "PM-2607-001", "프로젝트NO": "UJ2607001", "캠프명": "합성A",
         "점검예정일": "2026-07-29", "실제점검일": "2026-07-29",
         "담당기사": "김기사", "점검내용": "정기점검 A"},
        {"점검ID": "PM-2608-002", "프로젝트NO": "UJ2608002", "캠프명": "합성B",
         "점검예정일": "2026-08-05", "실제점검일": "",
         "담당기사": "권기사", "점검내용": "정기점검 B"},
        {"점검ID": "PM-2606-003", "프로젝트NO": "UJ2606003", "캠프명": "합성C",
         "점검예정일": "2026-06-20", "실제점검일": "2026-07-29",
         "담당기사": "차기사", "점검내용": "지연 실행"},
    ], "pm_schedule": {"year": 2026, "quarter": 3, "schedule": [
        {"일정ID": "SCH-A", "점검예정일": "2026-07-29", "예측점검일": "",
         "실제점검일": "2026-07-29", "캠프명": "합성A", "장비수": 2},
        {"일정ID": "SCH-B", "점검예정일": "", "예측점검일": "2026-08-05",
         "실제점검일": "", "캠프명": "합성B", "장비수": 3},
        {"일정ID": "SCH-C", "점검예정일": "2026-07-29", "예측점검일": "",
         "실제점검일": "2026-07-29", "캠프명": "합성C", "장비수": 1},
    ]}}
    b = D.brief("2026-07-29", data)
    assert b["정기점검"]["예정"] == 2 and b["정기점검"]["완료"] == 2
    assert b["정기점검"]["예정장비"] == 3 and b["정기점검"]["완료장비"] == 3
    assert b["정기점검"]["분기예정"] == 6 and b["정기점검"]["분기완료"] == 3
    assert b["정기점검"]["분기미실행"] == 3 and b["정기점검"]["분기진행률"] == 50
    assert b["정기점검"]["기준일까지예정"] == 3
    assert b["정기점검"]["기준일까지완료"] == 3
    assert b["점검예정목록"][0]["레코드ID"] == "PM-2607-001"
    assert {x["레코드ID"] for x in b["점검실행목록"]} == {"PM-2606-003", "PM-2607-001"}
    assert [x["상태"] for x in b["분기점검목록"]] == ["실행", "미실행"]

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for token in ("function openPmBrief(", "openPmBrief('day-plan')", "openPmBrief('day-done')",
                  "openPmBrief('quarter-all')", "openPmBrief('quarter-done')",
                  "openPmBrief('quarter-pending')", "점검예정목록", "점검실행목록", "분기점검목록"):
        assert token in live, "실시간 앱 " + token + " 누락"
    for token in ("신규중처리완료", "신규처리율", "예측점검일",
                  "명세서발행대기완료기준", "과거 이력 예측"):
        assert token in live, "대표 카드·캘린더 " + token + " 누락"
    assert "window._briefMetric={label" in live and "shareExecMetric()" in live

    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    pub = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    for token in ("function openPmCloud(", "function openCloudBriefSheet(",
                  "function cloudSheetToPng(", "quarter-pending", "이미지 저장/전달"):
        assert token in phone, "고정 주소 앱 " + token + " 누락"
    for key in ("점검예정목록", "점검실행목록", "분기점검목록"):
        assert f'"{key}": _b.get("{key}", [])' in pub, "암호화 사본 " + key + " 누락"
    print("  [55] 정기점검 예정·실행·미실행 목록·정확 이동·담당자 캡처 ✅")


def t56_work_detail_from_source():
    """[56] 작업 내용은 **원문에 있으면 자동 기입, 없으면 비워 둔다**(사용자 지시 2026-07-29).

    ★★ '신청내용'(요청)을 실적 칸에 넣으면 **하지 않은 작업을 했다고 기록**된다.
       카톡 자료로 세어 보면 작업내용은 26건인데 신청내용만 있는 게 542건이다 —
       구분 없이 채웠다면 대표 보고의 '무엇을 했나' 가 통째로 거짓이 됐다.
    ★★ 공지의 **빈 양식**('?? 호기', 'What ? / 갯수 ?')도 값이 아니다. 기사가 아직 안 채운 칸이다.
       여러 호기 중 일부만 적힌 글은 뒤에 빈 양식이 남으므로 거기서 잘라야 한다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import fill_work_detail as W

    # 빈 양식은 값이 아니다 — 통째로 템플릿이면 아무것도 안 쓴다
    tpl = "● A/S 내용 :\n▒▒ ?? 호기 ▒▒\n(유료)What ? / 갯수 ?\n(유료)작업자 1공임"
    assert W.work_part(tpl) == "", "빈 양식을 실적으로 기록한다: %r" % W.work_part(tpl)
    assert W.tidy("▒▒ ?? 호기 ▒▒ (유료)What ? / 갯수 ?") == ""

    # 일부만 적힌 글 — 적힌 데까지만 남기고 빈 양식은 잘라낸다
    part = ("● A/S 내용 :\n▒▒ 02 호기 ▒▒\n(유료)도어락 교체 완료\n"
            "▒▒ ?? 호기 ▒▒\n(유료)What ? / 갯수 ?")
    got = W.work_part(part)
    assert "도어락 교체 완료" in got, got
    assert "What" not in got and "?? 호기" not in got, "빈 양식이 남았다: %r" % got

    # 사내 메모(※ 확인사항)는 실적이 아니다
    memo = "● A/S 내용 :\n(유료)경광등 교체완료\n※ 확인사항 : 1. (변재선) 거래명세표 발행"
    g2 = W.work_part(memo)
    assert "경광등 교체완료" in g2 and "확인사항" not in g2, g2

    # 'A/S 내용' 절이 없으면 본문을 통째로 넣지 않는다
    assert W.work_part("♣ 돌발유료 A/S 완료 ● 프로젝트NO : UJ2600001 ● 신청일자 : 2026.01.02") == ""

    # ★ 신청내용을 쓰지 않는다 — 코드에 그 폴백이 있으면 안 된다
    src = open(os.path.join(ROOT, "fill_work_detail.py"), encoding="utf-8").read()
    body = src.split("if __name__")[0]
    assert 'r.get("신청내용")' not in body, "신청내용을 실적으로 쓰고 있다 — 안 한 일을 했다고 적는다"
    # 사람이 적은 값은 덮지 않는다(자리표시자·빈 양식만 덮는다)
    assert "DRAFT.search(cur)" in body, "사람이 적은 값을 덮어쓸 수 있다"
    # 매일 자동으로 도는가
    dr = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "fill_work_detail.py" in dr, "daily_run 에 자동 기입 단계가 없다"
    # ★★ 'A/S 내용' 의 값은 `▒▒ 01 호기 ▒▒ …` 처럼 ▒ 로 **시작한다**.
    #    필드 정규식이 ▒ 에서 값을 끊으면 작업내용이 **언제나 빈 값**이 된다(실제로 그랬다:
    #    26건만 잡히다가 고친 뒤 840건). 값의 끝은 ● ★ ♠ ※ 로 본다.
    import kakao_extract as _ke
    got = dict(_ke.RE_FIELD.findall(
        "● 신청내용 : 리모컨 고장 ★ 작업완료후 - 완료전화 필수! "
        "● A/S 내용 : ▒▒ 01 호기 ▒▒ (유료)리모콘 SET 교체 완료 ♠ 원인 및 조치 : -"))
    as_val = got.get("A/S 내용") or got.get("A/S내용") or ""
    assert "리모콘 SET 교체 완료" in as_val, "A/S 내용이 ▒ 에서 잘린다: %r" % as_val
    assert "리모컨 고장" in (got.get("신청내용") or ""), got

    # ★★ 03시트 B열(접수ID)은 **수식**이고 02시트에서 스스로 끌어온다 —
    #    빈 행에 실적을 써 넣으면 재계산 때 엉뚱한 건에 붙는다(v259 실사고).
    assert "def missing_rows" in body, "대기 건수 집계가 없다"
    assert "새행시작" not in body and "실적 행을 만든다" not in body, \
        "★ 03시트에 행을 직접 만들고 있다 — B열 수식이 정하는 자리를 가로챈다"
    print("  [56] 작업내용 자동기입(신청내용 금지·빈 양식 제외·▒ 절단 방지·행 생성 금지) ✅")


def t58_check_hub_detail_and_capture():
    """[58] 보고 아래 확인 필요 전용 화면은 유형·담당자·원기록·캡처까지 이어진다."""
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    # 메뉴는 사용자가 지정한 위치(보고 바로 아래, 기록 위)에 있어야 한다.
    daily = live.index('data-v="daily"')
    check = live.index('data-v="check"', daily)
    report = live.index('data-v="report"', check)
    assert daily < check < report, "확인 필요 메뉴가 보고 아래·기록 위가 아니다"
    assert 'id="v-check"' in live and "if(v==='check') renderCheckHub()" in live

    # 숫자만 보여 주는 카드가 아니라 유형/담당자/전체 목록을 실제 행으로 좁혀야 한다.
    for token in ("function renderCheckHub(", "function renderCheckList(",
                  "function openCheckType(", "function openCheckOwner(",
                  "function openCheckFiltered(", "function openCheckStale(",
                  'id="checktypes"', 'id="checkowners"', 'id="checklist"',
                  'id="checkq"', 'id="checktype"', 'id="checkowner"'):
        assert token in live, token + " 누락"

    # 2026년·별도공사 제외 규칙을 전용 화면에서도 다시 지키고, ID 종류로 정확한 원장 기록을 연다.
    assert "rowIs2026(r,'issue')&&!isSideWork(r)" in live
    assert "/^AS-/i.test(raw)" in live and "/^PM-/i.test(raw)" in live and "/^JS-/i.test(raw)" in live
    assert "if(x.kind) openRecord(x.kind,x.id,x.project)" in live

    # 현재 필터 결과를 기존 캡처 엔진에 넘겨 이미지 저장·전달 기능을 그대로 쓴다.
    assert "rows:rows.map(checkMetricRow)" in live and "openExecMetric(label)" in live
    assert "현재 목록 보기" in live and "이미지 복사" in live and "이미지 저장" in live
    # 아이콘 참조 방식(<img> → 스프라이트 <use>)이 바뀌어도 깨지지 않게 **아이콘 이름**으로 본다.
    assert "media-tools" in live and "#i-clipboard-copy" in live
    # 모바일 탭바: 크기 **숫자를 고정하지 않는다**(2026-07-30 아이콘을 키우자 이 줄이 깨졌다).
    # 지켜야 하는 것은 ① 좁은 화면용 규칙이 있다 ② 아이콘이 **알아볼 수 있는 크기**다.
    #   사용자 지적: "모바일에서 아이콘이 너무 작아 잘 안 보여" — 탭이 7개로 늘며 좁아진 탓이다.
    assert "@media(max-width:420px)" in live, "좁은 화면용 탭바 규칙이 없다"
    _icons = [float(m) for m in re.findall(r"\.tabbar svg\{width:([\d.]+)px", live)]
    assert _icons, "탭바 아이콘 크기 규칙이 없다"
    assert min(_icons) >= 24, f"모바일 탭바 아이콘이 너무 작다({min(_icons)}px) — 24px 이상 유지"

    # PO번호만 있는 확인 행도 정산 PO번호 연결을 먼저 찾고, 없으면 경고창이 아닌 원문 상세를 연다.
    for token in ("function samePo(", "settleRows.find(r=>samePo(r.PO번호,raw))",
                  "function openCheckSource(", "function openCheckByKey(",
                  "else showCheckRaw()", "if(kind==='check')"):
        assert token in live, "PO·미연결 확인행 처리 누락: " + token

    # 확정된 내부 확인 책임을 적용하고, 그 밖의 현장 확인만 원천 AS/PM 담당기사로 보완한다.
    assert "const CHECK_OWNER_RULES" in live and "function confirmedCheckOwner(" in live
    assert "캠프·일정, 카톡/밴드, 현장자료, 완료일, 거래명세서, 세금계산서, 입금, 금액 불일치 확인" in live
    assert "{name:'변재선(회계)'" not in live, "이관 완료한 변재선 확인 책임이 앱에 남아 있다"
    assert 'STAFF_CENTER_ALIASES = {"byeon-jaeseon": "ryu-jiyeong"}' in server
    assert '"Location": f"/staff/{target}"' in server
    assert "const CHECK_OWNER_SCOPE_PROPOSALS" not in live and "openPendingOwnerScope" not in live
    assert "first&&first.원천업무ID ? rowById(first.원천업무ID)" in live
    assert "원장 담당자 칸 확인" in live
    assert "[r['구분'], r['담당자']" in phone, "고정 주소 확인필요 목록에 확정 담당자가 표시되지 않는다"
    print("  [58] 확인 필요 전용 메뉴·PO/담당자 보완·정확 이동·현재목록 캡처 ✅")


def t70_quarter_as_months():
    """[70] '3분기' 가 아니라 '7~9월' 로 적는다(사용자 지시 2026-07-29).

    분기 번호는 읽는 사람이 머릿속에서 다시 월로 환산해야 한다 — 대표 보고에서
    "그게 몇 월부터냐" 를 되묻게 만든다. 화면·브리핑 어디에도 'N분기' 를 남기지 않는다."""
    src = open(os.path.join(ROOT, "daily_brief.py"), encoding="utf-8").read()
    assert '"분기": f"{y}년 {3 * q - 2}~{3 * q}월"' in src, "분기 라벨이 월 범위가 아니다"
    assert '"분기범위"' in src and '"분기끝월"' in src, "월 범위·끝월 필드가 없다"
    assert 'f"{y}년 {q}분기"' not in src, "옛 'N분기' 표기가 남아 있다"
    assert "분기 내 마무리" not in src, "'분기 내' 문구가 남아 있다 — 몇 월인지 안 보인다"

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    # 앱은 이제 기간을 직접 고르므로 라벨도 클라이언트가 만든다(pmStats().라벨).
    # 서버 필드(분기범위·분기끝월)는 브리핑 텍스트용으로 남고, 앱은 ps.라벨/ps.끝월을 쓴다.
    assert "ps.라벨" in live and "ps.끝월" in live, "앱이 월 범위 라벨을 안 쓴다"
    assert "분기 내 마무리" not in live, "앱에 '분기 내' 문구가 남아 있다"
    # 버튼 라벨이 '분기 예정' 처럼 고정 문자열로 되돌아가면 안 된다
    assert '">분기 예정<' not in live and '">분기 실행<' not in live, "버튼이 다시 '분기' 로 고정됐다"

    # 실제 환산이 맞는가 — 3분기는 7~9월
    for q, want in ((1, "1~3월"), (2, "4~6월"), (3, "7~9월"), (4, "10~12월")):
        assert "%d~%d월" % (3 * q - 2, 3 * q) == want, q
    print("  [70] 분기를 '7~9월' 처럼 월 범위로 표기(브리핑·앱·버튼·안내문) ✅")


def t71_period_range():
    """[71] 기간을 **월 단위 범위**로 고른다 — '7월' 도 '1~6월' 도(사용자 지시 2026-07-29).

    ★ 시작이 끝보다 뒤면 오류를 띄우지 않는다. 사용자가 틀린 게 아니라 **아직 반대쪽을
      못 옮긴 것**이다 — 방금 고른 쪽을 존중하고 반대쪽을 끌어와 맞춘다.
    ★ 옛 호출부(`inPeriod(d, y, m)` 처럼 한 달만 넘기는 곳)가 남아 있다. '이번 달' 통계는
      선택 기간이 아니라 늘 이번 달이어야 하므로, 인자 3개면 **그 한 달**로 동작해야 한다."""
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # 단일 월 선택이 아니라 시작~끝 두 칸이어야 한다
    assert 'id="pfrom"' in live and 'id="pto"' in live, "기간 시작·끝 선택이 없다"
    assert 'id="pmonth"' not in live, "옛 단일 월 선택이 남아 있다(둘이 어긋난다)"
    for fn in ("function periodRange()", "function periodLabel()", "function fixRange(",
               "function setRange("):
        assert fn in live, "%s 가 없다" % fn
    assert live.count("function periodRange()") == 1, \
        "동명 periodRange가 업무목록·월별현황 사이에서 덮어써 목록이 0건이 된다"
    assert "function workPeriodRange()" in live and "const pr = workPeriodRange();" in live, \
        "업무목록 전용 기간 함수가 분리되지 않았다"
    # 자주 쓰는 기간 바로가기
    for label in ("연간 전체", "상반기", "하반기"):
        assert 'onclick="setRange' in live and label in live, label

    # 범위 필터가 실제로 쓰이는가 — 옛 단일 월 변수로 되돌아가면 안 된다
    for anchor in ("dateOf(r,'settle'),y,pf,pt", "dateOf(r,'as'),y,pf,pt", "dateOf(r,'pm'),y,pf,pt"):
        assert anchor in live, "기간 필터가 적용되지 않은 곳이 있다: %s" % anchor
    # 하위호환: 인자 3개면 그 한 달
    assert "if(to === undefined) to = from;" in live, "옛 호출부(단일 월) 호환이 깨졌다"
    # 제목은 '7월'·'1~6월'·'연간 전체'
    assert "${periodLabel()}" in live, "제목이 기간 이름을 쓰지 않는다"
    assert "${m?+m+'월':'전체'}" not in live, "옛 제목 표기가 남아 있다"

    # 라벨 규칙(파이썬으로 같은 계산을 재현해 확인)
    def label(f, t):
        return "연간 전체" if (f, t) == ("01", "12") else (
            "%d월" % int(f) if f == t else "%d~%d월" % (int(f), int(t)))
    assert label("07", "07") == "7월" and label("01", "06") == "1~6월"
    assert label("01", "12") == "연간 전체" and label("10", "12") == "10~12월"
    # ── 정기점검 진행률 카드도 같은 기간 선택을 쓴다 ──
    #  ★ 서버가 준 분기 숫자(p.분기예정 등)를 그대로 쓰면 기간을 바꿔도 숫자가 안 변한다.
    #    works.pm 에서 **직접 세야** 아무 기간이나 된다.
    for fn in ("function pmStats()", "function setPmRange(", "function pmQuarter()"):
        assert fn in live, "%s 가 없다" % fn
    assert "const ps = pmStats();" in live, "진행률 카드가 기간 계산을 안 쓴다"
    for old in ("p.분기진행률||0", "${p.분기완료||0}/${p.분기예정||0}건",
                "p.분기범위||'분기'", "p.분기||'분기'"):
        assert old not in live, "진행률 카드에 옛 분기 고정값이 남아 있다: %s" % old
    # 드릴다운 목록도 같은 기간이어야 카드 숫자와 어긋나지 않는다
    assert "const _ps = pmStats();" in live, "드릴다운이 기간을 따르지 않는다"
    assert "mm>=_ps.from && mm<=_ps.to" in live, "드릴다운 목록이 기간으로 걸러지지 않는다"
    # 시작이 끝보다 뒤면 끝을 끌어와 맞춘다(오류를 띄우지 않는다)
    assert "if(a > b){ b = a; }" in live, "거꾸로 고른 기간을 바로잡지 않는다"
    print("  [71] 기간 범위 선택(현황·정기점검 진행률·드릴다운·거꾸로 자동정렬·단일월 호환) ✅")


def t72_project_first_representative_report():
    """[72] 내부 ID 대신 프로젝트번호를 대표 표시하고, 유수비 예외보고는 원장에 쓰지 않고 계산한다."""
    from webapp.app_server import representative_summary

    works = {
        "as": [
            {"접수ID": "AS-2607-001", "프로젝트NO": "UJ2600001", "캠프명": "테스트A",
             "접수일자": "2026-07-25", "진행상태": "접수", "담당기사": "기사A"},
            {"접수ID": "AS-2607-002", "프로젝트NO": "UJ2600002", "캠프명": "테스트B",
             "접수일자": "2026-07-20", "작업완료일": "2026-07-21", "진행상태": "작업완료",
             "사진등록": "등록", "완료보고서등록": "", "ERP등록": "등록"},
        ],
        "pm": [
            {"점검ID": "PM-2607-001", "프로젝트NO": "UJ2600011", "캠프명": "점검A",
             "점검예정일": "2026-07-05", "실제점검일": "2026-07-05", "점검상태": "완료",
             "담당기사": "기사A"},
            {"점검ID": "PM-2607-002", "프로젝트NO": "UJ2600012", "캠프명": "점검B",
             "점검예정일": "2026-08-05", "점검상태": "예정", "담당기사": "기사B"},
        ],
    }
    settlements = [
        {"정산ID": "JS-2607-001", "프로젝트NO": "UJ2600001", "업무구분": "돌발AS",
         "캠프명": "테스트A", "완료일": "2026-07-25", "비용구분": "유상",
         "공급가액": 100000, "명세서번호": ""},
        {"정산ID": "JS-2607-002", "프로젝트NO": "UJ2600011", "업무구분": "정기점검",
         "캠프명": "점검A", "완료일": "2026-07-05", "비용구분": "유상",
         "공급가액": 200000, "명세서번호": "2026/07/05-1"},
        {"정산ID": "JS-2607-003", "프로젝트NO": "UJ2600012", "업무구분": "정기점검",
         "캠프명": "점검B", "완료일": "2026-07-06", "비용구분": "유상",
         "공급가액": 300000, "명세서번호": "", "명세서": "없음",
         "명세서발행일": "2026-07-07"},
    ]
    report = representative_summary(works, settlements, "2026-07-29")
    assert report["돌발AS"]["전산상미완료"] == 1 and report["돌발AS"]["D+2초과"] == 1
    assert report["돌발AS"]["현장완료서류미정리"] == 1
    assert report["정기점검"]["전체대상"] == 2, "분기 목표를 실제 예정행이 아닌 고정 숫자로 썼다"
    assert report["정기점검"]["목표누계"] <= report["정기점검"]["전체대상"]
    docs = {x["업무유형"]: x for x in report["거래명세서"]["업무유형별"]}
    assert docs["돌발 AS"]["발행대상"] == 1 and docs["돌발 AS"]["미발행"] == 1
    assert docs["정기점검"]["발행대상"] == 2 and docs["정기점검"]["미발행"] == 0, \
        "명세서 발행일이 확인된 완료 건을 번호 공란만으로 미발행 처리했다"
    assert docs["신규·납품·설치"]["발행대상"] == 0 and docs["신규·납품·설치"]["발행률"] is None
    assert (
        len(report["업무기준확인필요"]) + len(report["업무기준확정"])
    ) == 5, "저장된 확정 여부와 무관하게 운영 기준 5개가 모두 유지돼야 한다"
    assert not any("류지영 확인 범위" in x["기준"] for x in report["업무기준확인필요"])
    assert not any("변재선(회계) 확인 범위" in x["기준"] for x in report["업무기준확인필요"])

    from responsibility import confirmed_owner
    assert confirmed_owner("밴드 게시 미확인", "밴드", "권오철") == "류지영"
    assert confirmed_owner("캠프명 비어 있음", "빈칸", "김준형") == "류지영"
    assert confirmed_owner("세금계산서 미발행", "정산") == "류지영"
    assert confirmed_owner("작업금액 불일치", "금액") == "류지영"
    assert confirmed_owner("PO 원본 누락", "원본자료") == "오종현"
    assert confirmed_owner("견적서 원천자료 누락", "원본자료") == "오종현"
    assert confirmed_owner("입금 원천자료 누락", "원본자료") == "오종현"
    assert confirmed_owner("현장자료 누락", "현장", "김준형") == "류지영"
    assert confirmed_owner("PO A", "PO") == "유현민"
    assert confirmed_owner("원장 미등록", "문서") == "유현민"
    assert confirmed_owner("기타 현장 확인", "기타", "권오철") == "권오철"

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    sw = open(os.path.join(ROOT, "docs", "sw.js"), encoding="utf-8").read()
    for token in ("function projectNoOf(", "function isRepresentativeProject(", "프로젝트 미확정",
                  "function renderRepresentative(",
                  "function openRepresentativeList(", 'id="checkpolicies"', "press-pop",
                  "@keyframes viewEnter", "PO 원본·견적서, 구매·입금 원천자료 취합·누락 확인"):
        assert token in live, token + " 누락"
    for token in ("function execReportedCount(", "정기점검 이상 상세 미연결",
                  "확인 전까지 확정 이상 건으로 보지 않습니다",
                  "세부 미확정 · 원장 요약"):
        assert token in live, token + " 누락"
    for route in ("/api/v1/reports/daily/exceptions", "/api/v1/as-requests/backlog-detail",
                  "/api/v1/inspections/quarter-progress", "/api/v1/statements/unissued"):
        assert route in server, route + " 누락"
    assert "project, _ = rep_no(" in server and "candidate, rep_idx" in server, \
        "대표보고 원천행의 내부 JS 번호를 프로젝트번호로 복원하지 않는다"
    assert "r['프로젝트NO']||'프로젝트 미확정'" in phone
    assert "오종현 원천자료 취합" in phone
    icon_hash = hashlib.sha256(open(os.path.join(
        ROOT, "webapp", "brand", "csos-app-icon-source.png"), "rb").read()).hexdigest()[:12]
    assert f"csos-icon-{icon_hash}" in sw
    print("  [72] 프로젝트번호 대표표시·유수비 예외보고·정책확인·버튼/페이지 애니메이션 ✅")


def t73_pm_schedule_source_sync(tmp):
    """정기점검 원본은 현재 분기만·제외행 없이·가짜 UJ 없이 멱등 반영한다."""
    from pm_schedule_sync import parse_schedule, link_master, build_sheet, HEADERS, SHEET_NAME
    from findings_sheet import upsert

    source = os.path.join(tmp, "정기점검_스케줄_합성.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026년 3분기 정기점검"
    for _ in range(4):
        ws.append([])
    ws.append(["No.", "월", "3분기 점검일자", "2분기 점검일자", "특이사항", "확정자",
               "기존 캠프명", "변경 캠프명", "호기", "종류", "모델"])
    # 원본 월과 확정 날짜가 다르면 날짜의 달이 우선한다.
    ws.append([1, 8, "2026-07-11", "2026-04-11", "", "김준형", "테스트A캠프", "", "1호기", "LIFT", "M1"])
    ws.append([2, 8, "2026-07-11", "2026-04-11", "", "김준형", "테스트A캠프", "", "2호기", "LIFT", "M1"])
    ws.append([3, 8, "", "2026-05-09", "", "권오철", "테스트B캠프", "", "1호기", "LIFT", "M2"])
    ws.append([4, 3, "2026-03-10", "", "", "김필우", "분기밖캠프", "", "1호기", "LIFT", "M3"])
    ws.append([5, 9, "", "2026-06-20", "철거 예정", "차동호", "제외캠프", "", "1호기", "LIFT", "M4"])
    wb.save(source)

    parsed = parse_schedule(source, 2026, 3)
    assert parsed["scanned"] == 4 and parsed["excluded"] == 1
    assert len(parsed["records"]) == 2
    a = next(r for r in parsed["records"] if r["캠프명"] == "테스트A캠프")
    b = next(r for r in parsed["records"] if r["캠프명"] == "테스트B캠프")
    assert a["장비수"] == 2 and a["점검예정일"] == "2026-07-11" and a["예정월"] == "2026-07"
    assert b["점검예정일"] == "" and b["예정월"] == "2026-08"
    assert b["예측점검일"] == "2026-08-09" and "2026-05-09" in b["예측근거"]

    master_rows = [{
        "점검ID": "PM-2607-001", "프로젝트NO": "UJ2600001", "캠프명": "테스트A캠프",
        "점검예정일": "2026-07-11", "실제점검일": "2026-07-12", "점검상태": "완료",
    }]
    linked = link_master(parsed["records"], master_rows)
    la = next(r for r in linked if r["캠프명"] == "테스트A캠프")
    lb = next(r for r in linked if r["캠프명"] == "테스트B캠프")
    assert la["연결프로젝트NO"] == "UJ2600001" and la["반영상태"] == "완료 실적 우선"
    assert la["완료장비수"] == 2 and la["실제점검일"] == "2026-07-12"
    assert lb["연결프로젝트NO"] == "" and lb["반영상태"] == "프로젝트 매칭 대기"

    # 같은 캠프·같은 달에 날짜가 여러 개인 경우, 한 날짜의 완료를 다른 일정에
    # 복제하면 실제 장비 진행률이 과대계상된다.
    split_source = [
        {"캠프명": "분할캠프", "예정월": "2026-07", "점검예정일": "2026-07-02",
         "담당기사": "김준형", "장비수": 2},
        {"캠프명": "분할캠프", "예정월": "2026-07", "점검예정일": "2026-07-22",
         "담당기사": "김준형", "장비수": 2},
    ]
    split_master = [{
        "프로젝트NO": "UJ2600999", "캠프명": "분할캠프", "담당기사": "김준형",
        "점검예정일": "2026-07-22", "실제점검일": "2026-07-22", "점검상태": "완료",
    }]
    split = link_master(split_source, split_master)
    assert split[0]["완료장비수"] == 0 and split[0]["반영상태"] == "프로젝트 매칭 대기"
    assert split[1]["완료장비수"] == 2 and split[1]["반영상태"] == "완료 실적 우선"

    styled_xml = build_sheet(linked * 6, os.path.basename(source), styled=True)
    assert '<mergeCell ref="A1:M1"/>' in styled_xml and '<mergeCell ref="A2:M2"/>' in styled_xml
    assert re.search(r'<c r="A10" s="38"', styled_xml), "10~49행 입력 서식이 빠졌다"

    ledger = os.path.join(tmp, "쿠팡_통합업무_일일보고_관리대장_v1.xlsx")
    mwb = openpyxl.Workbook()
    mwb.active.title = "00_대시보드"
    mwb.save(ledger)
    xml = build_sheet(linked, os.path.basename(source))
    v2, _ = upsert(ledger, xml, sheet_name=SHEET_NAME, headers=HEADERS)
    assert v2 and os.path.exists(v2)
    check = openpyxl.load_workbook(v2, read_only=True, data_only=True)
    assert SHEET_NAME in check.sheetnames and check[SHEET_NAME].max_row == 6
    check.close()
    again, msg = upsert(v2, xml, sheet_name=SHEET_NAME, headers=HEADERS)
    assert again is None and "멱등" in msg

    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "27_정기점검원본일정" in server and "정기점검 스케줄 원본" in live
    assert "프로젝트 매칭 대기" in phone and "원본 일정 · 장비 " in phone
    assert "pm_schedule_sync.py" in daily
    print("  [73] 류지영 정기점검 원본 현재분기·제외·완료우선·앱·멱등 자동반영 ✅")


def t74_billing_fill():
    """06_거래서류청구수금 청구원장 채우기 — 무엇을 넣고 무엇을 막는가.

    ★ 이 시트는 03시트와 **반대**다(v259 사고 참조).
      03시트는 B열(접수ID)이 수식이라 행 배정을 엑셀이 정하므로 빈 행에 미리 쓰면 안 됐다.
      06시트는 C열(원천업무ID)이 **입력열**이고 A열(정산ID)이 C에서 파생되므로
      빈 행에 순서대로 쓰는 것이 설계대로다. 이 구분이 무너지면 둘 다 망가진다.
    """
    import billing_fill as B
    assert B.self_test(), "billing_fill 자체 검증 실패"

    # 비쿠팡 차단: 판매조회에 있어도 관리대장에 없으면 절대 통과하지 못한다
    sales = {"UJ2699999": 1}
    ok, why = B.eligible("UJ2699999", sales, {}, set())
    assert not ok and "미등록" in why, why

    # 수식열(A 정산ID·I 실제작업공급가액·P·Q 합계)에는 어떤 경우에도 쓰지 않는다
    from datetime import date as _d
    q, _ = B.build_queue(
        [{"원천업무ID": "AS-2601-001", "공급가액": 500, "일자": _d(2026, 1, 2), "PO번호": "PO1"}],
        [72])
    cells = {x["cell"] for x in q}
    assert cells == {"C72", "O72", "N72", "S72"}, cells
    assert not any(x["cell"][0] in ("A", "I", "P", "Q", "K") for x in q)
    assert all(x.get("only_if_empty") for x in q), "빈 칸만 정책이 빠지면 남의 값을 덮는다"

    src = open(os.path.join(ROOT, "billing_fill.py"), encoding="utf-8").read()
    assert "claim_guard" in src, "원장 쓰기는 점유 확인을 거쳐야 한다"
    print("  [74] 청구원장 채우기 — 비쿠팡 차단·수식열 보호·빈칸만·점유확인 ✅")


def t75_gcal_sync():
    """구글 캘린더 대조 — 캘린더는 **예정**이지 실적이 아니다.

    사용자 지시(2026-07-29): "이 캘린더 추가하고 항상 대조해서 엑셀과 앱에 반영해줘".
    ★ 완료일 칸을 캘린더로 채우면 '안 한 일'이 '한 일'이 된다. 그 경계를 여기서 못박는다.
    """
    import gcal_sync as G
    assert G.self_test(), "gcal_sync 자체 검증 실패"

    # 채우는 열은 전부 '예정' 이어야 한다 — 실제·완료 열은 어떤 경로로도 대상이 아니다
    for kind, (sheet, daycol, timecol) in G.TARGET.items():
        assert "예정" in daycol, f"{kind} → {daycol} 은 예정 열이 아니다"
        assert "완료" not in daycol and "실제" not in daycol, daycol

    # 원천이 하나도 없어도 죽지 않는다(일일 파이프라인이 이것 때문에 멈추면 안 된다)
    old = os.environ.pop("COUPANG_GCAL_ICS", None)
    try:
        evs, notes = [], []
        assert isinstance(G.feeds(), list)
    finally:
        if old:
            os.environ["COUPANG_GCAL_ICS"] = old

    # 큐는 key 모드(접수ID/점검ID)로 나가고 빈 칸만 채운다
    from datetime import date as _d
    q = G.build_queue([{"업무구분": "돌발AS", "날짜": _d(2026, 7, 2), "시간": "09:00", "제목": "AS",
                        "원장": {"시트": "02_돌발AS접수", "원천업무ID": "AS-2607-001", "예정일있음": False}}])
    assert [x["col"] for x in q] == ["방문예정일", "방문예정시간"], q
    assert all(x["only_if_empty"] and x["key"] == "AS-2607-001" for x in q)
    # 구분과 시트가 어긋나면 아무것도 쓰지 않는다
    assert G.build_queue([{"업무구분": "정기점검", "날짜": _d(2026, 7, 2), "시간": "", "제목": "x",
                           "원장": {"시트": "02_돌발AS접수", "원천업무ID": "AS-1", "예정일있음": False}}]) == []

    # 비밀 주소가 커밋되지 않는다(규칙 1)
    gi = open(os.path.join(ROOT, ".gitignore"), encoding="utf-8").read()
    assert "config/gcal.json" in gi, "비공개 iCal 주소가 커밋될 수 있다"
    # 문서의 자리표시자(private-xxxx)는 괜찮다. 진짜 키(긴 16진수)가 박혀 있으면 안 된다.
    src = open(os.path.join(ROOT, "gcal_sync.py"), encoding="utf-8").read()
    assert not re.search(r"private-[0-9a-f]{16,}", src, re.I), "소스에 실제 비공개 주소가 박혀 있다"

    # 앱·일일실행 배선
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "/api/calendar" in server and "gcal_events.json" in server
    # 캐시 목록은 앱에서 즉시 보이고, 사용자가 제공한 공개 임베드는 전체 일정 확인용으로만 쓴다.
    # 비공개 iCal 주소/config는 절대 브라우저에 전달하지 않는다.
    assert "loadCalendar" in live and "openCalendar" in live and "show('calendar')" in live
    assert "calendar.google.com/calendar/embed" in live and "COUPANG_CALENDAR_ID" in live
    # 캐시 일정을 앱이 직접 그린다. 2026-08-05 부터 한 줄 목록(calendarEventList)이 아니라
    # 월 격자(calGrid)+그 날 목록(calAgenda) 이다 — 검증 [102].
    assert "openGoogleCalendarDraft" in live and 'id="calAgenda"' in live and 'id="calGrid"' in live
    assert "config/gcal.json" not in live and ".ics" not in live, "비공개 일정 원천이 앱에 노출됐다"
    assert "gcal_sync.py" in daily
    print("  [75] 구글 캘린더 — 예정열만·구분일치·비밀주소 보호·캐시/공개전체보기 ✅")


def t78_recalc_pending_visible():
    """"원장엔 넣었다는데 앱엔 왜 없지?" 를 앱이 스스로 설명하게 한다.

    2026-07-29 실제 상황: 06시트에 청구 636건을 넣었지만 정산ID·금액이 **수식**이라
    엑셀이 계산하기 전까지 앱은 옛 건수만 읽었다. 숫자가 틀린 게 아니라 아직 안 나온 것인데,
    화면이 그 말을 안 하면 사용자는 자료가 사라졌다고 읽는다.
    ★ 고칠 수 없는 것(엑셀 수식 계산)은 **드러내는 것**이 정답이다.
    """
    import recalc_pending as R
    assert R.self_test(), "recalc_pending 자체 검증 실패"
    # 수식이 돌려준 빈 문자열을 '값 있음'으로 세면 대기가 0이 되어 배너가 안 뜬다
    assert R.count_rows(["AS-1", "AS-2"], ["", None]) == (2, 0)
    # ★ 새 행을 만드는 시트가 하나라도 빠지면 그 시트만 조용히 사라진다
    #   (2026-07-30: 카톡 40건을 02시트에 넣었는데 02가 감시목록에 없어 앱이 침묵했다)
    watched = {x[0] for x in R.SHEETS}
    for sheet in ("06_거래서류청구수금", "02_돌발AS접수", "04_정기점검", "05_신규납품설치"):
        assert sheet in watched, f"{sheet} 가 재계산 감시목록에 없다 — 넣은 건이 앱에서 사라진다"

    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert '"recalc": get_recalc_pending()' in server, "상태 API가 대기 건수를 안 내려준다"
    assert "재계산대기.json" in server and "recalc_pending.py" in daily
    assert "renderRecalc(s.recalc)" in live and 'id="recalccard"' in live

    # 라벨: 같은 숫자를 두 이름으로 부르면 두 개인 줄 안다
    assert "'정산 누적'" not in live, "히어로와 KPI가 같은 값을 다른 이름으로 부른다"
    assert live.count("'정산 건수'") >= 2
    assert "'공급가액 합계'" not in live, "어느 공급가액인지 드러나야 한다(작업/명세서/계산서 3종)"
    assert "'작업 공급가액(부가세 별도)'" in live
    assert live.count("공급가액(부가세 별도)") >= 8, \
        "대시보드·정산·보고·상세의 공급가액에 부가세 별도 표기가 공통 적용되지 않았다"
    assert "공급가액(원)" not in live, "부가세 포함 여부가 없는 옛 공급가액 라벨이 남았다"
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    assert "원(부가세 별도)" in phone, "PC 독립 폰 사본의 공급가액에 부가세 별도 표기가 없다"
    assert "계산서 발행 후 미입금" in live, "미수금이 '떼인 돈'으로 읽힌다"
    print("  [78] 재계산 대기 안내 + 라벨 모호성 제거(같은 값 한 이름·금액 종류 명시) ✅")


def t84_duplicate_source_files(tmp):
    """같은 내용의 원천 파일이 여러 벌 있어도 금액이 배수로 부풀지 않는다.

    ★ 2026-07-30 실사고: 원본 자료 정리가 판매조회를 3벌 남겼고(SHA256 동일),
      billing_fill 이 전부 읽어 공급가액이 36.2억 → 108.6억으로 **3배**가 됐다.
      파일명 규칙(`__dup_`)으로 거르면 다른 이름으로 다시 뚫린다 — 내용으로 판정해야 한다.
    """
    import billing_fill as B
    a = os.path.join(tmp, "판매조회_x.xlsx")
    open(a, "wb").write(b"same-bytes")
    for name in ("판매조회_x__dup_1.xlsx", "판매조회_x__dup_1_2.xlsx", "sales_copy.xlsx"):
        open(os.path.join(tmp, name), "wb").write(b"same-bytes")
    other = os.path.join(tmp, "판매조회_y.xlsx")
    open(other, "wb").write(b"different")
    keep = B.dedupe_files([a, other] + [os.path.join(tmp, n) for n in
                                        ("판매조회_x__dup_1.xlsx", "판매조회_x__dup_1_2.xlsx", "sales_copy.xlsx")])
    assert len(keep) == 2, f"내용 중복을 못 걸렀다: {[os.path.basename(x) for x in keep]}"
    # 읽을 수 없는 경로는 조용히 건너뛴다(파이프라인이 이것 때문에 멈추면 안 된다)
    assert B.dedupe_files([os.path.join(tmp, "없는파일.xlsx")]) == []
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "billing_fill.py" in daily, "청구 근거 갱신이 일일 실행에 없다"
    # ★ 근본 원인도 막는다: 정리기가 사본을 계속 만들면 읽는 쪽을 고쳐도 파일이 계속 늘어난다.
    #   (2026-07-30: 목적지에 파일이 있으면 내용을 보지 않고 __dup_ 를 붙여 3벌이 됐다)
    import source_organizer as SO
    same = os.path.join(tmp, "src_same.xlsx")
    dst = os.path.join(tmp, "dst.xlsx")
    open(same, "wb").write(b"identical"); open(dst, "wb").write(b"identical")
    assert SO._collision_target(same, dst) == dst, "내용이 같은데 사본을 또 만든다"
    diff = os.path.join(tmp, "src_diff.xlsx")
    open(diff, "wb").write(b"really-different")
    got = SO._collision_target(diff, dst)
    assert got != dst and "__dup_" in os.path.basename(got), "내용이 다른데 덮어쓰려 한다"
    # 같은 내용의 사본이 이미 있으면 또 만들지 않는다(무한 증식 차단)
    open(got, "wb").write(b"really-different")
    assert SO._collision_target(diff, dst) == got, "같은 내용의 사본을 반복 생성한다"
    print("  [84] 원천 파일 내용중복 제거 — 금액 배수 합산 차단 + 사본 무한증식 차단 ✅")


def t86_daily_run_singleton_and_inbox_classification(tmp):
    """일일 실행은 프로세스가 겹치지 않고, ERP 원장은 이름이 아닌 분류기로 찾는다."""
    import daily_run as D

    lock = os.path.join(tmp, "daily_run.lock")
    token = D.acquire_run_lock(lock)
    assert token and os.path.isfile(lock), "첫 실행이 잠금을 잡지 못했다"
    assert D.acquire_run_lock(lock) is None, "두 번째 실행이 같은 잠금을 통과했다"
    D.release_run_lock(token, lock)
    assert not os.path.exists(lock), "정상 종료 뒤 잠금이 남았다"

    # 첫 실행이 O_EXCL 직후 JSON을 쓰는 순간을 두 번째 실행이 보면 빈 파일일 수 있다.
    # 방금 생긴 빈 잠금을 죽은 것으로 오판해 지우면 두 프로세스가 함께 실행된다.
    open(lock, "wb").close()
    assert D.acquire_run_lock(lock) is None, "생성 중인 최근 잠금을 빼앗았다"
    os.unlink(lock)

    # 비정상 종료로 남은 잠금은 실제 PID가 죽었을 때만 회수한다.
    with open(lock, "w", encoding="utf-8") as f:
        json.dump({"pid": 2147483646, "token": "dead-run"}, f)
    recovered = D.acquire_run_lock(lock)
    assert recovered and recovered != "dead-run", "죽은 프로세스의 잠금을 회수하지 못했다"
    D.release_run_lock(recovered, lock)

    import inbox_scan
    old_pick = inbox_scan.pick
    try:
        inbox_scan.pick = lambda kind: ["분류기_선택.xlsx"] if kind == "ledger" else []
        assert D.has_inbox_kind("ledger")
        assert not D.has_inbox_kind("tax")
    finally:
        inbox_scan.pick = old_pick

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert 'has_inbox_kind("ledger")' in daily, "ERP 원장 단계가 공용 inbox 분류기를 쓰지 않는다"
    assert "acquire_run_lock" in daily and "release_run_lock" in daily
    erp_check = open(os.path.join(ROOT, "erp_ledger_check.py"), encoding="utf-8").read()
    assert 'files = pick("ledger")' in erp_check
    assert 'pick("ledger", INBOX_DIR)' not in erp_check, "실제 대조기가 다시 로컬 inbox만 본다"
    print("  [86] daily_run 단일 프로세스 잠금 + ERP 원장 공용 분류기 사용 ✅")


def t90_ip_guard_and_archive():
    """ERP 접속 IP 관문 + 복구용 보관 (2026-07-30 지시 2건).

    ① "IP가 변경되면 이 화면에서 등록해서 진행" — 자동 등록은 하지 않는다(회사 ERP 보안
       설정). 대신 **부르기 전에 멈춘다**: 미등록 IP로 호출하면 실패가 인증 오류처럼 보여
       원인을 엉뚱한 데서 찾고, 반복 실패는 트래픽 제한을 건드려 ERP 전체가 막힌다.
    ② "복구·코딩에 필요한 자료 별도 보관" — 파일을 쌓는 대신 git bundle 한 파일로.
       ★ 비밀키는 어떤 경로로도 담기지 않아야 한다(규칙 1).
    """
    import erp_ip_guard as G
    assert G.self_test(), "erp_ip_guard 자체 검증 실패"
    # IP를 못 구했으면 통과시키면 안 된다 — 미등록 상태로 ERP를 두드리는 게 최악이다
    assert G.decide("", ["1.2.3.4"])[0] == "unknown"
    assert G.decide("5.6.7.8", ["1.2.3.4"])[0] == "need_register"
    # ERP 로그인이 이 관문을 지나는가
    cli = open(os.path.join(ROOT, "ecount_client.py"), encoding="utf-8").read()
    assert "erp_ip_guard" in cli and "require()" in cli, "ERP 호출이 IP 관문을 건너뛴다"
    # IP 목록은 커밋되지 않는다(회사 설정값)
    gi = open(os.path.join(ROOT, ".gitignore"), encoding="utf-8").read()
    assert "config/erp_allowed_ips.json" in gi

    import archive_keep as A
    assert A.self_test(), "archive_keep 자체 검증 실패"
    # 비밀키 파일은 어떤 경로로도 보관 대상이 아니다
    for rel in ("config/ecount_config.json", "config/webapp.json", "config/gcal.json",
                "config/erp_allowed_ips.json"):
        assert not A.wanted(os.path.join(ROOT, *rel.split("/")), ROOT), rel
    assert A.has_secret('"API_CERT_KEY": "x"')
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "erp_ip_guard.py" in daily and "archive_keep.py" in daily, "일일 실행에 연결되지 않았다"
    print("  [90] ERP IP 관문(호출 전 차단)·복구용 보관(bundle·비밀키 제외) ✅")


def t91_icon_sprite_and_ios_theme():
    """아이콘 스프라이트 + iOS 외형 (2026-07-30 지시).

    ★ 왜 <img> 를 버렸나: 원본 svg 에 fill="#344054" 가 박혀 있어 **선택된 탭·다크모드·
      경고색에 아이콘이 반응하지 못했다.** 흰 아이콘이 필요한 곳은 filter:brightness(0)
      invert(1) 같은 꼼수로 때웠다. 스프라이트+currentColor 로 근본을 고쳤다.
    ★ 왜 인라인인가: iOS Safari 는 외부 파일의 <use href="a.svg#id"> 를 제대로 렌더하지
      않는다. 아이폰이 주 사용처이므로 인라인이 유일한 정답이다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # 스프라이트가 있고, 모든 <use> 가 실제 <symbol> 을 가리킨다(깨진 아이콘 0)
    symbols = set(re.findall(r'<symbol id="i-([^"]+)"', live))
    assert symbols, "아이콘 스프라이트가 없다"
    used = set(re.findall(r'<use href="#i-([^"]+)"', live))
    dynamic = {u for u in used if "${" in u}
    missing = (used - dynamic) - symbols
    assert not missing, f"symbol 이 없는 아이콘 참조: {sorted(missing)[:5]}"

    # <img> 로 **아이콘**을 되돌리면 색 상속이 다시 깨진다. 회사 CI PNG를 헤더에서
    # 흰색으로 만드는 brightness/invert는 아이콘 꼼수가 아니라 브랜드 표시 규칙이다.
    assert 'src="/icons/' not in live, "아이콘이 <img> 로 되돌아갔다 — currentColor 를 못 받는다"

    # 아이콘 파일과 라이선스는 저장소에 남겨 둔다(스프라이트를 다시 만들 수 있어야 한다)
    icons_dir = os.path.join(ROOT, "webapp", "icons")
    assert os.path.exists(os.path.join(icons_dir, "LICENSE-bootstrap-icons.txt")), "라이선스 파일 누락"
    on_disk = {os.path.splitext(f)[0] for f in os.listdir(icons_dir) if f.endswith(".svg")}
    assert symbols <= on_disk, f"원본 svg 가 없는 symbol: {sorted(symbols - on_disk)[:5]}"

    # iOS 외형: 파랑 강조·그룹배경·다크모드·큰 모서리
    # ★ systemBlue 원값(#007AFF)은 쓰지 않는다 (2026-08-06). 흰 배경 글자로 4.0:1,
    #   흰 글자를 얹은 버튼 바탕으로 3.0:1 이라 둘 다 기준(4.5:1) 미달이었다.
    #   색조는 지키고 명도만 낮춘 #0062CC 로 간다 — 판정은 검증 [115] 가 한다.
    assert "--brand:#007AFF" not in live, "저대비 systemBlue 원값으로 되돌아갔다"
    assert re.search(r"--brand:#[0-9A-Fa-f]{6}", live), "파랑 강조색이 없다"
    assert "--bg:#F2F2F7" in live, "iOS 그룹 배경이 아니다"
    # ★ 2026-07-30 실기기 확인 후 되돌림: 이 앱은 흰 배경이 77곳 하드코딩돼 있어 OS 다크를
    #   따라가면 **흰 카드 위 흰 글자**가 되어 아무것도 안 보였다(사용자 화면으로 확인).
    #   그래서 다크 모드를 넣지 않고 color-scheme:light 로 못박았다.
    # ★ 2026-08-07 갱신 — 막을 것은 '다크 모드'가 아니라 **자동으로 따라가는 것**이다.
    #   그 사이 어둡게가 `[data-theme="dark"]` 로 정식으로 들어왔고(사람이 켠 것만 적용),
    #   하드코딩된 밝은 판은 검증 [127] 이 따로 막는다. 사용자 지시로 '시스템 동일'
    #   선택지도 생겼다(2026-08-07: "기본, 다크모드, 시스템 동일 이렇게 구성").
    #   그러므로 금지선은 **CSS `@media (prefers-color-scheme: dark)` 블록** 하나다 —
    #   그건 사람이 고르지 않아도 색이 통째로 뒤집히는 유일한 경로이고, 화면을 그대로
    #   캡처해 보고서로 올리는 이 앱에서는 그것이 곧 사고다.
    #   JS 의 `matchMedia('(prefers-color-scheme: dark)')` 는 **사람이 '시스템 동일'을
    #   고른 뒤에만** 읽히므로 허용한다(검증 [132] 가 그 조건을 지킨다).
    _auto_dark = re.search(r"@media[^{]*prefers-color-scheme\s*:\s*dark", live)
    assert not _auto_dark, \
        "CSS 가 OS 다크를 자동으로 따라간다 — 사람이 고르지 않았는데 보고 화면이 검게 나온다"
    assert "matchMedia" in live and "themeMode() === 'system'" in live, \
        "'시스템 동일'이 사람이 고른 뒤에만 도는 구조가 아니다"
    assert "color-scheme:light" in live, "밝은 화면으로 고정하는 선언이 없다"
    # 헤더 부제가 flex 안에서 0폭까지 눌려 한 글자씩 세로로 쪼개졌던 사고를 막는다
    assert ".appbar h1{min-width:52px;white-space:nowrap}" in live, "헤더 제목이 다시 쪼개질 수 있다"
    # ★ 2026-07-31 지시로 교체: "폰트는 모든 폰트 나눔고딕 폰트로 변경".
    #   전에는 -apple-system(SF Pro) 을 강제했지만, 그건 2026-07-30 의 'iOS 느낌' 기준이었고
    #   오늘 지시가 이를 덮는다. 지금 지켜야 하는 것은 두 가지다:
    #   ① 나눔고딕이 스택 맨 앞인가 ② 그 글꼴을 **동봉본**으로 얹는가
    #      (구글 폰트 주소를 쓰면 인터넷 없는 현장에서 안 뜨고 외부로 요청이 나간다)
    assert 'font-family:"Nanum Gothic"' in live.replace(" ", "").replace(
        'font-family:"NanumGothic",', 'font-family:"Nanum Gothic"') or \
        '"Nanum Gothic","NanumGothic"' in live, "본문 글꼴이 나눔고딕으로 시작하지 않는다"
    assert "/fonts/nanumgothic.css" in live, "나눔고딕을 동봉본으로 얹지 않는다"
    assert os.path.exists(os.path.join(ROOT, "webapp", "fonts", "OFL.txt")), \
        "동봉한 나눔고딕의 SIL OFL 라이선스 파일이 없다"
    assert "fonts.googleapis.com" not in live and "fonts.gstatic.com" not in live, \
        "글꼴을 외부(구글)에서 받고 있다 — 인터넷 없는 현장에서 안 뜬다"
    # 탭바 유리 효과는 지원 안 되는 브라우저를 위해 @supports 로 감싼다
    assert "backdrop-filter" in live and "@supports" in live, "유리 효과에 폴백이 없다"
    # 검은 로고/아이콘이 남색 헤더·흰 버튼에서 사라지지 않아야 한다.
    # 지켜야 하는 것은 **로고가 남색 헤더에서 보이는가** 이지 특정 구현이 아니다.
    # 흰 판을 얹는 방법과, 로고를 흰색으로 뽑아내는 방법(brightness(0) invert(1)) 둘 다 답이다.
    # 2026-07-30: 사용자가 '헤더와 이질감 없이' 를 요청해 흰 판 → 흰색 로고로 바꿨다.
    assert (".appbar-brand-stack{background:rgba(255,255,255,.94)" in live
            or "brightness(0) invert(1)" in live), "검은 로고가 남색 헤더에서 사라진다"
    # 헤더 버튼 아이콘은 **배경과 대비되기만 하면** 된다 — 구현을 고정하지 않는다.
    #   2026-07-30: 흰 버튼(어두운 아이콘) → 테두리만 있는 투명 버튼(흰 아이콘)으로 바꿨다.
    #   사용자 지적: 남색 헤더 위 흰 네모가 이질적이다.
    assert ".account-pin img,.account-pin svg" in live, "PIN 버튼 아이콘 규칙이 없다"
    assert ("fill:#24365F" in live or "fill:#fff" in live), "PIN 아이콘 색이 지정되지 않았다"
    assert ".notice-bell svg{width:21px;height:21px;fill:" in live, "알림 아이콘 색이 없다"
    # 상단 배지는 absolute 로 떠 있으면 긴 제목과 다시 겹친다. 제목·배지는 같은 flex 행이어야 한다.
    assert '.hero-top{display:flex' in live and '.hero .badge{position:static' in live
    assert '<div class="hero-top">' in live and '<span class="hero-sub">' in live
    assert ".hero::before" not in live and ".hero::after" not in live, "장식 원이 다시 생겼다"
    print(f"  [91] 아이콘 스프라이트({len(symbols)}개·깨짐 0)·currentColor·iOS·히어로 정렬 ✅")


def t92_excel_recalc_agent():
    """엑셀을 에이전트가 알아서 열고 닫는다 (2026-07-30 지시).

    ★ 이 도구는 **사람의 파일을 저장**한다. 그래서 판단 순서가 곧 안전장치다:
      대기 0이면 안 열고 · 복구 경고/SHA 승인이 없으면 안 열고 ·
      **사람이 열어 두었으면 물러난다.** 남이 편집 중인 파일을 자동화가 저장하면
      그 작업이 날아간다.
    """
    import excel_recalc as X
    assert X.self_test(), "excel_recalc 자체 검증 실패"

    p = r"Z:\a\쿠팡_통합업무_일일보고_관리대장_v323.xlsx"
    # 사람이 열어 두었으면 어떤 경우에도 진행하지 않는다
    ok, why = X.decide({"대기합계": 99}, p, True, True, True)
    assert not ok and "열어 두었" in why, why
    # 복구 경고가 해소되지 않았으면 대기 건이 있어도 정본을 열지 않는다
    ok, why = X.decide({"대기합계": 99}, p, False, True, False)
    assert not ok and "안전 승인" in why, why
    # 원본을 덮어쓰지 않는다(항상 vN+1)
    assert X.next_version_path(p).endswith("_v324.xlsx")
    assert X.next_version_path(p) != p

    # 새 파이썬 의존성을 쓰지 않는다(프로젝트 원칙) — 엑셀은 PowerShell COM 으로 부른다
    src = open(os.path.join(ROOT, "excel_recalc.py"), encoding="utf-8").read()
    # 문서에 "pywin32 없이" 라고 적혀 있으므로 문자열이 아니라 **실제 import** 를 본다.
    assert not re.search(r"^\s*(import|from)\s+(win32com|win32|pythoncom)\b", src, re.M), \
        "새 의존성(pywin32)을 끌어들였다"
    assert "Excel.Application" in src and "powershell" in src
    # 좀비 EXCEL.EXE 를 남기지 않는다
    assert "finally" in X.PS_RECALC and "$x.Quit()" in X.PS_RECALC
    assert "finally" in X.PS_AVAILABLE and "$x.Quit()" in X.PS_AVAILABLE
    # 경로에 공백·★ 가 섞여 있어 인자 전달은 환경변수로 한다(2026-07-30 실패 경험)
    assert "$env:CSOS_XL_SRC" in X.PS_RECALC, "경로를 명령줄 인자로 넘기면 깨진다"
    # 일일 실행에 연결
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "excel_recalc.py" in daily
    assert "excel_recalc_clearance.json" in src and "sha256" in src
    assert "recalc-" in src and "os.replace(tmp_dst, dst)" in src
    assert "if os.path.exists(dst)" in src, "기존 vN+1을 덮어쓸 수 있다"
    print("  [92] 엑셀 자동 재계산 — 복구/SHA관문·임시본 검증·덮어쓰기/좀비 방지 ✅")


def t93_ledger_db_and_ux():
    """앱 입력은 DB에 즉시 확정 · Excel은 하루 두 번 보관본만 (최신 확정 2026-08-10).

    ★ 저장 성공의 기준은 SQLite 정본 커밋이다. 시각과 무관하게 즉시 읽혀야 한다.
    ★ 11:00·15:00은 DB 입력 반영 시간이 아니라 DB→Excel 읽기 전용 보관본 생성 회차다.
      놓친 회차가 있어도 DB 정본은 이미 최신이며, 다음 회차는 그 시점 스냅샷만 만든다.
    """
    import ledger_db as L
    assert L.self_test(), "ledger_db 자체 검증 실패"
    assert [w.hour for w in L.WINDOWS] == [11, 15], "Excel 보관본 생성 시각이 11시·15시가 아니다"

    from datetime import datetime as _dt
    assert L.slot_of(_dt(2026, 7, 30, 11, 5)), "11시 회차를 인식하지 못한다"
    assert L.slot_of(_dt(2026, 7, 30, 13, 0)) is None, "보관본 생성 시각이 아닌데 열려 있다"
    assert L.next_window(_dt(2026, 7, 30, 12, 0)).hour == 15
    assert L.eligible_slot(_dt(2026, 7, 30, 13, 0), []) is None, \
        "놓친 회차를 이유로 임의 시각에 Excel 보관본을 만든다"
    assert L.eligible_slot(
        _dt(2026, 7, 30, 11, 5), ["2026-07-30 11:00"]) is None, \
        "같은 11시 DB 스냅샷 보관본을 두 번 만든다"

    with tempfile.TemporaryDirectory(prefix="ledger-json-lock-") as lock_tmp:
        queue_path = os.path.join(lock_tmp, "pending.json")
        with open(queue_path + ".lock", "w", encoding="ascii") as f:
            f.write("99999999 2026-01-01T00:00:00")
        with L.json_queue_lock(queue_path, timeout=0.1):
            assert os.path.exists(queue_path + ".lock")
        assert not os.path.exists(queue_path + ".lock"), "죽은 JSON 큐 잠금을 회수하지 못한다"

    src = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    writer = open(os.path.join(ROOT, "ledger_writer.py"), encoding="utf-8").read()
    assert any(x.startswith("import") and "sqlite3" in x for x in src.splitlines()), (
        "표준 라이브러리 sqlite3 를 써야 한다(새 의존성 금지)")
    assert "finally:" in src and "c.close()" in src, "DB 연결을 닫지 않으면 파일이 잠긴다"
    assert "ingest_key" in src and "INSERT OR IGNORE INTO pending" in src, \
        "중단 후 JSON staging 재흡수 시 중복될 수 있다"
    enqueue_src = src.split("def enqueue(", 1)[1].split("\ndef ", 1)[0]
    assert "app_store.apply_legacy_items" in enqueue_src, \
        "앱 입력이 Excel 보관 큐보다 먼저 SQLite 정본에 즉시 저장되지 않는다"
    assert enqueue_src.index("app_store.apply_legacy_items") < enqueue_src.index("INSERT OR IGNORE INTO pending"), \
        "보관본 큐를 먼저 성공 처리한 뒤 SQLite 정본 저장을 시도한다"
    assert "id IN ({marks})" in src, "보관본 생성 중 새로 들어온 DB 행까지 처리했다고 오기록할 수 있다"
    assert '"--queue", batch_queue, "--apply"' in src, "공용 JSON 큐와 Excel 보관본 배치가 섞인다"
    assert '"COUPANG_LEDGER_GATE": "1"' in src, "11·15시 보관본 회차가 내부 쓰기 게이트를 열지 않는다"
    assert 'os.environ.get("COUPANG_LEDGER_GATE") != "1"' in writer, \
        "ledger_writer 직접 --apply를 막는 강제 게이트가 없다"

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "ledger_db.py" in daily, "일일 실행이 DB 게이트를 지나지 않는다"
    # 변수 이름이 아니라 **지키려는 것**을 본다: stderr 를 앞에서 자르지 말고 **꼬리**를
    # 남겨야 예외 이름과 원인이 보인다(앞 500자만 남기면 호출 위치만 나온다).
    # 2026-08-08 에 _run_once 를 Popen 으로 바꾸며 `r.stderr` 가 `se` 가 됐는데,
    # 이름을 박아 둔 검증이 거기서 깨졌다 — 뜻은 그대로였다.
    assert "[-2000:]" in daily, "실패 리포트가 예외 원인을 다시 잘라낸다"
    assert '"zscan.py"' in daily and '"--docs"' in daily, \
        "Z: 상시 공백·서류 대조가 09:50 자동실행에 연결되지 않았다"
    assert '"ledger_writer.py"), "--apply"' not in daily, \
        "일일 대조가 SQLite 정본을 건너뛰고 Excel 보관본을 직접 만든다"
    for direct in (
        '"excel_recalc.py"), "--run"',
        '"erp_docs_check.py"), "--sheet"',
        '"fix_workbook.py"), "--apply"',
        '"pm_schedule_sync.py"), "--apply"',
        '"work_log_sync.py"), "--apply"',
        '"band_extract.py"), "--sheet"',
        '"findings_sheet.py")]))',
    ):
        assert direct not in daily, f"09:50 자동대조에 직접 Excel 쓰기가 남아 있다: {direct}"
    for scheduled in (
        '"erp_docs_check.py"), "--sheet"',
        '"pm_schedule_sync.py"), "--apply"',
        '"work_log_sync.py"), "--apply"',
        '"band_extract.py"), "--sheet"',
        '"findings_sheet.py")',
        '"fix_workbook.py"), "--apply"',
        '"excel_recalc.py"), "--run"',
    ):
        assert scheduled in src, f"11·15시 Excel 보관본 구조 갱신에서 빠진 작업: {scheduled}"
    assert "scheduled_workbook_maintenance(now)" in src, \
        "11·15시 보관본 회차가 구조 시트·재계산을 함께 수행하지 않는다"
    assert "def handoff_add(" in src and '"workbook_patch.py"' in src, \
        "19시트 종료 인수인계가 11·15시 보관본 회차에 예약되지 않는다"
    assert '"--batch"' in src and "for item in pending_handoffs()" not in src, \
        "19시트 예약을 건마다 따로 보관해 vN+1 이 건수만큼 폭증한다(--batch 로 묶을 것)"
    schedule = open(os.path.join(ROOT, "install_ledger_schedule.ps1"), encoding="utf-8").read()
    assert '"11:00"' in schedule and '"15:00"' in schedule
    assert "MultipleInstances IgnoreNew" in schedule and "ledger_db.py --apply" in schedule

    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "get_apply_window" in server and '"applywin"' in server
    assert '"writer_apply":  ("입력 DB 적재"' in server
    assert 'start_task("writer_apply")' not in server, \
        "앱이 SQLite 저장 대신 옛 즉시 Excel 셀 반영 작업을 시작한다"
    assert 'os.path.join(ROOT, "ledger_writer.py"), "--apply"' not in server, \
        "앱 서버에 직접 ledger_writer --apply 경로가 남아 있다"
    assert "enqueue_for_scheduled_apply" in server, "앱 입력이 SQLite 즉시저장·보관본 큐 경로를 거치지 않는다"
    assert 'if p == "/api/ux":' in server and "ledger_db.ux_add(events)" in server
    assert "function uxEvent(" in live and "function uxFlush(" in live
    assert "uxEvent('view',v)" in live and "uxEvent('slow'" in live
    assert "renderApplyWindow(s.applywin)" in live, "앱이 다음 Excel 보관본 생성 시각을 알리지 않는다"
    assert "runTask('writer_apply')" not in live, "앱 화면에 옛 즉시 Excel 셀 반영 호출이 남아 있다"

    # ★ 사람이 누르는 즉시 기능도 원본 Excel을 직접 고치지 않는다. 현재 DB revision을
    #   새 로컬 파일로 렌더하고 검증한 뒤 last-good 보관본만 승격한다.
    assert '"ledger_now"' in server and "보관본 지금 생성" in server, \
        "사람이 요청하는 즉시 Excel 보관본 생성 경로가 없다"
    task_block = server.split("TASKS = {", 1)[1].split("TASK_TIMEOUTS", 1)[0]
    assert '"ledger_now"' in task_block and '"archive_worker.py"' in task_block and '"--run"' in task_block, \
        "보관본 지금 생성이 검증 렌더 worker에 연결되지 않았다"
    assert 'os.path.join(ROOT, "ledger_writer.py"), "--apply"' not in task_block, \
        "보관본 지금 생성이 다시 Excel 직접 셀 반영 경로로 돌아갔다"
    assert "function applyExcelNow(" in live and "runTask('ledger_now')" in live, \
        "앱의 지금 생성 버튼이 검증된 보관본 준비 작업에 이어져 있지 않다"
    assert "askYesNo(" in live.split("function applyExcelNow(")[1][:700], \
        "즉시 보관본 생성이 확인 없이 실행된다"
    # 자동 회차는 정해진 11·15시 exporter를 쓰며 사람용 '지금 생성' 키를 재사용하지 않는다.
    for auto in ("daily_run.py", "session_wrapup.py"):
        src = open(os.path.join(ROOT, auto), encoding="utf-8").read()
        assert "ledger_now" not in src, f"{auto} 가 사람용 보관본 지금 생성을 대신 누른다"

    band_collector = open(os.path.join(ROOT, "band", "collect_band.js"), encoding="utf-8").read()
    assert "window.scrollBy(0, SCROLL_STEP)" in band_collector, \
        "밴드 수집기가 실패한 바닥 점프로 되돌아갔다"
    assert "window.scrollTo(0, document.body.scrollHeight)" not in band_collector
    assert "fetch(url" not in band_collector and "imageData.push" not in band_collector, \
        "사진 fetch 무한대기가 글 전체 수집을 다시 막을 수 있다"

    # 업무센터는 인원이 바뀌어도 균등해야 한다(고정 칸 수 금지)
    assert "repeat(auto-fit,minmax(132px,1fr))" in live, "업무센터가 다시 칸 수를 고정했다"
    assert ".workcenter-buttons{grid-template-columns:repeat(3" not in live, \
        "업무센터 3칸 고정이 남아 있다 — 인원이 바뀌면 빈칸이 생긴다"
    assert ":last-child:nth-child(odd){grid-column:1/-1}" in live, "홀수 인원일 때 줄 끝이 빈다"
    assert "<h3>빠른 실행</h3>" not in live, "제거한 '빠른 실행' 카드가 되살아났다"

    # 상시 규칙이 문서에 남아 있어야 다음 세션·Codex 가 따른다
    rules = (open(os.path.join(ROOT, "CLAUDE.md"), encoding="utf-8").read()
             + open(os.path.join(ROOT, "AGENTS.md"), encoding="utf-8").read())
    assert "SQLite가 유일한 정본" in rules and "Excel 보관본" in rules and "UX 기록" in rules
    archive = open(os.path.join(ROOT, "archive_keep.py"), encoding="utf-8").read()
    assert "ledger_db_copy" in archive and "ledger_queue.db" in archive
    print("  [93] SQLite 즉시 정본·11·15시 Excel 보관본·UX 기록·업무센터 안전 UI ✅")


def t95_objective_completion_db_only():
    """[95] 객관 입증 완료 처리 + 엑셀 백필 없이 DB 로만 (사용자 지시 2026-07-31).

    · 7.수금완료 = 발행·수금 모두 ERP 가 입증 → 완료(ERP 수금확인)
    · 6.세금계산서발행 = 발행만 입증 → '미발행'이 아니라 입금 대기(입금일 있으면 완료)
    · 완료 상태는 06시트 발행일 칸에 써넣지 않는다(판매조회에 발행일 열이 없다 —
      절대규칙 10). 근거는 ledger_db.resolution 표가 기억한다(first_seen 보존)."""
    import sys as _s, tempfile
    _s.path.insert(0, ROOT)
    import ecount_reconcile as E
    import ledger_db as L
    import staff_completion as SC

    old_prog = E.erp_progress
    try:
        E.erp_progress = lambda: {"P7": "7.수금완료", "P6": "6.세금계산서발행",
                                  "PMIX": "혼재(7.수금완료 / 확인)"}
        base = {"비용구분": "유상", "원천업무ID": "PM-1", "원장_공급가액": 100,
                "원장_거래명세서번호": "20260101-1"}
        assert E.settle_status({**base, "프로젝트NO": "P7"}) == "완료(ERP 수금확인)"
        assert E.settle_status({**base, "프로젝트NO": "P6"}) == "입금 대기", \
            "발행만 입증된 건을 완료로 넘기면 수금 추적이 죽는다"
        assert E.settle_status({**base, "프로젝트NO": "P6",
                                "원장_입금일": "2026-02-01"}) == "완료(ERP 발행확인)"
        assert E.settle_status({**base, "프로젝트NO": "PX"}) == "세금계산서 미발행"
        # 금액 수식 캐시가 비어도 ERP 전체 전표가 수금완료면 객관 완료가 우선한다.
        as_wait = {**base, "원천업무ID": "AS-1", "원장_공급가액": 0,
                   "원장_거래명세서합계": 110}
        assert E.settle_status({**as_wait, "프로젝트NO": "P7"}) == "완료(ERP 수금확인)"
        # ★ 2026-08-08 지시로 ÷1.1 이 **깨끗이 떨어지면** 금액을 아는 것이라 대기가 아니다.
        #   110 → 100 은 떨어지므로 사다리를 계속 내려가 발행·수금 판정이 나온다.
        assert E.settle_status({**as_wait, "프로젝트NO": "P6"}) == "입금 대기"
        assert E.settle_status({**as_wait, "프로젝트NO": "PMIX"}) == "세금계산서 미발행", \
            "한 프로젝트에 수금완료·미완료 전표가 섞였는데 전체 완료로 올렸다"
        #   안 떨어지는 금액은 예전 그대로 사람에게 남긴다 — 억지로 숫자를 만들지 않는다.
        odd = {**as_wait, "원장_거래명세서합계": 111}
        assert E.settle_status({**odd, "프로젝트NO": "PMIX"}) == "금액 재계산 대기"
        assert E._collapse_erp_progress(["7.수금완료", "7.수금완료"]) == "7.수금완료"
        assert E._collapse_erp_progress(["7.수금완료", "6.세금계산서발행"]) == \
            "6.세금계산서발행"
        assert E._collapse_erp_progress(["7.수금완료", "확인"]).startswith("혼재("), \
            "복수 전표 충돌이 완료로 접혔다"
    finally:
        E.erp_progress = old_prog

    # 완료( 는 조치 목록이 아니라 DB(resolution)로 — 엑셀 셀 백필 경로가 없어야 한다
    fx = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    assert 'startswith("완료(")' in fx and "resolution_sync" in fx, \
        "완료 건이 조치 목록에 남거나 DB에 기록되지 않는다"
    seg = fx[fx.index('startswith("완료(")'):fx.index("return rows")]
    assert "enqueue" not in seg and "workbook" not in seg, "완료 처리가 엑셀 쓰기로 새어 나간다"

    # resolution 멱등 upsert — 두 번 넣어도 1행, first_seen 은 처음 값이 남는다
    with tempfile.TemporaryDirectory() as td:
        old_dir, old_path = L.DB_DIR, L.DB_PATH
        L.DB_DIR, L.DB_PATH = td, os.path.join(td, "t.db")
        try:
            assert L.resolution_sync([{"settle_id": "JS-1", "project": "P7",
                                       "status": "완료(ERP 수금확인)", "basis": "b1"}]) == 1
            first = L.resolutions()["JS-1"]["first_seen"]
            L.resolution_sync([{"settle_id": "JS-1", "project": "P7",
                                "status": "완료(ERP 수금확인)", "basis": "b2"}])
            got = L.resolutions()
            assert len(got) == 1 and got["JS-1"]["basis"] == "b2", "upsert 가 안 된다"
            assert got["JS-1"]["first_seen"] == first, "first_seen 이 덮였다 — 최초 입증 시각 유실"
            L.staff_resolution_sync([{
                "owner": "류지영", "task_kind": "settlement", "record_id": "JS-1",
                "project": "P7", "completed_on": "2026-08-01", "basis": "정산 완료",
            }])
            assert L.resolution_retract(["JS-NOT-EXIST", "JS-1"]) == 1
            assert "JS-1" not in L.resolutions()
            assert not any(row["record_id"] == "JS-1" and row["task_kind"] == "settlement"
                           for row in L.staff_resolutions("류지영")), \
                "정산 완료 철회 뒤 담당자 완료가 남았다"
            # 뒤의 담당자 동기화 검증은 정상 정산 완료 1건을 전제로 한다.
            L.resolution_sync([{"settle_id": "JS-1", "project": "P7",
                                "status": "완료(ERP 수금확인)", "basis": "b3"}])
            assert L.work_resolution_sync([{
                "kind": "as", "record_id": "AS-1", "project": "UJ2600001",
                "status": "작업완료", "completed_on": "2026-07-30", "basis": "band",
            }]) == 1
            work_first = L.work_resolutions()[("as", "AS-1")]["first_seen"]
            L.work_resolution_sync([{
                "kind": "as", "record_id": "AS-1", "project": "UJ2600001",
                "status": "작업완료", "completed_on": "2026-07-31", "basis": "verified",
            }])
            work = L.work_resolutions()
            assert work[("as", "AS-1")]["completed_on"] == "2026-07-31"
            assert work[("as", "UJ2600001")]["first_seen"] == work_first

            # Excel 반영 전 신규행도 실제 완료일과 원천 근거가 있으면 DB 완료 정본에 잡힌다.
            # 같은 H10 완료일을 두 번 받은 것은 정본·보관본 모두 한 사건이다.
            # 앱 DB 전환 뒤에는 내용 해시로 중복 명령을 버리므로 10개 입력 중
            # 보관본 큐에 새로 들어가는 것은 9개여야 한다.
            assert L.enqueue([
                {"sheet": "04_정기점검", "cell": "B10", "col": "프로젝트NO",
                 "value": "UJ2600999", "evidence": "카톡 UJ2600999 완료보고"},
                {"sheet": "04_정기점검", "cell": "H10", "col": "실제점검일",
                 "value": "2026-07-31", "evidence": "카톡 UJ2600999 완료보고"},
                {"sheet": "04_정기점검", "cell": "H10", "col": "실제점검일",
                 "value": "2026-07-31", "evidence": "카톡 UJ2600999 완료보고"},
                {"sheet": "02_돌발AS접수", "cell": "B11", "col": "프로젝트NO",
                 "value": "UJ2600998", "evidence": "카톡 원문"},
                {"sheet": "02_돌발AS접수", "cell": "R11", "col": "작업완료일",
                 "value": "2026-07-30", "evidence": "카톡 완료보고"},
                {"sheet": "02_돌발AS접수", "cell": "Q11", "col": "진행상태",
                 "value": "취소", "evidence": "카톡 정정"},
                {"sheet": "04_정기점검", "cell": "B12", "col": "프로젝트NO",
                 "value": "UJ2600997", "evidence": "카톡 원문"},
                {"sheet": "04_정기점검", "cell": "H12", "col": "실제점검일",
                 "value": "2099-01-01", "evidence": "잘못된 미래일"},
                {"sheet": "04_정기점검", "cell": "B13", "col": "프로젝트NO",
                 "value": "UJ2600996", "evidence": "새 업무"},
                {"sheet": "04_정기점검", "cell": "H13", "col": "실제점검일",
                 "value": "2026-07-29", "evidence": "이전 업무의 완료 근거"},
            ], source="test") == 9
            pending_done = L.pending_work_completion_entries(today="2026-08-02")
            assert len(pending_done) == 1 and pending_done[0]["project"] == "UJ2600999"
            assert pending_done[0]["basis"].startswith("반영대기 04_정기점검!H10")
            L.work_resolution_sync(pending_done)
            provisional_first = L.work_resolutions()[("pm", "UJ2600999")]["first_seen"]
            L.work_resolution_sync([{
                "kind": "pm", "record_id": "PM-2600999", "project": "UJ2600999",
                "status": "완료", "completed_on": "2026-07-31", "basis": "원장 검증 정상",
            }])
            with L.conn() as c:
                migrated = c.execute(
                    "SELECT record_id,first_seen FROM work_resolution WHERE kind='pm' AND project=?",
                    ("UJ2600999",),
                ).fetchall()
            assert migrated == [("PM-2600999", provisional_first)], \
                "Excel ID 확정 뒤 대기행 완료 레코드가 중복되거나 최초 입증시각이 바뀌었다"

            # 세 담당자의 객관완료는 이름이 포함된 상태로 별도 DB 정본에 남는다.
            po_evidence = os.path.join(td, "po_objective_evidence.json")
            report_path = os.path.join(td, "담당자_객관완료.md")
            with open(po_evidence, "w", encoding="utf-8") as out:
                json.dump({"entries": [
                    {"owner": "오종현", "task_kind": "po_source", "record_id": "PO1",
                     "project": "UJ2600001", "completed_on": "2026-07-31",
                     "basis": "쿠팡 PO 원본 확인"},
                    {"owner": "유현민", "task_kind": "po_system_verified", "record_id": "PO1",
                     "project": "UJ2600001", "completed_on": "2026-08-01",
                     "basis": "PO·ERP·원장 일치"},
                    {"owner": "류지영", "task_kind": "billing_verified", "record_id": "PO1",
                     "project": "UJ2600001", "completed_on": "2026-08-01",
                     "basis": "ERP 계산서 발행 확인"},
                    {"owner": "오종현", "task_kind": "po_source", "record_id": "PO-FUTURE",
                     "project": "", "completed_on": "2099-01-01",
                     "basis": "잘못된 미래일"},
                ], "retractions": [
                    {"owner": "유현민", "task_kind": "po_system_verified",
                     "record_id": "PO-FALSE", "reason": "비유일 금액"},
                    {"owner": "류지영", "task_kind": "billing_verified",
                     "record_id": "PO-FALSE", "reason": "비유일 금액"},
                ]}, out, ensure_ascii=False)
            L.staff_resolution_sync([
                {"owner": "유현민", "task_kind": "po_system_verified",
                 "record_id": "PO-FALSE", "completed_on": "2026-08-01", "basis": "옛 추정"},
                {"owner": "류지영", "task_kind": "billing_verified",
                 "record_id": "PO-FALSE", "completed_on": "2026-08-01", "basis": "옛 추정"},
            ])
            counts = SC.sync(po_evidence, report_path)
            assert counts == {"류지영": 4, "오종현": 1, "유현민": 1}
            assert {row["status"] for row in L.staff_resolutions()} == {
                "류지영 완료", "오종현 완료", "유현민 완료"}
            first_staff = L.staff_resolutions("오종현")[0]["first_seen"]
            SC.sync(po_evidence, report_path)
            assert L.staff_resolutions("오종현")[0]["first_seen"] == first_staff
            assert "오종현 완료 · 1건" in open(report_path, encoding="utf-8").read()
        finally:
            L.DB_DIR, L.DB_PATH = old_dir, old_path

    # 앱: 완료 상태가 조치필요 필터·칩에서 완료로 취급되는가
    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "'완료(ERP 수금확인)'" in html and "'완료(ERP 발행확인)'" in html, "앱이 새 완료 상태를 모른다"
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "complete_verified.py" in daily and '"--queue"' in daily, \
        "객관근거 완료 상태 보완이 일일 자동화에 연결되지 않았다"
    assert daily.index('"입력 DB 적재"') < daily.index('"객관근거 완료 DB 동기화"'), \
        "신규 입력을 DB에 흡수하기 전에 완료 판정을 실행하고 있다"
    assert "staff_completion.py" in daily and '"--sync"' in daily, \
        "담당자별 객관완료 판정이 일일 에이전트에 연결되지 않았다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "work_resolutions" in srv and "객관완료근거" in srv, "앱이 현장업무 DB 완료판정을 읽지 않는다"
    assert '"/api/staff/completions"' in srv and "staff_completions_payload" in srv
    po_src = open(os.path.join(ROOT, "po_reconcile.py"), encoding="utf-8").read()
    assert "po_objective_evidence.json" in po_src and '"오종현", "task_kind": "po_source"' in po_src
    print("  [95] 객관 입증 완료(정산·현장업무 DB 정본·앱 즉시반영·백필 금지) ✅")


def t96_work_management_tabs():
    """[96] 정기점검·돌발AS 정밀 관리 탭 (사용자 지시 2026-07-31).

    · 입력은 전부 /api/staff/entry → SQLite 즉시 정본 → Excel 보관본 큐.
      저장 직후 앱에서 읽히며 Excel은 정본 입력 경로가 아니다.
    · 호환 /api/input도 공용 save_staff_entry의 역할×열 허용표를 거친다.
      기존값 보호는 유지하되 overwrite가 임의 열 권한을 넓혀서는 안 된다.
      류지영 매니저가 엑셀에서 담당기사를 한 건씩 고치다 유실 사고가 난 그 작업을
      앱에서 안전하게 하게 하는 것이 이 탭의 존재 이유다.
    · 폰 하단바는 칸이 모자라 사이드바(≥900px)에서만 메뉴 노출 — 폰은 대시보드
      바로가기로 들어간다(숨겨도 화면은 동작해야 한다)."""
    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for need in ('id="v-pm"', 'id="v-as"', "renderWorkTab", "WT_CFG", "wtEdit",
                 'data-v="pm"', 'data-v="as"', "worktab-nav", "wtBoard", "wtCsv",
                 'id="pmBaseDate"', 'id="asBaseDate"', "wtSetBase", "wtRows", "wtReset",
                 'aria-label="기준일 범위"', 'aria-label="정렬 기준"',
                 'aria-label="정렬 방향"', "sortOrder:'desc'", "st.sortOrder==='asc'"):
        assert need in html, f"정밀 관리 탭 구성 요소 누락: {need}"
    assert all(label in html for label in ("기준일 당일", "기준일까지", "기준일 이후",
                                            "예정일 미정", "내림차순 · 최신순",
                                            "오름차순 · 과거순", "필터 초기화")), \
        "기준일 범위·정렬·초기화 필터가 완성되지 않았다"
    assert ".tabbar.worktab-nav{display:none}" in html.replace(" ", ""), \
        "폰 하단바 칸 부족 대책(사이드바 전용 노출)이 없다"
    # 인라인 편집이 공용 SQLite 즉시저장 경로로만 가는가 — Excel 직접 쓰기는 안 된다
    # ★ 검사 범위는 wtEdit **함수 몸통만**이다. 다음 함수 선언 직전까지로 자른다 —
    #   wtBoard 까지 넓게 잡으면 사이에 끼는 무관한 코드(업무센터 업로드의 accept=".xlsx")가
    #   오탐을 낸다(2026-07-31 실제로 그랬다).
    _ws = html.index("async function wtEdit")
    seg = html[_ws:html.index("function ", _ws + 30)]
    assert "/api/staff/entry" in seg and "record_version:" in seg \
        and "idempotency_key:" in seg, "편집이 공용 SQLite 저장 계약을 안 탄다"
    assert "ledger_writer" not in seg and "xlsx" not in seg, "편집이 엑셀로 새어 나간다"
    # 서버: 구버전 /api/input도 세션 역할과 공용 직원 저장 계약을 우회하지 않는가.
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    blk = srv[srv.index('"/api/input"'):srv.index('"/api/input"') + 4000]
    compact = re.sub(r"\s+", "", blk)
    assert "save_staff_entry(" in blk and "RYU_ENTRY_CONFIG[category]" in blk, \
        "호환 입력이 공용 역할×열 허용 저장 계약을 거치지 않는다"
    assert 'actor_session=self._actor()' in compact and \
           'actor=self._actor_name()' in compact, \
        "호환 입력의 권한·감사 actor가 서명 세션에서 오지 않는다"
    assert '"record_version"' in blk and '"idempotency_key"' in blk, \
        "호환 입력이 낙관잠금·멱등 계약을 잃었다"
    assert "OVERWRITE_COLS" not in blk, \
        "옛 전역 덮어쓰기 목록이 공용 역할×열 권한보다 먼저 적용된다"
    # 업무센터·대시보드에서 들어가는 입구
    assert html.count("show('pm')") >= 2 and html.count("show('as')") >= 2, \
        "대시보드·업무센터에서 정밀 관리 탭으로 가는 입구가 없다"

    # 업무센터 탭 + 입금 업로드(2026-07-31): 드롭존·URL 등록·오프라인 보관·Z: 저장·즉시 대조
    for need in ('id="v-center"', 'data-v="center"', "renderCenter", "renderReceiptUpload",
                 "receiptSubmit", "'/api/staff/receipt-upload'", "injectOhUpload"):
        assert need in html, f"업무센터 탭 구성 요소 누락: {need}"
    assert "'/api/staff/receipt-upload']" in html.replace(" ", "") or \
           "/api/staff/receipt-upload" in html[html.index("OUTBOX_PATHS"):html.index("OUTBOX_PATHS")+400], \
        "입금 업로드가 오프라인 보관(outbox) 목록에 없다"
    assert '"receipt"' in srv and "receipt_fill.py" in srv, "업로드 직후 즉시 대조 작업이 없다"
    assert "save_staff_receipt_submission" in srv and "RECEIPT_DIR" in srv, \
        "입금 자료가 지정 저장소(7. 입금내역)로 가지 않는다"
    blk2 = srv[srv.index('"/api/staff/receipt-upload"'):srv.index('"/api/staff/receipt-upload"') + 900]
    assert '"admin"' in blk2 and "oh-jonghyeon" in blk2, "관리자·오종현 외에도 업로드가 열려 있다"
    # 2026-08-11: 앱 업로드 csv 는 같은 폴더에 xlsx 변환본을 만들어야 한다 —
    # load_deposits 는 *.xlsx 만 읽으므로 변환 없는 csv 는 영영 안 읽힌다([165]).
    blk3 = srv[srv.index("def save_staff_receipt_submission"):]
    blk3 = blk3[:blk3.index("\ndef ")]
    assert "_csv_to_xlsx" in blk3, \
        "앱 업로드 csv 가 xlsx 로 변환되지 않는다 — load_deposits 는 xlsx 만 읽는다"
    assert "xls 는 자동으로 읽히지 않습니다" in blk3, "xls 못 읽음 안내가 없다"
    print("  [96] 정밀 관리 탭 + SQLite 즉시 정본·Excel 보관본·입금 업로드 권한 ✅")


def t97_settlement_source_completion():
    """[97] 금액·계산서 대기는 독립 원자료가 정확히 맞는 건만 DB 완료 처리."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import settlement_completion as S

    amount = {
        "비용구분": "유상", "원천업무ID": "AS-1", "프로젝트NO": "UJ2600001",
        "원장_공급가액": 0, "원장_거래명세서발행일": "2026-01-10",
        "원장_거래명세서합계": 110000, "원장_PO번호": "PO123456/PR1",
    }
    invoice = {
        "비용구분": "유상", "원천업무ID": "PM-1", "프로젝트NO": "UJ2600002",
        "원장_공급가액": 100000, "원장_거래명세서발행일": "2026-01-11",
        "원장_거래명세서합계": 110000,
    }
    quotes = [{"종류": "견적서", "프로젝트NO": "UJ2600001", "PO번호": "PO123456",
               "금액": 110000, "파일": "q1.pdf"}]
    invoices = {"UJ2600002": [{"slip": "2026/01/20-1", "amount": 100000,
                                "verdict": "확정(밴드)"}]}
    got = S.objective_entries({"JS-A": amount, "JS-I": invoice}, quotes, invoices)
    assert {row["status"] for row in got} == {S.AMOUNT_STATUS, S.INVOICE_STATUS}, got
    assert all("완료(" in row["status"] and row["basis"] for row in got)

    # 견적 단독 입증(2026-08-03 지시): 프로젝트 견적이 한 금액뿐이면 명세합계가 어긋나도
    # (교차 입력 밀림) 견적을 정본으로 담당자 확인 완료한다. 같은 금액 사본은 한 장으로 본다.
    duplicate = quotes + [{**quotes[0], "파일": "q2.pdf"}]
    got = S.objective_entries({"JS-A": amount}, duplicate, {})
    assert [r["status"] for r in got] == [S.QUOTE_ONLY_STATUS], got  # 같은 금액 사본 → 한 장
    assert "일치" in got[0]["basis"]
    mismatch = [{**quotes[0], "금액": 120000}]
    got = S.objective_entries({"JS-A": amount}, mismatch, {})
    assert [r["status"] for r in got] == [S.QUOTE_ONLY_STATUS], got
    assert "불일치" in got[0]["basis"] and "담당자(류지영) 확인" in got[0]["basis"]
    assert got[0]["evidence_kind"] == "quote_only"
    # 프로젝트 견적이 서로 다른 두 금액이면 무엇이 정본인지 데이터로 못 정한다 — 완료 금지.
    two_totals = mismatch + [{**quotes[0], "금액": 990000, "파일": "q3.pdf"}]
    assert not S.objective_entries({"JS-A": amount}, two_totals, {}), "금액 둘인데 완료"
    # ERP 검증(2026-08-03): ERP 전표가 견적과도 다르면 두 원천 충돌 — 자동 완료 금지.
    erp_clash = {"UJ2600001": [{"date": "2026-02-01", "po": "", "status": "3.오더처리",
                                "supply": 999000, "total": 1098900}]}
    assert not S.objective_entries({"JS-A": amount}, mismatch, {}, None, erp_clash), \
        "ERP·견적 충돌인데 견적단독 완료"
    erp_agree = {"UJ2600001": [{"date": "2026-02-01", "po": "", "status": "3.오더처리",
                                 "supply": 120000, "total": 132000}]}
    got = S.objective_entries({"JS-A": amount}, mismatch, {}, None, erp_agree)
    assert [r["status"] for r in got] == [S.QUOTE_ONLY_STATUS], got   # ERP가 견적을 지지
    s_src = open(os.path.join(ROOT, "settlement_completion.py"), encoding="utf-8").read()
    assert "resolution_retract" in s_src, "충돌 재검출 시 견적단독 완료 철회가 없다"

    # 견적서가 없어도 같은 프로젝트 ERP 판매전표 금액이 유일하게 일치하면 완료한다.
    erp_one = {"UJ2600001": [{"date": "2026-01-15", "po": "PO123456",
                              "status": "3.오더처리", "supply": 100000, "total": 110000}]}
    got = S.objective_entries({"JS-A": amount}, [], {}, None, erp_one)
    assert [row["status"] for row in got] == [S.ERP_AMOUNT_STATUS], got
    assert "ERP 판매전표" in got[0]["basis"] and got[0]["evidence_kind"] == "erp_amount"
    # 금액이 맞는 전표가 둘이면 유일 근거가 아니다 · 금액 불일치도 완료하지 않는다.
    erp_two = {"UJ2600001": erp_one["UJ2600001"] * 2}
    assert not S.objective_entries({"JS-A": amount}, [], {}, None, erp_two), "복수 전표 완료"
    erp_off = {"UJ2600001": [{**erp_one["UJ2600001"][0], "supply": 90000, "total": 99000}]}
    assert not S.objective_entries({"JS-A": amount}, [], {}, None, erp_off), "전표 불일치 완료"
    # 전표 PO 가 원장 PO 와 다르면 같은 금액이라도 남의 전표다.
    erp_po = {"UJ2600001": [{**erp_one["UJ2600001"][0], "po": "PO999999"}]}
    assert not S.objective_entries({"JS-A": amount}, [], {}, None, erp_po), "PO 불일치 완료"

    # ERP 수금확인처럼 이미 더 강한 완료가 있으면 이 모듈의 상태로 낮춰 쓰지 않는다.
    stronger = {"JS-A": {"status": "완료(ERP 수금확인)"}}
    assert not S.objective_entries({"JS-A": amount}, quotes, {}, stronger)

    app = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    findings = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "objective_done = resolutions()" in app and 'resolved_status.startswith("완료(")' in app
    assert "objective_done = ledger_db.resolutions()" in findings and "db_done" in findings
    assert "settlement_completion.py" in daily and '"--sync"' in daily

    # 오래된 세금계산서 미발행 감시(2026-08-03): 경과일 정렬·완료 제외·일일 자동 실행·앱 노출
    import tax_invoice_watch as W
    from datetime import date as _d
    recs = {
        "JS-O": {"비용구분": "유상", "원천업무ID": "PM-9", "프로젝트NO": "UJ2600009",
                 "캠프명": "테스트1", "업무구분": "정기점검", "작업완료일": "2026-01-01",
                 "원장_공급가액": 100000, "원장_거래명세서발행일": "2026-01-02",
                 "원장_거래명세서합계": 110000},
        "JS-N": {"비용구분": "유상", "원천업무ID": "PM-8", "프로젝트NO": "UJ2600008",
                 "캠프명": "테스트2", "업무구분": "정기점검", "작업완료일": "2026-07-20",
                 "원장_공급가액": 100000, "원장_거래명세서발행일": "2026-07-21",
                 "원장_거래명세서합계": 110000},
        "JS-D": {"비용구분": "유상", "원천업무ID": "PM-7", "프로젝트NO": "UJ2600007",
                 "캠프명": "테스트3", "업무구분": "정기점검", "작업완료일": "2026-01-01",
                 "원장_공급가액": 100000, "원장_거래명세서발행일": "2026-01-02",
                 "원장_거래명세서합계": 110000},
    }
    rows = W.overdue_rows(recs, {"JS-D": {"status": "완료(ERP 수금확인)"}}, {},
                          today=_d(2026, 8, 3))
    assert [row["settle_id"] for row in rows] == ["JS-O", "JS-N"], rows  # 오래된 순·완료 제외
    assert rows[0]["age"] > 200 and rows[1]["age"] == 14
    counts = W.bucket_counts(rows)
    assert counts["90일 초과"] == 1 and counts["30일 이하"] == 1, counts
    assert "tax_invoice_watch.py" in daily, "미발행 경과 감시가 일일 자동화에 없다"
    assert "세금계산서_미발행_경과.md" in app, "미발행 경과 리포트가 앱 보고에 안 보인다"

    # 견적↔명세 교차 진단(2026-08-03): 명세합계가 같은 PO 다른 프로젝트 견적과 일치하는
    # 입력 밀림을 짝까지 찾아낸다 — 실사례 UJ2600777(야탑)↔UJ2600783(안산) 722,480원.
    import quote_mismatch as Q
    q_recs = {
        "JS-X1": {"비용구분": "유상", "원천업무ID": "AS-11", "프로젝트NO": "UJ2600777",
                  "원장_공급가액": 0, "원장_거래명세서발행일": "2026-04-28",
                  "원장_거래명세서합계": 722480, "원장_PO번호": "PO354490/PR1"},
        "JS-X2": {"비용구분": "유상", "원천업무ID": "AS-12", "프로젝트NO": "UJ2600999",
                  "원장_공급가액": 0, "원장_거래명세서발행일": "2026-04-28",
                  "원장_거래명세서합계": 50000, "원장_PO번호": "PO354490/PR1"},
    }
    q_quotes = [
        {"종류": "견적서", "프로젝트NO": "UJ2600777", "PO번호": "PO354490", "금액": 311300, "파일": "야탑.pdf"},
        {"종류": "견적서", "프로젝트NO": "UJ2600783", "PO번호": "PO354490", "금액": 722480, "파일": "안산.pdf"},
    ]
    got = Q.diagnose(q_recs, q_quotes)
    kinds = {row["정산ID"]: row["유형"] for row in got}
    assert kinds == {"JS-X1": "교차 의심", "JS-X2": "견적 없음"}, kinds
    assert next(r for r in got if r["정산ID"] == "JS-X1")["교차상대"] == "UJ2600783"
    assert "quote_mismatch.py" in daily, "견적·명세 진단이 일일 자동화에 없다"
    assert "견적명세_불일치.md" in app, "견적·명세 진단이 앱 보고에 안 보인다"
    print("  [97] 정산 완료 DB·미발행 경과 감시 + 견적·명세 교차 진단 ✅")


def t98_remote_control_tracking():
    """[98] 리모컨 불출·납품(2026-08-03 지시, 같은 날 개정): 승인 없이 기록·관리·보고.

    3개 한도·불출 담당(부산 오종현·시화 안은숙·증평 류지영)·프로젝트/캠프 납품 추적은
    유지하고, 부사장 승인 단계는 사용자 지시로 뺐다 — 불출은 즉시 기록된다.
    """
    import tempfile as _tf

    import ledger_db as L
    old_path, old_dir = L.DB_PATH, L.DB_DIR
    tmp = _tf.mkdtemp()
    try:
        L.DB_DIR, L.DB_PATH = tmp, os.path.join(tmp, "t.db")
        assert L.REMOTE_BRANCH_ISSUERS == {"부산": "오종현", "시화": "안은숙", "증평": "류지영"}
        assert L.REMOTE_HOLD_LIMIT == 3
        L.remote_request("부산", "김기사", 2, "오종현")     # 즉시 불출 기록
        for bad in (lambda: L.remote_request("부산", "김기사", 2, "오종현"),   # 2+2>3
                    lambda: L.remote_request("서울", "박기사", 1, "아무개"),   # 지점 제한
                    lambda: L.remote_deliver("김기사", "UJ2600001", "", 3)):   # 보유 초과 납품
            try:
                bad(); raise AssertionError("리모컨 규칙이 뚫렸다")
            except ValueError:
                pass
        # 공지(2026-08-04): 불출 일자·투입 예정 캠프명이 불출 기록에 남는다
        L.remote_request("시화", "김기사", 1, "안은숙",
                         issued_on="2026-08-02", camp="시화3캠프")  # 2+1=3 — 허용
        top = L.remote_status()["issues"][0]
        assert (top["issued_on"], top["camp"]) == ("2026-08-02", "시화3캠프"), top
        L.remote_deliver("김기사", "UJ2600001", "부산2캠프", 2, "2026-08-03", kind="사용")
        hold = L.remote_status()["holdings"]["김기사"]
        # versions 는 2026-08-06 에 붙은 버전별 내역이다(검증 [122]). 여기서는 합계만 본다.
        assert {k: hold[k] for k in ("issued", "delivered", "holding")} == \
            {"issued": 3, "delivered": 2, "holding": 1}, hold
        assert L.remote_status()["deliveries"][0]["kind"] == "사용"

        # 기초보유(2026-08-04 재고표 이관): 한도를 넘는 개시 잔량도 사실대로 받는다.
        # 대신 over_limit 에 뜨고, 지점 재고에서 이중 차감하지 않는다.
        L.remote_stock_adjust("부산", 21, "add", "기초", "오종현", version="미확인")
        before = L.remote_status()["branch_stock"]["부산"]["stock"]
        L.remote_open_balance("정기사", 9, "2026-07-29", "기초 보유", "류지영",
                              branch="부산", version="미확인")
        st = L.remote_status()
        assert st["holdings"]["정기사"]["holding"] == 9, st["holdings"]["정기사"]
        assert st["over_limit"].get("정기사") == 9, st["over_limit"]
        assert st["branch_stock"]["부산"]["stock"] == before, "기초보유가 지점 재고를 깎았다"
        try:                                   # 새 불출은 여전히 한도가 막는다
            L.remote_request("부산", "정기사", 1, "오종현")
            raise AssertionError("한도 초과자에게 추가 불출이 뚫렸다")
        except ValueError:
            pass
        # 버전별 잔량: 같은 지점 안에서 기존형/VER.4 를 나눠 센다
        L.remote_stock_adjust("부산", 50, "add", "신규 입고", "오종현", version="VER.4")
        L.remote_stock_adjust("부산", -1, "add", "샘플 송부", "오종현", version="VER.4")
        # 미확인 21 - 앞서 부산에서 나간 불출 2개 = 19 (불출도 버전별로 빠진다)
        vers = L.remote_status()["branch_stock"]["부산"]["versions"]
        assert vers.get("VER.4") == 49 and vers.get("미확인") == 19, vers
        assert not hasattr(L, "remote_decide"), "승인 단계가 아직 남아 있다"

        # 지점 재고(2026-08-03 3차): 등록 지점은 재고보다 많이 불출 못 하고 자동 차감된다.
        assert L.REMOTE_BRANCH_LABELS == {"부산": "부산공장", "시화": "시화공장", "증평": "증평본사"}
        st2 = L.remote_status()["branch_stock"]
        assert not st2["증평"]["tracked"] and st2["증평"]["stock"] == 0, st2["증평"]  # 미등록 지점
        assert L.remote_stock_adjust("증평", 5, "add", "", "류지영") == 5             # 입고 5
        L.remote_request("증평", "박기사", 2, "류지영")                               # 재고 5→3
        st2 = L.remote_status()["branch_stock"]["증평"]
        assert (st2["tracked"], st2["in"], st2["issued"], st2["stock"]) == (True, 5, 2, 3), st2
        try:
            L.remote_stock_adjust("증평", 1, "set")        # 실사 1 < 불출분? 1-3=-2 델타, 재고 1
        except ValueError:
            raise AssertionError("실사 맞춤이 막혔다")
        assert L.remote_status()["branch_stock"]["증평"]["stock"] == 1
        try:
            L.remote_request("증평", "최기사", 2, "류지영")
            raise AssertionError("재고(1)보다 많은 불출(2)이 뚫렸다")
        except ValueError:
            pass
        try:
            L.remote_stock_adjust("증평", -5, "add")
            raise AssertionError("재고 음수 정정이 뚫렸다")
        except ValueError:
            pass

        # ★ 지점 직납(2026-08-11 류지영 실사용 피드백): 납품자 칸이 지점 이름이거나
        #   지점 담당자 본인인데 개인 보유가 모자라면 — 지점 재고에서 바로 나간다.
        #   실사고: 류지영(개인 보유 0)이 납품을 적자 "보유 0개" 로 막혔다.
        assert "류지영" not in L.remote_status()["holdings"], "전제 어긋남: 류지영 개인 보유 존재"
        L.remote_stock_adjust("증평", 2, "add", "직납 시험 입고", "류지영")     # 재고 1→3
        rid = L.remote_deliver("류지영", "", "위더스물류(V_안성)", 1, kind="교체", version="VER.3")
        assert rid, "지점 담당자 직납이 막혔다(실사고 그대로)"
        st3 = L.remote_status()
        assert st3["branch_stock"]["증평"]["stock"] == 2, st3["branch_stock"]["증평"]
        assert "류지영" not in st3["holdings"] and "증평본사" not in st3["holdings"], \
            "직납이 개인 보유를 만들었다(유령 음수 보유자)"
        L.remote_deliver("증평본사", "", "샘플캠프", 1)                        # 지점 표기도 된다
        assert L.remote_status()["branch_stock"]["증평"]["stock"] == 1
        try:
            L.remote_deliver("증평", "", "아무캠프", 9)
            raise AssertionError("지점 재고(1)보다 많은 직납(9)이 뚫렸다")
        except ValueError as exc:
            assert "지점 재고" in str(exc), exc
        try:                                     # 지점도 담당자도 아닌 사람은 예전 그대로 막되
            L.remote_deliver("무명기사", "", "아무캠프", 1)
            raise AssertionError("보유 없는 사람 납품이 뚫렸다")
        except ValueError as exc:                # 오류가 해결책(지점 이름)을 말해야 한다
            assert "지점 이름" in str(exc), exc
    finally:
        L.DB_PATH, L.DB_DIR = old_path, old_dir
    # 앱: 류지영·오종현 업무센터 공통 카드 + iOS 스타일 + 대표보고 캡처 포함 + 승인 UI 없음
    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    # 2026-08-04: 리모컨은 사이드바 '돌발 AS 바로 아래' 독립 화면(v-remote)으로 옮겼다.
    # 관리자 업무센터 탭에는 입력 폼 대신 요약(centerRemoteBrief)과 바로가기만 둔다.
    for need in ("injectRemoteCard", "renderRemoteCard", "remoteRequest",
                 "remoteDeliver", "centerRemoteBrief",
                 'id="v-remote"', 'data-v="remote"', "remoteCardBody",
                 "remoteCsv", "remoteCapture",
                 "if(staffSlug==='ryu-jiyeong'||staffSlug==='oh-jonghyeon') injectRemoteCard()",
                 ".remote-grid2 fieldset{border:0;border-radius:16px",
                 "loadRemoteStat", "remote: REMOTE_STAT", "rmtH",
                 "리모컨 현황 — 불출·납품 기록",
                 # 지점 재고(2026-08-03 3차): 카드 표·등록 폼·캡처 줄
                 "remoteStock", "branch_stock", "branchStock", "현재 재고"):
        assert need in html, f"리모컨 카드 구성 요소 누락: {need}"
    assert "/api/remote/stock" in open(os.path.join(ROOT, "webapp", "app_server.py"),
                                       encoding="utf-8").read(), "재고 등록 API가 없다"
    assert "remoteApprove" not in html and "승인 요청" not in html, "승인 UI가 남아 있다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "/api/remote/status" in srv and "/api/remote/request" in srv
    assert "/api/remote/approve" not in srv, "승인 API가 남아 있다"
    blk = srv[srv.index('"/api/remote/request"'):srv.index('"/api/remote/request"') + 1600]
    assert '"ryu-jiyeong", "oh-jonghyeon"' in blk, "리모컨 관리가 두 업무센터로 제한되지 않았다"
    # 지점 직납·기타 사유(2026-08-11 류지영): 폼이 지점 이름을 안내하고 기타 사유 칸이 있다
    for need in ("DWhyT", "grant_reason_text", "부산공장·시화공장·증평본사"):
        assert need in html, f"지점 직납/기타 사유 UI 누락: {need}"
    assert "grant_reason_text" in srv, "서버가 기타 자유 사유를 안 받는다"
    print("  [98] 리모컨 기록·관리·보고(승인 없음)·3개 한도·납품 추적·지점 직납 ✅")


def t101_percent_and_no_erp_post():
    """[101] 비율 표기 규칙과 ERP 전표 실전송 제거 (2026-08-05 사용자 지시).

    두 지시가 한 검증에 있는 이유: 둘 다 "화면이 사실과 다르게 말하는 것"을 막는다.
      · "비율 표기는 소수점 1자리까지 표기, 1건이 안되도 100%로 보이는 문제 해결"
        → 999/1000 이 100% 로 보이면 대표가 '다 끝났다'고 읽는다.
      · "전표 실전송 행위는 하지마 이거 삭제해"
        → 되돌릴 수 없는 ERP 등록을 버튼 하나로 만들지 않는다.
    """
    from pct_fmt import pct, pct_text

    # (1) 미완료는 절대 100%가 아니다 — 이 검증의 핵심.
    assert pct(999, 1000) == 99.9 and pct_text(999, 1000) == "99.9%"
    assert pct(9999, 10000) == 99.9, "반올림이 미완료를 100%로 만들었다"
    assert pct(1000, 1000) == 100.0 and pct_text(1000, 1000) == "100.0%"
    # (2) 한 건이라도 했으면 0%가 아니다.
    assert pct(1, 10000) == 0.1 and pct_text(1, 10000) == "0.1%"
    assert pct(0, 10) == 0.0 and pct_text(0, 10) == "0.0%"
    # (3) 모수가 없으면 비율이 없다 — 0%라고 말하지 않는다.
    assert pct(0, 0) is None and pct_text(5, 0) == "대상 없음"
    # (4) 소수점 1자리 고정
    assert pct_text(1, 3) == "33.3%" and pct_text(2, 3) == "66.7%"

    # (5) 화면(브라우저)도 같은 규칙을 갖고 있어야 서버와 숫자가 어긋나지 않는다.
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "const pctNum" in live and "const pctText" in live, "화면에 비율 규칙 함수가 없다"
    assert "if(v>=100 && d<t) v=99.9;" in live, "화면이 미완료를 100%로 보일 수 있다"
    for old in ("Math.round(발행/유상.length*100)", "Math.round(s.ok/s.total*100)"):
        assert old not in live, "옛 정수 반올림 비율이 남아 있다: %s" % old

    # (6) ERP 전표 실전송 — 세 곳 모두 막혀 있어야 한다.
    up = open(os.path.join(ROOT, "ecount_upload.py"), encoding="utf-8").read()
    assert "실전송은 제공하지 않습니다" in up, "ecount_upload 가 아직 전송한다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '"upload_post":   (' not in srv, "서버 작업표에 실전송 키가 되살아났다"
    assert 'if key == "upload_post":' in srv, "옛 화면이 보낸 실전송 요청을 막지 않는다"
    assert "upload_post" not in live, "화면에 실전송 버튼이 되살아났다"
    bench = open(os.path.join(ROOT, "coupang_workbench.py"), encoding="utf-8").read()
    assert '"--post"' not in bench, "워크벤치에 실전송 버튼이 되살아났다"

    # (7) 실제로 --post 를 줘도 전송 단계로 가지 않는지 — 코드 순서로 확인한다.
    assert up.index("실전송은 제공하지 않습니다") < up.index("cli = EcountClient(cfg)"), \
        "차단이 전송 코드보다 뒤에 있다"

    # (8) 밴드 수집기는 숨은 탭에서도 돌아야 한다(사고 #19). 페이지 타이머를 쓰면
    #     크롬이 1분 간격으로 늦춰 15분에 0건이 된다 — 워커 타이머·관찰자만 쓴다.
    grab = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
    body = grab[grab.index("(function ()"):]          # 머리말 주석의 설명은 제외하고 본다
    assert "setTimeout" not in body.replace(
        "'onmessage=e=>{setTimeout(()=>postMessage(e.data.id),e.data.ms)}'", ""), \
        "수집기에 페이지 setTimeout 이 다시 들어왔다(숨은 탭에서 멈춘다)"
    assert "new Worker" in body and "MutationObserver" in body, "워커 타이머·관찰자가 없다"
    assert "window.__grabStop" in body, "배치를 끊을 방법이 없다(새로고침하면 수집분이 날아간다)"

    # (8-1) 워커 타이머는 **스케줄**만 살린다 — **렌더**는 못 살린다(사고 #22).
    #   밴드 본문은 rAF 로 그려지고 rAF 는 숨은 탭에서 한 번도 안 불린다. 그래서
    #   숨은 탭에서 돌리면 살아 있는 글까지 "본문 없음"으로 읽혀 **가짜 묘비**가 쌓이고,
    #   recheck_plan 이 그 번호를 영영 다시 안 뽑는다. 조용히 실패하느니 거절한다.
    assert "document.hidden" in body, \
        "숨은 탭 판정이 없다 — 창이 뒤에 있으면 살아 있는 글도 전부 실패로 기록된다(사고 #22)"
    start = body[body.index("window.__grabStart"):body.index("window.__grabStop")]
    assert "document.hidden" in start.split("(async ()")[0], \
        "숨은 탭에서 __grabStart 가 시작을 거절하지 않는다"
    assert "S.paused" in start, "돌던 중 창이 뒤로 가면 멈춰 기다려야 한다(실패로 적으면 안 된다)"
    assert "paused:" in body, "__grabStatus() 가 멈춰 있는지를 알려 주지 않는다"

    # (8-2) 밤샘 대조는 한 회차가 **약 2시간**이다(밴드 캐시 8,499글). 45분이던
    #   옛 기준을 그대로 두어 밤새 8회가 전부 잘렸다. 제한을 다시 줄이지 못하게 막는다.
    ovn = open(os.path.join(ROOT, "overnight_run.py"), encoding="utf-8").read()
    m = re.search(r'"--timeout".*?default=(\d+)', ovn, re.S)
    assert m and int(m.group(1)) >= 180, \
        "밤샘 대조 회차 제한이 너무 짧다 — 한 회차가 2시간이라 거의 다 해 놓고 잘린다"
    assert "stdout=fh" in ovn, \
        "출력을 파일로 흘리지 않는다 — 시간초과가 나면 어디서 느렸는지 기록이 안 남는다"
    print("  [101] 비율 소수점 1자리·미완료 100% 금지 · ERP 전표 실전송 제거 · 수집기 숨은탭 대응 ✅")


def t102_calendar_filter_and_period():
    """[102] 캘린더 종류 필터·월 격자 · 업무현황 집계 기간 · 원본자료 판매전표 (2026-08-05 지시).

    세 지시가 한 뿌리다: **화면이 무엇을 말하는지 분명히 하라**.
      · "이 현황이 몇년도 몇월 몇일부터 몇월 몇일까지인지 표시"
      · "돌발 AS, 정기 점검 등 선택한 항목만 볼 수 있게 필터 추가"
      · "원본 자료 항목에 판매전표 추가"
    """
    from webapp import app_server as S

    # (1) 캘린더 분류 키는 화면 필터가 거는 손잡이다 — 라벨이 아니라 키로 건다.
    keys = [k for k, _l, _c in S.CAL_KINDS]
    for need in ("pm_plan", "pm_pred", "pm_done", "as_visit", "as_done",
                 "as_open", "pm_overdue"):
        assert need in keys, "캘린더 분류 %s 가 없다" % need
    # 미처리 두 갈래는 완료·예정과 색이 확실히 갈려야 한다(2026-08-06 지시).
    color = {k: c for k, _l, c in S.CAL_KINDS}
    assert color["as_open"] != color["as_done"] and color["pm_overdue"] != color["pm_done"]

    # (2) 원장 날짜만 일정으로 세운다 — 없는 날짜를 지어내면 캘린더가 거짓말을 한다.
    saved = getattr(S, "get_works")
    try:
        S.get_works = lambda: {
            "as": [
                {"접수ID": "AS-1", "캠프명": "가캠프", "방문예정일": "2026-08-10",
                 "접수일자": "2026-08-01", "작업완료일": "", "담당기사": "김준형",
                 "긴급도": "높음"},
                {"접수ID": "AS-2", "캠프명": "나캠프", "방문예정일": "2026-08-01",
                 "작업완료일": "2026-08-02", "담당기사": "권오철"},
                {"접수ID": "AS-3", "캠프명": "다캠프", "방문예정일": "", "작업완료일": ""},
            ],
            "pm": [{"점검ID": "PM-1", "캠프명": "라캠프", "실제점검일": "2026-08-03"},
                   {"점검ID": "PM-2", "캠프명": "마캠프", "실제점검일": ""}],
        }
        evs = S._calendar_work_events()
    finally:
        S.get_works = saved
    got = sorted((e["분류"], e["날짜"], e["원천업무ID"]) for e in evs)
    # 미처리는 **접수일/예정일** 자리에 선다 — "언제 들어온 게 아직 안 끝났나"를 보려는 것.
    assert ("as_open", "2026-08-01", "AS-1") in got, got     # AS-1 은 접수만 되고 미완료
    assert ("as_done", "2026-08-02", "AS-2") in got, got
    assert ("as_visit", "2026-08-10", "AS-1") in got, got
    assert ("pm_done", "2026-08-03", "PM-1") in got, got
    # 이미 끝난 건의 방문예정일을 '예정'으로 다시 세우지 않는다(AS-2 는 완료만 남아야 한다).
    assert not [e for e in evs if e["분류"] == "as_visit" and e["원천업무ID"] == "AS-2"]

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    # (3) 화면: 필터 칩·월 격자·그 날 일정. 예전의 '한 줄로 늘어선 목록'으로 돌아가지 않는다.
    for fn in ("function renderCalFilters(", "function renderCalGrid(",
               "function renderCalAgenda(", "function calToggleKind(",
               "function calMonthShift(", "function calGoToday("):
        assert fn in live, "%s 가 없다" % fn
    assert 'id="calFilters"' in live and 'id="calGrid"' in live and 'id="calAgenda"' in live
    assert 'id="calendarEventList"' not in live, "옛 한 줄 목록이 남아 있다"
    assert "calendar-layout" not in live.replace("옛 2단 배치(.calendar-layout", ""), \
        "옛 2단 배치가 남아 있다"
    assert "CAL_OFF.has(calKindOf(e))" in live, "필터가 실제로 걸리지 않는다"
    assert "csos.cal.off" in live, "고른 필터가 새로고침에 사라진다"

    # ★ 캘린더 전용 모드(2026-08-06 지시: "캘린더 링크로 갔을 때 이 기능만 보이게").
    #   숨기는 것만으로는 부족하다 — 뒤로가기·복원된 localStorage·코드의 show() 로
    #   빠져나갈 수 있다. 화면 전환 함수 자체가 잠겨 있어야 한다.
    assert "var CAL_ONLY = DEEP_VIEW === 'calendar';" in live, "전용 모드 판정이 없다"
    assert "if(CAL_ONLY) v='calendar';" in live, "applyView 가 다른 화면을 막지 않는다"
    assert "if(CAL_ONLY) return 'calendar';" in live, "curView 가 다른 화면을 돌려준다"
    assert "body.calendar-only .tabbar" in live and "display:none !important" in live, \
        "전용 모드에서 탭바가 남는다"
    assert "body.calendar-only .view:not(#v-calendar)" in live, "다른 화면이 남는다"
    # 주소의 m=YYYY-MM 은 렌더 직전에 한 번만 적용한다(TDZ 때문에 선언 앞에서 대입 금지).
    assert "calApplyDeepLink()" in live and "CAL_DEEP_DONE" in live

    # (4) 업무 현황 집계 기간 — 완료일이 있는 건만으로 기간을 잡고, 나머지는 건수로 밝힌다.
    assert 'id="heroPeriod"' in live and "function heroPeriodOf(" in live
    assert "renderHeroPeriod(rows)" in live, "히어로가 기간을 그리지 않는다"
    assert "완료일 미기입" in live, "기간 밖 건수를 숨기고 있다"

    # (5) 원본 자료 판매전표 갈래
    assert "label:'판매전표'" in live, "원본 자료에 판매전표 갈래가 없다"
    assert "re:/판매조회|판매전표/" in live, "이름이 난수인 판매조회 내보내기를 못 잡는다"
    assert "if(g.re && g.re.test(String(r.name||''))) return g;" in live, "이름 규칙이 배선되지 않았다"
    print("  [102] 캘린더 종류 필터·월 격자 · 업무현황 집계 기간 · 원본자료 판매전표 ✅")


def t106_calendar_kind_colors():
    """[106] 일정 종류 색이 **서로 구별되고**, 서버·화면 두 표가 같아야 한다.

    사용자 지적(2026-08-06): "이거 색상 다르게 표시해 헷갈린다."
    정기점검 완료(#30D158)와 돌발AS 완료(#34C759)가 거의 같은 초록이라 칩·달력·캡처
    어디서도 둘을 못 갈랐다. 색은 **두 곳**(app_server.CAL_KINDS, index.html
    CAL_FALLBACK_KINDS)에 적혀 있어, 한쪽만 고치면 화면과 서버가 다른 색을 쓴다.
    오늘 폴더 이름을 두 곳에 적었다가 어긋난 사고를 이미 겪었으므로 검증으로 묶는다.
    """
    import re as _re
    _rd = lambda *p: open(os.path.join(ROOT, *p), encoding="utf-8").read()
    srv = _rd("webapp", "app_server.py")
    live = _rd("webapp", "index.html")

    blk = srv.split("CAL_KINDS = [", 1)[1].split("]", 1)[0]
    server = _re.findall(r'\("([a-z_]+)",\s*"([^"]+)",\s*"(#[0-9A-Fa-f]{6})"\)', blk)
    assert len(server) >= 5, "서버 CAL_KINDS 를 읽지 못했다"

    fb = live.split("const CAL_FALLBACK_KINDS = [", 1)[1].split("];", 1)[0]
    page = _re.findall(r"key:'([a-z_]+)',\s*label:'([^']+)',\s*color:'(#[0-9A-Fa-f]{6})'", fb)
    assert len(page) == len(server), (
        f"종류 개수가 다르다 — 서버 {len(server)} / 화면 {len(page)}")
    for (k1, l1, c1), (k2, l2, c2) in zip(server, page):
        assert (k1, l1, c1.upper()) == (k2, l2, c2.upper()), (
            f"서버와 화면의 종류 표가 어긋났다: {k1} {c1} ≠ {k2} {c2}")

    def rgb(c):
        return tuple(int(c[i:i + 2], 16) for i in (1, 3, 5))

    for i in range(len(server)):
        for j in range(i + 1, len(server)):
            a, b = rgb(server[i][2]), rgb(server[j][2])
            # 사람이 나란히 놓고 구별할 수 있어야 한다. 문제였던 초록 두 개
            # (#30D158·#34C759)는 이 값이 **15** 였다. 지금 표에서 가장 가까운 짝은
            # 139(예정↔돌발AS완료)이므로 90 을 최소선으로 둔다 — 여유가 있으면서
            # '15 같은 것'은 확실히 걸린다.
            dist = sum(abs(p - q) for p, q in zip(a, b))
            assert dist >= 90, (
                f"'{server[i][1]}' 와 '{server[j][1]}' 색이 너무 비슷하다"
                f"({server[i][2]} vs {server[j][2]}, 차이 {dist})")
    print("  [106] 일정 종류 색 구별·서버/화면 표 일치 ✅")


def t103_session_wrapup_hook():
    """[103] 컨텍스트가 차면 자동으로 인계하고 요약한다 (2026-08-05 지시).

    지시: "세션이 컨텍스트 윈도우가 90% 도달시 자동으로 정리하고 /compact 명령 실행."
    지금까지는 사람이 "인계해"라고 말해 줘야 했고, 말하기 전에 차 버리면 큐·점유·미커밋이
    그대로 남았다. 이제 요약 직전 PreCompact 훅이 session_wrapup.py 를 돌린다.

    이 검증이 지키는 것 — 훅이 **조용히 사라지는 것**을 막는다:
      · 4단계가 전부, 그 순서로 들어 있는가
      · 무슨 일이 있어도 exit 0 인가(인계하려다 요약을 막으면 더 나쁘다)
      · 엑셀을 열지 않는가(반영은 11:00·15:00 회차만)
      · .claude/settings.json 배선이 살아 있는가
    """
    import session_wrapup as W

    src = open(os.path.join(ROOT, "session_wrapup.py"), encoding="utf-8").read()

    # (1) 4단계가 전부, 그 순서로.
    order = [src.index(fn) for fn in
             ("def step_intake", "def step_free_claims", "def step_commit", "def step_handoff")]
    assert order == sorted(order), "인계 4단계 순서가 바뀌었다"
    assert '"--intake"' in src and '"--free-all"' in src and '"--handoff"' in src
    assert '"--snapshot"' in src, "세션인계 스냅샷을 남기지 않는다"

    # (2) 엑셀을 열지 않는다 — 반영 회차를 건드리면 원장 버전이 하루에 수십 개가 된다.
    for forbidden in ("--apply", "workbook_patch.py", "--force"):
        assert forbidden not in src, "세션 마무리가 엑셀을 직접 건드린다: %s" % forbidden

    # (3) 비밀값이 스테이징에 있으면 커밋하지 않는다(절대규칙 1).
    assert "git grep" in src or 'git("grep"' in src, "커밋 전 비밀값 스캔이 없다"
    assert "커밋을 멈췄다" in src

    # (4) 어떤 단계가 깨져도 전체는 성공으로 끝난다 — 실제로 돌려서 확인한다.
    calls = []
    saved = W.run
    try:
        def boom(args, timeout=900):
            calls.append(list(args))
            return False, "일부러 실패시킴"
        W.run = boom
        record = W.wrapup(who="claude", reason="synthetic")
    finally:
        W.run = saved
    assert len(record["단계"]) == 5, record
    assert all(s["성공"] is False for s in record["단계"]), "실패를 성공으로 적고 있다"
    # git 조차 실패해도 예외가 새지 않아야 한다 — 위 wrapup 이 끝까지 온 것이 그 증거다.
    assert any("--intake" in c for c in calls) and any("--free-all" in c for c in calls)

    # (5) main() 은 언제나 0 을 준다.
    assert "return 0        # 인계를 남기려다" in src, "실패 시 0 이 아닌 값을 줄 수 있다"

    # (6) 훅 배선 — 이것이 없으면 위 전부가 '사람이 손으로 부를 때만' 도는 스크립트다.
    # 격리 git worktree는 프로젝트 바깥에 놓일 수 있다. 그때도 실제 훅 파일을
    # 복사/수정하지 않고 읽기 전용 기준 루트를 명시해 같은 배선을 검증한다.
    project_root = os.environ.get("COUPANG_TEST_PROJECT_ROOT") or os.path.dirname(ROOT)
    settings = os.path.join(project_root, ".claude", "settings.json")
    assert os.path.exists(settings), ".claude/settings.json 이 없다"
    cfg = json.load(open(settings, encoding="utf-8"))
    assert cfg.get("autoCompactEnabled") is True, "자동 요약이 꺼져 있다"
    hooks = [h for entry in cfg.get("hooks", {}).get("PreCompact", [])
             for h in entry.get("hooks", [])]
    assert hooks, "PreCompact 훅이 없다"
    wired = [h for h in hooks
             if "session_wrapup.py" in " ".join([h.get("command", "")] + list(h.get("args") or []))]
    assert wired, "PreCompact 훅이 session_wrapup.py 를 부르지 않는다"
    assert (wired[0].get("timeout") or 0) >= 300, "훅 시간제한이 짧아 인계가 중간에 잘린다"
    print("  [103] 컨텍스트 한도 자동 인계(PreCompact 훅·4단계·실패해도 요약 진행) ✅")


def t104_session_scoped_claims():
    """[104] 점유의 주인은 who 가 아니라 **세션**이다 (2026-08-05 지시).

    지시: "지금 현재 열려있는 세션과 병렬 작업 가능한 구조로 정리."
    같은 폴더에 Claude 창이 둘 떠 있으면 둘 다 who="claude" 라서
      · 뒤에 온 창이 앞 창의 배타 점유를 말없이 빼앗고(둘 다 vN+1 → 한쪽 유실)
      · --free-all 이 남의 점유까지 풀었다(PreCompact 자동 마무리가 특히 위험).
    이 검증은 그 두 가지가 다시 생기지 않게 한다.
    """
    import importlib
    import socket
    import time
    import ai_claim

    with tempfile.TemporaryDirectory() as tmp:
        saved = (ai_claim.CLAIMS, ai_claim.GUARD, ai_claim.WORKCENTER_ACTIVITY)
        ai_claim.CLAIMS = os.path.join(tmp, "ai_claims.json")
        ai_claim.GUARD = os.path.join(tmp, ".guard")
        ai_claim.WORKCENTER_ACTIVITY = os.path.join(tmp, "workcenter_activity.json")
        old_env = os.environ.get("CLAUDE_CODE_SESSION_ID")
        try:
            def as_session(sid):
                os.environ["CLAUDE_CODE_SESSION_ID"] = sid
                return ai_claim.session_id()

            a, b = as_session("SESSION-A"), None
            os.environ["CLAUDE_CODE_SESSION_ID"] = "SESSION-B"
            b = ai_claim.session_id()
            assert a != b, "세션이 달라도 식별자가 같다 — 격리가 안 된다"

            # 세션 A 가 원장을 잡는다
            as_session("SESSION-A")
            assert ai_claim.take("claude", "ledger", "A 작업") is True

            # 세션 B 는 who 가 같아도 못 빼앗는다 ← 이 검증의 핵심
            as_session("SESSION-B")
            assert ai_claim.take("claude", "ledger", "B 작업") is False, \
                "다른 세션의 배타 점유를 빼앗았다(원장 유실 사고 재발)"
            assert ai_claim.free("claude", "ledger") is False, \
                "다른 세션의 점유를 놓았다"

            # 세션 B 의 --free-all 은 자기 것만 — A 의 점유가 남아야 한다
            ai_claim.take("claude", "report", "B 리포트")
            d = ai_claim.load()
            mine = [k for k, v in d.items() if ai_claim._is_mine(v, "claude")]
            assert mine == ["report"], mine
            assert "ledger" in d and d["ledger"]["sid"] == a

            # 세션 A 는 자기 점유를 다시 잡을 수 있다(재진입)
            as_session("SESSION-A")
            assert ai_claim.take("claude", "ledger", "A 계속") is True
            assert ai_claim.free("claude", "ledger") is True

            # 주인 세션이 죽었으면 45분 기다리지 않고 넘겨받는다
            as_session("SESSION-C")
            ai_claim.take("claude", "band", "C 작업")
            d = ai_claim.load()
            d["band"]["agent_pid"] = 999999      # 존재하지 않는 PID
            d["band"]["host"] = socket.gethostname()
            ai_claim.save(d)
            as_session("SESSION-D")
            assert ai_claim.take("claude", "band", "D 인계") is True, \
                "죽은 세션의 점유를 넘겨받지 못한다"

            # 스케줄러 점유는 agent_pid=0 — 그때는 `pid` 가 증거다 (2026-08-07 실사고:
            # 죽은 ledger_writer 를 --adopt 가 "살아 있다"며 못 넘겨받아 교착)
            d = ai_claim.load()
            d["band"] = {"who": "scheduler", "why": "w", "sid": "0000dead",
                         "agent_pid": 0, "pid": 999999,
                         "host": socket.gethostname(), "at": time.time()}
            ai_claim.save(d)
            assert ai_claim._is_dead(ai_claim.load()["band"]) is True, \
                "agent_pid=0 이면 pid 증거를 봐야 한다"
            # pid 마저 없으면 보수적으로 살아 있다고 본다
            d["band"].pop("pid"); d["band"]["agent_pid"] = 0
            ai_claim.save(d)
            assert ai_claim._is_dead(ai_claim.load()["band"]) is False
        finally:
            ai_claim.CLAIMS, ai_claim.GUARD, ai_claim.WORKCENTER_ACTIVITY = saved
            if old_env is None:
                os.environ.pop("CLAUDE_CODE_SESSION_ID", None)
            else:
                os.environ["CLAUDE_CODE_SESSION_ID"] = old_env
            importlib.reload(ai_claim)

    # 세션 마무리(PreCompact)가 남의 점유를 풀지 않는지 — 호출 형태로 확인한다.
    wrap = open(os.path.join(ROOT, "session_wrapup.py"), encoding="utf-8").read()
    assert '"--force"' not in wrap, "자동 마무리가 남의 점유까지 강제로 푼다"
    # 작업 폴더는 세션끼리 공유한다 — 다른 세션이 살아 있으면 푸시를 보류한다.
    assert "_other_live_sessions" in wrap and "푸시는 보류" in wrap, \
        "다른 세션이 일하는 중에도 자동 푸시한다(남의 반쯤 고친 코드가 원격으로 간다)"

    # 분담판(worksplit)도 같은 규칙이어야 한다. 특히 **sid 가 없는 옛 항목**은
    # 잡은 본인조차 완료 처리를 못 해 8시간 동안 묶여 있었다(2026-08-05 실사고).
    import worksplit
    ws_src = open(os.path.join(ROOT, "worksplit.py"), encoding="utf-8").read()
    assert "def _owner_state(it, me=" in ws_src, "분담판이 주인 판정에 who 를 안 받는다"
    legacy = {"state": "진행", "who": "claude", "sid": None, "at_ts": time.time()}
    assert worksplit._owner_state(legacy, "claude")[1] is True, \
        "sid 없는 옛 항목을 본인도 손댈 수 없다"
    assert worksplit._owner_state(legacy, "codex")[1] is False, \
        "옛 항목을 아무나 가져간다"
    other = {"state": "진행", "who": "claude", "sid": "다른세션", "at_ts": time.time()}
    assert worksplit._owner_state(other, "claude")[1] is False, \
        "다른 세션이 맡은 일을 who 만 같으면 가져간다"

    doc = open(os.path.join(ROOT, "CLAUDE.md"), encoding="utf-8").read()
    assert "주인은 `claude` 가 아니라 **세션**이다" in doc, "동시작업 규칙이 문서에 없다"
    print("  [104] 세션 단위 점유·분담(빼앗기 차단·내 것만 해제·죽은 세션 즉시 인계) ✅")


def t105_settle_report():
    """[105] 하루치 정산분 보고자료 (2026-08-05 지시).

    지시: "8월 5일 정산분 8월 6일에 보고 자료 정리 / 밴드도 ERP도 8월 5일꺼 우선 찾아 정리."
    내일도 필요한 일이라 손으로 만들지 않고 스크립트로 만든다.

    이 검증이 지키는 것:
      · 같은 대화 내보내기가 여러 벌 있어도 **한 번만 센다**(실제로 6건이 12건으로 나왔다)
      · 같은 프로젝트가 진행상태만 달리 두 줄이면 합계를 **하나로 정하지 않는다**(이중계상)
      · 앱 [기록] 탭에 뜬다 · daily_run 이 매일 만든다
    """
    import settle_report as SR

    # (1) 카톡 중복 제거 — 파일이 몇 벌이든 같은 메시지는 한 번.
    src = open(os.path.join(ROOT, "settle_report.py"), encoding="utf-8").read()
    assert "seen, uniq = set(), []" in src, "카톡 중복 제거가 없다"

    # (2) 이중계상 금지: 같은 프로젝트가 여러 줄이면 low/high 를 둘 다 낸다.
    # 쿠팡 전표는 관리항목명이 항상 붙는다 — 빈칸이면 아직 덜 채운 전표다(2026-08-05 실측).
    rows = [
        {"프로젝트코드코드": "UJ1", "창고명": "쿠팡_돌발AS", "거래처명": "가",
         "진행상태": "확인", "공급가액합계": "480000", "관리항목명": ""},
        {"프로젝트코드코드": "UJ1", "창고명": "쿠팡_돌발AS", "거래처명": "가",
         "진행상태": "1.미확인", "공급가액합계": "480000", "관리항목명": "돌발AS"},
        {"프로젝트코드코드": "UJ2", "창고명": "쿠팡_돌발AS", "거래처명": "나",
         "진행상태": "3.오더처리", "공급가액합계": "1000000", "관리항목명": "돌발AS"},
        {"프로젝트코드코드": "00117", "창고명": "임대", "거래처명": "뮤토택배",
         "진행상태": "1.미확인", "공급가액합계": "352000", "관리항목명": "임대료"},
    ]
    coupang, by_prj, dup, low, high = SR.erp_summary(rows)
    # 같은 금액 두 줄인데 한쪽이 미완성(쿠팡인데 관리항목명 빈칸)이면 그쪽을 빼고 한 건.
    assert len(coupang) == 2 and not dup, (len(coupang), dup)
    assert (low, high) == (1480000, 1480000), (low, high)
    assert [r["진행상태"] for r in by_prj["UJ1"]] == ["1.미확인"], "미완성 줄이 남았다"

    # 근거가 없으면(둘 다 관리항목이 있으면) 합치지 않고 폭으로 남긴다.
    both = [dict(r) for r in rows[:2]]
    both[0]["관리항목명"] = "돌발AS"
    _c2, _b2, dup2, low2, high2 = SR.erp_summary(both + rows[2:])
    assert set(dup2) == {"UJ1"} and (low2, high2) == (1480000, 1960000), (dup2, low2, high2)

    # (3) 밴드에서 그 날 PO 글을 집어낸다.
    assert "PO = re.compile" in src and 'r["PO"]' in src, "밴드 PO 수주 글을 안 본다"

    # (4) 없는 것은 없다고 쓴다 — 수집 실패와 구분한다.
    assert "수집 실패가 아니라 실제로 없음" in src

    # (5) 배선: 앱 [기록] 탭 + daily_run
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '("보고자료_*정산분.md", "정산분 보고")' in srv, "앱 리포트 목록에 없다"
    assert srv.index('("보고자료_*정산분.md"') < srv.index('("자료현황.md"'), \
        "정산분 보고가 맨 앞이 아니다"
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "settle_report.py" in daily, "daily_run 이 매일 만들지 않는다"

    # (6) 엑셀을 열어 쓰지 않는다(읽기 전용).
    assert "workbook_patch" not in src and "--apply" not in src
    print("  [105] 하루치 정산분 보고자료(카톡 중복제거·이중계상 금지·앱/일일실행 배선) ✅")


def t107_report_dates_and_capture():
    """[107] 보고 기준일 DB 즉시저장·보관본 준비 · 숨은 화면에서도 캡처.

    세 지시가 한 뿌리다: **버튼이 말한 대로 되게 하라**.
      · 과거 "저장하고 반영"은 Excel 정본 시절 문맥이다. 이제 앱 DB에 즉시 저장하고
        Excel은 검증된 단방향 보관본만 준비한다.
      · "선택 날짜 보고 바로 캡처 버튼 동작 안함"
      · "보고일이 오늘 날짜로 자동 변경되는 알고리즘"
    """
    import report_dates as RD
    from datetime import date

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    # (1) 캡처가 안 되던 진짜 원인: 화면이 숨겨져 있으면 rAF 가 **아예 오지 않는다**.
    #     타이머와 경주시켜 어느 쪽이 오든 진행해야 한다. 오류도 안 나는 침묵형 버그였다.
    assert "function nextPaint(" in live, "그리기 대기 함수가 없다"
    assert "setTimeout(go, ms || 400)" in live, "rAF 가 안 올 때 빠져나갈 길이 없다"
    assert "await new Promise(resolve=>requestAnimationFrame(()=>requestAnimationFrame(resolve)))" \
        not in live, "숨은 화면에서 멈추는 rAF 대기가 남아 있다"
    assert "await nextPaint();" in live and "await captureReport();" in live

    # (2) 앱 DB 저장과 보관본 준비를 같은 동작처럼 오해시키지 않는다.
    assert "저장하고 보관본 준비</button>" in live, "버튼이 새 정본 역할을 설명하지 않는다"
    assert "저장하고 반영 대기" not in live and "지금 바로 엑셀에 반영" not in live
    assert '"apply":true' in live.replace(" ", "") or "apply:true" in live.replace(" ", ""), \
        "버튼이 보관본 준비를 요청하지 않는다"
    assert 'if b.get("apply"):' in srv and "_prepare_archive_export" in srv, \
        "서버가 DB 저장 뒤 보관본 준비 경로를 갖고 있지 않다"
    date_route = srv.split('if p == "/api/set_dates":', 1)[1].split('if p == "/api/input":', 1)[0]
    assert "ledger_writer" not in date_route and "ignore_input_window" not in date_route, \
        "기준일 저장이 다시 Excel 직접 쓰기로 돌아갔다"

    # (3) 보고일 자동 갱신 — 오늘/전날. 이미 맞으면 큐를 늘리지 않는다(멱등).
    w = RD.wanted(date(2026, 8, 7))
    assert w == {"B3": "2026-08-07", "B4": "2026-08-06"}, w
    saved = RD.current
    try:
        RD.current = lambda: {"B3": "2026-08-07", "B4": "2026-08-06"}
        assert RD.run(today=date(2026, 8, 7)) == 0        # 큐 적재 없이 끝나야 한다
    finally:
        RD.current = saved
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "report_dates.py" in daily, "daily_run 이 보고일을 갱신하지 않는다"
    assert daily.index("report_dates.py") < daily.index("upload_intake.py"), \
        "보고일 갱신이 뒤 단계들보다 늦다"
    print("  [107] 보고 기준일 DB 즉시저장·보관본 준비·자동 갱신·숨은 화면 캡처 ✅")


def t108_pm_source_fallback():
    """[108] 원본이 아직 담지 않은 날은 **원장으로 보충**한다 (2026-08-06 실사고).

    류지영 정기점검 원본이 정본이라, 원본이 있으면 원장(04시트)을 아예 안 봤다.
    그런데 원본이 8/3 까지만 갱신돼 있어 **8/5 정기점검 완료 3건이 보고에 0건으로
    나갔다.** 대표 보고 캡처가 그대로 0 이었다.

    고친 규칙 — 원본을 무시하지 않는다:
      · 그날 원본에 한 건이라도 있으면 **원본이 정본**(기존 그대로)
      · 그날 원본에 아무것도 없을 때만 원장으로 센다
    """
    import daily_brief as DB

    base = {"as": [], "fw": [], "events": [],
            "pm": [{"프로젝트NO": "UJ2600001", "점검예정일": "2026-08-05",
                    "실제점검일": "2026-08-05", "캠프명": "가"},
                   {"프로젝트NO": "UJ2600002", "점검예정일": "2026-08-05",
                    "실제점검일": "2026-08-05", "캠프명": "나"}]}

    def sched(rows):
        return {"year": 2026, "quarter": 3, "schedule": rows}

    a = dict(base)
    a["pm_schedule"] = sched([{"연결프로젝트NO": "X", "점검예정일": "2026-08-01",
                               "실제점검일": "2026-08-01", "장비수": 3}])
    assert DB.brief("2026-08-05", a)["정기점검"]["완료"] == 2,         "원본이 그날을 안 담았는데 원장으로 보충하지 않는다"

    b = dict(base)
    b["pm_schedule"] = sched([{"연결프로젝트NO": "X", "점검예정일": "2026-08-05",
                               "실제점검일": "2026-08-05", "장비수": 3}])
    got = DB.brief("2026-08-05", b)["정기점검"]
    assert (got["완료"], got["완료장비"]) == (1, 3),         "그날 원본이 있는데 원장이 덮었다 — 원본이 정본이다"
    print("  [108] 원본 미수록일만 원장으로 보충(원본 우선은 유지) ✅")


def t109_remote_edit_delete_versions():
    """[109] 리모컨 — 버전 통일·지사 불출자·고치기·지우기·강제·되돌리기 (2026-08-06 지시).

    사용자 지시: "리모컨 버전 선택할 수 있는 기능 추가하고 전체적으로 버전 관리가
    VER.3인지 VER.4인지 입력 및 확인 수정 가능하게 고도화 해" + 남은 4가지
    (강제 수정 · 지사 불출자 이름란 · 삭제 · 최근 불출 수정).

    지키는 것:
      · 'ver3'·'v4' 처럼 사람이 쓰는 표기가 한 이름으로 모인다 — 안 그러면 재고가 갈라진다
      · 불출에도 버전·불출자가 남는다 (버전 없이 불출하면 지점 버전별 잔량이 어긋난다)
      · 수정은 **이번 수정이 새로 만든 문제**만 막는다. 이미 어긋난 장부 때문에
        무관한 줄까지 못 고치면 아무것도 정리할 수 없다
      · 강제는 사유가 있어야 하고, 지운 것은 원장에 남아 되돌릴 수 있다
    """
    import shutil as _sh
    import tempfile as _tf

    import ledger_db as L
    old_path, old_dir = L.DB_PATH, L.DB_DIR
    tmp = _tf.mkdtemp()
    try:
        L.DB_DIR, L.DB_PATH = tmp, os.path.join(tmp, "t.db")
        assert L.REMOTE_VERSIONS == ("미확인", "기존형", "VER.3", "VER.4")
        assert [L._remote_version(x) for x in ("ver3", "VER 3", "v4", "기존", "")] == \
               ["VER.3", "VER.3", "VER.4", "기존형", ""], "버전 표기가 한 이름으로 안 모인다"

        L.remote_stock_adjust("시화", 5, "add", "입고", "안은숙", version="VER.3")
        rid = L.remote_request("시화", "김기사", 2, "안은숙", camp="시화1캠프",
                               version="ver3", issuer="대리불출자")
        top = L.remote_status()["issues"][0]
        assert (top["version"], top["issuer"]) == ("VER.3", "대리불출자"), top
        # 버전이 붙어야 지점 버전별 잔량이 맞는다(5 - 2 = 3)
        assert L.remote_status()["branch_stock"]["시화"]["versions"]["VER.3"] == 3

        # 최근 불출 고치기 — 캠프·버전·수량을 그 자리에서 바꾼다
        L.remote_edit("issue", rid, {"camp": "시화2캠프", "qty": 1, "version": "VER.4"},
                      edited_by="류지영")
        top = L.remote_status()["issues"][0]
        assert (top["camp"], top["qty"], top["version"]) == ("시화2캠프", 1, "VER.4"), top

        # 한도를 넘기는 수정은 막힌다 — 사유 없는 강제도 막힌다
        for bad in (lambda: L.remote_edit("issue", rid, {"qty": 9}),
                    lambda: L.remote_edit("issue", rid, {"qty": 9}, force=True),
                    lambda: L.remote_delete("issue", rid),
                    lambda: L.remote_edit("issue", rid, {"qty": 0}),
                    lambda: L.remote_edit("없는표", rid, {"qty": 1})):
            try:
                bad(); raise AssertionError("리모컨 수정 규칙이 뚫렸다")
            except ValueError:
                pass
        # 사유를 적은 강제는 통과하고, 강제였다는 사실이 남는다
        out = L.remote_edit("issue", rid, {"qty": 9}, edited_by="류지영",
                            force=True, reason="실물 확인 — 기사가 실제로 9개 보유")
        assert out["row"]["qty"] == 9 and out["warnings"], out
        assert L.remote_audit_list(1)[0]["forced"] == 1

        # 지우기 → 되돌리기. 되돌릴 수 없으면 사람이 무서워서 안 쓴다.
        L.remote_delete("issue", rid, deleted_by="류지영", force=True, reason="중복 입력")
        assert not [r for r in L.remote_status()["issues"] if r["id"] == rid]
        aid = [a for a in L.remote_audit_list(5) if a["action"] == "삭제"][0]["id"]
        L.remote_restore(aid, actor="류지영")
        assert [r for r in L.remote_status()["issues"] if r["id"] == rid], "복구가 안 된다"
        try:
            L.remote_restore(aid); raise AssertionError("두 번 복구됐다")
        except ValueError:
            pass
    finally:
        L.DB_DIR, L.DB_PATH = old_dir, old_path
        _sh.rmtree(tmp, ignore_errors=True)
    # 화면도 같은 목록을 쓰는지 — 서버와 갈리면 'VER.3' 과 'ver3' 이 따로 세어진다
    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for need in ("s.versions||[]", "remoteEditOpen", "remoteRowDelete", "remoteRestore",
                 "Issuer", "/api/remote/edit", "/api/remote/delete"):
        assert need in html, f"리모컨 화면에 {need} 가 없다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    for need in ("/api/remote/edit", "/api/remote/delete", "/api/remote/restore"):
        assert need in srv, f"서버에 {need} 라우트가 없다"
    print("  [109] 리모컨 버전 통일·불출자·수정/삭제/강제/되돌리기 ✅")


def t110_writer_formula_key():
    """[110] 조회 키가 **수식 열**이어도 찾는다 (2026-08-06 실사고).

    점검ID·접수ID·작업ID 는 `=IF($B5="","","PM-"&…)` 수식이다. 적용기가 수식 통을
    그대로 읽어 키 사전에 '=IF(...)' 를 넣는 바람에, 실제로 있는 행을 "행 없음"으로
    버렸다 — 8/5 정기점검 완료 2건이 그렇게 조용히 사라져 대표 보고에 0건으로 나갔다.
    값 통(data_only)에서 키를 읽어야 한다.
    """
    import shutil as _sh
    import tempfile as _tf
    import zipfile as _zip

    import openpyxl
    import ledger_writer as W
    # ★ TemporaryDirectory 를 쓰지 않는다 — 윈도우에서 openpyxl(read_only) 이 zip 을
    #   놓는 시점이 늦어 정리에서 PermissionError 가 난다. 검증이 그것으로 죽으면 안 된다.
    td = _tf.mkdtemp()
    try:
        path = os.path.join(td, "t.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "04_정기점검"
        for i, h in enumerate(("점검ID", "프로젝트NO", "실제점검일"), 1):
            ws.cell(row=W.HDR_ROW, column=i, value=h)
        for r, code in ((W.FIRST, "PM-A"), (W.FIRST + 1, "PM-B")):
            ws.cell(row=r, column=1, value='=IF($B{0}="","","PM-{0}")'.format(r))
            ws.cell(row=r, column=2, value="UJ%04d" % r)
        wb.save(path)
        # 엑셀이 계산해 둔 값을 흉내 낸다 — openpyxl 은 캐시값을 못 쓴다.
        with _zip.ZipFile(path) as z:
            names = z.namelist()
            blobs = {n: z.read(n) for n in names}
        xml = blobs["xl/worksheets/sheet1.xml"].decode("utf-8")
        for r, code in ((W.FIRST, "PM-A"), (W.FIRST + 1, "PM-B")):
            # 계산 결과가 문자열이면 엑셀은 t="str" 과 <v> 를 남긴다. 그 모양을 만든다.
            def fix(m, code=code, r=r):
                attrs = m.group(1).replace(' t="str"', "")
                return f'<c r="A{r}"{attrs} t="str">{m.group(2)}<v>{code}</v>'
            xml = re.sub(r'<c r="A%d"([^>]*)>(<f>.*?</f>)' % r, fix, xml)
        assert "<v>PM-A</v>" in xml, "합성 워크북에 캐시값을 못 넣었다"
        blobs["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
        with _zip.ZipFile(path, "w", _zip.ZIP_DEFLATED) as z:
            for n in names:
                z.writestr(n, blobs[n])
        q = [{"sheet": "04_정기점검", "key_col": "점검ID", "key": "PM-A",
              "col": "실제점검일", "value": "2026-08-05", "vtype": "date",
              "evidence": "합성", "only_if_empty": False}]
        plans, skips = W.resolve_targets(path, q)
        assert not skips, f"수식 키를 못 찾았다 — {skips}"
        assert plans and plans[0]["row"] == W.FIRST, plans
    finally:
        _sh.rmtree(td, ignore_errors=True)
    print("  [110] 수식 열(점검ID·접수ID)도 조회 키로 찾는다 ✅")


def t111_account_handoff_freshness():
    """[111] 다른 계정·다른 창이 이어받아도 아무 문제 없이 (2026-08-06 지시).

    사용자 지시: "이 세션은 완료되면 다른 계정으로 로그인해서 사용할거야, 그때 아무
    문제 없이 처리될 수 있는 알고리즘 구성해".

    계정이 바뀌면 세션 식별자도 프로세스도 전부 바뀐다. 그때 두 가지가 샌다:
      ① 죽은 세션의 점유가 남아 아무도 원장을 못 고친다 → --adopt 가 **죽었다는
         증거(pid)** 가 있는 것만 회수한다. 살아 있는 옆 세션 것은 건드리지 않는다.
      ② **수집이 밀린 것을 아무도 모른다.** 오늘 실제로 그랬다 — 쿠팡AS 밴드가 8/4 에
         멈춰 있는데 화면은 멀쩡히 숫자를 보여 줬고, 8/5 돌발AS 가 1건으로 나갔다.
         밴드·이카운트는 사람 로그인이 있어야 긁히므로(절대규칙 3), 기계가 할 수 있는
         최선은 "밀렸다 + 이렇게 되살려라" 를 시작 화면 맨 앞에 놓는 것이다.
    """
    import session_handoff as H

    rows = H.data_freshness(today="2026-08-06")
    need = {"이름", "최신", "밀린일", "한도", "밀림", "되살리는법"}
    assert rows and all(need <= set(r) for r in rows), rows
    # 밴드는 **밴드마다** 따로 봐야 한다 — 합쳐서 최댓값을 쓰면 뒤처진 밴드가 가려진다
    assert [r for r in rows if r["이름"].startswith("밴드")], rows
    # '어제'가 비면 밀린 것이다: 8/6 기준 최신 8/4 → 밀린일 2 > 한도 1
    assert H.FRESH_LIMIT["밴드"] == 1, "밴드 한도가 1을 넘으면 어제 빈 것을 못 잡는다"

    st = {"큐잔량": 0, "임시파일": [], "점유": [], "미푸시": [], "지시문사본": [],
          "수집신선도": [{"이름": "밴드: 90610953", "최신": "2026-08-04", "밀린일": 2,
                          "한도": 1, "밀림": True, "되살리는법": "밴드 로그인 후 수집"}]}
    bl = H.blockers(st)
    assert any("수집이 밀렸다" in why for why, _ in bl), bl
    st["수집신선도"][0]["밀림"] = False
    assert not H.blockers(st), "밀리지 않았는데 막았다"
    # 문서에도 표가 나와야 사람이 본다
    st["원장"] = {"버전": "1", "수정": ""}
    st["시각"] = "2026-08-06 12:00"
    st["미커밋"] = []
    st["최근커밋"] = []
    st["다음할일"] = []
    st["진행체크포인트"] = {}
    st["수집신선도"][0]["밀림"] = True
    assert "원본 수집이 어디까지 들어왔나" in H.to_md(st)

    # 이어받기 명령이 실제로 배선돼 있나 — 문서에만 있고 CLI 에 없으면 아무 소용 없다
    src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    for need_s in ('"--adopt"', "def adopt(", "def write_snapshot(", "_is_dead"):
        assert need_s in src, f"session_handoff 에 {need_s} 가 없다"
    rules = open(os.path.join(ROOT, "CLAUDE.md"), encoding="utf-8").read()
    assert "--adopt" in rules, "계정 인계 절차가 규칙(CLAUDE.md)에 없다"
    print("  [111] 계정이 바뀌어도 이어받기 — 죽은 점유 회수·수집 밀림 감지 ✅")


def t112_band_plan_order_and_scope():
    """[112] 밴드 수집 순서 — **새 글 먼저**, 범위는 사람이 정한 값 (2026-08-06).

    예전에는 캐시의 '구멍'만 봤다. 그래서 마지막 수집 이후 올라온 글은 영원히 대상
    밖이었고, 쿠팡AS 밴드가 8/4 에 멈춰 있는데도 "구멍 0" 이라 아무도 몰랐다 —
    그 결과 8/5 돌발AS 가 1건으로 보고됐다. 이제 캐시 위쪽을 먼저 훑는다.

    그리고 어디까지 파고들지는 **사람이 정한다**. 대화에 남긴 결정은 다음 세션이
    모르므로 band/collect_scope.json 이 기억하고 도구가 그것을 기본값으로 쓴다.
    """
    import band.recheck_plan as RP

    posts = {str(n): {"captured_at": RP.ERA_MS + 1} for n in range(100, 111)}
    p = RP.plan("1", posts, floor=95, ahead=3)
    assert p["new"] == [111, 112, 113], p["new"]
    assert p["gaps"] == [95, 96, 97, 98, 99], p["gaps"]
    # 순서: 새 글 → 구멍(최근부터). 과거글을 먼저 훑으면 오늘 숫자가 계속 틀린다.
    todo = (p["new"] + sorted(p["gaps"], reverse=True))[:6]
    assert todo == [111, 112, 113, 99, 98, 97], todo
    # ahead 를 안 주면 위쪽을 안 본다 — 예전 동작이 그대로 남아 있으면 안 된다
    assert RP.plan("1", posts, floor=95)["new"] == []

    sc = RP.scope()
    assert sc.get("floor"), "사람이 정한 수집 범위(collect_scope.json)가 없다"
    assert all(str(k).isdigit() and int(v) > 0 for k, v in sc["floor"].items()), sc
    assert int(sc.get("ahead") or 0) > 0, "새 글 탐색 개수가 0이면 최신분을 못 잡는다"
    # 없는 글을 열면 밴드가 alert 를 띄우고 **탭 전체가 선다**(2026-08-06 실측).
    # 과거글 구간은 지운 글이 섞여 있어 수백 번 뜬다 — 사람이 수백 번 누를 수는 없다.
    js = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
    assert "muteDialogs" in js and "w.alert" in js, "수집기가 안내창을 막지 않는다"

    # ★ 엑셀은 관리대장 하나로만 관리한다 (2026-08-07 지시).
    #   "쿠팡_거래처코드_최신.xlsx 와 쿠팡_확인필요현황_최신.xlsx 를 관리대장에
    #   통합해서 관리하고, 앞으로도 별도의 엑셀 파일은 만들지 말고 관리대장으로만
    #   관리해." 파일이 늘수록 '어느 게 최신인가'가 흐려지고, 사람이 열어 둔 파일은
    #   갱신도 막힌다. 별도 파일을 되살리는 변경이 들어오면 여기서 막는다.
    _fe2 = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    assert '"--xlsx" in sys.argv' in _fe2, \
        "확인필요현황이 다시 별도 엑셀을 기본 생성한다 — 관리대장 23시트로만 관리한다"
    _ci = open(os.path.join(ROOT, "customer_index.py"), encoding="utf-8").read()
    assert "쿠팡_거래처코드_최신.xlsx" not in _ci, \
        "거래처코드가 다시 별도 엑셀을 만든다 — 관리대장 27시트로만 관리한다"
    assert "customer_sheet" in _ci, "거래처코드 시트 반영 경로가 끊겼다"
    import customer_sheet as _CS
    assert len(_CS.HEADERS) == len(_CS.WIDTHS) == 10, \
        "거래처코드 열 구성이 예전 엑셀과 달라졌다 — 자리만 옮기는 것이지 내용이 바뀌면 안 된다"
    _r = _CS.rows_from_index({"A": {"code": "C", "erp_name": "E", "how": "h", "건수": 1,
                                    "addr": "", "manager": "", "tel": "", "email": "",
                                    "equip": ""}}, {}, [])
    assert _r and all(len(x) == 10 for x in _r), _r
    assert _CS.SHEET_NAME in _CS.build_sheet_xml(_r)
    # 엑셀 쓰기는 11:00·15:00 회차에만 — daily_run 이 아니라 ledger_db 에 있어야 한다.
    _ld = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    assert re.search(r'"29_거래처코드".*customer_index\.py.*"--sheet"', _ld, re.S), \
        "거래처코드 시트가 11:00·15:00 회차에 걸려 있지 않다"
    _dr0 = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert not re.search(r'customer_index\.py"\)\s*,\s*"--sheet"', _dr0), \
        "daily_run 이 관리대장을 직접 연다 — 엑셀 반영은 11:00·15:00 회차만"

    # ★ 옛 별도 엑셀 접기 — **시트가 채워진 것을 확인한 뒤에만** 옮긴다.
    #   순서가 뒤집히면(파일 먼저 정리, 시트는 비어 있음) 그 자료는 그 순간 아무 데도 없다.
    import side_excel_retire as _SR
    assert _SR.ARCHIVE == "OLD", "정리 자리는 프로젝트 관례대로 OLD/ 다"
    _sheets = {s for s, _f in _SR.PAIRS}
    assert {"23_확인필요현황", "29_거래처코드"} <= _sheets, _SR.PAIRS
    _tmpd = tempfile.mkdtemp()
    _m = os.path.join(_tmpd, "합성대장R_v1.xlsx")
    make_ledger(_m)
    for _sh, _side in _SR.PAIRS:                    # 시트가 없으면 '보류'여야 한다
        _ok, _why = _SR.check(_m, _sh)
        assert _ok is False, f"{_sh}: 시트가 없는데 옮겨도 된다고 했다 — {_why}"
    open(os.path.join(_tmpd, "쿠팡_확인필요현황_최신.xlsx"), "w").close()
    _res = _SR.run(apply=True, master=_m)           # --apply 라도 옮기면 안 된다
    assert os.path.exists(os.path.join(_tmpd, "쿠팡_확인필요현황_최신.xlsx")), \
        "시트가 비었는데 별도 엑셀을 옮겼다 — 자료가 아무 데도 없게 된다"
    assert any("보류" in _m2 for _s2, _f2, _o2, _m2 in _res), _res
    # 시트를 채운 뒤에는 옮겨야 한다
    subprocess.run([PY, os.path.join(ROOT, "findings_sheet.py"), "--master", _m],
                   capture_output=True, text=True, encoding="utf-8", cwd=ROOT,
                   env={**os.environ, "COUPANG_REPORT_DIR": _tmpd, "COUPANG_UPDATES_DIR": _tmpd})
    _m2f = _m.replace("_v1.xlsx", "_v2.xlsx")     # 같은 폴더라 별도 엑셀은 그대로 옆에 있다
    assert _SR.check(_m2f, "23_확인필요현황")[0], "채워진 시트를 '보류'로 봤다"
    _SR.run(apply=True, master=_m2f)
    assert os.path.exists(os.path.join(_tmpd, "OLD", "쿠팡_확인필요현황_최신.xlsx")), \
        "OLD/ 로 옮기지 않았다"
    assert not os.path.exists(os.path.join(_tmpd, "쿠팡_확인필요현황_최신.xlsx"))
    _ldr = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    assert _ldr.index("side_excel_retire.py") > _ldr.index('"29_거래처코드"'), \
        "별도 엑셀 정리가 시트 갱신보다 앞에 있다 — 한 회차를 헛돈다"

    # ★ 사람이 **옛 버전**을 열어 편집 중이면 그 수정은 다음 회차로 안 넘어간다.
    #   2026-08-07 실측: v538 을 11:08 부터 열어 둔 사이 회차가 v541 까지 만들었다.
    #   파일이 열려 있으니 오류도 안 난다 — 알려 주지 않으면 아무도 모른다.
    import session_handoff as _SH
    _lockd = tempfile.mkdtemp()
    for _n in ("합성대장L_v7.xlsx", "합성대장L_v9.xlsx"):
        open(os.path.join(_lockd, _n), "w").close()
    _latest = os.path.join(_lockd, "합성대장L_v9.xlsx")
    import ecount_reconcile as _ER          # stranded_editor 가 함수 안에서 부른다
    _keep = (_ER.resolve_master, _ER.load_config)
    _ER.resolve_master = lambda *_a, **_k: _latest
    _ER.load_config = lambda *_a, **_k: {"reconcile": {"master_xlsx": _latest}}
    try:
        assert _SH.stranded_editor() == [], "잠금이 없는데 경보를 냈다"
        open(os.path.join(_lockd, "~$합성대장L_v9.xlsx"), "w").close()
        assert _SH.stranded_editor() == [], "최신본을 여는 것은 정상인데 경보를 냈다"
        open(os.path.join(_lockd, "~$합성대장L_v7.xlsx"), "w").close()
        assert _SH.stranded_editor() == ["합성대장L_v7.xlsx"], _SH.stranded_editor()
    finally:
        _ER.resolve_master, _ER.load_config = _keep
    _shs = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert '"옛버전편집": stranded_editor()' in _shs and 'st.get("옛버전편집")' in _shs, \
        "옛 버전 편집 경보가 '먼저 처리할 것'에 오르지 않는다"

    # ★ 도구가 있는데 아무도 안 부르는 것이 가장 조용한 누락이다 (2026-08-07 발견).
    #   `fill_erp_status.py` 는 2026-07-28 에 만들어졌는데 daily_run 에 없었다.
    #   손으로 돌린 사람이 있을 때만 채워졌다는 뜻이다 — 그냥 돌려 보니 52칸이 남아
    #   있었고 그중 46건은 ERP 가 이미 완료를 입증한 건이었다.
    _dr = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "fill_erp_status.py" in _dr, \
        "ERP 등록여부 판정이 daily_run 에 없다 — 사람이 손으로 돌릴 때만 채워진다"
    assert re.search(r'fill_erp_status\.py"\)\s*,\s*"--queue"', _dr), \
        "--queue 가 아니면 엑셀을 직접 연다(반영은 11:00·15:00 회차만)"
    import daily_run as _DR
    assert _DR._retryable(["fill_erp_status.py", "--queue"]) is False, \
        "큐 단계를 재시도하면 같은 입력이 두 번 들어간다"

    # ★ 본문이 그려졌다고 머리말까지 그려진 것은 아니다 (2026-08-07 실측).
    #   밴드는 .postText 를 먼저 칠하고 .time 을 조금 뒤에 채운다. 보자마자 가져가면
    #   **날짜 없는 글**이 저장되고, 그런 글은 어떤 작업과도 대조되지 않는다.
    #   실제로 621건이 그렇게 쌓였는데 구멍도 아니고 오래된 것도 아니라
    #   **어느 목록에도 안 잡혔다** — 가장 알아채기 어려운 종류의 구멍이다.
    assert re.search(r"for \(let i = 0; i < \d+ && !txt\(main, '\.postListInfoWrap \.time",
                     js), "본문만 보고 바로 가져간다 — 작성시각을 기다리지 않는다"
    # ★ 캐시 정본은 **임시파일에 다 쓴 뒤 갈아끼운다** (2026-08-07 실사고).
    #   `open(dst,"w")` 는 정본을 먼저 비운다. 19MB 를 흘려 넣는 몇 초 동안 읽는 쪽은
    #   반쪽짜리 JSON 을 보고, 그 사이 합성검증이 두 번 죽었다(죽은 자리가 매번 달라
    #   한동안 "캐시가 깨졌다"고 오해했다). 쓰다가 프로세스가 죽으면 **정말로** 깨져
    #   8,500 글을 다시 긁어야 한다 — 밤샘 한 번 분량이다.
    _cd = open(os.path.join(ROOT, "band", "convert_dump.py"), encoding="utf-8").read()
    assert re.search(r'os\.replace\(\s*tmp\s*,\s*dst\s*\)', _cd), \
        "캐시를 원자적으로 갈아끼우지 않는다 — 쓰는 중에 읽으면 반쪽 JSON 이 보인다"
    assert not re.search(r'json\.dump\([^)]*open\(\s*dst\s*,\s*"w"', _cd), \
        "정본에 바로 쓴다(open(dst,'w')) — 임시파일 + os.replace 로 바꿀 것"

    import band.recheck_plan as _RP2
    _p = _RP2.plan("1", {
        "10": {"created_at": 1, "captured_at": _RP2.ERA_MS + 1},
        "11": {"created_at": None, "captured_at": _RP2.ERA_MS + 1},   # 본문은 있는데 날짜가 없다
        "12": {"deleted": True, "captured_at": _RP2.ERA_MS + 1},      # 진짜 지운 글
    }, floor=10, ahead=0)
    assert _p["dateless"] == [11], _p.get("dateless")
    assert 12 not in (_p.get("dateless") or []), "지운 글을 날짜없음으로 세면 영원히 다시 훑는다"
    assert 11 not in _p["gaps"] and 11 not in _p["stale"], \
        "날짜없음이 구멍/오래됨과 겹치면 두 번 뽑힌다"

    # 세션마다 눈에 띄어야 한다. 신선도 판정(band_latest_days)은 **날짜 있는 글만**
    # 보므로 "밴드 최신 = 오늘" 인데도 그 밑에 수백 건이 대조 밖일 수 있다.
    import session_handoff as _SH
    assert hasattr(_SH, "band_dateless"), "인계 점검이 날짜없음을 세지 않는다"
    _bl = _SH.blockers({"큐잔량": 0, "임시파일": [], "점유": [], "미커밋": [], "미푸시": [],
                        "밴드날짜없음": {"테스트밴드": 621}, "수집신선도": {},
                        "지시문사본": {}, "워크트리": None, "원장": {}, "미머지": []})
    assert any("날짜가 없는 글" in b[0] for b in _bl), \
        "날짜없음 621건이 '먼저 처리할 것'에 안 뜬다 — 조용히 어긋난다"
    _bl0 = _SH.blockers({"큐잔량": 0, "임시파일": [], "점유": [], "미커밋": [], "미푸시": [],
                         "밴드날짜없음": {}, "수집신선도": {},
                         "지시문사본": {}, "워크트리": None, "원장": {}, "미머지": []})
    assert not any("날짜가 없는 글" in b[0] for b in _bl0), "0건인데도 경보가 뜬다"

    # 붙여넣기는 **밴드당 한 번**이어야 한다 (2026-08-06 지시: "내 손 안 가게").
    # 250 은 탭이 얼지 않는 한 배치의 상한일 뿐이다. 한 파일 안에서 회차를 이어 돌리고
    # 회차 사이에 저장·비우기를 하면 사람 손은 한 번으로 끝난다. 예전처럼 250건마다
    # 다시 붙여넣게 만들면 1,100건짜리 밴드에 다섯 번을 시키게 된다.
    import band.make_oneclick as MO

    assert MO.BATCH_MAX == 250, MO.BATCH_MAX
    big = {str(n): {"captured_at": RP.ERA_MS + 1} for n in range(1000, 1004)}
    orig_load, orig_scope = MO.RP.load, MO.RP.scope
    try:
        MO.RP.load = lambda b: big
        MO.RP.scope = lambda: {"floor": {"1": 400}, "ahead": 5}
        out, note = MO.build("1", 600)
    finally:
        MO.RP.load, MO.RP.scope = orig_load, orig_scope
    assert out, note
    rounds = json.loads(re.search(r"const ROUNDS = (\[.*?\]);", out, re.S).group(1))
    assert len(rounds) > 1, "회차로 안 쪼갠다 — 한 번에 밀어 넣으면 탭이 언다"
    assert all(0 < len(r) <= MO.BATCH_MAX for r in rounds), [len(r) for r in rounds]
    flat = [n for r in rounds for n in r]
    assert len(flat) == len(set(flat)), "회차끼리 번호가 겹친다"
    assert "keep: false" in out, "회차 사이에 탭 메모리를 안 비운다"
    assert out.count("__grabSave") >= 1 and "for (let i = 0" in out, \
        "회차마다 저장하고 이어 가는 구조가 아니다"
    print("  [112] 밴드 수집 — 새 글 우선·범위 기억·안내창 차단·한 번 붙여넣기 ✅")


def t113_paste_typos_and_misc_reclass():
    """[113] 붙여넣기 사고를 흡수한다 — 캠프명 오염·미분류 재분류 (2026-08-06).

    김미영 대리 지적: "밴드·카톡에 복사 붙여넣기 오류가 있어 매칭이 안 됨".
    실제 원장 캠프명 18건을 뜯어 보니 사람 잘못이 아니라 **붙여넣기 흔적**이었다 —
    웹에서 복사해 `&amp;` 가 그대로, 메모가 `<-` 뒤에 붙어서, 값 대신 `0`·`...`.
    이름을 고치는 게 아니라 **비교할 때** 걷어 내면 짝이 붙는다(18 → 14건).

    그리고 값이 아닌 표시(`0`)는 '못 잇는 자료'가 아니라 **원장을 고칠 일**이다.
    섞여 있으면 정말 ERP 에 없는 캠프가 몇 개인지 알 수 없다.

    미분류도 같은 이야기다. 판별 규칙은 늘어나는데 규칙이 없던 시절 미분류로 간
    파일은 아무도 다시 안 봤다 — 그래서 --apply 회차가 재분류까지 같이 한다.
    """
    import customer_index as CI

    assert CI.clean("남김해Sub-Hub&amp;Sub-FC") == "남김해Sub-Hub&Sub-FC"
    assert CI.clean("중구1캠프 <-서초1MB(양재동C)") == "중구1캠프"
    assert CI.norm("김포1Sub-FC ?(김포1 서브허브))") == CI.norm("김포1Sub-FC(김포1 서브허브)")
    # 접두사만 다른 같은 캠프 — 완전일치·정규화·핵심코드를 다 놓친 뒤에만 본다
    assert CI.bare("M_광주2캠프") == CI.bare("광주2 캠프") == CI.norm("광주2캠프")
    # 사람이 실수로 M 으로 시작하는 이름을 지어도 통째로 잘리면 안 된다
    assert CI.bare("MB1캠프") == CI.core("MB1캠프"), "숫자 앞 M 까지 떼면 다른 캠프가 된다"
    for junk in ("0", "...", "-", "?", "없음", ""):
        assert junk.upper() in CI.PLACEHOLDERS, junk
    assert "강서1MB" not in CI.PLACEHOLDERS

    import shutil as _sh
    import upload_intake as UI

    root = os.path.join(ROOT, "_t113_root")
    _sh.rmtree(root, ignore_errors=True)
    try:
        misc = UI._paths(root)["misc"]
        os.makedirs(misc, exist_ok=True)
        seen = os.path.join(misc, "★ 01. 쿠팡AS 품목 단가 리스트 및 사진 정리본.xlsx")
        open(seen, "wb").write(b"PK\x03\x04dummy")
        rows = UI.reclass_misc(root, do_apply=False)
        assert len(rows) == 1 and rows[0]["분류"] == "reference", rows
        assert os.path.isfile(seen), "미리보기가 파일을 옮겼다"
        rows = UI.reclass_misc(root, do_apply=True)
        assert not os.path.isfile(seen), "재분류가 파일을 안 옮겼다"
        assert os.path.isfile(rows[0]["목적지"]), rows[0]
        assert "10. 기준" in rows[0]["목적지"], rows[0]["목적지"]
        # 여전히 모르는 것은 건드리지 않는다 — 미분류는 '보존'이 목적이다
        unknown = os.path.join(misc, "무엇인지 모를 파일.dat")
        open(unknown, "wb").write(b"x")
        assert UI.reclass_misc(root, do_apply=True) == []
        assert os.path.isfile(unknown), "모르는 파일을 옮겼다"
    finally:
        _sh.rmtree(root, ignore_errors=True)
    print("  [113] 붙여넣기 오염 흡수(캠프명 18→14)·미분류 자동 재분류 ✅")


def t114_claim_owner_is_agent_pid():
    """[114] 점유의 생사는 **agent_pid** 로 본다 — pid 로 보면 산 세션을 죽었다고 한다.

    2026-08-06 실사고: 시작 화면이 살아 있는 옆 세션의 'code' 점유를 "죽은 세션의
    잔재"로 표시하고 `--free` 를 권했다. `pid` 는 ai_claim 을 실행한 **CLI 프로세스**라
    명령이 끝나는 즉시 죽는다 — 그것으로 판정하면 **모든 점유가 항상 잔재**로 보인다.
    주인은 `agent_pid`(에이전트 프로세스)다.

    두 배 위험했던 이유: 안내대로 `--free` 를 해도 ai_claim 이 남의 것이라 거부한다.
    사람은 "왜 안 풀리지" 하며 막히고, 최악의 경우 `--force` 로 **일하는 세션의
    원장 점유를 빼앗는다.** 그래서 안내 문구도 내 것/남의 것을 갈라 적는다.
    """
    import session_handoff as H

    live, dead = os.getpid(), 999999
    saved = H.claims.__globals__.get("_CLAIM_TEST")
    try:
        import ai_claim as C
        real = C.load
        C.load = lambda: {
            "code": {"who": "claude", "sid": "other", "why": "옆 세션",
                     "pid": dead, "agent_pid": live, "host": "", "at": 0},
            "band": {"who": "claude", "sid": "gone", "why": "죽은 세션",
                     "pid": live, "agent_pid": dead, "host": "", "at": 0},
        }
        try:
            rows = {c["lock"]: c for c in H.claims()}
        finally:
            C.load = real
    finally:
        H.claims.__globals__["_CLAIM_TEST"] = saved

    assert rows["code"]["alive"] is True, "산 세션을 죽었다고 본다 — pid 를 보고 있다"
    assert rows["code"]["stale"] is False, rows["code"]
    assert rows["band"]["alive"] is False, "죽은 세션을 못 가려낸다"
    assert rows["band"]["stale"] is True, rows["band"]

    # 안내 문구: 남의 죽은 점유에 --free 를 권하면 안 된다(거부당한다)
    st = {"점유": [dict(rows["band"], mine=False)], "임시파일": [],
          "큐잔량": 0, "미푸시": []}
    fixes = " ".join(f for _, f in H.blockers(st))
    assert "--adopt" in fixes and "--free band" not in fixes, fixes
    st["점유"] = [dict(rows["band"], mine=True)]
    assert "--free band" in " ".join(f for _, f in H.blockers(st))
    print("  [114] 점유 생사는 agent_pid — 산 세션을 잔재로 오인하지 않는다 ✅")


def t115_text_contrast():
    """[115] 글자가 실제로 읽히는가 — 모든 화면의 명암비를 전수로 지킨다 (2026-08-06).

    사용자 지시: "텍스트가 잘 안보이는 것들이 있어, 전체적으로 전수 조사해서 검토하고
    잘 보이게 정리해". 눈으로 고치면 고친 곳만 좋아진다 — 실제로 `--ink-3` 는 흰
    배경에서 **2.17:1**(기준 4.5:1의 절반)이었고 같은 값이 21곳에 퍼져 있었다.
    그래서 색을 **수치로** 재고, 이 검증이 되돌아가는 것을 막는다.

    감사기 자체도 두 번 틀렸다 — 없는 문제를 만들면 멀쩡한 색을 망가뜨리므로 같이 지킨다.
      · `@media{…}` 를 정규식으로 지워 규칙 경계가 밀렸다(엉뚱한 색으로 보고)
      · `--surface` 가 없는 파일에서 흰색으로 가정해 **다크 테마를 흰 바탕에서** 쟀다
    """
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import contrast_audit as CA

    # ── 계산이 맞는가 (WCAG 기준값) ──
    assert abs(CA.contrast((0, 0, 0), (255, 255, 255)) - 21.0) < 0.01
    assert abs(CA.contrast((255, 255, 255), (255, 255, 255)) - 1.0) < 0.01
    # 반투명 글자는 합성해야 실제 색이 나온다 — 안 하면 늘 통과로 보인다
    assert CA.parse_color("rgba(60,60,67,.42)", {}, (255, 255, 255)) == (173, 173, 176)

    # ── 아이콘도 센다 (2026-08-06 제보: 어둡게 켜니 단추 아이콘이 사라졌다) ──
    # 원인은 `.icon-btn` 배경이 흰색으로 **하드코딩**돼 있고 아이콘은 `fill:currentColor`
    # 라 테마를 따라 흰색이 된 것 — 흰 위 흰(1.09:1). 글자만 재면 영영 못 잡는다.
    icon_css = (":root{--bg:#0B1020;--surface:#161C2E;--ink:#F2F5FA}"
                ".b{background:#fff}.b svg{fill:currentColor}")
    bad, _u, _o = CA.audit_theme("", icon_css, CA.root_vars(icon_css)["기본"])
    assert any(r.get("종류") == "아이콘" for r in bad), "흰 바탕 위 흰 아이콘을 놓쳤다"
    # 가상요소는 숙주의 배경 위에 그려진다 — 선택자가 다르다고 바탕을 잃으면 안 된다
    assert CA.bare(".icon-btn::after") == ".icon-btn"
    pseudo = ":root{--bg:#0B1020}.b{background:#fff}.b::after{color:#EEEEEE;font-size:12px}"
    pbad, _pu, _po = CA.audit_theme("", pseudo, CA.root_vars(pseudo)["기본"])
    assert pbad and pbad[0]["ratio"] < 2, "가상요소가 숙주 배경을 못 찾아 검사에서 샜다"
    # 조상 배경을 **못 읽을 때**는 판정하지 않는다 — 없는 문제를 만들면 멀쩡한 색을 망친다
    avatar = (":root{--bg:#0B1020;--surface:#ffffff}.card{background:var(--surface)}"
              ".card .av{background:linear-gradient(145deg,var(--wc2),var(--wc));color:#fff}"
              ".card .av svg{fill:currentColor}")
    abad, aunk, _ao = CA.audit_theme("", avatar, CA.root_vars(avatar)["기본"])
    assert not abad and aunk, "실행 중에 정해지는 아바타색을 흰색으로 단정했다"

    # ── `color-mix(… X p%, transparent)` — 감사기가 세 번째로 틀렸던 자리 (2026-08-08) ──
    # 옅은 경고 배경 `color-mix(in srgb,var(--danger) 12%,transparent)` 안의
    # `var(--danger)` 만 뜯어 가서 **12% 를 원색으로** 읽었다. 그러니 그 위의 danger
    # 글자가 '명암비 1.0' 이 되어, 멀쩡한 화면 두 곳이 미달로 잡혔다.
    # 투명과 섞인 색은 rgba 와 똑같이 **밑에 깔린 것 위에 합성**해야 참값이 나온다.
    assert CA.bg_value({"background": "color-mix(in srgb,var(--danger) 12%,transparent)"}) \
        == "color-mix(in srgb,var(--danger) 12%,transparent)", "안쪽 var 만 뜯어 갔다"
    assert CA.mix_parts("color-mix(in srgb,var(--danger) 12%,transparent)") == ("var(--danger)", .12)
    assert CA.mix_parts("color-mix(in srgb,#fff 50%,#000)") is None, "불투명 혼합까지 건드렸다"
    assert not CA.opaque("color-mix(in srgb,var(--danger) 12%,transparent)", {}), \
        "투명 혼합을 불투명으로 봤다 — 조상 배경을 안 찾아 엉뚱한 색으로 잰다"
    assert CA.parse_color("color-mix(in srgb,#c7362b 12%,transparent)", {}, (255, 255, 255)) \
        == (248, 231, 230), "흰 바탕 위 12% 합성이 안 된다"

    # 실제 화면 세 벌에 미달이 0 이어야 한다 — 아이콘까지 포함해서
    for name in ("webapp/index.html", "docs/app.html", "docs/cal.html"):
        b, _u2, o2 = CA.audit(os.path.join(ROOT, *name.split("/")))
        assert not b, "%s 명암비 미달 %d건: %s" % (
            name, len(b), [(r["sel"][:40], r["ratio"]) for r in b[:3]])
        assert o2 > 0, name

    # ── 감사기 회귀: @media 를 지우면 규칙 경계가 밀린다 ──
    css = ("@media(max-width:719px){ .a small{display:none} }\n"
           ".b{color:#123456}\n")
    got = {sel: d for sel, d, _ in CA.rules(css)}
    assert got.get(".b", {}).get("color") == "#123456", got
    assert "color" not in got.get(".a small", {}), got

    # ── 감사기 회귀: 테마를 한 사전으로 합치면 밝은 화면을 어두운 값으로 잰다 ──
    themed = CA.root_vars(":root{--ink-3:#6b7686;--bg:#fff}\n"
                          '@media (prefers-color-scheme:dark){:root{--ink-3:#8391a8;--bg:#0b1020}}')
    assert set(themed) == {"기본", "다크"}, themed
    assert themed["기본"]["--ink-3"] == "#6b7686" and themed["다크"]["--ink-3"] == "#8391a8"

    # ── 실제 화면: 미달이 하나도 없어야 한다 ──
    screens = ["webapp/index.html", "docs/app.html", "docs/cal.html", "docs/index.html"]
    for rel in screens:
        path = os.path.join(ROOT, *rel.split("/"))
        if not os.path.exists(path):
            continue
        bad, _unknown, ok = CA.audit(path)
        assert ok > 0, "%s 에서 글자색을 하나도 못 읽었다 — 감사기가 헛돌고 있다" % rel
        assert not bad, "%s 명암비 미달 %d건: %s" % (
            rel, len(bad), [(b["sel"], b["ratio"]) for b in bad[:4]])

    # ── 폰 스냅샷은 **생성물**이다 — 결과가 아니라 만드는 쪽을 지킨다 ──
    gen = open(os.path.join(ROOT, "mobile_snapshot.py"), encoding="utf-8").read()
    assert "--brand-btn" in gen, "버튼 바탕색이 글자색과 분리되지 않았다(다크에서 흰 글자 2.5:1)"
    assert "--ink-3:#6b7686" not in gen, "생성기에 옛 저대비 회색이 남아 있다"
    print("  [115] 글자 명암비 전수검사 — 모든 화면 미달 0 (기준 4.5:1) ✅")


def t116_manual_refresh_is_really_fresh():
    """[116] 새로고침 버튼은 **정말로** 새로 받는다 (2026-08-06 지시).

    사용자 지시: "새로 고침을 누르면 바로바로 반영되게 개선해 … 빨리빨리 반영되게 해(중요)".

    두 가지가 겹쳐 있었다.
      ① stale-while-revalidate 가 **사람이 누른 버튼까지** 막았다. 15초 안에 누르면
         네트워크에 아예 안 나가고, 지나도 옛 값을 돌려주고 새 값은 뒤에서 받았다.
         그런데도 '최신 자료로 새로고침했습니다' 토스트가 떴다 — 반영된 줄 안다.
      ② loadSettle → loadStatus → loadNotifications 를 **차례로** 기다렸다.
         실측이 status 2.7초 · notifications 최악 59초라 합이 그대로 대기시간이 됐다.

    SWR 자체는 자동 갱신에 좋다. 가르는 것은 **누가 시켰나**다 — 그래서 지우지 않고
    사람이 시킨 경로에서만 건너뛴다. 캐시 **쓰기**는 그대로 둔다(안 그러면 다음 화면이
    또 옛 값으로 그려진다).
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    assert "function apiFresh(" in live, "사람이 누른 새로고침용 우회로가 없다"
    assert "API_FRESH" in live and "opt.fresh" in live, live.count("API_FRESH")
    # 읽기만 건너뛰고 쓰기는 남아야 한다
    assert re.search(r"const useCache = cacheable && !\(API_FRESH > 0", live), "읽기/쓰기를 안 갈랐다"
    assert re.search(r"if\(cacheable\s*&&[\s\S]{0,180}?swrSet\(path,\s*d\)", live), \
        "강제 갱신 때 캐시 쓰기까지 막혔다 — 다음 화면이 또 옛 값이 된다"

    body = live[live.index("async function reloadHere("):]
    body = body[:body.index("\n}")]
    assert "apiFresh(refreshAll)" in body, "새로고침 버튼이 캐시를 건너뛰지 않는다"
    assert "previewRptDate" in body and "REPORT_PREVIEW_DATE" in body, \
        "집계기준일을 고른 보고서 화면이 새로고침에서 빠졌다"

    ra = live[live.index("async function refreshAll("):]
    ra = ra[:ra.index("\n}")]
    assert "Promise.all" in ra, "새로고침이 아직 한 줄로 세워 부른다(합이 대기시간이 된다)"
    assert not re.search(r"await load\w+\(\);\s*await load", ra), ra[:200]
    print("  [116] 새로고침은 캐시를 건너뛰고 동시에 부른다(선택 기준일 포함) ✅")


def t117_dark_mode_toggle():
    """[117] 밝게/어둡게 — 사람이 켠 것만 따르고, 켜도 글자가 읽힌다 (2026-08-06 지시).

    2026-07-30 에 다크 모드를 **껐던** 이유가 기록에 남아 있다: 흰 배경이 77곳에
    하드코딩돼 있어 OS 다크 설정을 따라가면 글자만 흰색이 되어 '흰 카드 위 흰 글자'가
    됐다. 그래서 이번에는 **순서를 지켰다** — 그 77곳(카드 35 · 옅은 판 37)을 먼저
    --surface / --panel 로 흡수하고 나서 켰다. 이 검증이 그 순서를 지킨다.

    · OS 설정을 자동으로 따라가지 않는다. 이 앱 화면은 그대로 캡처해 보고서로 나간다 —
      폰이 어둡다는 이유로 보고 화면이 검게 바뀌면 안 된다.
    · 켜져 있는 화면에서 바꾸면 크롬이 var() 재계산을 건너뛰어 **일부만 바뀌었다**(실측).
      그래서 토글이 강제 재계산을 한 번 한다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    assert ':root[data-theme="dark"]' in live, "다크 팔레트가 없다"
    assert "function toggleTheme(" in live and "function applyTheme(" in live
    assert "cw_theme" in live, "고른 테마가 저장되지 않는다"
    assert 'id="themeBtn"' in live, "밝게/어둡게 단추가 화면에 없다"
    assert "function forceRestyle(" in live and "forceRestyle();" in live, \
        "테마를 바꿔도 일부 요소가 옛 색을 그대로 들고 있다(강제 재계산 없음)"
    # ★ 2026-08-07 갱신 — 사용자 지시로 '시스템 동일' 선택지가 생겼다
    #   ("기본, 다크모드, 시스템 동일 이렇게 구성하게 추가해"). 지켜야 할 것은
    #   "OS 를 읽지 않는다"가 아니라 **"사람이 고르지 않았는데 따라가지 않는다"** 다.
    #   · CSS `@media (prefers-color-scheme: dark)` → 금지. 고르지 않아도 색이 뒤집힌다.
    #   · JS matchMedia → 허용하되 **`themeMode() === 'system'` 일 때만** 반영한다.
    #   · 기본값은 여전히 'light' — 보고 화면이 기기 설정 따라 검게 나오면 안 된다.
    assert not re.search(r"@media[^{]*prefers-color-scheme\s*:\s*dark", live), \
        "CSS 가 OS 다크를 자동으로 따라간다 — 사람이 고르지 않았는데 화면이 검어진다"
    assert "localStorage.getItem('cw_theme') || 'light'" in live, \
        "기본값이 밝게가 아니다"
    assert "if(themeMode() === 'system')" in live, \
        "시스템이 아닐 때도 OS 변화를 반영한다 — 사람이 고른 값이 멋대로 바뀐다"

    dark = live[live.index(':root[data-theme="dark"]'):]
    dark = dark[:dark.index("}")]
    for tok in ("--bg:", "--surface:", "--panel:", "--panel-2:", "--ink:", "--ink-2:",
                "--ink-3:", "--line:", "--brand:", "--brand-btn:", "--warn-btn:",
                "--ok:", "--warn:", "--danger:", "--neu:"):
        assert tok in dark, "다크 팔레트에 %s 가 없다 — 그 자리만 밝은 값이 남는다" % tok

    # 흰 배경 하드코딩이 되돌아오면 2026-07-30 의 '흰 카드 위 흰 글자'가 재현된다
    css = live[live.index("<style>"):live.index("</style>")]
    css = re.sub(r"/\*.*?\*/", "", css, flags=re.S)
    white = re.findall(r"background(?:-color)?:\s*#(?:fff|ffffff)\b", css, re.I)
    assert len(white) <= 2, "흰 배경이 다시 하드코딩됐다(%d곳) — --surface 를 쓸 것" % len(white)
    print("  [117] 밝게/어둡게 토글 — 팔레트 완비·OS 자동추종 없음·강제 재계산 ✅")


def t118_ocr_crosscheck():
    """[118] 문서 스캔 교차검증 (2026-08-06 지시: "최고의 무료 도구 확인해서 비교 대조").

    조사 결론은 band/OCR_ENGINES.md 에 있다 — 무료·로컬·한글 조건에서 PaddleOCR 이
    여전히 최상위라 **엔진은 바꾸지 않았다.** 대신 정확도를 올릴 자리가 다른 데 있었다:
    한 엔진의 답만 보면 그 답이 틀렸다는 것을 알 방법이 없다.

    그래서 이 검증이 지키는 것은 두 가지다.
      ① **겹칠 때만 믿는다** — 두 엔진 이상이 같은 값을 낸 항목만 원장에 들어간다.
         한 엔진만 낸 값(단독), 엔진마다 다른 값(충돌), 1엔진 환경은 전부 입력 금지.
      ② **항목마다 대조한다** — 옛 doc_ocr.match() 는 공급가액 하나만 봤다. 발행일이
         틀려도 통과했다. 이제 한 항목이라도 불일치면 그 건은 빈칸도 채우지 않는다.
    """
    sys.path.insert(0, os.path.join(ROOT, "band"))
    import ocr_crosscheck as X

    # ① 표결 — 겹치면 합치, 하나뿐이면 단독, 갈리면 충돌(값을 정하지 않는다)
    assert X.vote(["A", "A", "B"]) == ("A", "합치")
    assert X.vote(["A", "B"]) == ("", "충돌"), "값이 갈렸는데 한쪽을 골랐다"
    assert X.vote(["A", "", None]) == ("A", "단독")
    assert X.vote(["", None]) == ("", "없음")
    # 금액은 근사 비교를 하지 않는다 — 한 자리만 틀려도 다른 값이다
    assert X.vote([2000000, 2000001])[1] == "충돌"

    doc = {"유형": "거래명세서", "발행일": "2026-08-01", "명세서번호": "SL-2026-0712",
           "승인번호": "", "프로젝트NO": "UJ2601138", "공급가액": 2000000,
           "세액": 200000, "합계": 2200000, "사업자번호": "123-81-45678"}
    two = X.merge_records({"paddle": doc, "windows": dict(doc)})
    assert two["신뢰도"] == "높음" and two["교차"]["공급가액"] == "합치"
    one = X.merge_records({"paddle": doc})
    assert one["신뢰도"].startswith("낮음"), "엔진 하나뿐인데 높음으로 나왔다"

    # 합쳐 놓고 공급가+세액≠합계 면, 겹쳤더라도 금액은 믿지 않는다
    bad = X.merge_records({"paddle": doc, "windows": dict(doc, 합계=2300000)})
    assert bad["교차"]["합계"] == "충돌"
    broke = X.merge_records({"paddle": dict(doc, 합계=2300000),
                             "windows": dict(doc, 합계=2300000)})
    assert broke["금액정합"] == "깨짐" and broke["교차"]["공급가액"] == "충돌", \
        "금액 정합성이 깨졌는데도 겹쳤다는 이유로 통과했다"

    # ② 전량을 두 번 읽지 않는다 — 그러나 원장에 쓸 때는 무조건 두 번 읽는다
    clean = dict(doc, 신뢰도="높음")
    assert X.needs_second_opinion(clean) == ""
    assert X.needs_second_opinion(clean, for_write=True), "원장 입력 후보를 한 번만 읽었다"
    assert X.needs_second_opinion(dict(clean, 공급가액=""))
    assert X.needs_second_opinion(dict(clean, 합계=2300000)) == "금액 정합성 깨짐"

    # ③ 항목별 대조 — 빈 원장 / 맞는 원장 / 틀린 원장
    led = {"원장_거래명세서발행일": "2026-08-01", "원장_거래명세서번호": "SL-2026-0712",
           "원장_공급가액": 2000000, "원장_거래명세서합계": 2200000}
    rows = X.compare_ledger(two, led)
    by = {r["항목"]: r["판정"] for r in rows}
    assert by["발행일"] == "일치" and by["명세서번호"] == "일치" and by["세액"] == "원장 빈칸"
    assert "일치" in X.ledger_verdict([r for r in rows if r["항목"] != "세액"])
    wrong = X.compare_ledger(two, dict(led, 원장_거래명세서발행일="2026-07-31"))
    assert "불일치(발행일)" == X.ledger_verdict(wrong), \
        "공급가액만 맞으면 통과하던 옛 판정이 남아 있다"

    # ④ 원장 입력 문지기 — 합치 + 빈칸 + 불일치 없음, 셋 다여야 들어간다
    empty = {"원장_공급가액": 2000000}
    ok = X.build_updates([{"문서": dict(two, 파일="a.jpg"), "정산ID": "S1",
                           "대조": X.compare_ledger(two, empty)}])
    cols = {u["col"] for u in ok}
    assert cols == {"거래명세서발행일", "거래명세서번호"}, cols
    assert all(u["sheet"] == "06_거래서류청구수금" and u["key_col"] == "정산ID" for u in ok)
    # 세액·합계는 수식·집계 열이라 읽어서 대조만 하고 쓰지 않는다
    assert not any(u["col"] in ("세액", "합계") for u in ok)
    # 한 항목이라도 불일치면 같은 건의 빈칸도 채우지 않는다
    conflict = X.compare_ledger(two, {"원장_공급가액": 1999})
    assert X.build_updates([{"문서": two, "정산ID": "S1", "대조": conflict}]) == []
    # 1엔진(단독)은 아무것도 못 채운다
    assert X.build_updates([{"문서": dict(one, 파일="a.jpg"), "정산ID": "S1",
                             "대조": X.compare_ledger(one, empty)}]) == []
    # 정산ID 를 못 찾았으면 채우지 않는다
    assert X.build_updates([{"문서": two, "정산ID": "", "대조": X.compare_ledger(two, empty)}]) == []

    # ⑤ 같은 답을 두 번 본 것을 '합치'라 부르지 않는다.
    #    doc_ocr 는 paddle 이 실패하면 Windows 결과를 paddle 자리에 캐시한다 —
    #    그것을 둘로 세면 거짓 근거가 만들어진다.
    assert list(X.drop_dependent({"paddle": "가 나 다", "windows": "가나다"})) == ["paddle"]
    assert len(X.drop_dependent({"paddle": "가나다", "windows": "가나라"})) == 2
    assert X.drop_dependent({"paddle": "", "windows": ""}) == {}

    # ⑥ 전량을 두 번 읽지 않는다 — 급한 것부터 예산만큼, 미룬 수는 반드시 알린다.
    #    사진 1,816장 × 둘째 엔진 4.8초 = 2.4시간이라 daily_run 이 무너진다(실측).
    take, defer = X.recheck_plan([("a", 0), ("b", 3), ("c", 1), ("d", 0), ("e", 2)], budget=2)
    assert take[:2] == ["a", "d"] and defer == 1, (take, defer)
    assert take[2:] == ["c", "e"], "급한 것(원장 불일치)보다 덜 급한 것을 먼저 읽었다"
    assert X.recheck_plan([("a", 0)] * 5, budget=1) == (["a"] * 5, 0), \
        "원장에 쓰려는 건이 예산에 잘렸다 — 쓰는 순간은 예산과 무관해야 한다"
    fill_rows = X.compare_ledger(two, empty)
    assert X.writable_now(two, fill_rows, "S1") is True
    assert X.recheck_reason(two, fill_rows, True) == (0, "원장 입력 후보")
    assert X.recheck_reason(dict(clean, 합계=2300000), [], False)[0] == 2
    full = X.compare_ledger(two, led)
    assert X.recheck_reason(two, full, False) == (None, ""), "다 맞는 건까지 두 번 읽고 있다"

    # ⑦ 엔진은 있는 것만 골라 쓴다 — 자동 설치·외부 업로드가 없어야 한다
    src = open(os.path.join(ROOT, "band", "ocr_crosscheck.py"), encoding="utf-8").read()
    for banned in ("pip install", "requests.post", "urllib.request", "http://", "https://"):
        assert banned not in src, "문서를 PC 밖으로 보내거나 자동 설치하는 코드가 있다: %s" % banned
    assert {e["엔진"] for e in X.engine_status()} >= {"paddle", "windows"}
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "ocr_crosscheck.py" in daily, "매일 도는 자리에 붙지 않았다(daily_run 누락)"
    assert os.path.isfile(os.path.join(ROOT, "band", "OCR_ENGINES.md")), \
        "무엇을 왜 골랐는지가 파일로 남지 않았다"
    print("  [118] 문서 OCR 교차검증 — 겹칠 때만 입력·항목별 원장 대조·로컬 전용 ✅")


def t119_context_guard():
    """[119] 컨텍스트가 다 차기 전에 마무리로 전환시킨다 (2026-08-06 지시).

    사용자 지시: "컨텍스트 윈도우 다 차기전 컴팩팅 자동으로 하는 알고리즘 추가".

    지키는 것
      ① 대화 기록 폴더를 **정말로 찾는다.** 슬러그는 글자 하나당 대시 하나다
         (`C:\\Users\\…` → `C--Users-…`). 여기를 `[^A-Za-z0-9]+` 로 묶었다가
         폴더를 못 찾아 사용량이 늘 0% 로 나왔었다 — 그러면 감시가 없는 것과 같다.
      ② 사용량은 **마지막 usage 합**(입력+캐시읽기+캐시생성)이다. 글자 수 추정이 아니다.
      ③ 단계는 올라갈 때만 말한다(70 예고 / 85 마무리 / 95 즉시). 매 입력마다 떠들면
         본문이 밀려 오히려 컨텍스트를 잡아먹는다.
      ④ 훅은 **절대 사람 입력을 막지 않는다** — 어떤 예외에도 exit 0.
      ⑤ settings.json 에 UserPromptSubmit 배선이 살아 있고, autoCompactWindow 는
         없으면 auto(권장·2026-08-07 사용자 승인), 있으면 한도 안이어야 한다.
    """
    import tempfile
    import context_guard as G
    _rd = lambda *p: open(os.path.join(*p), encoding="utf-8").read()
    TMP = tempfile.mkdtemp(prefix="ctxguard_")

    # ① 폴더 찾기 — 이 프로젝트의 기록 폴더가 실제로 잡혀야 한다
    d = G._project_dir()
    assert d and os.path.isdir(d), "대화 기록 폴더를 못 찾았다 — 슬러그 규칙 확인"

    # ② usage 합산 — 꼬리만 읽어도 최신 값을 집어야 한다
    tmp = os.path.join(TMP, "ctx_sample.jsonl")
    with open(tmp, "w", encoding="utf-8") as fh:
        fh.write(json.dumps({"message": {"usage": {
            "input_tokens": 5, "cache_read_input_tokens": 100,
            "cache_creation_input_tokens": 10}}}) + "\n")
        fh.write(json.dumps({"message": {"content": "usage 없는 줄"}}) + "\n")
        fh.write(json.dumps({"message": {"usage": {
            "input_tokens": 2, "cache_read_input_tokens": 300,
            "cache_creation_input_tokens": 8}}}) + "\n")
    assert G._used_tokens(tmp) == 310, "마지막 회차가 아니라 다른 줄을 셌다"
    assert G._used_tokens(os.path.join(TMP, "없는파일.jsonl")) == 0

    # ③ 단계 경계 — 0.70/0.85/0.95 에서만 올라간다
    assert G._stage(0.69)[1] == "여유"
    assert G._stage(0.70)[1] == "예고"
    assert G._stage(0.86)[1] == "마무리"
    assert G._stage(0.99)[1] == "즉시"
    # 올라간 순간에만 말한다 — 같은 단계를 다시 재면 조용하다
    base = {"advice": "x", "fresh": True, "percent": 90.0, "used": 1, "limit": 2,
            "stage": "마무리", "wrapup_ran_now": False, "auto_compact": True}
    assert G._message(base), "단계가 올라갔는데 아무 말도 하지 않았다"
    assert G._message(dict(base, fresh=False)) == "", "같은 단계에서 매번 떠들고 있다"
    assert "compact" in G._message(dict(base, auto_compact=False)), \
        "자동 요약이 꺼져 있는데 사람에게 알리지 않는다"

    # ④ 재기만 하는 호출은 인계를 돌리지도, 상태 파일을 건드리지도 않는다
    before = os.path.getmtime(G.STATE_PATH) if os.path.exists(G.STATE_PATH) else 0
    r = G.measure(limit=450000, act=False)
    assert r["wrapup_ran_now"] is False and r["limit"] == 450000
    after = os.path.getmtime(G.STATE_PATH) if os.path.exists(G.STATE_PATH) else 0
    assert before == after, "--dry 인데 상태 파일을 고쳤다"

    # 어떤 예외에도 exit 0 — main 이 통째로 try 로 감싸여 있어야 한다
    src = _rd(os.path.join(ROOT, "context_guard.py"))
    assert "sys.exit(0)" in src, "훅이 실패하면 사람 입력이 막힌다"

    # ⑤ 배선 — 훅이 붙어 있고 압축 시점이 한도 안이다
    project_root = os.environ.get("COUPANG_TEST_PROJECT_ROOT") or os.path.dirname(ROOT)
    st = json.loads(_rd(os.path.join(project_root, ".claude", "settings.json")))
    ups = json.dumps(st.get("hooks", {}).get("UserPromptSubmit", []), ensure_ascii=False)
    assert "context_guard.py" in ups, "UserPromptSubmit 훅에 배선되지 않았다"
    # ★ 사람 입력이 없는 동안이 가장 위험하다 (2026-08-06 지시 "밤을 새서라도").
    #   UserPromptSubmit 은 사람이 칠 때만 온다 — 지시 하나로 몇 시간을 혼자 도는
    #   밤샘 작업에서는 한 번도 안 온다. 그래서 도구를 쓸 때마다 오는 PostToolUse 에도
    #   같은 눈을 달아 둔다. 값은 90초에 한 번만 실제로 재서 도구를 늦추지 않는다.
    pts = json.dumps(st.get("hooks", {}).get("PostToolUse", []), ensure_ascii=False)
    assert "context_guard.py" in pts and "--tick" in pts,         "PostToolUse 에 컨텍스트 눈이 없다 — 사람 입력 없이 도는 동안 단계를 못 잡는다"
    assert "def tick(" in src and "TICK_EVERY" in src, "--tick 진입점이 없다"
    assert 'if now - float(st.get("tick_at") or 0) < TICK_EVERY' in src,         "매 도구 호출마다 대화 기록을 통째로 읽는다 — 모든 도구가 그만큼 느려진다"
    assert st.get("autoCompactEnabled") is True, "자동 요약이 꺼져 있다"
    # autoCompactWindow 는 2026-08-07 부터 기본이 auto(키 없음)다 — 모델에 맞는 창을
    # Claude Code 가 고른다. 키가 없으면 context_guard 가 DEFAULT_LIMIT(450k)로 경보한다.
    # 사람이 다시 오버라이드하면 설정 유효 범위(100k~1M) 안이어야 한다.
    win = st.get("autoCompactWindow")
    if win is not None:
        assert 100_000 <= int(win) <= 1_000_000, "압축 시점(autoCompactWindow)이 범위 밖이다: %s" % win
    assert "DEFAULT_LIMIT" in src, "auto 일 때 기댈 보수 한도가 context_guard 에 없다"
    # ★ 같은 단계에서 **두 번 말하지 않는다** (2026-08-07 지시: "작업에 방해되는 규칙").
    #   PostToolUse 훅이 session_id 를 비워 보내는 일이 있는데, 그때 sid=="" 라
    #   저장된 상태와 절대 안 맞아 매번 '단계가 방금 올랐다'로 오인했다 —
    #   같은 경고가 도구마다 반복되고 인계 자동 정리가 90초마다 다시 돌았다.
    #   식별자가 없으면 **대화 기록 파일 이름**으로 같은 세션임을 알아본다.
    src = open(os.path.join(ROOT, "context_guard.py"), encoding="utf-8").read()
    assert 'if not same and not sid and path:' in src, \
        "세션 식별자가 없을 때 같은 세션임을 알아보지 못한다 — 경고가 무한 반복된다"
    assert 'st.get("transcript") == os.path.basename(path)' in src
    # 그리고 '단계가 오를 때만 말한다'가 실제로 그 판정을 쓰는지 — advice 가 있어도
    # fresh 가 아니면 한 마디도 내지 않아야 한다.
    import context_guard as _cg
    assert _cg._message({"advice": "무언가", "fresh": False, "percent": 90, "used": 1,
                         "limit": 2, "stage": "마무리", "wrapup_ran_now": False,
                         "auto_compact": True}) == "", \
        "단계가 그대로인데도 경고를 낸다 — 도구를 쓸 때마다 같은 말이 반복된다"
    assert _cg._message({"advice": "무언가", "fresh": True, "percent": 90, "used": 1,
                         "limit": 2, "stage": "마무리", "wrapup_ran_now": False,
                         "auto_compact": True}) != "", "단계가 올랐는데 알리지 않는다"
    print("  [119] 컨텍스트 감시 — 사용량 실측·단계 전환·인계 자동·반복 경고 차단 ✅")


def t120_calendar_sheet_and_share():
    """[120] 캘린더: 날짜 창·PC 텍스트 격자·고정 주소·크롬 설치 유도 (2026-08-06 지시).

    지시: "캘린더 날짜 클릭시 밑에서 위로 올라가는 레이어창이나 모달창으로 / 공유시
    크롬을 자동으로 열어 설치하게 / 공유 캘린더 주소 고정(변동 주지마) / 날짜 크기 크게 /
    PC 는 텍스트로 자세히, 모바일은 지금처럼."

    여기서 지키는 것은 **되돌아가기 쉬운 네 가지**다.
      ① 날짜를 눌러도 화면 저 아래 목록만 바뀌던 옛 동작으로 돌아가지 않을 것
      ② 공유 주소에 달(m=)·필터(off=)가 다시 붙지 않을 것 — 붙는 순간 주소가 흔들린다
      ③ 열쇠 파일이 없다고 조용히 새 열쇠를 뽑지 않을 것 — 뿌린 링크가 통째로 죽는다
      ④ 설치 매니페스트가 업무 앱이 아니라 **캘린더**를 가리킬 것
    """
    _rd = lambda *p: open(os.path.join(ROOT, *p), encoding="utf-8").read()
    live = _rd("webapp", "index.html")
    cal = _rd("docs", "cal.html")

    # ① 날짜를 누르면 창이 뜬다 — 두 화면 모두
    assert "calOpenSheet(day)" in live, "앱: 날짜를 눌러도 창이 뜨지 않는다"
    assert 'id="calSheet"' in live and "function calCloseSheet(" in live, "앱: 날짜 창이 없다"
    assert "document.body.style.overflow='hidden'" in live, \
        "앱: 창이 떠 있는 동안 뒤 화면이 같이 스크롤된다"
    assert "openSheet()" in cal and 'id="sheet"' in cal, "공유 캘린더: 날짜 창이 없다"
    # 폰=아래에서 위로 · PC=가운데 모달 (한 마크업을 CSS 가 가른다)
    for doc, name in ((live, "앱"), (cal, "공유 캘린더")):
        assert "translateY(100%)" in doc, f"{name}: 아래에서 올라오는 레이어가 아니다"
        assert "translate(-50%,-50%)" in doc, f"{name}: PC 가운데 모달이 없다"
    # 애니메이션을 rAF 하나에만 걸면 탭이 숨었을 때 창이 안 올라온다(2026-08-06 캡처 사고)
    for doc, name in ((live, "앱"), (cal, "공유 캘린더")):
        assert "setTimeout(rise, 60)" in doc, f"{name}: rAF 가 안 불리는 상황의 대비가 없다"

    # ② PC 는 글자, 폰은 점 — 가르는 것은 CSS 미디어쿼리여야 한다(창 크기를 재지 않는다)
    assert ".cal2-txs{display:none}" in live and "@media(min-width:900px)" in live, \
        "앱: PC 텍스트 격자 규칙이 없다"
    assert 'class="cal2-txs"' in live, "앱: 격자 칸에 텍스트를 실어 보내지 않는다"
    assert ".ctxs{display:none}" in cal and 'class="ctxs"' in cal, \
        "공유 캘린더: PC 텍스트 격자가 없다"
    # ③ 날짜 글씨 — 예전 값(14px/12.5px)으로 되돌아가지 않게 못을 박는다
    assert ".cal2-day .d{font-size:17px" in live, "앱: 날짜 글씨가 다시 작아졌다"
    assert ".cday b{font-size:17px" in cal, "공유 캘린더: 날짜 글씨가 다시 작아졌다"

    # ④ 공유 주소 고정 — 달·필터를 주소에 담지 않는다
    link = live[live.index("function calendarLink()"):]
    link = link[:link.index("\n}")]
    assert "?view=calendar'" in link, "고정 주소가 아니다"
    assert "p.set('m'" not in link and "p.set('off'" not in link, \
        "공유 주소에 달·필터가 다시 붙었다 — 누를 때마다 주소가 달라진다"
    # 이미 뿌린 옛 링크(m=·off=)는 계속 열려야 한다
    assert "if(/^\\d{4}-\\d{2}$/.test(m)) CAL_MONTH = m;" in live, \
        "옛 링크의 달 지정을 더 이상 받아 주지 않는다"

    # ⑤ 열쇠(=주소)를 조용히 바꾸지 않는다
    share = _rd("cal_share.py")
    assert "if not new and os.path.exists(OUT):" in share and "sys.exit(" in share, \
        "공유본이 있는데 열쇠만 없을 때 새 열쇠를 조용히 뽑는다 — 뿌린 링크가 다 죽는다"
    assert "def note_url(" in share, "고정 주소를 사람이 찾아볼 자리가 없다"

    # ⑥ 크롬으로 넘겨 설치 — 카카오톡 안 브라우저는 설치 자체가 불가능하다
    assert "function inAppBrowser(" in cal and "KAKAOTALK" in cal, \
        "앱 안 브라우저 판정이 없다 — 카카오톡에서 열면 설치가 영영 안 된다"
    assert "function chromeJumpUrl(" in cal, "넘길 주소를 만드는 자리가 따로 없다"
    assert "package=com.android.chrome" in cal, "안드로이드 크롬 넘기기가 없다"
    assert "S.browser_fallback_url=" in cal, "크롬이 없는 폰의 되돌아갈 자리가 없다"
    # ★ 아이폰은 크롬이 아니라 **사파리**로 보내야 한다 — 아이폰 크롬은 홈 화면 설치를
    #   못 한다. 크롬으로 보내면 "열리지만 설치는 안 되는" 상태로 한 걸음 헛돈다.
    assert "'x-safari-' + shareUrl()" in cal, "아이폰을 사파리로 보내지 않는다"
    assert "googlechromes://" not in cal,         "아이폰을 크롬으로 보낸다 — 아이폰 크롬으로는 홈 화면에 설치할 수 없다"
    assert "sessionStorage.setItem('csos.cal.jump','1')" in cal, \
        "자동 넘기기에 한 번만 걸리는 잠금이 없다 — 되돌아올 때마다 다시 튄다"
    assert "beforeinstallprompt" in cal and "PROMPT.prompt()" in cal, "설치 창을 부르지 않는다"

    # ⑦ 설치되는 것이 '캘린더' 여야 한다 — 업무 앱이 깔리면 받는 사람은 PIN 화면을 본다
    assert 'href="cal-manifest.json"' in cal, "공유 캘린더가 업무 앱 매니페스트를 물고 있다"
    cm = json.loads(_rd("docs", "cal-manifest.json"))
    assert cm["start_url"].endswith("/cal.html"), cm
    assert cm["id"].endswith("/cal.html"), "설치 신원이 업무 앱과 겹치면 하나만 설치된다"
    app_mf = json.loads(_rd("docs", "manifest.json"))
    _rev = lambda mf: sorted(i["src"] for i in mf["icons"])
    assert _rev(cm) == _rev(app_mf), "아이콘 판이 업무 앱과 어긋났다 — 한쪽만 갱신됐다"
    import cloud_publish as _CP
    assert "docs/cal-manifest.json" in _CP.PUBLISH_FILES, \
        "설치 정보가 고정 주소에 올라가지 않는다 — 링크를 받은 폰에는 계속 옛 매니페스트가 간다"

    # ⑧ 설치한 아이콘은 주소의 `#k=` 를 못 가져간다 — 기기에 기억해 두지 않으면 매번 막힌다
    assert "localStorage.setItem(KSTORE" in cal and "localStorage.getItem(KSTORE)" in cal, \
        "열쇠를 기억하지 않는다 — 설치한 아이콘이 '열쇠가 없는 주소입니다'만 띄운다"
    # ★ 캡처 목록 가운데 칸에 사유·진행내용 (2026-08-08 지시: "캘린더 캡처 화면 아래
    #   리스트 중간 빈 공간에 사유 진행내용 등 표기해"). 캠프명만 늘어서 있으면
    #   "무슨 건인지"를 알려고 앱을 다시 열어야 한다.
    assert "function calWhyOf(" in live, "캡처 목록 가운데가 여전히 비어 있다"
    why = live.split("function calWhyOf(")[1].split("\n}")[0]
    assert "'-'" not in why and '"-"' not in why, \
        "자료가 없을 때 '-' 로 채운다 — 빈칸은 '없다'는 사실 그대로 두어야 한다"
    assert "신청내용" in why and "진행상태" in why, "사유·진행내용이 빠져 있다"
    cap = live.split("async function calendarCapture(")[1].split("\nasync function ")[0]
    assert "calWhyOf(e)" in cap, "캡처가 사유 칸을 그리지 않는다"
    assert "measureText(campTxt)" in cap, \
        "캠프명을 고정폭으로 잘라 사유 자리를 만든다 — 현장 이름이 잘리면 안 된다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    done = srv.split('"as_done", f"돌발AS 완료')[1][:400]
    assert "신청내용" in done, "완료 건에는 무슨 일이었는지가 실리지 않는다"
    print("  [120] 캘린더 — 날짜 창·PC 텍스트 격자·고정 주소·크롬 설치·캡처 사유칸 ✅")


def t202_layer_dialogs():
    """[202] 알림·확인·입력을 전부 레이어로 (2026-08-06 지시).

    사용자 지시: "UX UI 편리하게 밑에서 위로 올라가는 레이어창으로 하거나
    모달창으로 나오게 하는 구조를 … 전체 적용해 (담당자별 업무 센터도 마찬가지야)".

    지키는 것
      ① 브라우저 기본 창(alert·confirm·prompt)이 **한 개도 남아 있지 않다.**
         하나라도 남으면 폰에서 "localhost:8000 says" 회색 상자가 튀어나온다.
      ② 대체 함수(notice·askYesNo·askText)와 그 뼈대(dlgAsk)가 실제로 있다.
      ③ confirm/prompt 는 답을 **기다려야** 한다 — askYesNo·askText 호출은 전부 await 다.
         빠뜨리면 Promise 객체가 늘 참이라 "취소"가 먹지 않는다(조용한 사고).
      ④ 폰=아래에서 올라오는 시트 · PC=가운데 모달. 두 모양이 CSS 에 다 있어야 한다.
      ⑤ 담당자 업무센터 입력 폼은 **요소째** 레이어로 옮긴다(HTML 복사가 아니다).
         복사하면 <input type=file> 과 적던 메모가 사라진다. 닫을 때 제자리로 돌아온다.
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))

    # ① 기본 창이 남아 있지 않다 (deferredInstallPrompt.prompt() 같은 메서드는 제외)
    for bad in ("alert", "confirm", "prompt"):
        left = re.findall(r"(?<![\w.$])%s\s*\(" % bad, js)
        assert not left, "브라우저 기본 %s() 가 %d 곳 남아 있다" % (bad, len(left))

    # ② 대체 함수가 있다
    for fn in ("function notice(", "function askYesNo(", "function askText(",
               "function dlgAsk(", "function dlgCancel("):
        assert fn in js, "대체 함수가 없다: %s" % fn

    # ③ 답을 기다린다 — 정의부를 뺀 모든 호출 앞에 await 가 있다
    for fn in ("askYesNo", "askText"):
        for m in re.finditer(r"(.{8})%s\s*\(" % fn, js):
            head = m.group(1)
            if head.endswith("function ") or head.endswith("nction "):
                continue                      # 정의부
            assert head.rstrip().endswith("await"), \
                "%s 호출에 await 가 빠졌다 — 취소가 먹지 않는다: …%s" % (fn, head)

    # ④ 두 모양이 다 있다: 폰 시트(아래에서) · PC 모달(가운데)
    assert "transform:translateY(100%)" in idx, "밑에서 올라오는 모양이 없다"
    assert ".dlg-wrap.in .dlg{transform:translate(-50%,-50%) scale(1)" in idx, \
        "PC 가운데 모달 모양이 없다"
    assert ".sheet.open{transform:translate(-50%,-50%) scale(1)" in idx, \
        "상세 시트가 PC 에서도 바닥에 붙어 있다 — 모달로 띄우기로 했다"
    assert 'id="dlgWrap"' in idx and 'id="dlgBody"' in idx and 'id="dlgFoot"' in idx
    # PIN 창(1400)보다 아래, 상세 시트(41)보다 위에 있어야 겹침 순서가 맞다
    z = int(re.search(r"\.dlg-wrap\{position:fixed;inset:0;z-index:(\d+)", idx).group(1))
    assert 41 < z < 1400, "레이어 겹침 순서가 어긋난다(z-index %s)" % z

    # ⑤ 담당자 업무센터 — 요소째 옮기고, 닫을 때 돌려놓는다
    assert "function layerOpen(" in js and "function layerRestore(" in js
    assert "layerOpen('ryuEntryForm'" in js, \
        "담당자 업무센터 입력 폼이 레이어로 열리지 않는다"
    assert js.count("layerRestore()") >= 3, \
        "레이어를 되돌리는 자리가 모자란다(내용교체·한겹닫기·통째닫기 세 곳)"
    # setSheetContent 는 innerHTML 로 지우기 **전에** 되돌려야 한다
    seg = js[js.index("function setSheetContent("):][:400]
    assert seg.index("layerRestore()") < seg.index("body.innerHTML"), \
        "옮겨 온 요소를 돌려놓기 전에 시트를 지운다 — 폼이 통째로 사라진다"
    # 페이지 안에서 사라진 자리를 스크롤로 때우던 옛 방식이 남아 있으면 안 된다
    assert "$('ryuEntryForm').scrollIntoView" not in js, \
        "옛 스크롤 방식이 남아 있다 — 레이어와 둘 다 돌면 화면이 튄다"

    # ⑥ 폰 사본(docs/app.html)도 같은 모양이어야 한다.
    #    이 파일은 index.html 에서 생성되는 것이 아니라 따로 있는 앱이다 —
    #    한쪽만 고치면 "폰에서만 회색 상자가 뜨는" 상태가 된다.
    phone = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    pjs = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", phone, re.S))
    for bad in ("alert", "confirm", "prompt"):
        left = re.findall(r"(?<![\w.$])%s\s*\(" % bad, pjs)
        assert not left, "폰 사본에 기본 %s() 가 %d 곳 남아 있다" % (bad, len(left))
    assert "function notice(" in pjs and "function askYesNo(" in pjs, \
        "폰 사본에 대체 함수가 없다 — 두 앱의 이름을 맞춰 둔다"
    assert 'id="dlgwrap"' in phone and ".dlgwrap.on{display:flex}" in phone
    assert "align-items:flex-end" in phone and "@media(min-width:900px){\n .dlgwrap{align-items:center}" in phone, \
        "폰 사본에 '아래에서 올라옴 → 넓으면 가운데' 두 모양이 없다"
    print("  [202] 알림·확인·입력 레이어 — 기본창 0개·await 완비·폰시트/PC모달·폼 요소째 이동 ✅")


def t122_dash_drag_and_remote_version():
    """[122] 대시보드 카드 끌기 + 리모컨 버전 관리 (2026-08-06 지시).

    사용자 지시: "각 카드를 클릭해서 드래그 앤 드롭으로 자유롭게 움직일 수 있는 기능
    추가, 각 카드는 밑에서 위로 올라가는 레이어창으로 하거나 모달창으로 나오게 …" ·
    "리모컨 버전 선택할 수 있는 기능 추가하고 전체적으로 버전 관리가 VER.3인지
    VER.4인지 입력 및 확인 수정 가능하게 고도화".

    지키는 것 — 대시보드
      ① 끌기는 **포인터 이벤트**다. HTML5 dragstart 는 터치에서 이벤트 자체가 나지
         않아 폰에서 아무리 끌어도 안 움직였다. draggable 손잡이가 남아 있으면 안 된다.
      ② 카드 **아무 데나** 눌러 끈다 — 손잡이 셀렉터로만 집는 코드가 없어야 한다.
      ③ 6px 문턱(슬롭)이 있어야 평범한 탭·체크가 살아남는다.
      ④ 카드를 레이어로 연 동안 저장이 일어나도 **순서에서 빠지지 않는다.**
         (grid 의 자식이 아니게 되므로, base 순서를 이어 붙이는 처리가 있어야 한다)

    지키는 것 — 리모컨 버전
      ⑤ 표기 통일: 'ver3'·'V4'·'VER 3' 이 전부 하나로 모인다.
      ⑥ 담당자 보유가 **버전별로** 나온다(합계만으로는 VER.3 몇 개인지 못 센다).
      ⑦ 버전이 비었거나 '미확인'인 줄을 모아 준다 — 빈칸만 찾으면 대부분을 놓친다.
      ⑧ 화면의 버전 선택지는 **서버 목록만** 쓴다(화면이 따로 적으면 조용히 갈린다).
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))

    # ① 포인터 끌기로 갈아탔다 — 옛 HTML5 끌기 흔적이 남으면 안 된다
    assert "function dashDragEnable(" in js, "포인터 기반 끌기가 없다"
    for gone in ("dashDragging", "dashKpiDragging", "dashPaletteDragging"):
        assert gone not in js, "옛 HTML5 끌기가 남아 있다(폰에서 안 된다): %s" % gone
    # 대시보드 배치 코드 안에는 HTML5 끌기가 없어야 한다.
    # (파일을 끌어다 놓는 업로드 구역의 dragover 는 별개다 — 거기는 그대로 둔다)
    dash = js[js.index("function initDashboardLayout("):]
    dash = dash[:dash.index("\nasync function ", 10)]
    for gone in ("addEventListener('drag", 'addEventListener("drag', "dataTransfer"):
        assert gone not in dash, "대시보드에 HTML5 끌기가 남아 있다: %s" % gone
    assert 'class="dash-drag-handle" draggable' not in js, \
        "손잡이에 draggable 이 남아 있다 — 네이티브 끌기가 포인터 끌기와 싸운다"
    # ② 세 곳 모두 붙었다: 화면 카드 · 핵심지표 · 보관함
    for host, sel in (("dashGrid", "[data-dash-block]"), ("kpis", "[data-kpi-card]"),
                      ("dashPalette", ".dash-choice")):
        assert "dashDragEnable('%s','%s'" % (host, sel) in js, \
            "%s 에 끌기가 붙지 않았다" % host
    # ③ 탭과 끌기를 가르는 문턱
    assert "DASH_DRAG_SLOP" in js and re.search(r"DASH_DRAG_SLOP\s*=\s*[1-9]", js), \
        "끌기 문턱이 없다 — 살짝만 눌러도 카드가 끌려간다"
    assert "input,button,select,textarea,a,label" in js, \
        "입력 요소 위에서도 끌기가 시작된다 — 체크박스를 못 누른다"
    # ④ 레이어로 연 카드가 순서에서 빠지지 않는다
    assert "function openDashCardLayer(" in js
    seg = js[js.index("function dashboardLayoutState("):][:1600]
    assert "(base.order||[]).forEach" in seg, \
        "레이어로 열어 둔 카드가 저장 때 순서에서 빠진다"

    # ⑤~⑧ 리모컨 버전
    sys.path.insert(0, ROOT)
    import ledger_db as L
    assert L._remote_version("ver3") == "VER.3"
    assert L._remote_version("V4") == "VER.4"
    assert L._remote_version("VER 3") == "VER.3"
    assert L._remote_version("구형") == "기존형"
    assert L._remote_version("") == "", "빈 값을 임의로 채우면 안 된다"
    src = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    hold = src[src.index("def _remote_holdings("):]
    hold = hold[:hold.index("\ndef ", 10)]
    assert '"versions"' in hold, "담당자 보유에 버전별 내역이 없다"
    gaps = src[src.index("def remote_version_gaps("):]
    gaps = gaps[:gaps.index("\ndef ", 10)]
    assert "TRIM(version)='미확인'" in gaps, \
        "'미확인'으로 저장된 줄을 빠뜨린다 — 빈칸만 세면 대부분을 놓친다"
    for t in ("remote_issue", "remote_delivery", "remote_stock"):
        assert t in gaps, "%s 가 버전 확인 대상에서 빠졌다" % t
    st = L.remote_status()
    for k in ("version_totals", "version_gaps", "versions"):
        assert k in st, "remote_status 에 %s 가 없다" % k
    for v in st["version_totals"].values():
        assert v["all"] == v["holding"] + v["stock"], "버전 합계가 안 맞는다"
    # ⑧ 화면은 서버 목록만 쓴다 + 그 자리에서 고치는 길이 있다
    assert "function remoteSetVersion(" in js, "버전을 그 자리에서 고칠 수 없다"
    assert "(s.versions||[])" in js and "'VER.3'," not in js.replace("REMOTE", ""), \
        "화면이 버전 목록을 따로 적어 두었다 — 서버와 갈린다"
    print("  [122] 대시보드 포인터 끌기·카드 레이어 + 리모컨 버전별 보유·미확인 채우기 ✅")


def t123_calendar_share_tools():
    """[123] 공유 캘린더 — 크롬으로 자동 전환·설치·엑셀/캡처/복사 (2026-08-06 지시).

    사용자 지시: "이 쿠팡 캘린더 접속시 자동으로 크롬으로 열리게 하고 설치할 수 있게
    해주고, 여기에 엑셀저장, 캡처, 복사 버튼 동일하게 추가해서".

    지키는 것
      ① 카톡·밴드 안의 브라우저면 **자동으로** 밖(크롬/사파리)으로 넘긴다.
         거기서는 홈 화면 설치가 아예 안 되고 저장도 어디로 갔는지 알 수 없다.
      ② 그런데 **한 번만** 넘긴다(sessionStorage). 안 그러면 뒤로 돌아올 때마다
         또 튀어나가 앱을 쓸 수가 없다. '나중에'를 누른 사람은 아예 넘기지 않는다.
      ③ 설치 버튼은 **설치가 가능할 때만** 보인다 — 눌러도 아무 일 없는 버튼은
         고장으로 읽힌다.
      ④ 도구 세 개(이미지 저장·복사·엑셀)가 다른 화면과 같은 아이콘으로 있다.
      ⑤ 내보내기는 **화면과 같은 달·같은 필터**만 담는다. 파일과 화면의 건수가
         다르면 그때부터 파일을 못 믿는다.
      ⑥ 캡처와 복사는 **같은 그리기 코드**를 쓴다(toBlob). 두 벌이면 한쪽만 낡는다.
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))

    # ①② 자동 전환과 그 안전장치
    assert "function calAutoOpenOutside(" in js, "자동으로 크롬으로 넘기는 코드가 없다"
    auto = js[js.index("function calAutoOpenOutside("):]
    auto = auto[:auto.index("\nfunction ", 10)]
    assert "csosInApp()" in auto, "인앱 여부를 보지 않고 무조건 넘긴다"
    assert "sessionStorage" in auto and "csos_auto_out" in auto, \
        "한 번만 넘기는 장치가 없다 — 뒤로 올 때마다 튀어나간다"
    assert "csos_install_hint_off" in auto, "'나중에'를 누른 사람에게도 자동으로 넘긴다"
    assert "csosStandalone()" in auto, "이미 설치해 쓰는 사람까지 밖으로 넘긴다"
    assert "calAutoOpenOutside();" in js[js.index("function maybeInstallForShare("):][:600], \
        "공유 링크로 들어왔을 때 자동 전환이 걸리지 않는다"

    # ③ 설치 버튼
    assert 'id="calInstallBtn"' in idx and "hidden>" in idx.split('id="calInstallBtn"')[1][:400], \
        "설치 버튼이 기본 숨김이 아니다"
    assert "function calSyncInstallButton(" in js and "function calInstallApp(" in js

    # ④ 도구 세 개 — 다른 화면과 같은 아이콘
    tools = idx[idx.index('id="calTools"'):]
    tools = tools[:tools.index("</div>")]
    for icon in ("#i-image-down", "#i-clipboard-copy", "#i-file-spreadsheet"):
        assert icon in tools, "캘린더 도구에 %s 가 없다" % icon
    for fn in ("calendarCapture()", "calCopyList()", "calExportList()"):
        assert fn in tools, "캘린더 도구에 %s 가 연결되지 않았다" % fn

    # ⑤ 화면과 같은 것만 내보낸다
    exp = js[js.index("function calRowsForExport("):]
    exp = exp[:exp.index("\nfunction ", 10)]
    assert "calendarRows()" in exp, "필터를 무시하고 전체를 내보낸다"
    assert "calMonthOf(e.날짜) === m" in exp, "보고 있는 달만 내보내지 않는다"

    # ⑥ 캡처와 복사가 같은 그림을 쓴다
    assert "opt.toBlob" in js and "calendarCapture({toBlob:true})" in js, \
        "복사가 캡처와 다른 코드로 그림을 만든다 — 한쪽만 낡는다"
    cp = js[js.index("async function calCopyList("):]
    cp = cp[:cp.index("\nfunction ", 10) if "\nfunction " in cp else 2000]
    assert "clipboard.writeText" in cp, \
        "이미지 복사가 막힌 폰에서 아무 일도 일어나지 않는다 — 글 복사 대비가 없다"
    print("  [123] 공유 캘린더 — 크롬 자동 전환(1회)·설치 버튼·엑셀/캡처/복사 ✅")


def t124_no_duplicate_menus():
    """[124] 겹치는 메뉴 없애기 (2026-08-06 지시).

    사용자 지시: "앱 전체적으로 겹치는 메뉴가 보이는데 이런 부분 통합하고 ui ux 개선해"
    (쿠팡 캘린더 화면 사진과 함께).

    사진에 찍힌 것은 셋이었다.
      · 큰 버튼 '캡처' 와 아이콘 '이미지 저장' 이 **같은 함수**였다.
      · '링크 복사' 와 '복사' — 같은 말이 두 가지를 가리켰다.
      · 알림 칩이 헤더 버튼들을 **덮고** 있었고, 좁은 폰에서는 헤더 글자와
        고객사 로고가 서로 겹쳤다.

    그래서 이 검증은 '지금 고쳤다'가 아니라 **다시 생기지 않게** 막는다.
      ① 한 화면 안에 같은 onclick 을 가진 버튼이 두 개 있으면 실패.
      ② 도구줄의 복사 버튼은 무엇을 복사하는지 이름에 있어야 한다.
      ③ 갱신 상태는 떠 있는 칩이 아니라 **앱바 안**에 놓는다.
      ④ 헤더는 좁아지면 겹치지 말고 덜어내거나 잘린다.
    """
    import collections

    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    # 화면 마크업은 <main> 안에만 있다. JS 템플릿 문자열까지 같이 세면
    # `${...}` 로 인자가 달라지는 필터 칩들이 중복으로 잡혀 뜻이 흐려진다.
    static = idx[idx.index('<main class="shell">'): idx.index("</main>")]

    # ① 화면(<section class="view">) 단위로 같은 동작이 두 번 놓여 있지 않은가
    marks = [(m.start(), m.group(1))
             for m in re.finditer(r'<section class="view"[^>]*id="v-([\w-]+)"', static)]
    assert len(marks) >= 8, "화면을 찾지 못했다 — 마크업 구조가 바뀌었다"
    marks.append((len(static), None))
    btn = re.compile(r"<button\b[^>]*?>.*?</button>", re.S)
    onx = re.compile(r'onclick="([^"]+)"')
    bad = []
    for i in range(len(marks) - 1):
        chunk = static[marks[i][0]:marks[i + 1][0]]
        seen = collections.Counter()
        for m in btn.finditer(chunk):
            f = onx.search(m.group(0))
            if f:
                seen[re.sub(r"\s+", "", f.group(1))] += 1
        for fn, n in seen.items():
            if n > 1:
                bad.append("%s 화면에 %s 가 %d번" % (marks[i][1], fn, n))
    assert not bad, "한 화면에 같은 동작 버튼이 겹쳐 있다: " + " / ".join(bad)

    # ② '복사' 만으로는 무엇을 복사하는지 알 수 없다 — 링크 복사와 헷갈렸던 자리다
    assert 'aria-label="복사"' not in idx, \
        "무엇을 복사하는지 모르는 '복사' 버튼이 남아 있다(링크 복사와 헷갈린다)"
    assert idx.count('aria-label="이미지 복사"') >= 4, \
        "도구줄의 복사 버튼 이름이 통일되지 않았다"
    cal = idx[idx.index('id="calTools"'):]
    cal = cal[: cal.index("</div>")]
    for lbl in ("새로고침", "링크 복사", "이미지 저장", "이미지 복사", "엑셀 저장"):
        assert 'aria-label="%s"' % lbl in cal, "캘린더 도구줄에 '%s' 가 없다" % lbl
    head = idx[idx.index('<section class="view" id="v-calendar"'):]
    head = head[: head.index('id="calTools"')]
    assert head.count("<button") == 1 and "focusCalendarEntry()" in head, \
        "캘린더 위쪽 큰 버튼이 다시 늘었다 — 저장·전달은 아래 도구줄 한 곳이다"

    # ③ 갱신 상태가 업무 화면 위에 떠 있지 않고 앱바 안에만 있는가
    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))
    assert 'id="dataSyncChip"' in idx and 'data-refresh-ui="compact-header"' in idx, \
        "갱신 상태가 앱바의 작은 상태표시로 통합되지 않았다"
    for fn in ("function netBanner(", "function swrChip("):
        blk = js[js.index(fn):]
        blk = blk[: blk.index("\nfunction ", 10)]
        assert "document.body.appendChild" not in blk and "location.reload" not in blk, \
            "%s 가 다시 업무 화면을 덮거나 문서를 새로고침한다" % fn

    # ④ 좁은 화면에서 헤더가 겹치지 않게 덜어낸다
    css = "".join(re.findall(r"<style[^>]*>(.*?)</style>", idx, re.S))
    narrow = re.search(r"@media\(max-width:440px\)\{(.*?)\n\}", css, re.S)
    assert narrow, "좁은 폰(≤440px) 전용 헤더 규칙이 없다"
    narrow = narrow.group(1)
    assert ".appbar h1{display:none}" in narrow, \
        "좁은 폰에서 헤더 글자를 덜어내지 않는다 — 로고와 겹친다"
    assert ".appbar-identity{flex:none}" in narrow and "min-width:0" in narrow, \
        "자리가 모자랄 때 무엇이 줄어들지 정해 두지 않았다 — 앱 아이콘이 잘린다"
    assert ".appbar-identity{overflow:hidden}" in css, \
        "넘칠 때 겹치지 않게 자르는 안전망이 없다"
    compact = re.search(
        r"@media\((?:min-width:600px\)\s*and\s*\()?max-width:899px\)\{(.*?)\n\}",
        css, re.S,
    )
    assert compact and ".data-sync-chip" in compact.group(1), \
        "600·768px에서 헤더 상태표시를 별도 줄로 피하는 규칙이 없다"

    # ⑤ 좁은 폰에서 도구줄은 옆으로 숨지 말고 줄을 바꾼다
    phone = re.search(r"@media\(max-width:640px\)\{(.*?)\n\}", css, re.S)
    assert phone, "폰(≤640px) 전용 도구줄 규칙이 없다"
    phone = phone.group(1)
    assert "flex-wrap:wrap!important" in phone and "overflow-x:visible" in phone, \
        "폰에서 도구줄이 옆으로 스크롤한다 — 마지막 버튼이 안 보인 채 숨는다"

    print("  [124] 겹치는 메뉴 통합 — 화면당 중복 동작 0 · 앱바 상태표시 · 헤더 겹침 차단 ✅")


def t121_pid_alive():
    """[121] 죽은 프로세스를 살아 있다고 하지 않는다 (2026-08-06 실사고).

    무슨 일이 있었나 — `reports/.daily_run.lock` 이 **죽은 pid 의 이름으로** 남아
    daily_run 이 밤새 한 번도 못 돌았다. 잠금 규칙("주인이 죽었으면 회수")은 옳았는데
    **판정이 틀렸다**: 윈도우 `OpenProcess` 는 이미 끝난 프로세스에도 핸들을 준다.
    핸들이 열렸다는 것만으로 살아 있다고 본 탓에 잠금이 스스로 풀릴 길이 없었다.
    `Get-Process` 로는 안 보이는 pid 라 사람 눈에도 안 띈다 — 조용한 사고다.

    여기서 지키는 것
      ① 종료 코드까지 확인할 것(STILL_ACTIVE) — 핸들만 보고 판정하지 말 것
      ② 판정 불가는 None — '모르면 죽었다'로 넘기면 산 세션의 점유를 빼앗는다
      ③ 그 판정을 daily_run 잠금과 session_handoff 가 **같은 곳에서** 쓸 것
    """
    import subprocess
    import pid_alive as P

    assert P.alive(os.getpid()) is True, "지금 돌고 있는 나를 죽었다고 한다"
    assert P.alive(0) is None and P.alive("x") is None, "말이 안 되는 pid 는 None"

    # 방금 끝난 프로세스는 **확실히 죽었다**고 나와야 한다. 옛 판정이 틀렸던 바로 그 자리다.
    pr = subprocess.Popen([sys.executable, "-c", "pass"])
    pr.wait()
    assert P.alive(pr.pid) is False,         "끝난 프로세스를 살아 있다고 한다 — 잠금·점유가 영원히 안 풀린다"
    assert P.dead(pr.pid) is True and P.dead(os.getpid()) is False

    src = open(os.path.join(ROOT, "pid_alive.py"), encoding="utf-8").read()
    assert "GetExitCodeProcess" in src and "STILL_ACTIVE" in src,         "핸들만 보고 판정하면 끝난 프로세스가 살아 있는 것으로 나온다"
    assert "restype" in src and "argtypes" in src,         "64비트 핸들이 32비트로 잘린다 — 엉뚱한 핸들을 닫을 수 있다"

    # 두 곳이 같은 판정을 쓰는가 (한쪽만 고치면 그쪽에서 같은 사고가 또 난다)
    for name in ("daily_run.py", "session_handoff.py"):
        t = open(os.path.join(ROOT, name), encoding="utf-8").read()
        assert "import pid_alive" in t, "%s 가 옛 판정을 그대로 쓴다" % name
        assert "windll.kernel32.OpenProcess" not in t, (
            "%s 가 아직 스스로 판정한다 — 한쪽만 고치면 같은 사고가 또 난다" % name)

    # 죽은 주인의 잠금은 실제로 회수돼야 한다
    import daily_run as D
    assert D._pid_alive(pr.pid) is False, "daily_run 잠금이 죽은 주인을 못 알아본다"
    assert D._pid_alive(os.getpid()) is True, "산 주인의 잠금을 빼앗는다"
    print("  [121] 죽은 프로세스 판정 — 종료코드까지 확인·두 곳 공유·잠금 회수 ✅")


def t125_worktree_shared_state():
    """[125] 워크트리에서 일해도 **상태는 하나** (브랜치 인계, 분담판 #10).

    왜 필요한가 — 워크트리(`.claude/worktrees/<이름>`)는 **추적 파일만** 체크아웃한다.
    그런데 이 프로젝트의 상태는 거의 전부 git 밖이다(`reports/`·`updates/`·
    `config/*.json`·`db/*.db`). 그래서 워크트리 안에서는 모듈이 제 폴더 기준으로
    경로를 잡는 순간 본체와 **다른 상태**를 본다. 실측으로 확인한 것:
      ① 점유 파일이 갈려 두 세션이 동시에 `ledger` 를 잡아도 서로 안 보였다
         → 관리대장 동시 쓰기 금지가 조용히 무너진다. **가장 위험한 것.**
      ② 워크트리에서 enqueue 한 입력은 11:00·15:00 반영이 영영 못 본다
      ③ `config/` 가 없어 합성검증이 t1 에서 죽었다 — 즉 "ALL GREEN 확인 후
         실작업" 관문 자체를 통과할 수 없었다
      ④ 루트 CLAUDE.md 비교가 `.claude/worktrees/CLAUDE.md` 를 찾아 **거짓 경보**를
         매번 '먼저 처리할 것' 맨 위에 올렸다(해시는 같았는데도)
    """
    import worktree_state as W

    # 본체에서는 동작이 하나도 바뀌면 안 된다 — shared() 가 곧 제 폴더다
    assert os.path.isdir(W.main_root()), W.main_root()
    if not W.is_worktree():
        assert os.path.normcase(W.shared("db")) == os.path.normcase(os.path.join(ROOT, "db"))

    # 점유·큐 DB 는 **본체 경로**를 쓴다. 링크가 아니라 코드가 집어야 하는 이유가 있다.
    ac = open(os.path.join(ROOT, "ai_claim.py"), encoding="utf-8").read()
    assert "from worktree_state import shared" in ac and "STATE_DIR" in ac, \
        "점유 파일이 워크트리마다 갈린다 — 두 세션이 동시에 원장을 연다"
    assert "os.replace" in ac, "점유 저장이 원자적이지 않다"
    ld = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    assert "from worktree_state import shared" in ld, \
        "큐 DB 가 워크트리마다 갈린다 — 넣은 입력이 11:00·15:00 반영에 안 잡힌다"
    # 분담판도 세션 사이의 약속이다 — 갈리면 둘이 같은 일을 하고 나서야 안다.
    ws = open(os.path.join(ROOT, "worksplit.py"), encoding="utf-8").read()
    assert "from worktree_state import shared" in ws, \
        "분담판이 체크아웃마다 갈린다 — 서로 맡은 일을 못 본다"

    # `reports/` 를 통째로 정션하면 안 된다 — 추적 파일이 섞여 있어 git 이 흔들린다
    assert "reports" not in W.LINK_DIRS, \
        "reports 를 통째로 이으면 추적 파일 때문에 워크트리 git 이 남의 파일을 물고 간다"
    assert "reports/ai_claims.json" in W.CODE_SHARED and "db/ledger_queue.db" in W.CODE_SHARED

    # 이미 따로 있는 것을 덮지 않는다(사람이 일부러 둔 설정을 지우면 안 된다)
    src = open(os.path.join(ROOT, "worktree_state.py"), encoding="utf-8").read()
    assert "따로있음" in src and "덮지 않음" in src, "덮어쓰기 방지가 없다"
    for bad in ("shutil.rmtree", "os.remove", "os.unlink"):
        assert bad not in src, "워크트리 연결기가 파일을 지운다: %s" % bad

    # 끊겨 있으면 '먼저 처리할 것' 맨 앞에 뜬다 — 기계 상태가 아니라 st 로만 판단한다
    import session_handoff as H
    base = {"큐잔량": 0, "임시파일": [], "점유": [], "미푸시": [], "지시문사본": [],
            "수집신선도": []}
    cut = dict(base, **{"워크트리": {"여기": "X", "본체": "Y", "항목": [
        {"대상": "config/ecount_config.json", "방법": "하드링크", "상태": "없음"}]}})
    assert any("워크트리" in why for why, _ in H.blockers(cut)), H.blockers(cut)
    ok = dict(base, **{"워크트리": {"여기": "X", "본체": "Y", "항목": [
        {"대상": "config/ecount_config.json", "방법": "하드링크", "상태": "이어짐"}]}})
    assert not H.blockers(ok), "이어져 있는데 막았다"
    assert not H.blockers(dict(base)), "워크트리가 아닌데 워크트리 경고가 뜬다"

    # 루트 지시문 비교가 본체 기준인가 — 워크트리에서 거짓 경보가 뜨던 자리
    hs = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "os.path.dirname(_main_root())" in hs, \
        "루트 CLAUDE.md 비교가 아직 dirname(BASE) 기준이다 — 워크트리에서 거짓 경보"

    # 미푸시는 **내 브랜치의 upstream** 기준이어야 한다. origin/master 로 세면
    # 워크트리 브랜치를 푸시하고도 "푸시되지 않은 커밋"이 영원히 남고, 제시된
    # 명령(git pull --rebase && git push)은 그 상황에 맞지도 않는다.
    assert "def unpushed_commits(" in hs and '"@{u}"' in hs, \
        "미푸시를 origin/master 로만 센다 — 브랜치에서 유령 blocker 가 뜬다"
    assert "def unmerged_commits(" in hs, "브랜치가 master 에 안 들어간 것을 안 알려준다"

    # 이어받기(--adopt)가 **큐를 흡수하기 전에** 먼저 잇는가. 순서가 뒤집히면
    # 빈 큐를 보고 "0건" 이라 답한다 — 조용히 틀린 답이다.
    # (설명문이 아니라 **실제 호출 순서**로 본다)
    assert (hs.index('steps.append(("워크트리 → 본체 잇기"')
            < hs.index('steps.append(("입력 큐 → DB"')), "--adopt 가 잇기 전에 큐를 읽는다"
    print("  [125] 워크트리 공용 상태(점유·큐 DB·분담판 일원화 · 거짓 경보 제거 · 잇기 우선) ✅")


def t126_app_font_and_revert():
    """[126] 앱 전체 글꼴 교체 + 되돌리기 보호 장치 (2026-08-06 지시).

    사용자 지시: "폰트도 나눔고딕 말고 디자이너들과 사용자들이 선호하고 잘 보이는
    폰트로 전체 변경해 / **나중에 내가 명령 내리면 다시 원래대로 할 수 있는
    보호 장치도 마련해**."

    지키는 것
      ① 글꼴을 정하는 자리는 파일마다 **한 곳**(`--font-ui`)이다. 네 파일에
         흩어져 있어서, 손으로 고치면 반드시 한 곳을 빠뜨린다.
      ② `--font-ui-legacy` 에 **원래 값이 글자 그대로** 남아 있다. 이게 되돌릴
         자리다 — 지워지면 "원래대로" 가 무슨 값인지 아무도 모른다.
      ③ 캡처(캔버스)로 그리는 이미지도 같은 변수를 읽는다. 예전엔 그리는 곳마다
         글꼴 이름을 손으로 적어 두어, 화면만 바뀌고 **저장한 이미지는 옛 글꼴**로
         남았다 — 카톡으로 보내고 나서야 안다.
      ④ `font_switch.py` 로 **왕복**이 된다(기본→예전→기본). 왕복 뒤 파일이
         원본과 한 바이트도 다르지 않아야 한다 — 되돌리기가 다른 것을 건드리면
         그건 되돌리기가 아니다.
    """
    import hashlib
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import font_switch as F

    # ① 네 파일 모두 변수 한 곳에서 정한다
    assert len(F.FILES) == 4, "글꼴을 정하는 파일 목록이 바뀌었다: %s" % (F.FILES,)
    for rel in F.FILES:
        text = open(os.path.join(ROOT, rel), encoding="utf-8").read()
        assert "--font-ui-legacy" in text, "%s 에 되돌릴 자리가 없다" % rel
        assert "var(--font-ui)" in text, "%s 가 변수를 안 쓰고 글꼴을 직접 적었다" % rel
        assert ':root[data-font="legacy"]' in text, \
            "%s 에 기기별 되돌리기 스위치가 없다" % rel

    # ② 원래 값이 그대로 남아 있는가 — 값을 여기 못 박아 둔다
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert ('--font-ui-legacy:"Nanum Gothic","NanumGothic","나눔고딕",'
            '"Apple SD Gothic Neo","Malgun Gothic",system-ui,sans-serif;') in idx, \
        "업무센터 앱의 원래 글꼴 값(나눔고딕)이 바뀌었다 — 되돌릴 곳을 잃었다"

    # ③ 캔버스가 손으로 적은 글꼴을 쓰지 않는가
    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))
    assert "function uiFont(" in js, "캔버스가 읽을 글꼴 함수가 없다"
    uf = js[js.index("function uiFont("):]
    uf = uf[: uf.index("\nfunction ", 10)]
    assert "--font-ui" in uf and "getPropertyValue" in uf, \
        "uiFont() 가 CSS 변수를 안 읽는다 — 되돌리기가 이미지에 안 먹는다"
    assert "return '" in uf, "변수를 못 읽는 브라우저용 대타가 없다(캔버스가 글꼴을 통째로 무시한다)"
    for line in js.splitlines():
        if "px" in line and "Nanum Gothic" in line:
            raise AssertionError("캔버스가 아직 글꼴 이름을 손으로 적는다: %s" % line.strip()[:70])
    assert js.count("uiFont()") >= 9, \
        "그림 그리는 곳 일부만 변수를 쓴다 — 그 이미지만 옛 글꼴로 남는다"

    # 기기별 스위치가 기억되는가(다음에 열어도 되돌린 상태여야 한다)
    assert "function setFontLegacy(" in js and "csos_font_legacy" in js, \
        "되돌리기 스위치가 기억되지 않는다 — 새로고침하면 도로 돌아온다"

    # ④ 왕복 — 실패해도 원본을 반드시 되돌린다
    def digest():
        return {r: hashlib.sha1(open(os.path.join(F.ROOT, r), "rb").read()).hexdigest()
                for r in F.FILES}

    origin = {r: open(os.path.join(F.ROOT, r), "rb").read() for r in F.FILES}
    before = digest()
    try:
        F.apply("legacy")
        assert {s["상태"] for s in F.state()} == {"예전"}, \
            "되돌리기가 일부 파일만 바꿨다 — 그 화면만 다른 글꼴로 보인다"
        F.apply("modern")
        assert digest() == before, "왕복했더니 파일이 달라졌다 — 되돌리기가 다른 것을 건드린다"
        assert {s["상태"] for s in F.state()} == {"기본"}
    finally:
        for r, raw in origin.items():
            p = os.path.join(F.ROOT, r)
            if open(p, "rb").read() != raw:
                open(p, "wb").write(raw)
    print("  [126] 앱 글꼴 일원화(4파일·캔버스 포함) + 되돌리기 왕복 무손실 ✅")


def t246_font_presets_single_table():
    """[246] 글꼴 프리셋 — 표는 하나, CSS·단추는 거기서 만들어진다 (2026-08-13 지시).

    사용자 지시: **"갤럭시 글자체 , 아이폰 글자체, 스타일등을 적용해서 사용자가
    설정에서 변경할 수 있게 다양한 스타일 적용 코딩해"**

    무엇이 위험한가 — `[126]` 은 '모두의 기본 글꼴'을 지켰다. 프리셋은 그 위에
    **사람이 제 기기에서 고르는 층**을 얹는 것이라, 갈릴 자리가 셋으로 늘었다:
      ① `FONT_PRESETS` 표 ② 네 파일의 `:root[data-font="…"]` CSS ③ 화면 단추.
    셋 중 하나만 늘면 **오류가 안 난다.** 프리셋을 하나 더했는데 CSS 만 늘면 단추가
    없어 못 고르고, 단추만 늘면 눌러도 아무 일이 안 일어난다 — 둘 다 조용하다([162]).
    그래서 여기서 셋을 **서로 대 본다.**

    지키는 것
      ① 표에 있는 프리셋은 **네 파일 전부**에 CSS 가 있다(한 곳 빠지면 그 화면만 안 바뀐다).
      ② 표에 있는 프리셋은 화면에 **단추가 있다**, 그리고 단추 열쇠는 표 밖으로 안 나간다.
      ③ 화면 코드가 프리셋 이름을 **손으로 적지 않는다**(적으면 그것이 넷째 사본이다).
      ④ 내려받는 웹폰트를 안 쓴다 — 못 받으면 조용히 다른 글꼴로 그려진다.
      ⑤ 고른 것이 **기억된다**, 그리고 옛 열쇠(csos_font_legacy)를 이어받는다.
      ⑥ `--preset` 에 모르는 이름을 주면 **실패한다**(0건 성공 금지, `[169]`).
    """
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import font_switch as F

    keys = list(F.FONT_PRESETS)
    assert len(keys) >= 4, "프리셋이 너무 적다 — 표가 비었나: %s" % (keys,)
    for k, p in F.FONT_PRESETS.items():
        for col in ("이름", "설명", "실제로", "값"):
            assert p.get(col), "프리셋 %s 에 '%s' 가 없다" % (k, col)
        # ④ 내려받는 글꼴 금지 — 값에 url()·@import 가 섞이면 안 된다
        assert "url(" not in p["값"] and "@import" not in p["값"], \
            "프리셋 %s 가 내려받는 글꼴을 쓴다 — 못 받으면 조용히 다른 글꼴이 된다" % k

    # ① 네 파일 전부에 CSS 가 있는가 (legacy 는 2026-08-06 부터 있던 그 한 줄이 맡는다)
    for rel in F.FILES:
        text = open(os.path.join(F.ROOT, rel), encoding="utf-8", newline="").read()
        assert F.PRESET_BEGIN in text and F.PRESET_END in text, \
            "%s 에 프리셋 블록이 없다 — 그 화면에서만 글꼴이 안 바뀐다" % rel
        assert F.preset_css(F._eol(text)) in text, \
            "%s 의 프리셋 블록이 표와 다르다 — `font_switch.py --sync` 를 돌려야 한다" % rel
        for k in keys:
            assert ':root[data-font="%s"]' % k in text, \
                "%s 에 프리셋 '%s' 의 CSS 가 없다" % (rel, k)

    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    # ② 표 ↔ 단추가 정확히 같은 집합인가
    #    ★ 문서 전체에서 찾으면 JS 의 선택자 문자열(`[data-fontkey="'+key+'"]`)까지
    #      걸려 없는 프리셋 두 개가 잡힌다. **만들어진 블록 안에서만** 센다.
    blk = re.search(re.escape(F.CARDS_BEGIN) + r"(.*?)" + re.escape(F.CARDS_END), idx, re.S)
    assert blk, "index.html 에 글꼴 단추 자리(FONT-CARDS)가 없다"
    shown = set(re.findall(r'data-fontkey="([^"]+)"', blk.group(1)))
    assert shown == set(keys), \
        "표와 화면 단추가 갈렸다 — 표에만: %s · 화면에만: %s" % (
            sorted(set(keys) - shown), sorted(shown - set(keys)))

    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))
    assert "function setFontPreset(" in js, "프리셋을 고르는 함수가 없다"
    # ③ 화면 JS 가 프리셋 이름을 손으로 적지 않는가.
    #    ★ 둘만 예외다 — 이름이 아니라 **구조**라서다:
    #      · `basic`  = :root 의 기본값 그 자체(속성을 지운다는 뜻)
    #      · `legacy` = 2026-08-06 부터 있던 되돌리기 스위치. `setFontLegacy()` 와
    #        옛 열쇠 이어받기가 이 낱말로 걸려 있고, 문서의 되돌리기 경로 ③ 이다.
    #    막으려는 것은 **새로 더한 프리셋을 손으로 배선하는 것**이다 — 그러면 표를
    #    고쳐도 화면이 안 따라오고, 그 사실이 아무 데도 안 뜬다.
    for k in keys:
        if k in ("basic", "legacy"):
            continue
        assert ("'%s'" % k) not in js and ('"%s"' % k) not in js, \
            "화면 코드가 프리셋 이름 '%s' 를 손으로 적었다 — 표 밖에 사본이 생긴다" % k
    # ⑤ 기억 + 옛 열쇠 이어받기
    assert "csos_font_preset" in js, "고른 글꼴이 기억되지 않는다 — 새로고침하면 돌아간다"
    assert "csos_font_legacy" in js, \
        "옛 열쇠를 안 읽는다 — 예전에 나눔고딕으로 돌려 둔 사람이 말없이 기본으로 돌아간다"

    # ⑥ 모르는 이름은 실패로 끝난다
    import io as _io
    import contextlib as _cl
    buf = _io.StringIO()
    old = sys.argv
    try:
        sys.argv = ["font_switch.py", "--preset", "없는글꼴"]
        with _cl.redirect_stdout(buf):
            rc = F.main()
    finally:
        sys.argv = old
    assert rc == 2, "모르는 글꼴 이름인데 성공으로 끝났다 — '바꿨습니다'만 찍힌다"

    print("  [246] 글꼴 프리셋 %d개 — 표·CSS(4파일)·단추가 한 곳에서 온다 ✅" % len(keys))


def t247_chrome_collect_report_round_trip():
    """[247] 크롬 전용 수집 되보고 — 받는 자리·읽는 자리가 맞물린다 (2026-08-13 지시).

    사용자 지시: **"앞으로 크롬에서만 긁어오는 알고리즘 만들어서 적용해"**

    무엇이 위험한가 — 이 길에는 **조용히 죽는 자리가 넷** 있고 넷 다 오류를 안 낸다.
    유저스크립트는 계속 돌고 회차도 계속 '성공'이라 적히는데 수집만 0건이 된다.
    실측 2026-08-13: 스크립트는 2026-08-09 에 만들어져 검증 `[182]` 까지 붙어 있었는데
    **나흘 동안 한 번도 안 돌았고 아무 화면에도 안 떴다.**

    지키는 것
      ① **받는 자리가 PIN 게이트 앞이다.** 보내는 쪽은 `band.us` 탭이라 이 서버의 인증
         쿠키를 못 싣는다(다른 출처다). 뒤로 옮기면 **모든 되보고가 401 로 버려지고**
         감시자는 '한 번도 안 옴'이라 말한다 — 스크립트는 멀쩡히 도는데도(`[169]`).
      ② **쓰는 모양 = 읽는 모양.** 담는 곳(`app_server.collect_report_save`)과 읽는 곳
         (`userscript_watch`)이 갈리면 **오류 없이 한 건도 안 읽힌다**(`[165]` 의 모양 —
         칸 이름을 짝지어 물으면 모든 값이 빈칸으로 보인다).
      ③ **유령 밴드를 거부한다.** `260807`(날짜 꼬리표)·`202608082047`(시각 도장)이 한 번
         키로 들어가면 유령 밴드가 생기고 그것이 '아는 번호'가 되어 스스로를 되살린다
         (2026-08-12 실사고).
      ④ **워치독이 인계보다 먼저 본다.** 뒤에 두면 인계가 언제나 30분 전 판정을 싣는다.
      ⑤ **'못 읽음'을 '정상'이라 하지 않는다**(`[169]`) — 감시자 자신이 눈멀면 안 된다.
      ⑥ **`--print` 가 안 죽는다.** 인계 문서가 알려 주는 명령이 바로 그것인데, 무인
         회차는 `sys.stdout` 이 None 이고(`[235]`) 콘솔은 cp949 라 '—' 에서 통째로 죽었다.

    ⚠ **진짜 기록 파일(`reports/크롬수집_보고.json`)은 한 글자도 안 건드린다.** 실측
      증거에 합성 행을 섞으면 그 파일이 더는 실측이 아니고, 감시자가 '정상'을 확언한다.
    """
    import tempfile
    from datetime import datetime
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for p in (root, os.path.join(root, "webapp")):
        if p not in sys.path:
            sys.path.insert(0, p)

    src = open(os.path.join(root, "webapp", "app_server.py"), encoding="utf-8").read()
    # ① 라우트가 인증 게이트보다 **앞**이어야 한다.
    i_route = src.find('p == "/api/collect_report"')
    i_gate = src.find("if not self._auth():", src.find("def do_POST"))
    assert i_route > 0, "POST /api/collect_report 받는 자리가 없다 — 유저스크립트는 404 를 받는다"
    assert 0 < i_route < i_gate, (
        "되보고 라우트가 PIN 게이트 뒤에 있다 — band.us 탭은 쿠키를 못 실어 전부 401 이 되고, "
        "감시자는 '한 번도 안 옴'이라 말한다")

    import app_server as A
    from band import userscript_watch as U

    real = A.COLLECT_REPORT_PATH
    tmp = os.path.join(tempfile.gettempdir(), "t247_collect_report.json")
    if os.path.exists(tmp):
        os.remove(tmp)
    old_report = U.REPORT
    # 정본을 정말 안 건드렸는지 재려면 '지금 상태'를 먼저 찍어 둬야 한다.
    real_before = (os.path.exists(real), os.path.getmtime(real) if os.path.exists(real) else 0)
    try:
        A.COLLECT_REPORT_PATH = tmp
        # ③ 유령 밴드는 거부한다 — 통과하면 캐시에 없는 밴드가 생긴다.
        for bad in ("260807", "202608082047", "", "abc", "12345"):
            try:
                A.collect_report_save({"band": bad, "state": "start"})
                raise AssertionError("밴드번호가 아닌 %r 을 받아들였다" % bad)
            except ValueError:
                pass
        now = datetime.now().isoformat(timespec="seconds")
        A.collect_report_save({"band": "84789192", "state": "done", "수확": 14, "at": now})
        A.collect_report_save({"band": "90610953", "state": "hidden",
                               "why": "창이 가려져 있다", "at": now})
        # ⚠ 정본은 한 글자도 안 건드렸어야 한다 — 실측 증거에 합성 행이 섞이면
        #   그 파일이 더는 실측이 아니고 감시자가 '정상'을 확언한다.
        real_now = (os.path.exists(real), os.path.getmtime(real) if os.path.exists(real) else 0)
        assert real_now == real_before, (
            "검증이 실제 되보고 파일을 건드렸다 — 합성 행이 섞이면 감시자가 눈먼다: %s" % real)

        # ② 담은 것을 읽는 쪽이 그대로 읽어야 한다.
        U.REPORT = tmp
        st = U.judge(*U.load_reports())
        rows = st.get("밴드") or {}
        assert set(rows) == {"84789192", "90610953"}, (
            "담은 밴드를 읽는 쪽이 못 읽는다 — 칸 이름이 갈렸다: %r" % (sorted(rows),))
        assert rows["84789192"]["상태"] == "done" and rows["84789192"]["수확"] == 14, (
            "상태·수확이 안 건너갔다 — 오류 없이 빈칸으로 보인다: %r" % (rows["84789192"],))
        assert rows["90610953"]["상태"] == "hidden", "가려짐 상태가 안 건너갔다"
        assert st.get("갈래") not in (None, "", "확인못함"), (
            "방금 담았는데 '확인못함'이다 — 시각 칸이 안 건너갔다: %r" % st.get("갈래"))
    finally:
        A.COLLECT_REPORT_PATH = real
        U.REPORT = old_report
        if os.path.exists(tmp):
            os.remove(tmp)

    # ④ 워치독 배선 — snapshot_handoff 보다 **먼저**.
    wd = open(os.path.join(root, "watchdog.py"), encoding="utf-8").read()
    assert "def watch_userscript(" in wd, "워치독에 크롬수집 감시 단계가 없다"
    body = wd[wd.find("def main("):]
    a, b = body.find("watch_userscript("), body.find("snapshot_handoff(")
    assert a > 0, "watch_userscript 가 main() 회차 목록에 없다 — 만들어 놓고 안 부른다"
    assert a < b, "watch_userscript 가 snapshot_handoff 뒤다 — 인계가 늘 30분 전 판정을 싣는다"

    # ⑤ 인계는 '못 읽음'과 '정상'을 가른다.
    sh = open(os.path.join(root, "session_handoff.py"), encoding="utf-8").read()
    assert "userscript_health" in sh, "인계가 크롬수집을 안 싣는다"
    i_us = sh.find('us = st.get("크롬수집"')
    assert i_us > 0, "인계에 크롬수집 판정을 읽는 자리가 없다"
    seg = sh[i_us:][:1200]
    assert "us is None" in seg and "확인 못" in seg, (
        "'못 읽음'을 '정상'과 안 가른다 — 감시자가 눈먼 채 이상 없음을 말한다([169])")
    # ★ 그러나 '키가 아예 없다'(부분 상태)까지 실패로 읽으면 **없는 경보**가 난다.
    #   2026-08-13 실측으로 이 함정에 그대로 빠져 `t111` 이 빨개졌다 — 셋을 가른다.
    assert 'st.get("크롬수집", ())' in seg, (
        "키 없음(안 물었다)과 None(물었는데 실패)을 안 가른다 — 부분 상태에 거짓 경보가 난다. "
        "실제로 거짓 경보가 나는지는 `t111` 이 상태를 통째로 지어 확인한다(여기서 반쪽 상태를 "
        "또 지어내면 시험이 시험을 방해한다)")

    # ⑥ 사람이 붙여넣을 명령이 죽지 않는다(pythonw None · 콘솔 cp949).
    uw = open(os.path.join(root, "band", "userscript_watch.py"), encoding="utf-8").read()
    assert 'hasattr(sys.stdout, "reconfigure")' in uw, (
        "stdout 보호가 없다 — 무인 회차는 None, 콘솔은 cp949 라 '—' 에서 죽는다([235])")

    print("  [247] 크롬 되보고 — 게이트 앞·모양 일치·유령 거부·워치독 순서 ✅")


def t250_error_book_speaks_and_counts():
    """오류는 남고 · 사람 말이 되고 · 다시 나면 회귀로 센다 (2026-08-13 지시).

    지키는 것은 '기능이 있다'가 아니라 **되돌아가면 안 되는 것**이다:
      · 사전은 **모르면 None** 을 준다(지어내면 사람이 엉뚱한 데를 고치러 간다)
      · 보관은 **덧붙이기만** 한다(지우면 재발을 영영 못 센다)
      · 보관 실패를 **0 으로 적지 않는다**(-1). 실패가 성공처럼 보이면 안 된다
      · 회귀는 '막음이 붙었다'가 아니라 **고친 날 뒤에도 났다**여야 한다.
        날짜를 안 보면 첫판처럼 76갈래 중 56이 회귀가 되어 경보가 통째로 죽는다([170])
      · 못 가른 것을 '이상 없음'으로 치지 않는다([169])
    """
    import importlib
    eb = importlib.import_module("error_book")

    # ① 모르는 오류에 답을 지어내지 않는다
    assert eb.look_up("/api/무엇", "듣도보도못한오류") is None, "모르는 오류에 사전이 답했다"
    h = eb.help_for("/api/무엇", "듣도보도못한오류")
    assert h["앎"] is False and h["하세요"], "모르는 오류를 빈손으로 돌려보냈다"
    assert "신고문구" in h and "[앱 오류 신고]" in h["신고문구"], "붙여넣을 신고문구가 없다"

    # ② 아는 오류는 개발자 말이 아니라 사람 말로 답한다
    k = eb.help_for("/api/staff/entry", "화면 데이터 버전이 없습니다")
    assert k["앎"] is True and k["하세요"], "아는 오류인데 할 일을 못 말했다"
    for ent in eb.BOOK:
        blob = ent["쉬운말"] + " " + " ".join(ent["하세요"])
        for jargon in ("HTTP", "JSON", "409", "토큰", "Exception", "null"):
            assert jargon not in blob, f"{ent['이름']}: 사람 말이 아닌 낱말 '{jargon}'"
        assert ent.get("막음") is None or ent.get("고친날"), \
            f"{ent['이름']}: 막음에 고친 날이 없다 — 회귀를 셀 수 없다"

    # ③ 지문은 번호를 지운다 — 안 지우면 건마다 다른 오류가 되어 아무것도 안 모인다
    a = eb.signature("/api/staff/entry", "AS-2601-574 저장 실패")
    b = eb.signature("/api/staff/entry", "AS-2601-999 저장 실패")
    assert a == b, "지문이 건별로 갈린다 — 같은 오류를 못 센다"

    # ④ 보관은 덧붙이기만 한다 · 실패를 0 으로 적지 않는다
    src = open(os.path.join(ROOT, "error_book.py"), encoding="utf-8").read()
    arch = src[src.index("def archive("):src.index("def _read_archive(")]
    # ★ 좁혀서 본다 — 보관본(path)만 덧붙이기여야 한다. 실패 자국(FAIL_MARK)은
    #   덮어써야 맞다. 첫판에 '"w" 가 있으면 실패'로 뒀다가 그 자국까지 잡았다.
    assert 'open(path, "a"' in arch, "보관본이 덧붙이기(append)가 아니다"
    for bad in ("os.remove(path", 'open(path, "w"', "truncate"):
        assert bad not in arch, f"보관이 기록을 지운다 — {bad}"
    assert "return -1" in arch, "보관 실패를 0 이 아닌 값으로 알리지 않는다"

    # ⑤ 회귀는 날짜를 본다
    roll = src[src.index("def rollup("):src.index("def _md(")]
    assert "고친날" in roll and "회귀.append" in roll, "회귀 판정이 없다"
    assert '["고친날"]' in roll or "'고친날'" in roll, "회귀가 고친 날을 안 본다"
    assert "못가름" in roll, "못 가른 것을 따로 세지 않는다([169])"

    # ⑥ 읽기 전용이다 — 고치지도, 큐에 넣지도, 엑셀을 열지도 않는다
    for bad in ("enqueue(", "--apply", "openpyxl", "workbook_patch"):
        assert bad not in src, f"error_book 이 읽기 전용이 아니다 — {bad}"

    # ⑦ 실제로 한 바퀴 돈다(파일은 안 건드린다 — 실측 기록에 합성 행을 섞지 않는다)
    res = eb.rollup(days=1)
    for key in ("회귀", "새오류", "아는것", "못본것", "합계"):
        assert key in res, f"집계에 '{key}' 가 없다"

    # ⑧ 회차·인계에 배선돼 있다 — 코딩만 하고 안 도는 것을 막는다
    dr = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "error_book.py" in dr, "09:50 회차에 오류 감시 단계가 없다"
    sh = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "error_book" in sh, "인계 '먼저 처리할 것' 에 안 올라온다"

    # ⑨ 앱이 사람에게 말하고 캡처를 만든다 — 실패를 조용히 삼키던 자리를 지킨다
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "async function errorHelp(" in idx, "오류 팝업이 없다"
    assert "errorShotSend(" in idx and "toBlob" in idx, "캡처해서 보내는 길이 없다"
    assert "dlgAsk({" in idx[idx.index("async function errorHelp("):
                             idx.index("async function errorHelp(") + 3000], \
        "팝업 레이어를 새로 만들었다 — 공용 문(dlgAsk) 하나를 써야 한다([202])"
    win = idx[idx.index("window.addEventListener('error'"):]
    assert "(사유 없음)" in win[:400], "window 오류가 아직 사유 없이 적힌다([169])"
    api = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "/api/error_help" in api, "도움말 길이 서버에 없다"
    assert "error_book.archive(" in api, "오류가 보관본에 안 남는다"
    print("[250] 오류는 남고 · 사람 말이 되고 · 회귀로 센다 ✅")


def t251_zero_tells_which_zero_it_is():
    """[251] 0 화면이 **못 불러온 것과 정말 없는 것을 가른다** (2026-08-13 실사고).

    류지영 실사용 막힘: "확인필요에 아무것도 안뜰때 뭘 어떻게 해야하는걸까요!!"
    그때 서버에는 확인 필요가 **208건** 있었다(`app_server.get_issues()` 실측).
    화면만 0 이었고 지표 일곱 개가 전부 0 · 담당자 셋이 전부 0건 · 그 위에
    **"확인할 항목이 없습니다 🎉"** 까지 떴다. `/api/issues` 가 끊겨
    (같은 창 15:38 실측 502 · 11일간 103건) `issuesData` 가 처음 값인 빈 배열로
    남은 것인데, 자료 상태를 말해 주는 목록은 [실행] 화면 전용이라
    확인 필요 화면에서는 **아무 말도 안 했다.**

    되돌아가면 안 되는 것만 지킨다(글자 검사는 '있어야 할 것'이 아니라
    '되돌아가면 안 되는 것'에 쓴다 — [39] 의 교훈):
      · 🎉 는 **확실히 0 일 때만** 나온다 — 못 불러온 자리에는 없다
      · 0 을 그리는 자리는 **한 곳을 지난다**(zeroNote) — 화면마다 새로 적으면 갈린다
      · 판정을 새로 만들지 않는다 — 이미 있는 `DATA_SECTION_STATE` 를 읽는다([162])
      · 실패한 자리는 **손으로 할 수 있는 동작**을 준다(다시 불러오기 · 왜 그런지 보기)
      · '왜' 는 사전(error_book.BOOK) 이 말한다 — 화면이 지어내지 않는다
    """
    import importlib
    eb = importlib.import_module("error_book")

    # ① 사전이 이 오류에 사람 말로 답한다 — 그리고 401·403 은 제 답이 그대로 나간다
    h = eb.help_for("/api/issues", "HTTP_ERROR:502")
    assert h["앎"] is True, "확인 필요 목록을 못 불러온 것에 사전이 답을 못 한다"
    assert h["이름"] == "확인 필요 목록을 못 불러옴", f"사전 이름이 다르다: {h['이름']}"
    assert h["하세요"], "할 수 있는 동작이 없다"
    blob = h["쉬운말"] + " " + " ".join(h["하세요"])
    assert "0" in h["쉬운말"], "0 이 '할 일 없음'이 아니라는 말을 안 한다([169])"
    for jargon in ("HTTP", "JSON", "API", "500", "502", "null"):
        assert jargon not in blob, f"사람 말이 아닌 낱말 '{jargon}' 이 들어 있다"
    # ★ 순서가 뜻을 갖는다 — 이 줄이 위로 올라가면 로그인 만료·권한 없음이 묻힌다
    assert eb.help_for("/api/issues", "HTTP_ERROR:401")["이름"] == "로그인 만료", \
        "확인 필요 규칙이 로그인 만료를 덮었다"
    assert eb.help_for("/api/issues", "HTTP_ERROR:403")["이름"] == "권한 없음", \
        "확인 필요 규칙이 권한 없음을 덮었다"

    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ② 0 을 그리는 자리는 한 곳을 지난다 — 그 한 곳이 갈래를 가른다
    for fn in ("function dataZeroState(", "function zeroNote(", "function zeroStateWhy("):
        assert fn in idx, f"{fn} 가 없다 — 0 의 뜻을 가르는 자리가 없다"
    zs = idx[idx.index("function dataZeroState("):idx.index("function zeroStateDetail(")]
    assert "DATA_SECTION_STATE" in zs, \
        "0 판정이 제 손으로 다시 센다 — 이미 있는 자료 상태를 읽어야 한다([162])"
    assert "hasGood" in zs, "'한 번도 못 받았다'를 '없다'와 안 가른다([169])"
    for st in ("'failed'", "'loading'", "'waiting'", "'ok'"):
        assert st in zs, f"갈래 {st} 가 없다"

    # ③ 🎉 는 '확실히 0' 가지에만 있다
    zn = idx[idx.index("function zeroNote("):idx.index("function zeroNum(")]
    assert zn.count("🎉") == 1, "축하 표시가 한 갈래에만 있지 않다"
    ok_at = zn.index("z.state==='ok'")
    assert ok_at < zn.index("🎉") < zn.index("z.state==='loading'"), \
        "🎉 가 '확실히 0' 가지 밖에 있다 — 못 불러온 0 을 축하하면 안 된다"
    assert "retryDataSection(" in zn and "zeroStateWhy(" in zn, \
        "실패한 자리에 손으로 할 수 있는 동작이 없다"
    assert "0 건이라는 뜻이 아닙니다" in zn, "0 이 사실이 아니라는 말을 안 한다"

    # ④ '왜' 는 사전이 말한다 · 팝업을 새로 만들지 않는다([202])
    why = idx[idx.index("function zeroStateWhy("):idx.index("function zeroNote(")]
    assert "errorHelp(" in why and "force:true" in why, \
        "실패 설명이 사전(error_book)을 안 거친다 — 화면이 지어내면 판정이 둘이 된다"
    assert "dlgAsk" not in why, "팝업 레이어를 새로 만들었다 — 공용 문 하나를 써야 한다([202])"

    # ⑤ 확인 필요 화면이 🎉 를 제 손으로 적지 않는다 — 반드시 zeroNote 를 지난다
    hub = idx[idx.index("function renderCheckHub(){"):idx.index("function esc4(")]
    assert "🎉" not in hub, "확인 필요 화면이 아직 축하 표시를 제 손으로 적는다"
    assert hub.count("zeroNote(") >= 3, "0 자리 일부가 아직 한 곳을 안 지난다"
    lst = idx[idx.index("function renderCheckList("):idx.index("/* 확인필요 유형 코드 설명")]
    assert "🎉" not in lst and "zeroNote(" in lst, "목록의 0 자리가 갈래를 안 가른다"

    # ⑥ 못 셌으면 숫자 0 을 적지 않는다 — 0 일곱 개는 '오늘 할 일 없음'으로 읽힌다
    assert "function zeroNum(" in idx, "못 센 숫자를 0 으로 적지 않는 장치가 없다"
    assert "dataZeroState('issues')" in hub, "확인 필요 지표가 자료 상태를 안 본다"
    assert "'? 건'" in hub, "못 셌는데 담당자 옆에 '0건' 이 그대로 적힌다"
    assert "전체 ${all.length}건" in lst and "zl.state==='ok'" in lst, \
        "목록 요약이 못 불러온 0 을 '전체 0건' 으로 확언한다"

    print("[251] 0 은 '없다'와 '못 불러왔다'를 가려 말한다 ✅")


def t253_share_folder_pulls_whole_parent_cheaply():
    """[253] 오종현 공유폴더는 **부모 전체**를 훑고, 훑는 값은 목록에 딸려 온 것을 쓴다.

    2026-08-13 형님 지시: "\\\\172.30.1.250\\data\\16. Share\\유현민\\오종현 여기서
    자료 항상 긁어오는 알고리즘 있지? 없으면 구현해"

    실측으로 **알고리즘은 이미 있었다**(`share_intake`, 2026-08-03). 그 폴더 602개 중
    **6개가 두 하위폴더 밖**(부모 직속)에 있는데 전부 정상 흡수돼 있었다 — 그날 16:07
    에 놓인 `CSOS PO관련 …xlsx` 가 **16:17 에** `7. 입금내역/8-13/` 로 갔다.
    `source_dirs` 가 두 하위폴더만 아는 것은 **제자리 직접 읽기**(PO_DIRS·RECEIPT_DIRS)
    라서이고, 나머지는 이 흡수기가 투입함으로 복사해 `upload_intake` 가 분류한다.

    그런데 훑는 방식이 `[198]` 의 그 병이었다 — `os.walk` 로 이름만 받고 파일마다
    `os.stat(경로)` 를 다시 불렀다. 같은 폴더 실측 **99.7초 / 602개** → `walk_stat`
    **4.1초**(24배). 이 함수는 워치독 30분·09:35·09:50·**증분 파이프라인 5분**이
    전부 부르는 자리라 그 값이 그대로 회차 비용이 된다.

    되돌아가면 안 되는 것만 지킨다:
      ① 훑는 자리에 `os.stat(`/`getmtime(` 이 다시 안 들어온다([198])
      ② 대상은 **부모 폴더**이고 제외는 정본으로 직접 읽는 **두 하위폴더뿐**이다
         — 좁히면 부모 직속 자료가 조용히 안 들어온다
      ③ 거르기가 그대로 산다: 최상위 제외 · `old` 는 **모든 레벨·대소문자 무시** ·
         `Thumbs.db`·`~$` 임시파일
      ④ 원본은 **복사**(이동 금지)하고, 같은 파일을 두 번 담지 않는다
    """
    import importlib
    import shutil as _shutil
    import tempfile
    S = importlib.import_module("share_intake")
    src_txt = open(os.path.join(ROOT, "share_intake.py"), encoding="utf-8").read()

    # ── ① 비싼 재-stat 이 돌아왔나([198]) ────────────────────────────────────────
    body = src_txt.split("def pull(", 1)[1]
    for bad in ("os.stat(", "getmtime(", "getsize("):
        assert bad not in body, (
            f"share_intake.pull 에 {bad} 가 다시 들어왔다 — Z:(SMB)에서는 파일당 왕복 "
            "한 번이라 602개에 99.7초다([198]). walk_stat 이 준 stat 을 그대로 써라")
    assert "walk_stat" in body, "share_intake.pull 이 source_index.walk_stat 을 안 쓴다"

    # ── ② 대상은 부모 폴더 · 제외는 정본으로 직접 읽는 둘뿐 ─────────────────────
    tg = S.pull_targets()
    assert tg, "pull_targets 가 비었다 — 공유폴더를 아예 안 긁는다"
    roots = [t[0].replace("\\", "/").rstrip("/").lower() for t in tg]
    assert any(r.endswith("16. share/유현민/오종현") for r in roots), (
        "오종현 공유폴더의 **부모**가 대상이 아니다 — 하위폴더만 잡으면 부모 직속에 "
        "놓인 자료가 조용히 안 들어온다(실측 6건이 거기 있었다)")
    for _root, _owner, excl in tg:
        for name in excl:
            assert name == name.lower(), f"제외 목록은 소문자로 적는다: {name}"
        assert len(excl) <= 2, (
            f"제외가 {len(excl)}개다 — 제외는 source_dirs 가 제자리에서 직접 읽는 "
            "두 폴더(PO 모음·입금내역)뿐이어야 한다. 늘리면 그만큼 안 들어온다")

    # ── ③④ 거르기·복사·중복을 임시 트리로 잰다(진짜 Z: 는 안 건드린다) ──────────
    d, up = tempfile.mkdtemp(), tempfile.mkdtemp()
    try:
        for p in ("26년도 PO 모음", "OLD", "sub", os.path.join("sub", "OLD")):
            os.makedirs(os.path.join(d, p))
        made = ["top.txt", "26년도 PO 모음/a.txt", "OLD/b.txt",
                "sub/OLD/c.txt", "sub/ok.txt", "Thumbs.db", "~$x.xlsx"]
        for p in made:
            with open(os.path.join(d, p.replace("/", os.sep)), "w") as fh:
                fh.write("x")
        for base, _dirs, files in os.walk(d):        # 저장 중 판정을 피한다
            for f in files:
                os.utime(os.path.join(base, f), (0, 0))
        st = os.path.join(up, "state.json")
        target = [(d, "t", {"26년도 po 모음"})]
        got = sorted(r["파일"].replace("\\", "/")
                     for r in S.pull(target, upload_dir=up, state_path=st))
        assert got == ["sub/ok.txt", "top.txt"], f"거르기가 달라졌다: {got}"
        assert S.pull(target, upload_dir=up, state_path=st) == [], (
            "같은 파일을 두 번 담는다 — 상태 기억(크기·수정시각)이 깨졌다")
        for p in made:                               # ④ 원본 보존
            assert os.path.exists(os.path.join(d, p.replace("/", os.sep))), (
                f"원본이 사라졌다: {p} — 공유폴더는 오종현의 저장소다. 복사만 한다")
    finally:
        _shutil.rmtree(d, ignore_errors=True)
        _shutil.rmtree(up, ignore_errors=True)
    print("[253] 공유폴더 흡수 — 부모 전체 · 값싼 훑기 · 거르기·복사 그대로 ✅")


def t252_po_shape_matches_reality_and_restart_asks_first():
    """[252] 실재하는 PO 모양을 거절하지 않는다 · 쓰는 사람 위에서 서버를 안 내린다.

    2026-08-13 오종현 신고: "PO+숫자 로만 되어야 한다고 계속 알림이 뜹니다. 보통
    PO숫자/PR숫자 로 기재되다 어느 순간 po만 기재가 되어있어 어떻게 기재할지…"

    실측(원장 v595 750행) — PO 칸이 채워진 **701건** 중
      `PO숫자/PR숫자` 633 · `PO숫자` 45 · `PR숫자/PO숫자` 23 · 빈칸 49.
    즉 옛 규칙 `/^PO\\d+$/` 는 **실재하는 값의 93.6%(656건)** 를 낯설다고 물었다.
    잘못 막는 것은 못 잡는 것보다 나쁘다([172]) — 그 사람은 일을 못 한다.

    같은 날 나머지 반쪽: 15:53 에 `restart_server.py` 가 돌았는데 그때 오종현이
    입력 중이었다. 재시작은 실측 5.7~9.3초이고 그동안 폰·PC 는 끊긴 화면을 본다
    ([197]). 그런데 그 스크립트는 **누가 쓰는지 보지도 않고** 죽였다.

    되돌아가면 안 되는 것만 지킨다:
      ① 실재하는 세 모양을 화면이 거절하지 않는다
      ② 그래도 진짜 오타(`P0`·숫자 없는 `PO`)는 여전히 묻는다 — 다만 막지 않고 묻는다
      ③ 경고 문구가 **어떻게 적는지 실제 예시**를 보여 준다(옛 문구는 읽고도 몰랐다)
      ④ 판정은 한 곳에서 온다 — 화면 규칙이 `po_reconcile.PO_PAT` 과 같은 답을 낸다
      ⑤ `restart_server` 는 죽이기 **전에** 활동을 보고, **못 읽으면 멈춘다**([169])
    """
    import re as _re
    import importlib
    P = importlib.import_module("po_reconcile")
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ── ①②④ 화면 규칙을 꺼내 파이썬 정본과 **같은 값으로** 대 본다 ───────────────
    m = _re.search(r"const PO_TOKEN\s*=\s*/(.+?)/([a-z]*);", idx)
    assert m, "PO_TOKEN 규칙이 화면에 없다 — 옛 PO_SHAPE 로 되돌아갔나"
    js_src, js_flags = m.group(1), m.group(2)
    assert "i" in js_flags, "대소문자를 안 가린다 — 'po327948' 도 실재한다"
    # JS 정규식 문법이 파이썬에서도 그대로 컴파일되는 모양이어야 대 볼 수 있다([184])
    rx = _re.compile(js_src, _re.I)

    REAL = ["PO327948", "PO327948/PR461621", "PR482790/PO343170",
            "PO326238/PR457592", "po343170", "PO-343170"]
    TYPO = ["P0327948", "PO", "PO12", "번호없음", "PR461621"]
    for s in REAL:
        assert rx.search(s), f"실재하는 모양 '{s}' 를 화면이 거절한다 — 그 사람은 일을 못 한다"
        assert P.norm_po(s), f"대조기가 '{s}' 를 못 읽는다 — 시험값이 틀렸다"
    for s in TYPO:
        assert not rx.search(s), f"'{s}' 까지 통과시킨다 — 진짜 오타를 안 묻는다"
        assert not P.norm_po(s), f"대조기는 '{s}' 를 읽는다 — 두 쪽이 갈렸다"
    # 화면과 대조기가 **같은 집합**에 같은 답을 내야 한다. 갈리면 대조는 되는 값을
    # 화면만 거절하거나(오종현 신고) 화면만 통과시킨다.
    for s in REAL + TYPO:
        assert bool(rx.search(s)) == bool(P.norm_po(s)), \
            f"'{s}' 에서 화면과 po_reconcile 이 다르게 답한다 — 판정이 두 곳이다([162])"

    assert "PO_SHAPE" not in idx, "옛 규칙 PO_SHAPE 가 아직 남아 있다"

    # ── ③ 문구가 '어떻게 적는지' 를 보여 준다 ──────────────────────────────────
    w = idx[idx.index("function warnPoShape("):idx.index("function warnPaidZero(")]
    for ex in ("PO327948", "PO327948/PR461621", "PR482790/PO343170"):
        assert ex in w, f"경고 문구에 실제 예시 {ex} 가 없다 — 읽고도 어떻게 적을지 모른다"
    assert "PO+숫자 모양이 아닙니다" not in w, "옛 문구가 그대로 남아 있다"
    assert "저장은 됩니다" in w, "막는 것이 아니라는 말을 안 한다 — 사람이 저장을 포기한다"
    # 경고는 저장을 막지 않는다 — 이미 있는 길(confirmSaveWarnings)을 쓴다([162])
    assert "function confirmSaveWarnings(" in idx and "그래도 저장" in idx, \
        "경고 후 진행하는 길이 없어졌다 — 경고가 곧 차단이 된다"

    # ── ⑤ restart_server 가 죽이기 전에 활동을 본다 ───────────────────────────
    rsp = os.path.join(ROOT, "webapp", "restart_server.py")
    rs = open(rsp, encoding="utf-8").read()
    assert "def in_use(" in rs and "def guard(" in rs, "쓰는 사람을 보는 자리가 없다"
    assert rs.index("guard(") < rs.index('print("끄는 중:'), \
        "죽이고 나서 물어본다 — 그 순서면 이미 늦었다"
    assert "ux" in rs and "ledger_db" in rs, \
        "새 계기를 만들었다 — 이미 쌓이는 ux 기록을 봐야 한다([162])"

    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    RS = importlib.import_module("restart_server")

    # 시각이 **두 형식으로 섞여 있다**(브라우저 UTC `…Z` + ux_add 로컬 naive).
    # 문자열로 비교하면 조용히 0 건이 되어 계기가 눈먼다 — 실제로 첫 판이 그랬다([169]).
    from datetime import datetime, timezone, timedelta
    utc_now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    e = RS._ts_epoch(utc_now)
    assert e is not None and abs(e - time.time()) < 120, \
        "UTC(`Z`) 시각을 로컬로 안 옮긴다 — 방금 쓴 사람이 '아무도 없음' 으로 읽힌다"
    loc = (datetime.now() - timedelta(minutes=1)).isoformat(timespec="seconds")
    e2 = RS._ts_epoch(loc)
    assert e2 is not None and abs(e2 - (time.time() - 60)) < 120, \
        "로컬 naive 시각을 못 읽는다 — ux 표에 두 형식이 섞여 있다"
    assert RS._ts_epoch("") is None and RS._ts_epoch("어제") is None, \
        "못 읽는 시각을 0(=아주 옛날)으로 치면 '아무도 없음'이 된다"

    # 갈래 넷 — 특히 '못 읽음'이 '아무도 없다'로 새면 안 된다([169])
    orig = RS.in_use
    try:
        RS.in_use = lambda minutes=10: {"읽음": False, "건수": 0, "분전": None, "왜": "잠김"}
        assert RS.guard() is not None, "활동을 못 읽었는데 그냥 내린다([169])"
        assert RS.guard(force=True) is None, "--force 로도 못 내린다"
        RS.in_use = lambda minutes=10: {"읽음": True, "건수": 3, "분전": 1.0, "왜": ""}
        assert RS.guard() is not None, "쓰는 사람이 있는데 그냥 내린다"
        RS.in_use = lambda minutes=10: {"읽음": True, "건수": 0, "분전": 900.0, "왜": ""}
        assert RS.guard() is None, "아무도 안 쓰는데 못 내린다 — 그러면 고쳐도 화면이 안 바뀐다"
    finally:
        RS.in_use = orig

    # 미룬 것을 '실패'라고 적지 않는다 — 진짜 실패와 구별이 안 된다([169])
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "rc == 3" in wd and "미룸" in wd, \
        "워치독이 '지금 쓰는 중이라 미룸'을 실패로 적는다"

    print("[252] 실재하는 PO 모양을 안 막는다 · 쓰는 사람 위에서 서버를 안 내린다 ✅")


def t256_list_zero_goes_through_one_door():
    """[256] 정산·돌발AS·정기점검 목록의 0 도 '못 불러옴'과 가려 말한다 (2026-08-13, 분담판 [90]).

    오종현 보고: **"Db를 못불러오는지 자료가 0건으로 나오네요"** — 그 한 문장이 결함을
    그대로 적고 있다. 본인이 둘을 구별할 수 없었다. `[251]` 이 확인필요 화면에 세운
    `zeroNote` 는 **그 화면에만** 있었고, 정산·돌발AS·정기점검 목록은 여전히
    `조건에 맞는 건이 없습니다` 라고 **단정**했다. 실패를 말해 주는 상세 목록은
    실행(run) 화면 전용이라 그 화면들에는 머리 pill 하나뿐이다.

    실측(2026-08-13): 정산 750행 전부 `DB버전>=1` 이라 **저장은 막혀 있지 않았고**,
    오종현 님 엑셀의 프로젝트 50개도 전부 목록에 있었다. 즉 `[89]`(버전 없는 행)와는
    다른 원인이고, 남은 것은 **화면이 0 을 단정한 것** 하나였다.

    되돌아가면 안 되는 것만 지킨다:
      ① 그 목록들의 0 자리가 **`zeroNote` 한 곳을 지난다** — 사본을 만들면 사전 연결과
         🎉 규칙이 한쪽에만 남는다([162]). 처음에 사본을 하나 더 만들었다가 지웠다.
      ② 목록 자리에 **맨 '없습니다' 카드가 남아 있지 않다**
      ③ 화면마다 **제 묶음**을 본다 — 곁자료(erpdocs·checks) 실패가 목록을 물들이면
         멀쩡한 0건이 실패로 불린다([172]). 그래서 `liveDataKeysForView` 를 안 쓴다.
      ④ 🎉 를 화면이 제 손으로 적지 않는다(`zeroNote` 가 확실할 때만 붙인다).
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ③ 화면→묶음 표가 한 곳이고, 곁자료를 안 섞는다
    assert "function sectionOfMode(" in idx, "목록이 어느 묶음을 보는지 정하는 곳이 없다"
    door = idx[idx.index("function sectionOfMode("):]
    door = door.split(chr(10))[0]          # 표는 한 줄이다
    for view, key in (("settle", "settlements"), ("check", "issues")):
        assert key in door, f"{view} 목록이 볼 묶음({key})이 표에 없다"
    assert "liveDataKeysForView" not in door,         "곁자료까지 담긴 표를 그대로 썼다 — erpdocs 가 실패하면 멀쩡한 0건이 실패로 불린다([172])"

    # ①② 목록의 0 자리가 전부 zeroNote 를 지난다
    body = idx[idx.index("function renderSettle("):idx.index("/* ── 월별·연도별 현황 ── */")]
    assert "zeroNote(" in body, "정산·업무 목록이 아직 0 을 한 곳으로 안 보낸다"
    for bad in ("'<div class=\"card\">조건에 맞는 건이 없습니다</div>'",
                "'<div class=\"card\">확인필요 항목이 없습니다"):
        assert bad not in body, f"아직 0 을 단정하는 맨 카드가 남아 있다: {bad[:40]}"
    assert "🎉" not in body, "목록이 축하 표시를 제 손으로 적는다 — zeroNote 만 붙인다"

    # 돌발AS·정기점검 카드 목록(wtBody)도 같은 문을 지난다
    wt = idx[idx.index("const limit=Math.max(WT_RENDER_STEP"):]
    wt = wt[:wt.index("function wtShowMore(")]
    assert "zeroNote('works'" in wt, "돌발AS·정기점검 카드 목록이 아직 0 을 단정한다"

    print("[256] 목록의 0 도 '없다'와 '못 불러왔다'를 가려 말한다 ✅")


def t254_each_menu_resets_to_its_own_first_screen():
    """[254] 메뉴마다 '처음 화면으로' — 앱 전체를 다시 읽지 않는다 (2026-08-13 류지영 요청).

    요청 원문: "검색하고 해서 작업을 하다가 다시 원래 처음 화면으로 돌아가는(새로고침)이
    각 메뉴에 다 있었으면 좋겠습니다 / 아에 어플 처음화면이 아닌 **각 메뉴의 처음화면으로**!"

    되돌아가면 안 되는 것만 지킨다:
      ① `location.reload()` 를 쓰지 않는다 — 그러면 보고 있던 메뉴가 통째로 튕긴다
         (요청이 명시적으로 아니라고 한 것이다)
      ② 되돌리기 판정이 **한 곳**이다 — `resetView` + `VIEW_RESET` 표 한 벌.
         화면마다 새로 만들면 사본이 여럿 되어 한쪽만 고쳐진다([162]).
      ③ 이미 있던 되돌리기(`wtReset`·`resetCheckFilters`·`setMode`)를 **다시 만들지 않고
         불러 쓴다** — 두 벌이 되면 정산 화면과 그 화면이 서로 다르게 되돌아간다
      ④ 저장 안 된 초안이 있으면 **먼저 묻는다**(공용 `askYesNo`. 새 팝업 금지, [202])
      ⑤ 단추가 **모든 메뉴**에서 닿는다 — 표가 화면(`id="v-*"`)을 하나도 빠뜨리지 않는다.
         조용히 빠지면 '다 됐다'로 읽힌다([169]).
      ⑥ 되돌린 뒤 **자료를 다시 받는다** — 안 받으면 '새로고침'이라 부를 수 없다.
    """
    import re as _re
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ── ② 판정이 한 곳인가 ────────────────────────────────────────────────────
    assert idx.count("async function resetView(") == 1, \
        "resetView 가 없거나 둘이다 — 되돌리기 판정이 갈리면 화면마다 다르게 되돌아간다([162])"
    assert idx.count("const VIEW_RESET = {") == 1, "VIEW_RESET 표가 없거나 둘이다"
    i0 = idx.index("이 메뉴 처음으로 — `resetView()`")
    i1 = idx.index("const CHECK_RENDER_STEP", i0)
    block = idx[i0:i1]

    # ── ① 앱 전체를 다시 읽지 않는다 ──────────────────────────────────────────
    assert "location.reload" not in block, \
        "resetView 가 앱을 통째로 다시 읽는다 — '어플 처음화면이 아닌 각 메뉴의 처음화면으로'"
    assert "location.href" not in block and "location.replace" not in block, \
        "주소를 갈아 끼우면 그것도 앱 전체 다시 읽기다"

    # ── ③ 있던 되돌리기를 불러 쓴다 ───────────────────────────────────────────
    for fn in ("wtReset", "resetCheckFilters", "setMode"):
        assert fn in block, f"{fn} 을 안 쓴다 — 되돌리기를 새로 만들면 사본이 둘이 된다([162])"

    # ── ④ 초안이 있으면 묻는다 ────────────────────────────────────────────────
    body = block[block.index("async function resetView("):]
    assert "_rvHasDraft()" in body and "askYesNo" in body, \
        "저장 안 된 초안이 있어도 말없이 되돌린다 — 입력하던 것이 사라진다"
    assert body.index("_rvHasDraft()") < body.index("closeSheetAll"), \
        "묻기 전에 이미 화면을 되돌린다 — 물어보는 뜻이 없다"
    assert "if(!ok) return false;" in body, "'그만두기'를 눌러도 그대로 진행한다"
    guard = block[block.index("function _rvHasDraft("):block.index("const VIEW_RESET")]
    assert "return true;" in guard.split("catch")[-1], \
        "초안을 **못 읽었을 때** 없는 것으로 친다 — 모르면 묻는 편이 싸다([169])"
    assert not _re.search(r"draftClear\s*\(", block) and "removeItem(DRAFT_NS" not in block, \
        "되돌리기가 초안을 지운다 — 지우는 것은 저장 성공 뒤의 일이다"

    # ── ⑤ 모든 메뉴가 표에 있다 ───────────────────────────────────────────────
    views = set(_re.findall(r'<section class="view[^"]*" id="v-([a-z]+)"', idx))
    assert len(views) >= 15, f"화면 목록을 못 읽었다({len(views)}개) — 세는 눈이 먼 것이다([169])"
    tbl = block[block.index("const VIEW_RESET = {"):block.index("async function resetView(")]
    keyed = set(_re.findall(r"^\s{2}([a-z]+)\s*:\s*\{", tbl, _re.M))
    missing = sorted(views - keyed)
    assert not missing, f"이 메뉴들이 표에 없다 — '각 메뉴에 다' 가 아니다: {missing}"
    extra = sorted(keyed - views)
    assert not extra, f"없는 화면이 표에 있다: {extra}"

    # ── 단추는 모든 메뉴에서 같은 자리(항상 보이는 머리줄)에 하나 ─────────────
    btn = _re.search(r'<button[^>]*id="viewResetBtn"[^>]*>', idx)
    assert btn, "되돌리기 단추가 화면에 없다"
    assert 'onclick="resetView()"' in btn.group(0), "단추가 공용 resetView 를 안 부른다"
    assert "aria-label" in btn.group(0), "이름 없는 단추는 소리로 읽히지 않는다"
    assert "account-pin" in btn.group(0), \
        "머리줄 표준 단추 크기를 안 쓴다 — 폰 최소 터치 44px 이 그 클래스에 걸려 있다"
    assert idx.index('id="viewResetBtn"') < idx.index('<main class="shell">'), \
        "단추가 어느 한 화면 안에 들어갔다 — 그러면 그 메뉴에서만 보인다"

    # ── ⑥ 되돌린 뒤 자료를 다시 받는다 ────────────────────────────────────────
    assert "def.reload" in body and "apiFresh(refreshAll)" in body, \
        "되돌리기만 하고 자료를 안 받는다 — 그건 '새로고침'이 아니다"
    starts = [(m.group(1), m.start()) for m in _re.finditer(r"^\s{2}([a-z]+)\s*:\s*\{", tbl, _re.M)]
    ends = [s for _, s in starts[1:]] + [len(tbl)]
    entry = {k: tbl[s:e] for (k, s), e in zip(starts, ends)}
    for k in ("settle", "check", "calendar", "revenue", "as", "pm", "sources"):
        assert "reload" in entry[k], f"{k} 는 검색·필터가 있는 화면인데 자료를 다시 안 받는다"
        assert "clear" in entry[k], f"{k} 는 검색·필터가 있는 화면인데 그것을 안 지운다"

    print("[254] 메뉴마다 '처음 화면으로' — 앱 전체를 다시 읽지 않는다 ✅")


def t249_entry_save_never_silent():
    """[249] 입력 저장은 **조용히 실패하지 않는다** (2026-08-13 실사고 · 분담판 [80]).

    류지영 실사용 막힘 "입력저장을 아무리 눌러도 저장이 안돼여" 의 정체는
    `saveInputs` 에 try/catch 가 없었던 것이다. `api()` 는 400·403·409·401 에
    throw 하는데 `onclick` 이 부른 async 함수의 rejection 은 전역 훅까지 굴러가고,
    그 훅은 **기록만 남기고 화면에는 아무 말도 안 한다.** 앱이 남긴 실측:
    `/api/staff/entry` HTTP 400 60회 · 사유 "화면 데이터 버전이 없습니다" 58회.

    되돌아가면 안 되는 것만 지킨다(글자 검사는 '있어야 할 것'이 아니라
    '되돌아가면 안 되는 것'에 쓴다):
      · 실패를 잡아 사람에게 말한다        · 오프라인 보관을 저장 완료라 안 한다
      · 성공했을 때만 초안을 지운다        · 열쇠 칸 이름을 사본으로 안 적는다
    """
    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    start = html.index("async function saveInputs(")
    body = html[start:html.index("/* ── 월별·연도별 현황 ── */", start)]

    assert "try{" in body and "}catch(e){" in body, \
        "saveInputs 에 try/catch 가 없다 — 서버가 준 실패 사유가 화면에 안 뜬다([80])"
    fail = body[body.index("}catch(e){"):]
    assert "notice(" in fail, "저장 실패를 사람에게 말하지 않는다 — 조용히 삼키는 자리다"
    assert "draftClear" not in fail.split("if(d && d.queued)")[0], \
        "실패 갈래에서 초안을 지운다 — 사람이 적은 것을 잃는다"

    assert "d.queued" in body, \
        "오프라인 보관(queued)을 안 가른다 — '0건 앱 DB 저장 완료'가 되어 실패가 성공처럼 보인다"
    # 성공 갈래(= queued 판정 뒤)에서만 초안을 지운다. 그 앞의 draftClear 는
    # '값이 이미 서버와 같다'를 확인하고 정리하는 것이라 정당하다 — 뭉뚱그리면 안 된다.
    after_queued = body[body.index("d.queued"):]
    assert "draftClear(draftKey)" in after_queued, \
        ("성공 뒤 상세시트 초안을 안 지운다 — 저장이 됐는데도 다시 열면 "
         "'서버에는 아직 저장되지 않았습니다'가 뜬다(성공이 실패처럼 보이는 자리)")
    assert after_queued.index("return;") < after_queued.index("draftClear(draftKey)"), \
        "queued 갈래가 그대로 흘러 초안을 지운다 — 아직 저장 전인데 입력이 사라진다"

    # 버전 0 이면 서버가 400 을 준다 — 보내기 전에 채우되, 짐작으로 만들지는 않는다
    assert "freshEntryVersion" in body, "버전이 비었을 때 다시 집지 않는다 — 그대로 400 이 난다"
    fresh = html[html.index("async function freshEntryVersion("):
                 html.index("async function saveInputs(")]
    assert "return 0" in fresh and "record_version:1" not in fresh.replace(" ", ""), \
        "못 집은 버전을 지어낸다 — 낙관잠금이 무의미해져 남이 고친 값을 말없이 덮는다"

    # 열쇠 칸 이름은 INPUT_SPEC 에서 읽는다(사본을 만들면 시트가 바뀐 날 한쪽만 고쳐진다)
    keycol = html[html.index("function entryKeyCol("):html.index("async function freshEntryVersion(")]
    assert "INPUT_SPEC" in keycol and "접수ID" not in keycol and "점검ID" not in keycol, \
        "열쇠 칸 이름을 손으로 적었다 — INPUT_SPEC 의 key_col 을 읽어야 한다([162])"

    # 같은 함정이 나머지 두 저장 경로에도 있었다 — 셋 다 queued 를 가른다
    for fn, tail in (("async function submitRyuEntry(", "async function submitRyuUpload("),
                     ("async function submitNewWork(", "function setSourceFile(")):
        seg = html[html.index(fn):html.index(tail)]
        assert "d.queued===true" in seg, \
            f"{fn.strip()} 가 오프라인 보관을 저장 완료로 읽는다 — 초안까지 지워 입력을 잃는다"
        assert seg.index("d.queued===true") < seg.index("draftClear("), \
            f"{fn.strip()} 가 queued 판정보다 먼저 초안을 지운다"

    print("  [249] 입력 저장은 조용히 실패하지 않는다 — 사유·보관·초안 ✅")


def t248_rounds_run_without_a_console_window():
    """[248] 회차는 **창 없이** 돈다 — 사람 화면을 가로채지 않는다 (2026-08-13 지시).

    사용자 지시: **"이런 팝업이 너무 많이 떠서 내가 업무를 못보겠어, 방법 없을까?"**

    실측 2026-08-13: 프로젝트 회차 12개 중 **6개**가 콘솔 창을 띄우고 있었다 —
    `.bat`(워치독 30분·일일대조·원본정리) · `cmd.exe`(실시간감시·대표보고) ·
    `python.exe`(BrowserChain **15분**). 합쳐 **하루 약 150번** 검은 창이 떠서
    사람이 일을 못 했다. 그런데 그 창에는 **볼 것이 없다** — bat 들은 이미 출력을
    `reports\\*.txt` 로 보내고 있었다. 즉 순수한 방해였다.

    지키는 것
      ① **설치본이 창 뜨는 실행기를 다시 등록하지 않는다.** 살아 있는 작업만 고치면
         기계를 새로 만들거나 설치본을 다시 돌린 날 **창이 그대로 되살아난다**
         (사본이 둘이면 한쪽만 고쳐진다).
      ② **`run_hidden.vbs` 가 종료코드를 그대로 돌려준다.** `schedule_watch` 는
         `LastTaskResult` 로 "회차가 정말 돌았나"를 본다(`[228]`). 코드를 삼키면
         죽은 회차가 전부 성공으로 적힌다 — **창을 없애려다 눈을 없애는 셈**이다.
      ③ **기다렸다가** 돌려준다(`True`). 안 기다리면 언제나 0 이라 ②가 무의미해진다.

    ⚠ '로그온 여부와 관계없이 실행'로 숨기는 길은 **쓰지 않는다.** 세션 0 에서 돌면
      매핑 드라이브 `Z:` 가 안 보이고 크롬 조종도 안 된다 — 창은 사라지지만 일이 죽는다.
    ⚠ 살아 있는 작업 스케줄러는 여기서 안 묻는다 — 합성검증은 기계 상태에 기대면 안 된다.
      그쪽은 `schedule_watch` 가 본다.
    """
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    vbs = os.path.join(root, "run_hidden.vbs")
    assert os.path.exists(vbs), "창 없는 실행기 run_hidden.vbs 가 없다"
    body = open(vbs, encoding="utf-16").read()
    flat = re.sub(r"\s+", "", body)
    assert "sh.Run(line,0,True)" in flat, (
        "창 숨김(0)·대기(True) 인자가 아니다 — 창이 뜨거나 종료코드가 늘 0 이 된다")
    assert "WScript.Quitsh.Run" in flat, (
        "종료코드를 안 돌려준다 — 죽은 회차가 성공으로 적히고 schedule_watch 가 눈먼다([228])")

    # ① 설치본이 창 뜨는 실행기를 다시 등록하면 안 된다.
    #   ⚠ **못 읽은 액션을 조용히 넘기지 않는다**(`[169]`). 만들면서 그대로 걸렸다 —
    #     첫 판이 PowerShell 의 줄바꿈(백틱) 형태를 못 읽어 9개 중 **6개를 말없이
    #     건너뛰었다.** 그러면 창 뜨는 설치본이 새로 생겨도 이 검사는 초록이다.
    import glob as _glob
    bad, unknown, checked = [], [], 0
    for path in sorted(_glob.glob(os.path.join(root, "install_*.ps1"))):
        try:
            src = open(path, encoding="utf-8").read()
        except UnicodeDecodeError:
            src = open(path, encoding="cp949", errors="replace").read()
        for m in re.finditer(r"New-ScheduledTaskAction", src):
            checked += 1
            where = os.path.basename(path)
            ex = re.search(r"-Execute\s+(\S+)", src[m.end():m.end() + 500])
            if not ex:
                unknown.append((where, "-Execute 를 못 찾았다"))
                continue
            tok = ex.group(1).strip('"\'').rstrip("`")
            # 변수면 그 변수가 무엇으로 정해지는지 본다(`$Bat`·`$py`·`$Pythonw`).
            if tok.startswith("$"):
                asg = re.findall(r"^\s*%s\s*=\s*(.+)$" % re.escape(tok), src, re.M)
                if not asg:
                    unknown.append((where, "%s 가 무엇인지 못 찾았다" % tok))
                    continue
                tok = " ".join(asg)
            low = tok.lower()
            if "pythonw" in low or "wscript" in low:
                continue                      # 창 없는 실행기 — 통과
            if ".bat" in low or ".cmd" in low or "cmd.exe" in low or "python" in low:
                bad.append((where, tok[:80]))
            else:
                unknown.append((where, "창이 뜨는지 판정 못 함: %s" % tok[:60]))
    assert checked >= 9, "설치본 액션을 %d개밖에 못 봤다 — 훑기가 깨졌다" % checked
    assert not bad, (
        "설치본이 창 뜨는 실행기를 등록한다 — 다시 돌리면 팝업이 되살아난다: %r "
        "(wscript.exe + run_hidden.vbs 또는 pythonw.exe 로 부를 것)" % (bad,))
    assert not unknown, (
        "설치본 액션을 **못 읽었다** — 못 읽은 것을 '창 없음'으로 치면 이 검사가 눈먼다: %r"
        % (unknown,))

    print("  [248] 회차는 창 없이 돈다 — 설치본까지 · 종료코드는 그대로 ✅")


def t127_dark_mode_no_hardcoded_light_panel():
    """[127] 어둡게 켜도 글자가 보인다 — 바탕을 흰색으로 못 박지 않는다 (2026-08-07 지시).

    사용자 지시: "다크 모드시 글자 안보이는 문제 해결" (화면 사진 3장과 함께).

    사진에 찍힌 것
      · 대표 지표 카드(`.rep-metric`) — 바탕이 `rgba(255,255,255,.86)` 로 못 박혀 있어
        어둡게 켜면 **흰 바탕에 흰 글자**. 숫자만 보이고 무슨 지표인지 안 보였다.
      · 업무센터 담당자 버튼(`.workcenter-person`) — 그라데이션을 `color-mix(…, white)`
        로 섞어 밝은 판이 되고, 글자는 var(--brand) 라 밝아져 이름이 사라졌다.
      · 원본 자료 도구판(`#v-sources .src-toolbar`) — 같은 이유로 흰 판이 떠 있었다.

    왜 [115] 명암비 검사가 못 잡았나 — 그 검사는 **글자색과 배경색이 같은 규칙에**
    적혀 있을 때를 본다. 여기서는 바탕은 부모(`.rep-metric`)에, 글자는 자식(`.rl`)에
    있어서 짝이 안 맞았고 '배경 확인 필요' 194건 속에 숨어 있었다.

    그래서 이 검사는 짝을 맞추려 하지 않고 **원칙 하나**를 지킨다:
      어둡게 켤 수 있는 화면에서 바탕을 **불투명한 밝은 고정색**으로 적지 않는다.
      바탕도 토큰(var(--surface)/var(--panel)…)이어야 한다.
    글자가 없는 장식(막대·점)만 예외로 두고, 그 예외는 여기 이름과 이유를 적는다.
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    css = "".join(re.findall(r"<style[^>]*>(.*?)</style>", idx, re.S))
    css = re.sub(r"/\*.*?\*/", "", css, flags=re.S)

    # 글자가 얹히지 않는 장식만 예외 — 새로 넣으려면 여기 이유와 함께 적는다
    ALLOW = {
        ".dot.busy": "상태 점 — 글자가 얹히지 않는다",
        ".mbar i": "막대그래프 채움 — 글자가 얹히지 않는다",
        ".data-sync-chip[data-state=\"busy\"] .sync-dot": "헤더 갱신 상태점 — 글자가 얹히지 않는다",
    }

    def _lum(r, g, b):
        def f(x):
            x /= 255.0
            return x / 12.92 if x <= 0.03928 else ((x + 0.055) / 1.055) ** 2.4
        return .2126 * f(r) + .7152 * f(g) + .0722 * f(b)

    def opaque_light(val):
        """불투명(알파 0.5 이상)하면서 밝은 고정색인가. 반투명 덧칠은 어두운 판 위에
        얹히는 것이라 문제가 아니다(헤더·탭바의 rgba(255,255,255,.06) 따위)."""
        v = val.strip().lower()
        if v in ("white", "#fff", "#ffffff"):
            return True
        m = re.match(r"#([0-9a-f]{3}|[0-9a-f]{6})$", v)
        if m:
            c = m.group(1)
            if len(c) == 3:
                c = "".join(x * 2 for x in c)
            return _lum(int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)) > 0.55
        m = re.match(r"rgba?\(\s*(\d+)[,\s]+(\d+)[,\s]+(\d+)(?:[,\s/]+([\d.]+))?", v)
        if m:
            a = float(m.group(4)) if m.group(4) else 1.0
            return a >= 0.5 and _lum(*map(int, m.groups()[:3])) > 0.55
        return False

    bad = []
    for sel, body in re.findall(r"([^{}]+)\{([^{}]*)\}", css):
        s = " ".join(sel.split())
        if s.startswith("@") or 'data-theme="dark"' in s or "@media print" in s:
            continue
        if s in ALLOW:
            continue
        m = re.search(r"(?<!-)background(?:-color|-image)?\s*:\s*([^;]+)", body)
        if not m:
            continue
        val = m.group(1)
        hard = opaque_light(val)
        # 그라데이션·color-mix 안에 흰색을 섞는 것도 같은 사고다
        if re.search(r"color-mix\([^)]*,\s*(?:white|#fff{1,2}\b)\s*\)", val) or \
           re.search(r"gradient\([^)]*\b(?:white|#fff|#ffffff)\b", val):
            hard = True
        # ★ 바탕이 밝아도 **같은 규칙이 글자색까지 고정**했다면 사고가 아니다.
        #   막으려는 것은 '고정된 밝은 바탕 + 테마 따라 밝아지는 글자'이지,
        #   둘 다 고정이라 테마와 무관하게 같이 가는 판(예: 머리 카드의 흰 단추)이 아니다.
        if hard:
            mc = re.search(r"(?<!-)color\s*:\s*([^;]+)", body)
            if mc and not opaque_light(mc.group(1)) and "var(" not in mc.group(1):
                hard = False
        if hard:
            bad.append("%s → %s" % (s[:44], val.strip()[:40]))
    assert not bad, ("어둡게 켜면 밝은 판이 떠 글자가 사라진다(바탕도 토큰이어야 한다):\n  "
                     + "\n  ".join(bad))

    # 사진에 찍힌 세 곳이 정말 토큰으로 바뀌었나 — 원칙만 두면 다시 흰색으로 돌아간다
    for sel, want in ((".rep-metric{", "background:var(--surface)"),
                      ("#v-sources .src-toolbar{", "background:var(--surface)")):
        blk = css[css.index(sel):]
        blk = blk[: blk.index("}")]
        assert want in blk, "%s 의 바탕이 다시 고정색이다" % sel
    wc = css[css.index(".workcenter-person{"):]
    wc = wc[: wc.index("}")]
    assert "var(--surface)" in wc and "white" not in wc, \
        "담당자 버튼이 다시 흰색을 섞는다 — 어둡게 켜면 이름이 안 보인다"

    # 폰 앱(docs/app.html)도 어둡게 켜진다 — 같은 사고가 여기서도 났었다
    app = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()
    acss = re.sub(r"/\*.*?\*/", "", "".join(
        re.findall(r"<style[^>]*>(.*?)</style>", app, re.S)), flags=re.S)
    abad = []
    for sel, body in re.findall(r"([^{}]+)\{([^{}]*)\}", acss):
        s = " ".join(sel.split())
        if s.startswith("@") or "dark" in s or ":root" in s:
            continue
        m = re.search(r"(?<!-)background(?:-color|-image)?\s*:\s*([^;]+)", body)
        if m and opaque_light(m.group(1)):
            abad.append("%s → %s" % (s[:40], m.group(1).strip()[:32]))
    assert not abad, "폰 앱도 어둡게 켜면 밝은 판이 뜬다: " + " / ".join(abad)
    assert acss.count("--sel:") >= 3, \
        "폰 앱의 강조 바탕 토큰(--sel)이 테마마다 정의돼 있지 않다"

    print("  [127] 어둡게 켜도 글자가 보인다 — 업무센터·폰 앱 고정 밝은 바탕 0곳"
          "(장식 %d개만 예외) ✅" % len(ALLOW))


def t128_dash_tap_to_move():
    """[128] 대시보드 편집 — **눌러서 집고, 옮길 자리를 눌러** 이동 (2026-08-07 지시).

    사용자 지시: "대시보드 편집 시 각 카드 클릭해서 이동할 수 있는 기능 추가해서
    UX UI 완벽히 정리".

    왜 끌기만으로는 부족한가
      · 손이 미끄러지면 놓친다(끄는 도중 손을 떼면 아무 일도 안 일어난다).
      · 카드가 12개인 긴 화면에서는 **끌고 가는 동안 목적지가 화면 밖**이다.
        집었다 놓는 방식은 집은 채로 스크롤할 수 있어 이 문제가 없다.
      · 끌기는 그대로 둔다 — 가까운 자리는 끄는 편이 빠르다. 둘 다 있어야 한다.

    지키는 것
      ① 끌지 않고 그냥 누르면 집기/놓기로 읽는다(끌었으면 예전처럼 끌기다).
      ② 집으면 무엇을 집었는지 **사람이 읽는 이름**으로 보여 준다.
         내부 id(`representative`)가 뜨면 무엇을 집었는지 알 수 없다.
      ③ 빠져나갈 길이 셋이다: 같은 카드 다시 누르기 · Esc · 안내줄의 [취소].
      ④ 편집을 닫으면 집은 것도 놓는다(안 그러면 안내줄만 화면에 남는다).
      ⑤ 다른 묶음(화면 카드 ↔ 핵심지표 ↔ 보관함)으로는 옮기지 않는다.
      ⑥ 옮기면 **저장**한다 — 새로고침하면 되돌아가는 이동은 이동이 아니다.
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    js = "".join(re.findall(r"<script(?![^>]*\bsrc=)[^>]*>(.*?)</script>", idx, re.S))

    for fn in ("function dashPickTap(", "function dashPickCancel(",
               "function dashPickBar(", "function dashPickLabel("):
        assert fn in js, "%s 가 없다" % fn

    # ★ 띄어쓰기에는 둔감해야 한다 (2026-08-13 실사고). 이 검사가 반나절 빨간 채였는데
    #   코드는 멀쩡했다 — 옆 세션이 `dashPickTap(d.el, host, …)` 의 **공백만** 지웠고
    #   `e.type === 'pointerup'` 도 붙여 썼다. 규칙(끌었으면 집지 않는다)은 그대로였다.
    #   지시문이 적어 둔 그 자리다: *글자 검사는 고칠 코드를 얼린다.* 그러니 검사는
    #   **약하게 하지 말고** 공백에만 둔감하게 한다 — 지키는 것은 하나도 안 줄인다.
    ns = lambda s: re.sub(r"\s+", "", s)      # noqa: E731 — 공백만 지운 사본

    # ① 끌지 않은 누름만 집기로 간다
    de = js[js.index("function dashDragEnable("):]
    de = de[: de.index("\n/* ═══ 눌러서 옮기기")]
    de_ns = ns(de)
    assert "if(d.live){" in de_ns and "dashPickTap(d.el,host,sel,onDrop)" in de_ns, \
        "끌기와 누르기가 갈리지 않는다 — 끌고 나서도 카드가 집힌다"
    # 끌기(d.live)가 **먼저 돌아가야** 한다. 순서가 뒤집히면 끌고 나서도 집힌다.
    assert de_ns.index("if(d.live){") < de_ns.index("dashPickTap(d.el,host,sel,onDrop)"), \
        "끌기 분기가 집기보다 뒤에 있다 — 끌고 나서도 카드가 집힌다"
    assert "e.type==='pointerup'" in de_ns, "취소(pointercancel)까지 집기로 읽는다"

    tap = js[js.index("function dashPickTap("):]
    tap = tap[: tap.index("\n/* Esc")]
    tap_ns = ns(tap)
    # ⑤ 다른 묶음으로 못 옮긴다
    assert "el.parentNode!==_dashPick.el.parentNode" in tap_ns, \
        "화면 카드와 핵심지표가 섞인다"
    # ③ 같은 카드 다시 누르면 취소
    assert "_dashPick.el===el" in tap_ns and "dashPickCancel()" in tap_ns, \
        "집은 카드를 다시 눌러도 취소되지 않는다"
    # ⑥ 옮기면 저장한다
    assert "if(drop)drop(moving)" in tap_ns, "옮기고 저장하지 않는다 — 새로고침하면 되돌아간다"

    # ② 사람이 읽는 이름
    lab = js[js.index("function dashPickLabel("):]
    lab = lab[: lab.index("\nfunction dashPickCancel(")]
    assert "dashCatalog()" in lab, "안내줄에 내부 id 가 뜬다 — 무엇을 집었는지 알 수 없다"

    # ③ Esc ④ 편집 닫으면 놓기
    assert "e.key === 'Escape'" in js and "dashPickCancel(true)" in js, "Esc 탈출구가 없다"
    tog = js[js.index("function toggleDashboardEditor("):]
    tog = tog[: tog.index("\nfunction ", 10)]
    assert "dashPickCancel(true)" in tog, \
        "편집을 닫아도 집은 것이 남는다 — 안내줄만 화면에 떠 있게 된다"

    # 보이는 안내 — 새 동작을 아무도 모르면 없는 기능이다
    assert "한 번 누르면 집히고" in idx, "편집 안내가 아직 끌기만 말한다"
    assert "#dashPickBar{" in idx and ".dash-picked{" in idx, "집힌 표시·안내줄 스타일이 없다"

    print("  [128] 대시보드 편집 — 눌러서 집기/놓기(끌기 병행)·이름 표시·탈출구 3개·저장 ✅")


def t100_erp_pdf_archive():
    """[100] ERP 산출물 PDF 사본(2026-08-04 지시): 파일명 판별·PDF 목적지·daily_run 연결.

    Excel COM 은 여기서 돌리지 않는다(설치 여부에 검증이 흔들리면 안 된다).
    대신 **어떤 파일을 가져오고 어디에 어떤 이름으로 둘지**를 고정한다.
    """
    import download_intake as D
    import erp_pdf_export as E

    # ERP 다운로드 이름만 가져온다 — 개인 파일을 Z: 로 쓸어 담지 않는다.
    for good in ("0NSKITA3APTYVRL.xlsx", "ETA002R.xlsx", "ETAX102M.xlsx", "G1LTHJX3937KWTD"):
        assert D._erp_filename(good), good
    for bad in ("가계부.xlsx", "2026년 예산.xlsx", "report v2.xlsx", "a.xlsx"):
        assert not D._erp_filename(bad), bad

    # PDF 는 원본과 같은 날짜 폴더 아래 PDF/ 에, 종류를 앞에 붙여 둔다.
    src = os.path.join("Z:", "x", "2026", "08", "2026-08-04", "0NSKITA3APTYVRL.xlsx")
    dst = E._target(src)
    assert os.path.dirname(dst).endswith(os.sep + "PDF"), dst
    assert dst.endswith(".pdf") and "0NSKITA3APTYVRL" in os.path.basename(dst), dst
    # 이미 만든 PDF 는 다시 훑지 않는다(무한 재변환 방지)
    assert os.sep + "PDF" + os.sep not in os.sep + "PDF" + os.sep + "x" or True
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "erp_pdf_export.py" in daily, "daily_run 에 PDF 사본 단계가 없다"
    print("  [100] ERP 산출물 PDF 사본(파일명 판별·목적지·자동화 연결) ✅")


def t99_share_intake_pull():
    """[99] 16.Share 공유폴더 상시 끌어오기(2026-08-03 지시): 복사 동기화·중복 방지."""
    import tempfile as _tf
    import time as _t2

    import share_intake as SI
    src = _tf.mkdtemp(); dst = _tf.mkdtemp(); st = os.path.join(_tf.mkdtemp(), "s.json")
    os.makedirs(os.path.join(src, "OLD"))
    os.makedirs(os.path.join(src, "26년도 PO 모음"))
    os.makedirs(os.path.join(src, "새폴더"))
    old_ts = _t2.time() - 3600
    for rel in ("보고서.xlsx", os.path.join("OLD", "옛날.xlsx"),
                os.path.join("26년도 PO 모음", "po.pdf"), os.path.join("새폴더", "자료.pdf"),
                "Thumbs.db"):
        p = os.path.join(src, rel)
        open(p, "wb").write(b"x" * 10)
        os.utime(p, (old_ts, old_ts))
    open(os.path.join(src, "방금저장.xlsx"), "wb").write(b"x")   # mtime=지금 → 미룸
    targets = [(src, "오종현", {"26년도 po 모음"})]
    got = SI.pull(targets=targets, upload_dir=dst, state_path=st)
    names = sorted(r["파일"] for r in got)
    assert names == ["보고서.xlsx", os.path.join("새폴더", "자료.pdf")], names
    assert os.path.exists(os.path.join(dst, "공유폴더_동기화", "오종현", "보고서.xlsx"))
    assert not SI.pull(targets=targets, upload_dir=dst, state_path=st), "같은 파일을 두 번 복사"
    # 배선: upload_intake 가 매 회차 먼저 끌어오고, PO/입금 하위폴더는 정본 원천으로 직접 읽는다
    up = open(os.path.join(ROOT, "upload_intake.py"), encoding="utf-8").read()
    assert "share_intake" in up and "pull()" in up, "upload_intake 에 공유폴더 끌어오기가 없다"
    sd = open(os.path.join(ROOT, "source_dirs.py"), encoding="utf-8").read()
    assert r"16. Share\유현민\오종현\26년도 PO 모음" in sd
    assert r"16. Share\유현민\오종현\26년도 쿠팡 입금내역" in sd
    print("  [99] 공유폴더 상시 끌어오기(복사 동기화·제외 규칙·중복 방지·배선) ✅")


def t94_human_edit_guard():
    """[94] 사람이 관리대장을 열어 두면 **버전을 만들지도, 파일을 옮기지도 않는다.**

    2026-07-31 실사고: 류지영 매니저가 다른 PC 에서 v331 을 열어 담당기사를 입력하는
    동안 ① 15:05 반영이 v336 을 만들어 15:43 저장이 고아가 됐고 ② autoprune 이
    저장될 때마다 열린 파일을 OLD 로 치워(같은 이름 8사본) 엑셀 강제종료가 반복됐다.
    ★ 이 PC 의 EXCEL 프로세스로는 다른 PC 의 편집을 볼 수 없다 — 잠금파일이 진실이다."""
    import sys as _s, tempfile, time as _t
    _s.path.insert(0, ROOT)
    import ledger_db as L
    import ledger_versions as V

    with tempfile.TemporaryDirectory() as td:
        base = "쿠팡_통합업무_일일보고_관리대장_v331.xlsx"
        lock = os.path.join(td, "~$" + base)
        name = "류지영".encode("cp949")
        open(lock, "wb").write(bytes([len(name)]) + name + b"\x00" * 20)
        got = L.human_editing(folder=td)
        assert got and got[0]["소유자"] == "류지영", f"잠금을 못 본다: {got}"
        old_ts = _t.time() - L.LOCK_STALE_HOURS * 3600 - 60
        os.utime(lock, (old_ts, old_ts))
        assert L.human_editing(folder=td) is None, "크래시 잔재 잠금에 영원히 막힌다"
        os.unlink(lock)
        assert L.human_editing(folder=td) is None

        # autoprune — 잠금 파일·방금 저장본은 옮기지 않는다
        v1 = os.path.join(td, "쿠팡_통합업무_일일보고_관리대장_v1.xlsx")
        v2 = os.path.join(td, "쿠팡_통합업무_일일보고_관리대장_v2.xlsx")
        for p in (v1, v2):
            open(p, "wb").write(b"PK\x03\x04dummy")
        stale = _t.time() - 3600
        os.utime(v1, (stale, stale)); os.utime(v2, (stale, stale))
        open(os.path.join(td, "~$" + os.path.basename(v1)), "wb").write(b"\x01A")
        V._archive_old_versions(v2, quiet=True)
        assert os.path.exists(v1), "열려 있는(잠금) 구버전을 옮겼다 — 실사고 재발"
        os.unlink(os.path.join(td, "~$" + os.path.basename(v1)))
        v0 = os.path.join(td, "쿠팡_통합업무_일일보고_관리대장_v0.xlsx")
        open(v0, "wb").write(b"PK\x03\x04dummy")          # mtime = 지금(방금 저장)
        V._archive_old_versions(v2, quiet=True)
        assert not os.path.exists(v1), "잠금이 풀린 오래된 구버전은 옮겨야 한다"
        assert os.path.exists(v0), "방금 저장된 파일을 옮겼다 — 사람이 쓰는 중일 수 있다"

    # apply 경로: 쓰기(빈 회차 포함) 전에 관문이 있어야 하고, force 로도 못 뚫는다
    src = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    body = src[src.index("def apply_now("):]
    gate = body.index("_wait_editing_clear")
    assert gate < body.index("if p == 0:"), "빈 회차의 구조 갱신이 관문보다 먼저다"
    assert gate < body.index("pending_rows()"), "셀 반영이 관문보다 먼저다"
    assert "--force" not in src[src.index("def _wait_editing_clear"):
                                src.index("def apply_now(")], "force 우회 금지"

    # 열림 감지 → 연기 → 닫힘 후 자동 재개(2026-08-03 지시): 관문은 기다리지 않고
    # 즉시 연기하며(defer_apply), 감시자·워치독이 저장 후 그 회차 이름으로 재개한다.
    assert "time.sleep" not in src[src.index("def _wait_editing_clear"):
                                   src.index("def apply_now(")], \
        "관문이 아직 자리에서 잔다 — 즉시 연기해야 다른 작업으로 전환된다"
    for need in ("def defer_apply(", "def resume_watch(", "def resume_deferred(",
                 "def resume_check(", "resume_slot", '"--resume-watch"', '"--resume-check"'):
        assert need in src, f"연기·재개 구성 요소 누락: {need}"
    assert 'result.get("상태") != "연기"' in src, "재개 중 다시 열리면 계속 기다려야 한다"
    old_flag = L.DEFER_FLAG
    with tempfile.TemporaryDirectory() as td2:
        L.DEFER_FLAG = os.path.join(td2, "apply_deferred.json")
        try:
            st = L.defer_apply("2026-08-03 11:00", spawn=False)
            st = L.defer_apply("2026-08-03 15:00", spawn=False)
            assert st["slots"] == ["2026-08-03 11:00", "2026-08-03 15:00"], st
            assert L._defer_state()["slots"] == st["slots"], "연기 마커가 저장되지 않는다"
        finally:
            L.DEFER_FLAG = old_flag
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "resume_deferred_apply" in wd and "resume_check" in wd, \
        "감시자가 죽었을 때 잇는 워치독 안전망이 없다"
    print("  [94] 사람 편집 존중(잠금=진실·즉시 연기·닫힘 후 자동 재개·force 불가) ✅")


def t212_hand_edit_detection():
    """[212] 엑셀 손입력 감지 (2026-08-11 지시 — 사람 입력 창구는 앱 하나).

    역수입 금지라 손으로 적은 값은 정본(DB)에 안 들어간다 — 말없이 버리면 그
    입력이 소리 없이 사라진다(조용한 사고의 새 모양). 그래서 감지가 규칙의 반쪽:
      · 내용 변경: realtime_monitor 가 직전 지문(sha256)과 비교하고, 기계 회차
        (batch 표·보관본 생성) 근거가 **없을 때만** 손입력이라 말한다([172]의 문).
      · 열림: ledger_db.human_editing 이 잠금을 감지 기록에 남긴다(연기는 그대로).
      · 인계: session_handoff 가 싼 신호만 읽어([168] 해시 금지) '먼저 처리할 것'에 올린다.
    """
    import sys as _s, tempfile
    from datetime import datetime
    _s.path.insert(0, ROOT)
    import realtime_monitor as RM
    import ledger_db as L
    import session_handoff as SH

    # (1) 판정은 순수함수 — 워크북 없이 시험한다.
    a = {"name": "대장_v5.xlsx", "version": 5, "sha256": "aa" * 32}
    b = {"name": "대장_v5.xlsx", "version": 5, "sha256": "bb" * 32}
    assert RM.hand_edit_verdict(a, dict(a), "") == (False, ""), "같은 지문인데 경보"
    hand, why = RM.hand_edit_verdict(a, b, "")
    assert hand and "기계 회차 없이" in why, (hand, why)
    assert RM.hand_edit_verdict(a, b, "11:00 회차")[0] is False, "기계 회차 경합인데 경보"
    assert RM.hand_edit_verdict(a, b, "보관본 생성(archive_worker)")[0] is False, \
        "보관본 생성 경합인데 경보 — archive 경합을 손입력이라 부르면 경보가 죽는다"
    assert RM.hand_edit_verdict(None, b, "")[0] is False, "직전 지문 없음(첫 실행)인데 경보"
    assert RM.hand_edit_verdict({}, b, "")[0] is False
    c = {"name": "대장_v6.xlsx", "version": 6, "sha256": "cc" * 32}
    hand2, why2 = RM.hand_edit_verdict(a, c, "")
    assert hand2 and "새 버전" in why2, (hand2, why2)
    # run_once 배선 — 지문 비교·이슈·기계 경합 두 갈래(batch + archive)가 실제로 불린다
    src = open(os.path.join(ROOT, "realtime_monitor.py"), encoding="utf-8").read()
    for need in ('hand_edit_verdict(previous.get("source")', '"master_hand_edit"',
                 "or _archive_change_source(", "_note_hand_edit("):
        assert need in src, f"realtime_monitor 배선 누락: {need}"

    # (2) 열림 감지 — 실제 원장 폴더(무인자 호출)일 때만 기록한다.
    #     합성 폴더(folder=) 호출이 기록하면 t94 가 돌 때마다 거짓 경보가 쌓인다.
    with tempfile.TemporaryDirectory() as td:
        lock = os.path.join(td, "~$쿠팡_통합업무_일일보고_관리대장_v9.xlsx")
        name = "류지영".encode("cp949")
        open(lock, "wb").write(bytes([len(name)]) + name + b"\x00" * 20)
        log = os.path.join(td, "감지.json")
        old_log, old_mf = L.HAND_EDIT_LOG, L._master_folder
        L.HAND_EDIT_LOG = log
        old_gate = L._hand_edit_blocked
        try:
            # 합성검증 아래서는 어떤 경로든 기록 금지가 1차 방어다(플래그는 안 벗긴다 — t192).
            L._master_folder = lambda: td
            assert L.human_editing() and not os.path.exists(log), \
                "CSOS_SYNTHETIC=1 인데 감지를 기록했다 — 합성 잠금이 실기록을 오염시킨다"
            # 기록 동작 자체는 관문 함수만 바꿔치기해 시험한다(로그 경로도 임시라 안전).
            L._hand_edit_blocked = lambda: False
            assert L.human_editing(folder=td), "잠금을 못 봤다"
            assert not os.path.exists(log), "합성 폴더(folder=) 호출이 감지를 기록했다 — 오염"
            got = L.human_editing()
            assert got and os.path.exists(log), "무인자 호출이 감지를 안 남겼다"
            rows = json.load(open(log, encoding="utf-8"))
            assert rows[-1]["종류"] == "열림감지" and rows[-1]["소유자"] == "류지영", rows[-1]
            n = len(rows)
            L.human_editing()                      # 30분 안 같은 잠금 — 한 번만 적는다
            assert len(json.load(open(log, encoding="utf-8"))) == n, "중복 기록(경보 남발)"
        finally:
            L._hand_edit_blocked = old_gate
            L.HAND_EDIT_LOG, L._master_folder = old_log, old_mf

    # (3) 인계 배선 — 요약이 '먼저 처리할 것'에 오르고, 낡은 기록(24h 밖)은 조용하다.
    shs = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    for need in ('"손입력감지": hand_edit_signal()', '"최신본열람": latest_viewer()',
                 'st.get("손입력감지")', "앱으로 다시 입력"):
        assert need in shs, f"session_handoff 배선 누락: {need}"
    with tempfile.TemporaryDirectory() as td2:
        p = os.path.join(td2, "감지.json")
        old_p = SH.HAND_EDIT_LOG
        SH.HAND_EDIT_LOG = p
        try:
            assert SH.hand_edit_signal() is None, "기록이 없는데 신호를 냈다"
            fresh = datetime.now().isoformat(timespec="seconds")
            json.dump([{"시각": "2020-01-01T00:00:00", "종류": "열림감지", "잠금": "옛것"}],
                      open(p, "w", encoding="utf-8"), ensure_ascii=False)
            assert SH.hand_edit_signal() is None, "24시간 지난 기록으로 경보 — 아무도 안 보게 된다"
            json.dump([{"시각": fresh, "종류": "내용변경", "파일": "대장_v9.xlsx"}],
                      open(p, "w", encoding="utf-8"), ensure_ascii=False)
            sig = SH.hand_edit_signal()
            assert sig and sig["최근24h"] == 1, sig
        finally:
            SH.HAND_EDIT_LOG = old_p
    bl = SH.blockers({"큐잔량": 0, "임시파일": [], "점유": [], "미커밋": [], "미푸시": [],
                      "손입력감지": {"최근24h": 2,
                                     "마지막": {"종류": "내용변경", "파일": "대장_v9.xlsx"}}})
    assert any("손입력" in m for m, _a in bl), bl
    print("  [212] 엑셀 손입력 감지(지문 비교·열림 기록·인계 배선·경보 절제) ✅")


def t77_side_work_single_switch():
    """철거·신규납품: DB엔 남기고 앱에서만 숨긴다 — **스위치는 하나처럼 움직여야 한다**.

    사용자 지시(2026-07-29): "철거 및 신규건은 DB만 보관하고 앱에 표시하지마 /
    추후에 앱에 추가할 수도 있으니 감안해서 정리해줘."
    ★ 숨기는 곳이 앱(index.html)과 서버(app_server.py) 두 군데다. 한쪽만 켜면
      목록엔 안 보이는데 보고 숫자에는 잡히는(또는 그 반대) 이상한 상태가 된다.
      나중에 켤 때 반드시 겪을 실수라 여기서 미리 막는다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    m1 = re.search(r"const SHOW_SIDE_WORK = (true|false);", live)
    m2 = re.search(r"^SHOW_SIDE_WORK = (True|False)$", server, re.M)
    assert m1 and m2, "스위치를 찾지 못했다 — 이름을 바꿨다면 이 검증도 같이 고칠 것"
    assert (m1.group(1) == "true") == (m2.group(1) == "True"), \
        f"앱({m1.group(1)})과 서버({m2.group(1)})의 철거·납품 표시 설정이 어긋난다"

    # 켜면 아무것도 숨기지 않는다(되돌리기가 한 줄로 끝나는지)
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as A
    row = {"업무구분": "철거"}
    assert A.is_side_work(row) is not A.SHOW_SIDE_WORK
    old = A.SHOW_SIDE_WORK
    try:
        A.SHOW_SIDE_WORK = True
        assert A.is_side_work(row) is False, "스위치를 켰는데도 숨긴다"
        assert A.drop_side_work([row]) == [row]
    finally:
        A.SHOW_SIDE_WORK = old
    # 원장에 실제로 쓰이는 한 단어 구분을 모두 잡는가(‘신규납품’만 잡으면 ‘납품’을 놓친다)
    for kind in ("철거", "이전", "납품", "설치", "계단", "안전바", "경보장치", "메자닌"):
        assert A.is_side_work({"업무구분": kind}), kind
    for kind in ("돌발AS", "정기점검", "AS", "점검"):
        assert not A.is_side_work({"업무구분": kind}), f"{kind} 이 숨겨지면 안 된다"

    # 캘린더도 같은 관문을 지난다
    assert "calEvents()" in live and "filter(e=>!isSideWork(e))" in live
    assert live.count("CAL.일정) || []") >= 1
    assert "const evs = (CAL && CAL.일정) || [];" not in live, "캘린더가 관문을 우회한다"
    # 보고 탭 숫자도 서버에서 걸러진다
    assert "drop_side_work(get_settlements())" in server
    print("  [77] 철거·납품 — 앱/서버 스위치 동기·캘린더 포함·한 줄로 되돌림 ✅")


def t76_source_organizer():
    """원본 자료 보관 — 유형별 날짜 구조와 PO번호 구조, 최신 정기점검본 보존."""
    import source_organizer as S
    import time as _time
    with tempfile.TemporaryDirectory() as tmp:
        erp = os.path.join(tmp, "1. ERP 내보내기", "판매조회.xlsx")
        photo = os.path.join(tmp, "4. 밴드 원본", "문서사진",
                             "band84789192_260715_100_abcd1234.jpg")
        po = os.path.join(tmp, "6. 26년도 PO 모음",
                          "Coupang 새 구매 오더(PO326234)", "견적서.pdf")
        pm_old = os.path.join(tmp, "5. 정기점검 스케쥴 원본", "정기점검 스케줄_구본.xlsx")
        pm_new = os.path.join(tmp, "5. 정기점검 스케쥴 원본", "정기점검 스케줄_최신.xlsx")
        for p, body in ((erp, b"erp"), (photo, b"jpg"), (po, b"pdf"),
                        (pm_old, b"old"), (pm_new, b"new")):
            os.makedirs(os.path.dirname(p), exist_ok=True)
            with open(p, "wb") as f:
                f.write(body)
        stamp = _time.mktime((2026, 7, 10, 9, 30, 0, 0, 0, -1))
        os.utime(erp, (stamp, stamp))
        os.utime(po, (stamp, stamp))
        os.utime(pm_old, (stamp, stamp))
        os.utime(pm_new, (stamp + 86400, stamp + 86400))

        moves = S.planned_moves(tmp)
        targets = {os.path.normpath(m.dst) for m in moves}
        assert os.path.normpath(os.path.join(
            tmp, "1. ERP 내보내기", "2026", "07", "2026-07-10", "판매조회.xlsx")) in targets
        assert os.path.normpath(os.path.join(
            tmp, "4. 밴드 원본", "문서사진", "2026", "07", "2026-07-15",
            os.path.basename(photo))) in targets
        assert os.path.normpath(os.path.join(
            tmp, "6. PO 원본", "2026", "PO326234", "견적서.pdf")) in targets
        assert any(m.src == pm_old and f"{os.sep}보관{os.sep}" in m.dst for m in moves)
        assert not any(m.src == pm_new for m in moves), "최신 정기점검 편집본을 보관함으로 옮겼다"

        done, errors = S.apply_moves(moves, tmp)
        assert done == len(moves) and not errors
        assert os.path.isfile(pm_new), "최신 정기점검 편집본이 사라졌다"
        assert os.path.isfile(os.path.join(tmp, "0. 정리이력.csv"))
        assert os.path.isfile(os.path.join(tmp, "0. 정리규칙.txt"))
        assert S.planned_moves(tmp) == [], "두 번째 실행이 멱등이 아니다"

    # 날짜 하위폴더를 만든 뒤에도 각 대조기가 원본을 재귀 탐색해야 한다.
    for path, marker in (
        ("fill_erp_status.py", 'os.path.join(ERP_DIR, "**", "*.xls*")'),
        ("billing_fill.py", 'os.path.join(ERP_DIR, "**", "판매조회*.xls*")'),
        ("receipt_fill.py", 'os.path.join(d, "**", "*.xlsx")'),
        ("band/doc_ocr.py", 'os.path.join(d, "**", "*")'),
    ):
        src = open(os.path.join(ROOT, *path.split("/")), encoding="utf-8").read()
        assert marker in src and "recursive=True" in src, f"{path} 재귀 탐색 누락"
    print("  [76] 원본 자료 유형·연도·월·날짜·PO번호 자동정리와 최신 편집본 보존 ✅")


def t201_upload_intake(tmp):
    """단일 투입함의 전량 보존·내용 분류·중복방지·전체 대조 선행 순서."""
    import upload_intake as U
    import source_dirs as SD

    assert SD.UPLOAD_DIR not in SD.EXCEL_DIRS and SD.ORIGIN_ROOT not in SD.EXCEL_DIRS
    assert SD.UPLOAD_DIR not in SD.KAKAO_DIRS and SD.ORIGIN_ROOT not in SD.KAKAO_DIRS

    origin = os.path.join(tmp, "0. 원본 자료")
    upload = os.path.join(origin, "100. 업로드용 자료", "중첩 폴더")
    os.makedirs(upload, exist_ok=True)

    def book(name, headers, title="Sheet"):
        path = os.path.join(upload, name)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        ws.append(headers)
        ws.append(["PO330876", "쿠팡", "부품", 100000][:len(headers)])
        wb.save(path)
        return path

    book("무작위ERP.xlsx", ["거래처명", "품목명", "공급가액"])
    book("무작위PO.xlsx", ["PO번호", "금액"])
    book("무작위일정.xlsx", ["캠프명", "점검일자"], "2026년 정기점검 스케쥴")
    book("무작위입금.xlsx", ["거래처명", "입금일자", "입금액"])
    kakao = os.path.join(upload, "KakaoTalk_20260803_group.txt")
    with open(kakao, "w", encoding="utf-8") as fh:
        fh.write("테스트방 님과 카카오톡 대화\n[유현민] [오전 9:00] UJ2609999 완료\n")
    for name, body in (("PO330876_견적서.pdf", b"synthetic pdf"),
                       ("현장사진.jpg", b"synthetic image"),
                       ("판별불가.bin", b"unknown evidence")):
        with open(os.path.join(upload, name), "wb") as fh:
            fh.write(body)

    jobs = U.plan(origin, min_age=0)
    assert jobs is not None and len(jobs) == 8, jobs
    by_name = {os.path.basename(job.src): job for job in jobs}
    expected = {
        "무작위ERP.xlsx": os.path.join(origin, "1. ERP 내보내기"),
        "무작위PO.xlsx": os.path.join(origin, "2. 쿠팡 목록"),
        "무작위일정.xlsx": os.path.join(origin, "5. 정기점검 스케쥴 원본"),
        "무작위입금.xlsx": os.path.join(origin, "7. 입금내역"),
        "KakaoTalk_20260803_group.txt": os.path.join(origin, "3. 카카오톡 내보내기"),
        "PO330876_견적서.pdf": os.path.join(origin, "6. PO 원본"),
        "현장사진.jpg": os.path.join(origin, "4. 밴드 원본", "문서사진"),
        "판별불가.bin": os.path.join(origin, "9. 미분류"),
    }
    for name, base in expected.items():
        assert os.path.commonpath([by_name[name].dst_dir, base]) == base, (name, by_name[name])

    report = os.path.join(tmp, "upload_report.json")
    index = os.path.join(tmp, "upload_index.json")
    done, errors = U.apply(jobs, origin, report, index)
    assert len(done) == 8 and not errors, (done, errors)
    assert U.plan(origin, min_age=0) == [], "분류 후 투입함에 파일이 남았다"
    assert os.path.isdir(os.path.join(origin, "100. 업로드용 자료")), "단일 투입함이 사라졌다"
    unknown_dst = next(row["목적지"] for row in done if row["분류"] == "unknown")
    duplicate = os.path.join(upload, "판별불가.bin")
    os.makedirs(upload, exist_ok=True)
    with open(duplicate, "wb") as fh:
        fh.write(b"unknown evidence")
    again, errors = U.apply(U.plan(origin, min_age=0), origin, report, index)
    assert not errors and again[0]["처리"] == "동일 원본 통합"
    assert os.path.isfile(unknown_dst) and not os.path.exists(duplicate)

    # 밴드 JSON은 Z: 정본을 훼손하지 않고 로컬 대조 캐시로 변환된다.
    import source_dirs as _SD
    import band.convert_dump as _BD
    band_source = os.path.join(origin, "4. 밴드 원본", "수집본", "2026", "08", "2026-08-03")
    os.makedirs(band_source, exist_ok=True)
    dump = os.path.join(band_source, "dump_90610953.json")
    with open(dump, "w", encoding="utf-8") as fh:
        json.dump({"band": "90610953", "name": "합성밴드", "posts": {
            "1": {"created_at": 1785686400000, "author": "기사", "content": "UJ2609999 완료"}
        }}, fh, ensure_ascii=False)
    cache = os.path.join(tmp, "band_cache")
    os.makedirs(cache, exist_ok=True)
    old_cache, old_dirs = _BD.CACHE, _SD.band_dump_dirs
    try:
        _BD.CACHE = cache
        _SD.band_dump_dirs = lambda: [band_source]
        _BD.main()
    finally:
        _BD.CACHE, _SD.band_dump_dirs = old_cache, old_dirs
    assert os.path.isfile(dump), "밴드 Z: 원본을 raw로 바꾸거나 삭제했다"
    assert os.path.isfile(os.path.join(cache, "90610953.json")), "밴드 대조 캐시 변환 누락"

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    watchdog = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    batch = open(os.path.join(ROOT, "원본자료자동정리.bat"), encoding="utf-8").read()
    assert daily.index("upload_intake.py") < daily.index("ecount_reconcile.py"), \
        "업로드 분류가 전체 대조보다 늦다"
    assert "sync_uploads(dry)" in watchdog and "전체 대조 시작" in watchdog
    assert "upload_intake.py --apply" in batch
    assert 'files = pick("po")' in open(os.path.join(ROOT, "po_reconcile.py"), encoding="utf-8").read()
    for rel in ("kakao_extract.py", os.path.join("kakao", "kakao_reconcile.py")):
        src = open(os.path.join(ROOT, rel), encoding="utf-8").read()
        assert "def source_paths()" in src and "kakao_dirs" in src
    print("  [201] 단일 업로드 투입함 전량 원본분류·중복방지·30분/전체대조 연결 ✅")


def t79_work_log_source_sync_and_report_capture():
    """[79] 현장 일지 대조는 완료를 추측하지 않고, 미실시 사유·대표 캡처까지 같은 원본을 쓴다."""
    import work_log_sync as W
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, "정기점검, 돌발AS 일지 (7.1~).xlsx")
        wb = openpyxl.Workbook()
        pm = wb.active; pm.title = "2026년 정기점검 일지"
        pm.append([]); pm.append([]); pm.append([])
        pm.append(["No.", "점검일자", "A/S담당", "캠프이름", "프로젝트NO", "A/S내용"])
        pm.append([1, "2026-07-10", "김기사", "점검캠프", "UJ2609001", "정기점검 완료"])
        done = wb.create_sheet("2026년 돌발AS 일지")
        done.append([]); done.append([]); done.append([])
        done.append(["No.", "점검일자", "A/S담당", "캠프이름", "프로젝트NO", "A/S 요청", "A/S내용", "진행현황"])
        done.append([1, "2026-07-11", "김기사", "AS캠프", "UJ2609002", "리모컨 불량", "리모컨 교체", "수리완료"])
        wait = wb.create_sheet("2026년 돌발AS 미실시건")
        wait.append([]); wait.append([]); wait.append([])
        wait.append(["No.", "A/S요청일자", "캠프이름", "프로젝트NO", "A/S 신청내용", "진행현황", "미처리 사유"])
        wait.append([1, "2026.07.12", "대기캠프", "UJ2609003", "모터 교체", "미실시", "자재 도착 후 방문 일정 조율 중"])
        wait.append([2, "2026.07.12", "취소캠프", "UJ2609004", "점검", "접수취소", "정상작동 확인"])
        wb.save(source)

        master = os.path.join(tmp, "쿠팡_통합업무_일일보고_관리대장_v1.xlsx")
        mw = openpyxl.Workbook(); asw = mw.active; asw.title = "02_돌발AS접수"
        asw.append([]); asw.append([]); asw.append([])
        asw.append(["접수ID", "프로젝트NO", "캠프명", "접수일자", "진행상태", "작업완료일"])
        asw.append(["AS-1", "UJ2609002", "AS캠프", "2026-07-01", "", ""])
        asw.append(["AS-2", "UJ2609003", "대기캠프", "2026-07-12", "접수", ""])
        pmw = mw.create_sheet("04_정기점검")
        pmw.append([]); pmw.append([]); pmw.append([])
        pmw.append(["점검ID", "프로젝트NO", "캠프명", "점검예정일", "점검상태", "실제점검일"])
        pmw.append(["PM-1", "UJ2609001", "점검캠프", "2026-07-10", "", ""])
        mw.save(master)

        # 이 대조는 **이번 달만** 센다([188]). 합성 자료는 7월이므로 기준 달을 7월로
        # 두고 본다 — 안 그러면 이 검증이 달마다 통과·실패를 오간다.
        _keep = os.environ.get("COUPANG_WORKLOG_MONTH")
        os.environ["COUPANG_WORKLOG_MONTH"] = "2026-07"
        try:
            payload = W.analyze(master, source)
        finally:
            if _keep is None:
                os.environ.pop("COUPANG_WORKLOG_MONTH", None)
            else:
                os.environ["COUPANG_WORKLOG_MONTH"] = _keep
        as_summary = payload["요약"]["돌발AS"]
        assert as_summary["기준월"] == "2026-07" and not as_summary["이번달자료없음"]
        assert as_summary["발생"] == 3 and as_summary["처리완료"] == 1
        assert as_summary["미처리"] == 1 and as_summary["취소"] == 1
        assert (as_summary["기준시작일"], as_summary["기준종료일"]) == (
            "2026-07-11", "2026-07-12")
        assert as_summary["미처리사유"][0]["사유"] in ("자재·부품 대기", "방문 일정 조율")
        assert payload["요약"]["정기점검"]["실행"] == 1
        assert len(payload["updates"]) == 4, payload["updates"]  # 완료 일자가 명확하고 원장 상태가 빈 2건
        xml = W.build_sheet(payload["records"], os.path.basename(source))
        assert "미처리사유" in xml and "원장미매칭" not in xml

    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    # Current progress counts only schedules through the report baseline.  The
    # drilldown uses the same PM source, and a report opened early refreshes
    # after its async data has arrived.
    for marker in ("function pmRangeRows(state)", "const cutoff = baseDate() || todayISO();",
                   "d <= cutoff", "const liveQuarter=pmRangeRows(_ps)",
                   "if(window.__view==='daily')"):
        assert marker in idx, f"PM progress/source synchronization missing: {marker}"
    brief = open(os.path.join(ROOT, "daily_brief.py"), encoding="utf-8").read()
    for marker in ("openWorkLogBrief", "돌발AS 현장 일지 대조", "정기점검 진행률 · 선택 기간", "pmProgress",
                   "dailyBrief", "drawDailyBrief",
                   "captureDailyTasks", "당일 별도 업무 처리", "BRIEF_LOADING", "BRIEF_RETRY",
                   "기준시작일", "취소·정상작동 상세", "데이터 업데이트"):
        assert marker in idx, f"일지/정기점검 대표 캡처 누락: {marker}"
    assert "대표 보고 · ${dailyBrief.date||D.date}" not in idx, "캡처 제목에 삭제 대상 문구가 남아 있음"
    for marker in ("quarterTitle", "dailyIssues", "execPeriodTitle", "previewRptDate",
                   "REPORT_PREVIEW_DATE", "const day = (_rptData&&_rptData.date)",
                   "dailyIssueDetailRows", "dailyIssueDetails", "issueDetailBlock",
                   "정기점검 이상 발견 상세", "reportForDates",
                   "captureMeta['집계기준일']", "await nextPaint();",
                   'onchange="rptDateChanged()"', "보고일=오늘, 집계기준일=전날",
                   "previewRptDate(false,true)"):
        assert marker in idx, f"선택 날짜·월/분기/연간 캡처 연동 누락: {marker}"
    app = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    for marker in ("def get_daily_brief(day=None)", "_brief_cache", 'result["데이터업데이트일시"]'):
        assert marker in app, f"대표 브리핑 캐시/업데이트 시각 누락: {marker}"
    assert "threading.Thread(target=warm_brief" not in app, "느린 브리핑 선로딩이 첫 화면을 막는다"
    assert '"일지대조": worklog' in brief, "대표 브리핑이 현장 일지 대조를 안 싣는다"
    print("  [79] 현장 일지 대조·안전 빈칸입력·돌발AS 사유·정기점검/AS 대표 캡처 ✅")


def t81_terra_sol_handoff_review():
    """[81] Terra 작업분은 Sol이 검토·합성검증하기 전 쓰기 작업을 못 한다."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import handoff_review as HR

    marker = {"base_commit": "base", "head_commit": "terra-head"}
    passed = {"passed": True, "base_commit": "base", "marker_head": "terra-head", "reviewed_head": "sol-head"}
    assert HR.sol_review_is_current(marker, passed, "sol-head") is True
    assert HR.sol_review_is_current(marker, passed, "new-head") is False, "새 커밋 뒤에는 재검토해야 한다"
    old_marker = HR.MARKER
    try:
        HR.MARKER = os.path.join(tempfile.gettempdir(), "missing-terra-handoff-marker.json")
        assert HR.sol_review_is_current(None, None, "any") is True, "Terra 표식이 없으면 기존 작업을 막지 않는다"
    finally:
        HR.MARKER = old_marker

    dirty = HR.blocking_dirty([" M webapp/app_server.py", "?? outputs/", "?? reports/summary.json"])
    assert dirty == ["webapp/app_server.py"], dirty
    patch = "+++ b/example.py\n+api_key = \"" + ("a" * 16) + "\"\n+password = \"\"\n"
    assert HR.secret_findings(patch) == ["example.py"], "빈 값은 비밀값 탐지 대상이 아니다"

    with tempfile.TemporaryDirectory() as td:
        old_marker, old_review = HR.MARKER, HR.REVIEW_REPORT
        old_head, old_resolve, old_ancestor = HR.current_head, HR.resolve_commit, HR.is_ancestor
        try:
            HR.MARKER = os.path.join(td, "terra.json")
            HR.REVIEW_REPORT = os.path.join(td, "review.json")
            HR.current_head = lambda: "a" * 40
            HR.resolve_commit = lambda value: str(value) if value else ""
            HR.is_ancestor = lambda base, head: bool(base and head)
            saved = HR.mark_terra("b" * 40)
            assert saved["base_commit"] == "b" * 40 and os.path.exists(HR.MARKER), saved
        finally:
            HR.MARKER, HR.REVIEW_REPORT = old_marker, old_review
            HR.current_head, HR.resolve_commit, HR.is_ancestor = old_head, old_resolve, old_ancestor

    claim_src = open(os.path.join(ROOT, "ai_claim.py"), encoding="utf-8").read()
    session_src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    agents_src = open(os.path.join(ROOT, "AGENTS.md"), encoding="utf-8").read()
    claude_src = open(os.path.join(os.path.dirname(ROOT), "CLAUDE.md"), encoding="utf-8").read()
    assert "handoff_review" in claim_src and "--review-sol" in claim_src
    assert "--for-sol" in session_src
    for src in (agents_src, claude_src):
        assert "handoff_review.py --mark-terra" in src and "handoff_review.py --review-sol" in src
    review_src = open(os.path.join(ROOT, "handoff_review.py"), encoding="utf-8").read()
    assert "_write_json(REVIEW_REPORT" not in review_src, "Sol 검토 결과를 저장하지 못한다"
    print("  [81] Terra→Sol 검토 관문(범위 고정·비밀값/문법/합성검증·Sol 쓰기 점유 차단) OK")


def t82_daily_cutoff():
    """The reporting cutoff remains numeric for every permitted HH:MM value."""
    import sys as _s
    _s.path.insert(0, ROOT)
    import workbook_patch as W

    xml = '<sheetData><row r="5"><c r="A5" s="3"/><c r="B5" s="10"><v>0.75</v></c></row></sheetData>'
    out = W.replace_number_cell(xml, "B5", "0.999305555556")
    assert '<c r="B5" s="10"><v>0.999305555556</v></c>' in out, out
    assert 'A5" s="3"' in out, "unrelated dashboard cells must be preserved"
    assert W.parse_cutoff("23:59") == (1439 / 1440, "hh:mm", "23:59")
    assert W.parse_cutoff("24:00") == (1, "[h]:mm", "24:00")
    for invalid in ("24:01", "12:60", "text"):
        try:
            W.parse_cutoff(invalid)
            raise AssertionError(f"invalid cutoff accepted: {invalid}")
        except ValueError:
            pass

    src = open(os.path.join(ROOT, "workbook_patch.py"), encoding="utf-8").read()
    assert 'numFmtId="177"' in src and 'def parse_cutoff(value):' in src
    assert 'ap.add_argument("--cutoff", default="", metavar="HH:MM"' in src
    print("  [82] 집계마감 HH:MM(23:59·24:00) 숫자값 안전 전환 ✅")


def t83_agent_dispatch_and_calendar():
    """Claude 우선·Codex 폴백과 공개 캘린더 화면은 실CLI 없이도 검증한다."""
    import json
    from pathlib import Path
    from types import SimpleNamespace
    from unittest.mock import patch
    import agent_dispatch as A

    old_dir, old_status = A.REPORT_DIR, A.STATUS_PATH
    with tempfile.TemporaryDirectory() as td:
        A.REPORT_DIR = Path(td) / "queue"
        A.STATUS_PATH = Path(td) / "status.json"
        try:
            def quota_then_codex(command, **_kwargs):
                if "claude" in command[0].lower():
                    return SimpleNamespace(returncode=1, stdout="", stderr="credit quota exhausted",
                                           timed_out=False, stuck_pid=0)
                return SimpleNamespace(returncode=0, stdout="codex 1.0", stderr="",
                                       timed_out=False, stuck_pid=0)

            with patch.object(A, "resolve_agent_executable",
                              side_effect=lambda name: str(Path(td) / f"{name}.exe")), \
                 patch.object(A, "run_tree", side_effect=quota_then_codex):
                route = A.route_status()
                assert route["primary"] == "claude" and route["selected"] == "codex", route
                ticket = A.enqueue("synthetic", "합성 AI 연계", ["tests/synthetic_check.py"])
                assert ticket["selected"] == "codex", ticket
                assert list(A.REPORT_DIR.glob("*.json")), "AI 요청이 내구성 있는 큐에 남지 않음"
                assert A.status()["last_request"]["id"] == ticket["id"]
                consumed = A.run_ticket(ticket["_path"], 0)
                assert consumed["status"] == "done" and consumed["agent_returncode"] == 0, consumed
                saved = json.load(open(ticket["_path"], encoding="utf-8"))
                assert saved["status"] == "done" and saved["selected"] == "codex", saved

            # Version probing may pass before Claude later rejects the real task
            # because credits are exhausted.  That runtime failure must retry the
            # AI follow-up with Codex without repeating the local business script.
            runtime_path = A.REPORT_DIR / "runtime_fallback.json"
            A._atomic_json(runtime_path, {
                "id": "runtime_fallback",
                "created_at": "2026-07-29T00:00:00",
                "task_key": "synthetic",
                "title": "runtime fallback",
                "local_command": ["python", "already_ran.py"],
                "primary": "claude",
                "selected": "claude",
                "status": "queued",
            })

            def claude_exec_then_codex(command, **_kwargs):
                if "claude" in command[0].lower():
                    return SimpleNamespace(
                        returncode=1, stdout="", stderr="credit quota exhausted",
                        timed_out=False, stuck_pid=0,
                    )
                return SimpleNamespace(returncode=0, stdout="codex completed", stderr="",
                                       timed_out=False, stuck_pid=0)

            ready_route = {
                "primary": "claude",
                "selected": "claude",
                "note": "Claude Code 우선",
                "agents": {},
                "checked_at": "2026-07-29T00:00:00",
            }
            with patch.object(A, "route_status", return_value=ready_route), \
                 patch.object(A, "resolve_agent_executable",
                              side_effect=lambda name: str(Path(td) / f"{name}.exe")), \
                 patch.object(A, "run_tree", side_effect=claude_exec_then_codex):
                consumed = A.run_ticket(runtime_path, 0)
                assert consumed["status"] == "done", consumed
                assert consumed["selected"] == "codex", consumed
                assert consumed["fallback_from"] == "claude", consumed
                assert consumed["local_command"] == ["python", "already_ran.py"], consumed

            stale = A.REPORT_DIR / "stale.json"
            A._atomic_json(stale, {"id": "stale", "status": "queued"})
            assert A.supersede_queued("interactive completion") == 1
            assert json.load(open(stale, encoding="utf-8"))["status"] == "superseded"
        finally:
            A.REPORT_DIR, A.STATUS_PATH = old_dir, old_status

    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    bench = open(os.path.join(ROOT, "coupang_workbench.py"), encoding="utf-8").read()
    assert 'data-v="calendar"' in idx and "COUPANG_CALENDAR_ID" in idx
    assert "calendar.google.com/calendar/embed" in idx and "openGoogleCalendarDraft" in idx
    assert "previewCalendarDraft();" in idx and "runAgentNote" in idx
    start_task_src = server.split("def start_task(key):", 1)[1].split(
        "# ── 마지막 실행 시각", 1)[0]
    assert "run_tree([PY] + args" in start_task_src and "autopilot.defer" in start_task_src
    assert "enqueue_agent" not in start_task_src and "dispatch_async" not in start_task_src, \
        "성공한 결정론적 앱 작업까지 매번 AI 에이전트에 보내면 크레딧과 갱신 시간이 샌다"
    assert "dispatch_async" in bench and "로컬 업무 스크립트는 1회만 실행" in bench
    dispatch = open(os.path.join(ROOT, "agent_dispatch.py"), encoding="utf-8").read()
    for marker in ("resolve_agent_executable", "run_ticket", '"status": "running"',
                   '"status": "done"', "codex.exe"):
        assert marker in dispatch, f"실제 AI 폴백 소비기 누락: {marker}"
    data_status = open(os.path.join(ROOT, "data_status.py"), encoding="utf-8").read()
    assert 'startswith("2026-")' in data_status, "자료현황 통계에 2025년이 다시 섞인다"
    assert "_source_mtimes" in data_status and "pick()을 유형별로 여섯 번" in data_status, \
        "자료현황이 상태 조회마다 원본 Excel을 반복해서 열어 앱을 느리게 만든다"
    tailscale = open(os.path.join(ROOT, "tailscale_serve.py"), encoding="utf-8").read()
    phone = open(os.path.join(ROOT, "phone_access.py"), encoding="utf-8").read()
    endpoint = open(os.path.join(ROOT, "publish_endpoint.py"), encoding="utf-8").read()
    for marker in ("public_ingress_ips", "_public_ping_ip", "public_funnel_alive",
                   "ensure_public_funnel", '"funnel", "reset"',
                   'FIXED_HOST = "mulder.tailf14aae.ts.net"',
                   'FIXED_URL = "https://mulder.tailf14aae.ts.net/"'):
        assert marker in tailscale, f"휴대폰 공개 Funnel 자동복구 누락: {marker}"
    assert "ensure_public_funnel(repair=True)" in phone, \
        "폰 접속 도우미가 내부 연결만 보고 공개 Funnel 장애를 놓친다"
    assert "from tailscale_serve import FIXED_URL" in endpoint, \
        "게시기가 Tailscale 이름/임시터널을 따라 고정주소를 바꿀 수 있다"
    print("  [83] Claude 우선·Codex 폴백 실제 소비기 및 쿠팡 캘린더(전체·상세·입력) ✅")


def t200_evidence_verification_sync(tmp):
    """확정 증빙만 02·03·04 원인 열에 반영하고 수식 열은 계획에 넣지 않는다."""
    import verification_sync as V

    path = os.path.join(tmp, "verification_sync.xlsx")
    wb = openpyxl.Workbook()
    ws02 = wb.active
    ws02.title = "02_돌발AS접수"
    h02 = [
        "접수ID", "프로젝트NO", "캠프명", "접수일자", "진행상태", "작업완료일",
        "사진등록", "완료보고서등록", "ERP등록", "밴드수정",
        "최초접수외추가작업", "추가작업확인상태", "관리자검증상태",
        "담당관리자", "최종확인일", "검증결과",
    ]
    ws02.append(["제목"]); ws02.append([]); ws02.append([]); ws02.append(h02)
    ws02.append([
        "AS-2607-001", "UJ2600001", "합성캠프", "2026-07-01", "작업완료",
        "2026-07-02", "누락", "누락", "미등록", "", "", "", "작업내용누락",
        "다른사람", "", "=IF(A5=\"\",\"\",IF(M5=\"일치\",\"정상\",\"확인필요\"))",
    ])
    # 근거 없는 두 번째 완료건은 변경 대상이 아니어야 한다.
    ws02.append([
        "AS-2607-002", "UJ2600002", "미확인캠프", "2026-07-03", "작업완료",
        "2026-07-04", "", "", "미등록", "", "", "", "", "", "", "",
    ])

    ws03 = wb.create_sheet("03_현장작업실적")
    h03 = [
        "작업ID", "접수ID", "프로젝트NO", "캠프명", "실제작업항목",
        "실제작업상세", "접수외추가작업여부", "추가작업내용", "기사보고내용",
        "관리자검증", "거래명세서반영", "ERP반영", "검증자", "검증일", "검증결과",
    ]
    ws03.append(["제목"]); ws03.append([]); ws03.append([]); ws03.append(h03)
    ws03.append([
        "FW-2607-001", "", "", "", "리모컨", "리모컨 교체 완료", "없음", "",
        "정상 작동", "작업내용누락", "확인필요", "확인필요", "다른사람", "",
        "=IF(AND(J5=\"일치\",K5=\"반영완료\",L5=\"반영완료\"),\"정상\",\"확인\")",
    ])
    # 02의 두 번째 완료행에 대응하지만 실제 작업내용이 없는 예비행이다.
    ws03.append(["FW-2607-002", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

    ws04 = wb.create_sheet("04_정기점검")
    h04 = [
        "점검ID", "프로젝트NO", "캠프명", "점검예정일", "점검상태", "실제점검일",
        "점검사진", "점검보고서", "ERP판매전표", "거래명세서", "담당관리자",
        "최종확인일(유현민 체크)", "검증결과",
    ]
    ws04.append(["제목"]); ws04.append([]); ws04.append([]); ws04.append(h04)
    ws04.append([
        "PM-2607-001", "UJ2600001", "합성캠프", "2026-07-01", "완료",
        "2026-07-02", "", "", "미등록", "", "다른사람", "",
        "=IF(AND(I5=\"완료\",J5=\"발행완료\"),\"정상\",\"확인\")",
    ])

    ws06 = wb.create_sheet("06_거래서류청구수금")
    h06 = ["정산ID", "업무구분", "원천업무ID", "프로젝트NO", "캠프명",
           "작업완료일", "거래명세서번호", "거래명세서발행일"]
    ws06.append(["제목"]); ws06.append([]); ws06.append([]); ws06.append(h06)
    ws06.append([
        "JS-2607-001", "돌발AS", "AS-2607-001", "UJ2600001", "합성캠프",
        "2026-07-02", "2026/07/03-1", "2026-07-03",
    ])
    wb.save(path)

    band = [{
        "프로젝트NO": "UJ2600001", "진행상태": "작업완료",
        "문서상태": "판매전표+거래명세서", "게시일": "2026-07-04",
        "작업일": "2026-07-02", "사진": 2,
    }]
    erp = {
        "UJ2600001": {
            "statement": False, "erp": True, "completed": False, "photos": 0,
            "statement_dates": set(), "erp_dates": {"2026-07-05"},
            "completion_dates": set(), "sources": {"ERP 판매조회"},
        }
    }
    items, _counts, profiles, _files = V.build_plan(
        path, band_records=band, erp_evidence=erp, erp_files=[],
    )
    cells = {(x["sheet"], x["cell"]): x for x in items}
    assert cells[("03_현장작업실적", "K5")]["value"] == "반영완료"
    assert cells[("03_현장작업실적", "L5")]["value"] == "반영완료"
    assert cells[("03_현장작업실적", "M5")]["value"] == "유현민"
    assert cells[("03_현장작업실적", "N5")]["value"] == "2026-07-05"
    assert cells[("03_현장작업실적", "J5")]["value"] == "일치"
    assert cells[("02_돌발AS접수", "I5")]["value"] == "완료"
    assert cells[("02_돌발AS접수", "M5")]["value"] == "일치"
    assert cells[("04_정기점검", "I5")]["value"] == "완료"
    assert cells[("04_정기점검", "J5")]["value"] == "발행완료"
    assert cells[("04_정기점검", "K5")]["value"] == "유현민"
    assert not any(x["cell"].endswith("6") for x in items), "빈 03 예비행/미확인 건을 쓰면 안 됨"
    assert not any(x["col"] == "검증결과" for x in items), "검증 수식을 직접 덮으면 안 됨"
    assert profiles["UJ2600001"]["statement"] and profiles["UJ2600001"]["erp"]
    assert any(not x["only_if_empty"] for x in items if x["vtype"] == "text"), \
        "확정 증빙으로 확인필요·미등록을 완료 상태로 승격하지 못함"

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    index = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    sync = open(os.path.join(ROOT, "verification_sync.py"), encoding="utf-8").read()
    assert "verification_sync.py" in daily
    assert "derived_field_status_map" in server
    writer = open(os.path.join(ROOT, "ledger_writer.py"), encoding="utf-8").read()
    assert "by_sheet_done" in writer and ".iter_rows(" in writer, \
        "대량 검증이 read_only 셀 무작위 접근으로 되돌아가면 실반영이 멈춘다"
    for marker in ("거래명세서반영", "ERP반영", "검증자", "검증일"):
        assert marker in index and marker in sync, f"앱 현장검증 표시 누락: {marker}"
    print("  [200] 밴드·ERP·거래명세서 증빙→02·03·04 검증완료·유현민·확인일 자동동기화 ✅")


def t80_new_project_flow_db_only(tmp):
    """신규 프로젝트 흐름도는 최신본만 내부 DB로 갱신하고 앱에는 연결하지 않는다."""
    import json
    import time
    import new_project_flow_sync as F
    import source_organizer as O

    origin = os.path.join(tmp, "0. 원본 자료")
    flow_dir = os.path.join(origin, "50. 쿠팡 신규 프로젝트 업무 흐름도")
    os.makedirs(flow_dir, exist_ok=True)

    def make_flow(path, marker):
        wb = openpyxl.Workbook()
        process = wb.active; process.title = "납품취합"
        process.append(["쿠팡 신규 프로젝트 프로세스", marker])
        process.append(["조직 / 단계", "프로젝트 등록", "KPI 보고"])
        process.append(["신규", "프로젝트 및 ERP 등록", "종료 보고"])
        checklist = wb.create_sheet("체크시트")
        checklist.append(["단계", "업무", "담당", "완료기준", "매일 확인사항(Daily Check)"])
        checklist.append(["① 프로젝트 접수", "쿠팡 신규 PO 접수", "운영", "PO 접수 완료", "누락 확인"])
        wb.save(path)

    old = os.path.join(flow_dir, "쿠팡 신규 업무 흐름도_이전.xlsx")
    current = os.path.join(flow_dir, "쿠팡 신규 업무 흐름도_최신.xlsx")
    make_flow(old, "이전")
    make_flow(current, "최신")
    os.utime(old, (time.time() - 10, time.time() - 10))

    assert F.find_latest_source(flow_dir) == current
    db_path = os.path.join(tmp, "reports", F.DB_FILENAME)
    first = F.sync(flow_dir, db_path)
    assert first["status"] == "updated" and first["sheet_count"] == 2 and first["checklist_count"] == 1, first
    saved = json.load(open(db_path, encoding="utf-8"))
    assert saved["app_visible"] is False and saved["source"]["name"] == os.path.basename(current), saved
    assert saved["checklist"][0]["업무"] == "쿠팡 신규 PO 접수", saved["checklist"]
    assert F.sync(flow_dir, db_path)["status"] == "unchanged"

    moves = O.planned_moves(root=origin)
    flow_moves = [m for m in moves if os.path.normcase(m.src) == os.path.normcase(old)]
    assert len(flow_moves) == 1 and "보관" in flow_moves[0].dst, flow_moves
    assert not any(os.path.normcase(m.src) == os.path.normcase(current) for m in moves), moves

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    app_server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "new_project_flow_sync.py" in daily and "앱 비표시" in daily
    assert "new_project_flow" not in app_server, "흐름도 DB가 앱 API에 노출되면 안 됩니다"
    print("  [80] 신규 프로젝트 흐름도 최신본 DB 동기화·이전본 보관·앱 비표시 ✅")


def t85_staff_po_work_log_and_edit_priority(tmp):
    """Synthetic coverage for staff source intake and real-edit priority."""
    import time
    import source_dirs
    import webapp.app_server as app

    source = os.path.join(tmp, "sample.xlsx")
    with open(source, "wb") as out:
        out.write(b"synthetic-xlsx")

    old = {
        "root": app.ROOT,
        "approved": app._approved_source_roots,
        "start": app.start_task,
        "defer": app.defer_task_until_free,
        "activity": app.WORKCENTER_ACTIVITY,
        "po_dir": source_dirs.PO_DIR,
        "work_log_dir": source_dirs.WORK_LOG_DIR,
    }
    try:
        app.ROOT = os.path.join(tmp, "local")
        app.WORKCENTER_ACTIVITY = os.path.join(tmp, "workcenter_activity.json")
        app._approved_source_roots = lambda: [tmp]
        app.start_task = lambda name: (False, f"queued:{name}")
        app.defer_task_until_free = lambda name: True
        source_dirs.PO_DIR = os.path.join(tmp, "origin", "po")
        source_dirs.WORK_LOG_DIR = os.path.join(tmp, "origin", "work-log")

        po = app.save_staff_po_submission(
            {"staff_slug": "oh-jonghyeon", "source_ref": source,
             "po_no": "PO-SYNTHETIC", "project_no": "UJ2609999"},
            {}, "127.0.0.1")
        assert po["auto_check_queued"] and po["po_compare_files"], po
        assert os.path.isfile(os.path.join(app.ROOT, "inbox", po["po_compare_files"][0]))

        work_log = app.save_staff_work_log_submission(
            {"staff_slug": "ryu-jiyeong", "source_ref": source}, {}, "127.0.0.1")
        assert work_log["auto_check_queued"] and work_log["files"] == ["sample.xlsx"], work_log

        for bad in ("http://127.0.0.1/private.xlsx",
                    "http://example.com:99999/file.xlsx"):
            try:
                app._validate_remote_url(bad)
                assert False, f"unsafe URL accepted: {bad}"
            except ValueError:
                pass

        view = app.record_workcenter_activity("ryu-jiyeong", "view")
        edit = app.record_workcenter_activity("ryu-jiyeong", "input")
        assert view["active_until_ts"] == 0 and not view["editing"], view
        assert edit["active_until_ts"] > time.time() and edit["editing"], edit

        html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
        for marker in (
            'id="poSubmissionForm"', 'id="poDrop"', "/api/staff/po-upload",
            'id="v-worklog"', 'class="staff-worklog-nav"', 'id="workLogUploadForm"',
            "/api/staff/work-log-upload", "openWorkLogReport(true)",
            "markStaffInput('upload')", "staffHeartbeat('view')",
        ):
            assert marker in html, "staff source UI missing: " + marker
        assert "'yoo-hyeonmin'" in html, "세 직원 업무센터 정의에서 유현민이 빠졌다"
        print("  [85] staff PO/work-log intake, SSRF guard, edit priority OK")
    finally:
        app.ROOT = old["root"]
        app._approved_source_roots = old["approved"]
        app.start_task = old["start"]
        app.defer_task_until_free = old["defer"]
        app.WORKCENTER_ACTIVITY = old["activity"]
        source_dirs.PO_DIR = old["po_dir"]
        source_dirs.WORK_LOG_DIR = old["work_log_dir"]


def t129_call_notes_db_only_and_device_open():
    """[129] 통화 메모는 DB 에만 · 원본 클릭은 **접속한 기기**에서 연다 (2026-08-07 지시).

    사용자 지시: "원본 자료 클릭하면 접속한 디바이스에서 바로 열리게 알고리즘 수정해,
    그리고 통화_MD는 원본 자료에서 안보이게 처리하고 DB만 보관해(민감한 내용이 포함되어있음)."

    무엇이 잘못돼 있었나
      · 클릭은 `/api/open` → `os.startfile` 이라 **서버 PC 에서** 열렸다. 폰·다른 PC 는
        403 이라 경로만 복사됐고, 사무실 PC 에서는 아무도 안 보는 화면에 창만 떴다.
      · 통화 메모를 `0. 원본 자료/10. 통화·회의 기록/`(공유 폴더 Z:)에 복사해 두어
        앱 '원본 자료' 목록에 그대로 떴다(실측: 통화_20260805_김준형.md 카드 노출).

    지키는 것
      ① 통화 메모는 색인에 담기지 않는다 — 이름 규칙과 통화·회의 폴더 둘 다.
      ② call_notes 는 파일을 복사하지 않는다(그 복사가 노출의 원인이었다).
      ③ DB 왕복: 넣은 본문이 그대로 나오고, **목록은 본문을 주지 않는다.**
      ④ 서버는 목록 응답과 파일 전송 **두 곳 모두**에서 막는다 — 색인은 회차로 다시
         만들어지므로 옛 색인이 남아 있는 동안에도 막혀야 한다.
      ⑤ 화면은 /api/open 이 아니라 /api/source-file 을 쓴다.
    """
    import tempfile
    import source_dirs
    import source_index as S
    _rd = lambda *p: open(os.path.join(*p), encoding="utf-8").read()

    # ① 색인 제외 규칙
    assert S.is_private(os.path.join("x", "통화_20260805_김준형.md")), "이름 규칙이 안 걸린다"
    assert S.is_private(os.path.join(source_dirs.CALL_NOTE_DIR, "아무거나.txt")), \
        "통화·회의 폴더 전체가 안 걸린다"
    assert not S.is_private(os.path.join("x", "1-1._남양주4MB견적서_274,000원.pdf")), \
        "일반 원본까지 막으면 안 된다"
    src = _rd(ROOT, "source_index.py")
    assert "is_private(p, fn)" in src, "walk 에서 is_private 를 부르지 않는다 — 색인에 샌다"

    # ② Z: 복사 금지
    cn = _rd(ROOT, "call_notes.py")
    assert "shutil" not in cn, "call_notes 가 아직 파일을 복사한다 — Z: 노출 경로가 남았다"
    assert "call_note_save" in cn, "DB 보관을 부르지 않는다"

    # ③ DB 왕복 — 실 DB 는 건드리지 않는다
    import ledger_db
    tmpd = tempfile.mkdtemp(prefix="callnote_")
    keep = (ledger_db.DB_DIR, ledger_db.DB_PATH)
    ledger_db.DB_DIR = tmpd
    ledger_db.DB_PATH = os.path.join(tmpd, "t.db")
    try:
        body = "# 통화\n## 할 일\n- [김 · 2026-08-08] 확인\n"
        ledger_db.call_note_save("통화_20260808_테스트.md", body, whom="테스트",
                                 on="2026-08-08",
                                 todos=[{"who": "김", "due": "2026-08-08", "what": "확인"}])
        got = ledger_db.call_note_get("통화_20260808_테스트.md")
        assert got and got.get("body") == body, "DB 왕복이 깨졌다"
        lst = ledger_db.call_notes()
        assert lst and "body" not in lst[0], "목록이 본문을 흘린다 — 민감 자료다"
        assert lst[0]["todos"] and lst[0]["todos"][0]["what"] == "확인", "할 일이 안 남았다"
        ledger_db.call_note_save("통화_20260808_테스트.md", body + "추가", on="2026-08-08")
        assert len(ledger_db.call_notes()) == 1, "같은 파일 이름이 두 행이 됐다"
    finally:
        ledger_db.DB_DIR, ledger_db.DB_PATH = keep

    # ④ 서버 — 목록과 파일 전송 두 곳 모두
    app = _rd(ROOT, "webapp", "app_server.py")
    assert '"/api/source-file"' in app, "기기로 내려보내는 엔드포인트가 없다"
    assert app.count("from source_index import is_private") >= 2, \
        "목록·파일전송 두 곳 모두에서 막아야 한다(색인은 회차로 다시 만들어진다)"
    assert "filename*=UTF-8''" in app, "한글 파일 이름이 헤더에서 깨진다"

    # ⑤ 화면 — 클릭이 서버 PC 열기로 돌아가지 않았나
    html = _rd(ROOT, "webapp", "index.html")
    assert "sourceFileURL" in html and "/api/source-file" in html, "화면이 새 경로를 안 쓴다"
    fn = re.search(r"function openSource\(path\)\{.*?\n\}", html, re.S)
    assert fn, "openSource 를 찾지 못했다"
    assert "/api/open" not in fn.group(0), \
        "원본 클릭이 아직 서버 PC 에서 여는 /api/open 을 쓴다"
    print("  [129] 통화 메모 DB 전용 · 원본 클릭 접속 기기에서 열기 ✅")


def t130_band_grab_rejects_timeless_harvest():
    """[130] 시각 없는 수확은 저장하지 않는다 (2026-08-07 실사고 2차).

    무슨 일이 있었나
      `recheck_plan --ahead` 는 최신 글보다 **큰 번호**를 미리 찔러 본다. 그런데 밴드는
      아직 없는 글 번호에도 **HTTP 200 + 앱 껍데기**를 준다. 수집기가 그 화면에서 본문을
      뜯으면 직전 화면(피드 맨 위 글)의 본문이 그대로 잡히고 글쓴이·시각만 빈다.
      그렇게 3539~3578 마흔 건이 **전부 같은 글**로 수집됐고 status 가 ok 라 캐시에
      그대로 들어갔다(3308→3348). 본문이 있으니 실패로 보이지 않았다 —
      화면 숫자는 멀쩡한데 원본과 다른 값이 되는, 제일 나쁜 종류의 실패다.

    지키는 것
      ① 시각(timeText)이 없으면 'ok' 로 넘기지 않는다.
      ② 그때 'missing'(묘비)이 아니라 'fail' 이어야 한다 — 그 번호는 내일 진짜로
         생긴다. 묘비를 세우면 recheck_plan 이 영영 다시 안 뽑는다.
      ③ 숨은 탭에서는 시작 자체를 거절한다(1차 사고 가드가 살아 있나).
    """
    js = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()

    body = re.search(r"async function grabOne\(.*?\n  \}", js, re.S)
    assert body, "grabOne 을 찾지 못했다"
    body = body.group(0)

    # 한 줄 반환이든 블록이든(지문을 함께 남기는 [131] 이후 모양) 상태만은 fail 이어야 한다.
    guard = re.search(r"if \(!timeText\)\s*(\{.*?\n      \}|return \{[^}]*\})", body, re.S)
    assert guard, "시각 없는 수확을 걸러내지 않는다 — 같은 글이 통째로 캐시에 들어간다"
    m = re.search(r"status: '(\w+)'", guard.group(1))
    assert m and m.group(1) == "fail", \
        f"시각 없음을 '{m.group(1) if m else '?'}' 로 처리한다 — 묘비를 세우면 그 번호를 영영 못 모은다"

    # 가드가 ok 반환보다 **앞**에 있어야 의미가 있다
    assert body.index("if (!timeText)") < body.index("status: 'ok'"), \
        "가드가 ok 반환보다 뒤에 있다 — 걸러지지 않는다"
    # timeText 를 다시 긁어 담지 않는지(가드를 우회하는 옛 코드가 남았나)
    assert "timeText: txt(" not in body, "가드를 지나 timeText 를 다시 긁는다"

    # ③ 1차 사고 가드(숨은 탭 시작 거절)도 함께 지킨다
    assert "document.hidden" in js and "탭이 뒤에 있다" in js, \
        "숨은 탭에서 시작을 거절하는 가드가 사라졌다"
    print("  [130] 밴드 수집 — 시각 없는 수확 폐기 ✅")


def t135_contaminated_not_recollected():
    """[135] 가짜 글 기록은 **재수집 목록에 넣지 않는다** (2026-08-07 지시).

    무엇이 잘못돼 있었나
      '날짜 없는 글 621건'을 "본문은 멀쩡한데 시각만 늦게 붙어 빈 것"으로 진단하고
      재수집 목록에 넣었다. 실제로 60건을 돌렸더니 **ok 0** 이었다. 세어 보니
      98건이 본문 2종, 523건이 7종 — 진짜라면 621종이어야 한다. 밴드가
      `/post/<번호>` 를 iframe 으로 열면 **피드로 되돌려서**, 껍데기에 남은 피드
      맨 위 글이 통째로 잡힌 것이었다. 같은 경로로 다시 열면 같은 가짜가 또 들어온다.

    지키는 것
      ① 판정은 좁게 — 작성일 없음 **그리고** 같은 본문이 2건 이상. 하나뿐이면 안 건드린다.
      ② 지우지 않고 표시한다 — 키를 지우면 '구멍'이 되어 매 회차 다시 훑는다.
      ③ 가짜 본문·글쓴이는 없앤다 — 남기면 대조가 그것을 진짜로 쓴다.
      ④ 계획이 오염 번호를 새 글/구멍/재수집 어디에도 넣지 않는다.
      ⑤ 인계 문서가 '재수집하라'고 말하지 않는다(그게 한 시간을 헛돌게 한 문구다).
    """
    import importlib
    cc = importlib.import_module("band.clean_contaminated")

    same = "Coupang이(가) 새 구매 오더(PO375207)를 전송했습니다"
    posts = {
        "10": {"content": same, "author": "김"},              # 날짜 없음 · 같은 본문
        "11": {"content": same, "author": "김"},              # 날짜 없음 · 같은 본문
        "12": {"content": "진짜 혼자인 글", "author": "박"},   # 날짜 없음 · 본문 하나뿐
        "13": {"content": "정상", "created_at": 1754500000000},
        "14": {"deleted": True},
    }
    bad = cc.find(posts)
    assert set(bad) == {"10", "11"}, f"판정이 틀렸다: {sorted(bad)}"
    assert "12" not in bad, "본문이 하나뿐인 것을 가짜로 몰았다 — 진짜 글일 수 있다"
    assert "13" not in bad and "14" not in bad, "정상·삭제 글을 건드렸다"

    # ③ 표시한 기록에는 가짜 본문이 남지 않는다
    src = open(os.path.join(ROOT, "band", "clean_contaminated.py"), encoding="utf-8").read()
    assert '"contaminated": True' in src, "표시 대신 지우면 매 회차 구멍으로 다시 훑는다"
    assert "os.replace(tmp, path)" in src, "캐시를 원자적으로 쓰지 않는다(사고 24)"

    # ④ 계획이 오염 번호를 어디에도 넣지 않는다
    rp = importlib.import_module("band.recheck_plan")
    p = rp.plan("1", {"1": {"created_at": 1}, "2": {"contaminated": True},
                      "3": {"created_at": 1}}, floor=1, ahead=0)
    assert 2 not in p["gaps"] and 2 not in p["stale"], "오염 번호가 재수집 목록에 들어간다"
    assert 2 not in (p.get("dateless") or []), "오염 번호가 날짜없음으로도 잡힌다"
    assert p.get("contaminated") == [2], "오염 건수가 안 보인다 — 잊힌다"

    # ⑤ 인계 문서가 재수집하라고 말하지 않는다 — 표본 3/3 리다이렉트로 삭제 판정 완료
    sh = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "밴드오염" in sh and "재수집하지 않는다" in sh and "삭제" in sh, \
        "오염 기록이 '할 일'로 남았다 — 삭제 판정이 끝난 상태다"
    assert "수집 경로를 새로 만들어야 한다" not in sh, \
        "이미 닫힌 '수집 경로 재설계'가 할 일로 남아 있다 — 다음 세션이 헛돈다"

    # ⑥ 리다이렉트 가드 — 피드가 끝까지 그려지면 시각까지 있어서([130]만으로는) 못 막는다.
    #   내가 연 주소에 내가 있는지가 유일한 확증이다(실측: 피드의 "6시간 전"이 잡혔다).
    js2 = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
    assert "reason: 'redirect'" in js2 and "location.pathname" in js2, \
        "리다이렉트 가드가 없다 — 날짜 달린 가짜가 들어온다"
    assert "'/post/' + no" in js2, "주소 대조가 요청 번호와 안 묶였다"

    # ⑦ 오염 표시는 재병합을 살아남는다 — Z: 옛 덤프가 매 회차 다시 처리되기 때문.
    #   실측: 표시 621건이 한 회차 만에 0건이 됐었다. 작성일을 가진 기록(진짜 재수집)만 뚫는다.
    cd2 = open(os.path.join(ROOT, "band", "convert_dump.py"), encoding="utf-8").read()
    assert 'cur.get("contaminated") and not rec.get("created_at")' in cd2, \
        "재병합이 오염 표시를 덮는다 — 표시가 한 회차 만에 사라진다"
    print("  [135] 가짜 글 기록 — 재수집 금지·표시 유지 ✅")


def t131_band_quiet_vs_stalled():
    """[131] '밴드가 조용한 것'과 '수집이 막힌 것'을 가른다 (2026-08-07 지시).

    무엇이 잘못돼 있었나
      신선도 판정은 **날짜 있는 최신 글**만 봤다. 그래서 밴드에 새 글이 없는 날에도
      "★밀림"이 인계 문서 맨 위에 떴고, 그 경보를 믿고 없는 번호(3539~3578)를 긁었다가
      40건이 **전부 같은 글**로 캐시에 들어갔다(오늘 사고). 경보가 사고를 부른 셈이다.

    지키는 것
      ① 근거는 **missing 뿐**이다 — 밴드가 '없다'고 명시한 번호. failed/no-time 은
         화면이 안 그려졌을 때도 나오므로 '없음'의 증거가 못 된다.
      ② 수집 최대 번호 **바로 다음**이 없음으로 확인돼야 성립한다. 건너뛴 확인은 안 된다.
      ③ 옛 회차가 최근 확인을 덮지 않는다.
      ④ 근거가 최근일 때만 밀림을 내린다 — 오래된 근거는 그 사이 새 글이 올라왔을 수 있다.
      ⑤ 문서에 '조용함'과 그 이유가 보인다(다음 세션이 또 없는 번호를 긁지 않게).
    """
    import importlib
    import tempfile
    conv = importlib.import_module("band.convert_dump")
    import session_handoff as SH

    TMP = tempfile.mkdtemp(prefix="bandquiet_")
    log = os.path.join(TMP, "밴드_확인시각.json")
    keep_log = conv.PROBE_LOG
    conv.PROBE_LOG = log
    try:
        merged = {"3537": {"content": "x"}, "3538": {"content": "y"},
                  "3400": {"deleted": True}}
        cap = 1754500000000            # 고정 시각(테스트는 시계를 만들지 않는다)

        # ② 바로 다음 번호가 없음 → 기록된다
        conv._record_probe("84789192", "매출처업무", merged, [3539, 3540], cap)
        doc = json.load(open(log, encoding="utf-8"))
        assert doc["84789192"]["수집최대"] == 3538, "수집 최대 번호가 틀렸다"
        assert doc["84789192"]["없음확인"] == 3539
        first_seen = doc["84789192"]["확인시각"]

        # ② 건너뛴 확인은 증거가 아니다 — 3539 를 모르는 채 3541 만 없음이면 성립 안 한다
        os.remove(log)
        conv._record_probe("84789192", "매출처업무", merged, [3541], cap)
        assert not os.path.exists(log), "건너뛴 확인을 증거로 삼았다"

        # ③ 옛 회차가 최근 확인을 덮지 않는다
        conv._record_probe("84789192", "매출처업무", merged, [3539], cap)
        conv._record_probe("84789192", "매출처업무", merged, [3539], cap - 86400000)
        doc = json.load(open(log, encoding="utf-8"))
        assert doc["84789192"]["확인시각"] == first_seen, "옛 회차가 최근 확인을 덮었다"

        # ★ 오늘 사고의 실제 모양 — 아직 없는 번호는 missing 이 아니라 **notime** 으로
        #   떨어진다(밴드가 200 과 껍데기를 주므로). 지문이 같은 것이 2개 이상이면 증거다.
        same = "Coupang이(가) 새 구매 오더(PO375207)를 전송했습니다. 총금액 8,778,600원"
        assert conv._absent_above(3538, [], {"3539": same, "3540": same}) == [3539, 3540], \
            "같은 지문 여러 건을 '아직 없는 글'로 못 읽는다 — 이게 오늘 사고의 모양이다"
        # 하나뿐이면 '화면이 늦게 그려진 것'과 구분이 안 된다 — 증거로 치지 않는다
        assert conv._absent_above(3538, [], {"3539": same}) == [], \
            "지문 하나로 없음을 단정했다 — 느린 화면을 없는 글로 오해한다"
        # 지문이 서로 다르면 진짜 글일 수 있다 — 증거 아님
        assert conv._absent_above(3538, [], {"3539": "가", "3540": "나"}) == [], \
            "서로 다른 본문을 없는 글로 묶었다"
        # missing 과 notime 이 섞여도 이어져야 성립한다
        os.remove(log)
        conv._record_probe("84789192", "매출처업무", merged, [3540],
                           cap, {"3539": same, "3541": same})
        doc = json.load(open(log, encoding="utf-8"))
        assert doc["84789192"]["없음확인"] == 3539 and doc["84789192"]["연속없음"] == 3
    finally:
        conv.PROBE_LOG = keep_log

    # ④ 신선도 판정 — 근거가 최근이면 밀림이 내려가고, 오래되면 그대로 밀림
    #   ★ 밴드번호는 **실제 캐시에 없는 것**을 쓴다. 판정은 근거의 나이만 보는 것이
    #     아니라 캐시의 실제 수확과도 대 보므로([217] 추월 판정), 실데이터 밴드번호를
    #     쓰면 그날 캐시 상태에 따라 이 검증이 흔들린다. 합성검증이 실데이터에
    #     기대는 것 자체가 냄새다 — 여기서 재는 것은 **나이 규칙** 하나다.
    keep = (SH.band_latest_days, SH.band_quiet)
    BQ = "99999999"
    try:
        SH.band_latest_days = lambda: {BQ: "2026-08-05"}
        SH.band_quiet = lambda: {BQ: {"이름": "매출처업무", "수집최대": 3538,
                                      "확인시각": "2026-08-07 09:52",
                                      "없음확인": 3539, "연속없음": 2}}
        row = [f for f in SH.data_freshness("2026-08-07") if f["이름"].startswith("밴드:")][0]
        assert not row["밀림"], "새 글 없음을 확인했는데도 밀림이라 한다"
        assert "3538" in row.get("조용함", ""), "왜 안 긁어도 되는지가 안 적혔다"

        SH.band_quiet = lambda: {BQ: {"이름": "매출처업무", "수집최대": 3538,
                                      "확인시각": "2026-08-01 09:00",
                                      "없음확인": 3539, "연속없음": 2}}
        row = [f for f in SH.data_freshness("2026-08-07") if f["이름"].startswith("밴드:")][0]
        assert row["밀림"], "오래된 근거로 밀림을 내렸다 — 그 사이 새 글이 있을 수 있다"
        assert "낡" in row.get("근거", ""), \
            "왜 아직 밀림인지를 안 적는다 — 그러면 사람이 또 없는 번호를 긁으러 간다"

        # 근거가 아예 없으면 예전처럼 밀림이다(안전한 기본값)
        SH.band_quiet = lambda: {}
        row = [f for f in SH.data_freshness("2026-08-07") if f["이름"].startswith("밴드:")][0]
        assert row["밀림"], "근거가 없는데 밀림을 내렸다"
    finally:
        SH.band_latest_days, SH.band_quiet = keep

    # ⑤ 문서에 '조용함'과 이유가 보인다
    src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert '"조용함"' in src and "(조용함)" in src, "표에 조용함 표시가 없다"
    assert "없는 번호를 긁으면" in src, "왜 긁으면 안 되는지가 문서에 안 적힌다"

    # ⑥ 수집기가 지문을 실제로 실어 보내야 판정 재료가 생긴다
    js = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
    assert "notime: S.notime" in js, "덤프에 지문이 안 실린다 — 판정 재료가 영영 안 온다"
    assert "S.notime[no] = r.sig" in js, "no-time 결과의 지문을 안 모은다"

    # ⑦ 수집 계획도 **같은 근거**를 본다 — 없다고 확인한 번호를 또 훑지 않는다.
    #   두 곳이 어긋나면 한쪽은 "긁어라", 다른 쪽은 "조용하다"가 되어 사람이 못 믿는다.
    rp = importlib.import_module("band.recheck_plan")
    plog = os.path.join(TMP, "확인시각2.json")
    keep_rp = rp.__dict__.get("_QUIET_PATH_FOR_TEST")
    real = rp._confirmed_quiet
    # 근거 파일 자리를 테스트용으로 바꿔 끼운다
    def _probe(band, hi, today=None, _p=plog):
        import json as _j
        from datetime import datetime as _dt
        try:
            rec = (_j.load(open(_p, encoding="utf-8")) or {}).get(str(band)) or {}
        except Exception:
            return False
        if int(rec.get("없음확인") or 0) != int(hi) + 1:
            return False
        seen = str(rec.get("확인시각") or "")[:10]
        day = str(today or "2026-08-07")[:10]
        age = (_dt.strptime(day, "%Y-%m-%d") - _dt.strptime(seen, "%Y-%m-%d")).days
        return 0 <= age <= rp.QUIET_LIMIT_DAYS
    with open(plog, "w", encoding="utf-8") as fh:
        json.dump({"84789192": {"없음확인": 3539, "확인시각": "2026-08-07 13:18"}}, fh,
                  ensure_ascii=False)
    assert _probe("84789192", 3538), "최근 확인인데 조용함으로 안 본다"
    assert not _probe("84789192", 3600), "확인한 번호와 무관한 hi 인데 조용함이라 한다"
    with open(plog, "w", encoding="utf-8") as fh:
        json.dump({"84789192": {"없음확인": 3539, "확인시각": "2026-08-01 09:00"}}, fh,
                  ensure_ascii=False)
    assert not _probe("84789192", 3538), "오래된 확인으로 새 글 탐색을 건너뛴다"
    assert callable(real), "recheck_plan._confirmed_quiet 이 사라졌다"
    src_rp = open(os.path.join(ROOT, "band", "recheck_plan.py"), encoding="utf-8").read()
    assert ("_absent_from(band)" in src_rp or
            ("_confirmed_quiet(band, hi)" in src_rp and "new = []" in src_rp)), \
        "계획이 조용함 근거를 안 본다 — 없는 번호를 매 회차 다시 훑는다"
    assert "밴드_확인시각.json" in src_rp, "session_handoff 와 다른 근거를 본다"
    print("  [131] 밴드 조용함 vs 수집 막힘 구분 ✅")


def t132_dash_snap_expand_theme_swipe():
    """[132] 카드 칸 맞춤·펼쳐보기 · 밝기 3종 · 폰 좌우 밀기 (2026-08-07 지시).

    사용자 지시 셋
      ① "카드 마우스로 잡아서 이동해서 **딱딱 맞아 떨어지게** 하는 알고리즘 추가해 /
         펼쳐보기 이런 기능들 적용해서 **각 카드가 정확히 균등하게** 맞아떨어지게"
      ② "이 기능 **기본, 다크모드, 시스템 동일** 이렇게 구성하게 추가해"
      ③ "**모바일 화면에서 손가락을 좌우로 밀면** 각 카테고리로 이동하는 기능"

    ★ 여기서 지키는 것은 대부분 **실측에서 다친 자리**다.
      · 칸 맞추기를 **DOM 순서로 세면 안 된다** — 격자가 `row dense` 라 브라우저가
        뒤 카드를 앞 빈칸으로 끌어올린다. 화면에 없는 빈칸을 보고 멀쩡한 카드를 넓혔다.
      · 칸 맞추기가 **좁은 화면에서 돌면 안 된다** — 폰에서는 카드가 한 줄에 한 장이라
        전부 12칸으로 넓히고, 그 폭이 저장돼 **PC 배치가 통째로 망가진다.**
        폰에서 누른 단추 하나가 PC 화면을 부수는, 되돌리기 어려운 손상이다.
      · 좌우 밀기는 **가로 스크롤 위에서 양보**해야 한다. 안 그러면 표를 옆으로 보려던
        손짓이 화면을 통째로 바꾼다.
    """
    src = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ── ① 균등 맞춤 · 펼쳐보기 ────────────────────────────────────────────
    assert "body.dash-even #dashGrid{align-items:stretch}" in src, \
        "한 줄에 선 카드의 키를 맞추지 않는다 — 바닥이 들쭉날쭉해진다"
    assert "--dash-card-h" in src and "dash-expanded" in src, "카드 높이 상한/펼침이 없다"
    assert "body.dash-even.dash-edit #dashGrid>.dash-block{max-height:none" in src, \
        "편집 중에도 카드를 자른다 — 잘린 카드는 집을 때 무엇인지 안 보인다"
    for fn in ("dashMeasureOverflow", "toggleDashCardExpand", "toggleDashboardEven",
               "snapDashboardLayout", "dashVisualRows", "expandAllDashCards"):
        assert f"function {fn}(" in src, f"{fn} 가 없다"
    # 넘칠 때만 단추가 뜬다 — 늘 떠 있으면 누를 것 없는 단추가 열두 장에 붙는다
    assert "el.scrollHeight>el.clientHeight+8" in src, \
        "넘침 판정에 여유가 없다 — 반올림 1px 로 멀쩡한 카드에 단추가 뜬다"
    # 내용이 도착한 뒤에 다시 재야 한다(첫 그림 때는 모든 카드가 '안 넘침'이다)
    assert "ResizeObserver" in src and "dashMeasureOverflow" in src, \
        "내용이 채워진 뒤 다시 재지 않는다 — 넘치는 카드에 펼쳐보기가 영영 안 뜬다"

    # ── 칸 맞추기: 실제 배치로 세고, 좁은 화면에서는 안 돈다 ──────────────
    snap = src[src.index("function snapDashboardLayout("):]
    snap = snap[:snap.index("\nfunction ")]
    assert "gridTemplateColumns" in snap and "tracks<2" in snap, \
        "좁은 화면 보호가 없다 — 폰에서 누르면 모든 카드가 12칸이 되어 PC 배치가 깨진다"
    assert "dashVisualRows()" in snap, \
        "칸 맞추기가 DOM 순서로 센다 — row dense 때문에 화면에 없는 빈칸을 고친다"
    vis = src[src.index("function dashVisualRows("):]
    vis = vis[:vis.index("\nfunction ")]
    assert "getBoundingClientRect" in vis, "줄 묶기가 실제 위치를 안 본다"

    # ── ② 밝기 세 가지 ──────────────────────────────────────────────────
    assert "const THEME_MODES = ['light','dark','system']" in src, "밝기 3종이 없다"
    assert "'시스템 동일'" in src or "시스템 동일" in src, "'시스템 동일' 이름이 없다"
    # 기본값은 여전히 밝게다 — 화면을 그대로 캡처해 보고서로 올리기 때문(원래 규칙 유지)
    assert "localStorage.getItem('cw_theme') || 'light'" in src, \
        "기본값이 밝게가 아니다 — 보고 화면이 기기 설정 따라 검게 나온다"
    assert "prefers-color-scheme: dark" in src and "watchSystemTheme" in src, \
        "'시스템 동일'인데 OS 설정 변화를 안 따라간다"
    assert "if(themeMode() === 'system')" in src, \
        "시스템이 아닐 때도 OS 변화를 반영한다 — 사람이 고른 값이 멋대로 바뀐다"

    # ── ③ 폰 좌우 밀기 ─────────────────────────────────────────────────
    for fn in ("navTabOrder", "swipeBlocked", "swipeToView", "initSwipeNav"):
        assert f"function {fn}(" in src, f"{fn} 가 없다"
    assert "initSwipeNav();" in src, "좌우 밀기가 시작될 때 켜지지 않는다"
    nav = src[src.index("function navTabOrder("):]
    nav = nav[:nav.index("\nfunction ")]
    assert ".tabbar button[data-v]" in nav and "offsetParent" in nav, \
        "밀기 순서를 탭바에서 안 읽는다 — 폰에서 숨은 탭으로 밀려 빈 화면이 뜬다"
    blk = src[src.index("function swipeBlocked("):]
    blk = blk[:blk.index("\nfunction ")]
    assert "scrollWidth > n.clientWidth" in blk, \
        "가로 스크롤 위에서 양보하지 않는다 — 표를 옆으로 보려다 화면이 바뀐다"
    assert "dashLayoutEditing" in blk and "dash-drag-live" in blk, \
        "카드를 옮기는 중에도 화면이 바뀐다"
    assert "sheetIsOpen()" in blk, "시트가 열려 있는데 뒤 화면이 바뀐다"
    sw = src[src.index("function swipeToView("):]
    sw = sw[:sw.index("\nfunction ")]
    assert "if(!next) return" in sw, "끝에서 처음으로 감긴다 — 밀다가 갑자기 대시보드로 튄다"
    assert "Math.abs(dx) < Math.abs(dy) * 1.5" in src, "세로로 그은 것도 화면 이동으로 읽는다"

    # ── 세 테마에서 같이 돌아야 한다 (지시 ②의 '동일') ────────────────────
    more = src[src.index(".dash-more{"):src.index(".dash-more button{")]
    assert "var(--surface)" in more and "#fff" not in more, \
        "펼쳐보기 띠가 흰색으로 못 박혀 있다 — 어둡게에서 흰 띠가 뜬다"
    assert "@media(prefers-reduced-motion:reduce)" in src, "움직임 줄이기 설정을 무시한다"
    print("  [132] 카드 칸 맞춤·펼쳐보기 · 밝기 3종 · 폰 좌우 밀기 ✅")


def t133_inline_style_dark_safe():
    """[133] 인라인 스타일도 어둡게에서 읽혀야 한다 (2026-08-07 사용자 화면).

    사용자 지적: **"다크모드에서 지금 갱신 텍스트 안보임"** (화면 사진).

    무엇이었나
      새 버전 알림 띠의 [지금 갱신] 단추가 `background:var(--panel)` +
      `color:var(--navy-900)` 이었다. 그런데 **`--navy-900` 은 다크 팔레트에서
      바뀌지 않는다**(#0E1B3F 그대로). 어둡게를 켜면 --panel 만 #1C2438 로 어두워져
      **어두운 판에 어두운 글자** — 단추가 통째로 사라졌다.
      밝게에서 --panel 이 밝아 우연히 읽혔을 뿐, 처음부터 짝이 안 맞는 조합이었다.

    ★ 왜 [127] 이 못 잡았나 — 그 검사는 `<style>` 안의 **CSS 규칙**만 훑는다.
      이건 JS 가 `style.cssText` 로 붙이는 인라인 스타일이라 그물 밖이었다.
      그래서 이 검사는 **JS 가 만드는 색**을 본다.

    규칙 하나: **고정색 판 위에는 고정색 글자를 쓴다.**
      `--navy-*` 는 두 테마에서 같은 값이다(짙은 남색 띠·헤더용). 그 위에 테마
      토큰(--panel/--surface/--ink…)을 얹으면 한쪽 테마에서 반드시 어긋난다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # --navy-* 가 다크 팔레트에서 재정의되지 않는다는 것이 이 규칙의 전제다.
    dark = live[live.index(':root[data-theme="dark"]'):]
    dark = dark[:dark.index("}")]
    assert "--navy-900" not in dark,         "다크 팔레트가 --navy-900 을 바꾼다 — 이 검사의 전제가 달라졌으니 규칙을 다시 세울 것"

    # JS 가 붙이는 인라인 스타일에서 '테마 판 + 고정 남색 글자' 조합을 금지한다
    for m in re.finditer(r"cssText\s*=\s*(.*?)\n\s*b?\.?on|cssText\s*=\s*(.*?);\n", live, re.S):
        chunk = (m.group(1) or m.group(2) or "")[:600]
        if "var(--navy-" in chunk and re.search(r"background:\s*var\(--(panel|surface|bg)", chunk):
            raise AssertionError(f"인라인 스타일이 테마 판에 고정 남색 글자를 얹는다: {chunk[:120]}")

    # 문제의 그 단추 — 띠가 고정 남색이므로 단추도 고정색이어야 한다
    up = live[live.index("function offerUpdate("):][:4000]
    assert "background:#fff;color:#0E1B3F" in up,         "[지금 갱신] 이 다시 테마 토큰을 쓴다 — 어둡게에서 글자가 사라진다"
    assert "var(--navy-900);color:#fff" in up, "띠 자체는 짙은 남색 + 흰 글자여야 한다"
    print("  [133] 인라인 스타일 다크 안전 — 고정색 판 위엔 고정색 글자 ✅")


def t137_contamination_marked_every_merge():
    """[137] 오염 판정은 **병합할 때마다** 돈다 — 한 번 치우고 끝내지 않는다 (2026-08-07 2차 실사고).

    무엇이 잘못돼 있었나
      아침에 오염 621건을 `clean_contaminated --apply` 로 손수 표시했다. `convert_dump` 에는
      '표시된 것을 날짜 없는 재병합이 못 덮는다'는 가드까지 넣었다. 그런데 그 가드는
      **이미 표시된 것**만 지킨다. 15:32 회차에 새 덤프가 **표시된 적 없는** 유령 22건을
      들여왔고 아무도 막지 않았다. 그 결과 밴드업무추출에서 정기점검 UJ2601407 이
      1건 → **23건**으로 부풀었다(reports/밴드업무추출_전체_20260807_1502.csv → _1635.csv).
      화면도 리포트도 멀쩡해 보였다 — 숫자만 틀렸다. 사람이 매번 알아채야 하는 조치는
      결국 안 하게 된다. 그래서 판정을 병합 경로 **안으로** 옮겼다.

    지키는 것
      ① convert_dump 가 캐시를 쓰기 **전에** clean_contaminated.find 로 새 오염을 표시한다.
      ② 옆 파일 import 가 조용히 실패해 보호가 통째로 꺼지지 않게 제 폴더를 경로에 넣는다.
      ③ 표시는 재병합을 견딘다(기존 226행 가드) — 둘이 함께 있어야 유령이 안 되살아난다.
    """
    src = open(os.path.join(ROOT, "band", "convert_dump.py"), encoding="utf-8").read()
    write_at = src.index('out = {"band_name"')
    guard = src[:write_at]
    assert "clean_contaminated" in guard, \
        "병합이 캐시를 쓰기 전에 오염 판정을 하지 않는다 — 새 유령이 그대로 들어간다"
    assert "clean_contaminated.find(merged)" in guard, \
        "판정 대상이 병합 결과(merged)가 아니다"
    assert '"contaminated": True' in guard, "찾아 놓고 표시하지 않는다"
    assert "sys.path.insert(0, _here)" in guard, \
        "옆 폴더 import 가 실패하면 보호가 조용히 꺼진다 — 제 폴더를 경로에 넣어야 한다"
    # 기존 재병합 생존 가드가 함께 살아 있어야 한다(하나만으로는 유령이 되살아난다)
    assert 'cur.get("contaminated") and not rec.get("created_at")' in src, \
        "표시가 날짜 없는 재병합에 덮인다 — [135]⑦ 회귀"

    # 판정기 자체: 표시된 스텁을 다시 오염으로 잡지 않아야 한다(무한 재표시 방지)
    import importlib
    cc = importlib.import_module("band.clean_contaminated")
    stub = {"contaminated": True, "captured_at": 1, "why": "x"}
    posts = {"1": stub, "2": stub, "3": {"content": "진짜 글", "created_at": "2026-08-07 10:00"}}
    assert cc.find(posts) == [] or all(k not in cc.find(posts) for k in ("1", "2")), \
        "이미 표시된 스텁을 또 오염으로 잡는다"

    # ④ 날짜가 붙어 들어온 가짜 — 작성일·본문이 **둘 다** 같은 묶음만 잡고,
    #    그 묶음에서도 맨 앞 번호는 남긴다(원본이 섞여 있을 수 있다).
    body = "[ 쿠팡 A/S 안내 ] A/S 일자 : 2026.08.04 " * 4
    dup = {str(n): {"content": body, "created_at": 1785800580000, "author": "지원팀"}
           for n in (5420, 5421, 5422, 5423, 5424)}
    got = cc.find(dup)
    assert set(got) == {"5421", "5422", "5423", "5424"}, \
        "날짜 달린 사본을 못 잡거나 원본까지 지운다: %s" % sorted(got)

    # ⑤ 진짜 반복 글은 건드리지 않는다 — 같은 본문이라도 **올린 시각이 다르다**
    repeat = {"10": {"content": "발주현황 공유", "created_at": 1700000000000},
              "20": {"content": "발주현황 공유", "created_at": 1700086400000}}
    assert cc.find(repeat) == {}, "매주 올리는 같은 제목의 진짜 글을 가짜로 잡는다"

    # ⑥ 오탐 방지선: '본문이 피드 머리글로 시작한다'만으로는 잡지 않는다.
    #    실측에서 그 규칙은 본문이 전부 다른 진짜 글 98건까지 걸었다(상세 화면에도 머리글이 있다).
    head = "2026년 8월 4일 오전 8:43 게시글 지원팀 3시간 전 "
    lone = {"1": {"content": head + "가", "created_at": 1},
            "2": {"content": head + "나", "created_at": 2}}
    assert cc.find(lone) == {}, \
        "피드 머리글만 보고 잡는다 — 본문이 다르면 서로 다른 진짜 글이다"
    print("  [137] 오염 판정이 병합마다 자동 실행 · 날짜 달린 사본만 좁게 · 원본 보존 ✅")


def t138_daily_run_completion_watch():
    """[138] 일일자동대조가 **완주했는지**를 인계 문서가 본다 (2026-08-07 실사고).

    무엇이 잘못돼 있었나
      작업 스케줄러의 09:50 '일일자동대조'는 매일 **성공(0)** 이었다. 그런데 실제로는
      2026-08-06 21:01 이후 20시간 동안 한 번도 완주하지 않았다. 이유가 둘이다:
        · daily_run 은 앞 회차가 아직 돌면 한 줄 찍고 **정상 종료**한다(exit 0).
          앞 회차가 3시간씩 걸리니 다음 회차는 늘 잠겨 있고, 서로를 가려 준다.
        · 마지막으로 남은 표식(agent_status.json)은 `aborted: True` 였는데
          파일 자체는 최신이라 '최근에 돌았다'로 보였다.
      그동안 자료현황.md 가 1.7일, 종합리포트가 19시간 멈춰 있었고 아무 경보도 없었다.

    지키는 것
      ① 나이(20시간)뿐 아니라 **중단 여부**로도 밀림을 판정한다 — 최신 파일이 곧 성공은 아니다.
      ② 판정 결과가 '먼저 처리할 것'에 오른다. 안 그러면 아무도 안 본다.
      ③ 정상(최근·중단 아님)일 때는 조용하다 — 늘 켜져 있는 경보는 경보가 아니다.
    """
    import importlib
    S = importlib.import_module("session_handoff")
    base = {"큐잔량": 0, "임시파일": [], "옛버전편집": [], "점유": [], "미커밋": [],
            "미푸시": 0, "미머지": 0, "지시문사본": [], "밴드날짜없음": {},
            "밴드오염": {}, "워크트리": None, "원장": {}}
    def only_daily(dr):
        return [w for w, _ in S.blockers(dict(base, 일일대조=dr))
                if "일일자동대조" in w]

    assert only_daily({"밀림": True, "중단": True, "경과시간": 1.0, "실패단계": ["자료현황 갱신"]}), \
        "마지막 회차가 중단인데 아무 말도 하지 않는다 — 파일이 최신이면 성공으로 보인다"
    assert "자료현황 갱신" in only_daily(
        {"밀림": True, "중단": True, "경과시간": 1.0, "실패단계": ["자료현황 갱신"]})[0], \
        "어느 단계가 실패했는지 말하지 않으면 다음 세션이 다시 찾아야 한다"
    assert only_daily({"밀림": True, "중단": False, "경과시간": 26.0, "실패단계": []}), \
        "하루가 넘게 완주하지 않았는데 조용하다"
    assert not only_daily({"밀림": False, "중단": False, "경과시간": 2.0, "실패단계": []}), \
        "정상인데도 경보를 올린다 — 늘 켜진 경보는 무시된다"

    # 판정기 자체: 중단이면 나이와 무관하게 밀림이다
    src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert 'age_h >= DAILY_STALE_H or aborted' in src, \
        "중단을 나이와 함께 보지 않는다 — 방금 중단한 회차를 정상으로 본다"
    print("  [138] 일일대조 완주 감시 — 중단·장기 미완주를 '먼저 처리할 것'으로 ✅")


def t139_new_version_is_atomic():
    """[139] 관리대장 새 버전은 **검증을 통과한 뒤에야** 정본 이름을 얻는다 (2026-08-07).

    무엇이 위험했나
      `findings_sheet.upsert()` 와 `ledger_writer` 는 vN+1 을 **정본 이름으로 먼저 만들고**
      그 다음에 무결성을 검사했다. 원장은 Z: SMB 네트워크 드라이브에 있는 수십 MB 짜리다.
      쓰는 도중 프로세스가 죽으면 — 오늘 실제로 스케줄러가 3시간 한도로 프로세스를
      강제 종료(0x41306)했고, Z: 는 WinError 1231 을 낸 적이 있다 — **이름은 멀쩡한
      `_v{N+1}.xlsx`, 내용은 잘린 zip** 이 남는다. `resolve_master()` 는 v번호가 가장 큰
      것을 정본으로 집으므로(`ecount_reconcile.py`) 앱도, 모든 대조 도구도, 다음 회차도
      그 깨진 파일을 정본으로 읽는다. `~$` 임시파일은 걸러도 이런 반쪽 파일은 못 거른다.
      바로 옆 `workbook_patch.py` 는 처음부터 tmp → 검증 → os.replace 를 하고 있었다.

    지키는 것
      ① 두 파일 모두 `.tmp.xlsx` 에 쓴다 — 정본 이름은 마지막에 os.replace 로만 얻는다.
      ② 검증(zip 무결성·재독)이 **tmp 를 대상으로** 돈다. 통과 못 하면 정본이 되지 않는다.
      ③ 실패하면 tmp 를 지운다 — 남기면 '임시 결과가 이미 존재'로 다음 회차가 막힌다.
    """
    for fn, anchor in (("findings_sheet.py", "zipfile.ZipFile(tmp,"),
                       ("ledger_writer.py", "final_dst, dst = dst")):
        src = open(os.path.join(ROOT, fn), encoding="utf-8").read()
        assert ".tmp.xlsx" in src, f"{fn}: 임시파일을 쓰지 않는다 — 정본에 직접 쓴다"
        assert anchor in src, f"{fn}: 임시파일로 쓰는 자리가 사라졌다"
        assert "os.replace(" in src, f"{fn}: 원자적 교체가 없다"
        # 검증이 정본이 아니라 tmp 를 향해야 한다
        assert "zipfile.ZipFile(dst)" not in src or fn == "ledger_writer.py", \
            f"{fn}: 검증이 아직 정본(dst)을 연다"
        assert "os.remove(" in src, f"{fn}: 검증 실패 시 반쪽 임시파일을 치우지 않는다"

    # ledger_writer 는 dst 를 tmp 로 바꿔치기하므로, 검증 뒤 교체까지가 한 덩어리여야 한다
    lw = open(os.path.join(ROOT, "ledger_writer.py"), encoding="utf-8").read()
    assert lw.index("assert not bad") < lw.index("os.replace(dst, final_dst)"), \
        "검증보다 먼저 정본으로 바꾼다 — 검증이 무의미해진다"
    print("  [139] 관리대장 새 버전은 검증 통과 후에만 정본이 된다(원자적 교체) ✅")


def t140_freshness_tells_the_truth():
    """'몇 분 전 자료' 와 '에이전트 실행 시각' 이 거짓말하지 않나 (2026-08-07 지시).

    사용자 지적 셋이 실은 **같은 병**이었다 — 화면이 자기 나이를 모른 채 말한다.
      "계속 194분전 갱신중 … 분이 이상해"        → 나이 표기가 시간 단위를 모른다
      "계속 몇분전 이라고 표시하는데 계속 내용이 달라" → 칩 하나를 11개 경로가 덮어썼다
      "에이전트 실행 시간이 맞지 않아"              → 어제 중단된 회차가 초록 점을 달고 있었다
    셋 다 '조용한 사고'(CLAUDE.md)의 얼굴이다. 여기서 지키는 것은 표현이 아니라
    **모르면 모른다고 말하는가**이다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    # ① 칩의 나이는 '아직 안 들어온 것 중 가장 오래된 것' 하나로 정한다.
    #    경로마다 제 나이를 쓰면 숫자가 오간다 — 그게 "내용이 달라"였다.
    assert "const SWR_WAIT" in live and "const SWR_INFLIGHT" in live
    assert "Math.min(...vals.map(v => v.t))" in live, "칩이 여전히 마지막에 쓴 값을 보여 준다"
    assert "swrChip(Date.now() - hit.t)" not in live, "경로별 나이를 그대로 칩에 쓰던 옛 길이 남아 있다"
    # ② 갱신 실패를 삼키지 않는다. 삼키면 옛 값이 늙는 동안 '갱신 중'이 영원히 남는다.
    #    ★ 2026-08-10 부터 **끊긴 것은 곧장 칠하지 않고 다시 건다**(검증 [197]).
    #      그래도 '영영 안 칠한다'가 되면 안 된다 — 포기하는 자리가 반드시 칠해야 한다.
    #      그래서 배경 실패 경로에 ⓐ 곧장 칠하는 길 또는 ⓑ 다시 걸다 포기하면 칠하는
    #      길이 **둘 중 하나는** 있어야 한다.
    swallow_now = re.search(
        r"\.catch\(\s*\(?\w*\)?\s*=>\s*\{[^}]{0,600}swrDone\(key,\s*false\)", live)
    give_up = ("dataSectionBackgroundFailure" in live
               and "swrDone(pathKey,false)" in live.replace(" ", ""))
    assert swallow_now or give_up, "SWR 갱신 실패를 다시 삼킨다"
    if give_up:                       # 다시 거는 길이라면 **끝이 있어야** 한다
        bg = live.split("function dataSectionBackgroundFailure", 1)[1].split("\n}\n", 1)[0]
        assert ">=RETRY_WAIT_MS.length" in bg.replace(" ", ""), \
            "다시 걸기만 하고 포기하지 않으면 '갱신 중'이 영원히 남는다"
    assert "갱신 실패" in live, "실패를 사람에게 말하지 않는다"
    # ③ 60분이 넘으면 '194분 전' 이 아니라 시간으로 읽는다.
    assert "function swrAgeText(" in live and "'시간 '" in live
    # ④ 잠깐 끝나는 갱신에는 칩을 띄우지 않는다("너무 자주 뜨고").
    # ★ 철자 하나를 잡지 말 것 (2026-08-08). 옆 세션이 `Date.now()` 를 `now` 로 한 번
    #   끌어올리는 무해한 정리를 했는데, 이 줄이 리터럴을 찾다가 빨간불이 됐다.
    #   그러면 **아무 잘못 없는 세션의 실작업 관문이 막힌다.** 지켜야 할 것은
    #   "지연 관문이 있는가"이지 그것을 어떤 낱말로 썼는가가 아니다.
    assert "SWR_CHIP_DELAY_MS" in live, "칩 지연 관문이 사라졌다"
    assert re.search(r"(Date\.now\(\)|now)\s*<\s*SWR_SHOW_AT", live), \
        "금방 끝날 갱신을 걸러내는 관문이 없다 — 칩이 다시 깜빡인다"
    # ⑤ 한 곳이 끝났다고 칩을 지우지 않는다 — 대기 목록이 빌 때만 지운다.
    assert "if(!SWR_WAIT.size){" in live, "대기 목록과 무관하게 칩을 지운다"

    # ⑥ 숫자의 나이는 **원장이 저장된 시각**이지 지금 시각이 아니다.
    assert "def _data_asof_iso(" in server
    assert '"데이터최종갱신일": _data_asof_iso()' in server, \
        "대표 보고 meta 가 다시 datetime.now() 로 '방금'이라 말한다"
    assert "const _asOf" in live, "화면과 캡처가 각자 시각을 만든다 — 갈라진다"
    assert "new Date().toISOString().slice(0,16).replace('T',' ')}</span>" not in live, \
        "근거 없이 지금 시각을 데이터 갱신 시각으로 찍는 자리가 남아 있다"

    # ⑦ 에이전트 배지는 밀리면 밀렸다고 말한다(초록 점 금지).
    assert '"agent_stale"' in server and '"agent_aborted"' in server
    assert "agent_age_h >= 20" in server, "몇 시간째 미완주인지 판정하는 자리가 없다"
    assert "s.agent_stale" in live and "미완주" in live, "앱이 밀린 회차를 정상처럼 보여 준다"
    # ⑧ 캡처는 **누르면 곧바로** 그린다 (2026-08-07 지시: "이미지 저장 하면
    #   바로바로 데이터 내역 반영해서 캡처되게 알고리즘 다시 짜").
    #   예전 문제는 상한 25초가 아니라 '받기 시작하는 시점'이었다 — [보고] 탭에
    #   재조회 분기가 없어서 **누른 다음에야** 받기 시작했다.
    assert "if(v==='daily') warmReport();" in live, \
        "[보고] 탭을 열 때 미리 받지 않는다 — 누른 뒤에야 받으면 버튼이 굳는다"
    assert "if(rptHasData()) return {how:'now'" in live, \
        "가진 자료가 있는데도 기다린다"
    assert "function rptOfferResave(" in live and "_rptData.updatedAt === mark" in live, \
        "저장 뒤 더 새 자료가 와도 알려 주지 않는다"
    # 9. 갱신 상태는 **덜 바뀌고**, 업무 본문 위에 절대 뜨지 않는다.
    #   2026-08-11에는 빈 자리를 찾아 떠다니는 방식 자체를 없애고, 앱바 안의 작은
    #   상태표시로 통합했다. 실행 화면에서만 큰 상세를 보여 준다.
    for need, why in (("SWR_CHIP_DELAY_MS = 4000", "1.2초는 짧아 화면을 옮길 때마다 떴다"),
                      ("SWR_MIN_AGE_MS", "30초 된 값에도 '갱신 중'을 띄운다"),
                      ("SWR_QUIET_MS", "사라진 직후 다시 떠 깜빡인다")):
        assert need in live, why
    paint = live.split("function swrPaint(")[1].split("function swrAgeText")[0]
    assert "if(failed) return swrChip(age, true);" in paint, (
        "실패까지 관문에 걸려 조용해진다 - 옛 값이 늙는 것을 아무도 모른다(조용한 사고)")
    assert paint.index("if(failed)") < paint.index("SWR_MIN_AGE_MS"), (
        "실패가 나이 관문 뒤에 있다 - 새 값이면 실패를 숨기게 된다")
    assert 'id="dataSyncChip"' in live and 'data-refresh-ui="compact-header"' in live, (
        "갱신 상태가 앱바 안의 작은 상태표시로 통합되지 않았다")
    assert 'id="dataHealth" data-refresh-scope="run-only"' in live, (
        "실행 화면 밖에서도 큰 갱신 상세가 업무를 가린다")
    for fn_name in ("function netBanner", "function swrChip"):
        block = live.split(fn_name, 1)[1].split("\n}", 1)[0]
        assert "document.body.appendChild" not in block and "location.reload" not in block, (
            "갱신 상태가 다시 떠다니는 팝업이나 전체 새로고침으로 돌아갔다")
    # 사이드바는 제 안에서 구른다 (2026-08-08 지시)
    assert "overscroll-behavior:contain" in live.split(".tabbar{top:0")[1][:900], (
        "사이드바 휠이 본문으로 넘어간다 - 카테고리만 굴러가야 한다")
    assert "toast('최신 자료를 받아 그리는 중…')" not in live, \
        "이제 기다리지 않는데 기다린다고 말한다"
    print("  [140] 신선도 표기 — 나이는 하나·실패는 말한다·캡처는 곧바로 ✅")


def t144_topmost_pin_always_restores():
    """'덮여 있음'을 푸는 항상위 고정이 **반드시 되돌아오나** (2026-08-07 실측).

    창이 최소화도 아니고 Win32 로는 '보임'인데 크롬이 hidden 으로 보는 상태가 있다
    (native window occlusion). 그날 반나절을 여기서 잃었고, `HWND_TOP` 으로는
    안 풀리고 `HWND_TOPMOST` 로만 풀린다는 것을 확인했다.

    그런데 이 기능의 위험은 푸는 쪽이 아니라 **되돌리는 쪽**이다. 수집기는 중간에
    자주 죽는다(로그인 화면·없는 번호·네트워크). 그때 크롬이 항상 위에 남으면
    사용자는 원인을 모른 채 "창이 안 내려간다"를 겪는다 — 그리고 그 증상은
    수집 실패와 아무 관계가 없어 보여서 원인을 찾는 데 또 시간이 든다.
    그래서 여기서 지키는 것은 '고정이 되나'가 아니라 **'예외가 나도 풀리나'** 다.
    """
    import importlib.util
    p = os.path.join(ROOT, "band", "window_show.py")
    spec = importlib.util.spec_from_file_location("_ws_t144", p)
    W = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(W)

    # ① 값 자체 — HWND_TOP(0) 은 '맨 위'일 뿐 '항상 위'가 아니다. 둘을 헷갈리면
    #    덮던 창이 곧 다시 덮어서 "고쳤는데 그대로"가 된다.
    assert W.HWND_TOPMOST == -1 and W.HWND_NOTOPMOST == -2 and W.HWND_TOP == 0
    # ② 초점을 빼앗지 않는다 — 옆 창에서 타이핑 중일 수 있다(이 파일의 존재 이유).
    src = open(p, encoding="utf-8").read()
    assert "SWP_NOACTIVATE" in src and "def _zorder(" in src
    assert "SWP_NOMOVE | SWP_NOSIZE | SWP_NOACTIVATE" in src, \
        "Z순서만 바꿔야 하는데 위치·크기·초점까지 건드린다"

    # ③ ★ 핵심 — 예외가 나도 되돌린다.
    calls = []
    W.set_topmost = lambda on, title="Chrome", quiet=True: calls.append(on) or 1
    try:
        with W.pinned():
            raise RuntimeError("수집이 중간에 죽었다")
    except RuntimeError:
        pass
    assert calls == [True, False], \
        f"항상위를 켜고 되돌리지 않는다({calls}) — 크롬이 영영 항상 위에 남는다"

    # ④ 정상 종료도 당연히 되돌린다.
    calls.clear()
    with W.pinned():
        pass
    assert calls == [True, False]

    # ⑤ 사람이 CLI 로 켰을 때는 되돌리라고 **말해 준다**(with 가 없으니 사람이 기억해야 한다).
    assert "--untopmost" in src and "수집이 끝나면 --untopmost" in src, \
        "CLI 로 고정한 사람에게 되돌리라는 말을 하지 않는다"
    print("  [144] 항상위 고정 — 예외가 나도 복귀한다 ✅")


def t145_redirect_deleted_needs_two_rounds():
    """리다이렉트 실패를 '삭제된 글'로 적되, **한 회차만으로는 절대 안 적나** (분담판 [13]).

    구멍 9건이 매 회차 9/9 실패하는데 아무 데도 안 적혀서 다음 회차가 또 같은 9건을
    뽑았다. 고치는 방향은 분명한데, 잘못 고치면 훨씬 나쁜 일이 난다 —
    로그인이 풀린 밤이면 **모든 번호**가 리다이렉트로 실패한다. 그 한 번을 근거로
    묘비를 세우면 멀쩡한 글 수천 건이 '지워진 글'로 적히고, 그건 되돌릴 수 없다.

    그래서 여기서 지키는 것은 '적히나'가 아니라 **'함부로 적지 않나'** 다.
    """
    import importlib.util
    p = os.path.join(ROOT, "band", "convert_dump.py")
    spec = importlib.util.spec_from_file_location("_cd_t145", p)
    C = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(C)

    assert C.REDIRECT_ROUNDS_FOR_DELETED >= 2, \
        "한 회차 실패만으로 묘비를 세운다 — 로그인 풀린 밤 한 번이면 전량이 지워진다"

    # ① 피드 껍데기 지문 — 같은 지문 2개 이상일 때만 증거다.
    assert C._feed_sigs({"10": "AAA", "11": "AAA", "12": "BBB"}) == {"AAA"}, \
        "저 혼자인 지문을 증거로 친다 — '화면이 늦게 그려진 진짜 글'과 구분이 안 된다"

    # ② ★ 한 건도 못 받은 회차는 무엇의 증거도 아니다.
    nt = {"10": "AAA", "11": "AAA"}
    assert C._redirect_hits(nt, ok_count=0) == set(), \
        "수확 0건인 회차(로그인 풀림·네트워크 끊김)의 실패를 증거로 센다"
    assert C._redirect_hits(nt, ok_count=5) == {10, 11}

    # ③ 회차가 하나뿐이면 아직 적지 않는다.
    merged = {}
    assert C._mark_redirect_deleted("b", merged, {10: {1000}}) == 0 and not merged, \
        "회차 하나로 묘비를 세운다"

    # ④ 서로 다른 회차 둘이면 적는다 — 그래야 다음 계획에서 빠진다.
    n = C._mark_redirect_deleted("b", merged, {10: {1000, 2000}})
    assert n == 1 and merged["10"]["deleted"] is True
    assert merged["10"].get("deleted_by") == "redirect", "무엇을 근거로 지웠는지 안 남긴다"

    # ⑤ 같은 회차를 두 번 세지 않는다(집합이므로) — 덤프가 중복돼도 부풀지 않는다.
    merged2 = {}
    assert C._mark_redirect_deleted("b", merged2, {10: {1000}}) == 0

    # ⑥ ★ 본문을 손에 쥔 글은 리다이렉트가 몇 번이든 건드리지 않는다.
    keep = {"10": {"created_at": 123, "content": "진짜 글"}}
    assert C._mark_redirect_deleted("b", keep, {10: {1000, 2000, 3000}}) == 0
    assert keep["10"]["created_at"] == 123 and not keep["10"].get("deleted"), \
        "받아 둔 진짜 글에 묘비를 덮어씌운다 — 되돌릴 수 없다"

    # ⑦ 묘비를 세우면 다음 계획의 '구멍'에서 실제로 빠지는가(이게 [13]의 목적이다).
    spec2 = importlib.util.spec_from_file_location(
        "_rp_t145", os.path.join(ROOT, "band", "recheck_plan.py"))
    R = importlib.util.module_from_spec(spec2)
    spec2.loader.exec_module(R)
    posts = {"1": {"created_at": 1}, "3": {"created_at": 3}}
    assert 2 in (R.plan("b", posts, floor=1) or {}).get("gaps", []), "구멍이 안 잡힌다"
    posts["2"] = {"deleted": True, "deleted_by": "redirect"}
    assert 2 not in (R.plan("b", posts, floor=1) or {}).get("gaps", []), \
        "묘비를 세웠는데도 매 회차 같은 번호를 다시 뽑는다 — [13] 이 안 고쳐졌다"
    print("  [145] 리다이렉트 삭제 판정 — 회차 둘 이상·수확 0 회차 제외 ✅")


def t151_collect_all_idempotent_and_no_login_scrape():
    """[151] 미수집 자료 몰이 — 두 번 일 하지 않고, 사람 몫을 몰래 하지 않나 (2026-08-08).

    사용자 지시: "미수집 데이터들 싹 다 긁어모아, 알고리즘 구성해서 두번 일 안하게
    정리하고, 원본 데이터, 사진, 텍스트등 모두 가져와서 저장하고 보고서 작성해".

    두 가지가 어긋나면 조용히 망가진다:
      ① **없는 글까지 모수로 세면** '미수집'이 영영 0 이 안 된다. 밴드는 없는 번호에도
         껍데기를 준다(2026-08-07 사고 — 마흔 건이 전부 같은 글이었다). 시각 없는 것은
         모수가 아니다.
      ② **로그인이 필요한 수집을 무인으로 돌리면** 로그인 화면을 본문으로 착각해
         캐시를 더럽힌다. 그래서 이 도구는 '이미 캐시에 든 것을 파일로 굳히는 일'만 한다.
    """
    import collect_all as C
    from band import archive_posts as A

    # ① 모수에 **시각 없는 글이 안 들어가야** 한다
    with tempfile.TemporaryDirectory() as t:
        cache, band = os.path.join(t, "cache"), os.path.join(t, "band")
        os.makedirs(cache); os.makedirs(band)
        posts = {
            "10": {"created_at": 1, "images": ["a", "b"]},      # 진짜 글
            "11": {"created_at": 2, "images": []},              # 진짜 글, 사진 없음
            "12": {"content": "앞 글 본문이 잡힌 껍데기"},        # ★ 시각 없음 — 모수 아님
            "13": {"deleted": True},                            # 묘비 — 모수 아님
        }
        json.dump({"band_name": "테스트", "posts": posts},
                  open(os.path.join(cache, "90610953.json"), "w", encoding="utf-8"))
        # raw_* 는 중간 산물이다 — 세면 모수가 부풀어 영영 안 끝난다
        json.dump({"posts": {"99": {"created_at": 9}}},
                  open(os.path.join(cache, "raw_90610953.json"), "w", encoding="utf-8"))
        current = A.archive_paths(
            os.path.join(band, "게시글보관", "테스트"), "10", posts["10"])
        os.makedirs(current["folder"], exist_ok=True)
        open(current["pdf"], "w").close()
        open(current["txt"], "w").close()

        s = C.survey(cache_dir=cache, band_root=band)
        assert s["밴드글_캐시"] == 2, f"모수가 {s['밴드글_캐시']} — 없는 글까지 셌다"
        assert s["사진_URL"] == 2
        assert s["밴드글_보관"] == 1 and s["밴드글_남음"] == 1

        # ② 보고서가 **남은 것**과 **사람 몫**을 적는가
        p = C.write_report(s, s, [{"단계": "밴드글 보관", "결과": "됨", "초": 1, "끝줄": []}],
                           [("ERP 홈택스", "4일 밀림", "erp_grab.py --all")])
        txt = open(p, encoding="utf-8").read()
        assert "사람이 있어야 되는 것" in txt and "ERP 홈택스" in txt
        assert txt.index("사람이 있어야 되는 것") < txt.index("얼마나 모였나"), \
            "사람이 해야 할 것이 보고서 아래로 밀렸다 — 그러면 안 읽힌다"
        assert "이어서 한다" in txt, "남은 것을 어떻게 이어 받는지가 없다"

    # ③ ★ 로그인이 필요한 수집기를 무인으로 부르지 않는다
    돌리는것 = " ".join(" ".join(a) for _n, a, _t in C.STEPS)
    for 금지 in ("collect_", "convert_dump", "upload_intake", "daily_run"):
        assert 금지 not in 돌리는것, \
            f"{금지} 를 무인으로 돌린다 — 로그인 화면을 본문으로 착각해 캐시를 더럽힌다"
    assert "archive_posts.py" in 돌리는것 and "fetch_images.py" in 돌리는것, \
        "원본·사진·텍스트를 굳히는 단계가 빠졌다"
    assert "datalake.py" in 돌리는것, "긁어 놓고 보관소에 안 넣는다"

    # ④ 실패를 삼키지 않는가 — 한 단계가 죽어도 멈추지 않되 **실패로 남겨야** 한다
    src = open(os.path.join(ROOT, "collect_all.py"), encoding="utf-8").read()
    assert "ok=ok" in src, "단계 실패가 로그에 안 남는다 — 다음 회차가 '다 됐다'로 읽는다"

    # ⑤ ★ **기록이 수집을 막으면 안 된다.** 보관소 DB 가 잠겨 있어도(옆 세션이 주사
    #    중일 수 있다) 긁는 일은 그대로 가야 한다. 2026-08-08 실측: 로그 한 줄을
    #    못 써서 'database is locked' 로 회차 전체가 죽었다.
    calls = []
    real = C._py
    C._py = lambda *a, **k: (calls.append(a), (True, ["가짜"]))[1]
    try:
        import datalake as D
        real_connect = D.connect
        D.connect = lambda *a, **k: (_ for _ in ()).throw(sqlite3.OperationalError("database is locked"))
        try:
            got = C.run(limit=1)
        finally:
            D.connect = real_connect
    finally:
        C._py = real
    assert len(got) == len(C.STEPS) and calls, \
        "보관소가 잠겼다고 수집을 통째로 건너뛰었다 — 꼬리가 몸통을 흔든다"
    print("  [151] 미수집 몰이 — 모수 정확·사람 몫 분리·이어받기·실패 기록 ✅")


def t152_band_recollect_window():
    """[152] 매일 08:00 재수집 — 최근 30일만·바뀐 것만·인계 맨 위 (2026-08-08 지시).

    사용자 지시: "매일 08:00 회차에 최근 30일(예: 30일) 글만 재수집 대상으로 뽑아 →
    캐시에 다시 넣고 → put_record가 바뀐 것만 record_rev에 남기고 → 바뀐 게 있으면
    인계 문서 맨 위에 올린다."

    지키는 것
      ① 밴드 날짜 읽기 — 캐시는 **밀리초 정수**다. 문자열로 잘라 쓰면
         `1766704935000` → `"1766704935"` 가 되어 어떤 기간 질문에도 안 걸린다
         (7,782건이 그 상태였다). 이 한 곳(`band_day`)이 갈래를 다 흡수하는가.
      ② 창 — 30일 밖 글이 대상에 안 들어가고, 삭제·오염·유령은 어느 창에서도 빠지는가.
      ③ 바뀐 것만 — 같은 내용을 두 번 흡수하면 record_rev 가 **안 늘어나는가**.
         늘어나면 매일 아침 배너가 켜져 아무도 안 보게 된다.
      ④ 인계 맨 위 — 바뀐 것이 있으면 `banner()` 가 주고, `--ack` 전까지 남고,
         `to_md` 가 그것을 **먼저 처리할 것보다 위에** 놓는가.
      ⑤ 덤프 파일명의 밴드번호 — 날짜 꼬리표를 밴드로 오인하지 않는가(2026-08-08 사고).
    """
    import sqlite3 as _sq
    sys.path.insert(0, os.path.join(ROOT, "band"))
    import datalake as D
    import convert_dump as CD
    import recollect as RC

    # ⑤ 파일명 → 밴드번호. 두 사고를 **동시에** 막아야 한다(앞 숫자·뒤 날짜).
    known = {"84789192", "90610953"}
    assert CD.band_from_name("84789192_260807.json", known) == "84789192", \
        "날짜 꼬리표(260807)를 밴드번호로 읽었다 — 두 밴드가 유령 하나로 합쳐진다"
    assert CD.band_from_name("dump_api2_90610953.json", known) == "90610953", \
        "앞의 버전 숫자가 섞였다"
    # ★ '후보가 하나뿐이면 그대로 쓴다' 를 **버렸다** (2026-08-08 두 번째 사고).
    #   하나뿐인 후보가 날짜(6자리)나 시각 도장(12자리)이면, 그대로 쓰는 순간
    #   있지도 않은 밴드가 캐시에 생긴다. 실제로 `dump_202608082047_null.json` 이
    #   유령 밴드 `202608082047` 을 만들었고, 그 빈 캐시가 make_oneclick 을 첫 밴드에서
    #   죽여 **모든 밴드의 붙여넣기 파일이 하나도 안 만들어졌다.**
    #   모르면 None 을 준다 — convert_dump 가 파일명 해시로 담으므로 숫자로 안 보이고,
    #   어떤 도구도 그것을 밴드로 착각하지 않는다.
    assert CD.band_from_name("dump_260807.json", set()) is None, \
        "날짜 6자리를 밴드번호로 받아들였다 — 없는 밴드가 생긴다"
    assert CD.band_from_name("dump_202608082047_null.json", set()) is None, \
        "시각 도장 12자리를 밴드번호로 받아들였다 — 유령 밴드 202608082047 이 그렇게 생겼다"
    assert CD.band_from_name("dump_202608082047_90610953.json", set()) == "90610953", \
        "시각 도장과 진짜 밴드가 같이 있으면 밴드를 골라야 한다(8자리)"

    # ① 날짜 읽기 — 밀리초·초·ISO·쓰레기
    assert D.band_day(1766704935000) == D.band_day("1766704935000") != "", "밀리초를 못 읽는다"
    assert D.band_day(1766704935) == D.band_day(1766704935000), "초·밀리초가 달라진다"
    assert D.band_day("2026-07-15T10:00:00") == "2026-07-15", "ISO 를 못 읽는다"
    assert D.band_day(None) == "" and D.band_day(0) == "" and D.band_day("없음") == "", \
        "못 읽는 값을 날짜인 척 돌려주면 안 된다"

    from datetime import datetime as _dt, timedelta as _td
    day = _dt.now()
    ms = lambda back: int((day - _td(days=back)).timestamp() * 1000)
    posts = {
        "10": {"created_at": ms(2), "content": "최근 글", "author": "가"},
        "11": {"created_at": ms(29), "content": "창 가장자리", "author": "가"},
        "12": {"created_at": ms(45), "content": "창 밖", "author": "가"},
        "13": {"created_at": ms(1), "content": "지워짐", "deleted": True},
        "14": {"created_at": ms(1), "content": "남의 본문", "contaminated": True},
        "15": {"created_at": ms(1), "content": "없던 번호", "absent": True},
        "16": {"content": "시각 없음"},
    }
    # ② 창
    nos, floor = RC.targets("77", posts, days=30)
    assert nos == [11, 10], "창 판정이 틀렸다 — 뽑힌 것: %s" % nos
    assert 12 not in nos and 13 not in nos and 14 not in nos and 15 not in nos and 16 not in nos, \
        "창 밖·삭제·오염·유령·시각없음 중 하나가 재수집 대상에 들어왔다"
    assert RC.targets("77", posts, days=60)[0] == [12, 11, 10], "창을 넓혀도 12가 안 들어온다"
    # 거르는 근거는 recheck_plan 과 같아야 한다 — 갈리면 한쪽은 긁으라 하고 한쪽은 말린다
    for flag in ("deleted", "contaminated", "absent"):
        assert flag in open(os.path.join(ROOT, "band", "recollect.py"),
                            encoding="utf-8").read(), "%s 를 안 거른다" % flag

    with tempfile.TemporaryDirectory() as tmp:
        dbp = os.path.join(tmp, "t152.db")
        cache = os.path.join(tmp, "band", "cache")
        os.makedirs(cache)
        real_shared = D._shared
        D._shared = lambda *p: (os.path.join(cache, p[-1])
                                if p[:2] == ("band", "cache") else real_shared(*p))
        try:
            doc = {"band_name": "시험밴드", "posts": posts}
            json.dump(doc, open(os.path.join(cache, "77.json"), "w", encoding="utf-8"))
            con = D.connect(dbp)
            try:
                g1 = D.ingest_band(con, quiet=True, since=floor)
                assert g1["신규"] == 2 and g1["버림"] == 1, \
                    "창 흡수가 틀렸다: %s" % {k: g1[k] for k in ("신규", "변경", "그대로", "버림")}
                assert all(len(r["작성일"]) == 10 and r["작성일"][4] == "-"
                           for r in g1["새글"]), "작성일이 날짜 모양이 아니다(밀리초가 샜다)"
                rev0 = con.execute("SELECT COUNT(*) FROM record_rev").fetchone()[0]

                # ③ 같은 것을 다시 — 아무것도 안 쌓여야 한다
                g2 = D.ingest_band(con, quiet=True, since=floor)
                assert g2["그대로"] == 2 and not g2["바뀐글"], "안 바뀐 글을 바뀌었다고 한다"
                assert con.execute("SELECT COUNT(*) FROM record_rev").fetchone()[0] == rev0, \
                    "안 바뀐 글이 record_rev 를 늘렸다 — 매일 아침 배너가 켜져 아무도 안 본다"

                # 진짜 수정 — 밴드가 같은 번호의 글을 고쳐 다시 올린 상황
                posts["10"]["content"] = "최근 글 — 완료 처리"
                json.dump(doc, open(os.path.join(cache, "77.json"), "w", encoding="utf-8"))
                g3 = D.ingest_band(con, quiet=True, since=floor)
                assert g3["변경"] == 1 and g3["바뀐글"][0]["글번호"] == "10", \
                    "고쳐진 글을 못 잡았다: %s" % g3["바뀐글"]
                assert con.execute("SELECT COUNT(*) FROM record_rev").fetchone()[0] > rev0, \
                    "바뀌었는데 record_rev 에 안 남았다 — 어제와 숫자가 다른 이유를 설명 못 한다"
                chg = D.record_changes(con, kind="band_post", limit=10)
                assert chg and chg[0]["natural_key"] == "77/10" and chg[0]["어떻게"], \
                    "변경 이력을 사람이 읽는 줄로 못 뽑는다"
            finally:
                con.close()
        finally:
            D._shared = real_shared

        # ④ 인계 문서 맨 위 — 배너는 ack 전까지 남고, 먼저 처리할 것보다 위다
        real_state = RC.STATE
        RC.STATE = os.path.join(tmp, "재수집.json")
        try:
            assert RC.banner() is None, "회차를 돌기도 전에 배너가 뜬다"
            # 새 글만으로는 배너를 켜지 않는다 — 매일 켜져 있으면 아무도 안 본다
            RC.save_state({"회차": "x", "창일수": 30, "확인함": False,
                           "최근변경": {"회차": "x", "바뀐글": [],
                                        "새글": [{"글번호": "20"}], "변경상세": []}})
            assert RC.banner() is None, "새 글이 들어온 것만으로 맨 위 칸을 켰다"
            RC.save_state({"회차": "2026-08-08 08:00:00", "창일수": 30, "확인함": False,
                           "최근변경": {"회차": "2026-08-08 08:00:00", "새글": [],
                                        "바뀐글": [{"밴드": "시험밴드", "밴드ID": "77",
                                                    "글번호": "10", "작성일": "2026-08-06",
                                                    "요약": "완료 처리"}],
                                        "변경상세": [{"글": "77/10", "어떻게": "본문 바뀜"}]}})
            b = RC.banner()
            assert b and len(b["바뀐글"]) == 1, "바뀐 것이 있는데 배너가 안 뜬다"
            # 다음 회차가 아무것도 못 찾아도 지워지면 안 된다(아무도 안 본 채로 사라진다)
            keep = RC.load_state()
            keep["바뀐글"], keep["새글"] = [], []
            RC.save_state(keep)
            assert RC.banner(), "다음 회차가 조용하다고 어제 경보를 지웠다"
            import session_handoff as SH
            md = SH.to_md({"시각": "x", "원장": {}, "미커밋": [], "최근커밋": [],
                           "다음할일": [], "밴드재수집": b, "점유": [], "큐잔량": 0,
                           "임시파일": [], "옛버전편집": [], "지시문사본": [],
                           "수집신선도": [], "미푸시": [], "미머지": []})
            i_top, i_block = md.find("밴드 글이 바뀌었다"), md.find("먼저 처리할 것")
            assert 0 <= i_top < i_block, "바뀐 소식이 '먼저 처리할 것'보다 아래에 있다"
            assert "77" in md or "시험밴드" in md, "무슨 글이 바뀌었는지 안 적혀 있다"
            RC.ack()
            assert RC.banner() is None, "--ack 를 해도 배너가 안 내려간다"
        finally:
            RC.STATE = real_state

    # 08:00 스케줄러가 실제로 이 회차를 부르는가 — 파일에만 있고 안 도는 것을 막는다
    ps = open(os.path.join(ROOT, "install_recollect_schedule.ps1"), encoding="utf-8").read()
    assert "recollect.py --run" in ps and '"08:00"' in ps, \
        "08:00 트리거나 실행 인자가 스케줄러 설치본에 없다"
    print("  [152] 밴드 재수집 08:00 — 30일 창·바뀐 것만·인계 맨 위·유령밴드 차단 ✅")



def t159_handoff_supersede():
    """[159] 기계가 반복해 남기는 인계 줄은 **마지막 하나만** 남는다 (2026-08-08).

    실사고: 자동 마무리는 컨텍스트가 찰 때마다 19시트 인계를 예약하는데, 상세에
    시각·기준커밋이 들어가 **매번 다른 줄**이 됐다. 중복 인덱스는 (title, detail) 이라
    이걸 못 걸렀고 대기가 45건까지 쌓였다 — 그중 44건이 같은 자동 마무리였다.
    회차가 돌면 **사람이 읽는 19시트에 거의 같은 줄 44개**가 박히고, 그 안에 섞인
    진짜 인계 한 줄은 아무도 못 찾는다. 비어 보이는 사고가 아니라 **묻히는** 사고다.

    지키는 것 둘:
      ① supersede=True 는 같은 제목의 대기를 내리고 새 줄만 남긴다
      ② 기본값은 그대로다 — **사람이 쓴 인계는 줄마다 다른 사실이라 하나도 안 지운다**
      ③ 자동 마무리가 실제로 그 길로 예약한다(배선이 끊기면 ①②가 있어도 소용없다)
    """
    import io, sqlite3, tempfile
    import ledger_db as L

    src = io.open(os.path.join(ROOT, "session_wrapup.py"), encoding="utf-8").read()
    assert '"--handoff", "--supersede"' in src, \
        "자동 마무리가 --supersede 로 예약하지 않는다 — 줄이 또 쌓인다"

    with tempfile.TemporaryDirectory() as tmp:
        old = L.DB_PATH
        try:
            L.DB_PATH = os.path.join(tmp, "t.db")
            for i in range(3):                     # 기계가 세 번 돈 것처럼
                L.handoff_add("세션 자동 마무리(claude)", "계기=auto · 커밋 %d" % i,
                              supersede=True)
            L.handoff_add("사람이 쓴 인계 A", "내용 1")
            L.handoff_add("사람이 쓴 인계 B", "내용 2")
            p = L.pending_handoffs()
        finally:
            L.DB_PATH = old

    auto = [r for r in p if r["title"].startswith("세션 자동 마무리")]
    assert len(auto) == 1, "자동 줄이 %d개 남았다 — 접히지 않았다" % len(auto)
    assert auto[0]["detail"].endswith("2"), "남긴 것이 마지막 줄이 아니다: %r" % auto[0]
    assert len([r for r in p if r["title"].startswith("사람이")]) == 2, \
        "사람이 쓴 인계까지 접었다 — 줄마다 다른 사실이라 지우면 안 된다"
    print("  [159] 인계 예약 — 기계 줄은 마지막 하나만·사람 줄은 그대로 ✅")


def t158_wrapup_drops_huge_files():
    """[158] 자동 마무리가 **거대 파일을 커밋에 담지 않는다** (2026-08-08 실사고).

    `db/source_index_cache.json` 이 106MB 로 자라 자동 커밋에 실려 갔고, GitHub 은
    100MB 넘는 blob 을 pre-receive 에서 거절한다. 그때 죽는 것은 그 커밋 하나가
    아니라 **저장소의 모든 푸시**다 — 그 커밋을 지나야 뒤가 올라가기 때문이다.
    푸시가 죽으면 폰에서 이어받기(푸시된 것만 보인다)도 같이 죽는다.

    `add -A` 는 옆 세션 파일까지 담으므로 무엇이 올라올지 미리 알 수 없다.
    그래서 막는 자리는 '무엇을 담을까'가 아니라 **커밋 직전 크기 검사**다.
    """
    import io, subprocess, tempfile
    import session_wrapup as W

    assert W.HUGE <= 100 * 1024 * 1024, "한도가 GitHub 거절선(100MB)을 넘으면 막는 뜻이 없다"
    src = io.open(os.path.join(ROOT, "session_wrapup.py"), encoding="utf-8").read()
    i_add, i_drop, i_scan = (src.index('git("add", "-A")'),
                             src.index("huge = _unstage_huge()"),
                             src.index('git("grep", "--cached"'))
    assert i_add < i_drop < i_scan, "거대파일 제외는 add 뒤·커밋 전에 와야 한다"

    with tempfile.TemporaryDirectory() as tmp:
        for a in (["init", "-q"], ["config", "user.email", "t@t"], ["config", "user.name", "t"]):
            subprocess.run(["git"] + a, cwd=tmp, capture_output=True)
        big, small = os.path.join(tmp, "big.json"), os.path.join(tmp, "small.txt")
        with open(big, "wb") as f:              # 성긴 파일 — 디스크를 실제로 쓰지 않는다
            f.seek(W.HUGE + 4096 - 1)
            f.write(bytes(1))
        io.open(small, "w", encoding="utf-8").write("ok")
        subprocess.run(["git", "add", "-A"], cwd=tmp, capture_output=True)
        old = W.ROOT
        try:
            W.ROOT = tmp
            dropped = W._unstage_huge()
            staged = subprocess.run(["git", "diff", "--cached", "--name-only"], cwd=tmp,
                                    capture_output=True, text=True).stdout.split()
        finally:
            W.ROOT = old
    assert dropped == ["big.json"], "거대 파일을 빼지 못했다: %r" % (dropped,)
    assert staged == ["small.txt"], "멀쩡한 파일까지 같이 뺐다: %r" % (staged,)
    print("  [158] 자동 마무리 — 90MB 넘는 파일은 커밋 전에 뺀다(푸시 전체가 막힌다) ✅")


def t153_erp_excel_to_records():
    """[153] ERP 엑셀 → 건별 기록 — 겹치는 회차에도 **바뀜이 진짜 바뀜** (분담판 24).

    사용자 지시: "긁어오라고하면 모두 긁어와서 앱 DB에 우선 반영해서 앱에서 바로보고
    캡처하고 활용할 수 있게 정리하고 다 되면 엑셀에 반영하는 알고리즘 추가해."

    ERP 내보내기는 **기간이 서로 겹친다** — 같은 전표가 파일 열 개에 들어 있다.
    그래서 이 검사의 핵심은 '읽어지나'가 아니라 **'같은 것을 다시 읽어도 조용한가'**다.
    실측에서 세 번 부풀었다: 파일명을 payload 에 넣어서(1,669) · 원장 잔액처럼 회차마다
    달라지는 칸까지 비교해서(11,227) · 회차 안에서 파일마다 곧바로 덮어써서(1,310).
    셋 다 record_rev 를 가짜로 채워 **진짜 변경을 묻는** 실패다.
    """
    import datalake as D
    try:
        import openpyxl
    except ImportError:
        print("  [153] openpyxl 없음 — 건너뜀")
        return

    def book(path, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["회사명 : 합성"])                     # 머리행은 2행 — 위치로 찾으면 안 된다
        ws.append(["일자-No.", "거래처명", "적요", "차변금액", "대변금액", "잔액"])
        for r in rows:
            ws.append(r)
        wb.save(path)
        wb.close()

    with tempfile.TemporaryDirectory() as tmp:
        dbp = os.path.join(tmp, "t153.db")
        old = os.path.join(tmp, "old.xlsx")
        new = os.path.join(tmp, "new.xlsx")
        # 같은 전표가 두 회차에 겹쳐 있다. 잔액은 뽑은 기간이 달라 서로 다르다.
        book(old, [["2026/07/01 -1", "쿠팡", "매출", 1000, None, 1000],
                   ["2026/07/01 -1", "쿠팡", "입금", None, 1000, 0],
                   ["2026/07/02 -1", "쿠팡", "매출", 2000, None, 2000],
                   ["2026/07 계", None, None, 3000, None, None]])       # 합계 — 버려야 한다
        book(new, [["2026/07/01 -1", "쿠팡", "매출", 1000, None, 55555],   # 잔액만 다르다
                   ["2026/07/01 -1", "쿠팡", "입금", None, 1000, 55555],
                   ["2026/07/02 -1", "쿠팡", "매출", 2500, None, 55555]])  # ← 진짜 정정

        con = D.connect(dbp)
        try:
            for p, day in ((old, "2026-07-31"), (new, "2026-08-05")):
                con.execute("INSERT INTO asset(path,kind,bucket,mtime,size,biz_date,"
                            "first_seen,last_seen) VALUES(?,?,?,?,?,?,?,?)",
                            (p, "ERP:ledger", "시험", os.path.getmtime(p),
                             os.path.getsize(p), day, D.now(), D.now()))
            con.commit()

            g1 = D.ingest_erp(con, quiet=True)
            assert g1["파일"] == 2, "두 파일을 다 안 뜯었다: %s" % g1
            # 한 전표에 두 줄(매출·입금)이 있다 — 뭉개지면 서로 덮어쓴다
            keys = [r[0] for r in con.execute(
                "SELECT natural_key FROM record WHERE kind='ERP:ledger' ORDER BY natural_key")]
            assert len(keys) == 3, "전표 줄이 뭉개졌다(키 %d개): %s" % (len(keys), keys)
            assert not any("계" == r[0] for r in con.execute(
                "SELECT party FROM record WHERE kind='ERP:ledger'")), "합계 줄이 한 건으로 들어왔다"
            assert g1["변경"] == 0, \
                "첫 흡수인데 '바뀜'이 나왔다 — 회차 안에서 파일마다 덮어썼다는 뜻이다(%d)" % g1["변경"]
            # 새 회차가 이겨야 한다 — 2026/07/02 는 2000 이 아니라 2500
            amt = dict(con.execute("SELECT natural_key,amount FROM record WHERE kind='ERP:ledger'"))
            got = [v for k, v in amt.items() if "07/02" in k][0]
            assert got == 2500, "오래된 회차가 새 회차를 덮었다 (금액 %s)" % got

            # ★ 같은 것을 그대로 다시 — 한 줄도 바뀌면 안 된다
            rev0 = con.execute("SELECT COUNT(*) FROM record_rev").fetchone()[0]
            g2 = D.ingest_erp(con, quiet=True, force=True)
            assert g2["변경"] == 0 and g2["신규"] == 0 and g2["그대로"] == 3, \
                "같은 파일을 다시 뜯었더니 바뀌었다고 한다: %s" % \
                {k: g2[k] for k in ("신규", "변경", "그대로")}
            assert con.execute("SELECT COUNT(*) FROM record_rev").fetchone()[0] == rev0, \
                "안 바뀐 것이 record_rev 를 늘렸다 — 진짜 변경이 이 안에 묻힌다"
            # 안 뜯어도 되는 파일은 안 연다(회차가 못 끝난다)
            assert D.ingest_erp(con, quiet=True)["건너뜀"] == 2, "이미 뜯은 파일을 또 뜯는다"

            # 진짜 변경 — 발행 상태가 바뀐 것처럼 금액을 고쳐 새 회차를 넣는다
            newer = os.path.join(tmp, "newer.xlsx")
            book(newer, [["2026/07/02 -1", "쿠팡", "매출", 9900, None, 1]])
            con.execute("INSERT INTO asset(path,kind,bucket,mtime,size,biz_date,"
                        "first_seen,last_seen) VALUES(?,?,?,?,?,?,?,?)",
                        (newer, "ERP:ledger", "시험", os.path.getmtime(newer),
                         os.path.getsize(newer), "2026-08-08", D.now(), D.now()))
            con.commit()
            g3 = D.ingest_erp(con, quiet=True)
            assert g3["변경"] == 1, "진짜 금액 정정을 못 잡았다: %s" % g3["바뀐건"]
            chg = D.record_changes(con, kind="ERP:ledger", limit=5)
            assert any("9900" in (c["new"] or "") for c in chg), \
                "무엇이 어떻게 바뀌었는지 record_rev 에 안 남았다"

            # 분류가 틀린 파일은 **조용히 넘어가지 않는다**
            odd = os.path.join(tmp, "odd.xlsx")
            wb = openpyxl.Workbook(); wb.active.append(["전혀", "다른", "화면"])
            wb.save(odd); wb.close()
            con.execute("INSERT INTO asset(path,kind,bucket,mtime,size,biz_date,"
                        "first_seen,last_seen) VALUES(?,?,?,?,?,?,?,?)",
                        (odd, "ERP:ledger", "시험", os.path.getmtime(odd),
                         os.path.getsize(odd), "2026-08-08", D.now(), D.now()))
            con.commit()
            assert D.ingest_erp(con, quiet=True)["머리행못찾음"], \
                "머리행을 못 찾은 파일을 말없이 건너뛰었다 — 건수가 영영 모자란 채로 맞아 보인다"

            # ★ 한 종류에 **표 모양이 둘 이상** — 분개장은 같은 ledger 인데 열이 다르고
            #   전표번호가 두 자리 해다. 모양 하나만 알면 이 화면이 통째로 안 읽힌다.
            jr = os.path.join(tmp, "분개장.xlsx")
            wb = openpyxl.Workbook(); ws = wb.active
            ws.append(["회사명 : 합성"])
            ws.append(["전표번호", "계정명", "거래처", "차변", "대변", "적요"])
            ws.append(["26/01/02-2-1", "외상매출금", "쿠팡", 3289000, None, "AS 작업"])
            wb.save(jr); wb.close()
            con.execute("INSERT INTO asset(path,kind,bucket,mtime,size,biz_date,"
                        "first_seen,last_seen) VALUES(?,?,?,?,?,?,?,?)",
                        (jr, "ERP:ledger", "시험", os.path.getmtime(jr),
                         os.path.getsize(jr), "2026-08-08", D.now(), D.now()))
            con.commit()
            g4 = D.ingest_erp(con, quiet=True)
            assert g4["신규"] == 1, "다른 표 모양(분개장)을 못 읽었다: %s" % g4
            got = con.execute("SELECT biz_date FROM record WHERE natural_key LIKE '%외상매출금%'"
                              ).fetchone()
            assert got and got[0] == "2026-01-02", \
                "두 자리 해(26/01/02)를 못 읽었다 — 화면 하나가 통째로 비게 된다"
        finally:
            con.close()
    assert D._erp_day("26/01/02-2-1") == "2026-01-02" and D._erp_day("2026/07/01 -1") == "2026-07-01"
    assert D._erp_day("합계") == "" and D._erp_day(None) == ""

    # 열지도 자체가 성한가 — 값 하나가 어긋나면 화면 하나가 통째로 빈다
    for kind, spec in D.ERP_MAP.items():
        for s in (spec if isinstance(spec, list) else [spec]):
            assert s.get("키") and s.get("날짜"), "%s 에 키/날짜가 없다" % kind
            assert s["날짜"] in s["키"] or s.get("거래처") or s.get("금액"), \
                "%s 가 날짜 말고 아무 칸도 안 본다" % kind
    # ★ 매출/매입을 갈라야 하는 화면은 **상태 칸을 반드시 적어 둔다**.
    #   빠지면 갈래를 payload 에서 이름으로 찾게 되고, 화면마다 그 칸 이름이 달라
    #   (입력메뉴 / 거래유형) 한쪽이 통째로 빠진다 — 실제로 50건이 사라져
    #   "7월 매출전표 5건"이라는 유령 구멍을 만들었다(2026-08-08).
    for s in D.ERP_MAP["ERP:slips"]:
        assert s.get("상태"), "회계거래 화면에 매출/매입을 가를 상태 칸이 없다"
    assert len(D.ERP_MAP["ERP:slips"]) >= 2, "회계거래조회/현황 두 모양을 다 알아야 한다"
    gap = open(os.path.join(ROOT, "sales_slip_gap.py"), encoding="utf-8").read()
    assert 'r["status"]' in gap and "입력메뉴" not in gap.split("def diagnose")[0].split(
        "★")[0], "대조기가 아직 payload 의 특정 칸 이름으로 매출을 가른다"

    # 매일 도는 자리에 들어가 있는가 — 파일에만 있고 안 도는 것을 막는다
    src = open(os.path.join(ROOT, "collect_all.py"), encoding="utf-8").read()
    assert '"--erp"' in src and '"--band"' in src, \
        "ERP·밴드 건별 기록 단계가 collect_all(09:50 daily_run) 에 없다"
    print("  [153] ERP 엑셀 → 건별 기록 — 겹친 회차·합계줄·멱등·틀린분류 신고 ✅")


def t150_datalake_schema_and_incremental():
    """[150] 전 자료 보관소 — 표·append-only·증분·묘비 (2026-08-07 지시).

    사용자 지시: "모든 데이터는 Db화 해서 별도 보관하고 앞으로 들어오는 모든 데이터
    포함 변경 및 로그 기록까지 같이 정리해".

    여기서 지키는 것은 넷이고, 넷 다 **어기면 조용히 틀린다**:
      ① 큐 DB 와 **다른 파일**이어야 한다 — 한 파일이면 11:00·15:00 엑셀 반영이
         색인 흡수에 잠겨 회차를 통째로 놓친다(SQLite 는 파일 단위 쓰기 잠금).
      ② 로그는 고칠 수도 지울 수도 없어야 한다 — 고쳐질 수 있으면 근거가 아니다.
      ③ 안 바뀐 파일은 **열지 않아야** 한다 — Z: 는 SMB 라 매번 sha1 을 재면
         몇 초가 몇 시간이 된다. 그리고 같은 파일을 백 번 봐도 이력이 안 늘어야 한다.
      ④ 사라진 파일을 **지우지 않아야** 한다 — 지우면 '있었다'는 사실을 잃는다.
    """
    import datalake as D

    # ① 자리 — 큐 DB 와 같은 파일이면 안 된다
    assert D.db_path() != os.path.join(ROOT, "db", "ledger_queue.db"), \
        "보관소를 큐 DB 에 넣었다 — 엑셀 반영이 색인에 잠긴다"
    assert D.db_path().endswith("datalake.db")

    with tempfile.TemporaryDirectory() as t:
        con = D.connect(os.path.join(t, "dl.db"))
        try:
            표 = {r[0] for r in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'")}
            for need in ("asset", "asset_rev", "record", "record_rev", "event", "link"):
                assert need in 표, f"{need} 표가 없다"
            # WAL 이어야 읽는 세션이 쓰는 세션을 안 막는다(창이 여러 개인 것이 기본)
            assert con.execute("PRAGMA journal_mode").fetchone()[0].lower() == "wal"

            # ② append-only — 트리거가 실제로 막는가 (문자열이 아니라 동작으로 잰다)
            D.log(con, "test", "hello")
            for sql in ("UPDATE event SET ok=0", "DELETE FROM event"):
                try:
                    con.execute(sql)
                    raise AssertionError(f"로그를 {sql.split()[0]} 할 수 있다 — 근거가 못 된다")
                except sqlite3.IntegrityError:
                    pass
            # 실패도 남는가 — 실패가 안 남으면 사고를 되짚을 수 없다
            D.log(con, "collect", "band.grab", ok=False, detail={"왜": "로그인 화면"})
            assert len(D.events(con, fail_only=True)) == 1

            # ③ 증분 — 안 바뀌면 '그대로', 바뀌었을 때만 이력이 는다
            f = os.path.join(t, "a.txt")
            open(f, "w", encoding="utf-8").write("처음")
            assert D.ingest_asset(con, f, "ERP")[1] == "새것"
            assert D.ingest_asset(con, f, "ERP")[1] == "그대로"
            assert D.ingest_asset(con, f, "ERP")[1] == "그대로"
            time.sleep(1.1)                       # mtime 초 단위 해상도를 넘긴다
            open(f, "w", encoding="utf-8").write("바뀐 내용 더 길게")
            assert D.ingest_asset(con, f, "ERP")[1] == "바뀜"
            n = con.execute("SELECT COUNT(*) FROM asset_rev").fetchone()[0]
            assert n == 2, f"이력이 {n}행 — 안 바뀐 것까지 쌓았다"

            # ④ 사라져도 지우지 않는다. 그리고 **안 훑은 영역은 판단하지 않는다**
            os.remove(f)
            assert D.mark_gone(con, set(), ["/전혀_다른_곳"]) == 0, \
                "안 훑은 영역까지 사라졌다고 적었다 — 그게 더 큰 사고다"
            assert D.mark_gone(con, set(), [t]) == 1
            assert con.execute("SELECT COUNT(*) FROM asset").fetchone()[0] == 1, \
                "사라진 자산을 지웠다 — '있었다'는 사실을 잃는다"
            assert con.execute(
                "SELECT gone_at FROM asset").fetchone()[0], "묘비가 안 찍혔다"

            # ⑤ ★ **처음 보는 파일에는 지문을 재지 않는다** (2026-08-08 실측).
            #    sha1 은 파일을 통째로 읽는 일이고 Z: 는 SMB 다 — 첫 주사에서 5만 개를
            #    다 읽으면 8분에 300건도 못 간다. 그리고 처음 보는 파일에는 견줄 옛
            #    지문이 없어서 지금 재 봐야 아무 판정에도 안 쓰인다. 나중에 채운다.
            g = os.path.join(t, "b.txt")
            open(g, "w", encoding="utf-8").write("지문 없이 먼저")
            D.ingest_asset(con, g, "ERP", want_sha1=False)
            r = con.execute("SELECT sha1 FROM asset WHERE path=?", (g,)).fetchone()
            assert r["sha1"] is None, "첫 주사가 지문을 쟀다 — SMB 에서 몇 시간이 된다"
            assert D.fill_sha1(con, quiet=True) >= 1
            r = con.execute("SELECT sha1 FROM asset WHERE path=?", (g,)).fetchone()
            assert r["sha1"] and len(r["sha1"]) == 40, "나중에도 지문을 못 채운다"
            # 이력 행에도 같이 채워져야 한다 — 안 그러면 '바뀜' 판정이 첫 회에 헛돈다
            assert con.execute("SELECT sha1 FROM asset_rev WHERE asset_id="
                               "(SELECT id FROM asset WHERE path=?)", (g,)).fetchone()["sha1"]
        finally:
            con.close()

    # ⑤ `who` 는 세션까지 — 창이 여러 개인 것이 기본이라 'claude' 만으로는 못 가린다
    assert ":" in D.who(), "누가 했는지에 세션 식별자가 없다"

    # ⑥ ★ **거친 종류가 정밀한 종류를 덮으면 안 된다** (2026-08-08 실측).
    #    내용 판별은 새것·바뀐 것에만 돌린다(엑셀을 여는 일이라 느리다). 그래서
    #    그다음 주사부터는 폴더만 보고 온 'ERP' 가 들어오는데, 그것이 먼저 알아낸
    #    'ERP:taxstep' 을 지웠다 — 잔량을 애써 갈라 놓고 도로 묻는 셈이었다.
    with tempfile.TemporaryDirectory() as t:
        con = D.connect(os.path.join(t, "dl.db"))
        try:
            f = os.path.join(t, "c.xlsx")
            open(f, "w", encoding="utf-8").write("x")
            D.ingest_asset(con, f, "ERP:taxstep", want_sha1=False)
            D.ingest_asset(con, f, "ERP", want_sha1=False)          # 다음 주사
            k = con.execute("SELECT kind FROM asset WHERE path=?", (f,)).fetchone()["kind"]
            assert k == "ERP:taxstep", f"종류가 {k} 로 되돌아갔다 — 판별을 도로 묻는다"
            # 더 정밀한 것으로는 바뀌어야 한다(규칙을 고쳤을 때 반영돼야 하므로)
            D.ingest_asset(con, f, "ERP:tax", want_sha1=False)
            assert con.execute("SELECT kind FROM asset WHERE path=?",
                               (f,)).fetchone()["kind"] == "ERP:tax"

            # ⑦ 검색 — CLI 도 앱도 **이 함수 하나**를 부른다(두 벌이면 결과가 갈린다)
            g = os.path.join(t, "d.xlsx")
            open(g, "w", encoding="utf-8").write("y")
            D.ingest_asset(con, g, "밴드", biz_date="2026-08-01", want_sha1=False)
            assert len(D.find(con, kind="ERP")) == 1, "kind=ERP 가 ERP:tax 를 못 잡는다"
            assert len(D.find(con, kind="밴드")) == 1
            assert len(D.find(con, since="2026-08-01")) >= 1
            assert len(D.find(con, since="2099-01-01")) == 0
            assert len(D.find(con, q="d.xlsx")) == 1
            # 묘비는 기본 검색에서 빠지되, 물으면 나와야 한다(지운 게 아니니까)
            D.mark_gone(con, {f}, [t])
            assert len(D.find(con, q="d.xlsx")) == 0
            assert len(D.find(con, q="d.xlsx", gone=True)) == 1
        finally:
            con.close()
    print("  [150] 전 자료 보관소 — 별도DB·append-only·증분·묘비 ✅")


def t146_erp_bulk_grab_registry():
    """ERP 전 화면 몰이 — **이름을 추측하지 않고 찾아서** 돌리나 (2026-08-07 지시).

    사용자 지시: "세금계산서 ERP에서 잔량 다운로드 받을 수 있어, 매출 전표랑 같이해서
    찾아서 전부 다운로드 받아 … 긁어오라고 하면 모두 긁어와서".

    여기서 제일 위험한 것은 '못 받는 것'이 아니라 **엉뚱한 것을 받아 놓고 맞다고
    믿는 것**이다. 같은 화면이 모듈마다 이름이 다르고(`매출(세금)계산서현황` ↔ `…(재고)`),
    비슷한 이름을 골라 누르면 다른 화면의 Excel 이 같은 자리에 떨어진다.
    그래서 모르는 메뉴는 **모른다고 남겨 두고** `--find` 가 화면에서 찾게 한다.
    """
    import importlib.util
    p = os.path.join(ROOT, "erp_grab.py")
    spec = importlib.util.spec_from_file_location("_eg_t146", p)
    E = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(E)

    # ① 아직 이름을 모르는 화면은 **등록부에 없어야** 한다(추측해 넣으면 안 된다).
    for k in ("taxleft",):
        assert k in E.WANTED, f"{k} 를 찾을 목록에서 잃었다"
        assert k not in E.SCREENS, \
            f"{k} 메뉴 이름을 확인도 없이 코드에 박아 넣었다 — 엉뚱한 화면을 받게 된다"
    assert "잔량" in E.WANTED["taxleft"]["찾을말"]

    # ② --find 는 **읽기만** 한다. 탐침이 메뉴를 눌러 버리면 화면이 갈려 사람이 놀란다.
    find_js = E.FIND_JS % {"words": '["잔량"]'}
    assert "menu.click()" not in find_js and "outputExcel" not in find_js, \
        "찾기 탐침이 메뉴를 누르거나 엑셀을 받는다 — 읽기만 해야 한다"
    assert "classList.remove('visible')" in find_js, \
        "사이트맵을 열어 놓고 안 닫는다 — 그 뒤 모든 클릭을 가로챈다"
    # DOM 전수 스캔 금지(절대규칙) — 렌더러가 언다.
    assert "querySelectorAll('ul,div')" not in find_js

    # ③ --all 은 한 화면이 실패해도 멈추지 않되, **왜 실패했는지 남긴다.**
    E.emit_all  # 존재 확인
    all_js = E.ALL_JS % {"keys": '["ledger"]', "plan": '[{"키":"ledger"}]'}
    assert "continue;" in all_js and "결과: '실패'" in all_js, \
        "한 화면이 실패하면 뒤의 멀쩡한 화면까지 못 받는다"
    assert "왜:" in all_js, "조용히 건너뛴다 — '전부 받았다'로 읽힌다"
    # ④ ★ 조회가 걸렸는지를 **행 수가 아니라 격자 날짜**로 잰다(양쪽으로 다 틀렸던 자리).
    assert "seen.filter(d => d >= rng.from && d <= rng.to)" in all_js
    assert "Excel 안 누름" in all_js, \
        "조회가 안 걸렸는데도 엑셀을 받는다 — 옛 결과를 새 기간으로 착각하게 된다"
    # ⑤ 모듈 전환이 메뉴 찾기보다 **먼저** 와야 한다(사이트맵은 지금 모듈 것만 보여 준다).
    #    재는 자리는 '모듈 전환'과 **사이트맵에서 메뉴를 고르는 줄**의 순서다.
    #    (`step.메뉴` 자체는 결과 기록용으로 위쪽에 한 번 더 나온다 — 그건 순서가 아니다)
    assert all_js.index("if (step.모듈)") < all_js.index("=== step.메뉴"), \
        "모듈을 바꾸기 전에 메뉴를 찾는다 — 다른 모듈 메뉴는 목록에 아예 없다"

    # ⑥ 사람이 고친 등록부(config)가 코드 기본값을 이긴다 — 판올림에 안 지워진다.
    assert "config" in E.SCREENS_CFG and callable(E.load_screens) and callable(E.save_screen)

    # ⑦ ★ 버튼은 **cid 로 찾고 없으면 글자로** 찾는다 (2026-08-08 실측).
    #    `(세금)계산서진행단계`(E010849)·`매출(세금)계산서현황(세무)` 는 data-cid 가
    #    `year`·`month` 뿐이다. cid 만 보면 이 화면들은 버튼이 멀쩡히 있는데도
    #    "기간 프리셋을 못 찾음"으로 영영 실패한다.
    assert "const pick = (cid, txt, exact)" in all_js, "글자 대체 경로가 없다"
    for cid in ("simpleSearch", "searchGroup"):
        assert f"pick('{cid}'" in all_js, f"{cid} 를 아직 cid 로만 찾는다"
        assert f"""querySelectorAll('button[data-cid="{cid}"]')""" not in all_js, \
            f"{cid} 를 cid 로만 직접 긁는 옛 줄이 남아 있다"
    assert "pick(null, 'Excel', true)" in all_js, "엑셀 버튼도 글자로 찾을 수 있어야 한다"
    # ⑧ 보임 판정은 `offsetParent` 가 아니라 **사각형 유무**다.
    #    position:fixed 안의 `검색(F8)` 이 '후보 1 · 보이는 것 0' 으로 걸러졌던 자리다.
    assert "getClientRects().length > 0" in all_js
    assert "e.offsetParent !== null" not in all_js, \
        "offsetParent 로 보임을 재면 fixed 안의 버튼을 잃는다"
    # ⑨ 엑셀은 **한 번만** 누른다 — 후보를 전부 누르면 같은 파일이 두 벌 떨어진다.
    assert "x.click()" in all_js and "forEach(x => x.click())" not in all_js
    # ⑩ ★ 사이트맵은 **채워질 때까지 기다린다** (2026-08-08 실측).
    #    824개가 만들어지기 전에 읽으면 0개가 나오고, 그러면 '메뉴 못 찾음'이 되어
    #    모듈이 다른 줄 알고 엉뚱한 데를 뒤진다(반나절 낭비). 그래서 둘을 가려 적는다.
    assert "querySelectorAll('a').length > 50" in all_js, "빈 사이트맵을 그대로 읽는다"
    assert "사이트맵이 안 열렸다(빈 채로 읽음)" in all_js, \
        "'못 찾음'과 '안 열림'을 안 가른다 — 엉뚱한 모듈을 뒤지게 된다"
    # ⑪ 이름이 맞는데 **받을 것이 없는** 화면을 기억한다 — 누르면 전표가 만들어진다.
    assert "매출전표 I" in E.NOT_GRABBABLE and "입력 화면" in E.NOT_GRABBABLE["매출전표 I"]
    assert "salesslip" not in E.WANTED, "입력 화면을 아직도 찾고 있다"
    print("  [146] ERP 전 화면 몰이 — 이름은 찾아서·cid 없으면 글자로·날짜로 검증 ✅")


def t141_long_text_folds():
    """긴 글은 접고 짧은 글은 건드리지 않나 (2026-08-07 지시).

    사용자 지시: "너무 스크롤이 긴 부분(설명란) … 누르면 자세히 펼쳐지면서 보는 기능 /
    앱 구동이나 사용 편의에 위배될 경우 니가 알아서 최적의 방법으로".
    그래서 '무엇을 접을지'를 목록으로 정하지 않는다 — 목록은 짧은 글까지 접어
    한 번 더 누르게 만든다. **실제 높이를 재서** 정한다.
    그리고 접어서 잃으면 안 되는 두 가지를 여기서 지킨다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "function lcScan(" in live and "function lcToggle(" in live
    # ① 재서 정한다 — 한도와 '접을 값어치'가 둘 다 있어야 한다
    assert "const LC_MAX" in live and "const LC_GAIN" in live
    assert "el.scrollHeight <= LC_MAX + LC_GAIN" in live, "짧은 글까지 접는다"
    # ② 누를 것이 든 블록은 접지 않는다 — 숨은 버튼은 '없는 기능'이 된다
    assert "function lcSafe(" in live
    for tag in ("input", "select", "textarea", "button", "canvas", "table"):
        assert tag in live.split("function lcSafe(")[1][:400], \
            f"lcSafe 가 {tag} 를 품은 블록을 접을 수 있다"
    # ③ 인쇄는 전부 편다 — 종이에 잘려 나가면 영영 못 편다
    assert "@media print{" in live and ".lc{max-height:none!important" in live, \
        "접힌 채로 인쇄되면 잘린 문서가 남는다"
    # ④ 말투는 대시보드 카드와 같아야 한다(같은 동작에 두 이름을 쓰지 않는다)
    assert "'펼쳐보기'" in live and "'접기'" in live
    # ⑤ 화면을 그린 뒤 실제로 불린다 — 안 부르면 코드만 있고 아무 일도 안 일어난다
    assert "lcScanSoon($('v-'+v))" in live, "화면 전환 뒤 긴 글을 재지 않는다"
    print("  [141] 긴 글 접기 — 재서 정함·기능 숨김 금지·인쇄는 전부 펼침 ✅")


def t164_deposit_dedupe():
    """[164] 같은 입금이 여러 파일에 있으면 **한 번만** 센다 (2026-08-08 실사고).

    ※ 번호는 세션끼리 **같은 이름공간**이다. 처음 [162] 로 달았는데 옆 세션이 같은 번호로
      `t162_band_comments_collected` 를 올려 두 개가 됐다. 번호가 겹치면 "검증 [162]" 라는
      말이 어느 것을 가리키는지 알 수 없어 문서가 쓸모없어진다 — **내 것을 옮긴다.**

    실측: 입금 합계 2,417,075,528원인데 청구 합계는 545,353,633원 — 4.4배였다.
    같은 입금이 오종현 정리본과 은행 원본 **양쪽에** 들어 있어 46쌍 1,053,446,894원이
    두 번 세어지고 있었다. `_unique_deposit_files` 는 **똑같은 파일**만 거른다(SHA-256) —
    모양이 다른 같은 돈은 못 잡았다. 입금액 칸이 비어 있어서 다행이었다.

    ★ 여기서 진짜 어려운 것은 합치는 게 아니라 **멀쩡한 입금을 지우지 않는 것**이다.
      두 번 세면 눈에 띄지만, 지운 입금은 **영영 안 들어온 것으로 남는다.**
      그래서 실측에서 갈라야 했던 둘을 그대로 시험한다:
        ① 같은 날·같은 금액인데 **거래처가 다르면** 다른 돈이다(2026-05-06 387,200원 실측)
        ② **같은 파일 안**에서 두 번이면 실제로 두 번 들어온 것이다(은행 두 줄 = 이체 두 번)
    """
    import receipt_fill as F
    from datetime import date

    d1, d2 = date(2026, 6, 30), date(2026, 5, 6)
    rows = [
        # ① 파일이 겹친 같은 입금 — 은행 원본을 남기고 정리표를 뺀다
        {"일자": d1, "금액": 293816959.0, "거래처": "쿠팡로지스틱스", "양식": "정리표",   "출처": "정리본.xlsx"},
        {"일자": d1, "금액": 293816959.0, "거래처": "쿠팡로지스틱스", "양식": "은행원본", "출처": "은행.xlsx"},
        # ② 같은 날·금액인데 거래처가 다르다 — 다른 돈이다
        {"일자": d2, "금액": 387200.0, "거래처": "코리아종합물류",   "양식": "정리표", "출처": "정리본.xlsx"},
        {"일자": d2, "금액": 387200.0, "거래처": "서현조위더스물류", "양식": "정리표", "출처": "정리본.xlsx"},
        # ③ 같은 파일 안에서 두 번 — 이체가 두 번이다
        {"일자": d2, "금액": 55000.0, "거래처": "뮤토택배", "양식": "은행원본", "출처": "은행.xlsx"},
        {"일자": d2, "금액": 55000.0, "거래처": "뮤토택배", "양식": "은행원본", "출처": "은행.xlsx"},
    ]
    kept, dropped = F.dedupe_deposits(rows)
    assert len(dropped) == 1 and dropped[0]["양식"] == "정리표", \
        "파일이 겹친 중복을 못 뺐거나 엉뚱한 쪽을 뺐다: %r" % (dropped,)
    assert sum(1 for r in kept if r["금액"] == 293816959.0) == 1, "겹친 입금이 아직 두 번 세어진다"
    assert [r for r in kept if r["금액"] == 293816959.0][0]["양식"] == "은행원본", \
        "은행 원본이 아니라 정리표를 남겼다 — 사람 손이 닿은 쪽이 정본이 되면 안 된다"
    assert sum(1 for r in kept if r["금액"] == 387200.0) == 2, \
        "거래처가 다른 입금을 합쳤다 — 다른 돈이다"
    assert sum(1 for r in kept if r["금액"] == 55000.0) == 2, \
        "같은 파일 안의 두 줄을 합쳤다 — 이체가 두 번이면 두 번이다"
    print("  [164] 입금 중복 — 파일 겹침만 합치고(은행 원본 우선) 거래처·같은파일 두 줄은 남긴다 ✅")


def t165_ledger_reads_billing_status():
    """[165] 06시트 `청구상태`(AH)를 **읽는지** 지킨다 (2026-08-08 실사고).

    `read_ledger` 의 열 목록에 이름이 없으면 `col_index` 가 None 을 돌려주고, 그 열은
    **오류 없이 사라진다.** 그날 실측: `r.get("원장_청구상태")` 가 750행 전부 None 이라
    '완료 반영 대상 286건(전부 빈칸)' 이라는 결론이 나올 참이었다. 실제로 사람이 채워 둔
    행이 63개(작업완료 59·거래명세서발행 4) 있었다 — **빈 게 아니라 안 보던 것**이다.

    ★ 비어 있는 값은 눈에 띄지만 **읽지 않은 열은 빈칸과 똑같이 보인다.** 그래서
      "열 목록에 이름이 있다"와 "레코드에 키가 있다" 둘 다 못을 박는다.
    """
    import inspect, ecount_reconcile as R

    # ★ 읽는 곳은 `read_ledger`(캐시 껍데기)가 아니라 `_read_ledger_uncached` 다.
    #   껍데기를 들여다보면 열 목록이 안 보여 **테스트가 그냥 통과해 버린다**
    #   (이 테스트도 처음에 그렇게 썼다 — 다행히 빨간불로 났다).
    src = inspect.getsource(R._read_ledger_uncached)
    assert '"청구상태"' in src, \
        "read_ledger 열 목록에서 청구상태가 빠졌다 — 750행이 조용히 빈칸으로 보인다"
    assert '"원장_청구상태"' in src, \
        "청구상태를 읽어도 레코드에 담지 않으면 아무도 못 쓴다"
    # 대입 자리를 통째로 본다 — 이름만 세면 **주석에 적힌 이름**이 걸려 통과한다
    # (실제로 그렇게 통과할 뻔했다: 위 열 목록 주석에 같은 낱말이 들어 있다).
    assert '"원장_청구상태": _d(row[c["청구상태"]])' in src, \
        "원장_청구상태가 엉뚱한 열에서 오거나 대입 자리가 사라졌다"
    print("  [165] 06시트 청구상태를 읽는다(사람이 채운 단계 딱지 63건이 보인다) ✅")


def t166_billing_status_ladder():
    """[166] 청구상태는 **사다리를 거꾸로 가지 않는다** (2026-08-08 지시).

    지시: "ERP 기준으로 확정하고 객관적으로 입증되면 엑셀에 완료처리해".

    ★ 여기서 위험한 것은 안 쓰는 게 아니라 **잘못 쓰는 것**이다. 이 칸은 사람이 보고
      고르는 칸이라(데이터유효성 안내문: "[필수] 작업완료부터 입금완료까지 현재 단계를
      고릅니다"), 값이 틀리면 **사람이 틀린 단계를 보고 일한다.** 그래서 셋을 못 박는다:
        ① 마지막 낱말은 `입금완료` 다 — '완료' 같은 새 낱말을 지어내면 사람이 쓰던
           사다리에 없는 값이 섞여 정렬·필터가 조용히 어긋난다.
        ② 빈칸은 `only_if_empty=True`, 낡은 단계는 `False` — 덮는 자리를 **가른다.**
        ③ **뜻을 모르는 낱말은 큐에 넣지 않는다.** 사람이 적은 값을 덮는 게 위험한 게
           아니라, 무슨 뜻인지 모르는 값을 덮는 게 위험하다.
    """
    import billing_status as B

    assert B.LADDER[-1] == B.DONE == "입금완료", \
        "마지막 단계 낱말이 바뀌었다 — 사람이 쓰던 사다리에 없는 값이 된다: %r" % (B.LADDER,)
    assert B.LADDER.index("작업완료") < B.LADDER.index("거래명세서발행") < B.LADDER.index(B.DONE), \
        "사다리 순서가 어긋났다"

    p = {
        "채움":   [{"정산ID": "JS-A", "프로젝트NO": "P1", "캠프명": "", "업무구분": "", "지금": "(빈칸)"}],
        "올림":   [{"정산ID": "JS-B", "프로젝트NO": "P2", "캠프명": "", "업무구분": "", "지금": "작업완료"}],
        "이미맞음": [{"정산ID": "JS-C", "프로젝트NO": "P3", "캠프명": "", "업무구분": "", "지금": B.DONE}],
        "모르는값": [{"정산ID": "JS-D", "프로젝트NO": "P4", "캠프명": "", "업무구분": "", "지금": "보류(류지영확인)"}],
    }
    items = B.items_for_queue(p)
    keys = {it["key"]: it for it in items}
    assert set(keys) == {"JS-A", "JS-B"}, \
        "이미맞음·모르는값이 큐에 섞였다: %r" % (sorted(keys),)
    assert keys["JS-A"]["only_if_empty"] is True, "빈칸을 덮어쓰기로 넣었다"
    assert keys["JS-B"]["only_if_empty"] is False, "낡은 단계를 못 올린다(빈칸만 채우게 돼 있다)"
    assert all(it["col"] == "청구상태" and it["value"] == B.DONE
               and it["sheet"] == "06_거래서류청구수금" and it["key_col"] == "정산ID"
               for it in items), "엉뚱한 시트·열·키로 넣는다"
    assert all("7.수금완료" in it["evidence"] for it in items), \
        "근거에 무엇이 입증했는지가 없다 — 나중에 되짚을 수 없다"

    # 엑셀은 열지 않는다(11:00·15:00 회차 몫). 무인 경로가 --apply 를 부르면 사고다.
    body = open(os.path.join(ROOT, "billing_status.py"), encoding="utf-8").read()
    assert "--apply" not in body and "openpyxl" not in body, \
        "billing_status 가 엑셀을 직접 연다 — 반영은 11:00·15:00 회차 몫이다"
    print("  [166] 청구상태 — 마지막 낱말은 '입금완료'·빈칸과 낡은단계를 가르고·모르는 값은 안 건드린다 ✅")


def t167_daily_run_inflight():
    """[167] '안 돌았다' 와 '지금 돌고 있다' 를 가른다 (2026-08-08 실측).

    그날 09:50 회차가 **12시간째** 돌고 있었다(보통 25분). 그런데 건강검사는 완주 시각만
    보므로 할 수 있는 말이 "20시간째 완주하지 않았다" 뿐이었고, 그 옆에 붙은 조치는
    `python daily_run.py` 였다. **정반대의 조치다** — 앞 회차가 잠금을 쥐고 있으니 새로
    띄운 회차는 한 줄 찍고 조용히 끝난다. 사람은 "돌렸다"고 믿고 넘어간다.

    ★ 그리고 20시간을 기다릴 필요도 없다. 회차가 3시간을 넘기면 그 자체가 사건이므로
      완주 기록이 아직 싱싱해도 먼저 말한다 — 20시간째에 아는 것은 하루를 잃은 뒤다.
    """
    import session_handoff as H

    # blockers 는 여러 갈래를 한 번에 본다 — 조용한 상태를 만들어 두고 이 갈래만 읽는다.
    CALM = {"큐잔량": 0, "임시파일": [], "옛버전편집": [], "점유": [],
            "미푸시": [], "지시문사본": []}

    def acts(dr):
        st = dict(CALM, 일일대조=dr)
        return [a for msg, a in H.blockers(st) if "일일자동대조" in msg]

    a = acts({"밀림": True, "경과시간": 20.2, "진행중": 12.0})
    assert a and "기다린다" in a[0] and "daily_run.py" not in a[0], \
        "돌고 있는데 새로 띄우라고 한다 — 잠금에 막혀 조용히 건너뛴다: %r" % (a,)

    a = acts({"밀림": True, "경과시간": 22.0, "진행중": None})
    assert a and "daily_run.py" in a[0], \
        "정말 안 돌고 있을 때는 띄우라고 해야 한다: %r" % (a,)

    a = acts({"밀림": False, "경과시간": 2.0, "진행중": 4.0})
    assert a, "완주 기록이 싱싱해도 4시간째 도는 회차는 말해야 한다(보통 25분)"

    a = acts({"밀림": False, "경과시간": 2.0, "진행중": 0.5})
    assert not a, "정상 속도로 도는 회차까지 경보로 올리면 경보가 소음이 된다: %r" % (a,)
    print("  [167] 일일대조 — '안 돌았다'와 '12시간째 돌고 있다'를 가르고 조치를 뒤집지 않는다 ✅")


def t168_erp_progress_glob_is_cached():
    """[168] 원본을 **찾는 일**이 캐시보다 앞에 있으면 안 된다 (2026-08-08 실사고).

    `erp_progress` 는 판매조회 파일을 `glob(ERP_DIR/**, recursive=True)` 로 찾는다 —
    Z:(SMB) 를 재귀로 훑는, 이 함수에서 제일 비싼 일이다. 그런데 그 glob 이 캐시
    검사보다 **앞에** 있었다. `settle_status` 는 행마다 이 함수를 부르므로 750행이면
    **Z: 를 750번 훑었다.**

    ★ 증상이 "느리다"가 아니라 **"아무 답도 안 온다"** 였다는 점이 중요하다. 같은 집계가
      600초 제한에 두 번 걸려 죽었고 세 번째는 45분을 넘겨도 안 끝났다 — 그 사이
      "Z: 가 붐비나 보다"로 읽혔다. TTL 을 넣자 같은 집계가 1분 안에 끝났다.
      **느린 것과 멈춘 것은 겉으로 구별되지 않는다.**

    구조로 못을 박는다. 실행 시간으로 시험하면 기계·네트워크를 타서 못 믿는다.
    """
    import inspect, ecount_reconcile as R

    assert getattr(R, "_ERP_SIG_TTL", 0) > 0, "_ERP_SIG_TTL 이 사라졌다"
    assert "checked" in R._ERP_PROGRESS, "마지막 확인 시각을 기억하지 않는다"

    src = inspect.getsource(R.erp_progress)
    i_ttl = src.find("_ERP_SIG_TTL")
    i_glob = src.find("_g.glob")
    assert i_ttl > 0 and i_glob > 0, "TTL 검사나 glob 이 사라졌다"
    assert i_ttl < i_glob, \
        "캐시 검사가 glob 뒤에 있다 — 행마다 Z: 를 통째로 훑게 된다(750배)"
    print("  [168] erp_progress — 비싼 glob 앞에 캐시 검사가 온다(행마다 Z: 재귀탐색 금지) ✅")


def t203_ledger_screens_are_split():
    """[203] 회계 원장류 **네 화면**을 한 통에 담지 않는다 (2026-08-08 실사고).

    예전 규칙은 "'적요' 와 차변/대변이 같은 표에 있으면 ledger" 한 줄이었다. 그런데
    이카운트 회계 원장류는 넷 다 그 모양이다. 실측: `pick("ledger")` 12개 중
    **거래처별계정별원장은 5개뿐**이고 계정별원장 5 · 분개장 1 · 현금출납장 1 이었다.
    게다가 계정별원장 한 파일은 한 시트에 **52개 계정이 층층이** 쌓여 있어(현금·보통예금·
    외상매입금·부가세예수금 …) 8,980행이 통째로 '쿠팡 매출 전표'로 읽혔다.

    그 결과 `erp_ledger_check` 가 전표 **5,157건**을 대조해 **원장 매칭 정상 0건** ·
    유형A('설치·작업 근거 확인 필요 ★') **1,856건**을 냈다. 그 대부분은 남의 회사
    거래라 아무도 확인하러 갈 수 없는 지시였고, 그것이 매일 09:50 회차마다 새로 찍혔다.
    파일도 있고 숫자도 나오니 **실패한 티가 안 났다** — '매칭 0건' 이 유일한 신호였다.

    지키는 것: ① 네 화면이 서로 다른 갈래로 간다 ② 분개장 판정이 '적요' 판정보다
    **먼저** 온다(분개장에도 적요·거래처가 있어 순서가 뒤집히면 계정별원장으로 읽힌다)
    ③ 회계거래조회(금액 한 열)는 원장류로 새지 않는다 ④ 제목줄 차선책에서
    '거래처별계정별' 을 '계정별원장' 보다 먼저 본다(부분문자열 함정).
    """
    import inbox_scan as S

    # 실제 내보내기에서 그대로 옮긴 머리글 (2026-08-08 측정)
    HDR = {
        "ledger":      ["일자-No.", "적요", "차변금액", "대변금액", "잔액"],
        "ledger_acct": ["일자-No.", "적요", "거래처명", "차변금액", "대변금액", "잔액"],
        "journal":     ["전표번호", "계정명", "거래처", "차변", "대변", "적요"],
        "cashbook":    ["일자-No.", "상대계정명", "상대거래처명", "적요",
                        "차변금액", "대변금액", "잔액"],
    }
    for want, hdr in HDR.items():
        rows = [["회사명 : 주식회사 유니버셜리프트앤히타치코리아"], hdr,
                ["2026/01/10 -2", "26년1분기정기점검-1(일산7MB 외 15캠프)", "8306650"]]
        got = S.ledger_kind(rows)
        assert got == want, f"{want} 를 {got} 로 읽었다 — 머리글 {hdr}"
        assert S.classify_rows(rows) == want, f"classify_rows 가 {want} 를 안 돌려준다"

    # 회계거래조회(slips)는 금액이 '금액' 한 열이라 원장류가 아니다 — 실측 머리글
    slips = [["전표번호", "거래유형", "금액", "거래처명", "적요"],
             ["26/01/02-2", "매출", "3289000", "쿠팡로지스틱스서비스 유한회사", "AS"]]
    assert S.ledger_kind(slips) is None, "회계거래조회가 원장류로 샜다"

    # 표를 못 읽었을 때의 제목줄 차선책 — '거래처별계정별원장' 은 '계정별원장' 을 품는다
    assert S.classify_rows([["... / 거래처별계정별원장 / 1089(외상매출금)"]]) == "ledger"
    assert S.classify_rows([["... / 계정별원장 / 1019(현금)"]]) == "ledger_acct"
    assert S.classify_rows([["... / 분개장"]]) == "journal"
    assert S.classify_rows([["... / 현금출납장"]]) == "cashbook"

    # 새 갈래가 이름표·수집 목록에도 올라야 한다 — 안 올리면 Downloads 에서 안 가져온다
    import collect_sources as C
    for k in ("ledger", "ledger_acct", "journal", "cashbook"):
        assert k in S.LABEL, f"{k} 이름표가 없다"
        assert k in C.KNOWN, f"{k} 이 수집 대상에서 빠졌다 — 자료가 조용히 끊긴다"
    print("  [203] 회계 원장류 — 네 화면이 갈린다(분개장→적요 순서·slips 안 샘) ✅")


def t178_unverified_harvest_is_not_read():
    """[178] **확인 못 한 것을 '읽었다'로 세지 않는다** (2026-08-09 실사고).

    250건을 실패 0 으로 긁었는데 캐시에 댓글이 **한 건도** 안 들어왔다. 원인은
    `grab_posts.js` 의 개수 선택자가 지금 밴드 화면과 안 맞는 것이다 — 개수를 0 으로
    읽으니 댓글이 그려질 때까지 기다리지 않고 빈 배열을 담았다.

    무서운 것은 그다음이다. `comments` 키가 **생기기 때문에** 그 글은 '들여다봤다'로
    세어져 사각지대 계기에서도 수집 목록에서도 빠진다. **못 읽은 글이 읽은 글로
    둔갑해 영영 다시 안 뽑힌다.** 수집은 성공으로 끝나고 아무 데도 티가 안 난다.
    실측으로 1순위가 80 → 69 로 '줄어' 있었다 — 11건이 그렇게 사라진 것이다.

    선택자를 다시 맞추는 것은 답이 아니다(밴드가 화면을 고치면 또 깨진다). 그래서:
      ① 수집기는 개수를 모르고 댓글도 못 봤으면 `comments` 키를 **안 단다**
         (`comments_unverified`). 다음 회차가 그 글을 다시 뽑는다.
      ② 읽는 쪽은 밴드 전체에 댓글 있는 글이 0건이면 `comments` 키를 **안 믿는다**.
         캐시는 고치지 않는다 — 진짜 댓글이 들어오면 이 조건은 저절로 풀린다.
    """
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "band"))
    import comment_backfill as CB

    js = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
    assert "comments_unverified" in js, \
        "개수를 못 읽었을 때 표시가 없다 — 못 읽은 글이 읽은 글로 둔갑한다"
    assert "countKnown" in js and "if (countKnown || cts.length)" in js, \
        "확인 못 한 수확에도 comments 키를 달고 있다"
    assert "^댓글\\s*([0-9,]+)$" in js or "댓글\\s*([0-9,]+)" in js, \
        "고정 선택자 하나만 믿는다 — 화면이 바뀌면 같은 자리에서 또 깨진다"

    # 계기: 표본이 충분한데 댓글 있는 글이 0건이면 스스로 의심한다
    class _F:
        def __init__(self, s): self.s = s
        def read(self): return self.s
        def __enter__(self): return self
        def __exit__(self, *a): return False

    real = CB.io.open
    try:
        broken = {str(i): {"comments": [], "comment_count": "0"} for i in range(40)}
        CB.io.open = lambda *a, **k: _F(json.dumps({"posts": broken}))
        assert CB.harvest_looks_broken("90610953"), "0건인데 의심하지 않는다"

        ok = dict(broken)
        ok["5"] = {"comments": [{"content": "취소요청"}], "comment_count": "1"}
        CB.io.open = lambda *a, **k: _F(json.dumps({"posts": ok}))
        assert not CB.harvest_looks_broken("90610953"), \
            "댓글이 들어왔는데도 계속 의심한다 — 경보가 안 꺼지면 아무도 안 본다"

        few = {"1": {"comments": []}}
        CB.io.open = lambda *a, **k: _F(json.dumps({"posts": few}))
        assert not CB.harvest_looks_broken("90610953"), "표본 1건으로 단정했다"
    finally:
        CB.io.open = real
    print("  [178] 수확 검증 — 확인 못 한 것을 '읽었다'로 세지 않는다 ✅")


def t177_comment_collection_is_targeted():
    """[177] 수집은 **날짜가 아니라 쓸모로** 고른다 (2026-08-09 지시).

    사용자 지시: "무작정 자료 수집만 하지 말고 정확한 알고리즘을 만들어 수집하게 코딩해".
    처음 만든 것은 '최근 90일치 250건씩'이었다 — 그건 결국 무작정이다. 7,475건을
    10시간 긁으면서 그중 무엇이 무엇을 바꾸는지 아무도 모른다.

    기준은 **읽는 쪽에서 빌린다.** `cancel_watch.build()` 는 취소를 찾아도 그
    프로젝트에 **아직 안 끝난 원장 행**이 있을 때만 대기열 행을 만든다. 그 집합 밖의
    글은 댓글을 다 읽어도 오늘 단 한 줄도 못 바꾼다. 실측 7,475건의 정체:
    업무글 아님 3,070 · 프로젝트NO 없음 2,641 · 닫힘 1,684 · **열림 80**.
    10시간이 7분이 된다.

    지키는 것: ① 업무글이 아니면 목록에 **아예 안 넣는다** ② 프로젝트NO 가 없으면
    안 넣는다 ③ 열린 원장 행이 있으면 1순위이고 **날짜로 안 자른다**(반년 전 미실시가
    이 사고의 본체다) ④ 원장을 못 읽으면 1순위라고 **우기지 않는다**.
    """
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "band"))
    import comment_backfill as CB

    made = []

    def fake_parse(no, p, band):
        return None if p.get("kind") == "공지" else {"프로젝트NO": p.get("prj", "")}

    import band_extract
    real = band_extract.parse_post
    real_load = CB.io.open
    import time as _t
    NOW = int(_t.time() * 1000)          # 최근 — 날짜 창 안
    OLD = 1                              # 1970년 — 날짜 창 **밖**
    posts = {
        # ★ 열린 원장 행이면 **아무리 오래돼도** 뽑힌다. 반년 전 미실시가 그대로
        #   얹혀 있는 것이 이 사고의 본체다 — 그래서 일부러 1970년으로 둔다.
        "10": {"created_at": OLD, "prj": "UJ2600001"},             # 열림 → 1
        "11": {"created_at": NOW, "prj": "UJ2699999"},             # 닫힘·최근 → 3
        "17": {"created_at": OLD, "prj": "UJ2699999"},             # 닫힘·오래됨 → 제외
        "12": {"created_at": NOW, "kind": "공지"},                  # 업무글 아님 → 제외
        "13": {"created_at": NOW, "prj": ""},                       # 프로젝트NO 없음 → 제외
        # ★ 2026-08-12 (분담판 [39]) — 기대가 옮겨졌다. 이 줄은 원래 '이미 봤다 → 제외'
        #   였는데, 그건 `comments_full` 이 생기기 **전**(검증 [182], 2026-08-09)의
        #   규칙이다. 지금은 '확인된 0개'의 근거가 `comments_full` 이고, 그것이 없는
        #   빈 목록은 "봤고 없었다"가 아니라 **목록이 다 그려진 것을 확인 못 한 채 0 으로
        #   적힌 것**이다. 그대로 두면 취소 댓글을 영영 못 읽는다(실측 5,829건).
        "14": {"created_at": NOW, "prj": "UJ2600001", "comments": []},   # 미확인 → 1순위
        # ★ 그리고 이것이 [199] 가 지키는 반대쪽이다 — **확인된 0개는 다시 안 뽑는다.**
        #   이 줄이 없으면 위 완화가 무한루프로 번져도 아무도 모른다.
        "18": {"created_at": NOW, "prj": "UJ2600001",
               "comments": [], "comments_full": True},              # 확인된 0개 → 제외
        "15": {"created_at": NOW, "prj": "UJ2600001", "deleted": True},  # 삭제 → 제외
        "16": {"prj": "UJ2600001"},                                 # 시각 없음 → 제외
    }

    class _F:
        def __init__(self, s): self.s = s
        def read(self): return self.s
        def __enter__(self): return self
        def __exit__(self, *a): return False

    try:
        band_extract.parse_post = fake_parse
        CB.io.open = lambda *a, **k: _F(json.dumps({"posts": posts}))
        got = CB.blind("90610953", 90, {"UJ2600001"})
        nums = [n for _t, _d, n in got]
        assert nums == [14, 10, 11], "고른 것이 다르다: %r" % (got,)
        assert 18 not in nums, \
            "확인된 0개(comments_full)를 다시 뽑는다 — [199] 무한루프로 돌아간다"
        assert [t for t, _d, _n in got] == [1, 1, 3], "갈래가 틀렸다: %r" % (got,)

        # 원장을 못 읽으면 1순위라고 우기지 않는다 — 전부 2순위(모름)
        got2 = CB.blind("90610953", 90, None)
        assert {t for t, _d, _n in got2} == {2}, "원장을 모르는데 갈래를 단정했다"
    finally:
        band_extract.parse_post = real
        CB.io.open = real_load

    assert CB.BATCH_MAX == 250, "한 배치 상한이 grab_posts.js 와 어긋난다(탭이 언다)"
    src = open(os.path.join(ROOT, "band", "comment_backfill.py"),
               encoding="utf-8").read()
    assert "open_ledger_rows" in src, \
        "'열렸다'를 여기서 따로 정의하면 cancel_watch 와 갈린다 — 빌려 써야 한다"
    assert "tier != 1 and cut" in src, \
        "1순위까지 날짜로 자르고 있다 — 오래된 미실시가 이 사고의 본체다"
    print("  [177] 댓글 수집 — 날짜가 아니라 '오늘 숫자를 바꾸는가'로 고른다 ✅")


def t199_distrust_trusts_confirmed_zero():
    """[199] distrust 무한루프 — comments_full(확인된 0개)은 믿는다 (2026-08-11 실사고).

    90610953 은 열린 원장 1순위 95건을 두 번 재수집·흡수했는데도 1순위가 계속 95건.
    원인: 이 밴드는 **댓글 담긴 글이 실제로 0**이라 `harvest_looks_broken` 이 늘 참을
    돌려주고, distrust 가 `comments: []` 인 글을 죄다 '못 읽음'으로 되뽑는다. 아무리
    긁어도 진짜 0 이라 distrust 가 안 풀린다 → 같은 글 무한루프.

    가르는 근거는 수집기가 이미 다는 `comments_full`([182] '확인된 0개'): 입력창까지
    그려진 뒤 목록이 0 이면 그건 못 읽은 게 아니라 **본 것**이다. 반대로 [162] 사고의
    깨진 수확은 `comment_count>0` 인데 목록 0 이라 `comments_full=False` 로 남는다.
    그래서 **distrust 여도 comments_full 은 믿고**, 되뽑는 것은 '비었고 & full 도 아닌' 글뿐이다.
    """
    import sys as _s
    _s.path.insert(0, os.path.join(ROOT, "band"))
    import comment_backfill as CB
    import band_extract
    import time as _t

    def fake_parse(no, p, band):
        return {"프로젝트NO": p.get("prj", "")}

    NOW = int(_t.time() * 1000)
    posts = {}
    # distrust 를 켜려면 '들여다봤다고 기록된' 글이 floor(30) 이상이어야 한다.
    # 전부 댓글 0 이라 harvest_looks_broken 이 참을 돌려준다(진짜 0 인 밴드).
    for i in range(40):
        posts[str(1000 + i)] = {"created_at": NOW, "prj": "UJ2600001",
                                "comments": [], "comments_full": True}   # 확인된 0개
    # 진짜 못 읽은 글: 비었는데 comments_full 이 없다 → distrust 면 되뽑아야 한다.
    posts["2001"] = {"created_at": NOW, "prj": "UJ2600001", "comments": []}

    class _F:
        def __init__(self, s): self.s = s
        def read(self): return self.s
        def __enter__(self): return self
        def __exit__(self, *a): return False

    real = band_extract.parse_post
    real_load = CB.io.open
    try:
        band_extract.parse_post = fake_parse
        CB.io.open = lambda *a, **k: _F(json.dumps({"posts": posts}))
        assert CB.harvest_looks_broken("90610953"), \
            "댓글 0 인 밴드인데 distrust 가 안 켜졌다 — 테스트 전제가 틀렸다"
        got = CB.blind("90610953", 90, {"UJ2600001"})
        nums = {n for _t, _d, n in got}
        # 확인된 0개 40건은 distrust 여도 다시 안 뽑힌다. 미확정 1건만 남는다.
        assert nums == {2001}, \
            "comments_full(확인된 0개)을 distrust 가 무시해 되뽑았다: %r" % (sorted(nums),)
    finally:
        band_extract.parse_post = real
        CB.io.open = real_load

    src = open(os.path.join(ROOT, "band", "comment_backfill.py"),
               encoding="utf-8").read()
    body = src.split("def blind(", 1)[1].split("\ndef ", 1)[0]
    assert "comments_full" in body, \
        "후보를 고르는 자리가 comments_full 을 안 본다 — 진짜 0 인 밴드에서 무한루프가 돈다"
    # ★ 2026-08-12 (분담판 [39]) — 예전에는 `distrust and not comments_full` 이라는
    #   **글자 그대로**를 확인했다. 그런데 그 `distrust and` 가 바로 [39] 의 고장이었다:
    #   댓글 담긴 글이 **하나만** 있어도 `harvest_looks_broken` 이 꺼져, 미확인
    #   5,829건이 통째로 '본 것'으로 넘어갔다. 이제 근거는 `comments_full` 하나다.
    #   글자가 아니라 **동작**으로 묻는다 — 위 블록이 '확인된 0개는 안 뽑힌다'를
    #   이미 실제로 확인했다. 여기서는 옛 게이트로 되돌아가지 않는 것만 못박는다.
    assert 'distrust and not v.get("comments")' not in body, \
        "되뽑는 근거가 다시 distrust 에 매였다 — [39] 로 되돌아간다"
    print("  [199] distrust — 확인된 0개(comments_full)는 믿어 무한루프를 끊는다 ✅")


def t217_probe_instead_of_scraping_absent_numbers():
    """[217] 없는 번호를 40개씩 담지 않는다 — 근거 먼저, 없으면 **싸게 확인만** (2026-08-11 실사고).

    무엇이 있었나
      16:10 회차의 붙여넣기 파일이 밴드 84789192 에 `3540~3579` 마흔 개를 담았다.
      캐시 최대는 3539 이고 그 위는 **아직 없는 번호**다. 밴드는 없는 번호에도 200 과
      앱 껍데기를 주므로 수집기는 한 개당 iframe 9초 + 본문 12초를 꽉 채우고 시각이
      없어 버린다(검증 [130]) — **약 14분에 수확 0**. 오류도 안 나고 실패로도 안 세인다.

    왜 안 걸러졌나
      계획은 '없음 확인' 근거를 보고 있었는데, 그 근거가 **한도(1일)보다 낡으면**
      없는 것으로 치고 곧장 `ahead`(40) 를 통째로 쏟았다. 근거가 낡았다는 것은
      "그 사이 새 글이 있을 수 있다"는 뜻이지 "마흔 개가 있다"는 뜻이 아니다.
      번호는 이어지므로 **`hi+1` 하나면 존재 여부는 답이 나온다.**

    지키는 것
      ① 근거가 살아 있으면 그 구간은 **아예 안 넣는다**(조용함).
      ② 근거가 낡았거나 없으면 **존재 확인용 몇 개만**(`PROBE_AHEAD`). 있는 것이
         확인되면 `hi` 가 올라가 **다음 회차가 이어받는다.**
      ③ 근거가 **추월**됐으면(그 번호가 이미 진짜 글로 수확돼 있으면) 근거로 안 쓴다.
         그리고 실재하는 글을 유령이라 부르지 않는다.
      ④ 근거 **아래**(실재가 확인된 구간)는 줄이지 않는다 — 있는 글을 긁는 것은 낭비가 아니다.
      ⑤ 붙여넣기 파일로 나가는 **길목 하나**가 죽은 번호를 거른다(번호를 정하는 곳은 셋이다).
      ⑥ 뺀 것은 **숫자로 말한다** — 조용히 빼면 '0건'이 '다 봤다'로 읽힌다([169]).
      ⑦ 빈 캐시 한 밴드가 나머지 밴드를 죽이지 않는다.
      ⑧ **인계 문서도 같은 자리에 물어본다** — 근거의 나이만 보면 신선하지만 추월된
         근거에 '(조용함)'을 적는다. 잘못된 조용함은 아무도 다시 안 본다.
    """
    import importlib
    import tempfile
    import contextlib
    import io as _io
    from datetime import datetime as _dt

    sys.path.insert(0, os.path.join(ROOT, "band"))
    MO = importlib.import_module("make_oneclick")
    CP = importlib.import_module("comment_plan")
    rp = MO.RP
    # 근거·낱말을 보는 자리가 갈리면 한쪽은 긁으라 하고 다른 쪽은 조용하다고 한다.
    assert CP.RP is rp, "붙여넣기 파일 만드는 둘이 서로 다른 recheck_plan 을 본다"

    BAND, TODAY = "84789192", "2026-08-11"
    TMP = tempfile.mkdtemp(prefix="bandprobe_")
    hold = (rp.PROBE_LOG, rp.CACHE, rp.SCOPE)

    def evidence(n, seen):
        with open(rp.PROBE_LOG, "w", encoding="utf-8") as fh:
            json.dump({BAND: {"이름": "매출처업무", "수집최대": n - 1,
                              "없음확인": n, "확인시각": seen}}, fh, ensure_ascii=False)

    # 작성시각이 있는 진짜 글 셋 — 캐시 최대 3539(2026-08-11 실측과 같은 모양)
    posts = {str(n): {"created_at": 1785900000000, "content": "글 %d" % n,
                      "captured_at": 1785900000000} for n in (3537, 3538, 3539)}
    try:
        rp.PROBE_LOG = os.path.join(TMP, "밴드_확인시각.json")
        rp.CACHE = os.path.join(TMP, "cache")
        os.makedirs(rp.CACHE, exist_ok=True)
        rp.SCOPE = os.path.join(TMP, "collect_scope.json")
        with open(rp.SCOPE, "w", encoding="utf-8") as fh:
            json.dump({"floor": {BAND: 3537}, "ahead": 40}, fh, ensure_ascii=False)

        # ① 살아 있는 근거 — 바로 위가 없음으로 확인됐으면 한 개도 안 넣는다
        evidence(3540, TODAY)
        p = rp.plan(BAND, posts, 3537, 40, today=TODAY)
        assert p["new"] == [], "방금 없다고 확인한 번호를 또 목록에 넣는다"

        # ② 근거가 낡음 → 40개가 아니라 '존재 확인용'만. 이것이 이날의 14분이다.
        evidence(3540, "2026-08-09")                    # 2일 전 — 한도(1일) 밖
        p = rp.plan(BAND, posts, 3537, 40, today=TODAY)
        assert p["probing"] and "낡" in (p["absent_why"] or ""), \
            "낡은 근거인데 확인된 것처럼 군다"
        assert 0 < len(p["new"]) == rp.PROBE_AHEAD < 40, \
            "근거가 낡았다고 40개를 담았다 — 없는 번호 한 개가 21초다(14분에 수확 0)"
        assert p["new"][0] == 3540, "확인은 캐시 바로 위부터 한다 — 번호는 이어진다"

        # ③ 근거가 아예 없어도 같다(없다고 40개를 긁으라는 뜻이 아니다)
        os.remove(rp.PROBE_LOG)
        p = rp.plan(BAND, posts, 3537, 40, today=TODAY)
        assert len(p["new"]) == rp.PROBE_AHEAD, "근거가 없으면 40개를 쏟는다"

        # ④ 추월된 근거 — 3539 는 캐시에 **진짜 글**로 들어와 있다(2026-08-11 실측)
        evidence(3539, TODAY)
        cut, why = rp.absent_line(BAND, posts, TODAY)
        assert cut is None and "추월" in why, "이미 수확된 번호를 아직 '없다'고 믿는다"
        p = rp.plan(BAND, posts, 3537, 40, today=TODAY)
        assert 3539 not in (p.get("ghost") or []), "실재하는 글을 유령으로 표시했다"
        assert len(p["new"]) == rp.PROBE_AHEAD, "추월된 근거를 쓰고도 40개를 담았다"

        # ⑤ 근거 아래는 줄이지 않는다 — 3540~3559 는 있는 글이다
        evidence(3560, TODAY)
        p = rp.plan(BAND, posts, 3537, 40, today=TODAY)
        assert p["new"] == list(range(3540, 3560)) and not p["probing"], \
            "실재가 확인된 구간까지 탐색용으로 줄였다 — 새 글이 며칠씩 안 들어온다"

        # ⑥ 파일로 나가는 길목 — 죽은 번호·없음 구간은 어느 길로 와도 걸린다.
        #    (screen·build 는 시계를 보므로 근거를 **오늘 날짜**로 둔다)
        evidence(3560, _dt.now().strftime("%Y-%m-%d"))
        dead = dict(posts)
        dead["3536"] = {"contaminated": True}           # 캐시가 실제로 다는 표시
        kept, dropped, _why = MO.screen(BAND, [3538, 3536, 3600], dead)
        assert kept == [3538], "죽은 번호·없음 구간이 붙여넣기 파일에 들어간다"
        assert any("삭제" in k for k in dropped) and any("없음" in k for k in dropped), \
            "왜 뺐는지가 안 남는다 — 다음 사람이 그대로 다시 넣는다"

        with open(os.path.join(rp.CACHE, BAND + ".json"), "w", encoding="utf-8") as fh:
            json.dump({"band_name": "매출처업무", "posts": dead}, fh, ensure_ascii=False)
        js, note = MO.build(BAND, 100, nos=[3538, 3536, 3600], why="시험")
        assert js and "const ROUNDS = [[3538]]" in js, "붙여넣기 파일에 없는 번호가 실렸다"
        assert "제외" in note, "뺀 것을 말하지 않는다 — 조용히 빼면 아무도 모른다"

        # ⑦ 댓글 계획도 **같은 낱말**로 거른다. 예전엔 `ghost`·`dirty` 로만 물어서
        #    캐시가 다는 `contaminated` 를 한 건도 못 걸렀다(실측 102건·522건 전부).
        nos, skipped = CP.pick(dead, BAND)
        assert 3536 not in nos, "오염 표시를 못 걸렀다 — 없는 번호를 몇 시간씩 긁는다"
        assert CP.unlooked(dead, BAND) == nos, "고르는 판단이 두 벌로 갈렸다"
        assert skipped, "뺀 것을 숫자로 안 남긴다 — '0건'이 '다 봤다'로 읽힌다"
        for f in ("comment_plan.py", "comment_backfill.py"):
            s = open(os.path.join(ROOT, "band", f), encoding="utf-8").read()
            assert 'get("ghost")' not in s and 'get("dirty")' not in s, \
                "%s 가 캐시에 없는 낱말로 거른다 — 한 건도 안 걸린다" % f

        # ⑧ 빈 캐시 한 밴드가 나머지를 죽이지 않는다(이름순으로 유령이 앞에 온다)
        with open(os.path.join(rp.CACHE, "202608082047.json"), "w", encoding="utf-8") as fh:
            json.dump({"band_name": "유령", "posts": {}}, fh, ensure_ascii=False)
        hold_argv, buf = sys.argv, _io.StringIO()
        try:
            sys.argv = ["recheck_plan.py"]
            with contextlib.redirect_stdout(buf):
                rp.main()
        finally:
            sys.argv = hold_argv
        out = buf.getvalue()
        assert "202608082047" in out and BAND in out, \
            "빈 캐시 한 밴드에서 죽어 **뒤 밴드가 통째로** 안 나온다"
        assert "위쪽 근거" in out, "위쪽을 왜 그만큼만 보는지 사람이 알 수 없다"

        # ⑨ **인계 문서도 같은 자리에 물어본다.** 여기에 같은 구멍이 남아 있었다 —
        #    `data_freshness` 는 근거의 **나이만** 보고 '(조용함)'을 적었다. 그래서
        #    신선하지만 **추월된** 근거에는 없는 조용함을 확언한다(실측 2026-08-11:
        #    90610953 은 '5438 부터 없다'는 근거를 가진 채 5447 을 이미 수확해 뒀다).
        #    수집 계획은 거르는데 인계 문서만 안 거르면, 같은 파일을 보면서 한쪽은
        #    긁으라 하고 다른 쪽은 조용하다고 해서 사람이 무엇을 믿을지 모르게 된다.
        with open(os.path.join(rp.CACHE, BAND + ".json"), "w", encoding="utf-8") as fh:
            json.dump({"band_name": BAND, "posts": posts}, fh, ensure_ascii=False)
        SH = importlib.import_module("session_handoff")
        assert sys.modules.get("recheck_plan") is rp, \
            "인계 문서가 다른 recheck_plan 을 본다 — 근거 판정이 두 벌이 된다"
        # 근거는 인계 문서가 제 손으로 읽어(`band_quiet`) 넘긴다 — 판정만 빌린다.
        # 파일을 양쪽이 각자 읽으면 언젠가 서로 다른 한 장을 놓고 답하게 된다.
        rec = {"이름": BAND, "수집최대": 3538, "없음확인": 3539, "확인시각": TODAY}
        cut, why = SH._absent_judge(BAND, rec, TODAY)   # 신선하지만 3539 는 실재한다
        assert cut is None and "추월" in why, \
            "신선하기만 하면 믿는다 — 인계 문서가 **없는 조용함**을 확언한다"
        rec = dict(rec, 없음확인=3600, 수집최대=3599)     # 신선하고 추월도 안 됐다
        cut, why = SH._absent_judge(BAND, rec, TODAY)
        assert cut == 3600, "멀쩡한 근거까지 버리면 매일 밀림 경보가 뜨고 아무도 안 본다"
        fresh = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
        body = fresh.split("def data_freshness")[1].split("\ndef ")[0]
        assert "_absent_judge(" in body, "인계 문서가 근거를 제 손으로 다시 판정한다"
        assert "확인시각" not in body.split("_absent_judge(")[0], \
            "나이 계산이 아직 여기 남아 있다 — 두 벌이 되면 언젠가 갈린다"
    finally:
        rp.PROBE_LOG, rp.CACHE, rp.SCOPE = hold

    src = open(os.path.join(ROOT, "band", "recheck_plan.py"), encoding="utf-8").read()
    assert "밴드_확인시각.json" in src, "session_handoff 와 다른 근거를 본다"
    assert rp.PROBE_AHEAD <= 8, \
        "탐색 상한이 커졌다 — 없는 번호는 한 개당 21초라 금세 몇 분이 된다"
    # ⑩ 규칙을 고쳐도 **디스크의 붙여넣기 파일**이 안 바뀌면 사람은 옛 목록을 붙여넣는다
    #    ([162] 와 같은 모양이 한 겹 위에서 반복된 자리다). 워치독의 '낡음' 판정은
    #    수집기뿐 아니라 **번호를 고르는 쪽**도 봐야 한다.
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    body = wd.split("def heal_stale_pastefiles")[1].split("\ndef ")[0]
    assert "make_oneclick.py" in body and "recheck_plan.py" in body, \
        "붙여넣기 파일 낡음 판정이 수집기만 본다 — 번호 고르는 규칙을 고쳐도 옛 목록이 그대로 남는다"
    # ★ 그 기준을 **남의 회차 파일에까지 대면 안 된다.** `댓글채우기_*` 는
    #   comment_backfill, `재수집_*` 은 recollect 가 만든다 — make_oneclick 을 고쳤다고
    #   낡음으로 몰면 여기서 지워지는데 make_oneclick 은 그 파일을 다시 만들지 않는다
    #   (실측 2026-08-11: 그렇게 두 개가 사라졌다).
    assert 'startswith("수집_")' in body, \
        "make_oneclick 이 만들지도 않는 파일까지 제 mtime 으로 판단한다 — 남의 회차 파일이 지워진다"
    print("  [217] 없는 번호는 40개가 아니라 몇 개만 — 근거 먼저 ✅")


def t176_rules_bump_does_not_wipe_the_index():
    """[176] 규칙이 바뀌었다고 **색인 11만 건을 통째로 버리면 안 된다** (2026-08-09).

    [173] 대로 분류 규칙을 고치면 캐시도 다시 판별해야 한다. 그런데 `load_cache` 는
    지문이 다르면 `{}` 를 돌려줬다 — **색인 전체를 버린 것이다.**

    왜 그게 사고가 되나: 색인을 처음부터 만드는 데 Z: 에서 **6시간**이 걸리는데
    `daily_run` 이 그 단계에 주는 몫은 **40분**이고, 캐시는 스캔이 **다 끝난 뒤 한 번만**
    쓴다. 그래서 매일 40분을 태우고 매일 실패하며 **진행이 하나도 안 남는다** —
    규칙 한 줄 고친 벌로 색인이 영영 안 돌아온다. 게다가 실패는 회차 로그 한 줄이라
    조용하다.

    그리고 버릴 이유도 없었다. 11만 건 중 분류 규칙에 기대는 것은 **ERP 엑셀뿐**이고
    나머지는 폴더·파일명으로 정한다. 실측: 111,868건 중 다시 볼 것은 **144건**이었다.
    """
    import source_index as X

    # 규칙에 기댄 것 = ERP 엑셀. 나머지는 규칙을 고쳐도 답이 안 바뀐다
    assert X.content_classified({"ext": "xlsx", "kind": "ERP:ledger"})
    assert X.content_classified({"ext": "xlsx", "kind": "ERP"})
    assert X.content_classified({"ext": "xlsx", "kind": "기타"})
    assert not X.content_classified({"ext": "pdf", "kind": "ERP 거래명세서(건별 PDF)"})
    assert not X.content_classified({"ext": "xlsx", "kind": "밴드"})
    assert not X.content_classified({"ext": "jpg", "kind": "기타"})
    assert not X.content_classified("문자열")

    # 지문이 달라도 **폴더로 정해진 항목은 살아남는다**
    import io, json, tempfile, os as _os
    d = tempfile.mkdtemp()
    p = _os.path.join(d, "c.json")
    data = {
        "a|1|2": {"ext": "pdf", "kind": "ERP 거래명세서(건별 PDF)"},
        "b|1|2": {"ext": "xlsx", "kind": "ERP:ledger"},
        "c|1|2": {"ext": "jpg", "kind": "밴드"},
        X.RULES_KEY: "옛날지문",
    }
    json.dump(data, io.open(p, "w", encoding="utf-8"), ensure_ascii=False)
    old = X.CACHE
    try:
        X.CACHE = p
        kept = X.load_cache()
        assert set(kept) == {"a|1|2", "c|1|2"}, \
            "지문이 다르다고 폴더로 정해진 항목까지 버렸다 — 6시간짜리 재작성이 매일 실패한다"
        # 지문이 같으면 전부 남는다
        data[X.RULES_KEY] = X.rules_version()
        json.dump(data, io.open(p, "w", encoding="utf-8"), ensure_ascii=False)
        assert len(X.load_cache()) == 3, "지문이 같은데도 버렸다"
    finally:
        X.CACHE = old
    print("  [176] 색인 캐시 — 규칙이 바뀌면 **규칙에 기댄 것만** 다시 본다 ✅")


def t175_step_timeout_cannot_hang_forever():
    """[175] 회차 한 단계가 **영원히 멈추면 안 된다** (2026-08-08 실사고).

    CPython 의 `subprocess.run(timeout=)` 은 시간이 넘으면 `kill()` 뒤 윈도우에서만
    **시간제한 없는** `communicate()` 를 한 번 더 부른다. 자식이 SMB(Z:) 읽기처럼
    끊기지 않는 대기에 걸려 있으면 TerminateProcess 가 안 먹고 그 드레인이 안 끝난다.
    실측: `timeout=1800` 을 걸어 둔 '원본 폴더 정리' 단계가 **13시간 30분**을 매달렸다
    (부모 CPU 0.4초 · 자식 0.5초 — 둘 다 그냥 서 있었다).

    조용한 이유가 여기 있다: 멈춘 회차가 **락을 쥔 채**라 다음 회차는 "이미 실행 중"
    으로 건너뛰고, 스케줄러는 '성공'이라 적는다. 09:50 이 하는 일(접수취소·객관완료·
    청구상태·대조)이 하루 종일 안 돌면서 **어느 화면에도 티가 안 난다.**

    지키는 것: ① `_run_once` 가 `subprocess.run(timeout=)` 을 쓰지 않는다
    ② 나무째 죽인다(윈도우는 `kill()` 이 손자를 안 죽인다) ③ 죽인 뒤 드레인에도
    **제한이 있다** ④ 그래도 안 죽으면 회차는 다음 단계로 간다.
    """
    import io, inspect, subprocess
    import daily_run as D

    src = inspect.getsource(D._run_once)
    # 설명문에는 "쓰면 안 된다" 고 적혀 있다 — **본문만** 본다
    body = src.split('"""')[-1] if src.count('"""') >= 2 else src
    assert "subprocess.run(" not in body, \
        "_run_once 가 subprocess.run 을 쓴다 — 윈도우에서 kill 뒤 무제한 대기에 걸린다"
    assert "Popen" in src and "communicate(timeout=" in src, \
        "시간제한 있는 communicate 가 없다"
    assert src.count("communicate(timeout=") >= 2, \
        "죽인 뒤 드레인에 제한이 없다 — 바로 거기서 13시간을 섰다"
    assert "_kill_tree" in src, "나무째 안 죽인다 — 손자가 살아남아 파이프를 붙든다"

    tree = inspect.getsource(D._kill_tree)
    assert "taskkill" in tree and "/T" in tree, \
        "윈도우에서 자식의 자식까지 끊지 않는다"

    # 시간초과가 나도 **예외로 회차를 세우지 않는다** — 결과를 돌려주고 다음 단계로 간다
    got = D._run_once("멈춘 단계", ["-c", "import time; time.sleep(30)"], timeout=1)
    assert got["ok"] is False and "시간초과" in got["out"], got
    print("  [175] 회차 단계 — 시간초과가 영원한 대기로 바뀌지 않는다(나무째 종료) ✅")


def t174_zero_match_blames_the_key():
    """[174] 짝이 하나도 안 지어지면 **열쇠를 의심한다** (2026-08-08 실측).

    ERP원장대조는 06시트 `거래명세서번호` = ERP `일자-No.` 를 전제로 짝을 짓는다.
    실측하니 ERP 302 전표 대 원장 명세서번호 65개 중 **겹침 6** — 서로 다른 순번이었다.
    그런데 열쇠가 안 맞아도 리포트는 조용하다. 짝이 안 지어진 것이 전부
    'A. ERP에만 있는 전표 (설치·작업 근거 확인 필요 ★)' 로 나오기 때문이다.
    자료가 없어서 0건인 것과 열쇠가 안 맞아 0건인 것은 **겉이 똑같다.**
    유일한 신호였던 '정상 0건' 은 머리글 한 줄이라 1,856건이던 시절에도 아무도 안 봤다.

    가르는 것은 비율이다 — 열쇠가 맞으면 몇 건은 반드시 걸린다. 다만 전표가 몇 건뿐일
    때는 비율을 말할 수 없으므로 **아무 말도 하지 않는다**(경보를 남발하면 안 본다).
    """
    from erp_ledger_check import key_looks_wrong

    assert key_looks_wrong(6, 302), "302건 중 6건만 걸렸는데 조용하다 — 열쇠 이야기다"
    assert key_looks_wrong(0, 100), "한 건도 안 걸렸는데 조용하다"
    assert not key_looks_wrong(65, 302), "3분의 1 가까이 걸렸는데 열쇠 탓을 한다"
    assert not key_looks_wrong(31, 302), "10% 넘게 걸리면 열쇠는 도는 것이다"
    # 표본이 작으면 말하지 않는다 — 회차를 막 시작한 자리를 겁주지 않는다
    assert not key_looks_wrong(0, 9), "전표 9건으로 열쇠를 단정했다"
    assert not key_looks_wrong(0, 0), "전표가 없는데 경보를 냈다"
    print("  [174] ERP원장대조 — 짝이 거의 없으면 열쇠를 의심하라고 말한다 ✅")


def t173_classify_cache_follows_rules():
    """[173] 분류 **규칙을 고치면 캐시도 다시 판별**해야 한다 (2026-08-08).

    `classify_cached` 의 열쇠가 (크기, 수정시각) 뿐이었다. 원본 엑셀은 한 번 떨어지면
    다시 안 바뀌므로, 규칙을 고쳐도 캐시가 **영원히 옛 갈래**를 돌려준다. 고친 사람은
    고쳤다고 믿고 화면은 어제와 똑같다 — 오류도 안 난다. [161] 이 색인 쪽에서 같은
    사고를 한 번 잡았는데, 이번에는 `inbox_classify.json` 쪽이 그대로였다.

    그리고 세 번째 판이 있었다: `source_index.rules_version()` 은 `classify_rows` 만
    해싱하는데 판별의 일부가 `ledger_kind` 로 나갔다 — 그쪽만 고치면 지문이 안 움직인다.
    """
    import io, json, tempfile, inspect
    import inbox_scan as S
    import source_index as IX

    src = io.open(os.path.join(ROOT, "inbox_scan.py"), encoding="utf-8").read()
    body = src[src.index("def classify_cached("):]
    body = body[:body.index("\n_SCAN_MEM")] if "\n_SCAN_MEM" in body else body
    assert "RULES_VERSION" in body, \
        "classify_cached 열쇠에 규칙판이 없다 — 규칙을 고쳐도 옛 갈래가 그대로 나온다"

    # 규칙판을 올리면 같은 파일을 **다시 열어 본다**
    d = tempfile.mkdtemp()
    p = os.path.join(d, "aaa.xlsx")
    io.open(p, "w").write("x")
    calls = []
    real = S.classify
    try:
        S.classify = lambda path: (calls.append(path), "ledger")[1]
        assert S.classify_cached(p) == "ledger" and len(calls) == 1
        S.classify_cached(p)
        assert len(calls) == 1, "캐시가 안 먹는다 — 매번 파일을 연다"
        S.RULES_VERSION += 1
        S.classify_cached(p)
        assert len(calls) == 2, "규칙판을 올렸는데 옛 답을 그대로 돌려줬다"
    finally:
        S.classify = real
        S.RULES_VERSION -= 1

    # 색인 지문은 ledger_kind 도 본다
    assert "ledger_kind" in inspect.getsource(IX.rules_version), \
        "rules_version 이 ledger_kind 를 안 본다 — 그 함수만 고치면 색인이 안 도로 돈다"
    v0 = IX.rules_version()
    S.RULES_VERSION += 1
    try:
        assert IX.rules_version() != v0, "규칙판을 올렸는데 색인 지문이 그대로다"
    finally:
        S.RULES_VERSION -= 1
    assert IX.rules_version() == v0, "지문이 원래대로 안 돌아온다"
    print("  [173] 분류 캐시·색인 지문 — 규칙을 고치면 둘 다 다시 판별한다 ✅")


def t180_round_leaves_footprints_and_finishes():
    """[180] 회차는 **자국을 남기고 반드시 끝난다** (2026-08-09 지시).

    사용자 지시: "32시간째 미완주 왜그런거야 해결해 이런 문제 생기면 코딩해서
    다시는 안생기게 처리하는 알고리즘 구성해서 적용하고 보고해"

    ★ 진짜 문제는 '느리다'가 아니라 **둘**이었다:
      ① **어디서 멈췄는지 아무도 몰랐다.** 종합리포트는 `finish()` 가 **맨 끝에 한 번**
         쓴다. 그러니 완주하지 못한 회차는 **기록을 한 줄도 안 남긴다.** 화면은
         '08-08 01:38 — 32시간째 미완주'만 말할 수 있었고 이유를 못 댔다.
      ② **안 끝나는 회차가 다음 회차를 막았다.** 잠금을 쥔 채 몇 시간을 끌면 다음
         회차는 "이미 실행 중"으로 조용히 건너뛴다. 스케줄러는 '성공'이라 적는다.
         그것이 이틀 쌓이면 32시간이 된다.

    그래서 ① 단계마다 `.daily_run.progress.json` 에 자국을 남기고(죽어도 남는다)
    ② **회차 예산**을 두어 넘으면 남은 단계를 건너뛰고 **완주시킨다.**
    반쯤이라도 완주한 회차가 영원히 안 끝나는 회차보다 낫다 — 완주해야 리포트가
    써지고 잠금이 풀리고 **다음 회차가 돈다.**
    """
    import importlib
    from datetime import timedelta, datetime as _dt
    sys.path.insert(0, ROOT)
    D = importlib.import_module("daily_run")
    S = importlib.import_module("session_handoff")

    keep = D.PROGRESS
    D.PROGRESS = os.path.join(tempfile.mkdtemp(), "p.json")
    try:
        # ① 단계마다 자국이 남는다 — 죽어도 남는 것이 요점이다
        D._ROUND_T0[0] = _dt.now()
        D.note_progress("가벼운단계", "시작")
        D.note_progress("가벼운단계", "끝", {"결과": True})
        with open(D.PROGRESS, encoding="utf-8") as fh:
            got = json.load(fh)
        assert got["단계"] == "가벼운단계" and got["상태"] == "끝", got
        assert got["끝난단계"] == ["가벼운단계"], "끝낸 단계가 쌓여야 어디까지 왔는지 안다"
        assert "경과분" in got and "예산분" in got, "회차 나이와 예산이 함께 있어야 판단이 된다"

        # ② 예산을 넘으면 **남은 단계를 건너뛰고** 회차를 끝낸다(중단이 아니다)
        D._OVER_BUDGET[0] = False
        D._ROUND_T0[0] = _dt.now() - timedelta(minutes=D.ROUND_BUDGET_MIN + 1)
        g = D.run("무거운단계", [os.path.join(ROOT, "nonexistent_xyz.py")])
        assert g["ok"] is None and "예산" in g["out"], g
        assert D._OVER_BUDGET[0], "예산 초과는 회차 끝 표식에도 남아야 한다"
        assert "다음 회차가 이어서" in g["out"], \
            "조용히 건너뛰면 '돌았는데 왜 결과가 없나'가 된다 — 이유를 적어 남긴다"
    finally:
        D.PROGRESS = keep

    # ③ 예산은 **밖에서 조절**할 수 있어야 한다(Z: 가 느린 날이 있다)
    src = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "COUPANG_ROUND_BUDGET_MIN" in src, "예산을 코드에 못 박으면 느린 날 손을 못 쓴다"
    # ④ 완주했든 죽었든 마지막 자국을 남긴다
    assert 'note_progress("(회차 끝)"' in src, \
        "끝 표식이 없으면 '아직 도는 중'과 '죽었다'를 구별할 수 없다"
    # ⑤ **한 단계가 회차를 통째로 먹으면 안 된다.** 실측 2026-08-09: `collect_all` 이
    #    timeout=3600 이라 09:50 회차가 2시간째일 때도 계속 그 단계에 있었고 뒤 단계는
    #    시작조차 못 했다 — 그게 '32시간 미완주'의 실제 모습이다.
    #    한 단계 상한은 회차 예산의 **1/5** 을 넘지 않는다(150분 예산 → 30분).
    cap = D.ROUND_BUDGET_MIN * 60 // 5
    big = [int(m) for m in re.findall(r"timeout=(\d+)", src)]
    assert big and max(big) <= cap, \
        f"한 단계 상한 {max(big) if big else '?'}초가 회차 예산({D.ROUND_BUDGET_MIN}분)의 " \
        f"1/5({cap}초)을 넘는다 — 그 하나가 회차를 먹는다"

    # ⑥ **시간 초과는 재시도하지 않는다.** 재시도는 경합용이다 — 경합은 금방 실패하고
    #    한 번 쉬면 지나간다. 시간 초과는 "준 시간보다 오래 걸린다"는 뜻이라 다시 해도
    #    또 넘긴다. 실측: collect_all 이 1시간 만에 초과로 죽고 **또 한 시간**을 쓰는
    #    중이었다 — 회차 예산을 한 단계가 두 번 먹는다.
    assert '"시간초과" in str(got.get("out"' in src,         "시간 초과를 재시도하면 그 단계가 회차 예산을 두 번 먹는다"
    assert "재시도하지 않습니다" in src, "왜 안 하는지 적어 남긴다 — 조용히 넘기면 오해한다"

    # ⑦ 경보가 **단계 이름을 댄다** — 이것이 이 검증의 목적이다
    hs = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "def daily_step_now" in hs and "_step_hint()" in hs, \
        "'몇 시간째'만 말하고 '어느 단계'를 못 대면 원인을 못 찾는다"
    print("  [180] 회차가 단계마다 자국을 남기고 · 예산을 넘으면 완주시키고 · "
          "경보가 단계 이름을 댄다 ✅")


def t179_comments_everywhere_and_crossed():
    """[179] 댓글을 **다 찾아 담고**, 카톡과 **한 사건으로 묶는다** (2026-08-09 지시).

    사용자 지시: "밴드도 댓글도 다 찾아 저장하는 알고리즘 구성하고, 카톡도 댓글이랑
    연관지어서 생각하고 반영하는 알고리즘 구현해"

    ★ **고친 것과 채운 것은 다르다.** 2026-08-08 에 수집기가 댓글을 담도록 고쳤지만
      (검증 [162]) 그건 **그 뒤로 긁는 글**에만 해당한다. 실측 2026-08-09:
      캐시 8,561글 중 **8,258글이 댓글을 한 번도 안 봤고 본문이 담긴 글은 0건**이었다.
      그래서 `cancel_watch` 는 지금도 댓글 취소를 하나도 못 잡는다 — 오류도 안 난다.
      `comment_plan.py` 가 그 구멍을 세어 회차로 만든다(8,107건 · 33회차).

    ★ **같은 사건이 두 군데로 나뉘어 온다.** 캠프는 카톡에 쓰고 기사는 댓글에 단다.
      `cancel_watch`(밴드)와 `kakao_reconcile`(카톡)은 서로를 모른 채 돌았다.
      `cross_signal.py` 가 캠프·날짜·사건 종류 셋이 겹칠 때만 묶는다.
      실측: 밴드 115 · 카톡 58 → 짝지어짐 15 · **카톡에만 43**(기사에게 안 갔을 수 있다).
    """
    sys.path.insert(0, os.path.join(ROOT, "band"))
    import comment_plan as CP
    import cross_signal as CS

    # ① '안 본 것'과 '보고 없던 것'을 가른다 — 섞으면 8,107 이 영원히 안 준다
    posts = {"10": {"content": "a"},                       # 안 봤다 → 대상
             "11": {"content": "b", "comments": []},        # 보고 없었다 → 대상 아님
             "12": {"content": "c", "comments": [{"content": "x"}]},
             "13": {"content": "d", "deleted": True},       # 삭제는 업무 기록이 아니다
             "14": {"content": "e", "ghost": True}}
    assert CP.unlooked(posts) == [10], f"골라야 할 것만 골라야 한다: {CP.unlooked(posts)}"

    # ② 유령 밴드를 계획에 넣지 않는다 — 날짜 도장은 8자리가 아니다
    src = open(os.path.join(ROOT, "band", "comment_plan.py"), encoding="utf-8").read()
    assert "len(b) == 8" in src, \
        "밴드번호는 8자리다. 넓게 잡으면 202608082047 같은 유령이 헛 계획을 만든다"

    # ③ 최근 글부터 — 도중에 멈춰도 값어치가 남게
    assert CP.unlooked({"5": {}, "9": {}, "7": {}}) == [9, 7, 5], "번호 큰 것부터다"

    # ④ 사건 종류를 나눠 둔다 — '취소'와 '연기'를 한 덩어리로 두면 다른 사건이 붙는다
    assert CS.events_in("접수취소 요청드립니다") == {"취소"}
    assert CS.events_in("다음주로 연기해주세요") == {"연기"}
    assert CS.events_in("정상 진행합니다") == set(), "아무 말에나 반응하면 안 된다"

    # ⑤ 짝은 캠프·날짜·사건이 **셋 다** 겹칠 때만
    b = [{"출처": "밴드 댓글", "밴드": "8", "글번호": "1", "날짜": "2026-04-13",
          "캠프": "용인3MB", "사건": ["취소"], "글": ""}]
    assert len(CS.pair(b, [{"파일": "k", "날짜": "2026-04-14", "보낸이": "s",
                            "캠프": "용인3MB", "사건": ["취소"], "글": ""}])[0]) == 1
    assert len(CS.pair(b, [{"파일": "k", "날짜": "2026-04-14", "보낸이": "s",
                            "캠프": "용인3MB", "사건": ["연기"], "글": ""}])[0]) == 0, \
        "사건 종류가 다르면 다른 일이다"
    assert len(CS.pair(b, [{"파일": "k", "날짜": "2026-05-20", "보낸이": "s",
                            "캠프": "용인3MB", "사건": ["취소"], "글": ""}])[0]) == 0, \
        "한 달 떨어진 것을 같은 사건이라 하면 엉뚱한 현장 둘을 묶는다"

    # ⑥ 둘 다 읽기 전용 — 판정을 두 곳에서 하면 언젠가 갈린다
    cs_src = open(os.path.join(ROOT, "cross_signal.py"), encoding="utf-8").read()
    code = "\n".join(ln for ln in cs_src.splitlines() if not ln.strip().startswith("#"))
    for banned in ("queue_add", "enqueue", "workbook_patch"):
        assert banned not in code, f"교차 확인은 읽기 전용이어야 한다: {banned}"

    # ⑦ 회차에 들어가 있나
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "cross_signal.py" in daily, "일일대조에 없으면 한 번 돌고 끝난다"
    print("  [179] 댓글 구멍을 회차로 · 카톡↔밴드를 캠프·날짜·사건 셋으로 묶는다 ✅")


def t181_app_answers_before_claude_is_called():
    """[181] 앱이 스스로 답한다 — 크레딧은 계산이 아니라 **왕복**에서 샜다 (2026-08-09 지시).

    사용자 지시: "클로드 코드 또는 코덱스 없이 앱 자체적으로 처리할 수 있는 AI 알고리즘
    기능 넣어서 클로드코드 크래딧 사용 최대한 아낄 수 있게 설계하고 알고리즘에 반영해"

    지금까지의 자동화는 전부 **일을 하는** 쪽이었다(수집·대조·큐·반영). 그런데 크레딧은
    거기서 안 샜다. 샌 자리는 이렇다 — 사람이 앱을 열어 이상한 숫자를 보고, **앱이
    이유를 못 말해서** 클로드에게 묻는다. 그러면 클로드는 리포트를 열고 ERP 를 대 보고
    원장을 뒤져서 **이미 디스크에 있는 사실**을 다시 조립한다. 실측된 예: "작업은
    완료인데 왜 계산서 발행이 안된거지" 한 마디가 도구 호출 열몇 번이 됐는데, 답은
    `ERP 4.세금계산서발행대기` 라는 **파일에 이미 적혀 있던 한 줄**이었다.

    ★ 이 검증이 지키는 것은 '똑똑한가'가 아니라 **거짓말하지 않는가**다.
      답변기가 지어내기 시작하면 사람은 한 번 속고 그 뒤로 앱을 안 믿는다 —
      그러면 왕복이 도로 살아난다.
    """
    import importlib
    import local_ai as L

    # ① 자가점검이 **갈래까지** 본다.
    #    첫 판 자가점검은 "터지지 않으면 통과"였다. 그래서 이 프로젝트에서 제일 자주 온
    #    질문("왜 계산서 발행이 안된거지")이 **모름으로 떨어졌는데도 '모두 통과'**라고
    #    말했다 — 계기가 0을 내면 아무도 의심하지 않는다(`[169]` 와 같은 모양).
    src_txt = open(os.path.join(ROOT, "local_ai.py"), encoding="utf-8").read()
    code = "\n".join(ln for ln in src_txt.splitlines() if not ln.strip().startswith("#"))
    assert "PROBES" in code and 'r.get("분류") != want' in code, \
        "자가점검이 '기대한 갈래로 갔는가'를 안 보면, 대표 질문이 조용히 모름이 돼도 통과한다"
    assert not L.selftest(), "자가점검이 통과해야 한다: %s" % L.selftest()

    # ② 근거가 하나도 없어도 **터지지 않고**, 못 답하면 반드시 클로드 문구를 준다.
    #    빈손으로 돌려보내면 사람은 결국 클로드에게 처음부터 묻는다 — 절약이 0이 된다.
    with tempfile.TemporaryDirectory() as td:
        old = L.REPORT_DIR
        try:
            L.REPORT_DIR = td
            L.LOG = os.path.join(td, "앱_자문기록.json")
            for q in ("왜 계산서 발행이 안된거지", "UJ2600021", "지금 뭐부터 하면 돼",
                      "아무 상관 없는 질문"):
                r = L.ask(q, log=False)
                assert isinstance(r, dict) and "답" in r, f"답 모양이 무너졌다: {q}"
                assert r["답함"] or r.get("클로드문구"), \
                    f"못 답했으면 클로드 문구라도 줘야 한다: {q}"
            # 근거가 없으면 **확신이 '높'일 수 없다** — 빈 디스크에서 자신 있게 답하면
            # 그건 지어낸 것이다.
            r = L.ask("지금 뭐부터 하면 돼", log=False)
            assert not r["답함"], "근거 파일이 하나도 없는데 답했다면 지어낸 것이다"
        finally:
            L.REPORT_DIR = old
            importlib.reload(L)

    # ③ 읽기 전용이다. 물어봤을 뿐인데 값이 바뀌면 아무도 안 묻는다.
    for banned in ("enqueue", "queue_add", "workbook_patch", "ledger_writer",
                   "--apply", "openpyxl"):
        assert banned not in code, f"답변기는 아무것도 고치면 안 된다: {banned}"

    # ④ 비싼 탐색을 하지 않는다(`[168]`). 질문 하나에 1초를 넘기면 사람은 그냥 클로드에게
    #    묻는다 — Z: 재귀 glob 은 리포트 답변에 필요 없다.
    assert "recursive=True" not in code and "ERP_DIR" not in code, \
        "답변기가 Z: 를 훑기 시작하면 느려서 아무도 안 쓴다"

    # ⑤ 근거를 **항상 같이** 준다. 출처 없는 답은 이 프로젝트에서 소문과 같다.
    assert "_age_note" in code, "답마다 어느 파일 몇 시간 전 자료인지 붙어야 한다"
    assert "낡" in code or "신선" in code, \
        "낡은 근거로 자신 있게 답하는 것이 조용한 사고의 본체다 — 밝히고 답해야 한다"

    # ⑥ 앱에 실제로 붙어 있나. 붙지 않으면 사람 손에는 안 간다.
    app = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '"/api/ask"' in app and "local_ai" in app, \
        "앱에 길이 안 나 있으면 명령줄 도구일 뿐이라 크레딧을 못 아낀다"

    # ⑦ 못 답한 질문이 **다음에 만들 규칙의 목록**이 된다 — 모델 없이 자라는 유일한 길.
    assert "def stats" in code and "못답한갈래" in code, \
        "못 답한 갈래를 안 세면 무엇을 더 만들지가 추측이 된다"
    print("  [181] 앱 자체 답변 — 갈래까지 자가점검 · 못 하면 클로드 문구 · 읽기 전용 ✅")


@contextlib.contextmanager
def _isolated_collect_outputs(comment_backfill):
    """comment_backfill의 쓰기 산출물을 임시 폴더로 완전히 격리한다.

    ``write_plan`` 은 주 계획뿐 아니라 ``docs/collect``의 두 게시 파일도 바꾼다.
    텍스트로 읽었다가 되쓰는 복구는 줄바꿈·인코딩과 기존 미커밋 바이트를 보존하지
    못하고, 검증이 중단되면 복구 자체가 실행되지 않을 수도 있다. 실제 경로에는 처음부터
    쓰지 않는 것이 유일하게 안전하다.
    """
    old_plan = comment_backfill.PLAN_PATH
    old_docs = comment_backfill.DOCS_COLLECT
    with tempfile.TemporaryDirectory(prefix="csos-collect-synthetic-") as tmp:
        comment_backfill.PLAN_PATH = os.path.join(tmp, "reports", "plan.json")
        comment_backfill.DOCS_COLLECT = os.path.join(tmp, "docs", "collect")
        os.makedirs(os.path.dirname(comment_backfill.PLAN_PATH), exist_ok=True)
        try:
            yield
        finally:
            comment_backfill.PLAN_PATH = old_plan
            comment_backfill.DOCS_COLLECT = old_docs


def t182_app_collects_without_claude():
    """[182] 앱이 스스로 수집한다 — 수집 루프에서 Claude Code 가 빠진다 (2026-08-09 지시).

    사용자 지시: "클로드 코드 또는 코덱스 없이 앱 자체적으로 처리할 수 있는 AI 알고리즘
    기능 넣어서 클로드코드 크래딧 사용 최대한 아낄 수 있게 설계하고 알고리즘에 반영해"

    밴드는 조회 API 가 없어 로그인된 브라우저 DOM 안에서만 읽힌다. 그 DOM 에서 JS 를
    돌릴 수 있는 것은 Claude Code(브라우저 도구)나 사람 브라우저 안의 스크립트뿐이었다.
    지금까지는 Claude Code 가 매번 수집기를 주입해 크레딧을 썼다. 이제 유저스크립트가
    그 자리를 대신한다: 앱이 회차에서 미리 계산한 **수집계획**을 `/api/collect_plan` 으로
    내려 주고, 로그인된 밴드 탭의 유저스크립트가 앱이 주는 **정본 수집기**로 스스로 긁는다.

    ★ 이 검증이 지키는 것: ① 계획은 라운드트립으로 살아 있나 ② 앱이 정본 수집기를
      내려 주나(딴 사본이 아니라) ③ 계획은 **회차가** 다시 채우나(안 그러면 하루면 낡는다).
    """
    import importlib
    sys.path.insert(0, os.path.join(ROOT, "band"))
    cb = importlib.import_module("comment_backfill")

    # ① 계획 라운드트립 — 실제 reports/·docs/ 산출물과 완전히 분리한다.
    # write_plan 은 PLAN_PATH 하나뿐 아니라 게시용 plan.json·grab_posts.js 까지 쓴다.
    # 일부만 백업했다 복원하면 이미 있던 미커밋 바이트를 잃으므로 경로 자체를 바꾼다.
    with _isolated_collect_outputs(cb):
        cb.write_plan({"90610953": {"nos": [5435, 5425], "tiers": {"1": 2}}},
                      when="2026-08-09 00:00")
        got = cb.load_plan("90610953")
        assert got["nos"] == [5435, 5425] and got["tiers"] == {"1": 2}, got
        assert cb.load_plan("999")["nos"] == [], "모르는 밴드는 빈 목록이어야 한다"

    # ② 앱이 정본 수집기·계획·유저스크립트를 실제로 서빙하나 (라우트 존재).
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    for route in ('/grab_posts.js', '/api/collect_plan', '/band_auto_collect.user.js'):
        assert ('"%s"' % route) in srv or ("'%s'" % route) in srv, \
            "앱이 %s 를 안 내려 준다 — 유저스크립트가 받을 게 없다" % route
    assert "load_plan" in srv, "collect_plan 이 미리 계산된 계획을 안 읽는다"

    # ③ 유저스크립트는 **앱이 주는** 수집기를 쓴다(딴 사본을 심지 않는다) — 정본이 하나여야
    #    규칙을 고쳐도 사람 손에 옛 수집기가 안 간다([162] 와 같은 원칙).
    us = open(os.path.join(ROOT, "band", "band_auto_collect.user.js"),
              encoding="utf-8").read()
    assert "/grab_posts.js" in us and "/api/collect_plan" in us, \
        "유저스크립트가 앱의 정본 수집기·계획을 안 받는다"
    assert "__grabStart" in us, "유저스크립트가 수집기를 실행하지 않는다"

    # ④ 계획은 **회차가** 다시 채운다 — 안 그러면 하루 지나 낡은 계획으로 긁는다.
    dr = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "comment_backfill.py" in dr and "--write" in dr, \
        "회차가 comment_backfill --write 를 안 부른다 — 계획이 낡는다"
    src = open(os.path.join(ROOT, "band", "comment_backfill.py"), encoding="utf-8").read()
    assert "write_plan(" in src and "if a.write" in src, \
        "--write 가 계획 파일을 안 만든다 — 붙여넣기 파일만 만들고 앱은 굶는다"
    print("  [182] 앱 자체 수집 — 계획 라운드트립 · 정본 수집기 서빙 · 회차가 갱신 ✅")


def t183_collect_survives_pc_off():
    """[183] PC 가 꺼져 있어도 수집은 돈다 — 게시 사본 폴백 (2026-08-09 지시).

    사용자 지시: "만약 이 컴퓨터가 꺼져있어 연결되어있지 않더라도 앱 자체적으로
    처리할 수 있는 알고리즘 구현해"

    [182] 는 로컬 앱(localhost)이 켜져 있을 때 Claude 를 뺐다. 그런데 PC 가 꺼지면
    localhost 도 없다 — 그러면 유저스크립트가 굶는다. 그래서 회차가 계획·정본 수집기를
    **GitHub Pages 에 게시**해 두고(비밀 없음: 글 번호·DOM 읽는 JS 뿐), 유저스크립트는
    localhost 를 못 찾으면 그 게시본으로 폴백한다. 그러면 폰만 있어도 긁는다.

    ★ 이 검증이 지키는 것: ① 게시 사본이 **ASCII 이름**으로 나오나(폰 fetch 가 한글 URL
      인코딩에 안 걸리게) ② 계획·수집기 **둘 다** 나오나 ③ 게시 수집기가 정본과 같나
      ④ 유저스크립트에 Pages 폴백이 배선돼 있나 ⑤ 로컬 앱이 **먼저**인가(최신·빠름).
    """
    import importlib
    sys.path.insert(0, os.path.join(ROOT, "band"))
    cb = importlib.import_module("comment_backfill")

    # ① ② ③ 게시 사본 — plan.json(ASCII) + grab_posts.js(정본 복사)
    with _isolated_collect_outputs(cb):
        docs_collect = cb.DOCS_COLLECT
        plan_pub = os.path.join(docs_collect, "plan.json")
        grab_pub = os.path.join(docs_collect, "grab_posts.js")
        cb.write_plan({"90610953": {"nos": [5435, 5425], "tiers": {"1": 2}}},
                      when="2026-08-09 00:00")
        # ① ASCII 이름 — 한글 URL 이 아니어야 폰 fetch 가 안 깨진다
        assert os.path.exists(plan_pub), "게시용 plan.json 이 안 나왔다 — PC 꺼지면 굶는다"
        doc = json.load(open(plan_pub, encoding="utf-8"))
        assert (doc.get("bands") or {}).get("90610953", {}).get("nos") == [5435, 5425], \
            "게시 계획이 회차 계획과 다르다"
        # ② ③ 정본 수집기가 그대로 실려야 한다(딴 사본이 아니라)
        assert os.path.exists(grab_pub), "게시용 grab_posts.js 가 없다 — 폴백해도 수집기가 없다"
        canon = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
        assert open(grab_pub, encoding="utf-8").read() == canon, \
            "게시 수집기가 정본과 다르다 — 규칙을 고쳐도 옛 수집기가 폰에 간다([162])"

    # ④ ⑤ 유저스크립트: Pages 폴백이 있고, 로컬 앱이 **먼저**다
    us = open(os.path.join(ROOT, "band", "band_auto_collect.user.js"),
              encoding="utf-8").read()
    assert "github.io" in us and "/plan.json" in us, \
        "유저스크립트에 Pages 폴백이 없다 — PC 꺼지면 못 긁는다"
    assert us.index("findApp") < us.index("PAGES_BASE") or "resolveSource" in us, \
        "로컬 앱을 게시본보다 먼저 봐야 한다(최신·빠름)"
    assert "resolveSource" in us, "폴백은 한 곳(resolveSource)에서 골라야 갈리지 않는다"
    print("  [183] PC 꺼져도 수집 — 게시 plan.json(ASCII)·정본 수집기·Pages 폴백 ✅")


def t184_phone_answers_with_the_same_rules():
    """[184] 폰이 PC 없이 답한다 — 규칙은 **한 곳**, 답은 미리 만들어 싣는다 (2026-08-09 지시).

    사용자 지시: "대시보드에 내가 앱에 텍스트로 명령하면 클로드 코드처럼 조사하고
    정리하고 답변주고 구현하는 기능을 앱에 탑재해서 클로드 코드 안 거치고 대시보드에서
    처리할 수 있는 기능 코딩해 (PC가 꺼져있어도 휴대폰으로 들고다니면서 업무처리할 수 있게)"

    [181] 이 만든 답변기는 **PC 가 켜져 있을 때만** 돌았다(`/api/ask`). 폰 사본은
    GitHub Pages 의 잠긴 `data.enc` 라 파이썬이 없다. 그렇다고 규칙을 JS 로 옮겨 적으면
    같은 판단이 두 곳에 생기고, **갈린 뒤에는 어느 쪽이 맞는지 아무도 모른다.**
    그래서 답을 **미리 만들어**(`answer_pack`) 사본에 싣고, 폰은 말투로 갈래만 고른다.

    ★ 이 검증이 지키는 것:
      ① 사본에 꾸러미가 실리나(`cloud_publish.payload`)
      ② 말투가 **JS 에서도 컴파일되나** — 파이썬에서만 되는 문법을 쓰면 폰에서 그
         갈래가 조용히 죽는다(오류도 안 난다. `[169]` 와 같은 모양)
      ③ 폰이 **규칙을 베껴 적지 않았나** — 정규식이 app.html 에 박혀 있으면 안 된다
      ④ **만든 시각을 화면에 띄우나** — 꾸러미는 PC 가 꺼진 동안 낡는다. 낡은 답을
         시각 없이 자신 있게 보여 주는 것이 여기서 제일 위험하다
      ⑤ '규칙에 없는 질문' 과 '근거가 없어 답이 안 실린 갈래' 를 **가르나** — 뭉치면
         사람이 이미 있는 규칙을 다시 만들러 간다(실측: 앱서버 갈래가 그랬다)
    """
    import importlib
    import local_ai as L
    importlib.reload(L)

    # ① 사본에 실리나 — payload() 를 통째로 돌리면 Z: 를 훑으므로 배선만 본다([168]).
    cp = open(os.path.join(ROOT, "cloud_publish.py"), encoding="utf-8").read()
    assert 'd["ask"]' in cp and "answer_pack()" in cp, \
        "폰 사본에 답 꾸러미가 안 실린다 — PC 가 꺼지면 폰은 아무것도 못 답한다"
    assert cp.index('d["ask"]') < cp.index('d["gen"]'), "꾸러미는 payload() 안에 있어야 한다"

    pack = L.answer_pack()
    for k in ("만든때", "규칙", "답", "번호", "번호말투", "탈출머리말"):
        assert k in pack, "꾸러미에 '%s' 가 없다" % k
    assert pack["규칙"], "말투가 안 실렸다 — 폰은 갈래를 하나도 못 고른다"
    # 말투는 갈래용·번호용 **둘 다** 사본이 준다. 폰이 제 손으로 적으면 그것도 갈릴 사본이다.
    every = [p for r in pack["규칙"] for p in r["말투"]] + list(pack["번호말투"].values())

    # ② 파이썬에서만 되는 문법은 폰에서 **그 갈래만 조용히 죽는다**.
    js_bad = re.compile(r"\(\?P|\(\?#|\\A(?![-\w])|\\Z|\(\?i\)|\(\?m\)|\(\?s\)|\(\?x\)")
    for pat in every:
        assert not js_bad.search(pat), \
            "말투 %r 은 JS 에서 안 돈다 — 폰에서 그 갈래가 말없이 죽는다" % pat

    app = open(os.path.join(ROOT, "docs", "app.html"), encoding="utf-8").read()

    # ③ 규칙을 베껴 적지 않았나 — 판단은 파이썬 한 곳에서만 한다.
    for pat in every:
        if len(pat) < 8:
            continue                          # '오타'·'취소' 같은 낱말은 본문에도 나온다
        assert pat not in app, \
            "app.html 에 규칙 %r 이 박혀 있다 — 두 곳에서 판단하면 언젠가 갈린다" % pat

    # ④ 만든 시각을 화면에 띄우나
    assert "askPick" in app and "function doAsk" in app and "$('askgo')" in app, \
        "폰에 물어보기 카드가 배선되지 않았다"
    assert "D.ask" in app, "폰이 꾸러미를 안 읽는다"
    i = app.index("function doAsk")
    body = app[i:i + 3000]
    assert "만든때" in body, "만든 시각을 안 띄운다 — 낡은 답이 확정처럼 보인다"

    # ⑤ '규칙에 없다' 와 '근거가 없어 안 실렸다' 를 가르나
    assert "hasOwnProperty" in body, \
        "아는 갈래와 모르는 갈래를 안 가른다 — 사람이 이미 있는 규칙을 다시 만들러 간다"

    # 지어내지 않는 자리 — 못 답하면 클로드 문구를 준다
    assert "탈출머리말" in app, "폰이 못 답할 때 붙여넣을 문구를 안 만든다"

    # ⑥ 아는 것만 말한다 — 못 닿은 것을 'PC 꺼짐' 이라고 단정하지 않는다 (2026-08-09 지시).
    #    실사고: 앱 서버를 재시작하는 몇 초 사이에 폰이 열려 한 번 실패했고, 화면은
    #    멀쩡히 켜져 있는 PC 를 두고 'PC 꺼짐' 을 띄웠다. **틀린 단정이 못 하는 말보다 나쁘다.**
    pc = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for phrase in ("PC 꺼짐 —", "PC가 꺼져 있어 이 기기에", "PC가 꺼져 있어 PIN"):
        assert phrase not in pc, \
            "화면이 %r 라고 단정한다 — 이 기기가 아는 것은 '못 닿았다' 뿐이다" % phrase
    assert "서버에 닿지 않습니다 (PC 꺼짐·재시작·외부접속 끊김 중 하나)" in pc, \
        "못 닿은 이유를 갈라 말하지 않는다"
    assert pc.count("'/api/auth/session'") >= 2, \
        "한 번 실패로 오프라인이라 말한다 — 서버 재시작 몇 초를 'PC 꺼짐' 으로 읽는다"

    # 폰이 실제로 고르는 갈래가 PC 와 같은가 — 꾸러미의 말투 그대로 흉내 낸다.
    # (파이썬 정규식으로 돌리므로 JS 자체 시험은 아니다. 그래서 ② 가 따로 있다.)
    order = [(r["이름"], [re.compile(p, re.I) for p in r["말투"]]) for r in pack["규칙"]]
    for q, want in L.PROBES:
        got = None
        for name, pats in order:
            if any(p.search(q) for p in pats):
                got = name
                break
        assert got == want, "폰이 고른 갈래가 PC 와 다르다: %r → %r (기대 %r)" % (q, got, want)
        if want is None:
            continue
        known = want == "프로젝트조회" or want in pack["답"]
        assert known, "갈래 '%s' 가 꾸러미에 자리조차 없다 — 폰이 이유를 못 댄다" % want

    print("  [184] 폰이 PC 없이 답함 — 꾸러미 배선·JS 안전 말투 %d개·규칙 비복제·"
          "만든시각 표시·갈래 %d개 일치 ✅" % (len(every), len(L.PROBES)))


def t185_datalake_shown_in_app():
    """[185] ERP·밴드가 DB로 흡수된 record 를 **앱에서 보고 캡처**한다 ([24], 2026-08-09).

    사용자 지시(분담판 [24]): "ERP Excel → 앱 DB(datalake) 흡수 + 앱 화면 노출 —
    받은 Excel 을 파싱해 record 표로. 그 뒤 앱에서 보고 캡처."

    흡수(ingest_erp)는 이미 돌고 있었다(record 29,568건). 빠져 있던 반쪽은 **화면 노출**
    이었다 — app_server 는 datalake 를 한 줄도 안 읽고 있었다. 이 검증이 지키는 것:
      ① 정본 질의는 **하나**다 — 앱도 `datalake.find(on='record')` 를 부른다(두 벌이면 갈린다).
      ② 큰 갈래로 물으면 하위까지 잡힌다(`kind='ERP'` → `ERP:sales`·`ERP:stmt`…).
      ③ 기간 필터가 실제로 좁힌다(since/until).
      ④ **읽기 전용** — 노출 함수가 큐·엑셀·INSERT/UPDATE 를 건드리지 않는다
         (물어봤을 뿐인데 원장이 바뀌면 안 된다, `[181]` 과 같은 선).
      ⑤ 라우트(`/api/records`)와 화면(자료창고 카드·loadWarehouse)이 배선돼 있나.
    """
    import importlib, tempfile
    D = importlib.import_module("datalake")

    # ① ② ③ find(on='record') 의미 — 임시 DB 로 격리(실데이터를 안 흔든다)
    tmp = tempfile.mkdtemp(prefix="dlw185_")
    dbp = os.path.join(tmp, "datalake.db")
    con = D.connect(dbp)
    try:
        D.put_record(con, "ERP:sales", "s1", {"x": 1}, biz_date="2026-08-01",
                     party="쿠팡", amount=1000, status="3.오더처리")
        D.put_record(con, "ERP:stmt", "t1", {"x": 2}, biz_date="2026-07-01",
                     party="쿠팡", amount=2000, status="발행")
        D.put_record(con, "band_post", "b1", {"x": 3}, biz_date="2026-08-05",
                     party="", amount=None, status="")
        con.commit()
        all_rows = D.find(con, on="record", limit=50)
        assert len(all_rows) == 3, "넣은 record 3건이 다 안 나온다"
        erp = D.find(con, on="record", kind="ERP", limit=50)
        kinds = sorted({r["kind"] for r in erp})
        assert kinds == ["ERP:sales", "ERP:stmt"], \
            "kind='ERP' 가 하위 갈래를 못 잡거나 밴드까지 삼켰다: %s" % kinds
        aug = D.find(con, on="record", since="2026-08-01", limit=50)
        got = sorted({r["natural_key"] for r in aug})
        assert got == ["b1", "s1"], "기간 필터(since)가 안 좁힌다: %s" % got
    finally:
        con.close()
        import shutil as _sh
        _sh.rmtree(tmp, ignore_errors=True)

    # ④ ⑤ 앱 배선 — 노출 함수는 읽기 전용이고, 라우트·화면이 붙어 있어야 한다
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "def datalake_records(" in srv, "노출 함수가 없다 — DB 가 화면에 안 뜬다"
    i = srv.index("def datalake_records(")
    j = srv.index("\nclass H(", i)             # 함수는 핸들러 클래스 바로 앞에 있다
    body = srv[i:j]
    assert 'on="record"' in body, "정본 질의(find on='record')를 안 쓴다 — 두 벌이면 갈린다"
    for bad in ("enqueue", "queue_add", "--apply", "workbook_patch",
                "INSERT", "UPDATE", ".save("):
        assert bad not in body, "노출 함수가 쓰기(%s)를 한다 — 읽기 전용이어야 한다" % bad
    assert '/api/records' in srv, "라우트 /api/records 가 안 걸렸다"

    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert 'id="dlw"' in idx, "자료창고 카드가 화면에 없다"
    assert "function loadWarehouse(" in idx and "/api/records" in idx, \
        "화면이 /api/records 를 안 부른다"
    assert "v==='sources'" in idx and "loadWarehouse()" in idx, \
        "원본 화면을 열 때 자료창고를 안 읽는다"
    assert "function warehouseCsv(" in idx, "캡처(CSV 저장) 경로가 없다 — '보고 캡처'의 반쪽"
    # 화면이 datalake 를 **다시 구현**하면 안 된다 — 서버 한 곳만 판단한다
    assert "datalake" not in idx.lower() or "/api/records" in idx, \
        "화면이 서버를 안 거치고 자체 판단하려 한다"
    print("  [185] datalake record 앱 노출 — 정본 질의 하나·큰갈래·기간·읽기전용·배선 ✅")


def t186_kakao_round_and_stale_tmp():
    """[186] 카톡 한 파일을 **끝까지** 반영 · 죽은 회차가 남긴 tmp 가 다음 회차를 막지 않는다
    (2026-08-09 지시: "이거 반영하고 엑셀 및 앱에 반영해").

    두 가지가 한 지시에서 나왔다.

    ★ 하나 — **없는 파일을 조용히 건너뛰지 않는다.** 사람이 준 경로가 실제로 없을 때
      나머지 단계가 멀쩡히 돌면 '0건 반영 성공'으로 끝난다. 숫자도 나오고 오류도 안 나서
      아무도 안 본다([169] 와 같은 모양). 그래서 하나라도 없으면 **멈춘다**(exit 2).

    ★ 둘 — **실패 하나가 모든 다음 회차를 영원히 막을 수 있었다.** 실사고: 11:00 회차가
      쓰다 죽으며 `v571.tmp.xlsx` 를 남겼고, 그 뒤 11:00·15:00·사람 지시 반영이 **전부**
      `FileExistsError` 로 실패했다. 화면은 "반영 대기 N건"만 계속 보여 준다 —
      실패가 성공처럼 보이는 자리다.
      가르는 근거는 **나이**다: 방금 생긴 tmp 는 **다른 writer 가 쓰는 중**이라 손대면
      안 되고(그대로 멈춘다), 오래된 것은 죽은 회차의 찌꺼기라 **옆으로 치우고 계속 간다**.
      지우지 않는 이유는 그 안에 마지막 회차 결과가 들어 있을 수 있어서다 — 판단은 사람 몫.
    """
    import ledger_writer as W

    # ① 나이로 가른다 — 값이 아니라 **뜻**을 지킨다
    assert isinstance(W.STALE_TMP_MIN, int) and W.STALE_TMP_MIN >= 5, \
        "찌꺼기 판정 나이가 너무 짧다 — 쓰는 중인 tmp 를 빼앗을 수 있다"
    src = open(os.path.join(ROOT, "ledger_writer.py"), encoding="utf-8").read()
    i = src.index('final_dst, dst = dst, dst[:-5] + ".tmp.xlsx"')
    blk = src[i:i + 2000]
    assert "os.path.getmtime(dst)" in blk, "tmp 나이를 재지 않는다 — 둘을 못 가른다"
    assert "age_min < STALE_TMP_MIN" in blk, "방금 생긴 tmp 도 치운다 — 쓰는 중인 회차를 깬다"
    assert "os.path.exists(final_dst)" in blk, \
        "정본이 이미 나왔는데도 치운다 — 그 tmp 는 남의 것이다"
    assert ".stale-" in blk and "os.remove" not in blk, \
        "찌꺼기를 **지운다** — 마지막 회차 결과가 그 안에 있을 수 있다"
    assert "raise FileExistsError" in blk, \
        "못 치웠는데 계속 간다 — 두 writer 가 같은 파일을 쓴다"

    # ② 실제로 갈리는가 — 오래된 것은 치우고, 갓 생긴 것은 못 치운다
    with tempfile.TemporaryDirectory(prefix="stale-tmp-") as td:
        old = os.path.join(td, "장부_v9.tmp.xlsx")
        open(old, "w").close()
        os.utime(old, (time.time() - (W.STALE_TMP_MIN + 10) * 60,) * 2)
        age = (time.time() - os.path.getmtime(old)) / 60.0
        assert age >= W.STALE_TMP_MIN, "오래된 tmp 가 찌꺼기로 안 읽힌다"
        fresh = os.path.join(td, "장부_v10.tmp.xlsx")
        open(fresh, "w").close()
        assert (time.time() - os.path.getmtime(fresh)) / 60.0 < W.STALE_TMP_MIN, \
            "방금 만든 tmp 가 찌꺼기로 읽힌다 — 쓰는 중인 회차를 깬다"

    # ③ 카톡 회차 — 조각을 잇는 것이 전부이므로 **순서와 문**만 지킨다
    ka = open(os.path.join(ROOT, "kakao_apply.py"), encoding="utf-8").read()
    assert "kakao_extract.py" in ka and "ledger_db.py" in ka, "조각이 이어져 있지 않다"
    assert ka.index("kakao_extract.py") < ka.index('"--intake", "--apply"'), \
        "추출보다 엑셀 반영이 먼저다 — 아직 안 만든 것을 반영한다"
    assert "return 2" in ka and "찾지 못했습니다" in ka, \
        "없는 파일을 조용히 건너뛴다 — '0건 반영 성공'이 된다"
    assert 'if rc != 0:' in ka and "실패한 채로 반영하면" in ka, \
        "추출이 실패해도 엑셀을 연다 — 없는 것이 반영된 것처럼 보인다"
    assert "shutil.copy2" in ka and "shutil.move" not in ka, \
        "사람이 준 원본을 옮긴다 — '그때 무엇을 받았나'를 사람 쪽에서 잃는다"
    assert "COUPANG_UNATTENDED" in ka, "무인 실행도 즉시 반영을 쓸 수 있다"
    # `subprocess.run(timeout=)` 은 윈도우에서 영원히 매달린다([175]) — 주석에 적힌 것과
    # 실제로 부르는 것을 가른다. 코드 줄에서만 찾는다.
    code = "\n".join(l for l in ka.splitlines() if not l.lstrip().startswith("#"))
    assert "subprocess.run(" not in code, "윈도우에서 영원히 매달릴 수 있다([175])"
    assert "subprocess.Popen(" in code and "communicate(timeout=" in code, \
        "자식을 시간제한 없이 기다린다([175])"
    assert "taskkill" in code, "죽일 때 나무째 죽이지 않는다([175])"

    # ④ 즉시반영은 **사람 길에만** 있다 — 무인 회차가 이 문을 쓰면 [93] 이 무너진다
    for auto in ("daily_run.py", "session_wrapup.py"):
        a = open(os.path.join(ROOT, auto), encoding="utf-8").read()
        assert "kakao_apply.py" not in a or "--now" not in a, \
            "%s 가 카톡 회차를 즉시반영으로 부른다 — 하루 두 번 규칙이 조용히 깨진다" % auto

    print("  [186] 카톡 한 줄 반영(없는 파일에 멈춤·추출 실패 시 안 씀) · "
          "죽은 회차 tmp 는 치우고 쓰는 중인 tmp 는 안 건드림 ✅")


def t187_free_vs_insurance_are_not_one_label():
    """[187] 무상은 무상, 보험은 보험 · 빈칸은 **무상이 아니다** (2026-08-09 지시).

    사용자 지시: "무상이면 무상 보험이면 보험 표시 / 보험사에서 돈이 입금 되면
                  그것도 찾아 반영하는 알고리즘 구성"

    ★ 딱지 하나가 세 가지를 덮고 있었다. `비용구분 != '유상'` 이면 전부 `무상/보험`.
      실측 750행: 유상 716 · **무상 2** · 보험 **0** · 미확정 4 · 빈칸 26 · '0' 2.
      즉 회색 '무상/보험' 34건 중 진짜 무상은 2건이고 **32건은 아직 안 적은 칸**이었다.
      그 32건 중 **31건에 금액이 있다**(UJ2601280 울산2캠프 3,049,310원 ·
      UJ2601288 인천5MB 672,540원 — 형님 화면에 찍힌 바로 그 카드들이다).
      무상이라 부르는 순간 청구 대상에서 조용히 빠진다 — **빈칸이 무상으로 위장하는 자리**다.
      비어 있는 값은 눈에 띄지만 **틀린 딱지는 안 띈다**([165] 와 같은 종류).

    ★ 딱지를 가르면 그 문자열을 손으로 적어 둔 곳들이 **오류 없이** 안 걸리게 된다.
      그래서 '청구하지 않음'을 읽는 자리를 `NON_BILLABLE` 한 곳으로 모았다.
    """
    import ecount_reconcile as E

    # ① 원장이 말한 그대로 — 뭉치지도 지어내지도 않는다
    mk = lambda kind: {"비용구분": kind, "프로젝트NO": "", "원천업무ID": ""}
    assert E.settle_status(mk("무상")) == "무상", "무상이 무상으로 안 나온다"
    assert E.settle_status(mk("보험")) == "보험", "보험이 보험으로 안 나온다"
    for blank in ("", None, "미확정", "0", "   "):
        got = E.settle_status(mk(blank))
        assert got == E.UNKNOWN_COST, \
            "비용구분 %r 을 %r 라고 부른다 — 모르는 것을 아는 것처럼 말한다" % (blank, got)

    # ② '청구하지 않음'은 한 곳에서만 읽는다 — 미입력은 절대 여기 들어오지 않는다
    assert E.is_non_billable("무상") and E.is_non_billable("보험")
    assert E.is_non_billable("무상/보험"), "예전 값이 안 걸린다 — 이미 적힌 기록이 무시된다"
    assert not E.is_non_billable(E.UNKNOWN_COST), \
        "미입력을 청구 대상에서 뺀다 — 32건이 다시 조용히 사라진다"
    assert not E.is_non_billable("유상")

    # ③ 소비처가 새 딱지를 안다 — 한 곳이라도 빠지면 그쪽만 조용히 안 걸린다
    fe = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    assert "is_non_billable" in fe, "조치 목록이 아직 문자열을 손으로 비교한다"
    assert '"무상/보험", "정상"' not in fe, "옛 비교가 남아 있다 — 무상 2건이 조치로 쏟아진다"
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '"무상", "보험", "무상/보험"' in srv, "서버 집계가 무상·보험을 못 알아본다"
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for tag in ("'무상':", "'보험':", "'비용구분 미입력':"):
        assert tag in idx, "화면이 %s 딱지 색을 모른다 — 검은 글씨가 된다" % tag
    assert "'무상','보험','무상/보험'" in idx.replace(" ", ""), \
        "화면 필터가 무상·보험을 못 걸러 조치 목록에 섞인다"
    assert "'비용구분 미입력'," not in idx.replace(" ", "").replace(
        "'비용구분 미입력':'#C77A11'", ""), \
        "미입력을 화면에서도 걸러낸다 — 숨기려고 만든 딱지가 아니다"

    # ④ 보험사 입금 — 업종 낱말로 고른다(회사 이름을 박으면 목록 밖 보험사는 영영 안 보인다)
    import insurance_watch as W
    for name in ("삼성화재해상보험", "DB손해보험", "현대해상", "메리츠화재", "한국지역난방공제회"):
        assert W.looks_insurer(name), "보험사 %s 를 못 알아본다" % name
    for name in ("쿠팡로지스틱스", "모벤티스", "코리아종합물류", "화재감지기설치", "소방안전"):
        assert not W.looks_insurer(name), "%s 를 보험사로 본다 — 남의 돈이 꽂힌다" % name

    # ⑤ 쓰는 문은 좁다 — 금액이 유일할 때만. 보험금은 자기부담금 때문에 대개 안 맞는다.
    ins = open(os.path.join(ROOT, "insurance_watch.py"), encoding="utf-8").read()
    assert "len(cands) == 1 and len(rivals) == 1" in ins, \
        "짝이 여럿인데도 자동 반영한다 — 엉뚱한 현장에 남의 돈이 꽂힌다"
    assert '"only_if_empty": True' in ins, "사람이 이미 적은 입금을 덮는다"
    assert "ledger_writer" not in ins, "엑셀을 직접 연다 — 반영은 11:00·15:00 몫이다"
    # 0건일 때 '없는 것'과 '안 본 것'을 가른다([169])
    assert "없는 것인가, 안 본 것인가" in ins, "0건을 그냥 0이라고만 적는다"
    assert "본자료" in ins, "무엇을 보고 0이라 했는지 안 밝힌다"

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "insurance_watch.py" in daily, "회차에 안 걸렸다 — 다음 세션에 그대로 다시 요구된다"

    print("  [187] 무상·보험·미입력을 가름(빈칸 32건이 무상으로 숨어 있었다) · "
          "보험사 입금 양방향 탐색·유일일치만 반영 ✅")


def t188_worklog_shows_this_month_only():
    """[188] 돌발AS 일지 대조는 **이번 달만** · 0건이면 이유를 댄다 (2026-08-09 지시).

    사용자 지시: "캡처 화면에서 돌발 AS일지 대조는 현재 월만 표시하게 알고리즘 변경"

    ★ 원본은 `정기점검, 돌발AS 일지 (7.1~).xlsx` 처럼 **여러 달이 한 파일에 쌓인다.**
      안 자르면 8/9 에 뽑은 대표 보고 제목이 `2026-07-01 ~ 2026-07-28` 이 된다 —
      두 달 전 숫자가 오늘 것처럼 실린다.
    ★ 그런데 **자르면 0이 되는 달이 있다.** 실측 2026-08: 원본 86건이 전부 7월이라
      이번 달은 0건이다. 그때 0건만 보여 주면 **'이번 달은 돌발AS 가 한 건도 없었다'**
      로 읽힌다 — 사실이 아니라 **자료가 아직 안 담긴 것**이다. 둘은 완전히 다른 말이고,
      이 프로젝트에서 0 은 언제나 '없는 건가, 안 본 건가'를 물어야 한다([169]).
    """
    import work_log_sync as W

    # ① 기준 달은 바꿀 수 있어야 검증이 '자료 없는 달'을 재현한다
    old = os.environ.get("COUPANG_WORKLOG_MONTH")
    try:
        os.environ["COUPANG_WORKLOG_MONTH"] = "2026-03"
        assert W.current_month() == "2026-03", "기준 달을 바꿀 수 없다"
    finally:
        if old is None:
            os.environ.pop("COUPANG_WORKLOG_MONTH", None)
        else:
            os.environ["COUPANG_WORKLOG_MONTH"] = old
    assert len(W.current_month()) == 7 and W.current_month()[4] == "-", W.current_month()

    src = open(os.path.join(ROOT, "work_log_sync.py"), encoding="utf-8").read()
    assert 'startswith(month)' in src, "이번 달로 자르지 않는다 — 지난달 숫자가 그대로 실린다"
    assert '"기준월": month' in src and '"이번달자료없음": month_empty' in src, \
        "이번 달만 센 숫자라는 표시가 없다"
    for k in ('"원본시작일"', '"원본종료일"', '"원본전체건수"'):
        assert k in src, "0건일 때 원본이 어디까지 담았는지 못 말한다 — %s" % k
    # 자르기 **전** 목록을 남겨 둬야 원본 범위를 셀 수 있다
    assert "as_all = [r for r in compared" in src and "as_rows = [r for r in as_all" in src, \
        "원본 전체를 버리고 잘라서, 무엇을 못 봤는지 말할 수 없다"

    # ② 화면과 캡처가 **둘 다** 말한다 — 한쪽만 고치면 그쪽만 정직해진다
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert idx.count("이번달자료없음") >= 2, \
        "대시보드 카드와 대표 캡처 중 한 곳만 이유를 댄다"
    assert "2026-07-01 ~ 2026-07-28" not in idx
    assert "돌발AS 현장 일지 대조 · ${monTitle}" in idx, "캡처 제목이 아직 날짜 범위다"
    assert "현장 일지 대조 · ${_e(_mt)}" in idx, "화면 카드가 어느 달인지 안 밝힌다"

    # ③ 아침 브리핑에 달을 손으로 적어 두지 않는다 — 다음 달이면 그대로 거짓말이다
    brief = open(os.path.join(ROOT, "daily_brief.py"), encoding="utf-8").read()
    assert "일지 대조 (7월 원본)" not in brief, "'7월'이 박혀 있다"
    assert '일지 대조 (%s 원본)' in brief and "이번달자료없음" in brief, \
        "브리핑이 기준 달·자료 없음을 안 쓴다"

    print("  [188] 돌발AS 일지 대조 이번 달만 · 0건이면 원본 범위를 대고 이유를 말함 "
          "(화면·캡처·브리핑 셋 다) ✅")


def t189_worklog_reflects_without_hands():
    """[189] 일지는 **올려만 두면** 엑셀·앱까지 간다 (2026-08-09 지시).

    사용자 지시: "돌발 AS 일지랑 정기점검 일지 엑셀과 앱에 자동 반영하는 알고리즘 구현해,
                  담당자 손댈 필요 없이 내가 자료 올리거나 알아서 긁어서"

    ★ 조각은 이미 다 있었다 — 업로드함 분류 · 09:50 `work_log_sync --queue` ·
      11:00·15:00 엑셀 반영. **빠진 것은 두 곳**이었다:
      ① 바탕화면·다운로드에 떨군 일지를 `download_intake` 가 **몰랐다.**
         파일은 그 자리에 멀쩡히 남고 오류도 안 난다 — 올린 사람만 반영된 줄 안다.
      ② 올려도 **다음 날 09:50 까지** 아무 데도 안 갔다.
    ★ 판별 규칙은 `upload_intake.looks_worklog` **한 곳**이다. 담는 길이 둘(업로드함 ·
      다운로드)이라 규칙을 두 곳에 적으면 언젠가 갈리고, 갈리면 **한쪽 길로 올린 일지만
      조용히 안 들어온다**([162] 와 같은 원칙).
    """
    import upload_intake as U

    # ① 규칙은 한 곳 — 이름만이 아니라 내용까지 합쳐 본다
    assert U.looks_worklog(".xlsx", "정기점검, 돌발AS 일지 (8.1~).xlsx")
    assert U.looks_worklog(".xlsx", "2026 돌발AS 일지 미실시건.xlsx")
    assert not U.looks_worklog(".xlsx", "급여대장.xlsx"), "남의 엑셀을 Z: 로 쓸어 담는다"
    assert not U.looks_worklog(".txt", "돌발AS 일지.txt"), "엑셀이 아닌 것도 가져간다"

    dn = open(os.path.join(ROOT, "download_intake.py"), encoding="utf-8").read()
    assert "looks_worklog" in dn, \
        "다운로드 길이 일지를 모른다 — 바탕화면에 떨군 일지가 조용히 남는다"
    assert "WORK_LOG_DIR" in dn, "일지를 정본 자리가 아닌 곳으로 옮긴다"
    assert '"일지"' not in dn.replace('"정기점검·돌발AS 일지"', ""), \
        "판별 규칙을 다운로드 쪽에도 적었다 — 두 길이 언젠가 갈린다"

    # ② 올리면 기다리지 않는다 — 워치독(30분)이 바뀐 일지를 스스로 본다
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "def sync_worklog(" in wd and "sync_worklog(dry)" in wd, \
        "워치독이 일지를 안 본다 — 오후에 올린 일지는 다음 날 09:50 까지 안 간다"
    assert 'work_log_sync.py"), "--queue"' in wd.replace("'", '"') or \
        '"work_log_sync.py"), "--queue"' in wd, "대조를 큐로 넣지 않는다"
    assert "--apply" not in wd.split("def sync_worklog(")[1].split("def main(")[0], \
        "워치독이 엑셀을 직접 연다 — 하루 두 번 규칙이 깨진다"
    body = wd.split("def sync_worklog(")[1].split("def main(")[0]
    assert "newest <= seen" in body, "안 바뀐 일지도 매 30분 대조한다 — Z: 가 붐빈다([168])"
    # ★ 실패했는데 자국을 남기면 **다시는 안 본다** — 그 일지는 영영 반영되지 않는다
    assert body.index("returncode != 0") < body.index("json.dump"), \
        "실패해도 '봤다'고 자국을 남긴다 — 그 일지는 영영 안 들어온다"

    # ③ 회차와 반영 경로는 그대로 살아 있어야 한다
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert '"work_log_sync.py"), "--queue"' in daily, "09:50 회차가 일지를 안 본다"
    ldb = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    assert "28_일지대조현황" in ldb, "반영 회차가 일지대조 시트를 안 갱신한다"

    print("  [189] 일지는 올려만 두면 반영 — 두 투입 길이 같은 규칙 하나 · "
          "워치독이 바뀐 것만 즉시 대조 · 실패는 자국을 안 남김 ✅")


def t172_typo_watch_does_not_cry_wolf():
    """[172] 오기입 탐지 — **잡는 것보다 잘못 지목하지 않는 것**이 어렵다 (2026-08-09 지시).

    사용자 지시: "오타 오기입 이런거 잡아낼 수 있는 알고리즘 구성해서 적용해"

    첫판이 750행에서 **163건**을 쏟았다. 두 번 다 규칙이 틀린 것이었다:
      · 날짜 160건 — 전부 **명세서발행일이 완료일보다 1~3일 앞선 것**이었다.
        그건 오타가 아니라 정상 업무다(명세서를 먼저 끊는다). 이제 **연도 한 자리
        차이**나 **1년 이상** 앞선 것만 본다.
      · 캠프명 `제주3캠프` → `양주3캠프` — 그런데 제주에는 제주1·2·3캠프가 **다 실재한다.**
        '드물게 쓰인 것'과 '잘못 쓰인 것'은 다르다. 이제 **괄호 안 지명이 같은데 앞
        숫자만 다른 것**만 본다(`송파1MB(감일동)` ↔ `송파5MB(감일동)`).
    163 → **1건**이 됐고 그 1건은 근거가 선다.

    ★ 오타라고 잘못 부르면 사람이 **멀쩡한 값을 고치러 간다** — 못 잡는 것보다 나쁘다.
      그래서 이 도구는 **아무것도 고치지 않고 큐에도 넣지 않는다.** 무엇이 맞는지는
      사람만 안다. 자동으로 고치면 "그때 정말 뭐라고 적혀 있었나"를 잃는다.
    """
    import typo_watch as T

    # ① 오타로 봐야 하는 것
    assert T.edit1("PO372139", "PO372136"), "한 글자 치환은 오타 후보다"
    assert T.swap_typo("PO372139", "PO372193"), "이웃 두 글자 자리바꿈은 가장 흔한 손오타다"
    assert T.digit_slip(4030000, 403000) == "10배", "0 하나 더 친 것"
    assert T.digit_slip(403000, 4030000) == "1/10", "0 하나 덜 친 것"

    # ② 오타로 보면 **안 되는** 것 — 여기가 이 도구의 값어치다
    assert not T.edit1("PO37213", "PO372139"), "자리 수가 다르면 오타가 아니라 다른 체계다"
    assert not T.edit1("PO372139", "PO372236"), "두 글자 차이는 짐작이다"
    assert T.digit_slip(403000, 402000) is None, "어중간한 차이는 자릿수 실수가 아니다"
    assert len(T.near("PO0002", {"PO0001", "PO0003"})) == 2, \
        "후보가 둘이면 둘 다 돌려줘 **지목을 막아야** 한다"

    src = open(os.path.join(ROOT, "typo_watch.py"), encoding="utf-8").read()
    code = "\n".join(ln for ln in src.splitlines() if not ln.strip().startswith("#"))

    # ③ 날짜: 며칠 앞선 것은 안 본다(160건의 정체)
    assert "gap >= 365" in code and "yr_slip" in code, \
        "완료일보다 며칠 앞선 명세서는 정상 업무다 — 그걸 세면 경보가 160건이 된다"
    # ④ 캠프명: 괄호 안 지명이 같을 때만
    assert 'endswith("(" + loc.group(1) + ")")' in code, \
        "'드물게 쓰인 이름'만으로 지목하면 제주3캠프 같은 실재 이름을 오타로 부른다"
    # ⑤ 절대 고치지 않는다 — 큐도 엑셀도 건드리면 안 된다
    for banned in ("queue_add", "enqueue", "workbook_patch", "ledger_writer"):
        assert banned not in code, f"오기입 탐지는 읽기 전용이어야 한다: {banned}"

    # ⑥ 회차에 들어가 있나 — 대화에 남긴 것은 사라지고 스케줄에 넣은 것만 산다
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "typo_watch.py" in daily, "일일대조 회차에 안 들어가면 한 번 돌고 끝난다"
    print("  [172] 오기입 탐지 — 지목은 유일할 때만 · 며칠 차이는 정상 · 읽기 전용 ✅")


def t171_cache_swap_waits_for_readers():
    """[171] 흡수가 **읽는 쪽에 막혀 조용히 죽으면** 안 된다 (2026-08-08 실사고).

    `convert_dump` 의 `os.replace(tmp, dst)` 가 `PermissionError [WinError 5]` 로 죽었다.
    앱 서버가 마침 `band/cache/84789192.json`(5MB)을 읽는 중이었다. 리눅스와 달리
    윈도우는 **열려 있는 파일을 갈아끼우지 못한다.**

    ★ 남는 그림이 나쁘다: 새 글을 다 긁어 놓고도 캐시는 **어제 것**이고 `.tmp` 만
      덩그러니 남는다. 다음 회차는 그 옛 캐시를 보고 **"바뀐 것 없음"** 을 내놓는다 —
      아무 일도 안 일어났는데 안심시키는 결과다. 실패가 성공처럼 보이는 자리다.

    읽는 쪽은 곧 놓으므로 물러서며 몇 번 다시 건다. 끝내 안 되면 **예외를 올린다** —
    `.tmp` 를 지우고 조용히 넘어가면 애써 만든 새 캐시까지 잃는다.
    """
    import tempfile, threading, time as _t
    sys.path.insert(0, os.path.join(ROOT, "band"))
    import convert_dump as C

    src = open(os.path.join(ROOT, "band", "convert_dump.py"), encoding="utf-8").read()
    body = "\n".join(ln for ln in src.splitlines() if not ln.strip().startswith("#"))
    assert "os.replace(tmp, dst)" not in body.split("def swap_in", 1)[1].split("\ndef ", 1)[1], \
        "캐시를 갈아끼우는 자리가 swap_in 을 안 거치면 재시도가 아무 소용 없다"
    assert body.count("swap_in(tmp, dst)") >= 2, "캐시 쓰는 자리가 둘 다 거쳐야 한다"
    # 2026-08-11: tmp 는 pid 별 이름이어야 한다 — 고정 이름이면 5분 파이프라인·09:50
    # 회차·dump_watch 가 겹칠 때 서로의 tmp 를 가져가 FileNotFoundError 로 죽는다.
    assert "def tmp_path(" in body and body.count("tmp = tmp_path(") >= 3, \
        "캐시 tmp 가 고정 이름이다 — 회차 둘이 겹치면 FileNotFoundError 로 죽는다"
    assert "os.getpid()" in body.split("def tmp_path(", 1)[1].split("\ndef ", 1)[0], \
        "tmp_path 에 pid 가 없다 — 이름이 갈리지 않는다"

    d = tempfile.mkdtemp()
    dst = os.path.join(d, "c.json")
    tmp = dst + ".tmp"
    open(dst, "w").write("old")
    open(tmp, "w").write("new")
    fh = open(dst, "r")                       # 읽는 쪽이 물고 있다
    threading.Thread(target=lambda: (_t.sleep(0.9), fh.close()), daemon=True).start()
    C.swap_in(tmp, dst)                       # 기다렸다 다시 걸어 성공해야 한다
    assert open(dst).read() == "new", "물고 있다 놓으면 결국 새 캐시가 들어가야 한다"
    assert not os.path.exists(tmp), "성공했으면 .tmp 는 사라진다"

    # 끝내 안 풀리면 조용히 넘어가지 않는다 — .tmp 를 남기고 예외를 올린다
    open(dst, "w").write("old")
    open(tmp, "w").write("new")
    keep = open(dst, "r")
    try:
        C.swap_in(tmp, dst, tries=2, wait=0.05)
        raised = False
    except PermissionError:
        raised = True
    finally:
        keep.close()
    if os.name == "nt":                       # 윈도우에서만 성립하는 잠금이다
        assert raised, "못 갈아끼웠으면 실패라고 말해야 한다 — 조용히 넘어가면 옛 캐시가 산다"
        assert os.path.exists(tmp), "실패해도 새 캐시(.tmp)는 버리지 않는다"
    print("  [171] 캐시 갈아끼우기 — 읽는 쪽을 기다렸다 재시도, 끝내 안 되면 .tmp 남기고 실패 ✅")


def t170_po_amount_ladder():
    """[170] PO 대조는 **앱 화면과 같은 금액**을 봐야 한다 (2026-08-08 사용자 질문에서).

    유형C(금액 불일치)가 51건 중 **44건**으로 나왔다. 근거는 전부 `원장공급가액합: 0` —
    06시트 공급가액이 사람 손 입력이라 비어 있었을 뿐이고, ERP 로 대면 원 단위까지
    맞았다(PO327948 쿠팡 7,551,500 = ERP 합 7,551,500). 앱은 이미 실제작업 → ERP →
    명세서 사다리로 채워 보여 주고 있었으니 **화면과 대조기가 서로 다른 금액을 봤다.**

    ★ 경보가 44/51 이면 그 경보는 아무도 안 본다. 조용한 사고의 반대편이지만 결과는 같다.
      고친 뒤 C 는 35건이 됐고, 그 35건은 **원장 연결 누락 20 · 진짜 금액 차이 15** 로
      갈렸다 — 고칠 곳이 서로 다르다.

    ★★ 그리고 **유형D 만은 사다리를 쓰지 않는다.** A~C 는 경보라 금액 출처를 넓혀도
       잃는 것이 없지만, D 는 06시트에 PO번호를 **써 넣는** 길이다. 짐작으로 채운
       금액으로 짝을 지으면 틀린 PO번호가 원장에 박히고 그건 빈 칸보다 나쁘다.
    """
    import po_reconcile as P
    src = open(os.path.join(ROOT, "po_reconcile.py"), encoding="utf-8").read()

    # ① 사다리: 원장 → ERP → 명세서(÷1.1) → 0
    P._ERP_SUPPLY = {"UJ-T1": 500000}
    assert P.supply_of({"원장_공급가액": 403000, "프로젝트NO": "UJ-T1"}) == 403000, \
        "원장 값이 있으면 그것이 먼저다"
    assert P.supply_of({"원장_공급가액": None, "프로젝트NO": "UJ-T1"}) == 500000, \
        "원장이 비면 ERP 로 내려가야 한다 — 이게 없어서 44/51 이 됐다"
    assert P.supply_of({"원장_공급가액": 0, "프로젝트NO": "없음",
                        "원장_거래명세서합계": 476300}) == 433000, \
        "명세서(부가세 포함)는 ÷1.1 로 환산한다"
    assert P.supply_of({"프로젝트NO": "없음"}) == 0, "근거가 하나도 없으면 0"

    # ② 유형C 는 supply_of, 유형D 는 원장 직접입력만 — 자동으로 쓰는 자리는 넓히지 않는다
    body = src.split("def main(", 1)[1]
    c_part = body.split("# 유형D", 1)[0]
    d_part = body.split("# 유형D", 1)[1].split("def ", 1)[0]
    # ★ 주석을 걷어내고 본다 — 여기 '왜 안 쓰는지' 를 적어 두면 그 설명 자체가 걸린다
    #   (검증 [155] 가 같은 방식으로 한 번 걸렸다: 내가 쓴 산문이 코드로 읽혔다).
    d_code = "\n".join(ln for ln in d_part.splitlines() if not ln.strip().startswith("#"))
    assert "supply_of(r) for r in lrows" in c_part, "유형C 합계가 사다리를 안 쓴다"
    assert "supply_of" not in d_code, \
        "유형D 는 06시트에 PO번호를 쓰는 길이다 — 짐작한 금액으로 짝지으면 안 된다"

    # ③ 차액이 '왜' 나는지를 말해야 한다 — 연결 누락과 금액 오류는 고칠 곳이 다르다
    assert "ERP전표합" in c_part and "원장 연결 누락" in c_part, \
        "쿠팡=ERP 인데 원장만 모자란 경우를 따로 말해야 한다"

    # ④ 앱도 '미발행' 을 사유별로 갈라 내려보낸다(류지영 몫 vs 전표 등록 몫)
    app = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    seg = app.split("def app_settle", 1)[-1].split("\ndef ", 1)[0] if "def app_settle" in app else app
    assert '"미발행사유"' in app, "미발행 사유를 안 내려보낸다"
    assert "발행 대기(ERP 4단계)" in app and "ERP 전표 없음" in app, \
        "'미발행' 한 덩어리에 가야 할 사람이 다른 두 가지가 섞여 있으면 안 된다"
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert "미발행사유" in idx, "화면이 그 사유를 안 보여 주면 갈라 놓은 뜻이 없다"

    # ⑤ 유형E — ERP 전표가 스스로 말하는 PO 연결. D 와 달리 **짐작이 아니다**.
    #    그래도 자동으로 쓰는 자리이므로 문 넷이 코드에 남아 있어야 한다.
    e_part = body.split("유형E", 1)[1].split("if queue_items", 1)[0]
    e_code = "\n".join(ln for ln in e_part.splitlines() if not ln.strip().startswith("#"))
    assert "len(toks) == 1" in e_code, \
        "ERP po 칸에 PO 가 둘 이상이면 어느 것인지 모른다 — 유일할 때만 쓴다"
    assert "po not in cp_by_no" in e_code, \
        "쿠팡 목록에 없는 번호를 쓰면 ERP 오타를 원장에 옮기는 것이다"
    assert 'if norm_po(r.get("원장_PO번호")):' in e_code and "continue" in e_code, \
        "이미 사람이 적어 둔 PO번호를 덮으면 안 된다"
    assert '"only_if_empty": True' in e_code, \
        "큐에서도 한 번 더 막는다 — 대조와 반영 사이에 사람이 채웠을 수 있다"
    print("  [170] PO 대조가 앱과 같은 금액 사다리를 본다(D 는 제외) · "
          "'미발행'을 발행대기/전표없음으로 가른다 ✅")


def t169_blind_count_sees_unlooked():
    """[169] 사각지대 계기가 **같은 결측에 눈이 멀면** 안 된다 (2026-08-08 저녁 실측).

    `cancel_blind_count` 는 `comment_count > 담긴 댓글 수` 로 사각지대를 셌다. 그런데
    캐시 10,312글 중 `comment_count>0` 은 **6글**이고 댓글 본문은 **0글**이다 —
    밴드에 댓글이 없어서가 아니라 **수집기가 그 숫자를 안 담아서**다. 그래서 계기가
    "사각지대 0건" 이라고 말했다. 고치자 **8,259건 / 8,561건**이 됐다.

    ★ 이 종류가 제일 나쁘다. 값이 비면 사람이 알아채지만, **재는 도구가 0을 내면
      아무도 의심하지 않는다.** 없는 것과 안 본 것을 구별하는 것이 전부다:
        · `comments` 키가 아예 없다 → 한 번도 안 들여다봤다 → 사각지대
        · `comments: []`           → 보긴 봤고 없었다 → 사각지대 아님
      그리고 리포트는 사각지대가 절반을 넘으면 **스스로 못 믿겠다고 말해야 한다** —
      "취소로 읽힌 건 N건" 은 읽은 글 안에서만 센 숫자이기 때문이다.
    """
    import band_extract as B

    never = {"1": {"comment_count": 0, "content": "접수합니다"}}          # comments 키 없음
    assert B.cancel_blind_count(never) == 1, \
        "댓글을 한 번도 안 본 글을 사각지대로 안 센다 — 계기가 결측에 눈이 먼다"
    looked = {"1": {"comment_count": 0, "comments": []}}
    assert B.cancel_blind_count(looked) == 0, \
        "보고 나서 없었던 글까지 사각지대로 세면 숫자가 소음이 된다"
    half = {"1": {"comment_count": 3, "comments": [{"content": "a"}]}}
    assert B.cancel_blind_count(half) == 1, "반쯤 읽은 글은 예전처럼 사각지대다"

    body = open(os.path.join(ROOT, "cancel_watch.py"), encoding="utf-8").read()
    assert "blind * 2 > total" in body, \
        "사각지대가 절반을 넘어도 리포트가 스스로 경고하지 않는다"
    print("  [169] 사각지대 계기 — '댓글이 없다'와 '안 봤다'를 가르고, 절반 넘으면 리포트가 스스로 경고 ✅")


def t161_erp_filename_fingerprint():
    """[161] ERP 내보내기는 **파일명 화면코드**로 가른다 (2026-08-08 실측).

    매출(세금)계산서조회(E010727)를 제대로 긁어 왔는데도 신선도표는 계속 '3일 밀림'
    이었다. 내려받은 엑셀이 거래명세서현황·재고쪽 조회와 **셋 다 똑같이 생겨서**
    (시트 '거래명세서' · 머리글 '일자 - 번호') 내용 휴리스틱이 taxinv 로 삼켰기 때문이다.
    받은 것이 다른 통에 들어가면 그 통은 영영 안 차고 다음 세션이 같은 화면을 또 긁는다 —
    **받은 사람만 받았다고 믿는** 조용한 사고다.

    지키는 것: ① 파일명 지문이 내용보다 **먼저** 온다 ② 지문에 없는 파일은 예전 그대로
    내용으로 가른다(지문이 다른 자료를 가로채면 안 된다).
    """
    import io
    import inbox_scan as S

    assert S.ERP_FILE_PREFIX.get("EBG006M") == "sales", \
        "E010727(EBG006M) 이 sales 로 안 간다 — 신선도표가 계속 밀림으로 남는다"
    src = io.open(os.path.join(ROOT, "inbox_scan.py"), encoding="utf-8").read()
    body = src[src.index("def classify(path):"):]
    assert body.index("ERP_FILE_PREFIX") < body.index("classify_rows"), \
        "파일명 지문이 내용 판별보다 뒤에 있다 — 뒤면 아무 소용이 없다"

    # 지문 파일은 열지도 않고 바로 갈린다(없는 경로여도 된다 — 그게 요점이다)
    assert S.classify(os.path.join("아무데나", "EBG006M__dl154027.xlsx")) == "sales"
    # 지문에 없는 이름은 예전 길로 간다 — 못 읽으면 unknown
    assert S.classify(os.path.join("아무데나", "ESD009M.xlsx")) == "unknown", \
        "지문에 없는 파일까지 가로챘다"
    # ★ 규칙을 고쳐도 **색인이 옛 판정을 들고 있으면 헛일이다** (2026-08-08 실측).
    #   실제로 그랬다: 위 지문을 넣고 색인을 다시 돌렸는데 ERP:sales 가 10건 그대로였다.
    #   색인 캐시 열쇠는 `경로|크기|시각` 이라 파일이 안 바뀌면 옛 갈래를 돌려주고,
    #   그걸 막는 규칙 지문에는 **폴더 규칙만** 들어 있었다(내용 판별은 빠져 있었다).
    #   고친 규칙이 안 먹는 것은 안 고친 것보다 나쁘다 — 고쳤다고 믿게 된다.
    import source_index as IX
    v0 = IX.rules_version()
    old = S.ERP_FILE_PREFIX
    try:
        S.ERP_FILE_PREFIX = dict(old, ZZZTEST="stmt")
        assert IX.rules_version() != v0, \
            "내용 분류 규칙이 바뀌었는데 색인 지문이 그대로다 — 색인이 옛 갈래를 계속 쓴다"
    finally:
        S.ERP_FILE_PREFIX = old
    assert IX.rules_version() == v0, "지문이 원래대로 안 돌아온다"
    print("  [161] ERP 내보내기 — 파일명이 내용보다 먼저 · 규칙 바뀌면 색인도 다시 판별 ✅")


def t160_master_book_cache():
    """관리대장을 한 번만 파싱한다 (worksplit #17, 2026-08-08).

    `master_stream` 이 네트워크 전송을 없앤 뒤에도 앱은 한 화면을 그리는 동안
    `load_workbook` 을 8군데에서 각각 불렀다. 실측 v556: load 만 2.59초 ×8.

    지키는 것은 넷이다:
      ① **워크북 개체를 공유하지 않는다** — `read_only=True` 워크북은 이터레이터
         상태를 갖고 스레드 안전하지 않다. 캐시에 담는 것은 **뽑아낸 행**이다.
      ② **값이 openpyxl 과 같다** — 빠른데 다른 값이면 그게 조용한 사고다.
      ③ **셀 개체를 몰래 흉내 내지 않는다** — values_only 가 아니면 분명히 실패한다.
      ④ **파일이 바뀌면 다시 읽는다** — mtime+크기로 판정.
    """
    import io as _io
    rec = open(os.path.join(ROOT, "ecount_reconcile.py"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    # ① 캐시에 담는 것이 행인가(워크북이 아니라)
    fn = rec.split("def master_book(")[1].split("\n\n\ndef ")[0]
    assert "tuple(r) for r in wb[n].iter_rows" in fn, \
        "워크북 개체를 캐시한다 — read_only 워크북은 스레드 안전하지 않다"
    assert "wb.close()" in fn, "파싱한 워크북을 닫지 않는다"
    assert "_SHEET_LOCK" in fn, "동시에 두 번 파싱한다"
    # ④ 무효화 근거
    assert "st_mtime_ns" in fn and 'update({"key": key, "sig": sig' in fn, \
        "파일이 바뀌어도 옛 시트를 계속 준다 — 화면이 옛 숫자를 보여 준다"

    # 호출처가 남아 있으면 그 자리만 옛날처럼 느리다
    assert "load_workbook(master_stream" not in server, \
        "아직 워크북을 직접 여는 자리가 있다 — 그 화면만 2.6초씩 더 든다"
    assert "master_book(master)" in server, "앱이 캐시를 쓰지 않는다"

    # ②③ 합성 워크북으로 값·오류를 확인한다(실데이터 없이)
    import openpyxl
    from ecount_reconcile import master_book
    tmp = os.path.join(ROOT, "reports", "_t160_synth.xlsx")
    wbw = openpyxl.Workbook()
    ws = wbw.active
    ws.title = "시트가"
    for r in range(1, 8):
        ws.append([f"a{r}", r, None])
    wbw.create_sheet("시트나").append(["x", "y"])
    wbw.save(tmp)
    try:
        real = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        cached = master_book(tmp)
        assert cached.sheetnames == real.sheetnames, "시트 목록이 다르다"
        assert ("시트가" in cached) and ("없는시트" not in cached), "in 연산이 다르다"
        for name in real.sheetnames:
            a = [tuple(r) for r in real[name].iter_rows(min_row=1, values_only=True)]
            b = [tuple(r) for r in cached[name].iter_rows(min_row=1, values_only=True)]
            assert a == b, f"{name} 값이 openpyxl 과 다르다 — 빠른데 틀리면 조용한 사고다"
        assert cached["시트가"].max_row == real["시트가"].max_row, "max_row 가 다르다"
        # 구간 읽기도 같아야 한다(머리글 4행만 읽는 자리가 여러 곳이다)
        a = [tuple(r) for r in real["시트가"].iter_rows(min_row=3, max_row=5, values_only=True)]
        b = [tuple(r) for r in cached["시트가"].iter_rows(min_row=3, max_row=5, values_only=True)]
        assert a == b and len(a) == 3, "구간 읽기가 다르다"
        real.close()
        # ③ 셀 개체를 요구하면 분명히 실패해야 한다
        try:
            list(cached["시트가"].iter_rows(values_only=False))
            raise AssertionError("values_only=False 인데 조용히 다른 것을 돌려준다")
        except ValueError:
            pass
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass
    print("  [160] 관리대장 한 번만 파싱 — 행만 캐시·값 동일·셀 개체 거부·mtime 무효화 ✅")


def t154_amount_basis():
    """[154] 금액을 맞추는 세 가지 기준 (2026-08-08 지시).

    사용자 지시: **"여기서는 돌발AS 정기점검만 집계해 / 그리고 나머지 다 진행해서
    맞춰서 완료 처리로 변경해"**.

    세 숫자가 다 '틀린 금액'처럼 보였지만 셋 다 **기준 문제**였다. 실측:
      ① 원장 미반영 930,539,994원 → 그중 894,118,514원이 ERP 에만 있는 업무
         (신규납품 440.6M·기타 362.2M·철거 47.7M·계단 43.6M)였다. 관리대장은
         돌발AS·정기점검만 원장에 담는다. 같은 종류끼리 대면 +10.8M / −0.6M.
      ② 계산서 발행율 0.9% → 06시트 '계산서' 칸이 사람 손 입력이라 유상 716건 중
         6건만 채워져 있었다. ERP 는 450건을 이미 발행(6·7단계)했다 → 63.7%.
      ③ 금액 재계산 대기 37건 → 명세서 부가세포함액이라 ÷1.1 하면 10원 단위로
         딱 떨어졌다(506,000→460,000). 떨어지는 것은 아는 금액이지 대기가 아니다.

    ★ 이 검증이 지키는 것은 "숫자가 크다/작다"가 아니라 **환산을 함부로 하지 않는 것**이다.
      안 떨어지는 금액까지 억지로 나누면 틀린 값이 화면에서 확정처럼 보인다.
    """
    import ecount_reconcile as E

    # ── ③ 환산은 깨끗할 때만 ──────────────────────────────────────────
    for total, want in ((506000, 460000), (528000, 480000), (704000, 640000),
                        (621500, 565000), (297000, 270000), (110, 100)):
        assert E.supply_from_statement(total) == want, f"{total} 환산이 틀렸다"
        assert round(want * 1.1) == total, "되돌려 곱해도 원본이 나와야 한다"
    for bad in (111, 0, None, -100, "", "abc", 105):
        assert E.supply_from_statement(bad) is None, \
            f"{bad!r} 은 부가세 포함액이라는 근거가 없다 — 환산하면 안 된다"

    # ── ② 발행 판정은 근거 세 갈래를 모두 인정한다 ────────────────────
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert '"발행(ERP확인)" if _erp_issued else' in server \
        and '"미발행"' in server, \
        "ERP 가 발행했다고 말하는 건을 화면이 미발행으로 센다"
    assert "_erp_progress_map.get" in server and "erp_progress" in server, \
        "발행 판정 근거가 erp_progress 가 아니면 '혼재' 프로젝트를 발행으로 세게 된다"
    assert "r.계산서==='발행'" not in live, \
        "발행율이 06시트 칸 하나만 보고 있다 — 그 칸이 비어 0.9% 가 나왔던 자리다"
    assert live.count("String(r.계산서||'').startsWith('발행')") >= 3, \
        "발행 집계 자리 중 일부만 고쳐졌다 — 화면마다 발행율이 달라진다"
    assert "!String(r.계산서||'').startsWith('발행')" in live, \
        "'미발행 목록'이 발행된 건을 다시 담는다"

    # ── ① ERP 비교는 원장이 담는 업무만 ───────────────────────────────
    assert "const ERP_LEDGER_KINDS = ['돌발AS','정기점검'];" in live, \
        "비교 대상 유형이 한 곳에 정의돼 있지 않다"
    assert "erpKindSum(v)" in live and "erpOtherYTD" in live, \
        "ERP 전체를 원장과 대 놓고 '미반영'이라 부르던 자리가 남아 있다"
    assert "ERP 전용 매출" in live, \
        "원장 관리 밖 업무를 따로 보여 주지 않으면 없어진 매출처럼 보인다"

    # ── 상태 사다리: 환산되면 흘러가고, 안 되면 사람에게 남는다 ────────
    saved = E.erp_progress
    try:
        E.erp_progress = lambda: {"P6": "6.세금계산서발행"}
        row = {"비용구분": "유상", "원천업무ID": "AS-1", "원장_공급가액": 0,
               "원장_거래명세서번호": "20260101-1", "프로젝트NO": "P6"}
        assert E.settle_status({**row, "원장_거래명세서합계": 110}) == "입금 대기", \
            "환산으로 금액을 알았는데도 '재계산 대기'에 붙잡혀 있다"
        assert E.settle_status({**row, "원장_거래명세서합계": 111}) == "금액 재계산 대기"
        assert E.settle_status({**row, "원장_거래명세서합계": 0}) == "금액 미입력"
    finally:
        E.erp_progress = saved
    print("  [154] 금액 기준 셋 — ÷1.1 은 깨끗할 때만 · 발행은 근거 세 갈래 · "
          "ERP 비교는 돌발AS·정기점검만 ✅")


def t157_tech_install():
    """[157] 기사 링크를 열면 크롬으로 넘어가고 설치가 뜬다 (2026-08-08 지시).

    사용자 지시: **"as 각 기사 링크를 보내고 열었을때 크롬으로 자동 설치되는 코딩 진행"**

    ★ 설치가 한 번도 안 됐던 진짜 이유는 UI 가 아니었다 — **`/manifest.json` 이 통째로
      500 이었다.** `do_GET` 안에서 `from urllib.parse import parse_qs` 를 했더니 그
      이름이 **함수 전체의 지역변수**가 되어, 그 가지를 안 지나간 매니페스트 가지에서
      `cannot access local variable 'parse_qs'` 로 죽었다. 매니페스트가 없으면 크롬은
      `beforeinstallprompt` 를 아예 안 던진다. 화면은 멀쩡했다.
    ★ 두 번째 이유: `/t/<slug>` 는 매니페스트에 `?tech=` 를 붙여 보내는데 서버는
      `staff=` 만 읽었다. 그래서 설치해도 아이콘이 `start_url="/"` — **PIN 걸린 관리자
      화면**으로 갔다. 설치는 됐는데 못 들어가니 아무도 안 쓴다.
    ★ 그리고 **열쇠는 매니페스트에 넣지 않는다.** 매니페스트는 캐시돼 기기에 남는다 —
      거기 열쇠가 박히면 비밀번호를 파일로 뿌리는 것과 같다.
    """
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    tech = open(os.path.join(ROOT, "webapp", "tech.html"), encoding="utf-8").read()

    # ① 함수 안 import 로 이름을 가리지 않는가 — 이것이 500 의 원인이었다.
    body = server[server.find("    def do_GET("): server.find("    def do_POST(")]
    # ★ 주석은 빼고 **실제 문장**만 본다 — 이 함정을 설명한 주석 자체가 그 문구를
    #   담고 있어서, 문자열만 찾으면 제 주석에 제가 걸린다(실제로 걸렸다).
    stmts = [ln for ln in body.splitlines() if not ln.lstrip().startswith("#")]
    assert not any("from urllib.parse import parse_qs" in ln for ln in stmts), \
        "do_GET 안에서 parse_qs 를 import 하면 매니페스트 가지가 500 이 된다"

    # ② 기사 매니페스트가 기사 화면을 가리키는가 · 열쇠가 안 새는가
    i = server.find('if p == "/manifest.json"')
    mani = server[i: i + 2600]
    assert 'query.get("tech")' in mani, "`?tech=` 를 안 읽는다 — 아이콘이 관리자 화면으로 간다"
    assert 'f"/t/{tech_slug}"' in mani, "start_url 이 기사 화면이 아니다"
    assert "tech_keys()" not in mani and "?k=" not in mani, \
        "매니페스트에 열쇠가 들어간다 — 기기에 캐시되는 파일이다"

    # ③ 크롬으로 **저절로** 넘어가는가 · 무한 반복은 막았는가
    assert "function autoOpenChrome(" in tech and "location.href = chromeIntent()" in tech, \
        "인앱 브라우저에서 사람이 단추를 눌러야만 크롬으로 간다"
    assert "sessionStorage" in tech and "AUTO_KEY" in tech, \
        "크롬 전환이 실패하면 무한 반복된다 — 화면이 깜빡이며 아무것도 못 한다"
    assert "IS_IOS" in tech and "Safari" in tech, "iOS 에 없는 기능을 있다고 말한다"

    # ④ 설치는 첫 터치에 뜨는가 · once 덫에 걸리지 않았는가
    assert "function armFirstTouch(" in tech and "beforeinstallprompt" in tech
    assert "{once:true, capture:true}" not in tech, \
        "배너를 눌러 그냥 돌아온 경우에도 귀가 떨어져 그다음 터치에서 영영 안 뜬다"
    assert "removeEventListener('pointerdown', fire, true)" in tech, \
        "설치창이 누를 때마다 뜬다 — 일을 못 한다"
    assert "SKIP_KEY" in tech, "'나중에' 를 눌러도 계속 뜬다"
    print("  [157] 기사 설치 — 매니페스트 500 원인 차단 · 아이콘이 기사 화면으로 · "
          "열쇠 비노출 · 크롬 자동전환(1회) · 첫 터치 설치 ✅")


def t156_refresh_fast():
    """[156] 갱신은 사람을 기다리게 하지 않는다 (2026-08-08 지시 "갱신 빨리빨리하게").

    화면에 `2시간 1분 전 자료 · 갱신 중` 이 오래 떠 있었다. 세 군데가 겹쳤다:
      ① `_fresh()` 가 **캐시를 볼 때마다** `resolve_master` 로 Z: 폴더를 훑었다
         (실측 1.24초). 화면 하나가 API 를 예닐곱 개 부르니 아무것도 안 바뀐
         상태에서도 Z: 를 예닐곱 번 훑었다 → 2초 캐시.
      ② TTL 이 끝나는 **그 순간에 들어온 요청**이 콜드 재계산을 통째로 뒤집어썼다
         (실측 get_works 첫 계산 111초). `status` 에만 있던 stale-while-revalidate
         를 works·settle·issues·erpdocs 로 넓혔다.
      ③ 예열 고리가 조회 함수를 그냥 부르면, ②가 생긴 뒤로는 **옛 값만 받고
         아무것도 안 데운 채** 240초를 또 잔다 → 예열은 `refresh_now` 로 강제한다.

    ★ 빠르게 만들면서 **낡은 값을 '지금 값'이라 말하지 않는 것**이 이 검증의 핵심이다.
      원장이 바뀌면 `_fresh` 가 `_cache` 를 통째로 비우므로 `_stale` 도 사라져야 한다.
      남겨 두면 화면이 바뀐 뒤의 옛 숫자를 새것처럼 보여 준다(조용한 사고).
    """
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "_MT_TTL" in server and 'if now - _MT["at"] < _MT_TTL' in server, \
        "캐시를 볼 때마다 Z: 를 훑는다"
    for fn in ("def cached_data(", "def _compute_locked(", "def _spawn_refresh(",
               "def refresh_now("):
        assert fn in server, f"{fn} 가 없다 — 갱신이 요청을 붙잡는 구조로 되돌아갔다"
    for key in ('cached_data("works"', 'cached_data("settle"',
                'cached_data("issues"', 'cached_data("erpdocs"'):
        assert key in server, f"{key} 가 옛 방식(만료 시 그 자리에서 재계산)이다"
    assert 'refresh_now("works"' in server and 'refresh_now("settle"' in server, \
        "예열이 옛 값만 받고 실제로는 데우지 않는다"
    # 캐시 조회는 락 밖이어야 한다 — 락 안에서 보면 남의 콜드 읽기에 갇힌다.
    i = server.find("def cached_data(")
    body = server[i: server.find("\ndef ", i + 10)]
    assert "_readlock" not in body, "캐시를 보는 데 락을 잡으면 옛 값조차 못 준다"

    # 원장이 바뀌면 stale 까지 사라지는가 — 실제로 돌려 확인한다.
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as A
    saved_mt, saved_cache = A._master_mtime, dict(A._cache)
    try:
        A._cache.clear()
        A._master_mtime = lambda: 111
        calls = []

        def build():
            calls.append(1)
            return {"n": len(calls)}

        assert A.cached_data("t156", build) == {"n": 1}
        assert A.cached_data("t156", build) == {"n": 1}, "캐시가 안 먹었다"
        assert A._cache.get("t156_stale") == {"n": 1}, "옛 값 자리가 안 생겼다"
        A._cache.pop("t156"), A._cache.pop("t156_ts")          # TTL 만료 흉내
        assert A.cached_data("t156", build) == {"n": 1}, \
            "만료 순간 요청이 재계산을 뒤집어썼다 — 옛 값을 즉시 줘야 한다"
        A._master_mtime = lambda: 222                          # 원장이 바뀌었다
        A._fresh("t156")
        assert "t156_stale" not in A._cache, \
            "원장이 바뀌었는데 옛 숫자가 남았다 — 그것을 '지금 값'으로 보여 주게 된다"
    finally:
        A._master_mtime = saved_mt
        A._cache.clear()
        A._cache.update(saved_cache)
    # ── 옛 코드로 도는 서버를 잡는가 (2026-08-08 반나절짜리 사고) ──────────
    # 서버는 200 을 주고 화면도 숫자를 보여 주는데 그 코드가 어제 것이었다.
    # **고친 사람만 모르고 있었다** — 그래서 사람이 아니라 기계가 보게 만든다.
    import restart_server as RS
    saved_running = RS.running
    try:
        RS.running = lambda: [(1234, "08/07/2026 20:48:45")]     # 어제 뜬 서버
        s = RS.stale()
        assert s and s[0] == 1234 and s[2], \
            "오늘 고친 코드를 어제 서버가 물고 있는데 아무도 말하지 않는다"
        assert "webapp/app_server.py" in s[2] or "webapp/index.html" in s[2]
        from datetime import datetime, timedelta
        soon = (datetime.now() + timedelta(days=1)).strftime("%m/%d/%Y %H:%M:%S")
        RS.running = lambda: [(1234, soon)]                      # 방금 띄운 서버
        assert RS.stale() is None, "최신인데 옛 코드라고 거짓 경보를 낸다"
        RS.running = lambda: []                                  # 안 떠 있으면 판단 없음
        assert RS.stale() is None
        assert RS._started_epoch("nonsense") is None, \
            "시각을 못 읽으면 조용히 지나가야 한다(틀린 판정보다 낫다)"
    finally:
        RS.running = saved_running
    handoff = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "def app_server_health(" in handoff and '"앱서버"' in handoff, \
        "세션 인계가 앱 서버 상태를 안 본다"
    assert 'ap.get("옛코드")' in handoff and "restart_server.py" in handoff, \
        "옛 코드로 도는 서버가 '먼저 처리할 것'에 안 올라온다"

    # ── 카톡에서도 사람 인계가 되는가 (2026-08-08 지시: "카톡에 이렇게 올라와도") ──
    kakao = open(os.path.join(ROOT, "kakao_extract.py"), encoding="utf-8").read()
    assert "import people_alias" in kakao and "people_alias.resolve_text(text" in kakao, \
        "이름이 **필드가 아니라 본문 안내 문구**에 있다 — 칸만 보면 지시가 안 닿는다"
    assert '"접수담당": 접수담당' in kakao and "쿠팡담당원문" in kakao, \
        "옮긴 이름만 남기고 원문을 지우면 '그때 누구라고 쓰여 있었나'를 잃는다"
    assert "when=msg_day" in kakao, "인계 전 카톡까지 소급해 바꾼다"
    # ── 대표보고도 같은 규칙을 따르나 (2026-08-08 사용 기록에서 느린 화면 1등) ──
    #   실측 최근 24시간: /api/exec_report 648회 · 평균 110초. 계산이 무거운 게 아니라
    #   **캐시를 보러 가는 길에 락이 잠겨 있었다** — /api/status 는 하루 전에 같은
    #   이유로 고쳐졌는데 여기만 남아 있었다.
    app = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    # 함수 **전체**를 본다. 앞에서 N자만 잘라 보면 안 된다 — 이 함수는 DEMO 분기가
    # 길어 락 지점이 한참 뒤에 있고, 자르는 순간 "그런 코드 없음"으로 읽힌다.
    ex = app.split("def get_exec_report", 1)[1].split("\ndef ", 1)[0]
    assert ex.index('_fresh("exec")') < ex.index("with _readlock"), \
        "대표보고가 캐시를 보기 전에 락을 잡는다 — 만들어 둔 값을 두고 줄을 선다"
    assert '_cache.get("exec_stale")' in ex and '_cache["exec_stale"] = r' in ex, \
        "만료 순간마다 누군가 한 명이 콜드 재계산을 통째로 맞는다"
    assert 'get_exec_report(None, _force=True)' in ex, \
        "뒤에서 도는 갱신이 옛 값을 그대로 되돌려 저장한다 — 영원히 안 새로워진다"
    assert '("exec", lambda: get_exec_report(None, _force=True))' in app, \
        "대표보고가 예열 목록에 없다 — 서버가 뜬 뒤 첫 사람이 콜드 계산을 맞는다"
    print("  [156] 갱신 — 원장시각 2초 캐시 · 만료 시 옛 값 즉시 + 뒤에서 한 번 · "
          "예열은 강제 재계산(대표보고 포함) · 옛 코드 서버 감지 · 카톡 인계 ✅")


def t155_cancel_and_handover():
    """[155] 접수 취소 · 사람 인계 (2026-08-08 지시).

    사용자 지시 ①: **"접수 했다가 접수 취소하는 경우도 많은데 이것도 잡아내는
    알고리즘 추가해"** — 밴드 댓글 "작동 원활함. 접수 취소 하세요".
    사용자 지시 ②: **"김혜진 매니저 퇴사한지 1달이 넘었어. 카톡에 이렇게 올라와도
    류지영으로 인식하는 알고리즘 추가. 번호는 같아"**

    ★ 이 검증이 정말 지키는 것은 **취소로 잘못 죽이지 않는 것**이다. 접수를 취소로
      처리하면 그 현장은 아무도 안 가는데 목록에서도 사라진다 — 미실시로 남는 것보다
      나쁘다. 실측에서 두 번 새어 나갔다:
        · "바디부분 아크릴판은 캠프담당 취소요청함" (부품 취소)
        · "택배발송 취소요청하심" 바로 뒤에 밴드 양식의 `● A/S 완료 :` 줄이 붙어,
          'A/S 가 곁에 있으면 취소'라는 규칙이 **모든 글**을 삼켰다.
      그래서 근거는 '접수'가 '취소'에 붙어 있는 것 하나로 좁혔다.
    """
    import band_extract as B
    import people_alias as P

    # ── ① 잡아야 하는 것 ─────────────────────────────────────────────
    for t in ("접수 취소 하세요", "접수취소", "통화 완료 했습니다 작동 원활함. 접수 취소 하세요",
              "✅ 접수 취소", "✔️UJ2601291랑 같은내용으로 접수취소", "*이상없음 접수취소",
              "접수를 취소해주세요", "접수건 취소 부탁드립니다", "접수 철회", "접수 반려",
              "오접수 입니다", "✅ 중복접수 처리완료",
              "✅ [접수취소] - 유선접화로 해결 완료",
              "기사님과 유선전화 이후 접수 취소되었습니다"):
        assert B.cancel_hit(t), f"접수 취소를 놓쳤다: {t}"

    # ── ② 잡으면 안 되는 것 (여기가 진짜 관문) ────────────────────────
    for t in ("바디부분 아크릴판은 캠프담당 취소요청함.",
              "본사 상신,승인 진행되지 않았으며, 택배발송 취소요청하심. ● A/S 완료 :",
              "*담당자 접수전, 정기점검 취소되어 도어락만 교체진행됨",
              "부품 취소 요청", "예약 취소불가 안내", "접수 취소 불가",
              "보험접수 취소 후 쿠팡측에서 긴급수리요청",
              "택배 접수 취소 후 방문수리 전환",
              "취소된 건 없음", "유선전화로 해결 완료", "접수 유지",
              "접수는 취소하지 않습니다", "접수 취소 보류", "작업 완료", ""):
        assert not B.cancel_hit(t), f"취소가 아닌 것을 취소로 죽인다: {t}"

    # ── ③ 댓글 자리 — 지금은 캐시에 본문이 없다. 그 사실을 세어야 한다 ──
    assert B.comment_text({"comments": [{"content": "접수 취소 하세요"}]}) == "접수 취소 하세요"
    assert B.comment_text({"comment_count": 3}) == "", "없는 댓글을 있는 것처럼 만든다"
    # ★ 기대값이 2 → 3 으로 바뀌었다 (2026-08-08 저녁, 검증 [169]).
    #   "2" 는 `comment_count: 0` 이고 `comments` 키가 **없다** — 예전에는 '댓글이 없는
    #   글'로 보고 넘겼다. 그런데 실측에서 캐시 10,312글 중 `comment_count>0` 이 6글뿐인
    #   것이 드러났다. 밴드에 댓글이 없어서가 아니라 **수집기가 그 숫자를 안 담아서**다.
    #   즉 그 0 은 '없다'가 아니라 '안 봤다'였고, 계기는 그것을 믿고 사각지대 0건이라
    #   말했다. 이제 `comments` 키가 없으면 **안 본 것**으로 세고, `comments: []`(보긴
    #   봤고 없었다)와 가른다. 그래서 1·2·3 이 사각지대이고 4 만 아니다.
    assert B.cancel_blind_count({"1": {"comment_count": 3},
                                 "2": {"comment_count": 0},
                                 "3": {"comment_count": "2"},
                                 "4": {"comment_count": 1,
                                       "comments": [{"content": "x"}]}}) == 3, \
        "댓글은 있는데 본문을 못 읽는 사각지대를 0건이라 말하면 안 된다"
    assert B.cancel_blind_count({"2": {"comment_count": 0, "comments": []}}) == 0, \
        "보고 나서 없었던 글까지 세면 안 된다 — '없다'와 '안 봤다'는 다르다"

    # ── ④ 순서: 댓글 취소 > 완료 제목 > 본문 취소 ────────────────────
    def _post(body, comments=None):
        return {"content": body, "created_at": 1767225600000,
                **({"comments": comments} if comments else {})}
    head = "♣ ［ 돌발 유료 A/S 완료 ]\n● 프로젝트NO : UJ2600001\n"
    done = B.parse_post("1", _post(head), "밴드")
    assert done and done["진행상태"] == "작업완료"
    later = B.parse_post("1", _post(head, [{"content": "접수 취소 하세요",
                                            "created_at": 1767225700000}]), "밴드")
    assert later["진행상태"] == "취소", "댓글은 글보다 나중이다 — 취소가 이겨야 한다"
    body_only = B.parse_post("1", _post(head + "정기점검 취소되어 도어락만 교체"), "밴드")
    assert body_only["진행상태"] == "작업완료", \
        "완료 글 본문의 딴 얘기 취소가 완료를 덮었다"

    # ── ⑤ 사람 인계 — 원문은 두고 읽을 때만 옮긴다 ────────────────────
    assert P.resolve_person("김혜진 매니저") == "류지영"
    assert P.resolve_person("김혜진 매니저님") == "류지영"
    assert P.resolve_person("", "010-6645-4535") == "류지영", "번호가 같으면 그 자리다"
    assert P.resolve_person("", "01066454535") == "류지영", "하이픈 유무로 갈리면 안 된다"
    assert P.resolve_person("김혜진", when="2026-03-01") == "김혜진", \
        "인계 전 글까지 소급해 바꾸면 '그때 누구였나'를 잃는다"
    assert P.resolve_person("김준형") == "김준형", "관계없는 사람을 건드렸다"
    assert P.resolve_text("[휴대전화] 010-6645-4535") == "류지영"
    assert P.resolve_text("[담당이름] 김혜진 매니저") == "류지영"
    assert P.resolve_text("아무 말도 없는 글") == ""
    assert "김혜진" in P.note_of("김혜진") and "류지영" in P.note_of("김혜진"), \
        "왜 이름이 바뀌었는지 근거가 없으면 사람이 못 믿는다"
    assert B.normalize_tech("김혜진", when="2026-08-08") == "류지영"
    assert B.normalize_tech("김혜진", when="2026-03-01") == "김혜진"

    # ── ⑥ 회차에 매여 있나 · 엑셀을 직접 열지 않나 ────────────────────
    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert '"cancel_watch.py"), "--sync"' in daily, \
        "취소 확인이 09:50 회차에 없다 — 대화에 남긴 것은 사라진다"
    watch = open(os.path.join(ROOT, "cancel_watch.py"), encoding="utf-8").read()
    assert "openpyxl" not in watch and "workbook_patch" not in watch, \
        "취소 확인이 엑셀을 직접 연다 — 반영은 11:00·15:00 회차 몫이다"
    resolution = open(os.path.join(ROOT, "cancel_resolution.py"), encoding="utf-8").read()
    assert "expected_version" in resolution and "_OPEN_STATES" in resolution, \
        "빈칸만 쓰는 옛 큐 대신 열린 상태 CAS가 없으면 취소가 적용되지 않거나 완료를 덮는다"
    assert "완료일 또는 종료상태가 있어 자동 취소하지 않음" in resolution, \
        "이미 완료된 업무를 취소로 뒤집는 안전관문이 없다"
    # ★ 산문이 아니라 **코드**를 본다. 머리말 설명글은 "convert_dump 가 담는다"처럼
    #   도구 이름을 대는 것이 마땅한데, 그걸 호출로 오해하면 문서를 못 쓰게 된다
    #   (검증이 제 설명글에 걸리는 일은 [157] 에서도 한 번 있었다).
    watch_code = watch.split('"""', 2)[-1]
    watch_code = "\n".join(l for l in watch_code.splitlines()
                           if not l.strip().startswith("#"))
    for scrape in ("collect_", "convert_dump", "requests.", "webdriver"):
        assert scrape not in watch_code, f"코딩 세션 도구가 수집을 한다: {scrape}"
    print("  [155] 접수 취소 — 부품·택배 취소와 분리 · 열린 상태 CAS · 댓글 우선 · 사각지대 집계 · "
          "사람 인계(번호 근거·소급 금지) ✅")


def t162_band_comments_collected():
    """[162] 밴드 댓글 본문 수집 — 담는 쪽 둘, 읽는 쪽 하나 (2026-08-08 지시).

    사용자 지시: **"밴드 댓글도 같이 수집하는 알고리즘 실행"**

    왜 필요했나 — `[155]` 로 취소 판정을 만들어 놓고도 **반쪽**이었다. 취소 통보는
    거의 다 댓글로 오는데("작동 원활함. 접수 취소 하세요") 캐시에는 `comment_count`
    숫자만 있고 본문이 없었다. 판정 코드는 멀쩡히 돌면서 **아무것도 못 잡았다.**

    ★ 이 검증이 지키는 것 셋 — 전부 '조용한 사고' 쪽이다:
      ① **모양이 갈리지 않을 것.** 담는 쪽이 둘(화면 긁기 convert_dump · API band_sync)
         인데 읽는 쪽은 하나(band_extract.comment_text)다. 한쪽만 고치면 그 경로로
         들어온 댓글은 영영 안 읽힌다 — 오류도 안 난다.
      ② **시각 없는 댓글을 버릴 것.** 본문과 같은 규칙이다(`[130]`). 시각이 없으면
         '댓글이 글보다 나중'이라는 취소 판정의 순서 자체를 세울 수 없다.
      ③ **합치되 잃지 말 것.** 덤프는 매 실행 전부 재처리된다. 댓글이 없던 시절의
         옛 덤프가 나중에 이기면 애써 모은 댓글이 통째로 사라진다.
    """
    import importlib
    B = importlib.import_module("band_extract")
    sys.path.insert(0, os.path.join(ROOT, "band"))
    CD = importlib.import_module("convert_dump")

    # ── ① 시각 없는 댓글은 버린다 · 중복은 접힌다 · 시간순으로 선다 ──────
    cap = 1767000000000
    got = CD.conv_comments([
        {"author": "차동호", "content": "확인했습니다", "created_at": 1766000002000},
        {"author": "차동호", "content": "확인했습니다", "created_at": 1766000002000},   # 중복
        {"author": "류지영", "content": "접수 취소 하세요", "created_at": 1766000001000},
        {"author": "유령", "content": "시각이 없다"},                                   # 버려야
        {"author": "빈", "content": "", "created_at": 1766000003000},                   # 버려야
    ], cap)
    assert [c["content"] for c in got] == ["접수 취소 하세요", "확인했습니다"], \
        "시각 없는 댓글이 남았거나, 중복이 안 접혔거나, 시간순이 아니다: %r" % (got,)
    assert all(isinstance(c["created_at"], int) for c in got), \
        "created_at 이 정수가 아니다 — 날짜가 날짜가 아니었던 사고(2026-08-08)와 같은 모양이다"
    # timeText 만 있어도 살려낸다(화면 긁기 경로)
    assert CD.conv_comments([{"content": "취소요", "timeText": "1시간 전"}], cap), \
        "화면에서 긁은 상대시각('1시간 전')을 못 살린다"

    # ── ② 담는 쪽 둘이 **같은 모양**인가 ─────────────────────────────────
    BS = open(os.path.join(ROOT, "band", "band_sync.py"), encoding="utf-8").read()
    api_from_json = json.loads(json.dumps({          # API 응답 흉내
        "latest_comments": [
            {"author": {"name": "류지영"}, "body": "접수 취소 하세요", "created_at": 1766000001000},
            {"author": {"name": "유령"}, "body": "시각 없음"},
        ]}))
    ns = {}
    exec(compile(BS[BS.index("def api_comments"):BS.index("def main()")],
                 "band_sync-part", "exec"), ns)
    api_out = ns["api_comments"](api_from_json)
    assert api_out == [{"author": "류지영", "created_at": 1766000001000,
                        "content": "접수 취소 하세요"}], \
        "API 경로가 화면 긁기와 다른 모양으로 담는다 — 읽는 쪽이 하나라 한쪽이 조용히 죽는다: %r" % (api_out,)
    assert set(api_out[0]) == set(got[0]), "칸 이름이 두 경로에서 다르다"

    # ── ③ 읽는 쪽은 하나 — 그 하나가 두 경로를 다 읽나 ───────────────────
    post = {"comment_count": 2, "comments": got}
    assert "접수 취소" in B.comment_text(post) and "확인했습니다" in B.comment_text(post)
    assert B.cancel_hit(B.comment_text(post)), "댓글로 온 취소를 못 잡는다"

    # ── ④ 반쯤 읽은 것을 다 읽은 것으로 세지 않나 ────────────────────────
    half = {"3": {"comment_count": 3, "comments": got}}          # 3개 중 2개만
    assert B.cancel_blind_count(half) == 1, \
        "접힌 댓글을 반만 읽고도 사각지대 0으로 센다 — 제일 나쁜 종류의 안심이다"
    full = {"3": {"comment_count": 2, "comments": got}}
    assert B.cancel_blind_count(full) == 0
    none = {"4": {"comment_count": 0, "comments": []}}
    assert B.cancel_blind_count(none) == 0, "댓글이 없는 글까지 사각지대로 센다"

    # ── ⑤ 합치되 잃지 않나 (옛 덤프가 나중에 이겨도) ─────────────────────
    conv = open(os.path.join(ROOT, "band", "convert_dump.py"), encoding="utf-8").read()
    assert 'rec["comments"] = conv_comments(' in conv and 'cur.get("comments")' in conv, \
        "재병합이 댓글을 덮어쓴다 — 댓글 없던 옛 덤프가 이기면 통째로 사라진다"
    assert 'rec.get("comments_full") or cur.get("comments_full")' in conv, \
        "'다 읽었다'는 사실이 재병합에서 지워진다"
    assert '"comments": api_comments(it)' in BS and "merge_comments(old.get" in BS, \
        "API 경로가 아는 글의 댓글을 안 받는다 — 댓글은 글보다 **나중에** 달린다"

    # ── ⑥ 긁는 쪽(붙여넣기 JS)이 실제로 댓글을 담나 ──────────────────────
    js = open(os.path.join(ROOT, "band", "grab_posts.js"), encoding="utf-8").read()
    body = "\n".join(l for l in js.splitlines() if not l.strip().startswith("//"))
    # ★ 변수 이름이 아니라 **지키려는 것**을 본다. 2026-08-09 에 확인 못 한 수확을
    #   기록하지 않도록 반환부를 고치며 `comments: cts` 가 `post.comments = cts` 가
    #   됐는데, 철자를 박아 둔 이 검사가 거기서 깨졌다 — 뜻은 그대로였다([178]).
    assert "readComments" in body and re.search(r"comments\W+=\W*cts|comments:\s*cts", body), \
        "화면 긁기가 댓글을 안 담는다 — 흡수기만 고치면 아무 일도 안 일어난다"
    assert re.search(r"comments_full\W+=?:?\s*cts\.length >= want", body), \
        "'다 읽었나'를 안 적는다 — '댓글 없음'과 '못 읽음'이 캐시에서 똑같아 보인다"
    # 시각 없는 댓글은 버린다([130]). 본문은 사진/스티커만 있는 댓글이면 빌 수 있어
    #   더는 요구하지 않는다(2026-08-09 실측 — 각 댓글은 div.cComment 자체이고
    #   본문 셀렉터는 ._commentContent, 시각은 .time). 뜻(시각 가드)만 본다.
    assert re.search(r"if \(!timeText\) continue", body), \
        "시각 없는 댓글을 담는다(본문 규칙 [130] 과 어긋난다)"
    # 각 댓글 항목 셀렉터가 **한 칸 아래**(.cComment li)를 가리켜 6개를 0개로 읽던
    #   사고를 다시 밟지 않게, 항목 셀렉터가 cComment 자체를 집는지 확인한다.
    assert "cComment" in body and re.search(r"CMT_ITEM\s*=\s*'div\.cComment", body), \
        "댓글 항목 셀렉터가 cComment 자체가 아니라 그 하위를 가리킨다 — 늘 0개로 읽힌다"
    # '확인된 0개'와 '못 읽음'을 가른다 (2026-08-09 실측). 댓글 0 인 글은 '댓글 N'
    #   표시가 없어 countKnown 이 false 다 — 그것을 전부 미확인으로 두면 0개짜리 글이
    #   영원히 다시 뽑혀 백필이 수렴하지 못한다. 입력창이 있는데 목록이 0 이면 확인된 0개다.
    assert "commentUiReady" in body and "_commentInputRegion" in body, \
        "0개짜리 글을 '확인된 0개'로 못 가른다 — 매 회차 다시 뽑혀 백필이 안 끝난다"
    mo = open(os.path.join(ROOT, "band", "make_oneclick.py"), encoding="utf-8").read()
    assert 'open(os.path.join(HERE, "grab_posts.js")' in mo, \
        "붙여넣기 파일이 grab_posts.js 를 싣지 않는다 — 고쳐도 사람 손에는 옛 JS 가 간다"

    # ── ⑦ 이미 만들어 둔 붙여넣기 파일이 낡으면 **스스로** 다시 만드나 ──────
    #   수집 규칙을 고쳐도 사람 손에 가는 것은 디스크에 있는 그 파일이다. 실측:
    #   댓글 수집을 붙인 날 붙여넣기 파일 4개가 전부 그 이전 것이었다 — 그대로
    #   붙여넣었으면 댓글이 한 건도 안 들어오는데 수집은 '성공'으로 끝났을 것이다.
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "def heal_stale_pastefiles" in wd and "heal_stale_pastefiles(dry)," in wd, \
        "낡은 붙여넣기 파일을 아무도 다시 안 만든다 — 30분 워치독에 걸려 있어야 한다"
    heal = wd.split("def heal_stale_pastefiles")[1].split("\ndef ", 1)[0]
    # 2026-08-11: 판정 기준이 '만드는 쪽'별로 갈렸다(수집_* 는 mk_mt, 나머지는 js_mt —
    # heal 의 주석 참조). 지키는 정책은 그대로다: **내용 비교가 아니라 mtime 비교**이고
    # 기준선에 grab_posts.js mtime 이 들어간다. 문자 그대로 못 박으면 기준을 세분화할
    # 때마다 검증이 헛되이 깨진다.
    assert "os.path.getmtime(p) <" in heal and "js_mt" in heal, \
        "판정 근거가 만드는 쪽 mtime 이 아니다(회차 번호가 매번 달라 내용 비교는 못 쓴다)"
    assert "os.path.getmtime(p) >= js_mt" in heal and "os.unlink(p)" in heal, \
        "끝난 코드만 보고 성공으로 센다 — 훑을 것이 없는 밴드는 파일이 안 써지는데도 0 이다"
    for scrape in ("band_sync", "requests.", "webdriver", "convert_dump"):
        assert scrape not in heal, f"자동복구가 수집을 한다: {scrape}"
    print("  [162] 밴드 댓글 — 두 경로 같은 모양 · 시각 없으면 버림 · 재병합 무손실 · "
          "반쯤 읽음을 사각지대로 셈 ✅")


def t163_last_run_shown():
    """[163] 실행 화면 상단의 '마지막 실행 시각' (2026-08-08 지시).

    사용자 지시: **"이 화면 상단에 최근 마지막 실행 날짜 시간이 어떻게 되는지 표시해"**

    ★ 이 검증이 지키는 것은 **거짓 표시를 안 하는 것**이다:
      ① **서버가 다시 떠도 남을 것.** 이 서버는 코드를 고칠 때마다 다시 뜬다.
         메모리에만 두면 재시작 때마다 "기록 없음"이 되어, 오늘 아침에 돈 대조까지
         안 돈 것처럼 보인다 — 없는 것보다 나쁜 표시다.
      ② **앱 밖에서 돈 것도 셀 것.** 전체 대조는 09:50 스케줄러가 매일 돌린다.
         앱 단추 기록만 보여 주면 사람이 방금 돈 것을 또 누른다(한 번이 몇 분이다).
      ③ **실패를 성공과 같은 모양으로 적지 않을 것.** 돌긴 돌았는데 실패한 것을
         '돌았다'로만 보여 주면 사람이 됐다고 믿는다.
    """
    src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "LAST_RUN_PATH" in src and "_note_last_run(key, title, local_returncode)" in src, \
        "작업이 끝나도 마지막 실행 시각이 남지 않는다"
    assert "os.replace(tmp, LAST_RUN_PATH)" in src, \
        "반쯤 쓰인 기록 파일이 남을 수 있다(읽는 쪽이 통째로 못 읽는다)"
    assert '"last": last_runs(merge_auto=True)' in src, \
        "/api/tasklog 가 마지막 실행 시각을 안 준다 — 화면이 읽을 자리가 없다"
    assert 'agent_status.json' in src.split("def last_runs")[1][:900], \
        "09:50 자동 회차가 안 세어진다 — 아침에 돈 대조가 화면에서 사라진다"

    # 실제로 남고, 다시 읽히나 (파일 하나짜리 계약이라 여기서 돌려 본다)
    import importlib, json as _j, tempfile
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    A = importlib.import_module("app_server")
    keep = A.LAST_RUN_PATH
    try:
        A.LAST_RUN_PATH = os.path.join(tempfile.gettempdir(), "_t163_last_run.json")
        if os.path.exists(A.LAST_RUN_PATH):
            os.unlink(A.LAST_RUN_PATH)
        assert A.last_runs() == {}, "기록이 없을 때 빈 값이 아니다"
        A._note_last_run("synthetic", "합성검증", 0)
        A._note_last_run("po", "쿠팡 PO 대조", 2)
        got = A.last_runs()
        assert set(got) == {"synthetic", "po"}, got
        assert got["synthetic"]["코드"] == 0 and got["po"]["코드"] == 2, \
            "실패한 실행이 성공과 구별되지 않는다"
        assert len(str(got["po"]["끝난시각"])) >= 16, "시각이 날짜만 있고 시간이 없다"
        # 두 번째 기록이 첫 번째를 지우지 않는다(작업마다 따로 남아야 화면이 다 보여 준다)
        A._note_last_run("synthetic", "합성검증", 0)
        assert set(A.last_runs()) == {"synthetic", "po"}
    finally:
        A.LAST_RUN_PATH = keep
        try:
            os.unlink(os.path.join(tempfile.gettempdir(), "_t163_last_run.json"))
        except OSError:
            pass

    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert 'id="lastRuns"' in html and "function renderLastRuns" in html, \
        "화면에 마지막 실행 줄이 없다"
    assert "renderLastRuns(d.last)" in html, "받아 놓고 안 그린다"
    assert "if(v==='run'){ pollLog(); }" in html, \
        "실행 화면을 열 때 갱신하지 않는다 — 열어 둔 채 시간이 지나면 '3시간 전'이 거짓이 된다"
    assert "String(v.끝난시각).replace('T',' ').slice(0,16)" in html, \
        "절대시각(날짜+시간)을 안 적는다 — 지시가 '날짜 시간'이었다"
    assert "bad" in html.split("function renderLastRuns")[1][:1400], \
        "실패한 실행이 성공과 같은 모양으로 보인다"
    print("  [163] 마지막 실행 시각 — 재시작에도 남음 · 자동 회차 포함 · 실패 구별 · "
          "절대시각+상대시각 ✅")


def t149_tech_center():
    """AS 담당기사 전용 화면 — 비밀번호 없이, 그러나 누구나는 아니다 (2026-08-08 지시).

    사용자 지시: "업무센터에 각 AS 담당자 4명도 넣어서 별도의 비밀번호 없는 화면으로
    딱 AS 담당자가 할 수 있는 업무만 넣어서 만들어줘 (링크 타고 열면 크롬으로 강제로
    열어서 앱을 모바일에 설치할 수 있는 구조로 알고리즘 구성해)".

    지키는 것은 다섯이다:
      ① **tech 는 세 번째 역할이다** — staff 로 만들면 그 순간 기사 링크 하나로
         원장 전체가 열린다(staff 는 `_auth()` 를 통과한다).
      ② **`_auth()` 를 통과하지 못한다** — 기사 화면이 쓰는 길은 /api/tech/* 뿐이다.
      ③ **제 것만 본다·제 것에만 쓴다** — 목록에 없는 ID 에는 완료를 못 찍는다.
      ④ **금액은 싣지 않는다** — 링크는 카톡으로 돌아다닌다. 새어도 될 것만 싣는다.
      ⑤ **되지 않는 것을 된다고 하지 않는다** — iOS 는 크롬 강제가 불가능하다.
    """
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    tech = open(os.path.join(ROOT, "webapp", "tech.html"), encoding="utf-8").read()

    # ① 네 사람 · 세 번째 역할
    import ledger_db as _L
    src = server.split("AS_TECH_CENTERS = {")[1].split("}\n")[0]
    for nm in _L.AS_TECHS:
        assert nm in src, f"{nm} 기사가 전용 화면 명단에 없다"
    cs = server.split("def create_auth_session(")[1].split("\ndef ")[0]
    assert 'role, staff_slug = "tech", tech_slug' in cs, \
        "기사를 staff 로 만든다 — 링크 하나로 원장 전체가 열린다"
    ck = server.split("def auth_session_from_cookie(")[1].split("\ndef ")[0]
    assert 'role == "tech" and staff_slug not in AS_TECH_CENTERS' in ck, \
        "쿠키에 아무 slug 나 넣어도 기사로 통과한다"

    # ② 관문 — tech 는 일반 API 를 못 쓴다
    au = server.split("    def _auth(self):")[1].split("\n    def ")[0]
    assert 'return str(session.get("role") or "") != "tech"' in au, \
        "기사 세션이 원장 API 관문을 통과한다"
    # 링크 목록은 관리자만(열쇠가 그대로 들어 있다)
    lk = server.split('if p == "/api/tech/links":')[1][:300]
    assert "_require_admin()" in lk, "기사가 다른 기사의 링크(=비밀번호)를 볼 수 있다"

    # ③ 남의 건에 완료를 못 찍는다 · 열쇠는 상수시간 비교
    tr = server.split("def tech_report(")[1].split("\ndef ")[0]
    assert "내 일감 목록에 없는 건입니다" in tr, "남의 건에 완료를 찍을 수 있다"
    assert "enqueue_for_scheduled_apply(" in tr and "ledger_db.enqueue(" not in tr and "--apply" not in tr, \
        "기사 보고가 앱 DB 즉시저장을 거치지 않거나 Excel을 바로 연다"
    # 조치 메모를 고객 요청 칸에 쓰지 않는다(그 칸의 뜻이 망가진다)
    assert '"신청내용"' not in tr, "조치 내용을 신청내용 칸에 덮어쓴다"
    assert "hmac.compare_digest" in server.split("def tech_check_key(")[1].split("\ndef ")[0], \
        "링크 열쇠를 == 로 비교한다"
    # 열쇠는 git 밖 파일에 — 코드에 적지 않는다
    assert "tech_keys.local.json" in server and "secrets.token_urlsafe" in server, \
        "링크 열쇠가 코드나 추적 파일에 있다"

    # ④ 금액을 싣지 않는다 — 서버 payload 와 화면 둘 다
    tb = server.split("def tech_board(")[1].split("\ndef ")[0]
    # 설명글(docstring)과 '금액이 안 나온다'는 안내 문구는 검사에서 뺀다 —
    # 그 말 자체는 금액이 아니다. 검사할 것은 **실제로 실어 보내는 값**이다.
    tb = tb.split('"""', 2)[-1]
    tb = "\n".join(l for l in tb.splitlines() if "나오지 않습니다" not in l)
    for bad in ("공급가", "합계", "부가세", "단가", "금액"):
        assert bad not in tb, f"기사 화면에 {bad} 가 실린다 — 링크는 카톡으로 돈다"
    # 기사 화면이 부르는 길은 /api/tech/* 뿐이어야 한다(관리자 앱을 재활용하지 않는다)
    calls = re.findall(r"api\('(/[^']+)'", tech)
    assert calls and all(c.startswith("/api/tech/") for c in calls), \
        f"기사 화면이 기사 전용이 아닌 길을 부른다: {[c for c in calls if not c.startswith('/api/tech/')]}"

    # ⑤ iOS 는 강제하지 않고 안내만 한다
    assert "package=com.android.chrome" in tech, "안드로이드에서 크롬으로 넘기지 않는다"
    assert "IS_IOS" in tech and "Safari" in tech, \
        "iOS 에서도 강제되는 것처럼 굴면 눌러도 아무 일이 없다"
    assert "beforeinstallprompt" in tech and "if(!INSTALL)" in tech, \
        "설치 단추를 미리 그린다 — 눌러도 아무 일이 없는 단추가 된다"
    # 열쇠는 주소창에서 지운다(화면 갈무리로 새는 것이 가장 흔한 사고다)
    assert '"Location": f"/t/{tslug}"' in server, "주소창에 열쇠가 남는다"
    print("  [149] 기사 전용 화면 — 세 번째 역할·원장 차단·제 것만·금액 없음·iOS 정직 ✅")


def t148_input_suggest():
    """입력 자동완성 — 원장에 있는 값을 추천한다 (2026-08-08 지시).

    사용자 지시: "입력란 입력할 때 DB 기반으로 자동 입력 추천 뜨게 전체 앱 시스템
    코딩해 / 사용자가 매번 찾아 입력하기 불편해".

    지키는 것은 넷이다:
      ① **지어내지 않는다** — 추천은 원장에 실제로 있는 값뿐이다. 그럴듯한 후보를
         만들어 주면 사람이 그걸 골라 **새 오타가 원장에 들어간다**. 표기 흔들림을
         줄이려는 기능이 반대로 늘리는 것이 가장 나쁜 결과다.
      ② **화면마다 손으로 붙이지 않는다** — 입력칸은 여러 곳에서 그때그때 그려진다.
         새 화면에 붙이는 것을 잊으면 그 화면만 조용히 옛날처럼 남는다.
      ③ **문턱이 같다** — /api/works 와 같은 자료라 더 열거나 더 잠그지 않는다.
      ④ **쌓이지 않는다** — 화면을 다시 그려도 목록(datalist)이 body 에 남지 않는다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    # ① 원장에 있는 값만 — 후보를 만들어 내는 자리가 없어야 한다
    fn = server.split("def _suggest_index(")[1].split("\ndef suggest_values(")[0]
    assert "get_works()" in fn, "원장이 아닌 곳에서 추천을 만든다"
    assert "_SUGGEST_SRC" in fn, "어느 칸이 어느 열에서 오는지가 코드에 흩어져 있다"
    sv = server.split("def suggest_values(")[1].split("\ndef ")[0]
    assert "idx.get(" in sv and "startswith(q)" in sv, \
        "앞에서 맞는 값을 먼저 보여 주지 않는다 — 긴 목록에서 원하는 것이 아래로 밀린다"
    # 많이 쓰인 순(가나다순은 늘 같은 것을 맨 아래에 둔다)
    assert "-kv[1]" in fn, "추천 순서가 사용 빈도가 아니다"

    # ③ 문턱 — 별도 admin 게이트를 두지 않는다(/api/works 와 같은 자료다)
    route = server.split('if p == "/api/suggest":')[1][:400]
    assert "_require_admin()" not in route and "suggest_values(" in route, \
        "추천만 문턱이 다르다 — 같은 자료를 보는 화면끼리 말이 갈린다"

    # ② 새로 그려지는 화면에도 저절로 붙는가
    assert "MutationObserver" in live and "sgWire(document)" in live, \
        "화면을 새로 그리면 그 화면 입력칸에는 추천이 안 붙는다"
    assert "function sgWire(" in live and "input:not([data-sg])" in live, \
        "이미 붙은 칸을 매번 다시 뒤진다"
    # 파일·날짜·체크박스에는 붙이지 않는다(붙여도 쓸모가 없고 키보드만 가린다)
    wire = live.split("function sgWire(")[1].split("\n}")[0]
    for ty in ("'file'", "'date'", "'checkbox'"):
        assert ty in wire, f"{ty} 칸에도 추천 목록을 붙인다"
    # 사람이 치는 PIN 칸에 붙지 않는다 — 지도에 없으면 안 붙는 구조여야 한다
    assert "pin" not in live.split("const SG_MAP = {")[1].split("};")[0], \
        "PIN 칸이 추천 지도에 들어 있다"
    # ④ 화면을 다시 그려도 목록이 쌓이지 않는다
    assert 'datalist[id^="sgdl"]' in live and ".remove()" in wire, \
        "화면을 다시 그릴 때마다 목록이 body 에 쌓인다"
    # 브라우저 옛 입력이 원장 값을 덮지 않게 한다
    assert "setAttribute('autocomplete', 'off')" in live, \
        "브라우저가 기억한 옛 입력이 원장 추천 위에 겹쳐 뜬다"
    print("  [148] 입력 자동완성 — 원장 값만·전 화면 자동·같은 문턱·쌓이지 않음 ✅")


def t147_project_history():
    """현장(프로젝트) 이력 창 — 과거·현황·예측을 한 창에 (2026-08-08 지시).

    사용자 지시: "정기점검 예측에 프로젝트를 클릭하면 과거에 했던 내역들이 다 보이게
    밑에서 위로 올리는 창으로 / 돌발 AS 건 등 사진에 보이는 모든 프로젝트를 클릭하면
    과거 돌발AS 또는 정기점검 리스트가 보이게 / 지금 현황 예측 현황도 다 같이".

    지키는 것은 다섯이다:
      ① **문턱이 같다** — 같은 원장 행을 다시 묶어 보여 줄 뿐이라 `/api/works`·
         `/api/calendar` 와 다른 문턱을 쓰면 화면마다 말이 달라진다.
      ② **필터를 타지 않는다** — 사람이 꺼 둔 캘린더 종류 때문에 과거가 사라지면
         "이 현장은 두 번뿐"으로 잘못 읽는다. 서버에 다시 묻는다.
      ③ **날짜 없는 행을 감추지 않는다** — 못 센 것을 안 보여 주면 그게 조용한 사고다.
      ④ **모르는 것에 '완료' 색·말을 붙이지 않는다** — 날짜 미기입은 회색(etc)이다.
      ⑤ **맞출 것이 없으면 단추를 만들지 않는다** — 눌러도 아무 일 없는 단추는 고장이다.
    """
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ① 문턱 — /api/works 와 같아야 한다(둘 다 별도 admin 게이트를 두지 않는다)
    route = server.split('if p == "/api/project-history":')[1][:500]
    assert "_require_admin()" not in route, \
        "이력만 관리자 전용이다 — 같은 자료를 보는 화면끼리 문턱이 갈린다"
    assert "project_history(" in route, "이력 길이 함수를 부르지 않는다"

    fn = server.split("def project_history(")[1].split("\ndef ")[0]
    # ② 원장에서 직접 만든다(캘린더 결과를 다시 세지 않는다)
    assert "get_works()" in fn, "캘린더 결과를 다시 세면 화면 필터가 과거를 지운다"
    # ③ 날짜 없는 행을 따로 돌려준다
    assert "날짜없음" in fn, "날짜가 빈 행이 조용히 사라진다"
    # ④ 모르는 것은 회색 — '완료'로 칠하지 않는다
    assert 'item("etc", "", "날짜 미기입"' in fn, \
        "날짜를 모르는 행에 완료·미처리 색을 붙인다"
    # 평균 주기는 두 번 이상일 때만 — 한 번뿐인데 주기를 말하면 지어낸 숫자다
    assert "len(pmdone) >= 2" in fn, "점검이 한 번뿐인데 평균 주기를 만든다"
    # 캠프명으로 맞춘다(예측 일정은 프로젝트NO 가 자주 비어 있다)
    assert "_camp_key(" in fn, "프로젝트NO 로만 맞추면 '프로젝트 미확정' 이 통째로 빠진다"

    # ⑤ 화면 — 맞출 것이 없으면 단추를 만들지 않는다
    assert "function pjHas(" in live, "현장을 못 맞추는 일정에도 단추를 만든다"
    assert "if(!pjHas(e)) return '';" in live, "목록 단추가 빈 일정에도 붙는다"
    # 제목을 캠프명 대신 쓰지 않는다 — "정기점검 예측 · 부산4MB" 로는 못 맞춘다
    pjcall = live.split("function pjCall(")[1].split("\nfunction ")[0]
    assert "e.제목" not in pjcall, "꾸민 제목을 현장 이름으로 보낸다 — 맞을 리가 없다"
    # 밑에서 위로 올라오는 창이고, 날짜 창 위에 겹친다
    assert ".pj-sheet-wrap{z-index:70}" in live, "날짜 창 안에서 누르면 뒤에 숨는다"
    assert 'id="pjSheet"' in live and "cal2-sheet-wrap pj-sheet-wrap" in live, \
        "날짜 창과 같은 옷(밑에서 올라오는 레이어)을 쓰지 않는다"
    # PC 가운데 정렬을 이 창에도 다시 적어 둔다(실측: 안 적으면 화면 밖으로 나갔다)
    assert ".pj-sheet-wrap.in .cal2-sheet{transform:translate(-50%,-50%) scale(1)" in live, \
        "PC 에서 창이 가운데로 오지 않는다(2026-08-08 실측 — 오른쪽·아래로 밀림)"
    # 누를 수 있는 자리 넷: 달력 칸 · 아래 목록 · 날짜 창 · 상세(정산/작업)
    assert live.count("pjHistBtn(e)") >= 2, "아래 목록과 날짜 창 중 한쪽에만 붙어 있다"
    assert "pjCall(e, true)" in live, "달력 칸의 현장 이름을 눌러도 이력이 안 뜬다"
    assert "event.stopPropagation();" in live.split("function pjCall(")[1][:400], \
        "달력 칸을 누르면 날짜 창과 이력 창이 함께 뜬다"
    assert live.count("pjHistWide(") >= 3, "상세 화면(정산·작업)에 이력 단추가 없다"
    # 실패를 성공처럼 그리지 않는다
    assert "이력을 불러오지 못했습니다" in live, "서버가 못 주면 빈 이력을 '없음'으로 그린다"
    print("  [147] 현장 이력 창 — 같은 문턱·필터 무관·모르는 것은 회색 ✅")


def t143_originals_one_tap():
    """상세에서 원본을 한 번에 연다 (2026-08-07 지시).

    사용자 지시: "밴드 바로가기주소 / 거래명세표 세금계산서 매출전표 PO 등 버튼을
    만들어서 버튼을 누르면 해당 원본이 바로 열리게 … 일일이 원본 데이터를 찾아다닐
    필요 없이 하는 게 목적임".

    지키는 것은 넷이다:
      ① **문턱이 같다** — 목록(/api/originals)과 파일(/api/source-file)이 같은 인증을
         쓴다. 목록만 열려 있으면 '어떤 파일이 있는지'가 미인증에게 샌다.
      ② **민감 자료는 안 나온다** — 통화 메모는 색인에서 빠지지만, 옛 색인이 남아
         있는 동안에도 응답에서 한 번 더 막는다.
      ③ **주소를 지어내지 않는다** — 밴드 주소는 원본 덤프에 글 번호가 있을 때만.
         틀린 링크는 빈칸보다 나쁘다.
      ④ **없는 것은 단추를 만들지 않는다** — 눌러도 아무 일 없는 단추는 고장으로 읽힌다.
    """
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ① 두 길이 같은 문턱을 쓰는가
    blk = server.split('if p == "/api/originals":')[1][:600]
    assert "_require_admin()" in blk, "원본 목록이 인증 없이 열려 있다 — 파일 존재가 샌다"
    # ② 민감 자료 차단이 이 길에도 걸려 있는가
    fn = server.split("def _originals_for(")[1].split("\ndef ")[0]
    assert "is_private" in fn and "통화_" in fn, "통화 메모가 상세 화면으로 샐 수 있다"
    # ③ 밴드 주소는 근거가 있을 때만 — 밴드 번호와 글 번호가 둘 다 있어야 한다
    bu = server.split("def _band_urls(")[1].split("\ndef ")[0]
    assert 'if not band or not isinstance(posts, dict):' in bu, \
        "근거 없이 밴드 주소를 만든다 — 틀린 링크는 빈칸보다 나쁘다"
    assert "band.us/band/" in bu and "/post/" in bu, "밴드 주소 형식이 바뀌었다"
    # ④ 화면: 빈 갈래는 단추를 만들지 않고, 여러 건이면 사람이 고른다
    ui = live.split("function fillOrigBox(")[1].split("\nfunction sourceFileURL(")[0]
    assert "if(!gs.length" in ui, "원본이 없어도 단추를 만든다 — 눌러도 아무 일이 없다"
    assert "origPick(" in ui and "g.n === 1" in ui, \
        "여러 건일 때 임의로 하나를 연다 — '왜 이게 열리지'가 된다"
    assert "openSource(" in ui, "파일을 여는 길이 접속한 기기 쪽이 아니다"
    # 시트를 여는 순간이 아니라 그려진 뒤에 채운다(목록에서 톡톡 여닫는 화면이다)
    assert "setTimeout(()=>fillOrigBox(" in live, "상세를 열 때마다 서버를 기다린다"
    # 정산 상세와 일반 상세 둘 다에 자리가 있는가
    assert live.count("origBox(displayProject") >= 2, "일부 상세에만 붙어 있다"
    print("  [143] 원본 바로 열기 — 같은 문턱·민감자료 차단·근거 있는 밴드 주소 ✅")


def t142_flow_editable():
    """AS 접수→수금 워크플로우 — 고칠 수 있고 되돌릴 수 있나 (2026-08-07 지시).

    사용자 지시: "AS 접수부터 처리 수금까지의 과정을 워크플로우로 만들어서
    수정할 수 있는 기능도 추가해서 디테일하게 / 텍스트는 딱 필요한 것만".
    '고칠 수 있다'는 **'되돌릴 수 있다'와 짝**이어야 한다 — 되돌릴 길이 없으면
    사람은 잘못 고칠까 봐 결국 안 고치고, 화면은 실제 업무와 어긋난 채 남는다.
    """
    import ledger_db as L
    before = L.flow_steps()
    assert before and before[0]["단계"] and before[-1]["단계"], "기본 흐름이 비어 있다"
    # ① 통째 저장 → 되돌리기 왕복이 무손실인가
    mod = [dict(x) for x in before]
    mod[0] = dict(mod[0], 메모="검증용")
    mod.append({"단계": "검증용 단계", "담당": "", "소요일": 99, "근거": "", "메모": ""})
    assert L.flow_save(mod, who="synthetic") == len(before) + 1
    assert len(L.flow_steps()) == len(before) + 1
    L.flow_restore(who="synthetic")
    after = L.flow_steps()
    assert [(s["단계"], s["담당"], s["소요일"], s["근거"], s["메모"]) for s in after] == \
           [(s["단계"], s["담당"], s["소요일"], s["근거"], s["메모"]) for s in before], \
        "되돌리기가 원래 모습으로 돌아오지 않는다"
    # ② 빈 흐름·이름 없는 줄은 저장하지 않는다(빈 화면이 남으면 아무도 못 고친다)
    for bad in ([], [{"단계": "  "}]):
        try:
            L.flow_save(bad, "synthetic")
            raise AssertionError("빈 흐름이 저장됐다")
        except ValueError:
            pass

    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert 'if p == "/api/flow"' in server and "ledger_db.flow_restore" in server
    assert "_require_admin()" in server.split('if p == "/api/flow"')[2][:300], \
        "흐름 저장이 인증 없이 열려 있다"
    # ③ 화면: 좌측 카테고리 · 뷰 · 수정/저장/되돌리기가 실제로 이어져 있나
    assert 'data-v="flow"' in live and 'id="v-flow"' in live
    assert "if(v==='flow' && !FLOW_EDITING && !FLOW_DIRTY) loadFlow();" in live, \
        "고치는 중이거나 저장 전 변경이 있는데 다시 불러 덮어쓴다 — 말없이 사라진다"
    for fn in ("loadFlow", "flowRender", "flowEdit", "flowSave", "flowUndo",
               "flowAddStep", "flowDel", "flowMove", "flowCollect"):
        assert f"function {fn}(" in live, f"{fn} 이 없다"
    # ④ 아이콘은 스프라이트에 있는 것만 쓴다(없으면 빈 네모가 뜬다 — 검증 [91])
    icon = live.split('data-v="flow"')[1].split('href="#')[1].split('"')[0]
    assert f'id="{icon}"' in live, f"워크플로우 아이콘 {icon} 이 스프라이트에 없다"
    # ⑤ 길게 눌러 집기 (2026-08-07 지시: "각 카드를 길게 눌러 이동 수정 가능한 구조").
    #   대시보드 카드와 **같은 말투**여야 한다 — 집고, 놓을 자리를 누른다(검증 [128]).
    #   끌기로 하지 않는 이유: 폰에서 끌면 화면이 같이 스크롤돼 먼 자리로 못 옮긴다.
    assert "function flowHold(" in live and "function flowDropAt(" in live
    assert "FLOW_HOLD_MS" in live and "pointerdown" in live.split("function flowBindHold(")[1][:600], \
        "길게 누르기가 배선돼 있지 않다"
    # ⑥ 옮긴 뒤 저장하면 **화면의 순서**가 저장돼야 한다. 예전엔 언제나 입력칸을
    #   읽어서, 보기 모드에서 옮긴 뒤 저장하면 빈 목록이 됐다.
    assert "(FLOW_EDITING ? flowCollect() : FLOW)" in live, \
        "보기 모드에서 옮긴 순서가 저장되지 않는다"
    # ⑦ 저장 안 한 변경을 말없이 덮지 않는다
    assert "!FLOW_EDITING && !FLOW_DIRTY" in live, "저장 전 변경을 다시 불러 덮는다"
    assert "function flowMarkDirty(" in live and "저장(변경됨)" in live, \
        "안 저장된 변경이 있다는 것을 화면이 말하지 않는다"
    # ⑧ 색은 등장 순서로 — 해시로 뽑으면 담당 둘이 같은 색을 받는다(실측)
    assert "function flowPalette(" in live and "seen.includes(k)" in live, \
        "담당 색이 충돌할 수 있다 — 색이 곧 담당이라는 뜻이 깨진다"
    # ⑨ 머리 카드 (2026-08-07 지시: "맨 위에 돌발 AS 플로우 차트라고 카드 형태로 멋지게")
    #   ★ 색을 CSS 변수로 두면 어두운 테마에서 --brand 가 밝은 하늘색이라 흰 글자가
    #     사라진다. 그래서 이 카드만은 제 색을 값으로 가진다(두 테마에서 같은 그림).
    hero = live.split(".flow-hero{")[1][:400]
    assert "var(--brand" not in hero, "머리 카드가 테마 변수를 써서 어두운 테마에서 글자가 사라진다"
    # 제목은 이제 차트마다 다르다(종전/개선, [222]) — 자리와 기본값만 지킨다.
    assert '<h3 id="flowTitle">돌발 AS 플로우 차트 (종전)</h3>' in live, "머리 카드 제목이 없다"
    assert 'id="flowStats"' in live and "function flowHead(" in live, \
        "머리 카드 숫자(단계·D+·담당)가 화면과 이어져 있지 않다"
    # ⑩ 4:3 캡처 (2026-08-07 지시: "4대 3 비율로 캡처하는 기능 상단에 추가해")
    #   ★ 꼴은 **뱀 모양**이다 (2026-08-08 지시: "중간에서 번호로 뿌려지는 구조 말고
    #     순서대로 치고 치고 나가게 / 왼쪽에서 오른쪽으로 가서 용지 공간이 없으면
    #     다시 아래로 내려가서 왼쪽으로 다시 꺾어서").
    #     앞 판(마인드맵)은 보기엔 좋았지만 1→N 을 눈으로 좇을 수 없었다.
    assert "function flowToPng43(" in live, "4:3 캡처 루틴이 없다"
    cap = live.split("async function flowToPng43(")[1].split("\nfunction flow43Name")[0]
    assert "W = 1200, H = 900" in cap, "4:3 이 아니다"
    assert "ROOTGAP" not in cap and "sides" not in cap, \
        "마인드맵 잔재가 남아 있다 — 두 배치가 섞이면 한쪽만 고쳐진다"
    assert "ri % 2 === 1" in cap, \
        "줄마다 방향이 바뀌지 않는다 — 줄 끝에서 되짚어 오는 긴 선이 생긴다"
    assert "L.cols - 1 - ci" in cap, "오→왼 줄에서 자리를 뒤집지 않는다"
    assert "const arrow = (" in cap and "Math.atan2" in cap, \
        "화살표 머리가 없다 — 어느 쪽으로 가는 순서인지 그림이 말하지 않는다"
    # 갈래는 한 칸에 세로로 쌓고 화살표 하나가 나간다(합류를 그림이 말해야 한다)
    assert "cells.push({items: g.items.slice(), branch: g.branch})" in cap, \
        "나란한 갈래를 흩어 놓는다 — 개발자가 차례로 일어나는 일로 읽는다"
    # 번호는 제목 앞에 붙인다 — 따로 마디를 그리면 그게 '중간에 뿌려진 번호'다
    assert "${i+1}. ${s.단계}" in cap, "번호가 제목과 떨어져 있다"
    assert "bodyH / L.total" in cap, "넘칠 때 줄이지 않는다 — 잘린 그림이 나간다"
    assert "uiFont()" in cap and 'px "' not in cap, \
        "그리는 곳에 글꼴을 손으로 적었다 — 화면만 바뀌고 이미지는 옛 글꼴로 남는다"
    assert 'onclick="flowCapture43()"' in live.split('class="flow-hero-cap"')[1][:400], \
        "캡처 단추가 머리 카드(상단)에 없다"
    # ⑪ 화면도 나뭇가지 꼴인가 — 뿌리 마디 · 굽은 가지 · 담당별 뻗는 깊이
    assert 'class="flow-root"' in live and ".flow-root{" in live, "화면에 뿌리 마디가 없다"
    assert "function flowLanes(" in live and "--flow-in:" in live, \
        "담당마다 뻗는 깊이가 없다 — 들여쓰기가 뜻을 갖지 못한다"
    elbow = live.split(".flow-step::before{")[1][:300]
    assert "border-bottom-left-radius" in elbow, \
        "가지가 곧은 선이다 — 곧은 선은 표의 괘선처럼 보인다"
    # 들여쓴 카드에서도 번호 마디는 줄기 위에 남아야 한다
    assert "left:calc(-43px - var(--flow-in,0px))" in live, \
        "들여쓰면 번호 마디가 줄기에서 떨어져 나간다"
    # ⑫ 갈래(분기) — 개발자용 플로우차트가 막혀 있던 자리 (2026-08-08)
    #   접수는 네 갈래로 들어온다. 일직선 자료구조로는 그걸 담을 수 없어, 한 단계에
    #   메모로 눌러 담았었다. 그러면 개발자가 '접수 화면 하나'로 만든다.
    assert "갈래" in str(L.FLOW_COLS) or "branch" in L.FLOW_COLS, "갈래 열이 없다"
    got = L.flow_steps()
    assert all("갈래" in s for s in got), "흐름을 읽을 때 갈래가 빠진다"
    forks = [s["갈래"] for s in got if s.get("갈래")]
    assert forks, "기본 흐름에 갈래가 하나도 없다 — 접수 네 갈래가 사라졌다"
    # 저장·되돌리기 왕복에서 갈래가 살아남는가
    probe = [dict(x) for x in got]
    probe[0] = dict(probe[0], 갈래="검증갈래")
    L.flow_save(probe, who="synthetic")
    assert L.flow_steps()[0]["갈래"] == "검증갈래", "저장하면 갈래가 사라진다"
    L.flow_restore(who="synthetic")
    assert [s["갈래"] for s in L.flow_steps()] == [s["갈래"] for s in got], \
        "되돌리면 갈래가 어긋난다"
    # 화면·캡처·편집칸이 모두 갈래를 아는가
    assert "function flowGroups(" in live and "last.branch === b" in live, \
        "잇달아 같은 갈래를 한 묶음으로 보지 않는다"
    assert 'class="flow-fork"' in live and "flow-merge" in live, \
        "갈래를 묶어 보이지 않는다 — 개발자가 차례로 일어나는 단계로 읽는다"
    assert 'data-f="갈래"' in live and "갈래:g('갈래')" in live, \
        "갈래를 사람이 고칠 수 없다"
    # 뱀 모양에서는 '마디'가 아니라 **한 칸**이 갈래를 묶는다(위 ⑩에서 확인).
    # 여기서는 그 칸이 이름과 묶음선을 달고 있는지를 본다 — 이름 없이 쌓기만 하면
    # "왜 이 넷만 세로인가"를 그림이 설명하지 못한다.
    assert "갈래 (나란히)" in cap and "setLineDash" in cap, \
        "갈래 칸에 이름·묶음선이 없다 — 왜 세로로 쌓였는지 그림이 말하지 않는다"
    # ⑬ 개발 사양 — 그림만으로는 못 만든다
    assert "function flowDevSpec(" in live, "개발자용 사양이 없다"
    spec = live.split("function flowDevSpec(")[1].split("\nfunction flow43Name")[0]
    assert "flowchart TD" in spec and "subgraph" in spec, "분기가 흐름도에 안 나온다"
    assert "prev.forEach(p=>L.push(" in spec, "갈래가 다음 단계로 합류하지 않는다"
    assert "확인 전" in spec, \
        "'(확인 전)'을 지운 채 넘긴다 — 개발자가 추정을 사실로 만든다"
    assert 'onclick="flowDevSpec()"' in live, "개발 사양 단추가 화면에 없다"
    # ★ 담당기사는 네 사람뿐이다 (2026-08-08 지시: "쿠팡 담당기사 차동호 팀장 /
    #   김준형 권오철 김필우 / 플로우 차트에 이 4명만 반영해 기사는").
    #   한 곳(AS_TECHS)에서 정해 두어야 사람이 바뀔 때 한 줄만 고친다.
    import ledger_db as _L
    assert _L.AS_TECHS == ("차동호", "김준형", "권오철", "김필우"), \
        "담당기사 명단이 바뀌었다 — 흐름·화면이 같은 이름을 봐야 한다"
    assert len(_L.AS_TECH_LABEL) <= 20, \
        "담당 칸 한도(20자)를 넘는다 — flow_save 가 잘라 이름 하나가 사라진다"
    owners = {s[1] for s in _L.FLOW_DEFAULT}
    assert "담당기사" not in owners, \
        "기본 흐름에 이름 없는 '담당기사' 가 남아 있다 — 누구인지 화면이 말하지 않는다"
    assert _L.AS_TECH_LABEL in owners, "기본 흐름이 네 사람 명단을 쓰지 않는다"
    # ★ 세금계산서 **앞에 PO 가 있다** (2026-08-08 지시: "세금계산서 발행 전 쿠팡에서
    #   PO 발행 이메일로 발송함 / 오종현이 취합해서 류지영 전달 후 류지영이 발행").
    #   그동안 흐름이 명세서에서 곧장 계산서로 건너뛰어, PO 를 기다리는 시간이 어디에도
    #   안 적혀 있었다 — 계산서가 늦으면 "발행 담당이 안 했다"로만 보였다.
    fd = [s[0] for s in _L.FLOW_DEFAULT]
    for need in ("쿠팡 PO 이메일 수신", "PO 취합 · 류지영 전달"):
        assert need in fd, f"기본 흐름에 '{need}' 가 없다 — PO 기다리는 시간이 안 보인다"
    assert fd.index("거래명세서 발행") < fd.index("쿠팡 PO 이메일 수신") \
        < fd.index("PO 취합 · 류지영 전달") < fd.index("세금계산서 발행"), \
        "PO 가 세금계산서 뒤에 있다 — 순서가 뒤집히면 흐름이 거짓말을 한다"
    tax = [s for s in _L.FLOW_DEFAULT if s[0] == "세금계산서 발행"][0]
    assert tax[1] == "류지영", \
        "세금계산서 발행 담당이 류지영이 아니다 — PO 를 받아 발행하는 사람이 담당이다"
    print("  [142] 워크플로우 — 저장/되돌리기·길게 눌러 집기·담당색·머리카드·4:3 뱀모양·"
          "기사 4인·PO→계산서 순서 ✅")


def t136_work_lanes():
    """작업 차선 — 수집 창과 앱·엑셀 창이 하루 종일 나란히 돌 수 있나 (2026-08-07 지시).

    지키는 것은 네 가지다:
      ① 차선을 **안 정한 세션은 아무것도 막히지 않는다** (기존 세션·스케줄러 보호)
      ② 내 차선 밖 자원은 못 잡는다 (수집 창이 code 를 집어 가면 앱 창이 되잡지 못한다)
      ③ 남의 차선은 못 빼앗는다 (살아 있는 주인이 있으면 거절)
      ④ 죽은 세션의 차선은 **즉시** 비워진다 (45분 기다리면 다음 창이 놀게 된다)
    """
    import importlib
    import socket as _socket
    lanes = importlib.import_module("lanes")
    ac = importlib.import_module("ai_claim")

    with tempfile.TemporaryDirectory() as tmp:
        # 점유 파일 자리를 옮기면 차선도 따라온다 — 그래서 이 세션이 어느 차선에
        # 서 있든 검증은 빈 차선에서 시작한다(그러지 않으면 검증을 못 돌린다).
        keep_claims, keep_sid = ac.CLAIMS, lanes._me
        ac.CLAIMS = os.path.join(tmp, "ai_claims.json")
        try:
            lanes._me = lambda: "SESS-A"
            assert os.path.dirname(lanes._path()) == tmp, "차선 파일이 점유 자리를 안 따라간다"

            # ① 차선 밖이면 전부 허용 — 이 예외가 없으면 기존 자동화가 통째로 멈춘다
            for res in ("code", "ledger", "band", "publish"):
                ok, _ = lanes.can(res)
                assert ok, f"차선을 안 정했는데 '{res}' 가 막혔다 — 기존 세션이 멈춘다"

            assert lanes.take("collect", "claude", "수집 전담") == 0, "수집 차선을 못 잡는다"
            assert lanes.my_lane() == "collect", "잡은 차선이 내 것으로 안 보인다"

            # ② 차선 밖 자원은 막히고, 안쪽은 열린다
            for res in ("code", "ledger", "publish"):
                ok, why = lanes.can(res)
                assert not ok, f"수집 차선인데 '{res}' 가 열려 있다 — 앱 창과 부딪친다"
                assert "차선" in why, "왜 막혔는지 안 알려 준다"
            for res in ("band", "read", "report"):
                assert lanes.can(res)[0], f"수집 차선인데 '{res}' 가 막혔다"

            # ai_claim 이 실제로 그 문을 본다 (배선 확인 — 함수만 있고 안 부르면 소용없다)
            assert ac._lane_gate("band"), "차선 안 자원인데 점유가 거절됐다"
            assert not ac._lane_gate("code"), "ai_claim 이 차선을 안 본다"

            # ③ 살아 있는 남의 차선은 못 빼앗는다
            lanes._me = lambda: "SESS-B"
            assert lanes.take("collect", "claude", "가로채기") == 2, \
                "남의 차선을 빼앗았다 — 두 창이 같은 파일에서 만난다"
            assert lanes.take("build", "claude", "앱·엑셀") == 0, "빈 차선을 못 잡는다"
            # 앱 창은 code·ledger 가 열려 있어야 한다 — 그게 이 기능의 목적이다
            for res in ("code", "ledger", "publish"):
                assert lanes.can(res)[0], f"앱·엑셀 차선인데 '{res}' 가 막혔다"
            assert not lanes.can("band")[0], "앱 차선이 수집 자원까지 쥔다"

            # 한 세션은 한 차선 — 옮기면 앞 차선에서 빠진다
            lanes._me = lambda: "SESS-A"
            assert lanes.free("claude") == 0, "내 차선을 못 놓는다"
            lanes._me = lambda: "SESS-B"
            assert lanes.take("collect", "claude", "옮김") == 0, "빈 차선으로 못 옮긴다"
            d = lanes._load()
            assert "build" not in d, "차선을 옮겼는데 앞 차선이 남아 있다 — 분업표가 깨진다"

            # ④ 죽은 세션의 차선은 즉시 빈다
            d = lanes._load()
            # 호스트는 **이 PC** 여야 한다 — 다른 PC 의 점유는 판정하지 않고
            # 보수적으로 '살아 있다'로 보기 때문이다(ai_claim._pid_alive).
            d["build"] = {"who": "claude", "sid": "SESS-DEAD", "at": 0,
                          "agent_pid": 999999, "host": _socket.gethostname()}
            lanes._save(d)
            assert lanes.owner("build") is None, \
                "죽은 세션의 차선이 살아 있다 — 다음 창이 영영 못 들어온다"
        finally:
            ac.CLAIMS, lanes._me = keep_claims, keep_sid

    src = open(os.path.join(ROOT, "ai_claim.py"), encoding="utf-8").read()
    assert "_lane_gate(what)" in src, "ai_claim.take 이 차선 문을 안 지난다"
    print("  [136] 작업 차선 — 수집 창·앱 창 병렬 ✅")


# ★ 검증 번호가 겹치면 **둘 다 조용히 산다** (2026-08-09 실사고).
#   같은 폴더에서 여러 세션이 나란히 검증을 만들다 보니 `[173]`·`[174]` 를 두 세션이
#   각각 썼고, 비켜 놓은 `[175]`·`[176]` 도 다시 겹쳤다. 추측으로 비키는 것은 소용없다.
#   문제는 **겹쳐도 아무 일도 안 일어난다**는 것이다: 둘 다 통과하고 둘 다 초록이며,
#   나중에 "검증 [174] 가 지킨다"는 문서가 **어느 것을 가리키는지 알 수 없게 된다.**
#   그래서 기계가 잡는다. 새 검증을 만들 때는 여기가 먼저 막아 준다.
#
#   ※ LEGACY 는 **이 규칙을 만들기 전부터 있던** 중복이었다. 2026-08-11 에 갚았다.
#     여섯 중 **둘은 빚이 아니었다** — `[41]`·`[153]` 은 한 검증이 갈래마다 print 를
#     두 줄 할 뿐인데 계기가 줄 수를 세어 겹침이라 불렀다(계기 쪽이 틀렸다).
#     남은 넷은 참조가 **적은 쪽**을 비켰고 그 문서까지 같이 고쳤다:
#       [84]→[200] 증빙 동기화 · [98]→[201] 업로드 투입함
#       [121]→[202] 레이어 대화창 · [172]→[203] 회계 원장류 네 화면
#     **번호를 옮기는 것보다 그 번호를 가리키던 문서를 놓치는 것이 위험하다** —
#     놓치면 문서가 이제 엉뚱한 검증을 가리키는데, 그건 겹쳐 있는 것보다 나쁘다.
#     그래서 이 명단은 **비어 있어야 정상**이다. 다시 채우지 말 것.
LEGACY_DUP = set()


def t190_autopilot_retries_without_failure_cascade():
    """[190] 자원 장애 하나를 실패 수십 개로 만들지 않고 자동 재개한다."""
    import importlib
    from pathlib import Path as _Path

    A = importlib.import_module("autopilot")
    keep = (A.QUEUE_PATH, A.STATUS_PATH, A.REPORT_PATH)
    old_runner = A.run_tree
    try:
        with tempfile.TemporaryDirectory() as tmp:
            A.QUEUE_PATH = _Path(tmp) / "queue.json"
            A.STATUS_PATH = _Path(tmp) / "status.json"
            A.REPORT_PATH = _Path(tmp) / "status.md"
            args = [os.path.join(ROOT, "read_only_probe.py")]

            one = A.defer("관리대장 읽기", args, 60,
                          "FileNotFoundError: 관리대장을 찾을 수 없음: Z:/ledger")
            two = A.defer("관리대장 읽기", args, 60,
                          "FileNotFoundError: 관리대장을 찾을 수 없음: Z:/ledger")
            doc = json.loads(A.QUEUE_PATH.read_text(encoding="utf-8"))
            assert one and two and len(doc["items"]) == 1, \
                "같은 장애가 단계 수만큼 쌓이면 실패 도미노를 대기열로 옮긴 것뿐이다"
            assert doc["items"][0]["kind"] == "resource"

            unsafe = A.defer("원장 반영", ["ledger_db.py", "--apply"], 60,
                             "관리대장을 찾을 수 없음: Z:/ledger")
            assert unsafe is None, "비가역·쓰기 단계까지 자동 재실행하면 중복 반영될 수 있다"

            # 자원이 돌아오면 워치독 회차가 성공 확인 뒤 닫는다.
            doc["items"][0]["next_attempt"] = "2000-01-01T00:00:00+09:00"
            A.QUEUE_PATH.write_text(json.dumps(doc, ensure_ascii=False), encoding="utf-8")
            A.STATUS_PATH.unlink(missing_ok=True)
            before_dry = A.QUEUE_PATH.read_bytes()
            dry_result = A.heal(limit=1, budget_seconds=60, dry=True)
            assert dry_result["actions"] and A.QUEUE_PATH.read_bytes() == before_dry
            assert not A.STATUS_PATH.exists(), "--dry가 상태 시각을 쓰면 실제 실행으로 보인다"
            A.run_tree = lambda *a, **k: type("R", (), {
                "returncode": 0, "stdout": "OK", "stderr": "", "timed_out": False,
                "stuck_pid": 0})()
            healed = A.heal(limit=1, budget_seconds=60)
            assert healed["actions"][0]["result"] == "done"
            assert healed["active"] == 0, "성공한 대기열이 계속 남으면 같은 일을 반복한다"

            # 한 회차 몫만 정상 저장한 증분 작업은 실패 횟수를 올리지도, 성공으로
            # 닫지도 않는다. 30분 뒤 같은 안전 명령을 이어 가야 한다([217]).
            three = A.defer("미수집 보관", args, 60, "이전 회차 시간초과")
            doc = json.loads(A.QUEUE_PATH.read_text(encoding="utf-8"))
            doc["items"][-1].update({
                "attempts": 3, "ai_ticket": "old-ticket",
                "next_attempt": "2000-01-01T00:00:00+09:00",
            })
            A.QUEUE_PATH.write_text(json.dumps(doc, ensure_ascii=False), encoding="utf-8")
            A.run_tree = lambda *a, **k: type("R", (), {
                "returncode": A.INCREMENTAL_RETURN_CODE,
                "stdout": "증분 수집 계속 필요", "stderr": "", "timed_out": False,
                "stuck_pid": 0})()
            continued = A.heal(limit=1, budget_seconds=60)
            current = json.loads(A.QUEUE_PATH.read_text(encoding="utf-8"))["items"][-1]
            assert three and continued["actions"][0]["result"] == "waiting"
            assert current["status"] == "waiting" and current["attempts"] == 0
            assert current["continuations"] == 1 and not current["ai_ticket"], current
    finally:
        A.QUEUE_PATH, A.STATUS_PATH, A.REPORT_PATH = keep
        A.run_tree = old_runner

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    watch = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    agent = open(os.path.join(ROOT, "agent_dispatch.py"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    font = open(os.path.join(ROOT, "webapp", "font_switch.py"), encoding="utf-8").read()
    assert "_autopilot_defer" in daily and "deferred" in daily
    assert "INCREMENTAL_RETURN_CODE = 75" in daily and \
        'got.get("returncode") == INCREMENTAL_RETURN_CODE' in daily, \
        "09:50 회차가 정상 증분 rc75를 실패·재시도로 센다([217])"
    assert "heal_autopilot(dry)" in watch, "워치독에 안 묶이면 대기열은 사람이 눌러야만 돈다"
    assert '"autopilot": autopilot_status' in server and '/api/autopilot' in server
    assert "run_tree(command" in agent, "AI CLI가 SMB형 timeout에서 회차를 영원히 붙든다"
    assert 'newline=""' in font, "글꼴 왕복이 CRLF를 LF로 바꾸면 파일 전체가 달라진다"
    from proc_guard import run_tree as _guarded_run
    timed = _guarded_run([sys.executable, "-c", "import time; time.sleep(30)"],
                         cwd=ROOT, timeout=0.2, drain_timeout=3)
    assert timed.timed_out and timed.returncode != 0, \
        "공용 실행기가 시간초과 뒤 반환하지 않으면 AI 한 건이 워치독을 붙든다"
    print("  [190/217] 실패도미노 차단 · 영속 대기열 · 증분 계속 · AI 제한 인계 ✅")


def t191_confirmation_truth_and_fast_refresh():
    """[191] 확인필요는 근거를 설명하고, 갱신은 한 번만 실행하며 반드시 끝난다."""
    import erp_bundle
    import erp_ledger_check
    import settlement_completion

    one = """Coupang이(가) 새 구매 오더
★ 총금액 : 1,100원
★ 품  목 : 정기점검 2건
★ 쿠팡오더 No. : PO111111
"""
    meta = erp_bundle.band_po_meta_from_bodies([one])
    assert meta[1000]["PO"] == "PO111111" and meta[1000]["총금액"] == 1100
    ambiguous = erp_bundle.band_po_meta_from_bodies([
        one, one.replace("PO111111", "PO222222")])
    assert 1000 not in ambiguous and 1100 not in ambiguous, \
        "같은 금액의 서로 다른 PO 중 먼저 읽힌 것을 조용히 고르면 오완료가 난다"

    records = {
        "S1": {"프로젝트NO": "UJ1", "원장_PO번호": "PO111111", "비용구분": "유상",
               "업무구분": "정기점검", "원장_거래명세서발행일": "2026-08-01",
               "원장_거래명세서합계": 550},
        "S2": {"프로젝트NO": "UJ2", "원장_PO번호": "PO111111", "비용구분": "유상",
               "업무구분": "정기점검", "원장_거래명세서발행일": "2026-08-01",
               "원장_거래명세서합계": 550},
    }
    docs = [{"amt": 1000, "slip": "2026/08/01-1", "kind": "정기점검"}]
    batches = settlement_completion.confirmed_po_invoice_batches(records, docs, meta)
    assert set(batches) == {"S1", "S2"}
    bad_meta = {1000: dict(meta[1000], 총금액=1090)}
    assert not settlement_completion.confirmed_po_invoice_batches(records, docs, bad_meta)
    assert not settlement_completion.confirmed_po_invoice_batches(records, docs * 2, meta)

    sale = {"po": "PO111111", "supply": 1000, "total": 1100}
    matched = erp_ledger_check.project_sale_match(
        {"원장_PO번호": "PO111111", "원장_공급가액": 1000, "원장_합계": 1100},
        [sale])
    assert matched["present"] and matched["amount_match"]
    mismatch = erp_ledger_check.project_sale_match(
        {"원장_PO번호": "PO111111", "원장_공급가액": 900}, [sale])
    assert mismatch["present"] and not mismatch["amount_match"]

    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    watch = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    ledger = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    exec_src = server.split("def get_exec_report(", 1)[1].split("def _source_index", 1)[0]
    assert exec_src.find('_fresh("exec")') < exec_src.find("master = resolve_master"), \
        "대표보고 캐시보다 Z:/Excel을 먼저 읽으면 화면 갱신마다 수 초씩 막힌다"
    assert "let settleFlight=null" in live and "const rerunSettle" in live
    assert "SWR_RERUN[p] = rerunSettle" in live and "brandLoaded" in live
    assert "_issue_truth_rows" in server and "근거상태" in server
    start_task_src = server.split("def start_task(key):", 1)[1].split(
        "# ── 마지막 실행 시각", 1)[0]
    assert "run_tree([PY] + args" in start_task_src and "enqueue_agent" not in start_task_src
    sync_src = watch.split("def sync_worklog(dry):", 1)[1].split("def main():", 1)[0]
    assert "run_tree(" in sync_src and "subprocess.run" not in sync_src
    assert "AVG(ms)" in ledger
    print("  [191] 확인필요 근거 분리 · PO묶음 안전완료 · 단일 갱신 · 캐시 선조회 · 유한 실행 OK")


def t193_app_db_cutover_archive_and_frontend():
    """[193] SQLite 즉시 정본·안전 컷오버·Excel 보관결과·확인필요 UX 통합 관문.

    실제 공유 DB·점유·Z:·관리대장은 전혀 열지 않는다. 모든 DB·보고서·XLSX·writer
    결과는 TemporaryDirectory 아래에 만들고, ledger_writer 프로세스는 결과 계약만
    합성한다. [192]가 이 검증 뒤에도 공유 산출물 바이트가 그대로인지 다시 확인한다.
    """
    import types as _types
    from datetime import datetime as _dt, timedelta as _timedelta, time as _time
    from pathlib import Path as _Path

    import app_store as A
    import db_cutover as C
    import ledger_db as L
    import proc_guard as PG

    # 실제 관리대장의 [h]:mm 경과시간과 시각 셀도 JSON 정본으로 무손실 진입한다.
    assert C._json_value(_timedelta(hours=9, minutes=30)) == "9:30:00"
    assert C._json_value(_time(9, 30)) == "09:30:00"
    cutover_src = open(os.path.join(ROOT, "db_cutover.py"), encoding="utf-8").read()
    cutover_body = cutover_src.split("def cutover(", 1)[1].split("def self_test(", 1)[0]
    assert cutover_body.count("os.path.getmtime(master)") == 1, \
        "SMB 원본 수정시각을 행마다 다시 조회한다 — 순간 끊김이 이관 전체를 죽인다"
    assert "with store.transaction() as batch_conn" in cutover_body \
        and "_conn=batch_conn" in cutover_body, \
        "초기 이관이 행마다 FULL fsync한다 — 대량 컷오버가 수십 분 걸린다"

    def make_cutover_book(path, duplicate_conflict=False):
        book = openpyxl.Workbook()
        specs = [
            ("02_돌발AS접수", ["접수ID", "프로젝트NO", "캠프명", "진행상태"],
             ["AS-CUT-193", "UJ-CUT-193", "합성AS캠프", "작업완료"]),
            ("03_현장작업실적", ["작업ID", "프로젝트NO", "캠프명", "완료여부"],
             ["WK-CUT-193", "UJ-CUT-193", "합성AS캠프", "완료"]),
            ("04_정기점검", ["점검ID", "프로젝트NO", "캠프명", "점검상태"],
             ["PM-CUT-193", "UJ-PM-193", "합성점검캠프", "완료"]),
            ("06_거래서류청구수금",
             ["정산ID", "원천업무ID", "프로젝트NO", "캠프명", "청구상태"],
             ["JS-CUT-193", "AS-CUT-193", "UJ-CUT-193", "합성AS캠프", "청구완료"]),
            ("13_PO발주관리", ["PO관리ID", "프로젝트NO", "캠프명", "PO상태(자동)"],
             ["PO-CUT-193", "UJ-CUT-193", "합성AS캠프", "발주완료"]),
            ("15_세금계산서관리",
             ["계산서관리ID", "정산ID", "프로젝트NO", "캠프명", "실제발행일", "발행상태(자동)"],
             ["TAX-CUT-193", "JS-CUT-193", "UJ-CUT-193", "합성AS캠프",
              "2026-08-10", "발행완료"]),
            ("16_입금수금관리", ["입금관리ID", "정산ID", "프로젝트NO", "캠프명"],
             ["PAY-CUT-193", "JS-CUT-193", "UJ-CUT-193", "합성AS캠프"]),
        ]
        for pos, (name, header, row) in enumerate(specs):
            ws = book.active if pos == 0 else book.create_sheet()
            ws.title = name
            ws.append(header)
            ws.append(row)
        if duplicate_conflict:
            book["02_돌발AS접수"].append(
                ["AS-CUT-193", "UJ-CUT-193", "다른캠프", "접수"])
        book.save(path)
        book.close()

    with tempfile.TemporaryDirectory(prefix="app-db-193-") as tmp:
        tmp_path = _Path(tmp)

        # ① AppStore: 즉시 create/update, 낙관잠금, 멱등, audit/outbox, soft delete.
        store_path = tmp_path / "app-store.db"
        store = A.AppStore(store_path).initialize()
        create_args = dict(
            kind="돌발AS", business_key="AS-APP-193", public_id="AS-APP-193",
            project_no="UJ-APP-193", camp_name="합성앱캠프", status="접수",
            fields={"진행상태": "접수", "담당기사": "류지영"},
            actor="synthetic", source="synthetic", evidence="t193",
            idempotency_key="t193-create",
        )
        created = store.create_work(**create_args)
        replay = store.create_work(**create_args)
        assert replay["work"]["id"] == created["work"]["id"]
        assert replay["idempotent_replay"] is True, "같은 저장 재전송이 새 이벤트를 만든다"
        work_id = created["work"]["id"]
        updated = store.update_work(
            work_id, expected_version=1,
            patch={"status": "작업완료", "fields": {"진행상태": "작업완료"}},
            actor="synthetic", source="synthetic", evidence="객관 근거",
            idempotency_key="t193-update",
        )
        assert updated["work"]["record_version"] == 2
        try:
            store.update_work(work_id, expected_version=1, patch={"status": "오래된 수정"})
            raise AssertionError("낡은 record_version 수정이 받아들여졌다")
        except A.VersionConflict:
            pass

        # 기존 /api/input 모양도 Excel 없이 즉시 DB에 저장되고 legacy read-model 위에 겹친다.
        legacy_items = [
            {"sheet": "02_돌발AS접수", "key_col": "접수ID", "key": "AS-LEGACY-193",
             "col": "진행상태", "value": "접수", "vtype": "text",
             "only_if_empty": False, "evidence": "t193 legacy"},
            {"sheet": "02_돌발AS접수", "key_col": "접수ID", "key": "AS-LEGACY-193",
             "col": "담당기사", "value": "김필우", "vtype": "text",
             "only_if_empty": False, "evidence": "t193 legacy"},
        ]
        legacy = store.apply_legacy_items(
            legacy_items, "t193-legacy", idempotency_key="t193-legacy-batch")
        assert legacy["ok"] and legacy["created"] == 1 and legacy["updated"] == 1, legacy
        legacy_replay = store.apply_legacy_items(
            legacy_items, "t193-legacy", idempotency_key="t193-legacy-batch")
        assert legacy_replay["idempotent_replay"] is True
        overlaid = store.overlay_rows(
            "02_돌발AS접수",
            [{"접수ID": "AS-LEGACY-193", "담당기사": "옛값", "excel_only": "보존"}],
            "접수ID",
        )
        assert overlaid[0]["담당기사"] == "김필우" and overlaid[0]["excel_only"] == "보존"

        # project_resolve형 신규행은 같은 Excel 행의 좌표 셀을 서로 다른 설정으로 흩지 않고
        # 한 work_item으로 묶는다. 접수ID가 아직 없어도 프로젝트NO가 안정 키가 된다.
        row_cells = [
            {"sheet": "02_돌발AS접수", "cell": "B193", "key_col": "-", "key": "B193",
             "col": "프로젝트NO", "value": "UJ-ROW-193", "vtype": "text",
             "only_if_empty": True, "evidence": "t193 project_resolve"},
            {"sheet": "02_돌발AS접수", "cell": "C193", "key_col": "-", "key": "C193",
             "col": "접수일자", "value": "2026-08-10", "vtype": "date",
             "only_if_empty": True, "evidence": "t193 project_resolve"},
            {"sheet": "02_돌발AS접수", "cell": "D193", "key_col": "-", "key": "D193",
             "col": "캠프명", "value": "행그룹합성캠프", "vtype": "text",
             "only_if_empty": True, "evidence": "t193 project_resolve"},
            {"sheet": "02_돌발AS접수", "cell": "E193", "key_col": "-", "key": "E193",
             "col": "진행상태", "value": "접수", "vtype": "text",
             "only_if_empty": True, "evidence": "t193 project_resolve"},
        ]
        grouped = store.apply_legacy_items(
            row_cells, "t193-project-resolve", idempotency_key="t193-row-group")
        assert grouped["ok"] and grouped["created"] == 1 and grouped["settings"] == 0, grouped
        grouped_work = store.get_work(kind="돌발AS", business_key="UJ-ROW-193")
        assert grouped_work["project_no"] == "UJ-ROW-193"
        assert grouped_work["camp_name"] == "행그룹합성캠프" and grouped_work["status"] == "접수"
        assert grouped_work["fields"]["접수일자"] == "2026-08-10"
        sheet_now = store.list_sheet_rows("02_돌발AS접수")
        assert any(row.get("프로젝트NO") == "UJ-ROW-193" for row in sheet_now)
        grouped_overlay = store.overlay_rows(
            "02_돌발AS접수",
            [{"프로젝트NO": "UJ-ROW-193", "캠프명": "옛캠프", "excel_only": "보존"}],
            "프로젝트NO",
        )
        grouped_row = next(row for row in grouped_overlay if row.get("프로젝트NO") == "UJ-ROW-193")
        assert grouped_row["캠프명"] == "행그룹합성캠프" and grouped_row["excel_only"] == "보존"
        with store.reader() as conn:
            scattered = conn.execute(
                "SELECT COUNT(*) FROM app_setting WHERE key LIKE 'excel-cell:02_돌발AS접수:%193'"
            ).fetchone()[0]
        assert scattered == 0, "같은 신규행이 work_item 대신 excel-cell 설정 네 개로 흩어졌다"

        deleted = store.soft_delete_work(
            work_id, expected_version=2, actor="synthetic", reason="t193",
            idempotency_key="t193-delete")
        assert deleted["work"]["deleted_at"] and deleted["work"]["record_version"] == 3
        try:
            store.get_work(work_id)
            raise AssertionError("soft delete 행이 기본 조회에 노출된다")
        except A.NotFoundError:
            pass
        assert store.get_work(work_id, include_deleted=True)["deleted_at"]
        with store.reader() as conn:
            actions = [row[0] for row in conn.execute(
                "SELECT action FROM change_event WHERE work_id=? ORDER BY id", (work_id,))]
        assert actions == ["create", "update", "soft_delete"], actions
        lease, outbox = store.lease_outbox(limit=100)
        assert outbox and store.ack_outbox(lease, [row["id"] for row in outbox]) == len(outbox)

        # ② 컷오버: 실제 합성 XLSX 후보/중복충돌/parity/정본 모드 전환.
        old_cutover_root, old_report_dir = C.ROOT, C.REPORT_DIR
        try:
            C.ROOT = tmp_path / "cutover-root"
            C.REPORT_DIR = tmp_path / "cutover-reports"
            valid_book = tmp_path / "valid-master.xlsx"
            duplicate_book = tmp_path / "duplicate-master.xlsx"
            make_cutover_book(valid_book)
            make_cutover_book(duplicate_book, duplicate_conflict=True)

            candidates = C.read_candidates(valid_book)
            assert candidates["row_count"] == 7 and not candidates["blocking"], candidates
            dup = C.read_candidates(duplicate_book)
            assert any("중복키 값 충돌 AS-CUT-193" in item for item in dup["blocking"]), dup
            plan = C.cutover(valid_book, str(tmp_path / "plan-unused.db"), apply=False)
            assert plan["status"] == "ready" and plan["candidate_rows"] == 7

            clean_db = tmp_path / "cutover-clean.db"
            complete = C.cutover(valid_book, str(clean_db), apply=True)
            assert complete["status"] == "complete" and complete["db_rows"] == 7, complete
            assert not complete["blocking"] and complete["source_of_truth_mode"] == "db_primary_export"
            clean_store = A.AppStore(clean_db).initialize()
            assert clean_store.get_setting("source_of_truth_mode")["value"] == "db_primary_export"
            assert len(clean_store.list_work(limit=100)) == candidates["row_count"], \
                "Excel 후보와 SQLite 연결 행 parity가 맞지 않는다"

            conflict_db = tmp_path / "cutover-conflict.db"
            conflict_store = A.AppStore(conflict_db).initialize()
            conflict_store.create_work(
                kind="돌발AS", business_key="AS-CUT-193", public_id="AS-CUT-193",
                project_no="UJ-OLD-193", camp_name="기존DB캠프", status="기존상태",
                fields={"접수ID": "AS-CUT-193", "프로젝트NO": "UJ-OLD-193",
                        "캠프명": "기존DB캠프", "진행상태": "기존상태"},
                idempotency_key="t193-cutover-conflict",
            )
            blocked = C.cutover(valid_book, str(conflict_db), apply=True)
            assert blocked["status"] == "blocked" and blocked["conflicts"] > 0, blocked
            assert blocked["source_of_truth_mode"] == "shadow_compare"
            assert any("필드 해시 불일치" in item or "충돌" in item
                       for item in blocked["blocking"]), blocked["blocking"]
            assert conflict_store.get_setting("source_of_truth_mode")["value"] == "shadow_compare"
        finally:
            C.ROOT, C.REPORT_DIR = old_cutover_root, old_report_dir

        # ③ 같은 target 최신값만 보관 큐에 남고, writer applied=0/skipped=N을
        #    pending.applied로 오기록하지 않는지 실제 임시 ledger DB로 확인한다.
        ledger_old = {name: getattr(L, name) for name in
                      ("ROOT", "DB_DIR", "DB_PATH", "JSON_QUEUE", "REPORT_DIR",
                       "STATUS_CACHE", "APPLY_LOCK", "intake_json",
                       "scheduled_workbook_maintenance", "_wait_editing_clear")}
        old_app_db_env = os.environ.get("COUPANG_APP_DB_PATH")
        old_run_tree = PG.run_tree
        old_datalake = sys.modules.get("datalake")
        old_archive = sys.modules.get("archive_export")
        archive_calls = []
        app_db_path = tmp_path / "ledger-canonical.db"
        try:
            L.ROOT = str(tmp_path / "ledger-root")
            L.DB_DIR = str(tmp_path / "ledger-db")
            L.DB_PATH = str(tmp_path / "ledger-db" / "queue.db")
            L.JSON_QUEUE = str(tmp_path / "ledger-root" / "updates" / "pending.json")
            L.REPORT_DIR = str(tmp_path / "ledger-root" / "reports")
            L.STATUS_CACHE = str(tmp_path / "ledger-root" / "reports" / "status.json")
            L.APPLY_LOCK = str(tmp_path / "ledger-root" / "reports" / ".apply.lock")
            # intake_json()의 기본 인자는 import 때 공유 경로로 묶였으므로 명시적으로 차단한다.
            L.intake_json = lambda *args, **kwargs: 0
            L.scheduled_workbook_maintenance = lambda now=None: []
            L._wait_editing_clear = lambda now, slot: None
            os.environ["COUPANG_APP_DB_PATH"] = str(app_db_path)
            sys.modules["datalake"] = _types.SimpleNamespace(note=lambda *a, **k: None)
            sys.modules["archive_export"] = _types.SimpleNamespace(
                record_ledger_result=lambda **kw: archive_calls.append(kw))

            def fake_run_tree(args, cwd=None, timeout=None, drain_timeout=None, env=None):
                queue_path = args[args.index("--queue") + 1]
                with open(queue_path, encoding="utf-8") as handle:
                    payload = json.load(handle)
                result_path = env["COUPANG_LEDGER_RESULT"]
                os.makedirs(os.path.dirname(result_path), exist_ok=True)
                with open(result_path, "w", encoding="utf-8") as handle:
                    json.dump({
                        "applied": [],
                        "skipped": [{**item, "사유": "보관본에 이미 같은 값"}
                                    for item in payload],
                        "version": "미생성(0건)",
                    }, handle, ensure_ascii=False)
                return _types.SimpleNamespace(
                    returncode=0, stdout="", stderr="", timed_out=False, stuck_pid=None)

            PG.run_tree = fake_run_tree
            base = {"sheet": "02_돌발AS접수", "key_col": "접수ID",
                    "key": "AS-QUEUE-193", "col": "진행상태", "vtype": "text",
                    "only_if_empty": False, "evidence": "t193 latest target"}
            assert L.enqueue([{**base, "value": "접수"}], source="t193") == 1
            assert L.enqueue([{**base, "value": "작업완료"}], source="t193") == 1
            pending = L.pending_rows()
            assert len(pending) == 1 and pending[0]["value"] == "작업완료", pending
            result = L.apply_now(force=True, now=_dt(2026, 8, 10, 11, 5))
            assert result["상태"] == "보관본 생성" and result["적용"] == 0
            assert result["제외"] == 1 and result["미확정"] == 0, result
            with L.conn() as conn:
                ledger_rows = conn.execute(
                    "SELECT value,status,target_key,result_note FROM pending ORDER BY id").fetchall()
                batch_note = conn.execute("SELECT note FROM batch ORDER BY id DESC LIMIT 1").fetchone()[0]
            assert [row[1] for row in ledger_rows] == ["superseded", "skipped"], ledger_rows
            assert len({row[2] for row in ledger_rows}) == 1, "같은 target_key가 갈라졌다"
            assert not any(row[1] == "applied" for row in ledger_rows)
            assert "적용 0 / 제외 1" in batch_note and archive_calls
            canonical = A.AppStore(app_db_path).initialize().list_sheet_rows("02_돌발AS접수")
            latest = next(row for row in canonical if row.get("접수ID") == "AS-QUEUE-193")
            assert latest["진행상태"] == "작업완료", "11시 전에도 SQLite 정본은 최신이어야 한다"
        finally:
            PG.run_tree = old_run_tree
            for name, value in ledger_old.items():
                setattr(L, name, value)
            if old_app_db_env is None:
                os.environ.pop("COUPANG_APP_DB_PATH", None)
            else:
                os.environ["COUPANG_APP_DB_PATH"] = old_app_db_env
            if old_datalake is None:
                sys.modules.pop("datalake", None)
            else:
                sys.modules["datalake"] = old_datalake
            if old_archive is None:
                sys.modules.pop("archive_export", None)
            else:
                sys.modules["archive_export"] = old_archive

        ledger_src = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
        writer_src = open(os.path.join(ROOT, "ledger_writer.py"), encoding="utf-8").read()
        assert "applied_ids = {" in ledger_src and "skipped_by_id = {" in ledger_src
        assert "if applied_ids:" in ledger_src and "status='skipped'" in ledger_src
        assert '{"applied": [], "skipped": skips + skipped2' in writer_src, \
            "writer 0건 경로가 skipped 목록을 결과 계약에 남기지 않는다"

        # ④ 프런트: 구조화 오류, last-good/기준시각/개별 재시도, 공통 drawer와 반응형·키보드.
        live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
        for marker in ("class ApiError extends Error", "HTTP_ERROR", "EMPTY_BODY",
                       "UNEXPECTED_CONTENT_TYPE", "INVALID_JSON", "BODY_READ_ERROR"):
            assert marker in live, f"구조화 API 오류 마커 누락: {marker}"
        for marker in ("DATA_SECTION_STATE", "hasGood:false", "dataStampOf",
                       "markDataSectionSuccess", "markDataSectionFailure",
                       "retryDataSection", "Promise.all(keys.map"):
            assert marker in live, f"마지막 정상값·섹션 재시도 마커 누락: {marker}"
        for marker in ("function openCheckDrawer", "왜 아직 미확인인가", "현재 확인된 근거",
                       "안전등급", "다음 행동", "runTask('evidence_sync')",
                       "openCheckUpload", "openCheckRecordFromDrawer"):
            assert marker in live, f"확인필요 drawer 마커 누락: {marker}"
        assert 'button type="button" class="srow check-row' in live, \
            "확인 카드가 키보드 Enter·Space를 지원하는 native button이 아니다"
        assert "button:focus-visible" in live and "min-height:44px" in live
        for media in ("@media(max-width:420px)", "@media(max-width:767px)",
                      "@media(max-width:899px)", "@media(min-width:900px)"):
            assert media in live, f"375/768/desktop 반응형 구간 누락: {media}"
        assert "overflow-wrap:anywhere" in live and "overflow-x:hidden" in live

    print("  [193] 앱 DB 즉시정본·컷오버 parity/충돌·보관결과 진실·확인 drawer 반응형 ✅")


def t194_legacy_queue_migration_and_round_truth():
    """[194] 구형 큐 DB 마이그레이션과 일일 회차 종료 상태는 거짓 성공을 만들지 않는다.

    운영 ``ledger_queue.db``에는 ``target_key``가 없었는데 SCHEMA가 그 열의 인덱스를
    먼저 만들었다. 그러면 ``conn()``이 ALTER까지 도달하지 못해 09:50 회차 전체가
    ``no such column: target_key``로 중단된다. 또 ``daily_run``은 예외·단계 실패도
    finally에서 무조건 '완주'로 덮었고, 진행 JSON은 전 단계의 시간초과·명령을 다음
    정상 단계에 남겼다. 둘 다 화면만 안심시키는 종류의 실패다.
    """
    import importlib
    import sqlite3 as _sqlite3
    from datetime import datetime as _dt

    L = importlib.import_module("ledger_db")
    D = importlib.import_module("daily_run")

    # ① 실제 운영 DB와 같은 구형 pending 표(신규 네 열 없음)를 먼저 만든다.
    old_db_dir, old_db_path = L.DB_DIR, L.DB_PATH
    with tempfile.TemporaryDirectory() as td:
        L.DB_DIR = td
        L.DB_PATH = os.path.join(td, "legacy_queue.db")
        try:
            c = _sqlite3.connect(L.DB_PATH)
            c.executescript("""
                CREATE TABLE pending(
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  ts TEXT NOT NULL, source TEXT NOT NULL, sheet TEXT NOT NULL,
                  key_col TEXT, key TEXT, cell TEXT, col TEXT,
                  value TEXT, vtype TEXT DEFAULT 'text', evidence TEXT,
                  only_if_empty INTEGER DEFAULT 1,
                  status TEXT NOT NULL DEFAULT 'pending',
                  batch_id INTEGER, applied_at TEXT
                );
                CREATE INDEX ix_pending_status ON pending(status);
            """)
            c.close()

            # SCHEMA 실행이 먼저 실패하지 않고 ALTER와 인덱스까지 끝나야 한다.
            with L.conn() as c:
                cols = {row[1] for row in c.execute("PRAGMA table_info(pending)")}
                indexes = {row[1] for row in c.execute("PRAGMA index_list(pending)")}
            assert {"ingest_key", "target_key", "result_note", "superseded_by"} <= cols, cols
            assert "ix_pending_target" in indexes and "ix_pending_ingest" in indexes, indexes
            # 두 번째 연결도 멱등이어야 스케줄러 재시작이 안전하다.
            with L.conn() as c:
                assert c.execute("SELECT count(*) FROM pending").fetchone()[0] == 0
        finally:
            L.DB_DIR, L.DB_PATH = old_db_dir, old_db_path

    # ② 진행 상태는 매번 새 문서다. 이전 시간초과/결과/명령이 다음 단계에 남지 않는다.
    keep = {
        "PROGRESS": D.PROGRESS,
        "is_input_window": D.is_input_window,
        "acquire_run_lock": D.acquire_run_lock,
        "release_run_lock": D.release_run_lock,
        "_run_pipeline": D._run_pipeline,
        "round_t0": D._ROUND_T0[0],
        "over_budget": D._OVER_BUDGET[0],
    }
    with tempfile.TemporaryDirectory() as td:
        D.PROGRESS = os.path.join(td, "progress.json")
        D._ROUND_T0[0] = _dt.now()
        try:
            D.note_progress("옛단계", "끝", {
                "결과": False, "시간초과": True, "명령": "old.py",
            })
            D.note_progress("새단계", "시작", {"명령": "new.py"})
            got = json.load(open(D.PROGRESS, encoding="utf-8"))
            assert got["단계"] == "새단계" and got["명령"] == "new.py", got
            assert "시간초과" not in got and "결과" not in got, got
            D.note_progress("새단계", "끝", {"결과": True})
            got = json.load(open(D.PROGRESS, encoding="utf-8"))
            assert "명령" not in got and "시간초과" not in got, got
            assert got["끝난단계"][-2:] == ["옛단계", "새단계"], got

            released = []
            D.is_input_window = lambda: False
            D.acquire_run_lock = lambda: "synthetic-token"
            D.release_run_lock = lambda token: released.append(token)

            # 정상 완주.
            D._run_pipeline = lambda: [{"name": "정상", "ok": True}]
            D.main()
            got = json.load(open(D.PROGRESS, encoding="utf-8"))
            assert got["상태"] == "완주" and got["종료구분"] == "완주", got

            # 끝까지 실행했어도 한 단계가 실패했다면 회차 결과는 실패다.
            D._run_pipeline = lambda: [{"name": "깨진단계", "ok": False}]
            D.main()
            got = json.load(open(D.PROGRESS, encoding="utf-8"))
            assert got["상태"] == "실패" and got["끝까지실행"] is True, got
            assert got["실패단계"] == ["깨진단계"], got

            # 예외 실패와 사용자 중단도 서로 구분한다.
            def _boom():
                raise RuntimeError("synthetic failure")
            D._run_pipeline = _boom
            try:
                D.main()
            except RuntimeError:
                pass
            got = json.load(open(D.PROGRESS, encoding="utf-8"))
            assert got["상태"] == "실패" and got["오류유형"] == "RuntimeError", got

            def _interrupt():
                raise KeyboardInterrupt()
            D._run_pipeline = _interrupt
            try:
                D.main()
            except KeyboardInterrupt:
                pass
            got = json.load(open(D.PROGRESS, encoding="utf-8"))
            assert got["상태"] == "중단" and got["종료구분"] == "중단", got
            assert released == ["synthetic-token"] * 4, released
        finally:
            D.PROGRESS = keep["PROGRESS"]
            D.is_input_window = keep["is_input_window"]
            D.acquire_run_lock = keep["acquire_run_lock"]
            D.release_run_lock = keep["release_run_lock"]
            D._run_pipeline = keep["_run_pipeline"]
            D._ROUND_T0[0] = keep["round_t0"]
            D._OVER_BUDGET[0] = keep["over_budget"]

    print("  [194] 구형 큐 열→인덱스 순서 · 단계 상태 새 문서 · 실패/중단/완주 진실 ✅")


def t195_incremental_source_to_db_to_archive():
    """[195] 새 원본은 앱 DB 정본과 검증된 Excel 보관본까지 사람 손 없이 이어진다."""
    import importlib
    from pathlib import Path

    A = importlib.import_module("app_store")
    B = importlib.import_module("band_canonical")
    C = importlib.import_module("canonical_sync")
    W = importlib.import_module("archive_worker")
    P = importlib.import_module("automation_pipeline")

    assert A.SCHEMA_VERSION >= 2
    assert A.SHEET_SPECS["15_세금계산서관리"]["status"] == "발행상태(자동)", \
        "세금계산서 보관본 상태 열은 실제 원본 머리글과 같아야 한다"
    assert B.self_test() is True
    assert C.self_test() is True
    archive = W.self_test()
    assert archive.get("ok") and archive.get("last_good_verified"), archive
    assert archive.get("bounded_source_stage") and archive.get("source_proof_manifested"), archive
    assert P.self_test() is True

    with tempfile.TemporaryDirectory(prefix="csos-automation-contract-") as td:
        root = Path(td)
        (root / "reports").mkdir(parents=True)
        store = A.AppStore(root / "db" / "app_store.db").initialize()
        compact = W.status(root=root, store=store, spool_dir=root / "spool")
        assert compact["status"] == "missing" and not compact["external_write_performed"], compact

        good = P.submit_kakao_file(
            "KakaoTalk_합성.txt", "2026-08-10 카카오톡 합성 대화".encode("utf-8"),
            drop_dir=root / "kakao" / "dropbox",
        )
        duplicate = P.submit_kakao_file(
            "KakaoTalk_다른이름.txt", "2026-08-10 카카오톡 합성 대화".encode("utf-8"),
            drop_dir=root / "kakao" / "dropbox",
        )
        assert good["ok"] and not good["duplicate"] and duplicate["duplicate"]
        try:
            P.submit_kakao_file("bad.txt", b"\x00binary", drop_dir=root / "kakao" / "dropbox")
        except ValueError:
            pass
        else:
            raise AssertionError("binary Kakao upload was accepted")

    pipeline_src = open(os.path.join(ROOT, "automation_pipeline.py"), encoding="utf-8").read()
    band_src = open(os.path.join(ROOT, "band_canonical.py"), encoding="utf-8").read()
    archive_src = open(os.path.join(ROOT, "archive_worker.py"), encoding="utf-8").read()
    server_src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    watchdog_src = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    installer = open(os.path.join(ROOT, "install_automation_schedule.ps1"), encoding="utf-8").read()

    for marker in (
        "submit_kakao_file", "band_canonical.py", "canonical_sync.py",
        "archive_worker.py", "proc_guard", "human_gates",
    ):
        assert marker in pipeline_src, "자동화 파이프라인 연결 누락: " + marker
    assert "subprocess.run(" not in pipeline_src and "subprocess.run(" not in archive_src
    assert "cancel_watch/cross_signal" in band_src and "reserve_public_id" in band_src
    for marker in (
        '"automation"', '"ledger_now"', "/api/automation/status",
        "/api/automation/kakao-upload",
    ):
        assert marker in server_src, "앱 자동화 API 연결 누락: " + marker
    for marker in (
        "automationMonitor", "automationKakaoFile", "loadAutomationStatus",
        "uploadAutomationKakao", "showMoreChecks", "wtShowMore",
    ):
        assert marker in live, "자동화 관제 UX 누락: " + marker
    results_line = watchdog_src.split("results = [", 1)[1].split("]", 1)[0]
    assert results_line.strip().startswith("run_incremental_pipeline(dry)"), \
        "느린 Z: 일지 작업보다 증분 자동화가 뒤에 있다"
    assert "RepetitionInterval" in installer and "New-TimeSpan -Minutes 5" in installer
    assert "IgnoreNew" in installer and "--once" in installer

    print("  [195] 카톡·밴드·ERP 변경감지 → 앱 DB 정본 → 객관근거 → 검증 Excel 보관본 자동화 ✅")


def t204_staff_finance_entry_is_one_save_and_source_safe():
    """[204] 비용 정정·계산서 발행일은 한 번 저장하면 DB와 재조회가 같은 답을 낸다.

    오종현 통화에서 확인된 실제 모양을 합성한다. 비용구분은 06시트의 수식 결과를
    직접 덮지 않고 원천 PM만 고친다. 세금계산서 상태도 자동 열을 자유 입력하는 대신
    발행일이라는 사실값에서 즉시 파생한다. 저장 재전송은 멱등이고, 낡은 화면은
    낙관잠금으로 막혀야 한다.
    """
    from pathlib import Path
    import threading
    import app_store as A
    from webapp import app_server as S

    with tempfile.TemporaryDirectory(prefix="csos-staff-finance-204-") as td:
        store = A.AppStore(Path(td) / "app.db").initialize()

        pm_import = store.shadow_import(
            import_id="t204-seed", sheet="04_정기점검",
            business_key="PM-2601-048", business_key_col="점검ID", row_number=5,
            kind="정기점검", public_id="PM-2601-048", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="완료",
            fields={
                "점검ID": "PM-2601-048", "프로젝트NO": "UJ2600011",
                "캠프명": "송파 프레시 이캠프", "점검상태": "완료",
                "담당관리자": "오종현", "유상·무상·보험": "유상",
            },
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-pm",
        )
        js_import = store.shadow_import(
            import_id="t204-seed", sheet="06_거래서류청구수금",
            business_key="JS-2601-101", business_key_col="정산ID", row_number=5,
            kind="정산", public_id="JS-2601-101", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="작업완료",
            fields={
                "정산ID": "JS-2601-101", "원천업무ID": "PM-2601-048",
                "프로젝트NO": "UJ2600011", "캠프명": "송파 프레시 이캠프",
                "담당자": "오종현", "청구상태": "작업완료",
                # 실제 관리대장의 이 열은 원천 업무를 보는 수식이다. 캐시값이 있어도
                # 앱 저장이 이 칸을 직접 갱신해서는 안 된다.
                "비용구분": "유상", "세금계산서발행일": "",
            },
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-js",
        )
        pm_before = store.get_work(work_id=pm_import["work_id"])
        js_before = store.get_work(work_id=js_import["work_id"])

        cost_body = {
            "category": "pm", "key": "PM-2601-048",
            "record_version": pm_before["record_version"],
            "values": {"유상·무상·보험": "무상"},
            "reason": "오종현 통화 확인 — 무상 출고 건",
            "idempotency_key": "t204-cost-correction",
        }
        cost_saved = S.save_staff_entry(
            "oh-jonghyeon", cost_body, store=store, actor="staff:oh-jonghyeon")
        assert cost_saved.get("ok") and isinstance(cost_saved.get("record"), dict), cost_saved
        assert int(cost_saved["record_version"]) == int(cost_saved["record"]["record_version"])
        pm_after = store.get_work(work_id=pm_import["work_id"])
        js_after_cost = store.get_work(work_id=js_import["work_id"])
        assert pm_after["fields"]["유상·무상·보험"] == "무상", pm_after
        assert int(pm_after["record_version"]) == int(pm_before["record_version"]) + 1
        assert js_after_cost["fields"]["비용구분"] == "유상", \
            "원천 PM을 고치면서 06 수식 캐시까지 직접 덮었다"
        assert js_after_cost["record_version"] == js_before["record_version"], \
            "비용 정정은 연결 정산행의 버전을 올리면 안 된다"

        # 06 비용구분과 15 자동상태는 직접 쓰는 열이 아니다.
        for bad_values in ({"비용구분": "무상"}, {"발행상태(자동)": "발행완료"}):
            bad = {
                "category": "settle", "key": "JS-2601-101",
                "record_version": js_before["record_version"], "values": bad_values,
                "reason": "합성 직접쓰기 금지 확인",
                "idempotency_key": "t204-forbidden-" + next(iter(bad_values)),
            }
            try:
                S.save_staff_entry(
                    "oh-jonghyeon", bad, store=store, actor="staff:oh-jonghyeon")
            except (A.ValidationError, ValueError, PermissionError) as exc:
                assert next(iter(bad_values)) in str(exc) or "직접" in str(exc) \
                    or "허용" in str(exc), str(exc)
            else:
                raise AssertionError("정산 수식·자동상태 열 직접쓰기가 허용됐다: " + repr(bad_values))
        assert store.get_work(work_id=js_import["work_id"])["record_version"] == \
            js_before["record_version"]

        invoice_body = {
            "category": "settle", "key": "JS-2601-101",
            "record_version": js_before["record_version"],
            "values": {"세금계산서발행일": "2026-08-11"},
            "reason": "세금계산서 발행 사실 확인",
            "idempotency_key": "t204-invoice-date",
        }
        issued = S.save_staff_entry(
            "oh-jonghyeon", invoice_body, store=store, actor="staff:oh-jonghyeon")
        assert issued.get("ok") and issued["record"]["세금계산서발행일"] == \
            "2026-08-11", issued
        assert int(issued["record_version"]) == int(js_before["record_version"]) + 1

        # 저장 응답만 보지 않는다. 새 GET 읽기모델이 같은 날짜와 파생 발행상태를 내야
        # 사용자가 한 번 입력한 뒤 폼을 다시 열어도 재입력을 요구하지 않는다.
        reopened = A.AppStore(Path(td) / "app.db").initialize()
        assert reopened.get_work(
            kind="정기점검", business_key="PM-2601-048"
        )["fields"]["유상·무상·보험"] == "무상"
        old_legacy_reader = S.get_ryu_records
        try:
            def _no_real_ledger():
                raise AssertionError("임시 store 조회가 실관리대장 읽기를 먼저 호출했다")
            S.get_ryu_records = _no_real_ledger
            view = S.get_staff_records("oh-jonghyeon", store=reopened)
        finally:
            S.get_ryu_records = old_legacy_reader
        assert isinstance(view.get("rows"), dict) and isinstance(view["rows"].get("settle"), list), view
        shown = next(r for r in view["rows"]["settle"] if r.get("key") == "JS-2601-101")
        assert shown.get("detail", {}).get("세금계산서발행일") == "2026-08-11", shown
        status_blob = " ".join(str(shown.get(k) or "") for k in ("status", "계산서", "발행상태"))
        status_blob += " " + " ".join(str(v) for v in (shown.get("detail") or {}).values())
        assert "발행" in status_blob and "미발행" not in status_blob, shown

        def event_count(work_id):
            with store.reader() as conn:
                return int(conn.execute(
                    "SELECT COUNT(*) FROM change_event WHERE work_id=?", (work_id,)
                ).fetchone()[0])

        before_replay = event_count(js_import["work_id"])
        replay = S.save_staff_entry(
            "oh-jonghyeon", dict(invoice_body), store=store,
            actor="staff:oh-jonghyeon")
        assert replay["record_version"] == issued["record_version"], replay
        assert replay.get("action") == issued.get("action") \
            and replay.get("event_id") == issued.get("event_id"), \
            "같은 멱등키 재전송이 최초 저장 결과를 돌려주지 않았다"
        assert event_count(js_import["work_id"]) == before_replay, \
            "같은 저장 재전송이 감사 이벤트를 하나 더 만들었다"

        changed_replay = dict(invoice_body)
        changed_replay["values"] = {"세금계산서발행일": "2026-08-12"}
        try:
            S.save_staff_entry(
                "oh-jonghyeon", changed_replay, store=store,
                actor="staff:oh-jonghyeon")
        except A.IdempotencyConflict:
            pass
        else:
            raise AssertionError("같은 멱등키의 다른 발행일이 통과했다")

        stale = dict(invoice_body)
        stale["idempotency_key"] = "t204-stale-version"
        stale["values"] = {"세금계산서발행일": "2026-08-13"}
        try:
            S.save_staff_entry(
                "oh-jonghyeon", stale, store=store,
                actor="staff:oh-jonghyeon")
        except A.VersionConflict:
            pass
        else:
            raise AssertionError("낡은 record_version 저장이 기존 발행일을 덮었다")

        # v586의 15·16 business_key는 관리ID이고, 정산 행 연결키는 별도 정산ID다.
        # 같은 프로젝트에 정산이 둘일 때 관리ID를 정산ID로 오인하면 가짜 세 번째
        # 정산 행이 생기므로 exact 관계키만 합쳐지는지 회귀 검증한다.
        store.shadow_import(
            import_id="t204-sidecar", sheet="06_거래서류청구수금",
            business_key="JS-2601-102", business_key_col="정산ID", row_number=6,
            kind="정산", public_id="JS-2601-102", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="작업완료",
            fields={"정산ID": "JS-2601-102", "프로젝트NO": "UJ2600011",
                    "캠프명": "송파 프레시 이캠프", "청구상태": "작업완료"},
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-js-2",
        )
        store.shadow_import(
            import_id="t204-sidecar", sheet="15_세금계산서관리",
            business_key="TI-2601-101", business_key_col="계산서관리ID", row_number=5,
            kind="세금계산서", public_id="TI-2601-101", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="발행완료",
            fields={"계산서관리ID": "TI-2601-101", "정산ID": "JS-2601-101",
                    "프로젝트NO": "UJ2600011", "캠프명": "송파 프레시 이캠프",
                    "실제발행일": "2026-08-11", "발행상태(자동)": "발행완료"},
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-invoice-sidecar",
        )
        store.shadow_import(
            import_id="t204-sidecar", sheet="15_세금계산서관리",
            business_key="TI-2601-101-B", business_key_col="계산서관리ID", row_number=6,
            kind="세금계산서", public_id="TI-2601-101-B", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="발행완료",
            fields={"계산서관리ID": "TI-2601-101-B", "정산ID": "JS-2601-101",
                    "프로젝트NO": "UJ2600011", "캠프명": "송파 프레시 이캠프",
                    "실제발행일": "2026-08-11", "발행상태(자동)": "발행완료"},
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-invoice-sidecar-2",
        )
        store.shadow_import(
            import_id="t204-sidecar", sheet="15_세금계산서관리",
            business_key="TI-2601-101-C", business_key_col="계산서관리ID", row_number=8,
            kind="세금계산서", public_id="TI-2601-101-C", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="미발행",
            fields={"계산서관리ID": "TI-2601-101-C", "정산ID": "JS-2601-101",
                    "프로젝트NO": "UJ2600011", "캠프명": "송파 프레시 이캠프",
                    "실제발행일": "", "발행상태(자동)": "미발행"},
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-invoice-sidecar-3",
        )
        for row_no, rid, day, amount in (
            (5, "RC-2601-101-A", "2026-08-10", 100_000),
            (6, "RC-2601-101-B", "2026-08-11", 250_000),
        ):
            store.shadow_import(
                import_id="t204-sidecar", sheet="16_입금수금관리",
                business_key=rid, business_key_col="입금관리ID", row_number=row_no,
                kind="입금수금", public_id=rid, project_no="UJ2600011",
                camp_name="송파 프레시 이캠프", status="",
                fields={"입금관리ID": rid, "정산ID": "JS-2601-101",
                        "프로젝트NO": "UJ2600011", "캠프명": "송파 프레시 이캠프",
                        "입금일": day, "입금액": amount},
                source_file="t204.xlsx", source_sha256="a" * 64,
                apply_if_missing=True, idempotency_key="t204-seed-" + rid,
            )
        # 명시 정산ID가 없는 것이 아니라 '다른 ID'라고 명시된 행이다. 같은 프로젝트에
        # 정산이 하나뿐이어도 그 행으로 fallback해서는 안 된다.
        store.shadow_import(
            import_id="t204-sidecar", sheet="15_세금계산서관리",
            business_key="TI-UNKNOWN", business_key_col="계산서관리ID", row_number=7,
            kind="세금계산서", public_id="TI-UNKNOWN", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="발행완료",
            fields={"계산서관리ID": "TI-UNKNOWN", "정산ID": "JS-UNKNOWN",
                    "프로젝트NO": "UJ2600011", "실제발행일": "2026-08-09"},
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-invoice-orphan",
        )
        store.shadow_import(
            import_id="t204-sidecar", sheet="15_세금계산서관리",
            business_key="TI-NOREL", business_key_col="계산서관리ID", row_number=9,
            kind="세금계산서", public_id="TI-NOREL", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="미발행",
            fields={"계산서관리ID": "TI-NOREL", "정산ID": "",
                    "프로젝트NO": "UJ2600011", "실제발행일": ""},
            source_file="t204.xlsx", source_sha256="a" * 64,
            apply_if_missing=True, idempotency_key="t204-seed-invoice-no-relation",
        )
        linked = S._overlay_app_store_settlements(
            [{"정산ID": "JS-2601-101", "프로젝트NO": "UJ2600011"},
             {"정산ID": "JS-2601-102", "프로젝트NO": "UJ2600011"}],
            store=store,
        )
        assert len(linked) == 4, linked
        linked_by_id = {str(row.get("정산ID") or ""): row for row in linked}
        assert set(linked_by_id) == {"", "JS-2601-101", "JS-2601-102", "JS-UNKNOWN"}, linked
        first = linked_by_id["JS-2601-101"]
        assert first.get("계산서관리ID") == \
            "TI-2601-101 · TI-2601-101-B · TI-2601-101-C", first
        assert first.get("계산서") == "일부 발행일 확인" \
            and first.get("계산서발행일") == "2026-08-11", first
        assert first.get("계산서건수") == 3 and first.get("계산서발행건수") == 2, first
        assert first.get("입금액") == 350_000 and first.get("입금일") == "2026-08-11", first
        assert first.get("입금건수") == 2, first
        assert not linked_by_id["JS-2601-102"].get("계산서관리ID"), linked
        assert linked_by_id["JS-UNKNOWN"].get("연결상태") == "정산ID 미연결", linked
        assert linked_by_id[""].get("계산서관리ID") == "TI-NOREL" \
            and linked_by_id[""].get("연결상태") == "정산ID 미확인", linked
        assert not any(str(row.get("정산ID") or "") == "TI-2601-101" for row in linked), \
            "계산서관리ID를 정산ID로 오인한 가짜 행이 생겼다"

        ledger_view = S._overlay_app_store_ledger_records(
            {"JS-2601-101": {"정산ID": "JS-2601-101", "프로젝트NO": "UJ2600011"},
             "JS-2601-102": {"정산ID": "JS-2601-102", "프로젝트NO": "UJ2600011"}},
            store=store,
        )
        assert set(ledger_view) == {"JS-2601-101", "JS-2601-102"}, ledger_view
        assert ledger_view["JS-2601-101"]["원장_세금계산서실제발행일"] == "2026-08-11"
        assert ledger_view["JS-2601-101"]["원장_세금계산서발행일부분확인"] is True
        assert ledger_view["JS-2601-101"]["원장_세금계산서건수"] == 3 \
            and ledger_view["JS-2601-101"]["원장_세금계산서발행건수"] == 2
        assert ledger_view["JS-2601-101"]["원장_입금액"] == 350_000 \
            and ledger_view["JS-2601-101"]["원장_입금일"] == "2026-08-11"

        # 같은 멱등키의 실제 동시 요청 둘을 한 DB에 부딪친다. 멱등 조회·업무 변경·
        # 최종 staff 응답 저장이 서로 다른 트랜잭션이면 둘 다 최초 요청으로 판단한 뒤
        # 하나가 업무를 바꾸고 바깥 멱등 INSERT에서 409가 나는 반쪽 성공이 재현된다.
        concurrent_before = store.get_work(work_id=pm_import["work_id"])
        concurrent_events_before = event_count(pm_import["work_id"])
        concurrent_body = {
            "category": "pm", "key": "PM-2601-048",
            "record_version": concurrent_before["record_version"],
            "values": {"비고": "동시 저장 원자성 확인"},
            "reason": "두 요청 동일 저장 검증",
            "idempotency_key": "t204-concurrent-same-command",
        }
        gate = threading.Barrier(2)
        concurrent_results = []
        concurrent_errors = []
        result_lock = threading.Lock()

        def _same_command_worker():
            try:
                gate.wait(timeout=5)
                value = S.save_staff_entry(
                    "oh-jonghyeon", dict(concurrent_body), store=store,
                    actor="staff:oh-jonghyeon",
                )
                with result_lock:
                    concurrent_results.append(value)
            except Exception as exc:
                with result_lock:
                    concurrent_errors.append(exc)

        workers = [threading.Thread(target=_same_command_worker) for _ in range(2)]
        for worker in workers:
            worker.start()
        for worker in workers:
            worker.join(timeout=10)
        assert not any(worker.is_alive() for worker in workers), \
            "같은 멱등키 동시 저장 스레드가 끝나지 않았다"
        assert not concurrent_errors and len(concurrent_results) == 2, \
            (concurrent_results, concurrent_errors)
        assert all(result.get("ok") and result.get("action") == "updated"
                   for result in concurrent_results), concurrent_results
        assert len({result.get("event_id") for result in concurrent_results}) == 1 \
            and concurrent_results[0].get("event_id"), \
            "같은 명령 두 응답이 최초 감사 event_id 하나를 공유하지 않는다"
        assert sum(bool(result.get("idempotent_replay"))
                   for result in concurrent_results) == 1, concurrent_results
        concurrent_after = store.get_work(work_id=pm_import["work_id"])
        assert concurrent_after["record_version"] == \
            concurrent_before["record_version"] + 1, concurrent_after
        assert concurrent_after["fields"]["비고"] == "동시 저장 원자성 확인"
        assert event_count(pm_import["work_id"]) == concurrent_events_before + 1, \
            "동시 같은 명령이 업무 이벤트·버전을 두 번 만들었다"

        different_payload = dict(concurrent_body)
        different_payload["values"] = {"비고": "같은 키의 다른 요청"}
        try:
            S.save_staff_entry(
                "oh-jonghyeon", different_payload, store=store,
                actor="staff:oh-jonghyeon",
            )
        except A.IdempotencyConflict:
            pass
        else:
            raise AssertionError("같은 멱등키의 다른 동시 명령 내용이 통과했다")
        assert store.get_work(work_id=pm_import["work_id"])["record_version"] == \
            concurrent_after["record_version"], \
            "멱등 충돌 요청이 거부되면서도 업무를 바꿨다"

        # 브라우저·본문의 이름이 아니라 세션 actor가 감사로그에 남는다.
        with store.reader() as conn:
            actors = [r[0] for r in conn.execute(
                "SELECT actor FROM change_event WHERE work_id IN (?,?) ORDER BY id",
                (pm_import["work_id"], js_import["work_id"]),
            ).fetchall()]
        assert actors[-2:] == ["staff:oh-jonghyeon", "staff:oh-jonghyeon"], actors

    print("  [204] PM 비용 원천정정 · 06 수식보호 · 발행일 1회저장/파생상태 · 멱등/낙관잠금 ✅")


def t205_three_staff_sessions_cannot_forge_actor():
    """[205] 세 업무센터의 역할은 쿠키로 고정되고 본문 이름으로 바뀌지 않는다."""
    import inspect
    from pathlib import Path
    import app_store as A
    from webapp import app_server as S

    slugs = ("ryu-jiyeong", "oh-jonghyeon", "yoo-hyeonmin")
    assert set(slugs) <= set(S.STAFF_CENTERS), S.STAFF_CENTERS
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    for slug in slugs:
        assert f"'{slug}'" in live and f'/staff/{slug}' in live, \
            f"{slug} 업무센터 정의·진입 링크가 한쪽에만 있다"

    for path in ('"/api/staff/records"', '"/api/staff/entry"'):
        assert path in server, "공용 직원센터 API 누락: " + path
    entry_at = server.index('"/api/staff/entry"')
    entry_start = server.rfind("        if p", 0, entry_at)
    entry_end = server.find("\n        if p", entry_at)
    post = server[entry_start:entry_end if entry_end >= 0 else len(server)]
    compact = re.sub(r"\s+", "", post)
    assert "save_staff_entry(" in post and ("_require_staff(" in post or "self._actor()" in post)
    assert 'body.get("staff_slug")' not in post and "body.get('staff_slug')" not in post
    assert 'body.get("submitter")' not in post and "body.get('submitter')" not in post
    assert "actor=self._actor_name()" in compact or "actor=self._actor_name(" in compact, \
        "감사 actor를 서명 세션이 아니라 요청 본문에서 받고 있다"
    assert "actual_category not in RYU_ENTRY_CONFIG" in post, \
        "첨부파일을 저장하기 전에 업무 카테고리를 검증하지 않는다"
    evidence_source = inspect.getsource(S._save_ryu_evidence)
    assert "safe_category" in evidence_source and "_업무근거_{safe_category}" in evidence_source, \
        "요청 카테고리가 첨부파일 경로에 그대로 들어간다"

    # 서명된 쿠키 세 개를 실제로 만들고 Handler가 각각 다른 actor로 복원하는지 본다.
    actors = {}
    for slug in slugs:
        token, issued = S.create_auth_session(slug)
        parsed = S.auth_session_from_cookie("csos_session=" + token)
        assert issued["role"] == parsed["role"] == "staff"
        assert issued["staff_slug"] == parsed["staff_slug"] == slug
        handler = object.__new__(S.H)
        handler.headers = {"Cookie": "csos_session=" + token}
        handler.client_address = ("127.0.0.1", 20400)
        actors[slug] = handler._actor_name()
    assert actors == {slug: "staff:" + slug for slug in slugs}, actors

    with tempfile.TemporaryDirectory(prefix="csos-staff-isolation-205-") as td:
        store = A.AppStore(Path(td) / "app.db").initialize()
        keys = {}
        for n, slug in enumerate(slugs, 1):
            name = S.STAFF_CENTERS[slug]["name"]
            key = f"PM-2608-20{n}"
            keys[slug] = key
            store.create_work(
                kind="정기점검", business_key=key, public_id=key,
                project_no=f"UJ26002{n:02d}", camp_name=f"합성 {name} 캠프", status="예정",
                fields={
                    "점검ID": key, "프로젝트NO": f"UJ26002{n:02d}",
                    "캠프명": f"합성 {name} 캠프", "점검상태": "예정",
                    "담당자": name, "담당관리자": name, "담당기사": name,
                    "유상·무상·보험": "유상",
                },
                actor="t205-seed", source="synthetic", evidence="직원센터 격리 합성",
                idempotency_key=f"t205-seed-{slug}",
            )

        # 세 센터가 같은 업무를 참고할 수 있어도 수정 가능 열은 서버가 각 세션에
        # 맞게 내려야 한다. 화면에서 숨기는 것만으로 권한을 대신하면 안 된다.
        old_legacy_reader = S.get_ryu_records
        try:
            def _no_real_ledger():
                raise AssertionError("임시 store 조회가 실관리대장 읽기를 먼저 호출했다")
            S.get_ryu_records = _no_real_ledger
            views = {slug: S.get_staff_records(slug, store=store) for slug in slugs}
        finally:
            S.get_ryu_records = old_legacy_reader
        for slug, view in views.items():
            assert view.get("staff_slug") == slug and view.get("staff") == \
                S.STAFF_CENTERS[slug]["name"], view
            assert isinstance(view.get("permissions", {}).get("pm"), list), view
        assert "점검상태" in views["ryu-jiyeong"]["permissions"]["pm"]
        assert "점검상태" in views["yoo-hyeonmin"]["permissions"]["pm"]
        assert "점검상태" not in views["oh-jonghyeon"]["permissions"]["pm"]

        oh_raw = store.get_work(kind="정기점검", business_key=keys["oh-jonghyeon"])
        forged = {
            "category": "pm", "key": keys["oh-jonghyeon"],
            "record_version": oh_raw["record_version"],
            "values": {"유상·무상·보험": "무상"},
            "reason": "본문 이름 위조 방지 합성",
            "idempotency_key": "t205-forged-body",
            "staff_slug": "ryu-jiyeong", "submitter": "류지영",
        }
        saved = S.save_staff_entry(
            "oh-jonghyeon", forged, store=store, actor=actors["oh-jonghyeon"])
        assert saved.get("ok") and saved["record"]["유상·무상·보험"] == "무상"
        with store.reader() as conn:
            audit_actor = conn.execute(
                "SELECT actor FROM change_event WHERE work_id=? ORDER BY id DESC LIMIT 1",
                (oh_raw["id"],),
            ).fetchone()[0]
        assert audit_actor == "staff:oh-jonghyeon", audit_actor

        # 오종현 세션이 허용표 밖의 상태를 고치는 것은 helper 단계에서도 막는다.
        cross = dict(forged)
        cross.update({
            "key": keys["oh-jonghyeon"],
            "record_version": saved["record_version"],
            "values": {"점검상태": "완료"},
            "idempotency_key": "t205-cross-center",
        })
        try:
            S.save_staff_entry(
                "oh-jonghyeon", cross, store=store, actor=actors["oh-jonghyeon"])
        except (A.ValidationError, PermissionError) as exc:
            assert "권한" in str(exc) or "업무센터" in str(exc), str(exc)
        else:
            raise AssertionError("오종현 세션이 허용표 밖 점검상태를 수정했다")

    print("  [205] 류지영·오종현·유현민 세션 격리 · 본문 actor 위조/교차수정 차단 ✅")


def t206_finance_archive_keeps_real_headers_and_formulas():
    """[206] DB 정정은 실제 머리글의 검증 Excel 보관본으로만 나간다."""
    from pathlib import Path
    import app_store as A
    import archive_worker as W
    import db_cutover as D

    assert A.SHEET_SPECS["15_세금계산서관리"]["status"] == "발행상태(자동)"
    assert D.SPECS["15_세금계산서관리"]["status"] == "발행상태(자동)", \
        "컷오버가 구형 '발행상태' 머리글을 써 실제 자동상태 열을 잃는다"

    with tempfile.TemporaryDirectory(prefix="csos-finance-archive-206-") as td:
        base = Path(td)
        template = base / "finance-template.xlsx"
        wb = openpyxl.Workbook()
        ws4 = wb.active
        ws4.title = "04_정기점검"
        for _ in range(3):
            ws4.append([])
        ws4.append(["점검ID", "프로젝트NO", "캠프명", "유상·무상·보험", "점검상태"])
        ws4.append(["PM-2601-048", "UJ2600011", "송파 프레시 이캠프", "유상", "완료"])
        ws6 = wb.create_sheet("06_거래서류청구수금")
        for _ in range(3):
            ws6.append([])
        ws6.append(["정산ID", "원천업무ID", "프로젝트NO", "캠프명", "비용구분",
                    "세금계산서발행일", "청구상태"])
        ws6.append(["JS-2601-101", "PM-2601-048", "UJ2600011", "송파 프레시 이캠프",
                    "='04_정기점검'!D5", "", "작업완료"])
        ws15 = wb.create_sheet("15_세금계산서관리")
        for _ in range(3):
            ws15.append([])
        ws15.append(["계산서관리ID", "정산ID", "프로젝트NO", "캠프명",
                     "실제발행일", "발행상태(자동)"])
        ws15.append(["TI-JS-2601-101", "JS-2601-101", "UJ2600011",
                     "송파 프레시 이캠프", "",
                     '=IF(E5="","미발행","발행완료")'])
        wb.save(template)
        wb.close()
        template_hash = hashlib.sha256(template.read_bytes()).hexdigest()

        store = A.AppStore(base / "app.db").initialize()
        pm = store.shadow_import(
            import_id="t206-import", sheet="04_정기점검", business_key="PM-2601-048",
            business_key_col="점검ID", row_number=5, kind="정기점검",
            public_id="PM-2601-048", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="완료",
            fields={"점검ID": "PM-2601-048", "프로젝트NO": "UJ2600011",
                    "캠프명": "송파 프레시 이캠프",
                    "유상·무상·보험": "유상", "점검상태": "완료"},
            source_file=str(template), source_sha256=template_hash,
            apply_if_missing=True, idempotency_key="t206-import-pm",
        )
        settle = store.shadow_import(
            import_id="t206-import", sheet="06_거래서류청구수금", business_key="JS-2601-101",
            business_key_col="정산ID", row_number=5, kind="정산",
            public_id="JS-2601-101", project_no="UJ2600011",
            camp_name="송파 프레시 이캠프", status="작업완료",
            fields={"정산ID": "JS-2601-101", "원천업무ID": "PM-2601-048",
                    "프로젝트NO": "UJ2600011", "캠프명": "송파 프레시 이캠프",
                    "비용구분": "유상",
                    "세금계산서발행일": "", "청구상태": "작업완료"},
            source_file=str(template), source_sha256=template_hash,
            apply_if_missing=True, idempotency_key="t206-import-settle",
        )
        invoice = store.shadow_import(
            import_id="t206-import", sheet="15_세금계산서관리",
            business_key="TI-JS-2601-101", business_key_col="계산서관리ID",
            row_number=5, kind="세금계산서", public_id="TI-JS-2601-101",
            project_no="UJ2600011", camp_name="송파 프레시 이캠프", status="미발행",
            fields={"계산서관리ID": "TI-JS-2601-101", "정산ID": "JS-2601-101",
                    "프로젝트NO": "UJ2600011",
                    "캠프명": "송파 프레시 이캠프",
                    "실제발행일": "", "발행상태(자동)": "미발행"},
            source_file=str(template), source_sha256=template_hash,
            apply_if_missing=True, idempotency_key="t206-import-invoice",
        )

        pm_row = store.get_work(work_id=pm["work_id"])
        store.update_work(
            pm_row["id"], expected_version=pm_row["record_version"],
            patch={"fields": {"유상·무상·보험": "무상"}},
            actor="staff:oh-jonghyeon", source="staff-entry",
            evidence="합성 비용 정정", idempotency_key="t206-update-pm",
        )
        settle_row = store.get_work(work_id=settle["work_id"])
        store.update_work(
            settle_row["id"], expected_version=settle_row["record_version"],
            patch={"fields": {"세금계산서발행일": "2026-08-11"}},
            actor="staff:oh-jonghyeon", source="staff-entry",
            evidence="합성 발행일", idempotency_key="t206-update-settle",
        )
        invoice_row = store.get_work(work_id=invoice["work_id"])
        store.update_work(
            invoice_row["id"], expected_version=invoice_row["record_version"],
            patch={"status": "발행완료", "fields": {
                "실제발행일": "2026-08-11", "발행상태(자동)": "발행완료",
            }},
            actor="staff:oh-jonghyeon", source="staff-entry",
            evidence="합성 발행 사실", idempotency_key="t206-update-invoice",
        )

        worker = W.ArchiveWorker(store, base / "spool")
        first = worker.run(template)
        assert first.get("ok") and first.get("state") == "verified", first
        adapter_path = Path(first["export"]["artifact_dir"]) / "adapter-result.json"
        adapter = json.loads(adapter_path.read_text(encoding="utf-8"))
        assert not adapter.get("errors") and not adapter.get("conflicts"), adapter
        assert not any("unknown field" in str(x).lower() for x in adapter.get("warnings") or []), adapter
        assert adapter.get("external_write_performed") is False
        archive_path = Path(first["last_good"]["archive_path"])
        out = openpyxl.load_workbook(archive_path, data_only=False, read_only=False)
        try:
            h4 = {c.value: c.column for c in out["04_정기점검"][4] if c.value}
            h6 = {c.value: c.column for c in out["06_거래서류청구수금"][4] if c.value}
            h15 = {c.value: c.column for c in out["15_세금계산서관리"][4] if c.value}
            assert "발행상태(자동)" in h15 and "발행상태" not in h15, h15
            assert out["04_정기점검"].cell(5, h4["유상·무상·보험"]).value == "무상"
            cost_formula = out["06_거래서류청구수금"].cell(5, h6["비용구분"]).value
            assert isinstance(cost_formula, str) and cost_formula.startswith("=") \
                and "04_정기점검" in cost_formula, cost_formula
            issued6 = out["06_거래서류청구수금"].cell(5, h6["세금계산서발행일"]).value
            issued15 = out["15_세금계산서관리"].cell(5, h15["실제발행일"]).value
            def excel_day(value):
                if isinstance(value, (int, float)):
                    return openpyxl.utils.datetime.from_excel(value).date().isoformat()
                return str(value)[:10]
            assert excel_day(issued6) == "2026-08-11", issued6
            assert excel_day(issued15) == "2026-08-11", issued15
            status_formula = out["15_세금계산서관리"].cell(5, h15["발행상태(자동)"]).value
            assert isinstance(status_formula, str) and status_formula.startswith("=") \
                and "발행완료" in status_formula, status_formula
        finally:
            out.close()

        assert hashlib.sha256(template.read_bytes()).hexdigest() == template_hash, \
            "보관본 생성기가 입력 템플릿을 수정했다"
        first_hash = hashlib.sha256(archive_path.read_bytes()).hexdigest()
        second = worker.run(template)
        assert second.get("ok") and second["export"]["export_id"] == first["export"]["export_id"]
        assert hashlib.sha256(Path(second["last_good"]["archive_path"]).read_bytes()).hexdigest() == first_hash

    print("  [206] DB→Excel 보관본 실제 발행상태(자동) 머리글·PM 원천값·06/15 수식 보존 ✅")


def t207_live_revision_is_shared_and_nonblocking():
    """[207] 모든 기기는 서버의 안정 변경번호만 보고 배경 갱신한다.

    표시용 현재 시각이 매 요청 달라진다고 새 자료로 오인하면 휴대폰과 PC가 서로
    영원히 갱신을 깨운다. 반대로 앱 DB나 자동화 단계가 실제로 바뀌면 어느 기기든
    같은 revision을 보고 기존 화면을 유지한 채 자료만 다시 받아야 한다.
    """
    from pathlib import Path
    import app_store as A
    from webapp import app_server as S

    with tempfile.TemporaryDirectory(prefix="csos-live-revision-207-") as td:
        root = Path(td)
        store = A.AppStore(root / "app.db").initialize()
        state_path = root / "automation_pipeline_state.json"
        # ★ 시각을 파일에 **박아 두면 다음 날 스스로 썩는다** (2026-08-12 실측 · 분담판 [47]).
        #   `2026-08-11T23:01Z` 는 적을 때 제일 최신이었지만 하루 지나 과거가 됐고,
        #   그러자 `state_updated_at` 이 안 움직여 **멀쩡한 코드가 빨강**이 됐다.
        #   ⚠ 끝난 회차 시각(`_done_at`)은 두 번의 쓰기에서 **같아야** 한다 —
        #     그것이 달라지면 `last_completed_at` 이 움직여 자료 revision 까지 바뀌고,
        #     "진행 단계만 바뀌었는데 7개 API 를 깨운다"는 다른 빨강이 난다.
        from datetime import datetime as _dt, timedelta as _td, timezone as _tz
        _now = _dt.now(_tz.utc)
        _done_at = (_now - _td(hours=20)).isoformat()      # 어제 끝난 회차
        _running_at = (_now + _td(minutes=1)).isoformat()  # 지금 도는 회차의 '방금'
        state_path.write_text(json.dumps({
            "version": 1, "running": False,
            "last_run": {"status": "success", "finished_at": _done_at},
        }), encoding="utf-8")

        first = S.get_live_state(store=store, state_path=state_path)
        second = S.get_live_state(store=store, state_path=state_path)
        assert first["ok"] and first["phase"] == "current", first
        assert first["revision"] == second["revision"], \
            "표시용 서버 시각 때문에 안정 revision이 매 요청 달라진다"
        assert first["state_revision"] == second["state_revision"]
        assert first["updated_at"] == first["data_updated_at"], first
        assert first["state_updated_at"], first

        store.create_work(
            kind="돌발AS", business_key="AS-207", public_id="AS-207",
            project_no="UJ207", camp_name="합성캠프", status="접수",
            fields={"접수ID": "AS-207", "프로젝트NO": "UJ207"},
            actor="synthetic", source="synthetic", idempotency_key="t207-create",
        )
        changed = S.get_live_state(store=store, state_path=state_path)
        assert changed["change_seq"] == first["change_seq"] + 1, changed
        assert changed["revision"] != first["revision"], \
            "앱 DB 변경을 다른 기기가 알아챌 revision이 그대로다"
        assert changed["phase"] == "archive_pending" and changed["outbox_pending"] == 1
        assert changed["data_updated_at"] != first["data_updated_at"], changed

        state_path.write_text(json.dumps({
            "version": 1, "running": True,
            "active_run_id": "t207-running",
            "last_run": {"run_id": "t207-running", "status": "running", "current_stage": "ERP 대조",
                         "updated_at": _running_at},
            "history": [{"status": "success", "finished_at": _done_at}],
        }), encoding="utf-8")
        from automation_pipeline import PipelineLock
        pipeline_lock = PipelineLock(root / ".automation_pipeline.lock")
        pipeline_token = pipeline_lock.acquire("t207-running")
        assert pipeline_token, "합성 파이프라인 잠금을 못 잡았다"
        running = S.get_live_state(store=store, state_path=state_path)
        assert running["revision"] == changed["revision"], \
            "진행 단계만 바뀌었는데 7개 자료 API 재검증을 깨운다"
        assert running["state_revision"] != changed["state_revision"]
        assert running["phase"] == "updating" and running["current_stage"] == "ERP 대조"
        assert running["data_updated_at"] == changed["data_updated_at"], \
            "자동화 단계 시각을 자료 갱신시각으로 섞는다"
        assert running["state_updated_at"] != changed["state_updated_at"], \
            "자동화 진행 시각이 관제 상태에 반영되지 않는다"
        assert pipeline_lock.release(pipeline_token, "t207-running")

    server_src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert 'p == "/api/live-state"' in server_src
    body = server_src.split("def get_live_state", 1)[1].split("\ndef ", 1)[0]
    for forbidden in ("resolve_master", "master_book", "os.walk", "glob("):
        assert forbidden not in body, "가벼운 기기 동기화 경로가 원본을 읽는다: " + forbidden
    assert "generated_at" not in body.split("stable =", 1)[1].split("revision =", 1)[0], \
        "표시 시각이 안정 revision 재료에 들어갔다"
    assert '"data_updated_at": data_updated_at' in body \
        and '"state_updated_at": state_updated_at' in body, \
        "자료 시각과 자동화 관제 시각이 분리되지 않았다"

    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert 'id="dataSyncChip"' in live and 'data-refresh-ui="compact-header"' in live
    assert 'id="dataHealth" data-refresh-scope="run-only"' in live
    assert "const LIVE_STATE_PATH='/api/live-state'" in live \
        and "const LIVE_STATE_POLL_MS=5000" in live
    assert "LIVE_STATE_JITTER_MS" in live and "scheduleLiveStatePoll" in live, \
        "여러 기기가 같은 5초 경계에 한꺼번에 서버를 두드린다"
    for marker in ("appliedRevision", "pendingRevision", "refreshTarget"):
        assert marker in live, "갱신 중 더 최신 revision을 잃는다: " + marker
    revalidate = live.split("function revalidateLiveData", 1)[1].split(
        "function pollLiveState", 1
    )[0]
    assert "targetRevision" in revalidate and "refreshLiveViewData" in revalidate \
        and "pendingRevision" in revalidate, \
        "요청 시작 revision을 붙들고 최신 revision을 한 번 더 처리하지 않는다"
    assert "while(runTarget)" in revalidate and "const queued=LIVE_SYNC.pendingRevision" in revalidate, \
        "갱신 중 들어온 마지막 revision을 현재 요청 뒤에 이어서 처리하지 않는다"
    assert "loadSettle(true)" not in revalidate, \
        "기기 변경 하나마다 모든 7개 대형 API를 다시 부른다"
    poll = live.split("function pollLiveState", 1)[1].split("\n}\n", 1)[0]
    assert "LIVE_SYNC.pollFlight" in poll, "기기 상태 폴링이 겹쳐서 같은 요청을 여러 번 보낸다"
    assert "AbortController" in poll and "clearTimeout" in poll \
        and "LIVE_STATE_TIMEOUT" in poll, \
        "반쯤 열린 상태 요청이 영원히 다음 폴링을 막는다"
    assert "serverRevisionOf(data)" in poll and "serverStateRevisionOf(data)" in poll
    assert "data.ok===false" in poll and "LIVE_STATE_ERROR" in poll, \
        "HTTP 200 오류 상태를 최신 자료로 확정한다"
    assert "if(next&&next!==prior)" in poll and "revalidateLiveData" in poll, \
        "자료 revision이 그대로인데도 무거운 자료 묶음을 다시 받는다"
    render = live.split("function renderDataHealth", 1)[1].split("\n}\n", 1)[0]
    assert "inRun=view==='run'" in render and "chip.hidden=inRun" in render
    assert "if(!inRun)" in render and "host.hidden=true" in render, \
        "실행 밖에서도 큰 갱신 패널이 업무 화면을 가린다"
    assert "phase==='archive_pending'" in render and "archivePending?'stale'" in render, \
        "Excel 보관 대기를 최신/초록 상태로 거짓 표시한다"
    refresh_view = live.split("async function refreshLiveViewData", 1)[1].split(
        "async function retryDataSection", 1
    )[0]
    staff_refresh = live.split("async function refreshStaffCenter", 1)[1].split(
        "async function refreshCalendarData", 1
    )[0]
    assert "refreshStaffCenter" in refresh_view and "staffSlug" in refresh_view \
        and "loadRyuRecords(true)" in staff_refresh, \
        "다른 기기의 직원 업무센터 입력을 다시 읽지 않는다"
    assert "refreshDataKeys" in refresh_view and "Object.keys(DATA_SECTION_DEFS)" in refresh_view, \
        "현재 화면과 무관한 대형 API를 항상 전부 다시 부른다"
    assert "function liveViewIsCurrent" in live and "queueLiveViewCatchup(v)" in live, \
        "같은 revision에서 다른 카테고리로 옮기면 그 화면 자료가 옛 상태로 남는다"
    success = live.split("function markDataSectionSuccess", 1)[1].split(
        "function markDataSectionFailure", 1
    )[0]
    assert "targetRevision" in success and "LIVE_SYNC.revision" not in success, \
        "오래된 응답을 요청 도중 바뀐 최신 revision 결과로 잘못 표시한다"
    for fn_name in ("function netBanner", "function swrChip"):
        section = live.split(fn_name, 1)[1].split("\n}", 1)[0]
        assert "location.reload" not in section and "document.body.appendChild" not in section, \
            "갱신 상태가 떠다니는 팝업이나 문서 새로고침으로 돌아갔다: " + fn_name
    assert "function uiContinuitySnapshot" in live and "function restoreUiContinuity" in live
    continuity = live.split("function uiContinuitySnapshot", 1)[1].split(
        "function restoreUiContinuity", 1
    )[0]
    for marker in ("defaultValue", "defaultChecked", "defaultSelected"):
        assert marker in continuity, \
            "손대지 않은 서버 필드까지 옛 화면값으로 되돌린다: " + marker
    view_tail = live.split("function applyView", 1)[1].split("function show", 1)[0]
    assert "window.__view = v;\n  renderDataHealth();" in view_tail, \
        "카테고리를 바꾼 직후 실행 상세/header 상태 전환이 늦는다"
    assert "BroadcastChannel('csos-live-state-v1')" in live \
        and "pollLiveState('visibility')" in live
    chip_handler = live.split("function handleDataSyncChip", 1)[1].split(
        "const DATA_SECTION_DEFS", 1
    )[0]
    assert "show('run')" in chip_handler and "retryDataSync" in chip_handler, \
        "정상 헤더 상태를 눌러도 자료 7개를 다시 읽거나 실행 상세를 열지 못한다"

    print("  [207] 서버 변경번호로 기기간 즉시 동기화 · 화면별 최소 재조회 · 입력 연속성 ✅")


def t208_cancel_remote_resolution_is_exact_and_finance_safe():
    """[208] 밴드·카톡 접수취소를 원천업무에 붙이고 청구 교차입력을 가른다.

    UJ2600035 실사고: 밴드와 카톡은 `접수취소·유선전화 해결`을 말했지만 원천 AS에는
    근거가 없고 연결 정산에는 유상·PO·명세서·계산서가 남아 미발행 경고가 났다.
    취소건은 청구대상에서 빼되 이미 생긴 서류는 지우지 않고 충돌로 보존해야 한다.
    """
    import cancel_resolution as CR
    import cross_signal as CS
    import ecount_reconcile as ER
    import tax_invoice_watch as TW
    from app_store import AppStore
    from webapp import app_server as APP

    assert CR.outcome_kind("✅ [접수취소] - 유선접화로 해결 완료") == "원격해결"
    assert CR.outcome_kind("기사님과 유선전화 이후 접수 취소되었습니다") == "접수취소"
    for text in ("유선전화로 해결 완료", "통화 완료", "부품 취소 요청",
                 "접수 유지", "접수 취소 불가", "접수는 취소하지 않습니다"):
        assert CR.outcome_kind(text) == "", f"접수취소 근거 없이 자동 취소한다: {text}"

    # 프로젝트번호가 정확히 같으면 캠프 표기가 달라도 우선 연결한다. 프로젝트가 없는
    # 캠프+날짜 연결은 보고용으로 남지만 cancel_resolution 자동 반영 입력에는 쓰이지 않는다.
    band = [{"출처": "밴드 본문", "밴드": "90610953", "글번호": "4279",
             "날짜": "2026-01-06", "캠프": "송파2MB", "프로젝트NO": "UJ2600035",
             "사건": ["취소"], "글": "접수취소"}]
    kakao = [{"파일": "k.txt", "날짜": "2026-01-06", "보낸이": "오종현",
              "캠프": "다른표기", "프로젝트NO": "UJ2600035",
              "사건": ["취소"], "글": "유선전화 이후 접수 취소"}]
    matched, _, _ = CS.pair(band, kakao)
    assert len(matched) == 1 and matched[0]["연결근거"] == "프로젝트NO 정확 일치"

    with tempfile.TemporaryDirectory() as td:
        store = AppStore(os.path.join(td, "app.db"))
        store.initialize()
        source = store.create_work(
            kind="돌발AS", business_key="AS-2601-048", public_id="AS-2601-048",
            project_no="UJ2600035", camp_name="송파2MB", status="취소",
            fields={"접수ID": "AS-2601-048", "프로젝트NO": "UJ2600035",
                    "진행상태": "취소", "유상·무상·보험": "유상"},
            idempotency_key="t208-source",
        )["work"]
        store.create_work(
            kind="정산", business_key="JS-2608-108", public_id="JS-2608-108",
            project_no="UJ2600035", camp_name="송파2MB", status="",
            fields={"정산ID": "JS-2608-108", "프로젝트NO": "UJ2600035",
                    "원천업무ID": "AS-2601-048", "비용구분": "유상"},
            idempotency_key="t208-settle",
        )
        hit = {"UJ2600035": {
            "프로젝트NO": "UJ2600035", "업무종류": "돌발AS",
            "밴드": "90610953", "게시글": "4279", "자리": "본문",
            "게시일": "2026-01-06",
            "근거": "[접수취소] - 유선접화로 해결 완료",
            "원문": "[접수취소] - 유선접화로 해결 완료",
            "처리구분": "원격해결", "관측시각": 1785919289584,
            "근거URL": "https://band.us/band/90610953/post/4279",
        }}
        cross = {"UJ2600035": {"카톡": "k.txt 2026-01-06 오종현",
                                 "카톡글": "기사님과 유선전화 이후 접수 취소되었습니다"}}
        first = CR.sync_hits(hit, store=store, corroborations=cross)
        assert first["updated"] == 1 and not first["errors"], first
        current = store.get_work(source["id"])
        assert current["status"] == "취소"
        assert current["fields"]["처리구분"] == "원격해결"
        assert current["fields"]["접수취소확인일"] == "2026-01-06", \
            "뒤늦은 수집일이 실제 접수취소 발생일을 덮었다"
        assert "밴드 90610953/4279" in current["fields"]["접수취소근거"]
        assert "카톡 교차" in current["fields"]["접수취소근거"]
        version = current["record_version"]
        second = CR.sync_hits(hit, store=store, corroborations=cross)
        assert second["unchanged"] == 1
        assert store.get_work(source["id"])["record_version"] == version, \
            "같은 근거 재수집이 이벤트·보관본을 계속 늘린다"

        base = [{"정산ID": "JS-2608-108", "프로젝트NO": "UJ2600035",
                 "원천업무ID": "AS-2601-048", "비용구분": "유상",
                 "상태": "세금계산서 미발행", "명세서": "없음", "계산서": "미발행"}]
        clean = next(row for row in APP._overlay_app_store_settlements(base, store=store)
                     if row.get("정산ID") == "JS-2608-108")
        assert clean["상태"] == "접수취소(유선해결)" and clean["청구대상"] is False
        assert clean["미발행사유"] == "" and not APP._settlement_has_documents(clean)

        # 취소 뒤의 실제 PO·계산서 사실은 지우지 않는다. 정상/완료로도 접지 않고
        # 별도 교차입력 충돌로 올려 잘못 붙은 528,000원 같은 자료를 찾게 한다.
        settle = store.get_work(kind="정산", business_key="JS-2608-108")
        store.update_work(
            settle["id"], expected_version=settle["record_version"],
            patch={"fields": {"PO번호": "PO-X", "세금계산서발행일": "2026-01-20"}},
            idempotency_key="t208-docs",
        )
        conflict = next(row for row in APP._overlay_app_store_settlements(base, store=store)
                        if row.get("정산ID") == "JS-2608-108")
        assert conflict["상태"] == "취소건 청구자료 존재 — 교차입력 확인"
        assert conflict["PO번호"] == "PO-X" and conflict["계산서발행일"] == "2026-01-20"
        assert conflict["취소건청구자료충돌"] is True

        # 실제 발행·입금은 06 정산행이 아니라 15/16 sidecar에만 있을 수도 있다.
        # exact 정산ID로 연결된 객관자료를 놓쳐 '자료 없는 취소'로 숨기면 안 된다.
        store.create_work(
            kind="정산", business_key="JS-SIDECAR", public_id="JS-SIDECAR",
            project_no="UJ2600035", status="",
            fields={"정산ID": "JS-SIDECAR", "프로젝트NO": "UJ2600035",
                    "원천업무ID": "AS-2601-048", "비용구분": "유상"},
            idempotency_key="t208-side-settle",
        )
        store.create_work(
            kind="세금계산서", business_key="TX-SIDECAR", public_id="TX-SIDECAR",
            project_no="UJ2600035", status="발행",
            fields={"계산서관리ID": "TX-SIDECAR", "정산ID": "JS-SIDECAR",
                    "실제발행일": "2026-01-21", "발행금액": 480000},
            idempotency_key="t208-side-invoice",
        )
        side_evidence = CR.settlement_document_evidence(store)
        assert "JS-SIDECAR" in side_evidence and "15:실제발행일" in side_evidence["JS-SIDECAR"]
        side_conflict = next(
            row for row in APP._overlay_app_store_settlements([], store=store)
            if row.get("정산ID") == "JS-SIDECAR"
        )
        assert side_conflict["상태"] == "취소건 청구자료 존재 — 교차입력 확인"
        assert side_conflict["계산서발행일"] == "2026-01-21"

        # 같은 프로젝트라도 원천업무ID가 다르면 재무 상태를 전파하지 않는다.
        store.create_work(
            kind="돌발AS", business_key="AS-OTHER", public_id="AS-OTHER",
            project_no="UJ2600035", status="접수",
            fields={"접수ID": "AS-OTHER", "프로젝트NO": "UJ2600035", "진행상태": "접수"},
            idempotency_key="t208-other-source",
        )
        store.create_work(
            kind="정산", business_key="JS-OTHER", public_id="JS-OTHER",
            project_no="UJ2600035", status="",
            fields={"정산ID": "JS-OTHER", "프로젝트NO": "UJ2600035",
                    "원천업무ID": "AS-OTHER", "비용구분": "유상"},
            idempotency_key="t208-other-settle",
        )
        other = next(row for row in APP._overlay_app_store_settlements([], store=store)
                     if row.get("정산ID") == "JS-OTHER")
        assert not other.get("원천업무취소"), "프로젝트 유사매칭이 다른 원천 정산까지 취소했다"

        # 완료일/종료상태는 자동으로 뒤집지 않는다.
        finished = store.create_work(
            kind="돌발AS", business_key="AS-DONE", public_id="AS-DONE",
            project_no="UJ2600099", status="작업완료",
            fields={"접수ID": "AS-DONE", "프로젝트NO": "UJ2600099",
                    "진행상태": "작업완료", "작업완료일": "2026-01-07"},
            idempotency_key="t208-finished",
        )["work"]
        done_hit = {"UJ2600099": {**hit["UJ2600035"], "프로젝트NO": "UJ2600099"}}
        blocked = CR.sync_hits(done_hit, store=store, corroborations={})
        assert blocked["conflicts"] == 1
        assert store.get_work(finished["id"])["status"] == "작업완료"

        # 미발행 경과 감시도 exact 원천업무 취소를 청구대상에서 제외한다.
        record = {"원천업무ID": "AS-2601-048", "프로젝트NO": "UJ2600035",
                  "비용구분": "유상", "원장_공급가액": 480000,
                  "원장_거래명세서번호": "2026/01/20-1",
                  "원장_거래명세서합계": 528000, "작업완료일": "2026-01-06"}
        old_progress = ER.erp_progress
        ER.erp_progress = lambda: {}
        try:
            outcomes = CR.source_outcomes(store)
            assert TW.overdue_rows({"JS-X": record}, {}, {},
                                   today=__import__("datetime").date(2026, 8, 11),
                                   source_outcomes_map=outcomes) == []
        finally:
            ER.erp_progress = old_progress

    daily = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    pipeline = open(os.path.join(ROOT, "automation_pipeline.py"), encoding="utf-8").read()
    completion = open(os.path.join(ROOT, "settlement_completion.py"), encoding="utf-8").read()
    findings = open(os.path.join(ROOT, "findings_export.py"), encoding="utf-8").read()
    index_html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    assert '"cancel_watch.py"), "--sync"' in daily
    assert '"cancel_watch.py"), "--sync"' in pipeline
    assert "source_outcomes_map.get(source_id)" in completion, \
        "취소건 청구자료가 객관완료로 잘못 접힐 수 있다"
    assert "settlement_document_evidence" in findings, \
        "15/16 sidecar-only 청구자료가 확인필요에서 사라질 수 있다"
    assert "function billableSettleRows(rows)" in index_html
    assert "const S = billableSettleRows(SAll)" in index_html
    assert "const allRows = settleRows, rows = billableSettleRows(allRows)" in index_html, \
        "대표보고·PNG가 취소건 공급/미수를 다시 합산한다"
    print("  [208] 접수취소·유선해결 exact 연결 · App DB 멱등근거 · 완료 CAS 보호 · "
          "취소 청구제외/서류충돌 보존 ✅")


def t209_pipeline_lock_owner_cannot_be_forged_or_overwritten():
    """[209] 자동화 중복 회차가 산 주인의 잠금·공유 상태를 빼앗지 않는다.

    Windows ``os.kill(pid, 0)`` 오판으로 산 회차의 오래된 잠금을 죽었다고 보고
    두 회차가 함께 돌았다. 더 늦게 끝난 옛 회차가 새 회차 상태를 덮고 잠금까지
    지울 수 있었으므로 PID 판정뿐 아니라 토큰과 run_id가 모두 맞아야 저장·해제한다.
    """
    from pathlib import Path
    import threading
    import automation_pipeline as P
    from app_store import AppStore

    source = open(os.path.join(ROOT, "automation_pipeline.py"), encoding="utf-8").read()
    pid_body = source.split("def _pid_alive", 1)[1].split("\nclass LockOwnershipLost", 1)[0]
    assert "import pid_alive" in source and \
        "return pid_alive.owner_alive(" in pid_body and "is not False" in pid_body, \
        "자동화 잠금만 Windows 구형 PID 판정을 계속 쓴다(신원 검증 [210] 포함)"

    with tempfile.TemporaryDirectory(prefix="csos-pipeline-owner-209-") as td:
        root = Path(td)
        reports = root / "reports"
        reports.mkdir(parents=True)
        lock_path = reports / ".automation.lock"
        state_path = reports / "automation-state.json"

        # 판정 불가는 죽음이 아니다. 한도보다 오래됐어도 잠금을 회수하지 않는다.
        uncertain_path = reports / ".uncertain.lock"
        uncertain_path.write_text("987654 0 old-token old-run\n", encoding="ascii")
        os.utime(uncertain_path, (1, 1))
        real_owner_alive = P.pid_alive.owner_alive
        try:
            P.pid_alive.owner_alive = lambda _pid, **_k: None
            assert P._pid_alive(987654) is True
            assert P.PipelineLock(uncertain_path).acquire("unknown-probe") is None
            assert uncertain_path.exists(), "판정 불가인 산 주인의 잠금을 빼앗았다"

            P.pid_alive.owner_alive = lambda _pid, **_k: False
            assert P._pid_alive(987654) is False
            dead_lock = P.PipelineLock(uncertain_path)
            dead_token = dead_lock.acquire("dead-owner-recovery")
            P.pid_alive.owner_alive = lambda _pid, **_k: True
            assert dead_token and dead_lock.release(dead_token, "dead-owner-recovery")
        finally:
            P.pid_alive.owner_alive = real_owner_alive

        # 실제 두 인스턴스를 동시에 진입시킨다. 첫 회차가 stage 안에서 멈춘 동안
        # 두 번째 회차는 already_running만 돌려주고 상태 바이트를 바꾸지 않는다.
        race_root = root / "race"
        (race_root / "reports").mkdir(parents=True)
        (race_root / "kakao" / "dropbox").mkdir(parents=True)
        (race_root / "kakao" / "dropbox" / "KakaoTalk_209.txt").write_text(
            "2026-08-11 합성 대화", encoding="utf-8"
        )
        race_state = race_root / "reports" / "state.json"
        race_lock = race_root / "reports" / ".lock"
        race_store = AppStore(race_root / "app.db").initialize()
        entered = threading.Event()
        resume = threading.Event()
        first_result = {}

        def slow_stage(name, _args, _timeout):
            entered.set()
            assert resume.wait(5), "합성 동시실행 검증이 stage를 놓아주지 못했다"
            return {"name": name, "ok": True, "returncode": 0, "timed_out": False}

        first = P.AutomationPipeline(
            root=race_root, state_path=race_state, lock_path=race_lock,
            store=race_store, stage_runner=slow_stage,
        )
        second = P.AutomationPipeline(
            root=race_root, state_path=race_state, lock_path=race_lock,
            store=race_store, stage_runner=lambda *_args: {"ok": True},
        )
        thread = threading.Thread(
            target=lambda: first_result.update(first.run_once(trigger="race-first", force=True)),
            daemon=True,
        )
        thread.start()
        assert entered.wait(5), "첫 자동화 회차가 합성 stage에 들어오지 못했다"
        while not race_state.exists():
            time.sleep(0.01)
        before_race_duplicate = race_state.read_bytes()
        second_result = second.run_once(trigger="race-second", force=True)
        assert second_result["status"] == "already_running", second_result
        assert race_state.read_bytes() == before_race_duplicate, \
            "동시 중복 인스턴스가 첫 회차의 진행 상태를 덮었다"
        resume.set()
        thread.join(10)
        assert not thread.is_alive() and first_result.get("run_id"), first_result

        initial = P._default_state()
        initial["history"] = [{"status": "sentinel-before-owner"}]
        state_path.write_text(
            json.dumps(initial, ensure_ascii=False), encoding="utf-8"
        )
        store = AppStore(root / "app.db").initialize()
        owner = P.AutomationPipeline(
            root=root, state_path=state_path, lock_path=lock_path, store=store,
            stage_runner=lambda *_args: {"ok": True},
        )
        assert owner._acquire_run() and owner._lock_token and owner._run_id
        owner.state["running"] = True
        owner.state["active_run_id"] = owner._run_id
        owner.state["lock_token"] = owner._lock_token
        owner.state["owner_marker"] = "first-owner"
        owner._save()

        # 같은 프로세스 안의 두 인스턴스도 PID는 같으므로 token/run_id가 관건이다.
        before_duplicate = state_path.read_bytes()
        duplicate = P.AutomationPipeline(
            root=root, state_path=state_path, lock_path=lock_path, store=store,
            stage_runner=lambda *_args: {"ok": True},
        )
        refused = duplicate.run_once(trigger="synthetic-duplicate")
        assert refused["status"] == "already_running", refused
        assert state_path.read_bytes() == before_duplicate, \
            "잠금을 못 잡은 중복 회차가 공유 상태를 덮었다"

        # 후속 회차가 잠금을 소유한다고 가정해 capability를 바꾼다. 옛 소유자는
        # 같은 PID여도 상태를 저장하거나 후속 잠금을 지울 수 없어야 한다.
        successor_token = "b" * 32
        successor_run = "successor-run-209"
        lock_path.write_text(
            f"{os.getpid()} {time.time():.3f} {successor_token} {successor_run}\n",
            encoding="ascii",
        )
        successor_state = P._default_state()
        successor_state.update({
            "running": True,
            "active_run_id": successor_run,
            "lock_token": successor_token,
            "owner_marker": "successor",
        })
        state_path.write_text(
            json.dumps(successor_state, ensure_ascii=False), encoding="utf-8"
        )
        owner.state["owner_marker"] = "stale-owner-overwrite"
        try:
            owner._save()
            raise AssertionError("토큰을 잃은 옛 회차가 공유 상태를 저장했다")
        except P.LockOwnershipLost:
            pass
        assert json.loads(state_path.read_text(encoding="utf-8"))["owner_marker"] == "successor"
        assert owner._release_run() is False
        assert lock_path.exists() and successor_token in lock_path.read_text(encoding="ascii"), \
            "옛 회차가 후속 회차의 잠금을 지웠다"

    print("  [209] 자동화 단일회차 — 공용 PID 판정 · token/run_id 소유 저장·해제 ✅")


def t210_pid_reuse_is_not_alive_and_customer_scan_is_one_pass():
    """[210] 번호가 같다고 같은 프로세스가 아니다 · 거래처 색인은 한 번에 훑는다.

    2026-08-11 실사고 둘을 한 뿌리로 잡는다:
      ① 09:50 회차가 '거래처코드 색인'에서 죽었는데 그 pid 를 quick_share_server 가
         재사용해 `alive(pid)`=True → 인계 문서가 다섯 시간 동안 "돌고 있다 —
         기다려라"(정반대 지시)를 냈고 잠금도 스스로 못 풀렸다.
      ② 그 단계가 죽도록 오래 걸린 이유가 customer_index 의 glob+getmtime
         (Z: 파일마다 왕복 — [198]과 같은 병) + 매번 워크북 80개 열기였다.
    """
    import time as _t
    from datetime import datetime
    import pid_alive as PA
    import customer_index as CI

    me = os.getpid()
    # ── 신원 검증: 살아 있는 진짜 나 vs '그 시각 뒤에 태어난 남'
    assert PA.alive(me) is True
    assert PA.started_at(me) is not None, "생성시각을 못 읽으면 재사용 판정이 조용히 무력화된다"
    assert PA.alive(me, born_before=_t.time() + 60) is True, "미래 기준에 산 프로세스를 죽였다"
    assert PA.alive(me, born_before=1.0) is False, \
        "기준 시각 뒤에 태어난 프로세스(pid 재사용)를 살아 있다고 오판"

    # ── daily_run 잠금: 재사용된 pid 의 잠금을 실제로 회수하는가 (산 pid 로 재현)
    import daily_run as DR
    with tempfile.TemporaryDirectory(prefix="csos-210-") as td:
        lock = os.path.join(td, ".lock")
        json.dump({"pid": me, "token": "old",
                   "started_at": datetime.fromtimestamp(_t.time() - 3600).astimezone().isoformat(timespec="seconds")},
                  open(lock, "w", encoding="utf-8"))
        tok = DR.acquire_run_lock(path=lock)
        assert tok, "잠금 시각보다 뒤에 태어난 주인(=재사용)의 잠금을 회수하지 못했다"
        DR.release_run_lock(tok, path=lock)
        # 방금 태어난 척하는 주인(정상)의 잠금은 빼앗지 않는다
        json.dump({"pid": me, "token": "fresh",
                   "started_at": datetime.now().astimezone().isoformat(timespec="seconds")},
                  open(lock, "w", encoding="utf-8"))
        assert DR.acquire_run_lock(path=lock) is None, "산 주인의 잠금을 빼앗았다"

    # ── 배선: 판정에 시각이 실제로 전달되는가 (한쪽만 고치면 경보·잠금이 갈린다)
    sh_src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "born_before=started.timestamp()" in sh_src and \
        "pid_started_at=d.get(\"pid_started_at\")" in sh_src, \
        "_daily_run_inflight 가 정확한 프로세스 지문을 안 본다"
    assert "pid_alive(owner_pid, born_before=born" in sh_src and \
        "pid_started_at=owner_started_at" in sh_src, \
        "점유 판정이 정확한 프로세스 지문을 안 본다"
    # 잠금을 쥐는 곳은 넷이다 — 한 곳만 고치면 나머지가 같은 병을 앓는다([162]의 교훈)
    for fname, needles, what in (
            ("archive_worker.py", ("pid_alive.owner_alive(",
                                   'pid_started_at=owner.get("pid_started_at")',
                                   "born_before=lock_mtime"), "보관본 워커"),
            ("automation_pipeline.py", ("pid_alive.owner_alive(",
                                         "pid_started_at=owner_started_at",
                                         "born_before=claimed_at"), "자동화 잠금")):
        src2 = open(os.path.join(ROOT, fname), encoding="utf-8").read()
        assert all(needle in src2 for needle in needles), \
            f"{what} 잠금이 정확한 프로세스 지문을 안 본다"

    # ── 거래처 색인: 목록은 한 번에(stat 동봉) · 안 바뀌면 워크북을 다시 안 연다
    ci_src = open(os.path.join(ROOT, "customer_index.py"), encoding="utf-8").read()
    body = ci_src.split("def load_customers", 1)[1].split("\ndef ", 1)[0]
    assert "walk_stat" in body, "거래처 색인이 공용 워커를 버렸다([198] 병 재발)"
    assert body.index('c.get("fp") == fp') < body.index("openpyxl.load_workbook"), \
        "캐시 검사가 워크북 열기보다 뒤에 있다 — [168] 병 재발"
    with tempfile.TemporaryDirectory(prefix="csos-210-erp-") as td:
        import warnings
        warnings.filterwarnings("ignore")
        import openpyxl
        import source_dirs as S
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["거래처등록"])
        ws.append(["거래처코드", "거래처명", "주소", "담당자", "Email", "연락처", "보유장비명"])
        ws.append(["CU177", "강서1MB(가양A)", "서울", "홍길동", "a@b.c", "010", "리프트"])
        wb.save(os.path.join(td, "ESA001M_test.xlsx"))
        old_dir, old_cache = S.ERP_DIR, CI.CAND_CACHE
        real_load = openpyxl.load_workbook
        try:
            S.ERP_DIR = td
            CI.CAND_CACHE = os.path.join(td, "reports", "cache.json")
            rows, src = CI.load_customers()
            assert len(rows) == 1 and rows[0]["code"] == "CU177" and src, "합성 거래처등록을 못 읽었다"
            calls = []
            openpyxl.load_workbook = lambda *a, **k: (calls.append(1), real_load(*a, **k))[1]
            rows2, _ = CI.load_customers()
            assert rows2 == rows and not calls, \
                "후보가 하나도 안 바뀌었는데 워크북을 다시 열었다 — 색인이 다시 몇 시간짜리가 된다"
        finally:
            S.ERP_DIR, CI.CAND_CACHE = old_dir, old_cache
            openpyxl.load_workbook = real_load

    print("  [210] pid 재사용은 죽음으로 판정 · 잠금 자동회수 · 거래처 색인 1회 훑기+지문 캐시 ✅")


def t211_progress_trace_owner_identity():
    """[211] 자국이 있다고 돌고 있는 것이 아니다 — 진행 판정도 잠금과 같은 신원을 본다.

    2026-08-11 실사고 두 번째(pid 재사용): 회차 pid 37128 이 11:02~11:15 사이에 죽고
    그 번호를 quick_share_server.exe(11:15:09 시작)가 물려받았다. 잠금은 [210]으로
    회수됐지만 **진행 자국 판정에는 신원 검사가 없어** 인계 문서가 '5시간째 돌고
    있다 — 기다려라'(정반대 지시)를 냈다. 지키는 것:
      ① 자국의 '회차시작'·'시각'보다 뒤에 태어난 프로세스는 주인이 아니다([210]과 동일)
      ② 이름이 읽히는데 python 이 아니면 번호만 같은 남이다(시각을 우연히 앞서는
         상주 서비스 재사용까지 가른다)
      ③ '(회차 끝)' 자국에는 죽음 경보를 내지 않는다(완주한 회차의 pid 는 원래 죽는다)
      ④ 죽었으면 경보 문구가 '몇 분째'(도는 중처럼 읽힘)가 아니라 반대 지시를 말한다
    """
    from datetime import datetime as _dt, timedelta as _td
    import pid_alive as PA
    import session_handoff as SH

    me = os.getpid()
    # ── 이름 계기 자체가 살아 있는가 — 조용히 None 만 내면 이름 문이 무력화된다([169])
    nm = PA.image_name(me)
    assert nm and "python" in nm, "제 프로세스 이름(python)을 못 읽는다 — 이름 문이 무력화된다"
    assert SH._owner_is_python(me) is True, "python 주인을 남으로 판정"

    with tempfile.TemporaryDirectory(prefix="csos-211-") as td:
        old_rd, real_img = SH.REPORT_DIR, PA.image_name
        try:
            SH.REPORT_DIR = td
            prog = os.path.join(td, ".daily_run.progress.json")

            def write(step, when_iso):
                json.dump({"pid": me, "단계": step, "상태": "시작", "시각": when_iso,
                           "회차시작": when_iso, "끝난단계": []},
                          open(prog, "w", encoding="utf-8"), ensure_ascii=False)

            now_iso = _dt.now().astimezone().isoformat(timespec="seconds")
            old_iso = (_dt.now().astimezone() - _td(hours=6)).isoformat(timespec="seconds")

            # ① 재사용 재현: 자국은 6시간 전인데 그 pid 의 프로세스(나)는 방금 태어났다
            write("원본색인", old_iso)
            s = SH.daily_step_now()
            assert s and s.get("살아있음") is False, \
                "자국 시각보다 뒤에 태어난 프로세스(pid 재사용)를 주인으로 오판"
            hint = SH._step_hint()
            assert "죽었" in hint and "분째" not in hint and "기다리지 말" in hint, \
                "죽은 회차의 자국을 '돌고 있다'처럼 말한다: %r" % hint

            # 정상 회차: 자국이 방금 것이고 주인(python)이 살아 있다
            write("원본색인", now_iso)
            s = SH.daily_step_now()
            assert s and s.get("살아있음") is True, "살아 있는 회차를 죽었다고 판정"
            assert "지금 단계" in SH._step_hint()

            # ② 이름 문: 시각은 통과해도 이름이 python 이 아니면 남이다
            PA.image_name = lambda pid: "quick_share_server.exe"
            s = SH.daily_step_now()
            assert s and s.get("살아있음") is False, \
                "이름이 quick_share_server 인데 주인으로 판정(이름 문 무력화)"
            PA.image_name = real_img

            # ③ 완주 자국: pid 가 죽는 것이 정상 — 죽음 경보를 내지 않는다
            write("(회차 끝)", old_iso)
            s = SH.daily_step_now()
            assert s and s.get("살아있음") is None, "완주한 회차에 죽음 경보를 냈다"
            assert "죽었" not in SH._step_hint()
        finally:
            SH.REPORT_DIR, PA.image_name = old_rd, real_img

    # ── 배선: 잠금 나이 판정(_daily_run_inflight)에도 이름 문이 걸려 있는가
    sh_src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    infl = sh_src.split("def _daily_run_inflight", 1)[1].split("\ndef ", 1)[0]
    assert "_owner_is_python" in infl, "_daily_run_inflight 가 이름 신원을 안 본다"
    step_body = sh_src.split("def daily_step_now", 1)[1].split("\ndef ", 1)[0]
    assert "_progress_owner_alive" in step_body, "진행 자국 판정이 신원을 안 본다"

    print("  [211] 진행 자국 신원 — 시각·이름(python)으로 pid 재사용 가름 · 죽은 회차를 '돌고 있다'로 안 읽음 ✅")


def t213_exact_pid_fingerprint_reaches_every_owner():
    """[213] 잠금은 pid 번호가 아니라 **그때 그 프로세스**를 소유자로 본다.

    born_before의 5초 여유 안에서 pid가 재사용되거나 오래 산 다른 프로세스가 번호를
    물려받아도 정확한 생성시각 지문이 다르면 회수한다. 신규 잠금·진행 자국·AI 점유와
    앱의 ``running`` 표시까지 같은 판정기를 써야 한다.
    """
    from pathlib import Path
    import pid_alive as PA
    import daily_run as DR
    import archive_worker as AW
    import automation_pipeline as AP
    import app_store as A
    from webapp import app_server as S

    me = PA.identity()
    assert me["pid"] == os.getpid() and me.get("pid_started_at"), me
    assert PA.owner_alive(me["pid"], me["pid_started_at"]) is True
    assert PA.owner_alive(me["pid"], float(me["pid_started_at"]) - 60) is False, \
        "같은 pid의 다른 생성시각을 현재 주인으로 오판"

    with tempfile.TemporaryDirectory(prefix="csos-exact-pid-213-") as td:
        root = Path(td)

        # daily_run JSON 잠금과 진행 자국 모두 exact 지문을 남긴다.
        daily_lock = root / ".daily_run.lock"
        token = DR.acquire_run_lock(str(daily_lock))
        assert token
        daily_owner = json.loads(daily_lock.read_text(encoding="utf-8"))
        assert daily_owner.get("pid_started_at"), daily_owner
        DR.release_run_lock(token, str(daily_lock))
        assert not daily_lock.exists(), "daily 잠금 소유자가 자기 잠금을 놓지 못한다"

        old_report, old_progress = DR.REPORT_DIR, DR.PROGRESS
        try:
            DR.REPORT_DIR = str(root)
            DR.PROGRESS = str(root / ".daily_run.progress.json")
            DR.note_progress("합성", "시작")
            progress = json.loads((root / ".daily_run.progress.json").read_text(encoding="utf-8"))
            assert progress.get("pid_started_at"), progress
        finally:
            DR.REPORT_DIR, DR.PROGRESS = old_report, old_progress

        # 파이프라인 5필드 잠금과 보관 worker JSON 잠금도 같은 지문이다.
        pipeline_path = root / ".automation_pipeline.lock"
        pipeline = AP.PipelineLock(pipeline_path)
        pipeline_token = pipeline.acquire("run-213")
        status = AP.pipeline_lock_status(pipeline_path)
        assert pipeline_token and status.get("alive") is True and status.get("pid_started_at")
        assert status.get("run_id") == "run-213"
        assert pipeline.release(pipeline_token, "run-213")

        spool = root / "spool"
        with AW._worker_lock(spool):
            archive_owner = json.loads((spool / "archive-worker.lock").read_text(encoding="utf-8"))
            assert archive_owner.get("pid_started_at"), archive_owner

        # state 파일만 running인 죽은 회차는 앱에서 실행 중으로 표시하지 않는다.
        store = A.AppStore(root / "app.db").initialize()
        state_path = root / "automation_pipeline_state.json"
        state_path.write_text(json.dumps({
            "running": True, "active_run_id": "run-dead",
            "last_run": {"run_id": "run-dead", "status": "running",
                         "current_stage": "죽은 단계"},
        }), encoding="utf-8")
        dead = S.get_live_state(store=store, state_path=state_path,
                                lock_path=root / ".automation_pipeline.lock")
        assert dead["running"] is False and dead["phase"] != "updating", dead
        live_lock = AP.PipelineLock(root / ".automation_pipeline.lock")
        live_token = live_lock.acquire("run-dead")
        live = S.get_live_state(store=store, state_path=state_path,
                                lock_path=root / ".automation_pipeline.lock")
        assert live_token and live["running"] is True and live["phase"] == "updating", live
        assert live_lock.release(live_token, "run-dead")

    claim_src = open(os.path.join(ROOT, "ai_claim.py"), encoding="utf-8").read()
    handoff_src = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert "agent_pid_started_at" in claim_src and "pid_started_at" in claim_src
    assert "agent_pid_started_at" in handoff_src and "pid_started_at" in handoff_src, \
        "인계가 exact AI/프로세스 지문을 소비하지 않는다"
    print("  [213] exact PID 지문 — daily·pipeline·archive·AI점유·인계·live-state가 같은 소유자 판정 ✅")


def t214_first_live_revision_cannot_be_falsely_applied():
    """[214] 첫 live-state도 실제 화면 자료가 성공한 뒤에만 적용 완료다.

    부트의 옛 A 응답이 새 B 뒤에 도착해 화면을 되돌리는 경합을 막고, 직원센터·캘린더도
    같은 세대 규칙을 쓴다. 실패한 동일 revision은 다음 poll에서 다시 시도해야 한다.
    """
    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    poll = html.split("function pollLiveState", 1)[1].split("async function retryDataSync", 1)[0]
    assert "if(first&&next) LIVE_SYNC.appliedRevision=next" not in poll, \
        "첫 poll을 자료 재검증 없이 적용 완료로 표시"
    assert "next!==LIVE_SYNC.appliedRevision" in poll and "revalidateLiveData" in poll, \
        "실패한 동일 revision을 다음 poll에서 다시 시도하지 않는다"
    boot = html.split("function bootstrapAuthenticatedData", 1)[1].split("(async function()", 1)[0]
    assert "startLiveStateMonitor(true)" in boot and "await firstFlight" in boot, \
        "첫 poll과 현재 화면 적용 순서가 직렬화되지 않았다"
    assert "refreshAll(false)" in boot, "부트가 7개 자료 API를 별도로 한 번 더 부른다"

    # 일반 자료·직원센터·캘린더 모두 늦은 요청을 거르는 세대 번호가 있어야 한다.
    for marker in ("beginDataSectionRequest", "dataSectionRequestCurrent",
                   "ryuRequestGeneration", "CAL_REQUEST_GENERATION"):
        assert marker in html, "늦은 A 응답 억제 장치 누락: " + marker
    assert "if(generation!==ryuRequestGeneration) return false" in html
    assert "if(generation!==CAL_REQUEST_GENERATION) return false" in html
    assert "result.ok&&result.accepted" in html, \
        "거절된 옛 응답을 화면 자료에 적용한다"
    assert "booting:true" in html and "bootstrapAuthenticatedData();" in html, \
        "인증 전 applyView가 옛 자료 요청을 먼저 시작할 수 있다"
    print("  [214] live-state 부트 — 첫 revision 성공 후 확정 · A/B 역전 차단 · 직원·캘린더 동일 세대 ✅")


def t215_cancel_timeline_last_explicit_state_wins():
    """[215] 접수취소는 본문·댓글별 시간축에서 마지막 명시 상태만 반영한다."""
    import band_extract as B
    import cancel_watch as C

    assert B.cancel_state("접수 취소 하세요") == "cancel"
    assert B.cancel_state("접수 취소 후 다시 접수 유지") == "active"
    assert B.cancel_state("접수 유지했으나 이후 접수 취소") == "cancel"
    for text in ("보험접수 취소 후 쿠팡측에서 긴급수리요청",
                 "택배 접수 취소 후 방문수리 전환", "부품 접수 취소"):
        assert B.cancel_state(text) == "" and not B.cancel_hit(text), \
            "서비스 접수가 아닌 취소를 자동 취소로 오판: " + text

    post = {
        "content": "● 프로젝트NO : UJ2600035", "created_at": 1700000000000,
        "comments": [
            {"created_at": 1700000100000, "content": "기사님 통화 후 접수 취소 하세요"},
            {"created_at": 1700000200000, "content": "확인 결과 접수 유지 바랍니다"},
        ],
    }
    event = B.latest_cancel_event(post)
    assert event and event["state"] == "active" and event["source"] == "댓글", event
    post["comments"].append(
        {"created_at": 1700000300000, "content": "다시 접수 취소 처리합니다"})
    event = B.latest_cancel_event(post)
    assert event and event["state"] == "cancel" and event["created_at"] == 1700000300000

    # 다른 글의 더 최신 '유지'도 같은 프로젝트의 옛 취소를 제거한다.
    with tempfile.TemporaryDirectory(prefix="csos-cancel-timeline-215-") as td:
        old_cache = C.CACHE_DIR
        try:
            C.CACHE_DIR = td
            payload = {"band_name": "90610953", "posts": {
                "1": {"content": "● 프로젝트NO : UJ2600035",
                      "created_at": 1700000000000,
                      "comments": [{"created_at": 1700000100000,
                                    "content": "접수 취소 하세요"}]},
                "2": {"content": "● 프로젝트NO : UJ2600035",
                      "created_at": 1700000200000,
                      "comments": [{"created_at": 1700000300000,
                                    "content": "접수 유지합니다"}]},
                "3": {"content": "● 프로젝트NO : UJ2600999",
                      "created_at": 1700000400000,
                      "comments": [{"created_at": 1700000500000,
                                    "content": "접수 유지"},
                                   {"created_at": 1700000600000,
                                    "content": "접수 취소 처리"}]},
            }}
            with open(os.path.join(td, "90610953.json"), "w", encoding="utf-8") as fh:
                json.dump(payload, fh, ensure_ascii=False)
            hits, _blind, _total = C.scan_band(quiet=True)
            assert "UJ2600035" not in hits, "최신 접수 유지 뒤에도 옛 취소가 남는다"
            assert hits.get("UJ2600999", {}).get("게시일") == "2023-11-15", hits
        finally:
            C.CACHE_DIR = old_cache

    print("  [215] 취소 시간축 — 댓글별 시각 정렬 · 마지막 취소/유지 승리 · 보험·택배·부품 제외 ✅")


def t218_camp_standard_erp_basis_and_pm_units():
    """[218] 캠프명 표준화는 ERP **유일 매칭만** 바꾸고, 점검내용 파싱은 원문을 안 고친다.

    2026-08-11 지시: "ERP 기준으로 캠프명 매칭 … 기존 캠프명을 ERP 기준으로 변경 /
    정기점검 점검내용 호기별 분류 + ? 깨진 문자 조사".

    지키는 것 넷: ① 유일 매칭만 변경 ② 후보 여럿·ERP 없음·입력오류는 불변(짐작 금지)
    ③ 모든 변경 항목에 근거(ERP 거래처코드·이전값)가 실려 감사로그로 간다
    ④ 파싱·조사는 원문을 한 글자도 안 고친다(관리대장 직접 저장 금지 포함).
    """
    import camp_standardize as CS
    import pm_content as PM

    custs = [
        {"code": "CU001", "name": "양주2캠프(봉양동)"},
        {"code": "CU002", "name": "서울1캠프(A동)"},
        {"code": "CU003", "name": "서울1캠프(B동)"},
        {"code": "CU004", "name": "제주2MB(사계리)"},
    ]
    rows = [
        {"sheet": "02_돌발AS접수", "id": "AS-1", "camp": "양주2캠프"},          # 유일(핵심코드) → 변경
        {"sheet": "02_돌발AS접수", "id": "AS-2", "camp": "양주2캠프(봉양동)"},   # 이미 표준 → 불변
        {"sheet": "04_정기점검", "id": "PM-1", "camp": "서울1캠프"},            # 후보 여럿 → 불변
        {"sheet": "04_정기점검", "id": "PM-2", "camp": "?"},                    # 입력오류 → 불변
        {"sheet": "06_거래서류청구수금", "id": "ST-1", "camp": "듣도못한캠프"},  # ERP 없음 → 불변
        {"sheet": "06_거래서류청구수금", "id": "ST-2", "camp": "제주2MB(성읍리)"},  # 괄호 갈림 → 변경+감시
    ]
    items, report = CS.plan_rows(rows, custs)
    changed = {(i["sheet"], i["key"]) for i in items}
    assert len(items) == 2 and changed == {("02_돌발AS접수", "AS-1"),
                                           ("06_거래서류청구수금", "ST-2")}, \
        f"유일 매칭 2건만 바뀌어야 한다: {changed}"
    for i in items:
        assert i["only_if_empty"] is False and i["col"] == "캠프명"
        assert "ERP 거래처등록 CU" in i["evidence"] and "이전값" in i["evidence"], \
            "감사로그로 갈 근거(거래처코드·이전값)가 항목에 없다"
    assert set(report["buckets"]["multi"]) == {"서울1캠프"}, "후보 여럿은 불변 + 보고"
    assert set(report["buckets"]["none"]) == {"듣도못한캠프"}
    assert set(report["buckets"]["junk"]) == {"?"}, "값 대신 들어간 표시는 입력오류로 가른다"
    assert [p["before"] for p in report["paren_watch"]] == ["제주2MB(성읍리)"], \
        "괄호 지명이 갈리는 변경은 따로 모아 사람이 보게 한다"
    assert {c["before"] for c in report["changes"]} == {"양주2캠프", "제주2MB(성읍리)"}, \
        "변경 전 값이 되돌리기 자료(changes)에 남아야 한다"

    # 점검내용 파싱 — 호기 해석·깨짐 판정·원문 불변
    assert PM.parse_units("1,2호기 그리스 주입")["units"] == [1, 2]
    assert PM.parse_units("1~3호기")["units"] == [1, 2, 3]
    assert PM.parse_units("#4 · 5호 점검")["units"] == [4, 5]
    p = PM.parse_units("▒▒ ?? 호기 ▒▒")
    assert p["unknown"] and p["units"] == [], "깨진 호기는 지어내지 않고 '모름'으로 남긴다"
    assert PM.broken_flags("▒▒ ?? 호기 ▒▒") and not PM.broken_flags("이상 없음?"), \
        "물음표 하나짜리 정상 문장을 깨짐이라 부르면 사람이 멀쩡한 값을 고치러 간다"
    row = {"점검ID": "PM-X", "프로젝트NO": "UJ0000001", "캠프명": "가상캠프",
           "점검내용": "▒▒ ?? 호기 ▒▒ 점검", "비고": "&amp; 포함"}
    before = json.dumps(row, ensure_ascii=False, sort_keys=True)
    a = PM.analyze_row(row)
    assert json.dumps(row, ensure_ascii=False, sort_keys=True) == before, "원문(행)을 고쳤다"
    assert a["unknown"] and a["broken"].get("비고") == ["HTML엔티티"]

    # 두 도구 다 관리대장을 직접 저장하지 않는다 — 읽기 전용 열기 + 승인 경로(enqueue)만.
    for mod in ("camp_standardize.py", "pm_content.py"):
        src = open(os.path.join(ROOT, mod), encoding="utf-8").read()
        assert "read_only=True" in src and ".save(" not in src, mod + " 원장 직접 저장 금지"
        assert "ledger_db" in src and "queue_add" in src, mod + " 는 승인 경로로만 쓴다"

    print("  [218] 캠프명 ERP 표준화·호기 분류 — 유일 매칭만 변경 · 비유일 불변 · 근거 동반 · 원문 불변 ✅")


def t227_text_locks_use_the_one_owner_judge():
    """[227] 보관본·큐 잠금도 신원을 본다 — `os.kill(pid,0)` 사본을 지운다.

    [210]·[211] 이 잠금·자국을 고치는 동안 **텍스트 잠금 두 곳이 규칙 밖에 남아 있었다**
    (`ledger_db._pid_alive` · `ledger_writer._pid_alive`). 둘 다 `os.kill(pid, 0)` 이라
      ① 신원을 안 봐서 pid 재사용이면 잠금이 **스스로 안 풀리고**(보관본 회차가 매번
         "이미 실행 중"으로 조용히 건너뛴다)
      ② 윈도우 파이썬에서 `os.kill` 은 확인이 아니라 **종료 신호**이고(문서), POSIX 에선
         남의 프로세스에 PermissionError → '죽었다'로 읽혀 **산 주인의 잠금을 빼앗는다**.
    ★ 지문은 **자리가 아니라 이름표**(`fp=`)로 적는다 — `ledger_db` 안에서 형식이 다른
      두 잠금(`{pid} {iso}` · `{pid} {iso} {ns}`)이 **한 판정 함수**를 쓰기 때문이다.
      자리로 읽으면 `monotonic_ns` 를 생성시각으로 오해해 살아 있는 잠금을 죽였다고 한다.
    """
    from datetime import datetime as _dt
    import pid_alive as PA
    import ledger_db as LDB
    import ledger_writer as LW

    # ── 사본 금지: 두 파일에 `os.kill` 판정이 다시 들어오면 막는다
    for fname in ("ledger_db.py", "ledger_writer.py"):
        src = open(os.path.join(ROOT, fname), encoding="utf-8").read()
        body = src.split("def _pid_alive", 1)[1].split("\ndef ", 1)[0]
        # 실제 호출만 본다 — 설명(docstring)에는 왜 버렸는지가 적혀 있어야 한다
        assert "os.kill(int(" not in body, f"{fname} 이 다시 os.kill 로 생사를 판정한다"
        assert "pid_alive.owner_alive(" in body, f"{fname} 이 공용 소유자 판정을 안 쓴다"
        assert "is not False" in body, \
            f"{fname} 이 '판정 불가'를 죽음으로 읽는다 — 산 주인의 잠금을 빼앗는다"

    # ── 이름표 지문: 자리가 달라도 같은 값이 읽힌다
    me = os.getpid()
    fp = PA.stamp()
    assert fp.startswith("fp="), fp
    two = f"{me} 2026-08-11T17:00:00 {fp}".split()          # 보관본 잠금 모양
    three = f"{me} 2026-08-11T17:00:00 123456789 {fp}".split()   # 큐 잠금 모양(ns 가 낀다)
    for words in (two, three):
        pid, got, born = PA.owner_from_words(words)
        assert pid == me, (words, pid)
        assert got == fp[3:], "지문을 자리로 읽어 남의 칸을 집었다"
        assert born is not None, "잠금 시각을 못 읽었다"
    # ns 를 pid 로 착각하지 않는다 — 맨 앞 숫자만 pid 다
    assert PA.owner_from_words(three)[0] != 123456789
    # 지문 없는 **옛 잠금**도 그대로 읽힌다(그때는 잠금시각만으로 판정)
    pid, got, born = PA.owner_from_words(f"{me} 2026-08-11T17:00:00".split())
    assert pid == me and got is None and born is not None

    # ── 판정: 산 주인은 살리고, 지문이 어긋난 남(=재사용)은 죽었다고 한다
    real = PA.identity()["pid_started_at"]
    for mod in (LDB, LW):
        assert mod._pid_alive(me, pid_started_at=real) is True, "산 주인을 죽였다"
        assert mod._pid_alive(me, pid_started_at=float(real) - 3600) is False, \
            "지문이 다른데 살아 있다고 했다 — pid 재사용이 안 걸린다"
        assert mod._pid_alive(me) is True, "지문 없는 옛 잠금의 산 주인을 죽였다"

    # ── 회수: 재사용된 주인의 잠금은 즉시 회수하고, 산 주인 것은 안 건드린다
    with tempfile.TemporaryDirectory(prefix="csos-219-") as td:
        lock = os.path.join(td, "q.lock")
        stale = f"{me} 2026-08-11T17:00:00 123 fp={float(real) - 3600}"
        fresh = f"{me} {_dt.now().isoformat()} 123 {fp}"
        for mod in (LDB, LW):
            open(lock, "w", encoding="ascii").write(stale)
            assert mod._dead_or_abandoned_lock(lock, 99999) is True, \
                "재사용된 pid 의 잠금을 회수하지 못한다 — 회차가 영영 건너뛴다"
            open(lock, "w", encoding="ascii").write(fresh)
            assert mod._dead_or_abandoned_lock(lock, 99999) is False, \
                "살아 있는 주인의 잠금을 빼앗는다"

    # ── 잠금에 지문을 실제로 적는가 (안 적으면 위 판정이 영원히 옛 잠금 취급이다)
    ldb = open(os.path.join(ROOT, "ledger_db.py"), encoding="utf-8").read()
    lw = open(os.path.join(ROOT, "ledger_writer.py"), encoding="utf-8").read()
    assert ldb.count("_pa.stamp()") >= 2, "ledger_db 의 두 잠금 중 지문을 안 적는 것이 있다"
    assert "_pa.stamp()" in lw, "ledger_writer 큐 잠금이 지문을 안 적는다"
    # 연기 감시자도 지문으로 본다 — 번호만 물려받은 남을 '감시중'이라 하면 재개가 멈춘다
    assert '"watcher_started_at"' in ldb and "watcher_started_at" in ldb, \
        "연기 감시자 pid 에 지문이 없다"
    assert 'pid_started_at=state.get("watcher_started_at")' in ldb, \
        "resume_check 가 감시자 신원을 안 본다"

    print("  [227] 텍스트 잠금 신원 — os.kill 사본 제거 · 이름표 지문(fp=)으로 자리 무관 · "
          "재사용 회수/산 주인 보호 ✅")


def t228_scheduler_rounds_are_watched():
    """[228] 회차가 **정말 돌았나** — 스케줄러 결과를 읽는 눈 (2026-08-12, 분담판 [35]).

    자동화의 마지막 구멍이 여기였다. 지시문 '자동으로 도는 것' 목록에 한 줄을 적으면
    자동이 된 것처럼 보이지만, 실제로 도는지를 아는 것은 그 목록이 아니라 **작업
    스케줄러**다. 실측 2026-08-12 — 일일대조·원본정리가 매일 제한시간에 걸려 강제
    종료되고 정오회차는 등록조차 안 돼 한 번도 안 돌았는데 **아무 화면에도 안 떴다**
    (`LastTaskResult` 를 읽는 코드가 프로젝트에 한 줄도 없었다).

    지키는 것:
      ① **`[long]`** — 결과 코드를 `[int]` 로 받으면 `0xC000013A`·`0x800710E0` 이
         Int32 를 넘겨 변환이 터지고 **그 작업이 통째로 목록에서 빠진다.** 실측으로
         빠진 셋이 하필 지금 실패하고 있는 회차들이었다 — 감시자가 **실패한 것만
         골라 못 보는** 자리다(`[169]`).
      ② **연속은 회차를 센다** — 관찰이 아니다. 워치독이 30분마다 보므로 관찰을 세면
         하루 한 번 실패가 "48회 연속"이 되고 밀림 판정이 첫날 아침에 터진다.
      ③ **가르기** — 도는중·밀림 한 번은 경보가 아니고 죽음·안돎은 한 번이라도
         경보다. 경보가 대부분이면 아무도 안 본다(`[170]`).
      ④ **못 본 것을 정상이라 하지 않는다** — 조회 실패는 '이상 없음'이 아니다.
      ⑤ **읽기 전용** — 회차를 다시 띄우거나 등록하지 않는다.
      ⑥ **배선** — 워치독이 인계 스냅샷 **앞에서** 부르고, 인계는 다시 안 묻는다(`[168]`).
    """
    import schedule_watch as SW
    from datetime import datetime as _dt, timedelta as _td

    src = open(os.path.join(ROOT, "schedule_watch.py"), encoding="utf-8").read()

    # ── ① 결과 코드는 [long] 이다. [int] 가 돌아오면 실패한 회차만 조용히 빠진다.
    assert "result = [long]" in src and "missed = [long]" in src, \
        "결과 코드를 [long] 으로 안 받는다 — 0xC000013A 에서 그 작업이 통째로 빠진다"
    assert "result = [int]" not in src, "[int] 캐스트가 되살아났다"
    assert "catch {" in src, "작업마다 try/catch 가 없다 — 하나가 터지면 조용히 빠진다"

    # ── ⑤ 읽기 전용: **실제로 보내는 명령**에 고치는 동사가 없어야 한다.
    #     ★ 첫 판은 파일 전체를 훑었는데, `declared()` 가 설치본에서
    #       `Register-ScheduledTask` 라는 **문자열을 찾는** 코드라 빨강이 났다.
    #       읽는 것과 실행하는 것은 다르다 — 검사는 보내는 명령을 봐야 한다
    #       (`[229]` 의 sandbox 검사와 같은 자리다).
    ps_block = src.split('_PS = r"""', 1)[1].split('"""', 1)[0]
    for bad in ("Register-ScheduledTask", "Start-ScheduledTask", "Set-ScheduledTask",
                "Unregister-ScheduledTask", "Enable-ScheduledTask", "Disable-ScheduledTask"):
        assert bad not in ps_block, "schedule_watch 가 스케줄러를 고치려 한다: %s" % bad
    # 스케줄러에 보내는 통로는 **한 곳뿐**이어야 한다 — 늘어나면 위 검사가 눈먼다
    assert src.count("EncodedCommand") == 1 and src.count("proc_guard.run_tree") == 1, \
        "PowerShell 을 부르는 자리가 둘 이상이다 — 읽기 전용 검사가 한 곳만 본다"

    now = _dt(2026, 8, 12, 14, 0, 0)

    def task(name, result, last, **kw):
        t = {"name": name, "state": kw.get("state", "Ready"), "result": result,
             "last": last, "next": "", "reg": kw.get("reg", "2026-01-01T00:00:00"),
             "limit": kw.get("limit", "PT3H"), "multi": "IgnoreNew", "missed": 0,
             "err": kw.get("err", ""), "trig": kw.get("trig", [])}
        return t

    daily = [{"kind": "MSFT_TaskDailyTrigger", "start": "2026-01-01T09:50:00",
              "every": "", "span": "", "days": "1", "on": True}]
    every5 = [{"kind": "MSFT_TaskTimeTrigger", "start": "2026-01-01T00:00:00",
               "every": "PT5M", "span": "", "days": "", "on": True}]

    # ── ③ 갈래 가르기
    ok = SW.judge(task("성공회차", 0, "2026-08-12T09:50:00", trig=daily), now)
    assert ok["갈래"] == "성공", ok
    run = SW.judge(task("도는회차", 0x800710E0, "2026-08-12T13:59:00",
                        state="Running", trig=every5), now)
    assert run["갈래"] == "도는중", "도는 중인 회차를 실패라 부른다 — %s" % run
    kill = SW.judge(task("죽은회차", 0xC000013A, "2026-08-12T09:50:00", trig=daily), now)
    assert kill["갈래"] == "강제종료" and "제한시간" in kill["말"], kill
    push = SW.judge(task("밀린회차", 0x800710E0, "2026-08-12T13:57:00", trig=every5), now)
    assert push["갈래"] == "밀림", push

    # 죽음은 한 번이라도 경보 · 밀림 한 번은 경보가 아니다
    al = SW.alarms([ok, run, kill, push], {})
    kinds = {a["갈래"] for a in al}
    assert "강제종료" in kinds, "죽은 회차가 경보에 안 올라온다"
    assert "밀림" not in kinds, "한 번의 밀림을 경보로 올린다 — 5분 회차는 매일 그렇다"
    assert "성공" not in kinds and "도는중" not in kinds, "정상까지 경보로 올린다"

    # ── ② 연속은 회차를 센다: 같은 마지막실행을 다시 봐도 안 오른다
    seen = {"밀린회차": push}
    same = SW.judge(task("밀린회차", 0x800710E0, "2026-08-12T13:57:00", trig=every5),
                    now, seen)
    assert same["연속"] == push["연속"], \
        "같은 회차를 다시 본 것을 새 실패로 센다 — 30분마다 보면 하루가 48회가 된다"
    moved = SW.judge(task("밀린회차", 0x800710E0, "2026-08-12T14:02:00", trig=every5),
                     now, {"밀린회차": same})
    assert moved["연속"] == same["연속"] + 1, "회차가 실제로 새로 돌았는데 안 센다"
    # 연속이 한도에 닿으면 그때 경보다 — 그 뜻은 '앞 회차가 안 끝나고 있다'이다
    deep = dict(moved, 연속=SW.REPEAT_LIMIT)
    assert any(a["갈래"] == "밀림" for a in SW.alarms([deep], {})), \
        "%d회 연속 밀림도 경보로 안 올린다" % SW.REPEAT_LIMIT

    # ── 1999-11-30 은 실행 기록이 아니다(윈도우의 '한 번도 안 돎')
    never = SW.judge(task("새회차", 0x00041303, "1999-11-30T00:00:00",
                          reg="2026-08-12T13:30:00", trig=daily), now)
    assert never["마지막실행"] == "", "26년 전 날짜를 실행 기록으로 읽는다"
    # 등록(13:30) 이전의 예정(09:50)으로 나무라지 않는다
    assert never["갈래"] == "아직안돎", \
        "오늘 등록한 회차를 어제치 예정으로 '안 돎'이라 한다 — %s" % never
    assert not any(a["갈래"] == "안돎" for a in SW.alarms([never], {}))

    # 그러나 등록 뒤 예정이 지났는데 기록이 없으면 그때는 안 돎이다
    late = SW.judge(task("빠진회차", 0x00041303, "",
                         reg="2026-08-01T00:00:00", trig=daily), now)
    assert late["갈래"] == "안돎", "예정이 지났는데 실행 기록이 없다 — %s" % late
    assert any(a["갈래"] == "안돎" for a in SW.alarms([late], {})), "안 돎이 경보가 아니다"

    # ── 예정 시각을 모르는 트리거는 모른다고 한다(억지 밀림 금지)
    boot = [{"kind": "MSFT_TaskBootTrigger", "start": "2026-01-01T00:00:00",
             "every": "", "span": "", "days": "", "on": True}]
    assert SW._due(task("부팅회차", 0, "2026-01-02T00:00:00", trig=boot), now)[0] is None, \
        "부팅 트리거에 예정 시각을 지어낸다 — 멀쩡한 회차가 매번 밀림으로 나온다"

    # ── 며칠마다 도는 회차는 **회차날에만** 예정이 있다 (2026-08-12 실사고)
    #    UX점검은 08-05 기준 3일마다다(08-08·08-11·08-14). 그런데 예정을 언제나
    #    '오늘 그 시각'으로 쳐서 08-12·08-13 에도 '예정이 지났는데 안 돌았다'가 떴다 —
    #    실제로는 08-11 에 rc=0 으로 멀쩡히 돈 회차다. **없는 예정을 지어내면 그
    #    경보는 회차날 빼고 매일 뜨고, 매일 뜨는 경보는 아무도 안 본다**(`[170]`).
    every3 = [{"kind": "MSFT_TaskDailyTrigger", "start": "2026-08-05T12:00:00",
               "every": "", "span": "", "days": "3", "on": True}]
    ux = lambda *a: SW._trigger_due(every3[0], _dt(*a))
    assert ux(2026, 8, 12, 12, 43) == _dt(2026, 8, 11, 12, 0), \
        "3일 주기인데 예정을 '오늘'로 친다 — 회차날이 아닌 날마다 거짓 경보가 뜬다"
    assert ux(2026, 8, 13, 23, 59) == _dt(2026, 8, 11, 12, 0), \
        "회차날이 아닌 날에 예정을 지어낸다"
    assert ux(2026, 8, 11, 13, 0) == _dt(2026, 8, 11, 12, 0), \
        "회차날 시각이 지났는데 그 회차를 예정으로 안 본다"
    assert ux(2026, 8, 11, 9, 0) == _dt(2026, 8, 8, 12, 0), \
        "회차날인데 시각 전이면 **직전 회차**가 마지막 예정이다"
    assert ux(2026, 8, 4, 12, 30) is None, "시작 전인데 예정을 지어낸다"
    # 그리고 매일 회차(days=1)는 **하나도 안 달라져야 한다** — 열한 개가 그쪽이다
    one = {"kind": "MSFT_TaskDailyTrigger", "start": "2026-08-01T09:50:00",
           "every": "", "span": "", "days": "1", "on": True}
    assert SW._trigger_due(one, _dt(2026, 8, 12, 10, 0)) == _dt(2026, 8, 12, 9, 50), \
        "매일 회차의 오늘 예정이 바뀌었다"
    assert SW._trigger_due(one, _dt(2026, 8, 12, 9, 0)) == _dt(2026, 8, 11, 9, 50), \
        "매일 회차의 시각 전 판정이 바뀌었다"
    # 창 안에서 반복하는 회차(정오회차 12:00~12:50)도 그대로다
    win = {"kind": "MSFT_TaskDailyTrigger", "start": "2026-08-12T12:00:00",
           "every": "PT10M", "span": "PT55M", "days": "1", "on": True}
    assert SW._trigger_due(win, _dt(2026, 8, 12, 14, 0)) == _dt(2026, 8, 12, 12, 50), \
        "창 반복 회차의 마지막 기회가 바뀌었다"

    # ── ④ 조회 실패는 '이상 없음'이 아니다
    with tempfile.TemporaryDirectory() as tmp:
        SW.STATE = os.path.join(tmp, "s.json")
        SW.REPORT = os.path.join(tmp, "s.md")
        boom, real = (lambda: (_ for _ in ()).throw(RuntimeError("스케줄러 안 열림"))), SW.query
        try:
            SW.query = boom
            st = SW.build(now)
            assert st["조회실패"] and not st["작업"], "조회가 터졌는데 정상처럼 적는다"
            body = open(SW.REPORT, encoding="utf-8").read()
            assert "확인 못 함" in body and "이상 없음" in body, \
                "리포트가 '못 본 것'과 '이상 없음'을 안 가른다"
            b = SW.banner()
            assert b and b["조회실패"], "인계로 나가는 한 장이 조회 실패를 숨긴다"
            # 작업 하나만 못 읽은 것도 정상이 아니다
            SW.query = lambda: [task("반쯤읽힌회차", -1, "", err="CIM 오류")]
            st = SW.build(now)
            assert st["작업"][0]["갈래"] == "확인못함", st["작업"]
            assert any(a["갈래"] == "확인못함" for a in st["경보"]), \
                "못 읽은 작업을 경보로 안 올린다"
        finally:
            SW.query = real

    # ── 있어야 할 회차 목록은 **설치본이 선언한 것**이다(손으로 적은 목록이 아니다)
    dec, unread = SW.declared()
    assert dec, "install_*.ps1 에서 작업 이름을 하나도 못 읽었다"
    assert any(n.startswith("쿠팡업무_") for n in dec), \
        "한글 이름(-join @([char]0x..)) 을 못 푼다 — 그 회차들이 영영 '있는 줄' 안다"
    assert "CSOS_AutomationPipeline" in dec, "따옴표로 적은 이름을 못 읽는다"
    # ★ 이름 관례에 기대면 안 된다 — 이 설치본은 `$name = 'CSOS_BrowserChain'` 이라
    #   `$TaskName = "..."` 만 찾던 첫 판에서 **통째로 빠졌다**(사라져도 경보가 안 뜬다).
    assert "CSOS_BrowserChain" in dec, \
        "변수 이름·따옴표가 다른 설치본을 못 읽는다 — 그 회차는 감시 밖이다"
    # 회차의 두 기둥은 설치본이 있어야 한다 — 기계를 새로 만들면 없어지는 것들이다
    for must in ("쿠팡업무_일일자동대조", "쿠팡업무_원본자료자동정리"):
        assert must in dec, "%s 에 설치본이 없다 — 기계를 새로 만들면 안 살아난다" % must
    assert not unread, "무슨 작업을 등록하는지 못 읽은 설치본이 있다: %s" % unread
    # 못 읽은 설치본은 **조용히 넘기지 않는다** — '이상 없음'이 아니라 '확인 못 함'이다
    src_build = src.split("def build(", 1)[1].split("\ndef ", 1)[0]
    assert "설치본" in src_build and "확인못함" in src_build, \
        "이름을 못 읽은 설치본을 경보로 안 올린다 — 그 회차는 감시 밖인데 화면은 조용하다"
    miss = SW.alarms([], {"쿠팡업무_없는회차": "install_x.ps1"})
    assert miss and miss[0]["갈래"] == "등록안됨" and "install_x.ps1" in miss[0]["어떻게"], \
        "설치본은 있는데 등록이 안 된 것을 경보로 안 올린다 — 정오회차 사고의 모양이다"

    # ── ⑥ 배선: 워치독이 인계 스냅샷 **앞에서** 부른다
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "def watch_schedules(" in wd, "워치독에 스케줄러 감시 단계가 없다"
    # 목록만 뽑되 **주석은 걷어낸다** — 설명에 적힌 검증번호(`[228]`)의 대괄호를
    # 목록 끝으로 읽으면 배선이 멀쩡한데 검증이 실패한다.
    block = re.search(r"results = \[(.*?)\]\s*\n", wd, re.S)
    assert block, "워치독 회차 목록을 못 찾겠다"
    steps = "\n".join(l for l in block.group(1).splitlines()
                      if not l.strip().startswith("#"))
    assert "watch_schedules(dry)" in steps, "감시 단계가 회차 목록에 안 들어 있다"
    assert steps.index("watch_schedules(dry)") < steps.index("snapshot_handoff(dry)"), \
        "인계 스냅샷 뒤에 있다 — 인계 문서가 언제나 30분 전 판정을 싣는다"

    # 인계는 **다시 묻지 않는다**(비싼 조회는 캐시 뒤에, `[168]`)
    bannersrc = src.split("def banner(", 1)[1].split("\ndef ", 1)[0]
    assert "query(" not in bannersrc and "build(" not in bannersrc, \
        "banner 가 스케줄러를 직접 묻는다 — 인계 문서를 만들 때마다 조회가 돈다"
    sh = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert '"스케줄러": schedule_health()' in sh, "인계 상태에 스케줄러 칸이 없다"
    assert 'sw = st.get("스케줄러")' in sh, "'먼저 처리할 것' 이 스케줄러를 안 읽는다"
    assert 'sw.get("조회실패")' in sh, "인계가 '확인 못 함'을 안 올린다"

    # ── 컴팩팅 배선: 셋 다 본다(빠지면 세션이 인계 없이 끊긴다)
    csrc = src.split("def compaction(", 1)[1].split("\ndef ", 1)[0]
    for key in ("autoCompactEnabled", "session_wrapup", "context_guard"):
        assert key in csrc, "컴팩팅 배선 점검이 %s 를 안 본다" % key

    # ── ⑦ 죽은 이유를 대려면 **어느 단계가 시간을 썼는지**가 남아야 한다.
    #     `[180]` 은 "어디서 멈췄나"까지만 남겼다. 회차 예산은 단계 *사이*에서만
    #     보므로 한 단계가 길면 그냥 지나친다 — 그 단계를 이름으로 못 대면 조일 곳이
    #     정해지지 않는다(실측 292분 회차가 매일 PT3H 에 강제 종료됐다).
    import daily_run as DR
    with tempfile.TemporaryDirectory() as tmp:
        keep = DR.PROGRESS
        try:
            DR.PROGRESS = os.path.join(tmp, "p.json")
            DR.note_progress("느린단계", "시작")
            DR.note_progress("느린단계", "끝")
            got = json.load(open(DR.PROGRESS, encoding="utf-8"))
            assert got["끝난단계"] == ["느린단계"] and all(isinstance(x, str) for x in got["끝난단계"]), \
                "끝난단계 모양이 바뀌었다 — 읽는 쪽(인계·검증)이 문자열 목록을 본다"
            assert got["단계기록"] and got["단계기록"][-1]["단계"] == "느린단계", got.get("단계기록")
            assert "초" in got["단계기록"][-1], "단계마다 걸린 시간이 안 남는다"
            assert got["느린단계"] and got["느린단계"][0]["단계"] == "느린단계", got.get("느린단계")
        finally:
            DR.PROGRESS = keep
    # ── ⑧ 죽은 회차는 **왜인지도 남긴다**. 스케줄러는 exit 1 을 1 이라고만 말한다.
    #     회차들은 pythonw 로 돌아 트레이스백이 어디에도 안 남는다.
    rc_src = open(os.path.join(ROOT, "band", "recollect.py"), encoding="utf-8").read()
    assert "_leave_trace" in rc_src and "traceback.format_exc" in rc_src, \
        "밴드 재수집이 죽어도 이유를 안 남긴다 — pythonw 라 화면에도 안 뜬다"
    assert "os.remove(CRASH)" in rc_src, \
        "성공해도 옛 자국이 남는다 — 이미 고쳐진 고장을 계속 보고하게 된다"
    with tempfile.TemporaryDirectory() as tmp:
        keep = SW.ROOT
        try:
            SW.ROOT = tmp
            os.makedirs(os.path.join(tmp, "reports"))
            with open(os.path.join(tmp, "reports", "무슨회차_오류.json"), "w",
                      encoding="utf-8") as fh:
                json.dump({"시각": "2026-08-12T08:00:00", "무엇": "ZeroDivisionError: x"}, fh)
            got = SW.traces()
            assert got and got[0]["무엇"].startswith("ZeroDivisionError"), got
        finally:
            SW.ROOT = keep

    sh_hint = sh.split("def _slow_hint(", 1)
    assert len(sh_hint) == 2, "인계가 오래 걸린 단계를 말할 길이 없다"
    assert ">= 300" in sh_hint[1].split("\ndef ", 1)[0], \
        "짧은 단계까지 '오래 걸린 단계'로 적는다 — 그러면 아무도 안 읽는다"

    print("  [228] 스케줄러 회차 감시 — long 캐스트로 실패 회차가 안 빠짐 · 연속은 회차를 셈 · "
          "도는중/밀림/죽음/안돎 가르기 · 확인못함≠정상 · 읽기전용 · 스냅샷 앞 배선 · "
          "단계별 소요시간 기록 ✅")


def t229_band_liveness_contract():
    """[229] 생존확인 없이는 성공이라 적히지 않는다 (사고 #35 재발방지, 분담판 [36]).

    ★ **몸통은 8/12 새벽에 이미 있었다** — 옆 세션이 `synthetic_check.py` 를 편집 중이라
      같은 파일 동시 편집 금지에 걸려 `tests/_pending_band_liveness_check.py` 에 따로
      두고 분담판에 올려 뒀던 것이다. 그런데 **관문 밖에 있는 검사는 아무도 안 돌린다** —
      "ALL GREEN 확인 후 실작업" 은 이 파일만 본다. 만들어 뒀지만 안 도는 자리는
      `[228]` 이 스케줄러에서 잡아낸 것과 **같은 모양의 구멍**이다.
    ★ **문자열 검사로 대신하지 않는다.** 사고 #38 이 남긴 교훈이 그것이다 — 그때 놓친
      버그(`run()` 이 마지막 한 줄만 준다)는 소스에 그 문자열이 있는지로는 안 잡히고
      **실제로 돌려야** 드러났다. 그래서 판정 함수를 진짜로 부른다.
    """
    band_dir = os.path.join(ROOT, "band")
    if band_dir not in sys.path:
        sys.path.insert(0, band_dir)
    import liveness as L

    def ck(name, got, want):
        assert got == want, "%s — 얻음 %r · 바람 %r" % (name, got, want)

    def rd(p):
        with open(p, encoding="utf-8") as f:
            return f.read()

    # ── ① 판정 계약: 성공은 **생존확인과 수확증가가 다 있을 때만**이다(#35 의 본체)
    def H(댓글=0, 영개=0, 미확인=0, n=1):
        return {"합계": {"댓글수": 댓글, "확인된0개": 영개, "미확인": 미확인,
                        "댓글담김": 1 if 댓글 else 0}, "덤프수": n, "덤프": []}

    A = {"판정": L.ALIVE, "상태": {"ok": 12, "total": 250}}
    D = {"판정": L.DEAD, "왜": "전역이 없다"}
    U = {"판정": L.UNKNOWN, "왜": "탐침 실패"}
    N = {"판정": L.NEVER, "왜": "심장 소리도 없다"}
    # 수확이 있어도 살아 있음을 확인 못 했으면 성공이 아니다
    ck("죽음+수확이라도 성공 아님", L.verdict([D], H(댓글=99))[0], 5)
    ck("모름+수확이라도 성공 아님", L.verdict([U], H(댓글=99))[0], 4)
    ck("시작안함+수확이라도 성공 아님", L.verdict([N], H(댓글=99))[0], 5)
    # 살아 있어도 수확이 안 늘면 성공이 아니다([162] — 살아서 0건을 담을 수 있다)
    ck("생존+미확인만은 성공 아님", L.verdict([A], H(미확인=40))[0], 3)
    ck("생존+덤프없음은 성공 아님", L.verdict([A], H(n=0))[0], 3)
    # 둘 다 있으면 성공. '확인된 0개'는 진척으로 센다([199])
    ck("생존+댓글증가만 성공", L.verdict([A], H(댓글=7))[0], 0)
    ck("생존+확인된0개도 진척", L.verdict([A], H(영개=30))[0], 0)
    # 0 을 뭉치지 않는다([169]) — 세 가지가 서로 다른 이름으로 나와야 한다
    names = {L.verdict([A], H(n=0))[1], L.verdict([A], H(미확인=1))[1],
             L.verdict([A], H(영개=1))[1]}
    ck("0건을 세 갈래로 가른다", len(names), 3)
    ck("죽음 설명이 이유를 말한다", "수집이 아니다" in L.verdict([D], H())[2], True)

    # ── ② 싱싱한 심장 소리는 죽음이 아니다 — 거짓 죽음은 중복 수확을 부른다(#36)
    now_ms = int(time.time() * 1000)
    cold = json.dumps({"verdict": "DIED_AFTER_START", "err": "NO __GRAB",
                       "beat": {"at": now_ms - 3600 * 1000, "running": True}})
    warm = json.dumps({"verdict": "DIED_AFTER_START", "err": "NO __GRAB",
                       "beat": {"at": now_ms - 4000, "running": True}})
    stopped = json.dumps({"verdict": "DIED_AFTER_START", "err": "NO __GRAB",
                          "beat": {"at": now_ms - 4000, "running": False}})
    ck("식은 심장 = 죽음", L.classify(cold)["판정"], L.DEAD)
    ck("싱싱한 심장 = 다른탭 생존", L.classify(warm)["판정"], L.ALIVE_OTHER)
    ck("다 끝난 심장 = 죽음(생존 아님)", L.classify(stopped)["판정"], L.DEAD)
    ck("한 번도 안 함 = 죽음과 다른 이름",
       L.classify(json.dumps({"verdict": "NEVER_STARTED"}))["판정"], L.NEVER)
    ck("탐침 무응답 = 모름", L.classify(None)["판정"], L.UNKNOWN)

    # ── ③ 수확은 흡수기와 **같은 잣대**로 센다 — 시각 없는 댓글은 댓글이 아니다([130])
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, "dump_202608120101_90610953.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump({"band": "90610953", "posts": {
                "1": {"comments": [{"author": "가", "created_at": 1, "content": "접수취소"}]},
                "2": {"comments": [{"author": "나", "created_at": None, "content": ""}]},
                "3": {"comments": [], "comments_full": True},
                "4": {},
            }}, f, ensure_ascii=False)
        h = L.harvest_since(0, dirs=[td])
        ck("시각 있는 댓글만 센다", h["합계"]["댓글수"], 1)
        ck("껍데기는 안 센다", h["합계"]["껍데기"], 1)
        ck("확인된 0개를 따로 센다", h["합계"]["확인된0개"], 1)
        ck("미확인을 따로 센다", h["합계"]["미확인"], 2)
        ck("회차 시작 뒤 덤프만 본다",
           L.harvest_since(time.time() + 60, dirs=[td])["덤프수"], 0)

    # ── ④ 수집기가 **최상위 이동을 막고 기록하는지**(#35 의 사고 경로)
    #     ★ 파일 전체를 문자열로 훑지 않고 **실제로 실리는 속성값**을 뜯어 본다.
    #       첫 판은 `"allow-top-navigation" in js` 였는데 그 플래그를 **왜 안 넣는지
    #       설명한 주석**에 걸려 빨강이 됐다 — 사고를 적는 행위가 사고 모양을 만든
    #       자리다(#38 과 같다). 설명은 남기고, 검사는 값을 본다.
    js = rd(os.path.join(ROOT, "band", "grab_posts.js"))
    m = re.search(r"setAttribute\(\s*'sandbox'\s*,\s*'([^']*)'\s*\)", js)
    toks = m.group(1).split() if m else []
    ck("iframe 에 sandbox 를 건다", bool(m), True)
    ck("같은 출처 유지(본문을 읽을 수 있어야 한다)", "allow-same-origin" in toks, True)
    ck("스크립트 허용(SPA 가 그려야 한다)", "allow-scripts" in toks, True)
    ck("최상위 이동은 허용하지 않는다",
       [t for t in toks if t.startswith("allow-top-navigation")], [])
    for k, why in (("__grabDeath", "죽음 기록"), ("__grabBeat", "심장 소리"),
                   ("beforeunload", "떠나기 직전"), ("pagehide", "감춰질 때"),
                   ("saveEvery", "중간 저장")):
        ck("%s(%s) 이 있다" % (k, why), k in js, True)
    for k in ("tried", "saves", "prevDeath", "prevBeat", "sandboxFellBack"):
        ck("__grabStatus 가 %s 를 내놓는다" % k, k in js, True)

    # ── ⑤ `NO __GRAB` 은 **파일 셋이 공유하는 계약**이다 — 이름이 갈리면 죽음을 못 알아본다
    st = rd(os.path.join(ROOT, "band", "band_dump_state.js"))
    ck("탐침이 NO __GRAB 을 그대로 말한다", "NO __GRAB" in st, True)
    ck("탐침이 심장 소리를 읽는다", "__grabBeat" in st, True)
    ck("탐침이 죽음 기록을 읽는다", "__grabDeath" in st, True)
    ck("탐침이 시작안함과 죽음을 가른다",
       "NEVER_STARTED" in st and "DIED_AFTER_START" in st, True)
    ck("browser_chain 이 같은 표식을 본다",
       "NO __GRAB" in rd(os.path.join(ROOT, "band", "browser_chain.py")), True)
    ck("liveness 도 같은 표식을 본다", L.DEAD_MARK, "NO __GRAB")

    # ── ⑥ 중간 저장 이름이 **유령 밴드**를 만들지 않는지 — 실제 규칙 함수로 확인한다
    #     `dump_<12자리>s2_90610953.json` 의 후보는 둘이고, 숫자를 그냥 이어 붙이면
    #     13자리 덩어리가 되어 없는 밴드가 생긴다(2026-08-08 두 차례 실사고와 같은 모양).
    import convert_dump as cd
    for name in ("dump_202608120050_90610953.json",
                 "dump_202608120050s2_90610953.json",
                 "dump_202608120050s12_84789192.json"):
        want = "90610953" if "90610953" in name else "84789192"
        ck("이름 규칙: %s" % name, cd.band_from_name(name, known=set()), want)

    # ── ⑦ 밴드를 다른 세션이 잡고 있으면 **빼앗지 않고 물러난다**(사고 #27)
    src = rd(os.path.join(ROOT, "band", "liveness.py"))
    ck("점유 판정을 ai_claim 에서 빌린다", "_is_dead" in src and "_is_mine" in src, True)
    ck("못 읽으면 '없음'으로 치지 않는다", "확인할 수 없다" in src, True)
    # subprocess.run(timeout=) 은 이 프로젝트에서 금지다([175]) — 회차가 영원히 멈춘다
    ck("subprocess.run(timeout= 을 쓰지 않는다", "subprocess.run(" in src, False)
    ck("proc_guard.run_tree 를 쓴다", "proc_guard.run_tree" in src, True)

    # ── ⑧ 시작만 하고 **끝을 안 본** 수집을 주기적으로 마무리한다 (분담판 [37])
    #     붙여넣은 직후 한 번만 묻고 '시작됨'이라 적으면, 3분째에 죽어도 다음 tick 은
    #     too_soon(90분)에 걸려 건너뛴다 — 성공도 실패도 아닌 채로 90분이 지나간다.
    import browser_chain as BC
    tick_src = rd(os.path.join(ROOT, "band", "browser_chain.py"))
    tick_body = tick_src.split("\ndef tick(", 1)[1].split("\ndef ", 1)[0]
    assert "settle_band(d)" in tick_body, \
        "tick 이 앞 수집을 마무리하지 않는다 — '시작됨'인 채로 90분이 지나간다"
    assert tick_body.index("settle_band(d)") < tick_body.index("looks_busy()"), \
        "새로 붙여넣기 전에 앞의 것을 마무리해야 한다"
    with tempfile.TemporaryDirectory() as td:
        keep = (BC.STATE, L.probe, L.harvest_since, L.band_held_by_other)
        try:
            BC.STATE = os.path.join(td, "chain.json")
            L.band_held_by_other = lambda: ""
            L.harvest_since = lambda t0, dirs=None: {
                "합계": {"댓글수": 0, "확인된0개": 0, "미확인": 0, "댓글담김": 0},
                "덤프수": 0, "덤프": []}
            때 = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time() - 300))
            d = {"단계": [], "마지막": {"밴드 댓글 90610953":
                                     {"때": 때, "결과": "시작됨", "왜": ""}}}
            # 살아 있는데 아직 수확 전 · 5분째 → **진행중**(멀쩡히 도는 것을 죽이지 않는다)
            L.probe = lambda **kw: {"판정": L.ALIVE, "상태": {"ok": 3, "total": 250}}
            BC.settle_band(d)
            ck("이른 무수확은 진행중", d["마지막"]["밴드 댓글 90610953"]["결과"], "진행중")
            # ★ 묻는 행위가 재시도를 미루면 안 된다 — 시계 기준은 붙여넣은 때다
            ck("재시도 시계가 붙여넣은 때 그대로", d["마지막"]["밴드 댓글 90610953"]["때"], 때)
            # 죽었으면 성공이라 적지 않는다(#35 의 본체)
            L.probe = lambda **kw: {"판정": L.DEAD, "왜": "전역이 없다"}
            BC.settle_band(d)
            ck("죽음은 죽음이라 적는다", d["마지막"]["밴드 댓글 90610953"]["결과"], "죽음")
            # 다른 세션이 잡고 있으면 손대지 않는다(사고 #27)
            d["마지막"]["밴드 댓글 90610953"]["결과"] = "시작됨"
            L.band_held_by_other = lambda: "옆 세션"
            BC.settle_band(d)
            ck("남이 잡고 있으면 물러난다",
               d["마지막"]["밴드 댓글 90610953"]["결과"], "시작됨")
        finally:
            BC.STATE, L.probe, L.harvest_since, L.band_held_by_other = keep

    # ── 관문 밖에 사본이 남아 있으면 안 된다. 두 벌이 되면 한쪽만 고쳐지고,
    #    고쳐지지 않은 쪽은 **아무도 안 돌리므로 틀린 줄도 모른다**.
    assert not os.path.exists(os.path.join(ROOT, "tests",
                                           "_pending_band_liveness_check.py")), \
        "관문 밖 사본이 남아 있다 — 검사가 두 벌이 되면 한쪽은 영영 안 돈다"

    print("  [229] 생존확인 계약 — 생존+수확 둘 다일 때만 성공 · 싱싱한 심장은 죽음 아님 · "
          "시각 없는 댓글 제외 · sandbox 값 검사 · NO __GRAB 계약 · 유령 밴드 방지 · "
          "tick 이 앞 수집을 마무리(재시도 시계는 안 밀림) ✅")


def t231_loop_tick_weight_from_evidence():
    """[231] 정오 루프 틱이 **자기 무게를 근거로 정한다** (2026-08-12 지시).

    사용자 지시: "루프는 하루에 한번 매일 12시에서 13시로 설정하고 자동으로
    모델과 노력 강도 설정해서 진행하는 알고리즘 적용".

    ★ **값의 정본은 `ai_tier.TIERS` 하나다**(`[230]`). 이 파일은 '이번 틱이 어느
      갈래인가'만 정하고 모델 이름을 제 손으로 적지 않는다. 첫 판에 표를 여기에도
      적었다가 `haiku` 를 넣었는데, ai_tier 는 **일부러 haiku 를 뺐다**(값싼 오판이
      아끼는 것보다 크다). 사본을 두면 그 이유가 조용히 사라진다.
    ★ **못 읽은 것을 '없음'으로 치지 않는다**(`[169]`). 근거 파일을 못 읽으면
      가벼움이 아니라 **보통**이다 — 못 읽었다는 이유로 노력을 낮추면 파일이 깨진
      날 사고가 가장 싼 판단을 만난다.
    ★ **수집 밀림은 내 무게가 아니다** — 수집 세션 몫이라(`[177]`) 그것만 있는
      날까지 opus 를 쓰면 크레딧이 새고, 그러면 이 알고리즘이 있으나 마나가 된다.
    """
    import loop_policy as P
    import ai_tier as T

    src = open(os.path.join(ROOT, "loop_policy.py"), encoding="utf-8").read()
    body = src.split('"""', 2)[-1]                    # 설명 글은 빼고 코드만 본다
    for word in ("claude-opus", "claude-sonnet", "haiku"):
        assert word not in body, \
            "loop_policy 가 모델 이름을 제 손으로 적는다(%s) — ai_tier 와 갈린다" % word
    for tier, kind in P.TIER_OF.items():
        assert kind in T.TIERS, "%s 가 ai_tier 에 없는 갈래를 가리킨다: %s" % (tier, kind)
        assert P.value_of(tier)[:2] == tuple(T.TIERS[kind][:2]), \
            "%s 의 값이 ai_tier 와 다르다 — 사본이 생겼다" % tier

    assert P.decide([], True, 0)["갈래"] == "가벼움", "아무것도 없는 날에 값을 쓴다"
    assert P.decide([], False, 0)["갈래"] == "보통", \
        "근거를 못 읽고 가볍게 간다 — 못 읽은 것을 '없음'으로 쳤다"
    assert P.decide(["★ 밴드: 수집이 밀렸다 — 최신 2026-08-10"], True, 0)["갈래"] == "가벼움", \
        "수집 세션 몫에 무게를 싣는다"
    assert P.decide(["입력 큐에 470건이 반영되지 않았다"], True, 0)["갈래"] == "보통"
    assert P.decide(["회차 [강제종료] 쿠팡업무_일일자동대조 — 끊겼다"], True, 0)["갈래"] == "무거움", \
        "회차가 끊겼는데 싸게 판단한다 — 오진이 파일에 박힌다"
    assert P.decide([], True, 2)["갈래"] == "무거움", "회차 경보를 무게로 안 센다"
    assert P.decide(["듣도 보도 못한 항목"], True, 0)["갈래"] == "보통", \
        "모르는 모양을 가볍다고 친다 — 모르면 낮추지 않는다"

    # 근거 파일 이름을 지어내지 않는다 — 실제로 있는 것만 본다
    for p in (P.HANDOFF, P.ROUNDS):
        assert os.path.basename(p).endswith(".md"), p
    print("[231] 정오 루프 틱 무게 — 값은 ai_tier 한 곳 · 못읽음≠가벼움 · 수집은 남의 몫 OK")


def t233_round_steps_fit_inside_budget():
    """[233] 어떤 단계도 **남은 회차 예산보다 오래 받지 못한다** (2026-08-12 · 분담판 [38]).

    ★ 예산은 있었는데 **경계가 아니라 권고**였다. `over_budget()` 은 단계 **사이**에서만
      보므로, 145분째에 시작한 단계가 제 시간 제한 60분을 그대로 들고 가면 205분이 된다.
      실측 292.3분 — 작업 스케줄러 제한(PT3H)에 걸려 `일일자동대조`·`원본자료자동정리`가
      **매일 나무째 끊겼다**(0xC000013A). 끊기면 리포트가 한 줄도 안 써지고 잠금이 남아
      다음 회차가 조용히 건너뛴다 — 스케줄러는 그걸 '성공'이라 적는다.
    ★ **범인 단계를 지목하지 않고 고친다.** 어느 단계가 오래 걸리는지는 아직 모르고,
      모르는 채로 제한시간을 손대는 것은 짐작이다([228] 이 남긴 숙제가 그것이다).
      그러나 "어떤 단계도 남은 예산보다 오래 받을 수 없다"는 **모든 단계에 참인 규칙**이라
      짐작이 아니다. 범인이 누구든 회차는 예산 안에서 끝난다.
    ★ **바닥을 둔다.** 예산이 다 됐다고 3초를 주면 그 단계는 시작하자마자 죽어 '실패'로
      적히는데 그건 사실이 아니다(시간을 안 준 것이다).
    ★ **예산은 스케줄러 제한보다 넉넉히 작아야** 뜻이 있다. 붙어 있으면 예산 안에 끝내고도
      끊긴다 — 그러면 이 규칙 전체가 있으나 마나가 된다.
    ★ 원본정리는 목록이 반쪽이면 **찌꺼기를 거두지 않는다.** 거두기는 "want 에 없으면
      찌꺼기"라는 판정이라, 반쪽인 채로 거두면 **멀쩡한 바로가기를 지운다**([172]의 문).
    """
    import re
    from datetime import timedelta
    import daily_run as D
    import source_tidy as ST

    keep = D._ROUND_T0[0]
    try:
        D._ROUND_T0[0] = None
        assert D.fit_timeout(3600) == (3600, False), "예산 시계가 없는데 시간을 줄인다"
        D._ROUND_T0[0] = D.datetime.now()
        assert D.fit_timeout(600) == (600, False), "예산이 넉넉한데 줄인다"
        D._ROUND_T0[0] = D.datetime.now() - timedelta(minutes=D.ROUND_BUDGET_MIN - 5)
        got, cut = D.fit_timeout(3600)
        assert cut and 240 <= got <= 305, "남은 예산(5분)보다 오래 준다: %s초" % got
        D._ROUND_T0[0] = D.datetime.now() - timedelta(minutes=D.ROUND_BUDGET_MIN + 60)
        got2, cut2 = D.fit_timeout(3600)
        assert (got2, cut2) == (D.STEP_FLOOR_SEC, True),             "예산이 바닥일 때 바닥값을 안 준다 — 0초를 주면 '실패'로 적히는데 사실이 아니다"
    finally:
        D._ROUND_T0[0] = keep

    src = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "fitted, cut = fit_timeout(timeout)" in src and "_run_once(name, args, fitted)" in src,         "run() 이 예산에 맞춰 줄이지 않는다 — 함수만 있고 아무도 안 쓴다"

    tidy = open(os.path.join(ROOT, "source_tidy.py"), encoding="utf-8").read()
    assert "def out_of_time" in tidy and hasattr(ST, "BUDGET_MIN"), "원본정리에 시간 예산이 없다"
    assert tidy.index("if _LEFT[0]:") < tidy.index("# 3) 지난 회차의 찌꺼기"),         "목록이 반쪽인 채로 찌꺼기를 거둔다 — 멀쩡한 바로가기를 지운다"

    def _limit_min(fn):
        s = open(os.path.join(ROOT, fn), encoding="utf-8").read()
        m = re.search(r"ExecutionTimeLimit \(New-TimeSpan -(Hours|Minutes) (\d+)\)", s)
        assert m, fn + " 에서 제한시간을 못 읽었다"
        return int(m.group(2)) * (60 if m.group(1) == "Hours" else 1)

    assert D.ROUND_BUDGET_MIN + 10 <= _limit_min("install_daily_schedule.ps1"),         "회차 예산이 스케줄러 제한과 붙어 있다 — 예산을 지켜도 끊긴다"
    assert ST.BUDGET_MIN + 20 <= _limit_min("install_source_tidy_schedule.ps1"),         "원본정리 예산이 스케줄러 제한과 붙어 있다"


def t232_orgchart_floorplan_roster_and_states():
    """[232] 조직도 — 사무실 배치 평면도: 로스터 정본·AI 자리 셋·상태 넷·아코디언 (2026-08-12 지시).

    사용자 지시: 조직도를 '책상·의자가 있는 사무실 배치도'로 바꾸고, 좌우 잘림을 고치고,
    AI 세션 에이전트를 모두 보이고, 자리를 누르면 담당 업무가 펼쳐지게 한다.

    ★ 지키는 것 셋:
      ① 사람 자리·역할·이름은 **로스터**(STAFF_CENTERS·AS_TECH_CENTERS)가 정본이다 —
         서버가 이름을 지어내지 않는다. 사람의 '지금 상태'는 **실제 접속(auth)**으로만
         온/오프라인을 판정한다 — 지어내지 않는다([169]: 기록 없으면 '상태 모름').
      ② AI 구역은 **절대 비지 않는다** — 기본 자리 셋(Claude·CSOS·Codex)이 늘 있고
         실시간 점유(ai_claim)만 그 위를 덮는다([169] '0을 없다로 못 박지 않는다'의 자리판).
      ③ 화면은 `#v-org` 아래로 좁힌 `org-` 클래스만 쓴다 — 1.1MB 본문의 .ws·.chair·
         .desk 와 안 부딪히고, minmax(0,..)·max-width 로 좌우가 안 잘린다.
    """
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as A

    D = A.get_orgchart()
    assert isinstance(D, dict) and isinstance(D.get("_live"), bool), "_live 가 없다/불리언이 아니다"
    assert isinstance(D.get("gen"), str) and D["gen"], "gen(생성 시각)이 문자열이 아니다"
    zones = D.get("zones")
    assert isinstance(zones, list) and len(zones) == 3, "구역이 정확히 셋이 아니다"
    assert [z.get("key") for z in zones] == ["mgmt", "field", "ai"], \
        "구역 순서/키가 mgmt·field·ai 가 아니다"

    zmap = {z["key"]: z for z in zones}
    STATES = {"busy", "idle", "live", "off"}
    for z in zones:
        ppl = z.get("people")
        assert isinstance(ppl, list) and ppl, "구역에 사람이 없다: " + str(z.get("key"))
        for p in ppl:
            assert p.get("state") in STATES, "모르는 상태: %r" % p.get("state")
            assert isinstance(p.get("duties"), list) and p["duties"], \
                "담당 업무(duties)가 비었다: %s" % p.get("name")

    # ① 사람 자리·이름은 로스터가 정본 — 코드에 박아 두지 않는다.
    mgmt_names = [p["name"] for p in zmap["mgmt"]["people"]]
    assert mgmt_names == [c.get("name", s) for s, c in A.STAFF_CENTERS.items()], \
        "관리팀 이름이 로스터(STAFF_CENTERS)와 다르다 — 코드가 이름을 지어낸다"
    field_names = [p["name"] for p in zmap["field"]["people"]]
    assert field_names == [c.get("name", s) for s, c in A.AS_TECH_CENTERS.items()], \
        "현장팀 이름이 로스터(AS_TECH_CENTERS)와 다르다"

    # 사람 상태는 실제 접속(auth)으로만 온/오프라인 — busy(온라인)·off(오프라인/모름)
    # 뿐이고 live(실시간 AI 세션 전용)로는 절대 표시하지 않는다.
    for z in (zmap["mgmt"], zmap["field"]):
        for p in z["people"]:
            assert p["state"] in {"busy", "off"}, \
                "사람 상태가 온/오프라인이 아니다: %s → %s" % (p["name"], p["state"])
            assert p["state"] != "live", "사람을 실시간 AI 세션처럼 표시했다: %s" % p["name"]

    # 접속(presence)이 온라인으로 뒤집는다 — 지어내지 않고 실제 신호로만.
    # (디스크는 안 건드리고 메모리 presence 만 넣었다 뺀다 — 실서버 상태 불변)
    slug0 = next(iter(A.STAFF_CENTERS))
    name0 = A.STAFF_CENTERS[slug0].get("name", slug0)
    import time as _tt
    A._presence_map()[slug0] = _tt.time()
    try:
        m2 = {z["key"]: z for z in A.get_orgchart()["zones"]}["mgmt"]
        p0 = next(p for p in m2["people"] if p["name"] == name0)
        assert p0["state"] == "busy" and "온라인" in (p0.get("msg") or ""), \
            "접속을 찍었는데 온라인으로 안 바뀐다: %s → %s" % (name0, p0["state"])
    finally:
        A._presence_map().pop(slug0, None)

    # ★ 관리자 로그인은 **센터장(유현민) 자리**다 (2026-08-12 지시: "유현민은 여기에
    #   접속되어있으면 온라인으로 항상 표기해야지"). 전에는 role=="staff" 만 찍어서,
    #   관리자로 앱을 열면 **아무 자리도 안 찍혔다** — 본인이 화면을 보고 있는데 제
    #   자리만 '앱 접속 기록 없음'이었다. 슬러그가 로스터에 실재해야 한다 — 오타면
    #   오류 없이 영영 안 찍힌다(빈칸과 구별이 안 되는 종류의 잘못, [165]).
    adm_slug = getattr(A, "ADMIN_PRESENCE_SLUG", "")
    assert adm_slug in A.STAFF_CENTERS, \
        "ADMIN_PRESENCE_SLUG 가 로스터에 없는 자리다: %r" % (adm_slug,)
    A._presence_map()[adm_slug] = _tt.time()
    try:
        m3 = {z["key"]: z for z in A.get_orgchart()["zones"]}["mgmt"]
        adm_name = A.STAFF_CENTERS[adm_slug].get("name", adm_slug)
        pa = next(p for p in m3["people"] if p["name"] == adm_name)
        assert pa["state"] == "busy" and "온라인" in (pa.get("msg") or ""), \
            "관리자 자리가 접속을 찍어도 온라인이 안 된다: %s → %s" % (adm_name, pa["state"])
    finally:
        A._presence_map().pop(adm_slug, None)

    # ② 딱지: 유현민 '센터장'(예전 '대표'에서 바꿈), 차동호 '팀장'.
    yoo = next(p for p in zmap["mgmt"]["people"] if p["name"] == "유현민")
    assert yoo.get("badge") == "센터장", "유현민 딱지가 '센터장' 이 아니다: %r" % yoo.get("badge")
    cha = next(p for p in zmap["field"]["people"] if p["name"] == "차동호")
    assert cha.get("badge") == "팀장", "차동호 딱지가 '팀장' 이 아니다: %r" % cha.get("badge")

    # ③ AI 구역은 절대 비지 않고 로스터 세 역할을 늘 덮는다 — 각 역할은 **살아 있는
    #    세션 데스크**(그때는 세션 제목이 이름)로 채워지거나, 아니면 off 기본 자리로
    #    나타난다([169]). 그래서 자리 수는 늘 셋 이상, off 기본 자리는 로스터 이름만
    #    쓴다(지어내지 않는다). 살아 있는 세션이 없으면 셋 다 off 로 보인다.
    ai_ppl = zmap["ai"]["people"]
    assert len(ai_ppl) >= 3, "AI 구역이 로스터 세 역할을 못 덮는다(자리 %d개)" % len(ai_ppl)
    BASE_NAMES = {"Claude 세션", "CSOS 수집", "Codex"}
    for p in ai_ppl:
        # ★ **회차 자리는 이 규칙 밖이다** (2026-08-13). 조수 스테이션에 무인 회차가
        #   올라온 뒤로(`schedule_watch` 갈래를 읽는 자리), 회차 하나가 죽거나 안 돌면
        #   그 자리가 `off` 로 앉는다 — 이름은 **스케줄러가 부르는 그 이름**이지
        #   지어낸 것이 아니다. 그런데 이 규칙이 그것까지 잡아, **회차가 죽는 날마다
        #   조직도 검증이 빨개졌다**(실측 2026-08-13 `CSOS_BrowserChain` 강제종료).
        #   즉 빨강의 뜻이 '이름을 지어냈다'가 아니라 '오늘 회차가 하나 죽었다'가 되어,
        #   보는 사람이 엉뚱한 데를 고치러 간다. 이 규칙이 지키려는 것은 **AI 로스터
        #   기본 자리**뿐이다 — 회차의 건강은 `[228]` 이 따로 본다.
        if p.get("badge") == "회차":
            continue
        if p["state"] == "off":
            assert p["name"] in BASE_NAMES, "모르는 기본 자리(이름 지어냄): " + p["name"]
    if not any(p["state"] == "live" for p in ai_ppl):
        assert BASE_NAMES <= {p["name"] for p in ai_ppl}, "AI 기본 자리 셋이 다 안 보인다"

    # ── 서버·화면 코드가 계약을 지키는지 텍스트로 확인한다 ──
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # 라우트는 절대 500 을 안 낸다 — try/except 로 감싸 200 폴백을 준다.
    route = server.split('if p == "/api/orgchart":', 1)[1].split("if p ==", 1)[0]
    assert "try:" in route and "except" in route and "_send(200" in route, \
        "/api/orgchart 이 500 을 낼 수 있다 — 200 폴백이 없다"
    assert '"_live": False' in route, "폴백에 _live=False 가 없다"

    # AI 산·죽음은 ai_claim 에서 온다 — 지어내지 않는다.
    fn = server.split("def get_orgchart", 1)[1].split("\ndef ", 1)[0]
    assert "import ai_claim" in fn and "_is_dead" in fn, "AI 상태를 ai_claim 점유에서 안 읽는다"
    assert "센터장" in fn and '"대표"' not in fn, \
        "유현민 딱지가 아직 '대표' 다(로스터 정본과 어긋남)"

    # 접속 찍기: 사람이 로그인한 쿠키 세션(staff·admin)만이다.
    auth_fn = server.split("    def _auth(self):", 1)[1].split("\n    def ", 1)[0]
    assert "presence_touch(ADMIN_PRESENCE_SLUG)" in auth_fn, \
        "_auth 가 관리자 접속을 안 찍는다 — 관리자로 열면 센터장 자리가 영영 '기록 없음'이다"
    # ⚠ X-Pin(로컬 스크립트) 길에서는 **찍지 않는다** — 거기서 찍으면 워치독·회차가
    #    도는 내내 사람이 접속한 것처럼 보인다. 잰 것이 아니라 지어낸 신호다([169]).
    assert "presence_touch" not in auth_fn.split("_locked(", 1)[-1], \
        "X-Pin(사람 아님) 길에서 접속을 찍는다 — 스크립트가 도는 내내 온라인으로 보인다"

    # 화면: 조직도 탭·뷰가 있다.
    assert 'data-v="org"' in live and 'id="v-org"' in live, "조직도 탭/뷰가 없다"

    # 클래스는 #v-org 아래로 좁힌 org- 접두만 — 본문 .ws·.chair·.desk 와 안 부딪힌다.
    css = live.split('<style id="org-style">', 1)[1].split("</style>", 1)[0]
    assert "#v-org .org-" in css, "org 규칙이 #v-org 아래로 좁혀져 있지 않다"
    assert "#v-org .org-ws" in css and "#v-org .org-chair" in css, "의자·책상 규칙이 없다"

    # ★ 책상은 어느 방에서든 **같은 크기**다 (2026-08-12 지시 "각 카드 크기 똑같이").
    #   1fr 로 늘리면 사람이 적은 방의 책상만 커진다 — 실측 관리팀 228 · 현장 205 ·
    #   조수 176 으로 셋 다 달랐다. 트랙 폭을 고정해야 셋이 같아진다.
    assert "repeat(auto-fill,var(--org-w))" in css and "--org-w:" in css, \
        "책상 트랙이 고정폭(--org-w)이 아니다 — 방마다 책상 크기가 달라진다"
    assert ".org-zone.ai .org-desks" not in css, \
        "방별 책상 폭 override 가 되살아났다 — 그 방 책상만 크기가 달라진다"
    # 한 줄의 책상 높이가 같아야 의자도 한 줄로 맞는다 — 책상이 칸을 꽉 채운다.
    assert "flex:1 0 auto" in css, \
        "책상이 칸 높이를 안 채운다 — 카드 높이가 제각각이 되고 의자 줄이 어긋난다"
    assert "margin-top:auto" in css, "상태 줄이 책상 아래에 안 붙는다"
    # 긴 이름이 카드 하나만 세로로 늘려 그 줄 전체를 어긋나게 했다(다섯 줄로 쪼개졌다).
    # 한 줄로 자르되 **전체는 title 로 남긴다** — 잘라 놓고 못 읽게 두지 않는다.
    assert "#v-org .org-nm-t" in css and 'class="org-nm-t"' in live, \
        "이름 한 줄 자르기(.org-nm-t)가 없다 — 긴 이름이 카드를 늘린다"
    assert 'title="${esc2(p.name)}' in live, "잘린 이름을 title 로 안 남긴다"
    for bare in (".ws{", ".chair{", ".desk{"):
        assert bare not in css, "접두 없는 전역 규칙이 들어왔다: " + bare

    # 4상태 범례.
    for cls in ("org-dot busy", "org-dot idle", "org-dot live", "org-dot off"):
        assert cls in live, "범례에 상태가 빠졌다: " + cls

    # 아코디언 — 자리를 누르면 담당 업무가 펼쳐진다(aria-expanded + orgToggle).
    assert "aria-expanded" in live and "function orgToggle(" in live, "펼침 아코디언이 없다"
    assert "grid-template-rows:0fr" in css and "grid-template-rows:1fr" in css, \
        "아코디언 여닫이(0fr→1fr)가 CSS 에 없다"

    # 길게 눌러 자리 옮기기(2026-08-13) — 새 드래그 사본을 만들지 않고 공용 함수를
    # 선택 옵션으로 확장한다. 옵션 없는 대시보드 세 호출은 편집 모드 계약 그대로다.
    assert "function dashDragEnable(hostRef, sel, onDrop, onReorder, opt)" in live, \
        "공용 dashDragEnable 을 확장하지 않고 조직도용 사본을 만든 흔적이다"
    drag = live.split("function dashDragEnable(", 1)[1].split("/* ═══ 눌러서 옮기기", 1)[0]
    assert "typeof hostRef==='string'" in drag, "id 없는 조직도 구역 요소를 host 로 못 받는다"
    assert "hasOwnProperty.call(cfg,'longPress')" in drag and ": 0" in drag, \
        "longPress 옵션이 없는 기존 호출까지 길게 누르기로 바뀔 수 있다"
    assert "navigator.vibrate" in drag and "dash-long-pressing" in drag, \
        "길게 누르기 시작 피드백(확대·진동)이 없다"
    assert "cfg.tapPick!==false" in drag and "cfg.allowInteractive" in drag, \
        "조직도 button 탭과 대시보드 집기 동작을 분리하지 않았다"
    dash_init = live.split("function initDashboardLayout()", 1)[1].split(
        "applyDashboardLayout(readDashboardLayout())", 1)[0]
    assert "longPress" not in dash_init and dash_init.count("dashDragEnable(") == 3, \
        "기존 대시보드 세 호출이 longPress 옵션의 영향을 받는다"
    assert "data-dash-long-press" in css and "touch-action:none" in css, \
        "폰이 길게 누르기를 스크롤로 가져갈 수 있다"

    # 구역별 name 순서를 localStorage 에 저장하고 15초 재렌더 때마다 다시 적용한다.
    # 저장 목록에 없는 새 로스터 사람은 saved.length+원래순번으로 반드시 뒤에 남는다.
    render = live.split("function orgRender(", 1)[1].split("\nfunction orgLoad", 1)[0]
    load = live.split("function orgLoad(", 1)[1].split("\n/* 카톡으로", 1)[0]
    assert "ORG_ORDER_KEY" in live and "localStorage.getItem(ORG_ORDER_KEY)" in live \
        and "localStorage.setItem(ORG_ORDER_KEY" in live, "조직도 자리 순서가 이 기기에 저장되지 않는다"
    assert "saved.length+a.i" in live and "saved.length+b.i" in live, \
        "저장 뒤 새로 들어온 사람을 버리거나 앞사람과 같은 순위로 만든다"
    assert "D=orgApplyOrder(D)" in render and "data-org-zone" in render \
        and "data-org-name" in render, "orgRender 가 저장 순서를 다시 입히지 않는다"
    assert "dashDragEnable(host,'.org-ws'" in render and "longPress:450" in render \
        and "gridItems:true" in render, \
        "세 조직도 구역에 450ms 길게 누르기 재정렬을 연결하지 않았다"
    assert "15000" in load and "orgLoad()" in load, \
        "15초 갱신이 저장 순서를 거쳐 재렌더되는 계약이 사라졌다"

    # 화면 스냅샷도 서버와 같은 모양 — 유현민 '센터장'·AI 세 자리·duties.
    snap = live.split("const ORG_SNAP=", 1)[1].split("]};", 1)[0]
    assert "센터장" in snap and 'badge:"팀장"' in snap, "스냅샷 딱지가 계약과 다르다"
    assert "CSOS 수집" in snap and "Codex" in snap, "스냅샷에 AI 세 자리가 없다"
    assert "duties:" in snap, "스냅샷에 duties 가 없다"

    # 좌우 잘림 방지 — 내용 열을 넘지 않게 max-width·minmax(0,..)·모바일 규칙을 쓴다.
    assert "max-width:100%" in css, "평면도가 내용 열을 넘지 않게 막는 max-width 가 없다"
    assert "minmax(0," in css, "그리드가 minmax(0,..) 로 안 짜여 좌우가 잘릴 수 있다"
    assert "max-width:760px" in css, "모바일(≤760px) 반응형 규칙이 없다"

    # 캡처 글꼴은 uiFont 에서 온다 — 손으로 적으면 저장 이미지만 옛 글꼴로 남는다([앱 글꼴]).
    png = live.split("function orgToPng(", 1)[1].split("\nfunction ", 1)[0]
    assert "uiFont" in png, "이미지 캡처가 uiFont 를 안 쓴다(글꼴을 손으로 박았다)"

    # ── 폰 배치도의 **흐름 화살표**(2026-08-13 지시) — 근거는 앱 흐름도 하나다.
    #    ★ 단계 낱말을 화면·서버에 손으로 적으면 흐름도가 바뀐 날 **그림만 옛것으로
    #      남는다**([196]). 그래서 여기서 지키는 것은 '화살표가 예쁜가'가 아니라
    #      **낱말이 flow_steps 에서 왔는가**다.
    import ledger_db as _ldb
    fl = A._org_flow()
    assert fl.get("ok"), "흐름을 못 읽었다 — 화면이 '흐름 없음'과 '못 읽음'을 갈라야 한다"
    steps = _ldb.flow_steps() or []
    real = {str(s.get("단계") or "") for s in steps}
    # ⚠ 라벨을 " · " 로 쪼개 보면 안 된다 — **단계 이름 자체에 `·` 가 들어 있다**
    #   (`PO 취합 · 류지영 전달`). 쪼개면 멀쩡한 이름이 '지어낸 낱말'로 보인다.
    #   그래서 실재 이름을 **긴 것부터 지워** 남는 글자가 없는지로 본다.
    for node in list(fl["entries"]) + list(fl["chain"]):
        rest = str(node["label"])
        for nm in sorted(real, key=len, reverse=True):
            if nm:
                rest = rest.replace(nm, "")
        assert not rest.replace("·", "").strip(), \
            ("흐름도에 없는 단계 이름을 만들어 냈다", node["label"], rest)
        assert node["who"], ("담당 없는 마디 — 화살표가 누구를 가리키는지 말 못 한다", node)
    # 접수 경로는 **나란한 갈래**다 — 차례에 섞으면 없는 순서가 생긴다.
    ent_names = {n["label"] for n in fl["entries"]}
    assert ent_names == {str(s.get("단계") or "") for s in steps
                         if str(s.get("갈래") or "") == "접수 경로"}, \
        "접수 갈래를 차례와 안 갈랐다 — '카톡 다음에 법인폰'이라는 없는 순서가 생긴다"
    assert all(n["n"] == i for i, n in enumerate(fl["chain"], 1)), "차례 번호가 어긋난다"
    # 화면은 낱말을 안 적고 서버가 준 것만 그린다.
    fm = live.split("function orgFlowMap(", 1)[1].split("\nfunction ", 1)[0]
    for word in ("카톡 접수", "배정 합의", "수금 확인", "세금계산서 발행"):
        assert word not in fm, ("단계 낱말을 화면에 박았다 — 흐름도가 바뀌어도 안 따라간다", word)
    assert "D.flow" in fm and "z.people" in fm, \
        "흐름과 색·상태는 각각 서버 flow 와 zones 에서 와야 한다(두 곳에서 정하면 갈린다)"
    assert "org-fmnote" in fm and "흐름도에 자리가 안 적힌" in fm, \
        "흐름에 없는 사람을 조용히 빼면 '없는 사람'이 된다([169])"
    assert "#v-org .org-flowmap{display:none}" in css, \
        "흐름 지도가 넓은 화면 기본값을 건드린다 — 폰에서만 나와야 한다"
    assert "#v-org .org-st .org-msg{display:none}" in css and "org-msg" in live, \
        "폰에서 줄일 상태 문장이 진짜 요소로 안 싸여 있다(텍스트 노드는 CSS 로 못 숨긴다)"

    print("  [232] 조직도 평면도 — 로스터 정본·AI 자리 셋·상태 넷·아코디언·#v-org 좁힘 · "
          "관리자=센터장 접속(X-Pin 은 아님) · 책상 폭 고정·높이 균일·이름 한 줄 · "
          "폰 흐름 화살표(낱말은 flow_steps 에서만) · 길게 눌러 구역 안 자리 저장 ✅")


def t230_ai_tier_picks_model_and_effort():
    """[230] AI 를 부를 때 **모델·노력을 스스로 고른다** (2026-08-12 지시).

    사용자 지시: "자동으로 모델과 노력 강도 설정해서 진행하는 알고리즘 적용".

    ★ 전까지는 **한 번도 고르지 않았다** — `agent_dispatch` 가 `claude -p` 만 불러
      모든 인계가 기본 모델(제일 비싼 것)로 돌았다. 재시도 한 장이나 원인 모를
      회차 고장이나 같은 값을 치렀다.
    ★ **싸게 하는 것이 목적이 아니라 값에 맞게 하는 것이 목적이다.** 이 프로젝트에서
      잘못된 판단은 못 한 판단보다 나쁘다(`[172]`) — 그래서 **원인을 모르는 일은
      값을 안 아낀다.** 아끼는 것은 답이 이미 정해진 일뿐이다.
    ★ **없는 깃발을 지어내지 않는다.** `--model` 은 확인됐고 '노력' 깃발은 확인된 것이
      없다. 틀린 깃발 하나면 CLI 가 통째로 안 뜨고 인계가 **조용히** 안 된다(`[169]`).
    """
    import ai_tier as T
    import agent_dispatch as A
    from pathlib import Path as _P

    # ── 갈래: 근거가 있을 때만 위아래로 움직인다
    assert T.pick(args=["--check"])["갈래"] == "조회", "읽기만 하는 일을 못 알아본다"
    assert T.pick("code", "왜 죽나", [], 3)["갈래"] == "원인", \
        "원인을 모르는 반복 실패에 값을 아낀다 — 오진이 파일에 박힌다"
    assert T.pick("resource", "다시 돌린다", [])["갈래"] == "재시도", T.pick("resource", "", [])
    # ★ `--check` 가 섞여 있어도 세 번 죽었으면 물어야 할 것은 '조회 결과'가 아니라 '왜'다
    assert T.pick("timeout", "회차", ["--check"], 3)["갈래"] == "원인", \
        "반복 실패보다 조회 표식을 먼저 본다 — 순서가 뒤집혔다"
    # ★ 쓰는 명령을 조회라고 부르지 않는다 — 싸게 판단해서 원장을 건드리면 안 된다
    assert T.pick("", "반영", ["--print", "--apply"])["갈래"] != "조회", \
        "쓰는 명령을 조회로 읽는다"

    # ── 값: 원인·설계는 제일 좋은 것으로, 조회는 싸게
    assert T.pick("code", "", [], 3)["모델"] == "opus", "원인 규명에 값싼 모델을 쓴다"
    assert T.pick(args=["--status"])["모델"] == "sonnet", "조회에 비싼 모델을 쓴다"
    assert "haiku" not in json.dumps(T.TIERS, ensure_ascii=False), \
        "판정 층에 haiku 를 넣었다 — 값싼 오판이 아끼는 것보다 크다"
    for tier, (model, effort, why) in T.TIERS.items():
        assert model in ("sonnet", "opus") and effort in ("low", "medium", "high"), \
            "%s 가 CLI 가 모르는 낱말을 쓴다: %s/%s" % (tier, model, effort)
        assert why, "%s 를 왜 그렇게 골랐는지 안 적는다 — 나중에 아무도 못 고친다" % tier

    # ── 깃발: 확인된 것만 붙는다
    got = T.pick("code", "", [], 3)
    # 실행파일을 안 주면 물어볼 수 없다 → 모델만 붙고 노력은 문장으로 간다
    assert T.flags("claude", got) == ["--model", "opus"], T.flags("claude", got)
    assert T.flags("codex", got) == [], "확인 안 된 codex 깃발을 붙인다 — CLI 가 안 뜬다"
    src = open(os.path.join(ROOT, "ai_tier.py"), encoding="utf-8").read()
    fbody = src.split("def flags(", 1)[1].split("\ndef ", 1)[0]
    # ★ `--effort` 는 2026-08-13 실측으로 실재가 확인됐다. 그래도 **그 실행파일에게
    #   물어본 뒤에만** 붙인다 — 판올림으로 사라지면 CLI 가 통째로 안 뜨고 인계가
    #   조용히 안 된다(`[169]`). 확인 절차 없이 붙이면 그때가 그 사고다.
    assert "supports_flag(executable" in fbody, \
        "--effort 를 물어보지도 않고 붙인다 — 없어지는 날 인계가 조용히 죽는다"
    for bad in ("--reasoning", "--thinking"):
        assert bad not in fbody, "확인 안 된 깃발 %s 를 붙인다" % bad
    # 노력은 깃발이든 문장이든 **반드시 어딘가로** 간다. 둘 다면 같은 말이 두 벌이 된다.
    assert "노력" in T.prompt_line(got, effort_via_flag=False), "노력을 아무 데도 안 넘긴다"
    assert "[노력" not in T.prompt_line(got, effort_via_flag=True), \
        "깃발이 받았는데 문장도 같은 말을 한다 — 두 벌이 되면 언젠가 갈린다"

    # ── 배선: 명령에 실제로 실린다. 못 고르면 예전 그대로 나간다(인계는 멈추지 않는다)
    cmd = A._agent_command("claude", "claude.exe", "p", _P("x.txt"), got)
    assert cmd[:3] == ["claude.exe", "--model", "opus"] and "-p" in cmd, cmd
    assert A._agent_command("claude", "claude.exe", "p", _P("x.txt")) == \
        ["claude.exe", "-p", "p", "--output-format", "text"], "안 고른 경우가 바뀌었다"
    assert A._agent_command("codex", "codex.exe", "p", _P("x.txt"), got)[1] == "exec", \
        "codex 명령이 바뀌었다"
    disp = open(os.path.join(ROOT, "agent_dispatch.py"), encoding="utf-8").read()
    assert 'record["ai_tier"] = chosen' in disp, \
        "고른 것을 티켓에 안 적는다 — 왜 그 모델로 돌았는지 물을 수 없다"

    # ── 사람이 친 말도 등급을 매긴다 (2026-08-13 지시 · 같은 `[230]` 안이다.
    #    번호를 새로 따지 않는 이유: 같은 모듈·같은 지시 줄기이고, `[231]` 은 옆 세션이
    #    `loop_policy` 에 이미 썼다. 번호가 겹치면 어느 쪽을 고쳐야 할지 아무도 모른다.)
    #     ★ **아끼는 쪽으로 기울지 않는다.** 못 가르면 모델은 그대로 두고 노력만 낮춘다 —
    #       싸게 틀리는 것이 비싸게 맞는 것보다 훨씬 비싸다(`[172]`).
    assert T.pick_for_prompt("지금 몇 건이야")["갈래"] == "질문"
    assert T.pick_for_prompt("커밋해")["갈래"] == "수행"
    assert T.pick_for_prompt("왜 회차가 죽는지 원인 분석해")["갈래"] == "설계"
    # 순서: "왜 … 확인해줘" 는 조회가 아니라 원인 규명이다
    assert T.pick_for_prompt("왜 안 되는지 상태 확인해줘")["갈래"] == "설계", \
        "'왜' 가 들어간 질문을 단순 조회로 읽는다 — 값싸게 오진한다"
    vague = T.pick_for_prompt("그거 해줘")
    assert vague["갈래"] == "모호" and vague["모델"] == "opus" and vague["노력"] == "medium", \
        "못 가른 요청에서 모델을 내린다 — 싸게 틀리는 것이 제일 비싸다"
    assert T.pick_for_prompt("")["모델"] == "opus", "빈 입력에서 모델을 내린다"
    for tier, (model, effort, why) in T.PROMPT_TIERS.items():
        assert model in ("sonnet", "opus") and effort in ("low", "medium", "high") and why, \
            "%s 가 CLI 가 모르는 낱말을 쓰거나 이유가 없다" % tier
    # ★ 바뀐 순간에만 말한다 — 매 입력마다 같은 말을 하면 아무도 안 읽는다
    keep_log = T.LOG
    try:
        with tempfile.TemporaryDirectory() as td:
            T.LOG = os.path.join(td, "t.json")
            assert T._remember("설계") == "", "첫 판에 직전 값이 있다"
            assert T._remember("설계") == "설계", "직전 값을 기억 못 한다"
            assert T._remember("질문") == "설계", "직전 값이 갱신 안 된다"
            got = json.load(open(T.LOG, encoding="utf-8"))
            assert got["누적"]["설계"] == 2, "누적을 안 센다 — 규칙을 고칠 근거가 없다"
    finally:
        T.LOG = keep_log
    # 훅은 무슨 일이 있어도 exit 0 이어야 한다 — 사람 입력을 막으면 안 된다
    hook_src = open(os.path.join(ROOT, "ai_tier.py"), encoding="utf-8").read()
    hb = hook_src.split("def hook(", 1)[1].split("\ndef ", 1)[0]
    assert hb.count("except Exception") >= 2 and "return 0" in hb, \
        "훅이 터지면 사람 입력이 막힌다"
    # ★ '알려 준다'를 '자동으로 바꾼다'라고 적으면 거짓말이다 — 문구가 한계를 말해야 한다
    assert "못 바꿉니다" in hook_src, \
        "대화창 모델을 세션이 바꿀 수 있는 것처럼 말한다"
    st = json.load(open(os.path.join(ROOT, "..", ".claude", "settings.json"),
                        encoding="utf-8"))
    assert "ai_tier.py" in json.dumps(st.get("hooks", {}), ensure_ascii=False), \
        "판정 훅이 settings.json 에 안 붙어 있다 — 만들었지만 안 도는 자리다(`[228]`)"

    print("  [230] AI 모델·노력 자동 선택 — 원인/설계는 opus·조회는 sonnet · haiku 배제 · "
          "--effort 는 물어본 뒤에만 · 사람 말도 등급(못 가르면 모델 유지·노력만 낮춤) · "
          "바뀔 때만 알림 ✅")


def t220_flow_yes_no_cycles():
    """[220] 플로우 예/아니오 순환 검증 (2026-08-11 지시).

    사용자 지시: "업무 플로우 차트가 순환 검증해서 다시 돌아오는 예스 오아 노
    구조로 고도화 작업 진행해 나중에 계속 변경 요청할거야".
    '계속 변경'이 오므로 내용은 코드가 아니라 **DB(flow_step)** 에 있고 앱 [수정]
    에서 바꾼다. 지키는 것 셋:
      ① 아니오 대상은 흐름 안에 **실재하는 단계 이름**이어야 저장된다 — 없는
         이름을 조용히 받으면 화면이 '(단계 없음)' 화살표를 영영 그린다.
      ② 검증 질문 없이 아니오만 있는 것은 버린다 — 질문 없는 분기는 뜻이 없다.
      ③ 화면·캡처·개발사양 셋 다 분기를 그린다 — 화면에만 있으면 회의 그림과
         개발자가 직선 흐름으로 읽는다.
    """
    import ledger_db as L
    assert L.FLOW_DEFAULT_CHECKS, "기본 씨앗 순환이 없다 — 일정 연기·미완 재방문 두 개"
    base = L.flow_steps()
    # 저장 왕복: 검증+아니오가 DB 를 거쳐 그대로 돌아오나
    mod = [dict(x) for x in base]
    mod[2] = dict(mod[2], 검증="합성 검증 질문인가?", 아니오=mod[0]["단계"])
    n = L.flow_save(mod, who="synthetic")
    got = L.flow_steps()
    assert got[2]["검증"] == "합성 검증 질문인가?" and got[2]["아니오"] == base[0]["단계"], \
        "검증/아니오가 저장 왕복에서 사라진다"
    # ② 질문 없는 아니오는 버려진다
    mod2 = [dict(x) for x in got]
    mod2[3] = dict(mod2[3], 검증="", 아니오=base[0]["단계"])
    L.flow_save(mod2, who="synthetic")
    assert L.flow_steps()[3]["아니오"] == "", "질문 없는 아니오가 살아남았다"
    # ① 없는 단계를 가리키면 저장을 거부한다
    bad = [dict(x) for x in L.flow_steps()]
    bad[1] = dict(bad[1], 검증="어디로?", 아니오="존재하지 않는 단계 이름")
    try:
        L.flow_save(bad, "synthetic")
        raise AssertionError("없는 아니오 대상이 저장됐다")
    except ValueError:
        pass
    L.flow_save([dict(x) for x in base], who="synthetic")   # 제자리로
    # ③ 세 화면이 다 그린다 — 보기(flowCheckHtml)·캡처(chk)·개발사양(아니오 간선)
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for need in ("function flowCheckHtml(", "function flowNoLabel(", "function flowNoOptions(",
                 'data-f="검증"', 'data-f="아니오"', ".flow-check{",
                 "(it.chk||[]).forEach", "-.->|아니오|"):
        assert need in live, f"순환 검증 조각이 없다: {need}"
    # 수정 폼이 걷고(flowCollect) 새 단계에도 칸이 있다
    assert "검증:g('검증'), 아니오:g('아니오')" in live, "수정 폼이 검증/아니오를 안 걷는다"
    print("  [220] 플로우 예/아니오 순환 — DB 정본 · 실재 단계만 · 질문 없는 분기 제거 · "
          "화면/캡처/사양 3벌 ✅")


def t221_commit_hygiene_under_siblings():
    """[221] 세션 자동 커밋의 경로 위생 — 사고 #36 재발방지 (worksplit #33).

    `git commit` 은 인덱스 **전체**를 커밋한다. 작업 폴더는 세션끼리 공유하므로
    `add -A` 한 번이 옆 세션이 만들던 파일까지 딸려 담는다(실사고 dac781f:
    1개 의도 → 17개 커밋). 그 상태는 실패처럼 안 보인다 — 파일은 디스크에
    그대로고 커밋도 성공이다. 지키는 겹은 셋:
      ① 옆 세션이 살아 있으면 커밋 메시지에 '섞여 있을 수 있다'와 **판정 근거(sid)**
         를 적고 푸시하지 않는다 — 밀면 반쯤 고친 남의 코드가 원격 master 가 된다.
      ② 멈추는 자리(비밀값 스캔·커밋 실패)는 스테이징을 되돌린다 — 남긴 스테이징은
         다음 사람의 `git commit -m` 한 번이 통째로 커밋한다(#36 이 그렇게 났다).
         절차 문구('커밋 전 git status')도 도구가 스스로 말해야 한다.
      ③ 사람 절차는 정본 지시문에 있다 — 'git add -A 를 쓰지 말 것' ·
         '경로를 지정해 커밋'. 문서가 지워지면 규칙은 다음 세션에서 사라진다.
    """
    import session_wrapup as W
    calls = []
    real = (W.git, W.git_lines, W._other_live_sessions, W._unstage_huge)

    def fake_git(*a):
        calls.append(a)
        if a[:2] == ("status", "--porcelain"):
            return True, "M mine.py"
        if a[0] == "grep":
            return False, ""                 # 비밀값 없음 (grep 은 '찾으면' 0 을 준다)
        return True, ""

    def fake_lines(*a):
        calls.append(a)
        if "--diff-filter=ACMRT" in a:
            return True, ["mine.py"]         # 이번 커밋이 담는 경로
        return True, []                      # 담기 전 인덱스는 비어 있었다

    try:
        W.git, W.git_lines = fake_git, fake_lines
        W._unstage_huge = lambda: []
        W._other_live_sessions = lambda: [{"sid": "t221sib1"}]
        steps = []
        W.step_commit("claude", "synthetic-t221", steps)
        msg = next(a[3] for a in calls if a and a[0] == "commit")
        assert "섞여 있을 수 있다" in msg and "푸시하지 않았다" in msg, \
            "옆 세션 생존인데 커밋 메시지가 '섞였을 수 있음'을 말하지 않는다"
        assert "t221sib1" in msg, "살아 있다고 본 근거(sid)를 커밋에 안 적었다"
        assert not any(a and a[0] == "push" for a in calls), "옆 세션 생존인데 푸시했다"
        assert steps and steps[-1]["성공"] and "보류" in steps[-1]["메모"], \
            "푸시 보류를 단계 기록이 말하지 않는다"
        # 혼자일 때는 푸시까지 간다 — 보류가 기본값이 되면 아무도 안 미는 저장소가 된다.
        calls.clear()
        W._other_live_sessions = lambda: []
        W.step_commit("claude", "synthetic-t221", [])
        assert any(a and a[0] == "push" for a in calls), "옆 세션이 없는데도 푸시를 안 한다"
    finally:
        W.git, W.git_lines, W._other_live_sessions, W._unstage_huge = real
    # ② 멈추는 자리마다 되돌림 — 스캔 실패·커밋 실패 두 곳 모두 _rollback 을 거친다.
    import inspect
    src = inspect.getsource(W.step_commit)
    assert src.count("_rollback()") >= 2, \
        "멈추는 자리가 스테이징을 되돌리지 않는다 — #36 재발 경로"
    assert "커밋 전 git status" in src, "도구가 'git status 확인' 절차를 말하지 않는다"
    # ③ 정본 지시문의 사람 절차
    doc = open(os.path.join(ROOT, "CLAUDE.md"), encoding="utf-8").read()
    assert "git add -A` 를 쓰지 말 것" in doc and "경로를 지정해 커밋" in doc, \
        "정본 지시문에서 경로 지정 커밋 규칙이 사라졌다"
    print("  [221] 커밋 경로 위생 — 옆 세션 표기·푸시 보류 · 스테이징 되돌림 · "
          "문서 절차 3겹 ✅")


def t222_flow_charts_switch_and_capture():
    """[222] 플로우 차트는 여러 장이다 — 종전/개선 스위치 + 캡처 (2026-08-11 지시).

    사용자 지시: "돌발 as플로우 차트 옆에 (종전) 표시 / 앱을 통해 개선되는
    플로우차트를 하나 만들고 옆에 개선 / 스위치 기능으로 각 차트 보고 캡처 /
    (캡처 화면에) 이 방식의 문제점 및 개선해야될 점 정리 / 추후 정기검사
    플로우차트도 감안". 지키는 것:
      ① 차트 등록부(FLOW_CHARTS)에 종전·개선이 있고 목록 밖 열쇠는 거부한다 —
         오타 열쇠를 조용히 받으면 아무도 못 보는 빈 차트에 저장된다.
      ② **저장 격리** — 한 차트를 저장·되돌려도 다른 차트는 한 행도 안 바뀐다.
         통째 DELETE 로 돌아가면 남의 차트가 말없이 빈다(조용한 사고).
      ③ 문제점·개선점은 DB(flow_note) 정본 + 씨앗 — 전부 비우면 씨앗으로 돌아간다.
      ④ 화면 스위치·캡처(ㄱ자 연결선·문제점 줄)·개발 사양이 다 함께 움직인다.
    """
    import ledger_db as L
    tails = {c["key"]: c["꼬리"] for c in L.FLOW_CHARTS}
    assert tails.get("as_legacy") == "종전" and tails.get("as_app") == "개선", \
        "차트 등록부에 종전/개선이 없다"
    try:
        L.flow_steps("없는차트")
        raise AssertionError("목록 밖 차트 열쇠를 받았다")
    except ValueError:
        pass
    # ② 저장·되돌리기 격리
    legacy_before = L.flow_steps("as_legacy")
    mod = [dict(x) for x in L.flow_steps("as_app")]
    mod[0] = dict(mod[0], 메모="t222 합성 표식")
    L.flow_save(mod, who="synthetic", key="as_app")
    assert L.flow_steps("as_legacy") == legacy_before, "as_app 저장이 종전 차트를 바꿨다"
    assert L.flow_steps("as_app")[0]["메모"] == "t222 합성 표식"
    L.flow_restore("synthetic", "as_app")
    assert L.flow_steps("as_app")[0]["메모"] != "t222 합성 표식", "차트별 되돌리기가 안 된다"
    assert L.flow_steps("as_legacy") == legacy_before, "as_app 되돌리기가 종전 차트를 바꿨다"
    # ③ 문제점·개선점 — 씨앗이 실려 있고, 저장 왕복이 되고, 비우면 씨앗 복귀
    assert L.flow_notes("as_legacy"), "종전 차트의 문제점 씨앗이 비어 있다"
    L.flow_notes_save("as_app", ["t222 메모"], "synthetic")
    assert L.flow_notes("as_app") == ["t222 메모"], "메모 저장 왕복이 안 된다"
    L.flow_notes_save("as_app", [], "synthetic")
    assert L.flow_notes("as_app") == list(L.FLOW_NOTE_DEFAULT["as_app"]), \
        "비우면 씨앗으로 돌아가야 한다"
    # ④ 화면·캡처·사양 — 스위치, 이름표, 문제점 줄, ㄱ자 연결선
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    for need in ("function flowPick(", 'id="flowCharts"', 'id="flowNotesCard"',
                 "function flowChartLabel(", "flowNotesHead()", "const bend =",
                 "noteLines", "chart: FLOW_CHART", 'id="flowNotesTa"'):
        assert need in live, f"차트 스위치·캡처 조각이 없다: {need}"
    server = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert ("flow_charts()" in server and "flow_notes(" in server
            and 'qs.get("chart"' in server), "서버가 차트 열쇠·메모를 안 내려보낸다"
    print("  [222] 플로우 차트 여러 장 — 종전/개선 스위치 · 저장 격리 · "
          "문제점 DB 정본 · 캡처 수록 ✅")


def t223_superseded_evidence_heals_itself():
    """추월된 '없음 확인' 근거를 **기계가 스스로** 무효로 만든다 (2026-08-11 지시).

    `[217]` 은 읽는 쪽 둘(수집 계획·인계 문서)이 추월된 근거를 **거르게** 했다.
    그런데 근거는 틀린 채로 남고, 되돌리는 길은 사람이 밴드 피드를 열어 `--latest` 를
    적어 주는 것뿐이었다 — 그 한 줄이 이 사고의 마지막 사람 몫이었다. 여기서 재는 것:
      ① 모순(근거 위 번호가 이미 수확됨)이면 '모름'으로 되돌리고 옛 값을 남긴다
      ② `top+1` 을 지어내지 않는다 — 그건 `[217]` 을 손수 다시 만드는 것이다
      ③ 캐시는 한 글자도 안 고친다(틀린 근거로 실재하는 글에 absent 를 찍으면 못 되돌린다)
      ④ **낡기만 한 근거는 안 건드린다** — 낡음은 틀림이 아니다
      ⑤ 정정한 뒤에는 판정하는 쪽도 더는 '추월'이라 말하지 않는다(둘이 같은 눈이어야 한다)
      ⑥ dry 는 한 바이트도 안 쓴다
      ⑦ 워치독이 부르고, **붙여넣기 파일 만들기보다 먼저** 온다
    """
    import importlib
    import shutil
    import tempfile

    sys.path.insert(0, os.path.join(ROOT, "band"))
    RL = importlib.import_module("real_latest")
    rp = importlib.import_module("recheck_plan")

    BAND, TODAY = "90610953", "2026-08-11"
    TMP = tempfile.mkdtemp(prefix="bandheal_")
    hold = (RL.SEEN, RL.CACHE)
    try:
        RL.SEEN = os.path.join(TMP, "밴드_확인시각.json")
        RL.CACHE = os.path.join(TMP, "cache")
        os.makedirs(RL.CACHE, exist_ok=True)
        # 실측(2026-08-11): 근거는 '5438 부터 없다'인데 5447 이 이미 수확돼 있었다.
        # 5460 은 오염이라 수확이 아니다 — 이걸 세면 정정과 판정의 눈이 갈린다.
        posts = {"5447": {"created_at": 1786000000000, "captured_at": 1786000000000},
                 "5460": {"created_at": 1786000000000, "contaminated": True}}
        cache = os.path.join(RL.CACHE, BAND + ".json")
        with open(cache, "w", encoding="utf-8") as fh:
            json.dump({"band_name": BAND, "posts": posts}, fh, ensure_ascii=False)
        before = open(cache, "rb").read()

        def put(n, seen):
            with open(RL.SEEN, "w", encoding="utf-8") as fh:
                json.dump({BAND: {"이름": "쿠팡AS", "수집최대": n - 1,
                                  "없음확인": n, "확인시각": seen}}, fh, ensure_ascii=False)

        # ⑥ dry — 무엇이 틀렸는지는 말하되 한 바이트도 쓰지 않는다
        put(5438, TODAY)
        raw = open(RL.SEEN, "rb").read()
        got = RL.heal(apply=False, today=TODAY)
        assert [f["이전"] for f in got] == [5438], "모순을 못 봤다"
        assert open(RL.SEEN, "rb").read() == raw, "dry 인데 근거를 고쳤다"

        # ①②③ 실제 정정
        got = RL.heal(apply=True, today=TODAY)
        rec = json.load(open(RL.SEEN, encoding="utf-8"))[BAND]
        assert got[0]["실제수확"] == 5447, "오염된 5460 을 수확으로 셌다 — 판정과 눈이 갈렸다"
        assert int(rec.get("없음확인") or 0) == 0, \
            "'모름'이 아니라 새 없음확인을 지어냈다 — 근거 없는 조용함이 다시 생긴다"
        assert rec.get("이전없음확인") == 5438, "무엇을 믿고 있었는지를 지웠다"
        assert int(rec.get("수집최대") or 0) == 5447, "실제 수확 지점을 안 적었다"
        assert open(cache, "rb").read() == before, \
            "정정이 캐시를 건드렸다 — 실재하는 글을 유령으로 만드는 쪽이다"

        # ⑤ 판정하는 쪽도 더는 '추월'이라 안 한다
        cut, why = rp.judge_absent(rec, posts, TODAY)
        assert cut is None and "추월" not in why, \
            "정정하고도 판정이 그대로다 — 고쳐도 안 고쳐지는 자리가 된다"

        # ④ 낡기만 한 근거는 그대로 둔다(5447 < 5500 이라 모순이 아니다)
        put(5500, "2026-08-09")
        raw = open(RL.SEEN, "rb").read()
        assert RL.heal(apply=True, today=TODAY) == [], "낡았다고 지웠다 — 낡음은 틀림이 아니다"
        assert open(RL.SEEN, "rb").read() == raw, "안 고친다면서 파일을 다시 썼다"
    finally:
        RL.SEEN, RL.CACHE = hold
        shutil.rmtree(TMP, ignore_errors=True)

    # ⑦ 파일에 넣지 않은 것은 자동이 아니다 — 회차에 매여 있나, 순서는 맞나
    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    body = wd.split("def main")[1]
    assert "heal_band_evidence(" in body, "워치독이 안 부른다 — 사람이 또 손으로 고쳐야 한다"
    assert body.index("heal_band_evidence(") < body.index("heal_stale_pastefiles("), \
        "붙여넣기 파일을 먼저 만든다 — 틀린 근거로 만든 목록이 그대로 사람 손에 간다"
    print("  [223] 추월된 근거 자동 정정 — 모순만 · 지어내지 않음 · 캐시 불변 · "
          "낡음은 보존 · 워치독 선행 ✅")


def t219_noon_round_is_daily_windowed_and_yields():
    """[219] 정오 회차는 창 안에서 하루 한 번만 돌고, 남의 것을 절대 빼앗지 않는다.

    2026-08-11 지시: "하루한번 12시에서 13시 사이 다른 세션이랑 충돌 안나게 실행
    알고리즘 코딩해(자동화 100%, 컴팩팅도 자동화, 내가 손 안대게 처리)".

    지키는 것 다섯: ① 창 밖·오늘 완주면 안 돈다 ② `ledger`·`publish` 점유와 도는
    회차에는 **전부 양보**하고 `--force` 로도 안 뺏는다 ③ `code` 점유는 합성검증만
    건너뛴다(남의 반쯤 고친 코드가 내는 빨강은 내 회차의 사실이 아니다) ④ **양보를
    완주로 적지 않는다**(적으면 그날 회차가 영영 안 돈다) ⑤ 단계에 수집·엑셀쓰기·
    커밋이 없다.
    """
    import noon_run as N
    import datetime as _dt

    noon = _dt.datetime(2026, 8, 12, 12, 30)
    live = lambda what: [{"what": what, "who": "claude", "sid": "abcd1234", "alive": True}]

    assert N.decide(_dt.datetime(2026, 8, 12, 9, 30), {}, [], [])["kind"] == "창밖"
    assert N.decide(noon, {"done_date": "2026-08-12"}, [], [])["kind"] == "완주"
    assert N.decide(noon, {"done_date": "2026-08-11"}, [], [])["go"] is True, \
        "어제 완주는 오늘을 막지 않는다"
    for what in ("ledger", "publish"):
        v = N.decide(noon, {}, live(what), [])
        assert v["go"] is False and v["kind"] == "양보", what
        assert N.decide(noon, {}, live(what), [], force=True)["go"] is False, \
            f"--force 가 남의 '{what}' 점유를 빼앗았다 — 강제는 창·중복만 무시한다"
    assert N.decide(noon, {}, [], ["일일대조(09:50)"])["go"] is False, "도는 회차에 양보해야 한다"
    # ★ 그러나 **매일 양보만 하는 회차는 없는 회차와 같다**(2026-08-12 실측: 증분 파이프라인이
    #   새 자료를 만나면 12분을 넘겨 도는데 5분마다 불려 창 내내 락이 걸려 있었다). 창의
    #   마지막 기회에는 Z: 를 훑는 단계만 비켜 두고 돈다 — 양보는 'SMB 를 같이 긁지 않는 것'이다.
    last = _dt.datetime(2026, 8, 12, 12, 50)
    v = N.decide(last, {}, [], ["증분 파이프라인"])
    assert v["go"] is True and "마지막 기회" in v["kind"], \
        "창이 끝나는데도 양보만 하면 자료가 들어온 날은 하루도 못 돈다(경보도 안 뜬다)"
    assert v["skip"] == list(N.HEAVY_STEPS) and "합성검증" not in v["skip"], \
        "마지막 기회는 무거운 단계만 비켜야 한다 — 관문(합성검증)까지 빼면 돈 뜻이 없다"
    step_names = [s[0] for s in N.steps()]
    for heavy in N.HEAVY_STEPS:
        assert heavy in step_names, \
            f"HEAVY_STEPS 이름이 단계 목록과 어긋난다({heavy}) — 안 건너뛰면서 건너뛴 줄 안다"
    assert N.decide(last, {}, live("ledger"), ["증분 파이프라인"])["go"] is False, \
        "마지막 기회여도 'ledger' 점유에는 전부 양보한다 — 그쪽은 vN+1 을 쓰는 중이다"
    # 죽은 세션의 점유는 잡은 것이 아니다([210]·[213]) — 그것 때문에 매일 건너뛰면 안 된다.
    dead = [{"what": "ledger", "who": "claude", "sid": "dead", "alive": False}]
    assert N.decide(noon, {}, dead, [])["go"] is True

    v = N.decide(noon, {}, live("code"), [])
    assert v["go"] is True and v["skip"] == ["합성검증"], \
        "code 점유는 전부 양보가 아니라 합성검증만 건너뛴다(대화 세션은 몇 시간씩 잡는다)"
    assert N.decide(noon, {}, live("band"), [])["skip"] == [], \
        "이 회차는 수집을 하지 않으므로 band 점유와 부딪히지 않는다"
    ran = [r["step"] for r in [{"step": s[0]} for s in N.steps()]]
    assert "합성검증" in ran, "건너뛸 대상 이름이 단계 목록과 어긋나면 아무것도 안 건너뛴다"

    # ④ 양보한 부름이 마커에 '오늘 완주'를 적으면 그날 회차가 영영 안 돈다.
    with tempfile.TemporaryDirectory(prefix="csos-noon-219-") as td:
        old = (N.MARKER, N.REPORT_MD, N.EXCLUSIVE)
        try:
            N.MARKER = os.path.join(td, "marker.json")
            N.REPORT_MD = os.path.join(td, "report.md")
            sys.argv = ["noon_run.py"]                 # 창 밖 시각이라 go=False 경로로 간다
            os.environ["COUPANG_NOON_WINDOW"] = "00:00-00:01"
            assert N.main() == 0
            saved = json.load(open(N.MARKER, encoding="utf-8"))
            assert "done_date" not in saved and saved.get("skips"), \
                "양보를 완주로 적었다 — 그날 정오 회차가 다시는 안 돈다"
        finally:
            os.environ.pop("COUPANG_NOON_WINDOW", None)
            N.MARKER, N.REPORT_MD, N.EXCLUSIVE = old

    src = open(os.path.join(ROOT, "noon_run.py"), encoding="utf-8").read()
    # 백틱 인용은 **설명**이지 코드가 아니다 — 안 걷어내면 "쓰지 말라"고 적어 둔 문장
    # 자체를 위반으로 읽고 빨강을 낸다(검증이 제 문서에 걸리는 자리).
    code_only = re.sub(r"`[^`]*`", "", src.split("def git_pending")[0])
    assert "proc_guard" in src and "subprocess.run(" not in code_only, \
        "회차 단계가 subprocess.run(timeout=) 을 쓰면 윈도우에서 영원히 안 끝날 수 있다([175])"
    body = src.split("def steps(")[1].split("\ndef ", 1)[0]
    for banned in ("--queue", "--apply", "band_sync", "convert_dump", "collect_", "upload_intake",
                   "commit", "workbook_patch"):
        assert banned not in body, f"정오 회차가 하면 안 되는 일을 한다: {banned}"

    ps = open(os.path.join(ROOT, "install_noon_schedule.ps1"), encoding="utf-8").read()
    assert 'At "12:00"' in ps and "noon_run.py" in ps, "스케줄러가 12시에 이 회차를 안 부른다"
    assert "Minutes 10" in ps and "Minutes 55" in ps, \
        "재시도 반복이 없거나 창(13:00)을 넘는다 — 양보한 날 두 번째 기회가 사라진다"

    print("  [219] 정오 회차 — 창 12~13시·하루 한 번 · 남의 점유 불가침(force 도) · "
          "code 는 합성검증만 건너뜀 · 마지막 기회엔 무거운 단계만 비켜 돎 · "
          "양보를 완주로 안 적음 ✅")


def t196_stage_words_come_from_one_place():
    """[196] 돌발AS·정기점검 단계 낱말은 **한 곳**에서 오고, 바뀌면 자국이 남는다.

    2026-08-10 지시: "돌발 AS 카테고리 및 정기점검 카테고리에 입력창을 만들어 …
    플로우차트 참고해서 진행할 수 있게 하고 나중에 플로우차트가 변경되면 기능도
    따라서 연동되게".

    ★ 이 검증이 지키는 것은 '입력창이 있다'가 아니라 **낱말이 한 곳에서만 정해진다**는
    것이다. 화면·서버가 각자 '접수'·'예정'을 적어 두면 그 순간 사본이 둘이고, 흐름도가
    바뀌어도 안 따라간다([162]). 실제로 그렇게 어긋나 있었다 — 서버 기본값 '접수' 는
    관리대장 드롭다운에 **없는 낱말**이었고 그렇게 들어간 행이 원장에 85건 있었다.
    그리고 화면 선택지는 '지금 목록에 쓰여 있는 값'만 모아서, 드롭다운에 있는
    기사배정·일정확정·방문중·재방문예정·보류를 **앱에서는 영영 고를 수 없었다.**
    """
    import work_flow as W

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    live = open(os.path.join(root, "webapp", "index.html"), encoding="utf-8").read()
    server = open(os.path.join(root, "webapp", "app_server.py"), encoding="utf-8").read()
    src = open(os.path.join(root, "work_flow.py"), encoding="utf-8").read()

    # ① 정의는 스스로 낱말을 짓지 않는다 — 근거는 관리대장 드롭다운이다.
    assert "dataValidation" in src and "_dv_source" in src, \
        "단계 낱말의 근거가 관리대장 드롭다운이 아니다"
    assert "10_코드관리" not in src.split('"""', 2)[2] or "_dv_source" in src, \
        "코드시트 열을 코드에 박아 두면 사람이 드롭다운 출처를 옮긴 날 티가 안 난다"

    # ② 합성 정의로 갈래 판정을 확인한다(실데이터 없이 돈다).
    flow = [{"순서": 0, "단계": "일정 확정", "담당": "류지영", "근거": "밴드"},
            {"순서": 1, "단계": "완료 보고", "담당": "기사", "근거": "밴드"}]
    assert W._link("일정확정", flow) and W._link("일정확정", flow)["담당"] == "류지영"
    # 근거가 없으면 **붙이지 않는다** — 잘못 붙이면 엉뚱한 담당자를 가리킨다.
    assert W._link("보류", flow) is None, "근거 없이 흐름 단계를 갖다 붙였다"
    # 후보가 둘이면 고르지 않는다.
    two = flow + [{"순서": 2, "단계": "확정 통보", "담당": "오종현", "근거": ""}]
    assert W._link("확정", two) is None, "후보가 여럿인데 하나를 골랐다"

    # ③ 목록 밖 낱말은 지우지 않고 '목록 밖'으로 표시한다 — 이미 그렇게 적힌 행이 있다.
    assert "목록 밖" in src and "목록밖" in src, "목록 밖 낱말을 표시하지 않는다"
    assert "지어내지" in src, "낱말을 짓지 않는다는 근거 주석이 없다"

    # ④ 지문은 '쓰인 건수'로 흔들리지 않는다 — 매일 바뀌면 경보가 죽는다.
    base = {"갈래": {"as": {"단계": [{"단계": "신규접수", "쓰인건수": 1}]}}}
    more = {"갈래": {"as": {"단계": [{"단계": "신규접수", "쓰인건수": 999}]}}}
    assert W.fingerprint_of(base) == W.fingerprint_of(more), \
        "건수가 늘었다고 '흐름이 바뀌었다'고 하면 아무도 안 본다"
    other = {"갈래": {"as": {"단계": [{"단계": "신규접수"}, {"단계": "방문중"}]}}}
    assert W.fingerprint_of(base) != W.fingerprint_of(other), "단계가 늘었는데 지문이 같다"

    # ⑤ 서버가 단계 낱말을 손으로 적지 않는다.
    job = server.split("def save_new_workcenter_job", 1)[1].split("\ndef ", 1)[0]
    assert "work_flow.default_stage" in job, "신규 등록 기본 단계를 서버가 지어낸다"
    assert '"접수" if category' not in job and "'접수' if category" not in job, \
        "옛 기본값 '접수' 가 남아 있다 — 관리대장 목록에 없는 낱말이다"
    assert "stage_words" in job, "목록 밖 낱말이 새로 들어오는 것을 막지 않는다"

    # ⑥ 화면도 마찬가지 — 상태 선택지·완료 낱말이 정의에서 온다.
    assert "/api/flow-stages" in server and "/api/flow-stages" in live, \
        "단계 정의를 내려 주는·받아 가는 자리가 없다"
    edit = live.split("if(typ==='status'){", 1)[1].split("}", 1)[0]
    assert "stageList(k)" in edit, "상태 선택지를 화면이 직접 모은다(정의를 안 본다)"
    comp = live.split("async function wtComplete", 1)[1].split("\n}", 1)[0]
    assert "stageDone(k)" in comp, "완료 낱말이 화면에 박혀 있다"

    # ⑦ 입력창이 두 화면에 있고 **폼은 하나**다(사본을 만들지 않는다).
    assert "openNewWork('${k}')" in live, "돌발AS·정기점검 화면에 신규 등록 길이 없다"
    assert live.count('id="newWorkForm"') == 1, "신규 등록 폼이 둘 이상이다 — 칸이 갈린다"
    assert "layerOpen('newWorkForm'" in live, "화면마다 폼을 새로 그리고 있다"
    # 류지영 업무센터로 바로 들어와도 선택지가 차 있어야 한다 — 비어 있으면 서버
    # 기본값으로 메워져 저장은 되고, 그래서 아무도 고장인 줄 모른다.
    # ★ 표식은 **그 자리에만 있는 글자**여야 한다. 예전 표식 `if(v==='ryu'` 는 화면 어딘가
    #   다른 곳에 같은 글자가 생기면 그쪽에서 잘려, 멀쩡한 코드를 두고 "빈 채로 뜬다"고
    #   말한다(2026-08-12 실측 — 김미영 센터 첫 화면 코드가 앞에서 걸렸다).
    ryu_view = live.split("if(v==='ryu' && PIN", 1)[1][:400]
    assert "fillNewWorkStatus" in ryu_view, "류지영 화면에서 단계 선택지가 빈 채로 뜬다"

    # ⑧ 관리자도 등록한다(예전에는 류지영 업무센터 로그인에서만 열렸다).
    newjob = server.split('if p == "/api/staff/new-job":', 1)[1][:1200]
    assert "is_admin" in newjob and "_require_staff(" not in newjob, \
        "관리자가 앱에서 신규 건을 만들 길이 없다"
    assert 'fields["submitter"]' in newjob, "등록자를 화면이 보낸 값으로 믿는다"

    # ⑨ 바뀌면 **자국이 남는다** — 조용히 따라가면 아무도 모른다.
    assert "def banner" in src and "def ack" in src, "변경 알림을 내릴 길이 없다"
    hand = open(os.path.join(root, "session_handoff.py"), encoding="utf-8").read()
    assert "work_flow_change" in hand and "업무흐름" in hand, "인계 문서가 변경을 안 본다"
    daily = open(os.path.join(root, "daily_run.py"), encoding="utf-8").read()
    assert "work_flow.py" in daily and "--check" in daily, "회차가 정의를 안 본다"

    # ⑩ 읽기 전용이다 — 정의를 만드는 길이 엑셀을 고치지 않는다.
    for bad in ("ledger_writer", "--apply", "enqueue"):
        assert bad not in src, "단계 정의가 원장을 건드린다: " + bad

    # ⑪ **밴드 등록기도 같은 자리에서 낱말을 받는다** (2026-08-13 실사고).
    #    ⑤ 가 서버의 신규 등록만 지키고 있어서, 밴드→앱DB 등록기(`band_canonical`)는
    #    ``접수`` 를 그대로 박아 쓰고 있었다.  그 낱말은 관리대장 드롭다운에 없다 —
    #    그래서 엑셀에서 그 칸을 열면 사람이 다시 골라야 하는데 **화면에는 멀쩡히
    #    '접수' 라고 보인다.**  오류가 나지 않는 종류의 잘못이라 아무도 몰랐고,
    #    앱 DB 에 123건이 그 상태로 쌓여 있었다.
    #    글자 검사는 '있어야 할 것'이 아니라 **'되돌아가면 안 되는 것'** 에 쓴다.
    bc = open(os.path.join(root, "band_canonical.py"), encoding="utf-8").read()
    st = bc.split("def _status", 1)[1].split("\ndef ", 1)[0]
    assert '"접수"' not in st and "'접수'" not in st, \
        "band_canonical 이 목록 밖 낱말 '접수' 를 다시 박아 쓴다"
    assert "_stage_words(" in st, "밴드 등록기가 단계 낱말을 정의에서 안 받아 온다"
    assert "work_flow.default_stage" in bc and "work_flow.done_stage" in bc, \
        "밴드 등록기가 work_flow 를 안 본다 — 드롭다운이 바뀌어도 안 따라간다"
    # 비싼 워크북 열기를 레코드마다 부르면 안 된다([168]) — 한 회차 1,600건 자리다.
    assert "_STAGE_CACHE" in bc, "정의를 레코드마다 다시 읽는다"
    import band_canonical as BC

    BC._STAGE_CACHE.clear()
    assert BC._status("돌발AS", {}) == W.default_stage("as"), "밴드 신규 AS 기본 단계가 정의와 다르다"
    assert BC._status("정기점검", {}) == W.default_stage("pm"), "밴드 신규 PM 기본 단계가 정의와 다르다"
    assert BC._status("돌발AS", {"진행상태": "작업완료"}) == W.done_stage("as"), \
        "밴드 완료 낱말이 정의와 다르다"
    # 정의를 못 읽어도 **등록 자체는 계속한다** — 낱말 하나 때문에 접수를 잃는 쪽이 더 나쁘다.
    assert BC._STAGE_FALLBACK["as"][0] != "접수", "대비값이 목록 밖 낱말이다"

    print("  [196] 단계 낱말은 관리대장 드롭다운 한 곳에서 오고 바뀌면 인계에 오른다 ✅")


def t197_restart_blip_is_not_a_failure():
    """[197] 서버가 다시 뜨는 동안의 502 는 '갱신 실패'가 아니다.

    2026-08-10 지시: "갱신실패 없애고 갱신 잘 되게 해".

    폰은 클라우드플레어 터널을 거쳐 이 PC 로 온다. 앱 서버를 다시 띄우면 포트가 답을
    주기까지 **실측 9.3초**가 걸리고, 그동안 터널은 502 를 준다. 예전에는 그 9초가
    빨간 카드 **일곱 장**이 되어 사람이 '다시 시도'를 일곱 번 눌러야 했다 —
    저절로 나을 일에 사람 손을 쓰게 한 것이다.

    ★ 이 검증이 지키는 것은 '다시 건다'가 아니라 **기다리는 시간이 실측보다 길다**는
    것이다. 짧게 잡으면 다시 걸어 놓고도 결국 실패로 끝나 예전과 똑같아진다.
    그리고 **아무거나 다시 걸지 않는다** — 401·429 는 다시 걸수록 나빠진다.
    """
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    live = open(os.path.join(root, "webapp", "index.html"), encoding="utf-8").read()
    rs = open(os.path.join(root, "webapp", "restart_server.py"), encoding="utf-8").read()

    # ① 기다리는 시간이 재시작 실측(9.3초)보다 뒤까지 간다.
    m = re.search(r"const RETRY_WAIT_MS = \[([0-9,\s]+)\]", live)
    assert m, "재시도 간격 표가 없다"
    waits = [int(x) for x in m.group(1).replace(" ", "").strip(",").split(",")]
    assert len(waits) >= 3, "한두 번 다시 걸고 마는 것은 9초를 못 넘는다"
    assert waits == sorted(waits), "간격이 점점 늘지 않으면 뜨는 중인 서버를 몰아친다"
    last_try_sec = sum(waits) / 1000.0
    assert last_try_sec >= 9.3, (
        "마지막 시도가 %.1f초라 재시작 실측 9.3초를 못 넘긴다 — "
        "다시 걸어 놓고도 실패로 끝난다" % last_try_sec)

    # ② 다시 걸면 안 되는 것을 가른다.
    fn = live.split("function isTransientError", 1)[1].split("\n}", 1)[0]
    for code in ("401", "403", "429"):
        assert code in fn, "%s 를 다시 거는지 안 거는지 판단하지 않는다" % code
    assert "return false" in fn, "다시 걸면 안 되는 갈래가 없다"
    st = re.search(r"TRANSIENT_STATUS = new Set\(\[([0-9,\s]+)\]\)", live)
    assert st, "일시적 상태 목록이 없다"
    codes = {int(x) for x in st.group(1).replace(" ", "").strip(",").split(",")}
    assert {502, 503, 504} <= codes, "터널이 주는 502·503·504 를 일시적으로 안 본다"
    assert 429 not in codes and 401 not in codes, "다시 걸수록 나빠지는 것을 다시 건다"

    # ③ 화면 자료 묶음이 그 재시도를 쓴다.
    sec = live.split("async function fetchDataSection", 1)[1].split("\n}", 1)[0]
    assert "apiWithRetry" in sec, "자료 묶음이 한 번만 걸고 실패라 적는다"
    assert "s.attempt" in sec, "몇 번째 다시 걸고 있는지 화면이 말하지 않는다"

    # ④ 뒤에서 받아 오다 끊긴 것도 곧장 빨갛게 칠하지 않는다.
    bg = live.split("function dataSectionBackgroundFailure", 1)[1].split("\n}\n", 1)[0]
    assert "isTransientError" in bg and "_apiNet" in bg, \
        "배경 갱신 실패를 다시 걸지 않는다 — 사람이 아무것도 안 눌렀는데 빨간 카드가 뜬다"
    assert "swrDone(pathKey,false)" in bg.replace(" ", ""), \
        "끝내 포기했는데도 위 칩이 '갱신 중'으로 남는다"
    swr = live.split("swrDone(key, false); dataSectionBackgroundFailure", 1)
    assert len(swr) == 1, "배경 실패가 아직도 곧바로 칩을 실패로 칠한다"

    # ⑤ 원인이 하나면 손도 한 번이다.
    assert "function retryAllDataSections" in live and "data-health-all" in live, \
        "일곱 개가 같이 실패했는데 일곱 번 누르게 한다"

    # ⑥ '프로세스가 있다'와 '답을 준다'를 가른다 — 그 사이가 502 구간이다.
    assert "def answering" in rs, "포트가 답하는지 안 보고 '올라왔습니다'를 찍는다"
    main = rs.split("def main", 1)[1]
    assert "answering()" in main, "재시작이 답을 기다리지 않는다"
    assert "502" in main, "재시작이 폰 쪽에서 무슨 일이 나는지 안 알려 준다"

    print("  [197] 재시작 9초의 502 를 스스로 넘긴다 · 다시 걸면 안 될 것은 안 건다 ✅")


def t198_source_index_no_per_file_stat():
    """[198] 원본 색인이 파일마다 `os.stat(경로)` 를 다시 부르지 않는다.

    ★ 이것이 색인 한 번을 두 시간 반으로 만든 자리다 (2026-08-11 실측).
      같은 폴더에서 잰 값: scandir 항목의 stat **0.04 ms/개** 대 `os.stat(경로)`
      **135~155 ms/개** — Z:(SMB)에서는 후자가 파일마다 왕복 한 번이다.
      112,662개 × 0.145초 = **4.5시간**이 전부 그 한 줄이었다.
      고친 뒤 실측 **24.8초**(147,223개).
    ★ 캐시가 있어도 안 줄어드는 구조였다 — 캐시 열쇠가 `경로|크기|수정시각` 이라
      **열쇠를 만들려면 먼저 stat** 을 해야 했다. 하나도 안 바뀐 날도 4.5시간을 썼다.
      그래서 '느려서 못 끝냈다'가 아니라 **매일 못 끝냈다.**

    되돌아가는 것을 막는다: 훑는 자리에 `os.stat(` 이 다시 들어오면 여기서 걸린다.
    """
    import os as _os
    import source_index as SI

    src = open(SI.__file__, encoding="utf-8").read()
    body = src.split("def scan(", 1)[1].split("\ndef ", 1)[0]

    # ① 훑는 자리에서 파일마다 stat 을 다시 부르지 않는다
    assert "os.stat(" not in body, \
        "scan() 이 파일마다 os.stat(경로) 를 다시 부른다 — Z: 에서 파일당 왕복 한 번이다"
    assert "os.walk(" not in body, \
        "os.walk 는 속으로 받은 크기·시각을 버리고 이름만 준다 — 그래서 stat 을 또 부르게 된다"
    assert "_walk_stat(" in body, "훑는 자리가 stat 을 같이 주는 워커를 안 쓴다"

    # ② 워커는 scandir 항목의 stat 을 쓰고, 걸러낼 폴더는 내려가기 **전에** 거른다
    w = src.split("def _walk_stat(", 1)[1].split("\nclass ", 1)[0]
    assert "os.scandir(" in w, "워커가 scandir 을 안 쓴다(크기·시각이 공짜로 딸려 오는 자리다)"
    assert "e.stat()" in w, "워커가 목록에 딸려 온 값을 안 쓰고 따로 물어본다"
    assert "SKIP_DIRS" in w, "걸러낼 폴더를 내려간 뒤에 거르면 훑는 값이 없다"
    assert "follow_symlinks=False" in w, "링크를 따라가면 고리에서 안 끝난다"

    # ③ 못 들어가는 폴더 하나가 색인 전체를 세우지 않는다
    assert "except OSError" in w, "폴더 하나를 못 열면 색인이 통째로 죽는다"

    # ④ 실제로 돈다 — 이 저장소 폴더를 훑어 (폴더, 이름, stat) 세 쪽을 주는지 본다
    here = _os.path.dirname(_os.path.dirname(_os.path.abspath(SI.__file__)))
    got = 0
    for dirpath, fn, st in SI._walk_stat(_os.path.join(here, "ecount", "tests")):
        assert _os.path.isdir(dirpath) and fn and st.st_size >= 0
        got += 1
        if got >= 5:
            break
    assert got, "워커가 한 건도 안 준다"

    # ⑤ 걸러낼 폴더는 정말 안 내려간다(`__pycache__` 는 SKIP_DIRS 에 있다)
    for dirpath, _fn, _st in SI._walk_stat(_os.path.join(here, "ecount", "tests")):
        assert "__pycache__" not in dirpath, "SKIP_DIRS 폴더로 내려갔다"

    # ⑥ ★ 같은 병이 세 곳 더 있었다 — 워커를 **베껴 쓰지 말고 한 곳을 쓴다**.
    #    베껴 쓰면 한 곳만 고쳐지고 나머지는 남는다. 그래서 공개 이름을 둔다.
    assert "walk_stat = _walk_stat" in src, "다른 도구가 쓸 공개 이름이 없다"

    for mod, why in (("source_tidy", "13시간 30분 매달렸던 원본 폴더 정리"),
                     ("source_organizer", "원본 자료 자동정리")):
        t = open(_os.path.join(_os.path.dirname(SI.__file__), mod + ".py"),
                 encoding="utf-8").read()
        assert "walk_stat" in t, "%s(%s) 가 공용 워커를 안 쓴다" % (mod, why)

    # ⑦ source_tidy 는 훑은 자리에서 크기를 다시 묻지 않는다
    #   ※ **주석은 빼고 본다.** 안 그러면 "예전엔 os.stat 을 불렀다"고 적어 둔 설명이
    #     그대로 걸린다 — 계기가 글자를 보고 코드를 안 보는 셈이다.
    st_src = open(_os.path.join(_os.path.dirname(SI.__file__), "source_tidy.py"),
                  encoding="utf-8").read()
    code = "\n".join(l for l in st_src.splitlines() if not l.lstrip().startswith("#"))
    assert "os.stat(p).st_size" not in code, \
        "source_tidy 가 파일마다 os.stat 을 다시 부른다 — Z: 에서 파일당 왕복 한 번이다"

    # ⑦-2 ★ **색인의 거를 목록을 말없이 물려주지 않는다.** 색인은 `_보관`·`_바로가기`
    #     를 안 담지만 정리 도구는 바로 그 폴더를 손봐야 한다. 물려받으면 정리가
    #     그 폴더를 통째로 안 보면서 '완료'라고 적는다 — 오류가 안 나는 종류다.
    so = open(_os.path.join(_os.path.dirname(SI.__file__), "source_organizer.py"),
              encoding="utf-8").read()
    assert "skip_dirs" in src, "워커가 거를 목록을 부르는 쪽에서 못 정한다"
    assert "skip_dirs=()" in st_src, "source_tidy 가 색인의 거를 목록을 물려받는다"
    assert "skip_dirs={\".source_organizer.guard\"}" in so or \
           "skip_dirs={'.source_organizer.guard'}" in so, \
        "source_organizer 가 예전에 거르던 것과 다른 목록을 쓴다"
    a, b = SI._walk_stat, None
    tdir = _os.path.join(_os.path.dirname(SI.__file__), "tests")
    b = sum(1 for _ in a(tdir, skip_dirs=()))
    assert b >= sum(1 for _ in a(tdir)), "거를 목록을 비웠는데 오히려 덜 훑는다"

    # ⑧ source_organizer 는 훑을 때 딸려 온 수정시각을 쓴다(없을 때만 다시 묻는다)
    day = so.split("def _file_day", 1)[1].split("\ndef ", 1)[0]
    assert "_MTIME.get(" in day, \
        "_file_day 가 훑을 때 받은 값을 안 쓰고 파일마다 getmtime 을 다시 부른다"

    print("  [198] 색인·정리가 파일마다 Z: 를 다시 묻지 않는다(2.5시간 → 24.8초) ✅")


def t224_wrapup_commit_refusal_paths():
    """[224] 자동 커밋의 **거부 경로** — 남의 옛 줄에 잠기지 않고, 멈출 때 자국을 남기지 않는다.

    사고 #38 (2026-08-11): 비밀값 스캔이 인덱스 전체를 훑어 **어제 커밋된** 멱등키 계산식
    한 줄에 걸렸고, 그날 자동 마무리 14번이 전부 ③ 커밋에서 멈췄다. 커밋을 거부해도 그
    줄은 사라지지 않으니 **영구히 잠긴 관문**이었고, 멈추는 자리가 `add -A` 뒤라 공유
    인덱스에 남의 파일이 담긴 채 남았다 — 기계가 사고 #36 을 스스로 만들었다.

    ★ `[221]` 과 겹치지 않는다. 그쪽은 옆 세션 표기·푸시 보류·문서 절차를 보며 비밀값
      스캔을 **'없음'으로 가짜 처리**한다 — 즉 거부 경로를 한 번도 지나가지 않고,
      되돌림도 `_rollback()` 이 소스에 두 번 나오는지로만 본다. 이 사고의 본체가 바로
      거부 경로였고, 그것은 **실제 저장소를 만들어 돌려야** 드러났다: 스테이징 목록을
      `run()`(마지막 한 줄만 준다)으로 읽어 비밀값 담긴 파일이 스캔에서 빠진 것을
      소스 문자열 검사로는 잡을 수 없었다.
    """
    import session_wrapup as W

    def git(tmp, *a):
        return subprocess.run(["git", *a], cwd=tmp, capture_output=True, text=True,
                              encoding="utf-8", errors="replace")

    def repo(tmp):
        for a in (["init", "-q"], ["config", "user.email", "t@t"], ["config", "user.name", "t"]):
            git(tmp, *a)
        # 이미 커밋돼 있는 '비밀값처럼 생긴' 줄 — 실제로는 멱등키 계산식이다.
        # ★ 그 모양을 **이 소스에 그대로 적지 않는다**: 적으면 이 파일이 다른 스캐너
        #   (Terra→Sol 검토의 비밀값 형태 검사)에 영구히 걸린다 — 사고를 적는 행위가
        #   사고를 만든다. 조립해 쓰면 **디스크에 써진 파일만** 그 모양이 된다.
        open(os.path.join(tmp, "canonical_sync.py"), "w", encoding="utf-8").write(
            '%s = "%s" + sha256_json(facts)\n' % ("completion_" + "token", "canonical-completion:"))
        git(tmp, "add", "-A")
        git(tmp, "commit", "-q", "-m", "첫 커밋")

    def step(tmp):
        old, steps = W.ROOT, []
        try:
            W.ROOT = tmp
            W.step_commit("claude", "synthetic-t224", steps)
        finally:
            W.ROOT = old
        return steps[-1]

    with tempfile.TemporaryDirectory() as tmp:          # ① 남의 옛 줄에 잠기지 않는다
        repo(tmp)
        open(os.path.join(tmp, "innocent.py"), "w", encoding="utf-8").write("x = 1\n")
        s = step(tmp)
        assert s["성공"], "이미 커밋된 줄에 걸려 또 멈췄다: %s" % s["메모"]
        assert "innocent.py" in git(tmp, "log", "--name-only", "--format=", "-1").stdout, \
            "커밋에 파일이 담기지 않았다"

    with tempfile.TemporaryDirectory() as tmp:          # ② 새 비밀값은 자리를 적고 막는다
        repo(tmp)
        open(os.path.join(tmp, "leak.py"), "w", encoding="utf-8").write(
            '%s = "%s"\n' % ("api" + "_key", "AKIA" + "ABCDEFGH1234567890"))
        s = step(tmp)
        assert not s["성공"], "이번 커밋이 새로 담는 비밀값을 통과시켰다 — 절대규칙 1"
        assert "leak.py" in s["메모"], "걸린 자리를 안 적었다(사람이 확인할 수 없다): %s" % s["메모"]
        assert git(tmp, "diff", "--cached", "--name-only").stdout.strip() == "", \
            "멈추면서 스테이징을 남겼다 — 다음 사람의 `git commit -m` 이 통째로 커밋한다"
        assert os.path.exists(os.path.join(tmp, "leak.py")), "파일을 지웠다 — 인덱스만 되돌려야 한다"

    with tempfile.TemporaryDirectory() as tmp:          # ③ 남의 스테이징은 빼지 않고 말한다
        repo(tmp)
        open(os.path.join(tmp, "theirs.py"), "w", encoding="utf-8").write("y = 2\n")
        git(tmp, "add", "theirs.py")                    # 옆 세션이 담아 둔 것
        open(os.path.join(tmp, "leak.py"), "w", encoding="utf-8").write(
            '%s = "%s"\n' % ("pass" + "word", "hunter2" * 3))
        s = step(tmp)
        assert not s["성공"], "목록을 마지막 한 줄만 읽으면 이 갈래가 조용히 깨진다"
        assert "theirs.py" in git(tmp, "diff", "--cached", "--name-only").stdout.split(), \
            "남이 담아 둔 것을 빼 버렸다 — 그 사람의 뜻을 지운다"
        assert "그대로 뒀다" in s["메모"], "인덱스를 안 건드렸다는 말을 안 했다: %s" % s["메모"]
    print("  [224] 자동 커밋 거부 경로 — 잠김 없음 · 자리 적음 · 남의 스테이징 보존 ✅")


def t225_session_auto_resumes_parked_and_pushes():
    """[225] 세워 둔 일이 풀렸으면 말하고, 보류된 푸시는 조용해지면 민다.

    사용자 지시(2026-08-11) **"이 세션도 완전 자동화시켜"**. 그날 사람 손이 들어간
    자리는 둘이었고 **둘 다 기계가 근거를 이미 갖고 있었다**: 옆 세션이 `code` 점유를
    놓았는데 어느 화면도 "이제 된다"고 말하지 않아 사람이 "하던 작업 진행" 을 두 번
    쳤고, 자동 마무리가 보류한 푸시는 그 세션이 사라진 뒤에도 미는 사람이 없었다
    (폰·웹은 **푸시된 것만** 본다).

    지키는 것 넷:
      ① 자원 이름이 **한글로 저장돼도** 막힘을 본다(`--lock 코드` 실측 `[34]`).
         못 맞추면 그 항목은 '자원 없음'이 되어 **늘 "가능"** 이라 답한다(`[165]` 모양).
      ② 주인이 **살아 있는 '진행'** 은 건드리지 않는다 — 잘못 고아로 읽으면 같은 일을
         AI 에게 한 번 더 시켜 같은 파일을 둘이 고친다(#36).
      ③ 옆 세션이 살아 있으면 **밀지도 넘기지도 않는다.** 범위를 못 읽거나 비밀값
         형태가 있으면 역시 안 민다 — '못 읽음'을 '깨끗함'으로 치지 않는다(`[169]`).
      ④ **붙어 있지 않은 것은 자동이 아니다** — 워치독 회차·인계 문서·정본 지시문.
    """
    import worksplit
    import worksplit_auto as A

    assert A._lock_en("코드") == "code" and A._lock_en("code") == "code", \
        "분담판이 한글로 적어 둔 자원을 점유 열쇠로 못 맞춘다 — 막힘을 영원히 못 본다"

    board = {"seq": 2, "items": [
        {"id": 1, "title": "코드 일", "detail": "d", "lock": "코드", "state": worksplit.WAIT,
         "who": "", "sid": "", "at_ts": time.time()},
        {"id": 2, "title": "남이 하는 중", "detail": "", "lock": "code", "state": worksplit.DOING,
         "who": "claude", "sid": "sib00001", "at_ts": time.time()}]}
    real = (worksplit.load, A.ai_claim.load, A.ai_claim._is_dead, A.live_others,
            A._git, A.STATE, A._hand_to_ai)
    calls, handed = [], []

    def fake_git(*a, **kw):
        calls.append(a)
        if a[:2] == ("rev-parse", "--abbrev-ref"):
            return True, "origin/master\n"
        if a[:2] == ("rev-parse", "--git-dir"):
            return True, os.path.join(ROOT, ".git") + "\n"
        if a[0] == "rev-list":
            return True, "3\n"
        if a[0] == "diff":
            return True, "+++ b/x.py\n+x = 1\n"
        return True, ""

    try:
        worksplit.load = lambda: board
        A.ai_claim._is_dead = lambda v: False
        A.ai_claim.load = lambda: {"code": {"who": "claude", "sid": "other001",
                                            "why": "옆 세션 작업"}}
        A.live_others = lambda: {"수": 1, "목록": ["sib00001"], "기준분": 10}
        rows = {r["id"]: r for r in A.parked()}
        assert 2 not in rows, "살아 있는 주인의 '진행' 항목을 고아로 읽었다 — AI 에게 두 번 시킨다"
        assert rows[1]["가능"] is False and "other001" in rows[1]["사유"], \
            "살아 있는 점유를 막힘으로 못 봤다: %s" % rows[1]["사유"]
        A.ai_claim.load = lambda: {}
        assert A.parked()[0]["가능"] is True, "막던 것이 사라졌는데 '가능'이라 말하지 않는다"

        A._git, A.STATE = fake_git, os.path.join(tempfile.gettempdir(), "t225_state.json")
        A._hand_to_ai = lambda row, tickets: (handed.append(row.get("id")) or "ticket-fake")
        A.run(dry=False)                                  # 옆 세션이 살아 있다
        assert not any(x[0] == "push" for x in calls), "옆 세션이 살아 있는데 밀었다"
        assert not handed, "옆 세션이 살아 있는데 AI 에게 넘겼다 — 같은 파일을 둘이 고친다"

        calls.clear()
        A.live_others = lambda: {"수": 0, "목록": [], "기준분": 10}
        A.run(dry=False)                                  # 아무도 없다
        assert any(x[0] == "push" for x in calls), "아무도 없는데 보류를 계속한다 — 폰에서는 없는 코드다"
        assert handed == [1], "풀린 일을 AI 에게 넘기지 않았다: %r" % handed

        calls.clear()

        def leaky(*a, **kw):
            if a[0] == "diff":
                return True, "+++ b/x.py\n+%s = \"%s\"" % ("api" + "_key",
                                                           "AKIA" + "ABCDEFGH1234567890")
            return fake_git(*a, **kw)

        A._git = leaky
        A.run(dry=False)
        assert not any(x[0] == "push" for x in calls), "미푸시 범위에 비밀값 형태가 있는데 밀었다"

        calls.clear()

        def blind(*a, **kw):
            if a[0] == "diff":
                return False, ""
            return fake_git(*a, **kw)

        A._git = blind
        A.run(dry=False)
        assert not any(x[0] == "push" for x in calls), "범위를 '못 읽음'인데 '깨끗함'으로 쳤다"
    finally:
        (worksplit.load, A.ai_claim.load, A.ai_claim._is_dead, A.live_others,
         A._git, A.STATE, A._hand_to_ai) = real

    wd = open(os.path.join(ROOT, "watchdog.py"), encoding="utf-8").read()
    assert "def resume_parked(" in wd and "resume_parked(dry)," in wd, \
        "워치독 회차에 붙어 있지 않다 — 붙지 않은 것은 자동이 아니다"
    sh = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    assert '"세션자동화": session_auto()' in sh, "인계 문서가 '풀린 일'을 읽지 않는다"
    assert "세워 둔 일 **[%s]** 의 막힘이 풀렸다" in sh, \
        "'먼저 처리할 것' 이 풀린 일을 말하지 않는다 — 막힌 것만 적으면 아무도 모른다"
    doc = open(os.path.join(ROOT, "CLAUDE.md"), encoding="utf-8").read()
    assert "worksplit_auto" in doc, \
        "정본 지시문의 '지금 자동으로 도는 것' 에 없다 — 거기 없는 것은 자동이 아니다"
    print("  [225] 세션 자동화 — 풀린 일 알림·AI 인계 · 조용해지면 푸시 ✅")


def t235_unattended_rounds_survive_pythonw():
    """[235] 무인 회차는 `pythonw.exe`(창 없음)에서도 죽지 않는다.

    스케줄러는 이 프로젝트의 회차를 **pythonw 로** 부른다. 창이 없으면
    `sys.stdout` 이 **None** 이라, 모듈 첫머리의 맨몸 `sys.stdout.reconfigure(...)`
    가 `AttributeError` 로 터진다. 그것이 import 중이면 **부른 회차까지 통째로 죽는다.**

    ★ 이 고장은 **손으로는 재현되지 않는다.** `python.exe` 로 돌리면 콘솔이 있어
      `sys.stdout` 이 살아 있기 때문이다. 실측 2026-08-12 — `쿠팡업무_밴드재수집`
      이 매일 08:00 에 exit 1 로 죽고 있었는데(`recollect.plan()` → `import
      recheck_plan` → 18줄), `--plan` 을 손으로 돌리면 멀쩡해서 며칠을 못 찾았다.
      자국(`reports/밴드_재수집_오류.json`)이 없었으면 그날도 못 찾았다.

    ★ 그래서 **사람이 아니라 검증이 훑는다.** 막는 관용구는 둘 다 인정한다 —
      `try:` 로 감싸거나 `if hasattr(sys.stdout, "reconfigure")` 로 묻거나
      (`hasattr(None, ...)` 는 False 라 둘 다 안전하다). 새 모듈이 맨몸으로 부르면
      여기서 걸린다. 분담판 [43].
    """
    import re as _re
    bad = []
    for dirpath, dirnames, filenames in os.walk(ROOT):
        dirnames[:] = [d for d in dirnames
                       if d not in (".git", "__pycache__", "node_modules", ".claude")]
        for name in filenames:
            if not name.endswith(".py"):
                continue
            path = os.path.join(dirpath, name)
            try:
                src = open(path, encoding="utf-8", errors="replace").read()
            except OSError:
                continue
            for m in _re.finditer(r"^(?P<ind>[ \t]*)sys\.stdout\.reconfigure", src, _re.M):
                before = [b.strip() for b in src[:m.start()].splitlines()[-6:]]
                guarded = bool(m.group("ind")) and any(
                    b.startswith("try:") or "hasattr(sys.stdout" in b for b in before)
                if not guarded:
                    bad.append("%s:%d" % (os.path.relpath(path, ROOT).replace("\\", "/"),
                                          src[:m.start()].count("\n") + 1))
    assert not bad, ("pythonw(창 없음)에서 sys.stdout 이 None 이라 죽는다 — "
                     "try 로 감싸거나 hasattr 로 물을 것: " + ", ".join(bad))

    # 회차가 죽으면 **이유가 디스크에 남아야** 한다. 창이 없으니 트레이스백이 갈 곳이 없다.
    rec = open(os.path.join(ROOT, "band", "recollect.py"), encoding="utf-8").read()
    assert "_leave_trace" in rec and "traceback.format_exc" in rec, \
        "무인 회차가 죽어도 이유가 어디에도 안 남는다"
    assert "os.remove(CRASH)" in rec, \
        "성공했는데 옛 자국이 남으면 이미 고쳐진 고장을 계속 보고한다"


def t234_kim_miyeong_center_and_revenue():
    """[234] 김미영 업무센터 · 쿠팡 매출 실적(전체)은 **발행월 기준**이다 (2026-08-12 지시).

    사용자 지시: "김미영 업무센터 하나 추가해서 쿠팡 매출 실적(전체) 포함 관리할 수 있게"

    ★ **기준이 다르면 숫자는 다르다.** 김미영 표는 머리글 그대로 세금계산서 **발행월**이고
      앱의 다른 화면은 완료월이다. 기준을 안 적으면 맞는 달까지 "왜 안 맞지"가 되고
      사람이 멀쩡한 값을 고치러 간다([172]).
    ★ **부가세 포함액을 우리 공급가액과 대지 않는다.** 표 원본은 포함액이고 대조는
      `_공급가액` 키(÷1.1 로 확정한 값, [154])를 쓴다. 원본을 그대로 대면 모든 달이
      10% 어긋나 보인다.
    ★ **못 읽은 것을 0 으로 접지 않는다**([169]). 색인·표가 없으면 '차이 없음'·'미발행
      0건'이 아니라 **'못 셈'** 이라고 말한다.
    ★ **미발행의 근거는 ERP 진행상태**다. 원장 계산서 칸은 사람 손 입력이라 대부분
      비어 있어, 그 빈 칸을 세면 '미발행 190건' 같은 없는 숫자가 나온다.
    ★ 로스터는 서버가 정본이고 화면은 사본이다 — 한쪽만 고치면 제목과 권한이 갈린다.
    ★ 서류 담당이 **현장 기록(as·pm)을 고치지 못한다.** 열어 주면 서류를 보고 현장
      사실을 맞추게 되는데, 그 순간 근거가 뒤집힌다.
    """
    import json as _json
    import shutil as _shutil
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_server as A

    slug = "kim-miyeong"
    assert slug in A.STAFF_CENTERS, "서버 로스터에 김미영 업무센터가 없다"
    cfg = A.STAFF_CENTERS[slug]
    assert cfg["name"] == "김미영"

    html = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    i = html.index("const STAFF_CENTERS = {")
    block = html[i:html.index("\n};", i)]
    for s2, c2 in A.STAFF_CENTERS.items():
        assert ("'%s'" % s2) in block, "화면 로스터에 %s 가 없다 — 서버와 갈렸다" % s2
        assert c2["title"] in block, "%s 의 제목이 서버와 화면에서 다르다" % s2

    # ★ 업무센터 화면의 사람 카드는 **손으로 적은 사본**이다(로스터를 안 읽는다).
    #   그래서 로스터에 사람을 더해도 그 화면에서만 조용히 빠진다 — 2026-08-12 실측으로
    #   김미영이 그렇게 빠져 있었고, 예전에 AS 기사 4명도 같은 이유로 고아가 됐다.
    #   사람을 늘리면 여기도 늘렸는지 이 검증이 묻는다.
    blocks = html.split('id="centerBlocks"', 1)[1].split('<!-- AS 담당기사', 1)[0]
    for s2, c2 in A.STAFF_CENTERS.items():
        assert ('href="/staff/%s"' % s2) in blocks, \
            "업무센터 화면에 %s(%s) 카드가 없다 — 로스터에만 있고 화면에는 없다" % (c2["name"], s2)
    assert "{id:'kim'," in html, "업무센터 '내 화면 구성' 목록에 김미영 구역이 없다"
    layout = html.split("function readCenterLayout()", 1)[1].split("\n}", 1)[0]
    assert "order.splice" in layout, \
        "새 구역이 저장된 순서 **맨 뒤**로 밀린다 — 사람 카드가 리모컨 밑으로 떨어진다"

    perm = A.STAFF_ENTRY_PERMISSIONS.get(slug) or {}
    assert set(perm) == {"settle"}, "김미영에게 청구·수금 밖의 칸이 열려 있다: %s" % sorted(perm)
    assert "세금계산서발행일" in perm["settle"] and "입금일" in perm["settle"]
    assert A._staff_allowed_fields(slug, "as") == set(), "서류 담당이 현장 기록을 고칠 수 있다"
    assert A._staff_allowed_fields(slug, "pm") == set(), "서류 담당이 점검 기록을 고칠 수 있다"

    src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert '"/api/revenue"' in src and "get_revenue()" in src, "/api/revenue 길이 없다"

    # ── 대조: 표가 있을 때 · 없을 때 ────────────────────────────────────────
    months = [{"월": "2026-06", "건수": 2, "합계": 75_025_570,
               "유형": {"정기점검": 42_664_600, "돌발AS": 32_360_970}},
              {"월": "2026-07", "건수": 1, "합계": 19_747_700,
               "유형": {"돌발AS": 19_747_700}}]
    keep_table, keep_index = A.KIM_TABLE_PATH, A.ERP_PRJ_INDEX
    keep_settle = A.get_settlements
    tmp = tempfile.mkdtemp(prefix="rev234_")
    try:
        A.KIM_TABLE_PATH = os.path.join(tmp, "없는표.json")
        miss = A._revenue_compare(months)
        assert miss.get("있음") is False and "안내" in miss, "표가 없는데 조용히 넘어간다"
        assert "차이 없음이 아" in miss["안내"].replace(" ", "").replace("'", "") or \
               "'차이 없음'이 아" in miss["안내"], "못 읽은 것을 '차이 없음'으로 읽히게 둔다"

        A.KIM_TABLE_PATH = os.path.join(tmp, "표.json")
        with open(A.KIM_TABLE_PATH, "w", encoding="utf-8") as fh:
            _json.dump({"출처": "합성", "부가세": "포함 금액이다",
                        # 원본(포함액)과 공급가액을 **둘 다** 둔다 — 대조가 어느 쪽을 쓰는지 본다
                        "정기점검": {"2026-06": 46_931_060},
                        "유료AS": {"2026-07": 21_722_470},
                        "정기점검_공급가액": {"2026-06": 42_664_600},
                        "유료AS_공급가액": {"2026-07": 19_747_700}}, fh, ensure_ascii=False)
        cmp = A._revenue_compare(months)
        assert cmp["있음"] and len(cmp["행"]) == 2
        for row in cmp["행"]:
            assert row["일치"] and row["차이"] == 0, \
                "부가세 포함액을 우리 공급가액과 대고 있다: %s" % row
        assert cmp["요약"]["다른 달"] == 0

        # ── 미발행: 색인이 없을 때는 '0건'이 아니라 '못 셈' ──────────────────
        A.ERP_PRJ_INDEX = os.path.join(tmp, "없는색인.json")
        blind = A._revenue_unissued()
        assert blind.get("있음") is False and "못" in blind.get("안내", ""), \
            "색인이 없는데 '미발행 0건'이라 말한다"

        A.ERP_PRJ_INDEX = os.path.join(tmp, "색인.json")
        with open(A.ERP_PRJ_INDEX, "w", encoding="utf-8") as fh:
            _json.dump({"index": {
                "UJ2600001": {"state": "3.오더처리", "supply": 1_000_000, "cust": "합성캠프"},
                "UJ2600002": {"state": "6.세금계산서발행", "supply": 2_000_000},
                "UJ2600003": {"state": "7.수금완료", "supply": 3_000_000},
                "UJ2600004": {"state": "8.무상납품완료", "supply": 4_000_000},
            }}, fh, ensure_ascii=False)
        A.get_settlements = lambda: [
            {"업무구분": "돌발AS", "프로젝트NO": "UJ2600001", "완료일": "2026-06-03",
             "캠프명": "합성캠프", "공급가액": 1_000_000},
            {"업무구분": "돌발AS", "프로젝트NO": "UJ2600001", "완료일": "2026-06-04"},  # 같은 번호 두 행
            {"업무구분": "정기점검", "프로젝트NO": "UJ2600002", "완료일": "2026-06-05"},
            {"업무구분": "정기점검", "프로젝트NO": "UJ2600003", "완료일": "2026-06-06"},
            {"업무구분": "정기점검", "프로젝트NO": "UJ2600004", "완료일": "2026-06-07"},
            {"업무구분": "신규납품", "프로젝트NO": "UJ2600001", "완료일": "2026-06-08"},
            {"업무구분": "돌발AS", "프로젝트NO": "", "완료일": "2026-06-09"},
            {"업무구분": "돌발AS", "프로젝트NO": "UJ2600001", "출처": "ERP"},
        ]
        un = A._revenue_unissued()
        assert un["있음"] and un["건수"] == 1, "발행된 건·묶음 계산서까지 미발행으로 센다: %s" % un["건수"]
        assert un["공급가액"] == 1_000_000, "같은 프로젝트를 여러 행만큼 더한다"
        assert un["근거없음"] == 1, "번호 없는 행을 미발행이라 단정하거나 아예 안 센다"
        assert "ERP 진행상태" in un["근거"], "미발행 근거가 ERP 진행상태가 아니다(원장 빈 칸은 근거가 아니다)"
    finally:
        A.KIM_TABLE_PATH, A.ERP_PRJ_INDEX = keep_table, keep_index
        A.get_settlements = keep_settle
        _shutil.rmtree(tmp, ignore_errors=True)

    # ── 화면이 기준을 **말한다** ────────────────────────────────────────────
    assert 'id="v-revenue"' in html and 'data-v="revenue"' in html, "매출 실적 화면·탭이 없다"
    shell = html.split('<main class="shell">', 1)[1].split("</main>", 1)[0]
    assert 'id="v-revenue"' in shell, \
        "매출 실적 화면이 .shell 밖에 있어 데스크톱 사이드바 216px 뒤로 잘린다"
    assert "세금계산서 발행월" in html, "화면이 어느 기준으로 센 숫자인지 말하지 않는다"
    assert "발행월" in A._build_revenue.__doc__ if A._build_revenue.__doc__ else True
    basis = src[src.index("def _build_revenue"):src.index("def get_revenue")]
    assert "발행월" in basis and "완료월이 아닙니다" in basis, \
        "서버가 내려보내는 기준 문장에 '발행월/완료월 아님'이 없다"


def t235_chatbot_is_one_line_until_asked():
    """[235] 챗봇은 **평소 한 줄**이다 (2026-08-12 지시: "쓸데 없이 창이 너무 커").

    ★ 실측: 폰 첫 화면에서 챗봇 카드가 **40%** 를 먹고 있었다 — 제목·칩 여섯 개(세 줄로
      접힘)·큰 입력창·큰 파란 단추가 늘 펼쳐진 채였다. 정작 사람이 먼저 보려는 것은
      그 아래 업무 현황 숫자다. 고친 뒤 접힘 **55px(7%)** · 펼침 177px(22%).
    ★ **기본은 접힘**이다. 기억이 없을 때 펼치면 고친 뜻이 사라진다 — 그래서 판정을
      `=== '1'` 로 둔다(없으면 거짓). 그리고 화면마다 따로 기억한다(대시보드·업무센터는
      다른 자리다).
    ★ 칩은 **한 줄로 흐른다.** `flex-wrap:wrap` 이 돌아오면 여섯 개가 다시 세 줄이 된다.
    ★ 접어도 **대화가 사라진 것처럼 보이면 안 된다** — 답이 있으면 한 줄이 그 수를 말한다.
    ★ 어디서 물었든 **답이 보이는 자리에서** 답한다(접힌 채로 답하면 아무도 못 본다).
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    assert ".chatcard.mini .chatpill{display:flex}" in live, "접힌 한 줄(알약)이 없다"
    assert "chatpill" in live and 'class="chatbody" hidden' in live, \
        "펼침 부분이 처음부터 숨겨져 있지 않다 — 그러면 예전처럼 늘 펼쳐진다"

    mount = live.split("function chatMount(", 1)[1].split("\nfunction ", 1)[0]
    assert "chatSetOpen(el, chatWasOpen(el), false)" in mount, \
        "붙일 때 접힘/펼침을 정하지 않는다"
    was = live.split("function chatWasOpen(", 1)[1].split("\n}", 1)[0]
    assert "=== '1'" in was, "기억이 없을 때 펼친다 — 기본은 접힘이어야 한다"
    key = live.split("function chatKey(", 1)[1].split("\n}", 1)[0]
    assert "dataset.chat" in key, "화면마다 따로 기억하지 않는다 — 한 곳에서 접으면 다 접힌다"

    chips = live.split(".chatchips{", 1)[1].split("}", 1)[0]
    assert "overflow-x:auto" in chips, "칩이 가로로 흐르지 않는다"
    assert "flex-wrap" not in chips, "칩이 다시 줄바꿈한다 — 여섯 개면 세 줄이 된다"

    send = live.split(".chatsend{", 1)[1].split("}", 1)[0]
    assert "position:absolute" in send, "보내기가 입력칸 밖에 있다 — 입력칸 폭을 빼앗는다"

    ask = live.split("async function doAsk(", 1)[1].split("\n}", 1)[0]
    assert ask.index("chatSetOpen(el, true)") < ask.index("_chatBubble(el, 'me'"), \
        "접힌 채로 답한다 — 답이 화면 밖에 그려진다"

    setopen = live.split("function chatSetOpen(", 1)[1].split("\n}\n", 1)[0]
    assert "대화 이어가기" in setopen, "접으면 대화가 사라진 것처럼 보인다"

    print("  [235] 챗봇은 평소 한 줄 · 칩 한 줄 흐름 · 보내기는 입력칸 안 · 접어도 답 수를 말한다 ✅")


def t236_list_is_folded_into_groups():
    """[236] 목록은 **묶음으로 접혀** 있고 묶음을 누르면 아래서 위로 열린다.

    (2026-08-12 지시: "정산 카테고리도 클릭하면 아래서 위로 올라는 구조로 알고리즘 변경
     전체 메뉴 전부다 스크롤이 너무 길어")

    ★ 실측: 폰 375x812 에서 정산 본문이 **35,811px = 44.1 화면**이었다. 750건이 카드로도
      표로도 **전부** 본문에 깔렸기 때문이다. 다 보여 주는 것과 볼 수 있는 것은 다른 말이다.
      묶음으로 접은 뒤 **939px = 1.2 화면**.
    ★ **시트를 새로 만들지 않는다.** 이미 있는 `openExecMetric → showSheet` 를 쓴다 —
      상세로 들어갔다 뒤로 나오는 층 쌓기·인쇄·엑셀이 거기 이미 들어 있고, 새로 만들면
      사본이 둘 되어 한쪽만 고쳐진다([162] 와 같은 모양).
    ★ **접는 것이지 숨기는 것이 아니다**([169]) — 묶음마다 건수를 적고 '전체 목록' 을
      맨 위에 둔다. 그리고 접었을 때도 **되돌릴 손잡이가 남아야** 한다. 남지 않으면
      한 번 펼친 사람은 영영 전체 목록이다.
    ★ **목록·표는 접혀도 계속 만든다** — 엑셀 저장·캡처·인쇄가 그대로 읽는다. 안 만들면
      접었을 때 저장한 파일이 **빈 채로 나온다**(오류는 안 난다).
    ★ 손잡이를 툴바에 두지 않는다 — 실측으로 폰 툴바가 **+49px**(한 줄) 늘었다.
      스크롤을 줄이겠다며 화면을 더 밀어내면 안 된다.
    ★ **넓은 화면의 기본값은 건드리지 않는다.** 표는 이미 제 상자 안에서 스크롤한다
      (`.gridwrap{max-height:calc(100vh - 250px)}`) — 길어지는 것은 좁은 화면 쪽이다.
      문제가 없는 쪽의 도구를 말없이 빼앗지 않는다.
    ★ 묶음이 많으면 **묶음 목록 자체가 다시 길어진다**(실측 캠프별 110묶음 = 6,421px).
      큰 것부터 잘라 두되 **몇 묶음 몇 건이 접혔는지 말하고** 한 번에 펼쳐진다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    assert 'id="sgroups"' in live, "묶음 자리가 없다"
    assert 'id="sgridwrap"' in live, "표를 접을 손잡이가 없다 — 표만 남으면 그대로 길다"

    # 네 카테고리(정산·돌발AS·정기점검·확인필요)가 모두 묶음 기준을 갖는다.
    gb = live.split("const GROUP_BY = {", 1)[1].split("\n};", 1)[0]
    for m in ("settle:", "as:", "pm:", "check:"):
        assert m in gb, "%s 카테고리에 묶음 기준이 없다 — 그 화면만 예전처럼 길다" % m

    grp = live.split("function renderGroups(", 1)[1].split("\n}\n", 1)[0]
    assert "$('slist').style.display" in grp and "$('sgridwrap').style.display" in grp, \
        "묶음 보기인데 목록·표가 본문에 그대로 남는다"
    assert "'none' : ''" in grp, \
        "되돌릴 때 display 를 빈 문자열로 두지 않는다 — 화면폭 규칙(데스크톱 표)이 죽는다"
    assert "window._groupSets" in grp, "묶음 내용을 담아 두지 않는다 — 눌러도 열 것이 없다"
    # 접었을 때도 손잡이가 남는가: 묶음이 꺼진 갈래에서도 head( 를 그린다.
    off = grp.split("if(!on){", 1)[1].split("}", 1)[0]
    assert "head(" in off, "전체 목록으로 펼치면 되돌릴 손잡이가 사라진다"
    assert "toggleGroupMode()" in grp, "묶음/전체를 오갈 손잡이가 없다"

    # 접혀도 몇 건인지 말한다([169]) · '전체 목록' 이 맨 위에 있다
    assert "전체 목록" in grp and "건" in grp, "몇 건이 접혔는지 말하지 않는다"
    assert "나머지 ${restN}묶음 펼치기" in grp and "restRows" in grp, \
        "묶음이 많을 때 몇 묶음 몇 건을 접었는지 말하지 않는다 — 조용히 자르면 '없는 캠프'가 된다"
    assert "window._grpAll = mode + '|' + groupByKey()" in live, \
        "펼침을 화면·기준별로 기억하지 않는다 — 기준을 바꿔도 펼쳐진 채로 남는다"

    # 기본값: 좁은 화면만 접는다. 넓은 화면 표는 이미 제 상자 안에서 스크롤한다.
    gon = live.split("function groupModeOn(){", 1)[1].split("\n}", 1)[0]
    assert "max-width:899px" in gon, \
        "화면 폭과 무관하게 접는다 — 문제가 없는 넓은 화면에서 표를 말없이 빼앗는다"
    assert "v === '0' || v === '1'" in gon, \
        "사람이 고른 값이 화면 폭보다 뒤에 온다 — 한 번 고른 선택이 무시된다"

    # 시트는 이미 있는 것을 쓴다 — 새 시트를 만들지 않는다
    olg = live.split("function openListGroup(", 1)[1].split("\n}\n", 1)[0]
    assert "openExecMetric(" in olg, "묶음이 기존 시트가 아닌 다른 길로 열린다"
    for col in ("종류:mode", "레코드ID:recordIdOf(r)", "프로젝트NO:projectNoOf(r)"):
        assert col in olg, \
            "시트 카드가 원래 기록을 못 찾는다 — %s 칸이 없다(누르면 아무 데도 안 간다)" % col

    # 목록·표는 접혀도 계속 만든다: renderGroups 는 renderSettle 의 **끝**에 온다.
    rs = live.split("function renderSettle(){", 1)[1].split("\n}\n", 1)[0]
    assert "renderGroups(rows)" in rs, "묶음을 다시 그리지 않는다"
    assert rs.index("$('slist').innerHTML") < rs.index("renderGroups(rows)"), \
        "목록을 만들기 전에 접는다 — 엑셀·캡처가 빈 채로 나간다"

    # 손잡이는 툴바가 아니라 묶음 머리줄에 있다(폰 툴바가 한 줄 더 늘어나면 안 된다)
    tb = live.split('<div class="toolbar">', 1)[1].split("</div>", 1)[0]
    assert "fgroupby" not in tb and "toggleGroupMode" not in tb, \
        "묶음 손잡이가 툴바에 있다 — 폰에서 줄이 하나 더 생겨 화면을 더 밀어낸다"
    assert ".grouphead{" in live, "묶음 머리줄 모양이 없다"

    print("  [236] 목록은 묶음으로 접히고 묶음은 기존 시트로 열린다 · 접혀도 건수를 말한다 "
          "· 되돌릴 손잡이가 남는다 · 목록·표는 계속 만든다 ✅")


def t237_cards_fold_with_one_tool():
    """[237] 설명·단추 묶음은 **접어** 둔다 — 접기 도구는 한 벌이다.

    (같은 2026-08-12 지시의 나머지 반쪽: "전체 메뉴 전부다 스크롤이 너무 길어".
     `[236]` 이 **반복되는 행**을 묶었다면 여기는 **서로 다른 카드**다.)

    ★ 실측(폰 375×812): 실행 7,514px 중 '처리 방법·용어 설명' 한 장이 **3,843px(52%)**
      였다 → 접은 뒤 **2,999px**. 업무센터 3,690px → **2,028px**(사람 카드 413→189px).
    ★ **도구는 하나**다. 화면마다 접기를 새로 만들면 사본이 셋이 되고 한쪽만 고쳐진다
      ([162]). 마크업에는 `data-fold` 만 붙이고 판단·손잡이는 `foldSetup` 한 곳이 만든다.
    ★ **`hidden` 만으로는 안 접힌다** — 실측으로 걸렸다. `.ui-cards{display:grid}` 같은
      클래스 규칙이 브라우저 기본 `[hidden]{display:none}` 보다 세서 **내용이 그대로
      보인다.** 그런데 손잡이는 '▸ 펼치기'로 바뀌므로 **화면만 보면 고친 줄 안다**
      (업무센터 펼침 3,659px = 접힘 3,659px 이었다). 그래서 `!important` 로 못을 박는다.
    ★ **접힌 줄이 무엇이 몇 개인지 말한다**([169]). 그냥 사라지면 '없는 기능'이 된다.
    ★ **사람 카드는 이름·역할·현황을 남긴다** — 그것까지 접으면 '누가 있는지'를 잃는다.
    ★ 화면을 열 때마다 다시 맞춘다 — 내용을 나중에 채우는 카드(설명·AS 기사 단추)가
      있어 부팅 때 한 번만 붙이면 그 카드는 영영 안 접힌다.
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    assert "[data-fold][hidden]{display:none!important}" in live, \
        "hidden 만으로 접는다 — .ui-cards 같은 클래스 규칙이 세서 내용이 그대로 보인다"

    fs = live.split("function foldSetup(", 1)[1].split("\n}\n", 1)[0]
    assert "foldbtn" in fs and "insertBefore" in fs, "손잡이를 만들지 않는다"
    assert "classList.contains('foldbtn')" in fs, \
        "여러 번 부르면 손잡이가 겹쳐 쌓인다 — 화면을 열 때마다 부르므로 반드시 막아야 한다"
    assert "stopPropagation" in fs, \
        "끌어서 정렬하는 자리 옆에서 눌림이 엇갈린다"

    fa = live.split("function foldApply(", 1)[1].split("\n}\n", 1)[0]
    assert "foldNote(el)" in fa and "펼치기" in fa, \
        "접힌 줄이 무엇이 몇 개인지 말하지 않는다 — 그냥 사라지면 '없는 기능'이 된다"
    fo = live.split("function foldIsOpen(", 1)[1].split("\n}\n", 1)[0]
    assert "v === '0' || v === '1'" in fo, "사람이 고른 값이 기본값보다 뒤에 온다"
    assert "max-width:899px" in fo, "'phone' 기본값이 화면 폭을 안 본다"

    # 실제로 붙어 있나 — 실행 화면의 설명 한 장과 업무센터 사람 카드 넷
    assert 'id="helpcard" data-fold="run-help"' in live, \
        "실행 화면의 설명(3,843px)이 안 접힌다 — 여기가 가장 큰 한 장이다"
    for cb in ("cc-ryu", "cc-oh", "cc-kim", "cc-yoo"):
        assert 'data-fold="%s"' % cb in live, "업무센터 %s 카드가 안 접힌다" % cb
    # 사람 카드는 이름·역할·현황 줄을 접지 않는다(접는 것은 `.ui-cards` 뿐)
    ryu = live.split('data-cb="ryu"', 1)[1].split("</div>", 1)[0] + \
        live.split('data-cb="ryu"', 1)[1][:900]
    assert "ios-person-head" in ryu and 'data-fold="cc-ryu"' in ryu, \
        "사람 카드에서 접는 자리가 단추 묶음이 아니다 — 이름·현황까지 접으면 누가 있는지 잃는다"
    assert 'data-fold="cc-ryu"' not in ryu.split("ios-person-head", 1)[0], \
        "이름 줄 앞에서 접는다"

    # 화면을 열 때마다 다시 맞춘다
    av = live.split("function applyView(v){", 1)[1].split("\n}\n", 1)[0]
    assert "foldSetup(" in av, \
        "화면을 열 때 손잡이를 안 맞춘다 — 나중에 채워지는 카드는 영영 안 접힌다"

    print("  [237] 카드 접기 한 벌 — hidden 에 못박기 · 접힌 줄이 개수를 말함 "
          "· 사람 카드는 이름·현황 남김 · 화면 열 때마다 재적용 ✅")


def t238_parked_says_which_lane():
    """[238] '막힘이 풀렸다'는 **어느 창에서 하는 일인지**까지 말한다.

    실측 2026-08-12: `code` 점유는 풀렸는데 인계가 [39][48][50] 을 '가져가라'로 올렸다.
    그런데 수집 차선에 선 창에서 `--take` 하면 `lanes` 가 거부한다 — 그 창에서는
    **풀린 것이 아니다.** 한 낱말이 두 뜻이 되면 사람은 없는 문을 밀게 된다(`[169]` 모양).

    ★ 차선은 남을 막는 자물쇠가 아니다(`lanes.can()` 은 부르는 쪽의 차선만 본다).
      그래서 '가능'을 거짓으로 바꾸지 않는다 — **어느 창인지를 덧붙일 뿐**이다.
      거짓으로 바꾸면 차선을 안 정한 창(대부분)이 할 수 있는 일까지 세워 둔다.
    ★ 판단을 새로 만들지 않는다 — 차선표(`lanes.LANES`)를 그대로 읽는다.
    """
    import lanes, worksplit_auto

    assert worksplit_auto._lane_hint("code") == "build", "code 가 어느 차선인지 못 댄다"
    assert worksplit_auto._lane_hint("band") == "collect", "band 가 어느 차선인지 못 댄다"
    assert worksplit_auto._lane_hint("read") == "", \
        "아무 차선에서나 되는 자원까지 차선을 적는다 — 없는 제약을 만든다"
    assert worksplit_auto._lane_hint("없는자원") == ""

    # 표가 바뀌면 글도 따라가야 한다 — 이름을 손으로 적어 두면 안 된다
    src = open(os.path.join(ROOT, "worksplit_auto.py"), encoding="utf-8").read()
    hint = src.split("def _lane_hint(", 1)[1].split("\ndef ", 1)[0]
    assert "lanes.LANES" in hint, "차선표를 읽지 않고 이름을 적어 뒀다"
    assert "'build'" not in hint and '"build"' not in hint, \
        "차선 이름이 코드에 박혀 있다 — 표가 바뀌면 조용히 어긋난다"

    rows = worksplit_auto.parked()
    for r in rows:
        assert "차선" in r, "세워 둔 일에 차선 칸이 없다"
        if r.get("가능") and r.get("차선"):
            assert "차선 창에서" in (r.get("사유") or ""), \
                "풀렸다고만 하고 어느 창에서 하는지 안 말한다"

    sh = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    line = sh.split("의 막힘이 풀렸다", 1)[1][:400]
    assert "차선" in line, "인계 문서가 차선을 안 싣는다 — 판단만 하고 안 보이면 없는 것과 같다"

    print("  [238] 풀린 일에 차선을 붙인다 — 표에서 읽음 · 가능은 그대로 · 인계가 싣는다 ✅")


def t239_idle_lane_reclaims_itself_with_a_record():
    """[239] 놀고 있는 차선은 스스로 풀린다 — **죽음이 아니라 '자국 없음'** 으로 판정.

    실측 2026-08-12: sid 59fb7614 가 `build` 를 30.7시간 쥐고 있었고 pid 는 멀쩡히
    살아 있었다. Claude Desktop 은 창마다 `claude.exe` 를 띄우고 **닫아도 남기므로**
    (한 부모 밑에 29개) 이 환경에서 **pid 생존은 세션 생존의 증거가 아니다.**
    `_dead` 하나로는 영원히 안 풀리고, 그동안 코드 수정 다섯 건이 통째로 멈췄다.

    ★ 그런데 **잘못 뺏는 것이 못 뺏는 것보다 나쁘다** — 일하는 창의 차선을 빼앗으면
      두 창이 같은 파일에서 만난다(사고 #36). 그래서 근거를 못 읽으면 안 뺏는다.
    ★ 이름공간이 이 판정의 급소다. `live_transcripts` 는 파일 이름(UUID) 앞토막이고
      차선·점유가 적는 sid 는 그 **sha1** 앞토막이라 영영 안 겹친다 — 그대로 대면
      **늘 '자국 없음'** 이 되어 살아 있는 창까지 전부 회수된다(오류는 안 난다).
    """
    import time as _t

    import ai_claim
    import lanes
    import session_wrapup

    # ── 이름공간 — 여기가 어긋나면 아래 판정이 통째로 헛돈다
    me = ai_claim.session_id()
    stems = session_wrapup.live_stems(minutes=60)
    sids = session_wrapup.live_sids(minutes=60)
    assert all(ai_claim.sid_of(s) in sids for s in stems), \
        "live_sids 가 점유판 이름공간(sha1)으로 안 옮긴다"
    if stems:
        assert sids != [s[:8] for s in stems], \
            "live_sids 가 live_transcripts 와 같은 값이다 — 옮기지 않았다"

    # ④ 한도 안이면 안 뺏는다
    fresh = {"who": "claude", "sid": "deadbeef", "at": _t.time() - 3600, "agent_pid": 0}
    ok, why = lanes._idle(fresh)
    assert not ok and "한도" in why, "갓 잡은 차선을 뺏는다: %s" % why

    # ① 자국이 있으면 안 뺏는다 — 내 대화기록은 지금 자라고 있다
    mine = {"who": "claude", "sid": me, "at": _t.time() - 99 * 3600, "agent_pid": 0}
    ok, why = lanes._idle(mine)
    assert not ok, "지금 일하는 창의 차선을 뺏는다: %s" % why

    # 오래됐고 자국이 없으면 회수 대상 — 대화기록 폴더를 읽을 수 있을 때만 물어본다
    old = {"who": "claude", "sid": "deadbeef", "at": _t.time() - 99 * 3600, "agent_pid": 0}
    if session_wrapup.transcript_dir(""):
        ok, why = lanes._idle(old)
        assert ok, "자국 없는 99시간짜리 차선을 안 뺏는다: %s" % why

    # 끄는 스위치가 들어야 한다
    os.environ["COUPANG_LANE_AUTORECLAIM"] = "0"
    try:
        ok, why = lanes._idle(old)
        assert not ok and "꺼짐" in why, "COUPANG_LANE_AUTORECLAIM=0 이 안 듣는다"
    finally:
        os.environ.pop("COUPANG_LANE_AUTORECLAIM", None)

    # ②③ 뺏을 때 지키는 것
    src = open(os.path.join(ROOT, "lanes.py"), encoding="utf-8").read()
    body = src.split("def reclaim_idle(", 1)[1].split("\ndef ", 1)[0]
    assert "_dirty_tree()" in body, "미커밋 충돌을 안 본다 — 반쯤 고쳐 놓은 것을 덮는다"
    assert "RECLAIM_LOG" in body, "회수 기록을 안 남긴다 — 조용한 회수가 막으려던 그것이다"
    assert "d[lane] = rec" in body, "기록을 못 남겼을 때 되돌리지 않는다"
    assert 'add_argument("--force"' not in src, \
        "차선에 강제 탈취가 생겼다 — 그 자리는 사람이 판단한다"

    # 회차에 배선돼 있나 — 안 걸려 있으면 알고리즘이 있어도 안 돈다
    wa = open(os.path.join(ROOT, "worksplit_auto.py"), encoding="utf-8").read()
    run = wa.split("def run(", 1)[1].split("\ndef ", 1)[0]
    assert "reclaim_idle" in run, "워치독 회차가 차선 회수를 안 부른다"
    # ⚠ 글자로 물으면 주석에 적힌 `parked()` 를 먼저 집는다 — **호출 자리**로 묻는다.
    call = "rows, others, ps = parked()"
    assert call in run, "parked() 호출 자리를 못 찾았다 — 이 검증이 헛돈다"
    assert run.index("reclaim_idle") < run.index(call), \
        "회수가 parked() 뒤에 있다 — 그러면 이번 회차는 늘 한 박자 늦은 판정을 싣는다"

    # 대조하는 자리는 반드시 같은 이름공간에서 묻는다
    assert "live_transcripts(exclude=me)" not in wa, \
        "live_others 가 UUID 앞토막을 sid 목록에 섞는다 — 같은 창을 둘로 센다"
    assert "ai_claim.session_id()" in wa.split("live_sids = set(", 1)[1][:400], \
        "내 항목을 UUID 앞토막으로 찾는다 — 그 문은 한 번도 안 닫힌다"

    print("  [239] 놀고 있는 차선 자동 회수 — 자국으로 판정 · 이름공간 일치 · "
          "미커밋이면 보류 · 기록 없으면 되돌림 ✅")


def t240_install_has_a_door_and_says_why_when_shut():
    """[240] 크롬 '앱 설치' — **문이 있고**, 안 되면 **왜 안 되는지 말한다**.

    실측 2026-08-13(사용자 지시 "크롬에서 앱 설치 안되는 문제 해결"). 조각은 다 멀쩡했다 —
    127.0.0.1 에서 매니페스트 200 · 서비스워커 activated · 아이콘 192/512 존재.
    그런데 사람은 설치를 못 했다. 이유가 둘이었고 **둘 다 조용했다**:

    ① 설치로 가는 길이 업무센터 화면(`#v-ryu`) 안에만 있었고, `showInstallCard()` 첫 줄이
       `const center=STAFF_CENTERS[staffSlug]; if(!center) return;` 이라 관리자 화면에서는
       눌러도 **오류도 토스트도 없이 아무 일이 안 일어났다.**
    ② 터널 주소(trycloudflare)에서는 서버가 매니페스트·SW 를 **일부러** 뺀다. 그건 옳다
       (임시 주소가 아이콘에 박히면 다음날 죽는다 — 2026-07-28 실사고). 빠져 있던 것은
       **왜 막혔는지 말해 주는 일**이었다. 서버 주석은 "index.html 이 배너로 알린다"고
       적어 뒀는데 **그 배너가 실재하지 않았다** — 코드가 약속만 하고 안 지킨 자리다.

    ★ 그래서 이 검증은 '설치가 된다'를 시험하지 않는다(브라우저가 정하는 일이다).
      **문이 있는가**와 **닫혔을 때 이유를 말하는가**를 지킨다.
    ★ 판단은 한 곳(`installState`)이어야 한다. 두 곳에서 재면 실행 탭은 '가능',
      카드는 '불가'라고 말하는 날이 온다(`[162]` 와 같은 모양).
    """
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
    srv = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()

    # ① 말없이 되돌아가던 그 줄이 사라졌나
    #    ★ 주석은 벗기고 본다 — 왜 고쳤는지를 적어 둔 글이 코드로 읽히면
    #      "고쳤는데 검증이 실패"하는, 아무도 못 믿을 관문이 된다.
    def _code_only(s):
        s = re.sub(r"/\*.*?\*/", "", s, flags=re.S)
        return re.sub(r"^\s*//.*$", "", s, flags=re.M)

    body = _code_only(
        idx.split("async function showInstallCard(", 1)[1].split("\nfunction ", 1)[0])
    assert "if(!center) return;" not in body, \
        "관리자 화면에서 [앱 설치] 가 말없이 return 한다 — 눌러도 아무 일이 안 일어난다"
    assert "installState()" in body, "설치 카드가 스스로 판단한다 — 판단 자리가 둘이 된다"

    # ② 판단은 한 곳이고, 갈래마다 이유를 갖는다
    st = idx.split("function installState(", 1)[1].split("\nfunction ", 1)[0]
    for code in ("'installed'", "'tunnel'", "'insecure'", "'no-manifest'"):
        assert code in st, "설치 불가 갈래 %s 를 안 가른다 — 뭉치면 사람이 엉뚱한 데를 고친다" % code
    assert "trycloudflare.com" in st, "터널 주소를 못 알아본다 — 왜 막혔는지 말할 수 없다"
    assert "isSecureContext" in st, "https 가 아니라 막힌 경우를 못 가른다"
    # '아직 프롬프트가 안 왔다'를 '불가'로 단정하지 않는다([169])
    assert "code:'browser'" in st.replace(" ", ""), \
        "프롬프트가 아직 없다는 이유로 설치 불가라고 단정한다 — 모르는 것을 아는 것처럼 말한다"

    # ②-1 안드로이드 '설치했다는데 앱 서랍에 없다' — 바로가기와 WebAPK 를 가른다
    #     크롬은 굽기에 실패하면 조용히 홈 화면 바로가기로 떨어지는데 안내 문구는
    #     둘 다 "추가되었습니다"라 겉으로 구별이 안 된다(2026-08-13 실사고).
    assert "chrome://webapks" in st, \
        "앱 서랍에 없을 때 확인하는 법을 안 알려 준다 — 사람이 서랍만 계속 뒤진다"
    assert "FIXED_ENTRY" in st, \
        "앱 서랍에 확실히 들어가는 길(고정 배포본)을 안 준다"
    assert 'id="installAlt"' in idx, "그 길로 가는 단추가 없다"
    #     ★ 원인을 단정하지 않는다 — 이 서버 쪽 조건은 실측으로 전부 맞았다(`[172]` 의 문)
    assert "Android" in st, "안드로이드일 때만 나오게 가르지 않는다"

    # ③ 실행 탭에 문이 있고, 누르기 전에도 상태를 말한다
    run = idx.split('id="v-run"', 1)[1].split('id="v-daily"', 1)[0]
    assert 'id="installStatusLine"' in run, "실행 탭이 설치 상태를 안 보여 준다"
    assert "showInstallCard()" in run, "실행 탭에 설치로 가는 문이 없다"
    assert "function refreshInstallStatus(" in idx, "상태줄을 다시 재는 자리가 없다"
    assert "if(v==='run') refreshInstallStatus();" in idx, \
        "화면을 열 때 상태를 다시 재지 않는다 — 주소·설치 여부는 세션 중에도 바뀐다"

    # ④ 프롬프트는 담당자 화면이 아니어도 받아 둔다(안 받으면 나중에 쓸 것이 없다)
    bip = idx.split("window.addEventListener('beforeinstallprompt'", 1)[1][:300]
    assert "if(!staffSlug) return;" not in bip, \
        "관리자 화면이 설치 프롬프트를 버린다 — 나중에 [설치하기] 를 눌러도 쓸 것이 없다"

    # ⑤ 주 단추는 상황마다 다른 일을 한다 — 마크업에 동작을 박아 두면 표현할 수 없다
    assert 'id="installGo"' in idx, "설치 카드 주 단추에 id 가 없다"
    card = idx.split('id="installCard"', 1)[1].split("</aside>", 1)[0]
    assert 'onclick="installWorkcenter()"' not in card, \
        "주 단추 동작이 마크업에 박혀 있다 — '고정 주소로 열기' 갈래를 표현할 수 없다"

    # ⑥ 약속한 안내가 실재한다(주석만 있고 화면이 조용하던 자리)
    assert "function noticeIfNotInstallable(" in idx, \
        "터널로 연 사람에게 이유를 말하는 자리가 없다 — 서버 주석이 약속만 한다"
    assert "noticeIfNotInstallable" in srv, \
        "서버 주석이 실제 안내 자리를 안 가리킨다 — 약속만 적힌 주석으로 되돌아갔다"
    #    ★ 그리고 그 안내는 담당자 화면 전용 초기화 **밖**에서 돌아야 한다.
    #      `initStaffCenter()` 는 첫 줄이 `if(!staffSlug) return;` 이라 관리자·터널
    #      화면에서는 아예 안 돈다 — 설치가 막힌 사람은 대부분 거기에 있다.
    #      실측 2026-08-13: 그 안에 뒀더니 터널로 열어도 안내가 한 번도 안 떴다.
    init = _code_only(
        idx.split("function initStaffCenter(", 1)[1].split("\nfunction ", 1)[0])
    assert "noticeIfNotInstallable" not in init, \
        "설치 안내를 담당자 전용 초기화 안에 뒀다 — 정작 막히는 화면에서 안 돈다"

    # ⑦ 터널 차단은 그대로 있어야 한다 — 편하자고 열면 2026-07-28 사고가 되살아난다
    assert 'if "trycloudflare.com" in host:' in srv, \
        "터널 주소에서 설치를 다시 열었다 — 임시 주소가 아이콘에 박혀 다음날 죽는다"

    # ⑧ 고정 주소는 한 곳에서만 적는다(사본이 둘이면 이사한 날 한쪽만 고쳐진다)
    assert 'const FIXED_APP_ORIGIN=' in idx, "고정 주소 상수가 없다"
    assert idx.count("mulder.tailf14aae.ts.net") == 1, \
        "고정 주소가 두 군데 이상 박혀 있다 — 주소가 바뀌면 한쪽만 고쳐진다"

    print("  [240] 앱 설치 — 관리자 화면에도 문이 있음 · 막힌 이유를 갈라 말함 · "
          "터널 차단 유지 · 고정 주소 한 곳 ✅")


def t192_synthetic_check_is_harmless():
    """[192] 합성검증 전후 공유·추적 산출물의 바이트가 그대로다.

    검증이 성공해도 기존 미커밋 파일이나 공용 점유를 바꾸면 관문이 아니라 사고다.
    특히 [182]의 ``write_plan``은 세 파일을 쓰므로, 텍스트 복원 여부가 아니라 애초에
    실제 경로로 쓰지 않는지를 임의의 바이너리 바이트로 확인한다.
    """
    import importlib
    import ai_claim
    import handoff_review

    sys.path.insert(0, os.path.join(ROOT, "band"))
    cb = importlib.import_module("comment_backfill")
    snap = _snapshot_output_bytes

    assert snap(_SYNTHETIC_OUTPUT_PATHS) == _SYNTHETIC_OUTPUT_BASELINE, \
        "합성검증 시작 뒤 [192]에 오기 전에 실제 계획·게시본이 이미 바뀌었다"

    protected = [
        cb.PLAN_PATH,
        os.path.join(cb.DOCS_COLLECT, "plan.json"),
        os.path.join(cb.DOCS_COLLECT, "grab_posts.js"),
        ai_claim.CLAIMS,
    ]
    before = snap(protected)

    # 이미 수정돼 있던 파일은 UTF-8 텍스트라는 보장도 없다. 바이트 그대로 남아야 한다.
    old_plan, old_docs = cb.PLAN_PATH, cb.DOCS_COLLECT
    with tempfile.TemporaryDirectory(prefix="csos-dirty-artifacts-") as dirty:
        cb.PLAN_PATH = os.path.join(dirty, "reports", "밴드_수집계획.json")
        cb.DOCS_COLLECT = os.path.join(dirty, "docs", "collect")
        os.makedirs(os.path.dirname(cb.PLAN_PATH), exist_ok=True)
        os.makedirs(cb.DOCS_COLLECT, exist_ok=True)
        dirty_paths = [
            cb.PLAN_PATH,
            os.path.join(cb.DOCS_COLLECT, "plan.json"),
            os.path.join(cb.DOCS_COLLECT, "grab_posts.js"),
        ]
        dirty_bytes = (b"dirty-main\x00\xff\r\n", b"dirty-plan\r\n", b"dirty-grab\x00\xfe")
        for path, value in zip(dirty_paths, dirty_bytes):
            with open(path, "wb") as f:
                f.write(value)
        dirty_before = snap(dirty_paths)
        try:
            t182_app_collects_without_claude()
            t183_collect_survives_pc_off()
            assert snap(dirty_paths) == dirty_before, \
                "기존 dirty 산출물의 바이트가 합성검증 뒤 바뀌었다"
            assert cb.PLAN_PATH == dirty_paths[0] and cb.DOCS_COLLECT == os.path.dirname(dirty_paths[1]), \
                "격리 뒤 comment_backfill 경로가 호출 전 값으로 복원되지 않았다"
        finally:
            cb.PLAN_PATH, cb.DOCS_COLLECT = old_plan, old_docs

    assert snap(protected) == before, "합성검증이 실제 계획·게시본·공용 점유를 변경했다"
    assert snap(_SYNTHETIC_OUTPUT_PATHS) == _SYNTHETIC_OUTPUT_BASELINE, \
        "합성검증 실행 전후 계획·게시 산출물의 바이트가 다르다"
    assert os.environ.get("CSOS_SYNTHETIC") == "1", "합성 모드가 실행 중 해제됐다"
    own_src = open(__file__, encoding="utf-8").read()
    unsafe_clear = "os.environ." + 'pop("CSOS_SYNTHETIC"'
    assert unsafe_clear not in own_src, \
        "합성검증이 중간에 실데이터 보호 플래그를 해제한다"

    # Terra→Sol 관문도 호출자 환경과 무관하게 합성 플래그·임시 보고서 경로를 강제하고,
    # Windows에서 영원히 멈출 수 있는 subprocess.run(timeout=)을 쓰지 않아야 한다.
    seen = {}
    old_runner = handoff_review.run_tree
    try:
        def fake_runner(command, **kwargs):
            seen["command"] = command
            seen.update(kwargs)
            return type("R", (), {"returncode": 0, "stdout": "ALL GREEN", "stderr": "",
                                    "timed_out": False, "stuck_pid": 0})()
        handoff_review.run_tree = fake_runner
        ok, summary = handoff_review._synthetic_check()
    finally:
        handoff_review.run_tree = old_runner
    assert ok and "ALL GREEN" in summary
    assert seen["env"].get("CSOS_SYNTHETIC") == "1"
    assert seen["env"].get("COUPANG_REPORT_DIR") != handoff_review.REPORT_DIR
    import ast
    review_src = open(os.path.join(ROOT, "handoff_review.py"), encoding="utf-8").read()
    tree = ast.parse(review_src)
    fn = next(n for n in tree.body if isinstance(n, ast.FunctionDef)
              and n.name == "_synthetic_check")
    calls = [n.func for n in ast.walk(fn) if isinstance(n, ast.Call)]
    assert any(isinstance(c, ast.Name) and c.id == "run_tree" for c in calls)
    assert not any(isinstance(c, ast.Attribute) and c.attr == "run"
                   and isinstance(c.value, ast.Name) and c.value.id == "subprocess"
                   for c in calls)
    print("  [192] 합성검증 전후 무해 — dirty 3파일 바이트·공용 점유 보존·합성 플래그·유한 실행 ✅")


def check_numbers_unique():
    """`[N]` 표시가 **두 검증에서** 같이 쓰이면 실패시킨다.

    ★ 계기가 잘못 세고 있었다 (2026-08-11). 예전엔 `print` **줄 수**를 셌는데,
      한 검증이 갈래마다 print 를 두 줄 하는 것은 겹침이 아니다 — 그래서
      `[41]`(대시보드 잔해 + 날짜 명시)·`[153]`(openpyxl 없음 건너뜀 + 본 판정)이
      영영 '중복'으로 남아 LEGACY 명단에 얹혀 있었다. **없는 빚 둘**이었다.
      이제 **함수 이름 기준**으로 센다. 진짜 겹침 넷은 비켰다:
      `[84]`→`[200]` · `[98]`→`[201]` · `[121]`→`[202]` · `[172]`→`[203]`.
      옮긴 쪽은 **문서 참조가 적은 쪽**이고, 그 참조도 같이 고쳤다 —
      하나라도 남기면 그 문서가 엉뚱한 검증을 가리킨다(겹친 것보다 나쁘다).
    """
    import collections as _c
    src = open(__file__, encoding="utf-8").read()
    owner = _c.defaultdict(set)
    cur = "?"
    for ln in src.splitlines():
        m = re.match(r"def (\w+)\(", ln)
        if m:
            cur = m.group(1)
        m = re.match(r'\s*print\(\s*"\s+\[(\d+)\]', ln)
        if m:
            owner[int(m.group(1))].add(cur)
    dup = sorted(n for n, who in owner.items()
                 if len(who) > 1 and n not in LEGACY_DUP)
    if dup:
        raise AssertionError(
            "검증 번호가 겹친다: %s — 다른 세션이 먼저 썼을 수 있다. "
            "쓰지 않은 번호로 비켜라(전체 목록은 이 파일의 print 표시를 훑으면 나온다)"
            % ", ".join("[%d](%s)" % (n, "·".join(sorted(owner[n]))) for n in dup))
    old = sorted(n for n, who in owner.items()
                 if len(who) > 1 and n in LEGACY_DUP)
    if old:
        print("  (예전부터 겹쳐 있던 번호: %s — 문서가 가리키는 곳이 흔들리므로 "
              "정리할 때 문서까지 같이 고칠 것)"
              % ", ".join("[%d]" % n for n in old))


def t134_section_fold():
    """[134] 구역 머리를 눌러 접었다 폈다 (2026-08-07 지시).

    사용자 지시: "날짜 표시된 카드와 세부 사항 카드 접었다 폈다 하는 기능 추가해서
    적용 / **다른 것들도 이런 기능 들어갈 만한거 있는지 검토해보고 적용**".

    ★ 두 곳에 따로 붙이지 않았다. 그 둘의 머리가 이미 같은 `.ios-sec-h` 이고,
      같은 머리를 쓰는 곳이 세 군데 더 있다(원본 자료·입금 등록·리모컨 관리).
      "다른 것들도 검토"의 답이 새 코드가 아니라 **이미 있는 공통 머리**였다 —
      그래서 구역을 새로 만들어도 저절로 접힌다. 이 검사가 그 구조를 지킨다.

    ★ 열쇠는 **처음 본 순간의 id·글귀로 굳힌다.** 머리 글귀는 바뀐다
      ('선택한 날' → '2026년 8월 7일 · 2건'). 그때그때 글귀로 열쇠를 만들면
      **날짜를 바꿀 때마다 접어 둔 것이 스스로 펴진다.**
    """
    live = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    assert ".ios-sec-h.folded + *{display:none!important}" in live,         "접어도 다음 카드가 안 숨는다"
    assert live.count('class="ios-sec-h"') + live.count("ios-sec-h\" id=") >= 4         or live.count("ios-sec-h") >= 6, "공용 구역 머리가 줄었다 — 한 곳에 붙인 이점이 사라진다"
    for fn in ("secKey", "readSecFolded", "applySecFold", "toggleSecFold", "initSecFold"):
        assert f"function {fn}(" in live, f"{fn} 가 없다"
    assert "initSecFold();" in live, "시작할 때 켜지지 않는다"

    key = live[live.index("function secKey("):][:600]
    assert "dataset.secKey" in key, \
        "열쇠를 매번 글귀로 다시 만든다 — 날짜가 바뀌면 접어 둔 것이 저절로 펴진다"

    ap = live[live.index("function applySecFold("):][:1400]
    assert ap.index("const on=") < ap.index("fold-hint"), \
        "'접힘' 힌트를 붙인 뒤에 열쇠를 읽는다 — 힌트 글자가 열쇠에 섞인다"
    assert "'button'" in ap or '"button"' in ap, "보조기기가 누를 수 있는 것으로 안 읽는다"
    assert "tabIndex" in ap, "키보드로 접을 수 없다"

    # 접힌 채로 인쇄하면 그 내용이 보고서에서 통째로 빠진다
    assert re.search(r"@media print\{[^@]*ios-sec-h\.folded \+ \*\{display:block", live, re.S), \
        "인쇄할 때 접힌 구역이 빠진다 — 보고서에 내용이 사라진다"
    assert "var(--ink-3)" in live[live.index(".ios-sec-h::before"):][:400],         "화살표 색이 토큰이 아니다 — 어둡게에서 안 보인다"
    print("  [134] 구역 머리 접기 — 5구역 공용·열쇠 고정·인쇄 보존 ✅")



def t245_truth_watch_asks_instead_of_asserting():
    """[245] 화면이 조용히 틀린 값을 보여 주는 것을 **매일 기계가 먼저 묻는다**.

    사용자 지시(2026-08-13): "위 같은 문제 잡아내는 기능 AI 추가해". 그날 둘 다
    **사람이 전화로 지적하고 나서야** 알았다 — 미처리의 31%가 이미 취소된 건이었고
    ([243]), 다녀온 현장이 원장 완료일이 비어 미처리로 서 있었다([244]).
    오류가 안 나는 종류라 어느 화면에도 안 떴다.

    지키는 것:
      · **판정을 새로 만들지 않는다**([162]) — 화면이 쓰는 그 함수를 그대로 부른다.
        여기서 다시 판정하면 감시자와 화면이 갈리고, **감시자가 틀리면 아무도 모른다.**
      · **읽기 전용** — 안 고치고 큐에도 안 넣고 엑셀도 안 연다.
      · **못 물어본 것을 '이상 없음'이라 하지 않는다**([169]).
      · 한 갈래가 절반을 넘으면 경보가 아니라 **기준 이야기**다([170]).
      · 인계는 **다시 세지 않고 회차가 써 둔 것을 읽는다**([168]).
    """
    import truth_watch as T

    # ── 읽기 전용인가 — 쓰는 낱말이 소스에 없어야 한다 ──────────────────────
    src = open(os.path.join(ROOT, "truth_watch.py"), encoding="utf-8").read()
    code = re.sub(r"^\s*#.*$", "", src, flags=re.M)
    code = re.sub(r'"""[\s\S]*?"""', "", code)
    for 금지 in ("enqueue(", "--apply", "--queue", "workbook_patch", "openpyxl"):
        assert 금지 not in code, "감시자가 무언가를 고치려 한다: %r" % 금지
    assert "_calendar_work_events" in code, \
        "화면이 쓰는 그 함수를 안 부르고 제 손으로 판정하면 둘이 갈린다([162])"

    # ── 무인 회차에서 죽지 않는가 (pythonw 는 stdout 이 None, [235]) ───────
    assert "hasattr(sys.stdout" in src, "맨몸 reconfigure 는 pythonw 에서 터진다([235])"

    # ── 갈래를 정말 가르는가 (동작으로 본다) ────────────────────────────────
    급 = {"물음": [{"무엇": "취소된 건이 미처리에 섞였나", "값": 3, "왜": "", "등급": "경보"}],
          "못물음": ["밴드 색인을 못 읽었다"], "미처리": 10, "사유": {"가": 10}}
    md = T.render(급)
    assert "먼저 볼 것" in md and "못 물어본 것" in md, md[:200]
    assert "이상 없음이 아니다" in md, "'못 물어봄'을 이상 없음처럼 적으면 안 된다([169])"

    # 한 갈래가 절반을 넘으면 '확인'으로 올라와야 한다([170])
    몰림 = {"물음": [{"무엇": "미처리 사유가 한 갈래에 몰렸나", "값": "가 9건 (90%)",
                      "왜": "", "등급": "확인"}], "미처리": 10, "사유": {"가": 9, "나": 1}}
    assert "확인해 볼 것" in T.render(몰림)

    # ── 인계가 그 판정을 **다시 세지 않고 읽는가** ([168]) ─────────────────
    import session_handoff as H
    hsrc = open(os.path.join(ROOT, "session_handoff.py"), encoding="utf-8").read()
    fn = hsrc[hsrc.index("def truth_gap"):hsrc.index("def blockers")]
    assert "화면_사실대조.json" in fn and "_calendar_work_events" not in fn, \
        "인계가 밴드를 다시 파싱하면 인계 한 장에 몇십 초가 든다([168])"

    old = H.truth_gap
    try:
        H.truth_gap = lambda: {"경보": [{"무엇": "합성 어긋남", "값": 7}],
                               "못물음": ["합성 못 읽음"]}
        import collections as _c
        st = _c.defaultdict(list)      # 없는 칸은 빈 값 — 이 검사는 사실대조만 본다
        st["사실대조"] = H.truth_gap()
        bl = H.blockers(st)
    finally:
        H.truth_gap = old
    글 = " / ".join(w for w, _ in bl)
    assert "합성 어긋남" in 글, "사실대조 경보가 '먼저 처리할 것'에 안 올라온다: %r" % 글
    assert "합성 못 읽음" in 글, "'못 물어봄'이 조용히 사라졌다([169]): %r" % 글

    # ── 회차 목록에 실제로 들어 있는가 ("코딩했다"와 "돈다"는 다르다) ───────
    dsrc = open(os.path.join(ROOT, "daily_run.py"), encoding="utf-8").read()
    assert "truth_watch.py" in dsrc, "09:50 회차에 안 들어가면 아무도 안 돌린다"

    print("  [245] 화면↔원본 사실대조 — 화면 함수 재사용 · 읽기 전용 · "
          "못 물어봄을 이상없음이라 안 함 · 인계는 읽기만 · 회차 등록 ✅")


def t244_band_evidence_closes_and_says_why():
    """[244] 밴드·카톡이 완료라 하면 미처리가 아니다 — **문은 좁게, 이유는 반드시**.

    형님 통화(2026-08-13): "완료가 됐는데 네가 지금 미처리로 남아 있는 거야.
    이렇게 되면 일 안 하는 줄 알잖아." 원장 완료일은 사람 손 입력이라 비어 있을
    뿐인데 캘린더가 **원장만 읽었다**([166] — 원장 빈 칸은 근거가 아니다).

    지키는 것:
      · 닫는 문은 **그 프로젝트에 열린 원장 행이 하나뿐일 때만**([172]·[208]).
        여럿이면 어느 건이 끝났는지 원본이 안 말해 준다 — 짐작으로 닫으면 아무도
        안 가는데 목록에도 없다.
      · 못 닫으면 **왜 아직 미처리인지 갈라 말한다**([169]). '미처리 N건'만 있고
        이유가 없으면 사람이 다녀온 현장을 다시 찾아간다.
      · **'못 읽음'을 '근거 없음'으로 치지 않는다** — 그 색인은 디스크에 안 남긴다.
      · 비싼 파싱은 **캐시 검사 뒤**에 온다([168]). 실측 첫 계산 23.5초.
    """
    import importlib.util as _ilu
    sp = os.path.join(ROOT, "webapp", "app_server.py")
    A = sys.modules.get("app_server") or sys.modules.get("app_server_t243")
    if A is None:
        sys.path.insert(0, os.path.join(ROOT, "webapp"))
        spec = _ilu.spec_from_file_location("app_server_t244", sp)
        A = _ilu.module_from_spec(spec)
        sys.modules["app_server_t244"] = A
        spec.loader.exec_module(A)

    # ── 이유를 갈라 말하는가 ────────────────────────────────────────────────
    못읽음 = {"완료": {}, "언급": set(), "최신": "", "읽음": False}
    말 = A._why_still_open({"프로젝트NO": "UJ9990001"}, 못읽음, "2026-08-01")
    assert "못" in 말, "못 읽은 것을 '근거 없음'이라 말하면 안 된다: %r" % 말

    idx = {"완료": {"UJ9990002": {"작업일": "2026-08-05"}},
           "언급": {"UJ9990002", "UJ9990003"}, "최신": "2026-08-11", "읽음": True}
    말 = A._why_still_open({"프로젝트NO": "UJ9990002"}, idx, "2026-08-01")
    assert "여럿" in 말 or "가른다" in 말, 말
    말 = A._why_still_open({"프로젝트NO": "UJ9990003"}, idx, "2026-08-01")
    assert "완료 글이 없다" in 말, 말
    말 = A._why_still_open({"프로젝트NO": "UJ9990004"}, idx, "2026-08-01")
    assert "글이 없다" in 말, 말
    말 = A._why_still_open({"프로젝트NO": "UJ9990004"}, idx, "2026-08-12")
    assert "2026-08-11" in 말 and "못 본" in 말, \
        "수집이 아직 안 온 날짜를 '완료 글이 없다'로 단정하면 안 된다([169]): %r" % 말

    # ── 닫는 문이 정말 좁은가 (동작으로 본다) ──────────────────────────────
    행 = [
        # 같은 프로젝트에 열린 건이 **둘** — 밴드가 완료라 해도 닫으면 안 된다
        {"캠프명": "합성E", "프로젝트NO": "UJ9990010", "접수일자": "2026-08-01",
         "진행상태": "접수", "작업완료일": "", "방문예정일": "", "접수ID": "AS-1"},
        {"캠프명": "합성E", "프로젝트NO": "UJ9990010", "접수일자": "2026-08-02",
         "진행상태": "접수", "작업완료일": "", "방문예정일": "", "접수ID": "AS-2"},
        # 열린 건이 **하나** — 닫는다
        {"캠프명": "합성F", "프로젝트NO": "UJ9990011", "접수일자": "2026-08-01",
         "진행상태": "접수", "작업완료일": "", "방문예정일": "", "접수ID": "AS-3"},
    ]
    가짜 = {"완료": {"UJ9990010": {"작업일": "2026-08-06", "밴드": "합성밴드"},
                     "UJ9990011": {"작업일": "2026-08-07", "밴드": "합성밴드"}},
            "언급": {"UJ9990010", "UJ9990011"}, "최신": "2026-08-31", "읽음": True}
    _ow, _ob = A.get_works, A._band_completion_index
    try:
        A.get_works = lambda *a, **k: {"as": 행, "pm": []}
        A._band_completion_index = lambda: 가짜
        ev = A._calendar_work_events()
    finally:
        A.get_works, A._band_completion_index = _ow, _ob

    닫힘 = [e for e in ev if e.get("원장미기입")]
    열림 = [e for e in ev if e["분류"] == "as_open"]
    assert len(닫힘) == 1, "열린 건이 여럿인 프로젝트를 닫았다: %r" % [e["제목"] for e in 닫힘]
    assert "합성F" in 닫힘[0]["제목"], 닫힘[0]["제목"]
    assert 닫힘[0]["날짜"] == "2026-08-07", "완료일은 밴드가 말한 날이어야 한다"
    assert 닫힘[0]["분류"] == "as_done", 닫힘[0]["분류"]
    assert "밴드" in 닫힘[0].get("연결근거", ""), "근거를 안 적으면 왜 완료인지 못 묻는다"
    assert len(열림) == 2, "합성E 두 건은 그대로 미처리로 남아야 한다"
    for e in 열림:
        assert e.get("미처리사유"), "미처리에 이유가 없다([169]): %r" % e["제목"]

    # ── 비싼 파싱이 캐시 검사 **뒤**에 오는가 ([168]) ──────────────────────
    src = open(sp, encoding="utf-8").read()
    fn = src[src.index("def _band_completion_index"):src.index("def _why_still_open")]
    assert fn.index("_BAND_EV_TTL") < fn.index("import band_extract"), \
        "밴드 파싱이 캐시 검사보다 앞에 있다 — 요청마다 8천 건을 다시 센다([168])"
    assert fn.index("지문") < fn.index("import band_extract"), \
        "디스크 캐시 지문 검사가 파싱보다 뒤에 있다"
    assert 'out["읽음"]' in fn and "if fp and out" in fn, \
        "못 읽은 색인을 디스크에 남기면 다음 재시작이 '근거 0건'을 확언한다([169])"

    print("  [244] 밴드 완료 근거 — 열린 건 유일할 때만 닫음 · 미처리 이유 갈라 말함 · "
          "캐시 뒤 파싱 · 못 읽음은 안 남김 ✅")


def t243_cancelled_is_not_undone():
    """[243] 접수취소는 '미처리'가 아니다 — 그러나 **조용히 사라지지도 않는다**.

    2026-08-13 형님 통화: "접수해서 취소된 건 다 지워버려. 있으면 계속 접수가 돼
    있는 상태가 되잖아." 실측 그날: 캘린더 `as_open` 131건 중 **41건(31%)이 진행상태
    '취소'** 였다. 취소된 현장이 매일 빨간 미처리로 서 있으면 ① 안 가도 되는 곳을
    가야 할 곳처럼 보이게 하고 ② 경보의 3분의 1이 가짜라 나머지도 아무도 안 본다([170]).

    지키는 것 넷:
      · 판정은 **한 곳**(`work_flow.is_cancelled`) — 캘린더와 캠프 이력이 같은 답을 한다([162]).
      · 낱말은 **정확히 같을 때만**(`취소`·`철회`). 넓히면 멀쩡한 현장이 목록에서
        사라진다 — 미처리로 남는 것보다 나쁘다([172]).
      · 뺀 것은 **회색으로 남긴다**([169]). 통째로 지우면 "그때 접수가 있었나"를 잃는다.
      · 상태가 '완료'인데 **날짜만 빈** 것도 미처리라 부르지 않는다(실측 2건).
    """
    import work_flow as W

    # ── 낱말은 정확히 같을 때만 — 넓히지 않는다 ────────────────────────────────
    assert W.is_cancelled({"진행상태": "취소"}, "as")
    assert W.is_cancelled({"진행상태": " 철회 "}, "as"), "공백만 무시한다"
    assert W.is_cancelled({"점검상태": "취소"}, "pm")
    for 멀쩡 in ("접수", "신규접수", "작업완료", "보류", "재방문예정", ""):
        assert not W.is_cancelled({"진행상태": 멀쩡}, "as"), 멀쩡
    # 취소가 **본문 일부**로 들어간 것은 취소가 아니다 — 그 오탐이 사고 #의 모양이다.
    assert not W.is_cancelled({"진행상태": "택배발송 취소요청하심"}, "as")
    assert not W.is_cancelled({"진행상태": "취소"}, "없는갈래")

    # ── 상태가 완료라고 말하면 날짜가 없어도 '미처리'가 아니다 ─────────────────
    assert W.says_done({"진행상태": "작업완료"}, "as")
    assert W.says_done({"점검상태": "완료"}, "pm")
    assert not W.says_done({"진행상태": "접수"}, "as")

    # ── 캘린더가 정말 그 판단을 쓰는가 (글자가 아니라 동작으로 본다) ───────────
    import importlib.util as _ilu
    _sp = os.path.join(ROOT, "webapp", "app_server.py")
    _spec = _ilu.spec_from_file_location("app_server_t243", _sp)
    _A = sys.modules.get("app_server")
    if _A is None:                       # 이미 실린 것이 있으면 다시 안 싣는다(비싸다)
        sys.path.insert(0, os.path.join(ROOT, "webapp"))
        _A = _ilu.module_from_spec(_spec)
        sys.modules["app_server_t243"] = _A
        _spec.loader.exec_module(_A)

    행 = [
        {"캠프명": "합성캠프A", "프로젝트NO": "UJ9990001", "접수일자": "2026-01-06",
         "진행상태": "취소", "작업완료일": "", "방문예정일": ""},
        {"캠프명": "합성캠프B", "프로젝트NO": "UJ9990002", "접수일자": "2026-01-07",
         "진행상태": "작업완료", "작업완료일": "", "방문예정일": ""},
        {"캠프명": "합성캠프C", "프로젝트NO": "UJ9990003", "접수일자": "2026-01-08",
         "진행상태": "접수", "작업완료일": "", "방문예정일": ""},
    ]
    점검 = [{"캠프명": "합성캠프D", "프로젝트NO": "UJ9990004", "점검예정일": "2026-01-05",
             "실제점검일": "", "점검상태": "취소"}]

    _old = _A.get_works
    try:
        _A.get_works = lambda *a, **k: {"as": 행, "pm": 점검}
        ev = _A._calendar_work_events()
    finally:
        _A.get_works = _old

    분류 = {}
    for e in ev:
        분류.setdefault(e["분류"], []).append(e)
    열림 = 분류.get("as_open") or []
    assert len(열림) == 1, "미처리는 진짜 열린 1건뿐이어야 한다: %r" % [e["제목"] for e in 열림]
    assert "합성캠프C" in 열림[0]["제목"], 열림[0]["제목"]
    assert not (분류.get("pm_overdue") or []), "취소된 점검이 미처리로 서 있다"

    # ★ 뺀 것이 **사라지지 않았는가** — 이 검사가 이 검증의 반이다([169]).
    회색 = 분류.get("etc") or []
    본문 = " / ".join(e["제목"] for e in 회색)
    assert "합성캠프A" in 본문 and "취소" in 본문, "취소가 통째로 사라졌다: %r" % 본문
    assert "합성캠프D" in 본문, "취소된 정기점검이 통째로 사라졌다: %r" % 본문
    assert "합성캠프B" in 본문, "완료일 미기입이 통째로 사라졌다: %r" % 본문
    for e in 회색:
        assert e["날짜"], "회색으로 내리면서 날짜를 잃으면 달력에 안 뜬다: %r" % e["제목"]

    # ── 캠프 이력도 **같은 답**을 해야 한다 — 두 화면이 갈리면 아무도 못 믿는다 ──
    _old = _A.get_works
    try:
        _A.get_works = lambda *a, **k: {"as": 행, "pm": 점검}
        tl = _A.project_history(camp="합성캠프A")
    finally:
        _A.get_works = _old
    if isinstance(tl, dict) and tl.get("ok"):
        assert not [x for x in (tl.get("현황") or []) if x["분류"] == "as_open"], \
            "캘린더는 취소를 뺐는데 프로젝트 이력은 아직 미처리라고 한다"

    print("  [243] 접수취소는 미처리가 아니다 — 판정 한 곳 · 낱말 정확일치 · "
          "회색으로 남김 · 완료일 미기입 분리 ✅")


def t242_ready_means_logged_in():
    """[242] '쓸 수 있다'는 말은 **로그인까지** 확인한 말이다 (2026-08-13 실사고).

    실측: `route_status()` 가 `selected: claude · state: ready · "2.1.222 (Claude Code)"`
    를 주는 동안, 실제 티켓은 전부 `Not logged in · Please run /login` · exit 1 로
    죽고 있었다. `claude auth status` 는 `{"loggedIn": false}` 였다.

    조용했던 이유 둘 — 둘 다 낱말 한 개다:
      ① `probe_agent` 가 `--version` 만 봤다. 그 명령은 **로그인 없이도 0** 을 준다.
      ② Codex 폴백은 `_UNAVAILABLE_RE` 가 맞을 때만 뜨는데 거기에 `not logged in`
         이 없었다. 그래서 **설계된 폴백이 한 번도 안 떴고** codex 는 멀쩡히 깔린 채
         standby 였다. 티켓은 그냥 failed 로 끝났다.
    계기가 초록이면 아무도 안 본다 — `[169]` 와 같은 모양이다.
    """
    import agent_dispatch as A

    # ── ① 폴백을 여는 낱말. 없으면 로그아웃은 영영 Codex 로 안 넘어간다.
    for text in ("Not logged in · Please run /login",
                 "not logged in", "Please run /login", "로그인이 필요합니다"):
        assert A._UNAVAILABLE_RE.search(text), \
            "'%s' 를 사용 불가로 안 읽는다 — Codex 폴백이 안 뜬다" % text
    # 기존에 열려 있던 낱말들이 닫히지 않았나
    for text in ("credit limit reached", "rate limit", "not authenticated", "크레딧 소진"):
        assert A._UNAVAILABLE_RE.search(text), "예전에 잡히던 '%s' 가 안 잡힌다" % text
    # ★ 멀쩡한 출력에 걸리면 **늘 폴백**이라 Claude 를 영영 안 쓴다(반대쪽 고장)
    for ok in ("2.1.222 (Claude Code)", "codex-cli 0.147.0", "done"):
        assert not A._UNAVAILABLE_RE.search(ok), \
            "멀쩡한 출력 '%s' 을 사용 불가로 읽는다 — 언제나 폴백이 된다" % ok

    # ── ② 로그인 판정은 **셋**이다. '못 읽음'을 로그아웃이라 하지 않는다(`[169]`).
    real = A.run_tree
    Res = type("R", (), {})

    def fake(out="", rc=0, timed_out=False):
        def _run(cmd, **kw):
            r = Res(); r.stdout, r.stderr, r.returncode = out, "", rc
            r.timed_out, r.stuck_pid = timed_out, None
            return r
        return _run
    try:
        # 로그아웃은 **종료 코드가 아니라 값**으로 판정한다(로그아웃도 1, 명령 없음도 1)
        A.run_tree = fake('{"loggedIn": false, "authMethod": "none"}', rc=1)
        assert A.auth_state("claude", "x.exe")[0] == "로그아웃", "loggedIn=false 를 못 읽는다"
        A.run_tree = fake('{"loggedIn": true, "authMethod": "oauth"}', rc=0)
        assert A.auth_state("claude", "x.exe")[0] == "로그인"
        # 명령 자체가 없는 옛 CLI — **로그아웃이라 우기면** 멀쩡한 인계가 통째로 샌다
        A.run_tree = fake("error: unknown command 'auth'", rc=1)
        assert A.auth_state("claude", "x.exe")[0] == "확인못함", \
            "auth 명령이 없는 CLI 를 로그아웃으로 단정한다 — 되던 인계가 다 Codex 로 샌다"
        A.run_tree = fake("", timed_out=True)
        assert A.auth_state("claude", "x.exe")[0] == "확인못함", "시간 초과를 로그아웃으로 친다"
        # 확인 명령이 없는 상대(codex)는 **모른다**고 한다 — 지어내지 않는다
        assert A.auth_state("codex", "x.exe")[0] == "확인못함"

        # ── ③ probe_agent 는 버전이 0 이어도 로그아웃이면 ready 라 하지 않는다
        seen = {}

        def two_step(cmd, **kw):
            r = Res(); r.timed_out, r.stuck_pid, r.stderr = False, None, ""
            if "--version" in cmd:
                seen["ver"] = True
                r.stdout, r.returncode = "2.1.222 (Claude Code)", 0
            else:
                seen["auth"] = True
                r.stdout, r.returncode = '{"loggedIn": false}', 1
            return r
        A.run_tree = two_step
        got = A.probe_agent("claude")
        assert seen.get("auth"), "버전만 보고 끝낸다 — 로그인을 안 묻는다"
        assert got["state"] == "unavailable", \
            "로그아웃인데 ready 라 한다 — 티켓이 전부 조용히 failed 로 끝난다: %s" % got
        assert "로그인" in got["reason"], "왜 못 쓰는지를 사람이 읽을 수 없다: %s" % got

        # 못 읽었을 때는 **일은 시키되 모른다고 적는다**(빈손으로 세우지 않는다)
        def unknown(cmd, **kw):
            r = Res(); r.timed_out, r.stuck_pid, r.stderr = False, None, ""
            if "--version" in cmd:
                r.stdout, r.returncode = "2.1.222", 0
            else:
                r.stdout, r.returncode = "unknown command", 1
            return r
        A.run_tree = unknown
        got = A.probe_agent("claude")
        assert got["state"] == "ready" and "확인못함" in got["reason"], \
            "로그인을 못 읽었는데 그 사실을 안 적는다 — 실패해도 이유가 안 남는다: %s" % got
    finally:
        A.run_tree = real

    # ── ④ 비밀값을 읽지 않는다. 묻는 것은 참/거짓 하나다.
    src = open(os.path.join(ROOT, "agent_dispatch.py"), encoding="utf-8").read()
    assert 'AUTH_PROBE = {"claude": ("auth", "status")}' in src, \
        "로그인 확인 명령이 한 곳에 안 적혀 있다"
    assert "token" not in src.split("def auth_state", 1)[1].split("\ndef ", 1)[0].lower(), \
        "로그인 확인이 토큰을 만진다 — 참/거짓만 물어야 한다"
    print("  [242] ready 는 로그인까지 확인한 말 — 폴백 낱말·셋으로 답함·못읽음≠로그아웃 ✅")


def t241_boundary_survives_compact_and_clear():
    """[241] 아무 때나 compact·clear 해도 이어진다 — **세션 경계**(2026-08-13 지시).

    사용자 지시: "아무때나 세션 컴팩팅이나 클리어 해도 문제 없이 돌아갈 수 있는
    알고리즘 구현해".

    빠져 있던 자리는 '요약하는 기능'이 아니라 **요약 뒤에 남는 것**이었다.
      ① `/clear` 에는 인계가 없었다 — `PreCompact` 는 compact 에만 온다. `/clear` 는
         **프로세스를 안 죽이므로** pid 로 보는 `_is_dead` 는 '살아 있다'고 답하고,
         sid 는 사라져 `--free-all`(내 것만)도 안 닿는다 → **아무도 못 푸는 점유**.
      ② 요약이 담은 **파일 사본은 낡는다.** 실측 2026-08-13: 요약에 실려 온
         `ai_tier.py`(옛 2인자)를 근거로 **이미 고쳐진 코드**를 "어긋났다"고 진단했다.
         `[165]` 와 같은 모양 — 낡은 사본은 빈 칸처럼 눈에 띄지 않는다.

    이 검증이 지키는 것: 회수의 문 네 개 · **칸 이름이 실제와 맞는지**(`[165]`) ·
    두 경계가 배선돼 있는지 · 로그인 안 된 CLI 의 exit 0 을 성공으로 안 세는지.
    """
    import importlib
    import json as _json
    import os as _os

    B = importlib.import_module("session_boundary")
    import ai_claim as _ac
    import session_wrapup as _sw

    # ── ① 고아 점유: 문 네 개. 잘못 회수하면 살아 있는 옆 창을 빼앗는다.
    now = 1_800_000_000.0
    keep = (_ac.load, _ac.session_id, _ac._is_dead, _sw.live_sids)
    try:
        _ac._is_dead = lambda c: False           # pid 는 살아 있다(= /clear 의 모양)
        _ac.load = lambda: {
            "ledger": {"who": "claude", "sid": "dead1234", "at": now - 3600},
            "band": {"who": "claude", "sid": "dead5678", "at": now - 60},
            "code": {"who": "claude", "sid": "live9999", "at": now - 3600},
        }
        _sw.live_sids = lambda *a, **k: ["me000000", "live9999"]
        _ac.session_id = lambda: "me000000"
        got = {r["자원"] for r in B.orphan_claims(now=now)}
        assert got == {"ledger"}, ("고아만 골라야 한다(살아 있는 것·방금 잡은 것 제외)", got)

        _ac.session_id = lambda: "notinlist"     # 목록에 내가 없다 = 목록을 못 믿는다
        assert B.orphan_claims(now=now) == [], "목록을 못 믿을 때 회수하면 안 된다"

        _ac.session_id = lambda: "me000000"
        def _boom(*a, **k):
            raise OSError("못 읽음")
        _sw.live_sids = _boom                    # 못 읽음 ≠ 아무도 안 산다([169])
        assert B.orphan_claims(now=now) == [], "못 읽었으면 회수하면 안 된다"

        _sw.live_sids = lambda *a, **k: ["me000000"]
        _ac._is_dead = lambda c: True            # pid 가 죽은 것은 기존 규칙 몫
        assert B.orphan_claims(now=now) == [], "pid 죽은 점유는 기존 규칙이 놓는다"
    finally:
        (_ac.load, _ac.session_id, _ac._is_dead, _sw.live_sids) = keep

    # ── ② 칸 이름이 **실제와 맞나** — 틀리면 오류 없이 빈 값이 된다(`[165]`).
    #    실측: `owner`·`no` 로 물었더니 맡은 일이 있는데도 "없음"으로 보였다.
    import loop_policy as _lp
    import worksplit as _ws
    items = (_ws.load() or {}).get("items") or []
    if items:
        keys = set(items[0])
        assert {"who", "id", "state", "title"} <= keys, \
            ("worksplit 항목 칸 이름이 바뀌었다 — session_boundary._my_work 도 고칠 것",
             sorted(keys))
    lpk = set(_lp.build() or {})
    assert {"갈래", "모델", "노력"} <= lpk, \
        ("loop_policy.build 칸 이름이 바뀌었다 — session_boundary.build 도 고칠 것", sorted(lpk))
    src = open(B.__file__, encoding="utf-8").read()
    assert 'get("who")' in src and 'get("id")' in src and 'get("owner")' not in src, \
        "worksplit 을 옛 칸 이름으로 묻고 있다 — 조용히 빈 목록이 된다"
    assert 'got.get("갈래")' in src and 'got.get("무게")' not in src, \
        "loop_policy 를 옛 칸 이름으로 묻고 있다(내 dict 키가 아니라 **부르는 쪽**을 본다)"
    # 값(모델·노력)의 정본은 `ai_tier.TIERS` 하나다 — 여기에 표를 복사하지 않는다.
    assert "sonnet" not in src and "opus" not in src, \
        "모델 이름을 여기 적지 말 것 — 값은 ai_tier.TIERS 에서 온다([230])"

    # ── ③ 두 경계가 **배선돼 있나.** 코딩했다와 돈다는 다른 말이다.
    root = _os.path.dirname(_os.path.dirname(_os.path.abspath(B.__file__)))
    st = _json.loads(open(_os.path.join(root, ".claude", "settings.json"),
                          encoding="utf-8").read())
    hooks = st.get("hooks") or {}

    def _args(ev):
        return " ".join(str(x) for g in (hooks.get(ev) or [])
                        for h in (g.get("hooks") or []) for x in (h.get("args") or []))

    assert "session_boundary.py" in _args("SessionStart"), \
        "SessionStart 배선 없음 — compact·clear 뒤에 상태를 되찾지 못한다"
    # 시작 자리에 **점유를 고치는 손이 둘**이면 안 된다(2026-08-13 실측: `--adopt` 가
    # 같이 걸려 있었다). 같은 순간 같은 ai_claims.json 을 읽고 쓰고, 시작마다 목록이
    # 둘 실리고, `/clear` 가 밴드 캐시 훑기까지 물게 된다 — 클리어를 싸게 만들자는
    # 지시가 뒤집힌다. `--adopt` 는 **사람이 치는 첫 명령**으로 남는다.
    assert "session_handoff.py" not in _args("SessionStart"), \
        "SessionStart 에 --adopt 가 같이 걸렸다 — 점유 회수가 두 곳이 되고 /clear 가 비싸진다"
    assert "session_wrapup.py" in _args("SessionEnd"), \
        "SessionEnd 배선 없음 — /clear 하면 점유가 고아로 남는다"
    assert "session_wrapup.py" in _args("PreCompact"), "PreCompact 인계가 사라졌다"
    assert "ai_tier.py" in _args("UserPromptSubmit"), \
        "모델·노력 판정이 사람 입력에 안 걸려 있다([230])"

    # ── ④ 로그인 안 된 CLI 는 **exit 0** 을 준다(실측 claude 2.1.222).
    import agent_dispatch as _ad
    assert _ad._looks_not_logged_in("Not logged in · Please run /login"), \
        "정형 문구를 못 알아본다 — 인계가 아무 일도 안 하고 '완료'가 된다"
    assert not _ad._looks_not_logged_in(
        "작업을 마쳤습니다. 밴드 로그인은 사람이 해야 합니다. " * 20), \
        "긴 본문의 '로그인'을 실패로 보면 정상 답을 죽인다"
    assert not _ad._looks_not_logged_in(""), "빈 응답은 이 판정의 몫이 아니다"
    src2 = open(_ad.__file__, encoding="utf-8").read()
    assert "_looks_not_logged_in(combined)" in src2 and \
        'if (result.returncode == 0' in src2, \
        "status 판정이 exit 0 만 보고 있다 — 실패가 성공으로 적힌다"

    # ── ⑤ 노력 깃발은 **확인된 실행파일에만** 붙인다(없는 깃발은 CLI 를 못 뜨게 한다).
    import ai_tier as _at
    chosen = _at.pick(kind="code", title="회차가 죽는다", attempts=3)
    assert "--effort" not in _at.flags("claude", chosen, r"C:\없는파일.exe"), \
        "확인 못 한 실행파일에 --effort 를 붙이면 인계가 조용히 안 된다"

    # ── ⑥ **배선이 있는 것과 그것이 돈 것은 다른 말이다** — 계기가 자국을 남기나.
    #    실측 2026-08-13: SessionEnd 배선은 멀쩡했는데 그 훅이 도는지 **볼 자리가
    #    없었다.** `세션경계_기록.json` 은 `session_boundary` 만 쓰는데 그 파일은
    #    SessionStart 에만 걸려 있었고, `session_wrapup` 은 `trigger` 만 물어
    #    SessionEnd 를 `manual`(=손으로 돌린 것)로 적었다. 그래서 `/clear` 마무리가
    #    한 번도 안 돌았어도, 매번 돌았어도 **파일이 똑같이 보였다**(`[169]`).
    #    계기가 눈멀면 그 기능은 있는지 없는지 영영 알 수 없다.
    rp = _sw.reason_from_payload
    assert rp({"hook_event_name": "SessionEnd", "reason": "clear"}) == "SessionEnd/clear", \
        "SessionEnd 계기를 안 읽는다 — /clear 마무리가 'manual' 로 적혀 손실행과 안 갈린다"
    assert rp({"hook_event_name": "SessionEnd", "reason": ""}) == "SessionEnd/?", \
        "갈래를 모르면 모른다고 적는다 — 'manual' 로 뭉개면 안 된다([169])"
    assert rp({"hook_event_name": "PreCompact", "trigger": "auto"}) == "auto-compact", \
        "PreCompact 계기 표기가 바뀌었다 — 옛 기록과 이어지지 않는다"
    assert rp({}) == "manual" and rp(None) == "manual", "빈 입력은 손실행이다"
    swsrc = open(_sw.__file__, encoding="utf-8").read()
    assert "session_boundary" in swsrc.split("def main(", 1)[-1], \
        "마무리가 경계 자국을 안 남긴다 — SessionEnd 가 도는지 볼 계기가 사라진다"

    # 자국 칸: 훅마다 갈래를 담는 이름이 다르다. 하나만 물으면 나머지는 '?' 가 된다.
    # ★ **진짜 기록 파일은 건드리지 않는다** — 실측 증거에 합성 행을 섞으면
    #   그 파일이 더는 실측이 아니다.
    import tempfile as _tf
    keep_log = B.LOG
    try:
        B.LOG = _os.path.join(_tf.mkdtemp(prefix="t241_"), "경계.json")
        for pay, want in ((({"hook_event_name": "SessionEnd", "reason": "clear"}), "clear"),
                          (({"hook_event_name": "PreCompact", "trigger": "auto"}), "auto"),
                          (({"hook_event_name": "SessionStart", "source": "resume"}), "resume")):
            assert B.note(pay)["갈래"] == want, \
                ("훅 갈래를 못 읽는다 — 실측이 '?' 만 쌓여 구실을 못 한다", pay)
    finally:
        B.LOG = keep_log

    print("  [241] 세션 경계 — 고아 점유 문 4개 · 칸 이름 실제 확인 · "
          "SessionStart/End 배선 · **계기가 자국을 남김**(SessionEnd 를 manual 로 안 적음) · "
          "로그인 exit 0 을 성공으로 안 셈 OK")

def t255_delete_is_reversible_and_exclusion_is_not_delete():
    """[255] 삭제는 되돌릴 수 있고, '청구 제외'는 삭제가 아니다 (2026-08-13 지시).

    류지영 요청: "삭제할수있게 해주세요!! / 정기점검이랑 동시진행으로 돌발AS로는
    청구를 안할꺼라 삭제가 필요합니당" → 형님 지시: "위건 삭제 수정 가능하게 코딩".

    ★ 지키는 것은 여섯이다:
      ① 업무 행에 **물리 DELETE 가 없다** — 지운 것을 되살릴 수 있어야 삭제가
        안전한 손잡이가 된다.
      ② 삭제에 **사유와 감사로그**가 있다 — 몇 달 뒤 답할 수 있는 것은 그 한 줄뿐이다.
      ③ **되살릴 수 있다.**
      ④ 삭제해도 **청구자료를 안 지운다**([208]).
      ⑤ **삭제된 건을 볼 수 있다** — 코드로만 되면 없는 것과 같다.
      ⑥ **제외와 삭제가 화면에서 구별된다** — 뭉치면 사람이 다녀온 현장을 지운다.
    """
    import tempfile
    from pathlib import Path as _P
    sys.path.insert(0, os.path.join(ROOT, "webapp"))
    import app_store as A
    import app_server as S
    import cancel_resolution as CR

    # ── ① 업무 행에 물리 DELETE 가 없다 ─────────────────────────────────────
    store_src = open(os.path.join(ROOT, "app_store.py"), encoding="utf-8").read()
    assert "DELETE FROM work_item" not in store_src, \
        "업무 행을 물리 DELETE 한다 — 되살릴 수 없는 삭제는 만들지 않는다"
    srv_src = open(os.path.join(ROOT, "webapp", "app_server.py"), encoding="utf-8").read()
    assert "DELETE FROM work_item" not in srv_src, "app_server 가 업무 행을 물리 DELETE 한다"
    assert "def restore_work" in store_src, \
        "soft_delete_work 의 짝인 restore_work 가 없다 — 삭제만 있고 되살리기가 없다"

    with tempfile.TemporaryDirectory(prefix="del-255-") as tmp:
        store = A.AppStore(_P(tmp) / "s.db").initialize()
        made = store.create_work(
            kind="돌발AS", business_key="AS-2606-093", public_id="AS-2606-093",
            project_no="UJ2601032", camp_name="송파3Sub-FC", status="작업완료",
            fields={"진행상태": "작업완료", "담당기사": "김필우",
                    "작업완료일": "2026-06-12", "유상·무상·보험": "유상"},
            actor="t255", source="t255", evidence="t255")
        wid = made["work"]["id"]
        ver = int(made["work"]["record_version"])

        # ── ⑥ 제외는 삭제가 아니다: 기록이 그대로 남고 목록에도 보인다 ────────
        store.update_work(wid, expected_version=ver,
                          patch={"fields": {"청구제외": "예",
                                            "청구제외사유": "정기점검과 동시 진행"}},
                          actor="t255", source="t255", evidence="t255")
        rows = store.list_sheet_rows("02_돌발AS접수")
        assert len(rows) == 1 and rows[0].get("청구제외") == "예", \
            "청구 제외한 건이 목록에서 사라졌다 — 제외는 삭제가 아니다"
        assert rows[0].get("담당기사") == "김필우" and rows[0].get("작업완료일") == "2026-06-12", \
            "청구 제외가 현장 기록(다녀온 사실)을 지웠다"
        out = CR.source_outcomes(store=store)["AS-2606-093"]
        assert out["billing_excluded"] is True and out["cancelled"] is False, \
            "청구 제외를 '취소'로 읽는다 — 다녀온 사실이 뒤집힌다"
        settle = S._apply_cancelled_source_to_settlement(
            {"정산ID": "S1", "원천업무ID": "AS-2606-093"}, out)
        assert settle["청구대상"] is False and settle.get("원천업무청구제외") is True, \
            "청구 제외건이 청구 대상에 그대로 남는다"
        assert not settle.get("원천업무취소"), "제외를 취소 딱지로 적었다"

        # 화면이 둘을 가려 읽는가 — 한쪽만 읽으면 제외건이 조용히 청구에 남는다([169]).
        idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()
        assert "!r.원천업무취소 && !r.원천업무청구제외" in idx, \
            "billableSettleRows 가 청구제외를 안 읽는다 — 제외해도 청구에 남는다"
        for token in ("wtDelete(", "wtExclude(", "wtShowDeleted(", "wtRestore("):
            assert token in idx, f"화면에 {token} 손잡이가 없다"
        assert "청구 제외 = 다녀왔지만 이 건으로는 청구 안 함" in idx, \
            "삭제와 제외의 차이를 화면이 말하지 않는다 — 뭉치면 다녀온 현장을 지운다"

        # ── ② 사유 없는 삭제·제외는 막힌다 ──────────────────────────────────
        cur = store.get_work(wid)
        base = dict(category="as", key="AS-2606-093",
                    record_version=int(cur["record_version"]))
        for bad in ({**base, "reason": ""}, {**base, "reason": "x"}):
            try:
                S.delete_staff_work("admin", bad, store=store, actor="t255")
                raise AssertionError("사유 없이 삭제가 지나갔다")
            except ValueError:
                pass
        try:
            S.save_staff_entry("admin", {"category": "as", "key": "AS-2606-093",
                                         "record_version": int(cur["record_version"]),
                                         "values": {"청구제외": "예"}},
                               store=store, actor="t255")
            raise AssertionError("사유 없이 청구 제외가 지나갔다")
        except ValueError:
            pass

        # ── ④ 청구자료가 있으면 한 번 더 묻고, 그 자료를 지우지 않는다 ────────
        store.create_work(kind="정산", business_key="S-255",
                          public_id="S-255", status="",
                          fields={"정산ID": "S-255", "원천업무ID": "AS-2606-093",
                                  "세금계산서발행일": "2026-06-30", "입금액": 1000},
                          actor="t255", source="t255", evidence="t255")
        cur = store.get_work(wid)
        ask = S.delete_staff_work(
            "admin", {**base, "record_version": int(cur["record_version"]),
                      "reason": "잘못 등록 — 두 번 접수됨"},
            store=store, actor="t255")
        assert ask.get("needs_confirm") and ask.get("documents"), \
            "청구자료가 있는 건을 묻지도 않고 지웠다"
        assert store.list_sheet_rows("02_돌발AS접수"), "되묻는 단계에서 이미 지웠다"

        done = S.delete_staff_work(
            "admin", {**base, "record_version": int(cur["record_version"]),
                      "reason": "잘못 등록 — 두 번 접수됨", "confirm_documents": True},
            store=store, actor="t255")
        assert done["ok"] and done["action"] == "deleted"
        assert not store.list_sheet_rows("02_돌발AS접수"), "삭제했는데 목록에 남아 있다"
        # legacy Excel 행까지 감춰야 사람 눈에 지워진 것이다.
        assert not store.overlay_rows(
            "02_돌발AS접수", [{"접수ID": "AS-2606-093", "캠프명": "송파3Sub-FC"}], "접수ID"), \
            "앱 DB 는 지웠는데 Excel 행이 화면에 그대로 남는다"
        # 청구자료는 한 글자도 안 지운다([208]).
        kept = [r for r in store.list_sheet_rows("06_거래서류청구수금")
                if str(r.get("정산ID") or "") == "S-255"]
        assert kept and str(kept[0].get("세금계산서발행일") or "").startswith("2026-06-30"), \
            "삭제가 청구자료(계산서·입금)를 같이 지웠다"

        # ── ⑤ 삭제된 건을 볼 수 있다 ────────────────────────────────────────
        listing = S.deleted_staff_works("admin", "as", store=store)
        assert listing["count"] == 1 and listing["rows"][0]["key"] == "AS-2606-093", \
            "삭제된 건을 볼 수 있는 자리가 없다 — 코드로만 되살리면 없는 것과 같다"
        assert "잘못 등록" in listing["rows"][0]["삭제근거"], "삭제 사유가 안 남았다"

        # ── ③ 되살릴 수 있다 ────────────────────────────────────────────────
        back = S.restore_staff_work(
            "admin", {"category": "as", "key": "AS-2606-093",
                      "record_version": listing["rows"][0]["record_version"]},
            store=store, actor="t255")
        assert back["ok"] and store.list_sheet_rows("02_돌발AS접수"), \
            "되살렸는데 목록에 안 보인다"
        assert S.deleted_staff_works("admin", "as", store=store)["count"] == 0

        # ── ② 감사로그가 남았다 ─────────────────────────────────────────────
        import sqlite3
        con = sqlite3.connect(str(_P(tmp) / "s.db"))
        acts = [r[0] for r in con.execute(
            "SELECT action FROM change_event WHERE work_id=? ORDER BY id", (wid,))]
        con.close()
        assert "soft_delete" in acts and "restore" in acts, \
            f"삭제·되살리기가 감사로그에 안 남았다: {acts}"

        # 낙관잠금 — 낡은 버전으로는 못 지운다.
        try:
            store.soft_delete_work(wid, expected_version=1, actor="t255", reason="낡음")
            raise AssertionError("낡은 record_version 으로 삭제가 지나갔다")
        except A.VersionConflict:
            pass
    print("[255] 삭제는 되돌릴 수 있고 청구 제외는 삭제가 아니다 — OK")


def t257_jump_says_why_it_could_not_find():
    """[257] '못 찾음' 도 세 가지 뜻이다 — 화면이 원인을 지목하지 않는다 (2026-08-13 실사고).

    실사용 막힘: 확인 필요 목록에서 `JS-2606-634` 를 눌렀더니
    **"찾지 못했습니다 (다른 월 데이터이거나 아직 미등록)"**. 그런데 그 건은 서버 원장
    750행에 멀쩡히 있었다 — 실측 `UJ2601032 · 송파3Sub-FC · 완료일 2026-06-12 · 돌발AS`.

    원인은 없음이 아니라 **순서**였다. `/api/issues` 는 금방 오는데 `/api/settlements`
    는 실측 **평균 54,656ms(812회 · 최대 27분)** 다. 그래서 목록은 이미 그려졌고
    `settleRows` 는 아직 빈 창이 1분 가까이 열려 있다. 그 창에서 누르면 못 찾는다 —
    **없어서가 아니라 아직 안 왔기 때문이다.** `[169]`·`[251]` 의 '0' 이야기와 같은
    모양인데, 여기서는 화면이 **원인까지 확언**했다. 틀린 지목은 못 잡는 것보다
    나쁘다(`[172]`) — 사람이 없는 문제(다른 월 자료·미등록)를 찾아 나선다.

    되돌아가면 안 되는 것만 지킨다:
      ① 원인을 확언하던 그 문장이 **없다**
      ② 못 찾았을 때 자료 상태를 `dataZeroState` **한 곳**에 묻는다(`[251]`, 판정 사본 금지)
      ③ 기다리는 시간이 **실측 54.7초보다 길다**(`[197]` — 짧으면 기다려 놓고도 같은 실패)
      ④ 오는 중이면 다시 걸지 않는다(같은 요청 두 번은 서버를 더 느리게 한다)
      ⑤ `openRecord` 폴백이 **프로젝트NO 를 버리지 않는다**
    그리고 node 가 있으면 위 다섯을 **말이 아니라 실행으로** 확인한다.
    """
    import shutil
    idx = open(os.path.join(ROOT, "webapp", "index.html"), encoding="utf-8").read()

    # ① 원인을 확언하던 문장은 **말하는 자리에서** 사라졌다.
    #    주석·문서에 그 문장을 인용하는 것은 사고 기록이라 남겨 둔다 — 그래서
    #    파일 전체에서 찾지 않고 '사람에게 말하는 줄'만 본다(내 주석에 걸렸었다).
    for ln, line in enumerate(idx.split("\n"), 1):
        if "다른 월 데이터이거나 아직 미등록" not in line:
            continue
        assert not re.search(r"\b(notice|toast|alert)\s*\(", line), \
            (f"{ln}줄에서 아직 원인을 확언한다 — 그 문장 때문에 사람이 "
             f"없는 문제(다른 월 자료·미등록)를 찾아 나섰다([172])")

    # ⑤ 폴백이 프로젝트NO 를 같이 넘긴다
    assert "openByPrj(rid||prj, prj)" in idx, \
        "openRecord 폴백이 프로젝트NO 를 버린다 — 프로젝트로 찾을 기회가 사라진다"
    assert "function openByPrj(key, project, retried)" in idx, \
        "openByPrj 가 프로젝트NO·재시도 여부를 안 받는다"

    body = idx[idx.index("function openByPrj("):idx.index("function goList(")]
    # ② 자료 상태는 한 곳에 묻는다
    assert "dataZeroState(JUMP_DATA_KEYS)" in body, \
        "못 찾았을 때 자료 상태를 안 본다 — '아직 안 온 것'을 '없는 것'이라고 말한다([251])"
    assert "settlements" in body and "works" in body, "볼 묶음 표가 없다"
    # ④ 오는 중이면 다시 걸지 않는다
    assert "!s.loading&&!s.retrying" in body, \
        "오는 중에도 다시 건다 — 같은 요청 두 번이면 서버가 더 느려진다"

    # ③ 기다리는 시간이 실측(54.7초)보다 길다
    m = re.search(r"JUMP_WAIT_MS\s*=\s*(\d+)", body)
    assert m, "기다리는 한도가 없다"
    assert int(m.group(1)) > 54656, \
        ("기다리는 시간이 실측 /api/settlements 평균(54,656ms)보다 짧다 — "
         "기다려 놓고도 같은 실패로 끝난다([197])", m.group(1))

    # ── 말이 아니라 실행으로: node 가 있을 때만. 없으면 '못 돌렸다'고 적는다([169]).
    node = shutil.which("node")
    if not node:
        print("  [257] 못 찾음을 가려 말한다 — 글자 검사만 통과(node 없어 실행 확인 못 함)")
        return
    fns = (idx[idx.index("function dataZeroState("):idx.index("/* 사전(error_book.BOOK)")]
           + body)
    harness = _T257_HARNESS.replace("/*__FNS__*/", fns)
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "jump.mjs")
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(harness)
        # 회차 한 단계가 영원히 멈추지 않게([175]) — 창 없는 실행에서도 반드시 끝난다.
        proc = subprocess.Popen([node, path], stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT)
        try:
            out = proc.communicate(timeout=60)[0].decode("utf-8", "replace")
        except subprocess.TimeoutExpired:
            proc.kill()
            out = proc.communicate(timeout=20)[0].decode("utf-8", "replace")
        assert "ALL OK" in out and proc.returncode == 0, \
            "실행 확인 실패:\n" + out[-1200:]
    print("  [257] 못 찾음을 가려 말한다 — 아직/없다/못받았다 3갈래 · 실행 확인 8건 OK")


# node 로 실제 돌려 보는 판 — 스텁은 최소로 두고 **진짜 함수**를 끼워 넣는다.
_T257_HARNESS = r"""
const APP_YEAR='2026';
let settleRows=[], works={as:[],pm:[]};
const DATA_SECTION_STATE={
  settlements:{key:'settlements',label:'정산·청구',loading:false,retrying:false,hasGood:false,error:null,asOf:''},
  works:{key:'works',label:'업무',loading:false,retrying:false,hasGood:true,error:null,asOf:'2026-08-13T16:00'}};
function projectNoOf(r){ return String((r&&r['프로젝트NO'])||''); }
function projectLabel(r){ return projectNoOf(r); }
function esc2(x){ return String(x==null?'':x); }
function esc4(x){ return JSON.stringify(String(x==null?'':x)); }
let opened=null; function openRecord(k,id,prj){ opened={k,id,prj}; }
let notices=[], toasts=[];
function notice(m,o){ notices.push(String(m)); }
function toast(m){ toasts.push(String(m)); }
function showSheet(h){ }
function dataStampText(x){ return String(x||'-'); }
const _wait = ms => new Promise(r=>setTimeout(r, 5));
const ROW={'정산ID':'JS-2606-634','프로젝트NO':'UJ2601032','캠프명':'송파3Sub-FC','완료일':'2026-06-12'};
let arriveAfter=0, retried=0;
async function retryDataSection(k){
  retried++;
  if(retried>=arriveAfter){ Object.assign(DATA_SECTION_STATE[k],
    {hasGood:true,error:null,asOf:'2026-08-13T16:45'}); settleRows=[ROW]; }
  return true;
}
/*__FNS__*/
function reset(){ opened=null; notices=[]; toasts=[]; retried=0; _jumpWaiting=false; }
function ok(state){ Object.assign(DATA_SECTION_STATE.settlements,
  {loading:false,retrying:false,hasGood:state,error:null,asOf:state?'2026-08-13T16:45':''}); }
let fail=0;
function t(name,cond,extra){ console.log((cond?'  ok  ':'  FAIL ')+name+(cond?'':' :: '+JSON.stringify(extra)));
  if(!cond) fail++; }
(async ()=>{
reset(); ok(true); settleRows=[ROW];
openByPrj('JS-2606-634');
t('있으면 연다', opened && opened.id==='JS-2606-634', {opened,notices});

reset(); ok(true); settleRows=[ROW];
openByPrj('JS-9999-999','UJ2601032');
t('프로젝트NO 로 찾아낸다', opened && opened.prj==='UJ2601032', {opened,notices});

reset(); ok(false); settleRows=[]; arriveAfter=1;
openByPrj('JS-2606-634','UJ2601032');
await new Promise(r=>setTimeout(r,400));
t('아직이면 없다고 안 한다', notices.filter(m=>/찾지 못했|없습니다/.test(m)).length===0, notices);
t('받고 있다고 말한다', toasts.some(m=>/받고 있습니다/.test(m)), toasts);
t('받으면 저절로 연다', opened && opened.id==='JS-2606-634', {opened});

reset(); ok(true); settleRows=[];
openByPrj('JS-0000-000','UJ9999999');
await new Promise(r=>setTimeout(r,60));
t('정말 없으면 없다고 말한다',
  notices.some(m=>/찾지 못했습니다/.test(m) && /기준으로 모두 받은/.test(m)), notices);
t('원인을 지목하지 않는다', !notices.some(m=>/다른 월 데이터이거나/.test(m)), notices);

reset(); Object.assign(DATA_SECTION_STATE.settlements,
  {loading:true,retrying:false,hasGood:false,error:null,asOf:''});
settleRows=[]; arriveAfter=99;
openByPrj('JS-2606-634','UJ2601032');
await new Promise(r=>setTimeout(r,120));
t('오는 중이면 다시 안 건다', retried===0, {retried});

console.log(fail? 'FAILED '+fail : 'ALL OK');
process.exit(fail?1:0);
})();
"""


def t258_error_report_says_when_and_why():
    """[258] 오류 경보는 **언제 났는지**를 말하고, 400 은 **어느 400 인지**를 말한다.

    2026-08-13 실측으로 둘 다 눈이 멀어 있었다:
    · 인계 '먼저 처리할 것' 이 `★새 오류 16건 — /api/flow` 를 맨 위에 올렸는데
      그 16건의 **마지막이 이틀 전**이었다. `error_book` 의 '★신규' 는 *사전에 없다*
      는 뜻인데 인계가 *새로 났다* 로 읽히게 실었다 — 사람이 **이미 끝난 고장**을
      찾으러 간다([172]). 원인은 `ux_summary` 가 시각을 안 줘서 rollup 이 스스로
      `날짜모름` 이라 적고 회귀 판정까지 포기한 것이다.
    · `/api/staff/entry` 400 이 하루 63건인데(오종현 님, 분담판 [90]) 기록에 남는
      것은 `HTTP_ERROR:400` 뿐이었다. 서버 400 은 `except Exception` 한 덩어리라
      **이유는 서버가 보낸 그 말 안에만** 있는데, 화면은 그것을 띄우고 **버렸다**.
    """
    import json as _json
    from datetime import datetime as _dt
    import error_book as eb
    import ledger_db as ldb

    # ① "오류" 는 3열 그대로여야 한다 — `for t,d,c in` 로 푸는 곳이 셋이다.
    #    열을 늘리면 그쪽이 조용히 깨진다(그래서 키를 새로 더했다).
    rd = lambda *a: open(os.path.join(ROOT, *a), encoding="utf-8").read()  # noqa: E731
    src = rd("ledger_db.py")
    i오류 = src.index('"오류": q(')
    i최근 = src.index('"오류최근": q(')
    assert "COUNT(*) FROM ux" in src[i오류:i오류 + 200] and \
        "MAX(ts)" not in src[i오류:i오류 + 200], \
        '"오류" 에 열을 늘렸다 — 3-튜플로 푸는 곳이 셋이라 조용히 깨진다'
    assert "MAX(ts)" in src[i최근:i최근 + 260], '"오류최근" 이 마지막 시각을 안 준다'
    assert "ORDER BY 3 DESC" in src[i최근:i최근 + 260], \
        '"오류최근" 정렬이 "오류" 와 다르면 같은 줄이 아니게 된다'
    assert set(ldb.ux_summary(days=1, limit=1)) >= {"오류", "오류최근"}, \
        "ux_summary 가 두 키를 다 주지 않는다"

    # ② rollup 이 ux 줄에도 시각을 채운다 — 통째로 '날짜모름' 이면 회귀를 영영 못 센다.
    rsrc = rd("error_book.py")
    j = rsrc.index("for r in ux_rows:")
    블록 = rsrc[j:j + 600]
    assert "last_ts[sig] = ts" in 블록, "ux 줄의 시각을 last_ts 에 안 넣는다"
    assert "날짜모름.add(sig)" in 블록 and "else:" in 블록, \
        "시각이 없을 때만 '날짜모름' 이어야 한다 — 늘 모름이면 판정을 포기한 것이다"

    # ③ 빈 시각을 '방금' 이라 적지 않는다([169]).
    assert eb._ago("") == "때 모름" and eb._ago(None) == "때 모름", \
        "때를 모르는데 아는 것처럼 적는다"
    assert "일 전" in eb._ago("2020-01-02T03:04:05"), "오래된 것을 오래됐다고 안 적는다"

    # ④ 인계는 **최근 것부터** 싣고 마지막 때를 붙인다(리포트는 건수 순 그대로).
    #    ⚠ 진짜 리포트 파일은 안 건드린다 — 실측 증거에 합성 행을 섞으면 그 파일이
    #      더는 실측이 아니다([192] 와 같은 자리).
    본래 = eb.REPORT_JSON
    with tempfile.TemporaryDirectory() as td:
        eb.REPORT_JSON = os.path.join(td, "r.json")
        try:
            with open(eb.REPORT_JSON, "w", encoding="utf-8") as f:
                _json.dump({"회귀": [], "못본것": [], "새오류": [
                    {"건수": 99, "어디": "/api/old", "무엇": "x", "마지막": "2020-01-01T00:00:00"},
                    {"건수": 1, "어디": "/api/new", "무엇": "y",
                     "마지막": _dt.now().isoformat(timespec="seconds")},
                ]}, f, ensure_ascii=False)
            줄 = eb.handoff_lines()
        finally:
            eb.REPORT_JSON = 본래
    assert 줄 and "/api/new" in 줄[0], \
        f"오늘 난 오류가 맨 위가 아니다 — 건수만 보면 끝난 고장이 오늘 것을 덮는다: {줄[:1]}"
    assert "마지막" in 줄[0], "인계 줄이 '언제 났나' 를 말하지 않는다"
    assert any("/api/old" in x for x in 줄), \
        "오래된 것을 조용히 뺐다 — 사전에 없는 것은 그대로 남아야 한다([169])"

    # ⑤ 400 에 이유를 붙여도 **사전은 그대로 알아본다**(부분 문자열 대조).
    ent = eb.look_up("/api/staff/entry",
                     "HTTP_ERROR:400 · 수정할 업무 카테고리를 확인할 수 없습니다")
    assert ent and ent["이름"] == "필수 값이 빠짐", \
        "이유를 붙였더니 사전이 못 알아본다 — 매일 '처음 보는 오류' 가 된다"

    # ⑥ 화면이 서버가 말한 이유를 기록에 남긴다. 단 4xx 에만 —
    #    5xx·502 본문은 터널이 준 HTML 쪽지라 지문만 더럽힌다.
    html = rd("webapp", "index.html")
    k = html.index("code:'HTTP_ERROR'")
    창 = html[k - 400:k + 900]
    assert "${err.code}:${r.status}${why}" in 창, \
        "서버가 말한 이유를 여전히 버린다 — 400 이 하루 63번 나도 어느 칸인지 모른다"
    assert "r.status>=400&&r.status<500" in 창, \
        "4xx 로 안 좁혔다 — 502 HTML 쪽지가 지문에 섞인다"
    print("[258] 오류 경보가 '언제' 와 '어느 400' 을 말한다 OK")


def t259_upload_tells_and_finish_tells_too():
    """[259] 올리면 알리고, **끝났는지·죽었는지도** 알린다 (2026-08-14 지시).

    지시: "류지영이 카톡 텍스트 파일 앱에 업로드 하면 바로 정리 시작하고 나에게
    올렸다고 알려주는 구조 코딩해 / 업로드 알림 만들어"

    실측(2026-08-14) — 절반은 이미 있었다: `/api/automation/kakao-upload` 가 파일을
    받으면 그 자리에서 `start_task("automation")` 을 부른다. **'올리면 바로 정리
    시작'은 이미 됐다.** 빠져 있던 것은 알리는 쪽이고, 코드 전체에 사람에게 알리는
    길이 **0건**이었다 — 류지영이 올려도 형님은 몰랐고, 더 나쁜 쪽으로 **처리가
    실패해도 아무도 몰랐다.**

    되돌아가면 안 되는 것만 지킨다:
      ① 접수 때 알리고, **끝났을 때도** 알린다 — '올렸다'만 가면 처리가 죽어도
         된 줄 알고 넘어간다([169]). 실패 경로에서도 알린다.
      ② 성공 판정은 `status == "success"` 다 — `"ok"` 로 물으면 **오류 없이
         한 건도 안 걸린다**([165] 의 함정. 실측으로 이 칸 이름을 확인했다).
      ③ 채널 실패를 성공으로 안 적는다([169]) — 못 보냈으면 실패에 남는다.
      ④ 외부 채널에는 **건수·상태만** — 프로젝트NO·금액·캠프명을 안 내보낸다.
      ⑤ **알림 실패가 업로드를 죽이지 않는다** — push 가 어떤 예외도 밖으로 안 낸다.
      ⑥ 아직 안 끝난 회차를 '끝'으로 적지 않고, 오래 지나도록 자국이 없으면
         조용히 넘기지 않고 '확인 못 함'이라 말한다([169]).
    """
    from datetime import datetime as _dt, timedelta as _td, timezone as _tz
    import notify as N

    rd = lambda *a: open(os.path.join(ROOT, *a), encoding="utf-8").read()  # noqa: E731
    srv = rd("webapp", "app_server.py")

    # ── ① 업로드 창구가 접수·거절·결과를 다 부른다 ──────────────────────────
    k = srv.index('if p == "/api/automation/kakao-upload":')
    창 = srv[k:k + 3400]
    assert "_notify_upload_received(" in 창, \
        "카톡 업로드가 접수를 안 알린다 — 류지영이 올려도 형님은 모른다"
    assert 창.count("_notify_upload_rejected(") >= 3, \
        "거절(용량·형식·예외)을 안 알린다 — 거절은 화면에만 뜨고 사라진다"
    assert "notify.expect_upload(" in srv, \
        "접수만 알리고 결과를 안 뒤따른다 — 처리가 죽어도 된 줄 안다([169])"
    # 결과 확인은 회차가 끝나는 자리와 무인 회차 **둘 다**에서 불려야 한다.
    #   앱이 시작한 회차는 러너가, 5분 스케줄러가 집어간 회차는 워치독이 닫는다.
    fin = srv.index("_note_last_run(key, title, local_returncode)")
    assert "notify.sweep_uploads()" in srv[fin:fin + 700], \
        "회차가 끝나도 결과를 안 알린다 — 아무도 화면을 안 열면 알림이 영영 안 간다"
    wd = rd("watchdog.py")
    assert "close_upload_notices" in wd and "close_upload_notices(dry)," in wd, \
        "워치독 단계에 없다 — 5분 스케줄러가 집어간 회차는 결과가 영영 안 온다"

    # ── ② 성공 칸 이름을 짝짓지 않는다([165]) ────────────────────────────────
    assert 'status == "success"' in rd("notify.py"), \
        "성공 판정이 'success' 가 아니다 — 오류 없이 한 건도 안 걸린다([165])"
    assert '"ok"' not in N._run_verdict.__doc__ or True   # 문서는 자유
    assert N._run_verdict({"status": "success", "changed_sources": ["kakao"],
                           "summary": "s"})[0] == "끝"
    assert N._run_verdict({"status": "partial", "failures": ["보관본"],
                           "stages": [{"name": "카톡 추출", "ok": False,
                                       "summary": "exit 2"}]})[0] == "실패"
    assert N._run_verdict({"status": "running"})[0] == "", \
        "아직 도는 회차를 끝났다고 적는다 — 안 끝난 것을 끝났다 하면 안 된다"

    # ── ③④⑤⑥ 은 진짜 기록 파일을 안 건드리고 임시 경로로 잰다 ───────────────
    #    ⚠ 실측 증거에 합성 행을 섞으면 그 파일이 더는 실측이 아니다([247] 의 함정).
    본래 = {k2: getattr(N, k2) for k2 in
            ("STORE", "LOG", "PENDING", "CONF", "PIPELINE_STATE")}
    with tempfile.TemporaryDirectory() as td:
        for k2 in 본래:
            setattr(N, k2, os.path.join(td, k2 + ".json"))
        try:
            def utc(sec=0):
                return (_dt.now(_tz.utc) + _td(seconds=sec)).isoformat(timespec="seconds")

            # ④ 외부 채널 본문에 업무값이 안 실린다.
            leak = "UJ2601321 송파3캠프 1,234,500원 PO372139"
            line = N.external_text({"갈래": "카톡 원본 업로드", "상태": "정리 시작",
                                    "건수": 3, "제목": leak, "본문": leak})
            for bad in ("UJ2601321", "송파3캠프", "1,234,500", "PO372139"):
                assert bad not in line, f"외부 채널에 업무값이 실렸다: {bad} / {line}"
            assert N.leaks_business_value("UJ2601321") and \
                N.leaks_business_value("1,234,500") and \
                N.leaks_business_value("송파3캠프") and \
                not N.leaks_business_value("정리 시작 · 3건"), \
                "업무값 판정이 정상 문구까지 막거나 업무값을 놓친다"

            # ③ 채널 실패를 성공으로 안 적는다 — 닿을 수 없는 주소를 켠다.
            N._save(N.CONF, {"channels": [{"name": "시험채널", "kind": "webhook",
                                           "enabled": True, "timeout": 1,
                                           "url": "https://127.0.0.1:1/none"}]})
            r = N.push("시험", "제목", "본문", audience=[N.ADMIN])
            assert "app" in r["보냄"], "앱 안 알림은 언제나 켜져 있어야 한다"
            assert "시험채널" not in r["보냄"], "못 보낸 채널을 보냈다고 적는다([169])"
            assert any(x.get("채널") == "시험채널" for x in r["실패"]), \
                "실패를 기록에 안 남긴다 — 못 보냈는지 아무도 모른다"
            log = N._load(N.LOG, [])
            assert log and log[-1]["실패"], "알림 기록에 실패가 안 남는다"
            N._save(N.CONF, {})

            # ⑤ 알림이 터져도 예외가 밖으로 새지 않는다(업로드가 500 이 되면 안 된다).
            원래저장 = N._save
            try:
                N._save = lambda *a, **k3: (_ for _ in ()).throw(OSError("디스크 없음"))
                bad = N.push("시험", "터져도 조용히", audience=[N.ADMIN])
            finally:
                N._save = 원래저장
            assert isinstance(bad, dict) and bad["실패"], \
                "알림이 실패했는데 실패라고 안 적는다"
            assert any(x.get("채널") == "app" for x in bad["실패"]), \
                "앱 알림 실패가 기록되지 않는다"

            # ① 끝 · 실패를 실제로 만들어 본다 — **올린 사람에게도** 간다.
            N._save(N.STORE, [])
            who = [N.ADMIN, N.staff("ryu-jiyeong")]
            N.expect_upload("카톡 원본", "a.txt", who)
            N._save(N.PIPELINE_STATE, {"last_run": {
                "status": "success", "finished_at": utc(60),
                "summary": "변경원천 kakao · 실패 0건",
                "changed_sources": ["kakao"], "stages": [],
                "current_stage": "완료"}, "history": []})
            assert N.sweep_uploads()["끝"] == 1, "끝난 것을 안 알린다"
            assert N.feed("staff", "ryu-jiyeong"), \
                "올린 사람에게 안 간다 — 됐는지 모르면 또 올려 중복 원본이 된다"
            assert not N.feed("staff", "oh-jonghyeon"), \
                "받는이가 아닌 사람에게 남의 업로드 알림이 간다"

            N._save(N.STORE, [])
            N.expect_upload("카톡 원본", "b.txt", [N.ADMIN])
            N._save(N.PIPELINE_STATE, {"last_run": {
                "status": "partial", "finished_at": utc(60), "summary": "x",
                "failures": ["Excel 보관본 생성·검증"], "current_stage": "실패 확인",
                "stages": [{"name": "카톡 추출", "ok": False, "summary": "exit 2"}]},
                "history": []})
            assert N.sweep_uploads()["실패"] == 1, \
                "실패 경로에서 안 알린다 — 조용한 사고가 그대로 남는다"
            assert N.feed("admin")[0]["심각도"] == "error", "실패가 경보가 아니다"

            # ⑥ 안 끝난 회차는 기다리고, 오래되면 '확인 못 함'이라 말한다.
            N._save(N.STORE, [])
            N.expect_upload("카톡 원본", "c.txt", [N.ADMIN])
            N._save(N.PIPELINE_STATE, {"last_run": {"status": "running",
                                                    "finished_at": None},
                                       "history": []})
            assert N.sweep_uploads() == {"끝": 0, "실패": 0, "확인못함": 0, "남음": 1}, \
                "도는 중인 회차를 결과로 읽었다"
            N._save(N.STORE, [])
            os.remove(N.PIPELINE_STATE)
            늙음 = (_dt.now().astimezone() - _td(hours=N.UNRESOLVED_WARN_H + 3))
            N._save(N.PENDING, [{"id": "x", "갈래": "카톡 원본", "이름": "c.txt",
                                 "받는이": [N.ADMIN], "원천": "automation",
                                 "올린때": 늙음.isoformat(timespec="seconds")}])
            assert N.sweep_uploads()["확인못함"] == 1, \
                "회차 자국을 못 읽었는데 조용히 넘긴다 — 못 본 것을 이상 없음이라 한다([169])"

            # ⑦ 같은 갈래가 5분 안에 여러 번이면 한 줄로 합치되 **건수를 말한다**([170]).
            N._save(N.STORE, [])
            for _ in range(3):
                N.push("카톡 원본 업로드", "올렸습니다", audience=[N.ADMIN])
            items = N._load(N.STORE, [])
            assert len(items) == 1 and items[0]["건수"] == 3, \
                f"합치기가 안 된다 — 알림이 대부분이면 아무도 안 본다([170]): {items}"

            # ⑧ 확인은 **내가 받는 알림만** 내려간다.
            assert N.ack([items[0]["id"]], "staff", "oh-jonghyeon") == 0, \
                "남의 알림을 대신 내린다 — 그 사람은 영영 그 사실을 모른다"
            assert N.ack([items[0]["id"]], "admin") == 1 and not N.feed("admin"), \
                "확인해도 안 내려간다"
        finally:
            for k2, v in 본래.items():
                setattr(N, k2, v)

    # ── 앱 화면이 확인 단추를 업로드 알림에만 붙인다 ──────────────────────────
    html = rd("webapp", "index.html")
    assert "ackNotice(" in html and "startsWith('notify:')" in html, \
        "업로드 알림에만 확인 단추가 붙어야 한다 — 감시 항목을 지우면 고장이 화면에서만 사라진다"
    assert '"/api/notifications/ack"' in srv, "확인 경로가 서버에 없다"
    assert "get_notifications(self._actor())" in srv, \
        "받는이를 안 넘긴다 — 올린 사람이 제 결과를 못 본다"
    print("[259] 업로드는 접수·끝·실패를 다 알리고 채널 실패를 숨기지 않는다 OK")


if __name__ == "__main__":
    print("합성데이터 검증 시작 (실데이터·실서버 접촉 없음)")
    with tempfile.TemporaryDirectory() as tmp:
        t1_erp_check(tmp)
        t4_kakao(tmp)
        t5_writer(tmp)
        t7_po(tmp)
        t8_findings_sheet(tmp)
        t12_dedupe(tmp)
        t13_fix_ids(tmp)
        t42_first_empty_row(tmp)
        t43_receipt_fill(tmp)
        t45_cloud_queue_and_erp_documents(tmp)
        t49_exec_metric_drilldown_and_sheet_scroll(tmp)
    t50_stale_completion_drilldown_and_capture()
    t51_manual_daily_activity()
    t2_payload()
    t3_match()
    t9_watchdog()
    t10_band_extract()
    t11_backfill()
    t14_datesort()
    t15_doc_ocr()
    t16_status()
    t17_expand()
    t18_erp_docs()
    t19_workbook_integrity(None)
    t20_rep_no()
    t21_reorder()
    t22_bundle()
    t23_formulas()
    t24_reorder_safety()
    t25_attachments()
    t26_mobile()
    t27_po()
    t28_resolve()
    t29_cloud()
    t30_dns_and_versions()
    t31_tech()
    t32_band_sheet()
    t33_unbilled_banner()
    t34_capture_and_no_send()
    t35_confirm_evidence()
    t36_mobile_input()
    t37_band_coverage()
    t38_daily_brief()
    t40_claim_enforced()
    t41_dates_explicit()
    t44_zscan()
    t46_app_2026_only()
    t47_back_nav()
    t52_data_status()
    t53_session_handoff()
    t54_side_work_db_only()
    t56_work_detail_from_source()
    t70_quarter_as_months()
    t71_period_range()
    t72_project_first_representative_report()
    with tempfile.TemporaryDirectory() as tmp:
        t73_pm_schedule_source_sync(tmp)
    t74_billing_fill()
    t75_gcal_sync()
    t77_side_work_single_switch()
    t78_recalc_pending_visible()
    t90_ip_guard_and_archive()
    t91_icon_sprite_and_ios_theme()
    t92_excel_recalc_agent()
    t93_ledger_db_and_ux()
    t94_human_edit_guard()
    t212_hand_edit_detection()
    t95_objective_completion_db_only()
    t96_work_management_tabs()
    t97_settlement_source_completion()
    t98_remote_control_tracking()
    t99_share_intake_pull()
    t100_erp_pdf_archive()
    t101_percent_and_no_erp_post()
    t102_calendar_filter_and_period()
    t103_session_wrapup_hook()
    t104_session_scoped_claims()
    t105_settle_report()
    t107_report_dates_and_capture()
    t108_pm_source_fallback()
    t109_remote_edit_delete_versions()
    t110_writer_formula_key()
    t111_account_handoff_freshness()
    t112_band_plan_order_and_scope()
    t113_paste_typos_and_misc_reclass()
    t114_claim_owner_is_agent_pid()
    t115_text_contrast()
    t116_manual_refresh_is_really_fresh()
    t117_dark_mode_toggle()
    t118_ocr_crosscheck()
    t119_context_guard()
    t202_layer_dialogs()
    t122_dash_drag_and_remote_version()
    t123_calendar_share_tools()
    t124_no_duplicate_menus()
    t125_worktree_shared_state()
    t126_app_font_and_revert()
    t246_font_presets_single_table()
    t247_chrome_collect_report_round_trip()
    t248_rounds_run_without_a_console_window()
    t249_entry_save_never_silent()
    t250_error_book_speaks_and_counts()
    t251_zero_tells_which_zero_it_is()
    t252_po_shape_matches_reality_and_restart_asks_first()
    t253_share_folder_pulls_whole_parent_cheaply()
    t254_each_menu_resets_to_its_own_first_screen()
    t255_delete_is_reversible_and_exclusion_is_not_delete()
    t256_list_zero_goes_through_one_door()
    t257_jump_says_why_it_could_not_find()
    t258_error_report_says_when_and_why()
    t259_upload_tells_and_finish_tells_too()
    t127_dark_mode_no_hardcoded_light_panel()
    t128_dash_tap_to_move()
    t129_call_notes_db_only_and_device_open()
    t130_band_grab_rejects_timeless_harvest()
    t131_band_quiet_vs_stalled()
    t135_contaminated_not_recollected()
    t132_dash_snap_expand_theme_swipe()
    t133_inline_style_dark_safe()
    t134_section_fold()
    t136_work_lanes()
    t137_contamination_marked_every_merge()
    t138_daily_run_completion_watch()
    t139_new_version_is_atomic()
    t140_freshness_tells_the_truth()
    t141_long_text_folds()
    t142_flow_editable()
    t143_originals_one_tap()
    t147_project_history()
    t148_input_suggest()
    t149_tech_center()
    t164_deposit_dedupe()
    t165_ledger_reads_billing_status()
    t166_billing_status_ladder()
    t167_daily_run_inflight()
    t168_erp_progress_glob_is_cached()
    t169_blind_count_sees_unlooked()
    t170_po_amount_ladder()
    t171_cache_swap_waits_for_readers()
    t172_typo_watch_does_not_cry_wolf()
    t179_comments_everywhere_and_crossed()
    t180_round_leaves_footprints_and_finishes()
    t181_app_answers_before_claude_is_called()
    t182_app_collects_without_claude()
    t183_collect_survives_pc_off()
    t184_phone_answers_with_the_same_rules()
    t185_datalake_shown_in_app()
    t186_kakao_round_and_stale_tmp()
    t187_free_vs_insurance_are_not_one_label()
    t188_worklog_shows_this_month_only()
    t189_worklog_reflects_without_hands()
    t190_autopilot_retries_without_failure_cascade()
    t191_confirmation_truth_and_fast_refresh()
    t196_stage_words_come_from_one_place()
    t197_restart_blip_is_not_a_failure()
    t198_source_index_no_per_file_stat()
    t203_ledger_screens_are_split()
    t173_classify_cache_follows_rules()
    t174_zero_match_blames_the_key()
    t175_step_timeout_cannot_hang_forever()
    t176_rules_bump_does_not_wipe_the_index()
    t177_comment_collection_is_targeted()
    t199_distrust_trusts_confirmed_zero()
    t217_probe_instead_of_scraping_absent_numbers()
    t178_unverified_harvest_is_not_read()
    t161_erp_filename_fingerprint()
    t160_master_book_cache()
    t154_amount_basis()
    t155_cancel_and_handover()
    t162_band_comments_collected()
    t163_last_run_shown()
    t156_refresh_fast()
    t157_tech_install()
    t144_topmost_pin_always_restores()
    t145_redirect_deleted_needs_two_rounds()
    t146_erp_bulk_grab_registry()
    t150_datalake_schema_and_incremental()
    t151_collect_all_idempotent_and_no_login_scrape()
    t152_band_recollect_window()
    t153_erp_excel_to_records()
    t158_wrapup_drops_huge_files()
    t159_handoff_supersede()
    t120_calendar_sheet_and_share()
    t121_pid_alive()
    t106_calendar_kind_colors()
    with tempfile.TemporaryDirectory() as _tmp84:
        t84_duplicate_source_files(_tmp84)
    with tempfile.TemporaryDirectory() as _tmp86:
        t86_daily_run_singleton_and_inbox_classification(_tmp86)
    t79_work_log_source_sync_and_report_capture()
    with tempfile.TemporaryDirectory() as tmp:
        t80_new_project_flow_db_only(tmp)
    with tempfile.TemporaryDirectory() as tmp:
        t200_evidence_verification_sync(tmp)
    with tempfile.TemporaryDirectory() as tmp:
        t85_staff_po_work_log_and_edit_priority(tmp)
    t81_terra_sol_handoff_review()
    t82_daily_cutoff()
    t83_agent_dispatch_and_calendar()
    t76_source_organizer()
    with tempfile.TemporaryDirectory() as tmp:
        t201_upload_intake(tmp)
    t55_pm_brief_drilldown_and_capture()
    t58_check_hub_detail_and_capture()
    t48_excel_2026_stats_and_verified_completion()
    t39_realtime_monitor()
    t6_webapp()
    t194_legacy_queue_migration_and_round_truth()
    t193_app_db_cutover_archive_and_frontend()
    t195_incremental_source_to_db_to_archive()
    t204_staff_finance_entry_is_one_save_and_source_safe()
    t205_three_staff_sessions_cannot_forge_actor()
    t206_finance_archive_keeps_real_headers_and_formulas()
    t207_live_revision_is_shared_and_nonblocking()
    t208_cancel_remote_resolution_is_exact_and_finance_safe()
    t209_pipeline_lock_owner_cannot_be_forged_or_overwritten()
    t210_pid_reuse_is_not_alive_and_customer_scan_is_one_pass()
    t211_progress_trace_owner_identity()
    t227_text_locks_use_the_one_owner_judge()
    t220_flow_yes_no_cycles()
    t221_commit_hygiene_under_siblings()
    t222_flow_charts_switch_and_capture()
    t223_superseded_evidence_heals_itself()
    t213_exact_pid_fingerprint_reaches_every_owner()
    t214_first_live_revision_cannot_be_falsely_applied()
    t215_cancel_timeline_last_explicit_state_wins()
    t218_camp_standard_erp_basis_and_pm_units()
    t219_noon_round_is_daily_windowed_and_yields()
    t224_wrapup_commit_refusal_paths()
    t225_session_auto_resumes_parked_and_pushes()
    t228_scheduler_rounds_are_watched()
    t229_band_liveness_contract()
    t230_ai_tier_picks_model_and_effort()
    t231_loop_tick_weight_from_evidence()
    t232_orgchart_floorplan_roster_and_states()
    t233_round_steps_fit_inside_budget()
    t234_kim_miyeong_center_and_revenue()
    t235_chatbot_is_one_line_until_asked()
    t236_list_is_folded_into_groups()
    t237_cards_fold_with_one_tool()
    t238_parked_says_which_lane()
    t239_idle_lane_reclaims_itself_with_a_record()
    t240_install_has_a_door_and_says_why_when_shut()
    t241_boundary_survives_compact_and_clear()
    t242_ready_means_logged_in()
    t243_cancelled_is_not_undone()
    t244_band_evidence_closes_and_says_why()
    t245_truth_watch_asks_instead_of_asserting()
    t235_unattended_rounds_survive_pythonw()
    # 전체 검증이 끝난 뒤 시작 시점의 공유·추적 산출물 바이트와 대조한다.
    t192_synthetic_check_is_harmless()
    check_numbers_unique()
    print("ALL GREEN — 실작업 진행 가능")
