# -*- coding: utf-8 -*-
"""
band_reconcile.py — 밴드 게시글 ↔ 관리대장 자동 대조
======================================================
cache/ 의 밴드 게시글과 관리대장 02_돌발AS접수·04_정기점검을 대조해
"작업완료됐는데 밴드 게시가 확인되지 않는 건"과 "원장에 없는 밴드 게시"를 리포트한다.

매칭 규칙(보수적):
  1순위  게시글 본문에 프로젝트NO(UJ26...) 포함
  2순위  캠프명 핵심부(괄호 앞부분) 포함 + 게시일이 작업완료일 ±3일
  보조   작성자가 담당기사와 일치하면 신뢰도 ↑ 표기
원장은 read-only로만 연다(수정 없음). 결과는 reports/ 에만 출력.

실행:  python band_reconcile.py
"""
import sys, os, json, csv, re
from datetime import datetime, timedelta

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE_DIR)
CONFIG_PATH = os.path.join(ROOT, "config", "ecount_config.json")
CACHE_DIR = os.path.join(BASE_DIR, "cache")
REPORT_DIR = os.path.join(ROOT, "reports")

HDR_ROW, FIRST = 4, 5

cfg = json.load(open(CONFIG_PATH, encoding="utf-8"))
DATE_TOL = int(cfg.get("band", {}).get("date_tolerance_days", 3))


def load_posts():
    posts = []
    if not os.path.isdir(CACHE_DIR):
        return posts
    for fn in os.listdir(CACHE_DIR):
        if not fn.endswith(".json") or fn.startswith(("raw_", "dump_")):
            continue
        c = json.load(open(os.path.join(CACHE_DIR, fn), encoding="utf-8"))
        for pk, p in c.get("posts", {}).items():
            ts = p.get("created_at")
            d = datetime.fromtimestamp(ts / 1000).date() if isinstance(ts, (int, float)) else None
            posts.append({"band": c.get("band_name", fn[:8]), "post_key": pk, "date": d,
                          "author": p.get("author", ""), "content": p.get("content", ""),
                          "photos": p.get("photo_count", 0)})
    return posts


def camp_core(camp):
    return re.split(r"[(\s]", str(camp or ""))[0].strip()


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    s = str(v or "")[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def read_rows():
    import openpyxl
    sys.path.insert(0, ROOT)
    from ecount_reconcile import resolve_master
    wb = openpyxl.load_workbook(resolve_master(cfg["reconcile"]["master_xlsx"]), read_only=True, data_only=True)
    out = []
    spec = {  # 시트: (ID, 프로젝트NO, 캠프명, 담당기사, 완료일, 사진등록, 완료보고서, 밴드수정, 상태)
        "02_돌발AS접수": ("접수ID", "프로젝트NO", "캠프명", "담당기사", "작업완료일", "사진등록", "완료보고서등록", "밴드수정", "진행상태"),
        "04_정기점검":   ("점검ID", "프로젝트NO", "캠프명", "담당기사", "실제점검일", None, None, None, "점검상태"),
    }
    for sheet, cols in spec.items():
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        def gv(row, name):
            return row[idx[name]] if name and name in idx and idx[name] < len(row) else None
        for row in ws.iter_rows(min_row=FIRST, values_only=True):
            # ID 열은 수식이라 새로 추가된 행은 엑셀을 열기 전까지 캐시값이 없다(None).
            # 그런 행도 대조 대상이어야 하므로 프로젝트NO를 대체 키로 쓴다.
            rid = gv(row, cols[0]) or gv(row, cols[1])
            if not rid:
                continue
            done = to_date(gv(row, cols[4]))
            if not done:
                continue        # 완료된 건만 대조
            out.append({"시트": sheet, "ID": str(rid), "프로젝트NO": str(gv(row, cols[1]) or ""),
                        "캠프명": str(gv(row, cols[2]) or ""), "담당기사": str(gv(row, cols[3]) or ""),
                        "완료일": done,
                        "사진등록": str(gv(row, cols[5]) or ""), "완료보고서": str(gv(row, cols[6]) or ""),
                        "밴드수정": str(gv(row, cols[7]) or "")})
    wb.close()
    return out


def photo_updates(results):
    """확인된 밴드 사진을 안정적인 프로젝트NO 키의 빈칸 입력 큐로 만든다."""
    return [{"sheet": "02_돌발AS접수",
             "key_col": "프로젝트NO" if r["프로젝트NO"] else "접수ID",
             "key": r["프로젝트NO"] or r["ID"], "col": "사진등록",
             "value": "등록", "vtype": "text", "only_if_empty": True,
             "evidence": f"밴드 게시 {r['게시일']} {r['게시자']} 사진{r['사진수']}장"}
            for r in results
            if r["시트"] == "02_돌발AS접수" and r["밴드게시"] == "확인"
            and str(r["사진수"]).isdigit() and int(r["사진수"]) > 0]


def main():
    posts = load_posts()
    if not posts:
        sys.exit("밴드 캐시가 없습니다. 먼저  python band_auth.py → python band_sync.py  를 실행하세요.")
    rows = read_rows()
    print(f"밴드 게시글 {len(posts)}건 / 원장 완료건 {len(rows)}건 대조 시작")

    matched_posts = set()
    results = []
    # 밴드 수집이 닿는 가장 이른 날.
    # ★ 밴드별로 따로 본 뒤 **가장 늦은 시작일**을 쓴다. 두 밴드를 합쳐 최소값을 잡으면
    #   매출처 밴드(2025-11-10~)가 쿠팡AS 밴드(2025-12-16~)의 빈 구간을 덮어 버린다.
    #   AS·점검 완료 보고는 쿠팡AS 밴드에 올라가므로, 그 밴드가 안 닿는 기간은
    #   '게시 안 함'이라고 말할 수 없다 — 우리가 못 긁어온 것뿐이다.
    _first = {}
    for _p in posts:
        if _p["date"] and (_p["band"] not in _first or _p["date"] < _first[_p["band"]]):
            _first[_p["band"]] = _p["date"]
    oldest = max(_first.values()) if _first else None
    if oldest:
        print(f"밴드 수집 시작일(가장 늦은 밴드 기준) {oldest} — 그 이전 작업은 '수집범위밖'으로 둔다")
    for r in rows:
        best, how = None, ""
        for p in posts:
            if p["date"] is None:
                continue
            near = abs((p["date"] - r["완료일"]).days) <= DATE_TOL
            if r["프로젝트NO"] and r["프로젝트NO"] in p["content"]:
                best, how = p, "프로젝트NO"
                break
            core = camp_core(r["캠프명"])
            if near and core and len(core) >= 3 and core in p["content"]:
                if best is None:
                    best, how = p, "캠프명+날짜"
        author_ok = bool(best and r["담당기사"] and r["담당기사"] in best["author"])
        if best:
            matched_posts.add(best["post_key"])
        results.append({**{k: r[k] for k in ("시트", "ID", "프로젝트NO", "캠프명", "담당기사")},
                        "완료일": r["완료일"].isoformat(),
                        # ★ 밴드 수집이 닿지 않는 **이전 기간**은 '미확인'이 아니다.
                        #   캐시는 2025-12-16부터인데 12-08 작업까지 '게시 안 함'으로 몰면
                        #   기사들을 헛되이 추궁하게 된다 — 우리가 못 긁어온 것뿐이다.
                        #   (2026-07-28: 엑셀 저장으로 옛 행 상태가 드러나며 28건이 그렇게 떴다)
                        "밴드게시": ("확인" if best else
                                     ("수집범위밖" if (oldest and r["완료일"] < oldest) else "미확인")),
                        "매칭근거": how + ("+기사일치" if author_ok else ""),
                        "게시일": best["date"].isoformat() if best else "",
                        "게시자": best["author"] if best else "",
                        "사진수": best["photos"] if best else "",
                        "원장_사진등록": r["사진등록"], "원장_밴드수정": r["밴드수정"]})

    unmatched = [p for p in posts if p["post_key"] not in matched_posts]

    os.makedirs(REPORT_DIR, exist_ok=True)
    base = os.path.join(REPORT_DIR, f"밴드대조_{datetime.now():%Y%m%d_%H%M}")
    cols = list(results[0].keys()) if results else []
    with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(results)

    miss = [r for r in results if r["밴드게시"] == "미확인"]
    with open(base + ".md", "w", encoding="utf-8") as f:
        f.write("# 밴드 ↔ 관리대장 대조 리포트\n\n")
        f.write(f"- 생성: {datetime.now():%Y-%m-%d %H:%M}\n")
        f.write(f"- 원장 완료건 {len(results)} / 밴드게시 확인 {len(results)-len(miss)} / 미확인 {len(miss)}\n")
        f.write(f"- 원장과 연결 안 된 밴드 게시글: {len(unmatched)}건\n\n")
        f.write("## 밴드 게시 미확인 건 (작업완료인데 게시글 못 찾음)\n\n")
        f.write("| ID | 캠프 | 기사 | 완료일 | 원장 사진등록 | 원장 밴드수정 |\n|---|---|---|---|---|---|\n")
        for r in miss:
            f.write(f"| {r['ID']} | {r['캠프명']} | {r['담당기사']} | {r['완료일']} | {r['원장_사진등록']} | {r['원장_밴드수정']} |\n")
        f.write("\n## 원장에 연결 안 된 밴드 게시글 (최근 30건)\n\n")
        f.write("| 밴드 | 게시일 | 작성자 | 내용(앞 60자) | 사진 |\n|---|---|---|---|---|\n")
        for p in sorted(unmatched, key=lambda x: (x["date"] or datetime.min.date()), reverse=True)[:30]:
            f.write(f"| {p['band']} | {p['date']} | {p['author']} | {p['content'][:60].replace('|',' ')} | {p['photos']} |\n")

    print(f"확인 {len(results)-len(miss)} / 미확인 {len(miss)} / 미연결 게시글 {len(unmatched)}")
    print("리포트:", base + ".md")

    # 확정 사실 → 관리대장 자동입력 큐: 사진 있는 밴드 게시 확인 건의 02 사진등록(빈 칸만)
    try:
        sys.path.insert(0, ROOT)
        from ledger_writer import queue_add
        # ID는 수식 캐시라 행 재배치·엑셀 재계산 전후에 오래된 값일 수 있다.
        # 원문과 원장을 이미 프로젝트NO로 정확히 연결했으므로 큐도 안정적인 프로젝트NO를 우선한다.
        ups = photo_updates(results)
        if ups:
            print("자동입력 큐 적재:", queue_add(ups), "건 (02 사진등록)")
    except Exception as e:
        print("(자동입력 큐 적재 실패:", e, ")")


if __name__ == "__main__":
    main()
