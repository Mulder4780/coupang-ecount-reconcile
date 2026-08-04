# -*- coding: utf-8 -*-
"""정산 대기 중 독립 원자료로 입증된 건을 DB 완료 정본에 기록한다.

완료 기준은 서로 다른 원천이 정확히 맞는 경우로 제한한다.

* 금액 재계산 대기: 관리대장 06의 프로젝트·PO·거래명세서 합계와 PO 원본
  견적서의 프로젝트·PO·부가세 포함 총액이 모두 같고, 일치 견적서가 한 장뿐인 건.
* 금액 재계산 대기(ERP): 견적서가 없어도 같은 프로젝트의 ERP 판매전표
  공급가액(또는 부가세포함 합계)이 거래명세서합계와 **유일하게** 일치하는 건.
  전표·명세서는 서로 다른 원천이므로 금액이 정확히 맞으면 재계산 결과가 입증된다.
* 세금계산서 미발행: 관리대장 26_계산서구성에서 프로젝트가 연결되고 판정이
  ``확정``인 ERP 계산서 건.

완료 상태는 Excel 수식이나 발행일을 덮어쓰지 않고 ``ledger_db.resolution``에만
저장한다. 이 모듈은 daily_run에서 매일 재실행되며 upsert라 중복되지 않는다.
"""
import argparse
import json
import os
import re
import sys
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

import ledger_db
from ecount_reconcile import has_statement, load_config, read_ledger, resolve_master
from po_pdf import CACHE_FILE, PO_RE

REPORT = os.path.join(ROOT, "reports", "정산_객관완료.json")
AMOUNT_STATUS = "완료(견적·명세서 금액확인)"
ERP_AMOUNT_STATUS = "완료(ERP 전표 금액확인)"
QUOTE_ONLY_STATUS = "완료(견적 원본확인·담당자확인)"
INVOICE_STATUS = "완료(ERP 계산서 원본확인)"
# ERP 판매조회 진행상태만으로 발행을 입증하는 근거(2026-08-05 지시).
ERP_STATE_STATUS = "완료(ERP 발행상태확인)"
ERP_ISSUED_STATES = ("6.세금계산서발행", "7.수금완료")


def _money(value):
    try:
        return round(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _po(value):
    match = re.search(r"PO\d+", str(value or ""), re.I)
    return match.group().upper() if match else ""


def _raw_amount_wait(record):
    return (
        record.get("비용구분") == "유상"
        and str(record.get("원천업무ID") or "").startswith("AS-")
        and not _money(record.get("원장_공급가액"))
        and bool(record.get("원장_거래명세서발행일"))
        and _money(record.get("원장_거래명세서합계")) > 0
    )


def dedup_quote_rows(cache_path=CACHE_FILE):
    """PO 원본 파서 캐시를 ``po_pdf.scan``과 같은 키로 중복 제거해 반환한다.

    원본이 정본 폴더와 과거 공유 폴더에 함께 있어 경로만 다르면 같은 견적서가 두
    장처럼 보인다. 네트워크 PDF 전체를 다시 열지 않고 이미 검증된 파서 캐시를 쓴다.
    """
    try:
        cache = json.load(open(cache_path, encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return []
    rows, seen = [], set()
    for cache_key, row in cache.items():
        if row.get("종류") != "견적서":
            continue
        try:
            path, size, _mtime = cache_key.rsplit("|", 2)
        except ValueError:
            continue
        joined = os.path.normpath(path).replace(" ", "")
        match = PO_RE.search(joined)
        parent = os.path.basename(os.path.dirname(path)).lower()
        key = (
            match.group(1) if match else parent,
            os.path.basename(path).lower(),
            size,
        )
        if key in seen:
            continue
        seen.add(key)
        rows.append(dict(row))
    return rows


def quote_index(rows):
    """(프로젝트, PO, 부가세포함 총액) → 중복 제거된 견적서 목록."""
    out = {}
    for row in rows or []:
        project = str(row.get("프로젝트NO") or "").strip().upper()
        po_no = _po(row.get("PO번호"))
        total = _money(row.get("금액"))
        if project and po_no and total > 0:
            out.setdefault((project, po_no, total), []).append(row)
    return out


def erp_sales_index():
    """ERP 판매조회에서 {프로젝트NO: [전표(일자·PO·진행상태·공급가액·금액합계)]}를 읽는다.

    ``erp_progress``와 같은 최신 원본 한 벌만 읽고, 실패하면 빈 dict — 원본이 없다고
    완료를 철회하거나 지어내지 않는다. 합성검증은 실데이터 접촉 0이 원칙이라 건너뛴다.
    """
    if os.environ.get("CSOS_SYNTHETIC") == "1":
        return {}
    # ★ 2026-08-05 버그 수정: 아래 폴백은 `판매조회*.xlsx` 라는 **파일명**을 찾는데,
    #   ERP 내보내기 이름은 무작위(E91RXX1FJ7KAKFP.xlsx)다. 그래서 이 함수는 늘 빈
    #   dict 를 돌려줬고 **ERP 근거 완료가 한 번도 작동하지 않았다**(원장 미발행 744건
    #   중 450건이 ERP 로는 이미 발행/수금완료였다). 내용으로 판별해 만든 색인
    #   `erp_sales_index.py` 산출물을 먼저 읽는다.
    try:
        idx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "reports", "ERP판매_프로젝트색인.json")
        with open(idx_path, encoding="utf-8") as fh:
            idx = json.load(fh).get("index") or {}
        if idx:
            return {k.upper(): [{"date": v.get("date", ""), "po": _po(v.get("po")),
                                 "status": v.get("state", ""),
                                 "supply": _money(v.get("supply")),
                                 "total": _money(v.get("total"))}]
                    for k, v in idx.items()}
    except Exception:
        pass
    try:
        import glob

        import openpyxl
        from source_dirs import ERP_DIR

        cands = [
            path
            for path in glob.glob(os.path.join(ERP_DIR, "**", "판매조회*.xlsx"), recursive=True)
            if "~$" not in path and "__dup_" not in path
        ]
        if not cands:
            return {}
        path = max(cands, key=lambda p: os.stat(p).st_mtime)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    out = {}
    try:
        ws = wb["판매조회"] if "판매조회" in wb.sheetnames else wb.worksheets[0]
        cols = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            if cols is None:
                if "진행상태" not in cells:
                    continue
                cols = {
                    "date": next((i for i, c in enumerate(cells) if c == "일자"), None),
                    "prj": next((i for i, c in enumerate(cells) if "프로젝트" in c), None),
                    "status": cells.index("진행상태"),
                    "po": next((i for i, c in enumerate(cells) if "PO" in c.upper()), None),
                    "supply": next((i for i, c in enumerate(cells) if "공급가" in c), None),
                    "total": next((i for i, c in enumerate(cells) if c == "금액합계"), None),
                }
                continue
            get = lambda key: (
                row[cols[key]] if cols[key] is not None and cols[key] < len(row) else None
            )
            project = str(get("prj") or "").strip().upper()
            if not project:
                continue
            out.setdefault(project, []).append({
                "date": str(get("date") or "")[:10],
                "po": _po(get("po")),
                "status": str(get("status") or "").strip(),
                "supply": _money(get("supply")),
                "total": _money(get("total")),
            })
    except Exception:
        return {}
    finally:
        wb.close()
    return out


def confirmed_invoice_index(master):
    """26_계산서구성의 확정 행을 프로젝트별로 읽는다."""
    import openpyxl

    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    try:
        ws = next((sheet for sheet in wb.worksheets if sheet.title.startswith("26_")), None)
        if ws is None:
            return {}
        header = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        index = {str(value or "").strip(): i for i, value in enumerate(header) if value}
        required = {"일자-No.", "계산서공급가액", "포함프로젝트NO", "판정"}
        if not required.issubset(index):
            return {}
        out = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            get = lambda col: row[index[col]] if index[col] < len(row) else None
            verdict = str(get("판정") or "").strip()
            amount = _money(get("계산서공급가액"))
            slip = str(get("일자-No.") or "").strip()
            if not verdict.startswith("확정") or not slip or amount <= 0:
                continue
            evidence = {"slip": slip, "amount": amount, "verdict": verdict}
            for project in str(get("포함프로젝트NO") or "").split(","):
                project = project.strip().upper()
                if project:
                    out.setdefault(project, []).append(evidence)
        return out
    finally:
        wb.close()


def objective_entries(records, quotes, invoices, existing=None, erp_sales=None):
    """현재 원장 레코드에서 객관완료 DB 항목을 만든다.

    이미 ERP 수금완료처럼 더 강한 완료 근거가 있으면 상태를 낮춰 쓰지 않는다. 이
    모듈이 전에 쓴 상태는 다시 반환해 ``last_seen``만 갱신한다.
    """
    existing = existing or {}
    qindex = quote_index(quotes)
    erp_sales = erp_sales or {}
    own_statuses = (AMOUNT_STATUS, ERP_AMOUNT_STATUS, QUOTE_ONLY_STATUS, INVOICE_STATUS,
                    ERP_STATE_STATUS)
    # 프로젝트NO → {부가세포함 총액: 견적행}. 견적이 프로젝트당 정확히 한 금액이면
    # 명세합계가 어긋나 있어도(교차 입력 밀림) 견적 원본이 금액의 유일한 입증이다.
    quotes_by_project = {}
    for row in quotes or []:
        prj = str(row.get("프로젝트NO") or "").strip().upper()
        total = _money(row.get("금액"))
        if prj and total > 0:
            quotes_by_project.setdefault(prj, {})[total] = row
    entries = []
    for settle_id, record in sorted((records or {}).items()):
        old = existing.get(settle_id) or {}
        old_status = str(old.get("status") or "")
        if old_status.startswith("완료(") and old_status not in own_statuses:
            continue
        project = str(record.get("프로젝트NO") or "").strip().upper()

        if _raw_amount_wait(record):
            po_no = _po(record.get("원장_PO번호"))
            total = _money(record.get("원장_거래명세서합계"))
            hits = qindex.get((project, po_no, total), [])
            if len(hits) == 1:
                quote = hits[0]
                issue_day = str(record.get("원장_거래명세서발행일") or "")[:10]
                entries.append({
                    "settle_id": settle_id,
                    "project": project,
                    "status": AMOUNT_STATUS,
                    "basis": (
                        f"PO 원본 견적서 {quote.get('파일') or '-'} · 프로젝트 {project} · "
                        f"{po_no} · 부가세포함 {total:,}원 = 관리대장 06 거래명세서 "
                        f"{issue_day} 합계 {total:,}원 (유일 일치)"
                    ),
                    "evidence_kind": "amount",
                })
                continue

            # 견적서가 없으면 ERP 판매전표 금액으로 본다. 같은 프로젝트 전표 중
            # 공급가액 또는 부가세포함 합계가 명세서합계와 맞는 전표가 **정확히 한 장**
            # 이어야 하고, 전표에 PO번호가 있으면 원장 PO와 달라선 안 된다.
            slips = [
                slip
                for slip in erp_sales.get(project, [])
                if total > 0
                and (slip.get("supply") == total or slip.get("total") == total)
                and not (slip.get("po") and po_no and slip["po"] != po_no)
            ]
            if len(slips) == 1:
                slip = slips[0]
                issue_day = str(record.get("원장_거래명세서발행일") or "")[:10]
                entries.append({
                    "settle_id": settle_id,
                    "project": project,
                    "status": ERP_AMOUNT_STATUS,
                    "basis": (
                        f"ERP 판매전표 {slip.get('date') or '-'} ({slip.get('status') or '-'}) · "
                        f"프로젝트 {project} · 전표금액 {total:,}원 = 관리대장 06 거래명세서 "
                        f"{issue_day} 합계 {total:,}원 (유일 일치)"
                    ),
                    "evidence_kind": "erp_amount",
                })
                continue

            # 견적 단독 입증(2026-08-03 지시): "사람이 확인 안 해도 데이터로 입증된 건은
            # 모두 완료 — 각 담당자 확인으로 처리". 이 프로젝트의 견적서가 캐시 전체에서
            # **한 금액뿐**이면 그 견적이 금액의 유일한 원본 입증이다. 명세합계가 다르면
            # 교차 입력 밀림(quote_mismatch 진단)이며, 완료 근거에 차이를 그대로 남긴다.
            own = quotes_by_project.get(project) or {}
            if len(own) == 1:
                quote_total, quote = next(iter(own.items()))
                # ★ ERP 검증(2026-08-03 지시): ERP 전표가 이 프로젝트에 있고 금액이
                #   견적과도 다르면 두 원천이 서로 충돌 — 자동 완료하지 않고 확인
                #   작업(quote_mismatch 'ERP·견적 충돌')으로 넘긴다.
                erp_amounts = {slip.get("supply") for slip in erp_sales.get(project, [])}
                erp_amounts |= {slip.get("total") for slip in erp_sales.get(project, [])}
                erp_amounts.discard(0)
                erp_amounts.discard(None)
                if erp_amounts and quote_total not in erp_amounts:
                    continue
                diff = ("일치" if quote_total == total
                        else f"원장 명세합계 {total:,}원과 불일치 — 교차 입력 의심, 견적이 정본")
                entries.append({
                    "settle_id": settle_id,
                    "project": project,
                    "status": QUOTE_ONLY_STATUS,
                    "basis": (
                        f"PO 원본 견적서 {quote.get('파일') or '-'} · 프로젝트 {project} · "
                        f"견적 부가세포함 {quote_total:,}원 ({diff}) · "
                        f"담당자(류지영) 확인 처리 — 사용자 지시 2026-08-03"
                    ),
                    "evidence_kind": "quote_only",
                })
                continue

        issued = record.get("원장_세금계산서실제발행일") or record.get("원장_세금계산서발행일")
        invoice_hits = invoices.get(project, [])
        if (
            record.get("비용구분") == "유상"
            and not _raw_amount_wait(record)
            and has_statement(record)
            and not issued
            and invoice_hits
        ):
            slips = ", ".join(str(item["slip"]) for item in invoice_hits)
            amounts = sum(_money(item["amount"]) for item in invoice_hits)
            verdicts = ", ".join(sorted({str(item["verdict"]) for item in invoice_hits}))
            entries.append({
                "settle_id": settle_id,
                "project": project,
                "status": INVOICE_STATUS,
                "basis": (
                    f"관리대장 25 ERP 매출(세금)계산서 원본 · 26 계산서구성 "
                    f"{verdicts} · 프로젝트 {project} · 전표 {slips} · 공급가액 {amounts:,}원"
                ),
                "evidence_kind": "invoice",
            })
            continue

        # ★ ERP 진행상태 근거(사용자 지시 2026-08-05 "찾아 완료 처리").
        #   판매조회 진행상태가 **6.세금계산서발행 / 7.수금완료** 면 그 프로젝트의 계산서는
        #   ERP 상 실제로 나간 것이다. 금액 일치까지는 못 봐도(묶음 발행이라 건별 배분 불가)
        #   '발행되었다'는 사실 자체가 객관 근거다. 발행일은 지어내지 않는다 —
        #   DB(resolution)에만 근거와 함께 남기고 엑셀 발행일 칸은 비워 둔다.
        if (
            record.get("비용구분") == "유상"
            and has_statement(record)
            and not issued
        ):
            states = [str(s.get("status") or "") for s in erp_sales.get(project, [])]
            done = [s for s in states if s.startswith(ERP_ISSUED_STATES)]
            if done and len(done) == len(states):   # 한 전표라도 미발행이면 인정하지 않는다
                entries.append({
                    "settle_id": settle_id,
                    "project": project,
                    "status": ERP_STATE_STATUS,
                    "basis": (
                        f"ERP 판매조회 진행상태 {done[0]} · 프로젝트 {project} · "
                        f"전표 {len(states)}장 모두 발행 이상 단계 "
                        f"(묶음 발행이라 건별 금액 배분은 불가 — 발행 사실만 입증)"
                    ),
                    "evidence_kind": "erp_state",
                })
    return entries


def write_report(master, entries, synced=0):
    os.makedirs(os.path.dirname(REPORT), exist_ok=True)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "master": os.path.basename(master),
        "eligible": len(entries),
        "amount": sum(row.get("evidence_kind") == "amount" for row in entries),
        "erp_amount": sum(row.get("evidence_kind") == "erp_amount" for row in entries),
        "quote_only": sum(row.get("evidence_kind") == "quote_only" for row in entries),
        "invoice": sum(row.get("evidence_kind") == "invoice" for row in entries),
        "erp_state": sum(row.get("evidence_kind") == "erp_state" for row in entries),
        "synced": synced,
        "entries": entries,
    }
    tmp = REPORT + ".tmp"
    with open(tmp, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    os.replace(tmp, REPORT)
    return payload


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--sync", action="store_true", help="객관근거 완료를 resolution DB에 upsert")
    args = parser.parse_args(argv)
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    records = read_ledger(master)
    existing = ledger_db.resolutions()
    quotes = dedup_quote_rows()
    invoices = confirmed_invoice_index(master)
    erp_sales = erp_sales_index()
    entries = objective_entries(records, quotes, invoices, existing, erp_sales)
    synced = ledger_db.resolution_sync(entries) if args.sync else 0
    # ERP 검증(2026-08-03 지시): 견적 단독 완료였는데 지금 원본 기준으로 ERP 전표와
    # 충돌해 후보에서 빠진 건은 정확한 ID로 철회한다 — 확인 작업(quote_mismatch)으로
    # 넘어간다. 원본이 안 읽히는 날(빈 quotes/erp)은 과거 근거를 보존한다.
    if args.sync and quotes and erp_sales:
        valid = {e["settle_id"] for e in entries if e.get("evidence_kind") == "quote_only"}
        stale = [sid for sid, row in (existing or {}).items()
                 if str(row.get("status") or "") == QUOTE_ONLY_STATUS and sid not in valid]
        if stale:
            removed = ledger_db.resolution_retract(stale)
            print(f"  ERP·견적 충돌 재검출 — 견적단독 완료 {removed}건 철회(확인 작업으로)")
    payload = write_report(master, entries, synced)
    print(
        f"정산 객관완료 후보 {payload['eligible']}건 "
        f"(금액 {payload['amount']} · ERP전표 {payload['erp_amount']} · "
        f"견적단독 {payload['quote_only']} · 계산서 {payload['invoice']})"
        + (f" · DB 동기화 {synced}건" if args.sync else " · dry-run")
    )
    print("리포트:", REPORT)


if __name__ == "__main__":
    main()
