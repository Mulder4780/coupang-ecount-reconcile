# -*- coding: utf-8 -*-
"""신규 프로젝트 업무 흐름도 원본을 앱 비표시 내부 DB로 동기화한다.

원본 위치: ``0. 원본 자료/50. 쿠팡 신규 프로젝트 업무 흐름도``

이 모듈은 관리대장 시트·앱 API·고정 주소 사본을 전혀 변경하지 않는다. 폴더의 최신
Excel 파일을 안정적으로 읽어 모든 비어 있지 않은 셀과 체크시트 표를 ``reports/``의
로컬 DB(JSON)에 보존한다. 이후 신규 업무 기준이 추가돼도 앱 화면을 바꾸지 않은 채
원본 버전과 구조를 비교·활용할 수 있다.

실행:
  python new_project_flow_sync.py          # 최신 원본 분석만
  python new_project_flow_sync.py --apply  # 내부 DB 갱신(내용 동일 시 미기록)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from source_dirs import NEW_PROJECT_FLOW_DIR

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR", os.path.join(ROOT, "reports"))
DB_FILENAME = "new_project_flow_db.json"
APP_VISIBLE = False


def _json_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _digest(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_latest_source(folder: str = NEW_PROJECT_FLOW_DIR) -> str | None:
    """폴더 바로 아래 최신본을 우선하고, 없을 때만 보관본을 찾는다."""
    if not os.path.isdir(folder):
        return None

    direct, archived = [], []
    for base, _dirs, files in os.walk(folder):
        for name in files:
            if name.startswith("~$") or not name.lower().endswith((".xlsx", ".xlsm")):
                continue
            path = os.path.join(base, name)
            try:
                item = (os.path.getmtime(path), os.path.getsize(path), path)
            except OSError:
                continue
            rel = os.path.relpath(path, folder).split(os.sep)
            (direct if len(rel) == 1 else archived).append(item)
    candidates = direct or archived
    return max(candidates)[2] if candidates else None


def stable_snapshot(source: str, temp_dir: str) -> tuple[str, str]:
    """열려 있는 Excel도 마지막 저장본만 읽고, 저장 중 변경되면 다음 실행으로 넘긴다."""
    before = os.stat(source)
    copied = os.path.join(temp_dir, os.path.basename(source))
    shutil.copy2(source, copied)
    after = os.stat(source)
    if (before.st_size, before.st_mtime_ns) != (after.st_size, after.st_mtime_ns):
        raise RuntimeError("신규 업무 흐름도 원본이 저장 중입니다. 다음 자동 실행에서 다시 반영합니다.")
    return copied, _digest(copied)


def _nonempty_rows(ws) -> list[dict]:
    rows: list[dict] = []
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [
            {"column": column_number, "value": _json_value(value)}
            for column_number, value in enumerate(row, start=1)
            if value not in (None, "")
        ]
        if cells:
            rows.append({"row": row_number, "cells": cells})
    return rows


def _checksheet_table(rows: list[dict]) -> list[dict]:
    """체크시트의 표는 향후 신규 업무 추가 시 바로 검색할 수 있게 행 사전으로도 보관한다."""
    if not rows:
        return []
    first = {cell["column"]: cell["value"] for cell in rows[0]["cells"]}
    if not {"단계", "업무", "담당"}.issubset(set(first.values())):
        return []
    headers = {value: column for column, value in first.items() if isinstance(value, str) and value}
    records = []
    for row in rows[1:]:
        values = {cell["column"]: cell["value"] for cell in row["cells"]}
        record = {name: values.get(column, "") for name, column in headers.items()}
        if any(value not in (None, "") for value in record.values()):
            record["원본행"] = row["row"]
            records.append(record)
    return records


def parse_workflow(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheets = []
        checks = []
        for ws in wb.worksheets:
            rows = _nonempty_rows(ws)
            sheet = {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_row_count": len(rows),
                "rows": rows,
            }
            sheets.append(sheet)
            if ws.title.replace(" ", "") == "체크시트":
                checks = _checksheet_table(rows)
        return {"sheets": sheets, "checklist": checks}
    finally:
        wb.close()


def _load_db(path: str) -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def _write_db(path: str, payload: dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, temporary = tempfile.mkstemp(prefix=".new_project_flow_", suffix=".json", dir=os.path.dirname(path))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except OSError:
            pass
        raise


def sync(folder: str = NEW_PROJECT_FLOW_DIR, db_path: str | None = None) -> dict:
    """최신 원본을 읽어 DB를 바꾸고, 결과/변경 여부를 반환한다."""
    source = find_latest_source(folder)
    target = db_path or os.path.join(REPORT_DIR, DB_FILENAME)
    if not source:
        return {"status": "skipped", "reason": "원본 없음", "db_path": target}

    with tempfile.TemporaryDirectory(prefix="new_project_flow_") as temp_dir:
        snapshot, digest = stable_snapshot(source, temp_dir)
        previous = _load_db(target)
        prior = previous.get("source", {}) if isinstance(previous, dict) else {}
        if prior.get("sha256") == digest and prior.get("path") == source:
            return {
                "status": "unchanged", "source": source, "sha256": digest,
                "db_path": target, "sheet_count": len(previous.get("sheets", [])),
            }
        parsed = parse_workflow(snapshot)

    stat = os.stat(source)
    payload = {
        "schema_version": 1,
        "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "app_visible": APP_VISIBLE,
        "source": {
            "name": os.path.basename(source),
            "path": source,
            "sha256": digest,
            "size": stat.st_size,
            "modified_at": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
        },
        **parsed,
    }
    _write_db(target, payload)
    return {
        "status": "updated", "source": source, "sha256": digest, "db_path": target,
        "sheet_count": len(parsed["sheets"]), "checklist_count": len(parsed["checklist"]),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="내부 DB를 실제 갱신")
    parser.add_argument("--folder", default=NEW_PROJECT_FLOW_DIR, help="원본 폴더(검증용)")
    parser.add_argument("--db", default="", help="DB 저장 경로(검증용)")
    args = parser.parse_args()

    source = find_latest_source(args.folder)
    if not source:
        print("신규 프로젝트 업무 흐름도 DB — 원본 없음(기존 DB 유지)")
        return 0
    if not args.apply:
        with tempfile.TemporaryDirectory(prefix="new_project_flow_preview_") as temp_dir:
            snapshot, digest = stable_snapshot(source, temp_dir)
            parsed = parse_workflow(snapshot)
        print(f"신규 프로젝트 업무 흐름도 분석 — {os.path.basename(source)} · "
              f"시트 {len(parsed['sheets'])}개 · 체크 {len(parsed['checklist'])}건 · 해시 {digest[:12]}")
        return 0

    result = sync(args.folder, args.db or None)
    if result["status"] == "skipped":
        print("신규 프로젝트 업무 흐름도 DB — 원본 없음(기존 DB 유지)")
    elif result["status"] == "unchanged":
        print(f"신규 프로젝트 업무 흐름도 DB — 변경 없음 · 시트 {result['sheet_count']}개")
    else:
        print(f"신규 프로젝트 업무 흐름도 DB 갱신 — 시트 {result['sheet_count']}개 · "
              f"체크 {result['checklist_count']}건 · 앱 비표시")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
