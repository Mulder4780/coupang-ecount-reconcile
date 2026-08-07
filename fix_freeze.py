# -*- coding: utf-8 -*-
"""
fix_freeze.py — 틀 고정을 바로잡아 **프로젝트NO가 항상 보이게** 한다
================================================================================
사용자 지시(2026-07-28): "프로젝트 넘버를 맨 좌측으로 보내 고정해줘, 제일 중요해."

무엇이 잘못돼 있었나
  02시트 고정이 `D536`, 04시트 `D371`, 07시트 `E282` 였다. **열 지정(A~C)은 맞는데
  행이 틀렸다** — 아래로 스크롤한 상태에서 틀 고정을 눌러 535행까지가 통째로 얼었다.
  그래서 오른쪽으로 밀면 프로젝트NO가 같이 사라졌다.

왜 열을 물리적으로 옮기지 않는가
  프로젝트NO를 A열로 **실제로 이동**하면 27개 시트의 수식이 전부 어긋난다
  (02시트만 공유수식 5,321칸, 07·17시트는 다른 시트를 열 문자로 참조한다).
  프로젝트NO는 이미 B(또는 C)열이므로 **고정만 제대로 하면 항상 왼쪽에 남는다.**
  정말 물리적으로 옮기려면 엑셀에서 '잘라내기 → 잘라낸 셀 삽입'을 써야 한다
  (그 방법만 참조를 같이 옮겨 준다).

규칙: 머리글 4행을 고정하고(행 5부터 스크롤), 프로젝트NO 열까지는 반드시 고정한다.

사용
  python fix_freeze.py            # 어떻게 바꿀지만 (파일 안 건드림)
  python fix_freeze.py --apply    # vN+1 생성
"""
import os
import re
import sys
import zipfile

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path  # noqa: E402

HDR = 4
KEY = ("프로젝트NO", "정산ID", "원천업무ID")     # 이 열까지는 꼭 보이게 둔다


def _s(v):
    return "" if v is None else str(v).strip()


def plan(path):
    import openpyxl
    from openpyxl.utils import get_column_letter as GL, column_index_from_string as CI
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sn in wb.sheetnames:
        # 표가 아닌 안내·보고 시트 제외. 10_코드관리도 뺀다 — 드롭다운 목록 모음이라
        # 프로젝트NO가 S열에 있어 규칙대로 하면 18열이 얼어 못 쓰게 된다.
        if not sn[:2].isdigit() or sn[:2] in ("00", "01", "08", "09", "10",
                                              "12", "18", "19", "20", "22"):
            continue
        ws = wb[sn]
        # ★ 에이전트가 만든 보고 시트(23_확인필요현황·27_거래처코드)는 건드리지 않는다.
        #   틀고정은 그 시트 XML 안에 이미 들어 있고, 여기서 고쳐 쓰면 다음 갱신 때
        #   `findings_sheet` 가 "바뀌었다"고 보고 vN+1 을 만든다 — 둘이 서로 되돌리며
        #   버전만 끝없이 늘어난다(2026-08-07, 자기잠식 사고와 같은 갈래).
        from findings_export import is_agent_sheet
        if is_agent_sheet(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())):
            continue
        try:
            hdr = [_s(h) for h in next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))]
        except StopIteration:
            continue
        if not any(hdr):
            continue
        need = 0
        for k in KEY:
            if k in hdr:
                need = max(need, hdr.index(k) + 1)
        if not need:
            continue
        cur = _s(ws.freeze_panes)
        curcol = 0
        m = re.match(r"([A-Z]+)(\d+)", cur)
        if m:
            curcol = CI(m.group(1)) - 1               # 얼어 있는 열 수
        col = max(curcol, need)                       # 원래 의도는 살리되 프로젝트NO는 반드시
        new = f"{GL(col + 1)}{HDR + 1}"
        if new != cur:
            out.append((sn, cur or "(없음)", new, hdr[need - 1]))
    wb.close()
    return out


def patch_pane(xml, target):
    """<pane .../> 를 새 고정 위치로 바꾼다. 없으면 sheetView 안에 넣는다."""
    from openpyxl.utils import column_index_from_string as CI
    m = re.match(r"([A-Z]+)(\d+)", target)
    x, y = CI(m.group(1)) - 1, int(m.group(2)) - 1
    pane = (f'<pane xSplit="{x}" ySplit="{y}" topLeftCell="{target}" '
            f'activePane="bottomRight" state="frozen"/>')
    if re.search(r"<pane\b[^>]*/>", xml):
        return re.sub(r"<pane\b[^>]*/>", pane, xml, count=1)
    # 고정이 아예 없던 시트 — sheetView 바로 뒤에 넣는다
    return re.sub(r"(<sheetView\b[^>]*>)", r"\1" + pane, xml, count=1)


def main():
    do = "--apply" in sys.argv
    m = latest_master()
    src = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src)}\n")

    jobs = plan(src)
    if not jobs:
        print("고칠 것이 없습니다 — 이미 제대로 고정돼 있습니다.")
        return 0
    print(f"  {'시트':<22}{'지금':>10} → {'바꿀 값':<8} (이 열까지 보이게)")
    for sn, cur, new, key in jobs:
        print(f"  {sn:<22}{cur:>10} → {new:<8} {key}")
    print(f"\n합계 {len(jobs)}개 시트 · 머리글 {HDR}행 + 프로젝트NO 열까지 항상 보이게 합니다.")
    print("  ※ 열을 물리적으로 옮기지는 않습니다 — 27개 시트의 수식이 전부 어긋납니다.")
    if not do:
        print("\n실제로 고치려면:  python fix_freeze.py --apply")
        return 0

    zin = zipfile.ZipFile(src)
    changed = {}
    for sn, _cur, new, _k in jobs:
        sp = sheet_xml_path(zin, sn)
        changed[sp] = patch_pane(zin.read(sp).decode("utf-8"), new).encode("utf-8")

    mm = re.search(r"_v(\d+)\.xlsx$", src)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm.group(1)) + 1}.xlsx", src)
    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close(); zin.close()

    try:
        verify(src, tmp, jobs, set(changed))
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise
    os.replace(tmp, dst)
    print(f"\n{len(jobs)}개 시트 고정 수정 → {os.path.basename(dst)}")
    return 0


def verify(src, out, jobs, allowed):
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"
    wa = openpyxl.load_workbook(src, data_only=True)
    wb_ = openpyxl.load_workbook(out, data_only=True)
    for sn, _c, new, _k in jobs:
        got = str(wb_[sn].freeze_panes or "")
        assert got == new, f"{sn} 고정이 {got!r} (기대 {new!r})"
    # ★ 고정은 보기 설정일 뿐 — **값은 한 칸도 바뀌면 안 된다**
    for sn, _c, _n, _k in jobs:
        sa, sb = wa[sn], wb_[sn]
        assert sa.max_row == sb.max_row and sa.max_column == sb.max_column, f"{sn} 크기가 변했다"
        for r in range(1, min(sa.max_row, 60) + 1):
            for c in range(1, min(sa.max_column, 40) + 1):
                assert sa.cell(r, c).value == sb.cell(r, c).value, \
                    f"{sn}!{sa.cell(r, c).coordinate} 값이 변했다"
    assert wa.sheetnames == wb_.sheetnames, "시트 구성이 바뀌었다"
    wa.close(); wb_.close()
    zs = zipfile.ZipFile(src)
    other = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not other, f"의도치 않은 파트 변경: {other}"
    zs.close(); z.close()
    print("  검증 통과 — 값 변경 없음 · 다른 파트 변경 없음")


if __name__ == "__main__":
    sys.exit(main())
