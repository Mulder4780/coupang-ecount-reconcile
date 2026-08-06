# -*- coding: utf-8 -*-
"""data_status.py — **자료현황 한 장**. "그거 지금 몇 건이지?" 를 매번 다시 세지 않게.

사용자 지시(2026-07-29): "물어본 것들은 DB에 반영해서 다음에 빨리빨리 불러올 수 있게."

무엇을 답하나 — 그동안 반복해서 물어본 것들
  · 밴드에서 뭘 얼마나 가져왔나 (글·사진·원장 반영)
  · 카톡·ERP·쿠팡PO 자료를 지금 갖고 있나
  · 원장이 얼마나 채워졌나 (시트별 핵심 열)
  · 입금은 얼마나 들어왔나
  · 접속 주소는 무엇인가

설계 원칙 — **빨라야 쓴다.**
  · 느린 것(Z: 폴더 2만 개 순회)은 여기서 다시 돌지 않는다. `zscan`·`receipt_fill` 이
    남긴 리포트에서 숫자만 읽는다. 그 리포트들은 daily_run 이 갱신한다.
  · 원장은 read_only 로 한 번만 연다.

내보내는 것
  reports/자료현황.md    — 앱 [기록] 탭에 그대로 뜬다(app_server.latest_reports 에 등록)
  reports/자료현황.json  — 다른 도구가 쓰기 좋게

실행
  python data_status.py            # 집계 + 저장
  python data_status.py --print    # 화면에만
"""
import glob
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE, "reports")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def _num(s, default=0):
    try:
        return int(str(s).replace(",", ""))
    except (TypeError, ValueError):
        return default


def from_report(pattern, *patterns):
    """이미 만들어 둔 리포트에서 숫자만 긁어온다(느린 재계산을 피한다)."""
    fs = sorted(glob.glob(os.path.join(REPORT_DIR, pattern)))
    if not fs:
        return {}
    text = open(fs[-1], encoding="utf-8").read()
    out = {"_file": os.path.basename(fs[-1])}
    for key, rx in patterns:
        m = re.search(rx, text)
        out[key] = _num(m.group(1)) if m else None
    return out


def image_tree_stats(root, year="2026"):
    """서버 문서사진의 해당 연도 파일 수·용량을 하위 폴더까지 센다.

    앱 통계에는 2026년만 표시한다. 사진은 YYYY/MM/날짜 구조로 보관되므로
    연도 폴더부터 훑어 과거 자료와 무관한 폴더를 건드리지 않는다.
    """
    base = os.path.join(root, str(year))
    if not os.path.isdir(base):
        return 0, 0.0
    count, total = 0, 0
    pending = [base]
    while pending:
        folder = pending.pop()
        try:
            entries = list(os.scandir(folder))
        except OSError:
            continue
        for entry in entries:
            try:
                if entry.is_dir(follow_symlinks=False):
                    pending.append(entry.path)
                elif entry.is_file(follow_symlinks=False) and os.path.splitext(entry.name)[1].lower() in (
                    ".jpg", ".jpeg", ".png"
                ):
                    count += 1
                    total += entry.stat(follow_symlinks=False).st_size
            except OSError:
                continue
    return count, total / 1e6


def band_status():
    from band_extract import load_records
    from source_dirs import DOC_PHOTO_DIRS
    # 앱·보고 통계는 사용자 지시대로 2026년만 집계한다. 과거 원본은
    # 보존하되 레코드 수·기간·월별 숫자에 섞지 않는다.
    recs = [
        row for row in load_records()
        if str(row.get("게시일") or "").startswith("2026-")
    ]
    days = sorted(x.get("게시일", "") for x in recs if x.get("게시일"))
    months = Counter((x.get("게시일") or "")[:7] for x in recs if x.get("게시일"))
    kinds = Counter(x.get("업무유형", "") for x in recs)
    photos, mb = image_tree_stats(DOC_PHOTO_DIRS[0], year="2026")
    return {"레코드": len(recs), "기간": "%s ~ %s" % (days[0], days[-1]) if days else "-",
            "월별": dict(sorted(months.items())), "업무유형": dict(kinds.most_common()),
            "사진": photos, "사진MB": round(mb), "OCR캐시": len(glob.glob(os.path.join(BASE, "band", "ocr_cache", "*")))}


def inbox_status():
    """지금 어떤 원천 자료를 갖고 있나 — 없으면 뭘 넣어야 하는지 바로 보인다."""
    labels = {
        # ★ 'ledger' 는 **거래처별**계정별원장이다(2026-08-06). 예전엔 이 이름표가
        #   '계정별원장' 이라 분개장·현금출납장까지 한 칸에 뭉쳐 세고 있었다 —
        #   화면은 "원장 6건 있음" 이라 답하는데 정작 대조에 쓸 수 있는 건 3건이었다.
        "ledger": "거래처별계정별원장",
        "acctledger": "계정별원장",
        "journal": "분개장",
        "cashbook": "현금출납장",
        "tax": "세금계산서현황",
        "stmt": "거래명세서현황",
        "slips": "회계거래",
        "taxinv": "매출계산서조회",
        "po": "쿠팡PO·판매조회",
    }
    out = {label: 0 for label in labels.values()}
    try:
        from inbox_scan import scan, INBOX_DIR
        from source_dirs import ORIGIN_ROOT, excel_dirs
        watched = []
        for folder in [*excel_dirs(), INBOX_DIR]:
            if os.path.isdir(folder) and folder not in watched:
                watched.append(folder)
        source_mtimes = {
            os.path.normcase(folder): os.stat(folder).st_mtime_ns for folder in watched
        }

        # 자료 개수는 내용 대조 결과가 아니라 "보유 여부" 안내용이다. 같은
        # 네트워크 Excel을 상태 화면을 열 때마다 다시 여는 대신 폴더가 바뀌지
        # 않았으면 직전 집계를 즉시 재사용한다.
        previous_path = os.path.join(REPORT_DIR, "자료현황.json")
        if os.path.exists(previous_path):
            previous = json.load(open(previous_path, encoding="utf-8"))
            cached = previous.get("보유자료") or {}
            if cached and (
                not cached.get("_source_mtimes")
                or cached.get("_source_mtimes") == source_mtimes
            ):
                for label in labels.values():
                    out[label] = cached.get(label, 0)
                out["_source_mtimes"] = source_mtimes
                out["카톡 내보내기"] = len(
                    glob.glob(os.path.join(BASE, "kakao", "inbox", "*.txt"))
                )
                return out

        # pick()을 유형별로 여섯 번 부르면 원본 네트워크 폴더의 같은 Excel을
        # 여섯 번 다시 열어 1분 넘게 걸린다. 원본+로컬을 한 번씩만 분류한다.
        seen = set()
        for folder in (ORIGIN_ROOT, INBOX_DIR):
            if not os.path.isdir(folder):
                continue
            for path, kind in scan(folder):
                key = (os.path.basename(path).lower(), os.path.getsize(path))
                if key in seen:
                    continue
                seen.add(key)
                if kind in labels:
                    out[labels[kind]] += 1
        out["_source_mtimes"] = source_mtimes
    except Exception:
        out = {label: None for label in labels.values()}
    out["카톡 내보내기"] = len(glob.glob(os.path.join(BASE, "kakao", "inbox", "*.txt")))
    return out


def ledger_status():
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from workbook_patch import latest_master
    path, ver = latest_master()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"버전": ver, "시트": {}}
    spec = {
        "02_돌발AS접수": ["ERP등록", "완료보고서등록", "사진등록", "밴드 바로가기"],
        "04_정기점검": ["ERP판매전표", "거래명세서", "점검사진"],
        "06_거래서류청구수금": ["거래명세서합계", "세금계산서합계", "입금액"],
    }
    for sheet, cols in spec.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        if "프로젝트NO" not in hdr:
            continue
        jp = hdr.index("프로젝트NO")
        idx = {c: hdr.index(c) for c in cols if c in hdr}
        rows = 0
        fill = {c: 0 for c in idx}
        for r in ws.iter_rows(min_row=5, values_only=True):
            if not r[jp]:
                continue
            rows += 1
            for c, j in idx.items():
                if j < len(r) and r[j] not in (None, ""):
                    fill[c] += 1
        out["시트"][sheet] = {"행": rows, "열": {c: {"채움": v, "빈칸": rows - v} for c, v in fill.items()}}
    wb.close()
    return out


def endpoint_status():
    ep = os.path.join(BASE, "docs", "endpoint.json")
    url = ""
    try:
        url = json.load(open(ep, encoding="utf-8")).get("url", "")
    except Exception:
        pass
    return {"고정 진입점": "https://mulder4780.github.io/coupang-ecount-reconcile/",
            "전체기능 앱": url or "(미게시)", "PC 로컬": "http://localhost:8899"}


def collect():
    st = {"기준시각": datetime.now().strftime("%Y-%m-%d %H:%M")}
    for key, fn in (("밴드", band_status), ("보유자료", inbox_status),
                    ("원장", ledger_status), ("접속주소", endpoint_status)):
        try:
            st[key] = fn()
        except Exception as e:
            st[key] = {"오류": str(e)[:80]}
    st["입금"] = from_report("입금현황.md",
                             ("건수", r"입금\s*([\d,]+)건"), ("합계", r"합계\s*([\d,]+)원"))
    st["Z폴더서류"] = from_report("Z폴더_서류대조.md",
                                  ("서류PDF", r"서류 PDF\s*\*{0,2}?([\d,]+)"),
                                  ("확정", r"1:1 확정 \*{0,2}([\d,]+)"),
                                  ("짝없음", r"짝이 없는 서류 \(([\d,]+)"))
    return st


def to_md(st):
    L = ["# 자료현황 — 한 장으로 보기", "",
         "- 기준: %s · 관리대장 v%s" % (st["기준시각"], (st.get("원장") or {}).get("버전", "?")),
         "- 이 문서는 `data_status.py` 가 만든다(daily_run 이 매일 갱신). 매번 다시 세지 말 것.", ""]
    b = st.get("밴드", {})
    L += ["## 밴드", "", "| 항목 | 값 |", "|---|---:|",
          "| 추출 레코드 | %s건 |" % b.get("레코드"),
          "| 기간 | %s |" % b.get("기간"),
          "| 사진(서버 보관) | %s장 · %s MB |" % (b.get("사진"), b.get("사진MB")),
          "| OCR 처리 | %s건 |" % b.get("OCR캐시"), ""]
    if b.get("월별"):
        L += ["월별: " + " · ".join("%s %d" % kv for kv in b["월별"].items()), ""]
    if b.get("업무유형"):
        L += ["유형: " + " · ".join("%s %d" % kv for kv in b["업무유형"].items()), ""]
    L += ["> 밴드 사진은 OCR 로 값이 거의 안 나온다(현장 사진 위주).",
          "> 서류 금액·번호는 Z폴더 PDF 가 정본이다.", ""]

    z = st.get("Z폴더서류") or {}
    if z.get("서류PDF"):
        L += ["## Z폴더 서류(거래명세서·세금계산서)", "", "| 항목 | 값 |", "|---|---:|",
              "| 서류 PDF | %s장 |" % z.get("서류PDF"),
              "| 원장과 1:1 확정 | %s건 |" % z.get("확정"),
              "| 원장에 짝 없음 | %s건 |" % z.get("짝없음"), ""]

    r = st.get("입금") or {}
    if r.get("건수"):
        L += ["## 입금", "", "| 항목 | 값 |", "|---|---:|",
              "| 입금 | %s건 |" % format(r["건수"], ","),
              "| 합계 | %s원 |" % format(r.get("합계") or 0, ","), ""]

    L += ["## 지금 갖고 있는 원천 자료", "", "| 자료 | 파일 수 |", "|---|---:|"]
    for k, v in (st.get("보유자료") or {}).items():
        if str(k).startswith("_"):
            continue
        L.append("| %s | %s |" % (k, "확인불가" if v is None else v))
    L.append("")

    led = st.get("원장") or {}
    if led.get("시트"):
        L += ["## 관리대장 채움 현황 (v%s)" % led.get("버전"), ""]
        for sheet, info in led["시트"].items():
            L += ["**%s** — %d행" % (sheet, info["행"]), "", "| 열 | 채움 | 빈칸 |", "|---|---:|---:|"]
            for c, v in info["열"].items():
                L.append("| %s | %d | %d |" % (c, v["채움"], v["빈칸"]))
            L.append("")

    e = st.get("접속주소") or {}
    L += ["## 접속 주소", "", "| 용도 | 주소 |", "|---|---|"]
    for k, v in e.items():
        L.append("| %s | %s |" % (k, v))
    L.append("")
    return "\n".join(L)


def main():
    st = collect()
    md = to_md(st)
    if "--print" in sys.argv:
        print(md)
        return 0
    os.makedirs(REPORT_DIR, exist_ok=True)
    open(os.path.join(REPORT_DIR, "자료현황.md"), "w", encoding="utf-8").write(md)
    json.dump(st, open(os.path.join(REPORT_DIR, "자료현황.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1, default=str)
    b = st.get("밴드", {})
    led = st.get("원장", {})
    print("자료현황 갱신 — 밴드 %s건·사진 %s장 · 관리대장 v%s"
          % (b.get("레코드"), b.get("사진"), led.get("버전")))
    print("  reports/자료현황.md · 자료현황.json (앱 [기록] 탭에서 바로 보입니다)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
