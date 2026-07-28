# -*- coding: utf-8 -*-
"""밴드·카톡 완료보고의 문서 발행 근거를 ERP 상태 칸에 반영한다.

근거:
  - 프로젝트NO가 정확히 일치하는 완료보고
  - 보고 본문 문서상태에 '판매전표' 또는 '거래명세서'가 명시됨
  - 관리대장 업무 자체가 이미 완료 상태임

빈칸만 채우며 수식·사람 입력은 덮지 않는다. 반영은 ZIP 내부 시트 XML만 패치한다.
"""
from __future__ import annotations

import os
import re
import sys
import zipfile
from collections import Counter

from operation_window import input_window_label, is_input_window
from workbook_patch import esc, latest_master, sheet_xml_path

ROOT = os.path.dirname(os.path.abspath(__file__))
TARGETS = (
    ("02_돌발AS접수", "ERP등록", "진행상태", "작업완료", "판매전표", "완료"),
    ("04_정기점검", "ERP판매전표", "점검상태", "완료", "판매전표", "완료"),
    ("04_정기점검", "거래명세서", "점검상태", "완료", "거래명세서", "발행완료"),
)


def _s(value):
    return "" if value is None else str(value).strip()


def document_evidence():
    from band_extract import load_records

    evidence = {}
    for row in load_records():
        project = _s(row.get("프로젝트NO"))
        if not re.fullmatch(r"UJ\d{7}", project):
            continue
        if _s(row.get("진행상태")) != "작업완료":
            continue
        docs = {_s(x) for x in _s(row.get("문서상태")).split("+") if _s(x)}
        if docs:
            evidence.setdefault(project, set()).update(docs)
    return evidence


def plan(path):
    import openpyxl
    from openpyxl.utils import get_column_letter

    evidence = document_evidence()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fills, counts = {}, Counter()
    for sheet, column, status_column, done_value, document, result in TARGETS:
        ws = wb[sheet]
        headers = [_s(v) for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        index = {name: i for i, name in enumerate(headers) if name}
        letter = get_column_letter(index[column] + 1)
        for row_no, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            if not row[0]:
                continue
            get = lambda name: _s(row[index[name]]) if index[name] < len(row) else ""
            if get(status_column) != done_value or get(column):
                continue
            project = get("프로젝트NO")
            if document not in evidence.get(project, set()):
                continue
            fills[f"{sheet}!{letter}{row_no}"] = result
            counts[f"{sheet}:{column}"] += 1
    wb.close()
    return fills, counts, evidence


def replace_cells(xml, fills):
    seen = set()

    def repl(match):
        ref, attrs = match.group(1), match.group(2) or ""
        if ref not in fills or 't="shared"' in match.group(0) or 't="array"' in match.group(0):
            return match.group(0)
        seen.add(ref)
        attrs = re.sub(r'\s+t="[^"]*"', "", attrs)
        return f'<c r="{ref}"{attrs} t="inlineStr"><is><t>{esc(fills[ref])}</t></is></c>'

    out = re.sub(
        r'<c r="([A-Z]+\d+)"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)',
        repl,
        xml,
        flags=re.S,
    )
    return out, seen


def verify(src, out, by_sheet, changed):
    import openpyxl

    archive = zipfile.ZipFile(out)
    assert archive.testzip() is None, "zip 무결성 실패"
    result = openpyxl.load_workbook(out, read_only=True, data_only=True)
    for sheet, cells in by_sheet.items():
        for ref, value in cells.items():
            assert _s(result[sheet][ref].value) == value, f"{sheet}!{ref} 반영 실패"
    result.close()
    formulas = openpyxl.load_workbook(out, read_only=True, data_only=False)
    for ref in ("AB5", "AD5"):
        assert str(formulas["04_정기점검"][ref].value or "").startswith("="), f"{ref} 수식 손상"
    formulas.close()
    original = zipfile.ZipFile(src)
    other = [
        name for name in original.namelist()
        if name not in changed and original.read(name) != archive.read(name)
    ]
    assert not other, f"의도치 않은 파트 변경: {other}"
    original.close()
    archive.close()


def apply(path, fills):
    from dashboard_clean import force_recalc

    by_sheet = {}
    for key, value in fills.items():
        sheet, ref = key.split("!", 1)
        by_sheet.setdefault(sheet, {})[ref] = value
    source = zipfile.ZipFile(path)
    changed = {}
    for sheet, cells in by_sheet.items():
        part = sheet_xml_path(source, sheet)
        xml, seen = replace_cells(source.read(part).decode("utf-8"), cells)
        missing = set(cells) - seen
        assert not missing, f"{sheet} XML 셀 누락: {sorted(missing)[:5]}"
        changed[part] = xml.encode("utf-8")
    changed["xl/workbook.xml"] = force_recalc(
        source.read("xl/workbook.xml").decode("utf-8")
    ).encode("utf-8")

    version = int(re.search(r"_v(\d+)\.xlsx$", path).group(1))
    out = re.sub(r"_v\d+\.xlsx$", f"_v{version + 1}.xlsx", path)
    if os.path.exists(out):
        raise FileExistsError(os.path.basename(out))
    temp = out[:-5] + ".tmp.xlsx"
    archive = zipfile.ZipFile(temp, "w", zipfile.ZIP_DEFLATED)
    for item in source.infolist():
        archive.writestr(item.filename, changed.get(item.filename, source.read(item.filename)))
    archive.close()
    source.close()
    try:
        verify(path, temp, by_sheet, set(changed))
    except BaseException:
        try:
            os.remove(temp)
        except OSError:
            pass
        raise
    os.replace(temp, out)
    return out


def main():
    if is_input_window():
        print(f"입력 보호시간({input_window_label()}) — ERP 문서 상태 반영 생략")
        return
    do_apply = "--apply" in sys.argv
    do_queue = "--queue" in sys.argv
    source = latest_master()[0]
    fills, counts, evidence = plan(source)
    print(f"원본: {os.path.basename(source)}")
    print(f"완료보고 문서 근거: {len(evidence)}개 프로젝트")
    print(f"반영 대상: {len(fills)}칸")
    for name, count in counts.items():
        print(f"  {name} {count}칸")
    if not do_apply and not do_queue:
        print("미리보기만 완료 — 큐 추가: python fill_erp_documents.py --queue")
        print("즉시 반영: python fill_erp_documents.py --apply")
        return
    if do_queue:
        import ledger_writer

        items = []
        for key, value in fills.items():
            sheet, ref = key.split("!", 1)
            items.append({
                "sheet": sheet,
                "row": int(re.search(r"\d+$", ref).group(0)),
                "col": {
                    "02_돌발AS접수": {"ERP등록": "ERP등록"},
                    "04_정기점검": {
                        "ERP판매전표": "ERP판매전표",
                        "거래명세서": "거래명세서",
                    },
                }[sheet][next(
                    column for target_sheet, column, *_rest in TARGETS
                    if target_sheet == sheet and value in (
                        "발행완료" if column == "거래명세서" else "완료",
                    )
                )],
                "value": value,
                "vtype": "text",
                "only_if_empty": True,
                "evidence": "밴드·카톡 완료보고 문서상태",
            })
        added = ledger_writer.queue_add(items)
        print(f"자동입력 큐 추가 {added}건")
        return
    from claim_guard import require

    require("ledger", "fill_erp_documents")
    if not fills:
        print("새로 반영할 칸 없음")
        return
    out = apply(source, fills)
    print(f"검증 통과 — {len(fills)}칸 반영 → {os.path.basename(out)}")


if __name__ == "__main__":
    main()
