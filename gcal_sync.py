# -*- coding: utf-8 -*-
"""
gcal_sync.py — 구글 캘린더(COUPANG 설치+납품+AS)를 원장·앱과 상시 대조
===============================================================================
사용자 지시(2026-07-29): "이 캘린더 추가하고 **항상** 대조해서 엑셀과 앱에 반영해줘"
  → daily_run 이 매일 돌리고, 결과는 관리대장 예정일 칸과 앱 '일정' 탭에 들어간다.

무엇을 어디에 넣나 (일정 = **예정**이지 실적이 아니다 — 완료 칸은 절대 건드리지 않는다)
  · 02_돌발AS접수   N 방문예정일 · O 방문예정시간
  · 04_정기점검     D 점검예정일 · E 점검예정시간
  · 05_신규납품설치 R 납품예정일 · T 설치예정일 · V 철거·이전예정일
  실제완료일(02 R·04 H·05 S/U/W)은 캘린더로 채우지 않는다. 캘린더는 "하기로 한 날"이고
  완료는 밴드·카톡·ERP가 증거다. 둘을 섞으면 "안 한 일이 한 일"이 된다.

접속 방법 — 공개 전환이 필요 없다
  이 캘린더는 비공개 공유라 `public/basic.ics` 는 404다(확인함). 공개로 돌리면 전 세계에
  일정이 열리므로 그렇게 하지 않는다. 대신 구글 캘린더 설정 맨 아래 **비공개 주소(iCal)**
  한 줄만 `config/gcal.json` 에 넣으면 된다. 그 주소는 비밀키다 — config/ 는 커밋 금지(규칙 1).

    config/gcal.json
    {"calendars": [{"name": "COUPANG 설치+납품+AS", "id": "...@group.calendar.google.com",
                    "ics": "https://calendar.google.com/calendar/ical/.../private-xxxx/basic.ics"}]}

  주소가 아직 없어도 동작한다: `.ics` 파일을 내려받아 인박스에 놓아도 되고(아래 ICS_DIRS),
  나중에 공개로 바뀌면 공개 피드를 자동으로 시도한다. **원천이 없으면 조용히 넘어간다** —
  daily_run 이 이것 때문에 멈추면 안 된다.

사용
  python gcal_sync.py                 # 대조 결과만 보기
  python gcal_sync.py --queue         # 예정일 빈 칸 채우기 큐 적재
  python gcal_sync.py --self-test     # 합성 검증
"""
import sys, os, re, json, glob, urllib.parse, urllib.request
from datetime import datetime, date, timedelta

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

HDR, FIRST = 4, 5
YEAR = 2026                      # ★ 2025는 DB만 보관하고 작업하지 않는다(사용자 지시)
PRJ = re.compile(r"UJ\d{7}")
CACHE = os.path.join(ROOT, "reports", "gcal_events.json")

# 일정 종류 판정 — 앞에 오는 것이 이긴다(철거가 '납품 후 철거'보다 강하다)
KIND_RULES = [
    ("철거", re.compile(r"철거|이전|반출|회수")),
    ("납품", re.compile(r"납품|설치|입고|출고")),
    ("정기점검", re.compile(r"정기\s*점검|정기|점검|PM")),
    ("돌발AS", re.compile(r"A/?S|고장|수리|보수|긴급|돌발")),
]
# 업무구분 → (시트, 예정일 열, 예정시간 열)
TARGET = {
    "돌발AS":  ("02_돌발AS접수",   "방문예정일",     "방문예정시간"),
    "정기점검": ("04_정기점검",     "점검예정일",     "점검예정시간"),
    "납품":    ("05_신규납품설치", "납품예정일",     None),
    "철거":    ("05_신규납품설치", "철거·이전예정일", None),
}


def _s(v):
    return "" if v is None else str(v).strip()


# ── ICS 파서 (표준 라이브러리만) ──────────────────────────────────
def unfold(text):
    """RFC5545 접힘 줄 풀기 — 다음 줄이 공백/탭으로 시작하면 앞 줄에 이어 붙는다.
    이걸 빼먹으면 긴 SUMMARY 가 중간에서 잘려 캠프명·UJ번호를 놓친다."""
    out = []
    for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if line[:1] in (" ", "\t") and out:
            out[-1] += line[1:]
        else:
            out.append(line)
    return out


def parse_dt(val, params):
    """DTSTART 값 → (date, 'HH:MM' 또는 '')
    · VALUE=DATE  : 20260701          → 종일 일정
    · TZID/UTC    : 20260701T090000Z  → 시각 있음(한국시간으로 본다)"""
    v = val.strip()
    if params.get("VALUE") == "DATE" or (len(v) == 8 and v.isdigit()):
        try:
            return date(int(v[:4]), int(v[4:6]), int(v[6:8])), ""
        except ValueError:
            return None, ""
    m = re.match(r"^(\d{4})(\d{2})(\d{2})T(\d{2})(\d{2})(\d{2})(Z?)$", v)
    if not m:
        return None, ""
    y, mo, d, hh, mi, _ss, z = m.groups()
    dt = datetime(int(y), int(mo), int(d), int(hh), int(mi))
    if z == "Z":                                  # UTC → 한국시간(고정 +9, DST 없음)
        dt += timedelta(hours=9)
    return dt.date(), f"{dt.hour:02d}:{dt.minute:02d}"


def parse_ics(text):
    """VEVENT 목록 → [{uid, 제목, 설명, 장소, 날짜, 시간, 취소여부}]"""
    events, cur = [], None
    for line in unfold(text):
        if line == "BEGIN:VEVENT":
            cur = {"uid": "", "제목": "", "설명": "", "장소": "", "날짜": None,
                   "시간": "", "취소": False}
            continue
        if line == "END:VEVENT":
            if cur is not None:
                events.append(cur)
            cur = None
            continue
        if cur is None or ":" not in line:
            continue
        head, val = line.split(":", 1)
        bits = head.split(";")
        name = bits[0].upper()
        params = {}
        for b in bits[1:]:
            if "=" in b:
                k, v = b.split("=", 1)
                params[k.upper()] = v
        val = val.replace("\\n", "\n").replace("\\,", ",").replace("\\;", ";").replace("\\\\", "\\")
        if name == "UID":
            cur["uid"] = val.strip()
        elif name == "SUMMARY":
            cur["제목"] = val.strip()
        elif name == "DESCRIPTION":
            cur["설명"] = val.strip()
        elif name == "LOCATION":
            cur["장소"] = val.strip()
        elif name == "STATUS":
            cur["취소"] = val.strip().upper() == "CANCELLED"
        elif name == "DTSTART":
            cur["날짜"], cur["시간"] = parse_dt(val, params)
    return events


# ── 일정 해석 ────────────────────────────────────────────────
def kind_of(text):
    for k, rx in KIND_RULES:
        if rx.search(text or ""):
            return k
    return ""


def project_of(text):
    m = PRJ.search(text or "")
    return m.group(0) if m else ""


def enrich(ev):
    """제목+설명+장소를 한 덩어리로 보고 업무구분·프로젝트NO를 뽑는다."""
    blob = " ".join((ev.get("제목", ""), ev.get("설명", ""), ev.get("장소", "")))
    return {**ev, "업무구분": kind_of(blob), "프로젝트NO": project_of(blob)}


def in_scope(ev):
    """2026년 · 취소 아님 · 날짜 있음 — 셋 다여야 대조 대상."""
    d = ev.get("날짜")
    return bool(d) and not ev.get("취소") and d.year == YEAR


def norm_camp(s):
    """캠프명 비교용 정규화 — 괄호·공백·중점 제거. '송파5MB(감일동)' ↔ '송파5MB 감일동'."""
    s = re.sub(r"[\s·\-_]", "", _s(s))
    return re.sub(r"[()\[\]]", "", s)


def match_row(ev, rows, slack=3):
    """일정 하나 → 원장 행 하나. 근거가 강한 순서로만 붙인다.

    1) 프로젝트NO 일치 — 유일하고 확실하다
    2) 캠프명 + 업무구분 + 날짜 ±slack일 — 후보가 **정확히 하나일 때만** 붙인다
    후보가 여럿이면 붙이지 않는다. 애매한 걸 붙이면 남의 건에 예정일이 들어간다."""
    prj = ev.get("프로젝트NO")
    if prj:
        hit = [r for r in rows if r["프로젝트NO"] == prj]
        if len(hit) == 1:
            return hit[0], "프로젝트NO"
    camp, kind, d = norm_camp(ev.get("장소") or ev.get("제목")), ev.get("업무구분"), ev.get("날짜")
    if not (camp and kind and d):
        return None, ""
    cand = [r for r in rows
            if r["업무구분"] == kind and norm_camp(r["캠프명"])
            and (norm_camp(r["캠프명"]) in camp or camp in norm_camp(r["캠프명"]))
            and r["날짜"] and abs((r["날짜"] - d).days) <= slack]
    return (cand[0], "캠프명+날짜") if len(cand) == 1 else (None, "")


# ── 원천 ─────────────────────────────────────────────────────
def feeds():
    """(이름, 주소종류, 값) 목록. 비밀주소 > 인박스 .ics > 공개피드 순."""
    out = []
    cfg = os.path.join(ROOT, "config", "gcal.json")
    if os.path.exists(cfg):
        try:
            for c in json.load(open(cfg, encoding="utf-8")).get("calendars", []):
                if c.get("ics"):
                    out.append((c.get("name", "캘린더"), "url", c["ics"]))
                elif c.get("id"):
                    out.append((c.get("name", "캘린더"), "url", public_ics(c["id"])))
        except Exception:
            pass
    env = os.environ.get("COUPANG_GCAL_ICS", "").strip()
    if env:
        out.append(("환경변수", "url", env))
    for d in ics_dirs():
        for p in sorted(glob.glob(os.path.join(d, "*.ics"))):
            out.append((os.path.basename(p), "file", p))
    return out


def ics_dirs():
    try:
        from source_dirs import existing, LEDGER_DIR
        return existing([os.path.join(LEDGER_DIR, "0. 원본 자료", "3. 캘린더"),
                         os.path.join(ROOT, "inbox")])
    except Exception:
        return [d for d in (os.path.join(ROOT, "inbox"),) if os.path.isdir(d)]


def public_ics(cal_id):
    return ("https://calendar.google.com/calendar/ical/"
            + urllib.parse.quote(cal_id) + "/public/basic.ics")


def fetch(kind, value, timeout=30):
    if kind == "file":
        return open(value, encoding="utf-8", errors="replace").read()
    req = urllib.request.Request(value, headers={"User-Agent": "coupang-work/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", "replace")


def collect():
    """모든 원천에서 2026년 일정을 모은다 → (이벤트, 원천보고)"""
    evs, notes = {}, []
    for name, kind, value in feeds():
        try:
            text = fetch(kind, value)
        except Exception as e:
            notes.append(f"{name}: 읽기 실패 ({type(e).__name__})")
            continue
        raw = [enrich(e) for e in parse_ics(text)]
        keep = [e for e in raw if in_scope(e)]
        for e in keep:
            evs[e["uid"] or f"{e['제목']}|{e['날짜']}"] = e
        notes.append(f"{name}: 전체 {len(raw)}건 → 2026년 {len(keep)}건")
    return sorted(evs.values(), key=lambda e: (e["날짜"], e["제목"])), notes


# ── 원장 ─────────────────────────────────────────────────────
def ledger_rows(path):
    """02/04/05 → 대조용 평면 목록"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    spec = (("02_돌발AS접수", "돌발AS", "접수일자", "방문예정일"),
            ("04_정기점검", "정기점검", "점검예정일", "점검예정일"),
            ("05_신규납품설치", "납품", "요청일", "납품예정일"))
    rows = []
    for sn, kind, datecol, planned in spec:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdr = next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))
        H = {_s(h): i for i, h in enumerate(hdr) if h is not None}
        for i, r in enumerate(ws.iter_rows(min_row=FIRST, values_only=True), FIRST):
            wid = _s(r[0]) if r else ""
            if not wid:
                continue
            def g(col):
                j = H.get(col)
                return r[j] if j is not None and j < len(r) else None
            d = g(planned) or g(datecol)
            d = d.date() if isinstance(d, datetime) else (d if isinstance(d, date) else None)
            rows.append({"행": i, "시트": sn, "업무구분": kind, "원천업무ID": wid,
                         "프로젝트NO": _s(g("프로젝트NO")), "캠프명": _s(g("캠프명")),
                         "날짜": d, "예정일있음": g(planned) is not None})
    wb.close()
    return rows


def reconcile(events, rows):
    matched, only_cal = [], []
    for ev in events:
        row, how = match_row(ev, rows)
        (matched.append({**ev, "원장": row, "근거": how}) if row else only_cal.append(ev))
    return matched, only_cal


# ── 합성 검증 ─────────────────────────────────────────────────
SAMPLE = """BEGIN:VCALENDAR
BEGIN:VEVENT
UID:a1
SUMMARY:[정기점검] 김해2캠프 UJ26011
 41
DTSTART;VALUE=DATE:20260701
LOCATION:김해2캠프
END:VEVENT
BEGIN:VEVENT
UID:a2
SUMMARY:송파5MB(감일동) A/S 방문
DTSTART;TZID=Asia/Seoul:20260702T090000
LOCATION:송파5MB 감일동
END:VEVENT
BEGIN:VEVENT
UID:a3
SUMMARY:철거 반출 - 삼선동OFC
DTSTART:20260703T000000Z
STATUS:CANCELLED
END:VEVENT
BEGIN:VEVENT
UID:a4
SUMMARY:작년 점검
DTSTART;VALUE=DATE:20251201
END:VEVENT
END:VCALENDAR"""


def self_test():
    bad = 0
    evs = [enrich(e) for e in parse_ics(SAMPLE)]
    by = {e["uid"]: e for e in evs}
    # 접힘 줄: UJ2601141 이 잘리지 않아야 한다
    if by["a1"]["프로젝트NO"] != "UJ2601141":
        print("  [FAIL] 접힘 줄 →", by["a1"]["프로젝트NO"]); bad += 1
    if by["a1"]["업무구분"] != "정기점검" or by["a1"]["시간"] != "":
        print("  [FAIL] 종일 일정", by["a1"]["업무구분"], by["a1"]["시간"]); bad += 1
    if by["a2"]["업무구분"] != "돌발AS" or by["a2"]["시간"] != "09:00":
        print("  [FAIL] 시각 파싱", by["a2"]["시간"]); bad += 1
    # UTC 09:00Z → 한국 18:00 (00:00Z → 09:00)
    if by["a3"]["시간"] != "09:00":
        print("  [FAIL] UTC 변환", by["a3"]["시간"]); bad += 1
    if by["a3"]["업무구분"] != "철거":
        print("  [FAIL] 철거 우선", by["a3"]["업무구분"]); bad += 1
    # 범위: 취소·2025년은 빠진다
    scope = [e["uid"] for e in evs if in_scope(e)]
    if scope != ["a1", "a2"]:
        print("  [FAIL] 범위", scope); bad += 1

    rows = [{"행": 5, "시트": "04_정기점검", "업무구분": "정기점검", "원천업무ID": "PM-2607-001",
             "프로젝트NO": "UJ2601141", "캠프명": "김해2캠프", "날짜": date(2026, 7, 1), "예정일있음": False},
            {"행": 6, "시트": "02_돌발AS접수", "업무구분": "돌발AS", "원천업무ID": "AS-2607-001",
             "프로젝트NO": "UJ2600975", "캠프명": "송파5MB(감일동)", "날짜": date(2026, 7, 3), "예정일있음": False},
            {"행": 7, "시트": "02_돌발AS접수", "업무구분": "돌발AS", "원천업무ID": "AS-2607-009",
             "프로젝트NO": "UJ2600976", "캠프명": "송파5MB(감일동)", "날짜": date(2026, 7, 4), "예정일있음": False}]
    r, how = match_row(by["a1"], rows)
    if not r or how != "프로젝트NO" or r["원천업무ID"] != "PM-2607-001":
        print("  [FAIL] 프로젝트NO 매칭", how); bad += 1
    # 같은 캠프 후보가 둘이면 붙이지 않는다 — 애매하면 안 붙이는 게 규칙이다
    r2, how2 = match_row(by["a2"], rows)
    if r2 is not None:
        print("  [FAIL] 중복 후보를 붙였다", r2["원천업무ID"]); bad += 1
    r3, how3 = match_row(by["a2"], rows[:2])
    if not r3 or how3 != "캠프명+날짜" or r3["원천업무ID"] != "AS-2607-001":
        print("  [FAIL] 캠프명+날짜 매칭", how3); bad += 1
    # 실적(완료) 열은 어떤 경로로도 대상이 아니다
    if any(c in str(TARGET) for c in ("완료일", "실제")):
        print("  [FAIL] 완료 열이 대상에 들어 있다"); bad += 1
    print("gcal_sync self-test:", "OK" if not bad else f"{bad}건 실패")
    return bad == 0


# ── 실행 ─────────────────────────────────────────────────────
def main():
    if "--self-test" in sys.argv:
        sys.exit(0 if self_test() else 1)
    from workbook_patch import latest_master
    path, ver = latest_master()
    events, notes = collect()
    print(f"관리대장 v{ver} · 캘린더 원천 {len(feeds())}개")
    for n in notes:
        print("  ·", n)
    if not events:
        print("\n일정 원천이 아직 없습니다 — config/gcal.json 에 비공개 iCal 주소를 넣거나")
        print("  .ics 파일을", (ics_dirs() or ["<인박스 없음>"])[0], "에 놓아 주세요.")
        save_cache([], [], notes, ver)
        return
    rows = ledger_rows(path)
    matched, only_cal = reconcile(events, rows)
    fillable = [m for m in matched if not m["원장"]["예정일있음"]]
    print(f"\n2026년 일정 {len(events)}건")
    print(f"  원장과 연결 {len(matched)}건 (예정일 빈 칸 {len(fillable)}건 → 채울 수 있음)")
    print(f"  캘린더에만 있음 {len(only_cal)}건 ← 원장 누락 후보")
    from collections import Counter
    print("  구분:", dict(Counter(e["업무구분"] or "미상" for e in events)))
    save_cache(events, matched, notes, ver)
    write_report(events, matched, only_cal, ver)

    if "--queue" in sys.argv:
        import claim_guard
        claim_guard.require("ledger")
        q = build_queue(fillable)
        if q:
            add_queue(q)
        print(f"큐 적재 {len(q)}셀")


def build_queue(fillable):
    q = []
    for m in fillable:
        tgt = TARGET.get(m["업무구분"])
        if not tgt:
            continue
        sheet, daycol, timecol = tgt
        if sheet != m["원장"]["시트"]:      # 구분이 어긋나면 넣지 않는다
            continue
        base = {"sheet": sheet, "key_col": _keycol(sheet), "key": m["원장"]["원천업무ID"],
                "evidence": f"구글캘린더 {m['제목'][:40]}", "only_if_empty": True}
        q.append({**base, "col": daycol, "value": m["날짜"].isoformat(), "vtype": "date"})
        if timecol and m["시간"]:
            q.append({**base, "col": timecol, "value": m["시간"], "vtype": "text"})
    return q


def _keycol(sheet):
    return {"02_돌발AS접수": "접수ID", "04_정기점검": "점검ID", "05_신규납품설치": "납품ID"}.get(sheet, "ID")


def add_queue(items):
    qp = os.path.join(ROOT, "updates", "pending_updates.json")
    os.makedirs(os.path.dirname(qp), exist_ok=True)
    old = []
    if os.path.exists(qp):
        try:
            old = json.load(open(qp, encoding="utf-8"))
        except Exception:
            old = []
    json.dump(old + items, open(qp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)


def save_cache(events, matched, notes, ver):
    """앱이 읽는 캐시. 앱은 네트워크를 타지 않는다 — 폰에서 즉시 뜬다."""
    os.makedirs(os.path.dirname(CACHE), exist_ok=True)
    link = {id(m): m for m in matched}
    out = []
    for e in events:
        m = next((x for x in matched if x["uid"] == e["uid"]), None)
        out.append({"날짜": e["날짜"].isoformat(), "시간": e["시간"], "제목": e["제목"],
                    "장소": e["장소"], "업무구분": e["업무구분"], "프로젝트NO": e["프로젝트NO"],
                    "원천업무ID": (m or {}).get("원장", {}).get("원천업무ID", ""),
                    "연결근거": (m or {}).get("근거", "")})
    json.dump({"갱신": datetime.now().isoformat(timespec="seconds"), "관리대장": ver,
               "원천": notes, "일정": out},
              open(CACHE, "w", encoding="utf-8"), ensure_ascii=False, indent=1)


def write_report(events, matched, only_cal, ver):
    rp = os.path.join(ROOT, "reports", "캘린더_대조.md")
    with open(rp, "w", encoding="utf-8") as fh:
        fh.write(f"# 구글 캘린더 대조 (v{ver}, {datetime.now():%Y-%m-%d %H:%M})\n\n")
        fh.write(f"- 2026년 일정 {len(events)}건 / 원장 연결 {len(matched)}건 / 캘린더에만 {len(only_cal)}건\n\n")
        if only_cal:
            fh.write("## 캘린더에만 있는 일정 (원장 누락 후보)\n\n")
            for e in only_cal[:120]:
                fh.write(f"- {e['날짜']} {e['시간']} [{e['업무구분'] or '미상'}] {e['제목']} / {e['장소']}\n")
    print(f"\n리포트: {rp}")


if __name__ == "__main__":
    main()
