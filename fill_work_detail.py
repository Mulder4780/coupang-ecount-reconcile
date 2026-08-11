# -*- coding: utf-8 -*-
"""fill_work_detail.py — 작업 내용이 비어 있으면 **원문(카톡·밴드)에서 채운다**.

사용자 지시(2026-07-29): "작업 내용이 안 적힌 완료건인데, 밴드나 카톡에서 확인되면
자동 기입해서 정리하는 방향으로. **모든 데이터가 마찬가지야.**"

원칙 (이 프로젝트의 표준 경로다 — 새 칸을 채울 때도 이 순서를 따른다)
  1. 원문에 있으면 **자동으로 채운다.** 사람이 옮겨 적게 하지 않는다.
  2. 원문에 없으면 **비워 두고 확인 목록에 남긴다.** 지어내지 않는다(절대규칙 10).
  3. 사람이 적어 둔 값은 **절대 덮지 않는다.** 단 '(자동 초안)' 자리표시자는 값이 아니다.
  4. 무엇을 근거로 채웠는지 함께 적는다.

무엇을 채우나
  03_현장작업실적 · 실제작업상세   ← 카톡 **'A/S 내용'**(완료 글) → 밴드 **완료 글** 본문
  (이 칸이 비면 대표 브리핑에 "작업 내용이 안 적힌 완료건" 으로 뜬다 — daily_brief 의 '무엇')

★★ **'신청내용' 은 쓰지 않는다.** 그건 "무엇을 해 달라" 이지 "무엇을 했다" 가 아니다.
   카톡 자료를 세어 보면 작업내용(A/S 내용)은 26건뿐이고 신청내용만 있는 게 542건이다.
   신청내용을 실제작업상세에 넣으면 **하지 않은 작업을 했다고 기록**하는 셈이고,
   대표 보고의 '무엇을 했나' 가 통째로 거짓이 된다. 신청내용은 이미 02시트에 있다.
   같은 이유로 밴드도 **완료 글만** 쓴다(접수 글 본문도 요청이다).

실행
  python fill_work_detail.py            # 미리보기
  python fill_work_detail.py --apply    # 큐 적재 + 원장 반영(vN+1)
"""
import os
import re
import sys

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SHEET = "03_현장작업실적"
COL = "실제작업상세"
# 값이 아니라 '아직 안 적었다' 는 표시 — 덮어써도 된다.
#  · (자동 초안)…            : 이 프로젝트 도구가 넣어 둔 자리표시자
#  · What ? / 갯수 ? / ?? 호기 : 공지 서식의 **빈 양식**(기사가 아직 안 채운 칸)
#    ★ 이걸 값으로 보면 "작업을 했다" 고 기록돼 버린다. 실제로 그런 글이 있다.
DRAFT = re.compile(r"\(자동\s*초안\)|실제 작업내용을 입력하세요|What\s*\?|갯수\s*\?|\?\?\s*호기")
# 미기입 양식 조각 — 한 글에 여러 호기가 있고 일부만 적힌 경우 뒤쪽에 그대로 남는다
BLANK_UNIT = re.compile(r"\?\?\s*호기.*$", re.S)
MAXLEN = 200


def _s(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def kakao_text():
    """프로젝트NO → **무엇을 했나**(카톡 완료 글의 'A/S 내용').
    ★ 신청내용은 쓰지 않는다 — 요청이지 실적이 아니다(모듈 설명 참고)."""
    out = {}
    try:
        import kakao_extract as ke
        for r in ke.extract():
            k = r.get("프로젝트NO")
            if not k:
                continue
            t = _s(r.get("작업내용"))
            if t and len(t) > len(_s(out.get(k, ""))):
                out[k] = t
    except Exception:
        pass
    return out


def band_text():
    """프로젝트NO → 글 본문(밴드). 완료 글을 우선한다.

    `band_extract.load_records()` 는 본문을 들고 있지 않다(게시글 필드는 출처 id 다).
    그래서 캐시 원문에서 직접 읽는다."""
    import glob
    import json
    out = {}
    RE = re.compile(r"UJ\d{7}")
    for f in sorted(glob.glob(os.path.join(BASE, "band", "cache", "dump_*.json")) +
                    glob.glob(os.path.join(BASE, "band", "cache", "raw_*.json"))):
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        posts = d.get("posts") or {}
        for _pid, p in (posts.items() if isinstance(posts, dict) else enumerate(posts)):
            body = _s(p.get("content"))
            if not body:
                continue
            m = RE.search(body)
            if not m:
                continue
            k = m.group(0)
            # ★ **완료 글만** 쓴다. 접수 글 본문은 '해 달라' 이지 '했다' 가 아니다.
            if "완료" not in body[:60]:
                continue
            cur = out.get(k)
            if cur is None or len(body) > len(cur):
                out[k] = body
    return out


RE_AS = re.compile(r"[●∙•]\s*A\s*/?\s*S\s*내용\s*[:：]\s*(.*?)(?=\s*[●∙•]\s*[가-힣A-Za-z]|$)", re.S | re.I)


def work_part(text):
    """공지 본문에서 **'A/S 내용'** 만 뽑는다.

    ★ 본문을 통째로 넣으면 신청일자·담당자·캠프주소까지 실적 칸에 들어간다.
      그러면 '무엇을 했나' 가 아니라 공지 전문이 실린다 — 읽는 사람이 더 헷갈린다.
      그 절이 없으면 **아무것도 돌려주지 않는다**(지어내지 않는다)."""
    m = RE_AS.search(_s(text))
    if not m:
        return ""
    t = _s(m.group(1))
    t = re.sub(r"★ ?작업(시작전|완료후).*?필수!?", " ", t)
    t = re.sub(r"https?://\S+", " ", t)
    # '※ 확인사항' 뒤는 서류 발행·보고 지시(사내 메모)다 — 작업 실적이 아니다.
    t = re.split(r"※\s*확인사항", t)[0]
    t = re.sub(r"▒+\s*", " ", t)          # ▒▒ 01 호기 ▒▒ → 01 호기
    # ★ 여러 호기 중 일부만 적힌 글은 뒤에 빈 양식이 그대로 남는다. 거기서 자른다.
    t = BLANK_UNIT.sub(" ", t)
    t = re.sub(r"\(유료\)\s*What\s*\?[^()]*", " ", t)
    t = _s(t)
    # 자르고 나서 실제 내용이 없으면 채우지 않는다 — 빈 양식만 있던 글이다.
    return "" if (not t or DRAFT.search(t) or len(t) < 4) else t[:MAXLEN]


def tidy(text):
    """카톡 추출기가 이미 'A/S 내용' 만 담아 준 값 — 다듬기만 한다.
    빈 양식(What ? / ?? 호기)만 있으면 채우지 않는다."""
    t = _s(text)
    t = re.sub(r"★ ?작업(시작전|완료후).*?필수!?", " ", t)
    t = re.sub(r"https?://\S+", " ", t)
    t = re.split(r"※\s*확인사항", t)[0]
    t = re.sub(r"▒+\s*", " ", t)
    t = BLANK_UNIT.sub(" ", t)
    t = _s(t)
    return "" if (not t or DRAFT.search(t) or len(t) < 4) else t[:MAXLEN]


def missing_rows(wb, kt, bt):
    """02시트에서 **작업완료인데 03시트에 아직 안 올라온** 건 — 세기만 한다.

    ★★ 2026-07-29 실사고: 여기서 **행을 만들면 안 된다.**
      03시트 B열(접수ID)은 값이 아니라 **수식**이고, 그 수식이
        "02시트에서 작업완료 + 완료일 있음 + 아직 03에 없는 접수ID" 를
      **스스로 하나씩 끌어온다.** 즉 **행을 늘리기만 하면 엑셀이 알아서 채운다.**
      그런데 내가 빈 행에 실적 내용을 내 순서대로 써 넣었더니, 엑셀이 재계산하는 순간
      그 행의 접수ID가 **다른 건**으로 정해져 **엉뚱한 작업에 남의 실적이 붙을** 상태가 됐다.
      (v259 를 만들자마자 되돌렸다)
    → 올바른 순서: ① 행 확장(expand_rows) ② **엑셀을 한 번 열어 재계산**(B열이 채워진다)
      ③ 그 다음에 이 도구가 **프로젝트NO 기준**으로 실제작업상세를 채운다."""
    ws3 = wb["03_현장작업실적"]
    h3 = [_s(x) for x in next(ws3.iter_rows(min_row=4, max_row=4, values_only=True))]
    j3 = h3.index("프로젝트NO")
    have, last = set(), 4
    for i, r in enumerate(ws3.iter_rows(min_row=5, values_only=True), start=5):
        if any(v not in (None, "") for v in r[1:]):
            last = i
            if j3 < len(r) and r[j3]:
                have.add(_s(r[j3]))
    ws2 = wb["02_돌발AS접수"]
    h2 = [_s(x) for x in next(ws2.iter_rows(min_row=4, max_row=4, values_only=True))]
    g = {c: h2.index(c) for c in ("프로젝트NO", "진행상태", "작업완료일", "캠프명",
                                  "담당기사", "신청내용", "접수ID") if c in h2}
    out = []
    for r in ws2.iter_rows(min_row=5, values_only=True):
        k = _s(r[g["프로젝트NO"]]) if g.get("프로젝트NO") is not None else ""
        if not k or k in have:
            continue
        if _s(r[g["진행상태"]]) != "작업완료":
            continue
        txt = tidy(kt.get(k, "")) or work_part(bt.get(k, ""))
        if not txt:
            continue                       # 무엇을 했는지 모르면 행을 만들지 않는다
        d = r[g["작업완료일"]] if g.get("작업완료일") is not None else None
        out.append({"프로젝트NO": k, "캠프명": _s(r[g.get("캠프명", 0)]),
                    "담당기사": _s(r[g.get("담당기사", 0)]), "접수ID": _s(r[g.get("접수ID", 0)]),
                    "최초접수내용": _s(r[g.get("신청내용", 0)])[:MAXLEN],
                    "작업일자": d.date().isoformat() if hasattr(d, "date") else _s(d)[:10],
                    "실제작업상세": txt,
                    "_src": "카톡 완료글 A/S내용" if kt.get(k) else "밴드 완료글 A/S내용"})
        have.add(k)
    return out, last + 1, h3


def plan(master):
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from openpyxl.utils import get_column_letter as GL
    kt, bt = kakao_text(), band_text()
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    ws = wb[SHEET]
    hdr = [_s(h) for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    if COL not in hdr or "프로젝트NO" not in hdr:
        wb.close()
        return [], {}
    jc, jp = hdr.index(COL), hdr.index("프로젝트NO")
    L = GL(jc + 1)
    items, tally = [], {"채움": 0, "이미 있음": 0, "근거 없음": 0}
    misses = []
    for n, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        k = _s(row[jp]) if jp < len(row) else ""
        if not k:
            continue
        cur = _s(row[jc]) if jc < len(row) else ""
        if cur and not DRAFT.search(cur):
            tally["이미 있음"] += 1
            continue                       # 사람이 적은 값은 건드리지 않는다
        src, txt = "", ""
        if kt.get(k):
            src, txt = "카톡 완료글 A/S내용", tidy(kt[k])
        if not txt and bt.get(k):
            txt = work_part(bt[k])          # 'A/S 내용' 절이 없으면 빈 문자열 → 비워 둔다
            src = "밴드 완료글 A/S내용" if txt else ""
        if not txt:
            tally["근거 없음"] += 1
            misses.append(k)
            continue                       # 원문에 없으면 비워 둔다 — 지어내지 않는다
        tally["채움"] += 1
        items.append({"sheet": SHEET, "key_col": "-", "key": "%s%d" % (L, n),
                      "cell": "%s%d" % (L, n), "col": COL, "value": txt, "vtype": "text",
                      "evidence": "%s (%s)" % (src, k),
                      # 자리표시자를 덮어야 하므로 '빈 칸만' 이 아니다.
                      # 위에서 **사람이 적은 값은 이미 걸렀다.**
                      "only_if_empty": False})
    # 03시트에 아직 안 올라온 완료건 — **세기만 한다.** 행은 엑셀 B열 수식이 만든다.
    news, start, _h3 = missing_rows(wb, kt, bt)
    room = ws.max_row - start + 1
    wb.close()
    return items, {"집계": tally, "근거없음샘플": misses[:8],
                   "대기": len(news), "여유행": room}


def main():
    do = "--apply" in sys.argv
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    items, info = plan(master)
    t = info.get("집계", {})
    print("%s · %s — 채움 %d · 이미 있음 %d · 근거 없음 %d"
          % (SHEET, COL, t.get("채움", 0), t.get("이미 있음", 0), t.get("근거 없음", 0)))
    for it in items[:6]:
        print("   %s  %s" % (it["evidence"], it["value"][:56]))
    if info.get("대기"):
        print("   ★ 03시트에 아직 안 올라온 완료건 %d건 (원문에 작업내용 있음) · 빈 행 여유 %d"
              % (info["대기"], info.get("여유행", 0)))
        print("     → 03시트 B열 수식 재계산은 자동입니다(excel_recalc·보관본 회차).")
        print("        스스로 끌어옵니다(행만 있으면 됩니다). 그 뒤 이 도구를 다시 돌리면")
        print("        프로젝트NO 기준으로 실제작업상세가 채워집니다.")
        if info.get("대기", 0) > info.get("여유행", 0):
            print("     → 빈 행이 모자랍니다: python expand_rows.py --sheet %s --add %d --apply"
                  % (SHEET, info["대기"] - info.get("여유행", 0) + 20))
    if info.get("근거없음샘플"):
        print("   [비워 둠] 원문에 없음:", ", ".join(info["근거없음샘플"]))
    if not do:
        print("\n실제로 채우려면:  python fill_work_detail.py --apply")
        return 0
    if not items:
        print("채울 것 없음")
        return 0
    from ledger_writer import queue_add
    print("큐 적재:", queue_add(items), "개 셀 → ledger_db --intake 후 11:00·15:00 원장 반영")
    return 0


if __name__ == "__main__":
    sys.exit(main())
