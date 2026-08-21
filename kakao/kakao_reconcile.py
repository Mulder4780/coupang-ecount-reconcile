# -*- coding: utf-8 -*-
"""
kakao_reconcile.py — 카카오톡 대화 내보내기(.txt) ↔ 관리대장 자동 대조
========================================================================
카카오는 채팅방 읽기 API를 제공하지 않으므로(보내기 API만 존재),
공식 기능인 [대화 내보내기]로 저장한 .txt를 kakao/inbox/ 에 넣으면 자동 처리한다.
(PC 카톡: 채팅방 → ☰ → 대화 내용 → 내보내기 / 모바일: 채팅방 설정 → 대화 내보내기)

지원 형식(자동 감지):
  PC:    --------------- 2026년 7월 20일 월요일 ---------------
         [유현민] [오후 2:59] 메시지 내용
  모바일: 2026년 7월 20일 오후 2:59, 유현민 : 메시지 내용

대조: 관리대장 02(돌발AS)·04(정기점검) 완료 건 ↔ 카톡 메시지
  1순위 프로젝트NO 포함 / 2순위 캠프명 핵심부 + 날짜 ±3일 / 보조 발신자=담당기사
원장은 read-only. 결과는 reports/ 에만 출력.

실행:  python kakao/kakao_reconcile.py            # kakao/inbox/*.txt 전체
       python kakao/kakao_reconcile.py --file 경로 [--master 경로]   # 테스트용
"""
import sys, os, re, csv, json, glob, hashlib
from datetime import datetime, date

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE_DIR)
CONFIG_PATH = os.path.join(ROOT, "config", "ecount_config.json")
INBOX_DIR = os.path.join(BASE_DIR, "inbox")
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")
HDR_ROW, FIRST = 4, 5

cfg = json.load(open(CONFIG_PATH, encoding="utf-8"))
DATE_TOL = int(cfg.get("kakao", {}).get("date_tolerance_days", 3))

RE_PC_DATE = re.compile(r"-+\s*(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일.*-+")
RE_PC_MSG = re.compile(r"^\[([^\]]+)\]\s*\[(오전|오후)\s*(\d{1,2}):(\d{2})\]\s*(.*)$")
RE_MO_MSG = re.compile(r"^(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*(오전|오후)\s*(\d{1,2}):(\d{2}),\s*([^:]+?)\s*:\s*(.*)$")
RE_MO_DATE = re.compile(r"^(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*\S요일\s*$")


def source_paths():
    """카톡 원본 정본+로컬 inbox. 경로 중복과 내용이 같은 사본을 제거한다."""
    # 신규등록·대표보고와 같은 대화방 최신본 선택기를 쓴다. 카톡 내보내기는
    # 누적 전체대화라 최신 두 방이면 이력이 모두 있고, Z: 73개를 다시 열 필요가 없다.
    try:
        import band_extract
        latest = band_extract.kakao_source_paths()
        if len(latest) == 2:
            return latest
    except Exception:
        pass
    sys.path.insert(0, ROOT)
    try:
        from source_dirs import kakao_dirs
        folders = kakao_dirs()
    except Exception:
        folders = [INBOX_DIR] if os.path.isdir(INBOX_DIR) else []
    paths, seen_path, seen_hash = [], set(), set()
    for folder in folders:
        for path in sorted(glob.glob(os.path.join(folder, "**", "*.txt"), recursive=True)):
            key = os.path.normcase(os.path.abspath(path))
            if key in seen_path:
                continue
            seen_path.add(key)
            try:
                with open(path, "rb") as fh:
                    digest = hashlib.sha256(fh.read()).digest()
            except OSError:
                continue
            if digest in seen_hash:
                continue
            seen_hash.add(digest)
            paths.append(path)
    return paths


def parse_export(path):
    """카톡 내보내기 txt → [{date, time, sender, text}] (여러 줄 메시지는 직전 메시지에 병합)"""
    msgs, cur_date = [], None
    # 내보내기 인코딩은 보통 UTF-8, 구버전은 cp949
    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            lines = open(path, encoding=enc).read().splitlines()
            break
        except UnicodeDecodeError:
            continue
    else:
        raise UnicodeDecodeError("kakao", b"", 0, 0, "지원 인코딩 아님")
    for ln in lines:
        m = RE_PC_DATE.match(ln) or RE_MO_DATE.match(ln)
        if m:
            g = m.groups()
            cur_date = date(int(g[0]), int(g[1]), int(g[2]))
            continue
        m = RE_PC_MSG.match(ln)
        if m and cur_date:
            sender, ap, hh, mm, text = m.groups()
            h = int(hh) % 12 + (12 if ap == "오후" else 0)
            msgs.append({"date": cur_date, "time": f"{h:02d}:{mm}", "sender": sender.strip(), "text": text})
            continue
        m = RE_MO_MSG.match(ln)
        if m:
            y, mo, d, ap, hh, mm, sender, text = m.groups()
            cur_date = date(int(y), int(mo), int(d))
            h = int(hh) % 12 + (12 if ap == "오후" else 0)
            msgs.append({"date": cur_date, "time": f"{h:02d}:{mm}", "sender": sender.strip(), "text": text})
            continue
        if msgs and ln.strip() and not ln.startswith("저장한 날짜"):
            msgs[-1]["text"] += " " + ln.strip()      # 여러 줄 메시지 병합
    return msgs


def camp_core(camp):
    return re.split(r"[(\s]", str(camp or ""))[0].strip()


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def manual_kakao_done(v):
    """원장 조치내용에 사용자가 명시한 카톡 확인 종료 기록이 있는지 본다.

    원문을 찾지 못했더라도 사용자가 개별 건을 완료 처리하라고 확정한 경우,
    그 지시를 02/04 시트의 빈 ``조치내용`` 칸에 남긴다. 이후 일일 대조가
    다시 실행돼도 같은 건을 '미확인'으로 되살리지 않기 위한 영구 근거다.
    """
    text = re.sub(r"\s+", "", str(v or ""))
    return ("카톡보고미확인완료처리" in text or
            "카톡보고확인완료" in text)


def read_rows(master):
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    spec = {
        "02_돌발AS접수": ("접수ID", "프로젝트NO", "캠프명", "담당기사", "작업완료일", "진행상태"),
        "04_정기점검":   ("점검ID", "프로젝트NO", "캠프명", "담당기사", "실제점검일", "점검상태"),
    }
    out = []
    for sheet, cols in spec.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        def gv(row, name):
            return row[idx[name]] if name in idx and idx[name] < len(row) else None
        for row in ws.iter_rows(min_row=FIRST, values_only=True):
            # ID 열은 수식이라 새로 추가된 행은 엑셀을 열기 전까지 캐시값이 없다(None).
            # 그런 행도 대조 대상이어야 하므로 프로젝트NO를 대체 키로 쓴다.
            rid = gv(row, cols[0]) or gv(row, cols[1])
            done = to_date(gv(row, cols[4]))
            if not rid or not done:
                continue
            out.append({"시트": sheet, "ID": str(rid), "프로젝트NO": str(gv(row, cols[1]) or ""),
                        "캠프명": str(gv(row, cols[2]) or ""), "담당기사": str(gv(row, cols[3]) or ""),
                        "완료일": done,
                        "수동완료": manual_kakao_done(gv(row, "조치내용"))})
    wb.close()
    return out


def main():
    args = sys.argv[1:]
    # config에는 처음 만든 파일명(v20)이 그대로 적혀 있다. 실제로 봐야 할 건 **최신본**이다.
    # 다른 도구들은 전부 resolve_master를 쓰는데 여기만 빠져 있어 파일 없음으로 죽었다.
    sys.path.insert(0, ROOT)
    from ecount_reconcile import resolve_master
    # 합성검증·명시 실행은 --master를 진실의 원천으로 쓴다. 기존처럼 설정파일을
    # 먼저 해석하면 네트워크 원본 폴더가 잠시 끊긴 상황에서도 로컬 대조가 시작조차 못 한다.
    master = (args[args.index("--master") + 1] if "--master" in args
              else resolve_master(cfg["reconcile"]["master_xlsx"]))
    if "--file" in args:
        files = [args[args.index("--file") + 1]]
    else:
        files = source_paths()
    if not files:
        sys.exit("kakao/inbox/ 에 카톡 내보내기 .txt가 없습니다. (PC 카톡: 채팅방 → ☰ → 대화 내용 → 내보내기)")

    msgs = []
    for f in files:
        part = parse_export(f)
        room = os.path.splitext(os.path.basename(f))[0]
        for m in part:
            m["room"] = room
        msgs += part
        print(f"'{os.path.basename(f)}': 메시지 {len(part)}건 파싱")

    rows = read_rows(master)
    print(f"카톡 메시지 {len(msgs)}건 / 원장 완료건 {len(rows)}건 대조")

    # ★ 자료가 없는 기간을 '미확인'이라고 부르지 않는다 (2026-08-07 지시).
    #   사용자 지시: "단톡방에 내가 들어간게 7월인 것 같은데, 그 이전 자료 기록은
    #   찾기 힘들어. 4원천 검증에서 그 이전 카톡은 확인이 안되어도 다른 곳에서
    #   확인되면 완료처리할 수 있게" —
    #   카톡 내보내기는 **방에 들어간 뒤**부터만 나온다. 그 전 작업은 아무리 잘해도
    #   카톡에서 못 찾는다. 그걸 '미확인'으로 세면 영원히 안 지워지는 빨간 줄이
    #   쌓이고, 진짜 누락(자료는 있는데 보고가 없는 건)이 그 속에 묻힌다.
    #   ★ 7월로 못박지 않는다 — **가진 메시지 중 가장 이른 날**을 자료의 시작으로
    #     삼는다. 나중에 옛 내보내기가 들어오면 경계가 저절로 내려간다.
    #     못박아 두면 자료가 생겨도 계속 '자료없음'이라 답한다.
    floor = min((m["date"] for m in msgs), default=None)
    if floor:
        print(f"카톡 자료 시작: {floor.isoformat()} — 이보다 이른 완료건은 "
              f"'미확인'이 아니라 '자료없음'으로 가른다")

    matched_keys = set()
    results = []
    for r in rows:
        best, how = None, ""
        manual_done = bool(r.get("수동완료"))
        core = camp_core(r["캠프명"])
        if manual_done:
            how = "사용자완료처리"
        else:
            for m in msgs:
                near = abs((m["date"] - r["완료일"]).days) <= DATE_TOL
                if r["프로젝트NO"] and r["프로젝트NO"] in m["text"]:
                    best, how = m, "프로젝트NO"
                    break
                if near and core and len(core) >= 3 and core in m["text"] and best is None:
                    best, how = m, "캠프명+날짜"
        sender_ok = bool(best and r["담당기사"] and r["담당기사"] in best["sender"])
        if best:
            matched_keys.add(id(best))
        # 자료가 시작되기 **전**에 끝난 일은 카톡에서 찾을 방법이 없다.
        # '미확인'(찾아봤는데 없다)과 '자료없음'(찾아볼 자료가 없다)은 다른 말이다.
        before_data = bool(floor and not best and not manual_done
                           and r["완료일"] < floor)
        if before_data:
            verdict = "자료없음"
            how = f"카톡 자료 시작({floor.isoformat()}) 이전 — 다른 원천으로 판정"
        else:
            verdict = "확인" if (best or manual_done) else "미확인"
        results.append({**{k: r[k] for k in ("시트", "ID", "프로젝트NO", "캠프명", "담당기사")},
                        "완료일": r["완료일"].isoformat(),
                        "카톡보고": verdict,
                        "매칭근거": how + ("+기사일치" if sender_ok else ""),
                        "메시지일": best["date"].isoformat() if best else "",
                        "발신자": best["sender"] if best else "",
                        "방": best["room"] if best else ""})

    miss = [r for r in results if r["카톡보고"] == "미확인"]
    nodata = [r for r in results if r["카톡보고"] == "자료없음"]
    os.makedirs(REPORT_DIR, exist_ok=True)
    base = os.path.join(REPORT_DIR, f"카톡대조_{datetime.now():%Y%m%d_%H%M}")
    cols = list(results[0].keys()) if results else []
    with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(results)
    with open(base + ".md", "w", encoding="utf-8") as f:
        f.write("# 카카오톡 ↔ 관리대장 대조 리포트\n\n")
        ok_n = sum(1 for r in results if r["카톡보고"] == "확인")
        f.write(f"- 생성 {datetime.now():%Y-%m-%d %H:%M} / 완료건 {len(results)} / "
                f"카톡보고 확인 {ok_n} / 미확인 {len(miss)} / 자료없음 {len(nodata)}\n")
        if floor:
            f.write(f"- 카톡 자료 시작 **{floor.isoformat()}** — 이보다 이른 완료건은 "
                    f"카톡에서 찾을 방법이 없다. '미확인'이 아니라 **자료없음**이며, "
                    f"밴드·ERP·쿠팡PO 중 하나로 확인되면 완료로 본다.\n")
        f.write("\n## 카톡 보고 미확인 건 (자료는 있는데 보고가 없다 — 진짜 확인 대상)\n\n"
                "| ID | 캠프 | 기사 | 완료일 |\n|---|---|---|---|\n")
        for r in miss:
            f.write(f"| {r['ID']} | {r['캠프명']} | {r['담당기사']} | {r['완료일']} |\n")
        if nodata:
            f.write(f"\n## 자료없음 (카톡 시작 이전) — {len(nodata)}건\n\n"
                    "카톡으로는 확인할 수 없다. 다른 원천으로 판정한다.\n\n"
                    "| ID | 캠프 | 완료일 |\n|---|---|---|\n")
            for r in nodata[:60]:
                f.write(f"| {r['ID']} | {r['캠프명']} | {r['완료일']} |\n")
    print(f"확인 {sum(1 for r in results if r['카톡보고']=='확인')} / "
          f"미확인 {len(miss)} / 자료없음 {len(nodata)}")
    print("리포트:", base + ".md")
    return results


if __name__ == "__main__":
    main()
