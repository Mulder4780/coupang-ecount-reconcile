# -*- coding: utf-8 -*-
"""
dedupe_rows.py — 관리대장 중복 행 확인·정리 (근거 기반, 값 비우기)
=====================================================================
같은 프로젝트NO가 여러 행에 입력된 '중복 접수'를 찾아 정리한다.

판정 근거 (추측 금지 — 실제 원천으로 확인):
  · 밴드 게시글 수 = 실제 작업 건수. 원장 행 수가 그보다 많으면 초과분이 중복.
  · 남길 행: 밴드 게시일/작업일과 접수일이 일치하는 행 > 가장 이른 행.
  · 핵심 필드(접수일자·담당기사·신청내용·작업완료일)가 모두 같은 행만 중복으로 인정.

안전 원칙:
  1. 행을 삭제하지 않고 **값만 비운다** — 행 구조·수식(ID·검증·집계)이 그대로 살아 빈 행으로 처리됨
  2. 수식 셀은 절대 건드리지 않음 (`<f>` 포함 셀 skip)
  3. 항상 vN+1 새 파일 생성 — 원본 보존, 언제든 롤백
  4. 밴드 근거가 없으면 정리하지 않고 '확인필요'로만 보고
  5. --apply 없이는 리포트만

실행:
  python dedupe_rows.py            # 조사·리포트만
  python dedupe_rows.py --apply    # 중복 행 값 비우기(vN+1)
"""
import sys, os, re, csv, json, glob, zipfile
from datetime import datetime, date
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")
CACHE_DIR = os.path.join(BASE_DIR, "band", "cache")

# same = '동일 작업'을 판정하는 필드. 접수일자/예정일은 제외한다 —
# 같은 작업을 다른 날 중복 접수한 사례가 실제로 있어(UJ2600983: 6/05·7/13) 접수일 차이는 중복을 부정하지 않는다.
SHEETS = {
    "02_돌발AS접수": {"key": "프로젝트NO", "same": ["담당기사", "신청내용", "작업완료일"],
                    "date": "접수일자", "id": "접수ID"},
    "04_정기점검": {"key": "프로젝트NO", "same": ["담당기사", "실제점검일"],
                  "date": "점검예정일", "id": "점검ID"},
}


def band_counts():
    """프로젝트NO → 밴드 게시 건수·게시일 목록 (실제 작업 건수의 근거)"""
    out = defaultdict(list)
    for f in glob.glob(os.path.join(CACHE_DIR, "*.json")):
        if os.path.basename(f).startswith(("raw_", "dump_")):
            continue
        d = json.load(open(f, encoding="utf-8"))
        for p in d.get("posts", {}).values():
            c = p.get("content") or ""
            ts = p.get("created_at")
            day = datetime.fromtimestamp(ts / 1000).date() if ts else None
            for prj in set(re.findall(r"\bUJ\d{6,}\b", c)):
                out[prj].append(day)
    return out


def norm(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip() if v is not None else ""


def scan(master):
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    bands = band_counts()
    groups, review = [], []
    for sheet, spec in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [norm(h) for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        idx = {h: i for i, h in enumerate(hdr) if h}
        by = defaultdict(list)
        for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
            g = lambda k: norm(row[idx[k]]) if k in idx and idx[k] < len(row) else ""
            prj = g(spec["key"])
            if prj:
                by[prj].append({"row": 5 + i, "prj": prj, "id": g(spec["id"]),
                                "date": g(spec["date"]),
                                "sig": tuple(g(k) for k in spec["same"]),
                                "vals": {h: g(h) for h in idx}})
        for prj, rows in by.items():
            if len(rows) < 2:
                continue
            posts = bands.get(prj, [])
            real = len(posts)
            same_sig = len({r["sig"] for r in rows}) == 1
            item = {"sheet": sheet, "prj": prj, "rows": rows, "band": real,
                    "post_days": sorted({str(d) for d in posts if d}), "same_sig": same_sig}
            # 근거 있음: 밴드 1건 확인 + 핵심필드 동일 → 초과분 중복
            if real == 1 and same_sig and len(rows) > 1:
                bd = item["post_days"][0] if item["post_days"] else ""
                keep = next((r for r in rows if r["date"] and r["date"] == bd), None) or \
                       sorted(rows, key=lambda r: (r["date"] or "9999", r["row"]))[0]
                item["keep"] = keep
                item["drop"] = [r for r in rows if r["row"] != keep["row"]]
                groups.append(item)
            else:
                item["사유"] = ("밴드 게시 %d건 — 실제 다건 작업일 수 있음" % real if real != 1
                              else "핵심 필드가 서로 달라 별개 작업일 수 있음")
                review.append(item)
    wb.close()
    return groups, review


def blank_cells(master, targets):
    """targets: [(sheet, row, [col_letters])] — 값만 제거(수식 셀 skip), vN+1 생성"""
    from ledger_writer import sheet_file_map
    mv = re.search(r"_v(\d+)\.xlsx$", master)
    v = int(mv.group(1))
    dst = master.replace(f"_v{v}.xlsx", f"_v{v+1}.xlsx")
    if os.path.exists(dst):
        sys.exit(f"{os.path.basename(dst)} 이미 존재 — 중단")

    zin = zipfile.ZipFile(master)
    smap = sheet_file_map(zin)
    by_sheet = defaultdict(list)
    for sh, rn, cols in targets:
        by_sheet[smap[sh]].append((rn, cols))

    edits, cleared, kept_formula = {}, 0, 0
    for fname, items in by_sheet.items():
        xml = zin.read(fname).decode("utf-8")
        for rn, cols in items:
            mrow = re.search(r'(<row r="%d"[^>]*>)(.*?)(</row>)' % rn, xml, re.S)
            if not mrow:
                continue
            head, body, tail = mrow.groups()
            new_body = body
            for cl in cols:
                mopen = re.search(r'<c r="%s%d"[^>]*>' % (cl, rn), new_body)
                if not mopen:
                    continue
                if mopen.group(0).endswith("/>"):
                    continue                                  # 이미 빈 셀
                end = new_body.find("</c>", mopen.end())
                cxml = new_body[mopen.start():end + 4]
                if "<f" in cxml:                              # 수식 셀은 보존
                    kept_formula += 1
                    continue
                style = re.search(r' s="(\d+)"', mopen.group(0))
                empty = f'<c r="{cl}{rn}"{" s=" + chr(34) + style.group(1) + chr(34) if style else ""}/>'
                new_body = new_body[:mopen.start()] + empty + new_body[end + 4:]
                cleared += 1
            xml = xml[:mrow.start()] + head + new_body + tail + xml[mrow.end():]
        edits[fname] = xml.encode("utf-8")

    zout = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, edits.get(it.filename) or zin.read(it.filename))
    zout.close(); zin.close()

    z = zipfile.ZipFile(dst)
    assert z.testzip() is None, "zip 무결성 실패"
    zsrc = zipfile.ZipFile(master)
    diff = [n for n in zsrc.namelist() if n not in edits and zsrc.read(n) != z.read(n)]
    assert not diff, f"의도치 않은 파트 변경: {diff}"
    zsrc.close(); z.close()
    return dst, cleared, kept_formula, f"v{v} → v{v+1}"


def main():
    apply_mode = "--apply" in sys.argv
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    groups, review = scan(master)

    os.makedirs(REPORT_DIR, exist_ok=True)
    base = os.path.join(REPORT_DIR, f"중복정리_{datetime.now():%Y%m%d_%H%M}")
    lines = ["# 중복 행 확인·정리 리포트\n", f"- 생성 {datetime.now():%Y-%m-%d %H:%M} / 대상 {os.path.basename(master)}\n",
             f"- 근거 있는 중복 {len(groups)}그룹 / 확인필요 {len(review)}그룹\n\n## 정리 대상 (밴드 1건 확인 + 핵심필드 동일)\n"]
    rows_csv = []
    for g in groups:
        lines.append(f"\n### {g['sheet']} · {g['prj']} — 원장 {len(g['rows'])}행 / 밴드 {g['band']}건 "
                     f"(게시 {', '.join(g['post_days'])})\n")
        lines.append(f"- **유지**: {g['keep']['id'] or '(ID미계산)'} (행 {g['keep']['row']}, 접수 {g['keep']['date']})\n")
        for d in g["drop"]:
            lines.append(f"- 정리: {d['id'] or '(ID미계산)'} (행 {d['row']}, 접수 {d['date']})\n")
            rows_csv.append({"시트": g["sheet"], "프로젝트NO": g["prj"], "정리행": d["row"],
                             "정리ID": d["id"], "유지행": g["keep"]["row"], "유지ID": g["keep"]["id"],
                             "밴드게시": g["band"]})
    lines.append("\n## 확인필요 (자동 정리하지 않음)\n")
    for r in review:
        lines.append(f"- {r['sheet']} {r['prj']}: 원장 {len(r['rows'])}행 · 밴드 {r['band']}건 — {r['사유']}\n")
    open(base + ".md", "w", encoding="utf-8").writelines(lines)
    if rows_csv:
        with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows_csv[0].keys())); w.writeheader(); w.writerows(rows_csv)

    n_drop = sum(len(g["drop"]) for g in groups)
    print(f"근거 있는 중복 {len(groups)}그룹 / 정리 대상 {n_drop}행 / 확인필요 {len(review)}그룹")
    for g in groups:
        print(f"  {g['prj']}: {len(g['rows'])}행 → 1행 유지({g['keep']['row']}행), "
              f"{[d['row'] for d in g['drop']]}행 정리 (밴드 {g['band']}건)")
    print("리포트:", base + ".md")

    if not apply_mode:
        print("\n조사 모드 — 실제 정리: python dedupe_rows.py --apply")
        return
    if not n_drop:
        print("정리할 중복 없음"); return

    import openpyxl
    from ledger_writer import col_letter
    wb = openpyxl.load_workbook(master, read_only=True)
    targets = []
    for g in groups:
        ws = wb[g["sheet"]]
        hdr = [norm(h) for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        cols = [col_letter(i + 1) for i, h in enumerate(hdr) if h]
        for d in g["drop"]:
            targets.append((g["sheet"], d["row"], cols))
    wb.close()
    dst, cleared, kept, msg = blank_cells(master, targets)
    print(f"\n정리 완료: {msg} — {n_drop}행 / 셀 {cleared}개 비움 (수식 셀 {kept}개 보존)")
    print("   ", dst)


if __name__ == "__main__":
    main()
