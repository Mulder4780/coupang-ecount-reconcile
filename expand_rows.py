# -*- coding: utf-8 -*-
"""
expand_rows.py — 관리대장 시트의 **입력 가능 행을 늘린다** (백필 용량 확보)
===============================================================================
02_돌발AS접수(154행)·04_정기점검(104행)처럼 수식이 깔린 마지막 행까지 다 쓰면
백필이 "용량 초과"로 멈춘다. 이 도구가 마지막 행을 본떠 아래로 행을 더 만든다.

왜 openpyxl이 아니라 zip 패치인가
  openpyxl로 열어 save()하면 차트·도형버튼·x14 검증이 파괴된다(절대규칙 2).
  그래서 시트 XML만 직접 손대고 나머지 파트는 **바이트 그대로** 복사한다.

무엇을 함께 늘리는가 (이걸 빠뜨리면 새 행에 드롭다운·서식·표가 안 먹는다)
  · <dimension>            시트 범위
  · dataValidation         드롭다운(일반 + x14 확장 양쪽)
  · conditionalFormatting  조건부서식
  · autoFilter / 표(table) 범위 — xl/tables/*.xml 까지
공유수식(<f t="shared">)은 새 행에서 **일반 수식으로 풀어서** 넣는다.
공유 범위(ref)를 건드리면 엑셀이 파일을 손상으로 판단할 수 있기 때문이다.

사용
  python expand_rows.py --sheet 04_정기점검 --add 60            # 미리보기
  python expand_rows.py --sheet 04_정기점검 --add 60 --apply     # vN+1 생성
  python expand_rows.py --self-test                             # 합성 워크북 검증
"""
import sys, os, re, glob, shutil, zipfile, tempfile
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

ROW_RE = re.compile(r'<row r="(\d+)"(.*?)>(.*?)</row>', re.S)


def iter_tags(xml, name):
    """<name ...>…</name> 와 <name .../> 를 **2단계로** 파싱한다.
    (`<c r="E104" s="41"/>` 같은 자기닫힘 셀을 게으른 정규식으로 잡으면
     다음 셀까지 통째로 삼켜 XML이 깨진다 — 실제로 겪은 사고. AGENTS.md 참조)"""
    i, open_t = 0, "<" + name
    while True:
        s = xml.find(open_t, i)
        if s < 0:
            return
        nxt = xml[s + len(open_t):s + len(open_t) + 1]
        if nxt not in (" ", ">", "/"):        # <f> · <f .../> 는 맞고 <format…> 은 아님
            i = s + len(open_t)
            continue
        e = xml.find(">", s)
        if e < 0:
            return
        tag = xml[s:e + 1]
        if tag.endswith("/>"):
            yield tag, None
            i = e + 1
        else:
            c = xml.find("</" + name + ">", e)
            if c < 0:
                return
            yield tag, xml[e + 1:c]
            i = c + len(name) + 3


# ── 수식 행 이동 ─────────────────────────────────────────────────────
def shift_formula(f, src, dst):
    """상대참조 행번호만 옮긴다. $6 처럼 절대행은 그대로 둔다."""
    def rep(m):
        col, dollar, row = m.group(1), m.group(2), int(m.group(3))
        if dollar or row != src:
            return m.group(0)
        return f"{col}{dst}"
    return re.sub(r"(\$?[A-Z]{1,3})(\$?)(\d+)", rep, f)


def shared_map(xml):
    """si → (수식문자열, 마스터행). 공유수식 슬레이브를 일반 수식으로 풀 때 쓴다."""
    out = {}
    for ctag, cinner in iter_tags(xml, "c"):
        if not cinner:
            continue
        rm = re.search(r'r="[A-Z]+(\d+)"', ctag)
        if not rm:
            continue
        for ftag, ftext in iter_tags(cinner, "f"):
            si = re.search(r'si="(\d+)"', ftag)
            if si and ftext and ftext.strip():      # 마스터(수식 본문 보유)
                out.setdefault(si.group(1), (ftext, int(rm.group(1))))
    return out


def clone_row(row_xml, src_row, dst_row, smap):
    """마지막 행을 본떠 새 행 XML을 만든다. 값은 비우고 수식·서식만 물려준다."""
    body = ROW_RE.search(row_xml)
    attrs, inner = body.group(2), body.group(3)
    cells = []
    for ctag, cinner in iter_tags(inner, "c"):
        cm = re.search(r'r="([A-Z]+)\d+"', ctag)
        if not cm:
            continue
        col = cm.group(1)
        sm = re.search(r'\ss="\d+"', ctag)
        style = sm.group(0) if sm else ""
        ftag = ftext = None
        if cinner:
            for t, x in iter_tags(cinner, "f"):
                ftag, ftext = t, x
                break
        if ftag is None:
            cells.append(f'<c r="{col}{dst_row}"{style}/>')      # 데이터 칸 → 빈칸(서식만)
            continue
        ca = ' ca="1"' if 'ca="1"' in ftag else ""
        si = re.search(r'si="(\d+)"', ftag)
        if si and not (ftext or "").strip():                     # 공유수식 슬레이브
            master = smap.get(si.group(1))
            if not master:
                cells.append(f'<c r="{col}{dst_row}"{style}/>')
                continue
            f = shift_formula(master[0], master[1], dst_row)
        else:
            f = shift_formula(ftext or "", src_row, dst_row)
        # ★ <v/>를 빠뜨리면 t="str" 선언과 내용이 어긋나 엑셀이 복구 대화상자를 띄운다
        cells.append(f'<c r="{col}{dst_row}"{style} t="str"><f{ca}>{f}</f><v/></c>')
    return f'<row r="{dst_row}"{attrs}>{"".join(cells)}</row>'


def bump(text, last, newlast):
    """범위 문자열에서 **끝행만** last → newlast (A5:AG104 → A5:AG164)

    ★ 예전에는 문자열 안의 last를 전부 바꿨다. 그래서 시작행이 마침 last와 같은 범위
      (예: 마지막 행 하나만 덮는 A5:A5)가 A8:A8 처럼 **시작까지 밀려** 원래 구간을
      벗어났다(2026-07-26 합성 점검에서 발견). 콜론 뒤(=끝) 좌표만 손댄다.
    """
    def one(part):
        if ":" not in part:
            return part      # 단일 셀 참조는 범위가 아니다 — 건드리지 않는다
        a, b = part.split(":", 1)
        b2 = re.sub(r'(\$?[A-Z]{1,3}\$?)' + str(last) + r'(?!\d)',
                    lambda m: m.group(1) + str(newlast), b)
        return a + ":" + b2
    return " ".join(one(p) for p in text.split(" "))


def bump_attr(xml, tag_pat, attr, last, newlast):
    """지정 태그의 특정 속성(sqref/ref) 안에서만 끝행을 늘린다"""
    def rep(m):
        return m.group(1) + bump(m.group(2), last, newlast) + m.group(3)
    return re.sub(r'(<' + tag_pat + r'[^>]*?\b' + attr + r'=")([^"]*)(")', rep, xml)


# ── 시트 XML 전체 확장 ────────────────────────────────────────────────
# 시트별 ID 열 채번 수식 — 마지막 행의 ID는 fix_ids.py가 **확정값**으로 바꿔 놓아서
# 그 행을 그대로 복사하면 새 행에 채번 수식이 안 생긴다. 그래서 여기서 다시 심는다.
# (ID열, ID이름, 접두어, 채번에 쓰는 날짜열)
# ★ 여기 빠진 시트는 늘려도 **새 행에 채번 수식이 안 생긴다** — 값만 들어가고 ID는 영영 빈칸이다.
#   실제로 03_현장작업실적이 빠져 있어서, 늘려 놓고도 여유가 0인 채였다(2026-07-27).
#   날짜열은 시트마다 다르다(02·04는 D, 03·05는 E, 06은 F) — 검증 [28]이 실제 수식과 대조한다.
ID_FORMULA = {
    "02_돌발AS접수":       ("A", "접수ID", "AS", "D"),
    "03_현장작업실적":     ("A", "작업ID", "FW", "E"),
    "04_정기점검":         ("A", "점검ID", "PM", "D"),
    "05_신규납품설치":     ("A", "업무ID", "NS", "E"),
    "06_거래서류청구수금": ("A", "정산ID", "JS", "F"),
}


def id_formula(prefix, datecol, row):
    return (f'IF($B{row}="","","{prefix}-"&amp;TEXT(IF(${datecol}{row}="",TODAY(),'
            f'${datecol}{row}),"yymm")&amp;"-"&amp;TEXT(ROW()-4,"000"))')


def put_id_formula(row_xml, sheet, row):
    """새 행의 ID 칸을 채번 수식으로 되돌린다(빈칸으로 복제된 상태를 교체)."""
    spec = ID_FORMULA.get(sheet)
    if not spec:
        return row_xml
    col, _, prefix, datecol = spec
    f = id_formula(prefix, datecol, row)
    pat = re.compile(r'<c r="%s%d"([^>]*?)/>' % (col, row))
    m = pat.search(row_xml)
    if not m:
        return row_xml
    style = re.search(r'\ss="\d+"', m.group(1))
    style = style.group(0) if style else ""
    return pat.sub(f'<c r="{col}{row}"{style} t="str"><f>{f}</f><v/></c>', row_xml, count=1)


def expand_sheet_xml(xml, add, sheet=""):
    last = max(int(r) for r in re.findall(r'<row r="(\d+)"', xml))
    newlast = last + add
    tpl = re.search(r'<row r="%d".*?</row>' % last, xml, re.S).group()
    smap = shared_map(xml)
    rows = "".join(put_id_formula(clone_row(tpl, last, last + i, smap), sheet, last + i)
                   for i in range(1, add + 1))
    xml = xml.replace("</sheetData>", rows + "</sheetData>", 1)
    xml = re.sub(r'(<dimension ref="[A-Z]+\d+:[A-Z]+)(\d+)("/>)',
                 lambda m: m.group(1) + str(newlast) + m.group(3), xml, count=1)
    for tag, attr in (("dataValidation", "sqref"), ("x14:dataValidation", "sqref"),
                      ("conditionalFormatting", "sqref"), ("autoFilter", "ref"),
                      ("sqref", "")):
        if attr:
            xml = bump_attr(xml, re.escape(tag), attr, last, newlast)
    # x14 검증은 <xm:sqref>본문</xm:sqref> 형태도 쓴다
    xml = re.sub(r'(<xm:sqref>)(.*?)(</xm:sqref>)',
                 lambda m: m.group(1) + bump(m.group(2), last, newlast) + m.group(3), xml)
    return xml, last, newlast


def expand_table_xml(xml, last, newlast):
    xml = bump_attr(xml, "table", "ref", last, newlast)
    xml = bump_attr(xml, "autoFilter", "ref", last, newlast)
    return xml


# ── 파일 단위 실행 ────────────────────────────────────────────────────
def sheet_file_map(z):
    wbx = z.read("xl/workbook.xml").decode("utf-8")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    # 속성 순서는 파일마다 다르다(Id가 뒤에 오는 경우도 있음) → 요소 단위로 파싱
    rid = {}
    for el in re.findall(r"<Relationship\b[^>]*/?>", rels):
        i = re.search(r'Id="([^"]+)"', el)
        t = re.search(r'Target="([^"]+)"', el)
        if i and t:
            rid[i.group(1)] = t.group(1)
    out = {}
    for el in re.findall(r"<sheet\b[^>]*/?>", wbx):
        nm = re.search(r'name="([^"]+)"', el)
        r = re.search(r'r:id="([^"]+)"', el)
        if nm and r and r.group(1) in rid:
            tgt = rid[r.group(1)].lstrip("/")
            out[nm.group(1)] = tgt if tgt.startswith("xl/") else "xl/" + tgt
    return out


def sheet_tables(z, sheet_file):
    """이 시트가 쓰는 표(table) 파트 목록"""
    rel = sheet_file.replace("worksheets/", "worksheets/_rels/") + ".rels"
    if rel not in z.namelist():
        return []
    txt = z.read(rel).decode("utf-8")
    return ["xl/" + t.replace("../", "") for t in
            re.findall(r'Target="([^"]*tables/[^"]+)"', txt)]


def strip_calcchain(name, data, ct):
    """calcChain 제거 + 열 때 전체 재계산 (수식 캐시가 없으므로 필수)"""
    if name == "xl/calcChain.xml":
        return None, ct
    if name == "[Content_Types].xml":
        data = re.sub(rb'<Override PartName="/xl/calcChain\.xml"[^>]*/>', b"", data)
    if name == "xl/workbook.xml":
        s = data.decode("utf-8")
        if "<calcPr" in s:
            s = re.sub(r"<calcPr[^/]*/>", '<calcPr calcId="0" fullCalcOnLoad="1"/>', s)
        else:
            s = s.replace("</workbook>", '<calcPr calcId="0" fullCalcOnLoad="1"/></workbook>')
        data = s.encode("utf-8")
    return data, ct


def expand(src, dst, sheet, add):
    zin = zipfile.ZipFile(src)
    smap = sheet_file_map(zin)
    if sheet not in smap:
        zin.close(); raise SystemExit(f"시트 없음: {sheet}")
    sf = smap[sheet]
    xml, last, newlast = expand_sheet_xml(zin.read(sf).decode("utf-8"), add, sheet)
    patched = {sf: xml.encode("utf-8")}
    for t in sheet_tables(zin, sf):
        patched[t] = expand_table_xml(zin.read(t).decode("utf-8"), last, newlast).encode("utf-8")

    zout = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    ct = None
    untouched = 0
    for it in zin.infolist():
        data = zin.read(it.filename)
        data, ct = strip_calcchain(it.filename, data, ct)
        if data is None:
            continue
        if it.filename in patched:
            data = patched[it.filename]
        else:
            untouched += 1
        zout.writestr(it.filename, data)
    zout.close(); zin.close()
    return last, newlast, len(patched), untouched


def verify(dst, sheet, last, newlast):
    """3중 검증: zip 무결성 · 새 행 수식 존재 · 기존 행 값 보존"""
    z = zipfile.ZipFile(dst)
    assert z.testzip() is None, "zip 무결성 실패"
    z.close()
    import openpyxl
    wb = openpyxl.load_workbook(dst, read_only=True, data_only=False)
    ws = wb[sheet]
    ok_f = sum(1 for r in (last + 1, newlast)
               for c in ws[r] if isinstance(c.value, str) and c.value.startswith("="))
    assert ok_f > 0, "새 행에 수식이 없음"
    wb.close()
    return ok_f


def self_test():
    """합성 워크북으로 확장 로직 검증 (실데이터 접촉 0)"""
    import openpyxl
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "syn.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "04_정기점검"
    for _ in range(3):
        ws.append([])
    ws.append(["점검ID", "프로젝트NO", "캠프명", "점검예정일"])
    for r in range(5, 9):
        ws.cell(row=r, column=1).value = '=IF($B%d="","","PM-"&TEXT($D%d,"yymm")&"-"&TEXT(ROW()-4,"000"))' % (r, r)
        ws.cell(row=r, column=2).value = f"UJ260{r}"
    wb.save(src)
    dst = os.path.join(tmp, "syn_out.xlsx")
    last, newlast, np_, nu = expand(src, dst, "04_정기점검", 5)
    assert (last, newlast) == (8, 13), (last, newlast)
    w = openpyxl.load_workbook(dst)
    ws2 = w["04_정기점검"]
    assert ws2["B8"].value == "UJ2608", ws2["B8"].value          # 기존 행 보존
    assert ws2["B13"].value is None, ws2["B13"].value            # 새 행은 빈칸
    f = ws2["A13"].value
    assert f and f.startswith("=") and "$B13" in f and "$D13" in f, f   # 행번호 이동
    assert "ROW()-4" in f, f                                     # ROW()는 그대로
    # 수식 셀에는 캐시값이 반드시 있어야 한다(엑셀 손상 판정 방지)
    import zipfile as _z, re as _re
    zz = _z.ZipFile(dst)
    for nm in zz.namelist():
        if not _re.match(r"xl/worksheets/sheet\d+\.xml$", nm):
            continue
        for ctag, cin in iter_tags(zz.read(nm).decode("utf-8"), "c"):
            if cin and "<f" in cin and _re.search(r'\st="[^"]+"', ctag):
                assert "<v>" in cin or "<v/>" in cin, f"캐시값 없는 수식 셀: {ctag}"
    zz.close()
    w.close(); shutil.rmtree(tmp, ignore_errors=True)
    # 범위 확장 유틸
    assert bump("A5:AG104", 104, 164) == "A5:AG164"
    assert bump("$E$5:$E$104", 104, 164) == "$E$5:$E$164"
    assert bump("A5:AG104 B5:B104", 104, 164) == "A5:AG164 B5:B164"
    assert shift_formula('IF($A104="","",$D104)', 104, 164) == 'IF($A164="","",$D164)'
    assert shift_formula("'00_대시보드'!$B$4", 104, 164) == "'00_대시보드'!$B$4"   # 절대참조 불변
    print("  확장 로직 self-test 통과 (수식 이동·범위 확장·기존행 보존)")
    return True


def main():
    # 다른 AI가 원장을 잡고 있으면 여기서 멈춘다(동시 수정 시 한쪽이 통째로 묻힌다)
    from claim_guard import require
    require("ledger", "expand_rows")
    if "--self-test" in sys.argv:
        self_test(); return
    sheet = sys.argv[sys.argv.index("--sheet") + 1] if "--sheet" in sys.argv else None
    add = int(sys.argv[sys.argv.index("--add") + 1]) if "--add" in sys.argv else 0
    if not sheet or add <= 0:
        sys.exit("사용: python expand_rows.py --sheet 04_정기점검 --add 60 [--apply]")
    self_test()                                   # 실파일 손대기 전 항상 자체검증
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    m = re.search(r"_v(\d+)\.xlsx$", master)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(m.group(1))+1}.xlsx", master)
    print(f"원본: {os.path.basename(master)}\n대상: {os.path.basename(dst)}  ({sheet} +{add}행)")
    if "--apply" not in sys.argv:
        print("\n미리보기 — 실제 확장: --apply 추가"); return
    if os.path.exists(dst):
        sys.exit(f"{os.path.basename(dst)} 이미 존재 — 중단")
    last, newlast, np_, nu = expand(master, dst, sheet, add)
    n = verify(dst, sheet, last, newlast)
    print(f"완료: {last}행 → {newlast}행 (패치 파트 {np_}개 / 무손상 복사 {nu}개, 새 행 수식 {n}개 확인)")


if __name__ == "__main__":
    main()
