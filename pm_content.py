# -*- coding: utf-8 -*-
"""
pm_content.py — 정기점검 점검내용을 **호기별로 구조화**하고 깨진 문자를 조사한다 (2026-08-11 지시)
================================================================================
사용자 지시: "정기점검에서 점검내용에서 호기별 분류하고 내용에서 ?이런거 들어가있는거
조사해서 클로드 코드에서 엑셀에 반영된 내용 다 읽을 수 있게"

원칙:
  · **원문은 안 고친다.** 파싱 결과는 파생 DB(`db/pm_content.db`)에 따로 담는다 —
    파생 자료는 언제든 다시 만들 수 있어야 하므로 매 실행 전체 재구축(멱등)한다.
  · 읽는 원천은 정본(앱 DB `list_sheet_rows`) 우선, Excel 은 DB 에 없는 행만 보탠다.
    04시트 점검내용은 앱 DB 필드에 실려 있음을 실측으로 확인했다(내용 있는 행 65/700).
  · 깨진 문자('?? 호기'·치환문자·HTML 엔티티)는 **연속 물음표만** 깨짐으로 센다 —
    정상 문장의 물음표 하나를 깨짐이라 부르면 사람이 멀쩡한 값을 고치러 간다(typo_watch).
  · 원문 교정은 **근거(밴드 원본의 같은 건 온전한 텍스트)가 있을 때만**, 기존 승인
    경로(`ledger_db.enqueue` → 앱 DB 감사로그 + Excel 보관 큐)로만 한다. 근거의 문은
    셋: 같은 프로젝트NO · 깨진 글자를 걷어낸 원장 텍스트가 밴드 텍스트에 담겨 있음 ·
    후보 밴드 블록이 유일함.

  python pm_content.py           # 파싱 + 파생 DB 저장 + 리포트 (원문 불변)
  python pm_content.py --queue   # 근거가 선 깨진 점검내용만 교정 큐
"""
import sys, os, re, json, glob, sqlite3
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

DB_PATH = os.path.join(ROOT, "db", "pm_content.db")
REPORT_MD = os.path.join(ROOT, "reports", "정기점검_호기분류.md")
CACHE = os.path.join(ROOT, "band", "cache")
SHEET, ID_COL = "04_정기점검", "점검ID"
# 깨진 문자를 찾아볼 본문 열들 — 04시트에서 사람이 글을 적는 칸 전부
TEXT_COLS = ("점검내용", "이상내용", "추가작업내용", "문제내용", "조치내용", "비고")

UNKNOWN_RE = re.compile(r"[?�]{1,}\s*호\s*기")
CHUNK_RE = re.compile(r"((?:\d{1,2}\s*[,~\-·/]\s*)*\d{1,2})\s*호\s*기")
HASH_RE = re.compile(r"#\s*(\d{1,2})(?!\d)")
HO_RE = re.compile(r"(?<![\d.])(\d{1,2})\s*호(?!기|\d)")
BROKEN_PATTERNS = (
    ("치환문자", re.compile("�")),
    ("연속물음표", re.compile(r"\?{2,}")),
    ("박스문자", re.compile(r"[▒░▓]")),
    ("HTML엔티티", re.compile(r"&(?:amp|lt|gt|quot|nbsp|#\d+);")),
)


def parse_units(text):
    """점검내용 → {"units": [호기 번호들], "unknown": 깨져서 호기를 못 읽은 자리 여부}.

    '1,2호기'·'1~3호기'·'#4'·'3호' 를 다 받는다. 호기 번호는 1~99 만 인정한다 —
    날짜·금액 조각을 호기라고 우기지 않는다.
    """
    s = str(text or "")
    units, unknown = set(), bool(UNKNOWN_RE.search(s))
    for m in CHUNK_RE.finditer(s):
        chunk = m.group(1)
        parts = re.split(r"[,·/]", chunk)
        for p in parts:
            p = p.strip()
            r = re.fullmatch(r"(\d{1,2})\s*[~\-]\s*(\d{1,2})", p)
            if r:
                a, b = int(r.group(1)), int(r.group(2))
                if 1 <= a <= b <= 99 and b - a < 50:
                    units.update(range(a, b + 1))
                continue
            if p.isdigit() and 1 <= int(p) <= 99:
                units.add(int(p))
    stripped = CHUNK_RE.sub(" ", s)
    for m in HASH_RE.finditer(stripped):
        units.add(int(m.group(1)))
    for m in HO_RE.finditer(stripped):
        units.add(int(m.group(1)))
    return {"units": sorted(units), "unknown": unknown}


def broken_flags(text):
    """본문 하나에서 깨진 문자 갈래 목록. 물음표 하나짜리는 세지 않는다(정상 문장일 수 있다)."""
    s = str(text or "")
    return [name for name, pat in BROKEN_PATTERNS if pat.search(s)]


def analyze_row(row):
    """행 dict(원문 그대로) → 파생 결과. **row 를 고치지 않는다.**"""
    content = str(row.get("점검내용") or "").strip()
    parsed = parse_units(content)
    broken = {}
    for col in TEXT_COLS:
        f = broken_flags(row.get(col))
        if f:
            broken[col] = f
    return {"pid": str(row.get(ID_COL) or "").strip(),
            "project_no": str(row.get("프로젝트NO") or "").strip(),
            "camp": str(row.get("캠프명") or "").strip(),
            "has_content": bool(content),
            "units": parsed["units"], "unknown": parsed["unknown"],
            "broken": broken, "content": content}


def collect_rows():
    """정본(앱 DB) 우선 + Excel 은 DB 에 없는 행만. (camp_standardize 와 같은 원칙)"""
    rows, seen = [], set()
    import app_store
    for r in app_store.list_sheet_rows(SHEET):
        rid = str(r.get(ID_COL) or "").strip()
        # ★ 점검ID 없는 행도 버리지 않는다 — 실측 7행이 전부 깨진 점검내용이었다([169]).
        #   ID 있는 행만 seen 에 올려 Excel 중복 판정에 쓴다.
        if rid:
            seen.add(rid)
        rows.append(("db", r))
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    xlsx_only = 0
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        hdr = [str(h).strip() if h is not None else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        for raw in ws.iter_rows(min_row=5, values_only=True):
            g = lambda c: (raw[ix[c]] if c in ix and ix[c] < len(raw) else None)
            rid = str(g(ID_COL) or "").strip()
            if not rid or rid in seen:
                continue
            xlsx_only += 1
            rows.append(("xlsx", {c: g(c) for c in (ID_COL, "프로젝트NO", "캠프명") + TEXT_COLS}))
    wb.close()
    return rows, os.path.basename(master), xlsx_only


def save_db(results):
    """파생 DB 전체 재구축 — 원문은 발췌(200자)만 담는다. 정본은 앱 DB/원장이다."""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.executescript(
            "CREATE TABLE IF NOT EXISTS pm_content("
            " pid TEXT PRIMARY KEY, project_no TEXT, camp TEXT, has_content INT,"
            " units TEXT, unit_unknown INT, broken TEXT, src TEXT,"
            " content_excerpt TEXT, parsed_at TEXT);"
            "CREATE TABLE IF NOT EXISTS pm_units("
            " pid TEXT, unit INT, PRIMARY KEY(pid, unit));")
        now = datetime.now().astimezone().isoformat(timespec="seconds")
        with conn:
            conn.execute("DELETE FROM pm_content")
            conn.execute("DELETE FROM pm_units")
            for src, a in results:
                conn.execute(
                    "INSERT OR REPLACE INTO pm_content VALUES(?,?,?,?,?,?,?,?,?,?)",
                    (a["pid"], a["project_no"], a["camp"], int(a["has_content"]),
                     json.dumps(a["units"]), int(a["unknown"]),
                     json.dumps(a["broken"], ensure_ascii=False), src,
                     a["content"][:200], now))
                for u in a["units"]:
                    conn.execute("INSERT OR REPLACE INTO pm_units VALUES(?,?)", (a["pid"], u))
    finally:
        conn.close()


def band_repair_candidates(broken_rows):
    """깨진 점검내용의 교정 근거를 밴드 원본에서 찾는다.

    문 셋 — 셋 다 서야 후보가 된다:
      ① 같은 프로젝트NO 가 담긴 밴드 글일 것
      ② 깨진 글자([?▒�]·공백)를 걷어낸 원장 텍스트 조각들이 밴드 텍스트에 담겨 있을 것
      ③ 그런 밴드 블록이 **유일**할 것 (여럿이면 사람 몫 — 잘못 붙이면 더 나쁘다)
    """
    if not broken_rows:
        return []
    projects = {a["project_no"] for _, a in broken_rows if a["project_no"]}
    texts = {}                                     # project -> [본문 덩어리]
    for f in glob.glob(os.path.join(CACHE, "*.json")):
        b = os.path.basename(f)
        if b.startswith(("dump_", "raw_")):
            continue
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for p in (d.get("posts") or {}).values():
            body = p.get("content") or ""
            for blk in re.split(r"[♣✅]", body):
                for prj in projects:
                    if prj and prj in blk:
                        texts.setdefault(prj, []).append(blk.strip())
    out = []
    for src, a in broken_rows:
        cands = texts.get(a["project_no"]) or []
        pieces = [w for w in re.sub(r"[?▒░▓�]", " ", a["content"]).split() if len(w) >= 3]
        hits = []
        for blk in cands:
            flat = re.sub(r"\s+", "", blk)
            if pieces and all(re.sub(r"\s+", "", w) in flat for w in pieces):
                if "호기" in blk and not UNKNOWN_RE.search(blk):
                    hits.append(blk)
        uniq = sorted(set(hits))
        if len(uniq) == 1:
            fixed = re.sub(r"\s+", " ", uniq[0]).strip()
            out.append({"pid": a["pid"], "project_no": a["project_no"],
                        "before": a["content"], "after": fixed,
                        "units": parse_units(fixed)["units"],
                        "no_key": bool(a.get("no_key"))})
    return out


def main():
    rows, master, xlsx_only = collect_rows()
    results = [(src, analyze_row(r)) for src, r in rows]
    # ★ 점검ID 없는 행을 버리지 않는다 (2026-08-11 실측: ID 없는 7행이 **전부**
    #   '▒▒ ?? 호기 ▒▒' 깨진 내용이었다 — 버리면 깨짐 10건 중 3건만 보인다 [169]).
    #   앱 신규등록분이 ID 채번 전이라 생기는 과도기 상태다. 파생 DB 열쇠는
    #   프로젝트NO 로 대신 만들고, 교정 큐도 프로젝트NO 를 열쇠로 쓴다.
    fixed, seen_pid, no_key = [], set(), 0
    for s, a in results:
        if not a["pid"]:
            if not a["project_no"]:
                continue                     # ID 도 프로젝트NO 도 없으면 가리킬 방법이 없다
            a = dict(a)
            base = f"무ID:{a['project_no']}"
            pid, n = base, 2
            while pid in seen_pid:
                pid, n = f"{base}#{n}", n + 1
            a["pid"], a["no_key"] = pid, True
            no_key += 1
        seen_pid.add(a["pid"])
        fixed.append((s, a))
    results = fixed
    save_db(results)
    with_content = [(s, a) for s, a in results if a["has_content"]]
    with_units = [(s, a) for s, a in with_content if a["units"]]
    unknown = [(s, a) for s, a in with_content if a["unknown"]]
    broken = [(s, a) for s, a in results if a["broken"]]
    repairs = band_repair_candidates(
        [(s, a) for s, a in with_content if a["unknown"] or a["broken"].get("점검내용")])
    unit_count = {}
    for _, a in with_units:
        for u in a["units"]:
            unit_count[u] = unit_count.get(u, 0) + 1

    print(f"04시트 행 {len(results)} (DB 우선 · Excel 에만 {xlsx_only} · 점검ID 없음 {no_key}) · "
          f"점검내용 있음 {len(with_content)}")
    print(f"호기 읽힘 {len(with_units)} · 호기 깨짐(?? 호기) {len(unknown)} · 깨진 문자 행 {len(broken)}")
    print(f"밴드 근거로 교정 가능 {len(repairs)}건")

    os.makedirs(os.path.dirname(REPORT_MD), exist_ok=True)
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    L = ["# 정기점검 점검내용 — 호기 분류·깨진 문자 조사", "",
         f"- 생성: {now} · 원장: {master} · 파생 DB: `db/pm_content.db` (원문 불변)",
         f"- 행 {len(results)} (앱 DB 우선, Excel 에만 있던 행 {xlsx_only} · **점검ID 없는 행 {no_key}** — "
         f"앱 신규등록분 ID 채번 전, 프로젝트NO 로 추적) · 점검내용 있음 {len(with_content)}",
         f"- 호기 읽힘 {len(with_units)}행 · 호기 자리가 깨짐 {len(unknown)}행 · 깨진 문자 포함 {len(broken)}행",
         ""]
    if unit_count:
        L += ["## 호기 분포 (점검내용에서 읽은 것)", "",
              "| 호기 | 점검 행수 |", "|---|---|"]
        for u in sorted(unit_count):
            L.append(f"| {u}호기 | {unit_count[u]} |")
        L.append("")
    if broken or unknown:
        L += ["## 깨진 문자 행 — 원문은 그대로 두고 여기 적는다", ""]
        listed = set()
        for _, a in unknown + broken:
            if a["pid"] in listed:
                continue
            listed.add(a["pid"])
            what = " · ".join(f"{c}:{'/'.join(f)}" for c, f in a["broken"].items()) or "?? 호기"
            L.append(f"- `{a['pid']}` {a['camp']} ({a['project_no'] or '프로젝트NO 없음'}) — {what}")
            if a["content"]:
                L.append(f"  - 원문 발췌: `{a['content'][:100]}`")
        L.append("")
    if repairs:
        L += ["## 밴드 근거가 선 교정안 (문 셋 통과: 같은 프로젝트·조각 전부 포함·후보 유일)", ""]
        for r in repairs:
            L.append(f"- `{r['pid']}` ({r['project_no']}) 호기 {r['units']} — 교정안: `{r['after'][:100]}`")
        L += ["", "반영: `python pm_content.py --queue` (앱 DB 감사로그 + Excel 보관 큐)"]
    elif unknown or broken:
        L += ["밴드 원본에서 근거를 못 찾아 자동 교정은 없다 — 위 행들은 사람 몫이다.",
              "(수집 세션이 해당 프로젝트의 밴드 글을 받아 오면 다음 실행에서 다시 찾는다)"]
    open(REPORT_MD, "w", encoding="utf-8").write("\n".join(L) + "\n")
    print(f"리포트: {REPORT_MD}")

    if "--queue" in sys.argv and repairs:
        # 점검ID 없는 행은 프로젝트NO 를 열쇠로 쓴다(앱 DB 가 project_no 로 찾는다).
        items = [{"sheet": SHEET,
                  "key": (r["project_no"] if r["no_key"] else r["pid"]),
                  "key_col": ("프로젝트NO" if r["no_key"] else ID_COL),
                  "col": "점검내용", "value": r["after"], "vtype": "text",
                  "only_if_empty": False,
                  "evidence": f"밴드 원본({r['project_no']}) · 깨진 이전값 '{r['before'][:60]}'"}
                 for r in repairs]
        import ledger_writer as W
        import ledger_db
        print("큐 추가:", W.queue_add(items))
        print("DB 흡수:", ledger_db.intake_json(source="pm_content"))
    elif "--queue" in sys.argv:
        print("교정 근거가 선 것이 없어 큐에 넣지 않는다 — '0건 성공'이 아니라 근거 부재다")
    return 0


if __name__ == "__main__":
    sys.exit(main())
