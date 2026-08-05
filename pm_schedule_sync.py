# -*- coding: utf-8 -*-
"""
류지영 정기점검 스케줄 원본을 관리대장과 앱에 자동 반영한다.

원본은 장비(호기) 단위이고 UJ 프로젝트번호가 없다. 따라서 04_정기점검에
가짜 프로젝트를 만들지 않고, 27_정기점검원본일정에 캠프·일정·기사별로
집계한다. 04시트와 같은 캠프·같은 달인 건만 프로젝트번호를 연결하며,
완료/실제점검일은 원본의 예정일보다 항상 우선한다.

실행:
  python pm_schedule_sync.py             # 분석만
  python pm_schedule_sync.py --apply     # 27시트 갱신(vN+1, 내용 동일 시 미생성)
"""
from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import date, datetime

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from ecount_reconcile import load_config, resolve_master
from findings_sheet import build_generic_sheet, upsert
from source_dirs import PM_SCHEDULE_DIR
from zscan import camp_key

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SHEET_NAME = "27_정기점검원본일정"
HEADERS = [
    "일정ID", "기준연도", "분기", "예정월", "점검예정일", "캠프명", "담당기사",
    "장비수", "장비내역", "연결프로젝트NO", "반영상태", "원본행", "원본파일",
]
WIDTHS = [21, 10, 9, 11, 13, 28, 12, 9, 38, 26, 18, 14, 34]
EXCLUDE_WORDS = ("철거", "점검불가", "폐쇄", "매각")


def quarter_of(d: date) -> int:
    return (d.month - 1) // 3 + 1


def quarter_months(q: int) -> set[int]:
    start = (q - 1) * 3 + 1
    return {start, start + 1, start + 2}


def _norm_text(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _month(v) -> int | None:
    if isinstance(v, (datetime, date)):
        return v.month
    if isinstance(v, (int, float)) and 1 <= int(v) <= 12:
        return int(v)
    m = re.search(r"(?<!\d)(1[0-2]|0?[1-9])\s*월?", _norm_text(v))
    return int(m.group(1)) if m else None


def _iso_date(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _norm_text(v)
    m = re.search(r"(20\d{2})[./-]\s*(\d{1,2})[./-]\s*(\d{1,2})", s)
    if not m:
        return ""
    try:
        return date(*map(int, m.groups())).isoformat()
    except ValueError:
        return ""


def _stable_id(year: int, quarter: int, camp: str, when: str, tech: str) -> str:
    raw = f"{year}|{quarter}|{camp_key(camp)}|{when}|{_norm_text(tech)}"
    return f"SCH-{year}Q{quarter}-" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10].upper()


def _predicted_date(target_month: str, history: list[str]) -> tuple[str, str, str]:
    """동일 장비의 과거 실제 점검일을 대상 월로 옮긴 보수적 예측.

    원본에 3분기 날짜가 비어 있어도 같은 행에는 2·1분기 실제 점검일이 남아 있다.
    가장 최근 날짜의 일(day)을 예정월에 적용하되 월말을 넘기지 않는다. 공식 예정일을
    만들거나 덮어쓰지 않고, 앱에서만 ``예측``으로 구분해 보여 주기 위한 값이다.
    """
    try:
        year, month = (int(x) for x in target_month.split("-", 1))
        start = date(year, month, 1)
    except (TypeError, ValueError):
        return "", "", ""
    past = sorted({d for d in history if _iso_date(d) and _iso_date(d) < start.isoformat()})
    if not past:
        return "", "", ""
    latest = date.fromisoformat(past[-1])
    day = min(latest.day, calendar.monthrange(year, month)[1])
    predicted = date(year, month, day).isoformat()
    confidence = "높음" if len({date.fromisoformat(d).day for d in past}) == 1 else "중간"
    return predicted, f"동일 장비 과거 점검 {past[-1]}의 일자 패턴", confidence


def _fallback_predicted_date(target_month: str, history: list[str], label: str) -> tuple[str, str, str]:
    """동일 장비 이력이 없을 때 담당기사(없으면 전체)의 과거 점검일 중앙값을 사용."""
    try:
        year, month = (int(x) for x in target_month.split("-", 1))
        start = date(year, month, 1)
    except (TypeError, ValueError):
        return "", "", ""
    past = sorted({d for d in history if _iso_date(d) and _iso_date(d) < start.isoformat()})
    if not past:
        return "", "", ""
    days = sorted(date.fromisoformat(d).day for d in past)
    day = min(days[(len(days) - 1) // 2], calendar.monthrange(year, month)[1])
    return date(year, month, day).isoformat(), \
        f"{label} 과거 점검 {len(past)}건의 일자 중앙값", "낮음"


def find_latest_source(folder: str = PM_SCHEDULE_DIR) -> str:
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"정기점검 원본 폴더가 없습니다: {folder}")
    found = []
    for base, _dirs, files in os.walk(folder):
        for name in files:
            low = name.lower()
            if name.startswith("~$") or not low.endswith(".xlsx"):
                continue
            if "정기점검" not in name or not any(x in name for x in ("스케줄", "스케쥴", "schedule")):
                continue
            path = os.path.join(base, name)
            try:
                found.append((os.path.getmtime(path), os.path.getsize(path), path))
            except OSError:
                continue
    if not found:
        raise FileNotFoundError(f"정기점검 스케줄 xlsx가 없습니다: {folder}")
    return max(found)[2]


def stable_snapshot(source: str, temp_dir: str) -> tuple[str, str]:
    """열려 있는 Excel 파일도 마지막 저장본을 읽되, 복사 중 변경되면 중단한다."""
    before = os.stat(source)
    dst = os.path.join(temp_dir, os.path.basename(source))
    shutil.copy2(source, dst)
    after = os.stat(source)
    if (before.st_size, before.st_mtime_ns) != (after.st_size, after.st_mtime_ns):
        raise RuntimeError("원본이 저장 중이라 이번 반영을 건너뜁니다. 다음 실행에서 다시 확인합니다.")
    digest = hashlib.sha256()
    with open(dst, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return dst, digest.hexdigest()


def _sheet_for(wb, year: int, quarter: int):
    prefix = re.sub(r"\s+", "", f"{year}년{quarter}분기")
    candidates = [ws for ws in wb.worksheets
                  if prefix in re.sub(r"\s+", "", ws.title) and "정기점검" in ws.title]
    if not candidates:
        raise ValueError(f"{year}년 {quarter}분기 정기점검 시트를 찾지 못했습니다.")
    return candidates[0]


def _headers(ws) -> tuple[int, dict[str, int]]:
    required = ("월", "기존 캠프명")
    for rn in range(1, min(ws.max_row, 20) + 1):
        vals = [_norm_text(v) for v in next(ws.iter_rows(min_row=rn, max_row=rn, values_only=True))]
        if all(x in vals for x in required) and any("점검일자" in x for x in vals):
            return rn, {v: i for i, v in enumerate(vals) if v}
    raise ValueError(f"{ws.title}: 머리글 행을 찾지 못했습니다.")


def parse_schedule(path: str, year: int, quarter: int) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _sheet_for(wb, year, quarter)
        header_row, h = _headers(ws)
        date_header = next((k for k in h if re.sub(r"\s+", "", k).startswith(f"{quarter}분기점검일자")), "")
        if not date_header:
            raise ValueError(f"{ws.title}: {quarter}분기 점검일자 열이 없습니다.")

        def ix(*names):
            for n in names:
                if n in h:
                    return h[n]
            return None

        i_month, i_date = ix("월"), h[date_header]
        i_old, i_new = ix("기존 캠프명"), ix("변경 캠프명")
        i_tech, i_note = ix("확정자"), ix("특이사항")
        i_unit, i_kind, i_model = ix("호기"), ix("종류"), ix("모델")
        history_cols = [i for name, i in h.items()
                        if "점검일자" in re.sub(r"\s+", "", name) and name != date_header]
        months = quarter_months(quarter)
        grouped = defaultdict(lambda: {
            "rows": [], "units": [], "camp": "", "tech": "", "month": 0, "date": "",
            "history": [],
        })
        excluded = 0
        scanned = 0

        for rn, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                 start=header_row + 1):
            vals = list(row)
            get = lambda i: vals[i] if i is not None and i < len(vals) else None
            camp = _norm_text(get(i_new)) or _norm_text(get(i_old))
            month = _month(get(i_month))
            if not camp or month not in months:
                continue
            scanned += 1
            searchable = " ".join(_norm_text(v) for v in vals)
            if any(word in searchable for word in EXCLUDE_WORDS):
                excluded += 1
                continue
            exact = _iso_date(get(i_date))
            if exact:
                d = date.fromisoformat(exact)
                if d.year != year or d.month not in months:
                    exact = ""
                else:
                    # 원본의 '월'은 초안 분류이고 확정 점검일과 다른 행이 있다.
                    # 날짜가 있으면 그 실제 달을 예정월로 써야 8월인데 7월 완료처럼 보이지 않는다.
                    month = d.month
            tech = _norm_text(get(i_tech))
            when = exact or f"{year}-{month:02d}"
            key = (camp_key(camp), when, tech)
            g = grouped[key]
            g.update({"camp": camp, "tech": tech, "month": month, "date": exact})
            g["rows"].append(rn)
            for hi in history_cols:
                hist = _iso_date(get(hi))
                if hist and hist not in g["history"]:
                    g["history"].append(hist)
            equip = " ".join(x for x in (
                _norm_text(get(i_unit)), _norm_text(get(i_kind)), _norm_text(get(i_model))
            ) if x)
            if equip and equip not in g["units"]:
                g["units"].append(equip)

        records = []
        for g in grouped.values():
            when = g["date"] or f"{year}-{g['month']:02d}"
            predicted, prediction_basis, confidence = _predicted_date(
                f"{year}-{g['month']:02d}", g["history"]
            ) if not g["date"] else ("", "", "")
            rows = sorted(g["rows"])
            records.append({
                "일정ID": _stable_id(year, quarter, g["camp"], when, g["tech"]),
                "기준연도": year,
                "분기": f"{quarter}분기",
                "예정월": f"{year}-{g['month']:02d}",
                "점검예정일": g["date"],
                "예측점검일": predicted,
                "예측근거": prediction_basis,
                "예측신뢰도": confidence,
                "_과거점검일": sorted(g["history"]),
                "캠프명": g["camp"],
                "담당기사": g["tech"],
                "장비수": len(rows),
                "장비내역": " · ".join(g["units"]),
                "원본행": str(rows[0]) if len(rows) == 1 else f"{rows[0]}~{rows[-1]}",
            })
        # 동일 장비 이력이 없는 신규·이전 캠프도 달력에서 사라지지 않게 하되, 담당기사
        # 과거 패턴(담당자 미정이면 전체 중앙값)을 낮은 신뢰도로 명시한다.
        tech_history = defaultdict(list)
        all_history = []
        for r in records:
            for hist in r.get("_과거점검일") or []:
                if hist not in all_history:
                    all_history.append(hist)
                if hist not in tech_history[r["담당기사"]]:
                    tech_history[r["담당기사"]].append(hist)
        for r in records:
            if not r["점검예정일"] and not r["예측점검일"]:
                history = tech_history.get(r["담당기사"]) if r["담당기사"] else all_history
                label = f"담당기사 {r['담당기사']}" if r["담당기사"] else "전체 담당기사"
                r["예측점검일"], r["예측근거"], r["예측신뢰도"] = _fallback_predicted_date(
                    r["예정월"], history or all_history, label
                )
            r.pop("_과거점검일", None)

        records.sort(key=lambda r: (
            r["점검예정일"] or r["예측점검일"] or r["예정월"],
            camp_key(r["캠프명"]), r["담당기사"], r["일정ID"]
        ))
        return {
            "sheet": ws.title, "header_row": header_row, "scanned": scanned,
            "excluded": excluded, "records": records,
        }
    finally:
        wb.close()


def _master_pm(master: str) -> list[dict]:
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    try:
        if "04_정기점검" not in wb.sheetnames:
            return []
        ws = wb["04_정기점검"]
        hdr = [str(v or "").strip() for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        idx = {v: i for i, v in enumerate(hdr) if v}
        out = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            rec = {k: row[i] if i < len(row) else None for k, i in idx.items()}
            if rec.get("점검ID") or rec.get("프로젝트NO"):
                out.append(rec)
        return out
    finally:
        wb.close()


def link_master(records: list[dict], master_rows: list[dict]) -> list[dict]:
    by_camp_month = defaultdict(list)
    for r in master_rows:
        camp = camp_key(r.get("캠프명"))
        if not camp:
            continue
        planned = _iso_date(r.get("점검예정일"))
        actual = _iso_date(r.get("실제점검일"))
        month = (actual or planned)[:7]
        if month:
            by_camp_month[(camp, month)].append({
                "project": _norm_text(r.get("프로젝트NO")),
                "planned": planned,
                "actual": actual,
                "status": _norm_text(r.get("점검상태")),
                "tech": _norm_text(r.get("담당기사")),
            })

    # 같은 캠프가 한 달 안에 여러 날짜·기사 그룹으로 나뉜 실제 원본이 있다. 이때
    # 동월 완료 한 건을 모든 그룹에 복제하면 장비 진행률이 다시 과대계상된다.
    source_group_count = defaultdict(int)
    source_tech_count = defaultdict(int)
    for r in records:
        camp_month = (camp_key(r["캠프명"]), r["예정월"])
        source_group_count[camp_month] += 1
        source_tech_count[camp_month + (camp_key(r.get("담당기사")),)] += 1

    out = []
    for source in records:
        r = dict(source)
        camp_month = (camp_key(r["캠프명"]), r["예정월"])
        hits = by_camp_month.get(camp_month, [])
        all_completed = [x for x in hits if x["actual"] or x["status"] == "완료"]
        exact = [x for x in hits if r["점검예정일"] and
                 r["점검예정일"] in (x["planned"], x["actual"])]
        exact_completed = [x for x in exact if x in all_completed]
        tech = camp_key(r.get("담당기사"))
        tech_hits = [x for x in hits if tech and camp_key(x.get("tech")) == tech]
        tech_completed = [x for x in tech_hits if x in all_completed]

        if exact_completed:
            completed, link_hits = exact_completed, exact
        elif source_group_count[camp_month] == 1:
            completed, link_hits = all_completed, hits
        elif tech and source_tech_count[camp_month + (tech,)] == 1 and tech_completed:
            completed, link_hits = tech_completed, tech_hits
        else:
            # 여러 원본 일정 중 어느 것인지 날짜·기사로 특정할 수 없으면 완료를
            # 복제하지 않는다. 원장과 원본 모두 보존하고 매칭 대기로 남긴다.
            completed, link_hits = [], exact

        projects = sorted({x["project"] for x in link_hits if x["project"]})
        if completed:
            status = "완료 실적 우선"
        elif exact:
            status = "04 일정 일치"
        elif link_hits:
            status = "04 동월 연결"
        else:
            status = "프로젝트 매칭 대기"
        r["연결프로젝트NO"] = " · ".join(projects)
        r["반영상태"] = status
        r["실제점검일"] = max((x["actual"] for x in completed if x["actual"]), default="")
        r["완료장비수"] = int(r.get("장비수") or 0) if completed else 0
        out.append(r)
    return out


def sheet_rows(records: list[dict], source_name: str) -> list[tuple]:
    return [tuple(r.get(h, source_name if h == "원본파일" else "") for h in HEADERS)
            for r in records]


def build_sheet(records: list[dict], source_name: str, styled: bool = False) -> str:
    xml = build_generic_sheet(
        SHEET_NAME, HEADERS, WIDTHS, sheet_rows(records, source_name),
        "[자동동기화] 류지영 정기점검 스케줄 원본의 현재 분기 일정입니다. "
        "원본은 장비 단위라 캠프·예정일(또는 예정월)·기사별로 묶었습니다. "
        "UJ번호를 임의 생성하지 않으며, 04_정기점검의 완료·실제점검일이 항상 우선합니다.",
        empty_text="현재 분기 반영 대상 없음",
    )
    if not styled:
        return xml
    # 관리대장의 기존 제목/안내/머리글/입력행 스타일을 재사용한다.
    # styles.xml은 건드리지 않아 차트·도형·사용자 지정 서식을 보존한다.
    xml = xml.replace('<row r="1">', '<row r="1" ht="28" customHeight="1">')
    xml = xml.replace('<row r="2">', '<row r="2" ht="24" customHeight="1">')
    xml = xml.replace('<row r="4">', '<row r="4" ht="30" customHeight="1">')
    xml = re.sub(r'<row r="(\d+)"', lambda m:
                 f'<row r="{m.group(1)}" ht="32" customHeight="1"' if int(m.group(1)) >= 5 else m.group(0),
                 xml)
    xml = re.sub(r'<c r="([A-Z]+)1"', r'<c r="\g<1>1" s="135"', xml)
    xml = re.sub(r'<c r="([A-Z]+)2"', r'<c r="\g<1>2" s="136"', xml)
    xml = re.sub(r'<c r="([A-Z]+)4"', r'<c r="\g<1>4" s="23"', xml)
    xml = re.sub(
        r'<c r="([A-Z]+)(\d+)"',
        lambda m: f'<c r="{m.group(1)}{m.group(2)}" s="38"'
        if int(m.group(2)) >= 5 else m.group(0),
        xml,
    )
    # 긴 제목·안내문은 열 너비에 잘리지 않도록 표 전체 폭으로 합친다.
    xml = xml.replace(
        "</worksheet>",
        '<mergeCells count="2"><mergeCell ref="A1:M1"/>'
        '<mergeCell ref="A2:M2"/></mergeCells></worksheet>',
    )
    return xml


def report_payload(source: str, digest: str, parsed: dict, records: list[dict],
                   year: int, quarter: int) -> dict:
    statuses = defaultdict(int)
    for r in records:
        statuses[r["반영상태"]] += 1
    equipment_total = sum(int(r.get("장비수") or 0) for r in records)
    equipment_completed = sum(int(r.get("완료장비수") or 0) for r in records)
    compact = [{k: r.get(k, "") for k in (
        "일정ID", "예정월", "점검예정일", "예측점검일", "예측근거", "예측신뢰도",
        "실제점검일", "캠프명", "담당기사", "장비수", "완료장비수", "장비내역",
        "연결프로젝트NO", "반영상태",
    )} for r in records]
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "source_sha256": digest,
        "sheet": parsed["sheet"],
        "year": year,
        "quarter": quarter,
        "source_rows_in_quarter": parsed["scanned"],
        "excluded_rows": parsed["excluded"],
        "schedule_groups": len(records),
        "exact_date_groups": sum(bool(r["점검예정일"]) for r in records),
        "month_only_groups": sum(not bool(r["점검예정일"]) for r in records),
        "predicted_date_groups": sum(bool(r.get("예측점검일")) for r in records),
        "equipment_total": equipment_total,
        "equipment_completed": equipment_completed,
        "equipment_pending": max(0, equipment_total - equipment_completed),
        "status_counts": dict(sorted(statuses.items())),
        "schedule": compact,
        "pending": [{
            "일정ID": r["일정ID"], "예정월": r["예정월"], "캠프명": r["캠프명"],
            "담당기사": r["담당기사"], "장비수": r["장비수"],
        } for r in records if r["반영상태"] == "프로젝트 매칭 대기"],
    }


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--source")
    ap.add_argument("--source-dir", default=PM_SCHEDULE_DIR)
    ap.add_argument("--master")
    ap.add_argument("--year", type=int)
    ap.add_argument("--quarter", type=int, choices=(1, 2, 3, 4))
    ap.add_argument("--report", default=os.path.join(ROOT, "reports", "pm_schedule_sync.json"))
    args = ap.parse_args(argv)

    today = date.today()
    year = args.year or today.year
    quarter = args.quarter or quarter_of(today)
    source = args.source or find_latest_source(args.source_dir)
    master = args.master or resolve_master(load_config()["reconcile"]["master_xlsx"])

    with tempfile.TemporaryDirectory(prefix="pm_schedule_") as td:
        snapshot, digest = stable_snapshot(source, td)
        parsed = parse_schedule(snapshot, year, quarter)
    records = link_master(parsed["records"], _master_pm(master))
    payload = report_payload(source, digest, parsed, records, year, quarter)
    os.makedirs(os.path.dirname(args.report), exist_ok=True)
    with open(args.report, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"정기점검 원본: {os.path.basename(source)} · {parsed['sheet']}")
    print(f"  분기 원본 {parsed['scanned']}행 · 제외 {parsed['excluded']}행 · "
          f"일정 {len(records)}그룹(날짜확정 {payload['exact_date_groups']} / 월예정 {payload['month_only_groups']})")
    print("  " + " · ".join(f"{k} {v}" for k, v in payload["status_counts"].items()))
    if args.apply:
        xml = build_sheet(records, os.path.basename(source), styled=True)
        dst, msg = upsert(master, xml, sheet_name=SHEET_NAME, headers=HEADERS)
        print(f"  {SHEET_NAME}: {msg}")
        if dst:
            print("   ", dst)
    else:
        print("  분석만 완료 — 실제 반영은 --apply")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
