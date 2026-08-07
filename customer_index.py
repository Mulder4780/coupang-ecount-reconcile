# -*- coding: utf-8 -*-
"""
customer_index.py — ERP 거래처등록 → **거래처코드 색인**, 캠프명과 잇는다

사용자 지시(2026-08-05): "거래처 코드 전수 조사 및 앱과 엑셀에 추가해서 표기해."
앱 화면에는 캠프명(강서1MB(가양A))만 있고 ERP 거래처코드(CU177)가 없어, ERP·씨앗과
대조할 때마다 사람이 눈으로 찾아야 했다.

원천: ERP `거래처등록` 엑셀(ESA001M — 거래처코드·거래처명·주소·담당자·보유장비명).
이름이 무작위라 **머리글로 판별**한다(거래처코드+거래처명이 2행에 있는 파일).

캠프명 ↔ 거래처명 잇기
  · 1순위 완전 일치         강서1MB(가양A) = 강서1MB(가양A)
  · 2순위 괄호/공백 제거 후 일치   강서1MB(가양A) → 강서1MB가양A
  · 3순위 캠프 핵심코드 일치       강서1MB / M_강릉1 / 송파5캠프 같은 앞부분
  맞는 게 **하나일 때만** 확정한다. 여럿이면 후보로 남기고 추측하지 않는다.

산출: reports/거래처코드_색인.json  {캠프명 → {code, erp_name, addr, manager, equip}}
      reports/거래처코드_색인.csv   (사람이 엑셀에서 보는 용도)

  python customer_index.py
"""
import csv
import html
import glob
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

OUT_JSON = os.path.join(ROOT, "reports", "거래처코드_색인.json")
OUT_CSV = os.path.join(ROOT, "reports", "거래처코드_색인.csv")


# 값이 없다는 뜻으로 사람이 넣어 둔 것들. 이름이 아니므로 **미매칭이 아니라 입력오류**다.
# (2026-08-06 김미영 대리 지적: "밴드·카톡에 복사 붙여넣기 오류가 있어 매칭이 안 됨")
PLACEHOLDERS = {"", "0", "-", "--", "?", "??", ".", "..", "...", "N/A", "NA", "없음", "미정"}


def clean(s):
    """붙여넣기 사고를 걷어 낸다 — 이름 자체를 고치는 것이 아니라 **비교용**이다.

    실제로 원장에 들어와 있던 것들(2026-08-06 실측 18건):
      · `남김해Sub-Hub&amp;Sub-FC` — 웹에서 복사해 HTML 엔티티가 그대로 들어왔다
      · `중구1캠프 <-서초1MB(양재동C)` — 메모('←에서 옮김')가 이름 칸에 섞였다
      · `김포1Sub-FC ?(김포1 서브허브))` — 물음표·짝 안 맞는 괄호
    엔티티를 풀고 화살표 뒤를 떼면 나머지 정규화가 제 일을 할 수 있다.
    """
    t = html.unescape(str(s or ""))
    t = re.split(r"<-|->|←|→", t)[0]
    return re.sub(r"\s+", " ", t).strip()


def norm(s):
    """비교용 정규화 — 괄호·공백·기호를 지우고 대문자로."""
    return re.sub(r"[\s()\[\]_·\-/?]", "", clean(s)).upper()


def core(s):
    """캠프 핵심코드: '강서1MB(가양A)' → '강서1MB', 'M_강릉1' → 'M강릉1'."""
    s = clean(s).split("(")[0]
    return norm(s)


def bare(s):
    """ERP 표기 접두사를 뗀 형태 — 원장 `M_광주2캠프` ↔ ERP `광주2 캠프`.

    한쪽에만 붙는 표식이라 이것 때문에 같은 캠프가 둘로 갈렸다. 완전일치·정규화·
    핵심코드를 **다 놓친 뒤에만** 본다 — 먼저 보면 서로 다른 캠프를 붙일 수 있다.

    ★ 떼는 것은 `M_`·`M ` 처럼 **구분자가 붙은 접두사뿐**이다. 정규화가 밑줄을
      지운 뒤에 M 을 떼면 `MB1캠프` 가 `B1캠프` 가 된다 — 멀쩡한 캠프를 엉뚱한
      거래처에 붙이게 된다(검증 [113] 이 이 경우를 지킨다).
    """
    return core(re.sub(r"^M[_\s]+", "", clean(s)))


def load_customers():
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    import source_dirs as S
    cands = [p for p in glob.glob(os.path.join(S.ERP_DIR, "**", "*.xlsx"), recursive=True)
             if not os.path.basename(p).startswith(("~$", "ESD007E"))]
    cands.sort(key=os.path.getmtime, reverse=True)
    best = None
    for p in cands[:80]:
        try:
            wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
            ws = wb.active
            head = [str(c or "") for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
            j = "|".join(head)
            # ★ 2026-08-05: '거래처코드'만 보면 거래처관리대장(104행) 같은 다른 화면이
            #   먼저 걸린다. 거래처**등록**은 주소·담당자 열이 함께 있고 행이 압도적으로 많다.
            if "거래처코드" not in j or "거래처명" not in j or "주소" not in j:
                wb.close()
                continue
            if best and ws.max_row <= best[1]:
                wb.close()
                continue
            idx = {h: i for i, h in enumerate(head)}
            rows = []
            for r in ws.iter_rows(min_row=3, values_only=True):
                r = ["" if c is None else str(c).strip() for c in r]
                g = lambda k: (r[idx[k]] if k in idx and len(r) > idx[k] else "")
                code, name = g("거래처코드"), g("거래처명")
                if not code or not name:
                    continue
                rows.append({"code": code, "name": name, "addr": g("주소"),
                             "manager": g("담당자"), "email": g("Email"),
                             "tel": g("연락처"), "equip": g("보유장비명")})
            wb.close()
            if not best or len(rows) > len(best[2]):
                best = (os.path.basename(p), ws.max_row, rows)
        except Exception:
            continue
    return (best[2], best[0]) if best else ([], None)


def camps_from_ledger():
    """관리대장에서 실제 쓰이는 캠프명을 모은다(프로젝트NO 와 함께)."""
    from ecount_reconcile import read_ledger, load_config
    recs = read_ledger(load_config()["reconcile"]["master_xlsx"])
    out = {}
    for r in recs.values():
        camp = str(r.get("캠프명") or "").strip()
        if not camp:
            continue
        d = out.setdefault(camp, {"건수": 0, "프로젝트": set()})
        d["건수"] += 1
        if r.get("프로젝트NO"):
            d["프로젝트"].add(str(r["프로젝트NO"]))
    return out


def main():
    custs, src = load_customers()
    if not custs:
        print("거래처등록 원본을 찾지 못함 — ERP '거래처등록' 화면 엑셀이 필요하다")
        return 1
    by_exact, by_norm, by_core, by_bare = {}, {}, {}, {}
    for c in custs:
        by_exact.setdefault(c["name"], []).append(c)
        by_norm.setdefault(norm(c["name"]), []).append(c)
        by_core.setdefault(core(c["name"]), []).append(c)
        by_bare.setdefault(bare(c["name"]), []).append(c)

    camps = camps_from_ledger()
    linked, multi, none, junk = {}, {}, [], []
    for camp, info in camps.items():
        # 이름이 아닌 값은 거래처를 찾을 수 없다 — 찾으려 애쓰지 말고 **입력오류로 돌린다**.
        # 예전엔 이것들이 '미매칭'에 섞여 있어, 정말 못 찾은 캠프가 몇 개인지 알 수 없었다.
        if clean(camp).upper() in PLACEHOLDERS:
            junk.append({"camp": camp, "건수": info["건수"], "왜": "값 대신 들어간 표시"})
            continue
        for table, key, how in ((by_exact, camp, "완전일치"),
                                (by_norm, norm(camp), "정규화"),
                                (by_core, core(camp), "핵심코드"),
                                (by_bare, bare(camp), "접두사 제외")):
            hit = table.get(key) or []
            if len(hit) == 1:
                c = hit[0]
                linked[camp] = {"code": c["code"], "erp_name": c["name"], "how": how,
                                "addr": c["addr"], "manager": c["manager"],
                                "tel": c["tel"], "email": c["email"], "equip": c["equip"],
                                "건수": info["건수"]}
                break
            if len(hit) > 1:
                multi[camp] = {"how": how, "codes": [x["code"] for x in hit][:6],
                               "names": [x["name"] for x in hit][:6], "건수": info["건수"]}
                break
        else:
            none.append({"camp": camp, "건수": info["건수"]})

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    json.dump({"src": src, "customers": len(custs), "camps": len(camps),
               "linked": linked, "ambiguous": multi, "unmatched": none,
               "입력오류": junk},
              open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["캠프명", "거래처코드", "ERP거래처명", "연결방식", "원장건수",
                    "주소", "담당자", "연락처", "Email", "보유장비"])
        for camp, v in sorted(linked.items()):
            w.writerow([camp, v["code"], v["erp_name"], v["how"], v["건수"],
                        v["addr"], v["manager"], v["tel"], v["email"], v["equip"]])
        for camp, v in sorted(multi.items()):
            w.writerow([camp, "(후보 여럿)", " / ".join(v["names"]), v["how"], v["건수"],
                        "", "", "", "", ""])
        for x in sorted(none, key=lambda r: -r["건수"]):
            w.writerow([x["camp"], "(ERP에 없음)", "", "", x["건수"], "", "", "", "", ""])
    # ★ 2026-08-07 지시로 **별도 엑셀을 만들지 않는다** ("앞으로도 별도의 엑셀 파일은
    #   만들지 말고 관리대장으로만 관리해"). 같은 표가 관리대장 『27_거래처코드』
    #   시트로 들어간다 — 열 구성은 예전 엑셀과 **같게** 뒀다(자리만 옮기는 것이지
    #   내용이 바뀌는 게 아니어야 한다).
    #   2026-08-05 에 별도 파일을 택했던 이유는 "관리대장에 열을 늘릴 수 없어서"였는데,
    #   늘리는 게 아니라 **시트를 더하는 것**이면 차트·수식을 건드리지 않는다.
    #   ※ 관리대장을 openpyxl 로 열어 save() 하지 않는다(차트 파괴 — 절대규칙).
    #     `customer_sheet` 가 zip 파트 수술로 시트 XML 만 갈아끼운다.
    #   ※ 관리대장 쓰기는 **`--sheet` 를 줬을 때만** 한다. 그냥 돌리면 집계만 본다
    #     (엑셀 반영은 11:00·15:00 회차 규칙 그대로).
    if "--sheet" in sys.argv:
        try:
            import customer_sheet as CS
            rows, ok, msg = CS.apply(linked, multi, none)
            print(f"  관리대장 {CS.SHEET_NAME}: {len(rows)}행 — {msg}")
        except Exception as e:
            print(f"  ! 거래처코드 시트 갱신 실패: {str(e)[:90]}")
    else:
        print("  (관리대장 27_거래처코드 반영은 --sheet 로. 별도 엑셀은 더 만들지 않는다)")

    print(f"거래처 {len(custs)}개 · 원장 캠프 {len(camps)}개 → "
          f"코드 확정 {len(linked)} · 후보여럿 {len(multi)} · ERP에 없음 {len(none)} "
          f"→ reports/거래처코드_색인.json/csv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
