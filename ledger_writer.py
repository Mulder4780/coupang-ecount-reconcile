# -*- coding: utf-8 -*-
"""
ledger_writer.py — 관리대장 자동 입력 엔진 (확정 사실만, 빈 칸만, 새 버전으로)
================================================================================
대조 모듈들이 만든 확정 업데이트 큐(updates/pending_updates.json)를 관리대장에 반영한다.

안전 원칙:
  1. 빈 칸만 채운다 — 기존 값·수식이 있으면 그 항목은 '건너뜀'으로 기록 (--force 없음, 의도적으로 미구현)
  2. zip 단일파트 패치 — 차트·도형·x14 검증 무손상 (openpyxl save 금지 원칙 준수)
  3. 항상 vN+1 새 파일 생성 — 이전 버전 그대로 보존, 언제든 롤백 가능
  4. 반영 내역은 19_AI작업인수인계에 자동 기록 + updates/applied_*.json 보관
  5. calcChain 삭제 + fullCalcOnLoad 보장 → 엑셀이 열 때 전체 재계산(캐시 불일치 방지)

업데이트 항목 형식(JSON):
  {"sheet":"06_거래서류청구수금", "key_col":"정산ID", "key":"JS-2607-041",
   "col":"거래명세서번호", "value":"2026/07/24-3", "vtype":"text|date|number",
   "evidence":"ECOUNT SaveInvoiceAuto SlipNos", "only_if_empty":true}

실행:
  python ledger_writer.py            # 큐 미리보기(무엇을 쓸지, 쓸 수 있는지)
  python ledger_writer.py --apply    # 실제 반영(vN+1 생성)
"""
import sys, os, re, json, zipfile, glob, tempfile, time, html
from contextlib import contextmanager
from datetime import datetime, date

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPD_DIR = os.environ.get("COUPANG_UPDATES_DIR") or os.path.join(BASE_DIR, "updates")
PENDING = os.path.join(UPD_DIR, "pending_updates.json")
HDR_ROW, FIRST = 4, 5


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def date_serial(iso):
    d = datetime.strptime(str(iso)[:10], "%Y-%m-%d").date()
    return (d - date(1899, 12, 30)).days


def esc(t):
    return str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def load_queue():
    try:
        q = json.load(open(PENDING, encoding="utf-8"))
        return q if isinstance(q, list) else []
    except Exception:
        return []


@contextmanager
def queue_lock(timeout=30):
    """큐 읽기→수정→비우기를 한 임계구역으로 묶는 프로세스 간 잠금."""
    os.makedirs(os.path.dirname(PENDING) or ".", exist_ok=True)
    lock = PENDING + ".lock"
    started = time.monotonic()
    fd = None
    while fd is None:
        try:
            fd = os.open(lock, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, f"{os.getpid()} {datetime.now().isoformat()}".encode("ascii"))
        except FileExistsError:
            if time.monotonic() - started >= timeout:
                raise TimeoutError(f"자동입력 큐 잠금 대기 초과: {lock}")
            time.sleep(0.1)
    try:
        yield
    finally:
        os.close(fd)
        try:
            os.unlink(lock)
        except FileNotFoundError:
            pass


def atomic_json_dump(value, path):
    """중간에 프로세스가 종료돼도 반쪽짜리 JSON이 남지 않게 교체한다."""
    folder = os.path.dirname(path) or "."
    os.makedirs(folder, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=os.path.basename(path) + ".", suffix=".tmp", dir=folder)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(value, f, ensure_ascii=False, indent=1)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


def queue_add(items):
    """대조 모듈이 호출: 확정 업데이트를 큐에 중복 없이 추가"""
    with queue_lock():
        q = load_queue()
        seen = {(u["sheet"], u["key"], u["col"]) for u in q}
        added = 0
        for it in items:
            k = (it["sheet"], it["key"], it["col"])
            if k not in seen:
                q.append(it); seen.add(k); added += 1
        atomic_json_dump(q, PENDING)
        return added


def resolve_targets(master, queue):
    """openpyxl(read-only)로 각 항목의 (시트파일, 행번호, 열문자, 현재값) 확인"""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=False)
    plans, skips = [], []
    # 같은 시트에도 접수ID·프로젝트NO처럼 서로 다른 조회 키가 섞여 들어온다.
    # 시트명만 캐시 키로 쓰면 첫 항목의 열로 만든 사전을 뒤 항목에도 재사용해,
    # 실제 행이 있어도 "행 없음"으로 버린다.
    cache = {}
    for u in queue:
        sh = u["sheet"]
        if sh not in wb.sheetnames:
            skips.append({**u, "사유": "시트 없음"}); continue
        # 셀 직접 지정 모드 (예: 00_대시보드 B3 보고일) — 키 조회 없이 좌표로
        if u.get("cell"):
            mc = re.match(r"^([A-Z]+)(\d+)$", str(u["cell"]).upper())
            if not mc:
                skips.append({**u, "사유": f"셀 주소 오류 {u['cell']}"}); continue
            plans.append({**u, "row": int(mc.group(2)), "colL": mc.group(1)})
            continue
        cache_key = (sh, u["key_col"])
        if cache_key not in cache:
            ws = wb[sh]
            hdr = next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))
            hmap = {str(h).strip(): i + 1 for i, h in enumerate(hdr) if h is not None}
            keys = {}
            kcol = hmap.get(u["key_col"])
            if kcol:
                for i, row in enumerate(ws.iter_rows(min_row=FIRST, min_col=kcol, max_col=kcol, values_only=True)):
                    if row[0] is not None:
                        keys[str(row[0]).strip()] = FIRST + i
            cache[cache_key] = (hmap, keys)
        hmap, keys = cache[cache_key]
        if u["col"] not in hmap:
            skips.append({**u, "사유": f"열 '{u['col']}' 없음"}); continue
        rown = keys.get(str(u["key"]).strip())
        if not rown:
            skips.append({**u, "사유": f"{u['key_col']}={u['key']} 행 없음"}); continue
        cletter = col_letter(hmap[u["col"]])
        cur = wb[sh].cell(row=rown, column=hmap[u["col"]]).value if False else None
    # read_only 모드에선 개별 셀 접근이 비효율 — 현재값은 XML 단계에서 최종 확인
        plans.append({**u, "row": rown, "colL": cletter})
    wb.close()
    return plans, skips


def sheet_file_map(z):
    """시트명 → zip 내 파일경로. 속성 순서·절대/상대 Target 모두 대응(엑셀·openpyxl 산출물 공용)."""
    wbx = z.read("xl/workbook.xml").decode("utf-8")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rmap = {}
    for m in re.finditer(r"<Relationship\b[^>]*/?>", rels):
        tag = m.group(0)
        mid = re.search(r'Id="([^"]+)"', tag)
        mt = re.search(r'Target="([^"]+)"', tag)
        if mid and mt:
            rmap[mid.group(1)] = mt.group(1)
    out = {}
    for m in re.finditer(r"<sheet\b[^>]*/?>", wbx):
        tag = m.group(0)
        mn = re.search(r'name="([^"]+)"', tag)
        mr = re.search(r'r:id="([^"]+)"', tag)
        if mn and mr and mr.group(1) in rmap:
            t = rmap[mr.group(1)]
            t = t[1:] if t.startswith("/") else (t if t.startswith("xl/") else "xl/" + t)
            out[mn.group(1)] = t
    return out


def find_col_style(xml, colL):
    m = re.search(r'<c r="%s\d+" s="(\d+)"' % colL, xml)
    return m.group(1) if m else None


def cell_xml(colL, rown, style, value, vtype):
    s = f' s="{style}"' if style else ""
    ref = f'{colL}{rown}'
    if vtype == "date":
        return f'<c r="{ref}"{s}><v>{date_serial(value)}</v></c>'
    if vtype == "number":
        return f'<c r="{ref}"{s}><v>{value}</v></c>'
    return f'<c r="{ref}"{s} t="inlineStr"><is><t>{esc(value)}</t></is></c>'


def same_cell_value(cxml, value, vtype):
    """현재 XML 값과 쓸 값이 같은지 판정한다(멱등 버전 생성 방지)."""
    if "<f" in cxml:
        return False
    if vtype == "text":
        if 't="inlineStr"' not in cxml and "<is>" not in cxml:
            return False  # sharedStrings 인덱스는 이 함수만으로 원문을 알 수 없다.
        current = "".join(html.unescape(t) for t in re.findall(r"<t[^>]*>(.*?)</t>", cxml, re.S))
        return current == str(value)
    mv = re.search(r"<v[^>]*>(.*?)</v>", cxml, re.S)
    if not mv:
        return False
    current = mv.group(1).strip()
    expected = str(date_serial(value) if vtype == "date" else value).strip()
    try:
        return float(current) == float(expected)
    except ValueError:
        return current == expected


def apply_to_xml(xml, plan):
    """한 항목을 시트 XML에 반영. (성공여부, 새XML, 사유)"""
    rown, colL = plan["row"], plan["colL"]
    mrow = re.search(r'(<row r="%d"[^>]*>)(.*?)(</row>)' % rown, xml, re.S)
    if not mrow:
        mrow_empty = re.search(r'<row r="%d"[^>]*/>' % rown, xml)
        if not mrow_empty:
            return False, xml, "행 XML 없음"
        head = mrow_empty.group(0)[:-2] + ">"
        body, tail = "", "</row>"
        newrow_span = mrow_empty.span()
    else:
        head, body, tail = mrow.groups()
        newrow_span = mrow.span()

    # 셀 태그를 2단계로 정확히 잡는다: 여는 태그 → 자기닫힘(/>)이면 그 자체, 아니면 </c>까지.
    # (기존 단일 정규식은 <c r="M45" s="43"/> 의 '/'를 [^>]*가 삼켜 다음 셀까지 오매칭 — 실사고 원인)
    mopen = re.search(r'<c r="%s%d"[^>]*>' % (colL, rown), body)
    mcell = None
    if mopen:
        if mopen.group(0).endswith("/>"):
            span = mopen.span()
        else:
            mclose = body.find("</c>", mopen.end())
            span = (mopen.start(), mclose + 4)
        class _M:  # re.Match 인터페이스 최소 구현
            def __init__(s, a, b): s._a, s._b = a, b
            def group(s, i=0): return body[s._a:s._b]
            def start(s): return s._a
            def end(s): return s._b
        mcell = _M(*span)
    if mcell:
        cxml = mcell.group(0)
        has_val = bool(re.search(r"<v[\s/>]", cxml)) or ("<is>" in cxml) or ("<f" in cxml)
        if has_val and same_cell_value(cxml, plan["value"], plan.get("vtype", "text")):
            return False, xml, "동일 값(변경 없음)"
        if has_val and plan.get("only_if_empty", True):
            return False, xml, "이미 값·수식 있음(빈 칸만 정책)"
        style = (re.search(r' s="(\d+)"', cxml) or [None, find_col_style(xml, colL)])[1]
        new_cell = cell_xml(colL, rown, style, plan["value"], plan.get("vtype", "text"))
        new_body = body[:mcell.start()] + new_cell + body[mcell.end():]
    else:
        style = find_col_style(xml, colL)
        new_cell = cell_xml(colL, rown, style, plan["value"], plan.get("vtype", "text"))
        target_n = col_num(colL)
        ins = len(body)
        for mc in re.finditer(r'<c r="([A-Z]+)%d"' % rown, body):
            if col_num(mc.group(1)) > target_n:
                ins = mc.start(); break
        new_body = body[:ins] + new_cell + body[ins:]
    new_xml = xml[:newrow_span[0]] + head + new_body + tail + xml[newrow_span[1]:]
    return True, new_xml, ""


def strip_calcchain(name, data, ct, rels):
    """calcChain 제거(엑셀이 재생성) — 참조 3곳 동시 정리"""
    if name == "xl/calcChain.xml":
        return None, ct, rels
    if name == "[Content_Types].xml":
        ct = re.sub(r'<Override[^>]*calcChain[^>]*/>', "", data.decode("utf-8")).encode("utf-8")
        return ct, ct, rels
    if name == "xl/_rels/workbook.xml.rels":
        rels = re.sub(r'<Relationship[^>]*calcChain[^>]*/>', "", data.decode("utf-8")).encode("utf-8")
        return rels, ct, rels
    if name == "xl/workbook.xml":
        d = data.decode("utf-8")
        if "<calcPr" in d:
            if "fullCalcOnLoad" not in d:
                d = d.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
        else:
            d = d.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')
        return d.encode("utf-8"), ct, rels
    return data, ct, rels


def _main():
    apply_mode = "--apply" in sys.argv
    queue = load_queue()
    if not queue:
        print("큐 비어 있음 — 반영할 확정 업데이트가 없습니다."); return

    if "--master" in sys.argv:                    # 테스트용 마스터 오버라이드
        master = sys.argv[sys.argv.index("--master") + 1]
    else:
        from ecount_reconcile import load_config, resolve_master
        cfg = load_config()
        master = resolve_master(cfg["reconcile"]["master_xlsx"])
    mv = re.search(r"_v(\d+)\.xlsx$", master)
    v = int(mv.group(1))
    dst = master.replace(f"_v{v}.xlsx", f"_v{v+1}.xlsx")

    plans, skips = resolve_targets(master, queue)
    print(f"큐 {len(queue)}건 → 반영 가능 {len(plans)} / 사전 제외 {len(skips)}")
    # 셀 좌표 모드(`cell`)에는 key·col 이 없다. 미리보기 한 줄 때문에 반영이 통째로
    # 죽지 않게 한다 — 실제로 06시트 636건(2,513셀) 적재가 여기서 KeyError로 멈췄다.
    def _label(u):
        return u.get("key") or u.get("cell") or ""

    for s in skips:
        print(f"  [제외] {s['sheet']} {_label(s)} {s.get('col','')}: {s['사유']}")
    # 좌표 모드 대량 적재는 한 줄씩 찍으면 콘솔이 수천 줄로 덮인다(토큰 절약 규칙).
    verbose = len(plans) <= 60
    for p in plans if verbose else []:
        print(f"  [예정] {p['sheet']} {_label(p)} → {p.get('col','')}({p['colL']}{p['row']}) = {p['value']}  ⟵ {p.get('evidence','')}")
    if not verbose:
        from collections import Counter
        c = Counter((p["sheet"], p["colL"]) for p in plans)
        print("  [예정] " + ", ".join(f"{sh} {cl}열 {n}셀" for (sh, cl), n in sorted(c.items())))
    if not apply_mode:
        print("\n미리보기 모드 — 실제 반영: python ledger_writer.py --apply"); return
    if not plans:
        print("반영할 항목 없음"); return
    if os.path.exists(dst):
        sys.exit(f"{os.path.basename(dst)} 이미 존재 — 중단")

    zin = zipfile.ZipFile(master)
    smap = sheet_file_map(zin)
    by_file = {}
    for p in plans:
        by_file.setdefault(smap[p["sheet"]], []).append(p)

    done, skipped2 = [], []
    xmls = {}
    for fname, ps in by_file.items():
        xml = zin.read(fname).decode("utf-8")
        for p in sorted(ps, key=lambda x: -x["row"]):
            ok, xml, why = apply_to_xml(xml, p)
            (done if ok else skipped2).append({**p, "사유": why})
        xmls[fname] = xml

    # 실제 입력이 0건이면 새 버전을 만들지 않는다 (전부 이미 값 있음 = 원장과 이미 일치)
    if not done:
        zin.close()
        os.makedirs(UPD_DIR, exist_ok=True)
        stamp0 = datetime.now().strftime("%Y%m%d_%H%M")
        atomic_json_dump({"applied": [], "skipped": skips + skipped2, "version": "미생성(0건)"},
                         os.path.join(UPD_DIR, f"applied_{stamp0}.json"))
        atomic_json_dump([], PENDING)
        print(f"\n입력할 빈 칸 없음(전부 기존 값 보유) — 새 버전 미생성, 큐 {len(skipped2)}건 정리 완료")
        return

    # 인수인계 행도 같은 버전에 기록
    hand = smap.get("19_AI작업인수인계")
    if hand:
        hx = xmls.get(hand) or zin.read(hand).decode("utf-8")
        nr = max(int(r) for r in re.findall(r'<row r="(\d+)"', hx)) + 1
        summary = "; ".join(f"{d['sheet'].split('_')[0]} {_label(d)}.{d.get('col', d.get('colL',''))}={d['value']}" for d in done[:10])
        if len(done) > 10:
            summary += f" 외 {len(done)-10}건"
        # 스타일은 직전 행 A셀에서 상속(하드코딩 금지 — 파일마다 스타일 표가 다름)
        mprev = re.search(r'<c r="A%d" s="(\d+)"' % (nr - 1), hx)
        s_attr = f' s="{mprev.group(1)}"' if mprev else ""
        row = (f'<row r="{nr}" spans="1:3"><c r="A{nr}"{s_attr} t="inlineStr"><is><t>{esc(datetime.now().strftime("%Y-%m-%d"))} #auto입력</t></is></c>'
               f'<c r="B{nr}"{s_attr} t="inlineStr"><is><t>에이전트 자동 입력 {len(done)}건 (빈 칸만·근거 보유)</t></is></c>'
               f'<c r="C{nr}"{s_attr} t="inlineStr"><is><t>{esc(summary)} / 근거·전량은 updates/applied 파일 참조. ledger_writer.py 자동 기록.</t></is></c></row>')
        hx = hx.replace("</sheetData>", row + "</sheetData>", 1)
        hx = re.sub(r'<dimension ref="A1:C\d+"/>', f'<dimension ref="A1:C{nr}"/>', hx, count=1)
        xmls[hand] = hx

    ct = rels = None
    zout = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        data = zin.read(it.filename)
        data, ct, rels = strip_calcchain(it.filename, data, ct, rels)
        if data is None:
            continue
        if it.filename in xmls:
            data = xmls[it.filename].encode("utf-8")
        zout.writestr(it.filename, data)
    zout.close(); zin.close()

    # 검증
    z = zipfile.ZipFile(dst)
    assert z.testzip() is None, "zip 무결성 실패"
    import openpyxl
    w = openpyxl.load_workbook(dst, read_only=True, data_only=False)
    bad = []
    for d in done:
        cell = w[d["sheet"]].cell(row=d["row"], column=col_num(d["colL"])).value
        if d.get("vtype") == "date":
            okv = cell is not None
        else:
            okv = str(cell) == str(d["value"]) or (d.get("vtype") == "number" and cell == d["value"])
        if not okv:
            bad.append((d["key"], d["col"], cell))
    w.close(); z.close()
    assert not bad, f"검증 실패: {bad}"

    # 큐 정리·보관
    os.makedirs(UPD_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    atomic_json_dump({"applied": done, "skipped": skips + skipped2, "version": f"v{v+1}"},
                     os.path.join(UPD_DIR, f"applied_{stamp}.json"))
    atomic_json_dump([], PENDING)
    print(f"\n반영 완료: v{v} → v{v+1}, 입력 {len(done)}건 / 건너뜀 {len(skipped2)}건 (검증 통과)")
    print("   ", dst)
    for s in skipped2:
        print(f"  [건너뜀] {_label(s)} {s.get('col', s.get('colL',''))}: {s['사유']}")


def main():
    # 다른 AI가 원장을 잡고 있으면 여기서 멈춘다(동시 수정 시 한쪽이 통째로 묻힌다)
    from claim_guard import require
    require("ledger", "ledger_writer")
    global PENDING
    if "--queue" in sys.argv:                     # 테스트용 큐 경로 오버라이드
        PENDING = sys.argv[sys.argv.index("--queue") + 1]
    if "--apply" in sys.argv:
        with queue_lock():
            return _main()
    return _main()


if __name__ == "__main__":
    main()
