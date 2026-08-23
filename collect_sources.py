# -*- coding: utf-8 -*-
"""
collect_sources.py — 흩어져 있는 **원본 자료를 한 폴더로 모은다**
================================================================================
사용자 지시(2026-07-28): "0. 원본 자료 — 여기에 자료 모두 복사해서 붙여넣어주고
데이터 정리해줘. 구분해서 잘 보이게 깔끔하게 정리해줘."

지금까지 원본이 PC 여기저기(inbox·kakao/inbox·band/cache)에 흩어져 있었다.
PC가 꺼지거나 사람이 바뀌면 **무엇이 원본인지 아무도 모른다.** 그래서
관리대장 옆 '0. 원본 자료' 폴더에 종류별로 모아 둔다.

  0. 원본 자료/
      0. 수집안내.txt        ← 무엇이 언제 어디서 왔는지 (이 도구가 갱신)
      1~4, 7. 자료유형/YYYY/MM/날짜/
      5. 정기점검 스케쥴 원본/  ← 최신 편집본 + 이전본 날짜별 보관
      6. PO 원본/YYYY/PO번호/  ← 쿠팡 PO 통지문·견적서

원칙
  · **복사만 한다. 원본을 지우거나 옮기지 않는다.** PC 쪽은 그대로 둔다.
  · 크기·수정시각이 같으면 건너뛴다(Z: 네트워크 드라이브가 느리다).
  · 덮어쓰기 전에 내용이 다르면 이름 뒤에 날짜를 붙여 **둘 다 남긴다.**

사용
  python collect_sources.py            # 무엇을 복사할지만 보여 준다 (안전)
  python collect_sources.py --apply    # 실제 복사
"""
import os
import sys
import glob
import json
import shutil
import time
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from source_dirs import (  # noqa: E402
    ORIGIN_ROOT, ERP_DIR, COUPANG_DIR, KAKAO_DIR, BAND_DIR, MISC_DIR,
    PO_DIR, PO_DIRS, RECEIPT_DIR, RECEIPT_DIRS,
)
from source_organizer import dated_dir, po_dir_for  # noqa: E402

GUIDE = os.path.join(ORIGIN_ROOT, "0. 수집안내.txt")
GUIDE_YEAR_DIRS = {str(y) for y in range(2000, 2100)}


def _same(a, b, sa=None, sb=None):
    """이미 같은 파일인가 — 크기와 수정시각(초)으로 판단. 느린 드라이브에서 해시는 과하다.

    부르는 쪽이 이미 stat 을 갖고 있으면 그것을 준다([198]) — Z:(SMB)에서
    `os.stat(경로)` 는 파일당 왕복 한 번(135~155ms)이고 목록에 딸려 온 stat 은
    0.04ms 다. 안 주면 예전처럼 스스로 묻는다(옛 호출자를 안 깬다).
    """
    try:
        if sa is None:
            sa = os.stat(a)
        if sb is None:
            sb = os.stat(b)
    except OSError:
        return False
    return sa.st_size == sb.st_size and int(sa.st_mtime) == int(sb.st_mtime)


# 이카운트에서 'Excel' 을 누르면 파일이 Downloads 로 떨어지고 이름이 무작위다
# (`8W1JR7MGB50PHOP.xlsx`). 사람이 옮기지 않아도 되게 여기서 직접 집어 온다.
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")
DOWNLOAD_DAYS = 14      # 오래된 건 이미 반영됐거나 무관하다
# ★ Downloads 는 개인 폴더다. **아무 파일이나 퍼 오지 않는다** —
#   내용을 열어 아는 종류로 판별된 것만 가져온다.
#   ★ 2026-08-08: 회계 원장류를 넷으로 가르면서 여기도 같이 넓혔다. 안 넓히면 새 이름을
#     단 순간 계정별원장·분개장·현금출납장이 **'모르는 종류'가 되어 안 가져온다** —
#     분류를 고친 대가로 자료가 조용히 끊긴다.
KNOWN = ("ledger", "ledger_acct", "journal", "cashbook",
         "po", "sales", "tax", "stmt", "slips", "taxinv", "hometax", "receipt",
         "billing_status")


def _dest_base(k):
    """갈래 → 어느 통으로 갈지. **한 곳에서만 정한다**([162] — 예전엔 같은 식이 두 줄이었다).

    ★ billing_status(PO·청구 대조 현황표)는 `9. 미분류` 로 간다 — 원본은 남기되
      **소비하는 통에는 안 넣는다**(분담판 [91]). 입금 통에 두면 receipt_fill 이 읽고,
      PO 통에 두면 PO 대조가 읽는데 그쪽은 06시트에 **쓰는** 길이라 더 나쁘다([170]).
    """
    if k == "po":
        return COUPANG_DIR
    if k == "receipt":
        return RECEIPT_DIR
    if k == "billing_status":
        return MISC_DIR
    return ERP_DIR


def plan():
    """[(원본, 대상, 분류)] — 어디서 어디로 갈지. 판단은 여기서만 한다."""
    # 캐시는 **회차 한 번 동안만** 산다 — 오래 들고 있으면 그 사이 남이 넣은
    # 파일을 못 본다. 같은 프로세스에서 두 번 부르는 곳(검증·앱 서버)이 있어
    # 여기서 비운다(2026-08-22 에 만들면서 이 줄을 빠뜨렸다).
    _SRC_STAT.clear()
    _DST_ENTRIES.clear()
    jobs = []

    # ERP 내보내기 / 쿠팡 목록 — 파일명이 무작위일 수 있어 **내용으로** 가른다
    try:
        from inbox_scan import classify_cached as classify, LABEL
    except Exception:
        classify, LABEL = None, {}

    def kind_of(path):
        """★ 비싼 열기는 캐시 검사 뒤에 온다([168]).

        예전엔 맨몸 `classify` 라 회차마다 파일을 다시 열었다. 아래에서 **공유 폴더**
        (`Z:16. Share/…/오종현`)까지 내용으로 가르게 됐으므로 그대로 두면 매 회차
        네트워크 왕복이 는다. `classify_cached` 는 (크기·수정시각·규칙판)으로 먼저 거른다.
        """
        try:
            return classify(path) if classify else "unknown"
        except Exception:
            return "unknown"

    for src in sorted(glob.glob(os.path.join(BASE, "inbox", "*.xls*"))):
        if os.path.basename(src).startswith("~$"):
            continue
        k = kind_of(src)
        base = _dest_base(k)
        jobs.append((src, dated_dir(base, src), LABEL.get(k, "엑셀")))

    # Downloads 에 떨어진 이카운트 내보내기 — 아는 종류만, 최근 것만
    cutoff = __import__("time").time() - DOWNLOAD_DAYS * 86400
    for src in sorted(glob.glob(os.path.join(DOWNLOADS, "*.xls*"))):
        if os.path.basename(src).startswith("~$"):
            continue
        try:
            if os.path.getmtime(src) < cutoff:
                continue
        except OSError:
            continue
        k = kind_of(src)
        if k not in KNOWN:
            continue
        base = _dest_base(k)
        jobs.append((src, dated_dir(base, src),
                     LABEL.get(k, "엑셀") + " (Downloads)"))

    # 카카오톡 내보내기
    for src in sorted(glob.glob(os.path.join(BASE, "kakao", "inbox", "*.txt"))):
        jobs.append((src, dated_dir(KAKAO_DIR, src), "카톡 대화 내보내기"))

    # 밴드 캐시 사본 — **덤프 폴더에 두지 않는다**([334] 의 쓰는 쪽).
    #
    # 전에는 `수집본` 으로 넣고 "밴드 API 원문" 이라 적었다.  그런데 그것은
    # 원문이 아니라 **가공본**이고(바로 위 주석이 스스로 그렇게 적어 뒀다),
    # `source_dirs.band_dump_dirs()` 는 그 폴더를 "브라우저/API 밴드 JSON 원본의
    # 정본 탐색 경로" 로 정의한다.  그래서 흡수기가 제 출력을 입력으로 다시
    # 먹었다 — 옛 값이 나중에 처리되며 새 수확을 덮었다(2026-08-19 실사고).
    # 읽는 쪽은 [334] 가 막았지만 **쌓이는 것은 여기서 멈춘다**:
    # 실측 2026-08-20 그 폴더 json 159개 335.5MB 중 **42개 306.7MB(91%)** 가
    # 이 사본이었고 매일 26MB 씩 늘고 있었다.
    #
    # 사본 자체는 그대로 남긴다([172]) — 없애는 것이 아니라 **자리를 옮긴다**.
    # 이미 `수집본` 에 쌓인 것은 건드리지 않는다(지우는 것은 사람이 정한다).
    for src in sorted(glob.glob(os.path.join(BASE, "band", "cache", "*.json"))):
        # 일부러 격리해 둔 것을 다시 심지 않는다.  유령 캐시·가짜 묘비는 읽지
        # 말라고 이름을 바꿔 둔 파일인데, 그것을 원본 폴더로 복사하면 이름 규칙을
        # 모르는 다음 도구가 그대로 되살린다(2026-08-12 유령 밴드가 스스로를
        # 되살린 그 모양이다).
        name = os.path.basename(src)
        if "유령" in name or ".ghost-" in name or "격리" in name:
            continue
        # ★ **이 폴더에는 캐시만 있는 것이 아니다.**  `raw_api2_90610953.json`
        #   처럼 진짜 API 수집본도 같이 산다 — 그것을 캐시로 부르면 덤프 폴더
        #   밖으로 나가 흡수가 통째로 빠진다([172] 잘못 거르는 쪽이 더 나쁘다).
        #   정본 캐시의 이름은 쓰는 쪽(`convert_dump.swap_in`)이 정한 대로
        #   **정확히 `<밴드번호>.json`** 이다.  이름으로 후보를 좁힌 뒤
        #   **스키마로 확인한다**([334] 의 `looks_like_cache` 를 빌린다 — 여기서
        #   새로 판정하면 읽는 쪽과 갈린다).  못 읽으면 캐시라 우기지 않고
        #   예전 자리(덤프)로 둔다 — 모르는 쪽으로 옮기지 않는다([169]).
        is_cache = False
        stem = os.path.splitext(name)[0]
        try:
            from band import convert_dump as _cd
            if _cd.plausible_band(stem):
                with open(src, encoding="utf-8") as fh:
                    is_cache = _cd.looks_like_cache(json.load(fh))
        except Exception:
            is_cache = False
        if is_cache:
            jobs.append((src, dated_dir(os.path.join(BAND_DIR, "캐시사본"), src),
                         "밴드 캐시 사본(원문 아님)"))
        else:
            jobs.append((src, dated_dir(os.path.join(BAND_DIR, "수집본"), src),
                         "밴드 API 원문"))

    # 오종현 공유 폴더도 매번 취합한다. 공유 폴더의 파일은 지우지 않고 정본 보관소에 복사한다.
    # PO는 한 오더가 여러 프로젝트를 묶을 수 있어 프로젝트번호보다 PO번호가 안정적인 분류키다.
    for folder in PO_DIRS:
        try:
            if not os.path.isdir(folder) or os.path.commonpath(
                    [os.path.abspath(folder), os.path.abspath(ORIGIN_ROOT)]) == os.path.abspath(ORIGIN_ROOT):
                continue
        except (OSError, ValueError):
            continue
        # ★ 목록에 딸려 온 stat 을 버리지 않는다([198]). `os.walk` 는 **이름만**
        #   주므로 아래 `copy_one` 이 파일마다 `os.stat(src)` 를 다시 부른다 —
        #   Z:(SMB)에서는 그것이 파일당 왕복 한 번(135~155ms)이고 `scandir`
        #   항목의 `.stat()` 은 0.04ms 다(3,000배).
        #   실측 2026-08-23 09:35 회차: 이 폴더에서 온 판정이 **전부 `[동일]`**
        #   (이미 옮겨져 아무 일도 안 한다)인데 그것만으로 **40분 제한에 걸려
        #   끊겼다**(코드 -9 · 회차 80.7분). 2026-08-22 에 목적지 쪽 왕복 둘을
        #   없앴지만 **원본 쪽 하나가 남아** 회차가 그대로 죽었다 — 그날 스스로
        #   "다음 회차가 답한 뒤에 한다"고 적어 둔 자리이고, 그 회차가 답했다.
        # ⚠ `skip_dirs=()` 는 **일부러 비운 것**이다 — 색인의 기본 목록을 말없이
        #   물려받으면 `_보관`·`_바로가기` 안의 파일이 조용히 빠지면서 화면은
        #   "완료"라고 적는다([198] 의 ⚠). `os.walk` 도 아무것도 안 걸렀다.
        from source_index import walk_stat    # 늦게 — 순환 import 를 만들지 않는다
        found = []
        for base, name, st in walk_stat(folder, skip_dirs=()):
            if name.startswith("~$") or name.lower() in ("thumbs.db", ".ds_store"):
                continue
            found.append((os.path.join(base, name), st))
        # walk_stat 은 순서를 보장하지 않는다(스택이라 폴더가 역순이다) — 로그는
        # 사람이 읽는 목록이라 정렬한다. stat 객체는 비교할 수 없으므로 경로만 본다.
        found.sort(key=lambda t: t[0])
        for src, st in found:
            _SRC_STAT[src] = st
            jobs.append((src, po_dir_for(src), "오종현 PO 원본"))

    for folder in RECEIPT_DIRS:
        try:
            if not os.path.isdir(folder) or os.path.commonpath(
                    [os.path.abspath(folder), os.path.abspath(ORIGIN_ROOT)]) == os.path.abspath(ORIGIN_ROOT):
                continue
        except (OSError, ValueError):
            continue
        for src in sorted(glob.glob(os.path.join(folder, "*.xls*"))):
            if os.path.basename(src).startswith("~$"):
                continue
            # ★ **폴더는 종류의 증거가 아니다**(분담판 [91]). 오종현의 입금 공유 폴더에는
            #   'CSOS PO관련 누락 및 취소 건 현황'(PO·청구 대조표)이 같이 산다. 예전엔
            #   그 폴더의 엑셀을 **전부** `7. 입금내역` 으로 날라서 회차마다 `__dup_` 사본이
            #   쌓였다 — 실측 2026-08-20 그 통의 xlsx 11개 중 9개가 그 표였다.
            #   여기가 **사본이 늘어나던 자리**다. 갈래를 물어 `9. 미분류` 로 보낸다.
            k = kind_of(src)
            base = _dest_base(k) if k == "billing_status" else RECEIPT_DIR
            jobs.append((src, dated_dir(base, src),
                         "오종현 PO·청구 대조표(입금 원본 아님)"
                         if k == "billing_status" else "오종현 입금내역"))

    return jobs


#: 목적지 폴더 목록 — **폴더마다 한 번만** 훑는다([198]).
#: ★ 왜 — 예전에는 파일 하나에 Z: 왕복이 **셋**이었다: `os.path.exists(dst)` ·
#:   `os.stat(src)` · `os.stat(dst)`. 실측 2026-08-23 로그: 판정 12,786건 중
#:   **12,607건이 `[동일]`**, 곧 이미 옮겨져 있어 **아무 일도 안 하는** 파일이다.
#:   12,607 x 3 x 0.145초 = 약 91분 — 그래서 이 단계가 40분 제한에 걸려 끊겼고,
#:   뒤따르는 `원본 폴더 정리` 까지 같이 못 끝냈다.
#: ★ `scandir` 항목의 stat 은 목록에 딸려 오므로 폴더 하나에 왕복 한 번이면 된다.
#: ⚠ 회차 한 번 동안만 산다 — 오래 들고 있으면 남이 그 사이에 넣은 파일을 못 본다.
_DST_ENTRIES = {}

# ★ 목록에 딸려 온 원본 stat 을 여기 담아 `copy_one` 에 넘긴다([198]).
#   `jobs` 모양은 **그대로 3튜플**이다 — 검증(`t358`)이 `for s, d, why in jobs`
#   로 언패킹하므로 4튜플로 넓히면 그 검사가 그날부터 죽는다.
_SRC_STAT = {}


def _dst_entries(d):
    """그 폴더의 {파일이름: stat}. 못 읽으면 빈 사전이다(그러면 예전처럼 묻는다)."""
    v = _DST_ENTRIES.get(d)
    if v is None:
        v = {}
        try:
            with os.scandir(d) as it:
                for e in it:
                    try:
                        if e.is_file():
                            v[e.name] = e.stat()
                    except OSError:
                        pass
        except OSError:
            pass                      # 아직 없는 폴더 — 그대로 빈 사전
        _DST_ENTRIES[d] = v
    return v


def copy_one(src, dst_dir, apply=False, st=None):
    """반환: ('복사'|'동일'|'이름바꿈'|'실패', 대상경로)

    `st` 는 원본의 stat — 부르는 쪽이 목록에서 받아 두었으면 넘긴다([198]).
    """
    ents = _dst_entries(dst_dir)
    name = os.path.basename(src)
    dst = os.path.join(dst_dir, name)
    if name in ents:
        if _same(src, dst, st, ents[name]):
            return "동일", dst
        # 내용이 다르면 덮어쓰지 않는다 — 어느 쪽이 맞는지 사람이 봐야 한다
        # 목록을 만들 때 이미 받은 원본 시각을 다시 Z:에 묻지 않는다([198]).
        src_mtime = st.st_mtime if st is not None else os.path.getmtime(src)
        stamp = datetime.fromtimestamp(src_mtime).strftime("%y%m%d")
        root, ext = os.path.splitext(dst)
        dst = f"{root}_{stamp}{ext}"
        name2 = os.path.basename(dst)
        if name2 in ents:
            if _same(src, dst, st, ents[name2]):
                return "동일", dst
            # 밴드 캐시는 운영 정본이 아니라 **하루 한 번 남기는 복구용 사본**이다.
            # 본문·댓글이 들어올 때마다 두 캐시가 계속 바뀌는데, 같은 날짜 이름을
            # 매번 덮어쓰면 26MB를 느린 Z:에 반복 전송하며 원본정리 회차를 잡아먹는다.
            # 개별 글 정본은 `게시글보관`과 DB가 계속 갱신하므로, 오늘 사본이 이미
            # 있으면 다음 날 새 날짜 사본이 생길 때까지 다시 복사하지 않는다.
            try:
                cache_root = os.path.normcase(os.path.abspath(
                    os.path.join(BAND_DIR, "캐시사본")))
                here = os.path.normcase(os.path.abspath(dst_dir))
                is_band_cache_copy = (here == cache_root or
                                      here.startswith(cache_root + os.sep))
            except (OSError, ValueError, TypeError):
                is_band_cache_copy = False
            if is_band_cache_copy:
                return "동일", dst
        state = "이름바꿈"
    else:
        state = "복사"
    if apply:
        try:
            os.makedirs(dst_dir, exist_ok=True)
            shutil.copy2(src, dst)          # copy2 = 수정시각까지 보존 (다음 실행에서 '동일' 판정)
            # 방금 만든 것을 목록에 넣는다 — 같은 회차에 같은 이름이 또 오면
            # 목록에 없어서 '복사' 로 읽히고, 그러면 같은 파일을 두 번 쓴다.
            try:
                ents[os.path.basename(dst)] = os.stat(dst)
            except OSError:
                pass
        except OSError as e:
            return f"실패({e.strerror})", dst
    return state, dst


#: 행수 캐시 — **파일이 안 바뀌면 다시 열지 않는다**([168]).
#: 규칙(무엇을 한 행으로 세나)을 고치면 이 판을 **손으로 올린다** — 안 올리면
#: 원본이 그대로인 한 옛 답이 영원히 이긴다(이 프로젝트가 여러 번 겪은 모양:
#: `inbox_scan.RULES_VERSION` · `PM_SCHED_VER` · `_BAND_EV_VER`).
_ROWS_VER = 1
_ROWS_CACHE = os.path.join(BASE, "reports", ".수집안내_행수.json")
_rows_cache = None


def _rows_cache_load():
    global _rows_cache
    if _rows_cache is None:
        try:
            with open(_ROWS_CACHE, encoding="utf-8") as fh:
                d = json.load(fh)
            _rows_cache = d.get("행수") or {} if d.get("판") == _ROWS_VER else {}
        except Exception:
            _rows_cache = {}
    return _rows_cache


def _rows_cache_save():
    if _rows_cache is None:
        return
    try:
        os.makedirs(os.path.dirname(_ROWS_CACHE), exist_ok=True)
        tmp = _ROWS_CACHE + ".tmp"
        with open(tmp, "w", encoding="utf-8", newline="") as fh:
            json.dump({"판": _ROWS_VER, "행수": _rows_cache}, fh, ensure_ascii=False)
        os.replace(tmp, _ROWS_CACHE)
    except Exception:
        pass          # 캐시를 못 남겨도 회차를 세우지 않는다 — 다음에 다시 센다


#: 중간 저장 주기(초) — ★ **끝에 한 번만 저장하면 죽을 때 통째로 잃는다**([388]).
#: 실측 2026-08-22: 이 단계가 40분 제한에 걸려 SIGKILL(-9) 로 죽는 바람에
#: `_rows_cache_save()` 가 **한 번도 안 불렸고** 캐시 파일이 아예 없었다 —
#: 그래서 다음 회차도 처음부터 세고 또 죽는다. **스스로를 재현하는 고장**이다.
#: 상한은 개수가 아니라 **시간**으로 정한다 — 엑셀 크기가 제각각이라 개수로는
#: 시간을 못 잡는다([388] 이 밴드 수집기에서 배운 그대로).
#: ★ 자주 저장해도 잃는 것이 없다 — 이 캐시는 **버려도 안전한 값**이고
#:   갈아끼우기는 이미 원자적이다(os.replace · [171]).
_ROWS_SAVE_EVERY_S = 30.0
_rows_saved_at = 0.0
_rows_new = 0          # 이번 실행에서 **새로 센** 개수 — 자국에 숫자로 남긴다
_rows_hit = 0          # 캐시가 답한 개수


def _rows_touch():
    """새로 센 값을 캐시에 넣은 뒤 부른다 — 주기가 되면 디스크에 남긴다."""
    global _rows_saved_at
    now = time.time()
    if now - _rows_saved_at >= _ROWS_SAVE_EVERY_S:
        _rows_cache_save()
        _rows_saved_at = now


def count_rows(path, st=None):
    """엑셀은 행 수를, 텍스트는 줄 수를 센다 — '넣었는데 비어 있는' 파일을 잡기 위해.

    ★ **비싼 읽기는 캐시 검사 뒤에 온다**([168]).  이 함수는 Z:(SMB) 의 엑셀을
      **통째로 열어 모든 시트의 모든 행**을 훑는데, `main()` 이 원본 폴더의 파일
      **하나하나마다** 부른다 — 안 바뀐 파일까지 매 회차 다시 연다.  2026-08-21
      실측: '원본 모으기' 단계 **36.8분** · 그다음 '원본 폴더 정리'가 40분 제한에
      걸려 회차가 매일 exit 1 로 죽었다([324] 가 그 범인 단계를 대 줬다).
      ⚠ 그렇다고 **이것이 36.8분 전부라고 확언하지 않는다**([172]) — 구간 자국을
      같이 넣었으니 다음 회차가 숫자로 답한다.
    ★ 못 읽은 것(-1)은 **캐시하지 않는다**([169]) — 그때 한 번 못 읽은 것을
      '이 파일은 비었다'로 굳히면 다시는 안 열어 본다."""
    global _rows_new, _rows_hit
    ext = os.path.splitext(path)[1].lower()
    try:
        # ★ 목록에 딸려 온 stat 을 버리지 않는다([198]) — Z: 에서 `os.stat(경로)` 는
        #   파일당 왕복 한 번(135~155ms)인데 scandir 항목의 stat 은 0.04ms 다.
        #   부르는 쪽이 안 주면 예전처럼 스스로 묻는다(옛 호출자를 안 깬다).
        if st is None:
            st = os.stat(path)
        key = "%s|%d|%d" % (os.path.abspath(path), st.st_size, int(st.st_mtime))
    except OSError:
        return -1
    cache = _rows_cache_load()
    hit = cache.get(key)
    if isinstance(hit, int):
        _rows_hit += 1
        return hit
    try:
        if ext.startswith(".xls"):
            import openpyxl
            w = openpyxl.load_workbook(path, read_only=True, data_only=True)
            n = sum(1 for sn in w.sheetnames for r in w[sn].iter_rows(values_only=True)
                    if sum(1 for x in r if x not in (None, "")) >= 3)
            w.close()
            cache[key] = n
            _rows_new += 1
            _rows_touch()
            return n
        if ext == ".txt":
            with open(path, encoding="utf-8", errors="replace") as f:
                n = sum(1 for _ in f)
            cache[key] = n
            _rows_new += 1
            _rows_touch()
            return n
    except Exception:
        pass
    return -1


def write_guide(rows):
    """폴더를 처음 여는 사람이 읽을 안내문. 파일이 아니라 **뜻**을 적는다."""
    L = [
        "이 폴더는 '쿠팡 통합업무 자동화'가 쓰는 원본 자료 보관소입니다.",
        f"마지막 정리: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "■ 폴더 구분",
        "  1~4, 7번 폴더      자료유형 > 연도 > 월 > 수집일 순으로 자동 보관",
        "  5. 정기점검         최신 편집본은 바로 아래, 이전본은 날짜별 보관",
        "  6. PO 원본          연도 > PO번호 순으로 통지문·견적서 보관",
        "  9. 미분류           루트에 바로 둔 자료 중 자동 분류 근거가 없는 파일",
        "",
        "■ 자료를 새로 넣을 때",
        "  해당 폴더에 그냥 넣어 주시면 됩니다. 파일 이름은 아무래도 괜찮습니다 —",
        "  프로그램이 파일을 열어 내용으로 종류를 알아냅니다.",
        "  ★ 이카운트에서 'Excel' 로 내려받으면 Downloads 에 떨어지는데, 옮기지 않으셔도",
        "    됩니다. 최근 2주 안에 받은 것 중 아는 종류만 자동으로 가져옵니다.",
        "",
        "■ 이카운트에서 내보낼 화면 (매출 대조에 쓰는 것)",
        "  1) 재고 I > 영업관리 > 판매일괄회계반영 > 매출(세금)계산서조회(재고)",
        "  2) 재고 I > 영업관리 > 판매일괄회계반영 > 매출(세금)계산서현황(재고)  ← 품목·내역 단위",
        "  3) 회계 I > 전자(세금)계산서 > 홈택스자료조회 > 전자(세금)계산서",
        "     ('미반영'이 아니라 '전체' 를 고르고 기간을 올해로 잡아 주세요)",
        "  4) 회계 I > 전자(세금)계산서 > 이카운트 vs 홈택스 자료비교  ← 차이가 바로 보이는 화면",
        "",
        "■ 주의",
        "  · 이카운트에서 내보낼 때는 조회 조건을 넣어 **화면에 행이 보이는 상태**에서",
        "    내보내세요. 조건 없이 내보내면 회사명 한 줄만 든 빈 파일이 됩니다.",
        "  · 파일을 지우지 마세요. 옛 자료도 대조 근거로 계속 씁니다.",
        "",
        "■ 현재 보관 현황",
        "  ※ 아래는 연도 보관함 밖에 남은 신규·미분류 파일입니다. 연도별 과거 정본과",
        "     밴드 게시글은 다시 나열하지 않습니다. 전체 개별 목록은 앱의 원본 자료/",
        "     원본색인에서 확인합니다. 새로 들어온 빈 파일 검사는 별도로 계속합니다.",
    ]
    for d, items in rows:
        L.append(f"  [{os.path.basename(d)}]  {len(items)}개")
        for name, n in items:
            mark = "  ★ 내용 없음" if n == 0 else (f"  {n:,}행" if n > 0 else "")
            L.append(f"      {name}{mark}")
    try:
        os.makedirs(ORIGIN_ROOT, exist_ok=True)
        with open(GUIDE, "w", encoding="utf-8") as f:
            f.write("\n".join(L) + "\n")
    except OSError as e:
        print(f"  안내문 저장 실패: {e}")


def _guide_skip_dirs(folder):
    """안내문에서 **파일별 재나열하지 않을 과거 보관함**을 돌려준다.

    연도 폴더는 이미 앱 ``원본 자료/원본색인``에서 파일별로 찾는 정본이다. 사람이
    읽는 ``0. 수집안내.txt``에 그것을 매일 다시 풀어 쓰고 엑셀 6,820개를 전부 열어
    행 수를 세는 것은 검증이 아니다. 2026-08-23 실측으로 안내문만 12,400,687바이트,
    밴드 항목 152,547개였고 중간 캐시 1,934개를 만드는 데도 6분 넘게 들었다.

    새로 들어온 파일의 빈 내용 검사는 ``main``의 jobs에서 따로 한다. 따라서 여기를
    줄여도 '빈 파일을 알아채는 문'은 닫히지 않는다. ``_보관``·``_바로가기``처럼
    연도가 아닌 다른 폴더는 예전처럼 본다 — 넓게 빼서 자료를 숨기지 않는다.
    """
    try:
        same_band = (os.path.normcase(os.path.abspath(folder)) ==
                     os.path.normcase(os.path.abspath(BAND_DIR)))
    except (OSError, ValueError, TypeError):
        same_band = False
    skip = set(GUIDE_YEAR_DIRS)
    if same_band:
        skip.add("게시글보관")
    return skip


def main():
    apply = "--apply" in sys.argv
    if not os.path.isdir(ORIGIN_ROOT):
        print(f"원본 폴더에 접근할 수 없습니다(네트워크 드라이브 확인): {ORIGIN_ROOT}")
        return 2

    # * 구간 자국 - **어디가 오래 걸리나**를 회차가 스스로 대게 한다([324] 의 순서:
    #   짐작으로 제한시간부터 늘리지 않고, 먼저 범인을 대게 만든다).  회차 층에서는
    #   이미 이름을 댔다(2026-08-21 '원본 모으기' 36.8분) - 이제 단계 **안**이다.
    time_mod = __import__("time")
    _t0 = time_mod.time()
    jobs = plan()
    _t_plan = time_mod.time() - _t0
    print(f"{'복사 실행' if apply else '미리보기(복사 안 함)'} — 대상 {len(jobs)}개\n")
    tally = {}
    # 새로 들어오는 파일은 연도 보관함으로 곧장 들어가므로, 안내문에서 과거 연도를
    # 생략하기 **전에** 원본 자체를 검사한다. 그래야 빈 ERP/카톡 파일 경보는 그대로다.
    job_empty = set()
    for src, dst_dir, kind in jobs:
        # ★ 어제 만든 `st` 인자를 **부르는 쪽이 안 넘기고 있었다** — 그래서
        #   `copy_one` 안의 `_same()` 이 `os.stat(src)` 를 그대로 불렀다.
        #   인자를 만들었으면 **넘기는 자리까지가 한 벌**이다.
        state, dst = copy_one(src, dst_dir, apply, _SRC_STAT.get(src))
        tally[state] = tally.get(state, 0) + 1
        print(f"  [{state:6s}] {os.path.basename(dst_dir)}/{os.path.basename(dst)}   ({kind})")
        if apply and count_rows(src, _SRC_STAT.get(src)) == 0:
            job_empty.add(f"{kind}/{os.path.basename(src)}")

    _t_copy = time_mod.time() - _t0 - _t_plan
    print("\n집계: " + ", ".join(f"{k} {v}개" for k, v in sorted(tally.items())))
    if not apply:
        print("\n실제로 복사하려면:  python collect_sources.py --apply")
        return 0

    # 정리 결과를 폴더별로 세어 안내문에 남긴다
    rows, empty = [], list(job_empty)
    from source_index import walk_stat        # 늦게 — 순환 import 를 만들지 않는다
    for d in (ERP_DIR, COUPANG_DIR, KAKAO_DIR, BAND_DIR, RECEIPT_DIR):
        if not os.path.isdir(d):
            continue
        items = []
        # ★ 파일마다 Z: 에 다시 묻지 않는다([198]) — 목록이 크기·시각을 같이 준다.
        #   ⚠ `skip_dirs=()` 는 **일부러 비운 것**이다. 색인의 기본 목록을 말없이
        #     물려받으면 `_보관`·`_바로가기` 안의 파일이 안내문에서 조용히 빠지고
        #     '정리 완료'라고 적힌다 — 그 함수 주석이 경고한 바로 그 자리다.
        guide_skip = _guide_skip_dirs(d)
        # 다른 원본은 예전처럼 **아무 폴더도 말없이 빼지 않는다**. 이 명시적 갈래가
        # 있어야 색인의 SKIP_DIRS가 섞여 `_보관` 자료가 사라지는 회귀를 막는다.
        guide_walk = (walk_stat(d, skip_dirs=guide_skip) if guide_skip
                      else walk_stat(d, skip_dirs=()))
        for base, name, st in guide_walk:
            p = os.path.join(base, name)
            n = count_rows(p, st)
            items.append((os.path.relpath(p, d), n))
            if n == 0:
                empty.append(f"{os.path.basename(d)}/{os.path.relpath(p, d)}")
        items.sort()      # walk_stat 은 순서를 보장하지 않는다 — 사람이 읽는 목록이다
        rows.append((d, items))
    for d in (PO_DIR,):
        if os.path.isdir(d) and d.startswith(ORIGIN_ROOT):
            n = sum(len(f) for _b, _dd, f in os.walk(d))
            rows.append((d, [(f"(하위 폴더 포함 {n}개 파일)", -1)]))
    empty.sort()
    write_guide(rows)
    _rows_cache_save()
    _t_guide = time_mod.time() - _t0 - _t_plan - _t_copy
    print(f"안내문 갱신: {os.path.basename(GUIDE)}")
    # 이 줄이 다음 회차의 답이다 - 확언 대신 숫자로 말한다.
    print("  구간(초): 목록 %.1f · 복사 %.1f · 안내문(행수 세기) %.1f"
          "  [행수: 캐시 %d · 새로 셈 %d]"
          % (_t_plan, _t_copy, _t_guide, _rows_hit, _rows_new))
    if empty:
        print(f"\n★ 내용이 비어 있는 파일 {len(empty)}개 — 다시 내보내야 합니다")
        for x in empty:
            print(f"    {x}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
