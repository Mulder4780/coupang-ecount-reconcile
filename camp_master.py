# -*- coding: utf-8 -*-
"""
camp_master.py — **모든 캠프 한 장**: ERP 거래처 + 관리대장 + 밴드에서 캠프를 다 모은다

사용자 지시(2026-08-05): "캠프가 300개가 넘는다고 하는데 2026년 자료에서 확인이 안 될
경우 2025년 자료도 전부 가져와서 확인해보고…"

왜 필요한가(실측 2026-08-05)
  · ERP 거래처등록에 **캠프형 거래처 239개**(CU 코드 234개)가 있다.
  · 관리대장(2026년 작업)에는 **225개** 캠프만 나온다.
  · 둘을 합치면 300개 안팎이고, **ERP 에만 있는 캠프가 86개**다 — 2025년에 작업했거나
    아직 2026년 작업이 없는 곳이다. 한 곳에 모아 두지 않으면 "몇 개인지"조차 못 센다.

모으는 원천
  1. ERP 거래처등록(CU 코드·주소·담당자·보유장비)         ← 정본 키
  2. 관리대장 06/02/04 (2026년 작업 건수·최근 작업일)
  3. 밴드 캐시 본문에 등장하는 캠프명(작업 흔적)
     → 세 곳 중 어디에라도 있으면 캠프로 본다. 출처를 그대로 표시한다(추측 금지).

산출: reports/캠프마스터.json / .csv
"""
import csv
import glob
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

OUT_JSON = os.path.join(ROOT, "reports", "캠프마스터.json")
OUT_CSV = os.path.join(ROOT, "reports", "캠프마스터.csv")
CAMP = re.compile(r"(캠프|MB|MXC|SPA|FC|Sub-?hub|Sub-?FC|허브|HUB|물류센터|M_|V_|S-)", re.I)


def norm(s):
    return re.sub(r"[\s()\[\]_·\-/]", "", str(s or "")).upper()


def erp_customers():
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    import source_dirs as S
    cands = [p for p in glob.glob(os.path.join(S.ERP_DIR, "**", "*.xlsx"), recursive=True)
             if not os.path.basename(p).startswith(("~$", "ESD007E"))]
    cands.sort(key=os.path.getmtime, reverse=True)
    best = None
    for p in cands[:80]:
        try:
            wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
            ws = wb.active
            head = [str(c or "") for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
            j = "|".join(head)
            # 거래처**등록**(2,981행)만 쓴다 — '거래처코드'만 보면 거래처관리대장(104행)이
            # 먼저 걸려 캠프 매칭이 통째로 0이 된다(2026-08-05 실측).
            if "거래처코드" not in j or "거래처명" not in j or "주소" not in j:
                wb.close()
                continue
            idx = {h: i for i, h in enumerate(head)}
            out = []
            for r in ws.iter_rows(min_row=3, values_only=True):
                r = ["" if c is None else str(c).strip() for c in r]
                g = lambda k: (r[idx[k]] if k in idx and len(r) > idx[k] else "")
                if g("거래처코드") and g("거래처명"):
                    out.append({"code": g("거래처코드"), "name": g("거래처명"),
                                "addr": g("주소"), "manager": g("담당자"),
                                "tel": g("연락처"), "email": g("Email"),
                                "equip": g("보유장비명")})
            wb.close()
            if not best or len(out) > len(best[1]):
                best = (os.path.basename(p), out)
        except Exception:
            continue
    return (best[1], best[0]) if best else ([], None)


def ledger_camps():
    from ecount_reconcile import read_ledger, load_config
    recs = read_ledger(load_config()["reconcile"]["master_xlsx"])
    out = {}
    for r in recs.values():
        camp = str(r.get("캠프명") or "").strip()
        if not camp:
            continue
        d = out.setdefault(camp, {"건수": 0, "최근": "", "프로젝트": set()})
        d["건수"] += 1
        day = str(r.get("작업완료일") or "")[:10]
        if day > d["최근"]:
            d["최근"] = day
        if r.get("프로젝트NO"):
            d["프로젝트"].add(str(r["프로젝트NO"]))
    return out


def band_camps(known):
    """밴드 본문에서 이미 아는 캠프 이름이 몇 번 나오는지 — 작업 흔적의 증거."""
    hits = {}
    cache = os.path.join(ROOT, "band", "cache")
    keys = {norm(k): k for k in known}
    for f in glob.glob(os.path.join(cache, "*.json")):
        if not os.path.basename(f)[:-5].isdigit():
            continue
        try:
            posts = (json.load(open(f, encoding="utf-8")).get("posts") or {})
        except Exception:
            continue
        for p in posts.values():
            if not isinstance(p, dict):
                continue
            t = norm(p.get("content") or "")
            for nk, orig in keys.items():
                if len(nk) >= 4 and nk in t:
                    hits[orig] = hits.get(orig, 0) + 1
    return hits


def main():
    custs, src = erp_customers()
    erp_camps = {c["name"]: c for c in custs if CAMP.search(c["name"])}
    led = ledger_camps()
    all_names = set(erp_camps) | set(led)
    hits = band_camps(all_names)

    by_norm = {}
    for name in all_names:
        by_norm.setdefault(norm(name), []).append(name)

    rows, merged = [], set()
    for key, names in sorted(by_norm.items()):
        if key in merged:
            continue
        merged.add(key)
        erp = next((erp_camps[n] for n in names if n in erp_camps), None)
        lg = next((led[n] for n in names if n in led), None)
        band = max((hits.get(n, 0) for n in names), default=0)
        src_tags = []
        if erp:
            src_tags.append("ERP")
        if lg:
            src_tags.append("원장")
        if band:
            src_tags.append("밴드")
        rows.append({
            "캠프명": sorted(names, key=len)[-1],
            "별칭": " / ".join(sorted(set(names))[:4]) if len(set(names)) > 1 else "",
            "거래처코드": (erp or {}).get("code", ""),
            "ERP거래처명": (erp or {}).get("name", ""),
            "원장건수": (lg or {}).get("건수", 0),
            "최근작업일": (lg or {}).get("최근", ""),
            "밴드언급": band,
            "출처": "+".join(src_tags),
            "주소": (erp or {}).get("addr", ""),
            "담당자": (erp or {}).get("manager", ""),
            "연락처": (erp or {}).get("tel", ""),
            "보유장비": (erp or {}).get("equip", ""),
        })
    rows.sort(key=lambda r: (-r["원장건수"], r["캠프명"]))

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    json.dump({"src": src, "count": len(rows), "rows": rows},
              open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False)
    cols = ["캠프명", "거래처코드", "출처", "원장건수", "최근작업일", "밴드언급",
            "ERP거래처명", "별칭", "주소", "담당자", "연락처", "보유장비"]
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    both = sum(1 for r in rows if "ERP" in r["출처"] and "원장" in r["출처"])
    only_erp = sum(1 for r in rows if r["출처"] == "ERP")
    only_led = sum(1 for r in rows if "ERP" not in r["출처"])
    print(f"캠프 마스터 {len(rows)}곳 — ERP+원장 {both} · ERP에만 {only_erp} · "
          f"원장/밴드에만 {only_led} · 코드 있는 곳 {sum(1 for r in rows if r['거래처코드'])} "
          f"→ reports/캠프마스터.json/csv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
