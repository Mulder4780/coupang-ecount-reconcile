# -*- coding: utf-8 -*-
"""
tech_report.py — 누가 어디를 다녀왔는지 정리한다
================================================================================
'이 캠프는 누가 갔었지?'· '이번 달 김필우 기사가 몇 건 돌았지?'를 원장에서 매번 눈으로
찾고 있었다. 원장에는 이미 다 들어 있으니 모아서 보여 주기만 하면 된다.

세는 기준 — **실제로 다녀온 것**만 센다
  · 02_돌발AS접수  작업완료일이 있는 행
  · 04_정기점검    실제점검일이 있는 행
  예정일만 있고 완료일이 없는 행은 '아직 안 간 것'이므로 방문으로 세지 않는다.
  (이걸 섞으면 기사별 실적이 부풀고, 미방문 건이 눈에 안 띈다)

기사 두 명이 함께 간 건('김준형, 김필우')은 **양쪽 모두 1건씩** 센다 — 둘 다 갔기 때문이다.
합계가 행 수보다 큰 이유가 그것이라 리포트에 같이 적는다.

  python tech_report.py                 # 요약(콘솔)
  python tech_report.py --camp 창원     # 그 캠프에 누가 다녀갔는지
  python tech_report.py --who 김필우    # 그 기사가 어디를 다녀왔는지
  python tech_report.py --csv           # reports/기사별방문_YYYYMMDD.csv
"""
import sys, os, re, csv
from collections import Counter, defaultdict
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")

# (시트, 방문으로 인정하는 날짜열, 예정일열, 업무이름)
SRC = [("02_돌발AS접수", "작업완료일", "접수일자", "돌발AS"),
       ("04_정기점검", "실제점검일", "점검예정일", "정기점검")]


def split_tech(v):
    """'김준형, 김필우' → ['김준형','김필우']. 이름이 아닌 조각은 버린다."""
    from project_resolve import clean_tech
    out = []
    for part in re.split(r"[,./·]| 및 ", str(v or "")):
        t = clean_tech(part)
        if t:
            out.append(t)
    return out


def collect(master=None):
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    from project_resolve import _d
    master = master or resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)

    visits, pending, unknown = [], [], []
    for sheet, donec, planc, kind in SRC:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        for r in ws.iter_rows(min_row=5, values_only=True):
            g = lambda k: (r[ix[k]] if k in ix and ix[k] < len(r) else None)
            prj = str(g("프로젝트NO") or "").strip()
            if not prj:
                continue
            done, camp = _d(g(donec)), str(g("캠프명") or "").strip()
            techs = split_tech(g("담당기사"))
            row = {"프로젝트NO": prj, "업무": kind, "캠프명": camp,
                   "방문일": done, "예정일": _d(g(planc)),
                   "기사": ", ".join(techs) or ""}
            if not done:
                pending.append(row)          # 아직 안 간 건
            elif not techs:
                unknown.append(row)          # 다녀왔는데 누가 갔는지 안 적힌 건
            else:
                for t in techs:
                    visits.append({**row, "기사": t})
    wb.close()
    return visits, pending, unknown, master


def summary(visits):
    """기사별: 총 방문·업무별·최근월·다닌 캠프 수"""
    by = defaultdict(lambda: {"총": 0, "돌발AS": 0, "정기점검": 0,
                              "캠프": set(), "월": Counter(), "최근": ""})
    for v in visits:
        d = by[v["기사"]]
        d["총"] += 1
        d[v["업무"]] += 1
        if v["캠프명"]:
            d["캠프"].add(v["캠프명"])
        if v["방문일"]:
            d["월"][v["방문일"][:7]] += 1
            d["최근"] = max(d["최근"], v["방문일"])
    return by


def main():
    args = sys.argv[1:]
    getarg = lambda f: (args[args.index(f) + 1] if f in args and len(args) > args.index(f) + 1 else None)
    visits, pending, unknown, master = collect()

    camp_q, who_q = getarg("--camp"), getarg("--who")

    if camp_q:
        hit = [v for v in visits if camp_q in v["캠프명"]]
        print(f"'{camp_q}' 캠프 방문 {len(hit)}건 — " +
              " · ".join(f"{k} {n}" for k, n in Counter(v["기사"] for v in hit).most_common()))
        for v in sorted(hit, key=lambda x: x["방문일"])[-12:]:
            print(f"  {v['방문일']}  {v['기사']:<8} {v['업무']:<5} {v['캠프명'][:20]:<22} {v['프로젝트NO']}")
        return

    if who_q:
        hit = [v for v in visits if who_q in v["기사"]]
        camps = Counter(v["캠프명"] for v in hit if v["캠프명"])
        print(f"{who_q} 기사 방문 {len(hit)}건 · 캠프 {len(camps)}곳")
        print("  자주 간 곳:", " · ".join(f"{c[:14]} {n}" for c, n in camps.most_common(6)))
        for v in sorted(hit, key=lambda x: x["방문일"])[-12:]:
            print(f"  {v['방문일']}  {v['업무']:<5} {v['캠프명'][:20]:<22} {v['프로젝트NO']}")
        return

    by = summary(visits)
    rows = len({(v["프로젝트NO"], v["업무"]) for v in visits})
    print(f"기사별 방문 현황  ({os.path.basename(master)})")
    print(f"  방문 완료 {rows}건 / 연인원 {len(visits)}명"
          f"  · 동행 건이 있어 연인원이 더 많습니다")
    print(f"  {'기사':<8}{'총':>5}{'돌발AS':>8}{'정기점검':>9}{'캠프수':>7}  최근 방문")
    for t, d in sorted(by.items(), key=lambda x: -x[1]["총"]):
        print(f"  {t:<8}{d['총']:>5}{d['돌발AS']:>8}{d['정기점검']:>9}"
              f"{len(d['캠프']):>7}  {d['최근']}")

    m = sorted({v["방문일"][:7] for v in visits if v["방문일"]})[-3:]
    if m:
        print(f"\n  최근 3개월 ({', '.join(m)})")
        for t, d in sorted(by.items(), key=lambda x: -x[1]["총"])[:6]:
            print(f"    {t:<8}" + "  ".join(f"{mm} {d['월'].get(mm,0):>3}" for mm in m))

    print(f"\n  아직 안 간 건 {len(pending)}  ·  다녀왔는데 기사 미기입 {len(unknown)}")
    for u in unknown[:5]:
        print(f"    [기사 미기입] {u['방문일']} {u['업무']} {u['캠프명'][:18]} {u['프로젝트NO']}")
    if len(unknown) > 5:
        print(f"    … 외 {len(unknown)-5}건")

    if "--csv" in args:
        os.makedirs(REPORT_DIR, exist_ok=True)
        p = os.path.join(REPORT_DIR, f"기사별방문_{datetime.now():%Y%m%d}.csv")
        with open(p, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["기사", "방문일", "업무", "캠프명", "프로젝트NO", "예정일"],
                               extrasaction="ignore")
            w.writeheader()
            w.writerows(sorted(visits, key=lambda v: (v["기사"], v["방문일"])))
        print("\n리포트:", p)


if __name__ == "__main__":
    main()
