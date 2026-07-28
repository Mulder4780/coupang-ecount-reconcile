# -*- coding: utf-8 -*-
"""카톡 '대화 내보내기' txt → 관리대장 등록 후보 추출.

왜 필요한가: `kakao/kakao_reconcile.py` 는 **이미 원장에 있는 행**을 카톡과 맞춰 보는 대조기라
원장에 아예 없는 건은 찾아 주지 못한다. 류지영 매니저가 카톡방에 올리는
'♣ ［ 돌발유료 A/S 안내 ]' · '♣ ［ 2026년 0분기 3개월 유료 A/S 안내 ]' 공지가
접수의 원천이므로, 그 공지를 구조화해서 02·04 시트 등록 후보로 만든다.

절대 하지 않는 것
  · 원장에 쓰지 않는다 — 이 파일은 **읽기·추출 전용**이다(쓰기는 ledger_writer 큐를 거친다).
  · 원문에 없는 값을 채우지 않는다(AGENTS.md 절대규칙 10).
  · 값이 템플릿 그대로면(2026.00.00 · 빈 담당) 채우지 않고 비워 둔다 — 이게 '미정'의 진실이다.

사용
  python kakao_extract.py                      # 전체 미리보기(집계 한 줄 + 표)
  python kakao_extract.py --code UJ2601338     # 특정 건만 상세(눈으로 확인용)
  python kakao_extract.py --new                # 최신 원장에 없는 건만
  python kakao_extract.py --new --queue        # 없는 건을 ledger_writer 큐에 적재
"""
import argparse
import importlib.util
import json
import os
import re
import sys
from datetime import date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INBOX_DIR = os.path.join(BASE_DIR, "kakao", "inbox")
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")

sys.path.insert(0, BASE_DIR)


def _load_reconcile():
    """kakao/kakao_reconcile.py 의 parse_export 재사용 (내보내기 형식 파싱은 한 곳에만 둔다)."""
    path = os.path.join(BASE_DIR, "kakao", "kakao_reconcile.py")
    spec = importlib.util.spec_from_file_location("kakao_reconcile", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# 공지 본문은 '● 라벨 : 값' 의 나열이고, parse_export 가 여러 줄을 공백으로 이어 붙인다.
RE_FIELD = re.compile(r"[●∙•]\s*([^:：●]{1,20}?)\s*[:：]\s*(.*?)(?=\s*[●∙•]|\s*★|\s*▒|$)")
RE_CODE = re.compile(r"UJ\d{7}")
# 사람이 손으로 적는 칸이라 '2026.07.28' 과 '26.07.28' 이 섞여 온다.
# 두 자리 연도를 안 받으면 신청일자가 통째로 빈칸이 된다(실제 UJ2601345).
RE_DATE = re.compile(r"(?<!\d)(20\d{2}|\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})(?!\d)")
RE_PHONE = re.compile(r"01[016-9][-\s]?\d{3,4}[-\s]?\d{4}")

# 라벨 표기가 흔들려도(장비이름/프로젝트) 같은 칸으로 모은다.
ALIAS = {
    "프로젝트no": "프로젝트NO", "프로젝트 no": "프로젝트NO",
    "캠프이름": "캠프명", "캠프명": "캠프명",
    "캠프주소": "캠프주소",
    "장비이름": "설비", "프로젝트": "설비",
    "담당자명": "쿠팡담당자", "담당번호": "쿠팡담당번호",
    "e- mail": "쿠팡메일", "e-mail": "쿠팡메일", "e mail": "쿠팡메일", "mail": "쿠팡메일",
    "신청일자": "신청일자", "신청내용": "신청내용",
    "a/s 일자": "AS일자", "a/s일자": "AS일자",
    "a/s 담당": "AS담당", "a/s담당": "AS담당",
    "a/s 내용": "AS내용", "a/s내용": "AS내용",
    # 철거·납품 공지는 라벨이 다르지만 뜻은 같다(일정=방문일, 담당=기사)
    "철거일정": "AS일자", "납품일정": "AS일자", "설치일정": "AS일자",
    "철거담당": "철거담당", "납품담당": "철거담당", "설치담당": "철거담당",
}

TEMPLATE_BLANK = {"", "-", "미정", "2026.00.00", "0000.00.00"}


def norm_date(s):
    """'2026.07.27 (월요일)' → date. 템플릿(00.00)·미기입은 None."""
    if not s:
        return None
    m = RE_DATE.search(str(s))
    if not m:
        return None
    y, mo, d = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    if not (2000 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31):
        return None                      # 2026.00.00 같은 템플릿
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def clean(v):
    v = re.sub(r"\s+", " ", str(v or "")).strip(" .·-")
    return "" if v in TEMPLATE_BLANK else v


def head_of(text):
    """공지 머리글 = 첫 '●' 앞부분. 여기에 업무유형과 상태가 다 들어 있다."""
    return re.split(r"[●∙•]", text)[0][:60].strip()


def kind_of(head):
    """머리글로 업무유형을 정한다. 모르면 **찍지 않는다**(빈 문자열).

    억지로 02·04 중 하나로 몰면 엉뚱한 시트에 등록된다 — 철거·납품·외근 공유는
    02·04 어느 쪽도 아니다(AGENTS.md '기타 44건은 리졸버가 거부하는 게 맞다')."""
    if "돌발" in head:
        return "02_돌발AS접수"
    if "분기" in head or "정기" in head or "3개월" in head:
        return "04_정기점검"
    if "철거" in head:
        return "(철거·보관)"
    if "납품" in head:
        return "(납품설치)"
    if "외근" in head or "출장" in head:
        return "(외근공유)"
    return ""


def status_of(head):
    """같은 프로젝트NO 로 '안내'와 '완료'가 따로 온다 — 완료 글이 곧 작업완료 근거다."""
    if "취소" in head:
        return "취소"
    if "완료" in head:
        return "완료"
    if "안내" in head:
        return "접수"
    return ""


def extract(paths=None):
    kr = _load_reconcile()
    from band_extract import normalize_tech      # 기사명 정규화는 한 곳에만 둔다
    paths = paths or sorted(
        os.path.join(INBOX_DIR, f) for f in os.listdir(INBOX_DIR) if f.lower().endswith(".txt")
    )
    seen, out = {}, []
    for path in paths:
        for msg in kr.parse_export(path):
            text = msg["text"]
            if "프로젝트NO" not in text or not RE_CODE.search(text):
                continue
            fields = {}
            for label, value in RE_FIELD.findall(text):
                key = ALIAS.get(label.strip().lower())
                if key and key not in fields:          # 먼저 나온 값이 공지의 값
                    fields[key] = clean(value)
            code = (fields.get("프로젝트NO") or "").strip()
            m = RE_CODE.search(code) or RE_CODE.search(text)
            code = m.group(0) if m else ""
            if not code:
                continue
            head = head_of(text)
            state = status_of(head)
            raw_tech = fields.get("AS담당", "") or fields.get("철거담당", "")
            tech = normalize_tech(raw_tech)
            rec = {
                "프로젝트NO": code,
                "시트": kind_of(head),
                "상태": state,
                "캠프명": fields.get("캠프명", ""),
                "설비": fields.get("설비", ""),
                "캠프주소": fields.get("캠프주소", ""),
                "담당기사": tech,
                # 기사명이 아닌 원문(작업 메모·업체명)은 버리지 말고 남긴다 — 비고로 간다.
                "담당원문": "" if tech == raw_tech else raw_tech,
                "예정일": norm_date(fields.get("AS일자")),
                "신청일자": norm_date(fields.get("신청일자")),
                "완료일": msg["date"] if state == "완료" else None,
                "신청내용": fields.get("신청내용", ""),
                "작업내용": fields.get("AS내용", ""),
                "쿠팡담당자": fields.get("쿠팡담당자", ""),
                "쿠팡담당번호": fields.get("쿠팡담당번호", ""),
                "출처파일": os.path.basename(path),
                "카톡일시": "%s %s" % (msg["date"].isoformat(), msg["time"]),
                "발신자": msg["sender"],
                "머리글": head,
            }
            # 같은 건이 '안내 → 완료' 로 두 번 이상 온다. 나중 글이 최신 상태지만
            # 값은 **빈 칸만** 채워 합친다 — 완료 글에는 캠프주소·신청내용이 빠져 있다.
            prev = seen.get(code)
            if prev is None:
                seen[code] = len(out)
                out.append(rec)
            else:
                old = out[prev]
                newer = rec["카톡일시"] >= old["카톡일시"]
                base, other = (rec, old) if newer else (old, rec)
                merged = dict(other)
                for k, v in base.items():
                    if v:
                        merged[k] = v
                merged["글수"] = old.get("글수", 1) + 1
                out[prev] = merged
    out.sort(key=lambda r: (r["신청일자"] or r["예정일"] or date(1900, 1, 1), r["프로젝트NO"]))
    return out


def ledger_codes():
    """최신 원장(02·04)에 이미 있는 프로젝트NO 집합."""
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from workbook_patch import latest_master
    path, ver = latest_master()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    codes = set()
    for sn in ("02_돌발AS접수", "04_정기점검"):
        for row in wb[sn].iter_rows(min_row=5, min_col=2, max_col=2, values_only=True):
            if row[0]:
                codes.add(str(row[0]).strip())
    wb.close()
    return codes, ver


# 원장 열 ← 추출 필드. ID·자동계산 열은 **의도적으로 제외**한다(수식이 채운다).
SHEET_MAP = {
    "02_돌발AS접수": {
        "프로젝트NO": "프로젝트NO", "캠프명": "캠프명", "접수일자": "신청일자",
        "담당기사": "담당기사", "방문예정일": "예정일", "작업완료일": "완료일",
        "신청내용": "신청내용", "접수경로": "_경로", "비고": "_비고",
    },
    "04_정기점검": {
        "프로젝트NO": "프로젝트NO", "캠프명": "캠프명", "점검예정일": "예정일",
        "담당기사": "담당기사", "실제점검일": "완료일", "점검내용": "_점검내용",
        "비고": "_비고",
    },
}
DATE_COLS = {"접수일자", "방문예정일", "작업완료일", "점검예정일", "실제점검일"}


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _key(text):
    """내용 대조용 키 — 띄어쓰기·기호 차이를 무시한 앞부분."""
    return re.sub(r"[^0-9A-Za-z가-힣]", "", str(text or ""))[:18]


DATE_TOL_DAYS = 3          # kakao_reconcile 의 date_tolerance_days 와 같은 값


def _near(day_str, days):
    """원장 일자가 공지 일자들 중 하나와 ±3일 안이면 같은 건으로 본다."""
    if not day_str or not days:
        return False
    try:
        y, m, d = (int(x) for x in day_str[:10].split("-"))
        led = date(y, m, d)
    except ValueError:
        return False
    return any(abs((led - x).days) <= DATE_TOL_DAYS for x in days)


def sheet_state(master, sheet):
    """(첫 빈 행, 열이름→열번호, 용량행, 번호없는 기존행 목록)

    ★ 2026-07-28 실사고: 첫 빈 행을 **프로젝트NO 열만 보고** 정했더니,
      번호 없이 내용만 적혀 있던 547행(M_순천1·김필우)을 빈 행으로 보고
      전혀 다른 건(UJ2601347)의 번호를 얹었다. 빈 행 판정은 **행 전체**로 한다.
      ID(A열)는 수식이라 판정에서 뺀다.
    용량은 하드코딩하지 않는다 — expand_rows.py 로 늘리면 바로 반영돼야 한다."""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    cols = {str(h).strip(): i + 1 for i, h in enumerate(hdr) if h}
    jno = cols["프로젝트NO"] - 1
    date_col = cols.get("접수일자") or cols.get("점검예정일")
    text_col = cols.get("신청내용") or cols.get("점검내용")
    camp_col = cols.get("캠프명")
    last, orphans = 4, []
    for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
        rn = 5 + i
        if not any(v not in (None, "") for v in row[1:]):
            continue                                   # 완전한 빈 행
        last = rn
        if row[jno] in (None, ""):                     # 번호만 없는 행 = 사람이 먼저 적어 둔 행
            d = row[date_col - 1] if date_col else None
            orphans.append({
                "행": rn,
                "일자": d.date().isoformat() if hasattr(d, "date") else (str(d)[:10] if d else ""),
                "캠프명": str(row[camp_col - 1] or "") if camp_col else "",
                "내용키": _key(row[text_col - 1] if text_col else ""),
            })
    cap = ws.max_row
    wb.close()
    return last + 1, cols, cap, orphans


def build_queue(rows, master):
    """등록 대상 → ledger_writer 큐 셀 목록. 02·04 로 유형이 확정된 건만 만든다."""
    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        if r["시트"] in SHEET_MAP:
            groups[r["시트"]].append(r)
    queue, plan, held = [], [], []
    for sheet, recs in groups.items():
        start, cols, cap, orphans = sheet_state(master, sheet)
        room = cap - start + 1
        if len(recs) > room:
            held.append("%s: %d건 필요 / 여유 %d행 — 전량 보류 "
                        "(python expand_rows.py --sheet %s --add %d --apply 후 재실행)"
                        % (sheet, len(recs), room, sheet, len(recs) - room + 10))
            continue
        # 번호 없이 내용만 적혀 있던 행에 같은 건이 있으면 **새 행을 만들지 않고** 그 행을 채운다.
        # 새 행을 만들면 같은 작업이 두 행이 된다. 일자+내용이 **둘 다** 맞을 때만 인정한다.
        used, next_row = set(), start
        for r in recs:
            key = _key(r["신청내용"] or r["작업내용"])
            # 공지의 '신청일자'와 원장의 '접수일자'는 며칠 어긋난다(공지는 다음 날 올라온다).
            # 그래서 대조기와 같은 ±3일을 허용한다. 내용키 18자 완전일치가 주된 근거다.
            days = {d for d in (r["신청일자"], r["예정일"], r["완료일"]) if d}
            hit = next((o for o in orphans
                        if o["행"] not in used and key and o["내용키"] == key
                        and _near(o["일자"], days)), None)
            if hit:
                used.add(hit["행"])
                r["_행"] = hit["행"]
                print("  · %s → 기존 %d행에 번호 채움 (%s)" % (r["프로젝트NO"], hit["행"], hit["캠프명"][:18]))
            else:
                r["_행"] = next_row
                next_row += 1
        for r in recs:
            rn = r["_행"]
            note = "카톡 자동등록(%s)" % r["카톡일시"]
            if r.get("담당원문"):
                note += " / 담당 원문: %s" % r["담당원문"]
            if r.get("설비"):
                note += " / 설비: %s" % r["설비"]
            src = dict(r, _비고=note[:250], _경로="카톡",
                       _점검내용=(r["작업내용"] or r["신청내용"])[:200])
            for led_col, field in SHEET_MAP[sheet].items():
                if led_col not in cols:
                    continue
                v = src.get(field)
                if v in (None, ""):
                    continue                    # 원문에 없는 값은 채우지 않는다
                queue.append({
                    "sheet": sheet, "cell": "%s%d" % (col_letter(cols[led_col]), rn),
                    "key": "%s%d" % (col_letter(cols[led_col]), rn), "key_col": "-",
                    "col": led_col, "value": str(v),
                    "vtype": "date" if led_col in DATE_COLS else "text",
                    "evidence": "카톡 %s %s" % (r["프로젝트NO"], r["출처파일"]),
                    "only_if_empty": True,
                })
            plan.append({"시트": sheet, "행": rn, "프로젝트NO": r["프로젝트NO"],
                         "캠프명": r["캠프명"], "기사": r["담당기사"] or "",
                         "예정일": r["예정일"] or "", "신청일자": r["신청일자"] or ""})
        rowset = sorted(r["_행"] for r in recs)
        print("%s: %d건 → %s행 (첫 빈 행 %d · 여유 %d행)"
              % (sheet, len(recs), ",".join(str(x) for x in rowset), start, room))
    return queue, plan, held


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--code", help="특정 프로젝트NO 상세 출력")
    ap.add_argument("--new", action="store_true", help="최신 원장에 없는 건만")
    ap.add_argument("--codes", help="쉼표로 구분한 프로젝트NO 목록만 대상으로")
    ap.add_argument("--days", type=int, help="최근 N일 안에 올라온 글만 (매일 자동 실행용)")
    ap.add_argument("--queue", action="store_true", help="ledger_writer 큐에 적재(원장 반영은 별도)")
    ap.add_argument("--json", help="결과를 JSON 으로 저장할 경로")
    args = ap.parse_args()

    rows = extract()
    ver = None
    if args.days:
        # 매일 자동으로 도는 자리에서 과거분까지 한꺼번에 올리면 안 된다 —
        # 오래된 글은 파싱 품질을 사람이 한 번 보고 넣어야 한다(--new 로 따로).
        from datetime import timedelta
        cut = (date.today() - timedelta(days=args.days)).isoformat()
        before = len(rows)
        rows = [r for r in rows if r["카톡일시"][:10] >= cut]
        print("최근 %d일(%s~) %d건 / 전체 %d건" % (args.days, cut, len(rows), before))
    if args.codes:
        want = {c.strip().upper() for c in args.codes.split(",") if c.strip()}
        missing = want - {r["프로젝트NO"] for r in rows}
        rows = [r for r in rows if r["프로젝트NO"] in want]
        if missing:
            print("★ 카톡 원문에 없음:", ", ".join(sorted(missing)))
    if args.new or args.code or args.queue:
        known, ver = ledger_codes()
        if args.new or args.queue:
            dup = [r["프로젝트NO"] for r in rows if r["프로젝트NO"] in known]
            rows = [r for r in rows if r["프로젝트NO"] not in known]
            if dup and args.queue:
                print("★ 이미 원장에 있어 제외:", ", ".join(dup))

    if args.code:
        for r in rows:
            if r["프로젝트NO"] == args.code.strip().upper():
                for k, v in r.items():
                    print("%-12s %s" % (k, v))
                return 0
        print("없음:", args.code)
        return 1

    by_sheet = {}
    for r in rows:
        by_sheet[r["시트"] or "(유형미상)"] = by_sheet.get(r["시트"] or "(유형미상)", 0) + 1
    print("카톡 추출 %d건%s — %s" % (
        len(rows), (" (원장 v%d 미등록만)" % ver) if args.new and ver else "",
        " · ".join("%s %d" % kv for kv in sorted(by_sheet.items()))))
    print("%-11s %-13s %-4s %-16s %-12s %-10s %s" % (
        "프로젝트NO", "시트", "상태", "캠프명", "기사", "예정일", "신청내용"))
    for r in rows:
        print("%-11s %-13s %-4s %-16s %-12s %-10s %s" % (
            r["프로젝트NO"], (r["시트"] or "?")[:13], r["상태"] or "?", r["캠프명"][:16],
            r["담당기사"][:12] or "-", r["예정일"] or r["완료일"] or "-",
            (r["신청내용"] or r["작업내용"])[:55]))

    if args.json:
        os.makedirs(os.path.dirname(os.path.abspath(args.json)) or ".", exist_ok=True)
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2, default=str)
        print("저장:", args.json)

    if args.queue:
        from ecount_reconcile import load_config, resolve_master
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
        queue, plan, held = build_queue(rows, master)
        unmapped = [r for r in rows if r["시트"] not in SHEET_MAP]
        for r in unmapped:
            print("  [보류] %s %s — 대상 시트 없음(%s)" % (
                r["프로젝트NO"], r["캠프명"][:20], r["시트"] or "유형미상"))
        for h in held:
            print("  [보류]", h)
        print("\n등록 대상 %d건 / 셀 %d개" % (len(plan), len(queue)))
        if queue:
            from ledger_writer import queue_add
            print("큐 적재:", queue_add(queue), "개 셀 → python ledger_writer.py --apply 로 원장 반영")
    return 0


if __name__ == "__main__":
    sys.exit(main())
