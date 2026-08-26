# -*- coding: utf-8 -*-
"""
ecount_reconcile.py — 관리대장 ↔ 이카운트(ECOUNT) 대조기
==========================================================
관리대장(원장)에 기록된 "예상"과 이카운트에 실제 등록된 전표를 3개 문서층에서 대조:

  ① 거래명세서/판매전표  : 원장 06(정산) O·P·Q 금액  ↔  이카운트 판매전표
  ② 전자세금계산서        : 원장 15 실제발행일·발행금액·승인번호  ↔  이카운트 세금계산서
  ③ 수금/입금             : 원장 16 입금일·입금액  ↔  이카운트 수금

판정: 일치 · 금액불일치 · 이카운트미등록(원장에만) · 원장누락(이카운트에만)

원칙:
- 관리대장(마스터 xlsx)은 read-only 로만 연다. (openpyxl save 는 차트·도형·검증을 손상시키므로 절대 저장하지 않음)
- 결과는 별도 리포트로만 출력: reports/ 폴더에 CSV·Markdown·독립 xlsx.
- 이카운트 미연결(설정 미완/ IP 미등록)이어도 "원장 기준 대조 준비표"를 먼저 만들어 즉시 활용 가능.

실행:
    python ecount_reconcile.py            # 이카운트 연결 시도 후 대조, 실패하면 원장측만
    python ecount_reconcile.py --offline  # 이카운트 호출 없이 원장측 준비표만
    python ecount_reconcile.py --selftest # 설정/인증만 점검
"""
import sys, os, csv, json, time, threading
from datetime import datetime, date, timedelta
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
from ecount_client import EcountClient, load_config, days_until_expiry, EcountError

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.path.join(BASE_DIR, "reports")

# ★ '쿠팡에 청구하지 않는 상태'를 **한 곳에서** 정한다 (2026-08-09).
#   딱지를 '무상/보험' 하나에서 '무상'·'보험'으로 가르는 순간, 그 문자열을 손으로 적어
#   두었던 곳들(업로드·조치목록·화면 색)이 **아무 오류 없이** 안 걸리게 된다 —
#   무상 건이 갑자기 '조치 필요'로 쏟아지는 식이다. 그래서 읽는 자리를 하나로 모은다.
#   '무상/보험' 은 예전 값이라 남겨 둔다(리포트·DB 에 이미 적힌 것이 있다).
NON_BILLABLE = ("무상", "보험", "무상/보험")

# 비용구분을 아직 안 적은 것. **무상이 아니다** — 청구할지 말지를 아직 모르는 상태다.
UNKNOWN_COST = "비용구분 미입력"


def is_non_billable(status):
    """이 정산 상태가 '쿠팡에 청구하지 않음'인가. 미입력은 여기 들어오지 않는다."""
    return str(status or "").strip() in NON_BILLABLE

HDR_ROW = 4          # 관리대장 공통 상수 — 절대 변경 금지(수천 수식이 의존)
FIRST_DATA = 5


# ───────────────────────── 원장 읽기 ─────────────────────────
def _num(v):
    if v in (None, "", "-"):
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _money(v):
    """금액 표시는 모든 보고서에서 소수점 없이 천 단위 쉼표로 통일한다."""
    if v in (None, "", "-"):
        return ""
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return str(v)


def _d(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return "" if v is None else str(v)


def _latest_in(folder):
    """폴더에서 v번호가 가장 큰 관리대장. 구 버전 정리의 기준점으로만 쓴다."""
    import glob, re as _re
    best, bv = None, -1
    for p in glob.glob(os.path.join(folder, "쿠팡_통합업무_일일보고_관리대장_v*.xlsx")):
        m = _re.search(r"_v(\d+)\.xlsx$", p)
        if m and "~$" not in p and int(m.group(1)) > bv:
            best, bv = p, int(m.group(1))
    return best or folder


_MASTER_CACHE = {}          # xlsx_path -> (잰시각, 결과경로)
_MASTER_TTL = 30.0          # 초


def resolve_master(xlsx_path):
    """최신 관리대장을 찾되 **30초 안에는 다시 찾지 않는다**(2026-07-31 속도 개선).

    이 함수는 네트워크 드라이브에서 glob 을 돌리고 autoprune 까지 부른다. 그런데
    API 한 번에 7번씩 불리고 있었다(`real_works` 측정). 어느 파일인지만 고르는 일이라
    30초 캐시로 충분하다 — 내용이 바뀌었는지는 read_ledger 가 mtime 으로 따로 본다.
    새 프로세스는 항상 차갑게 시작하므로 낡은 경로를 오래 붙들 일은 없다.
    """
    now = time.monotonic()
    hit = _MASTER_CACHE.get(xlsx_path)
    if hit is not None and now - hit[0] < _MASTER_TTL:
        return hit[1]
    result = _resolve_master_uncached(xlsx_path)
    _MASTER_CACHE[xlsx_path] = (now, result)
    return result


def _resolve_master_uncached(xlsx_path):
    """설정된 경로가 없으면 같은 폴더에서 최신 v번호 파일을 자동 탐지한다
    (관리대장은 v19→v20→v21…로 계속 버전업되며 구버전은 OLD로 이동됨)."""
    import glob, re as _re
    # 구 버전을 OLD 로 접는다 — 사용자 지시(2026-07-28) "말 안 해도 들어가게".
    # 저장하는 쪽이 11군데라 여기(찾는 쪽) 한 곳에만 건다. 실패해도 본 작업은 그대로 간다.
    # autoprune 은 관리대장 이름이 아니면 스스로 아무것도 하지 않는다(합성검증 파일 보호).
    try:
        from ledger_versions import autoprune
        autoprune(xlsx_path if os.path.exists(xlsx_path)
                  else _latest_in(os.path.dirname(xlsx_path)))
    except Exception:
        pass
    if os.path.exists(xlsx_path):
        return xlsx_path
    folder = os.path.dirname(xlsx_path)
    cands = glob.glob(os.path.join(folder, "쿠팡_통합업무_일일보고_관리대장_v*.xlsx"))
    def ver(p):
        m = _re.search(r"_v(\d+)\.xlsx$", p)
        return int(m.group(1)) if m else -1
    cands = [c for c in cands if ver(c) >= 0 and "~$" not in c]
    if not cands:
        # ★ **'관리대장이 없다' 와 'Z: 에 못 닿았다' 는 다른 사실이다**([169]·[289]).
        #   예전 문구는 언제나 `(폴더에 v*.xlsx 없음)` 이라 **사람을 관리대장을
        #   찾으러 보냈다** — 2026-08-26 실측: 그날 자율복구 큐에서 이 문구를 단
        #   실패가 10건인데 진짜 원인은 Z: 가 안 붙은 것이었다(같은 회차의
        #   `원본 색인 갱신` 이 "원본 폴더에 **하나도 닿지 못했습니다**" 라 적었다).
        #   **틀린 지목은 못 잡는 것보다 나쁘다**([172]).
        # ★ **여기서 다시 물어보지 않는다** — 못 닿는 Z: 는 `isdir` 한 번에
        #   **43~156초**가 걸린다(2026-08-26 실측 · 평소 0.15초). 이 문구를 다는
        #   단계가 한 회차에 열 곳이라 그 확인만으로 회차가 몇십 분 는다
        #   ([168] — 비싼 탐색을 실패 경로에 더하지 않는다). **모르면 모른다고
        #   적는다**([169]) — 둘을 나란히 놓고 **싼 확인을 먼저** 가리킨다.
        # ★ 앞머리(`관리대장을 찾을 수 없음:`)는 **안 바꾼다** —
        #   `autopilot.classify_failure` 가 그 글자로 `resource` 를 가른다.
        #   낱말이 어긋나면 한 건도 안 걸리면서 오류도 안 난다([165]).
        # ★ 괄호 안의 `v*.xlsx` 도 **안 뺀다**(2026-08-26 실측) — 빼면
        #   `autopilot._RES_PATH_RE` 의 후보가 1개 → **0개**가 되어 `resource_back`
        #   이 영영 `None`(모름)이 되고 [424] 의 완화가 통째로 죽는다(지나간 자원
        #   실패가 경보에 계속 남는다). 같은 이유로 여기에 `Z:` 를 다시 쓰지
        #   않는다 — 슬래시가 붙으면 후보가 **둘**이 되어 결과가 같다.
        raise FileNotFoundError(
            f"관리대장을 찾을 수 없음: {xlsx_path}"
            " (v*.xlsx 를 못 찾았다 — 그 폴더에 못 닿았거나 폴더에 그 파일이 없다."
            " 네트워크 드라이브 연결부터 확인한다)")
    best = max(cands, key=ver)
    print(f"i 관리대장 최신본 자동 탐지: {os.path.basename(best)}")
    return best


_MASTER_BYTES = {}          # abspath -> ((mtime_ns, size), bytes)


def master_stream(xlsx_path):
    """관리대장을 **네트워크가 아니라 메모리에서** 열게 해 준다(2026-07-31 속도 개선).

    앱은 한 화면을 그리는 동안 `openpyxl.load_workbook(master, ...)` 를 9군데에서
    각각 부른다. 그때마다 **Z: 네트워크 드라이브에서 수십 MB 를 다시 끌어온다** — 느린
    것의 대부분이 파싱이 아니라 이 전송이다. 한 번 읽어 두고 BytesIO 로 건네면 파싱만 남는다.

    · 무효화는 mtime+크기. vN+1 이 생기면 키가 달라져 자동으로 다시 읽는다.
    · 새 버전을 읽을 때 캐시를 **비운다** — 안 비우면 버전이 쌓일수록 메모리를 계속 먹는다.
    · 파일을 못 재면 원래 경로를 그대로 돌려준다(호출부는 경로든 스트림이든 받는다).
    """
    import io
    key = os.path.abspath(xlsx_path)
    try:
        st = os.stat(key)
        sig = (st.st_mtime_ns, st.st_size)
    except OSError:
        return xlsx_path
    hit = _MASTER_BYTES.get(key)
    if hit is None or hit[0] != sig:
        with open(key, "rb") as f:
            data = f.read()
        _MASTER_BYTES.clear()
        _MASTER_BYTES[key] = (sig, data)
        hit = _MASTER_BYTES[key]
    return io.BytesIO(hit[1])


# ── 관리대장을 **한 번만 파싱**한다 (2026-08-08, worksplit #17) ──────────────
# `master_stream` 이 네트워크 전송을 없앤 뒤에도 남은 비용은 **파싱**이었다.
# 실측(v556, 30시트): `load_workbook` 만 2.59초 · 시트 하나 읽기 0.35초 ·
# 전 시트 읽기 5.26초. 그런데 앱은 한 화면을 그리는 동안 이 load 를 **8군데**에서
# 각각 부른다 — 화면 한 장에 파싱만 20초가 넘게 든다.
#
# ★ **워크북 개체를 공유하면 안 된다.** `read_only=True` 워크북은 이터레이터 상태를
#   갖고 스레드 안전하지 않다(두 요청이 같은 시트를 동시에 읽으면 행이 섞인다).
#   그래서 캐시에 담는 것은 워크북이 아니라 **뽑아낸 행(튜플)** 이다. 값만 담으므로
#   여러 스레드가 동시에 읽어도 안전하다.
# ★ 처음 한 번은 전 시트를 읽는다(약 7.9초). 대신 그 뒤로는 파일이 바뀌기 전까지
#   0초다. 관리대장은 하루 두 번(11:00·15:00) 바뀌므로 하루에 두 번 내는 값이다.
_MASTER_SHEETS = {"key": None, "sig": None, "sheets": {}, "stat_at": 0.0}
_SHEET_LOCK = threading.Lock()


class _CachedSheet:
    """openpyxl 시트에서 `values_only` 로 읽을 때와 같게 굴러가는 얇은 껍데기."""
    __slots__ = ("title", "_rows")

    def __init__(self, title, rows):
        self.title, self._rows = title, rows

    @property
    def max_row(self):
        return len(self._rows)

    @property
    def max_column(self):
        return max((len(r) for r in self._rows), default=0)

    def iter_rows(self, min_row=1, max_row=None, min_col=None, max_col=None,
                  values_only=True):
        # ★ 셀 개체는 주지 않는다. 필요해지면 캐시가 아니라 openpyxl 로 열어야 한다 —
        #   조용히 다른 것을 돌려주면 호출부가 왜 안 되는지 모른다.
        if not values_only:
            raise ValueError("캐시 시트는 values_only 로만 읽는다(셀 개체 없음)")
        lo = max(1, int(min_row or 1))
        hi = len(self._rows) if max_row is None else min(len(self._rows), int(max_row))
        for i in range(lo - 1, hi):
            row = self._rows[i]
            if min_col or max_col:
                row = row[(int(min_col or 1) - 1):(int(max_col) if max_col else len(row))]
            yield row


class _CachedBook:
    __slots__ = ("_sheets", "sheetnames")

    def __init__(self, sheets):
        self._sheets = sheets
        self.sheetnames = list(sheets)

    def __contains__(self, name):
        return name in self._sheets

    def __getitem__(self, name):
        return _CachedSheet(name, self._sheets[name])

    def close(self):
        pass                      # 캐시라 닫을 것이 없다(호출부를 안 고치려고 둔다)


def master_book(xlsx_path):
    """관리대장 워크북 — 같은 파일이면 파싱하지 않는다.

    `openpyxl.load_workbook(master_stream(p), read_only=True, data_only=True)` 자리에
    그대로 넣는다. `sheetnames` · `wb[시트]` · `iter_rows(values_only=True)` ·
    `max_row` · `close()` 가 같게 동작한다."""
    import openpyxl
    key = os.path.abspath(xlsx_path)
    # ★ Z: 는 네트워크 드라이브라 `os.stat` 한 번이 실측 0.35초다. 한 화면을 그리는
    #   동안 8번 물으면 그것만 2.8초 — 파싱을 없애 놓고 stat 에서 잃는다.
    #   그래서 몇 초 동안은 방금 잰 값을 그대로 쓴다. 관리대장은 11:00·15:00 회차에만
    #   바뀌므로 이 창 안에서 파일이 바뀔 일은 없다.
    now = time.time()
    if _MASTER_SHEETS["key"] == key and now - _MASTER_SHEETS.get("stat_at", 0) < 2.0:
        sig = _MASTER_SHEETS["sig"]
    else:
        try:
            st = os.stat(key)
            sig = (st.st_mtime_ns, st.st_size)
        except OSError:
            sig = None
        if sig is not None:
            _MASTER_SHEETS["stat_at"] = now
    with _SHEET_LOCK:
        if sig is not None and _MASTER_SHEETS["key"] == key and _MASTER_SHEETS["sig"] == sig:
            return _CachedBook(_MASTER_SHEETS["sheets"])
        wb = openpyxl.load_workbook(master_stream(xlsx_path), read_only=True,
                                    data_only=True)
        try:
            sheets = {n: [tuple(r) for r in wb[n].iter_rows(values_only=True)]
                      for n in wb.sheetnames}
        finally:
            wb.close()
        if sig is not None:
            # 새 버전이면 옛 시트를 **버린다** — 안 버리면 vN 이 쌓일수록 메모리를 먹는다.
            _MASTER_SHEETS.update({"key": key, "sig": sig, "sheets": sheets})
        return _CachedBook(sheets)


_LEDGER_CACHE = {}          # abspath -> ((mtime_ns, size), 원장dict)


def read_ledger(xlsx_path):
    """원장을 읽되, **파일이 그대로면 다시 열지 않는다**(2026-07-31 속도 개선).

    관리대장은 **네트워크 드라이브(Z:)** 에 있고 한 번 여는 데 ~1.0초가 든다. 그런데
    캐시가 없어서 API 한 번에 여러 번 다시 열고 있었다 — `real_works` 는 한 호출에
    7번 열었다(측정: settlements 2.5초, works 3.5초).

    무효화는 **파일의 mtime+크기**로 한다. 시간(TTL)으로 하면 vN+1 이 새로 생겼는데도
    남은 시간 동안 옛 숫자를 보여 주게 된다 — 원장은 그러면 안 된다. 파일이 바뀌는
    순간 키가 달라져 자동으로 다시 읽는다.

    돌려줄 때 **사본**을 준다. 호출자가 받은 dict 를 고쳐도 캐시가 오염되지 않아야 한다.
    """
    key = os.path.abspath(xlsx_path)
    try:
        st = os.stat(key)
        sig = (st.st_mtime_ns, st.st_size)
    except OSError:
        return _read_ledger_uncached(xlsx_path)      # 못 재면 그냥 읽는다
    hit = _LEDGER_CACHE.get(key)
    if hit is not None and hit[0] == sig:
        return _copy_ledger(hit[1])
    data = _read_ledger_uncached(xlsx_path)
    _LEDGER_CACHE[key] = (sig, data)
    return _copy_ledger(data)


def _copy_ledger(d):
    """원장은 {정산ID: {열: 값}} 2단이고 값은 전부 단순형이라 2단 복사면 충분하다.
       copy.deepcopy 는 datetime 까지 새로 만들어 느리다."""
    return {k: dict(v) for k, v in d.items()}


def _read_ledger_uncached(xlsx_path):
    """정산ID 기준으로 원장 예상값(3개 문서층)을 모은다."""
    import openpyxl
    xlsx_path = resolve_master(xlsx_path)
    wb = master_book(xlsx_path)

    def col_index(ws, names):
        row = next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))
        idx = {}
        for i, h in enumerate(row):
            if h is not None:
                idx[str(h).strip()] = i
        return {n: idx.get(n) for n in names}

    recs = {}   # 정산ID -> dict

    # 06 거래서류청구수금 : 기준 원장
    ws = wb["06_거래서류청구수금"]
    c = col_index(ws, ["정산ID", "업무구분", "원천업무ID", "프로젝트NO", "캠프명", "작업완료일",
                       "비용구분", "실제작업공급가액", "실제작업부가세", "실제작업합계",
                       "거래명세서번호", "거래명세서발행일", "거래명세서공급가액", "거래명세서합계",
                       "세금계산서발행일", "세금계산서합계", "입금일", "입금액",
                       "청구일", "지급예정일", "PO필요여부", "PO번호", "PO발행일",
                       # ★ 청구상태(AH)는 **사람이 손으로 채우는 단계 딱지**다. 여기 없으면
                       #   `r.get("원장_청구상태")` 가 750행 전부 None 을 돌려준다 —
                       #   비어 있는 게 아니라 **읽지도 않은 것**인데 세어 보면 '빈칸 750'
                       #   으로 나온다(2026-08-08 실사고: 그 숫자로 '완료 반영 대상 286건'
                       #   이라는 결론을 낼 참이었다. 실제로 채워진 행이 63개 있었다).
                       #   없는 열을 물어보면 col_index 가 None 을 주고 조용히 넘어가므로
                       #   **틀린 게 아니라 안 보이는** 종류의 사고가 된다.
                       "청구상태"])
    for row in ws.iter_rows(min_row=FIRST_DATA, values_only=True):
        sid = row[c["정산ID"]] if c["정산ID"] is not None else None
        if not sid:
            continue
        recs[str(sid)] = {
            "정산ID": str(sid),
            "업무구분": _d(row[c["업무구분"]]),
            "원천업무ID": _d(row[c["원천업무ID"]]),
            "프로젝트NO": _d(row[c["프로젝트NO"]]),
            "캠프명": _d(row[c["캠프명"]]),
            "작업완료일": _d(row[c["작업완료일"]]),
            "비용구분": _d(row[c["비용구분"]]),
            "원장_공급가액": _num(row[c["실제작업공급가액"]]),
            # 부가세는 **원장에 적힌 값**을 그대로 쓴다. 공급가액×10%로 계산하면 반올림 때문에
            # 원장·계산서와 1원씩 어긋날 수 있고, 화면 숫자가 서류와 달라지면 신뢰를 잃는다.
            "원장_부가세": _num(row[c["실제작업부가세"]]) if "실제작업부가세" in c else None,
            "원장_합계": _num(row[c["실제작업합계"]]),
            "원장_거래명세서번호": _d(row[c["거래명세서번호"]]),
            "원장_거래명세서발행일": _d(row[c["거래명세서발행일"]]),
            "원장_거래명세서합계": _num(row[c["거래명세서합계"]]),
            "원장_세금계산서발행일": _d(row[c["세금계산서발행일"]]),
            "원장_세금계산서합계": _num(row[c["세금계산서합계"]]),
            "원장_입금일": _d(row[c["입금일"]]),
            "원장_입금액": _num(row[c["입금액"]]),
            "원장_청구일": _d(row[c["청구일"]]) if c["청구일"] is not None else "",
            "원장_청구상태": _d(row[c["청구상태"]]) if c["청구상태"] is not None else "",
            "원장_지급예정일": _d(row[c["지급예정일"]]) if c["지급예정일"] is not None else "",
            "원장_PO필요여부": _d(row[c["PO필요여부"]]) if c["PO필요여부"] is not None else "",
            "원장_PO번호": _d(row[c["PO번호"]]) if c["PO번호"] is not None else "",
            "원장_PO발행일": _d(row[c["PO발행일"]]) if c["PO발행일"] is not None else "",
        }

    # 15 세금계산서관리 : 승인번호·실제발행일 보강
    ws = wb["15_세금계산서관리"]
    c = col_index(ws, ["정산ID", "실제발행일", "발행금액", "승인번호"])
    for row in ws.iter_rows(min_row=FIRST_DATA, values_only=True):
        sid = row[c["정산ID"]] if c["정산ID"] is not None else None
        if sid and str(sid) in recs:
            r = recs[str(sid)]
            r["원장_세금계산서실제발행일"] = _d(row[c["실제발행일"]])
            r["원장_세금계산서승인번호"] = _d(row[c["승인번호"]])
            if _num(row[c["발행금액"]]):
                r["원장_세금계산서합계"] = _num(row[c["발행금액"]]) or r.get("원장_세금계산서합계")

    # 16 입금수금관리 : 입금 보강
    ws = wb["16_입금수금관리"]
    c = col_index(ws, ["정산ID", "입금일", "입금액", "미수금액"])
    for row in ws.iter_rows(min_row=FIRST_DATA, values_only=True):
        sid = row[c["정산ID"]] if c["정산ID"] is not None else None
        if sid and str(sid) in recs:
            r = recs[str(sid)]
            if _d(row[c["입금일"]]):
                r["원장_입금일"] = _d(row[c["입금일"]])
            if _num(row[c["입금액"]]) is not None:
                r["원장_입금액"] = _num(row[c["입금액"]])
            r["원장_미수금액"] = _num(row[c["미수금액"]])

    wb.close()
    return recs


# ───────────────────────── 이카운트 값 정규화 ─────────────────────────
# 이카운트 조회 응답의 필드명은 계정/버전에 따라 다르므로 후보군에서 유연 추출
CAND = {
    "금액":   ["SUPPLY_AMT", "SUPPLY_AMT_KRW", "AMT", "TOTAL_AMT", "PRICE", "SUP_AMT", "TTL_AMT"],
    "일자":   ["IO_DATE", "BASE_DATE", "IO_DATE_KRW", "DATE", "WR_DATE", "REG_DATE"],
    "적요":   ["REMARKS", "E_TEXT", "U_MEMO", "WH_DES", "REMARKS_WIN", "PRJ_CD"],
    "거래처": ["CUST_DES", "CUST", "CUST_NM", "EMP_NM"],
    "번호":   ["IO_NO", "DOC_NO", "TAX_NO", "IV_NO", "SLIP_NO", "APPROVE_NO", "NTS_CONFIRM_NUM"],
}


def has_statement(r):
    """거래명세서가 실제로 있는가 — 번호 하나로 판단하지 않는다.

    billing_fill.py 가 ERP 판매조회를 근거로 채운 건들은 **발행일·금액은 있는데 번호가
    없다**(629건). 번호만 보면 이 전부가 '미청구(전표 없음)'로 잡히는데, 명세서는 실제로
    나갔고 번호만 원장에 안 옮겨진 것이라 거짓 경보다. 발행일+금액이 함께 있으면 있는 것으로 본다.
    """
    if str(r.get("원장_거래명세서번호") or "").strip():
        return True
    return bool(r.get("원장_거래명세서발행일")) and bool(r.get("원장_거래명세서합계"))


def supply_from_statement(total):
    """거래명세서합계(**부가세 포함**)에서 공급가액을 되돌린다.

    사용자 지시(2026-08-08): "부가세 포함 금액 ÷1.1 자동 환산". 실측 35건이 전부
    10원 단위로 딱 떨어졌다(506,000→460,000 · 621,500→565,000 · 297,000→270,000).

    ★ **깨끗하게 떨어질 때만** 돌려준다. 되돌려 곱한 값이 원본과 다르거나 10원 단위가
      아니면 `None` 이다 — 그런 건은 부가세 포함이 아니거나 다른 사연이 있는 것이라
      '금액 재계산 대기'로 남겨 사람이 본다. 환산이 안 되는 건까지 억지로 숫자를
      만들면, 틀린 금액이 화면에서 '확정'처럼 보인다(조용한 사고).
    """
    try:
        t = int(total or 0)
    except (TypeError, ValueError):
        return None
    if t <= 0:
        return None
    supply = round(t / 1.1)
    if supply % 10 or round(supply * 1.1) != t:
        return None
    return supply


_ERP_PROGRESS = {"sig": None, "map": {}, "statuses": {}, "checked": 0.0}

# ★ 원본이 바뀌었는지 **다시 확인하기까지** 기다리는 시간 (2026-08-08 실사고).
#   판매조회 파일을 찾는 일은 Z:(SMB) 를 **재귀로 훑는** glob 이다. 그런데 그 glob 이
#   캐시 검사보다 **앞에** 있어서, `settle_status` 를 750행 돌리면 Z: 를 750번 훑었다.
#   증상은 "왜 이렇게 느리지" 가 아니라 **아무 답도 안 오는 것**이었다 — 같은 집계가
#   600초 제한에 두 번 걸려 죽었고, 세 번째는 45분을 넘겨도 안 끝났다. 느린 것과
#   멈춘 것은 구별이 안 된다.
#   30초면 한 번의 집계 안에서는 한 번만 훑고, 회차와 회차 사이에는 새 파일을 본다.
_ERP_SIG_TTL = 30.0
# ERP 판매조회의 진행상태 중 '계산서가 이미 나갔다'는 뜻인 값들
_ERP_ISSUED = ("6.세금계산서발행", "7.수금완료")


def _collapse_erp_progress(statuses):
    """같은 프로젝트의 ERP 전표가 여러 개면 가장 보수적인 상태로 접는다.

    프로젝트번호가 같아도 별도 전표가 추가될 수 있다. 그중 하나만 수금완료인데 다른
    전표가 미확인/확인 단계라면 프로젝트 전체를 완료로 올리면 안 된다. 모든 전표가
    수금완료일 때만 ``7.수금완료``를 돌려준다.
    """
    values = {str(value).strip() for value in statuses or [] if str(value).strip()}
    if not values:
        return ""
    if values == {"7.수금완료"}:
        return "7.수금완료"
    if values.issubset(set(_ERP_ISSUED)):
        return "6.세금계산서발행"
    if len(values) == 1:
        return next(iter(values))
    return "혼재(" + " / ".join(sorted(values)) + ")"


def erp_progress():
    """ERP 판매조회에서 {프로젝트NO: 진행상태} 를 읽어 온다(없으면 빈 dict).

    ★ 왜 필요한가 — **이중발행을 막기 위해서다.**
      원장에 세금계산서 발행일이 안 적혔다는 이유만으로 '세금계산서 미발행'이라 부르면,
      ERP 에서는 **이미 발행되고 수금까지 끝난** 건까지 발행 대상 목록에 올라간다(206건).
      그걸 보고 다시 발행하면 이중발행이다. 원장이 모른다고 해서 안 나간 게 아니다.

    ★ 무엇을 하지 않는가 — **발행일을 지어내지 않는다.**
      판매조회에는 발행일 열이 **없다**(일자·진행상태만). 그래서 여기서 채울 수 있는 것은
      '나갔다/안 나갔다' 뿐이고, 실제 발행일·승인번호는 ERP [매출(세금)계산서현황] 을
      내보내야 들어온다. 원자료에 없는 값을 채우지 않는다(절대규칙 10).
    """
    # 합성검증은 실데이터·실서버 접촉 0이 원칙이다. 이전에는 t36이 여기서 Z: 전체를
    # 재귀 탐색해 10분 넘게 멈췄다. 합성 모드에서는 원본 진행상태가 없는 조건만 시험한다.
    if os.environ.get("CSOS_SYNTHETIC") == "1":
        return {}
    # ★ 원본을 찾는 glob 자체가 비싸다 — 그러니 **캐시 검사보다 앞에 두면 안 된다.**
    #   방금 확인했고 그때 읽은 것이 있으면 그대로 쓴다(자세한 이유는 _ERP_SIG_TTL).
    #   ★ '못 찾음'도 기억한다 — 판매조회가 아예 없을 때 750번 헛되이 훑는 것이
    #     제일 느리다(찾을 것이 없으면 glob 은 폴더를 **끝까지** 다 본다).
    import time as _t
    if _t.monotonic() - _ERP_PROGRESS["checked"] < _ERP_SIG_TTL:
        return _ERP_PROGRESS["map"]
    _ERP_PROGRESS["checked"] = _t.monotonic()
    try:
        import glob as _g
        from source_dirs import ERP_DIR
        cands = _g.glob(os.path.join(ERP_DIR, "**", "판매조회*.xlsx"), recursive=True)
        cands = [c for c in cands if "~$" not in c and "__dup_" not in c]
        if not cands:
            return {}
        path = max(cands, key=lambda p: os.stat(p).st_mtime)
        st = os.stat(path)
        sig = (path, st.st_mtime_ns, st.st_size)
    except Exception:
        return {}
    if _ERP_PROGRESS["sig"] == sig:
        return _ERP_PROGRESS["map"]
    grouped = {}
    try:
        import openpyxl as _ox
        wb = _ox.load_workbook(path, read_only=True, data_only=True)
        ws = wb["판매조회"] if "판매조회" in wb.sheetnames else wb.worksheets[0]
        hdr, ip, ist = None, None, None
        for row in ws.iter_rows(values_only=True):
            if hdr is None:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if "진행상태" in cells:
                    hdr = cells
                    # 열 이름이 '프로젝트코드코드' 로 나온다(ERP 내보내기 그대로) — 부분일치로 잡는다
                    ip = next((i for i, c in enumerate(cells) if "프로젝트" in c), None)
                    ist = cells.index("진행상태")
                continue
            if ip is None or ist is None:
                break
            prj = row[ip] if ip < len(row) else None
            stt = row[ist] if ist < len(row) else None
            if prj and stt:
                grouped.setdefault(str(prj).strip(), []).append(str(stt).strip())
        wb.close()
    except Exception:
        return {}
    out = {project: _collapse_erp_progress(statuses)
           for project, statuses in grouped.items()}
    _ERP_PROGRESS["sig"] = sig
    _ERP_PROGRESS["map"] = out
    _ERP_PROGRESS["statuses"] = {
        project: tuple(sorted(set(statuses))) for project, statuses in grouped.items()
    }
    return out


def erp_progress_statuses():
    """프로젝트별 ERP 원본 진행상태 전체를 반환한다(명시적 충돌 철회용)."""
    # 원본이 잠시 사라지거나 파싱에 실패하면 이전 캐시의 충돌값으로 철회하면 안 된다.
    if not erp_progress():
        return {}
    return _ERP_PROGRESS.get("statuses", {})


def settle_status(r):
    """정산 1건의 상태를 **한 곳에서** 판정한다.

    화면(webapp/app_server.py)과 엑셀(findings_export.py)이 각자 같은 if 사다리를 들고
    있어서 한쪽만 고치면 두 숫자가 어긋났다. 여기로 모은다.

    ★ 금액(06시트 I열 실제작업공급가액) 판정에 조건이 두 개 붙는 이유:
      ① I열은 750행 중 684행이 **수식**이고, 그 수식은
         `IF(LEFT(원천업무ID,3)="AS-", SUMIF('03_현장작업실적'!B:B, 원천업무ID, ...!Q:Q), "")`
         이다. 즉 **돌발AS(AS-)가 아니면 설계상 영원히 빈칸**이다. 정기점검(PM-) 305건을
         '금액 미입력'으로 잡던 것은 오탐이었고, 그 오탐이 if 사다리 맨 위에 있어서
         그 305건의 **진짜 문제(세금계산서 미발행)를 가리고 있었다.**
      ② AS- 인데 비어 있는 건은, 03시트에 실적 행이 아직 안 생겨서 수식이 0인 것이다.
         03시트 행 생성과 재계산은 지금 차단된 작업이다(AGENTS.md 0번). 거래명세서 금액이라는
         **근거가 있으면** '금액 재계산 대기'로 따로 부른다 — 사람이 손으로 넣을 일이 아니라
         재계산이 풀리면 저절로 채워질 건이라, 조치 목록에 섞어 두면 안 된다.
      ★ I열에 값을 직접 써넣지 말 것. 수식을 덮으면 AE(작업대비거래명세서차액)가 항상 0이 되어
        이 시트의 존재 이유인 '작업금액과 명세서금액의 차이 드러내기'가 죽는다.
    """
    # ★ '유상이 아니면 전부 무상/보험' 이었다 — 그래서 **비어 있는 칸까지 무상이라고 불렀다**
    #   (2026-08-09 지시: "무상이면 무상 보험이면 보험 표시"). 실측 750행 중 유상 716 ·
    #   무상 2 · **미확정 4 · 빈칸 26 · '0' 2** 이고 보험은 **0건**이다. 즉 회색 '무상/보험'
    #   딱지가 붙은 34건 중 진짜 무상은 2건뿐이고 나머지 32건은 **비용구분을 아직 안 적은 것**이다.
    #   그 32건에는 금액도 명세서도 있다(UJ2601288 정기점검 611,400원 · 명세서 2026/07/20-6).
    #   무상이라고 부르는 순간 청구 대상에서 조용히 빠진다 — **빈칸이 무상으로 위장하는 자리**다.
    #   빈 값은 눈에 띄지만 **틀린 딱지는 안 띈다**([165] 와 같은 종류).
    kind = str(r.get("비용구분") or "").strip()
    if kind in ("무상", "보험"):
        return kind                     # 원장이 말한 그대로 — 지어내지도 뭉치지도 않는다
    if kind != "유상":
        # 모르는 것을 아는 것처럼 부르지 않는다. '0'·'미확정'·빈칸이 다 여기로 온다.
        return "비용구분 미입력"
    # 실제작업공급가액 수식이 아직 계산되지 않았더라도 ERP가 **같은 프로젝트의 모든
    # 전표를 수금완료**로 확정했다면 정산은 객관적으로 끝났다. 종전에는 아래 금액
    # 재계산 검사가 먼저 실행돼 이런 143건이 완료 DB에서 가려졌다. 진행상태가 섞인
    # 프로젝트는 erp_progress가 '혼재(...)'로 돌려주므로 완료하지 않는다.
    project = str(r.get("프로젝트NO") or "").strip()
    prog = erp_progress().get(project)
    if prog == "7.수금완료":
        return "완료(ERP 수금확인)"
    src = str(r.get("원천업무ID") or "")
    if not r.get("원장_공급가액") and src.startswith("AS-"):
        # ★ 사용자 지시(2026-08-08): 부가세 포함 명세서 금액은 ÷1.1 로 **확정**한다.
        #   깨끗이 떨어지면 금액을 모르는 것이 아니라 아는 것이다 — 재계산을 기다릴
        #   이유가 없다. 안 떨어지는 건만 예전처럼 대기로 남는다.
        if supply_from_statement(r.get("원장_거래명세서합계")) is not None:
            pass
        elif r.get("원장_거래명세서합계"):
            return "금액 재계산 대기"
        else:
            return "금액 미입력"
    if not has_statement(r):
        return "미청구(전표 없음)"
    if not (r.get("원장_세금계산서실제발행일") or r.get("원장_세금계산서발행일")):
        # ★ 원장이 모른다고 해서 안 나간 게 아니다. ERP 가 '발행·수금완료' 라고 말하면
        #   그건 **이미 나간 것**이고, 발행 대상 목록에 올리면 이중발행으로 이어진다.
        #   발행일은 여전히 채우지 않는다 — 판매조회에 발행일 열이 없다(절대규칙 10).
        # ★ 사용자 지시(2026-07-31): 객관적으로 입증되는 건은 **완료 처리**하고,
        #   그 상태는 엑셀 셀 백필이 아니라 **분류·DB로만** 관리한다.
        #   · 7.수금완료  = 발행·수금까지 ERP 가 입증 → 완료
        #   · 6.세금계산서발행 = 발행만 입증 → '미발행'이 아니라 **입금 대기**가 정확하다
        if prog == "6.세금계산서발행":
            return "완료(ERP 발행확인)" if r.get("원장_입금일") else "입금 대기"
        return "세금계산서 미발행"
    if not r.get("원장_입금일"):
        return "입금 대기"
    return "정상"


def pick(row, kind):
    for k in CAND[kind]:
        if isinstance(row, dict) and row.get(k) not in (None, ""):
            return row[k]
    return None


def norm_ecount(rows):
    out = []
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        amt = pick(r, "금액")
        try:
            amt = round(float(str(amt).replace(",", "")), 2) if amt is not None else None
        except ValueError:
            amt = None
        out.append({
            "금액": amt,
            "일자": _d(pick(r, "일자")),
            "적요": _d(pick(r, "적요")),
            "거래처": _d(pick(r, "거래처")),
            "번호": _d(pick(r, "번호")),
            "_raw": r,
        })
    return out


def supply_effective(rec):
    """행의 **유효 공급가액**과 그 출처 — 원장 빈 칸을 ERP·명세서 사다리로 채운다([124]).

    ★ **한 곳에서 구한다**([170] — 같은 값을 두 곳에서 계산하면 언젠가 갈린다).
      매칭 열쇠·판정·불일치 문구·리포트 열·`key_warning` 의 빈금액이 다 이것을 본다.
      다섯 중 하나라도 옛 값을 보면 짝은 늘어나는데 판정이
      *"금액불일치(원장 None / EC …)"* 로 나와 **전보다 나쁜 경보**가 된다.
    ★ `supply_of` 는 못 찾으면 **0** 을 준다 — 0 을 열쇠로 쓰면 금액 0 짜리 전표에
      아무거나 붙는다. 그래서 **0 은 '금액 없음'(None)** 으로 돌려준다.
    ★ **원장 원값은 안 덮는다** — 리포트가 원값과 보정값을 나란히 낸다([169]).
      덮으면 "원장에 정말 뭐라고 적혀 있었나"를 잃는다.
    ⚠ **순환 import** — `po_reconcile` 이 이 모듈에서 가져다 쓰므로 반대 방향은
      **함수 안에서 늦게** 부른다. `erp_supply_index()` 는 그쪽 모듈 전역에 캐시돼
      Z: 를 한 번만 훑는다([168]).
    ⚠ 못 읽으면 '없음'이 아니라 **'못읽음'** 이다([169]) — 리포트가 그대로 적는다.
    """
    raw = rec.get("원장_공급가액")
    if raw:
        return raw, "원장"
    try:
        from po_reconcile import supply_of
        v = supply_of(rec)
    except Exception:
        return None, "못읽음"
    return (v, "보정(ERP·명세서)") if v else (None, "")


def match_project(ledger_rec, ecount_rows, tol_amt, filter_cust, exp_amt=None):
    """프로젝트NO(적요/번호 내 포함) 1순위, 금액 2순위로 이카운트 전표 1건 매칭.

    `exp_amt` 를 주면 그것을 2순위 열쇠로 쓴다([124] 사다리). 안 주면 예전처럼
    원장 값을 본다 — **세금계산서 층은 안 준다**(아래 부르는 자리 주석).
    """
    prj = (ledger_rec.get("프로젝트NO") or "").strip()
    if exp_amt is None:
        exp_amt = ledger_rec.get("원장_공급가액")
    cust_ok = lambda e: (not filter_cust) or (filter_cust in (e.get("거래처") or "") or not e.get("거래처"))
    # 1순위: 프로젝트NO 문자열 포함
    if prj:
        for e in ecount_rows:
            blob = f"{e.get('적요','')} {e.get('번호','')} {json.dumps(e.get('_raw',{}), ensure_ascii=False)}"
            if prj in blob and cust_ok(e):
                return e, "프로젝트NO"
    # 2순위: 금액 근접 — **0 은 '금액 없음'으로 다뢬다**([124]).
    #   0 을 열쇠로 쓰면 금액 0 짜리 전표에 아무거나 붙는다.
    #   그리고 **후보가 유일할 때만** 붙인다([172]·[314] 와 같은 문).
    #   ★ 실측 2026-08-19: 사다리로 빈 금액을 채우자 139건이 붙었는데 **전부**
    #     '금액만'이었고 **88%(122건)는 EC 전표번호조차 없었다.** 서로 다른 금액
    #     23개가 139건을 만들었다 — 380,000 하나에 51건이 붙었다. 판매 후보가
    #     8,781행이라 380,000·240,000 같은 흔한 금액은 반드시 부딪힌다. 그렇게 붙은
    #     '일치'는 근거가 아니라 **충돌**이고, 그 행은 '미등록' 목록에서도 빠져
    #     **아무도 안 챙기는데 화면에서는 끝난 것처럼 보인다** — 못 붙는 것보다 나쁘다.
    if exp_amt:
        cands = []
        for e in ecount_rows:
            if e.get("금액") is not None and abs(e["금액"] - exp_amt) <= tol_amt and cust_ok(e):
                cands.append(e)
                if len(cands) > 1:
                    break          # 둘이면 그것으로 끝 — 어느 쪽인지 원본이 안 말해 준다
        if len(cands) == 1:
            return cands[0], "금액"
    return None, None


# ───────────────────────── inbox(수동 내보내기) 로더 ─────────────────────────
INBOX_DIR = os.path.join(BASE_DIR, "inbox")

# 판매 후보로 **절대 넣지 않는** 갈래. 회계 원장류 넷은 `[203]` 이 가른 그대로다 —
# 예전엔 `ledger` 하나만 걸러 나머지 셋이 판매로 샜다.
SALE_NEVER = ("ledger_acct", "journal", "cashbook",
              "unknown", "taxstep", "quote", "receipt")

INBOX_HASH_CACHE = os.path.join(REPORT_DIR, "ecount_reconcile_hashes.json")
INBOX_PARSE_CACHE = os.path.join(REPORT_DIR, "ecount_reconcile_rows.json")


def _atomic_json(path, payload):
    """속도 캐시는 못 써도 본 대조를 막지 않는다."""
    tmp = path + ".%s.tmp" % os.getpid()
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
        os.replace(tmp, path)
    except OSError:
        try:
            os.remove(tmp)
        except OSError:
            pass


def _json_cache(path, default):
    try:
        data = json.load(open(path, encoding="utf-8"))
        return data if isinstance(data, dict) else default
    except Exception:
        return default


def _intake_excel_paths():
    """이번 회차가 막 옮긴 엑셀 경로. 최근 분류표보다 항상 우선해 합친다."""
    out = []
    for name in ("upload_intake.json", "download_intake.json"):
        data = _json_cache(os.path.join(REPORT_DIR, name), {})
        for row in data.get("이동", []):
            path = row.get("목적지") if isinstance(row, dict) else ""
            if str(path).lower().endswith((".xls", ".xlsx", ".xlsm")):
                out.append(path)
    return out


def _dated_excel_paths(roots, days=3):
    """오늘·직전 날짜 폴더만 짧게 본다. ERP 직접 다운로드도 즉시 흡수하기 위함이다."""
    import glob
    out = []
    today = datetime.now().date()
    for root in roots:
        if os.path.normcase(os.path.abspath(root)) == os.path.normcase(os.path.abspath(INBOX_DIR)):
            out.extend(glob.glob(os.path.join(root, "**", "*.xls*"), recursive=True))
            continue
        for back in range(max(1, int(days))):
            d = today - timedelta(days=back)
            folder = os.path.join(root, f"{d.year}", f"{d.month:02d}", d.isoformat())
            out.extend(glob.glob(os.path.join(folder, "*.xls*")))
    return out


def _inbox_inventory(roots):
    """(경로, 내용종류, [크기, 수정초]) 목록.

    원본 전체를 매번 다시 여는 대신 저장 후 불변인 과거 자료는 분류표를 쓰고, 오늘 자료와
    투입함 이동 결과만 실제 파일 상태를 확인한다. 분류 규칙판이 바뀌거나 캐시가 깨지면
    기존 전체 훑기로 자동 복귀한다.
    """
    import glob
    from inbox_scan import cached_inventory, classify_cached, SKIP_DIRS

    snap = cached_inventory(roots, max_age_s=366 * 86400)
    items = {}
    if snap is None:
        base_paths = []
        for root in roots:
            base_paths.extend(glob.glob(os.path.join(root, "**", "*.xls*"), recursive=True))
        refresh = base_paths
    else:
        for row in snap:
            if not str(row["path"]).lower().endswith(".xlsx"):
                continue
            key = os.path.normcase(os.path.abspath(row["path"]))
            items[key] = (row["path"], row["kind"], [row["size"], row["mtime"]])
        refresh = _dated_excel_paths(roots) + _intake_excel_paths()

    for path in refresh:
        if (not path or os.path.basename(path).startswith("~$")
                or not str(path).lower().endswith(".xlsx")):
            continue
        if any(seg in SKIP_DIRS for seg in path.replace("\\", "/").split("/")):
            continue
        key = os.path.normcase(os.path.abspath(path))
        # 정본 보관소는 같은 경로를 덮어쓰지 않고 충돌 이름을 새로 만든다. 최근 날짜
        # 폴더에서 이름이 이미 분류표에 있으면 네트워크 stat까지 다시 할 이유가 없다.
        if key in items:
            continue
        try:
            st = os.stat(path)
        except OSError:
            continue
        sig = [int(st.st_size), int(st.st_mtime)]
        kind = classify_cached(path)
        items[key] = (path, kind, sig)
    return list(items.values()), snap is not None


def _sale_kind(classified, name):
    if classified == "ledger" or classified in SALE_NEVER:
        return None
    if classified in ("tax", "taxinv", "hometax"):
        return "tax_invoice"
    if classified in ("sales", "stmt", "slips", "po"):
        return "sale"
    return "tax_invoice" if ("세금계산서" in name or "계산서" in name) else "sale"

def load_inbox(cfg):
    """이카운트 화면에서 내려받은 판매현황/세금계산서 엑셀을 inbox/에서 읽어
    API 조회 결과와 동일한 형태(CAND 키)로 정규화한다.
    파일 분류: 파일명에 '판매'→sale, '세금계산서' 또는 '계산서'→tax_invoice."""
    import openpyxl
    # ★ 2026-07-30 수정 이유: 여기는 `INBOX_DIR/*.xlsx` **비재귀** 글롭이라
    #   (1) 원본 자료 폴더(`0. 원본 자료/1. ERP 내보내기/2026/07/2026-07-25/...`)를 아예 못 봤고
    #   (2) 로컬 inbox 의 하위 폴더도 못 봤다. 실제로 판매조회(898행)가 원본 자료 폴더에만
    #   있어서 이 대조기는 그 자료를 한 번도 읽지 못했다.
    #   경로는 source_dirs 한 곳에서만 정한다는 원칙(AGENTS.md)에 맞춰 excel_dirs() 를 쓴다.
    from source_dirs import excel_dirs
    from billing_fill import dedupe_files
    roots = excel_dirs() or [INBOX_DIR]
    inventory, _fast_inventory = _inbox_inventory(roots)
    typed, sigs, skipped = {}, {}, {}
    for f, classified, sig in inventory:
        kind = _sale_kind(classified, os.path.basename(f))
        if kind is None:
            if classified in SALE_NEVER:
                skipped[classified] = skipped.get(classified, 0) + 1
            continue
        key = os.path.normcase(os.path.abspath(f))
        typed[key] = kind
        sigs[key] = sig
    cands = [row[0] for row in inventory
             if os.path.normcase(os.path.abspath(row[0])) in typed]
    # ★ 같은 내용의 사본을 여러 번 읽지 않는다. 판매조회가 SHA256 동일한 3벌 있었고
    #   (2026-07-30 3배 합산 사고) 파일명(`__dup_`)으로 거르면 다음번에 다른 이름으로 뚫린다.
    files = dedupe_files(cands, cache_path=INBOX_HASH_CACHE, signatures=sigs)
    if not files:
        return None
    out = {"sale": [], "tax_invoice": []}
    used = []
    parsed = _json_cache(INBOX_PARSE_CACHE, {"schema": 1, "files": {}})
    if parsed.get("schema") != 1 or not isinstance(parsed.get("files"), dict):
        parsed = {"schema": 1, "files": {}}
    parsed_files = parsed["files"]
    parsed_dirty = False
    for f in files:
        name = os.path.basename(f)
        # 파일명이 무작위인 이카운트 내보내기가 섞여 있다. **내용**으로 종류를 먼저 본다.
        # 원장(차변/대변)은 매출 후보로 쓰면 엉뚱한 금액이 매칭돼 거짓 '일치'가 되므로 뺀다.
        # ★ `po` 는 빼지 않는다: ERP 판매조회에 PO번호 열이 있어 classify() 가 이 파일을
        #   `po` 로 준다(2026-07-30 확인). 여기서 po 를 빼면 정작 필요한 판매 자료가 빠진다.
        #   쿠팡 PO 목록(오더 번호·고객·금액)은 아래 머리글 조건에 '공급가액/합계금액' 이
        #   없어 자연히 걸러진다 — 종류 판정에 이중으로 의존하지 않는다.
        cache_key = os.path.normcase(os.path.abspath(f))
        kind = typed[cache_key]
        sig = sigs[cache_key]
        hit = parsed_files.get(cache_key)
        if (hit and hit.get("sig") == sig and hit.get("kind") == kind
                and isinstance(hit.get("rows"), list)):
            out[kind].extend(hit["rows"])
            used.append(f"{name}→{kind}")
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except (OSError, ValueError, KeyError):
            # 최근 분류표 뒤에 사람이 원본을 지웠다면 다음 전체 훑기까지 옛 행을 쓰지 않는다.
            continue
        file_rows = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            # 머리글 행 탐색: '공급가액' 또는 '합계' + '거래처'류가 있는 행
            hdr_i, hdr = None, None
            for i, r in enumerate(rows[:15]):
                cells = [str(c) for c in r if c is not None]
                joined = " ".join(cells)
                if ("공급가액" in joined or "합계금액" in joined) and len(cells) >= 3:
                    hdr_i, hdr = i, r
                    break
            if hdr_i is None:
                continue
            def col(*keys):
                for j, h in enumerate(hdr):
                    if h is None:
                        continue
                    hs = str(h)
                    if any(k in hs for k in keys):
                        return j
                return None
            j_amt  = col("공급가액")
            j_tot  = col("합계금액", "합계")
            j_date = col("일자", "날짜", "작성일")
            j_no   = col("승인번호", "전표", "No", "번호")
            j_cust = col("거래처")
            for r in rows[hdr_i + 1:]:
                if r is None or all(c is None for c in r):
                    continue
                amt = r[j_amt] if j_amt is not None else (r[j_tot] if j_tot is not None else None)
                if amt is None or _num(amt) is None:
                    continue
                blob = " ".join(str(c) for c in r if c is not None)
                if "합계" in blob[:10]:      # 합계행 제외
                    continue
                file_rows.append({
                    "SUPPLY_AMT": _num(amt),
                    "IO_DATE": _d(r[j_date]) if j_date is not None else "",
                    "IO_NO": _d(r[j_no]) if j_no is not None else "",
                    "CUST_DES": _d(r[j_cust]) if j_cust is not None else "",
                    "REMARKS": blob,
                })
        wb.close()
        out[kind].extend(file_rows)
        parsed_files[cache_key] = {"sig": sig, "kind": kind, "rows": file_rows}
        parsed_dirty = True
        used.append(f"{name}→{kind}")
    if parsed_dirty:
        # 현재 원본만 남긴다. 경로를 바꾸거나 삭제한 옛 파일의 행이 캐시에서 되살아나지 않는다.
        active = {os.path.normcase(os.path.abspath(f)) for f in files}
        parsed["files"] = {k: v for k, v in parsed_files.items() if k in active}
        _atomic_json(INBOX_PARSE_CACHE, parsed)
    if not (out["sale"] or out["tax_invoice"]):
        return None
    if skipped:
        # 맨 앞에 둔다 — 뒤에 붙이면 파일 목록 백 줄에 묻혀 아무도 못 본다.
        used.insert(0, "판매 후보에서 뺀 것 %d개(%s) — 판매 자료가 아니다"
                    % (sum(skipped.values()),
                       " · ".join("%s %d" % (k, v)
                                  for k, v in sorted(skipped.items()))))
    out["_files"] = used
    out["_skipped"] = skipped
    return out


def key_warning(results, meta):
    """짝이 거의 안 지어지면 'ERP 미등록이 많다'가 아니라 **열쇠가 안 맞는다**.

    ★ 2026-08-19 실측: 750건 중 `이카운트미등록` 692건(92%)이었다. 경보가 대부분이면
      그 경보는 아무도 안 본다(`[170]`). 근거를 세어 보면 미등록이 아니라 **물어볼
      수가 없었던** 것이다 — 1순위 프로젝트NO 는 ERP 판매전표 본문에 그 번호가 없고,
      2순위 금액은 원장 06시트 공급가액이 사람 손 입력이라 대부분 비어 있다
      (금액허용오차 0 이라 빈 값으로는 한 건도 못 맞춘다).

    ★ **여기서 금액을 짐작해 채우지 않는다.** 그것은 `[170]` 의 `supply_of()` 사다리가
      할 일이고, 그 전에 판매 후보 오염(위 `SALE_NEVER`)을 먼저 걷어야 한다 —
      순서가 뒤집히면 남의 회사 금액에 붙어 **거짓 '일치'** 가 난다.

    판정은 `erp_ledger_check.key_looks_wrong` 을 그대로 빌린다(`[162]`).
    """
    try:
        from erp_ledger_check import key_looks_wrong
    except Exception:
        return []
    total = len(results)
    if not total:
        return []
    맞음 = sum(1 for r in results
              if "일치" in r["①판매/명세서_판정"]
              or "불일치" in r["①판매/명세서_판정"]
              or "금액없음" in r["①판매/명세서_판정"])
    if not key_looks_wrong(맞음, total):
        return []
    # ★ **유효 금액**으로 센다([124]). 원값으로 세면 사다리가 채운 뒤에도
    #   "빈 행이 N건이라 한 건도 못 맞춘다"를 **근거를 대며** 확언한다([169]).
    빈금액 = sum(1 for r in results if not r.get("공급가액(보정)"))
    후보 = meta.get("sale_rows")
    out = ["",
           "> ★ **이 집계는 'ERP 미등록'을 뜻하지 않는다 — 열쇠가 안 맞는다.**",
           "> 정산 %d건 중 이카운트 전표가 붙은 것이 %d건뿐이다. 열쇠가 맞으면 "
           "몇 건은 반드시 걸린다." % (total, 맞음),
           "> · 1순위 **프로젝트NO** — ERP 판매전표 본문에 그 번호가 안 적혀 있다.",
           "> · 2순위 **금액** — 원장 빈칸을 ERP·명세서 사다리로 채운 뒤에도 "
           "금액을 못 구한 행이 **%d건(%d%%)**. "
           "채워진 행도 **같은 금액 전표가 여럿이면 안 붙인다** — 그건 근거가 "
           "아니라 충돌이다(허용오차 0 · 후보가 유일할 때만)."
           % (빈금액, round(빈금액 * 100.0 / total)),
           "> 그러므로 아래 '이카운트미등록'을 **미청구·미발행 근거로 쓰지 말 것.**"]
    if 후보:
        out.insert(3, "> · 판매 후보 전표 %d행을 놓고 잰 값이다." % 후보)
    return out


# ───────────────────────── 대조 로직 ─────────────────────────
def reconcile(recs, ecount, cfg):
    rc = cfg["reconcile"]
    tol = rc.get("금액허용오차", 0)
    fc = rc.get("거래처_필터", "")
    sale = norm_ecount(ecount.get("sale"))
    tax = norm_ecount(ecount.get("tax_invoice"))
    online = ecount.get("_online", False)

    results = []
    for sid, r in sorted(recs.items()):
        # 원장이 유상이 아니면 세금계산서/판매 대조 제외 안내
        유상 = (r.get("비용구분") == "유상")

        # ① 판매/거래명세서 층
        #   ★ 금액은 **한 곳에서** 구한다([124]·[170]) — 매칭 열쇠·판정·불일치 문구·
        #     리포트 열·`key_warning` 의 빈금액이 전부 이 `sup` 하나를 본다.
        sup, sup_src = supply_effective(r)
        if online:
            m, how = match_project(r, sale, tol, fc, sup)
            if m:
                if not sup:
                    # ★ 금액을 모르면 **불일치라고 말하지 않는다**([169]). 여기 오는
                    #   것은 프로젝트NO 로 붙은 것뿐이다(금액으로 붙었다면 sup 이 있다)
                    #   — 근거는 오히려 센데 금액을 못 대는 것뿐이다.
                    s1 = "일치(프로젝트NO · 금액 확인 못 함)"
                elif m.get("금액") is None:
                    s1 = "이카운트금액없음"
                elif abs(m["금액"] - sup) <= tol:
                    # ★ **어떻게 붙었는지가 곧 근거의 세기다** (2026-08-19 실측).
                    #   `how` 는 여태 계산해 놓고 버려졌다. 금액으로만 붙은 것은
                    #   '금액이 우연히 같은 남의 전표'일 수 있어 근거가 약하다
                    #   (`[170]` 유형D 와 같은 자리). 실제로 그랬다 — 판매 후보
                    #   오염을 걷어내자 일치 23건 중 3건이 사라졌고, 그 셋은
                    #   원장 금액이 한 원도 안 바뀐 채 사라졌다(=금액 매치였다).
                    #   ⚠ 사다리가 빈 금액을 채우면 이 갈래가 는다 — 그것들이
                    #     'ERP 에 등록된 것'처럼 보이면 안 되므로 딱지를 그대로 둔다.
                    s1 = "일치" if how == "프로젝트NO" else "일치(금액만 · 근거 약함)"
                else:
                    # ★ **어느 금액과 비교했는지를 밝힌다** — '원장'이라고만 적으면
                    #   사다리로 채운 값까지 원장에 적힌 것처럼 읽힌다([169]).
                    s1 = f"금액불일치({sup_src} {_money(sup)} / EC {_money(m.get('금액'))})"
                ec_sale_no = m.get("번호")
            else:
                s1 = "이카운트미등록" if 유상 else "해당없음(무상)"
                ec_sale_no = ""
        else:
            s1 = "원장기준-EC미조회"
            ec_sale_no = ""

        # ② 세금계산서 층
        if online:
            # ⚠ **여기에는 사다리를 안 붙인다**([124]·[172]). 판매 층은
            #   `일치(금액만 · 근거 약함)` 딱지가 먼저 들어가 근거의 세기를
            #   말할 수 있는데, 이 층에는 그 딱지가 없어 약한 금액 매치가
            #   **그냥 '일치'** 로 적힌다. 딱지를 먼저 만든 뒤에 넓힌다.
            m2, _ = match_project(r, tax, tol, fc)
            if m2:
                ec_tax_no = m2.get("번호")
                led_no = r.get("원장_세금계산서승인번호") or ""
                if led_no and ec_tax_no and led_no != ec_tax_no:
                    s2 = f"승인번호불일치(원장 {led_no} / EC {ec_tax_no})"
                else:
                    s2 = "일치"
            else:
                led_issued = r.get("원장_세금계산서실제발행일") or ""
                if led_issued:
                    s2 = "원장발행-EC미확인"
                else:
                    s2 = "양측미발행" if 유상 else "해당없음(무상)"
                ec_tax_no = ""
        else:
            s2 = "원장기준-EC미조회"
            ec_tax_no = ""

        results.append({
            "정산ID": sid,
            "업무구분": r.get("업무구분"),
            "프로젝트NO": r.get("프로젝트NO"),
            "캠프명": r.get("캠프명"),
            "비용구분": r.get("비용구분"),
            "작업완료일": r.get("작업완료일"),
            "원장_공급가액": r.get("원장_공급가액"),
            "공급가액(보정)": sup,
            "금액출처": sup_src,
            "원장_거래명세서번호": r.get("원장_거래명세서번호"),
            "EC_판매번호": ec_sale_no,
            "①판매/명세서_판정": s1,
            "원장_세금계산서발행일": r.get("원장_세금계산서실제발행일") or r.get("원장_세금계산서발행일"),
            "원장_세금계산서승인번호": r.get("원장_세금계산서승인번호"),
            "EC_세금계산서번호": ec_tax_no,
            "②세금계산서_판정": s2,
            "원장_입금액": r.get("원장_입금액"),
            "원장_미수금액": r.get("원장_미수금액"),
        })
    return results


# ───────────────────────── 리포트 출력 ─────────────────────────
def write_reports(results, cfg, meta):
    os.makedirs(REPORT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base = os.path.join(REPORT_DIR, f"이카운트대조_{stamp}")

    # CSV
    cols = list(results[0].keys()) if results else []
    with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(results)

    # 집계
    from collections import Counter
    c1 = Counter(r["①판매/명세서_판정"].split("(")[0] for r in results)
    c2 = Counter(r["②세금계산서_판정"].split("(")[0] for r in results)

    # Markdown
    with open(base + ".md", "w", encoding="utf-8") as f:
        f.write(f"# 이카운트 ↔ 관리대장 대조 리포트\n\n")
        f.write(f"- 생성: {datetime.now():%Y-%m-%d %H:%M}\n")
        f.write(f"- 이카운트 연결: {'예(실시간 조회)' if meta.get('online') else '아니오(원장 기준 준비표)'}\n")
        f.write(f"- 대상 정산 건수: {len(results)}\n")
        if meta.get("expiry_days") is not None:
            f.write(f"- 인증키 만료까지: {meta['expiry_days']}일\n")
        if meta.get("note"):
            f.write(f"- 비고: {meta['note']}\n")
        f.write("\n## ① 판매/거래명세서 대조 집계\n\n")
        for k, v in c1.most_common():
            f.write(f"- {k}: {v}건\n")
        for ln in key_warning(results, meta):
            f.write(ln + "\n")
        f.write("\n## ② 세금계산서 대조 집계\n\n")
        for k, v in c2.most_common():
            f.write(f"- {k}: {v}건\n")
        f.write("\n## 상세(불일치·미등록 우선)\n\n")
        f.write("| 정산ID | 프로젝트NO | 캠프 | 공급가액(*=보정) | ①판매 | ②세금계산서 |\n")
        f.write("|---|---|---|--:|---|---|\n")
        def rank(r):
            bad = ("불일치", "미등록", "미확인", "미발행")
            return 0 if any(b in r["①판매/명세서_판정"] or b in r["②세금계산서_판정"] for b in bad) else 1
        for r in sorted(results, key=rank):
            f.write(f"| {r['정산ID']} | {r['프로젝트NO']} | {r['캠프명']} | "
                    f"{_money(r['공급가액(보정)'])}{'*' if str(r.get('금액출처','')).startswith('보정') else ''} | {r['①판매/명세서_판정']} | {r['②세금계산서_판정']} |\n")

    # 독립 xlsx(마스터와 무관한 별도 파일)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "이카운트대조"
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
        red = PatternFill("solid", fgColor="FFC7CE")
        grn = PatternFill("solid", fgColor="E2EFDA")
        for r in results:
            ws.append([r.get(c) for c in cols])
            row = ws[ws.max_row]
            j1 = cols.index("①판매/명세서_판정"); j2 = cols.index("②세금계산서_판정")
            for jj in (j1, j2):
                val = str(row[jj].value or "")
                row[jj].fill = grn if val.startswith("일치") or "해당없음" in val else red
        ws.freeze_panes = "A2"
        wb.save(base + ".xlsx")
    except Exception as e:
        print("  (xlsx 생성 생략:", e, ")")

    return base


# ───────────────────────── 메인 ─────────────────────────
def main():
    args = set(sys.argv[1:])
    cfg = load_config()
    expiry = days_until_expiry(cfg)
    if expiry is not None and expiry <= cfg["auth"].get("만료경고_남은일수", 30):
        print(f"⚠ 인증키 만료까지 {expiry}일 — 이카운트에서 유효기간 연장 필요")

    # 설정 완성도
    need = [k for k in ("COM_CODE", "USER_ID") if not (cfg["auth"].get(k) or "").strip()]

    if "--selftest" in args:
        print("=== 설정 점검 ===")
        print("인증키:", "설정됨" if cfg["auth"].get("API_CERT_KEY") else "없음")
        print("COM_CODE:", cfg["auth"].get("COM_CODE") or "(비어있음 — 필수)")
        print("USER_ID :", cfg["auth"].get("USER_ID") or "(비어있음 — 필수)")
        print("도메인  :", "테스트(sboapi)" if cfg["auth"].get("IS_TEST") else "실서비스(oapi)")
        if need:
            print("→ 미설정:", ", ".join(need)); return
        try:
            cli = EcountClient(cfg); cli.login()
            print("Zone:", cli.zone, "/ 로그인 성공, SESSION_ID 발급됨")
        except EcountError as e:
            print("로그인 실패:", e)
        return

    # 원장 읽기
    recs = read_ledger(cfg["reconcile"]["master_xlsx"])
    print(f"원장 정산 건수: {len(recs)}")

    ecount = {"_online": False}
    note = ""
    # 1순위: inbox/ 의 이카운트 수동 내보내기 파일 (판매·세금계산서 조회 API는 이카운트가 미제공)
    inbox = None if "--no-inbox" in args else load_inbox(cfg)
    if inbox:
        ecount["sale"] = inbox["sale"]
        ecount["tax_invoice"] = inbox["tax_invoice"]
        ecount["_online"] = True
        note = "이카운트 내보내기 파일(inbox) 기준 대조: " + ", ".join(inbox["_files"])
        print("inbox 로드:", note)
        print(f"  판매 {len(inbox['sale'])}건 / 세금계산서 {len(inbox['tax_invoice'])}건")
    elif "--offline" in args or need:
        note = ("이카운트 미조회 — " +
                ("COM_CODE/USER_ID 미설정" if need else "offline 모드") +
                ". 원장 기준 준비표만 생성.")
        print("i", note)
    else:
        note = ("이카운트 OAPI는 판매·세금계산서 '조회' API를 제공하지 않음(2026-07 확인). "
                "이카운트 화면에서 판매현황/세금계산서 목록을 엑셀로 내려받아 inbox/ 폴더에 넣으면 자동 대조됩니다. "
                "이번 실행은 원장 기준 준비표.")
        print("i", note)

    results = reconcile(recs, ecount, cfg)
    base = write_reports(results, cfg, {"online": ecount["_online"], "expiry_days": expiry, "note": note,
                                        # 열쇠 경고가 "무엇을 놓고 잰 값인가"를 말하려면 후보 수가 필요하다.
                                        "sale_rows": len(ecount.get("sale") or [])})
    print("리포트 생성:")
    for ext in (".md", ".csv", ".xlsx"):
        if os.path.exists(base + ext):
            print("  -", base + ext)


if __name__ == "__main__":
    main()
