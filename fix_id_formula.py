# -*- coding: utf-8 -*-
"""
fix_id_formula.py — 빈 행의 **ID 채번 수식**이 그 시트에 맞는지 보고 고친다
================================================================================
ID 열 수식은 시트를 만들 때 02_돌발AS접수에서 복사해 붙인 흔적이 남아 있다.
06_거래서류청구수금이 그 예다(2026-07-27 발견):

    현재  =IF($B72="","","AS-"&TEXT(IF($D72="",TODAY(),$D72),"yymm")&"-"&TEXT(ROW()-4,"000"))
                              ~~~~            ~~~~
                              접수ID 접두어    프로젝트NO(=UJ2601138)를 날짜로 읽는다
    맞음  =IF($B72="","","JS-"&TEXT(IF($F72="",TODAY(),$F72),"yymm")&"-"&TEXT(ROW()-4,"000"))
                              정산ID 접두어    작업완료일

지금은 빈 행이라 아무도 못 봤지만, **다음에 정산 한 건만 넣어도** 정산ID가 `AS-…`로 매겨져
02시트 접수ID와 이름이 겹친다. 06의 원천업무ID가 접수ID를 가리키는 구조라 사람이 봐도
어느 쪽 AS-2607-068인지 알 수 없게 된다.

값이 들어 있는 행은 건드리지 않는다 — 이미 매겨진 ID를 바꾸면 03·06이 값으로 들고 있는
참조가 끊긴다(AGENTS.md: ID는 다시 매기지 않는다).

  python fix_id_formula.py            # 미리보기
  python fix_id_formula.py --apply    # vN+1 생성
"""
import sys, os, re, zipfile

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from fix_formulas import sheet_map, parse_cells, esc
from project_resolve import SHEET_ID, HDR_ROW


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def scan(master):
    """고쳐야 할 (시트, 행, 기존수식, 새수식) 목록."""
    import openpyxl
    wbv = openpyxl.load_workbook(master, data_only=True)
    wbf = openpyxl.load_workbook(master, data_only=False)
    z = zipfile.ZipFile(master)
    smap = sheet_map(z)
    jobs, notes = [], []

    for sheet, (idc, pfx, datec, keyc) in SHEET_ID.items():
        if sheet not in wbv.sheetnames:
            continue
        wv, wf = wbv[sheet], wbf[sheet]
        hdr = [str(h).strip() if h else "" for h in
               next(wv.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))]
        if datec not in hdr or keyc not in hdr:
            notes.append(f"{sheet}: '{datec}'/'{keyc}' 열이 없어 건너뜀")
            continue
        dcol = col_letter(hdr.index(datec) + 1)

        for r in range(HDR_ROW + 1, wf.max_row + 1):
            f = wf.cell(row=r, column=1).value
            if not (isinstance(f, str) and f.startswith("=")):
                continue                       # 값으로 굳은 행 — 절대 건드리지 않는다
            if wv.cell(row=r, column=1).value not in (None, ""):
                continue                       # 이미 ID가 나와 있는 행(=데이터 행)
            body = f[1:]
            mp = re.search(r'"(\w{2,3})-"', body)
            md = re.search(r'TEXT\(IF\(\$([A-Z]+)\d+="",TODAY\(\),\$[A-Z]+\d+\),"yymm"\)',
                           body.replace(" ", ""))
            if not (mp and md):
                notes.append(f"{sheet} {r}행: 수식 모양이 달라 자동 판정 불가 — 사람이 볼 것")
                continue
            if mp.group(1) == pfx and md.group(1) == dcol:
                continue                       # 이미 맞다
            new = body
            new = new.replace(f'"{mp.group(1)}-"', f'"{pfx}-"')
            new = re.sub(r'(TEXT\(IF\(\$)[A-Z]+(\d+="",TODAY\(\),\$)[A-Z]+(\d+\),"yymm"\))',
                         lambda m: m.group(1) + dcol + m.group(2) + dcol + m.group(3), new)
            jobs.append({"sheet": sheet, "row": r, "old": body, "new": new,
                         "why": f"{mp.group(1)}→{pfx} · 날짜열 {md.group(1)}→{dcol}({datec})",
                         "file": smap[sheet]})
    z.close(); wbv.close(); wbf.close()
    return jobs, notes


def apply(master, jobs):
    from collections import defaultdict
    by = defaultdict(list)
    for j in jobs:
        by[j["file"]].append(j)

    zin = zipfile.ZipFile(master)
    edits = {}
    for fname, items in by.items():
        xml = zin.read(fname).decode("utf-8")
        cells = {(c[0], c[1]): c for c in parse_cells(xml)}
        # 뒤에서부터 고친다 — 앞을 먼저 바꾸면 뒤 셀의 위치(offset)가 밀린다
        for j in sorted(items, key=lambda x: -x["row"]):
            hit = cells.get((j["row"], "A"))
            if not hit:
                continue
            _, _, _old, s, e = hit
            chunk = xml[s:e]
            new = re.sub(r"(<f(?:\s[^>]*)?>)(.*?)(</f>)",
                         lambda m: m.group(1) + esc(j["new"]) + m.group(3),
                         chunk, count=1, flags=re.S)
            new = re.sub(r"<v[^>]*>.*?</v>|<v\s*/>", "<v/>", new, count=1, flags=re.S)
            xml = xml[:s] + new + xml[e:]
        edits[fname] = xml.encode("utf-8")

    mv = re.search(r"_v(\d+)\.xlsx$", master)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mv.group(1))+1}.xlsx", master)
    if os.path.exists(dst):
        sys.exit(f"{os.path.basename(dst)} 이미 존재 — 중단")
    zout = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, edits.get(it.filename) or zin.read(it.filename))
    zout.close()

    z = zipfile.ZipFile(dst)
    assert z.testzip() is None, "zip 무결성 실패"
    # 의도한 시트 말고 다른 파트가 바뀌지 않았는지 — 차트·도형 손상 조기 발견
    diff = [n for n in zin.namelist() if n not in edits and zin.read(n) != z.read(n)]
    z.close(); zin.close()
    assert not diff, f"의도치 않은 파트 변경: {diff[:3]}"
    return dst


def main():
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    jobs, notes = scan(master)

    from collections import Counter
    per = Counter(f"{j['sheet']} ({j['why']})" for j in jobs)
    print(f"채번 수식이 시트와 어긋난 빈 행 {len(jobs)}건")
    for k, v in per.items():
        print(f"  {k} — {v}행")
    for n in notes[:5]:
        print("  [확인]", n)
    if jobs:
        j = jobs[0]
        print(f"\n  예) {j['sheet']} {j['row']}행")
        print(f"    전: ={j['old'][:120]}")
        print(f"    후: ={j['new'][:120]}")

    if not jobs:
        print("고칠 것 없음 — 모든 시트의 채번 수식이 제자리입니다.")
        return
    if "--apply" not in sys.argv:
        print("\n미리보기 — 실제 반영: python fix_id_formula.py --apply")
        return
    dst = apply(master, jobs)
    print(f"\n반영 완료 → {os.path.basename(dst)} ({len(jobs)}행)")


if __name__ == "__main__":
    main()
