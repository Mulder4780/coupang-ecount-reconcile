# -*- coding: utf-8 -*-
"""
auto_status.py — 02_돌발AS접수의 **상태 칸을 근거로 자동 표기**한다
================================================================================
사용자 지시(2026-07-28): 사진·동영상·완료보고서·ERP·밴드수정·추가작업·관리자검증을
자동으로 채우고, 담당관리자는 전부 유현민, 관리자검증은 "어지간하면 일치".

■ 손대지 않는 것 — 검증결과·검증문제코드는 **수식**이다
  AK(검증결과) = AL 이 비면 "정상", AL = AN 에서 왔고 AN 이 원인을 모은다.
  즉 **원인 칸을 채우면 검증결과는 저절로 '정상'이 된다.** 수식을 값으로 덮으면
  다음부터 자동 판정이 죽는다 — 절대 덮지 않는다.

■ 드롭다운 허용값이 정해져 있다 (10_코드관리)
  사진·동영상·완료보고서 : 등록 / 누락 / 해당없음
  ERP등록               : 완료 / 미등록 / 해당없음     ← '등록' 아님
  밴드수정              : 완료 / 미수정 / 해당없음     ← '등록·누락' 아님
  최초접수외추가작업      : 있음 / 없음 / 확인필요
  추가작업확인상태        : 반영완료 / 미반영 / 확인필요
  관리자검증상태          : 일치 / 추가작업발생 / 작업내용누락 / 확인필요
  목록 밖 값을 넣으면 유효성 위반이고 검증 수식(Z="미등록" 등)도 안 맞는다.

■ 근거 (지어내지 않는다)
  사진      24_밴드업무추출의 사진 장수 > 0
  완료보고서 밴드에 '작업완료' 글이 있음
  밴드수정   밴드에 그 프로젝트 글이 있음
  ERP       06_거래서류청구수금에 거래명세서번호가 있음
  추가작업   03_현장작업실적의 추가작업내용
  ★ 동영상은 **근거원이 아예 없다.** 회사가 상시 요구하는 항목이 아니고 검증 수식도
    '누락'일 때만 문제 삼으므로 '해당없음'으로 둔다. 근거로 '등록'이라 적지 않는다.

■ 이미 사람이 적어 둔 값은 덮지 않는다 (담당관리자만 예외 — 전부 유현민 고정 지시)

사용
  python auto_status.py             # 무엇을 어떻게 채울지 집계만 (파일 안 건드림)
  python auto_status.py --sample    # 행 단위 예시도 같이
  python auto_status.py --apply     # vN+1 생성
"""
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path, esc  # noqa: E402

SHEET = "02_돌발AS접수"
HDR = 4
# 채울 열 → 허용값(유효성 목록). 목록에 없는 값은 쓰지 않는다.
ALLOWED = {
    "사진등록": ("등록", "누락", "해당없음"),
    "동영상등록": ("등록", "누락", "해당없음"),
    "완료보고서등록": ("등록", "누락", "해당없음"),
    "ERP등록": ("완료", "미등록", "해당없음"),
    "밴드수정": ("완료", "미수정", "해당없음"),
    "최초접수외추가작업": ("있음", "없음", "확인필요"),
    "추가작업확인상태": ("반영완료", "미반영", "확인필요"),
    "관리자검증상태": ("일치", "추가작업발생", "작업내용누락", "확인필요"),
    "담당관리자": None,          # 자유값 — 유현민 고정
}
MANAGER = "유현민"
OVERWRITE = {"담당관리자"}       # 사용자 지시로 기존 값도 덮는 열


def _s(v):
    return "" if v is None else str(v).strip()


def load(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def rows(sn):
        ws = wb[sn]
        hdr = [_s(h) for h in next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        out = []
        for n, r in enumerate(ws.iter_rows(min_row=HDR + 1, values_only=True), start=HDR + 1):
            if r[0] in (None, ""):
                continue
            g = {h: (r[i] if i < len(r) else None) for h, i in ix.items()}
            g["_행"] = n
            out.append(g)
        return hdr, out

    hdr02, as_rows = rows(SHEET)
    _, band = rows("24_밴드업무추출")
    _, fw = rows("03_현장작업실적")
    _, st = rows("06_거래서류청구수금")
    wb.close()

    # ── 근거를 프로젝트NO 로 모은다 ───────────────────────────────────────────
    photo, doneposted, posted = defaultdict(int), set(), set()
    for b in band:
        k = _s(b.get("프로젝트NO"))
        if not k:
            continue
        posted.add(k)
        try:
            photo[k] = max(photo[k], int(_s(b.get("사진")) or 0))
        except ValueError:
            pass
        if "완료" in _s(b.get("진행상태")):
            doneposted.add(k)

    extra, worked = set(), set()
    for f in fw:
        k = _s(f.get("프로젝트NO"))
        if not k:
            continue
        if _s(f.get("추가작업내용")):
            extra.add(k)
        d = _s(f.get("실제작업상세")) or _s(f.get("실제작업항목"))
        if d and "자동 초안" not in d:
            worked.add(k)

    erp = {_s(s.get("프로젝트NO")) for s in st if _s(s.get("거래명세서번호"))}
    erp.discard("")

    return hdr02, as_rows, {"사진": photo, "완료글": doneposted, "밴드글": posted,
                            "추가작업": extra, "작업내용": worked, "ERP": erp}


def decide(r, ev):
    """한 행에 대해 {열: 값}. 근거가 없으면 값을 만들지 않는다."""
    k = _s(r.get("프로젝트NO"))
    state = _s(r.get("진행상태"))
    cost = _s(r.get("유상·무상·보험"))
    out = {"담당관리자": MANAGER}          # 전부 고정 (사용자 지시)

    if state == "취소":
        # 취소된 건은 문서·등록 의무가 없다. 검증도 통과시킨다.
        out.update({"사진등록": "해당없음", "동영상등록": "해당없음",
                    "완료보고서등록": "해당없음", "ERP등록": "해당없음",
                    "밴드수정": "해당없음", "최초접수외추가작업": "없음",
                    "추가작업확인상태": "반영완료", "관리자검증상태": "일치"})
        return out

    if state != "작업완료":
        # 아직 안 끝난 건은 문서 항목을 판단할 때가 아니다. 검증 수식도
        # '작업완료'일 때만 이 칸들을 보므로 건드리지 않는다.
        return out

    out["사진등록"] = "등록" if ev["사진"].get(k, 0) > 0 else "누락"
    out["동영상등록"] = "해당없음"          # 근거원이 없다 — 위 설명 참고
    out["완료보고서등록"] = "등록" if k in ev["완료글"] else "누락"
    out["밴드수정"] = "완료" if k in ev["밴드글"] else "미수정"

    if k in ev["ERP"]:
        out["ERP등록"] = "완료"
    elif cost in ("무상", "보험"):
        out["ERP등록"] = "해당없음"          # 청구가 없으니 매출 등록 대상이 아니다
    else:
        out["ERP등록"] = "미등록"

    has_extra = k in ev["추가작업"]
    out["최초접수외추가작업"] = "있음" if has_extra else "없음"
    out["추가작업확인상태"] = "반영완료"       # 03시트에 이미 적혀 있으면 반영된 것이다
    # 사용자 지시: "어지간하면 일치". 추가작업이 있었던 건만 그 사실을 남긴다.
    out["관리자검증상태"] = "추가작업발생" if has_extra else "일치"
    return out


def plan(path):
    hdr, rows, ev = load(path)
    from openpyxl.utils import get_column_letter as GL
    col = {h: GL(hdr.index(h) + 1) for h in ALLOWED if h in hdr}
    missing = [h for h in ALLOWED if h not in hdr]
    if missing:
        raise RuntimeError(f"{SHEET} 에 없는 열: {missing}")

    fills, tally, skipped = {}, defaultdict(Counter), Counter()
    for r in rows:
        want = decide(r, ev)
        for h, v in want.items():
            allow = ALLOWED[h]
            if allow and v not in allow:
                raise RuntimeError(f"허용되지 않은 값 {h}={v!r}")
            cur = _s(r.get(h))
            if cur and h not in OVERWRITE:
                skipped[h] += 1                  # 사람이 적어 둔 값은 덮지 않는다
                continue
            if cur == v:
                continue
            fills[f"{col[h]}{r['_행']}"] = v
            tally[h][v] += 1
    return fills, tally, skipped, len(rows)


def apply_cells(xml, fills):
    """★ 6천 칸을 하나씩 찾으면 XML을 6천 번 훑는다. **한 번만 훑는다.**"""
    hit, seen = 0, set()

    def repl(m):
        nonlocal hit
        ref, attrs = m.group(1), m.group(2) or ""
        if ref not in fills:
            return m.group(0)
        if "t=\"shared\"" in m.group(0) or "t=\"array\"" in m.group(0):
            return m.group(0)                    # 수식 칸은 건드리지 않는다
        seen.add(ref)
        hit += 1
        a = re.sub(r'\s+t="[^"]*"', "", attrs)
        return f'<c r="{ref}"{a} t="inlineStr"><is><t>{esc(fills[ref])}</t></is></c>'

    out = re.sub(r'<c r="([A-Z]+\d+)"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)',
                 repl, xml, flags=re.S)
    return out, hit, [r for r in fills if r not in seen]


def main():
    do = "--apply" in sys.argv
    m = latest_master()
    src = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src)}\n")

    fills, tally, skipped, n = plan(src)
    print(f"대상 {n}행 · 채울 칸 {len(fills)}개\n")
    print(f"  {'열':<16}{'채움':>6}  값 분포                          {'기존값 유지':>10}")
    for h in ALLOWED:
        c = tally.get(h)
        if not c:
            continue
        dist = " · ".join(f"{v} {k}" for v, k in c.most_common())
        print(f"  {h:<16}{sum(c.values()):>6}  {dist:<34}{skipped.get(h, 0):>10}")
    print("\n  ※ 검증결과·검증문제코드는 **수식이라 건드리지 않습니다** —"
          " 원인 칸이 채워지면 저절로 '정상'이 됩니다.")
    print("  ※ 동영상등록은 근거원이 없어 '해당없음'으로 둡니다(없는 근거로 '등록'이라 적지 않습니다).")

    if not do:
        print("\n실제로 채우려면:  python auto_status.py --apply")
        return 0

    zin = zipfile.ZipFile(src)
    sp = sheet_xml_path(zin, SHEET)
    xml = zin.read(sp).decode("utf-8")
    new, hit, miss = apply_cells(xml, fills)
    if miss:
        print(f"\n★ XML에 칸이 없어 못 채운 곳 {len(miss)}개: {miss[:5]}")
    from dashboard_clean import force_recalc
    changed = {sp: new.encode("utf-8"),
               "xl/workbook.xml": force_recalc(
                   zin.read("xl/workbook.xml").decode("utf-8")).encode("utf-8")}

    mm = re.search(r"_v(\d+)\.xlsx$", src)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm.group(1)) + 1}.xlsx", src)
    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close()
    zin.close()

    try:
        verify(src, tmp, fills, set(changed))
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise
    os.replace(tmp, dst)
    print(f"\n{hit}칸 채움 → {os.path.basename(dst)}")
    print("검증결과 재계산은 자동입니다(excel_recalc·보관본 회차) — 엑셀을 열 필요 없습니다."
          " 값 수정은 앱에서만 합니다.")
    return 0


def verify(src, out, fills, allowed):
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"
    wa = openpyxl.load_workbook(src, data_only=True)
    wb_ = openpyxl.load_workbook(out, data_only=True)
    sa, sb = wa[SHEET], wb_[SHEET]

    for ref, v in fills.items():
        assert _s(sb[ref].value) == v, f"{ref} 이 {v!r} 로 안 들어갔다 ({sb[ref].value!r})"
    # 채운 칸 말고는 한 칸도 안 바뀌어야 한다
    diff = []
    for r in range(1, max(sa.max_row, sb.max_row) + 1):
        for c in range(1, max(sa.max_column, sb.max_column) + 1):
            ca = sa.cell(r, c)
            if ca.coordinate in fills:
                continue
            if ca.value != sb.cell(r, c).value:
                diff.append(ca.coordinate)
                if len(diff) > 5:
                    break
    assert not diff, f"의도치 않은 셀 변경: {diff}"
    # 수식은 살아 있어야 한다 — 이게 죽으면 자동 판정이 통째로 멈춘다
    wf = openpyxl.load_workbook(out, data_only=False)[SHEET]
    for ref in ("AK5", "AL5", "AN5", "A5", "M5"):
        assert str(wf[ref].value or "").startswith("="), f"{ref} 수식이 사라졌다"
    wa.close(); wb_.close()

    zs = zipfile.ZipFile(src)
    other = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not other, f"의도치 않은 파트 변경: {other}"
    zs.close(); z.close()
    print(f"  검증 통과 — {len(fills)}칸 반영 · 수식 보존 · 다른 셀·파트 변경 없음")


if __name__ == "__main__":
    sys.exit(main())
