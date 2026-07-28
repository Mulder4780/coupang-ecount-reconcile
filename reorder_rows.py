# -*- coding: utf-8 -*-
"""
reorder_rows.py — 관리대장 행을 **날짜 오름차순(과거→최근)** 으로 재배치
================================================================================
백필을 뒤에 이어 붙이다 보니 시트가 7월 → 6월 → 5월 → 4월 순으로 뒤섞여 있다.
이걸 실제 작업일 순서로 정렬해 맨 위가 가장 과거, 아래로 갈수록 최근이 되게 한다.

■ ID는 다시 매기지 않는다 (중요)
  접수ID·점검ID는 이미 연·월을 품고 있어(AS-2604-… < AS-2605-…) 행만 정렬하면
  번호도 자연히 오름차순이 된다. 반면 ID를 새로 매기면 03_현장작업실적·06_거래서류청구수금
  등 **다른 시트가 값으로 들고 있는 ID 참조가 전부 끊긴다**(AS 42건·PM 25건 등).
  → 정렬만 하고 번호는 그대로 둔다. ID와 날짜가 안 맞는 건은 리포트로 따로 알린다.

■ 수식 처리
  행을 옮기면 수식의 **상대참조 행번호는 이동량만큼 함께 움직여야** 한다.
    · 자기 행 참조:  $B5  → $B{새 행}
    · 직전 행 참조:  COUNT($AO$4:AO4)  → COUNT($AO$4:AO{새 행-1})   ← 누적 카운터
    · 절대참조($5):  그대로 둔다
  이 시트에서 자기 행 외 참조는 위 누적 카운터 하나뿐임을 확인하고 만들었다.

사용
  python reorder_rows.py                       # 계획만 보기(최신본)
  python reorder_rows.py --apply               # vN+1 생성
  python reorder_rows.py --sheet 02_돌발AS접수  # 특정 시트만
  python reorder_rows.py --self-test           # 합성 워크북 검증
"""
import sys, os, re, zipfile, tempfile, shutil
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
from expand_rows import sheet_file_map, strip_calcchain   # noqa: E402
from fix_workbook import iter_tags   # (시작,끝,태그,내용) 4-튜플 판을 쓴다

TARGETS = {
    "02_돌발AS접수":       ("접수ID", ("접수일자", "작업완료일")),
    "04_정기점검":         ("점검ID", ("점검예정일", "실제점검일")),
    # 03·06도 재배치한다. ID를 새로 매기는 게 아니라 **행만 옮기므로**
    # 이 시트가 값으로 들고 있는 접수ID·원천업무ID는 그대로 따라간다.
    # (다른 시트가 03·06을 상대 행번호로 참조하지 않는 것을 확인함 — 2026-07-26)
    "03_현장작업실적":     ("작업ID", ("작업일자",)),
    "06_거래서류청구수금": ("정산ID", ("작업완료일",)),
}
ROW_RE = re.compile(r'<row r="(\d+)"', re.S)


def shift_rows(text, delta):
    """수식 안의 **상대** 행번호를 delta만큼 옮긴다($가 붙은 행은 고정)."""
    def rep(m):
        col, dollar, row = m.group(1), m.group(2), int(m.group(3))
        if dollar:                      # $5 → 절대행, 건드리지 않음
            return m.group(0)
        return f"{col}{row + delta}"
    return re.sub(r"(\$?[A-Z]{1,3})(\$?)(\d+)", rep, text)


def unshare(xml):
    """공유수식(<f t="shared" si="3"/>)을 **평범한 수식으로 펼친다**.

    왜 필요한가 — 공유수식은 '이 그룹은 연속한 행'이라는 전제를 깔고 있다(master가 맨 위,
    ref 범위 안에 follower가 전부 들어 있어야 한다). 행을 섞으면 그 전제가 깨지고
    엑셀은 파일을 아예 열지 않는다("Workbooks.Open 실패", 2026-07-26 v101 실사고).
    구조 검사(fix_workbook)로는 안 잡힌다 — XML은 멀쩡하고 의미만 깨지기 때문이다.
    그래서 재배치 전에 공유를 풀어 각 셀이 자기 수식을 갖게 만든다.
    """
    from fix_formulas import parse_cells, normalize, instantiate, esc, unesc
    cells = parse_cells(xml)
    master = {}                      # si → (행, 수식)
    for rn, col, f, s, e in cells:
        m = re.search(r'si="(\d+)"', xml[s:e])
        if m and f.strip() and 'ref="' in xml[s:e]:
            master[m.group(1)] = (rn, f)
    edits = []
    for rn, col, f, s, e in cells:
        chunk = xml[s:e]
        m = re.search(r'si="(\d+)"', chunk)
        if not m:
            continue
        si = m.group(1)
        body = f
        if not body.strip():                       # follower — master 수식을 이 행으로
            if si not in master:
                continue
            mr, mf = master[si]
            body = instantiate(normalize(mf, mr), rn)
        # t="shared"·si·ref 를 떼고 평범한 <f>로 만든다
        new_f = "<f>" + esc(body) + "</f>"
        fixed = re.sub(r"<f(?:\s[^>]*)?(?:/>|>.*?</f>)", new_f, chunk, count=1, flags=re.S)
        edits.append((s, e, fixed))
    if not edits:
        return xml, 0
    out, last = [], 0
    for s, e, fixed in sorted(edits):
        out.append(xml[last:s]); out.append(fixed); last = e
    out.append(xml[last:])
    return "".join(out), len(edits)


def move_row(row_xml, new_r):
    """<row> 하나를 new_r 위치로 옮긴 XML을 만든다."""
    old_r = int(ROW_RE.search(row_xml).group(1))
    if old_r == new_r:
        return row_xml
    d = new_r - old_r
    out = re.sub(r'(<row r=")\d+(")', lambda m: m.group(1) + str(new_r) + m.group(2), row_xml, count=1)
    # 셀 참조
    out = re.sub(r'(<c r="[A-Z]+)\d+(")', lambda m: m.group(1) + str(new_r) + m.group(2), out)
    # 수식 상대참조 — ★ 자기닫힘 <f .../> 를 게으른 정규식으로 잡으면 다음 셀까지 삼켜
    #   r="B5" 같은 셀 참조까지 숫자를 바꿔 버린다(실제로 B-3 이 만들어졌다). 2단계 파싱 필수.
    parts, last = [], 0
    for fs, fe, ftag, finner in iter_tags(out, "f"):
        # ★ 배열수식 <f t="array" ref="Q69">…</f> 의 ref는 **자기 셀 주소**다.
        #   행을 옮기면서 이걸 안 따라가면 ref가 엉뚱한 행을 가리키고,
        #   엑셀은 그 파일을 아예 열지 않는다(2026-07-26 v101~v104 실사고, 108개).
        #   구조 검사로는 안 잡힌다 — XML은 멀쩡하고 의미만 어긋나기 때문이다.
        tag = re.sub(r'(ref=")([A-Z]{1,3})(\d+)((?::[A-Z]{1,3})?)(\d*)(")',
                     lambda m: (m.group(1) + m.group(2) + str(int(m.group(3)) + d) + m.group(4)
                                + (str(int(m.group(5)) + d) if m.group(5) else "") + m.group(6)),
                     ftag)
        if finner is None:          # <f .../> 자기닫힘 → 본문 없음, 태그만 고친다
            if tag != ftag:
                parts.append(out[last:fs]); parts.append(tag); last = fe
            continue
        parts.append(out[last:fs])
        parts.append(tag + shift_rows(finner, d) + "</f>")
        last = fe
    parts.append(out[last:])
    return "".join(parts)


def sheet_rows(xml):
    """[(행번호, 원문)] — sheetData 안의 <row> 전부"""
    i = xml.index("<sheetData>") + len("<sheetData>")
    j = xml.index("</sheetData>")
    body = xml[i:j]
    out = []
    for s, e, tag, inner in iter_tags(body, "row"):
        seg = body[s:e]
        out.append((int(ROW_RE.search(seg).group(1)), seg))
    return out, (xml[:i], xml[j:])


def shared_strings(z):
    """xl/sharedStrings.xml → 인덱스별 문자열. 머리글이 공유문자열로 저장돼 있다."""
    if "xl/sharedStrings.xml" not in z.namelist():
        return []
    x = z.read("xl/sharedStrings.xml").decode("utf-8")
    out = []
    for s, e, tag, inner in iter_tags(x, "si"):
        out.append("".join(re.findall(r"<t[^>]*>(.*?)</t>", inner or "", re.S)).strip())
    return out


def col_index(xml, header_row=4, sst=None):
    """머리글 행에서 '열이름 → 열문자'. inlineStr·공유문자열 둘 다 지원."""
    rows, _ = sheet_rows(xml)
    hdr = next((seg for r, seg in rows if r == header_row), None)
    if not hdr:
        return {}
    out = {}
    for s, e, ctag, cin in iter_tags(hdr, "c"):
        ref = re.search(r'r="([A-Z]+)\d+"', ctag)
        if not ref or not cin:
            continue
        name = ""
        t = re.search(r"<t[^>]*>(.*?)</t>", cin, re.S)
        if t:
            name = re.sub(r"<[^>]+>", "", t.group(1)).strip()
        elif 't="s"' in ctag and sst:                 # 공유문자열 인덱스
            v = re.search(r"<v[^>]*>(\d+)</v>", cin)
            if v and int(v.group(1)) < len(sst):
                name = sst[int(v.group(1))]
        if name:
            out[name] = ref.group(1)
    return out


def cell_text(row_xml, col):
    m = re.search(r'<c r="%s\d+"[^>]*>(.*?)</c>' % col, row_xml, re.S)
    if not m:
        return ""
    inner = m.group(1)
    t = re.search(r"<t[^>]*>(.*?)</t>", inner, re.S)
    if t:
        return re.sub(r"<[^>]+>", "", t.group(1)).strip()
    v = re.search(r"<v[^>]*>(.*?)</v>", inner, re.S)
    return (v.group(1).strip() if v else "")


def serial_to_iso(v):
    """엑셀 날짜 일련번호 → YYYY-MM-DD (문자열 날짜면 그대로 정규화)"""
    s = str(v or "").strip()
    if re.fullmatch(r"\d+(\.\d+)?", s):
        from datetime import date, timedelta
        return (date(1899, 12, 30) + timedelta(days=int(float(s)))).isoformat()
    m = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    return "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3))) if m else ""


def plan(xml, date_cols, id_col, sst=None):
    """(정렬된 새 순서, 변경건수, ID/날짜 불일치 목록)"""
    rows, _ = sheet_rows(xml)
    head = [(r, seg) for r, seg in rows if r <= 4]
    data = [(r, seg) for r, seg in rows if r > 4]
    cols = col_index(xml, sst=sst)
    if not cols:
        raise SystemExit("머리글을 읽지 못했습니다 — 중단(잘못 정렬하면 데이터가 섞입니다)")
    dcs = [cols[c] for c in date_cols if c in cols]
    icol = cols.get(id_col)

    def key(item):
        r, seg = item
        for dc in dcs:
            d = serial_to_iso(cell_text(seg, dc))
            if d:
                return (0, d, r)
        return (1, "", r)          # 날짜 없는 행은 맨 뒤, 원래 순서 유지

    ordered = sorted(data, key=key)
    moved = sum(1 for i, (r, _) in enumerate(ordered) if r != 5 + i)
    bad = []
    for r, seg in ordered:
        i = cell_text(seg, icol) if icol else ""
        d = serial_to_iso(cell_text(seg, dcs[0])) if dcs else ""
        m = re.match(r"(?:AS|PM)-(\d{2})(\d{2})-", i)
        if m and d and (d[2:4], d[5:7]) != (m.group(1), m.group(2)):
            bad.append((i, d))
    return head, ordered, moved, bad


def move_hyperlinks(post, mapping):
    """<hyperlink ref="AM9"/> 도 그 행을 따라 옮긴다.

    링크는 셀에 붙어 있는데 sheetData 밖(꼬리)에 적혀 있어서, 행만 정렬하면
    링크가 제자리에 남아 **다른 건의 밴드 글로 연결된다**(2026-07-26 AM9 실사례:
    9행이 UJ2501886으로 바뀌었는데 링크는 원래 있던 건의 게시글 5243을 그대로 가리켰다).
    """
    n = [0]

    def rep(m):
        col, r = m.group(2), int(m.group(3))
        new = mapping.get(r)
        if not new or new == r:
            return m.group(0)
        n[0] += 1
        return f'{m.group(1)}"{col}{new}"'

    out = re.sub(r'(<hyperlink\b[^>]*?\sref=)"([A-Z]{1,3})(\d+)"', rep, post)
    return out, n[0]


def rebuild(xml, head, ordered):
    rows, (pre, post) = sheet_rows(xml)
    parts = [seg for _, seg in head]
    mapping = {}
    for i, (old_r, seg) in enumerate(ordered):
        mapping[old_r] = 5 + i
        parts.append(move_row(seg, 5 + i))
    post, moved = move_hyperlinks(post, mapping)
    if moved:
        print(f"   하이퍼링크 {moved}개도 함께 이동")
    return pre + "".join(parts) + post


def self_test():
    import openpyxl
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "syn.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "02_돌발AS접수"
    for _ in range(3):
        ws.append([])
    ws.append(["접수ID", "프로젝트NO", "캠프명", "접수일자", "누적"])
    data = [("AS-2607-001", "UJ1", "가", "2026-07-10"),
            ("AS-2604-002", "UJ2", "나", "2026-04-02"),
            ("AS-2606-003", "UJ3", "다", "2026-06-05")]
    for n, (i, p, c, d) in enumerate(data, start=5):
        ws.cell(row=n, column=1).value = i
        ws.cell(row=n, column=2).value = p
        ws.cell(row=n, column=3).value = c
        ws.cell(row=n, column=4).value = d
        ws.cell(row=n, column=5).value = f'=IF($B{n}="","",COUNT($E$4:E{n-1})+1)'
    wb.save(src)
    z = zipfile.ZipFile(src)
    sf = sheet_file_map(z)["02_돌발AS접수"]
    xml = z.read(sf).decode("utf-8"); z.close()
    head, ordered, moved, bad = plan(xml, ("접수일자",), "접수ID")
    out = rebuild(xml, head, ordered)
    got = [cell_text(seg, "A") for seg in re.findall(r"<row r=\"[5-9]\".*?</row>", out, re.S)]
    assert got == ["AS-2604-002", "AS-2606-003", "AS-2607-001"], got   # 과거 → 최근
    # 누적 카운터가 '직전 행'을 가리키도록 따라 움직였는가
    for n in (5, 6, 7):
        seg = re.search(r'<row r="%d".*?</row>' % n, out, re.S).group()
        f = re.search(r"<f[^>]*>(.*?)</f>", seg, re.S)
        assert f and f"E{n-1})" in f.group(1), (n, f.group(1) if f else None)
        assert f"$B{n}=" in f.group(1), (n, f.group(1))                # 자기 행 참조도 이동
        assert "$E$4" in f.group(1), (n, f.group(1))                   # 절대참조는 고정
    shutil.rmtree(tmp, ignore_errors=True)
    print("  재배치 self-test 통과 (정렬·자기행·직전행·절대참조)")
    return True


def main():
    args = sys.argv[1:]
    if "--apply" in args:
        from claim_guard import require
        require("ledger", "reorder_rows 반영")
    if "--self-test" in args:
        self_test(); return
    self_test()
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    only = args[args.index("--sheet") + 1] if "--sheet" in args else None
    zin = zipfile.ZipFile(master)
    smap = sheet_file_map(zin)
    sst = shared_strings(zin)
    patched, report, report_unshare, report_counter = {}, [], [], []
    for sh, (idc, dcs) in TARGETS.items():
        if only and sh != only:
            continue
        if sh not in smap:
            continue
        xml = zin.read(smap[sh]).decode("utf-8")
        # ★ 재배치 전에 공유수식을 반드시 푼다 — 안 풀면 엑셀이 파일을 못 연다
        xml, nun = unshare(xml)
        head, ordered, moved, bad = plan(xml, dcs, idc, sst)
        if nun:
            report_unshare.append((sh, nun))
        report.append((sh, len(ordered), moved, bad))
        # 공유를 푼 것만으로도 파일은 바뀐다 — 움직인 행이 없어도 저장해 둔다
        if moved or nun:
            from fix_formulas import fix_cumulative_counters
            rebuilt = rebuild(xml, head, ordered)
            rebuilt, ncounter = fix_cumulative_counters(rebuilt)
            if ncounter:
                report_counter.append((sh, ncounter))
            patched[smap[sh]] = rebuilt.encode("utf-8")
    zin.close()

    for sh, n in report_unshare:
        print(f"{sh}: 공유수식 {n}개 펼침(재배치 안전화)")
    for sh, n in report_counter:
        print(f"{sh}: 누적번호 수식 {n}개를 목적행 기준으로 정렬")
    for sh, n, moved, bad in report:
        print(f"{sh}: {n}행 중 {moved}행 위치 변경")
        if bad:
            print(f"  ※ ID의 연·월과 실제 날짜가 다른 건 {len(bad)}개 — 번호를 바꾸지 않으므로 그대로 둡니다")
            for i, d in bad[:5]:
                print(f"     {i} → 실제 {d}")
    if "--apply" not in args:
        print("\n미리보기 — 실제 적용: python reorder_rows.py --apply")
        return
    if not patched:
        print("바꿀 것 없음"); return
    m = re.search(r"_v(\d+)\.xlsx$", master)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(m.group(1))+1}.xlsx", master)
    if os.path.exists(dst):
        sys.exit(f"{os.path.basename(dst)} 이미 존재 — 중단")
    tmp = dst + ".tmp"
    if os.path.exists(tmp):
        sys.exit(f"{os.path.basename(tmp)} 임시파일이 이미 존재 — 중단")
    try:
        zin = zipfile.ZipFile(master)
        zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
        ct = None
        for it in zin.infolist():
            data = zin.read(it.filename)
            data, ct = strip_calcchain(it.filename, data, ct)
            if data is None:
                continue
            zout.writestr(it.filename, patched.get(it.filename, data))
        zout.close(); zin.close()

        from fix_workbook import check
        problems, fixable = check(tmp)
        if problems or fixable:
            raise RuntimeError(f"재배치 결과 구조검사 실패: {problems[:3]}")
        from fix_formulas import parse_cells, is_broken, direct_self_refs
        with zipfile.ZipFile(tmp) as audit:
            amap = sheet_file_map(audit)
            bad_formula, self_refs = [], []
            for sh, path in amap.items():
                sx = audit.read(path).decode("utf-8")
                bad_formula += [f"{sh}!{col}{rn}" for rn, col, f, _s, _e in parse_cells(sx)
                                if is_broken(f)]
                self_refs += [f"{sh}!{ref}" for ref in direct_self_refs(sx)]
        if bad_formula or self_refs:
            raise RuntimeError(
                f"재배치 수식검사 실패: 깨진수식 {bad_formula[:5]}, 순환참조 {self_refs[:5]}"
            )
        os.replace(tmp, dst)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise
    print(f"\n생성: {os.path.basename(dst)}")


if __name__ == "__main__":
    main()
