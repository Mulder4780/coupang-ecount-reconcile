# -*- coding: utf-8 -*-
"""
fill_erp_status.py — 판매조회(ERP)를 근거로 **ERP 등록 여부**를 채운다
================================================================================
사용자 지시(2026-07-28): "ERP등록 판정 진행해".

근거가 생겼다
  이카운트 **판매조회** 화면에만 `프로젝트코드(UJ번호)` 가 있다. 이게 ERP 매출과
  우리 프로젝트NO 를 잇는 유일한 자료다. 그동안 06시트 거래명세서번호로 더듬느라
  365건이 '미등록'으로 남아 있었는데, 이제 건별로 맞출 수 있다.

판정 규칙 (지어내지 않는다)
  판매조회에 그 프로젝트가 있다        → 완료
  없는데 무상·보험이다                → 해당없음 (청구가 없으니 매출 등록 대상이 아님)
  없고 유상이다                      → 미등록
  아직 작업완료가 아닌 건은 건드리지 않는다 (검증 수식도 작업완료일 때만 본다)

■ 검증결과는 수식이다 — 직접 쓰지 않는다. 이 칸이 채워지면 저절로 정상이 된다.
■ 유효성 목록(10_코드관리) 밖의 값은 쓰지 않는다: 완료 / 미등록 / 해당없음

사용
  python fill_erp_status.py            # 집계만
  python fill_erp_status.py --apply    # vN+1 생성
"""
import os
import re
import sys
import glob
import zipfile
from collections import Counter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path, esc  # noqa: E402
from source_dirs import ERP_DIR  # noqa: E402
from billing_fill import dedupe_files  # noqa: E402

HDR = 4
ALLOWED = ("완료", "미등록", "해당없음")
PRJ = re.compile(r"^UJ\d{7}$")
# (시트, 채울 열, 상태열, 완료로 치는 값, 비용열)
TARGET = [
    ("02_돌발AS접수", "ERP등록", "진행상태", ("작업완료",), "유상·무상·보험"),
    ("04_정기점검", "ERP판매전표", "점검상태", ("완료",), "유상·무상·보험"),
]


def _s(v):
    return "" if v is None else str(v).strip()


def erp_projects():
    """판매조회 파일에서 프로젝트코드를 모은다. 여러 개면 전부 합친다."""
    import openpyxl
    found, files = {}, []
    # ★ 2026-07-30: 판매조회 사본이 SHA256 동일한 3벌 있어 958KB 파일을 3번 열고 있었다.
    #   여기는 프로젝트코드 dict 라 값이 부풀지는 않지만 읽는 시간이 3배다.
    #   파일명(`__dup_`)이 아니라 **내용 해시**로 거른다 — 이름 규칙은 다음번에 또 뚫린다.
    for p in dedupe_files(glob.glob(os.path.join(ERP_DIR, "**", "*.xls*"), recursive=True)):
        if os.path.basename(p).startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            hdr, hrow = None, 0
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
                cells = [_s(c) for c in row]
                if any("프로젝트코드" in c for c in cells):
                    hdr, hrow = cells, i
                    break
            if not hdr:
                continue
            ci = next(i for i, h in enumerate(hdr) if "프로젝트코드" in h)
            si = next((i for i, h in enumerate(hdr) if h == "진행상태"), None)
            n = 0
            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                k = _s(row[ci]) if ci < len(row) else ""
                if not PRJ.match(k):
                    continue
                st = _s(row[si]) if si is not None and si < len(row) else ""
                # 같은 프로젝트가 여러 줄이면 더 진행된 상태를 남긴다
                if k not in found or st > found[k]:
                    found[k] = st
                n += 1
            if n:
                files.append((os.path.basename(p), sn, n))
        wb.close()
    return found, files


def plan(path):
    import openpyxl
    from openpyxl.utils import get_column_letter as GL
    erp, files = erp_projects()
    wb = openpyxl.load_workbook(path, data_only=True)
    fills, tally, skipped, notdone = {}, {}, Counter(), Counter()
    contra = []          # ERP에 매출이 있는데 우리 상태는 아직 미완료인 건
    for sn, col, stcol, done_vals, costcol in TARGET:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hdr = [_s(h) for h in next(ws.iter_rows(min_row=HDR, max_row=HDR, values_only=True))]
        if col not in hdr:
            continue
        L = GL(hdr.index(col) + 1)
        ix = {h: i for i, h in enumerate(hdr) if h}
        for n, row in enumerate(ws.iter_rows(min_row=HDR + 1, values_only=True), start=HDR + 1):
            if row[0] in (None, ""):
                continue
            get = lambda h: _s(row[ix[h]]) if h in ix and ix[h] < len(row) else ""
            k = get("프로젝트NO")
            if not k:
                continue
            if get(stcol) not in done_vals:
                if k not in erp:
                    notdone[sn] += 1
                    continue                   # 아직 안 끝난 건은 판단할 때가 아니다
                # ★ 2026-07-28: ERP 판매조회에 그 프로젝트가 **있다**는 건 작업이 끝나고
                #   전표까지 끊겼다는 뜻이다. 그런데 우리 상태 칸이 아직 '접수'·'미점검'이라면
                #   그건 **상태 갱신이 누락된 것**이지 ERP 등록이 안 된 게 아니다.
                #   ERP 칸은 사실대로 적고(전표는 실재한다), 모순은 따로 보고한다.
                #   (상태 칸은 수식이 잡고 있어 여기서 건드리지 않는다 — 완료일이 들어와야 바뀐다)
                contra.append((sn, n, k, get(stcol)))
            if get(col):
                skipped[sn] += 1
                continue                       # 사람이 적어 둔 값은 덮지 않는다
            if k in erp:
                v = "완료"
            elif get(costcol) in ("무상", "보험"):
                v = "해당없음"
            else:
                v = "미등록"
            fills[f"{sn}!{L}{n}"] = v
            tally.setdefault(sn, Counter())[v] += 1
    wb.close()
    for c in tally.values():
        for v in c:
            assert v in ALLOWED, f"허용되지 않은 값 {v!r}"
    return fills, tally, skipped, notdone, erp, files, contra


def apply_cells(xml, fills):
    seen = set()

    def repl(m):
        ref, attrs = m.group(1), m.group(2) or ""
        if ref not in fills or 't="shared"' in m.group(0) or 't="array"' in m.group(0):
            return m.group(0)
        seen.add(ref)
        a = re.sub(r'\s+t="[^"]*"', "", attrs)
        return f'<c r="{ref}"{a} t="inlineStr"><is><t>{esc(fills[ref])}</t></is></c>'

    return re.sub(r'<c r="([A-Z]+\d+)"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)',
                  repl, xml, flags=re.S), seen


def main():
    do = "--apply" in sys.argv
    m = latest_master()
    src = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src)}\n")

    fills, tally, skipped, notdone, erp, files, contra = plan(src)
    print(f"판매조회에서 모은 프로젝트 {len(erp)}개")
    for f, sn, n in files:
        print(f"   {f} [{sn}] {n}행")
    print(f"\n채울 칸 {len(fills)}개")
    for sn, c in tally.items():
        print(f"  {sn:<14}{sum(c.values()):>5}  " + " · ".join(f"{v} {k}" for k, v in c.most_common())
              + f"   (기존값 유지 {skipped.get(sn,0)} · 미완료 제외 {notdone.get(sn,0)})")
    if contra:
        # ERP 에 매출이 잡혔는데 우리 상태는 아직 미완료 — 상태 갱신이 누락된 것이다.
        # 상태 칸은 수식이라 여기서 못 고친다(완료일이 들어와야 바뀐다). 사람이 볼 목록으로 남긴다.
        print(f"\n★ ERP 에 전표가 있는데 관리대장 상태는 아직 미완료 — {len(contra)}건")
        for sn, n, k, st in contra[:10]:
            print(f"    {sn[:2]} {n}행  {k}  상태={st}")
        rp = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE, "reports")
        os.makedirs(rp, exist_ok=True)
        out = os.path.join(rp, "ERP전표있는데_상태미완료.md")
        with open(out, "w", encoding="utf-8") as fh:
            fh.write("# ERP 판매전표는 있는데 관리대장 상태가 미완료인 건\n\n")
            fh.write("판매조회에 프로젝트코드가 있다 = 작업이 끝나고 전표까지 끊겼다는 뜻이다.\n")
            fh.write("그런데 관리대장 상태가 '접수'·'미점검'이면 **상태 갱신이 누락**된 것이다.\n")
            fh.write("상태 칸은 수식이라 도구가 못 고친다 — **완료일(작업완료일·실제점검일)이 들어와야** 바뀐다.\n\n")
            fh.write("| 시트 | 행 | 프로젝트NO | 현재 상태 |\n|---|---:|---|---|\n")
            for sn, n, k, st in contra:
                fh.write(f"| {sn} | {n} | {k} | {st} |\n")
        print("    목록:", out)
    print("\n  ※ 검증결과는 수식이라 건드리지 않습니다 — 이 칸이 채워지면 저절로 정상이 됩니다.")

    # ★ 2026-07-30 추가: 여태 이 도구는 `--apply`(vN+1 직접 생성)밖에 없어서
    #   "대조는 도구가, 원장 쓰기는 ledger_writer 가 한 번에" 라는 파이프라인에서 혼자 벗어나 있었다.
    #   여러 도구가 각자 vN+1을 만들면 한쪽 작업이 통째로 묻힌다(AGENTS.md 동시작업 규칙).
    #   다른 fill 도구들과 같은 `--queue` 를 붙여 큐로만 넘긴다.
    if "--queue" in sys.argv:
        if not fills:
            print("\n적재할 항목 없음")
            return 0
        colname = {sn: col for sn, col, *_ in TARGET}
        items = []
        for ref, v in fills.items():
            sn, cell = ref.split("!", 1)
            row = int(re.search(r"\d+$", cell).group(0))
            items.append({
                "sheet": sn,
                # 셀 좌표로 직접 지정한다. key/key_col 은 큐 중복판정(queue_add)용 식별자다.
                "key": f"{sn}@{cell}",
                "key_col": "프로젝트NO",
                "cell": cell,
                "col": colname[sn],
                "value": v,
                "vtype": "text",
                "only_if_empty": True,
                "evidence": f"ERP 판매조회 프로젝트코드 대조 ({row}행)",
            })
        from ledger_writer import queue_add
        print(f"\n큐 적재: {queue_add(items)}개 셀 → ledger_db --intake 후 11:00·15:00 원장 반영")
        return 0

    if not do:
        print("\n실제로 채우려면:  python fill_erp_status.py --queue  (또는 --apply)")
        return 0

    by = {}
    for ref, v in fills.items():
        sn, r = ref.split("!", 1)
        by.setdefault(sn, {})[r] = v
    zin = zipfile.ZipFile(src)
    changed = {}
    for sn, f in by.items():
        sp = sheet_xml_path(zin, sn)
        new, seen = apply_cells(zin.read(sp).decode("utf-8"), f)
        miss = [r for r in f if r not in seen]
        if miss:
            print(f"★ {sn}: 칸이 없어 못 채운 곳 {len(miss)}개 {miss[:4]}")
        changed[sp] = new.encode("utf-8")
    from dashboard_clean import force_recalc
    changed["xl/workbook.xml"] = force_recalc(
        zin.read("xl/workbook.xml").decode("utf-8")).encode("utf-8")

    mm = re.search(r"_v(\d+)\.xlsx$", src)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm.group(1)) + 1}.xlsx", src)
    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        os.remove(tmp)
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close(); zin.close()
    try:
        verify(src, tmp, by, set(changed))
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise
    os.replace(tmp, dst)
    print(f"\n{len(fills)}칸 채움 → {os.path.basename(dst)}")
    print("엑셀에서 한 번 열어 주세요 — 검증결과가 다시 계산됩니다.")
    return 0


def verify(src, out, by, allowed):
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"
    wa = openpyxl.load_workbook(src, data_only=True)
    wb_ = openpyxl.load_workbook(out, data_only=True)
    for sn, f in by.items():
        sa, sb = wa[sn], wb_[sn]
        for ref, v in f.items():
            assert _s(sb[ref].value) == v, f"{sn} {ref} 반영 실패"
        diff = []
        for r in range(1, max(sa.max_row, sb.max_row) + 1):
            for c in range(1, max(sa.max_column, sb.max_column) + 1):
                ca = sa.cell(r, c)
                if ca.coordinate in f:
                    continue
                if ca.value != sb.cell(r, c).value:
                    diff.append(f"{sn}!{ca.coordinate}")
                    if len(diff) > 5:
                        break
        assert not diff, f"의도치 않은 셀 변경: {diff}"
    wf = openpyxl.load_workbook(out, data_only=False)
    for sn, refs in (("02_돌발AS접수", ("AK5", "AN5")), ("04_정기점검", ("AB5", "AD5"))):
        for ref in refs:
            assert str(wf[sn][ref].value or "").startswith("="), f"{sn} {ref} 수식이 사라졌다"
    wa.close(); wb_.close()
    zs = zipfile.ZipFile(src)
    other = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not other, f"의도치 않은 파트 변경: {other}"
    zs.close(); z.close()
    print("  검증 통과 — 수식 보존 · 다른 셀·파트 변경 없음")


if __name__ == "__main__":
    sys.exit(main())
