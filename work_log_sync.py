# -*- coding: utf-8 -*-
"""정기점검·돌발AS 일지(미실시건)를 관리대장·대표보고에 대조 반영한다.

사용자가 계속 갱신하는 원본은 ``0. 원본 자료/8. 정기점검, 돌발AS 일지(미실시건)``
아래의 최신 xlsx다. 이 파일은 완료 실적과 미실시 사유를 모두 가진 현장 원본이라,
원장에 추측으로 날짜·상태를 덮어쓰지 않고 다음 세 갈래로 처리한다.

1) 프로젝트NO가 정확히 하나로 매칭되고 원장 상태가 비어 있으면 확정값만 입력 큐에 적재
2) 기존 값과 다르거나 완료일이 없는 ``처리완료``는 28_일지대조현황에 근거 그대로 보관
3) 대표보고 API는 이 모듈의 동일 분석 결과를 써서 발생·처리·미처리·사유를 함께 표시

실행:
  python work_log_sync.py             # 대조 결과만 출력·reports JSON 갱신
  python work_log_sync.py --queue     # 안전한 빈 칸 보완만 입력 큐에 적재
  python work_log_sync.py --apply     # 28_일지대조현황 시트 반영(vN+1)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.utils.datetime import from_excel

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from ecount_reconcile import load_config, resolve_master
from findings_sheet import build_generic_sheet, upsert
from source_dirs import WORK_LOG_DIR, work_log_dirs
from zscan import camp_key

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

APP_YEAR = "2026"
SHEET_NAME = "28_정기점검돌발AS일지대조"
HEADERS = [
    "구분", "원본시트", "원본행", "프로젝트NO", "캠프명", "원본일자", "원본상태",
    "미처리사유", "담당자", "요청내용", "실제조치", "원장매칭", "원장상태", "대조결과", "원본파일",
]
WIDTHS = [12, 24, 10, 16, 28, 14, 18, 44, 20, 44, 48, 15, 20, 24, 36]
PROJECT_RE = re.compile(r"\bUJ\s*[-_]?\s*(26\d{5})\b", re.I)


def _s(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (float, int)) and 20_000 <= float(value) <= 80_000:
        try:
            return from_excel(value).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            pass
    text = _s(value)
    match = re.search(r"(20\d{2})\s*[./년-]\s*(\d{1,2})\s*[./월-]\s*(\d{1,2})", text)
    if not match:
        return ""
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
    except ValueError:
        return ""


def _project(value) -> str:
    match = PROJECT_RE.search(_s(value))
    return "UJ" + match.group(1) if match else ""


def _header_row(ws) -> tuple[int, dict[str, int]]:
    for rn in range(1, min(ws.max_row, 16) + 1):
        values = [_s(x) for x in next(ws.iter_rows(min_row=rn, max_row=rn, values_only=True))]
        if "프로젝트NO" in values:
            return rn, {v: i for i, v in enumerate(values) if v}
    raise ValueError(f"{ws.title}: 프로젝트NO 머리글을 찾지 못했습니다.")


def _value(row: tuple, header: dict[str, int], *names: str):
    for name in names:
        idx = header.get(name)
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _reason_blocker(reason: str) -> str:
    """미실시 사유를 **책임 소재**로 가른다 (대표 지시 2026-08-04).

    대표 기준: "부품이 안 와서·합의된 일정이라 못 한 건 괜찮다. 그런데 **기사 스케줄을
    못 잡아 미루는 건 절대 안 된다.**" 그래서 건수만 세지 말고 둘을 갈라 보여 준다.

    핵심은 순서다 — "모터가 며칠 전에 도착, 방문 일정 조율 중"은 **자재 대기가 아니라
    스케줄 문제**다(물건은 이미 왔다). 기존 분류는 '모터'라는 단어만 보고 자재 대기로
    넣어 이런 건을 정당한 사유처럼 보이게 했다.
    """
    t = _s(reason)
    if not t:
        return "사유 미기재"                       # 사유가 없는 것도 확인 대상이다
    arrived = re.search(r"도착|입고|수령|왔|전일 도착", t)
    waiting = re.search(r"발주|주문|미도착|대기|수급|발송 예정|제작", t)
    scheduling = re.search(r"일정\s*조율|조율\s*중|조율\s*예정|일정조율|스케줄|"
                           r"조만간|추후|차주|다음주|이번주|예정임|바빠|여유", t)
    # ① 물건은 왔는데 일정만 미루는 중 → 스케줄 문제(핫이슈)
    if arrived and scheduling and not waiting:
        return "확인 필요(일정 미확정)"
    # ② 정기점검 등 **합의된 시점**에 함께 처리 → 정당
    if re.search(r"정기점검\s*(때|시)|정기점검에|점검일에|합의|협의 완료|"
                 r"운행에 지장이 없", t):
        return "정당(합의·차기 점검)"
    # ③ 부품·자재가 **아직 안 온 것**만 정당. 단순히 "○○교체예정"처럼 물건 이야기 없이
    #    '예정'만 있는 건 대표가 콕 집은 유형이다("조만간 교체 예정 — 이런 게 있으면 안 돼").
    if waiting:
        return "정당(자재·부품 대기)"
    if re.search(r"자재|모터|리모컨|부품", t) and not arrived:
        if re.search(r"예정|교체할|하기로", t) and not re.search(r"발주|주문|미도착|입고 예정", t):
            return "확인 필요(일정 미확정)"
        return "정당(자재·부품 대기)"
    # ④ 캠프·고객 사정 → 정당
    if re.search(r"캠프|고객|현장 사정|출입|휴무|폐쇄|공사", t):
        return "정당(현장 사정)"
    # ⑤ 날짜가 박혀 있으면 계획이 있는 것으로 본다
    if re.search(r"\d{1,2}\s*[/월]\s*\d{1,2}", t):
        return "정당(일정 확정)"
    if scheduling:
        return "확인 필요(일정 미확정)"
    return "확인 필요(사유 불명확)"


def is_hot_issue(reason: str) -> bool:
    """대표가 '있으면 안 된다'고 한 유형인가 — 스케줄 미확정·사유 불명확."""
    return _reason_blocker(reason).startswith(("확인 필요", "사유 미기재"))


def _reason_group(reason: str) -> str:
    t = _s(reason)
    if re.search(r"자재|모터|리모컨|부품|발송|도착|수급", t):
        return "자재·부품 대기"
    if re.search(r"일정|조율|방문|다음주|이번주|8월|9월", t):
        return "방문 일정 조율"
    if re.search(r"정기점검", t):
        return "정기점검 이관"
    if re.search(r"인력|단체|설치팀|상의|협의", t):
        return "인력·협의 필요"
    if re.search(r"취소|정상작동", t):
        return "취소·정상작동"
    return "기타 사유"


def _as_state(value: str) -> str:
    value = _s(value)
    if "취소" in value:
        return "취소"
    if "완료" in value or "수리완료" in value:
        return "완료"
    if "예정" in value:
        return "예정"
    return "미실시"


def _read_sheet(ws, kind: str) -> list[dict]:
    header_row, header = _header_row(ws)
    records = []
    for rn, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        project = _project(_value(row, header, "프로젝트NO"))
        if not project:
            continue  # 장비 추가행은 상위 프로젝트 행에 이미 묶여 있다.
        camp = _s(_value(row, header, "캠프이름", "캠프명"))
        when = _date(_value(row, header, "점검일자", "A/S요청일자", "요청일자"))
        tech = _s(_value(row, header, "A/S담당", "원담당", "담당기사"))
        request = _s(_value(row, header, "A/S 요청", "A/S 신청내용", "점검내용"))
        action = _s(_value(row, header, "A/S내용"))
        status = _s(_value(row, header, "진행현황"))
        reason = _s(_value(row, header, "미처리 사유"))
        note = _s(_value(row, header, "비고", "특이사항"))
        if kind == "pm":
            state = "실행" if when else "일자확인필요"
        elif kind == "as_done":
            state = _as_state(status or "수리완료")
        else:
            state = _as_state(status)
        records.append({
            "구분": "정기점검" if kind == "pm" else "돌발AS",
            "종류": kind,
            "원본시트": ws.title,
            "원본행": rn,
            "프로젝트NO": project,
            "캠프명": camp,
            "일자": when,
            "원본상태": status or state,
            "상태": state,
            "미처리사유": reason,
            "사유분류": _reason_group(reason) if reason else "",
            # 대표 지시(2026-08-04): 건수보다 **왜 못 했는지의 성격**이 중요하다.
            "책임구분": _reason_blocker(reason),
            "핫이슈": is_hot_issue(reason),
            "담당자": tech,
            "요청내용": request,
            "실제조치": action,
            "비고": note,
        })
    return records


def find_latest_source(folder: str | None = None) -> str:
    folders = [folder] if folder else work_log_dirs()
    candidates = []
    for base in folders:
        if not base or not os.path.isdir(base):
            continue
        for parent, _dirs, files in os.walk(base):
            for name in files:
                if name.startswith("~$") or not name.lower().endswith(".xlsx"):
                    continue
                if not ("정기점검" in name and "돌발" in name):
                    continue
                path = os.path.join(parent, name)
                try:
                    candidates.append((os.path.getmtime(path), os.path.getsize(path), path))
                except OSError:
                    pass
        # 정본 폴더에 하나라도 있으면 이전 호환 폴더보다 우선한다.
        if candidates and folder is None and os.path.normcase(os.path.abspath(base)) == os.path.normcase(os.path.abspath(WORK_LOG_DIR)):
            break
    if not candidates:
        raise FileNotFoundError(f"정기점검·돌발AS 일지 xlsx가 없습니다: {folders}")
    return max(candidates)[2]


def stable_snapshot(source: str, temp_dir: str) -> tuple[str, str]:
    before = os.stat(source)
    dst = os.path.join(temp_dir, os.path.basename(source))
    shutil.copy2(source, dst)
    after = os.stat(source)
    if (before.st_size, before.st_mtime_ns) != (after.st_size, after.st_mtime_ns):
        raise RuntimeError("일지 원본이 저장 중이라 이번 대조를 건너뜁니다. 다음 실행에서 재시도합니다.")
    digest = hashlib.sha256(open(dst, "rb").read()).hexdigest()
    return dst, digest


def read_journal(source: str) -> list[dict]:
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        records = []
        for ws in wb.worksheets:
            title = _s(ws.title)
            if "정기점검 일지" in title:
                records.extend(_read_sheet(ws, "pm"))
            elif "돌발AS 미실시" in title:
                records.extend(_read_sheet(ws, "as_open"))
            elif "돌발AS 일지" in title:
                records.extend(_read_sheet(ws, "as_done"))
        return [r for r in records if r["프로젝트NO"].startswith("UJ26")]
    finally:
        wb.close()


def read_master(master: str) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    try:
        out: dict[str, list[dict]] = {"as": [], "pm": []}
        for sheet, key, id_col, state_col, done_col in (
            ("02_돌발AS접수", "as", "접수ID", "진행상태", "작업완료일"),
            ("04_정기점검", "pm", "점검ID", "점검상태", "실제점검일"),
        ):
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = [_s(v) for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
            ix = {h: i for i, h in enumerate(headers) if h}
            for rn, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
                project = _project(_value(row, ix, "프로젝트NO"))
                if not project:
                    continue
                out[key].append({
                    "sheet": sheet, "row": rn, "프로젝트NO": project,
                    "ID": _s(_value(row, ix, id_col)), "캠프명": _s(_value(row, ix, "캠프명")),
                    "상태": _s(_value(row, ix, state_col)), "완료일": _date(_value(row, ix, done_col)),
                    "접수일": _date(_value(row, ix, "접수일자", "점검예정일")),
                    "상태열": state_col, "완료열": done_col, "키열": id_col,
                })
        return out
    finally:
        wb.close()


def analyze(master: str, source: str | None = None) -> dict:
    source = source or find_latest_source()
    records = read_journal(source)
    master_rows = read_master(master)
    index = {
        "as": defaultdict(list),
        "pm": defaultdict(list),
    }
    for key, rows in master_rows.items():
        for row in rows:
            index[key][row["프로젝트NO"]].append(row)

    updates, compared = [], []
    for r in records:
        key = "pm" if r["종류"] == "pm" else "as"
        hits = index[key].get(r["프로젝트NO"], [])
        x = dict(r)
        x["원장매칭"] = len(hits)
        x["원장상태"] = ""
        x["대조결과"] = ""
        if not hits:
            x["대조결과"] = "원장 미매칭"
        elif len(hits) > 1:
            x["대조결과"] = "원장 프로젝트 중복 — 자동입력 제외"
            x["원장상태"] = " · ".join(sorted({_s(h["상태"]) for h in hits if h["상태"]}))
        else:
            hit = hits[0]
            x["원장상태"] = " · ".join(v for v in (hit["상태"], hit["완료일"]) if v)
            source_done = (r["종류"] == "pm" and r["상태"] == "실행") or (r["종류"] == "as_done" and r["상태"] == "완료")
            if source_done and hit["완료일"] and r["일자"] and hit["완료일"] == r["일자"]:
                x["대조결과"] = "완료일 일치"
            elif source_done and hit["완료일"]:
                x["대조결과"] = "완료일 상이 — 원본 확인 필요"
            elif source_done:
                x["대조결과"] = "완료 근거 있음 · 원장 상태 보완 대기"
                # 완전히 빈 상태일 때만 자동 채운다. 이미 접수·진행중처럼 사람이 쓴 상태는
                # 같은 프로젝트의 재방문일 수 있으므로 28시트에 남겨 사람이 확인한다.
                if not hit["상태"] and r["일자"]:
                    updates.extend([
                        {"sheet": hit["sheet"], "key_col": hit["키열"], "key": hit["ID"],
                         "col": hit["상태열"], "value": "완료" if key == "pm" else "작업완료",
                         "vtype": "text", "evidence": f"일지 {os.path.basename(source)} {r['원본시트']} {r['원본행']}행", "only_if_empty": True},
                        {"sheet": hit["sheet"], "key_col": hit["키열"], "key": hit["ID"],
                         "col": hit["완료열"], "value": r["일자"], "vtype": "date",
                         "evidence": f"일지 {os.path.basename(source)} {r['원본시트']} {r['원본행']}행", "only_if_empty": True},
                    ])
            elif r["종류"] == "as_open":
                x["대조결과"] = "미실시 사유 반영" if r["상태"] in ("미실시", "예정") else "취소·완료일 확인 필요"
            else:
                x["대조결과"] = "원장 대조 완료"
        compared.append(x)

    as_rows = [r for r in compared if r["구분"] == "돌발AS"]
    pm_rows = [r for r in compared if r["구분"] == "정기점검"]
    as_done = [r for r in as_rows if r["종류"] == "as_done" and r["상태"] == "완료"]
    as_open = [r for r in as_rows if r["종류"] == "as_open" and r["상태"] in ("미실시", "예정")]
    as_cancel = [r for r in as_rows if r["종류"] == "as_open" and r["상태"] == "취소"]
    as_done_unknown = [r for r in as_rows if r["종류"] == "as_open" and r["상태"] == "완료"]
    as_dates = sorted({r["일자"] for r in as_rows if r.get("일자")})
    reason_counts = Counter(r["사유분류"] for r in as_open if r["사유분류"])
    reasons = [{"사유": key, "건수": count} for key, count in reason_counts.most_common()]
    # 대표 지시(2026-08-04): 미실시는 **정당한 사유가 있는 것과 스케줄을 못 잡아
    # 미루는 것**을 갈라 봐야 한다. 뒤엣것이 곧 '있으면 안 되는' 건이다.
    hot = [r for r in as_open if r.get("핫이슈")]
    blocker_counts = Counter(r.get("책임구분") or "사유 미기재" for r in as_open)
    blockers = [{"구분": k, "건수": v} for k, v in blocker_counts.most_common()]
    unmatched = sum(r["대조결과"] == "원장 미매칭" for r in compared)
    conflicts = sum("상이" in r["대조결과"] or "중복" in r["대조결과"] for r in compared)
    return {
        "ok": True,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "source_mtime": os.path.getmtime(source),
        "source_sha256": hashlib.sha256(open(source, "rb").read()).hexdigest(),
        "records": compared,
        "updates": updates,
        "요약": {
            "돌발AS": {
                "발생": len(as_rows), "처리완료": len(as_done), "미처리": len(as_open),
                "취소": len(as_cancel), "처리완료일확인": len(as_done_unknown),
                "기준시작일": as_dates[0] if as_dates else "",
                "기준종료일": as_dates[-1] if as_dates else "",
                "미처리사유": reasons, "미처리목록": as_open,
                "핫이슈": len(hot), "핫이슈목록": hot, "책임구분": blockers,
                "처리완료목록": as_done, "취소목록": as_cancel,
                "처리완료일확인목록": as_done_unknown,
            },
            "정기점검": {"실행": len(pm_rows), "원장미매칭": sum(r["대조결과"] == "원장 미매칭" for r in pm_rows)},
            "대조": {"전체": len(compared), "원장미매칭": unmatched, "충돌확인": conflicts,
                     "안전자동입력후보": len(updates)},
        },
    }


def sheet_rows(records: list[dict], source_name: str) -> list[tuple]:
    return [tuple([
        r["구분"], r["원본시트"], r["원본행"], r["프로젝트NO"], r["캠프명"], r["일자"], r["원본상태"],
        r["미처리사유"], r["담당자"], r["요청내용"], r["실제조치"], r["원장매칭"], r["원장상태"],
        r["대조결과"], source_name,
    ]) for r in records]


def build_sheet(records: list[dict], source_name: str) -> str:
    return build_generic_sheet(
        SHEET_NAME, HEADERS, WIDTHS, sheet_rows(records, source_name),
        "[자동대조] 정기점검·돌발AS 일지(미실시건) 원본을 프로젝트NO로 원장과 대조했습니다. "
        "완료일이 명확한 사실만 자동입력 후보로 만들며, 미실시·취소·일자상이는 근거와 사유를 그대로 남깁니다.",
        empty_text="대조 대상 없음",
    )


def write_report(payload: dict, path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source")
    ap.add_argument("--source-dir")
    ap.add_argument("--master")
    ap.add_argument("--queue", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--report", default=os.path.join(ROOT, "reports", "work_log_compare.json"))
    args = ap.parse_args(argv)
    master = args.master or resolve_master(load_config()["reconcile"]["master_xlsx"])
    source = args.source or find_latest_source(args.source_dir)
    with tempfile.TemporaryDirectory(prefix="work_log_") as td:
        snapshot, _digest = stable_snapshot(source, td)
        payload = analyze(master, snapshot)
    payload["source"] = source
    write_report(payload, args.report)
    summary = payload["요약"]
    a, p, c = summary["돌발AS"], summary["정기점검"], summary["대조"]
    print(f"일지 원본: {os.path.basename(source)}")
    print(f"  돌발AS 발생 {a['발생']} · 처리완료 {a['처리완료']} · 미처리 {a['미처리']} · 취소 {a['취소']} · 완료일확인 {a['처리완료일확인']}")
    print(f"  정기점검 실행 {p['실행']} · 원장 미매칭 {p['원장미매칭']} · 전체 미매칭 {c['원장미매칭']} · 충돌확인 {c['충돌확인']}")
    if a["미처리사유"]:
        print("  미처리 사유:", " · ".join(f"{x['사유']} {x['건수']}" for x in a["미처리사유"]))
    if args.queue and payload["updates"]:
        from ledger_writer import queue_add
        print("  안전 자동입력 큐:", queue_add(payload["updates"]), "셀")
    elif args.queue:
        print("  안전 자동입력 후보 없음")
    if args.apply:
        xml = build_sheet(payload["records"], os.path.basename(source))
        dst, msg = upsert(master, xml, sheet_name=SHEET_NAME, headers=HEADERS)
        print(f"  {SHEET_NAME}: {msg}")
        if dst:
            print("   ", dst)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
