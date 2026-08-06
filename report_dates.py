# -*- coding: utf-8 -*-
"""
report_dates.py — 00_대시보드의 **보고일·집계기준일을 날마다 저절로 오늘 것으로** 맞춘다

사용자 지시(2026-08-06): "보고일이 오늘 날짜로 자동 변경되는 알고리즘 구성해."

왜 필요한가
  B3(보고일)·B4(집계기준일)은 대표보고·캡처·화면이 모두 기준으로 삼는 값인데, 지금까지
  **사람이 앱에서 눌러 바꾸는 값**이었다. 안 바꾸면 어제 날짜로 오늘 보고가 나간다.
  날마다 반드시 하는 일이면 사람이 기억할 일이 아니다(프로젝트 상시 원칙).

규칙
  · 보고일 = 오늘, 집계기준일 = 전날. (보고는 다음 날 아침에 하므로 이 조합이 기본이다)
  · **이미 그 값이면 큐에 넣지 않는다.** 매일 같은 값을 다시 넣으면 반영 회차마다
    쓸데없이 vN+1 이 늘어난다.
  · 엑셀에 직접 쓰지 않는다 — `ledger_db` 큐에 넣고 11:00·15:00 회차가 기록한다.
    (사람이 앱에서 「저장하고 반영」을 누르면 그때 즉시 반영된다)
  · 사람이 과거 날짜로 일부러 바꿔 둔 경우를 덮지 않으려면 `--only-if-stale` 를 쓴다:
    B3 가 **오늘보다 과거일 때만** 갱신한다(미래로 맞춰 둔 것은 건드리지 않는다).

  python report_dates.py            # 오늘/전날로 큐에 넣는다(이미 같으면 아무것도 안 함)
  python report_dates.py --print    # 지금 원장에 무엇이 들어 있는지만 본다
"""
import argparse
import os
import sys
from datetime import datetime, timedelta

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SHEET = "00_대시보드"
CELLS = (("B3", "보고일"), ("B4", "집계기준일"))


def _norm(v):
    """엑셀이 날짜를 datetime 으로 주기도 하고 문자열로 주기도 한다 — 한 모양으로."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    if not s:
        return ""
    import re
    m = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    return "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3))) if m else s


def current():
    """원장의 지금 값(B3·B4). 못 읽으면 빈 값 — 그래도 갱신은 진행한다."""
    try:
        import openpyxl
        from ecount_reconcile import load_config, resolve_master
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
        wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
        ws = wb[SHEET] if SHEET in wb.sheetnames else None
        out = {c: _norm(ws[c].value) for c, _n in CELLS} if ws else {}
        wb.close()
        return out
    except Exception:
        return {}


def wanted(today=None):
    d = today or datetime.now().date()
    return {"B3": d.isoformat(), "B4": (d - timedelta(days=1)).isoformat()}


def run(only_if_stale=False, today=None):
    want, now = wanted(today), current()
    items = []
    for cell, label in CELLS:
        v = want[cell]
        if now.get(cell) == v:
            continue                      # 이미 맞다 — 큐를 늘리지 않는다
        if only_if_stale and now.get(cell) and now[cell] > v:
            continue                      # 사람이 앞날로 맞춰 둔 것은 건드리지 않는다
        items.append({"sheet": SHEET, "cell": cell, "key": cell, "key_col": "-",
                      "col": label, "value": v, "vtype": "date",
                      "evidence": "보고일 자동 갱신(report_dates)", "only_if_empty": False})
    if not items:
        print("보고일·집계기준일 이미 최신 — 큐에 넣지 않았습니다 (%s / %s)"
              % (now.get("B3", "?"), now.get("B4", "?")))
        return 0
    try:
        import ledger_db
        ledger_db.enqueue(items, source="report-dates")
    except Exception as exc:
        print("큐 적재 실패: %s" % exc)
        return 1
    print("보고일 %s · 집계기준일 %s 로 갱신 예약 (%d칸) — 다음 반영 회차에 기록됩니다"
          % (want["B3"], want["B4"], len(items)))
    return 0


def main():
    ap = argparse.ArgumentParser(description="보고일·집계기준일을 오늘/전날로 자동 갱신")
    ap.add_argument("--print", dest="show", action="store_true", help="지금 값만 본다")
    ap.add_argument("--only-if-stale", action="store_true",
                    help="원장 값이 오늘보다 과거일 때만 갱신(앞날 지정은 존중)")
    a = ap.parse_args()
    if a.show:
        now, want = current(), wanted()
        print("원장  보고일 %s · 집계기준일 %s" % (now.get("B3", "?"), now.get("B4", "?")))
        print("기대  보고일 %s · 집계기준일 %s" % (want["B3"], want["B4"]))
        return 0
    return run(only_if_stale=a.only_if_stale)


if __name__ == "__main__":
    sys.exit(main())
