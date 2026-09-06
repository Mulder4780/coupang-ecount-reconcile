# -*- coding: utf-8 -*-
"""밴드·ERP·거래명세서 근거를 02·03·04 검증 상태에 안전하게 동기화한다.

원칙
----
* 프로젝트NO가 정확히 일치하고 확인일이 있는 2026년 근거만 사용한다.
* 관리대장의 검증결과 수식 열은 직접 쓰지 않는다. 원인 열을 채워 Excel 수식이
  ``정상`` 또는 ``확인``으로 재계산되게 한다.
* 03시트의 접수ID·프로젝트NO는 02시트 완료행을 순서대로 끌어오는 배열수식이다.
  Excel 캐시가 비어 있어도 같은 조건을 재현하되, 실제 작업내용이 있는 행만 쓴다.
* 기존 ``확인필요·미반영·누락``은 확정 근거가 생겼을 때만 고친다. 수식과 사람이
  입력한 작업내용은 건드리지 않는다.
"""
from __future__ import annotations

import glob
import json
import os
import re
import sys
from collections import Counter
from datetime import date, datetime

from billing_fill import dedupe_files
from operation_window import input_window_label, is_input_window
from source_dirs import ERP_DIR
from workbook_patch import latest_master

ROOT = os.path.dirname(os.path.abspath(__file__))
ERP_CACHE = os.path.join(ROOT, "reports", "erp_sales_evidence_cache.json")
# 캐시 형식 버전. 내용해시 중복제거를 넣기 전(v1) 캐시는 같은 판매조회를 3벌 읽은
# 결과(파일 목록 3줄)를 담고 있어 그대로 재사용하면 고친 효과가 보이지 않는다.
# 버전이 다르면 캐시를 버리고 다시 읽는다.
ERP_CACHE_VERSION = 2
PROJECT_RE = re.compile(r"^UJ26\d{5}$", re.I)
HDR_ROW = 4
FIRST = 5


def _s(value):
    return "" if value is None else str(value).strip()


def _date_iso(value):
    """Excel 날짜·문자열을 YYYY-MM-DD로 정규화하고 2026년만 허용한다."""
    if isinstance(value, (datetime, date)):
        out = value.strftime("%Y-%m-%d")
    else:
        m = re.search(r"(?<!\d)(20\d{2})[-./년]\s*(\d{1,2})[-./월]\s*(\d{1,2})", _s(value))
        if not m:
            return ""
        try:
            out = date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ""
    return out if out.startswith("2026-") else ""


def _headers(ws):
    row = next(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))
    return {_s(value): i for i, value in enumerate(row) if _s(value)}


def _value(row, index, name):
    i = index.get(name)
    return _s(row[i]) if i is not None and i < len(row) else ""


def _profile(evidence, project):
    return evidence.setdefault(project, {
        "statement": False,
        "erp": False,
        "completed": False,
        "photos": 0,
        "statement_dates": set(),
        "erp_dates": set(),
        "completion_dates": set(),
        "sources": set(),
    })


def band_document_evidence(records=None):
    """완료된 밴드·카톡 보고의 문서상태와 게시/작업일을 프로젝트별로 모은다."""
    if records is None:
        from band_extract import load_records
        records = load_records()
    evidence = {}
    for row in records:
        project = _s(row.get("프로젝트NO")).upper()
        if not PROJECT_RE.fullmatch(project) or _s(row.get("진행상태")) != "작업완료":
            continue
        docs = {_s(x) for x in re.split(r"[+,/]", _s(row.get("문서상태"))) if _s(x)}
        when = _date_iso(row.get("게시일")) or _date_iso(row.get("작업일"))
        profile = _profile(evidence, project)
        profile["completed"] = True
        try:
            profile["photos"] = max(profile["photos"], int(row.get("사진") or 0))
        except (TypeError, ValueError):
            pass
        if when:
            profile["completion_dates"].add(when)
        if "거래명세서" in docs:
            profile["statement"] = True
            if when:
                profile["statement_dates"].add(when)
        if "판매전표" in docs:
            profile["erp"] = True
            if when:
                profile["erp_dates"].add(when)
        profile["sources"].add("밴드·카톡 완료보고")
    return evidence


def erp_sales_evidence(root=ERP_DIR):
    """ERP 판매조회 내보내기에서 프로젝트NO와 전표 확인일을 읽는다.

    ★ 2026-09-06 실측 - 이 함수가 09:50 회차 그 단계(600초 제한)의 시간을
      거의 다 쓴다. 캐시가 **없을 때 2,312.9초** · **있을 때 178.4초**(13배).
    ★ 그런데 그 178초는 일한 시간이 아니라 **캐시가 아직 맞나 물어본 시간**이다 -
      아래 for 문이 signature 298개를 `os.stat(path)` 로 하나씩 확인하는데
      Z:(SMB) 는 파일 하나가 왕복 한 번이라 **개당 614ms** 다([198]).
      `[409]`·`[441]` 이 같은 병을 세 번 고쳤다 - 목록을 받을 때 딸려 오는
      stat 을 버리지 말고 **폴더를 한 번 scandir** 하면 그 왕복이 사라진다.
    ⚠ 고칠 때 **결과가 한 톨도 바뀌면 안 된다**([441]) - 이 값은 원장에 쓰는
      큐를 만든다. 목록·순서를 앞뒤로 대 본다. 분담판 [385].
    ⚠ 콜드 2,313초는 **캐시가 없거나 무효일 때**다. 매일 죽는다면 캐시가 매일
      무효라는 뜻인데 **왜인지는 아직 안 쟀다**([169]) - root_mtime_ns 는
      2026-08-05 그대로였고 signature 298개도 전부 맞았다.
    """
    import openpyxl

    evidence, files = {}, []
    use_cache = (os.path.normcase(os.path.abspath(root))
                 == os.path.normcase(os.path.abspath(ERP_DIR))
                 and "--refresh-erp" not in sys.argv)
    if use_cache:
        try:
            cached = json.load(open(ERP_CACHE, encoding="utf-8"))
            known = cached.get("signature") or []
            root_now = os.stat(root).st_mtime_ns
            root_saved = cached.get("root_mtime_ns")
            same_known = bool(known) and cached.get("version") == ERP_CACHE_VERSION
            for path, size, mtime_ns in known:
                try:
                    st = os.stat(path)
                    if st.st_size != size or st.st_mtime_ns != mtime_ns:
                        same_known = False
                        break
                except OSError:
                    same_known = False
                    break
            # 이전 캐시에는 root_mtime_ns가 없을 수 있다. 이미 전체 스캔 뒤 생성된
            # 캐시이므로 현재값을 1회 보강하고, 다음부터 새 파일/폴더 추가도 감지한다.
            if same_known and (root_saved in (None, root_now) or root_saved == root_now):
                if root_saved is None:
                    cached["root_mtime_ns"] = root_now
                    temp = ERP_CACHE + ".tmp"
                    with open(temp, "w", encoding="utf-8") as f:
                        json.dump(cached, f, ensure_ascii=False, indent=1)
                    os.replace(temp, ERP_CACHE)
                for project, saved in cached.get("evidence", {}).items():
                    profile = _profile(evidence, project)
                    profile["erp"] = True
                    profile["erp_dates"].update(saved.get("dates") or ())
                    profile["sources"].add("ERP 판매조회")
                return evidence, [tuple(x) for x in cached.get("files", [])]
        except (OSError, ValueError, TypeError):
            pass

    patterns = os.path.join(root, "**", "*.xls*")
    paths = [p for p in sorted(glob.glob(patterns, recursive=True))
             if not os.path.basename(p).startswith("~$")]
    signature = []
    for path in paths:
        try:
            st = os.stat(path)
            signature.append([os.path.normcase(os.path.abspath(path)), st.st_size, st.st_mtime_ns])
        except OSError:
            continue
    # ★ 2026-07-30: 원본 정리가 SHA256이 같은 판매조회를 3벌 남겨 같은 파일을 3번 읽고 있었다.
    #   지금은 프로젝트NO 집합(set)이라 개수가 부풀지 않지만, 여기에 금액 합산이 붙는 순간
    #   billing_fill 에서 났던 3배 사고(36.2억→108.6억)가 그대로 재현된다.
    #   파일명(`__dup_`)으로 거르면 다음번에 다른 이름으로 다시 뚫리므로 **내용 해시**로 거른다.
    #   signature 는 중복까지 전부 담아 둔다 — 사본이 바뀌어도 캐시가 무효화되어야 하기 때문이다.
    for path in dedupe_files(paths):
        if os.path.basename(path).startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        used = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = None
            header_row = 0
            for row_no, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
                values = [_s(v) for v in row]
                if any("프로젝트코드" in v for v in values):
                    header, header_row = values, row_no
                    break
            if not header:
                continue
            project_i = next(i for i, value in enumerate(header) if "프로젝트코드" in value)
            date_i = next((i for i, value in enumerate(header) if value == "일자"), None)
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                project = _s(row[project_i] if project_i < len(row) else "").upper()
                if not PROJECT_RE.fullmatch(project):
                    continue
                when = _date_iso(row[date_i] if date_i is not None and date_i < len(row) else "")
                profile = _profile(evidence, project)
                profile["erp"] = True
                if when:
                    profile["erp_dates"].add(when)
                profile["sources"].add("ERP 판매조회")
                used += 1
        wb.close()
        if used:
            files.append((os.path.basename(path), used))
    if os.path.normcase(os.path.abspath(root)) == os.path.normcase(os.path.abspath(ERP_DIR)):
        os.makedirs(os.path.dirname(ERP_CACHE), exist_ok=True)
        temp = ERP_CACHE + ".tmp"
        payload = {
            "version": ERP_CACHE_VERSION,
            "signature": signature,
            "root_mtime_ns": os.stat(root).st_mtime_ns,
            "files": files,
            "evidence": {
                project: {"dates": sorted(profile["erp_dates"])}
                for project, profile in evidence.items()
            },
        }
        with open(temp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=1)
        os.replace(temp, ERP_CACHE)
    return evidence, files


def _merge_profiles(*groups):
    out = {}
    for group in groups:
        for project, source in group.items():
            target = _profile(out, project)
            for flag in ("statement", "erp", "completed"):
                target[flag] = target[flag] or bool(source.get(flag))
            target["photos"] = max(target["photos"], int(source.get("photos") or 0))
            for key in ("statement_dates", "erp_dates", "completion_dates", "sources"):
                target[key].update(source.get(key) or ())
    return out


def _ledger_statement_evidence(wb):
    """06시트의 실제 거래명세서번호와 발행일을 정본 근거로 사용한다."""
    evidence = {}
    sheet_name = "06_거래서류청구수금"
    if sheet_name not in wb.sheetnames:
        return evidence
    ws = wb[sheet_name]
    index = _headers(ws)
    for row in ws.iter_rows(min_row=FIRST, values_only=True):
        project = _value(row, index, "프로젝트NO").upper()
        number = _value(row, index, "거래명세서번호")
        if not PROJECT_RE.fullmatch(project) or not number:
            continue
        when = (_date_iso(row[index["거래명세서발행일"]])
                if "거래명세서발행일" in index else "")
        if not when:
            when = (_date_iso(row[index["작업완료일"]])
                    if "작업완료일" in index else "")
        profile = _profile(evidence, project)
        profile["statement"] = True
        if when:
            profile["statement_dates"].add(when)
        profile["sources"].add("06 거래명세서 원장")
    return evidence


def _verified_date(profile):
    dates = set(profile.get("statement_dates") or ()) | set(profile.get("erp_dates") or ())
    return max(dates) if dates else ""


def _completion_date(profile):
    dates = set(profile.get("completion_dates") or ())
    return max(dates) if dates else ""


def _evidence_text(profile):
    sources = "·".join(sorted(profile.get("sources") or ())) or "확정 증빙"
    when = _verified_date(profile)
    return f"{sources}" + (f" / 확인일 {when}" if when else "")


def _cell_item(sheet, row, column, letter, value, evidence, *, overwrite=False,
               vtype="text", project=""):
    """ledger_writer의 셀 직접 지정 모드용 항목."""
    return {
        "sheet": sheet,
        "key": f"{project or sheet}@{row}",
        "key_col": "프로젝트NO",
        "cell": f"{letter}{row}",
        "col": column,
        "value": value,
        "vtype": vtype,
        "only_if_empty": not overwrite,
        "evidence": evidence,
    }


def _different(current, value, vtype="text"):
    if vtype == "date":
        return _date_iso(current) != _date_iso(value)
    return _s(current) != _s(value)


def _eligible_as_rows(ws):
    """03 배열수식과 동일한 02 완료행 순서를 재현한다."""
    index = _headers(ws)
    rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=FIRST, values_only=True), start=FIRST):
        project = _value(row, index, "프로젝트NO").upper()
        completed = (_value(row, index, "진행상태") == "작업완료"
                     and bool(_date_iso(row[index["작업완료일"]])
                              if "작업완료일" in index else ""))
        if completed and PROJECT_RE.fullmatch(project):
            rows.append((row_no, project, row, index))
    return rows


def derived_field_status_map(wb):
    """앱도 쓸 수 있는 03행→프로젝트 매핑과 현재 검증 상태를 돌려준다."""
    if "02_돌발AS접수" not in wb.sheetnames or "03_현장작업실적" not in wb.sheetnames:
        return {}
    eligible = _eligible_as_rows(wb["02_돌발AS접수"])
    ws = wb["03_현장작업실적"]
    index = _headers(ws)
    out = {}
    for offset, row in enumerate(ws.iter_rows(min_row=FIRST, values_only=True)):
        if offset >= len(eligible):
            break
        project = eligible[offset][1]
        row_no = FIRST + offset
        detail = _value(row, index, "실제작업상세")
        item = _value(row, index, "실제작업항목")
        report = _value(row, index, "기사보고내용")
        if not any((detail, item, report)):
            continue
        out[project] = {
            "현장작업행": row_no,
            "실제작업항목": item,
            "실제작업상세": detail,
            "현장관리자검증": _value(row, index, "관리자검증"),
            "거래명세서반영": _value(row, index, "거래명세서반영"),
            "ERP반영": _value(row, index, "ERP반영"),
            "검증자": _value(row, index, "검증자"),
            "검증일": _date_iso(row[index["검증일"]]) if "검증일" in index else "",
        }
    return out


def build_plan(path, *, band_records=None, erp_evidence=None, erp_files=None):
    """업데이트 큐와 집계를 만든다. 실파일에는 쓰지 않는다."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    band = band_document_evidence(band_records)
    if erp_evidence is None:
        erp, discovered_files = erp_sales_evidence()
    else:
        erp, discovered_files = erp_evidence, (erp_files or [])
    profiles = _merge_profiles(band, erp, _ledger_statement_evidence(wb))
    items, counts = [], Counter()

    def add(ws, index, row_no, row, project, column, value, profile, *,
            overwrite=False, vtype="text"):
        current = row[index[column]] if index[column] < len(row) else None
        if not _different(current, value, vtype):
            return
        letter = get_column_letter(index[column] + 1)
        items.append(_cell_item(
            ws.title, row_no, column, letter, value, _evidence_text(profile),
            overwrite=overwrite, vtype=vtype, project=project,
        ))
        counts[f"{ws.title}:{column}"] += 1

    # 03 실제작업 내용을 먼저 인덱싱하여 02 관리자검증과 추가작업 상태에도 같은
    # 근거를 전파한다. 배열수식 캐시는 사용하지 않는다.
    ws02 = wb["02_돌발AS접수"]
    idx02 = _headers(ws02)
    eligible = _eligible_as_rows(ws02)
    ws03 = wb["03_현장작업실적"]
    idx03 = _headers(ws03)
    field_work = {}
    field_rows = list(ws03.iter_rows(min_row=FIRST, values_only=True))
    for offset, row in enumerate(field_rows[:len(eligible)]):
        project = eligible[offset][1]
        real_work = any(_value(row, idx03, name)
                        for name in ("실제작업항목", "실제작업상세", "기사보고내용"))
        if real_work:
            extra_flag = _value(row, idx03, "접수외추가작업여부")
            extra = (True if extra_flag == "있음" else
                     False if extra_flag == "없음" else None)
            revisit = _value(row, idx03, "재방문필요")
            field_work[project] = {
                "extra": extra,
                "revisit": revisit if revisit in ("예", "아니오") else "",
                "row": FIRST + offset,
            }

    # 02: 확정 문서/ERP 근거가 있는 완료건의 원인 열을 채운다.
    for row_no, project, row, _index in eligible:
        profile = profiles.get(project)
        if not profile:
            continue
        if profile["completed"]:
            if profile["photos"]:
                add(ws02, idx02, row_no, row, project, "사진등록", "등록", profile, overwrite=True)
            add(ws02, idx02, row_no, row, project, "완료보고서등록", "등록",
                profile, overwrite=True)
            add(ws02, idx02, row_no, row, project, "밴드수정", "완료",
                profile, overwrite=True)
        if profile["erp"]:
            add(ws02, idx02, row_no, row, project, "ERP등록", "완료", profile, overwrite=True)
        if profile["erp"] or profile["statement"]:
            add(ws02, idx02, row_no, row, project, "담당관리자", "유현민", profile, overwrite=True)
            when = _verified_date(profile)
            if when and not _value(row, idx02, "최종확인일"):
                add(ws02, idx02, row_no, row, project, "최종확인일", when, profile, vtype="date")
        if profile["statement"] and profile["erp"] and project in field_work:
            extra = field_work[project]["extra"]
            if extra is not None:
                add(ws02, idx02, row_no, row, project, "최초접수외추가작업",
                    "있음" if extra else "없음", profile, overwrite=True)
            if extra is True:
                add(ws02, idx02, row_no, row, project, "추가작업확인상태",
                    "반영완료", profile, overwrite=True)
            add(ws02, idx02, row_no, row, project, "관리자검증상태",
                "추가작업발생" if extra is True else "일치", profile, overwrite=True)
            if field_work[project]["revisit"]:
                add(ws02, idx02, row_no, row, project, "재방문여부",
                    field_work[project]["revisit"], profile, overwrite=True)

    # 03: 실제 작업내용이 있는 파생행만. 빈 예비행은 절대 쓰지 않는다.
    for offset, row in enumerate(field_rows):
        if offset >= len(eligible):
            break
        row_no = FIRST + offset
        project = eligible[offset][1]
        profile = profiles.get(project)
        if not profile:
            continue
        real_work = any(_value(row, idx03, name)
                        for name in ("실제작업항목", "실제작업상세", "기사보고내용"))
        if not real_work:
            continue
        if profile["completed"]:
            if profile["photos"] and "작업사진" in idx03:
                add(ws03, idx03, row_no, row, project, "작업사진", "등록",
                    profile, overwrite=True)
            if "완료보고서" in idx03:
                add(ws03, idx03, row_no, row, project, "완료보고서", "등록",
                    profile, overwrite=True)
        if profile["statement"]:
            add(ws03, idx03, row_no, row, project, "거래명세서반영", "반영완료",
                profile, overwrite=True)
        if profile["erp"]:
            add(ws03, idx03, row_no, row, project, "ERP반영", "반영완료",
                profile, overwrite=True)
        if profile["statement"] or profile["erp"]:
            add(ws03, idx03, row_no, row, project, "검증자", "유현민", profile, overwrite=True)
            when = _verified_date(profile)
            if when and not _value(row, idx03, "검증일"):
                add(ws03, idx03, row_no, row, project, "검증일", when, profile, vtype="date")
        if profile["statement"] and profile["erp"]:
            extra = field_work[project]["extra"]
            add(ws03, idx03, row_no, row, project, "관리자검증",
                "추가작업발생" if extra else "일치", profile, overwrite=True)

    # 04: 완료된 정기점검의 판매전표·거래명세서와 검증자/일을 동기화한다.
    ws04 = wb["04_정기점검"]
    idx04 = _headers(ws04)
    for row_no, row in enumerate(ws04.iter_rows(min_row=FIRST, values_only=True), start=FIRST):
        project = _value(row, idx04, "프로젝트NO").upper()
        actual = (_date_iso(row[idx04["실제점검일"]]) if "실제점검일" in idx04 else "")
        if (_value(row, idx04, "점검상태") != "완료"
                or not actual or not PROJECT_RE.fullmatch(project)):
            continue
        profile = profiles.get(project)
        if not profile:
            continue
        if profile["completed"]:
            if profile["photos"]:
                add(ws04, idx04, row_no, row, project, "점검사진", "등록",
                    profile, overwrite=True)
            add(ws04, idx04, row_no, row, project, "점검보고서", "등록",
                profile, overwrite=True)
        if profile["erp"]:
            add(ws04, idx04, row_no, row, project, "ERP판매전표", "완료",
                profile, overwrite=True)
        if profile["statement"]:
            add(ws04, idx04, row_no, row, project, "거래명세서", "발행완료",
                profile, overwrite=True)
        if profile["erp"] or profile["statement"]:
            add(ws04, idx04, row_no, row, project, "담당관리자", "유현민",
                profile, overwrite=True)
            when = _verified_date(profile)
            if when and not _value(row, idx04, "최종확인일(유현민 체크)"):
                add(ws04, idx04, row_no, row, project, "최종확인일(유현민 체크)",
                    when, profile, vtype="date")

    wb.close()
    return items, counts, profiles, discovered_files


def main():
    if is_input_window():
        print(f"입력 보호시간({input_window_label()}) — 검증상태 동기화 생략")
        return 0
    do_queue = "--queue" in sys.argv
    source = latest_master()[0]
    items, counts, profiles, files = build_plan(source)
    print("원본:", os.path.basename(source))
    print(f"확정 증빙 프로젝트: {len(profiles)}개 / 변경 후보: {len(items)}칸")
    for name, count in sorted(counts.items()):
        print(f"  {name}: {count}")
    if files:
        print("ERP 판매조회:", ", ".join(f"{name}({count})" for name, count in files))
    if not do_queue:
        print("미리보기 완료 — 반영 큐 추가: python verification_sync.py --queue")
        return 0
    import ledger_writer
    added = ledger_writer.queue_add(items)
    print(f"자동입력 큐 추가 {added}건")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
