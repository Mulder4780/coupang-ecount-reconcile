# -*- coding: utf-8 -*-
"""
po_pdf.py — 쿠팡 PO 원본 PDF에서 값을 뽑아 대조한다 (설치 0 · 표준 라이브러리만)
================================================================================
오종현이 모아 둔 PO 알림 메일 PDF 묶음이 **원본 근거**다. 지금까지 PO 대조는 쿠팡이 준
목록 엑셀에만 의존했는데, 목록은 사람이 만든 것이라 금액·날짜가 틀릴 수 있다.
원본을 읽어 두면 미청구 7건(348,698,500원) 같은 건을 원본으로 확인할 수 있다.

왜 직접 파서를 쓰나
  · 이 프로젝트는 외부 패키지 설치 0이 원칙이고, pypdf·pdfminer 가 깔려 있지 않다.
  · 이 PDF는 스캔본이 아니라 **글자가 들어 있는** 문서다(/Font 349개, ToUnicode 77개).
    OCR이 필요 없다 — 금융 문서를 외부 OCR에 올리지 않는다는 규칙(AGENTS.md 9)에도 맞다.

★ 한글·숫자는 CID 폰트라 바이트가 곧 글자가 아니다. **폰트별로** ToUnicode CMap을 읽어
  CID → 유니코드로 바꿔야 한다. 전부 합쳐서 쓰면 같은 CID가 폰트마다 다른 글자라서
  'Coupang'이 'Coupapg'가 되고 날짜 숫자가 통째로 어긋난다(2026-07-28 실제 증상).

  python po_pdf.py                 # 폴더 전체 파싱 → 요약
  python po_pdf.py --csv           # reports/PO원본_YYYYMMDD.csv
  python po_pdf.py --check         # 쿠팡 목록과 대조(다른 것만 보고)
"""
import sys, os, re, csv, zlib, glob
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")
# 경로는 source_dirs 한 곳에서만 정한다 — 도구마다 하드코딩하면 자료가 옮겨질 때
# 한쪽만 고쳐져 못 읽는 일이 생긴다(2026-07-28에 실제로 경로가 되돌아가 있었다).
from source_dirs import po_dirs

PO_RE = re.compile(r"PO\s*(\d{6})")
UJ_RE = re.compile(r"UJ\d{7}")


# ───────────────────────── PDF 텍스트 추출 ─────────────────────────
def _objects(data):
    """PDF 간접 객체: 번호 → 본문(바이트)."""
    objs = {}
    for m in re.finditer(rb"(\d+)\s+0\s+obj", data):
        e = data.find(b"endobj", m.end())
        objs[int(m.group(1))] = data[m.end(): e if e > 0 else len(data)]
    return objs


def _inflate(body):
    """객체 안의 stream 을 푼다(압축이 아니면 원문 그대로)."""
    m = re.search(rb"stream\r?\n", body)
    if not m:
        return b""
    e = body.find(b"endstream", m.end())
    raw = body[m.end(): e if e > 0 else len(body)]
    try:
        return zlib.decompress(raw)
    except Exception:
        return raw


def _u(hexstr):
    """UTF-16BE 16진 문자열 → 문자"""
    try:
        b = bytes.fromhex(hexstr if len(hexstr) % 2 == 0 else "0" + hexstr)
        return b.decode("utf-16-be", "replace")
    except Exception:
        return ""


def _parse_cmap(blob):
    """ToUnicode CMap → {CID: 문자}"""
    out = {}
    t = blob.decode("latin-1", "replace")
    for blk in re.findall(r"beginbfchar(.*?)endbfchar", t, re.S):
        for a, b in re.findall(r"<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>", blk):
            out[int(a, 16)] = _u(b)
    for blk in re.findall(r"beginbfrange(.*?)endbfrange", t, re.S):
        for a, b, c in re.findall(r"<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>", blk):
            lo, hi, dst = int(a, 16), int(b, 16), int(c, 16)
            for i in range(min(hi - lo + 1, 65536)):
                out[lo + i] = _u(format(dst + i, "04x"))
    return out


def _font_maps(data, objs):
    """폰트 이름(/F1 …) → CMap. 폰트별로 갈라야 글자가 안 어긋난다."""
    name2obj = {}
    for m in re.finditer(rb"/Font\s*<<(.+?)>>", data, re.S):
        for fm in re.finditer(rb"/(\w+)\s+(\d+)\s+0\s+R", m.group(1)):
            name2obj[fm.group(1).decode("latin-1")] = int(fm.group(2))
    maps = {}
    for name, onum in name2obj.items():
        body = objs.get(onum, b"")
        tu = re.search(rb"/ToUnicode\s+(\d+)\s+0\s+R", body)
        if not tu:
            df = re.search(rb"/DescendantFonts\s*\[\s*(\d+)\s+0\s+R", body)
            if df:
                tu = re.search(rb"/ToUnicode\s+(\d+)\s+0\s+R", objs.get(int(df.group(1)), b""))
        if tu:
            maps[name] = _parse_cmap(_inflate(objs.get(int(tu.group(1)), b"")))
    return maps


def _content(objs):
    """텍스트 연산자가 들어 있는 content stream 들."""
    out = []
    for body in objs.values():
        blob = _inflate(body)
        if b"Tj" in blob or b"TJ" in blob:
            out.append(blob)
    return out


TOK = re.compile(r"/(\w+)\s+[\d.]+\s+Tf|<([0-9A-Fa-f\s]+)>|\(((?:[^()\\]|\\.)*)\)")


def _text(streams, fmaps):
    """content stream 을 훑어 글자만 이어 붙인다. /Fx Tf 로 현재 폰트를 따라간다."""
    out = []
    for s in streams:
        t = s.decode("latin-1", "replace")
        cur = {}
        for m in TOK.finditer(t):
            if m.group(1) is not None:                  # 폰트 전환
                cur = fmaps.get(m.group(1), {})
            elif m.group(2) is not None:                # <hex> — CID 문자열
                h = re.sub(r"\s", "", m.group(2))
                step = 4 if (cur and len(h) % 4 == 0) else 2
                for i in range(0, len(h) - step + 1, step):
                    out.append(cur.get(int(h[i:i + step], 16), ""))
            else:                                        # (문자열) — 보통 ASCII
                out.append(re.sub(r"\\(.)", r"\1", m.group(3)))
        out.append("\n")
    return "".join(out)


def pdf_text(path):
    data = open(path, "rb").read()
    objs = _objects(data)
    return _text(_content(objs), _font_maps(data, objs))


# ───────────────────────── PO 값 뽑기 ─────────────────────────
# ★ 본문은 표를 그대로 이어 붙여서 나온다:
#     "구매 오더 세부 사항 오더 번호 금액 PO3445994,200,000 KRW"
#   PO번호와 금액 사이에 공백이 없다. 그래서 그냥 숫자를 긁으면 둘이 붙어
#   3,445,994,200,000원 같은 값이 나온다(2026-07-28: 18건이 그렇게 어긋났다).
#   PO번호 바로 뒤에 붙는 금액을 **그 구조 그대로** 읽는다.
# ★★ 문서 안에는 PO번호가 여러 번 나오고 숫자도 여기저기 있다. **'금액:' 라벨이 붙은
#   값만** 그 오더의 금액이다. 라벨을 안 보고 'PO번호 뒤 숫자'로 잡았더니 엉뚱한 값을
#   집어 목록과 다르다고 4건을 잘못 신고할 뻔했다(2026-07-28).
#   실제 원문: "구매 오더(신규)PO326234금액: 3,680,000 KRW"
AMT_LABEL = re.compile(r"금\s*액\s*[:：]\s*([\d,]{5,})\s*(?:KRW|원)")
AMT_AFTER_PO = re.compile(r"PO\d{6}\s*금\s*액\s*[:：]?\s*([\d,]{5,})\s*(?:KRW|원)")
AMT_GLUED = re.compile(r"PO(\d{6})([\d,]{5,})\s*(?:KRW|원)")
AMT_RE = re.compile(r"(?<!\d)(\d{1,3}(?:,\d{3})+)\s*(?:KRW|원)")
AMT_MAX = 10 ** 11        # 1,000억 — 이보다 크면 붙어서 생긴 값이다
DATE_RE = re.compile(r"(20\d{2})[-./년]\s*(\d{1,2})[-./월]\s*(\d{1,2})")


def _amount(t, po):
    """오더 금액. 이 PDF는 서식이 **두 가지**라 둘 다 봐야 한다.

      ① 라벨형 : "구매 오더(신규)PO326234금액: 3,680,000 KRW"
      ② 표  형 : "오더 번호금액PO3445994,200,000 KRW"   ← 번호와 금액이 붙어 있다

    ②를 그냥 숫자로 긁으면 번호와 금액이 붙어 3,445,994,200,000원이 되고,
    반대로 콤마 단위로만 끊으면 앞자리가 잘려 200,000원이 된다.
    그래서 **PO번호를 떼어 내고 남은 부분**을 금액으로 읽고, 떼어 낸 번호가
    파일의 PO번호와 같은지까지 확인한다(2026-07-28: 두 방식 모두에서 오판이 났다).
    """
    m = AMT_AFTER_PO.search(t) or AMT_LABEL.search(t)
    if m:
        return int(m.group(1).replace(",", ""))
    for m in AMT_GLUED.finditer(t):
        if not po or ("PO" + m.group(1)) == po:
            v = int(m.group(2).replace(",", ""))
            if 0 < v < AMT_MAX:
                return v
    amts = [int(a.replace(",", "")) for a in AMT_RE.findall(t)]
    amts = [a for a in amts if a < AMT_MAX]
    return max(amts) if amts else None


QUOTE_RE = re.compile(r"([\w가-힣()\-]+?)견적서_?([\d,]+)?원?")


def parse(path):
    name = os.path.basename(path)
    parent = os.path.basename(os.path.dirname(path))
    # PO번호는 파일명에 없으면 **상위 폴더명**에서 가져온다(첨부 견적서가 그렇다)
    po = PO_RE.search(name.replace(" ", "")) or PO_RE.search(parent.replace(" ", ""))
    quote = "견적서" in name
    rec = {"파일": name, "상위폴더": parent, "종류": "견적서" if quote else "PO통지",
           "PO번호": ("PO" + po.group(1)) if po else "",
           "변경본": ("변경" in name) or ("변경" in parent),
           "금액": None, "일자": "", "프로젝트NO": "", "본문요약": ""}
    if quote:
        # 견적서는 파일명이 이미 캠프·금액을 담고 있다: "1-1._부산4MB견적서_269,100원.pdf"
        q = QUOTE_RE.search(name)
        if q:
            rec["캠프힌트"] = re.sub(r"^[\d._\-]+", "", q.group(1))[:20]
            if q.group(2):
                rec["금액"] = int(q.group(2).replace(",", ""))
    try:
        t = pdf_text(path)
    except Exception as e:
        rec["본문요약"] = "읽기실패 " + type(e).__name__
        return rec
    t = re.sub(r"[ \t]+", " ", t)
    if not rec["PO번호"]:
        m = PO_RE.search(t)
        rec["PO번호"] = "PO" + m.group(1) if m else ""
    rec["금액"] = _amount(t, rec["PO번호"])
    d = DATE_RE.search(t)
    if d:
        rec["일자"] = "%s-%02d-%02d" % (d.group(1), int(d.group(2)), int(d.group(3)))
    u = UJ_RE.search(t)
    if u:
        rec["프로젝트NO"] = u.group()
    rec["본문요약"] = " ".join(t.split())[:140]
    return rec


def is_pdf(path):
    """확장자만 믿지 않는다. 이 폴더에는 **.pdf 가 안 붙은 PDF**가 섞여 있다
    (메일 본문을 저장하면서 확장자가 빠진 것으로 보인다). 머리 4바이트로 판별한다."""
    if not os.path.isfile(path):
        return False
    if path.lower().endswith(".pdf"):
        return True
    try:
        with open(path, "rb") as f:
            return f.read(4) == b"%PDF"
    except OSError:
        return False


def pdf_files(folder):
    """하위 폴더까지 훑는다. PO별로 폴더가 나뉘고 그 안에 견적서 첨부가 들어 있다."""
    out = []
    for base, _dirs, files in os.walk(folder):
        for fn in files:
            if fn.lower() == "thumbs.db":
                continue
            p = os.path.join(base, fn)
            if is_pdf(p):
                out.append(p)
    return sorted(out)


CACHE_FILE = os.path.join(REPORT_DIR, "po_pdf_cache.json")


def _load_cache():
    try:
        import json
        with open(CACHE_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_cache(c):
    try:
        import json
        os.makedirs(REPORT_DIR, exist_ok=True)
        tmp = CACHE_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(c, f, ensure_ascii=False)
        os.replace(tmp, CACHE_FILE)
    except Exception:
        pass


def scan(folder=None):
    """폴더 하나를 지정하면 그것만, 기본값이면 **PO_DIRS 전부**를 훑는다.

    ★ 같은 파일이 양쪽 폴더에 다 있을 수 있다(자료를 옮기는 중이다). 파일명+종류로
      한 번만 센다 — 중복을 그대로 두면 '고유 PO 개수'가 부풀고 대조 숫자가 어긋난다.
    """
    folders = [folder] if folder else po_dirs()
    # ★ 파싱 **전에** 중복을 거른다. 두 폴더가 같은 내용을 담고 있어(574개 동일)
    #   나중에 걸러내면 같은 PDF를 두 번 여는 셈이라 느린 네트워크 드라이브에서 두 배 걸린다.
    picked, seen = [], set()
    for d in folders:
        for f in pdf_files(d):
            # 정본 폴더는 ``2026/PO번호``이고 예전 공유 폴더는 긴 메일 제목 폴더다.
            # 부모 폴더명으로 중복을 가르면 같은 568개를 서로 다른 자료로 두 번 읽게 된다.
            # 경로 전체의 PO번호 + 파일명 + 크기로 정본/공유본을 한 번만 센다.
            joined = os.path.normpath(f).replace(" ", "")
            m = PO_RE.search(joined)
            try:
                size = os.path.getsize(f)
            except OSError:
                size = -1
            key = ((m.group(1) if m else os.path.basename(os.path.dirname(f)).lower()),
                   os.path.basename(f).lower(), size)
            if key in seen:
                continue
            seen.add(key)
            picked.append((d, f))
    # ★ Z: 는 네트워크 드라이브라 574개를 매번 다시 읽으면 몇 분씩 걸린다.
    #   (경로·크기·수정시각)이 그대로면 지난 파싱 결과를 그대로 쓴다.
    cache = _load_cache()
    recs, hit = [], 0
    for d, f in picked:
        try:
            st = os.stat(f)
            key = f"{f}|{st.st_size}|{int(st.st_mtime)}"
        except OSError:
            key = None
        if key and key in cache:
            r = dict(cache[key]); hit += 1
        else:
            r = parse(f)
            if key:
                cache[key] = r
        r["출처폴더"] = d
        recs.append(r)
    _save_cache(cache)
    if hit:
        print(f"(캐시 재사용 {hit}/{len(picked)}개 — 바뀐 파일만 다시 읽었습니다)")
    return recs, " + ".join(folders) if folders else str(folder)


def latest_by_po(recs):
    """PO번호별 최종 **통지문**. 변경 통지가 있으면 그것이 최종이다.

    ★ 견적서 첨부를 섞으면 안 된다. 견적서는 그 PO를 **캠프별로 쪼갠 일부**라
      금액이 훨씬 작다. 섞었더니 PO329774가 15,632,000 대신 506,000으로 잡혀
      '금액 불일치 46건'이 났다(2026-07-28). 견적서는 quotes_by_po 로 따로 본다.
    """
    out = {}
    for r in recs:
        if not r["PO번호"] or r.get("종류") != "PO통지":
            continue
        cur = out.get(r["PO번호"])
        if cur is None or (r["변경본"] and not cur["변경본"]):
            out[r["PO번호"]] = r
    return out


def quotes_by_po(recs):
    """PO번호 → 견적서 목록. 파일명이 캠프·금액을 담고 있어 **PO 내역**이 된다.
       묶음 계산서 구성을 추정이 아니라 원본으로 확인할 수 있는 유일한 근거다."""
    out = {}
    for r in recs:
        if r.get("종류") == "견적서" and r["PO번호"]:
            out.setdefault(r["PO번호"], []).append(r)
    return out


def check(byno):
    """쿠팡 목록 엑셀과 원본 PDF 대조 — **다른 것만** 보고한다."""
    from inbox_scan import pick
    import openpyxl
    files = pick("po")
    if not files:
        print("\ninbox 에 쿠팡 PO 목록이 없어 대조를 건너뜁니다.")
        return
    wb = openpyxl.load_workbook(files[0], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    wb.close()

    hdr, lst = None, {}
    for r in rows:
        v = [str(x).strip() if x is not None else "" for x in r]
        if hdr is None:
            if "오더 번호" in v:
                hdr = {h: i for i, h in enumerate(v) if h}
            continue
        m = PO_RE.search(v[hdr["오더 번호"]].replace(" ", "")) if "오더 번호" in hdr else None
        if not m:
            continue
        amt = None
        if "금액" in hdr and v[hdr["금액"]]:
            try:
                amt = int(float(re.sub(r"[^\d.]", "", v[hdr["금액"]]) or 0))
            except ValueError:
                amt = None
        lst["PO" + m.group(1)] = amt

    only_pdf = sorted(set(byno) - set(lst))
    only_lst = sorted(set(lst) - set(byno))
    # ★ 원본이 **부가세 포함** 금액을 적은 경우가 있다(정확히 목록×1.1).
    #   이걸 '금액 다름'으로 올리면 멀쩡한 건을 조사하게 된다 — 실제로 미청구 7건 중
    #   2건(PO354310·PO359404)이 여기 해당했다(2026-07-28). 단위 차이는 따로 표시한다.
    def _vat(a, b):
        return a and b and abs(b - round(a * 1.1)) <= 2
    pairs = [(k, lst[k], byno[k]["금액"]) for k in sorted(set(byno) & set(lst))
             if lst[k] is not None and byno[k]["금액"] is not None and lst[k] != byno[k]["금액"]]
    vat = [(k, a, b) for k, a, b in pairs if _vat(a, b)]
    diff = [(k, a, b) for k, a, b in pairs if not _vat(a, b)]

    print(f"\n쿠팡 목록 {len(lst)}건  vs  원본 PDF {len(byno)}건")
    print(f"  원본에만 있음 {len(only_pdf)} · 목록에만 있음 {len(only_lst)} · "
          f"금액 불일치 {len(diff)} · 부가세 표기차 {len(vat)}")
    for k, a, b in vat[:6]:
        print(f"    [부가세표기] {k} 목록 {a:,}(공급가) / 원본 {b:,}(VAT포함) — 같은 건")
    for k in only_pdf[:10]:
        a = byno[k]["금액"]
        print(f"    [원본에만] {k} {format(a, ',') + '원' if a else '(금액 미추출)'} {byno[k]['일자']}")
    for k in only_lst[:10]:
        print(f"    [목록에만] {k}  ← 원본 PDF가 아직 없습니다")
    for k, a, b in diff[:10]:
        print(f"    [금액다름] {k} 목록 {a:,} / 원본 {b:,} (차 {b - a:+,})")
    return {"only_pdf": only_pdf, "only_lst": only_lst, "diff": diff}


def main():
    args = sys.argv[1:]
    folder = args[args.index("--dir") + 1] if "--dir" in args else None
    recs, folder = scan(folder)
    if not recs:
        print("PDF를 찾지 못했습니다: " + str(folder))
        return

    ok = [r for r in recs if r["금액"]]
    byno = latest_by_po(recs)
    print(f"PO 원본 PDF {len(recs)}개 · 금액 추출 {len(ok)} · 실패 {len(recs) - len(ok)}")
    q = quotes_by_po(recs)
    print(f"고유 PO번호 {len(byno)}개(통지문 기준) · 변경 통지 {sum(1 for r in recs if r['변경본'])}건")
    print(f"견적서 첨부 {sum(len(v) for v in q.values())}건 · PO {len(q)}개에 붙음")
    for r in recs[:4]:
        print(f"  {r['PO번호'] or '?':10} {str(r['금액'] or '-'):>12} {r['일자'] or '-':10} {r['본문요약'][:52]}")

    if "--csv" in args:
        os.makedirs(REPORT_DIR, exist_ok=True)
        p = os.path.join(REPORT_DIR, f"PO원본_{datetime.now():%Y%m%d}.csv")
        with open(p, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(recs[0].keys()))
            w.writeheader()
            w.writerows(recs)
        print("리포트:", p)

    if "--check" in args:
        check(byno)


if __name__ == "__main__":
    main()
