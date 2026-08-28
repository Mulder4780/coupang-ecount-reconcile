# -*- coding: utf-8 -*-
"""두 관리대장 버전을 대 보고 **값이 바뀐 칸**을 짚는다. 읽기만 한다.

★ 왜 만들었나 — 손입력 경보가 거짓이었다(2026-08-28 실측).
  2026-08-27 21:43 에 `[P1] 관리대장 손입력 감지` 가 떴고, 그것이 형님 몫으로
  하루를 남아 있었다("앱에 다시 넣으셔야 합니다"). 그런데 v620 과 v621 을 실제로
  대 보니 **값이 바뀐 칸은 0개**이고 전부 수식이었다 —
  `05_신규납품설치` 66~70행의 빈 칸에 **채번 수식이 심어진 것**(`expand_rows`)이다.
  그것은 사람이 값을 적은 것이 아니라 **기계 도구가 행을 늘린 것**이다.

  범인은 판정 근거다. `realtime_monitor.hand_edit_verdict` 는 "기계 회차 근거 없이
  바뀐 최신본"을 손입력으로 보는데, 그 근거는 `batch` 표와 `archive_worker` 자국
  둘뿐이다 — **`expand_rows` 는 어느 쪽에도 자국을 안 남긴다.** 그래서 행을 늘릴
  때마다 거짓 경보가 나고, 거짓 경보는 **진짜 손입력을 덮는다**([170]).

★ 가르는 근거는 하나다 — **값이 바뀌었나, 수식·구조만 바뀌었나.**
  · 값이 바뀌었다 → 진짜 손입력이다. 역수입 금지라 앱에 다시 넣어야 하고,
    이 모듈이 **어느 시트 어느 칸인지** 짚어 준다(예전에는 "앱으로 입력" 한 마디뿐이라
    사람이 어느 칸인지 스스로 찾아야 했다).
  · 수식·구조만 바뀌었다 → 기계 도구가 한 일이다. 경보가 아니다([170]).

★ 지키는 것
  · **읽기만 한다.** 워크북을 열어 save() 하지 않는다(차트·도형 파괴 · 절대규칙).
  · **비싸다**(실측 12.6초 · Z: 워크북 둘). 그래서 **손입력 의심이 선 뒤에만** 부른다([168]).
  · **못 갈랐으면 못 갈랐다고 말한다**([169]) — 옛 버전이 없거나 못 읽으면
    `못함` 에 이유를 담는다. 모름을 '값 0'(경보 아님)으로 뭉개면 **진짜 손입력이
    조용히 사라진다** — 이 자리에서 기우는 방향은 **경보를 남기는 쪽**이다.
  · 배열 수식은 openpyxl 이 객체로 주어 `==` 가 언제나 거짓이다 — 정규화해서
    **없는 차이를 만들지 않는다**(실측: 정규화 전 180칸 → 후 72칸, 값은 그대로 0).
"""
import os
import re
import time

BUDGET_S = 120.0          # 이 안에 못 끝내면 '못함' 으로 돌려준다(회차를 안 세운다)
MAX_CELLS = 40            # 사람에게 보여 줄 칸 수 — 넘으면 몇 개인지 말한다([273])


def _norm(v):
    """배열 수식을 비교 가능한 모양으로 — 없는 차이를 만들지 않는다."""
    cls = type(v).__name__
    if cls == "ArrayFormula":
        return ("ARR", getattr(v, "text", None), getattr(v, "ref", None))
    return v


def previous_version(path):
    """`path`(vN)의 **바로 앞 버전(vN-1)** 을 찾는다. 없으면 None.

    같은 폴더와 보관 폴더(OLD)를 본다 — 옛 버전은 `ledger_versions` 가 OLD 로 옮긴다."""
    try:
        folder, base = os.path.split(path)
        m = re.search(r"v(\d+)", base)
        if not m:
            return None
        want = int(m.group(1)) - 1
        prev_name = base[:m.start(1)] + str(want) + base[m.end(1):]
        cands = [os.path.join(folder, prev_name)]
        try:
            import ledger_versions
            cands.append(os.path.join(ledger_versions.archive_folder(folder), prev_name))
        except Exception:
            cands.append(os.path.join(folder, "OLD", prev_name))
        for c in cands:
            if os.path.isfile(c):
                return c
    except Exception:
        return None
    return None


def changed_cells(old_path, new_path, budget_s=BUDGET_S):
    """두 버전을 대 본다. 돌려주는 값:

        {"값": [(시트, 칸, 옛값, 새값)...], "수식": n, "값수": n, "못함": 이유 or None}

    ⚠ `못함` 이 차 있으면 **'값 0' 을 믿으면 안 된다** — 안 센 것이지 없는 것이 아니다([169])."""
    out = {"값": [], "수식": 0, "값수": 0, "못함": None}
    if not old_path or not os.path.isfile(old_path):
        out["못함"] = "앞 버전을 못 찾았다"
        return out
    if not new_path or not os.path.isfile(new_path):
        out["못함"] = "새 버전을 못 찾았다"
        return out
    t0 = time.time()
    wa = wb = None
    try:
        import warnings
        warnings.filterwarnings("ignore")
        import openpyxl
        from openpyxl.utils import get_column_letter
        wa = openpyxl.load_workbook(old_path, data_only=False, read_only=True)
        wb = openpyxl.load_workbook(new_path, data_only=False, read_only=True)
        for name in [n for n in wb.sheetnames if n in wa.sheetnames]:
            if time.time() - t0 > budget_s:
                out["못함"] = "예산(%.0f초)을 넘겨 %s 부터 못 봤다" % (budget_s, name)
                break
            ra = {i: r for i, r in enumerate(wa[name].iter_rows(values_only=True), 1)}
            for i, row in enumerate(wb[name].iter_rows(values_only=True), 1):
                old = ra.get(i)
                if old == row:
                    continue
                n = max(len(old or ()), len(row))
                for c in range(n):
                    ov = _norm(old[c] if old and c < len(old) else None)
                    nv = _norm(row[c] if c < len(row) else None)
                    if ov == nv:
                        continue
                    of = isinstance(ov, str) and ov.startswith("=")
                    nf = isinstance(nv, str) and nv.startswith("=")
                    if isinstance(ov, tuple) or isinstance(nv, tuple) or of or nf:
                        out["수식"] += 1
                        continue
                    out["값수"] += 1
                    if len(out["값"]) < MAX_CELLS:
                        out["값"].append((name, "%s%d" % (get_column_letter(c + 1), i),
                                          ov, nv))
    except Exception as exc:
        out["못함"] = "%s: %s" % (type(exc).__name__, str(exc)[:120])
    finally:
        for w in (wa, wb):
            try:
                if w is not None:
                    w.close()
            except Exception:
                pass
    return out


def describe(res):
    """사람이 읽는 한 줄. **모르면 모른다고 적는다**([169])."""
    if res.get("못함"):
        return "손으로 고친 칸을 못 갈랐습니다(%s) — 값이 바뀌었는지 확인이 필요합니다." % res["못함"]
    if res.get("값수"):
        head = ", ".join("%s %s" % (s, ref) for (s, ref, _o, _n) in res["값"][:5])
        more = ""
        if res["값수"] > len(res["값"]):
            more = " (그 밖 %d칸)" % (res["값수"] - len(res["값"]))
        elif res["값수"] > 5:
            more = " (그 밖 %d칸)" % (res["값수"] - 5)
        return "손으로 고친 칸 %d개: %s%s — 이 값을 앱에서 다시 넣어 주세요." % (
            res["값수"], head, more)
    return ("값이 바뀐 칸은 없습니다(수식·구조만 %d칸) — "
            "행 늘리기 같은 기계 도구가 만든 차이입니다." % res.get("수식", 0))


def main(argv=None):
    import argparse
    ap = argparse.ArgumentParser(description="두 관리대장 버전의 값 차이(읽기 전용)")
    ap.add_argument("new", nargs="?", help="새 버전 경로(비우면 최신본)")
    ap.add_argument("--old", help="앞 버전 경로(비우면 vN-1 을 스스로 찾는다)")
    a = ap.parse_args(argv)
    new = a.new
    if not new:
        import workbook_patch
        new, _v = workbook_patch.latest_master()
    old = a.old or previous_version(new)
    res = changed_cells(old, new)
    print("앞: %s" % (os.path.basename(old) if old else "(못 찾음)"))
    print("새: %s" % os.path.basename(new))
    print(describe(res))
    for (s, ref, ov, nv) in res["값"]:
        print("  [%s] %s : %r -> %r" % (s, ref, str(ov)[:40], str(nv)[:40]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
