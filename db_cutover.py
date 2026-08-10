# -*- coding: utf-8 -*-
"""관리대장 Excel을 앱 DB로 한 번만 이관하고 정본을 안전하게 전환한다.

Excel은 읽기 전용 근거로만 사용한다. 이 스크립트는 workbook을 저장하지 않으며,
행별 출처·해시·충돌을 남긴 뒤 parity 관문을 통과한 경우에만
``source_of_truth_mode=db_primary_export``를 확정한다.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, time as time_value, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional

from app_store import AppStore, canonical_json, sha256_json


ROOT = Path(__file__).resolve().parent
REPORT_DIR = Path(os.environ.get("COUPANG_REPORT_DIR") or ROOT / "reports")

SPECS: Dict[str, Dict[str, Any]] = {
    "02_돌발AS접수": {
        "kind": "돌발AS", "keys": ("접수ID", "프로젝트NO"), "public_id": "접수ID",
        "project": "프로젝트NO", "camp": "캠프명", "status": "진행상태",
    },
    "04_정기점검": {
        "kind": "정기점검", "keys": ("점검ID", "프로젝트NO"), "public_id": "점검ID",
        "project": "프로젝트NO", "camp": "캠프명", "status": "점검상태",
    },
    "06_거래서류청구수금": {
        "kind": "정산", "keys": ("정산ID", "원천업무ID", "프로젝트NO"),
        "public_id": "정산ID", "project": "프로젝트NO", "camp": "캠프명",
        "status": "청구상태",
    },
    # 15·16시트는 같은 정산ID를 공유하므로 public_id 전역 유일키로 쓰지 않는다.
    "15_세금계산서관리": {
        "kind": "세금계산서", "keys": ("정산ID",), "public_id": None,
        "project": "프로젝트NO", "camp": "캠프명", "status": "발행상태",
    },
    "16_입금수금관리": {
        "kind": "입금수금", "keys": ("정산ID",), "public_id": None,
        "project": "프로젝트NO", "camp": "캠프명", "status": "수금상태",
    },
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).strip()


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    # Excel의 [h]:mm / 경과시간 셀은 openpyxl에서 timedelta로 온다.
    # 초 숫자로 바꾸면 사람이 읽기 어려우므로 Excel의 의미를 알아볼 수 있는
    # 문자열로 보관한다. 시각 셀도 같은 원칙으로 ISO 표기를 쓴다.
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, time_value):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def sha256_file(path: os.PathLike[str] | str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _header(ws, spec: Mapping[str, Any], scan_rows: int = 15):
    """행 번호 하드코딩 없이 키 열이 있는 가장 풍부한 머리글을 고른다."""
    candidates = []
    keys = set(spec["keys"])
    for row_no, values in enumerate(
            ws.iter_rows(min_row=1, max_row=min(scan_rows, ws.max_row), values_only=True), 1):
        labels = [_text(v) for v in values]
        nonempty = sum(bool(v) for v in labels)
        key_hits = sum(v in keys for v in labels)
        if key_hits:
            candidates.append((key_hits, nonempty, row_no, labels))
    if not candidates:
        raise ValueError(f"{ws.title}: 식별 키 머리글({', '.join(keys)})을 찾지 못함")
    _hits, _width, row_no, labels = max(candidates)
    return row_no, {label: idx for idx, label in enumerate(labels) if label}


def read_candidates(master: os.PathLike[str] | str) -> Dict[str, Any]:
    """Excel을 read_only/data_only로 한 번 읽어 정규화된 후보를 만든다."""
    import openpyxl

    book = openpyxl.load_workbook(master, read_only=True, data_only=True)
    result: Dict[str, Any] = {"rows": [], "sheets": {}, "blocking": []}
    try:
        for sheet, spec in SPECS.items():
            if sheet not in book.sheetnames:
                result["blocking"].append(f"필수 시트 없음: {sheet}")
                continue
            ws = book[sheet]
            try:
                header_row, columns = _header(ws, spec)
            except ValueError as exc:
                result["blocking"].append(str(exc))
                continue
            seen: Dict[str, str] = {}
            count = 0
            for row_no, values in enumerate(
                    ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                fields = {
                    name: _json_value(values[idx])
                    for name, idx in columns.items()
                    if idx < len(values) and values[idx] not in (None, "")
                }
                key_col = next((name for name in spec["keys"] if _text(fields.get(name))), "")
                if not key_col:
                    continue
                business_key = _text(fields[key_col])
                fingerprint = sha256_json(fields)
                if business_key in seen:
                    if seen[business_key] != fingerprint:
                        result["blocking"].append(
                            f"{sheet}: 중복키 값 충돌 {business_key} (행 {row_no})")
                    continue
                seen[business_key] = fingerprint
                result["rows"].append({
                    "sheet": sheet,
                    "row_number": row_no,
                    "kind": spec["kind"],
                    "business_key_col": key_col,
                    "business_key": business_key,
                    "public_id": _text(fields.get(spec["public_id"])) if spec["public_id"] else None,
                    "project_no": _text(fields.get(spec["project"])),
                    "camp_name": _text(fields.get(spec["camp"])),
                    "status": _text(fields.get(spec["status"])),
                    "fields": fields,
                    "fields_sha256": fingerprint,
                })
                count += 1
            result["sheets"][sheet] = {"header_row": header_row, "rows": count}
    finally:
        book.close()
    result["row_count"] = len(result["rows"])
    result["rows_sha256"] = sha256_json(result["rows"])
    return result


def _setting(store: AppStore, key: str, value: Any, actor: str, idem: str) -> None:
    current = store.get_setting(key)
    if current.get("value") == value:
        return
    store.set_setting(
        key, value, expected_version=int(current.get("record_version") or 0),
        actor=actor, idempotency_key=idem,
    )


def _write_report(payload: Mapping[str, Any]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = REPORT_DIR / f"app_db_cutover_{stamp}.json"
    temp = target.with_suffix(".json.tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, target)
    latest = REPORT_DIR / "app_db_cutover_latest.json"
    latest_temp = latest.with_suffix(".json.tmp")
    latest_temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(latest_temp, latest)
    return target


def cutover(master: os.PathLike[str] | str, db_path: Optional[str] = None,
            *, apply: bool = False) -> Dict[str, Any]:
    master = str(Path(master).resolve())
    # SMB 원본 메타데이터도 시작할 때 한 번만 고정한다. 행마다 getmtime을 다시
    # 조회하면 이미 워크북을 다 읽은 뒤 공유폴더가 순간 끊긴 것만으로 이관 전체가
    # 실패한다. 이 시각은 모든 행이 같은 원본 스냅샷에서 왔다는 증거이기도 하다.
    master_observed_at = datetime.fromtimestamp(
        os.path.getmtime(master)
    ).isoformat(timespec="seconds")
    master_sha = sha256_file(master)
    candidates = read_candidates(master)
    summary: Dict[str, Any] = {
        "format": "csos-db-cutover/v1",
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "mode": "apply" if apply else "plan",
        "master": master,
        "master_sha256": master_sha,
        "candidate_rows": candidates["row_count"],
        "candidate_sha256": candidates["rows_sha256"],
        "sheets": candidates["sheets"],
        "blocking": list(candidates["blocking"]),
    }
    if not apply:
        summary["status"] = "blocked" if summary["blocking"] else "ready"
        summary["report"] = str(_write_report(summary))
        return summary
    if summary["blocking"]:
        summary["status"] = "blocked"
        summary["report"] = str(_write_report(summary))
        return summary

    store = AppStore(db_path).initialize()
    backup = ROOT / "db" / "backups" / f"app_store_pre_cutover_{datetime.now():%Y%m%d_%H%M%S}.db"
    store.backup_to(backup)
    import_id = "xlsx-" + master_sha[:20]
    _setting(store, "source_of_truth_mode", "shadow_compare", "cutover",
             f"cutover:{master_sha}:mode-shadow")
    statuses = Counter()
    imported_ids = set()
    parity_errors = []
    # 1,853행을 행마다 FULL fsync하면 수십 분이 걸린다. 초기 이관은 한 원본 해시의
    # 단일 사건이므로 하나의 원자 트랜잭션으로 묶는다. 중간 실패 시 신규 행 전체가
    # 롤백되고, 이미 앞선 재시도에서 확정된 행은 idempotency replay로 건너뛴다.
    imported_by_locator: Dict[tuple[str, int], str] = {}
    with store.transaction() as batch_conn:
        for row in candidates["rows"]:
            response = store.shadow_import(
                import_id=import_id,
                sheet=row["sheet"],
                business_key=row["business_key"],
                business_key_col=row["business_key_col"],
                row_number=row["row_number"],
                kind=row["kind"],
                fields=row["fields"],
                public_id=row["public_id"] or None,
                project_no=row["project_no"] or None,
                camp_name=row["camp_name"] or None,
                status=row["status"],
                source_file=master,
                source_sha256=master_sha,
                observed_at=master_observed_at,
                evidence="앱 DB 컷오버 Excel 읽기 전용 원본",
                apply_if_missing=True,
                actor="cutover",
                idempotency_key=f"{import_id}:{row['sheet']}:{row['row_number']}",
                _conn=batch_conn,
            )
            statuses[str(response.get("status") or "unknown")] += 1
            work_id = response.get("work_id")
            if work_id:
                imported_ids.add(work_id)
                imported_by_locator[(row["sheet"], row["row_number"])] = work_id

    # parity는 커밋 뒤 읽기 전용 연결 하나로 전수 검사한다. 행마다 새 연결을 열지 않는다.
    with store.reader() as parity_conn:
        for row in candidates["rows"]:
            work_id = imported_by_locator.get((row["sheet"], row["row_number"]))
            if not work_id:
                continue
            stored = store._work_from_conn(parity_conn, work_id)
            stored_fields = {key: stored["fields"].get(key) for key in row["fields"]}
            if sha256_json(stored_fields) != row["fields_sha256"]:
                parity_errors.append(
                    f"{row['sheet']}!{row['row_number']} {row['business_key']}: 필드 해시 불일치")

    conflicts = store.list_import_conflicts(import_id=import_id)
    if len(imported_ids) != candidates["row_count"]:
        parity_errors.append(
            f"후보 {candidates['row_count']}행 중 DB 연결 {len(imported_ids)}행")
    blocking = [*parity_errors, *(
        f"충돌 {item['sheet']}!{item.get('row_number')} {item.get('field_key')}: {item['reason']}"
        for item in conflicts
    )]
    if not blocking:
        _setting(store, "source_of_truth_mode", "db_primary_export", "cutover",
                 f"cutover:{master_sha}:mode-primary")
        _setting(store, "legacy_excel_source", {
            "path": master, "sha256": master_sha, "import_id": import_id,
            "rows": candidates["row_count"], "cutover_at": datetime.now().isoformat(timespec="seconds"),
            "backup": str(backup), "direction": "db-to-excel-only",
        }, "cutover", f"cutover:{master_sha}:source")
    summary.update({
        "status": "complete" if not blocking else "blocked",
        "import_id": import_id,
        "backup": str(backup),
        "import_status": dict(statuses),
        "db_rows": len(imported_ids),
        "conflicts": len(conflicts),
        "blocking": blocking,
        "source_of_truth_mode": "db_primary_export" if not blocking else "shadow_compare",
        "finished_at": datetime.now().isoformat(timespec="seconds"),
    })
    summary["report"] = str(_write_report(summary))
    return summary


def self_test() -> bool:
    assert _text(datetime(2026, 8, 10, 9, 5)) == "2026-08-10"
    assert _json_value(10.0) == 10
    assert _json_value(date(2026, 8, 10)) == "2026-08-10"
    assert _json_value(timedelta(hours=9, minutes=30)) == "9:30:00"
    assert _json_value(time_value(9, 30)) == "09:30:00"
    assert len(SPECS) == 5 and SPECS["06_거래서류청구수금"]["kind"] == "정산"
    print("db_cutover self-test: OK")
    return True


def _resolve_master() -> str:
    from ecount_reconcile import load_config, resolve_master
    return resolve_master(load_config()["reconcile"]["master_xlsx"])


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--master")
    parser.add_argument("--db")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args(list(argv) if argv is not None else None)
    if args.self_test:
        return 0 if self_test() else 1
    result = cutover(args.master or _resolve_master(), args.db, apply=args.apply)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result.get("status") in {"ready", "complete"} else 2


if __name__ == "__main__":
    sys.exit(main())
