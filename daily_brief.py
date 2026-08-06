# -*- coding: utf-8 -*-
"""
daily_brief.py — 대표 보고용 **내용 중심** 일일 브리핑을 원장에서 만든다
================================================================================
2026-07-28 대표 통화 지시(요지):

  "숫자를 나한테 보고하라는 게 아니야."
  · 돌발이 어제 몇 건 접수됐고 오늘 몇 건 처리할 건지
  · 처리한 건은 **무엇 때문에 갔는데 무슨 작업을 했는지**
  · **유상인데 무상 부분이 포함**됐거나, 돌발이지만 **무료로 해줄 수밖에 없었던** 사정
  · 갔더니 **뭐가 더 있어서 유상을 추가**한 건
  · 정기점검은 루틴대로 가고 있는지, 분기 몇 %인지, 마무리 전망
  · 정기점검 갔다가 **유료가 발생한** 건
  · 미처리 건이 있는지

즉 대표가 알고 싶은 것은 '건수'가 아니라 **판단이 필요한 사정**이다.
그 내용은 이미 원장에 있다(신청내용 100% · 유·무상 98% · 실제작업상세 86%).
여기서 그걸 문장으로 엮어 준다. 없는 건 지어내지 않고 '미기입'으로 남긴다.

  python daily_brief.py                # 어제(집계기준일) 기준 브리핑
  python daily_brief.py --date 2026-07-27
  python daily_brief.py --md           # reports/일일브리핑_YYYYMMDD.md
"""
import sys, os, re, json
from datetime import datetime, date, timedelta
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

from pct_fmt import pct, pct_text          # 비율 표기 단일 규칙 (2026-08-05 지시)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(ROOT, "reports")
FREE = ("무상", "보험")
APP_YEAR = "2026"
MANUAL_EVENTS = os.path.join(REPORT_DIR, "manual_daily_events.json")
PM_SCHEDULE_REPORT = os.path.join(REPORT_DIR, "pm_schedule_sync.json")


def _d(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    m = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", str(v or ""))
    return "%s-%02d-%02d" % (m.group(1), int(m.group(2)), int(m.group(3))) if m else ""


def _s(v):
    return str(v).strip() if v not in (None, "") else ""


def load_manual_events():
    """프로젝트NO가 아직 없는 택배·부품발송 같은 당일 처리 근거.

    현장 AS 완료와 섞지 않고 별도 '업무 처리'로 보고한다. 파일이 없거나 깨져도
    원장 브리핑은 그대로 만들어져야 하므로 빈 목록으로 안전하게 돌아간다.
    """
    try:
        raw = json.load(open(MANUAL_EVENTS, encoding="utf-8"))
        rows = raw.get("events", []) if isinstance(raw, dict) else raw
        return [r for r in rows if isinstance(r, dict)]
    except (OSError, ValueError, TypeError):
        return []


def load_pm_schedule_report():
    """류지영 원본 스케줄의 전체 분기 장비 계획과 예측일 캐시."""
    try:
        raw = json.load(open(PM_SCHEDULE_REPORT, encoding="utf-8"))
        return raw if isinstance(raw, dict) else {}
    except (OSError, ValueError, TypeError):
        return {}


def load(master=None):
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = master or resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)

    def rows(sh):
        if sh not in wb.sheetnames:
            return []
        ws = wb[sh]
        hdr = [_s(h) for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        out = []
        for rn, r in enumerate(ws.iter_rows(min_row=5, values_only=True), 5):
            g = {h: (r[i] if i < len(r) else None) for h, i in ix.items()}
            g["_원천행"] = rn
            # 유수비 대표가 올린 M_순천1처럼 프로젝트NO 배정 전이라도 날짜·캠프·내용이
            # 있는 실제 접수는 브리핑에서 버리지 않는다. 빈 양식 행은 계속 제외한다.
            meaningful_orphan = (_d(g.get("접수일자")) or _d(g.get("점검예정일"))) and (
                _s(g.get("캠프명")) or _s(g.get("신청내용")) or _s(g.get("점검내용")))
            if _s(g.get("프로젝트NO")) or meaningful_orphan:
                out.append(g)
        return out

    d = {"as": rows("02_돌발AS접수"), "pm": rows("04_정기점검"),
         "fw": rows("03_현장작업실적"), "events": load_manual_events(),
         "pm_schedule": load_pm_schedule_report(),
         # 대표 브리핑이 일지 원본과 대조할 때 같은 관리대장만 읽게 한다.
         # 합성 데이터로 직접 brief()를 부르는 검증은 이 키가 없어 원본 접근을 하지 않는다.
         "_master": master}
    wb.close()
    return d, master


def _elapsed_pct(start, end, day):
    """기간이 몇 % 지났나 — 진행률이 '빠른지 늦은지'의 기준선.

    분기 절반이 지났는데 진행률이 30%면 늦은 것이고, 90%면 빠른 것이다. 이 기준선이
    없으면 대표가 매번 "지금 며칠째지?"를 머리로 계산해야 한다(2026-08-04 지시).
    """
    try:
        d = datetime.strptime(str(day)[:10], "%Y-%m-%d").date()
    except Exception:
        return 0
    total = (end - start).days + 1
    if total <= 0:
        return 0
    gone = min(max((d - start).days + 1, 0), total)
    return pct(gone, total) or 0


def brief(day=None, data=None):
    """하루치 브리핑. 반환값은 화면·이미지·문서 어디서나 같은 내용을 쓰도록 구조화한다."""
    data = data or load()[0]
    day = day or (date.today() - timedelta(days=1)).isoformat()
    if not str(day).startswith(APP_YEAR + "-"):
        data = {"as": [], "pm": [], "fw": []}

    def in_year(r, date_key, id_key):
        d = _d(r.get(date_key))
        if d:
            return d.startswith(APP_YEAR + "-")
        rid = _s(r.get(id_key))
        m = re.search(r"(?<![A-Za-z0-9])(?:AS|PM)-(\d{2})\d{2}(?:-|$)", rid)
        if m:
            return m.group(1) == APP_YEAR[-2:]
        return bool(re.fullmatch(r"UJ26\d{5}", _s(r.get("프로젝트NO")), re.I))

    A = [r for r in data["as"] if in_year(r, "접수일자", "접수ID")]
    P = [r for r in data["pm"] if in_year(r, "점검예정일", "점검ID")]
    E = [r for r in data.get("events", [])
         if _d(r.get("날짜")) == day and _d(r.get("날짜")).startswith(APP_YEAR + "-")]
    projects = {_s(r.get("프로젝트NO")) for r in A + P if _s(r.get("프로젝트NO"))}
    F = [r for r in data["fw"] if _s(r.get("프로젝트NO")) in projects]

    # 03시트(실제 작업 내용)를 프로젝트별로 붙여 둔다 — '무슨 작업을 했는지'가 여기 있다
    work = {}
    for r in F:
        k = _s(r.get("프로젝트NO"))
        if k and k not in work:
            work[k] = r

    def line(r, donec, kindc):
        """한 건을 '왜 갔고 무엇을 했고 유·무상은 어떻게'로 적는다."""
        prj = _s(r.get("프로젝트NO"))
        w = work.get(prj, {})
        why = _s(r.get("신청내용")) or _s(r.get("점검내용"))
        did = _s(w.get("실제작업상세")) or _s(w.get("실제작업항목"))
        cost = _s(r.get("유상·무상·보험")) or _s(w.get("비용구분"))
        extra = _s(r.get("최초접수외추가작업")) or _s(r.get("추가작업내용")) or _s(w.get("추가작업내용"))
        return {"프로젝트NO": prj, "캠프명": _s(r.get("캠프명")), "담당기사": _s(r.get("담당기사")),
                # 앱 목록에서 프로젝트NO가 중복돼도 정확한 원천 행으로 이동하도록
                # 화면 전용 레코드 식별자를 함께 보낸다.
                "레코드ID": _s(r.get("접수ID")) or _s(r.get("점검ID")),
                "레코드종류": "as" if kindc == "돌발AS" else "pm",
                "왜": why, "무엇": did, "비용": cost, "추가작업": extra,
                # 집계 숫자만으로는 대표가 "이상 3건이 무엇인가"를 알 수 없다.
                # 원천 열을 그대로 보존해 화면·캡처에서 건별 이상/조치 내용을 보여 준다.
                # 값이 비었을 때 다른 설명을 이상내용으로 추정하지 않는다.
                "이상내용": _s(r.get("이상내용")),
                "문제내용": _s(r.get("문제내용")),
                "조치내용": _s(r.get("조치내용")),
                "일자": _d(r.get(donec)), "구분": kindc,
                # ★ 사용자 지시(2026-07-28): "각 항목 옆에 날짜 붙여줘 — 금일이라고 하면
                #   어떤 날짜인지 헷갈림." 완료건은 '언제 접수돼 언제 끝났는지'가 같이 보여야
                #   대표가 "그거 언제 들어온 건데?" 를 되묻지 않는다.
                "접수일": _d(r.get("접수일자")) or _d(r.get("점검예정일")) or _d(r.get("요청일"))}

    def event_line(r):
        """택배·부품발송은 AS 완료가 아니라 그날 실제 처리한 업무로 분리한다."""
        return {
            "프로젝트NO": _s(r.get("프로젝트NO")),
            "캠프명": _s(r.get("캠프명")),
            "담당기사": _s(r.get("처리자")),
            "게시자": _s(r.get("게시자")),
            "왜": _s(r.get("신청내용")),
            "무엇": _s(r.get("처리내용")) or _s(r.get("상태")),
            "비용": _s(r.get("비용")),
            "추가작업": "",
            "일자": _d(r.get("날짜")),
            "구분": "업무처리",
            "접수일": _d(r.get("접수일")),
            "상태": _s(r.get("상태")),
            "근거": _s(r.get("근거")),
            "레코드ID": _s(r.get("레코드ID")),
            "레코드종류": _s(r.get("레코드종류")),
        }

    def pm_line(r, primary="plan"):
        """정기점검 한 건 — 예정일과 실행일을 함께 보존해 어느 목록에서도 같은 건을 연다."""
        actual = _d(r.get("실제점검일"))
        scheduled = _d(r.get("점검예정일"))
        x = line(r, "실제점검일" if primary == "done" else "점검예정일", "정기점검")
        raw = _s(r.get("점검상태"))
        state = "실행" if actual else (raw if raw in ("취소", "보류") else "미실행")
        x.update({
            "예정일": scheduled,
            "실행일": actual,
            "상태": state,
            "일자": actual if primary == "done" else scheduled,
        })
        return x

    def ordered(rows):
        """앱·캡처도 관리대장 원칙대로 과거→최근, 같은 날은 점검ID 순."""
        return sorted(rows, key=lambda x: (x.get("일자") or "9999", x.get("레코드ID") or ""))

    # ── 돌발AS ──
    def unique(rows, key):
        out, seen = [], set()
        for r in rows:
            k = key(r)
            if k in seen:
                continue
            seen.add(k)
            out.append(r)
        return out

    # 같은 프로젝트·같은 날·같은 신청내용이 새 접수ID로 중복 적재된 실제 사례가 있다.
    # 접수ID 개수 그대로 세면 2026-08-04는 5건으로 보이지만 실업무는 3건이다.
    as_new = unique(
        [r for r in A if _d(r.get("접수일자")) == day],
        lambda r: (_s(r.get("프로젝트NO")) or _s(r.get("접수ID")),
                   _d(r.get("접수일자")), re.sub(r"\s+", "", _s(r.get("신청내용")))),
    )
    as_done = [r for r in A if _d(r.get("작업완료일")) == day]
    fw_day = unique(
        [r for r in F if _d(r.get("작업일자")) == day],
        lambda r: (_s(r.get("프로젝트NO")), _d(r.get("작업일자")),
                   _s(r.get("실제작업상세")) or _s(r.get("실제작업항목"))),
    )
    fw_projects = {_s(r.get("프로젝트NO")) for r in fw_day if _s(r.get("프로젝트NO"))}
    new_processed = [r for r in as_new if _d(r.get("작업완료일")) == day
                     or _s(r.get("프로젝트NO")) in fw_projects]
    paid_as = [r for r in as_done if "유상" in (
        _s(r.get("유상·무상·보험")) or _s(work.get(_s(r.get("프로젝트NO")), {}).get("비용구분"))
    )]
    revisit = [r for r in A if _d(r.get("방문예정일")) == day
               and _s(r.get("재방문여부")) not in ("아니오", "없음", "N")]
    # ★ '완료일이 없다'와 '아직 안 갔다'는 다르다. 오래된 건은 거의 다 **기록 누락**이다
    #   (2026-07-28 확인: 84건 중 46건이 5월 이전). 둘을 뭉뚱그려 '미처리 84건'이라고
    #   보고하면 대표가 놀라고, 반대로 '없다'고 하면 최근 건을 놓친다. 갈라서 말한다.
    _open = [r for r in A if not _d(r.get("작업완료일"))
             and _s(r.get("진행상태")) not in ("취소", "보류")]
    _cut = (datetime.strptime(day, "%Y-%m-%d").date() - timedelta(days=30)).isoformat()
    as_open = [r for r in _open if (_d(r.get("접수일자")) or "0000") >= _cut]   # 최근 = 진짜 미처리
    as_stale = [r for r in _open if (_d(r.get("접수일자")) or "0000") < _cut]   # 오래됨 = 완료일 미기입
    # ── 정기점검 ──
    pm_done = [r for r in P if _d(r.get("실제점검일")) == day]
    pm_plan = [r for r in P if _d(r.get("점검예정일")) == day]

    done = [line(r, "작업완료일", "돌발AS") for r in as_done] + \
           [line(r, "실제점검일", "정기점검") for r in pm_done]

    # 대표가 콕 집어 물은 것들
    free = [x for x in done if any(k in x["비용"] for k in FREE)]
    extra = [x for x in done if x["추가작업"]]
    pm_paid = [line(r, "실제점검일", "정기점검") for r in pm_done
               if _s(r.get("유상추가작업발생")) in ("Y", "예", "발생", "있음")]
    to_as = [line(r, "실제점검일", "정기점검") for r in pm_done
             if _s(r.get("돌발AS전환여부")) in ("전환", "Y", "예")]
    abnormal = [line(r, "실제점검일", "정기점검") for r in pm_done
                if _s(r.get("이상발견여부")) in ("있음", "Y", "예")]

    # 정기점검 진행률 — 분기 기준(대표가 "이 분기 몇 프로"를 물었다)
    y = int(day[:4]); q = (int(day[5:7]) - 1) // 3 + 1
    qs, qe = date(y, 3 * q - 2, 1), (date(y + (q == 4), (3 * q) % 12 + 1, 1) - timedelta(days=1))
    inq = [r for r in P if qs.isoformat() <= (_d(r.get("점검예정일")) or "9999") <= qe.isoformat()]
    inq_done = [r for r in inq if _d(r.get("실제점검일"))]
    pm_plan_rows = ordered([pm_line(r, "plan") for r in pm_plan])
    pm_done_rows = ordered([pm_line(r, "done") for r in pm_done])
    pm_quarter_rows = ordered([pm_line(r, "plan") for r in inq])

    # ★ 분기 진행률의 분모는 04시트에 이미 들어온 프로젝트가 아니라 류지영 원본의
    # **전체 장비 대수**다. 04만 보면 완료된 행 위주라 54/58=93%로 과대 표시됐다.
    schedule_report = data.get("pm_schedule") if isinstance(data, dict) else {}
    source_schedule = []
    if (isinstance(schedule_report, dict)
            and int(schedule_report.get("year") or 0) == y
            and int(schedule_report.get("quarter") or 0) == q):
        source_schedule = [r for r in (schedule_report.get("schedule") or [])
                           if isinstance(r, dict)]

    def units(r):
        try:
            return max(0, int(r.get("장비수") or 0))
        except (TypeError, ValueError):
            return 0

    source_total = sum(units(r) for r in source_schedule)
    source_done = sum(units(r) for r in source_schedule
                      if _d(r.get("실제점검일")) and _d(r.get("실제점검일")) <= day)
    source_due = sum(units(r) for r in source_schedule
                     if (_d(r.get("점검예정일")) or _d(r.get("예측점검일")))
                     and (_d(r.get("점검예정일")) or _d(r.get("예측점검일"))) <= day)
    source_due_done = sum(units(r) for r in source_schedule
                          if (_d(r.get("점검예정일")) or _d(r.get("예측점검일")))
                          and (_d(r.get("점검예정일")) or _d(r.get("예측점검일"))) <= day
                          and _d(r.get("실제점검일")) and _d(r.get("실제점검일")) <= day)
    source_today = [r for r in source_schedule
                    if (_d(r.get("점검예정일")) or _d(r.get("예측점검일"))) == day]
    source_done_today = [r for r in source_schedule if _d(r.get("실제점검일")) == day]
    # ★ 원본이 그날을 아직 담고 있지 않으면 **원장으로 보충한다** (2026-08-06 실사고).
    #   류지영 원본이 정본이라 원본이 있으면 원장을 안 봤는데, 원본이 8/3 까지만
    #   갱신돼 있어 8/5 정기점검 완료 3건이 보고에 **0건**으로 나갔다. 원본을 무시하는
    #   것이 아니라, 원본이 아직 다루지 않는 날짜만 원장이 채운다 — 그날 원본에 한 건도
    #   없을 때만 갈아탄다(원본에 일부라도 있으면 그것이 정본이다).
    use_ledger_today = bool(source_schedule) and not source_today and not source_done_today
    if use_ledger_today:
        source_today, source_done_today = [], []
    # 비율은 pct_fmt 규칙(소수점 1자리·미완료는 100% 금지)만 쓴다 — 2026-08-05 지시.
    source_progress = pct(source_done, source_total) or 0
    source_due_rate = pct(source_due_done, source_due) or 0

    # 정기점검·돌발AS 일지는 완료 실적과 미실시 사유를 함께 적는 현장 정본이다.
    # 원장만으로는 '왜 아직 안 됐는지'가 보이지 않으므로 대표 보고에는 이 대조 결과도
    # 붙인다. 단 합성 검증처럼 master 경로가 없는 호출은 외부 원본에 닿지 않는다.
    worklog = {}
    master_hint = data.get("_master") if isinstance(data, dict) else None
    if master_hint:
        try:
            import work_log_sync as WLS
            worklog = WLS.analyze(master_hint).get("요약", {})
        except Exception:
            worklog = {}

    # 내용이 비어 있으면 숨기지 않고 '미기입'으로 남긴다 — 채워야 할 칸이다
    blank = [x for x in done if not x["무엇"]]
    handled = [event_line(r) for r in E]

    return {
        "기준일": day,
        "돌발AS": {"신규접수": len(as_new), "신규처리완료": len(new_processed),
                    "신규처리율": pct(len(new_processed), len(as_new)) or 0,
                    "완료": len(as_done), "현장작업": len(fw_day),
                    "유상발생": len(paid_as), "재방문예정": len(revisit),
                    "미처리": len(as_open), "완료일미기입": len(as_stale),
                    "업무처리": len(handled)},
        # 그날치는 원본 우선, 원본이 아직 그날을 안 담았으면 원장(04시트)으로 센다.
        "정기점검": {"예정": len(source_today) if (source_schedule and not use_ledger_today) else len(pm_plan),
                     "예정장비": sum(units(r) for r in source_today) if (source_schedule and not use_ledger_today) else len(pm_plan),
                     "완료": len(source_done_today) if (source_schedule and not use_ledger_today) else len(pm_done),
                     "완료장비": sum(units(r) for r in source_done_today) if (source_schedule and not use_ledger_today) else len(pm_done),
                     # ★ 사용자 지시(2026-07-29): "3분기라고 하면 모르겠고 몇월부터
                     #   몇월까지인지로 표기해줘." 분기 번호는 읽는 사람이 다시 환산해야 한다.
                     "분기": f"{y}년 {3 * q - 2}~{3 * q}월",
                     "분기범위": f"{3 * q - 2}~{3 * q}월", "분기끝월": f"{3 * q}월",
                     "분기예정": source_total or len(inq),
                     "분기완료": source_done if source_total else len(inq_done),
                     "분기미실행": (source_total - source_done) if source_total else len(inq) - len(inq_done),
                     "분기진행률": source_progress if source_total else (
                         pct(len(inq_done), len(inq)) or 0),
                     "분기일정그룹": len(source_schedule) if source_schedule else len(inq),
                     "기준일까지예정": source_due if source_total else sum(
                         1 for r in inq if (_d(r.get("점검예정일")) or "9999") <= day),
                     "기준일까지완료": source_due_done if source_total else sum(
                         1 for r in inq if (_d(r.get("점검예정일")) or "9999") <= day
                         and _d(r.get("실제점검일"))),
                     "기준일이행률": source_due_rate if source_total else 0,
                     "예측일정": sum(bool(r.get("예측점검일")) for r in source_schedule),
                     # ★ 대표 지시(2026-08-04): "이번 분기에 **일수를 따졌을 때** 몇 %
                     #   진행됐고 이상이 있는지 없는지". 진행률만 보면 빠른지 늦은지 모른다.
                     #   분기 경과일 비율을 '기대 진행률'로 두고 그 차이를 보여 준다.
                     "분기경과율": _elapsed_pct(qs, qe, day),
                     "분기진행격차": round((source_progress if source_total else
                                     (pct(len(inq_done), len(inq)) or 0))
                                     - _elapsed_pct(qs, qe, day), 1)},
        "완료내역": done, "무상건": free, "추가작업건": extra,
        "점검중유상": pm_paid, "AS전환": to_as, "이상발견": abnormal,
        "점검예정목록": pm_plan_rows,
        "점검실행목록": pm_done_rows,
        "분기점검목록": pm_quarter_rows,
        "내용미기입": blank,
        "당일처리목록": handled,
        "현장작업목록": [line(r, "작업일자", "돌발AS") for r in fw_day],
        "신규처리완료목록": [line(r, "작업완료일", "돌발AS") for r in new_processed],
        "재방문예정목록": [line(r, "방문예정일", "돌발AS") for r in revisit],
        "분기원본일정목록": source_schedule,
        "완료일미기입목록": [line(r, "접수일자", "돌발AS") for r in as_stale],
        "신규목록": [line(r, "접수일자", "돌발AS") for r in as_new],
        "일지대조": worklog,
    }


def md(s, base=""):
    """2026-07-27 → 07-27. 기준일과 해가 다르면 연도까지 적는다."""
    s = _s(s)
    if not s:
        return ""
    return s[5:] if base and s[:4] == base[:4] else s


def span(x, base=""):
    """완료건 꼬리표 — 접수 → 완료 경과일. 묵은 건이 눈에 띄게 한다."""
    a, b_ = x.get("접수일"), x.get("일자")
    if not (a and b_) or a == b_:
        return ""
    try:
        n = (datetime.strptime(b_, "%Y-%m-%d") - datetime.strptime(a, "%Y-%m-%d")).days
    except ValueError:
        return ""
    return f" (접수 {md(a, base)} · {n}일 만)" if n > 0 else ""


def tag(x, base=""):
    """항목 머리표 — '날짜 · 프로젝트NO · 캠프' 순서로 고정한다.

    ★ 사용자 지시(2026-07-28): "각 항목 옆에 날짜 붙여줘, 금일이라고 하면
      어떤 날짜인지 헷갈림." 어느 줄을 읽어도 날짜가 먼저 나오게 한다."""
    return f"{md(x.get('일자'), base) or '날짜미기입'} · {x.get('프로젝트NO') or '번호미기입'} · {x.get('캠프명') or '캠프미기입'}"


def text(b):
    """대표에게 그대로 읽어 드릴 수 있는 문장."""
    L = []
    d, a, p = b["기준일"], b["돌발AS"], b["정기점검"]
    wd = "월화수목금토일"[datetime.strptime(d, "%Y-%m-%d").weekday()]
    # ★ '금일·당일'이라는 말은 읽는 사람마다 다른 날을 떠올린다. 맨 위에 날짜를 못 박는다.
    L.append(f"[{d}({wd}) 실적 — 아래 날짜는 모두 실제 날짜입니다]")
    L.append(f"■ 돌발AS — 신규 접수 {a['신규접수']}건 중 {d} 처리 "
             f"{a['신규처리완료']}건({a['신규처리율']}%) · 전체 완료 {a['완료']}건 · "
             f"미처리 {a['미처리']}건(최근 30일)")
    # ★ 신규와 완료가 같은 모양으로 나오면 어느 게 처리된 건지 안 보인다.
    #   대표가 묻는 건 "완료한 건 무슨 작업을 했느냐"이므로 완료는 따로, 더 자세히 적는다.
    if b["신규목록"]:
        L.append(f"  ▸ 새로 접수 {len(b['신규목록'])}건  (접수일 {md(d, d)})")
        for x in b["신규목록"][:6]:
            L.append(f"      {tag(x, d)}"
                     + (f" · {x['담당기사']}" if x["담당기사"] else " · 기사 미배정"))
            L.append(f"         내용 : {x['왜'][:52] or '접수내용 미기입'}")

    doneA = [y for y in b["완료내역"] if y["구분"] == "돌발AS"]
    if doneA:
        L.append(f"  ▸ 완료 {len(doneA)}건 — 무엇 때문에 갔고 무슨 작업을 했는지  (완료일 {md(d, d)})")
        for x in doneA[:8]:
            head = (f"      {tag(x, d)}{span(x, d)}"
                    f" · {x['담당기사'] or '기사 미기입'} · {x['비용'] or '비용 미기입'}")
            L.append(head)
            L.append(f"         왜   : {x['왜'][:52] or '접수내용 미기입'}")
            # ★ 작업내용이 비었으면 조용히 넘기지 않는다. 대표가 그 자리에서 물어볼 항목이라
            #   '없다'는 사실 자체를 알고 있어야 한다.
            L.append(f"         작업 : {x['무엇'][:52]}" if x["무엇"]
                     else "         작업 : ★ 미기입 — 기사에게 확인해 03_현장작업실적에 입력 필요")
            if x["추가작업"]:
                L.append(f"         추가 : {x['추가작업'][:48]}")
    if b.get("당일처리목록"):
        L.append(f"  ▸ {d} 업무 처리 {len(b['당일처리목록'])}건 — 현장 AS 완료와 별도")
        for x in b["당일처리목록"]:
            L.append(f"      {tag(x, d)} · 처리 {x['담당기사'] or '담당 미기입'}"
                     + (f" · 게시 {x['게시자']}" if x.get("게시자") else ""))
            L.append(f"         요청 : {x['왜'][:52] or '신청내용 미기입'}")
            L.append(f"         처리 : {x['무엇'][:52] or '처리내용 미기입'}")

    L.append(f"\n■ 정기점검 — {md(d, d)} 완료 {p['완료']}그룹·{p['완료장비']}대 "
             f"(그날 예정 {p['예정']}그룹·{p['예정장비']}대)")
    L.append(f"   {p['분기']} 장비 진행률 {p['분기진행률']}% "
             f"(완료 {p['분기완료']}대 / 전체 예정 {p['분기예정']}대)")
    if p.get("기준일까지예정"):
        L.append(f"   기준일까지 예정 {p['기준일까지예정']}대 중 "
                 f"{p['기준일까지완료']}대 이행({p['기준일이행률']}%)")
    if p["분기예정"]:
        L.append("   " + (f"특별한 문제 없으면 {p['분기끝월']}까지 마무리 가능합니다."
                          if p["분기진행률"] >= 60 else "진행률이 낮아 일정 관리가 필요합니다."))

    if b["점검중유상"]:
        L.append(f"\n■ 정기점검 갔다가 유상 발생 {len(b['점검중유상'])}건")
        for x in b["점검중유상"][:5]:
            L.append(f"   {tag(x, d)} · {x['추가작업'][:36] or '내용 미기입'}")
    if b["무상건"]:
        L.append(f"\n■ 무상·보험 처리 {len(b['무상건'])}건 — 사유 확인 필요")
        for x in b["무상건"][:5]:
            L.append(f"   {tag(x, d)} · {x['왜'][:30]} [{x['비용']}]")
    if b["추가작업건"]:
        L.append(f"\n■ 접수 외 추가작업 {len(b['추가작업건'])}건")
        for x in b["추가작업건"][:5]:
            L.append(f"   {tag(x, d)} · {x['추가작업'][:40]}")
    if b["AS전환"]:
        L.append(f"\n■ 점검 중 돌발AS 전환 {len(b['AS전환'])}건")
    if a.get("완료일미기입"):
        # 접수일 범위를 적어 '언제부터 밀린 건지'를 한 줄로 보이게 한다
        ds = sorted(x["일자"] for x in b.get("완료일미기입목록", []) if x.get("일자"))
        rng = f" — 접수 {ds[0]} ~ {ds[-1]}" if ds else ""
        L.append(f"\n■ 완료일이 안 적힌 오래된 건 {a['완료일미기입']}건 (접수 후 30일 넘음){rng}")
        L.append("   실제로는 끝났을 가능성이 큽니다 — 완료일만 채우면 정리됩니다.")
    log = b.get("일지대조", {}) if isinstance(b, dict) else {}
    log_as = log.get("돌발AS", {}) if isinstance(log, dict) else {}
    if log_as:
        L.append("\n■ 정기점검·돌발AS 일지 대조 (7월 원본)")
        L.append(f"   돌발AS 발생 {log_as.get('발생', 0)}건 · 처리완료 {log_as.get('처리완료', 0)}건 · "
                 f"미처리 {log_as.get('미처리', 0)}건 · 취소 {log_as.get('취소', 0)}건")
        reasons = log_as.get("미처리사유", []) or []
        if reasons:
            L.append("   미처리 사유 : " + " · ".join(
                f"{x.get('사유', '기타')} {x.get('건수', 0)}건" for x in reasons[:5]))
        # ■ 현장 애로사항 — **누가 움직여야 풀리는가**로 묶어 보여 준다.
        #   (차동호 팀장 통화 2026-08-06: "자재보다 인력이 더 문제다")
        #   집계(책임구분)는 work_log_sync 가 이미 내고 있었는데 보고서에 자리가 없어
        #   대표는 '미처리 N건'만 봤다. 그러면 늘 현장이 미룬 것처럼 읽힌다.
        #   본사가 풀어야 할 것(인력·구매)과 현장이 잡아야 할 것(일정)을 갈라 놓는다.
        blockers = log_as.get("책임구분", []) or []
        if blockers:
            company = [x for x in blockers
                       if any(k in str(x.get("구분", "")) for k in ("인력", "구매", "조달"))]
            field = [x for x in blockers if x not in company
                     and str(x.get("구분", "")).startswith("확인 필요")]
            fair = [x for x in blockers if str(x.get("구분", "")).startswith("정당")]
            L.append("\n■ 현장 애로사항 — 못 한 이유를 '누가 풀어야 하나'로 나눔")
            if company:
                L.append("   [본사가 풀어야] " + " · ".join(
                    f"{x['구분']} {x['건수']}건" for x in company))
                L.append("      인력 배정·구매 승인이 막힌 건입니다. 현장을 재촉해도 풀리지 않습니다.")
            if field:
                L.append("   [현장이 잡아야] " + " · ".join(
                    f"{x['구분']} {x['건수']}건" for x in field))
            if fair:
                L.append("   [정당한 사유]   " + " · ".join(
                    f"{x['구분']} {x['건수']}건" for x in fair))
        if log_as.get("처리완료일확인"):
            L.append(f"   ※ 일지상 처리완료이나 완료일이 없는 건 {log_as['처리완료일확인']}건은 임의 완료일을 넣지 않고 확인 목록에 보관")
    if b["내용미기입"]:
        L.append(f"\n■ 작업 내용이 안 적힌 완료건 {len(b['내용미기입'])}건 — 기사에게 확인 필요")
        for x in b["내용미기입"][:5]:
            L.append(f"   {tag(x, d)} · {x['담당기사'] or '기사 미기입'} (완료 처리됨)")
    return "\n".join(L)


def _kakao_warning():
    """카톡 내보내기 1일 초과 시 브리핑 머리에 경고 한 줄(2026-08-04 지시).
    판정은 kakao/export_watch.py 가 남긴 JSON 만 읽는다."""
    try:
        k = json.load(open(os.path.join(ROOT, "reports", "카톡_내보내기_경과.json"),
                           encoding="utf-8"))
    except Exception:
        return ""
    if not k.get("stale"):
        return ""
    age = k.get("age_hours")
    d = f"{age/24:.1f}일 경과" if age else "파일 없음"
    return f"⚠ 카톡 내보내기 오래됨({d}) — 갱신 필요\n\n"


def main():
    args = sys.argv[1:]
    day = args[args.index("--date") + 1] if "--date" in args else None
    data, master = load()
    b = brief(day, data)
    warn = _kakao_warning()
    if warn:
        print(warn.strip())
    print(text(b))
    if "--md" in args:
        os.makedirs(REPORT_DIR, exist_ok=True)
        p = os.path.join(REPORT_DIR, f"일일브리핑_{b['기준일'].replace('-', '')}.md")
        open(p, "w", encoding="utf-8").write(
            f"# 일일 브리핑 {b['기준일']}\n\n{warn}원장: {os.path.basename(master)}\n\n```\n{text(b)}\n```\n")
        print("\n리포트:", p)


if __name__ == "__main__":
    main()
