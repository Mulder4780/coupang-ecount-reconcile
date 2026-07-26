# -*- coding: utf-8 -*-
"""
camp_fill.py — 캠프명이 빈 행을 **밴드 본문에서 찾아** 채운다
================================================================================
앱 카드에 '캠프 미상'으로 뜨는 건들이다. 어느 현장인지 모르면 계산서 구성 추정도,
캠프별 집계도 못 한다.

밴드 글은 이런 꼴이라 프로젝트NO와 캠프이름이 한 덩어리 안에 같이 있다:
    ♣ ［ 돌발유료 A/S 완료 ]
      ● 프로젝트NO : UJ2601132
      ● 캠프이름   : 창원1MB(팔용동)
한 글에 여러 건이 담기기도 해서 ♣·✅ 로 덩어리를 나눈 뒤 짝을 짓는다.

  python camp_fill.py           # 미리보기
  python camp_fill.py --apply   # ledger_writer 큐에 넣어 반영(빈칸만)
"""
import sys, os, re, json, glob

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

CACHE = os.path.join(ROOT, "band", "cache")
CAMP_RE = re.compile(r"캠프\s*이?름?\s*[:：]\s*([^\n●♣\[]{2,30})")
PRJ_RE = re.compile(r"UJ\d{7}")
LIST_RE = re.compile(r"\d+\.\s*(.+?)\(\d{1,2}/\d{1,2}\)\s*[:：][^\n]*?(UJ\d{7})")


def camp_book():
    """프로젝트NO → 캠프이름 (밴드 본문에서 추출)"""
    book = {}
    for f in glob.glob(os.path.join(CACHE, "*.json")):
        b = os.path.basename(f)
        if b.startswith(("dump_", "raw_")):
            continue
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for p in (d.get("posts") or {}).values():
            body = p.get("content") or ""
            for blk in re.split(r"[♣✅]", body):
                ps, cs = PRJ_RE.findall(blk), CAMP_RE.findall(blk)
                if ps and cs:
                    camp = cs[0].strip().rstrip("=·-").strip()
                    for x in ps:
                        book.setdefault(x, camp)
            # 매출처 밴드의 "세금계산서 발행 완료" 목록 글은 형식이 다르다:
            #   3. 송파1MB(감일동)(1/10) : 2R/T Mobile-lift 2EA 19,780,000원 / PO326259 UJ2501950
            # 캠프 이름 자체에 괄호가 있어 **맨 끝 (월/일)** 을 기준으로 잘라야 한다.
            for m in LIST_RE.finditer(body):
                book.setdefault(m.group(2), m.group(1).strip())
    return book


def plan():
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    book = camp_book()
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    items, missing = [], []
    for sheet in ("02_돌발AS접수", "04_정기점검", "05_신규납품설치", "06_거래서류청구수금"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
        if "프로젝트NO" not in idx or "캠프명" not in idx:
            continue
        for row in ws.iter_rows(min_row=5, values_only=True):
            g = lambda c: (row[idx[c]] if c in idx and idx[c] < len(row) else None)
            prj, camp = str(g("프로젝트NO") or "").strip(), g("캠프명")
            if not prj or camp:
                continue
            if prj in book:
                items.append({"sheet": sheet, "key": prj, "key_col": "프로젝트NO",
                              "col": "캠프명", "value": book[prj], "vtype": "text",
                              "only_if_empty": True, "evidence": "밴드 본문 캠프이름"})
            else:
                missing.append((sheet, prj))
    wb.close()
    return items, missing, len(book)


def main():
    items, missing, n = plan()
    print(f"밴드에서 캠프를 아는 프로젝트 {n}개")
    print(f"채울 수 있음 {len(items)}건 · 밴드에도 없어 못 채움 {len(missing)}건")
    for it in items[:5]:
        print(f"   {it['key']} → {it['value']}")
    if missing[:5]:
        print("  [남는 것]", ", ".join(p for _, p in missing[:8]))
    if "--apply" not in sys.argv:
        print("\n반영하려면: python camp_fill.py --apply")
        return
    if not items:
        return
    import ledger_writer as L, subprocess
    print("큐 추가:", L.queue_add(items))
    r = subprocess.run([sys.executable, os.path.join(ROOT, "ledger_writer.py"), "--apply"],
                       cwd=ROOT, capture_output=True, text=True,
                       encoding="utf-8", errors="replace",
                       env={**os.environ, "PYTHONIOENCODING": "utf-8"})
    for line in (r.stdout or "").splitlines():
        if "반영 완료" in line or "제외" in line:
            print(" ", line.strip())


if __name__ == "__main__":
    main()
