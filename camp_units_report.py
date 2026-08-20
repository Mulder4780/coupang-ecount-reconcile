# -*- coding: utf-8 -*-
"""쿠팡 캠프 정본 명단 + 캠프마다 들어간 **호기** 정리 (2026-08-20 지시).

형님 지시: "쿠팡 캠프 정확히 검증하고 합성해서 나열하고 그 옆에 그 캠프에
들어간 호기까지 나열해서 정리해봐".

★ 세는 곳이 셋이라 숫자가 셋이었다 — 그래서 **정본을 하나로 못 박는다**:
  · `정기점검 리스트(안전관리자).xlsx` 3분기 = 캠프명 있는 줄 160 · 고유 152
  · `(류지영) ★01 …원본.xlsx` = 분기 시트마다 고유 182 안팎(그중 몇은 주소가
    캠프명 칸에 들어온 줄이라 `_sched_camp_ok` 가 거른다)
  · 앱 `정기점검` 418 = 밴드·카톡에서 정기점검 글이 한 번이라도 잡힌 캠프
    (옛 이름·표기 변형 포함)
  정본은 **★01 워크북**이고, 한 캠프는 **가장 최신 분기 시트**의 것이 이긴다([326]).

★ **합치는 근거는 주소 하나다**([172]). 이름이 닮았다고 합치면
  `송파1MB(감일동)`·`송파5MB(감일동)` 처럼 **실재하는 다른 캠프**가 한 덩어리가 된다.
  그리고 **후보가 유일할 때만** 합친다 — 한 주소에 정본 캠프가 둘이면(서초1MB
  양재동A/B/C/D 처럼 진짜로 여러 곳) 어느 쪽인지 원본이 안 말해 주므로 안 합친다.

★ **호기의 뜻이 출처마다 다르다**([311]·[326]) — 정본에서 온 것은 **실제 설치 대수**
  (한 줄 = 한 대), 밴드에서 온 것은 '접수·점검 글에서 본 번호'다. 한 말로 뭉치면
  둘 중 하나에는 반드시 틀린 설명이 붙으므로 **열을 갈라** 싣는다.

★ **못 합친 것을 조용히 버리지 않는다**([169]) — 둘째 시트에 이유와 함께 남긴다.

읽기 전용이다 — 관리대장·앱 DB·큐에 한 글자도 안 쓴다. 새 xlsx 만 만든다.
"""
import datetime
import io
import json
import os
import re
import sys
import warnings

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "webapp"))

import camp_contacts as CC                                    # noqa: E402

OUT_XLSX = os.path.join(CC.REPORT_DIR, "쿠팡캠프_호기_정리.xlsx")
OUT_JSON = os.path.join(CC.REPORT_DIR, "쿠팡캠프_호기_정리.json")
MASTER_HINT = "★01"

# `_plain_xlsx` 는 (보이는 이름, 행의 칸 이름) 짝을 받는다 — 이름을 손으로 두 번
# 적으면 한쪽만 고쳐진다([162]). 그래서 이름 하나에서 짝을 만든다.
_C1 = ["캠프명", "지역", "주소", "호기수", "호기", "종류", "모델",
       "밴드에서 본 호기", "합친 다른 표기", "다른 표기 수", "근거분기"]
_C2 = ["캠프명", "주소", "왜 못 합쳤나", "최근작업일", "정기점검 글수"]
COLS = [(n, n) for n in _C1]
COLS_NO = [(n, n) for n in _C2]


def addr_key(value):
    """주소 비교 열쇠 — 띄어쓰기·괄호·하이픈만 흘려 읽는다.

    ⚠ 숫자는 한 글자도 안 건드린다. `202번길`과 `20번길`은 다른 곳이다.
    """
    return re.sub(r"[\s\-,()]+", "", str(value or ""))


def unit_no(value):
    """`01호기` -> 1. 못 읽으면 None — 지어내지 않는다([169])."""
    m = re.search(r"(\d{1,2})", str(value or ""))
    if not m:
        return None
    n = int(m.group(1))
    return n if 1 <= n <= 99 else None


def unit_text(nums):
    """`1~3,5호기 (사이 빔)` — 줄이되 **구멍은 그 자리에 드러낸다**([311])."""
    nums = sorted(set(nums))
    if not nums:
        return "모름"
    runs, start, prev = [], nums[0], nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        runs.append((start, prev))
        start = prev = n
    runs.append((start, prev))
    body = ",".join("%d" % a if a == b else "%d~%d" % (a, b) for a, b in runs)
    gap = " (사이 빔)" if len(nums) != nums[-1] - nums[0] + 1 else ""
    return body + "호기" + gap


def read_master():
    """정본 워크북을 분기 시트마다 읽는다 -> (경로, {(연,분기): {키: 캠프}})."""
    import openpyxl

    files = [p for p, _m, _z in CC._sched_files() if MASTER_HINT in os.path.basename(p)]
    if not files:
        return None, {}
    path = files[0]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    per_quarter = {}
    for sheet in wb.sheetnames:
        if "정기점검" not in sheet:
            continue
        ws = wb[sheet]
        head = None
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            if row and any(str(c or "").strip() == "기존 캠프명" for c in row):
                head = [str(c or "").strip() for c in row]
                break
        if not head:
            continue
        pos = {n: (head.index(n) if n in head else None) for n in
               ("변경 캠프명", "기존 캠프명", "캠프주소", "호기", "종류", "모델",
                "지역1", "지역2")}
        camps = {}
        for row in ws.iter_rows(min_row=6, values_only=True):
            def get(name):
                i = pos[name]
                return "" if i is None or i >= len(row) else str(row[i] or "").strip()

            name = get("변경 캠프명") or get("기존 캠프명")
            if not name or not CC._sched_camp_ok(name):
                continue
            camp = camps.setdefault(CC._norm(name), {
                "캠프명": name, "주소": get("캠프주소"),
                "지역": " ".join(x for x in (get("지역1"), get("지역2")) if x),
                "호기": [], "표기": {name},
            })
            camp["표기"].add(name)
            if not camp["주소"]:
                camp["주소"] = get("캠프주소")
            n = unit_no(get("호기"))
            if n is not None:
                camp["호기"].append({"번호": n, "종류": get("종류"), "모델": get("모델")})
        if camps:
            per_quarter[CC._quarter_of(sheet)] = camps
    wb.close()
    return path, per_quarter


def canonical(per_quarter):
    """캠프마다 **가장 최신 분기**의 줄이 이긴다([326]). 어느 분기였는지 적는다."""
    out = {}
    for quarter in sorted(per_quarter, reverse=True):
        for key, camp in per_quarter[quarter].items():
            if key in out:
                out[key]["표기"] |= camp["표기"]
                continue
            row = dict(camp)
            row["표기"] = set(camp["표기"])
            row["근거분기"] = ("%d년 %d분기" % quarter) if quarter[0] else "분기 모름"
            out[key] = row
    return out


def fold_app_names(camps):
    """앱이 아는 표기 변형을 **주소로** 합친다 — 후보가 유일할 때만([172])."""
    path = os.path.join(CC.REPORT_DIR, "캠프_담당자.json")
    try:
        rows = json.load(io.open(path, encoding="utf-8"))["rows"]
    except (OSError, ValueError, KeyError, TypeError):
        return {"읽음": False, "합침": 0, "못합침": [], "모호": 0}

    by_addr = {}
    for key, camp in camps.items():
        a = addr_key(camp["주소"])
        if a:
            by_addr.setdefault(a, []).append(key)

    folded, unmatched, ambiguous = 0, [], 0
    for row in rows:
        if not row.get("정기점검"):
            continue
        name = str(row.get("캠프명") or "").strip()
        band = [n for n in (unit_no(x) for x in re.findall(r"\d{1,2}", str(row.get("호기") or "")))
                if n is not None]
        key = CC._norm(name)
        if key in camps:
            camps[key]["표기"].add(name)
            camps[key].setdefault("밴드호기", set()).update(band)
            continue
        hits = by_addr.get(addr_key(row.get("캠프주소")), [])
        if len(hits) == 1:
            camps[hits[0]]["표기"].add(name)
            camps[hits[0]].setdefault("밴드호기", set()).update(band)
            folded += 1
            continue
        if len(hits) > 1:
            ambiguous += 1
            why = ("그 주소에 정본 캠프가 %d곳이다 — 어느 쪽인지 원본이 안 말한다"
                   % len(hits))
        elif addr_key(row.get("캠프주소")):
            why = "정본에 그 주소가 없다 — 옛 캠프이거나 이름이 바뀐 것"
        else:
            why = "주소를 모른다 — 못 가름"
        unmatched.append({"캠프명": name, "주소": row.get("캠프주소") or "모름",
                          "왜 못 합쳤나": why,
                          "최근작업일": row.get("최근작업일") or "모름",
                          "정기점검 글수": row.get("정기점검건수") or 0})
    return {"읽음": True, "합침": folded, "못합침": unmatched, "모호": ambiguous}


def build():
    path, per_quarter = read_master()
    if not path:
        return {"ok": False, "왜": "정기점검 스케줄 원본을 못 찾았다 — 화면 숫자를 믿지 말 것"}
    camps = canonical(per_quarter)
    app = fold_app_names(camps)
    rows = []
    for key in sorted(camps, key=lambda k: camps[k]["캠프명"]):
        c = camps[key]
        units = c["호기"]
        nums = {u["번호"] for u in units}
        band_only = sorted((c.get("밴드호기") or set()) - nums)
        others = sorted(x for x in c["표기"] if x != c["캠프명"])
        rows.append({
            "캠프명": c["캠프명"], "지역": c["지역"], "주소": c["주소"] or "모름",
            "호기수": len(nums), "호기": unit_text(nums),
            "종류": " / ".join(sorted({u["종류"] for u in units if u["종류"]})) or "모름",
            "모델": " / ".join(sorted({u["모델"] for u in units if u["모델"]})) or "모름",
            "밴드에서 본 호기": (",".join(str(n) for n in band_only) + "호기") if band_only else "",
            "합친 다른 표기": " | ".join(others) or "",
            "다른 표기 수": len(others),
            "근거분기": c["근거분기"],
        })
    return {"ok": True, "원본": os.path.basename(path), "분기시트": len(per_quarter),
            "캠프": rows, "앱": app}


def write(data):
    from app_server import _plain_xlsx

    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    total = sum(r["호기수"] for r in data["캠프"])
    meta = {"title": "쿠팡 캠프 · 캠프별 호기 정리",
            "sub": ("캠프 %d곳 · 호기 %d대 · 근거 %s(분기 시트 %d장 · 최신 분기가 이김)"
                    " · 만든 때 %s" % (len(data["캠프"]), total, data["원본"],
                                    data["분기시트"], stamp))}
    guide = [
        ["호기", "정본 스케줄표에 줄로 적힌 실제 설치 호기다(한 줄 = 한 대)."],
        ["밴드에서 본 호기", "정본에 없고 밴드·카톡 글에서만 관측된 번호다 — 설치 대수가 아니다."],
        ["합친 다른 표기", "같은 주소로 확인돼 한 캠프로 합친 옛 이름·다른 표기다."],
        ["모름", "값이 없다는 뜻이 아니라 아직 못 찾았다는 뜻이다."],
        ["못 합친 캠프", "둘째 표에 이유와 함께 있다 — 조용히 버리지 않는다."],
    ]
    blob, _ = _plain_xlsx("캠프별 호기", data["캠프"], COLS, guide_lines=guide, meta=meta)
    with open(OUT_XLSX, "wb") as fh:
        fh.write(blob)
    miss = data["앱"]["못합침"]
    if miss:
        meta2 = {"title": "정본에 못 합친 캠프 표기",
                 "sub": "%d개 · 이유별로 적었다 · 만든 때 %s" % (len(miss), stamp)}
        blob2, _ = _plain_xlsx("못 합친 표기", miss, COLS_NO, meta=meta2)
        with open(OUT_XLSX.replace(".xlsx", "_못합친것.xlsx"), "wb") as fh:
            fh.write(blob2)
    with io.open(OUT_JSON, "w", encoding="utf-8") as fh:
        json.dump({"만든때": stamp, **data}, fh, ensure_ascii=False, indent=1, default=list)
    return OUT_XLSX


def main():
    data = build()
    if not data.get("ok"):
        print("!", data.get("왜"))
        return 2
    rows, app = data["캠프"], data["앱"]
    total = sum(r["호기수"] for r in rows)
    print("근거:", data["원본"], "· 분기 시트", data["분기시트"], "장(최신 분기가 이김)")
    print("캠프 %d곳 · 호기 %d대" % (len(rows), total))
    if app["읽음"]:
        print("앱 표기 변형 %d개를 주소로 합쳤다 · 못 합친 것 %d개"
              "(그중 한 주소에 정본 캠프가 여럿이라 안 합친 것 %d)"
              % (app["합침"], len(app["못합침"]), app["모호"]))
    else:
        print("앱 캠프 목록을 못 읽었다 — 표기 변형은 안 합쳤다(0건이 아니라 못 셈)")
    zero = [r for r in rows if r["호기수"] == 0]
    if zero:
        print("호기를 못 읽은 캠프 %d곳 — '모름' 으로 적었다(0대가 아니다)" % len(zero))
    print("만든 파일:", write(data))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
