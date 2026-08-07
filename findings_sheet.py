# -*- coding: utf-8 -*-
"""
findings_sheet.py — 확인필요현황을 관리대장 안의 시트로 통합 (단일 엑셀 관리)
================================================================================
별도 파일 대신 관리대장 본체에 『23_확인필요현황』 시트를 추가/갱신한다.
- 최초: zip 4개 파트 수술로 새 시트 추가 (workbook.xml, rels, [Content_Types], 시트XML)
- 이후: 해당 시트 XML만 교체 — 차트·도형·x14 검증은 바이트 단위 무손상
- 내용이 이전과 동일하면 새 버전을 만들지 않음(멱등) / 변경 시 vN+1 생성
- 워크북 규약 준수: 1행 제목, 2행 [사용법], 4행 머리글(HDR_ROW=4), 5행부터 데이터, 자동필터·틀고정

실행:  python findings_sheet.py [--master 경로]   (daily_run이 매일 자동 호출)
"""
import sys, os, re, zipfile
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from ecount_reconcile import load_config, resolve_master
from findings_export import collect
from ledger_writer import sheet_file_map, esc, col_letter

SHEET_NAME = "23_확인필요현황"
# ★ 상세를 **열로 가른다** (2026-08-07 지시: "별도의 엑셀 파일은 만들지 말고
#   관리대장으로만 관리해"). 예전에는 명세서번호·PO번호·프로젝트NO·확인방법을
#   '내용·근거' 한 칸에 뭉쳐 넣었고, 그래서 `daily_run` 주석이 *"23시트는 평면
#   목록이라 유형별 상세열을 담지 못한다"* 며 **별도 엑셀을 따로 만들고 있었다.**
#   담지 못한 게 아니라 안 나눴을 뿐이다. 열로 가르면 걸러 보기·정렬이 되고
#   별도 파일이 필요 없어진다.
HEADERS = ["구분", "ID", "문제유형", "캠프명", "담당자", "금액", "일자",
           "프로젝트NO", "명세서번호", "PO번호", "확인방법", "내용·근거"]
WIDTHS = [11, 15, 22, 20, 10, 12, 12, 13, 15, 15, 30, 46]


def flatten(data):
    """수집 결과 → 통합 행. 열 순서는 HEADERS 와 **반드시** 같아야 한다."""
    rows = []

    def add(gubun, _id, kind, camp, who, amt, day,
            prj="", stmt="", po="", how="", note=""):
        rows.append((gubun, _id, kind, camp, who, amt, day,
                     prj or "", stmt or "", po or "", how or "", note or ""))

    for r in data.get("정산_조치필요", []):
        add("정산", r.get("정산ID"), r.get("문제유형"), r.get("캠프명"), r.get("담당자"),
            r.get("공급가액"), r.get("완료일"),
            prj=r.get("프로젝트NO"), stmt=r.get("명세서번호"), po=r.get("PO번호"),
            how=r.get("확인방법"),
            note=("명세서 없음 · " if not r.get("명세서번호") else "")
                 + ("PO 없음" if not r.get("PO번호") else ""))
    for r in data.get("밴드_미확인", []):
        add("밴드", r.get("ID"), "밴드 게시 미확인", r.get("캠프명"), r.get("담당자"),
            "", r.get("완료일"), prj=r.get("프로젝트NO"),
            note="작업완료인데 밴드 게시글 미발견 — 게시 여부 확인")
    for r in data.get("카톡_미확인", []):
        # '자료없음'(카톡 시작 이전)은 여기 오지 않는다 — findings_export 가
        # 정확히 '미확인'만 거른다. 찾아볼 자료가 없는 것을 확인 대상에 올리면
        # 영원히 안 지워지는 줄이 된다(2026-08-07 지시).
        add("카톡", r.get("ID"), "카톡 보고 미확인", r.get("캠프명"), r.get("담당자"),
            "", r.get("완료일"), prj=r.get("프로젝트NO"), how=r.get("매칭근거"),
            note="작업완료인데 카톡 보고 미발견")
    for r in data.get("ERP원장_문제", []):
        add("ERP", r.get("정산ID") or r.get("전표"), f"ERP {r.get('유형','')}",
            r.get("캠프명", ""), r.get("담당자"), r.get("ERP금액", ""), "",
            prj=r.get("프로젝트NO"), stmt=r.get("전표"), how=r.get("판정"),
            note=(r.get("적요") or "")[:60])
    for r in data.get("쿠팡PO_문제", []):
        add("PO", r.get("PO번호") or r.get("정산ID"), f"PO {r.get('유형','')}",
            "", r.get("담당자"), r.get("쿠팡금액", ""), r.get("발행일", ""),
            po=r.get("PO번호"), how=r.get("판정"),
            note=(r.get("후보정산") or r.get("내용") or "")[:60])
    for r in data.get("문서_원장미등록", []):
        add("문서", r.get("프로젝트NO"), "원장 미등록", "", r.get("담당자"),
            r.get("공급가액") or "", r.get("발행일") or "",
            prj=r.get("프로젝트NO"), stmt=r.get("명세서번호"), how=r.get("확인방법"))
    for r in data.get("금액_불일치", []):
        add("금액", r.get("정산ID"), "작업금액 불일치", r.get("캠프명"), r.get("담당자"),
            r.get("차액"), "", prj=r.get("프로젝트NO"), stmt=r.get("명세서번호"),
            how=r.get("확인방법"),
            note=f"작업 {r.get('작업금액'):,}원 / 명세서 {r.get('명세서금액'):,}원")
    for r in data.get("날짜_미상", []):
        add("빈칸", r.get("프로젝트NO"), f"{r.get('빈칸')} 비어 있음", r.get("캠프명"),
            r.get("담당자"), "", "", prj=r.get("프로젝트NO"), how=r.get("확인방법"))
    return rows


def build_generic_sheet(sheet_name, headers, widths, rows, usage_text, empty_text="해당 없음 ✅"):
    """워크북 규약(1행 제목·2행 사용법·4행 머리글·5행부터 데이터·필터·틀고정) 시트 XML 생성."""
    def cell(cL, rn, v):
        if v is None or v == "":
            return ""
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return f'<c r="{cL}{rn}"><v>{v}</v></c>'
        return f'<c r="{cL}{rn}" t="inlineStr"><is><t>{esc(v)}</t></is></c>'

    n_cols = len(headers)
    last = 4 + max(1, len(rows))
    x = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
         '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
         f'<dimension ref="A1:{col_letter(n_cols)}{last}"/>',
         '<sheetViews><sheetView workbookViewId="0"><pane ySplit="4" topLeftCell="A5" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>',
         '<sheetFormatPr defaultRowHeight="15"/>',
         '<cols>' + "".join(f'<col min="{i+1}" max="{i+1}" width="{w}" customWidth="1"/>'
                            for i, w in enumerate(widths)) + '</cols>',
         '<sheetData>']
    x.append(f'<row r="1">{cell("A",1, sheet_name + " (에이전트 자동 갱신 — 수기 입력 금지)")}</row>')
    x.append(f'<row r="2">{cell("A",2, usage_text)}</row>')
    hdr = "".join(cell(col_letter(i + 1), 4, h) for i, h in enumerate(headers))
    x.append(f'<row r="4">{hdr}</row>')
    if rows:
        for j, r in enumerate(rows):
            rn = 5 + j
            x.append('<row r="%d">%s</row>' % (rn, "".join(cell(col_letter(i + 1), rn, v)
                                                           for i, v in enumerate(r))))
    else:
        x.append(f'<row r="5">{cell("A",5,empty_text)}</row>')
    x.append('</sheetData>')
    x.append(f'<autoFilter ref="A4:{col_letter(n_cols)}{last}"/>')
    x.append('</worksheet>')
    return "".join(x)


def build_sheet_xml(rows):
    return build_generic_sheet(
        SHEET_NAME, HEADERS, WIDTHS, rows,
        "[사용법] 모든 대조(정산·밴드·카톡·ERP·쿠팡PO)의 확인필요 건이 여기 모입니다. 4행 필터로 구분·유형별 조회. "
        "처리 후 다음 에이전트 실행 때 자동 반영·소멸. 원본 근거는 reports/ 리포트 참조.",
        empty_text="확인필요 없음 — 전부 정상 ✅")


def upsert(master, new_xml, sheet_name=SHEET_NAME, headers=None):
    headers = headers or HEADERS
    zin = zipfile.ZipFile(master)
    smap = sheet_file_map(zin)
    exists = sheet_name in smap
    if exists and zin.read(smap[sheet_name]).decode("utf-8") == new_xml:
        zin.close()
        return None, "변경 없음(멱등) — 새 버전 미생성"

    mv = re.search(r"_v(\d+)\.xlsx$", master)
    v = int(mv.group(1))
    dst = master.replace(f"_v{v}.xlsx", f"_v{v+1}.xlsx")
    if os.path.exists(dst):
        zin.close()
        raise SystemExit(f"{os.path.basename(dst)} 이미 존재 — 중단")

    edits = {}
    if exists:
        target = smap[sheet_name]
        edits[target] = new_xml.encode("utf-8")
    else:
        wbx = zin.read("xl/workbook.xml").decode("utf-8")
        rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        ct = zin.read("[Content_Types].xml").decode("utf-8")
        new_n = max(int(m) for m in re.findall(r"worksheets/sheet(\d+)\.xml", " ".join(zin.namelist()))) + 1
        new_rid = max(int(m) for m in re.findall(r'Id="rId(\d+)"', rels)) + 1
        new_sid = max(int(m) for m in re.findall(r'sheetId="(\d+)"', wbx)) + 1
        target = f"xl/worksheets/sheet{new_n}.xml"
        edits[target] = new_xml.encode("utf-8")
        edits["xl/workbook.xml"] = wbx.replace(
            "</sheets>", f'<sheet name="{sheet_name}" sheetId="{new_sid}" r:id="rId{new_rid}"/></sheets>', 1).encode("utf-8")
        edits["xl/_rels/workbook.xml.rels"] = rels.replace(
            "</Relationships>",
            f'<Relationship Id="rId{new_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{new_n}.xml"/></Relationships>', 1).encode("utf-8")
        edits["[Content_Types].xml"] = ct.replace(
            "</Types>",
            f'<Override PartName="/{target}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>', 1).encode("utf-8")

    zout = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        data = edits.get(it.filename, None)
        zout.writestr(it.filename, data if data is not None else zin.read(it.filename))
    for name, data in edits.items():
        if name not in zin.namelist():          # 신규 시트 파트
            zout.writestr(name, data)
    zout.close(); zin.close()

    # 검증: 무결성 / 시트 판독 / 타 파트 동일성
    z = zipfile.ZipFile(dst)
    assert z.testzip() is None, "zip 무결성 실패"
    import openpyxl
    w = openpyxl.load_workbook(dst, read_only=True)
    assert sheet_name in w.sheetnames, "시트 미생성"
    ws = w[sheet_name]
    hdr = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4, max_col=len(headers)))]
    assert hdr == headers, f"머리글 불일치: {hdr}"
    w.close()
    zsrc = zipfile.ZipFile(master)
    diff = [n for n in zsrc.namelist() if n not in edits and zsrc.read(n) != z.read(n)]
    assert not diff, f"의도치 않은 파트 변경: {diff}"
    zsrc.close(); z.close()
    return dst, f"v{v} → v{v+1} ({'시트 갱신' if exists else '시트 신규 추가'})"


def main():
    args = sys.argv[1:]
    if "--master" in args:
        master = args[args.index("--master") + 1]
    else:
        master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    rows = flatten(collect(master))
    xml = build_sheet_xml(rows)
    dst, msg = upsert(master, xml)
    print(f"확인필요 {len(rows)}건 → {SHEET_NAME}: {msg}")
    if dst:
        print("   ", dst)


if __name__ == "__main__":
    main()
