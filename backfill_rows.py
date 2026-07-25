# -*- coding: utf-8 -*-
"""
backfill_rows.py — 밴드 추출분을 관리대장 신규 행으로 일괄 등록 (월별 백필)
==============================================================================
`band_extract` 레코드 중 **원장 미등록** 건을 02_돌발AS접수 / 04_정기점검 시트의
빈 행에 순차 배치한다. 실제 쓰기는 ledger_writer(빈 칸만·근거 필수·vN+1)를 통해서만 한다.

안전장치:
  1. 프로젝트NO 중복 제거 — 같은 건의 '안내→완료' 다중 게시는 1건으로(작업완료 우선)
  2. 원장에 이미 있는 프로젝트NO는 제외 (재등록 방지)
  3. ID 열은 건드리지 않음 — A열 수식이 접수일자(yymm) 기준으로 자동 채번(6월 → AS-2606-xxx)
  4. 자동계산 열(점검상태 등)도 건드리지 않음 — 수식 셀은 '빈 칸만' 정책이 자동 보호
  5. 시트 용량(수식이 깔린 마지막 행) 초과 시 그 시트는 **등록하지 않고 중단** 후 리포트

실행:
  python backfill_rows.py --month 2026-06            # 미리보기(큐 적재 없음)
  python backfill_rows.py --month 2026-06 --apply    # 큐 적재 + 원장 반영(vN+1)
"""
import sys, os, csv
from datetime import datetime
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.environ.get("COUPANG_REPORT_DIR") or os.path.join(BASE_DIR, "reports")

# 시트별: (용량행, 키열, {원장열: 추출필드})  — ID·자동계산 열은 의도적으로 제외
SPEC = {
    "돌발AS": {"sheet": "02_돌발AS접수", "cap": 154, "key_col": "프로젝트NO", "map": {
        "프로젝트NO": "프로젝트NO", "캠프명": "캠프명", "접수일자": "작업일", "담당기사": "담당기사",
        "진행상태": "_진행상태", "작업완료일": "_완료일", "유상·무상·보험": "비용구분",
        "신청내용": "_내용", "비고": "_출처"}},
    "정기점검": {"sheet": "04_정기점검", "cap": 104, "key_col": "프로젝트NO", "map": {
        "프로젝트NO": "프로젝트NO", "캠프명": "캠프명", "점검예정일": "작업일", "담당기사": "담당기사",
        "실제점검일": "_완료일", "유상·무상·보험": "비용구분", "비고": "_출처"}},
}
DATE_COLS = {"접수일자", "작업완료일", "점검예정일", "실제점검일"}


def dedupe(recs):
    """프로젝트NO 1건 = 원장 1행. 작업완료 > 정보량 많은 순으로 대표 선정."""
    by = defaultdict(list)
    for r in recs:
        if r.get("프로젝트NO"):
            by[r["프로젝트NO"]].append(r)
    out = []
    for prj, group in by.items():
        best = sorted(group, key=lambda x: (
            x["진행상태"] != "작업완료",                       # 완료 우선
            -sum(1 for v in x.values() if v),                  # 정보 많은 것 우선
            x["게시일"]))[0]
        merged = dict(best)
        if len(group) > 1:                                     # 완료일은 그룹 내 최신 정보로 보강
            done = [g["작업일"] for g in group if g["진행상태"] == "작업완료" and g["작업일"]]
            if done and not merged["작업일"]:
                merged["작업일"] = max(done)
            merged["_중복"] = len(group)
        out.append(merged)
    return sorted(out, key=lambda r: (r["작업일"] or r["게시일"], r["프로젝트NO"]))


def enrich(r):
    """원장 코드값으로 변환한 파생 필드"""
    done = r["진행상태"] == "작업완료"
    r["_진행상태"] = {"작업완료": "작업완료", "접수·예정": "접수", "취소": "취소"}.get(r["진행상태"], "접수")
    r["_완료일"] = r["작업일"] if done else ""
    r["_내용"] = (r.get("업무유형", "") + " " + (r.get("문서상태") or "")).strip()[:80]
    src = f"밴드 자동등록({r.get('게시일','')})"
    if r.get("_중복"):
        src += f" 게시 {r['_중복']}건 통합"
    r["_출처"] = src
    return r


def sheet_state(master, sheet):
    """(첫 빈 행, 열이름→열번호)"""
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    cols = {str(h).strip(): i + 1 for i, h in enumerate(hdr) if h}
    last = 4
    j = cols["프로젝트NO"]
    for i, row in enumerate(ws.iter_rows(min_row=5, min_col=j, max_col=j, values_only=True)):
        if row[0] not in (None, ""):
            last = 5 + i
    wb.close()
    return last + 1, cols


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def main():
    args = sys.argv[1:]
    month = args[args.index("--month") + 1] if "--month" in args else None
    apply_mode = "--apply" in args

    sys.path.insert(0, BASE_DIR)
    from band_extract import load_records, ledger_projects
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])

    recs = load_records()
    if month:
        recs = [r for r in recs if (r["작업일"] or r["게시일"]).startswith(month)]
    known = ledger_projects(master)
    recs = [enrich(r) for r in dedupe(recs) if r["프로젝트NO"] not in known]

    groups = defaultdict(list)
    for r in recs:
        kind = "정기점검" if "정기점검" in r["업무유형"] else ("돌발AS" if "돌발" in r["업무유형"] else "기타")
        groups[kind].append(r)

    queue, plan, skipped = [], [], []
    for kind, spec in SPEC.items():
        rows = groups.get(kind, [])
        if not rows:
            continue
        start, cols = sheet_state(master, spec["sheet"])
        need, room = len(rows), spec["cap"] - start + 1
        if need > room:
            skipped.append(f"{spec['sheet']}: {need}건 필요 / 여유 {room}행 — 용량 초과로 전량 보류"
                           f"(엑셀에서 마지막 행 복사→아래 붙여넣기로 확장 후 재실행)")
            continue
        for i, r in enumerate(rows):
            rn = start + i
            for led_col, field in spec["map"].items():
                if led_col not in cols:
                    continue
                v = r.get(field, "")
                if v in (None, ""):
                    continue
                queue.append({"sheet": spec["sheet"], "cell": f"{col_letter(cols[led_col])}{rn}",
                              "key": f"{col_letter(cols[led_col])}{rn}", "key_col": "-", "col": led_col,
                              "value": v, "vtype": "date" if led_col in DATE_COLS else "text",
                              "evidence": f"밴드 백필 {r['프로젝트NO']} ({month or '전체'})",
                              "only_if_empty": True})
            plan.append({"시트": spec["sheet"], "행": rn, "프로젝트NO": r["프로젝트NO"],
                         "캠프명": r["캠프명"], "일자": r["작업일"], "기사": r["담당기사"],
                         "상태": r["_진행상태"], "비용": r["비용구분"]})
        print(f"{spec['sheet']}: {need}건 → {start}~{start+need-1}행 (여유 {room}행)")

    for k in groups:
        if k not in SPEC:
            skipped.append(f"{k} {len(groups[k])}건 — 대상 시트 매핑 없음(수동 확인)")

    os.makedirs(REPORT_DIR, exist_ok=True)
    base = os.path.join(REPORT_DIR, f"백필_{month or '전체'}_{datetime.now():%Y%m%d_%H%M}")
    if plan:
        with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(plan[0].keys())); w.writeheader(); w.writerows(plan)
    print(f"\n등록 대상 {len(plan)}건 / 셀 {len(queue)}개")
    for s in skipped:
        print("  [보류]", s)
    print("리포트:", base + ".csv" if plan else "(없음)")

    if not apply_mode:
        print("\n미리보기 모드 — 실제 등록: python backfill_rows.py --month %s --apply" % (month or ""))
        return
    if not queue:
        print("등록할 항목 없음"); return

    from ledger_writer import queue_add
    print("큐 적재:", queue_add(queue), "개 셀 → ledger_writer --apply 실행")
    import subprocess
    r = subprocess.run([sys.executable, os.path.join(BASE_DIR, "ledger_writer.py"), "--apply"],
                       cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace",
                       env={**os.environ, "PYTHONIOENCODING": "utf-8"})
    print("\n".join(r.stdout.splitlines()[-6:]))


if __name__ == "__main__":
    main()
