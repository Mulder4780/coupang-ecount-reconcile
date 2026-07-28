# -*- coding: utf-8 -*-
"""
dashboard_clean.py — 00_대시보드의 **채우기 내림 사고 잔해**를 걷어낸다
================================================================================
사용자 지시(2026-07-28): "엑셀 대시보드 화면 깔끔하게 정리해줘."

무슨 일이 있었나
  누군가 K5 를 아래로 끌어 채웠다. 그래서 `=COUNTIF(...담당자 미배정...)` 이 K열 전체에
  복사돼 **56칸이 전부 '26'** 이 됐다. 같은 사고가 B·E·H 열에도 나 있어 빈 구역마다
  '0' 이 깔렸다. 화면이 지저분한 건 그 결과다.

미관 문제가 아니다 — 표 하나가 죽어 있었다
  H46 은 '조회일 직접입력(선택)' **사람이 비워 두는 칸**인데, 흘러든 수식 때문에 0 이 됐다.
  적용일 K46 = IF(H46="", 집계기준일, H46) → H46 이 빈칸이 아니라 0 이므로 **적용일 = 0
  (1900-01-00)**. 당일 실적 표는 전부 `COUNTIFS(..., $K$46)` 이라 **한 건도 안 잡힌다.**
  좌상단 '신규 접수 5건' 과 당일 '신규 접수 0건' 이 어긋나 보인 이유가 이것이다.

안전 장치
  · 관리대장은 openpyxl 로 저장하지 않는다(차트 5개·도형이 깨진다) — **zip 패치**만 쓴다.
  · 셀을 통째로 지우지 않고 **스타일(s)은 남기고 내용만 비운다** — 칸 색·테두리가 그대로다.
  · 공유수식/배열수식이 없음을 먼저 확인한다(있으면 한 칸만 지워도 나머지가 깨진다).
  · 지운 뒤 **남겨야 할 칸이 전부 그대로인지** 다시 열어 확인하고, 다른 zip 파트가
    바뀌지 않았는지도 확인한다.

사용
  python dashboard_clean.py            # 무엇을 지울지만 보여 준다 (파일 안 건드림)
  python dashboard_clean.py --apply    # vN+1 로 정리본 생성
"""
import os
import re
import sys
import shutil
import zipfile
import html

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from workbook_patch import latest_master, sheet_xml_path  # noqa: E402
from fix_workbook import iter_tags  # noqa: E402

SHEET = "00_대시보드"
LAST_ROW = 83

# ── 살려야 할 칸 ──────────────────────────────────────────────────────────────
# 값이 들어가는 열만 다룬다. 라벨 열(A·D·G·J)과 차트데이터(N·O)는 **손대지 않는다.**
def _r(col, a, b):
    return {f"{col}{i}" for i in range(a, b + 1)}


KEEP = set()
KEEP |= _r("B", 3, 7)                     # 보고일·집계기준일·마감시간·보고자·최종갱신
KEEP |= _r("B", 10, 23)                   # [돌발 AS]
KEEP |= {"B25"}                           # 세부 경고 항목 수
KEEP |= _r("B", 36, 44)                   # [문서발행·기한]
KEEP |= _r("B", 47, 58)                   # 당일 실적 (머리글 '당일' 포함)
KEEP |= _r("B", 62, 64) | _r("B", 67, 71) # 검증 보강 ①②③ / 미수금
KEEP |= {"B74"}                           # 조회 월 선택
KEEP |= _r("B", 77, 83)                   # 월별 집계 — AS 접수
KEEP |= _r("C", 1, LAST_ROW)              # C열은 전부 안내문·월별 AS완료 (지울 것 없음)
KEEP |= _r("E", 10, 18)                   # [정기점검]
KEEP |= _r("E", 77, 83)                   # 월별 집계 — 정기점검 완료
KEEP |= _r("F", 1, LAST_ROW)              # F열은 월별 집계 명세서 발행(건)뿐
KEEP |= _r("H", 10, 21)                   # [거래서류·청구·수금]
KEEP |= _r("H", 47, 57)                   # 리스크·잔여 ('값' 머리글 포함)
KEEP |= _r("H", 77, 83)                   # 월별 집계 — 입금액
KEEP |= _r("K", 10, 19)                   # [불일치·누락]  ← 여기만 남기고 K열은 전부 잔해
KEEP |= {"K46"}                           # 적용일 = IF(H46="",집계기준일,H46)

VALUE_COLS = ("B", "C", "E", "F", "H", "K")

# H46 은 '사람이 비워 두는 입력칸'이다. 지우는 게 아니라 **원래 상태로 되돌리는 것**이라
# 따로 표시해 둔다(리포트에서 이유를 밝히기 위해).
INPUT_CELLS = {"H46": "조회일 직접입력(선택) — 비어 있어야 적용일이 집계기준일로 잡힌다"}


def _cell_content(cinner):
    """셀의 실제 내용. 캐시가 없는 수식도 내용으로 판정한다."""
    if cinner is None:
        return None
    if re.search(r"<f(?:\s|>|/)", cinner):
        fm = re.search(r"<f[^>]*>(.*?)</f>", cinner, re.S)
        return "=" + html.unescape((fm.group(1) if fm else "<formula>").strip())
    texts = re.findall(r"<t[^>]*>(.*?)</t>", cinner, re.S)
    if texts:
        value = html.unescape("".join(texts)).strip()
        return value or None
    vm = re.search(r"<v[^>]*>(.*?)</v>", cinner, re.S)
    if vm:
        value = html.unescape(vm.group(1)).strip()
        return value or None
    return None


def survey_xml(xml):
    """대시보드 XML에서 지울 칸 목록 — 수식 캐시가 비어 있어도 놓치지 않는다."""
    out = []
    for _s, _e, ctag, cinner in iter_tags(xml, "c"):
        rm = re.search(r'r="([A-Z]{1,3})(\d+)"', ctag)
        if not rm:
            continue
        col, row = rm.group(1), int(rm.group(2))
        ref = f"{col}{row}"
        if col not in VALUE_COLS or row > LAST_ROW or ref in KEEP:
            continue
        value = _cell_content(cinner)
        if value is not None:
            out.append((ref, value, INPUT_CELLS.get(ref, "채우기 내림 잔해")))
    return out


def survey(path):
    """지울 칸 목록 — (ref, 현재값, 사유). 원본은 ZIP 읽기만 한다."""
    with zipfile.ZipFile(path) as z:
        sp = sheet_xml_path(z, SHEET)
        return survey_xml(z.read(sp).decode("utf-8"))


def blank_cells(xml, refs):
    """<c r="K5" s="16"><f>…</f><v>26</v></c> → <c r="K5" s="16"/>
    스타일만 남기고 내용을 비운다. 칸 색·테두리가 유지되므로 표 모양이 그대로다."""
    done = 0
    for ref in refs:
        pat = re.compile(r'<c r="%s"((?:\s+[a-zA-Z:]+="[^"]*")*)\s*(?:/>|>.*?</c>)' % ref, re.S)
        m = pat.search(xml)
        if not m:
            continue
        attrs = m.group(1) or ""
        attrs = re.sub(r'\s+t="[^"]*"', "", attrs)      # 자료형은 값이 사라지면 의미 없다
        xml = xml[:m.start()] + f'<c r="{ref}"{attrs}/>' + xml[m.end():]
        done += 1
    return xml, done


def force_recalc(xml):
    """수식 캐시를 무효화해 **파일을 열면 곧바로 다시 계산**하게 한다.

    칸을 비우면 그 칸을 참조하던 수식의 저장된 결과는 옛 값 그대로다. 이걸 안 해 두면
    사용자가 파일을 열었을 때 '적용일'이 여전히 1900-01-00 으로 보인다."""
    if "fullCalcOnLoad" in xml:
        return xml
    if "<calcPr" in xml:
        return re.sub(r"<calcPr([^>]*?)/>", r'<calcPr\1 fullCalcOnLoad="1"/>', xml, count=1)
    return xml.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')


def apply(src, dst, refs):
    zin = zipfile.ZipFile(src)

    # ① 공유·배열 수식이 있으면 칸 하나만 건드려도 나머지가 깨진다 — 먼저 막는다
    sp = sheet_xml_path(zin, SHEET)
    xml = zin.read(sp).decode("utf-8")
    for bad, why in (('t="shared"', "공유수식"), ('t="array"', "배열수식")):
        if bad in xml:
            raise RuntimeError(f"{why}가 있어 칸 단위로 지우면 위험합니다 — 중단합니다")

    xml, n = blank_cells(xml, refs)
    changed = {sp: xml.encode("utf-8")}

    wbp = "xl/workbook.xml"
    changed[wbp] = force_recalc(zin.read(wbp).decode("utf-8")).encode("utf-8")

    tmp = dst[:-5] + ".tmp.xlsx"
    if os.path.exists(tmp):
        raise FileExistsError(f"임시 결과가 이미 존재: {tmp}")
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        # 주의: 원본 ZipInfo를 그대로 넘기면 writestr 가 오프셋을 변조한다 — 이름만 넘긴다
        zout.writestr(it.filename, changed.get(it.filename, zin.read(it.filename)))
    zout.close()
    zin.close()

    verify(src, tmp, refs)
    os.replace(tmp, dst)
    return n


def verify(src, out, refs):
    """지운 건 비었고, **남겨야 할 건 하나도 안 변했는지** 확인한다."""
    import openpyxl
    z = zipfile.ZipFile(out)
    assert z.testzip() is None, "zip 무결성 실패"

    wa = openpyxl.load_workbook(src, data_only=False)
    wb_ = openpyxl.load_workbook(out, data_only=False)
    sa, sb = wa[SHEET], wb_[SHEET]

    for ref in refs:
        assert sb[ref].value in (None, ""), f"{ref} 이 안 비워졌다"
    kept = 0
    for ref in sorted(KEEP):
        try:
            before, after = sa[ref].value, sb[ref].value
        except Exception:
            continue
        assert before == after, f"남겨야 할 {ref} 가 바뀌었다: {before!r} → {after!r}"
        if before not in (None, ""):
            kept += 1

    # 차트·도형은 관리대장의 핵심이다. 개수가 줄면 즉시 중단한다.
    ca, cb = len(getattr(sa, "_charts", [])), len(getattr(sb, "_charts", []))
    assert ca == cb, f"차트가 사라졌다: {ca} → {cb}"
    # 다른 시트는 손대지 않았는지
    assert wa.sheetnames == wb_.sheetnames, "시트 구성이 바뀌었다"
    wa.close(); wb_.close()

    zs = zipfile.ZipFile(src)
    allowed = {sheet_xml_path(zs, SHEET), "xl/workbook.xml"}
    diff = [n for n in zs.namelist() if n not in allowed and zs.read(n) != z.read(n)]
    assert not diff, f"의도치 않은 파트 변경: {diff}"
    zs.close(); z.close()
    print(f"  검증 통과 — 지움 {len(refs)}칸 · 보존 {kept}칸 · 차트 {cb}개 · 다른 파트 변경 없음")


def main():
    args = sys.argv[1:]
    do = "--apply" in args
    if do:
        from claim_guard import require
        require("ledger", "dashboard_clean 반영")
    if "--file" in args:
        src = args[args.index("--file") + 1]
    else:
        m = latest_master()
        src = m[0] if isinstance(m, tuple) else m
    print(f"원본: {os.path.basename(src)}\n")

    refs = survey(src)
    by_col = {}
    for ref, v, why in refs:
        by_col.setdefault(ref[0], []).append((ref, v, why))
    for col in VALUE_COLS:
        got = by_col.get(col)
        if not got:
            continue
        vals = sorted({str(v)[:12] for _r_, v, _w in got})
        print(f"  {col}열 {len(got):>2}칸 지움  값: {', '.join(vals[:5])}"
              + ("…" if len(vals) > 5 else ""))
        print(f"        {', '.join(r for r, _v, _w in got)}")
    for ref, v, why in refs:
        if ref in INPUT_CELLS:
            print(f"\n  ★ {ref} (현재 {v}) — {why}")
    print(f"\n합계 {len(refs)}칸")

    if not do:
        print("\n실제로 정리하려면:  python dashboard_clean.py --apply")
        return 0

    mm = re.search(r"_v(\d+)\.xlsx$", src)
    dst = re.sub(r"_v\d+\.xlsx$", f"_v{int(mm.group(1)) + 1}.xlsx", src)
    print(f"\n정리본: {os.path.basename(dst)}")
    n = apply(src, dst, [r for r, _v, _w in refs])
    print(f"  {n}칸 비움 + 열 때 자동 재계산 설정")
    print("\n엑셀에서 한 번 열어 주세요 — 수식이 다시 계산되며 '당일 실적' 표가 살아납니다.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
