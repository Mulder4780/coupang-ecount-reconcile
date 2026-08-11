# -*- coding: utf-8 -*-
"""
camp_standardize.py — 캠프명을 **ERP 거래처등록 기준**으로 표준화한다 (2026-08-11 지시)
================================================================================
사용자 지시: "ERP 기준으로 캠프명 매칭하는 알고리즘 구현해, ERP가 기준이야,
캠프명이 이상한 경우 다 잡아내고, ERP에 있는 거래처 코드 적용,
기존 캠프명을 ERP 기준으로 변경해"

원칙 (typo_watch·customer_index 에서 배운 그대로):
  · 매칭 사다리는 **customer_index 를 그대로 빌린다** — 완전일치 → 정규화 → 핵심코드
    → 접두사제외. 같은 판단을 두 곳에 적으면 언젠가 갈린다.
  · **유일 매칭만 바꾼다.** 후보 여럿·ERP 에 없음·입력오류(값 대신 들어간 표시)는
    절대 짐작으로 바꾸지 않고 리포트에 남겨 사람 몫으로 돌린다 — 잘못 바꾸는 것이
    안 바꾸는 것보다 나쁘다.
  · 변경은 기존 승인 경로 하나다: `ledger_db.enqueue` → 앱 DB(정본) 즉시 저장
    (감사로그·낙관잠금·변경 전/후 보존) + Excel 보관 큐. **관리대장을 직접 열어
    저장하지 않는다.**
  · 변경 전 값은 감사로그 외에 `reports/캠프명_표준화.json` 에도 남긴다(사람이 보는
    되돌리기 자료). 거래처코드는 29_거래처코드 시트(customer_index --sheet)가 정본이다.

  python camp_standardize.py           # 조사 + 리포트만 (아무것도 안 바꿈)
  python camp_standardize.py --queue   # 유일 매칭만 정본 반영 + Excel 보관 큐
"""
import sys, os, re, json
from datetime import datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from customer_index import clean, norm, core, bare, PLACEHOLDERS, load_customers

REPORT_MD = os.path.join(ROOT, "reports", "캠프명_표준화.md")
REPORT_JSON = os.path.join(ROOT, "reports", "캠프명_표준화.json")

# 캠프명 열을 가진 시트와 그 시트의 행 열쇠(ID) 열. app_store.SHEET_SPECS 와 같다.
ID_COL = {
    "02_돌발AS접수": "접수ID",
    "03_현장작업실적": "작업ID",
    "04_정기점검": "점검ID",
    "05_신규납품설치": "업무ID",
    "06_거래서류청구수금": "정산ID",
    "13_PO발주관리": "PO관리ID",
    "15_세금계산서관리": "계산서관리ID",
    "16_입금수금관리": "입금관리ID",
}

_PAREN = re.compile(r"\(([^)]*)\)")


def build_tables(custs):
    """ERP 거래처 목록 → 매칭 사다리용 사전 4벌 (customer_index.main 과 같은 순서)."""
    by_exact, by_norm, by_core, by_bare = {}, {}, {}, {}
    for c in custs:
        by_exact.setdefault(c["name"], []).append(c)
        by_norm.setdefault(norm(c["name"]), []).append(c)
        by_core.setdefault(core(c["name"]), []).append(c)
        by_bare.setdefault(bare(c["name"]), []).append(c)
    return by_exact, by_norm, by_core, by_bare


def match_camp(camp, tables):
    """캠프명 하나를 ERP 에 대 본다.

    돌려주는 갈래: junk(값 대신 들어간 표시) · std(이미 ERP 표준명 그대로) ·
    diff(유일 매칭인데 이름이 다름 → 바꿀 대상) · multi(후보 여럿) · none(ERP 에 없음)
    """
    if clean(camp).upper() in PLACEHOLDERS:
        return {"status": "junk"}
    by_exact, by_norm, by_core, by_bare = tables
    for table, key, how in ((by_exact, camp, "완전일치"),
                            (by_norm, norm(camp), "정규화"),
                            (by_core, core(camp), "핵심코드"),
                            (by_bare, bare(camp), "접두사 제외")):
        hit = table.get(key) or []
        if len(hit) == 1:
            c = hit[0]
            if c["name"] == camp:
                return {"status": "std", "cust": c, "how": how}
            return {"status": "diff", "cust": c, "how": how}
        if len(hit) > 1:
            return {"status": "multi", "how": how,
                    "names": [x["name"] for x in hit][:6],
                    "codes": [x["code"] for x in hit][:6]}
    return {"status": "none"}


def paren_conflict(a, b):
    """양쪽 다 괄호 지명이 있는데 서로 다르면 True — 바꾸긴 바꾸되 눈에 띄게 보고한다.

    typo_watch 의 교훈: '제주2MB(성읍리)' → '제주2MB(사계리)' 같은 변경은 핵심코드가
    같아 유일 매칭이지만, 괄호 안이 갈리는 것은 캠프 이전/오기 어느 쪽일 수도 있다.
    ERP 가 기준이므로 바꾸되(사용자 지시) 사람이 다시 볼 수 있게 따로 모은다.
    """
    pa, pb = _PAREN.findall(clean(a)), _PAREN.findall(clean(b))
    if not pa or not pb:
        return False
    return norm("|".join(pa)) != norm("|".join(pb))


def plan_rows(rows, custs):
    """행 목록 + ERP 거래처 → (바꿀 큐 항목, 리포트 자료). 순수 함수 — 검증이 주입한다.

    rows: [{"sheet","id","camp","src"}]  · custs: load_customers() 모양
    """
    tables = build_tables(custs)
    verdicts = {}                       # camp -> match_camp 결과 (캠프명 단위 판정)
    items, changes = [], []
    buckets = {"std": 0, "diff": 0, "multi": {}, "none": {}, "junk": {}}
    paren_watch = []
    for r in rows:
        camp = str(r.get("camp") or "").strip()
        if not camp:
            continue
        v = verdicts.get(camp)
        if v is None:
            v = verdicts[camp] = match_camp(camp, tables)
        st = v["status"]
        if st == "std":
            buckets["std"] += 1
        elif st == "multi":
            d = buckets["multi"].setdefault(
                camp, {"건수": 0, "후보": v.get("names", []), "codes": v.get("codes", [])})
            d["건수"] += 1
        elif st == "none":
            buckets["none"].setdefault(camp, 0)
            buckets["none"][camp] += 1
        elif st == "junk":
            buckets["junk"].setdefault(camp, 0)
            buckets["junk"][camp] += 1
        elif st == "diff":
            buckets["diff"] += 1
            c = v["cust"]
            items.append({
                "sheet": r["sheet"], "key": str(r["id"]), "key_col": ID_COL[r["sheet"]],
                "col": "캠프명", "value": c["name"], "vtype": "text",
                "only_if_empty": False,
                "evidence": f"ERP 거래처등록 {c['code']} · {v['how']} · 이전값 '{camp}'",
            })
            changes.append({"sheet": r["sheet"], "id": str(r["id"]),
                            "before": camp, "after": c["name"],
                            "code": c["code"], "how": v["how"], "src": r.get("src", "")})
            if paren_conflict(camp, c["name"]) and camp not in [p["before"] for p in paren_watch]:
                paren_watch.append({"before": camp, "after": c["name"], "code": c["code"]})
    mapping = {camp: {"after": v["cust"]["name"], "code": v["cust"]["code"], "how": v["how"]}
               for camp, v in verdicts.items() if v["status"] == "diff"}
    return items, {"buckets": buckets, "changes": changes, "mapping": mapping,
                   "paren_watch": paren_watch}


def collect_rows():
    """실데이터 행 수집 — 정본(앱 DB) 우선, Excel 은 DB 에 없는 행만 보탠다."""
    rows, seen = [], set()
    import app_store
    for sheet, idcol in ID_COL.items():
        try:
            db_rows = app_store.list_sheet_rows(sheet)
        except Exception:
            db_rows = []
        for r in db_rows:
            rid = str(r.get(idcol) or "").strip()
            if not rid:
                continue
            seen.add((sheet, rid))
            camp = str(r.get("캠프명") or "").strip()
            if camp:
                rows.append({"sheet": sheet, "id": rid, "camp": camp, "src": "db"})
    # Excel 보강 — 아직 DB 에 못 들어온 옛 행. 관리대장은 **읽기 전용**으로만 연다.
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    from ecount_reconcile import load_config, resolve_master
    master = resolve_master(load_config()["reconcile"]["master_xlsx"])
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    for sheet, idcol in ID_COL.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [str(h).strip() if h is not None else "" for h in
               next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        ix = {h: i for i, h in enumerate(hdr) if h}
        if "캠프명" not in ix or idcol not in ix:
            continue
        for row in ws.iter_rows(min_row=5, values_only=True):
            rid = row[ix[idcol]] if ix[idcol] < len(row) else None
            rid = str(rid).strip() if rid is not None else ""
            if not rid or (sheet, rid) in seen:
                continue
            camp = row[ix["캠프명"]] if ix["캠프명"] < len(row) else None
            camp = str(camp).strip() if camp is not None else ""
            if camp:
                rows.append({"sheet": sheet, "id": rid, "camp": camp, "src": "xlsx"})
    wb.close()
    return rows, os.path.basename(master)


def write_reports(report, master, erp_src, n_rows, queued):
    os.makedirs(os.path.dirname(REPORT_MD), exist_ok=True)
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    b = report["buckets"]
    json.dump({"at": now, "master": master, "erp_src": erp_src, "rows": n_rows,
               "queued": queued, "mapping": report["mapping"],
               "changes": report["changes"],
               "ambiguous": b["multi"], "unmatched": b["none"], "junk": b["junk"],
               "paren_watch": report["paren_watch"]},
              open(REPORT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    L = ["# 캠프명 표준화 — ERP 거래처등록 기준", "",
         f"- 생성: {now} · 원장: {master} · ERP 원천: {erp_src}",
         f"- 캠프명 있는 행 {n_rows} · 이미 표준 {b['std']} · **바꿀 것(유일 매칭) {b['diff']}** · "
         f"후보 여럿 {sum(d['건수'] for d in b['multi'].values())} · "
         f"ERP 에 없음 {sum(b['none'].values())} · 입력오류 {sum(b['junk'].values())}",
         f"- 반영: {'큐에 넣음 — 앱 DB 즉시(감사로그) + Excel 보관 회차' if queued else '아직 안 함 (--queue 로 반영)'}",
         ""]
    if report["paren_watch"]:
        L += ["## ⚠ 괄호 지명이 갈리는 변경 — ERP 기준으로 바꾸지만 사람이 한번 볼 것", ""]
        for p in report["paren_watch"]:
            L.append(f"- `{p['before']}` → `{p['after']}` [{p['code']}] — 캠프 이전인지 오기였는지 확인")
        L.append("")
    if report["mapping"]:
        L += ["## 바꾸는 이름 (유일 매칭만 · 근거 = ERP 거래처코드)", "",
              "| 기존 | → ERP 표준명 | 코드 | 연결방식 |", "|---|---|---|---|"]
        for camp, m in sorted(report["mapping"].items()):
            L.append(f"| {camp} | {m['after']} | {m['code']} | {m['how']} |")
        L.append("")
    if report["buckets"]["multi"]:
        L += ["## 후보 여럿 — 사람 몫 (자동으로 바꾸지 않음)", ""]
        for camp, d in sorted(report["buckets"]["multi"].items()):
            L.append(f"- `{camp}` ({d['건수']}행) 후보: " +
                     " / ".join(f"{n}[{c}]" for n, c in zip(d["후보"], d["codes"])))
        L.append("")
    if report["buckets"]["none"]:
        L += ["## ERP 에 없음 — 사람 몫 (거래처 등록이 없거나 이름이 크게 다름)", ""]
        for camp, n in sorted(report["buckets"]["none"].items(), key=lambda x: -x[1]):
            L.append(f"- `{camp}` ({n}행)")
        L.append("")
    if report["buckets"]["junk"]:
        L += ["## 입력오류 — 값 대신 들어간 표시 (캠프를 알 수 없어 사람 몫)", ""]
        for camp, n in sorted(report["buckets"]["junk"].items()):
            L.append(f"- `{camp!r}` ({n}행)")
        L.append("")
    L.append("변경 전 값 전체는 `reports/캠프명_표준화.json` 의 `changes` 에 남아 있다 (되돌리기 자료).")
    open(REPORT_MD, "w", encoding="utf-8").write("\n".join(L) + "\n")


def main():
    custs, erp_src = load_customers()
    if not custs:
        print("ERP 거래처등록 원본을 찾지 못함 — 표준의 근거가 없어 아무것도 하지 않는다")
        return 1
    rows, master = collect_rows()
    items, report = plan_rows(rows, custs)
    b = report["buckets"]
    print(f"캠프명 있는 행 {len(rows)} (ERP 거래처 {len(custs)} · {erp_src})")
    print(f"이미 표준 {b['std']} · 바꿀 것 {len(items)}행/{len(report['mapping'])}종 · "
          f"후보 여럿 {len(b['multi'])}종 · ERP 없음 {len(b['none'])}종 · 입력오류 {len(b['junk'])}종")
    for c in report["changes"][:5]:
        print(f"   {c['sheet']} {c['id']}: '{c['before']}' → '{c['after']}' [{c['code']}]")
    queued = False
    if "--queue" in sys.argv and items:
        import ledger_writer as L
        import ledger_db
        print("큐 추가:", L.queue_add(items))
        print("DB 흡수:", ledger_db.intake_json(source="camp_standardize"))
        st = ledger_db.status()
        print(f"Excel 보관 대기 {st['대기']}건 · 다음 {st['다음반영']}")
        queued = True
    elif "--queue" in sys.argv:
        print("바꿀 것이 없다 — 큐에 넣을 항목 0건")
    write_reports(report, master, erp_src, len(rows), queued)
    print(f"리포트: {REPORT_MD}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
